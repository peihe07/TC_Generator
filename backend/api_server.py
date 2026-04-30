"""FastAPI server for frontend integration."""
from __future__ import annotations

import asyncio
import json
import logging
import os
import re
import tempfile
from datetime import datetime
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
from dotenv import load_dotenv

from generator import (
    DEFAULT_MODEL,
    GenerationError,
    calculate_cost,
    classify_test_sets,
    decompose_requirement,
    generate_quick_tc,
    generate_tcs_for_row,
)
from id_generator import generate_group_abbreviation, generate_tc_ids, normalize_tc_id
from job_store import SqliteJobStore, default_db_path
from parser import parse_tc_xlsx
from review_assistant import suggest_review_fix
from spec_matcher import build_spec_index, load_spec_index, match_spec_references
from spec_parser import detect_format, parse_docx, parse_pdf, parse_xlsx
from tools import (
    ToolError,
    aggregate_metrics_tool,
    generate_tc_tool,
    group_tests_tool,
    match_spec_tool,
    parse_workbook_tool,
    validate_tc_tool,
    write_excel_tool,
)
from tools.group import derive_test_set_name
from writer import build_output_path


def _tool_error_to_http(exc: ToolError) -> HTTPException:
    """Tool 例外 → HTTP 例外的統一翻譯。"""
    return HTTPException(status_code=exc.http_status, detail=exc.message)


def _is_quota_exceeded_tool_error(exc: ToolError) -> bool:
    return exc.code == "quota_exceeded"

load_dotenv()

app = FastAPI(title="tc-generator-api")

# CORS：開發預設只允 localhost，部署時用 TC_CORS_ORIGINS 覆寫（逗號分隔）。
# 設為 "*" 重現舊行為（不建議在公開環境）。
_default_origins = "http://localhost:3000,http://127.0.0.1:3000"
_cors_origins = [o.strip() for o in os.environ.get("TC_CORS_ORIGINS", _default_origins).split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

ALLOWED_RAW_EXTENSIONS = {".xlsx", ".xlsm"}
# Upload size guard（單檔 50 MB；TC_MAX_UPLOAD_MB env 可覆寫）
MAX_UPLOAD_BYTES = int(os.environ.get("TC_MAX_UPLOAD_MB", "50")) * 1024 * 1024
# SQLite-backed job registry — server 重啟後 jobs 仍可被檢索
JOB_REGISTRY = SqliteJobStore(default_db_path())

# Telemetry events — 從前端收 client-side analytics events，append-only JSONL。
# TC_EVENTS_LOG 可覆寫路徑（測試會用 monkeypatch 把它指到 tmp 檔）。
import threading

_TELEMETRY_LOCK = threading.Lock()
ALLOWED_TELEMETRY_EVENTS = {
    "experiment_exposure",
    "home_new_run_click",
    "builder_step_next",
    "builder_validation_fail",
    "run_execute_start",
    "run_execute_success",
    "run_execute_fail",
    "run_retry_click",
    "template_use_click",
    "template_save",
    "output_compare_open",
}
MAX_TELEMETRY_BATCH = 100
MAX_TELEMETRY_PROPS_BYTES = 4 * 1024  # 4 KB per event


def telemetry_log_path() -> Path:
    """目前 events log 路徑。每次 call 重算讓測試 monkeypatch env 生效。"""
    override = os.environ.get("TC_EVENTS_LOG")
    if override:
        return Path(override)
    return Path(__file__).resolve().parent.parent / "output" / "events.jsonl"


# ---------------------------------------------------------------------
# Job timeline — 在 SQLite job dict 內維護一個事件清單，給 Run Detail 使用
# ---------------------------------------------------------------------
JOB_TIMELINE_MAX = 50


VALIDATION_LOG_MAX = 500

WORKSPACE_HEADER = "x-workspace-id"
DEFAULT_WORKSPACE_ID = "default"


def _workspace_id(request: Request | None) -> str:
    """Read X-Workspace-Id header; fall back to ``default`` for backwards compat."""
    raw = request.headers.get(WORKSPACE_HEADER) if request else None
    value = (raw or "").strip()
    return value or DEFAULT_WORKSPACE_ID


def _record_stream_validation_failure(
    job_id: str,
    row: dict,
    message: str,
    *,
    field: str | None = None,
    severity: str = "error",
) -> None:
    """Inline record of a row failure into job["validationLog"].

    Mirrors POST /api/jobs/{id}/validation-logs semantics so reviewers see
    the same entries even before opening the Review step.
    """
    try:
        job = JOB_REGISTRY.get(job_id)
    except Exception:
        return
    if not job:
        return
    existing = job.get("validationLog")
    if not isinstance(existing, list):
        existing = []
    by_row: dict[str, dict] = {}
    for prior in existing:
        if isinstance(prior, dict) and prior.get("rowId"):
            by_row[str(prior["rowId"])] = prior
    row_id = str(row.get("id", "")) or f"row-{row.get('row_num', '?')}"
    by_row[row_id] = {
        "rowId": row_id,
        "reqId": str(row.get("req_id", "") or "") or None,
        "severity": severity,
        "field": field,
        "message": message,
        "ts": int(datetime.utcnow().timestamp() * 1000),
    }
    merged = list(by_row.values())
    if len(merged) > VALIDATION_LOG_MAX:
        merged = merged[-VALIDATION_LOG_MAX:]
    job["validationLog"] = merged
    try:
        JOB_REGISTRY[job_id] = job
    except Exception:
        pass


def _append_job_event(job_id: str, kind: str, message: str = "", **extra) -> None:
    """Append a timeline event to the job dict. Silent no-op if job missing."""
    try:
        job = JOB_REGISTRY.get(job_id)
    except Exception:
        return
    if not job:
        return
    timeline = job.get("timeline")
    if not isinstance(timeline, list):
        timeline = []
    event: dict = {
        "kind": kind,
        "ts": int(datetime.utcnow().timestamp() * 1000),
    }
    if message:
        event["message"] = message
    for k, v in extra.items():
        if v is None:
            continue
        event[k] = v
    timeline.append(event)
    # 上限保護，避免歷史無限長
    if len(timeline) > JOB_TIMELINE_MAX:
        timeline = timeline[-JOB_TIMELINE_MAX:]
    job["timeline"] = timeline
    try:
        JOB_REGISTRY[job_id] = job
    except Exception:
        # 寫入失敗不影響主流程
        pass

# Startup housekeeping：啟動時清掉超過 JOBS_MAX_AGE_DAYS（預設 30 天）的舊 job
# 並壓縮檔案。環境變數 `TC_JOBS_MAX_AGE_DAYS` 可覆寫（設 0 停用）。
try:
    _max_age_days = int(os.environ.get("TC_JOBS_MAX_AGE_DAYS", "30"))
    if _max_age_days > 0:
        _removed = JOB_REGISTRY.purge_older_than(_max_age_days * 86400)
        if _removed:
            JOB_REGISTRY.vacuum()
except (ValueError, Exception):
    pass

# 內建精簡規則：當 docs/ 規則檔案不存在時的 fallback
_FALLBACK_RULES = """
## ASPICE TC Writing Rules
- One behavior per TC, must match requirement intent
- Format: Condition/Trigger → Observable Outcome

## Pre-Conditions
- State or environment ONLY — never actions, checks/reads, or data-presence
  (e.g. "HU has 5,000 entries" → set up + read in a baseline step)
- Minimum necessary state, numbered list or NA

## Input Test Data
- Explicit deterministic values, or NA

## Test Procedure
- Setup steps → Transition steps → Final Step (verification)
- Each step: executable action + purpose
- Final step must include action + verification target
- Forbidden main verbs: observe / see if / check whether / confirm whether / verify / watch / monitor / inspect
  Use: Check that / Confirm that / Read / Record / Compare + explicit target

## Expected Result
- 1:1 mapping with procedure steps
- Observable, judgeable, no vague language

## Design Method (assign AFTER procedure+ER finalized)
- Judge from the ACTUAL flow via first-match on PRIMARY intent:
  Negative → Fault Injection → State Transition → Decision Table → EP → BVA → Combinatorial → Scenario → Functional

## Application Output Contract
- Priority is a workbook/tooling field, not an ASPICE rule in the instruction doc
- Return exactly P0 / P1 / P2 / P3
- P0: a feature's core/primary flow (the must-test happy path that defines the feature working at all); plus safety, boot/recovery, connection, audio output, eCall, vehicle-critical CAN signal, data loss risk. Default any "main functionality normal flow" test case to P0.
- P1: secondary or advanced operations of a major feature that are NOT the core primary flow — boundary/variation cases, key operational logic branches, non-primary user-facing flows
- P2: secondary/support functionality with limited major-feature impact
- P3: minor UI enhancement, low-impact customization, rare-use scenario, cosmetic detail
""".strip()

# 規則文件路徑（專案根目錄下 docs/）
_DOCS_DIR = Path(__file__).resolve().parent.parent / "docs"
_RULE_FILES = [
    _DOCS_DIR / "ASPICE_SWE6_AI_Instruction.md",
    _DOCS_DIR / "Test Case Design Method 判斷規則.md",
    _DOCS_DIR / "test_case_priority.md",
]


def _load_rules() -> str:
    """讀取 docs/ 下的規則 markdown 並串接；若全部缺失則退回精簡 fallback。"""
    sections: list[str] = []
    for path in _RULE_FILES:
        if path.is_file():
            try:
                text = path.read_text(encoding="utf-8").strip()
                if text:
                    sections.append(f"# {path.stem}\n\n{text}")
            except OSError:
                continue
    return "\n\n---\n\n".join(sections) if sections else _FALLBACK_RULES


RULES_SECTIONS = _load_rules()


@app.get("/api/health")
def healthcheck() -> dict:
    return {
        "status": "ok",
        "openai_configured": bool(os.environ.get("OPENAI_API_KEY")),
    }


class GenerateConfig(BaseModel):
    model: str
    batchSize: int = Field(ge=1)
    budget: float = Field(ge=0)
    strictValidation: bool = False
    # RULES.md §499 — skip rows that already have Pre-Cond / Procedure /
    # Expected content (reference examples that should not be overwritten).
    # Set True to force regeneration for all rows.
    regenerateAll: bool = False


class GenerateRow(BaseModel):
    id: str
    rowNum: int | None = None
    tcId: str | None = None
    reqId: str
    testItem: str
    originalRequirement: str | None = None
    testSet: str | None = None
    testSetHint: str | None = None
    specReference: str | None = None
    priority: str | None = None


class GenerateRequest(BaseModel):
    jobId: str | None = None
    rows: list[GenerateRow]
    config: GenerateConfig
    # Optional — when the user picked a template in the Builder, frontend
    # passes its name here so the backend can attribute usage analytics.
    templateId: str | None = None


class ExportRequest(BaseModel):
    jobId: str
    scope: str
    outputMode: str
    includeFrameworkSheet: bool = True
    selectedColumns: list[str] = Field(default_factory=list)
    rows: list[dict] = Field(default_factory=list)


class RegenerateRequest(BaseModel):
    rowIds: list[str]
    config: GenerateConfig | None = None
    rows: list[dict]
    regenerateReason: str | None = None
    # Optional — Rerun 當 backend JOB_REGISTRY 找不到 job 時用來重建 context。
    # 前端 Re-run flow 會帶這兩個欄位，regenerate 流程不使用。
    project: str | None = None
    testGroup: str | None = None


class GroupPreviewRequest(BaseModel):
    jobId: str
    rows: list[GenerateRow]
    forceRegroup: bool = False


class MatchPreviewRequest(BaseModel):
    jobId: str
    rows: list[GenerateRow]


class QuickGenerateRequest(BaseModel):
    testItem: str
    context: str | None = None
    testGroup: str | None = None
    testSet: str | None = None
    preConditions: str | None = None
    inputTestData: str | None = None
    testProcedure: str | None = None
    expectedResult: str | None = None
    # mode 保留做 backwards-compat，但目前統一走 auto-split 多筆 TC 路徑，
    # 傳任何值都會被忽略（包括舊前端的 "single" / "with_context" / "decompose"）。
    mode: str | None = None
    model: str = DEFAULT_MODEL


EXPORT_COLUMN_TO_FIELD = {
    "TC ID": "tc_id",
    "Test Set": "test_set",
    "TC Title": "tc_title",
    "Pre-Conditions": "pre_conditions",
    "Input Test Data": "input_test_data",
    "Test Procedure": "test_procedure",
    "Expected Result": "expected_result",
    "Priority": "priority",
    "Spec Reference": "spec_reference",
    "Design Method": "design_method",
    "AI 需求解讀": "reasoning",
    "AI Reasoning": "reasoning",
}


async def _read_with_limit(upload: UploadFile, label: str) -> bytes:
    """讀取 UploadFile 並限制總長度，超過 MAX_UPLOAD_BYTES 即 413。"""
    data = await upload.read()
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"{label} exceeds {MAX_UPLOAD_BYTES // (1024*1024)} MB limit",
        )
    return data


def _safe_upload_filename(filename: str | None, fallback: str) -> str:
    """Return a basename-only upload filename, rejecting traversal input."""
    raw = (filename or "").strip()
    if not raw:
        return fallback
    if "/" in raw or "\\" in raw:
        raise HTTPException(status_code=400, detail="filename must not contain path separators")
    safe = os.path.basename(raw)
    if safe in {"", ".", ".."}:
        raise HTTPException(status_code=400, detail="invalid filename")
    return safe


def _build_generate_job_id() -> str:
    return f"generate-{datetime.now().strftime('%Y%m%d-%H%M%S')}"


def _build_export_path(filename: str, output_mode: str) -> str:
    export_dir = os.path.join(tempfile.gettempdir(), "tc_generator_exports")
    os.makedirs(export_dir, exist_ok=True)
    filename = _safe_upload_filename(filename, "export.xlsx")
    if output_mode == "overwrite":
        return os.path.join(export_dir, filename)
    return build_output_path(os.path.join(export_dir, filename))


def _resolve_duplicate_of(raw: object, siblings: list[dict] | None) -> str:
    """正規化 AI 回的 duplicate_of 字串為 row_num（純數字字串），需對得到實際
    sibling 才回傳；hallucinate 或無 siblings 一律回 "" 讓 frontend 不顯示
    badge（避免「(已刪除或未找到)」這種誤導訊息）。

    AI 可能回的格式：``11``、``"11"``、``"row #11"``、``"row 11"``、舊版 uuid
    等。這裡掃 siblings 找對應的 row_num。
    """
    if not siblings or raw is None:
        return ""
    text = str(raw).strip()
    if not text:
        return ""
    # 抽出數字（兼容 "row #11" / "row 11" / "11"）。
    import re as _re

    digits_match = _re.search(r"\d+", text)
    digits = digits_match.group(0) if digits_match else ""
    for s in siblings:
        sib_row_num = s.get("row_num")
        sib_id = str(s.get("id") or "")
        # 1) 數字直接 match row_num
        if (
            digits
            and isinstance(sib_row_num, int)
            and sib_row_num > 0
            and digits == str(sib_row_num)
        ):
            return str(sib_row_num)
        # 2) AI 回了 uuid（舊格式）— 用 id match 後也回 row_num
        if sib_id and text == sib_id and isinstance(sib_row_num, int) and sib_row_num > 0:
            return str(sib_row_num)
    return ""


def _map_export_rows(
    rows: list[dict],
    scope: str,
    test_group: str | None,
    parsed_rows: list[dict] | None = None,
) -> tuple[list[dict], dict]:
    """Returns (exportable_rows, classification_usage).

    classification_usage 是 Phase 2 那次 `classify_test_sets` AI 呼叫的 cost /
    tokens；沒 AI 呼叫（所有 row 都已有 testSet）時為全零。Caller（export_job）
    需把這份 usage 累加回 job 的 persisted usage，否則使用者看不到匯出時的
    AI 開銷。
    """
    # 前端 payload 的 rowNum 有時會漏（存到 SQLite 就變 null），
    # 導致 writer 拿不到正確的 Excel 列號。這裡用 parsedData 做 fallback：
    # 先依 reqId 找回原列，然後把 row_num 補進 export 資料。
    parsed_by_req_id: dict[str, dict] = {}
    parsed_by_tc_id: dict[str, dict] = {}
    if parsed_rows:
        for pr in parsed_rows:
            tc_id = pr.get("tc_id")
            if tc_id and tc_id not in parsed_by_tc_id:
                parsed_by_tc_id[tc_id] = pr
            rid = pr.get("req_id")
            if rid and rid not in parsed_by_req_id:
                parsed_by_req_id[rid] = pr

    # Phase 1：先篩出要被匯出的列（scope filter），之後統一做 AI 分類。
    filtered: list[dict] = []
    for row in rows:
        review_status = row.get("reviewStatus")
        if scope == "accepted" and review_status != "accepted":
            continue
        if scope == "flagged" and review_status != "flagged":
            continue
        filtered.append(row)

    # Phase 2：收集「沒有 test_set」的 row → 一次 AI 分類（整份 review 後才跑）。
    # 改採 per-row 分類（不再以 req_id 去重），同一 Requirement ID 對應多列
    # 且內容不同時，每列各自獲得 Test Set，而非被迫共用第一筆。
    unresolved_rows: list[dict] = []
    for row in filtered:
        if str(row.get("testSet") or "").strip():
            continue
        row_id = str(row.get("id") or "").strip()
        req_id = str(row.get("reqId") or "").strip()
        if not row_id and not req_id:
            continue
        unresolved_rows.append(
            {
                "id": row_id or req_id,
                "req_id": req_id,
                "test_item": str(row.get("testItem") or "").strip(),
            }
        )

    classified: dict[str, str] = {}  # key = row id (or req_id fallback)
    classify_usage = {
        "cost": 0.0,
        "inputTokens": 0,
        "outputTokens": 0,
        "cacheCreationTokens": 0,
        "cacheReadTokens": 0,
    }
    if unresolved_rows:
        try:
            result = classify_test_sets(
                unresolved_rows,
                test_group=test_group,
            )
            classified = result.assignments
            classify_usage = {
                "cost": float(result.cost or 0.0),
                "inputTokens": int(result.input_tokens or 0),
                "outputTokens": int(result.output_tokens or 0),
                "cacheCreationTokens": int(result.cache_creation_tokens or 0),
                "cacheReadTokens": int(result.cache_read_tokens or 0),
            }
        except GenerationError:
            # AI 失敗就留空 test_set，讓 Framework sheet 的空白 group 做為可見提示。
            classified = {}

    exportable = []
    for row in filtered:
        generated = row.get("generated") or {}

        row_num = row.get("rowNum")
        tc_id = row.get("tcId", "")
        req_id = row.get("reqId", "")
        if not row_num and tc_id:
            base = parsed_by_tc_id.get(tc_id)
            if base:
                row_num = base.get("row_num")
        if not row_num and req_id:
            base = parsed_by_req_id.get(req_id)
            if base:
                row_num = base.get("row_num")

        test_set = row.get("testSet")
        if not str(test_set or "").strip():
            row_id = str(row.get("id") or "").strip()
            test_set = classified.get(row_id, "") or classified.get(req_id, "")

        input_test_data = generated.get("inputTestData", "")
        if not str(input_test_data or "").strip():
            input_test_data = "NA"

        # AI 解讀 (splitDecision.reasoning) 只放 primary TC（subIndex=0/undefined）；
        # 前端 payload 已遵守此約束，這裡照搬即可，sub TC 一律拿到空字串、
        # writer.py 會跳過該 cell。
        split_decision = row.get("splitDecision") or {}
        reasoning = ""
        if isinstance(split_decision, dict):
            reasoning = str(split_decision.get("reasoning") or "")

        exportable.append(
            {
                "row_num": row_num,
                "tc_id": row.get("tcId", ""),
                "req_id": req_id,
                "test_group": test_group or row.get("testGroup"),
                "test_set": test_set,
                "tc_title": generated.get("tcTitle", ""),
                "pre_conditions": generated.get("preConditions", ""),
                "input_test_data": input_test_data,
                "test_procedure": generated.get("testProcedure", ""),
                "expected_result": generated.get("expectedResult", ""),
                "priority": generated.get("priority", ""),
                "design_method": generated.get("designMethod", ""),
                "spec_reference": row.get("specReference") or generated.get("specReference"),
                "reasoning": reasoning,
            }
        )

    return exportable, classify_usage


_TC_ID_RE = re.compile(r"^(.+)-(.+)-(\d+)$")


def _resequence_export_tc_ids(rows: list[dict]) -> int:
    """Renumber tc_id sequentially within each (project, group_abbr) bucket.

    Reviewer-side delete leaves gaps in the original sequence (`-001`, `-002`,
    `-005`…). Export expects gap-free contiguous IDs, so right before write-
    back we sort each project / group bucket by row_num (ascending) and
    reassign `-001, -002, -003, …` in that order.

    Returns the number of rows whose tc_id changed (≥0). Rows with no tc_id
    or with an unparseable tc_id are left untouched.
    """
    buckets: dict[tuple[str, str], list[dict]] = {}
    for row in rows:
        original = str(row.get("tc_id") or "").strip()
        if not original:
            continue
        match = _TC_ID_RE.match(original)
        if not match:
            continue
        project, group_abbr, _ = match.groups()
        buckets.setdefault((project, group_abbr), []).append(row)

    changed = 0
    for (project, group_abbr), bucket in buckets.items():
        bucket.sort(key=lambda r: (int(r.get("row_num") or 0), str(r.get("tc_id") or "")))
        new_ids = generate_tc_ids(project, group_abbr, len(bucket))
        for row, new_id in zip(bucket, new_ids):
            if row.get("tc_id") != new_id:
                row["tc_id"] = new_id
                changed += 1
    return changed


def _build_framework_rows(rows: list[dict], test_group: str | None) -> list[dict]:
    # tc_count：每個 TC 算 1；req_count：同 req_id 僅計一次（拆分不重複計）。
    tc_counts: dict[tuple[str | None, str | None], int] = {}
    req_sets: dict[tuple[str | None, str | None], set[str]] = {}
    for row in rows:
        key = (test_group or row.get("testGroup"), row.get("test_set"))
        tc_counts[key] = tc_counts.get(key, 0) + 1
        req_id = str(row.get("req_id") or "").strip()
        if req_id:
            req_sets.setdefault(key, set()).add(req_id)

    framework_rows = []
    for key, tc_count in tc_counts.items():
        group_name, test_set = key
        framework_rows.append(
            {
                "test_group": group_name or "",
                "test_set": test_set or "",
                "tc_count": tc_count,
                "req_count": len(req_sets.get(key, set())) or tc_count,
            }
        )
    return framework_rows


def _selected_export_fields(selected_columns: list[str]) -> set[str] | None:
    mapped = {
        EXPORT_COLUMN_TO_FIELD[column]
        for column in selected_columns
        if column in EXPORT_COLUMN_TO_FIELD
    }
    return mapped or None


def _extract_existing_sequence_numbers(rows: list[dict]) -> list[int]:
    sequences = []
    for row in rows:
        tc_id = row.get("tc_id") or row.get("tcId")
        if not tc_id:
            continue
        parts = tc_id.rsplit("-", 1)
        if len(parts) == 2 and parts[1].isdigit():
            sequences.append(int(parts[1]))
    return sequences


# Multi-TC 下每個 req 平均會回 2–3 筆 TC（§1.4 列舉型需求更多）。output token
# 以原本 1 TC 預估 × 這個倍數做為保守估計，避免 budget gate 低估導致超支。
_AVG_TCS_PER_REQ = 2.5
_AVG_INPUT_TOKENS_PER_REQ = 1500
_AVG_OUTPUT_TOKENS_PER_TC = 800


def _estimate_cost(row_count: int, model: str, batch_size: int) -> float:
    # input token 只跟 req 數有關（prompt 依 req 組合），output 會因 AI 拆分變多。
    del batch_size
    avg_output = int(_AVG_OUTPUT_TOKENS_PER_TC * _AVG_TCS_PER_REQ)
    return calculate_cost(
        _AVG_INPUT_TOKENS_PER_REQ * row_count,
        avg_output * row_count,
        model,
    )


def _would_exceed_budget(current_cost: float, batch_size: int, model: str, budget: float) -> bool:
    estimated_batch_cost = _estimate_cost(batch_size, model, batch_size)
    return current_cost + estimated_batch_cost > budget


def _job_usage(job: dict | None) -> dict:
    usage = (job or {}).get("usage") or {}
    by_model_raw = usage.get("costByModel") or {}
    return {
        "cost": float(usage.get("cost") or 0.0),
        "inputTokens": int(usage.get("inputTokens") or 0),
        "outputTokens": int(usage.get("outputTokens") or 0),
        "cacheCreationTokens": int(usage.get("cacheCreationTokens") or 0),
        "cacheReadTokens": int(usage.get("cacheReadTokens") or 0),
        "costByModel": {
            str(k): float(v or 0.0) for k, v in by_model_raw.items()
        } if isinstance(by_model_raw, dict) else {},
    }


def _persist_job_usage(
    job_id: str,
    job: dict,
    *,
    cost: float,
    input_tokens: int,
    output_tokens: int,
    cache_creation_tokens: int,
    cache_read_tokens: int,
    cost_delta: float = 0.0,
    model: str | None = None,
) -> None:
    """Write back cumulative usage. Optional `model` + `cost_delta` attribute
    the delta to a specific model bucket so the frontend can show a per-model
    breakdown (e.g. classify_test_sets on gpt-5-mini vs generate on gpt-5).

    Callers that already know the incremental cost of THIS call should pass
    `cost_delta` + `model`; the bucket grows by `cost_delta` and the cumulative
    `cost` field continues to track the total as before.
    """
    existing = (job.get("usage") or {}).get("costByModel") or {}
    by_model: dict[str, float] = {
        str(k): float(v or 0.0) for k, v in existing.items()
    } if isinstance(existing, dict) else {}
    if model and cost_delta > 0:
        by_model[model] = round(by_model.get(model, 0.0) + float(cost_delta), 6)
    job["usage"] = {
        "cost": cost,
        "inputTokens": input_tokens,
        "outputTokens": output_tokens,
        "cacheCreationTokens": cache_creation_tokens,
        "cacheReadTokens": cache_read_tokens,
        "costByModel": by_model,
    }
    JOB_REGISTRY[job_id] = job


def _build_spec_index_for_job(job: dict) -> dict:
    spec_bytes = job.get("specBytes")
    spec_filename = job.get("specFileName")
    spec_format = job.get("specFormat")
    if not spec_bytes or not spec_filename or not spec_format:
        return {}

    parsers = {"pdf": parse_pdf, "docx": parse_docx, "xlsx": parse_xlsx}
    parser = parsers.get(spec_format)
    if parser is None:
        return {}

    with tempfile.TemporaryDirectory() as tmp_dir:
        spec_path = os.path.join(tmp_dir, _safe_upload_filename(spec_filename, "spec.bin"))
        with open(spec_path, "wb") as spec_file:
            spec_file.write(spec_bytes)
        return parser(spec_path)


def _build_reference_match_index_for_job(job: dict) -> dict:
    selected_spec_name = (job.get("selectedSpecName") or "").strip()
    if selected_spec_name:
        try:
            return load_spec_index(names=[selected_spec_name])
        except Exception:
            return {}

    reference_bytes = job.get("referenceWorkbookBytes")
    reference_filename = job.get("referenceWorkbookName")
    if not reference_bytes or not reference_filename:
        return {}

    with tempfile.TemporaryDirectory() as tmp_dir:
        reference_path = os.path.join(
            tmp_dir,
            _safe_upload_filename(reference_filename, "reference.xlsx"),
        )
        with open(reference_path, "wb") as reference_file:
            reference_file.write(reference_bytes)
        try:
            return build_spec_index(reference_path)
        except Exception:
            return {}


def _prepare_match_preview_rows(rows: list[GenerateRow], job: dict) -> list[dict]:
    prepared_rows = []
    parsed_rows = job.get("parsedData", {}).get("rows", [])
    parsed_by_row_num = {row["row_num"]: row for row in parsed_rows}
    parsed_by_req_id = {row.get("req_id"): row for row in parsed_rows}

    for raw_row in rows:
        dump = raw_row.model_dump()
        base_row = None
        if dump.get("rowNum") is not None:
            base_row = parsed_by_row_num.get(dump["rowNum"])
        if base_row is None and dump.get("reqId"):
            base_row = parsed_by_req_id.get(dump["reqId"])
        base_row = base_row or {}
        prepared_rows.append(
            {
                "id": dump["id"],
                "req_id": dump.get("reqId") or base_row.get("req_id", ""),
                "test_item": dump.get("testItem") or base_row.get("test_item", ""),
            }
        )
    return prepared_rows


def _prepare_generation_rows(job: dict) -> list[dict]:
    parsed_rows = job.get("parsedData", {}).get("rows", [])
    parsed_by_row_num = {row["row_num"]: row for row in parsed_rows}
    parsed_by_req_id = {row.get("req_id"): row for row in parsed_rows}

    prepared_rows = []
    for raw_row in job["rows"]:
        base_row = None
        if raw_row.get("rowNum") is not None:
            base_row = parsed_by_row_num.get(raw_row["rowNum"])
        if base_row is None and raw_row.get("reqId"):
            base_row = parsed_by_req_id.get(raw_row["reqId"])
        base_row = base_row or {}

        prepared_rows.append(
            {
                "id": raw_row["id"],
                "row_num": raw_row.get("rowNum") or base_row.get("row_num") or 0,
                "tc_id": normalize_tc_id(raw_row.get("tcId") or base_row.get("tc_id", "")),
                "req_id": raw_row.get("reqId") or base_row.get("req_id", ""),
                "test_item": raw_row.get("testItem") or base_row.get("test_item", ""),
                "original_requirement": raw_row.get("originalRequirement")
                or base_row.get("test_item", ""),
                "test_set": raw_row.get("testSet")
                if raw_row.get("testSet") is not None
                else base_row.get("test_set", ""),
                "spec_reference": raw_row.get("specReference")
                if raw_row.get("specReference") is not None
                else base_row.get("spec_reference"),
                "priority": raw_row.get("priority")
                if raw_row.get("priority") is not None
                else base_row.get("priority", ""),
                # Carry existing content from parse so §499 preserve-check
                # can run. GenerateRow schema does not round-trip these.
                "pre_conditions": base_row.get("pre_conditions", ""),
                "input_test_data": base_row.get("input_test_data", ""),
                "test_procedure": base_row.get("test_procedure", ""),
                "expected_result": base_row.get("expected_result", ""),
                "design_method": base_row.get("design_method", ""),
            }
        )

    # Sibling annotation：同一 Requirement ID 對應多列時，把其他列的 test_item
    # 帶進來當 prompt context，讓 AI 在寫 tc_title trigger 時做 sibling-distinction
    # （§6.1），並能主動 flag 重複場景（split_flag=true / split_reason）。
    rows_by_req: dict[str, list[dict]] = {}
    for prepared in prepared_rows:
        rid = str(prepared.get("req_id") or "").strip()
        if rid:
            rows_by_req.setdefault(rid, []).append(prepared)
    for prepared in prepared_rows:
        rid = str(prepared.get("req_id") or "").strip()
        peers = rows_by_req.get(rid, [])
        if len(peers) <= 1:
            continue
        prepared["siblings"] = [
            {
                "id": peer["id"],
                "row_num": peer.get("row_num") or 0,
                "test_item": str(peer.get("test_item") or "").strip(),
            }
            for peer in peers
            if peer["id"] != prepared["id"]
        ]

    reference_index = _build_reference_match_index_for_job(job)
    if reference_index:
        matched_rows = match_spec_references(prepared_rows, reference_index)
        for prepared_row, matched_row in zip(prepared_rows, matched_rows):
            prepared_row["spec_reference"] = matched_row.get("spec_reference")
            prepared_row["match_type"] = matched_row.get("match_type")
            if matched_row.get("match_score") is not None:
                prepared_row["match_score"] = matched_row.get("match_score")
            if matched_row.get("matched_spec_context"):
                prepared_row["matched_spec_context"] = matched_row.get("matched_spec_context")
            if matched_row.get("reference_candidate_context"):
                prepared_row["reference_candidate_context"] = matched_row.get("reference_candidate_context")

    existing_sequences = _extract_existing_sequence_numbers(parsed_rows) + _extract_existing_sequence_numbers(prepared_rows)
    missing_rows = [row for row in prepared_rows if not row.get("tc_id")]
    if missing_rows:
        project = job.get("parsedData", {}).get("project") or "project"
        test_group = job.get("parsedData", {}).get("test_group") or "TC"
        start = max(existing_sequences) + 1 if existing_sequences else 1
        tc_ids = generate_tc_ids(
            project,
            generate_group_abbreviation(test_group),
            len(missing_rows),
            start=start,
        )
        for row, tc_id in zip(missing_rows, tc_ids):
            row["tc_id"] = tc_id

    return prepared_rows


def _build_stream_row(row: dict, tc: dict) -> tuple[dict, bool]:
    result = validate_tc_tool(
        tc={
            "tc_id": row["tc_id"],
            "test_item": f"{row['original_requirement']}\n\n{tc['tc_title']}",
            "pre_conditions": tc["pre_conditions"],
            "test_procedure": tc["test_procedure"],
            "expected_result": tc["expected_result"],
            "design_method": tc["design_method"],
            "priority": tc["priority"],
        }
    )
    validation = result["issues"]
    has_warnings = result["hasWarnings"]
    return (
        {
            "id": row["id"],
            "rowNum": row["row_num"],
            "tcId": row["tc_id"],
            "reqId": row["req_id"],
            "testItem": row["test_item"],
            "originalRequirement": row["original_requirement"],
            "testSet": row.get("test_set", ""),
            "specReference": row.get("spec_reference"),
            "priority": tc["priority"],
            # Validation warnings should still land in Review as a generated row.
            # Only the SSE event type decides whether strict mode upgrades them
            # to `row.failed`; non-strict mode should not surface them as fail.
            "status": "ready",
            "reviewStatus": "pending",
            "generated": {
                "tcTitle": tc["tc_title"],
                "preConditions": tc["pre_conditions"],
                "inputTestData": tc["input_test_data"],
                "testProcedure": tc["test_procedure"],
                "expectedResult": tc["expected_result"],
                "designMethod": tc["design_method"],
                "priority": tc["priority"],
                "specReference": row.get("spec_reference"),
            },
            "validation": validation,
        },
        has_warnings,
    )


# §1.4 False-Pass 啟發式：需求文字若像在列舉多個支援項目但 AI 只回 1 筆 TC，
# 很大機率違反 Mistake #4，flag 一個 warning 讓使用者在 Review UI 注意。
_ENUM_HINTS_RE = re.compile(
    r"(\.\w{2,5}\s*[,/、]\s*\.\w{2,5})"             # .mp4, .avi 這類副檔名列舉
    r"|(\b(?:one of|any of|such as|including|e\.g\.|e\.g)\b)"
    r"|(\b(?:support|supported|supports)\b[^\n]{0,80}[,/、])",
    re.IGNORECASE,
)
_BOUNDARY_HINTS_RE = re.compile(
    r"\b(maximum|minimum|upper limit|lower limit|at most|at least|no more than|no less than)\b",
    re.IGNORECASE,
)


def _heuristic_split_warning(row: dict, tc_count: int) -> str:
    """回傳非空字串就表示此次拆分可能有問題（讓 Review UI 標 warning）。"""
    if tc_count >= 2:
        return ""
    text = str(row.get("test_item") or row.get("original_requirement") or "")
    if not text.strip():
        return ""
    if _ENUM_HINTS_RE.search(text):
        return (
            "需求像是列舉多個支援項目（§1.4 Mistake #4），但 AI 只產出 1 筆 TC。"
            "請確認是否遺漏對各項目的獨立驗證。"
        )
    if _BOUNDARY_HINTS_RE.search(text):
        return (
            "需求提及 maximum / minimum / 上下限字眼（§1.2 & BVA），"
            "建議至少拆成「= 上限」「> 上限」「< 上限」等獨立情境。"
        )
    return ""


def _row_has_existing_content(row: dict) -> bool:
    """True if the parsed row already carries AI-equivalent content.

    Used to honour RULES.md §499 — rows whose Pre-Cond / Procedure / Expected
    cells are already filled (typically reference examples) should not be
    regenerated and billed for.
    """
    for field in ("pre_conditions", "test_procedure", "expected_result"):
        value = row.get(field)
        if value and str(value).strip():
            return True
    return False


def _build_preserved_stream_row(row: dict) -> dict:
    """Return a row.completed-shaped dict that surfaces the row's existing
    content verbatim (no AI call). Shape matches `_build_stream_row` so the
    Review UI and the export writer treat it identically.
    """
    return {
        "id": row["id"],
        "rowNum": row["row_num"],
        "tcId": row["tc_id"],
        "reqId": row["req_id"],
        "testItem": row["test_item"],
        "originalRequirement": row["original_requirement"],
        "testSet": row.get("test_set", ""),
        "specReference": row.get("spec_reference"),
        "priority": row.get("priority", ""),
        "status": "ready",
        "reviewStatus": "pending",
        "generated": {
            # No rewrite — writer keeps Col I (Test Item) untouched.
            "tcTitle": "",
            "preConditions": row.get("pre_conditions", ""),
            "inputTestData": row.get("input_test_data", ""),
            "testProcedure": row.get("test_procedure", ""),
            "expectedResult": row.get("expected_result", ""),
            "designMethod": row.get("design_method", ""),
            "priority": row.get("priority", ""),
            "specReference": row.get("spec_reference"),
        },
        "validation": [],
        "preserved": True,
    }


def _build_failed_stream_row(row: dict, message: str) -> dict:
    return {
        "id": row["id"],
        "rowNum": row["row_num"],
        "tcId": row.get("tc_id", ""),
        "reqId": row.get("req_id", ""),
        "testItem": row.get("test_item", ""),
        "originalRequirement": row.get("original_requirement", ""),
        "testSet": row.get("test_set", ""),
        "specReference": row.get("spec_reference"),
        "priority": row.get("priority", ""),
        "status": "error",
        "reviewStatus": "pending",
        "generated": None,
        "validation": [
            {
                "id": f"runtime-{row.get('id', 'row')}",
                "severity": "warning",
                "field": "runtime",
                "message": message,
            }
        ],
    }


def _sse_event(payload: dict) -> str:
    return f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"


def _scenario_title(tc: dict, index: int) -> str:
    title = str(tc.get("tc_title") or "").strip()
    return title or f"TC {index}"


@app.post("/api/quick-generate/stream")
async def stream_quick_generate(payload: QuickGenerateRequest, request: Request) -> StreamingResponse:
    """Ad-hoc TC generation from manual input.

    統一走 multi-TC auto-split 路徑：AI 自動依 ASPICE SWE.6 §6.1（Test Item）、
    §9（False Pass 拆分）、§10.2（延伸情境）判斷需要幾筆 TC，先回
    reasoning + keyword 分析，再串流每筆 TC，最後收尾。
    不再保留舊的 single / with_context / decompose 三模式分歧。
    """

    async def event_generator():
        model = payload.model
        current_cost = 0.0

        yield _sse_event({"type": "job.started"})

        # 組成與 workbook row 相同語意的 context，讓 quick generate 也用同一套
        # multi-TC split 分析輸入。
        test_item = payload.testItem.strip()
        context_text = (payload.context or "").strip()
        test_group = (payload.testGroup or "QuickGenerate").strip() or "QuickGenerate"
        test_set = (payload.testSet or "Quick Generate").strip() or "Quick Generate"
        row = {
            "req_id": "QUICK",
            "test_group": test_group,
            "test_set": test_set,
            "test_item": test_item,
            "pre_conditions": (
                str(payload.preConditions).strip()
                if payload.preConditions and str(payload.preConditions).strip()
                else context_text
            ),
            "input_test_data": str(payload.inputTestData or "").strip(),
            "test_procedure": str(payload.testProcedure or "").strip(),
            "expected_result": str(payload.expectedResult or "").strip(),
        }
        ctx = {"project": "QuickGenerate", "test_group": test_group, "test_set": test_set}

        try:
            task = asyncio.create_task(
                asyncio.to_thread(
                    generate_tcs_for_row,
                    row=row,
                    context=ctx,
                    spec_index=None,
                    rules_text=RULES_SECTIONS,
                    model=model,
                )
            )
            while not task.done():
                if await request.is_disconnected():
                    task.cancel()
                    return
                await asyncio.sleep(0.1)
            result = task.result()
        except GenerationError as exc:
            if await request.is_disconnected():
                return
            yield _sse_event({"type": "job.failed", "message": str(exc)})
            return

        tcs = result.tc_data if isinstance(result.tc_data, list) else [result.tc_data]
        meta = result.split_meta[0] if result.split_meta else {}
        current_cost += result.cost
        total = len(tcs)

        # 1) 先送 analysis：reasoning + keyword coverage + synthetic scenarios list，
        #    讓前端的 DecomposeAnalysisPanel 維持原有 UI、不必大改。
        scenarios = [
            {
                "id": i + 1,
                "name": _scenario_title(tc, i + 1),
                "description": (tc.get("tc_title") or "").strip() or f"Generated TC {i + 1}",
                "test_item": test_item,
            }
            for i, tc in enumerate(tcs)
        ]
        # 正規化 keyword 欄位為前端格式（scenarios: number[]）。
        kw_out: list[dict] = []
        for k in (meta.get("keywords") or []):
            if not isinstance(k, dict):
                continue
            cb = k.get("covered_by") or k.get("coveredBy") or k.get("scenarios") or []
            if not isinstance(cb, list):
                cb = []
            kw_out.append(
                {
                    "keyword": str(k.get("keyword") or ""),
                    "meaning": str(k.get("meaning") or ""),
                    "scenarios": [int(n) for n in cb if isinstance(n, (int, float))],
                }
            )
        if await request.is_disconnected():
            return
        yield _sse_event(
            {
                "type": "decompose.analysis",
                "reasoning": str(meta.get("reasoning") or ""),
                "keywords": kw_out,
                "scenarios": scenarios,
                "stats": {"total": total, "processed": 0, "currentCost": round(current_cost, 4)},
            }
        )

        # 2) 依序送每筆 TC（沿用 tc.generating / tc.completed event shape，前端無需改）。
        for i, tc in enumerate(tcs):
            if await request.is_disconnected():
                return
            yield _sse_event(
                {
                    "type": "tc.generating",
                    "scenarioId": i + 1,
                    "scenarioName": _scenario_title(tc, i + 1),
                    "stats": {"total": total, "processed": i, "currentCost": round(current_cost, 4)},
                }
            )
            await asyncio.sleep(0.03)
            if await request.is_disconnected():
                return
            yield _sse_event(
                {
                    "type": "tc.completed",
                    "scenarioId": i + 1,
                    "scenarioName": _scenario_title(tc, i + 1),
                    "tc": tc,
                    "stats": {"total": total, "processed": i + 1, "currentCost": round(current_cost, 4)},
                }
            )

        if await request.is_disconnected():
            return
        yield _sse_event(
            {
                "type": "job.completed",
                "stats": {
                    "total": total,
                    "processed": total,
                    "currentCost": round(current_cost, 4),
                    "inputTokens": result.input_tokens,
                    "outputTokens": result.output_tokens,
                    "cacheCreationTokens": getattr(result, "cache_creation_tokens", 0),
                    "cacheReadTokens": getattr(result, "cache_read_tokens", 0),
                },
            }
        )

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )

@app.delete("/api/admin/reset")
def reset_all_state(request: Request) -> dict:
    """Wipe every SQLite job row — no recovery.

    Only reachable from the loopback interface so a misconfigured reverse
    proxy cannot invoke this remotely. GUI callers show a Win95Dialog
    confirmation before hitting this endpoint; see
    `WorkspaceMenu.tsx` for the user-facing flow.
    """
    client_host = request.client.host if request.client else None
    # "testclient" is the starlette TestClient default host — allow it so
    # pytest can exercise this endpoint without stubbing the socket.
    if client_host not in {"127.0.0.1", "::1", "localhost", "testclient"}:
        raise HTTPException(status_code=403, detail="reset only allowed from localhost")

    jobs_removed = JOB_REGISTRY.clear_all()
    JOB_REGISTRY.vacuum()

    return {
        "status": "ok",
        "jobsRemoved": jobs_removed,
    }


class TelemetryEvent(BaseModel):
    name: str
    props: dict = Field(default_factory=dict)
    experiments: dict = Field(default_factory=dict)
    ts: int | None = None


class TelemetryBatchRequest(BaseModel):
    events: list[TelemetryEvent]


@app.post("/api/events")
def collect_events(
    payload: TelemetryBatchRequest, request: Request
) -> dict:
    """Append client telemetry events to the JSONL log.

    - 每個 event 的 name 必須在 ALLOWED_TELEMETRY_EVENTS 內，否則整批拒絕
    - props 序列化後不能超過 4 KB（防 abuse）
    - 一次最多 MAX_TELEMETRY_BATCH 筆
    """
    events = payload.events or []
    if not events:
        raise HTTPException(status_code=400, detail="no events")
    if len(events) > MAX_TELEMETRY_BATCH:
        raise HTTPException(
            status_code=400,
            detail=f"batch too large (max {MAX_TELEMETRY_BATCH})",
        )

    workspace_id = _workspace_id(request)
    now_ms = int(datetime.utcnow().timestamp() * 1000)
    serialized: list[str] = []
    for event in events:
        if event.name not in ALLOWED_TELEMETRY_EVENTS:
            raise HTTPException(
                status_code=400, detail=f"unknown event: {event.name}"
            )
        ts = event.ts if event.ts and event.ts > 0 else now_ms
        record = {
            "name": event.name,
            "props": event.props,
            "experiments": event.experiments,
            "workspaceId": workspace_id,
            "ts": ts,
            "received_at": now_ms,
        }
        line = json.dumps(record, ensure_ascii=False, separators=(",", ":"))
        if len(line.encode("utf-8")) > MAX_TELEMETRY_PROPS_BYTES + 256:
            raise HTTPException(
                status_code=413,
                detail=f"event payload too large for {event.name}",
            )
        serialized.append(line)

    log_path = telemetry_log_path()
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with _TELEMETRY_LOCK:
        with open(log_path, "a", encoding="utf-8") as fh:
            for line in serialized:
                fh.write(line)
                fh.write("\n")

    return {"ok": True, "count": len(serialized)}


def _empty_event_bucket() -> dict:
    return {
        "eventCount": 0,
        "exposures": 0,
        "newRunClicks": 0,
        "runStarts": 0,
        "runSuccesses": 0,
        "runFailures": 0,
        "templateUses": 0,
        "templateSaves": 0,
        "retryClicks": 0,
        "comparesOpened": 0,
        "builderStepNexts": 0,
        "validationFails": 0,
        "completionRate": 0.0,
        "failureRate": 0.0,
        "templateReuseRate": 0.0,
        "rerunConversionRate": 0.0,
        "validationErrorRate": 0.0,
        "compareEngagementRate": 0.0,
    }


def _increment_event_bucket(bucket: dict, event_name: str) -> None:
    bucket["eventCount"] += 1
    if event_name == "experiment_exposure":
        bucket["exposures"] += 1
    elif event_name == "home_new_run_click":
        bucket["newRunClicks"] += 1
    elif event_name == "run_execute_start":
        bucket["runStarts"] += 1
    elif event_name == "run_execute_success":
        bucket["runSuccesses"] += 1
    elif event_name == "run_execute_fail":
        bucket["runFailures"] += 1
    elif event_name == "template_use_click":
        bucket["templateUses"] += 1
    elif event_name == "template_save":
        bucket["templateSaves"] += 1
    elif event_name == "run_retry_click":
        bucket["retryClicks"] += 1
    elif event_name == "output_compare_open":
        bucket["comparesOpened"] += 1
    elif event_name == "builder_step_next":
        bucket["builderStepNexts"] += 1
    elif event_name == "builder_validation_fail":
        bucket["validationFails"] += 1


def _finalize_event_bucket(bucket: dict) -> dict:
    terminal = bucket["runSuccesses"] + bucket["runFailures"]
    bucket["completionRate"] = (
        bucket["runSuccesses"] / terminal if terminal > 0 else 0.0
    )
    bucket["failureRate"] = (
        bucket["runFailures"] / terminal if terminal > 0 else 0.0
    )
    bucket["templateReuseRate"] = (
        bucket["templateUses"] / bucket["runStarts"]
        if bucket["runStarts"] > 0
        else 0.0
    )
    bucket["rerunConversionRate"] = (
        bucket["retryClicks"] / bucket["runFailures"]
        if bucket["runFailures"] > 0
        else 0.0
    )
    bucket["validationErrorRate"] = (
        bucket["validationFails"] / bucket["builderStepNexts"]
        if bucket["builderStepNexts"] > 0
        else 0.0
    )
    bucket["compareEngagementRate"] = (
        bucket["comparesOpened"] / bucket["runSuccesses"]
        if bucket["runSuccesses"] > 0
        else 0.0
    )
    return bucket


def _event_variant(record: dict, experiment: str | None) -> str | None:
    if not experiment:
        return None
    experiments = record.get("experiments")
    if isinstance(experiments, dict):
        variant = experiments.get(experiment)
        if isinstance(variant, str) and variant:
            return variant
    props = record.get("props")
    if (
        record.get("name") == "experiment_exposure"
        and isinstance(props, dict)
        and props.get("experiment") == experiment
        and isinstance(props.get("variant"), str)
    ):
        return props["variant"]
    return None


@app.get("/api/events/aggregate")
def events_aggregate(
    request: Request, experiment: str | None = None
) -> dict:
    """Read the append-only telemetry JSONL and aggregate experiment metrics.

    Tolerant of malformed lines (counted under ``malformedLines``).

    When the X-Workspace-Id header is present, only events tagged with that
    workspace are counted. Header absent = aggregate across all workspaces
    (useful for ops dashboards).
    """
    log_path = telemetry_log_path()
    variants: dict[str, dict] = {}
    unknown_variant = _empty_event_bucket()
    total_events = 0
    malformed_lines = 0
    workspace_filter = (
        request.headers.get(WORKSPACE_HEADER) if request else None
    )
    workspace_filter = (workspace_filter or "").strip() or None

    if not log_path.exists():
        return {
            "experiment": experiment,
            "workspaceId": workspace_filter,
            "totalEvents": 0,
            "malformedLines": 0,
            "variants": {},
            "unknownVariant": unknown_variant,
        }

    with _TELEMETRY_LOCK:
        lines = log_path.read_text("utf-8").splitlines()

    for line in lines:
        if not line.strip():
            continue
        try:
            record = json.loads(line)
        except json.JSONDecodeError:
            malformed_lines += 1
            continue
        if not isinstance(record, dict):
            malformed_lines += 1
            continue
        event_name = record.get("name")
        if not isinstance(event_name, str):
            total_events += 1
            continue
        # Workspace filter (Phase C-S2). Older rows without workspaceId map
        # to the default bucket so legacy data isn't silently dropped.
        if workspace_filter is not None:
            event_ws = record.get("workspaceId") or DEFAULT_WORKSPACE_ID
            if event_ws != workspace_filter:
                continue

        total_events += 1
        variant = _event_variant(record, experiment)
        if experiment:
            bucket = (
                variants.setdefault(variant, _empty_event_bucket())
                if variant
                else unknown_variant
            )
        else:
            bucket = variants.setdefault("all", _empty_event_bucket())
        _increment_event_bucket(bucket, event_name)

    finalized_variants = {
        variant: _finalize_event_bucket(bucket)
        for variant, bucket in sorted(variants.items())
    }
    return {
        "experiment": experiment,
        "workspaceId": workspace_filter,
        "totalEvents": total_events,
        "malformedLines": malformed_lines,
        "variants": finalized_variants,
        "unknownVariant": _finalize_event_bucket(unknown_variant),
    }


class ReviewFixSuggestRequest(BaseModel):
    tc: dict
    errors: list[dict] = Field(default_factory=list)
    model: str | None = None


@app.post("/api/review/suggest-fix")
def review_suggest_fix(payload: ReviewFixSuggestRequest) -> dict:
    """Single-shot AI fix suggestion for a TC that failed validation.

    Returns the explanation + a short imperative reason string the reviewer
    can paste into the Regenerate Reason input. Replaces the removed
    generic agent co-pilot with a narrow, no-session endpoint.
    """
    if not payload.errors:
        raise HTTPException(
            status_code=400,
            detail="suggest-fix requires at least one validation error",
        )
    try:
        result = suggest_review_fix(
            tc=payload.tc,
            errors=payload.errors,
            rules_text=RULES_SECTIONS,
            model=payload.model or DEFAULT_MODEL,
        )
    except GenerationError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    return {
        "problemRootCause": result["problem_root_cause"],
        "affectedFields": result["affected_fields"],
        "proposedChange": result["proposed_change"],
        "suggestedReason": result["suggested_reason"],
        "model": result["model"],
        "cost": result["cost"],
        "usage": result["usage"],
    }


class OutputCompareRequest(BaseModel):
    a: str
    b: str


_COMPARE_COLUMN_LABELS = {
    "tc_id": "TC ID",
    "test_group": "Test Group",
    "test_set": "Test Set",
    "pre_conditions": "Pre-Conditions",
    "input_test_data": "Input Test Data",
    "test_procedure": "Test Procedure",
    "expected_result": "Expected Result",
    "spec_reference": "Spec Reference",
    "priority": "Priority",
    "design_method": "Design Method",
}


def _read_tc_sheet_rows(xlsx_path: str) -> list[dict]:
    """Read every TC row (key = tc_id) from a generated workbook."""
    from openpyxl import load_workbook
    from writer import TC_SHEET_NAME, WRITE_COLUMNS

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if TC_SHEET_NAME not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[TC_SHEET_NAME]
    rows: list[dict] = []
    # 第 10 列以後是 data；header 在 row 9
    for raw in ws.iter_rows(min_row=10, values_only=True):
        if raw is None:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in raw):
            continue
        record: dict = {"reqId": raw[3] if len(raw) > 3 else None}
        for field, col in WRITE_COLUMNS.items():
            if field == "reasoning":
                continue
            idx = col - 1
            value = raw[idx] if idx < len(raw) else None
            record[field] = "" if value is None else str(value)
        rows.append(record)
    wb.close()
    return rows


def _diff_output_rows(rows_a: list[dict], rows_b: list[dict]) -> dict:
    """Diff two sets of rows keyed by tc_id."""
    by_a = {r["tc_id"]: r for r in rows_a if r.get("tc_id")}
    by_b = {r["tc_id"]: r for r in rows_b if r.get("tc_id")}
    keys = sorted(set(by_a) | set(by_b))
    diff_rows = []
    added = removed = changed = unchanged = 0
    for key in keys:
        a = by_a.get(key)
        b = by_b.get(key)
        if a and not b:
            removed += 1
            diff_rows.append(
                {
                    "tcId": key,
                    "status": "removed",
                    "reqId": a.get("reqId"),
                    "changes": [],
                }
            )
            continue
        if b and not a:
            added += 1
            diff_rows.append(
                {
                    "tcId": key,
                    "status": "added",
                    "reqId": b.get("reqId"),
                    "changes": [],
                }
            )
            continue
        changes = []
        for field, label in _COMPARE_COLUMN_LABELS.items():
            if field == "tc_id":
                continue
            va = (a or {}).get(field, "") or ""
            vb = (b or {}).get(field, "") or ""
            if va != vb:
                changes.append(
                    {"field": field, "label": label, "before": va, "after": vb}
                )
        if changes:
            changed += 1
            diff_rows.append(
                {
                    "tcId": key,
                    "status": "changed",
                    "reqId": (a or b or {}).get("reqId"),
                    "changes": changes,
                }
            )
        else:
            unchanged += 1
            diff_rows.append(
                {
                    "tcId": key,
                    "status": "unchanged",
                    "reqId": (a or {}).get("reqId"),
                    "changes": [],
                }
            )

    return {
        "summary": {
            "total": len(keys),
            "added": added,
            "removed": removed,
            "changed": changed,
            "unchanged": unchanged,
        },
        "rows": diff_rows,
    }


class BulkDownloadRequest(BaseModel):
    jobIds: list[str]


@app.post("/api/outputs/bulk-download")
def bulk_download_outputs(payload: BulkDownloadRequest) -> StreamingResponse:
    """Stream a zip archive containing each requested job's exported workbook."""
    import io
    import zipfile

    if not payload.jobIds:
        raise HTTPException(status_code=400, detail="jobIds required")
    if len(payload.jobIds) > 50:
        raise HTTPException(
            status_code=400, detail="too many jobs (max 50)"
        )

    seen: set[str] = set()
    paths: list[tuple[str, str]] = []
    for jid in payload.jobIds:
        if not isinstance(jid, str) or not jid or jid in seen:
            continue
        seen.add(jid)
        job = JOB_REGISTRY.get(jid)
        if not job:
            raise HTTPException(
                status_code=404, detail=f"job not found: {jid}"
            )
        export_path = job.get("exportPath")
        if not export_path or not os.path.isfile(export_path):
            raise HTTPException(
                status_code=409,
                detail=f"job {jid} has no exported workbook",
            )
        paths.append((jid, export_path))

    if not paths:
        raise HTTPException(status_code=400, detail="no usable jobs")

    used_names: set[str] = set()
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for _job_id, path in paths:
            base = os.path.basename(path)
            arcname = base
            stem, ext = os.path.splitext(arcname)
            counter = 1
            while arcname in used_names:
                counter += 1
                arcname = f"{stem}-{counter}{ext}"
            used_names.add(arcname)
            zf.write(path, arcname)
    buffer.seek(0)

    filename = (
        f"tc-outputs-bundle-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.zip"
    )
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/outputs/compare")
def compare_outputs(payload: OutputCompareRequest) -> dict:
    """Diff two exported workbooks row-by-row keyed by TC ID."""
    if payload.a == payload.b:
        raise HTTPException(
            status_code=400, detail="cannot compare a job with itself"
        )

    job_a = JOB_REGISTRY.get(payload.a)
    job_b = JOB_REGISTRY.get(payload.b)
    if not job_a:
        raise HTTPException(status_code=404, detail=f"job not found: {payload.a}")
    if not job_b:
        raise HTTPException(status_code=404, detail=f"job not found: {payload.b}")

    path_a = job_a.get("exportPath")
    path_b = job_b.get("exportPath")
    if not path_a or not os.path.isfile(path_a):
        raise HTTPException(
            status_code=409,
            detail=f"job {payload.a} has no exported workbook (run Export first)",
        )
    if not path_b or not os.path.isfile(path_b):
        raise HTTPException(
            status_code=409,
            detail=f"job {payload.b} has no exported workbook (run Export first)",
        )

    try:
        rows_a = _read_tc_sheet_rows(path_a)
        rows_b = _read_tc_sheet_rows(path_b)
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"failed to read workbooks: {exc}"
        ) from exc

    diff = _diff_output_rows(rows_a, rows_b)
    return {"a": payload.a, "b": payload.b, **diff}


@app.get("/api/jobs/{job_id}/timeline")
def get_job_timeline(job_id: str) -> dict:
    """Return ordered timeline events for a job (queued / running / completed / etc.)."""
    job = JOB_REGISTRY.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="job not found")
    timeline = job.get("timeline")
    if not isinstance(timeline, list):
        timeline = []
    return {"jobId": job_id, "events": timeline}


class ValidationLogEntry(BaseModel):
    rowId: str
    reqId: str | None = None
    severity: str
    field: str | None = None
    message: str
    ts: int | None = None


class ValidationLogBatch(BaseModel):
    entries: list[ValidationLogEntry]


@app.post("/api/jobs/{job_id}/validation-logs")
def append_validation_logs(job_id: str, payload: ValidationLogBatch) -> dict:
    """Frontend posts row-level validation issues for a finished run.

    Rows are aggregated (latest write replaces older ones for the same rowId),
    so re-posting after a Regenerate session keeps the latest state without
    unbounded growth.
    """
    job = JOB_REGISTRY.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")

    now_ms = int(datetime.utcnow().timestamp() * 1000)
    existing = job.get("validationLog")
    if not isinstance(existing, list):
        existing = []
    by_row: dict[str, dict] = {}
    for prior in existing:
        if isinstance(prior, dict) and prior.get("rowId"):
            by_row[str(prior["rowId"])] = prior
    for entry in payload.entries:
        record = {
            "rowId": entry.rowId,
            "reqId": entry.reqId,
            "severity": entry.severity,
            "field": entry.field,
            "message": entry.message,
            "ts": entry.ts if entry.ts and entry.ts > 0 else now_ms,
        }
        by_row[entry.rowId] = record
    merged = list(by_row.values())
    if len(merged) > VALIDATION_LOG_MAX:
        merged = merged[-VALIDATION_LOG_MAX:]
    job["validationLog"] = merged
    JOB_REGISTRY[job_id] = job
    return {"ok": True, "count": len(merged)}


@app.get("/api/jobs/{job_id}/validation-logs")
def get_validation_logs(job_id: str) -> dict:
    job = JOB_REGISTRY.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")
    log = job.get("validationLog")
    if not isinstance(log, list):
        log = []
    return {"jobId": job_id, "entries": log}


@app.get("/api/jobs/{job_id}/config")
def get_job_config(job_id: str) -> dict:
    """Return the resolved generation config snapshot recorded for this job.

    Run Detail uses this to show "what this run was actually configured with"
    without requiring the frontend to keep its own copy of the config.
    """
    job = JOB_REGISTRY.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")
    config = job.get("config")
    parsed = job.get("parsedData") or {}
    return {
        "jobId": job_id,
        "config": config if isinstance(config, dict) else None,
        "projectName": parsed.get("project") or job.get("projectName"),
        "testGroup": parsed.get("test_group") or job.get("testGroup"),
        "totalRows": job.get("totalRows"),
        "status": job.get("status"),
    }


@app.get("/api/jobs/{job_id}/usage")
def get_job_usage(job_id: str) -> dict:
    """Per-job usage breakdown including model-level cost attribution.

    Fast read endpoint for CostMeter dashboard popup — lets the UI show
    "classify: $0.02 (gpt-5-mini) / generate: $0.13 (gpt-5)" without
    dragging the full job record over the wire.
    """
    job = JOB_REGISTRY.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")
    usage = _job_usage(job)
    return {"jobId": job_id, **usage}


@app.get("/api/metrics/aggregate")
def metrics_aggregate(job_ids: str | None = None) -> dict:
    """跨 job 聚合指標（成本觀測面板用）。

    Query:
        job_ids: 逗號分隔 ID 子集；不帶參數 → 全部 job。
    """
    selected: list[str] | None = None
    if job_ids:
        selected = [jid.strip() for jid in job_ids.split(",") if jid.strip()]
    try:
        return aggregate_metrics_tool(job_ids=selected, job_store=JOB_REGISTRY)
    except ToolError as exc:
        raise _tool_error_to_http(exc) from exc


def _spec_manifest_path() -> Path:
    """Spec library manifest path; override via TC_SPEC_INDEX_MANIFEST."""
    override = os.environ.get("TC_SPEC_INDEX_MANIFEST")
    if override:
        return Path(override)
    return (
        Path(__file__).resolve().parent.parent / "spec-index" / "manifest.json"
    )


def _read_spec_manifest() -> dict:
    path = _spec_manifest_path()
    if not path.is_file():
        return {"specs": []}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"specs": []}
    if not isinstance(data, dict):
        return {"specs": []}
    return data


def _write_spec_manifest(data: dict) -> None:
    path = _spec_manifest_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _serialize_spec_entry(item: dict) -> dict:
    return {
        "name": item.get("name"),
        "sourceFile": item.get("source_file"),
        "entriesCount": item.get("entries_count"),
        "embeddingModel": item.get("embedding_model"),
        "updatedAt": item.get("updated_at"),
        "version": item.get("version"),
        "changelog": item.get("changelog") or [],
        "deprecated": bool(item.get("deprecated")),
    }


@app.get("/api/spec-library")
async def list_spec_library() -> dict:
    """回傳 ``spec-index/manifest.json`` 中已建好的 spec 清單，前端用來做 dropdown 選單。"""
    data = _read_spec_manifest()
    raw_specs = data.get("specs") or []
    specs = [
        _serialize_spec_entry(item)
        for item in raw_specs
        if isinstance(item, dict) and item.get("name")
    ]
    specs.sort(key=lambda s: (s.get("name") or "").lower())
    return {"specs": specs}


class ChangelogEntryRequest(BaseModel):
    version: str | None = None
    message: str
    author: str | None = None


@app.post("/api/spec-library/{name}/changelog")
def append_spec_changelog(
    name: str, payload: ChangelogEntryRequest, request: Request
) -> dict:
    """Append a changelog entry to a template; bump `version` if provided.

    Loopback-only — same posture as `/api/admin/reset` since it mutates the
    on-disk manifest. Tests run via Starlette's TestClient (host=testclient).
    """
    client_host = request.client.host if request.client else None
    if client_host not in {"127.0.0.1", "::1", "localhost", "testclient"}:
        raise HTTPException(
            status_code=403,
            detail="changelog edits only allowed from localhost",
        )

    if not payload.message.strip():
        raise HTTPException(status_code=400, detail="message required")

    manifest = _read_spec_manifest()
    specs = manifest.get("specs")
    if not isinstance(specs, list):
        raise HTTPException(
            status_code=404, detail=f"template not found: {name}"
        )
    target: dict | None = None
    for entry in specs:
        if isinstance(entry, dict) and entry.get("name") == name:
            target = entry
            break
    if target is None:
        raise HTTPException(
            status_code=404, detail=f"template not found: {name}"
        )

    changelog = target.get("changelog")
    if not isinstance(changelog, list):
        changelog = []
    new_entry = {
        "version": payload.version or target.get("version") or "1.0.0",
        "message": payload.message.strip(),
        "ts": int(datetime.utcnow().timestamp() * 1000),
    }
    if payload.author:
        new_entry["author"] = payload.author
    changelog.append(new_entry)
    target["changelog"] = changelog
    if payload.version:
        target["version"] = payload.version

    _write_spec_manifest(manifest)
    return {"ok": True, "entry": new_entry, "spec": _serialize_spec_entry(target)}


class TemplateDeprecateRequest(BaseModel):
    deprecated: bool


@app.patch("/api/spec-library/{name}")
def patch_spec_library_entry(
    name: str, payload: TemplateDeprecateRequest, request: Request
) -> dict:
    """Toggle the ``deprecated`` flag for a template (loopback only).

    Currently scoped to the deprecation flag. Other metadata (display name,
    embedding model overrides) belong in their own ingestion pipeline tools.
    """
    client_host = request.client.host if request.client else None
    if client_host not in {"127.0.0.1", "::1", "localhost", "testclient"}:
        raise HTTPException(
            status_code=403,
            detail="manifest edits only allowed from localhost",
        )

    manifest = _read_spec_manifest()
    specs = manifest.get("specs")
    if not isinstance(specs, list):
        raise HTTPException(
            status_code=404, detail=f"template not found: {name}"
        )
    target: dict | None = None
    for entry in specs:
        if isinstance(entry, dict) and entry.get("name") == name:
            target = entry
            break
    if target is None:
        raise HTTPException(
            status_code=404, detail=f"template not found: {name}"
        )

    target["deprecated"] = bool(payload.deprecated)
    _write_spec_manifest(manifest)
    return {"ok": True, "spec": _serialize_spec_entry(target)}


@app.get("/api/jobs/{job_id}/output-preview")
def get_job_output_preview(job_id: str, limit: int = 200) -> dict:
    """Read the exported workbook so the frontend can preview rows in-page.

    Returns at most ``limit`` rows (cap 1000 to keep payload sane). 409 when
    the job has no exported workbook yet — same contract as compare.
    """
    job = JOB_REGISTRY.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")
    export_path = job.get("exportPath")
    if not export_path or not os.path.isfile(export_path):
        raise HTTPException(
            status_code=409,
            detail="no exported workbook (run Export first)",
        )
    cap = max(1, min(int(limit) if limit else 200, 1000))
    try:
        all_rows = _read_tc_sheet_rows(export_path)
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"failed to read workbook: {exc}"
        ) from exc
    return {
        "jobId": job_id,
        "fileName": os.path.basename(export_path),
        "totalRows": len(all_rows),
        "limit": cap,
        "rows": all_rows[:cap],
    }


@app.get("/api/jobs/{job_id}/dataset")
def get_job_dataset(job_id: str) -> dict:
    """Return the parsed dataset rows so the Builder can re-hydrate without
    re-uploading the workbook.

    The response uses camelCase keys aligned with the frontend `TcRow` shape so
    the client can populate `useJobStore.tcRows` directly.
    """
    job = JOB_REGISTRY.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")

    parsed = job.get("parsedData") or {}
    parsed_rows = parsed.get("rows") or []
    test_group = parsed.get("test_group") or job.get("testGroup") or ""

    # /api/generate input rows (camelCase). 用來補 generated content 預設值。
    input_rows = job.get("rows") or []
    inputs_by_id = {
        str(r.get("id")): r
        for r in input_rows
        if isinstance(r, dict) and r.get("id")
    }

    rows: list[dict] = []
    for parsed_row in parsed_rows:
        row_num = parsed_row.get("row_num")
        # parser 產出 row_id；GenerateRow 的 id 則是前端自訂 — 兩者擇一。
        row_id = (
            parsed_row.get("id")
            or parsed_row.get("row_id")
            or f"row-{row_num}"
        )
        # 若有對應的 GenerateRow 輸入，順便還原它的 id 與 testSet
        gen_row = inputs_by_id.get(str(row_id))
        rows.append(
            {
                "id": str(row_id),
                "rowNum": row_num,
                "reqId": parsed_row.get("req_id") or "",
                "testItem": parsed_row.get("test_item") or "",
                "testSet": (
                    (gen_row or {}).get("testSet")
                    or parsed_row.get("test_set")
                    or ""
                ),
                "testGroup": test_group,
                "specReference": parsed_row.get("spec_reference"),
                "preConditions": "",
                "inputTestData": "",
                "steps": "",
                "expectedResults": "",
                "status": "pending",
            }
        )

    return {
        "jobId": job_id,
        "rowCount": len(rows),
        "projectName": parsed.get("project"),
        "testGroup": test_group,
        "rows": rows,
    }


@app.get("/api/spec-library/{name}/usage")
def spec_library_usage(name: str) -> dict:
    """Count completed runs that used the named template.

    Scans the SQLite job registry for jobs whose stored ``templateId`` matches.
    Only the latest 10 run IDs are returned to keep the payload light.
    """
    matches: list[dict] = []
    for job_id in JOB_REGISTRY.keys():
        job = JOB_REGISTRY.get(job_id)
        if not job or job.get("templateId") != name:
            continue
        timeline = job.get("timeline") if isinstance(job.get("timeline"), list) else []
        last_ts = max((e.get("ts") or 0 for e in timeline), default=0)
        matches.append({"jobId": job_id, "lastTs": last_ts, "status": job.get("status")})

    matches.sort(key=lambda m: m["lastTs"] or 0, reverse=True)
    recent = matches[:10]
    last_used_at = recent[0]["lastTs"] if recent else None
    return {
        "name": name,
        "usageCount": len(matches),
        "lastUsedAt": last_used_at if last_used_at else None,
        "recentRunIds": [m["jobId"] for m in recent],
    }


@app.post("/api/parse")
async def parse_workbook(
    request: Request,
    raw_file: UploadFile = File(...),
    reference_file: UploadFile | None = File(default=None),
    sys1_file: UploadFile | None = File(default=None),
    spec_file: UploadFile | None = File(default=None),
    selected_spec_name: str | None = Form(default=None),
) -> dict:
    # 先做 HTTP 層檢查：size limit、extension。extension 也會被 tool 再驗一次。
    raw_bytes = await _read_with_limit(raw_file, "raw_file")

    effective_reference_file = reference_file or sys1_file
    reference_bytes = (
        await _read_with_limit(effective_reference_file, "reference_file")
        if effective_reference_file
        else None
    )
    spec_bytes = await _read_with_limit(spec_file, "spec_file") if spec_file else None

    raw_ext = os.path.splitext(raw_file.filename or "")[1].lower() or ".xlsx"
    raw_filename = _safe_upload_filename(raw_file.filename, f"upload{raw_ext}")
    reference_filename = (
        _safe_upload_filename(effective_reference_file.filename, "reference.xlsx")
        if effective_reference_file else None
    )
    spec_filename = (
        _safe_upload_filename(spec_file.filename, "spec.bin")
        if spec_file else None
    )

    with tempfile.TemporaryDirectory() as tmp_dir:
        raw_path = os.path.join(tmp_dir, raw_filename)
        with open(raw_path, "wb") as handle:
            handle.write(raw_bytes)

        reference_path: str | None = None
        if effective_reference_file and reference_bytes is not None:
            reference_path = os.path.join(tmp_dir, reference_filename or "reference.xlsx")
            with open(reference_path, "wb") as handle:
                handle.write(reference_bytes)

        spec_path: str | None = None
        if spec_file and spec_bytes is not None:
            spec_path = os.path.join(tmp_dir, spec_filename or "spec.bin")
            with open(spec_path, "wb") as handle:
                handle.write(spec_bytes)

        try:
            result = parse_workbook_tool(
                raw_path=raw_path,
                raw_filename=raw_filename,
                reference_path=reference_path,
                reference_filename=reference_filename,
                spec_path=spec_path,
                spec_filename=spec_filename,
                selected_spec_name=selected_spec_name,
                job_store=JOB_REGISTRY,
            )
        except ToolError as exc:
            raise _tool_error_to_http(exc) from exc

        # Tag the freshly-created job with the caller's workspace so future
        # listings / analytics can scope by workspace (Phase C-S2).
        new_job_id = result.get("jobId") if isinstance(result, dict) else None
        if isinstance(new_job_id, str):
            stored = JOB_REGISTRY.get(new_job_id)
            if isinstance(stored, dict):
                stored["workspaceId"] = _workspace_id(request)
                JOB_REGISTRY[new_job_id] = stored
        return result


@app.post("/api/group")
async def preview_grouping(payload: GroupPreviewRequest) -> dict:
    job = JOB_REGISTRY.get(payload.jobId)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")

    parsed_rows = job.get("parsedData", {}).get("rows", [])
    parsed_by_row_num = {row.get("row_num"): row for row in parsed_rows}
    parsed_by_req_id = {row.get("req_id"): row for row in parsed_rows}
    rows_for_grouping: list[dict] = []
    for row in payload.rows:
        data = row.model_dump()
        base_row = None
        if data.get("rowNum") is not None:
            base_row = parsed_by_row_num.get(data["rowNum"])
        if base_row is None and data.get("reqId"):
            base_row = parsed_by_req_id.get(data["reqId"])
        original_test_set = str((base_row or {}).get("test_set") or "").strip()
        if original_test_set and not str(data.get("testSet") or "").strip():
            data["testSetHint"] = original_test_set
        rows_for_grouping.append(data)

    result = group_tests_tool(
        rows=rows_for_grouping,
        force_regroup=payload.forceRegroup,
    )
    usage = _job_usage(job)
    this_cost = float(result.get("cost") or 0.0)
    _persist_job_usage(
        payload.jobId,
        job,
        cost=usage["cost"] + this_cost,
        input_tokens=usage["inputTokens"] + int(result.get("inputTokens") or 0),
        output_tokens=usage["outputTokens"] + int(result.get("outputTokens") or 0),
        cache_creation_tokens=usage["cacheCreationTokens"] + int(result.get("cacheCreationTokens") or 0),
        cache_read_tokens=usage["cacheReadTokens"] + int(result.get("cacheReadTokens") or 0),
        cost_delta=this_cost,
        model=str(result.get("model") or ""),
    )
    return {"jobId": payload.jobId, **result}


@app.post("/api/match")
async def preview_spec_matching(payload: MatchPreviewRequest) -> dict:
    job = JOB_REGISTRY.get(payload.jobId)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")

    prepared_rows = _prepare_match_preview_rows(payload.rows, job)
    selected_spec_name = (job.get("selectedSpecName") or "").strip() or None
    reference_bytes = job.get("referenceWorkbookBytes")
    reference_filename = job.get("referenceWorkbookName")

    try:
        if selected_spec_name:
            result = match_spec_tool(
                rows=prepared_rows,
                reference_workbook_path=None,
                selected_spec_name=selected_spec_name,
            )
        elif reference_bytes and reference_filename:
            with tempfile.TemporaryDirectory() as tmp_dir:
                reference_path = os.path.join(
                    tmp_dir,
                    _safe_upload_filename(reference_filename, "reference.xlsx"),
                )
                with open(reference_path, "wb") as handle:
                    handle.write(reference_bytes)
                result = match_spec_tool(rows=prepared_rows, reference_workbook_path=reference_path)
        else:
            result = match_spec_tool(rows=prepared_rows, reference_workbook_path=None)
    except ToolError as exc:
        raise _tool_error_to_http(exc) from exc

    return {"jobId": payload.jobId, **result}


@app.post("/api/generate")
async def create_generate_job(payload: GenerateRequest, request: Request) -> dict:
    if not payload.rows:
        raise HTTPException(status_code=400, detail="rows must not be empty")

    job_id = payload.jobId or _build_generate_job_id()
    total_rows = len(payload.rows)
    stream_url = str(request.url_for("stream_generate_job")) + f"?jobId={job_id}"

    job_record = JOB_REGISTRY.get(job_id, {"jobId": job_id})
    update_fields = {
        "status": "queued",
        "rows": [row.model_dump() for row in payload.rows],
        "config": payload.config.model_dump(),
        "totalRows": total_rows,
        # Phase C-S2: tag with workspace if not already set; existing tag
        # (e.g. from /api/parse) wins so multiple stages stay consistent.
        "workspaceId": job_record.get("workspaceId") or _workspace_id(request),
    }
    if payload.templateId:
        update_fields["templateId"] = payload.templateId
    job_record.update(update_fields)
    JOB_REGISTRY[job_id] = job_record
    _append_job_event(job_id, "queued", rowCount=total_rows)

    return {
        "jobId": job_id,
        "status": "queued",
        "totalRows": total_rows,
        "streamUrl": stream_url,
    }


@app.get("/api/generate/stream", name="stream_generate_job")
async def stream_generate_job(jobId: str) -> StreamingResponse:
    job = JOB_REGISTRY.get(jobId)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")

    async def event_generator():
        all_rows = _prepare_generation_rows(job)
        total = len(all_rows)
        config = job.get("config", {})

        # 每一列都走 AI 拆分。Template 已填在 L/M 的 criteria / 預期拆解由
        # `_format_reviewer_hints` 當 prompt hints 傳給 AI，而不是跳過 AI。
        # （`regenerateAll` flag 仍保留接口但不再影響流程。）
        preserved_rows: list[dict] = []
        rows = all_rows

        context = {
            "project": job.get("parsedData", {}).get("project"),
            "test_group": job.get("parsedData", {}).get("test_group"),
            "test_set": "N/A",
        }
        spec_index = _build_spec_index_for_job(job)
        job["status"] = "running"
        JOB_REGISTRY[jobId] = job  # 回寫 SQLite，避免修改只停留在反序列化副本
        _append_job_event(jobId, "running", rowCount=total)

        # (A) 依最終顯示順序配 tc_id：主 TC 後面立刻接子 TC，再輪到下一個 req。
        # 不能只幫 sub-TC 補尾號，否則畫面/Excel 會變成 001, 006, 002, 007...
        # 這裡改成 stream 時依實際輸出順序連號。
        existing_seqs = _extract_existing_sequence_numbers(all_rows)
        next_seq = min(existing_seqs) if existing_seqs else 1
        tc_project = context["project"] or ""
        tc_group_abbr = (
            generate_group_abbreviation(context["test_group"])
            if context["test_group"] else ""
        )

        def _alloc_stream_tc_id() -> str:
            nonlocal next_seq
            if not tc_project or not tc_group_abbr:
                return ""
            new_id = generate_tc_ids(tc_project, tc_group_abbr, 1, start=next_seq)[0]
            next_seq += 1
            return new_id

        processed = 0
        usage = _job_usage(job)
        current_cost = usage["cost"]
        total_input_tokens = usage["inputTokens"]
        total_output_tokens = usage["outputTokens"]
        total_cache_creation_tokens = usage["cacheCreationTokens"]
        total_cache_read_tokens = usage["cacheReadTokens"]
        persisted_cost = usage["cost"]
        batch_size = config.get("batchSize", 1)
        model = config.get("model", DEFAULT_MODEL)
        strict_validation = config.get("strictValidation", False)
        budget = config.get("budget", 0)

        yield _sse_event(
            {
                "type": "job.started",
                "jobId": jobId,
                "stats": {
                    "total": total,
                    "processed": 0,
                    "currentCost": round(current_cost, 4),
                    "inputTokens": total_input_tokens,
                    "outputTokens": total_output_tokens,
                    "cacheCreationTokens": total_cache_creation_tokens,
                    "cacheReadTokens": total_cache_read_tokens,
                },
                "message": (
                    f"Backend generation started for {total} row(s) "
                    f"({len(rows)} to generate, {len(preserved_rows)} preserved)."
                ),
            }
        )

        # Emit preserved rows up front — zero cost, zero AI call.
        for preserved in preserved_rows:
            processed += 1
            yield _sse_event(
                {
                    "type": "row.completed",
                    "jobId": jobId,
                    "row": _build_preserved_stream_row(preserved),
                    "stats": {
                        "total": total,
                        "processed": processed,
                        "currentCost": round(current_cost, 4),
                    },
                    "message": f"Preserved existing content for row {preserved.get('row_num')}.",
                }
            )

        to_generate_count = len(rows)
        for i in range(0, to_generate_count, batch_size):
            batch = rows[i:i + batch_size]

            if budget and _would_exceed_budget(current_cost, len(batch), model, budget):
                for row in batch:
                    processed += 1
                    skip_msg = (
                        f"Skipped because the next batch would exceed the configured "
                        f"budget of ${budget:.2f}."
                    )
                    yield _sse_event(
                        {
                            "type": "row.failed",
                            "jobId": jobId,
                            "row": _build_failed_stream_row(row, skip_msg),
                            "stats": {
                                "total": total,
                                "processed": processed,
                                "currentCost": round(current_cost, 4),
                            },
                            "message": f"Skipped {row.get('req_id')} due to budget limit.",
                        }
                    )
                    _record_stream_validation_failure(
                        jobId, row, skip_msg, field="budget"
                    )
                continue

            try:
                batch_result = generate_tc_tool(
                    rows=batch,
                    context=context,
                    spec_index=spec_index,
                    rules_text=RULES_SECTIONS,
                    model=model,
                )
                tc_data_list = batch_result["tcData"]

                current_cost += batch_result["cost"]
                total_input_tokens += batch_result["inputTokens"]
                total_output_tokens += batch_result["outputTokens"]
                total_cache_creation_tokens += batch_result["cacheCreationTokens"]
                total_cache_read_tokens += batch_result["cacheReadTokens"]
                batch_delta = max(current_cost - persisted_cost, 0.0)
                if batch_delta > 0:
                    _persist_job_usage(
                        jobId,
                        job,
                        cost=current_cost,
                        input_tokens=total_input_tokens,
                        output_tokens=total_output_tokens,
                        cache_creation_tokens=total_cache_creation_tokens,
                        cache_read_tokens=total_cache_read_tokens,
                        cost_delta=batch_delta,
                        model=model,
                    )
                    persisted_cost = current_cost

                split_meta_list = batch_result.get("splitMeta") or []

                # tc_data_list 現在是 list[list[dict]]：每個 input row 對應 1..N 筆 TC。
                # tc_id 改成依實際輸出順序連號，而不是只讓子 TC 補尾號。
                for group_idx, (row, tc_group) in enumerate(zip(batch, tc_data_list)):
                    if not isinstance(tc_group, list) or not tc_group:
                        continue

                    # (B) 每個 req 只讓 processed +1（對應原始 rows 數），
                    # 當 AI 拆出 N 筆時把 total 往上加 N-1，這樣進度條與 stats 同步。
                    extras = max(len(tc_group) - 1, 0)
                    if extras > 0:
                        total += extras
                    processed += 1

                    # (F) 若 req 含多個 enumerated 項目但 AI 只回 1 筆，給個啟發式 warning
                    split_warning = _heuristic_split_warning(row, len(tc_group))

                    # 先送 req.split：告訴前端此 req 拆成幾筆 TC、AI 拆分 reasoning、keyword 分析
                    meta = split_meta_list[group_idx] if group_idx < len(split_meta_list) else {}
                    split_payload = {
                        "type": "req.split",
                        "jobId": jobId,
                        "rowId": row["id"],
                        "reqId": row.get("req_id") or meta.get("req_id") or "",
                        "tcCount": len(tc_group),
                        "reasoning": str(meta.get("reasoning") or ""),
                        "keywords": meta.get("keywords") or [],
                        "stats": {
                            "total": total,
                            "processed": processed,
                            "currentCost": round(current_cost, 4),
                            "inputTokens": total_input_tokens,
                            "outputTokens": total_output_tokens,
                            "cacheCreationTokens": total_cache_creation_tokens,
                            "cacheReadTokens": total_cache_read_tokens,
                        },
                        "message": (
                            f"{row.get('req_id') or row['id']}: "
                            f"AI split into {len(tc_group)} TC(s). "
                            f"{str(meta.get('reasoning') or '')[:200]}"
                        ).strip(),
                    }
                    if split_warning:
                        split_payload["splitWarning"] = split_warning
                    yield _sse_event(split_payload)

                    for sub_idx, tc in enumerate(tc_group):
                        is_primary = sub_idx == 0
                        sub_row = {
                            **row,
                            "tc_id": _alloc_stream_tc_id(),
                        }
                        if not is_primary:
                            sub_row["id"] = f"{row['id']}__tc{sub_idx + 1}"
                        updated_row, has_warnings = _build_stream_row(sub_row, tc)
                        if is_primary:
                            event_type = "row.failed" if strict_validation and has_warnings else "row.completed"
                        else:
                            # 子 TC：即使 strict mode 有 warning 也走 row.added，讓前端
                            # 保留整組、由 review UI 決定要不要接受。
                            event_type = "row.added"
                        if event_type == "row.failed":
                            updated_row["status"] = "error"
                            updated_row["generated"] = None
                        # (G) 把拆分資訊附到每筆 TC，sub 也能直接顯示「屬於 REQ-xxx 的 TC k/N」
                        updated_row["splitDecision"] = {
                            "reqId": row.get("req_id") or meta.get("req_id") or "",
                            "tcCount": len(tc_group),
                            "subIndex": sub_idx,
                            "parentId": row["id"],
                            # 完整 reasoning 僅放在 primary，sub 留空避免重複顯示
                            "reasoning": str(meta.get("reasoning") or "") if is_primary else "",
                            "keywords": (meta.get("keywords") or []) if is_primary else [],
                            # AI strict-marked equivalent sibling row id (Layer 1
                            # duplicate detection). Empty when not duplicate.
                            "duplicateOf": (
                                _resolve_duplicate_of(meta.get("duplicate_of"), row.get("siblings"))
                                if is_primary
                                else ""
                            ),
                            # AI 對 sibling 差異的明確聲明（B 方案），primary 才帶。
                            # axis: trigger_state / input_data / timing / boundary / mode / none；
                            # delta: 繁中差異描述含具體 token。沒 sibling 時 AI omit。
                            "distinguishingAxis": (
                                meta.get("distinguishing_axis") or {}
                                if is_primary
                                else {}
                            ),
                        }
                        if split_warning and is_primary:
                            updated_row["splitWarning"] = split_warning
                        if not is_primary:
                            updated_row["parentId"] = row["id"]
                            updated_row["subIndex"] = sub_idx
                        if event_type == "row.failed":
                            first_issue = next(
                                iter(updated_row.get("validation") or []), None
                            )
                            _record_stream_validation_failure(
                                jobId,
                                row,
                                (first_issue or {}).get("message")
                                or "Strict validation rejected this row.",
                                field=(first_issue or {}).get("field"),
                            )
                        yield _sse_event(
                            {
                                "type": event_type,
                                "jobId": jobId,
                                "row": updated_row,
                                "stats": {
                                    "total": total,
                                    "processed": processed,
                                    "currentCost": round(current_cost, 4),
                                    "inputTokens": total_input_tokens,
                                    "outputTokens": total_output_tokens,
                                    "cacheCreationTokens": total_cache_creation_tokens,
                                    "cacheReadTokens": total_cache_read_tokens,
                                },
                                "message": (
                                    f"Processed {processed}/{total} rows for "
                                    f"{row.get('req_id') or row.get('id')}"
                                    + (f" (TC {sub_idx + 1}/{len(tc_group)})" if len(tc_group) > 1 else "")
                                ),
                            }
                        )
            except ToolError as exc:
                if _is_quota_exceeded_tool_error(exc):
                    yield _sse_event(
                        {
                            "type": "job.failed",
                            "jobId": jobId,
                            "stats": {
                                "total": total,
                                "processed": processed,
                                "currentCost": round(current_cost, 4),
                            },
                            "message": exc.message,
                        }
                    )
                    return
                for row in batch:
                    processed += 1
                    fail_msg = f"Generation failed: {exc.message}"
                    yield _sse_event(
                        {
                            "type": "row.failed",
                            "jobId": jobId,
                            "row": _build_failed_stream_row(row, fail_msg),
                            "stats": {
                                "total": total,
                                "processed": processed,
                                "currentCost": round(current_cost, 4),
                            },
                            "message": f"Generation failed for {row.get('req_id') or row.get('id')}.",
                        }
                    )
                    _record_stream_validation_failure(
                        jobId, row, fail_msg, field="generation"
                    )

            await asyncio.sleep(0.15)

        job["status"] = "completed"
        # 把這次 stream 新增的成本歸進選定 model 的 bucket；classify 那邊的
        # gpt-5-mini 成本不會跑進這條 code path，所以 bucket 乾淨。
        stream_delta = max(current_cost - persisted_cost, 0.0)
        _persist_job_usage(
            jobId,
            job,
            cost=current_cost,
            input_tokens=total_input_tokens,
            output_tokens=total_output_tokens,
            cache_creation_tokens=total_cache_creation_tokens,
            cache_read_tokens=total_cache_read_tokens,
            cost_delta=stream_delta,
            model=model,
        )
        JOB_REGISTRY[jobId] = job  # 回寫 SQLite
        _append_job_event(
            jobId,
            "completed",
            rowCount=total,
            processed=processed,
            cost=round(current_cost, 4),
        )
        yield _sse_event(
            {
                "type": "job.completed",
                "jobId": jobId,
                "stats": {
                    "total": total,
                    "processed": total,
                    "currentCost": current_cost,
                    "inputTokens": total_input_tokens,
                    "outputTokens": total_output_tokens,
                    "cacheCreationTokens": total_cache_creation_tokens,
                    "cacheReadTokens": total_cache_read_tokens,
                },
                "message": "Backend generation complete. Review and export windows are ready.",
            }
        )

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.post("/api/jobs/{job_id}/regenerate/stream")
async def stream_regenerate(job_id: str, payload: RegenerateRequest) -> StreamingResponse:
    """Re-generate specific rows and stream results back as SSE.

    使用 reviewer 提供的 regenerate reason 重新產生選中 rows。AI 會先用完整
    pipeline 判斷是否需要拆分：primary 仍回 `row.regenerated`，新增 TC 回
    `row.added`，讓前端插在 parent row 後方。
    """
    # 和 rerun 一致：job 缺失時不 404，改從 payload 重建 context。
    # 舊行為是硬 404，使用者在 backend 重啟後會看到 regenerate 壞但 rerun 正常
    # 的 UX cliff。現在兩個路徑都能降級；UI 不用先判斷是哪一種錯誤。
    job = JOB_REGISTRY.get(job_id)
    job_available = bool(job)

    async def event_generator():
        nonlocal job, job_available
        # prepared_rows 來自 parsed data（原始 Excel 列），不含首次 stream 生成時
        # 以 row.added 事件補進的 sub-TC（id 形如 `foo__tc2`）。這些 sub-TC 只存
        # 在 frontend store，所以 regenerate 要改用 `payload.rows` 補齊，否則被
        # 使用者勾選的子 TC 永遠找不到對應 row，stream 不 yield → UI 卡 generating。
        # 若尚未跑過 /api/generate（job["rows"] 不存在），直接依 payload 處理。
        prepared_rows = _prepare_generation_rows(job) if job and job.get("rows") else []
        prepared_by_id = {r["id"]: r for r in prepared_rows}

        def _row_from_payload(raw: dict) -> dict:
            """把前端送來的 TcRow（camelCase）轉成 generator 吃的 snake_case shape。
            只覆蓋 regenerate 用得到的欄位；allow_split=False 不需要既有 pre/post。"""
            req_id = str(raw.get("reqId") or "")
            test_item = str(raw.get("testItem") or "")
            return {
                "id": str(raw.get("id") or ""),
                "row_num": raw.get("rowNum") or 0,
                "tc_id": normalize_tc_id(str(raw.get("tcId") or "")),
                "req_id": req_id,
                "test_item": test_item,
                "original_requirement": test_item,
                "test_set": raw.get("testSet") or "",
                "spec_reference": raw.get("specReference"),
                "priority": raw.get("priority") or "",
                "pre_conditions": str(raw.get("preConditions") or ""),
                "input_test_data": str(raw.get("inputTestData") or ""),
                "test_procedure": str(raw.get("steps") or ""),
                "expected_result": str(raw.get("expectedResults") or ""),
                "design_method": str(raw.get("designMethod") or ""),
                "regenerate_reason": str(raw.get("regenerateReason") or payload.regenerateReason or ""),
            }

        def _merge_payload_row(base: dict, raw: dict | None) -> dict:
            merged = dict(base)
            if raw:
                payload_row = _row_from_payload(raw)
                for key, value in payload_row.items():
                    if value is None:
                        continue
                    if isinstance(value, str) and not value.strip():
                        continue
                    merged[key] = value
            if payload.regenerateReason and str(payload.regenerateReason).strip():
                merged["regenerate_reason"] = str(payload.regenerateReason).strip()
            return merged

        payload_by_id = {
            str(r.get("id")): r for r in payload.rows if r.get("id")
        }

        rows_to_regen: list[dict] = []
        for rid in payload.rowIds:
            if rid in prepared_by_id:
                rows_to_regen.append(_merge_payload_row(prepared_by_id[rid], payload_by_id.get(rid)))
            elif rid in payload_by_id:
                rows_to_regen.append(_row_from_payload(payload_by_id[rid]))
        # 順序即 payload.rowIds 原順序，無需再排序。

        cfg = (
            payload.config.model_dump() if payload.config
            else (job.get("config", {}) if job_available else {})
        )
        parsed = job.get("parsedData", {}) if job_available else {}
        context = {
            "project": parsed.get("project") or "",
            "test_group": parsed.get("test_group") or "",
            "test_set": "N/A",
        }
        spec_index = _build_spec_index_for_job(job) if job_available else {}
        total = len(rows_to_regen)
        model = cfg.get("model", DEFAULT_MODEL)
        batch_size = cfg.get("batchSize", 1)
        processed = 0
        usage = _job_usage(job) if job_available else _job_usage(None)
        current_cost = usage["cost"]
        total_in = usage["inputTokens"]
        total_out = usage["outputTokens"]
        total_cache_create = usage["cacheCreationTokens"]
        total_cache_read = usage["cacheReadTokens"]
        persisted_cost = usage["cost"]

        all_payload_rows = list(payload_by_id.values()) + list(prepared_by_id.values())
        existing_seqs = _extract_existing_sequence_numbers(all_payload_rows)
        next_seq = (max(existing_seqs) + 1) if existing_seqs else 1

        prefix_counts: dict[str, int] = {}
        for row in all_payload_rows:
            tc_id_raw = row.get("tc_id") or row.get("tcId")
            if not tc_id_raw:
                continue
            parts = str(tc_id_raw).rsplit("-", 1)
            if len(parts) == 2 and parts[1].isdigit():
                prefix_counts[parts[0]] = prefix_counts.get(parts[0], 0) + 1
        tc_id_prefix = (
            max(prefix_counts.items(), key=lambda kv: kv[1])[0]
            if prefix_counts else None
        )

        tc_project = context["project"] or ""
        tc_group_abbr = (
            generate_group_abbreviation(context["test_group"])
            if context["test_group"] else ""
        )
        pad_width = 3
        for row in all_payload_rows:
            tc_id_raw = row.get("tc_id") or row.get("tcId")
            if not tc_id_raw:
                continue
            parts = str(tc_id_raw).rsplit("-", 1)
            if len(parts) == 2 and parts[1].isdigit():
                pad_width = max(pad_width, len(parts[1]))

        def _alloc_sub_tc_id() -> str:
            nonlocal next_seq
            if tc_id_prefix:
                new_id = f"{tc_id_prefix}-{str(next_seq).zfill(pad_width)}"
                next_seq += 1
                return new_id
            if tc_project and tc_group_abbr:
                new_id = generate_tc_ids(tc_project, tc_group_abbr, 1, start=next_seq)[0]
                next_seq += 1
                return new_id
            return ""

        def _stats() -> dict:
            return {
                "total": total,
                "processed": processed,
                "currentCost": round(current_cost, 4),
                "inputTokens": total_in,
                "outputTokens": total_out,
                "cacheCreationTokens": total_cache_create,
                "cacheReadTokens": total_cache_read,
            }

        yield _sse_event({"type": "regen.started", "jobId": job_id, "total": total, "stats": _stats()})

        for i in range(0, total, batch_size):
            batch = rows_to_regen[i : i + batch_size]
            try:
                batch_result = generate_tc_tool(
                    rows=batch,
                    context=context,
                    spec_index=spec_index,
                    rules_text=RULES_SECTIONS,
                    model=model,
                    allow_split=True,
                )
                tc_data_list = batch_result["tcData"]
                split_meta_list = batch_result.get("splitMeta") or []

                current_cost += batch_result["cost"]
                total_in += batch_result["inputTokens"]
                total_out += batch_result["outputTokens"]
                total_cache_create += batch_result["cacheCreationTokens"]
                total_cache_read += batch_result["cacheReadTokens"]
                persist_target = job if (job_available and job is not None) else {"jobId": job_id}
                batch_delta = max(current_cost - persisted_cost, 0.0)
                if batch_delta > 0:
                    _persist_job_usage(
                        job_id,
                        persist_target,
                        cost=current_cost,
                        input_tokens=total_in,
                        output_tokens=total_out,
                        cache_creation_tokens=total_cache_create,
                        cache_read_tokens=total_cache_read,
                        cost_delta=batch_delta,
                        model=model,
                    )
                    if not job_available or job is None:
                        job = JOB_REGISTRY.get(job_id)
                        job_available = job is not None
                    persisted_cost = current_cost
                for group_idx, (row, tc_group) in enumerate(zip(batch, tc_data_list)):
                    if not isinstance(tc_group, list) or not tc_group:
                        continue
                    processed += 1
                    extras = max(len(tc_group) - 1, 0)
                    if extras > 0:
                        total += extras

                    meta = split_meta_list[group_idx] if group_idx < len(split_meta_list) else {}
                    split_warning = _heuristic_split_warning(row, len(tc_group))
                    split_payload = {
                        "type": "req.split",
                        "jobId": job_id,
                        "rowId": row["id"],
                        "reqId": row.get("req_id") or meta.get("req_id") or "",
                        "tcCount": len(tc_group),
                        "reasoning": str(meta.get("reasoning") or ""),
                        "keywords": meta.get("keywords") or [],
                        "insertPlan": {
                            "needsInsert": extras > 0,
                            "insertAfterId": row["id"],
                            "newCount": extras,
                            "renumberRequired": extras > 0,
                        },
                        "stats": _stats(),
                    }
                    if split_warning:
                        split_payload["splitWarning"] = split_warning
                    yield _sse_event(split_payload)

                    for sub_idx, tc in enumerate(tc_group):
                        is_primary = sub_idx == 0
                        sub_row = dict(row)
                        if is_primary:
                            sub_row["tc_id"] = row.get("tc_id") or _alloc_sub_tc_id()
                        else:
                            sub_row["id"] = f"{row['id']}__tc{sub_idx + 1}"
                            sub_row["tc_id"] = _alloc_sub_tc_id()

                        updated_row, _ = _build_stream_row(sub_row, tc)
                        updated_row["splitDecision"] = {
                            "reqId": row.get("req_id") or meta.get("req_id") or "",
                            "tcCount": len(tc_group),
                            "subIndex": sub_idx,
                            "parentId": row["id"],
                            "reasoning": str(meta.get("reasoning") or "") if is_primary else "",
                            "keywords": (meta.get("keywords") or []) if is_primary else [],
                            # AI strict-marked equivalent sibling row id (Layer 1
                            # duplicate detection). Empty when not duplicate.
                            "duplicateOf": (
                                _resolve_duplicate_of(meta.get("duplicate_of"), row.get("siblings"))
                                if is_primary
                                else ""
                            ),
                            # AI 對 sibling 差異的明確聲明（B 方案），primary 才帶。
                            # axis: trigger_state / input_data / timing / boundary / mode / none；
                            # delta: 繁中差異描述含具體 token。沒 sibling 時 AI omit。
                            "distinguishingAxis": (
                                meta.get("distinguishing_axis") or {}
                                if is_primary
                                else {}
                            ),
                        }
                        if split_warning and is_primary:
                            updated_row["splitWarning"] = split_warning
                        if not is_primary:
                            updated_row["parentId"] = row["id"]
                            updated_row["subIndex"] = sub_idx

                        event_type = "row.regenerated" if is_primary else "row.added"
                        yield _sse_event(
                            {
                                "type": event_type,
                                "jobId": job_id,
                                "row": updated_row,
                                "stats": _stats(),
                            }
                        )
            except ToolError as exc:
                if _is_quota_exceeded_tool_error(exc):
                    yield _sse_event(
                        {
                            "type": "regen.failed",
                            "jobId": job_id,
                            "stats": _stats(),
                            "message": exc.message,
                        }
                    )
                    return
                for row in batch:
                    processed += 1
                    yield _sse_event(
                        {
                            "type": "row.regen_failed",
                            "jobId": job_id,
                            "row": _build_failed_stream_row(row, exc.message),
                            "stats": _stats(),
                        }
                    )
                    _record_stream_validation_failure(
                        job_id, row, exc.message, field="regenerate"
                    )
            await asyncio.sleep(0.05)

        # Job 缺失時也建一筆最小紀錄 persist usage，保持和 rerun 一致的
        # 語意；CostMeter 才看得到這次 regen 的花費。
        persist_target = job if (job_available and job is not None) else {"jobId": job_id}
        stream_delta = max(current_cost - persisted_cost, 0.0)
        _persist_job_usage(
            job_id,
            persist_target,
            cost=current_cost,
            input_tokens=total_in,
            output_tokens=total_out,
            cache_creation_tokens=total_cache_create,
            cache_read_tokens=total_cache_read,
            cost_delta=stream_delta,
            model=model,
        )

        yield _sse_event(
            {
                "type": "regen.completed",
                "jobId": job_id,
                "stats": _stats(),
            }
        )

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/jobs/{job_id}/rerun/stream")
async def stream_rerun(job_id: str, payload: RegenerateRequest) -> StreamingResponse:
    """Re-run selected rows through the FULL generation pipeline.

    和 `/regenerate/stream` 的差別：
      - 走 `allow_split=True` → AI 會重新判斷需求解讀與拆分
      - 一筆選中的 row 可能產出多筆 TC：第一筆以 `row.regenerated` 覆蓋原列，
        其餘以 `row.added`（含 `parentId`）串流回前端，由前端插在該 row 之後
      - 會 emit `req.split` 事件讓前端顯示拆分 reasoning / keyword 分析

    Stream event types:
      - `rerun.started`
      - `req.split`
      - `row.regenerated`（primary TC，覆蓋原 row）
      - `row.added`（sub TCs，前端 insert-after parentId）
      - `row.regen_failed`
      - `rerun.completed`
    """
    job = JOB_REGISTRY.get(job_id)
    # Rerun 容許 job 不存在：常發生在 backend 重啟、SQLite 被清空、或使用者
    # 切換過 session。此時完全依賴前端送來的 payload.rows + project/testGroup
    # 重建執行 context。Spec 索引會退化成空 dict（沒有 SYS1 匹配資訊），是刻意取捨。
    job_available = bool(job)

    async def event_generator():
        nonlocal job, job_available
        prepared_rows = _prepare_generation_rows(job) if job and job.get("rows") else []
        prepared_by_id = {r["id"]: r for r in prepared_rows}

        def _row_from_payload(raw: dict) -> dict:
            req_id = str(raw.get("reqId") or "")
            test_item = str(raw.get("testItem") or "")
            return {
                "id": str(raw.get("id") or ""),
                "row_num": raw.get("rowNum") or 0,
                "tc_id": normalize_tc_id(str(raw.get("tcId") or "")),
                "req_id": req_id,
                "test_item": test_item,
                "original_requirement": test_item,
                "test_set": raw.get("testSet") or "",
                "spec_reference": raw.get("specReference"),
                "priority": raw.get("priority") or "",
                "pre_conditions": str(raw.get("preConditions") or ""),
                "input_test_data": str(raw.get("inputTestData") or ""),
                "test_procedure": str(raw.get("steps") or ""),
                "expected_result": str(raw.get("expectedResults") or ""),
                "design_method": str(raw.get("designMethod") or ""),
            }

        def _merge_payload_row(base: dict, raw: dict | None) -> dict:
            if not raw:
                return base
            payload_row = _row_from_payload(raw)
            merged = dict(base)
            for key, value in payload_row.items():
                if value is None:
                    continue
                if isinstance(value, str) and not value.strip():
                    continue
                merged[key] = value
            return merged

        payload_by_id = {str(r.get("id")): r for r in payload.rows if r.get("id")}
        rows_to_run: list[dict] = []
        for rid in payload.rowIds:
            if rid in prepared_by_id:
                rows_to_run.append(_merge_payload_row(prepared_by_id[rid], payload_by_id.get(rid)))
            elif rid in payload_by_id:
                rows_to_run.append(_row_from_payload(payload_by_id[rid]))

        if not rows_to_run:
            yield _sse_event(
                {
                    "type": "rerun.completed",
                    "jobId": job_id,
                    "stats": {"total": 0, "processed": 0, "currentCost": 0},
                    "message": "No matching rows found in payload.",
                }
            )
            return

        cfg = (
            payload.config.model_dump() if payload.config
            else (job.get("config", {}) if job_available else {})
        )
        # Context 優先用 job parsedData（有完整 project/testGroup），否則 fall back 到
        # payload 送來的欄位或從 rows 推導（每 row 有 testSet，但沒有 project/group）。
        parsed = job.get("parsedData", {}) if job_available else {}
        context = {
            "project": parsed.get("project") or payload.project or "",
            "test_group": parsed.get("test_group") or payload.testGroup or "",
            "test_set": "N/A",
        }
        spec_index = _build_spec_index_for_job(job) if job_available else {}
        total = len(rows_to_run)
        model = cfg.get("model", DEFAULT_MODEL)
        batch_size = cfg.get("batchSize", 1)
        processed = 0
        usage = _job_usage(job) if job_available else _job_usage(None)
        current_cost = usage["cost"]
        total_in = usage["inputTokens"]
        total_out = usage["outputTokens"]
        total_cache_create = usage["cacheCreationTokens"]
        total_cache_read = usage["cacheReadTokens"]
        persisted_cost = usage["cost"]

        # 新 sub-TC 的 tc_id 分配：
        # 1. 從現有 rows 抽最多人用的 tc_id prefix（例如 `newR1L-DM-`），確保新
        #    sub-TC 和既有 TC 格式一致，不會一半 `newR1L-DM-045` 一半 `Proj-DM-046`。
        # 2. 編號從現有最大 +1 開始，不會撞號。
        # 3. 沒有 prefix 可用時才 fallback 到 project + group_abbr scheme。
        # Primary 沿用原 tc_id，不走這裡。
        all_payload_rows = list(payload_by_id.values()) + list(prepared_by_id.values())
        existing_seqs = _extract_existing_sequence_numbers(all_payload_rows)
        next_seq = (max(existing_seqs) + 1) if existing_seqs else 1

        prefix_counts: dict[str, int] = {}
        for row in all_payload_rows:
            tc_id_raw = row.get("tc_id") or row.get("tcId")
            if not tc_id_raw:
                continue
            parts = str(tc_id_raw).rsplit("-", 1)
            if len(parts) == 2 and parts[1].isdigit():
                prefix_counts[parts[0]] = prefix_counts.get(parts[0], 0) + 1
        tc_id_prefix = (
            max(prefix_counts.items(), key=lambda kv: kv[1])[0]
            if prefix_counts else None
        )

        tc_project = context["project"] or ""
        tc_group_abbr = (
            generate_group_abbreviation(context["test_group"])
            if context["test_group"] else ""
        )
        # 抽出既有 tc_id 的編號寬度（`045` → 3），避免 `newR1L-DM-46` vs `-046`
        pad_width = 3
        for row in all_payload_rows:
            tc_id_raw = row.get("tc_id") or row.get("tcId")
            if not tc_id_raw:
                continue
            parts = str(tc_id_raw).rsplit("-", 1)
            if len(parts) == 2 and parts[1].isdigit():
                pad_width = max(pad_width, len(parts[1]))

        def _alloc_sub_tc_id() -> str:
            nonlocal next_seq
            if tc_id_prefix:
                new_id = f"{tc_id_prefix}-{str(next_seq).zfill(pad_width)}"
                next_seq += 1
                return new_id
            if tc_project and tc_group_abbr:
                new_id = generate_tc_ids(tc_project, tc_group_abbr, 1, start=next_seq)[0]
                next_seq += 1
                return new_id
            return ""

        def _stats() -> dict:
            return {
                "total": total,
                "processed": processed,
                "currentCost": round(current_cost, 4),
                "inputTokens": total_in,
                "outputTokens": total_out,
                "cacheCreationTokens": total_cache_create,
                "cacheReadTokens": total_cache_read,
            }

        yield _sse_event({"type": "rerun.started", "jobId": job_id, "total": total, "stats": _stats()})

        for i in range(0, total, batch_size):
            batch = rows_to_run[i : i + batch_size]
            try:
                batch_result = generate_tc_tool(
                    rows=batch,
                    context=context,
                    spec_index=spec_index,
                    rules_text=RULES_SECTIONS,
                    model=model,
                    allow_split=True,
                )
                tc_data_list = batch_result["tcData"]
                split_meta_list = batch_result.get("splitMeta") or []

                current_cost += batch_result["cost"]
                total_in += batch_result["inputTokens"]
                total_out += batch_result["outputTokens"]
                total_cache_create += batch_result["cacheCreationTokens"]
                total_cache_read += batch_result["cacheReadTokens"]
                persist_target = job if (job_available and job is not None) else {"jobId": job_id}
                batch_delta = max(current_cost - persisted_cost, 0.0)
                if batch_delta > 0:
                    _persist_job_usage(
                        job_id,
                        persist_target,
                        cost=current_cost,
                        input_tokens=total_in,
                        output_tokens=total_out,
                        cache_creation_tokens=total_cache_create,
                        cache_read_tokens=total_cache_read,
                        cost_delta=batch_delta,
                        model=model,
                    )
                    if not job_available or job is None:
                        job = JOB_REGISTRY.get(job_id)
                        job_available = job is not None
                    persisted_cost = current_cost

                for group_idx, (row, tc_group) in enumerate(zip(batch, tc_data_list)):
                    if not isinstance(tc_group, list) or not tc_group:
                        continue
                    processed += 1
                    # 如果 AI 拆出 N>1 → total 往上加 N-1，進度條一致。
                    extras = max(len(tc_group) - 1, 0)
                    if extras > 0:
                        total += extras

                    split_warning = _heuristic_split_warning(row, len(tc_group))
                    meta = split_meta_list[group_idx] if group_idx < len(split_meta_list) else {}
                    split_payload = {
                        "type": "req.split",
                        "jobId": job_id,
                        "rowId": row["id"],
                        "reqId": row.get("req_id") or meta.get("req_id") or "",
                        "tcCount": len(tc_group),
                        "reasoning": str(meta.get("reasoning") or ""),
                        "keywords": meta.get("keywords") or [],
                        "insertPlan": {
                            "needsInsert": extras > 0,
                            "insertAfterId": row["id"],
                            "newCount": extras,
                            "renumberRequired": extras > 0,
                        },
                        "stats": _stats(),
                        "message": (
                            f"{row.get('req_id') or row['id']}: "
                            f"AI split into {len(tc_group)} TC(s). "
                            f"{str(meta.get('reasoning') or '')[:200]}"
                        ).strip(),
                    }
                    if split_warning:
                        split_payload["splitWarning"] = split_warning
                    yield _sse_event(split_payload)

                    for sub_idx, tc in enumerate(tc_group):
                        is_primary = sub_idx == 0
                        sub_row = dict(row)
                        if is_primary:
                            # 保留原 tc_id，primary 覆蓋既有列
                            sub_row["tc_id"] = row.get("tc_id") or _alloc_sub_tc_id()
                        else:
                            sub_row["id"] = f"{row['id']}__tc{sub_idx + 1}"
                            sub_row["tc_id"] = _alloc_sub_tc_id()
                        updated_row, _has_warnings = _build_stream_row(sub_row, tc)
                        updated_row["splitDecision"] = {
                            "reqId": row.get("req_id") or meta.get("req_id") or "",
                            "tcCount": len(tc_group),
                            "subIndex": sub_idx,
                            "parentId": row["id"],
                            "reasoning": str(meta.get("reasoning") or "") if is_primary else "",
                            "keywords": (meta.get("keywords") or []) if is_primary else [],
                            # AI strict-marked equivalent sibling row id (Layer 1
                            # duplicate detection). Empty when not duplicate.
                            "duplicateOf": (
                                _resolve_duplicate_of(meta.get("duplicate_of"), row.get("siblings"))
                                if is_primary
                                else ""
                            ),
                            # AI 對 sibling 差異的明確聲明（B 方案），primary 才帶。
                            # axis: trigger_state / input_data / timing / boundary / mode / none；
                            # delta: 繁中差異描述含具體 token。沒 sibling 時 AI omit。
                            "distinguishingAxis": (
                                meta.get("distinguishing_axis") or {}
                                if is_primary
                                else {}
                            ),
                        }
                        if split_warning and is_primary:
                            updated_row["splitWarning"] = split_warning
                        if not is_primary:
                            updated_row["parentId"] = row["id"]
                            updated_row["subIndex"] = sub_idx

                        event_type = "row.regenerated" if is_primary else "row.added"
                        yield _sse_event(
                            {
                                "type": event_type,
                                "jobId": job_id,
                                "row": updated_row,
                                "stats": _stats(),
                                "message": (
                                    f"Re-ran {row.get('req_id') or row['id']}"
                                    + (f" (TC {sub_idx + 1}/{len(tc_group)})" if len(tc_group) > 1 else "")
                                ),
                            }
                        )
            except ToolError as exc:
                if _is_quota_exceeded_tool_error(exc):
                    yield _sse_event(
                        {
                            "type": "rerun.failed",
                            "jobId": job_id,
                            "stats": _stats(),
                            "message": exc.message,
                        }
                    )
                    return
                for row in batch:
                    processed += 1
                    yield _sse_event(
                        {
                            "type": "row.regen_failed",
                            "jobId": job_id,
                            "row": _build_failed_stream_row(row, exc.message),
                            "stats": _stats(),
                        }
                    )
                    _record_stream_validation_failure(
                        job_id, row, exc.message, field="rerun"
                    )
            await asyncio.sleep(0.05)

        # Rerun 即使 job 原本不存在（backend 重啟 / DB 清空時的 fallback 路徑）
        # 也要建立 job 紀錄來 persist 這次跑的 usage，否則 CostMeter 會見鬼
        # —— 使用者看到 $$ 花了卻沒進累計。
        persist_target = job if (job_available and job is not None) else {"jobId": job_id}
        stream_delta = max(current_cost - persisted_cost, 0.0)
        _persist_job_usage(
            job_id,
            persist_target,
            cost=current_cost,
            input_tokens=total_in,
            output_tokens=total_out,
            cache_creation_tokens=total_cache_create,
            cache_read_tokens=total_cache_read,
            cost_delta=stream_delta,
            model=model,
        )

        yield _sse_event(
            {
                "type": "rerun.completed",
                "jobId": job_id,
                "stats": _stats(),
            }
        )

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


def _build_blank_template_workbook(rows: list[dict], test_group: str | None) -> bytes:
    """當 backend JOB_REGISTRY 已無原 Excel rawBytes 時，從前端 TcRow 產生一份
    最小可用的空白模板給 writer 寫入。

    產出的 workbook 結構符合 writer 期望：
      - Sheet "Test Case Specification&Result" 存在
      - Row 9 是 header（writer 不依賴 header 內容，但為了打開檔案閱讀方便還是寫上）
      - 每個 row_num 對應的列預先填好 Col D = req_id、Col I = test_item，
        writer 之後會 append AI rewrite 到 Col I 並寫其他欄位

    缺：原檔的 styling / Product Document sheet / 其他 untouched 欄位（B/C/E/O/...）。
    """
    from io import BytesIO
    from openpyxl import Workbook
    from writer import TC_SHEET_NAME

    wb = Workbook()
    ws = wb.active
    ws.title = TC_SHEET_NAME

    # Header row (col D / F / I 是 writer 會碰到的最少欄位)
    ws.cell(row=9, column=4, value="Requirement or Design ID")
    ws.cell(row=9, column=6, value="Test Case ID")
    ws.cell(row=9, column=7, value="Test Group")
    ws.cell(row=9, column=8, value="Test Set")
    ws.cell(row=9, column=9, value="Test Item")
    ws.cell(row=9, column=10, value="Pre-Conditions")
    ws.cell(row=9, column=11, value="Input Test Data")
    ws.cell(row=9, column=12, value="Test Procedure")
    ws.cell(row=9, column=13, value="Expected Result")
    ws.cell(row=9, column=14, value="Specification Reference")
    ws.cell(row=9, column=16, value="Priority")
    ws.cell(row=9, column=17, value="Test Case Design Method")

    # 預先把 req_id / test_item / test_group 寫入對應 row_num，writer 之後會 overlay。
    # 同 row_num 多筆（AI 拆分）只寫第一筆當 template，writer 會 insert_rows 補齊。
    written_rows: set[int] = set()
    for r in rows:
        rn = r.get("row_num")
        if not rn or rn in written_rows:
            continue
        written_rows.add(rn)
        if r.get("req_id"):
            ws.cell(row=rn, column=4, value=str(r["req_id"]))
        if r.get("test_item"):
            ws.cell(row=rn, column=9, value=str(r["test_item"]))
        if test_group:
            ws.cell(row=rn, column=7, value=test_group)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _prune_unexported_source_rows(
    source_path: str,
    export_rows: list[dict],
    parsed_rows: list[dict] | None,
) -> list[dict]:
    """Remove original workbook rows that are no longer part of the export.

    Review-side deletes only remove rows from the frontend payload. When export
    uses the original workbook as a template, those deleted source rows would
    otherwise remain in the output as stale, non-overwritten lines.
    """
    if not parsed_rows:
        return export_rows

    source_row_nums = {
        int(row["row_num"])
        for row in parsed_rows
        if isinstance(row.get("row_num"), int) and row.get("row_num") > 0
    }
    keep_row_nums = {
        int(row["row_num"])
        for row in export_rows
        if isinstance(row.get("row_num"), int) and row.get("row_num") > 0
    }
    delete_row_nums = sorted(source_row_nums - keep_row_nums)
    if not delete_row_nums:
        return export_rows

    from bisect import bisect_left
    from openpyxl import load_workbook
    from writer import TC_SHEET_NAME

    wb = load_workbook(source_path)
    ws = wb[TC_SHEET_NAME]
    for row_num in reversed(delete_row_nums):
        ws.delete_rows(row_num, 1)
    wb.save(source_path)
    wb.close()

    adjusted_rows: list[dict] = []
    for row in export_rows:
        row_num = row.get("row_num")
        if not isinstance(row_num, int) or row_num <= 0:
            adjusted_rows.append(row)
            continue
        adjusted = dict(row)
        adjusted["row_num"] = row_num - bisect_left(delete_row_nums, row_num)
        adjusted_rows.append(adjusted)
    return adjusted_rows


def _resolve_export_test_group(job: dict | None, rows: list[dict]) -> str | None:
    """Export context 的 test_group：優先用 parsedData，沒有就從前端 row.testGroup 推。

    多 test_group 混雜時（理論上一個 job 只有一個 group，但 payload 理論上可混）
    會 log 一條 warning，提醒 blank-template export 只能認一個 group。
    """
    if job and job.get("parsedData", {}).get("test_group"):
        return job["parsedData"]["test_group"]
    groups = {str(r.get("testGroup")) for r in rows if r.get("testGroup")}
    if len(groups) > 1:
        import logging
        logging.getLogger(__name__).warning(
            "export payload contains multiple testGroup values %s; "
            "blank-template export will use the first one only. "
            "Re-upload the original workbook per group for proper styling.",
            sorted(groups),
        )
    for row in rows:
        tg = row.get("testGroup")
        if tg:
            return str(tg)
    return None


@app.post("/api/export")
async def export_job(payload: ExportRequest, request: Request) -> dict:
    try:
        if payload.outputMode == "overwrite":
            raise HTTPException(status_code=400, detail="overwrite export mode is not supported")

        job = JOB_REGISTRY.get(payload.jobId)
        has_source_workbook = bool(
            job and job.get("rawBytes") and job.get("rawFileName")
        )

        test_group = _resolve_export_test_group(job, payload.rows)
        parsed_rows = (
            job.get("parsedData", {}).get("rows") if job and job.get("parsedData") else None
        )

        export_rows, classify_usage = _map_export_rows(
            payload.rows,
            payload.scope,
            test_group,
            parsed_rows=parsed_rows,
        )
        # Renumber TC IDs to fill any gaps left by reviewer-side deletes,
        # so the exported workbook stays contiguous (-001, -002, …).
        tc_ids_renumbered = _resequence_export_tc_ids(export_rows)
        if not export_rows:
            if not payload.rows:
                detail = "no rows in payload — re-select rows in Review before export."
            else:
                detail = (
                    f"no rows match scope='{payload.scope}' "
                    f"(input had {len(payload.rows)} row(s) but none qualify). "
                    f"Try scope='all' or accept / flag more rows in Review."
                )
            raise HTTPException(status_code=400, detail=detail)

        if classify_usage.get("cost", 0) > 0:
            from generator import CLASSIFICATION_MODEL
            existing = _job_usage(job)
            _persist_job_usage(
                payload.jobId,
                job if job else {},
                cost=existing["cost"] + classify_usage["cost"],
                input_tokens=existing["inputTokens"] + classify_usage["inputTokens"],
                output_tokens=existing["outputTokens"] + classify_usage["outputTokens"],
                cache_creation_tokens=existing["cacheCreationTokens"] + classify_usage["cacheCreationTokens"],
                cache_read_tokens=existing["cacheReadTokens"] + classify_usage["cacheReadTokens"],
                cost_delta=float(classify_usage["cost"]),
                model=CLASSIFICATION_MODEL,
            )
            if job is None:
                job = JOB_REGISTRY.get(payload.jobId)

        selected_fields = _selected_export_fields(payload.selectedColumns)
        framework_rows = (
            _build_framework_rows(export_rows, test_group)
            if payload.includeFrameworkSheet
            else None
        )

        if has_source_workbook:
            raw_bytes = job["rawBytes"]
            raw_filename = job["rawFileName"]
            export_path = _build_export_path(raw_filename, payload.outputMode)
        else:
            raw_bytes = _build_blank_template_workbook(export_rows, test_group)
            fallback_name = (
                (job and job.get("rawFileName"))
                or f"{test_group or 'TC'}_export.xlsx"
            )
            export_path = _build_export_path(fallback_name, payload.outputMode)

        with tempfile.TemporaryDirectory() as tmp_dir:
            source_path = os.path.join(tmp_dir, os.path.basename(export_path) or "source.xlsx")
            with open(source_path, "wb") as source_file:
                source_file.write(raw_bytes)

            rows_to_write = (
                _prune_unexported_source_rows(source_path, export_rows, parsed_rows)
                if has_source_workbook
                else export_rows
            )

            try:
                write_excel_tool(
                    source_path=source_path,
                    output_path=export_path,
                    rows=rows_to_write,
                    selected_fields=selected_fields,
                    framework_rows=framework_rows,
                )
            except ToolError as exc:
                raise _tool_error_to_http(exc) from exc

        if job is None:
            job = {}
        job["exportPath"] = export_path
        job.setdefault("rawFileName", os.path.basename(export_path))
        JOB_REGISTRY[payload.jobId] = job
        download_url = str(request.url_for("download_export", jobId=payload.jobId))
        from generator import CLASSIFICATION_MODEL as _CLASS_MODEL
        return {
            "jobId": payload.jobId,
            "status": "ready",
            "exportedRows": len(export_rows),
            "fileName": os.path.basename(export_path),
            "downloadUrl": download_url,
            "selectedColumns": payload.selectedColumns,
            "fallbackTemplate": not has_source_workbook,
            # Number of TC IDs renumbered to fill gaps from prior deletes.
            "tcIdsRenumbered": tc_ids_renumbered,
            # AI Test Set classification usage incurred during this export.
            # Frontend uses this to record an `export` entry in the workspace
            # job-history store; cost > 0 only when some rows arrived without
            # `testSet` and the backend ran `classify_test_sets` again.
            "classifyUsage": {
                **classify_usage,
                "model": _CLASS_MODEL,
            },
        }
    except HTTPException:
        raise
    except Exception as exc:
        logging.getLogger(__name__).exception(
            "Unexpected export failure for job %s", payload.jobId
        )
        raise HTTPException(
            status_code=500,
            detail=f"export failed: {type(exc).__name__}: {exc}",
        ) from exc


@app.post("/api/jobs/{job_id}/attach-raw")
async def attach_raw_workbook(job_id: str, raw_file: UploadFile = File(...)) -> dict:
    """為既有 job 補上原始 Excel rawBytes。

    使用情境：backend JOB_REGISTRY 仍留著 job 但 rawBytes 已遺失（例如
    匯入 .tcw.json workspace 後才要 export），前端偵測到 fallbackTemplate
    會 prompt 使用者上傳原始 Excel，再呼叫這支 endpoint 補回。
    """
    filename = _safe_upload_filename(raw_file.filename, "upload.xlsx")
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_RAW_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"unsupported extension '{ext}' — only .xlsx/.xlsm allowed",
        )
    raw_bytes = await _read_with_limit(raw_file, "raw_file")

    job = JOB_REGISTRY.get(job_id)
    if not job:
        # 允許在 backend 完全沒有 job 紀錄時，以最小 stub 建立
        job = {"jobId": job_id}

    job["rawBytes"] = raw_bytes
    job["rawFileName"] = filename
    JOB_REGISTRY[job_id] = job
    return {
        "jobId": job_id,
        "rawFileName": filename,
        "size": len(raw_bytes),
        "hasSource": True,
    }


@app.get("/api/jobs/{job_id}/source-status")
async def get_source_status(job_id: str) -> dict:
    """回報 job 是否仍保有 rawBytes，前端 export 前先用來判斷要不要 prompt 補上傳。"""
    job = JOB_REGISTRY.get(job_id)
    has_source = bool(job and job.get("rawBytes") and job.get("rawFileName"))
    return {
        "jobId": job_id,
        "hasSource": has_source,
        "rawFileName": (job or {}).get("rawFileName") if has_source else None,
    }


@app.get("/api/export/download/{jobId}", name="download_export")
async def download_export(jobId: str) -> FileResponse:
    job = JOB_REGISTRY.get(jobId)
    if not job or not job.get("exportPath") or not os.path.exists(job["exportPath"]):
        raise HTTPException(status_code=404, detail="export file not found")

    return FileResponse(
        job["exportPath"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(job["exportPath"]),
    )
