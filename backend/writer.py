"""Excel writer for generated TC results (RULES.md §10.3)."""
import os
import re
from copy import copy

from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment

TC_SHEET_NAME = "Test Case Specification&Result"
TC_SHEET_PREFIX = "Test Case Specification"
FRAMEWORK_SHEET_NAME = "Test Case Framework"

# Column mapping: field name -> 1-based column index
WRITE_COLUMNS = {
    "tc_id": 6,           # F
    "test_group": 7,      # G
    "test_set": 8,        # H
    "pre_conditions": 10, # J
    "input_test_data": 11,# K
    "test_procedure": 12, # L
    "expected_result": 13,# M
    "spec_reference": 14, # N
    "priority": 16,       # P
    "design_method": 17,  # Q
    # AI 對該 requirement 的解讀（splitDecision.reasoning）。Primary TC 才有；
    # sub TC 的 reasoning 為空字串會被 _write_tc_row 跳過。column R 在大多數
    # template 是空欄，_ensure_reasoning_header 會自動補一個 header 標籤。
    "reasoning": 18,      # R
}

CLEARABLE_FIELDS = {"test_set", "spec_reference"}

# 這四欄輸出時每行末尾都不留句點（"." / "。"）— 與 generator._normalize_tc_dict
# 保持一致，舊存量資料在匯出時也會被一併清掉。
_NO_TRAILING_PERIOD_FIELDS = {
    "pre_conditions",
    "input_test_data",
    "test_procedure",
    "expected_result",
}


def _strip_trailing_periods_per_line(text):
    """逐行移除末尾句點；非字串原樣返回。"""
    if not isinstance(text, str) or not text:
        return text
    cleaned_lines = []
    for line in text.split("\n"):
        s = line.rstrip()
        while s and s[-1] in (".", "。"):
            s = s[:-1].rstrip()
        cleaned_lines.append(s)
    return "\n".join(cleaned_lines)


# 之前寫入過的 rewrite 尾巴：跨行抓最後一段括號包住的文字（greedy 非必要）。
# 用 DOTALL 以便跨行比對多行 rewrite。
_REWRITE_TAIL_RE = re.compile(r"\n\n\(.*\)\s*$", re.DOTALL)


def find_tc_sheet_name(wb) -> str:
    """Resolve the TC sheet, accepting bilingual template sheet names."""
    if TC_SHEET_NAME in wb.sheetnames:
        return TC_SHEET_NAME

    prefix = TC_SHEET_PREFIX.lower()
    for name in wb.sheetnames:
        if name.strip().lower().startswith(prefix):
            return name

    raise KeyError(f"Worksheet {TC_SHEET_NAME} does not exist.")


def _wrap_rewrite(rewrite: str) -> str:
    """強制 rewrite 被一對外層 `()` 包住，避免 AI 只回純文字或只帶內層 `()`。

    策略：先剝掉前後空白，若首尾不是「合法外層括號 pair」就整段重新包。
    判斷合法：`text.startswith("(") and text.endswith(")") and` 前後括號數量
    平衡且最外層括號是匹配的。保守起見做法：只要外層不是一對單一括號就重包。
    """
    text = (rewrite or "").strip()
    if not text:
        return ""

    def _wrapped_ok(s: str) -> bool:
        # 首尾必須是括號；且首括號對應到最後一個字元（代表是最外層 pair）。
        if not (s.startswith("(") and s.endswith(")")):
            return False
        depth = 0
        for i, ch in enumerate(s):
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
                if depth == 0 and i != len(s) - 1:
                    # 中途就歸零 → 首尾不是同一對外括號（例：`(a)(b)` 或 `(a)b)` ）
                    return False
        return depth == 0

    if _wrapped_ok(text):
        return text
    # 去掉原有不平衡 / 錯位的首尾括號，再整段包起來。
    stripped = text
    while stripped.startswith("("):
        stripped = stripped[1:].lstrip()
    while stripped.endswith(")"):
        stripped = stripped[:-1].rstrip()
    return f"({stripped})"


def _sanitize_excel_text(value):
    """Strip control chars that OpenXML worksheets cannot store."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _merge_test_item_text(existing_text: str | None, rewrite: str) -> str:
    """在 Col I 的原始 test item 文字後面追加 AI rewrite。

    保留 template 原文（即使包含 `\\n\\n` 雙語分隔）— 只有當尾端真的是
    「舊的 `(... → ...)` rewrite」時才剝除，避免重跑時鎖鏈式增生。
    """
    original = _sanitize_excel_text(existing_text or "")
    # 移除「先前寫入過」的 rewrite 尾巴，不是所有 \n\n 都是 rewrite 分隔。
    original = _REWRITE_TAIL_RE.sub("", original).rstrip()
    wrapped = _wrap_rewrite(_sanitize_excel_text(rewrite))
    if not wrapped:
        return original
    return f"{original}\n\n{wrapped}"


# Suffixes that OS file managers append when duplicating a file. We strip
# these before tagging "_generated.xlsx" so the output filename stays clean
# and the `..._SWQT_{TestGroup}_{date}` pattern survives for re-import.
# Order matters: more specific patterns first so e.g. "bar - Copy" strips
# the whole " - Copy" (not just " Copy" leaving a dangling hyphen).
_DUPLICATE_SUFFIX_PATTERNS = [
    re.compile(r"\s*-\s*copy(?:\s*\(\d+\))?$", re.I),  # Windows: "foo - Copy" / "foo - Copy (2)"
    re.compile(r"\s*拷貝(?:\s*\d+)?$"),                 # macOS 繁中: "foo 拷貝" / "foo 拷貝 2"
    re.compile(r"\s*的副本(?:\s*\d+)?$"),                # macOS 簡中: "foo 的副本"
    re.compile(r"\s*copy(?:\s*\d+)?$", re.I),           # macOS 英文: "foo copy" / "foo copy 2"
    re.compile(r"\s*\(\d+\)$"),                         # 通用: "foo (2)"
]
_GENERATED_SUFFIX_RE = re.compile(r"(?:_generated)+$", re.I)


def _clean_basename(basename: str) -> str:
    """Strip common OS "duplicate file" suffixes (拷貝 / copy / (2)) off the tail."""
    cleaned = basename
    while True:
        before = cleaned
        for pattern in _DUPLICATE_SUFFIX_PATTERNS:
            cleaned = pattern.sub("", cleaned)
        cleaned = cleaned.rstrip()
        if cleaned == before:
            return _GENERATED_SUFFIX_RE.sub("", cleaned).rstrip()


def build_output_path(input_path: str, output_dir: str | None = None) -> str:
    """Build output file path: {input_filename}_generated.xlsx"""
    dirname = output_dir or os.path.dirname(input_path)
    basename = os.path.splitext(os.path.basename(input_path))[0]
    basename = _clean_basename(basename)
    return os.path.join(dirname, f"{basename}_generated.xlsx")


_TEMPLATE_CARRY_COLUMNS = (3, 4, 7, 8, 9)  # C/D/G/H/I: Polarion / ReqID / TestGroup / TestSet / TestItem


def _has_value(cell) -> bool:
    v = cell.value
    return v is not None and str(v).strip() != ""


def _copy_style(src, dst) -> None:
    """從 src cell 複製字體/填色/對齊/邊框到 dst cell（樣式物件需 copy 避免共享）。"""
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def _ensure_reasoning_header(ws) -> None:
    """在 column R 頂端自動補上 "AI 需求解讀" header。

    template 原本沒有這欄，但匯出時 reviewer 看到一整欄無標籤的長文會困惑，
    所以掃 J（Pre-Conditions）前 10 列找出 header 列、在同列 R 欄寫上中文標籤
    並複製格式。已存在內容（使用者自訂 template）一律不覆蓋。
    """
    header_row = 1
    for r in range(1, 11):
        header_text = " ".join(
            str(ws.cell(row=r, column=col).value or "").strip().lower()
            for col in (4, 9, 10)
        )
        if (
            "pre" in header_text
            or "前置" in header_text
            or "condition" in header_text
            or "requirement" in header_text
            or "test item" in header_text
        ):
            header_row = r
            break
    # 設 column R 欄寬與自動換行 — reasoning 是 2–5 句長文，不開 wrap 會橫向溢出。
    # column_dimensions 即使 column 未被存取仍能設定，openpyxl 會在儲存時生效。
    ws.column_dimensions["R"].width = 60

    target = ws.cell(row=header_row, column=18)
    if target.value:
        return
    target.value = "AI 需求解讀"
    src = ws.cell(row=header_row, column=10)
    if src.has_style:
        _copy_style(src, target)
    # header cell 強制 wrap_text — 即使 source J header 沒開，標題本身短不會吃虧；
    # 主要目的是保險，多數 template 的 J header 已開 wrap。
    target.alignment = Alignment(
        horizontal=target.alignment.horizontal or "left",
        vertical=target.alignment.vertical or "center",
        wrap_text=True,
    )


def _write_tc_row(
    ws,
    row_num: int,
    row_data: dict,
    selected_fields: set[str] | None,
    append_rewrite: bool,
    rewritten_req_ids: set[str] | None = None,
) -> None:
    """把單一 TC 的資料寫到指定的 Excel 列。

    `rewritten_req_ids` 參數保留做為介面相容，但 multi-TC 流程下每筆 TC 應各自
    帶有不同 scenario tag 的 rewrite（§1.2），不再做跨列 dedup。
    """
    del rewritten_req_ids  # 已廢棄；保留參數避免呼叫端同步改動

    for field, col_idx in WRITE_COLUMNS.items():
        if selected_fields is not None and field not in selected_fields:
            continue
        if field not in row_data:
            continue

        value = _sanitize_excel_text(row_data.get(field))
        if field in _NO_TRAILING_PERIOD_FIELDS:
            value = _strip_trailing_periods_per_line(value)
        cell = ws.cell(row=row_num, column=col_idx)
        has_value = value is not None and str(value).strip() != ""

        if has_value:
            cell.value = value
            # reasoning 是 2–5 句長文，column R 沒有 template 樣式可繼承，
            # 強制給 wrap_text + top-align 避免橫向溢出與垂直置中切字。
            if field == "reasoning":
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )
        elif field in CLEARABLE_FIELDS:
            cell.value = None
        # else: 空值且非 clearable 欄 → 保留 template 既有內容，不覆蓋。

    if append_rewrite:
        rewrite = row_data.get("tc_title")
        if rewrite and (selected_fields is None or "tc_title" in selected_fields):
            cell_i = ws.cell(row=row_num, column=9)
            cell_i.value = _merge_test_item_text(cell_i.value, rewrite)


def write_generated_results(
    input_path: str,
    generated_rows: list[dict],
    output_path: str,
    selected_fields: set[str] | None = None,
) -> None:
    """
    Write generated TC results back to xlsx.

    Preserves all original formatting. Only writes to specified columns.
    Col I (Test Item) gets rewrite appended, not overwritten.

    若多筆 generated_rows 指向同一個 row_num（AI 把一個 requirement 拆成 N 筆 TC），
    會在原列下方 `insert_rows` 補 N-1 列、複製 C/D/I 等欄位，再把每筆 TC 各自寫入。
    """
    wb = load_workbook(input_path)
    ws = wb[find_tc_sheet_name(wb)]
    _ensure_reasoning_header(ws)

    # 依 row_num 分組，保留原本傳入順序；row_num 為 None 的項目丟到尾端維持相容。
    groups: dict[int, list[dict]] = {}
    order: list[int] = []
    orphans: list[dict] = []
    for row_data in generated_rows:
        rn = row_data.get("row_num")
        if not rn:
            orphans.append(row_data)
            continue
        if rn not in groups:
            groups[rn] = []
            order.append(rn)
        groups[rn].append(row_data)

    rewritten_req_ids: set[str] = set()

    # 由上而下處理 row_num 以保留原始順序（尤其影響 rewrite dedup 的先後）。
    # 插列會讓後續列位移，所以維護一個累積 offset 套用到後面的 target row。
    cumulative_offset = 0
    for row_num in sorted(order):
        items = groups[row_num]
        extras = len(items) - 1
        base_row = row_num + cumulative_offset

        if extras > 0:
            # 在原列下方插入 N-1 列，複製 template 的 C/D/G/H/I 欄位過去。
            # Col I (Test Item) 特別處理：若 parent 已經寫過 rewrite tail，
            # 複製到 child 時要先剝掉，讓 child 之後能 append 屬於自己的 rewrite。
            ws.insert_rows(base_row + 1, amount=extras)
            for offset in range(1, extras + 1):
                for col in _TEMPLATE_CARRY_COLUMNS:
                    src = ws.cell(row=base_row, column=col)
                    dst = ws.cell(row=base_row + offset, column=col)
                    value = src.value
                    if col == 9 and isinstance(value, str):
                        # 剝掉 parent 留下的 `(... → ...)` rewrite 尾巴
                        value = _REWRITE_TAIL_RE.sub("", value).rstrip()
                    dst.value = _sanitize_excel_text(value)
                    _copy_style(src, dst)

        for idx, row_data in enumerate(items):
            _write_tc_row(
                ws,
                row_num=base_row + idx,
                row_data=row_data,
                selected_fields=selected_fields,
                append_rewrite=True,
                rewritten_req_ids=rewritten_req_ids,
            )

        cumulative_offset += extras

    # row_num 不明的退回舊行為：直接寫末尾，不插列。
    if orphans:
        append_start = ws.max_row + 1
        for i, row_data in enumerate(orphans):
            _write_tc_row(
                ws,
                row_num=append_start + i,
                row_data=row_data,
                selected_fields=selected_fields,
                append_rewrite=True,
                rewritten_req_ids=rewritten_req_ids,
            )

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    wb.close()


def write_framework_sheet(
    input_path: str,
    framework_data: list[dict],
    output_path: str,
) -> None:
    """
    Populate the Test Case Framework sheet.

    Columns: A=Test Group, B=Test Set, C=TC Count, D=Req Count.

    TC Count = 匯出到該 test set 的 TC 筆數（一個 req 被拆多筆時每筆各算 1）。
    Req Count = 原始 requirement 數（同一 req 的多筆拆分只算一次）。
    為了向後相容，若 entry 只帶 `req_count` 會當作同時是 TC 與 Req 數。
    """
    wb = load_workbook(input_path)

    if FRAMEWORK_SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(FRAMEWORK_SHEET_NAME)
    ws = wb[FRAMEWORK_SHEET_NAME]

    # Write header
    ws.cell(row=1, column=1, value="Test Group")
    ws.cell(row=1, column=2, value="Test Set")
    ws.cell(row=1, column=3, value="TC Count")
    ws.cell(row=1, column=4, value="Req Count")

    # Write data
    for i, entry in enumerate(framework_data):
        row = i + 2
        ws.cell(row=row, column=1, value=_sanitize_excel_text(entry["test_group"]))
        ws.cell(row=row, column=2, value=_sanitize_excel_text(entry["test_set"]))
        tc_count = entry.get("tc_count", entry.get("req_count", 0))
        req_count = entry.get("req_count", tc_count)
        ws.cell(row=row, column=3, value=tc_count)
        ws.cell(row=row, column=4, value=req_count)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    wb.close()


# Bilingual headers whose keywords match the parser's header resolution, so a
# generated workbook can be fed straight back into --review (closed loop).
_GEN_HEADERS = {
    4: "Requirement or Design ID 需求/設計 ID",
    6: "Test Case ID 測試用例ID",
    7: "Test Group 測試組",
    9: "Test Item 測試項目",
    10: "Pre-Conditions 先前條件",
    11: "Input Test Data 輸入條件",
    12: "Test procedure 測試程序",
    13: "Expected Result 預期結果",
    14: "Specification Reference 規格參考",
    16: "Test Case Priority 測試用例優先級別",
    17: "Test Case Design Methods 測試用例設計方法",
}
_GEN_FIELD_COL = {
    "req_id": 4, "tc_id": 6, "test_group": 7, "test_item": 9,
    "pre_conditions": 10, "input_test_data": 11, "test_procedure": 12,
    "expected_result": 13, "spec_reference": 14, "priority": 16,
    "design_method": 17,
}


def write_generated_tc_workbook(tcs, output_path, project="GEN",
                                test_group="Generated") -> None:
    """Write freshly generated TCs (gen_bridge output) into a NEW workbook using
    the team's header layout. Header row = 9, data from row 10 — re-reviewable."""
    from openpyxl import Workbook
    wb = Workbook()
    ws_pd = wb.active
    ws_pd.title = "Product Document"
    ws_pd.cell(row=3, column=2, value=project)

    ws = wb.create_sheet(TC_SHEET_NAME)
    for col, name in _GEN_HEADERS.items():
        ws.cell(row=9, column=col, value=name)

    for i, tc in enumerate(tcs):
        r = 10 + i
        row = dict(tc)
        row.setdefault("test_group", test_group)
        # test_item carries the requirement句 (Tier 1 anchor); prepend the title.
        title = str(row.get("tc_title") or "").strip()
        item = str(row.get("test_item") or "").strip()
        row["test_item"] = f"{title}\n{item}" if title and item else (item or title)
        for field, col in _GEN_FIELD_COL.items():
            val = row.get(field)
            if val:
                ws.cell(row=r, column=col, value=_sanitize_excel_text(str(val)))

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    wb.close()
