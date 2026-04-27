"""Spec reference matcher (RULES.md §2.4).

Layer 1   : PDM code exact match (regex, deterministic).
Layer 1.5 : Token Jaccard fuzzy match on spec description (offline, deterministic).
Layer 2   : Semantic match via pre-computed embeddings (cosine similarity)。

預處理流程（離線）
------------------
由 ``scripts/build_spec_index.py`` 解析 ``spec-index/cache/*.xlsx``，呼叫
:func:`build_spec_index` 產生 :class:`SpecIndex`，可再以 :func:`attach_embeddings`
批次算 embedding，最後 :func:`save_spec_index` 輸出 ``<basename>.json`` 於同一
資料夾，並經 :func:`update_manifest` 更新 ``manifest.json``。

執行階段
--------
以 :func:`load_spec_index` 從 ``spec-index/cache/`` 依名稱載入一份或多份 spec，
合併成單一 :class:`SpecIndex`，餵給 :func:`match_spec_references`。entries 若
皆帶 ``embedding`` 自動走 cosine；否則 fallback 到 Jaccard。
"""
from __future__ import annotations

import json
import math
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import load_workbook


# Regex for all PDM code patterns
PDM_PATTERN = re.compile(
    r"\b(PDM\d+\.?\d*|PDMS\d+\.?\d*|MPDM\d+\.?\d*|TD\d+\.?\d*"
    r"|DNDS\d+\.?\d*|PDEE\d+\.?\d*|APAC\d+\.?\d*|PSR\d+\.?\d*"
    r"|R1C\d+\.?\d*|CR\d+\.?\d*|ICE\d+\.?\d*|LS\d+\.?\d*|C\d+\.?\d*)\b"
)

# 通用英文 stop words 與無語意短詞
_STOP_WORDS = {
    "the", "and", "or", "but", "if", "then", "when", "where", "while",
    "is", "are", "was", "were", "be", "been", "being", "to", "of", "in",
    "on", "at", "for", "with", "by", "from", "this", "that", "these",
    "those", "it", "its", "as", "an", "a", "not", "no", "can", "will",
    "shall", "should", "must", "may", "do", "does", "did", "has", "have",
    "had", "one", "two", "any", "all", "some", "other", "user", "system",
    "via", "case", "which", "who", "what", "test", "item",
}

DEFAULT_EMBEDDING_MODEL = "text-embedding-3-large"


# ---------------------------------------------------------------------------
# Tokenization & similarity helpers
# ---------------------------------------------------------------------------


def _tokenize(text: str) -> set[str]:
    """切字：小寫、抽取 ≥3 字母的 alphabetic token，去 stop words。"""
    tokens = re.findall(r"[A-Za-z][A-Za-z0-9_]{2,}", (text or "").lower())
    return {t for t in tokens if t not in _STOP_WORDS}


def _jaccard(a: set[str], b: set[str]) -> float:
    """Jaccard similarity。空集合返回 0。"""
    if not a or not b:
        return 0.0
    inter = len(a & b)
    return inter / len(a | b)


def _cosine(a: list[float], b: list[float]) -> float:
    """Cosine similarity。任一向量為零向量或長度不符則返回 0。"""
    if not a or not b or len(a) != len(b):
        return 0.0
    dot = 0.0
    na = 0.0
    nb = 0.0
    for x, y in zip(a, b):
        dot += x * y
        na += x * x
        nb += y * y
    if na == 0.0 or nb == 0.0:
        return 0.0
    return dot / (math.sqrt(na) * math.sqrt(nb))


def _embed_source(outline: object, description: str) -> str:
    """Spec 側 embedding 輸入：outline + description，補足短描述上下文。"""
    parts = []
    if outline:
        parts.append(str(outline).strip())
    if description:
        parts.append(description.strip())
    return " \n".join(p for p in parts if p)


def extract_pdm_codes(text: str) -> list[str]:
    """Extract all PDM-family codes from text."""
    return PDM_PATTERN.findall(text)


# ---------------------------------------------------------------------------
# SpecIndex data structure
# ---------------------------------------------------------------------------


class SpecIndex(dict):
    """Dict mapping PDM code → entry，並夾帶 ``entries`` list 供 fuzzy match。

    繼承 dict 讓舊呼叫 ``index[code]`` / ``code in index`` / ``len(index)`` 維持
    原語意（len 只計 PDM code 數），新功能透過 ``.entries`` 屬性取得全部列。

    Attributes
    ----------
    entries : list[dict]
        每一列 spec 條目；含 ``nrl_id`` / ``outline`` / ``source_id`` /
        ``description`` / ``tokens``，預處理過後再加 ``embed_text`` 與 ``embedding``。
    embedding_model : str | None
        產生 embeddings 所用的模型名稱，None 代表此 index 不含 embedding。
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.entries: list[dict] = []
        self.embedding_model: str | None = None


# ---------------------------------------------------------------------------
# Build / persist / load
# ---------------------------------------------------------------------------


def build_spec_index(sys1_xlsx_path: str) -> SpecIndex:
    """解析 SYS1/HMI 的 xlsx（``Basic Report`` sheet）成 :class:`SpecIndex`。"""
    wb = load_workbook(sys1_xlsx_path, read_only=True, data_only=True)
    ws = wb["Basic Report"]

    index = SpecIndex()
    row_num = 2  # Skip header row
    while True:
        nrl_id = ws.cell(row=row_num, column=1).value
        if not nrl_id:
            break

        outline = ws.cell(row=row_num, column=3).value
        description = ws.cell(row=row_num, column=4).value or ""
        source_id = ws.cell(row=row_num, column=5).value or ""

        entry = {
            "nrl_id": nrl_id,
            "outline": outline,
            "source_id": source_id,
            "description": description,
            "tokens": _tokenize(description),
            "embed_text": _embed_source(outline, description),
        }
        index.entries.append(entry)

        for code in extract_pdm_codes(description):
            index[code] = entry

        row_num += 1

    wb.close()
    return index


def attach_embeddings(
    index: SpecIndex,
    *,
    model: str = DEFAULT_EMBEDDING_MODEL,
    embedder: Callable[[list[str], str], list[list[float]]] | None = None,
) -> SpecIndex:
    """對 index 所有 entries 計算 embedding 並寫入 ``entry["embedding"]``。"""
    if not index.entries:
        return index
    embed_fn = embedder or openai_embed
    texts = [e.get("embed_text") or e.get("description") or "" for e in index.entries]
    vectors = embed_fn(texts, model)
    if len(vectors) != len(index.entries):
        raise ValueError(
            f"embedder returned {len(vectors)} vectors for {len(index.entries)} entries"
        )
    for entry, vec in zip(index.entries, vectors):
        entry["embedding"] = list(vec)
    index.embedding_model = model
    return index


def openai_embed(texts: list[str], model: str = DEFAULT_EMBEDDING_MODEL) -> list[list[float]]:
    """以 OpenAI embeddings API 批次產生向量，空字串補 ``" "`` 佔位。"""
    from openai import OpenAI

    client = OpenAI()
    safe_texts = [t if t and t.strip() else " " for t in texts]
    out: list[list[float]] = []
    batch_size = 128
    for i in range(0, len(safe_texts), batch_size):
        batch = safe_texts[i : i + batch_size]
        resp = client.embeddings.create(model=model, input=batch)
        out.extend(d.embedding for d in resp.data)
    return out


def save_spec_index(
    index: SpecIndex,
    out_path: str | os.PathLike,
    *,
    name: str,
    source_file: str | os.PathLike,
) -> Path:
    """以 JSON 持久化 SpecIndex 到 ``out_path``。"""
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    src = Path(source_file)
    payload = {
        "name": name,
        "source_file": src.name,
        "source_mtime": src.stat().st_mtime if src.exists() else None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "embedding_model": index.embedding_model,
        "entries": [_entry_to_json_safe(entry) for entry in index.entries],
        "codes_map": {
            code: _entry_index(index.entries, entry)
            for code, entry in index.items()
        },
    }
    path.write_text(
        json.dumps(payload, ensure_ascii=True, separators=(",", ":"), sort_keys=True),
        encoding="utf-8",
    )
    return path


def _entry_to_json_safe(entry: dict) -> dict:
    out = dict(entry)
    tokens = out.get("tokens")
    if isinstance(tokens, set):
        out["tokens"] = sorted(tokens)
    return out


def _entry_from_json_safe(entry: object) -> dict:
    if not isinstance(entry, dict):
        raise ValueError("invalid spec index entry: expected object")
    out = dict(entry)
    tokens = out.get("tokens")
    if isinstance(tokens, list):
        out["tokens"] = {str(token) for token in tokens}
    elif tokens is None:
        out["tokens"] = _tokenize(str(out.get("description") or ""))
    elif not isinstance(tokens, set):
        out["tokens"] = {str(tokens)}
    return out


def _entry_index(entries: list[dict], target: dict) -> int:
    for i, e in enumerate(entries):
        if e is target:
            return i
    return -1


def load_spec_index(
    names: Iterable[str] | None = None,
    *,
    index_dir: str | os.PathLike = "spec-index/cache",
) -> SpecIndex:
    """從 ``index_dir`` 載入指定 specs 並合併成單一 SpecIndex。

    Parameters
    ----------
    names
        要載入的 spec basename（不含 ``.json``）。None 代表資料夾內全部 json。
    """
    base = Path(index_dir)
    if not base.exists():
        raise FileNotFoundError(f"Spec index dir not found: {base}")

    if names is None:
        index_files = sorted(base.glob("*.json"))
    else:
        index_files = [base / f"{n}.json" for n in names]
        missing = [p for p in index_files if not p.exists()]
        if missing:
            raise FileNotFoundError(
                "Missing spec index files: " + ", ".join(str(m) for m in missing)
            )

    merged = SpecIndex()
    models_seen: set[str] = set()
    for index_path in index_files:
        try:
            payload = json.loads(index_path.read_text(encoding="utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise ValueError(
                f"Invalid spec index cache {index_path}: expected JSON. "
                "Regenerate it with scripts/build_spec_index.py."
            ) from exc
        if not isinstance(payload, dict):
            raise ValueError(f"Invalid spec index cache {index_path}: top-level JSON must be an object")

        raw_entries = payload.get("entries")
        if not isinstance(raw_entries, list):
            raise ValueError(f"Invalid spec index cache {index_path}: entries must be an array")
        entries = [_entry_from_json_safe(entry) for entry in raw_entries]

        offset = len(merged.entries)
        merged.entries.extend(entries)
        for code, idx in (payload.get("codes_map") or {}).items():
            if isinstance(idx, int) and 0 <= idx < len(entries):
                merged[code] = merged.entries[offset + idx]
        for entry in entries:
            for code in extract_pdm_codes(str(entry.get("description") or "")):
                merged.setdefault(code, entry)

        model = payload.get("embedding_model")
        if model:
            models_seen.add(model)

    if len(models_seen) == 1:
        merged.embedding_model = next(iter(models_seen))
    elif len(models_seen) > 1:
        raise ValueError(
            "Cannot merge spec indices with different embedding models: "
            + ", ".join(sorted(models_seen))
        )
    return merged


def is_index_fresh(xlsx_path: str | os.PathLike, index_path: str | os.PathLike) -> bool:
    """Cache file exists and is not older than the source xlsx."""
    cache = Path(index_path)
    xlsx = Path(xlsx_path)
    if not cache.exists() or not xlsx.exists():
        return False
    return cache.stat().st_mtime >= xlsx.stat().st_mtime


def update_manifest(
    manifest_path: str | os.PathLike,
    *,
    name: str,
    source_file: str,
    entries_count: int,
    embedding_model: str | None,
) -> None:
    """把一筆 spec 資訊寫入 ``manifest.json``（存在則更新、不存在則建立）。"""
    path = Path(manifest_path)
    if path.exists():
        data = json.loads(path.read_text(encoding="utf-8"))
    else:
        data = {"specs": []}

    specs = [s for s in data.get("specs", []) if s.get("name") != name]
    specs.append({
        "name": name,
        "source_file": source_file,
        "entries_count": entries_count,
        "embedding_model": embedding_model,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    })
    data["specs"] = sorted(specs, key=lambda s: s["name"])
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------


# 不同模式採用不同門檻
JACCARD_THRESHOLD = 0.15
COSINE_THRESHOLD = 0.55
FUZZY_THRESHOLD = JACCARD_THRESHOLD  # 向下相容


def _has_embeddings(entries: list[dict]) -> bool:
    return bool(entries) and all("embedding" in e and e["embedding"] for e in entries)


def match_spec_references(
    rows: list[dict],
    spec_index: dict,
    fuzzy_threshold: float | None = None,
    *,
    query_embedder: Callable[[list[str], str], list[list[float]]] | None = None,
) -> list[dict]:
    """
    Match Test Items to spec references.

    Layer 1     ：Test Item 內出現的 PDM code 直接對應 → ``match_type="exact"``。
    Layer 1.5/2 ：若 Layer 1 無結果，對全部 entries 跑相似度比對——
        * entries 帶 embedding → cosine（``match_type="semantic"``）
        * 否則                  → token Jaccard（``match_type="fuzzy"``）

    Parameters
    ----------
    fuzzy_threshold
        None 代表依模式採用預設（Jaccard 0.15 / Cosine 0.55）。
    query_embedder
        semantic 模式下計算 test_item 向量的函式；未提供則走 :func:`openai_embed`。
    """
    def _entry_context(entry: dict) -> str:
        parts = []
        if entry.get("source_id"):
            parts.append(f"Source: {entry['source_id']}")
        if entry.get("outline"):
            parts.append(f"Outline: {entry['outline']}")
        if entry.get("description"):
            parts.append(f"Description: {entry['description']}")
        return " | ".join(str(p) for p in parts if p)

    def _candidate_context(scored_entries: list[tuple[float, dict]], limit: int = 3) -> str:
        parts = []
        for rank, (score, entry) in enumerate(scored_entries[:limit], 1):
            context = _entry_context(entry)
            if context:
                parts.append(
                    f"[Reference candidate {rank} - AI judgement required; "
                    f"score={round(score, 3)}] {context}"
                )
        return "\n".join(parts)

    def _row_query_text(row: dict) -> str:
        parts = [
            row.get("test_set") or row.get("testSet") or row.get("test_set_hint") or row.get("testSetHint") or "",
            row.get("test_item") or row.get("testItem") or "",
            row.get("test_procedure") or row.get("testProcedure") or "",
            row.get("expected_result") or row.get("expectedResult") or "",
        ]
        return " ".join(str(part).strip() for part in parts if str(part or "").strip())

    entries = getattr(spec_index, "entries", None) or []
    codes_map = spec_index

    use_semantic = _has_embeddings(entries) and (
        query_embedder is not None or bool(os.environ.get("OPENAI_API_KEY"))
    )
    threshold = fuzzy_threshold
    if threshold is None:
        threshold = COSINE_THRESHOLD if use_semantic else JACCARD_THRESHOLD

    query_vectors: dict[int, list[float]] = {}
    if use_semantic and rows:
        semantic_indices = []
        semantic_texts = []
        for row_idx, row in enumerate(rows):
            test_item = row.get("test_item", "") or ""
            codes = extract_pdm_codes(test_item)
            if any(code in codes_map for code in codes):
                continue
            semantic_indices.append(row_idx)
            semantic_texts.append(_row_query_text(row))
        if semantic_texts:
            model = getattr(spec_index, "embedding_model", None) or DEFAULT_EMBEDDING_MODEL
            embed_fn = query_embedder or openai_embed
            vectors = embed_fn(semantic_texts, model)
            query_vectors = {
                row_idx: vector
                for row_idx, vector in zip(semantic_indices, vectors)
            }

    results = []
    for idx, row in enumerate(rows):
        test_item = row.get("test_item", "") or ""
        codes = extract_pdm_codes(test_item)
        matched_refs = []
        matched_contexts = []

        for code in codes:
            if code in codes_map:
                entry = codes_map[code]
                matched_refs.append(entry["source_id"])
                context = _entry_context(entry)
                if context:
                    matched_contexts.append(f"[Reference exact: {code}] {context}")

        if matched_refs:
            seen: set[str] = set()
            unique_refs: list[str] = []
            for ref in matched_refs:
                if ref not in seen:
                    seen.add(ref)
                    unique_refs.append(ref)
            results.append({
                **row,
                "spec_reference": "; ".join(unique_refs),
                "match_type": "exact",
                "matched_spec_context": "\n".join(matched_contexts),
            })
            continue

        if entries:
            best_entry = None
            best_score = 0.0
            scored_entries: list[tuple[float, dict]] = []
            if use_semantic and idx in query_vectors:
                qvec = query_vectors[idx]
                for entry in entries:
                    score = _cosine(qvec, entry["embedding"])
                    scored_entries.append((score, entry))
                    if score > best_score:
                        best_score = score
                        best_entry = entry
                match_type = "semantic"
            else:
                item_tokens = _tokenize(_row_query_text(row))
                for entry in entries:
                    score = _jaccard(item_tokens, entry["tokens"])
                    scored_entries.append((score, entry))
                    if score > best_score:
                        best_score = score
                        best_entry = entry
                match_type = "fuzzy"
            scored_entries.sort(key=lambda item: item[0], reverse=True)

            if best_entry and best_score >= threshold:
                results.append({
                    **row,
                    "spec_reference": best_entry["source_id"],
                    "match_type": match_type,
                    "match_score": round(best_score, 3),
                    "matched_spec_context": (
                        f"[Reference {match_type}: score={round(best_score, 3)}] "
                        f"{_entry_context(best_entry)}"
                    ),
                })
                continue

            candidates = _candidate_context(scored_entries)
            result = {
                **row,
                "spec_reference": None,
                "match_type": "unmatched",
            }
            if candidates:
                result["reference_candidate_context"] = candidates
            results.append(result)
            continue

        results.append({
            **row,
            "spec_reference": None,
            "match_type": "unmatched",
        })

    return results
