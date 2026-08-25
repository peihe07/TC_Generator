#!/usr/bin/env python3
"""Zip-level surgical xlsx emit — R16-1/R16-2 remediation.

`openpyxl`'s `Workbook.save()` does not write the file it read; it writes a
file openpyxl can describe. Everything outside its object model is dropped or
regenerated. Measured on the AMFM delivery (R16 §2): 21 zip members lost,
10 added, all six `x14:dataValidation` elements gone — while the row contents
were correct and lint was green, which is exactly why the loss went unseen.

This module keeps openpyxl as the *calculation* layer and removes it from the
*emit* layer:

1. the caller mutates an openpyxl workbook as before
2. `surgical_save` diffs that workbook against a fresh read of the source,
   yielding a cell-level change set
3. the change set is applied to the SOURCE sheet XML as text, and every other
   zip member is copied byte-for-byte

What this preserves that `save()` does not: x14 (extension) data validations,
printer settings, SmartArt under `xl/diagrams/`, legacy VML comment drawings,
embedded media in its original encoding, `sharedStrings.xml`, and the shared
formula groups of untouched rows.

Deliberately NOT preserved, because the caller changed them: the target
sheets' XML. `verify_structure` enforces that the difference stops there —
zip member set equal, data-validation counts equal per sheet, only the
patched sheets' XML allowed to differ. A violation is canon §0 item 3
(invariant breach) and aborts; it is never downgraded to a warning.

Usage (see features/amfm/scripts/write_back.py):

    wb = openpyxl.load_workbook(src)
    ...mutate...
    report = surgical_save(wb, src, out)      # raises StructureError on drift
"""
from __future__ import annotations

import re
import shutil
import zipfile
from pathlib import Path
from xml.sax.saxutils import escape

# Cell reference: column letters + row digits.
CELL_RE = re.compile(r'<c r="([A-Z]+)(\d+)"([^>]*?)(?:/>|>(.*?)</c>)', re.S)
ROW_RE = re.compile(r'<row r="(\d+)"([^>]*?)(?:/>|>(.*?)</row>)', re.S)
DIMENSION_RE = re.compile(r'<dimension ref="[^"]*"/>')
CLASSIC_DV_RE = re.compile(r"<dataValidation[ >]")
X14_DV_RE = re.compile(r"<x14:dataValidation[ >]")


class StructureError(RuntimeError):
    """A structural invariant failed. Not recoverable by relaxing the check."""


# ------------------------------------------------------------ column helpers

def col_to_idx(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def idx_to_col(idx: int) -> str:
    s, n = "", idx
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


# ------------------------------------------------------------ sheet ↔ member

def sheet_members(path: Path) -> dict[str, str]:
    """Worksheet name -> zip member path, resolved through workbook rels.

    Resolved rather than assumed: sheet order in `workbook.xml` is display
    order, which is not the `sheetN.xml` numbering. The AMFM workbook happens
    to agree; relying on that would be luck.
    """
    with zipfile.ZipFile(path) as z:
        book = z.read("xl/workbook.xml").decode("utf-8")
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    target = {}
    for rel in re.finditer(r"<Relationship\b([^>]*)/>", rels):
        attrs = rel.group(1)
        rid = re.search(r'Id="([^"]+)"', attrs)
        tgt = re.search(r'Target="([^"]+)"', attrs)
        if rid and tgt:
            target[rid.group(1)] = tgt.group(1)
    out = {}
    for m in re.finditer(r"<sheet\b([^>]*?)/?>", book):
        attrs = m.group(1)
        name = re.search(r'name="([^"]*)"', attrs)
        rid = re.search(r'r:id="([^"]*)"', attrs)
        if not (name and rid and rid.group(1) in target):
            continue
        tgt = target[rid.group(1)].lstrip("/")
        if not tgt.startswith("xl/"):
            tgt = "xl/" + tgt
        out[_unescape(name.group(1))] = tgt
    return out


def _unescape(s: str) -> str:
    return (s.replace("&amp;", "&").replace("&lt;", "<")
             .replace("&gt;", ">").replace("&quot;", '"'))


# ------------------------------------------------------------------- diffing

def diff_cells(original, mutated) -> dict[str, dict[tuple[int, int], object]]:
    """{sheet name: {(row, col): new value}} for every cell that changed.

    Compares openpyxl values, so a formula compares as its string. Shared
    formulas expand on read, which is what makes an untouched shared-formula
    row compare equal and therefore stay untouched in the XML.
    """
    changes: dict[str, dict[tuple[int, int], object]] = {}
    for name in mutated.sheetnames:
        if name not in original.sheetnames:
            raise StructureError(f"sheet {name!r} is new; surgical emit only "
                                 "patches sheets that exist in the source")
        old, new = original[name], mutated[name]
        max_row = max(old.max_row, new.max_row)
        max_col = max(old.max_column, new.max_column)
        sheet_changes: dict[tuple[int, int], object] = {}
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                a = old.cell(r, c).value if r <= old.max_row and c <= old.max_column else None
                b = new.cell(r, c).value
                if a != b:
                    sheet_changes[(r, c)] = b
        if sheet_changes:
            changes[name] = sheet_changes
    return changes


# ------------------------------------------------------------- serialisation

def _cell_xml(coord: str, style_attr: str, value) -> str:
    """One `<c>` element. `style_attr` carries the source cell's `s="…"`.

    Strings go out as inline strings: the source `sharedStrings.xml` is copied
    verbatim, so appending to it would mean rewriting it — the one member the
    surgical path most wants to leave alone.
    """
    if value is None or value == "":
        return f'<c r="{coord}"{style_attr}/>'
    if isinstance(value, str) and value.startswith("="):
        return f'<c r="{coord}"{style_attr}><f>{escape(value[1:])}</f></c>'
    if isinstance(value, bool):
        return f'<c r="{coord}"{style_attr} t="b"><v>{int(value)}</v></c>'
    if isinstance(value, (int, float)):
        return f'<c r="{coord}"{style_attr}><v>{value}</v></c>'
    text = escape(str(value))
    return (f'<c r="{coord}"{style_attr} t="inlineStr">'
            f'<is><t xml:space="preserve">{text}</t></is></c>')


def _patch_row(row_xml: str, row_num: int,
               edits: dict[int, object]) -> str:
    """Rebuild one `<row>`'s cells, keeping every untouched cell verbatim."""
    cells: dict[int, str] = {}
    styles: dict[int, str] = {}
    for m in CELL_RE.finditer(row_xml):
        idx = col_to_idx(m.group(1))
        cells[idx] = m.group(0)
        s = re.search(r'\ss="\d+"', m.group(3) or "")
        styles[idx] = s.group(0) if s else ""
    for col, value in edits.items():
        cells[col] = _cell_xml(f"{idx_to_col(col)}{row_num}",
                               styles.get(col, ""), value)
    return "".join(cells[k] for k in sorted(cells))


def _new_row_xml(row_num: int, edits: dict[int, object]) -> str:
    body = "".join(_cell_xml(f"{idx_to_col(c)}{row_num}", "", v)
                   for c, v in sorted(edits.items()))
    return f'<row r="{row_num}">{body}</row>'


def patch_sheet_xml(xml: str, edits: dict[tuple[int, int], object]) -> str:
    """Apply a cell change set to one worksheet's XML."""
    by_row: dict[int, dict[int, object]] = {}
    for (r, c), v in edits.items():
        by_row.setdefault(r, {})[c] = v

    existing: dict[int, tuple[int, int, str, str]] = {}
    for m in ROW_RE.finditer(xml):
        existing[int(m.group(1))] = (m.start(), m.end(), m.group(2) or "",
                                     m.group(3) or "")

    # Rewrite existing rows from the end, so earlier offsets stay valid.
    out = xml
    for row_num in sorted(set(by_row) & set(existing), reverse=True):
        start, end, attrs, body = existing[row_num]
        out = (out[:start]
               + f'<row r="{row_num}"{attrs}>'
               + _patch_row(body, row_num, by_row[row_num])
               + "</row>"
               + out[end:])

    # Rows absent from the source are appended in order before </sheetData>.
    added = sorted(set(by_row) - set(existing))
    if added:
        highest = max(existing) if existing else 0
        if any(r < highest for r in added):
            raise StructureError(
                "surgical emit can only append rows past the last source row; "
                f"asked to insert {[r for r in added if r < highest][:5]} "
                f"before row {highest}. Inserting mid-sheet would shift every "
                "row below it, which this path deliberately cannot do")
        block = "".join(_new_row_xml(r, by_row[r]) for r in added)
        if "</sheetData>" in out:
            out = out.replace("</sheetData>", block + "</sheetData>", 1)
        else:
            out = out.replace("<sheetData/>", f"<sheetData>{block}</sheetData>", 1)

    return _fix_dimension(out)


def _fix_dimension(xml: str) -> str:
    """Restate `<dimension>` over the patched sheet's actual extent."""
    max_row, max_col = 0, 0
    for m in ROW_RE.finditer(xml):
        max_row = max(max_row, int(m.group(1)))
        for cm in CELL_RE.finditer(m.group(3) or ""):
            max_col = max(max_col, col_to_idx(cm.group(1)))
    if not max_row:
        return xml
    ref = f"A1:{idx_to_col(max_col or 1)}{max_row}"
    return DIMENSION_RE.sub(f'<dimension ref="{ref}"/>', xml, count=1)


# -------------------------------------------------------------- verification

def _dv_counts(path: Path) -> dict[str, tuple[int, int]]:
    counts = {}
    with zipfile.ZipFile(path) as z:
        for member in z.namelist():
            if member.startswith("xl/worksheets/sheet"):
                xml = z.read(member).decode("utf-8")
                counts[member] = (len(CLASSIC_DV_RE.findall(xml)),
                                  len(X14_DV_RE.findall(xml)))
    return counts


def verify_structure(src: Path, out: Path, patched: set[str]) -> dict:
    """R16 §3.2 invariant. Raises `StructureError`; never warns."""
    with zipfile.ZipFile(src) as a, zipfile.ZipFile(out) as b:
        src_members, out_members = set(a.namelist()), set(b.namelist())
    lost, added = sorted(src_members - out_members), sorted(out_members - src_members)
    if lost or added:
        raise StructureError(
            f"zip member set changed — lost {lost}, added {added}. "
            "The delivered file must carry every part the customer's file "
            "carried (R16-1)")

    before, after = _dv_counts(src), _dv_counts(out)
    bad = {m: (before[m], after[m]) for m in before if before[m] != after[m]}
    if bad:
        raise StructureError(
            f"data-validation counts changed (classic, x14): {bad}. "
            "The dropdowns are part of the controlled form (R16-2)")

    with zipfile.ZipFile(src) as a, zipfile.ZipFile(out) as b:
        differing = sorted(m for m in src_members
                           if a.read(m) != b.read(m))
    unexpected = [m for m in differing if m not in patched]
    if unexpected:
        raise StructureError(
            f"members differ that were not written: {unexpected}. "
            "Only the target sheets' XML may change")

    return {"members": len(src_members), "differing": differing,
            "dv_counts": {m: after[m] for m in sorted(after) if any(after[m])}}


# --------------------------------------------------------------------- emit

def surgical_save(mutated, src: Path, out: Path, *, verify: bool = True) -> dict:
    """Write `mutated`'s cell changes into a byte-for-byte copy of `src`."""
    import openpyxl

    src, out = Path(src), Path(out)
    original = openpyxl.load_workbook(src)
    changes = diff_cells(original, mutated)
    members = sheet_members(src)

    patched: dict[str, str] = {}
    for name, edits in changes.items():
        member = members.get(name)
        if member is None:
            raise StructureError(f"no zip member resolved for sheet {name!r}")
        with zipfile.ZipFile(src) as z:
            xml = z.read(member).decode("utf-8")
        patched[member] = patch_sheet_xml(xml, edits)

    out.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(
            out, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = (patched[info.filename].encode("utf-8")
                    if info.filename in patched else zin.read(info.filename))
            zout.writestr(info, data)

    report = {"sheets_patched": {n: len(e) for n, e in changes.items()},
              "members_patched": sorted(patched)}
    if verify:
        report.update(verify_structure(src, out, set(patched)))
    return report


def copy_unchanged(src: Path, out: Path) -> None:
    """Byte-for-byte copy, for the no-edits case."""
    shutil.copy(src, out)


# ------------------------------------------------- structural row insertion
#
# `surgical_save` is an append-only cell patcher: it refuses to insert a row
# mid-sheet, because shifting every row below it is not a cell change set (see
# the StructureError in `patch_sheet_xml`). Row insertion is therefore a
# separate, purely *structural* emit that runs BEFORE the content pass:
#
#     expanded = surgical_insert_rows(src, tmp, {11: 3, 12: 2, ...})
#     wb = openpyxl.load_workbook(tmp); ...mutate...
#     surgical_save(wb, tmp, out)
#
# It moves rows and the sheet's row-addressed references; it writes no values.
# Everything `surgical_save` preserves is preserved here for the same reason —
# every zip member but the target sheet's XML is copied byte-for-byte — and the
# same `verify_structure` invariant is asserted on the result.

ROW_ATTR_RE = re.compile(r'(<row r=")(\d+)(")')
CELL_REF_RE = re.compile(r'(<c r=")([A-Z]+)(\d+)(")')
A1_RE = re.compile(r"\$?([A-Z]+)\$?(\d+)")
SHEETDATA_RE = re.compile(r"(<sheetData[^>]*>)(.*)(</sheetData>)", re.S)


def build_shift(insertions: dict[int, int]):
    """Return `S(row) -> new row`, the offset a set of insertions imposes.

    Rows inserted "after row k" land at `S(k)+1 … S(k)+n`, so a source row `r`
    moves down by the number of rows inserted at anchors strictly before it.
    """
    anchors = sorted(insertions)

    def shift(row: int) -> int:
        return row + sum(insertions[a] for a in anchors if a < row)

    return shift


def _shift_ref_token(token: str, shift) -> str:
    """Shift one A1 or A1:B2 reference. Non-references pass through."""
    parts = token.split(":")
    out = []
    for part in parts:
        m = A1_RE.fullmatch(part)
        if not m:
            return token
        out.append(f"{m.group(1)}{shift(int(m.group(2)))}")
    return ":".join(out)


def shift_ref_list(refs: str, shift) -> str:
    """Shift a space-separated `sqref`/`ref` list (`P10:P221 Q10:Q11`)."""
    return " ".join(_shift_ref_token(t, shift) for t in refs.split())


def _clone_row(anchor_xml: str, anchor_attrs: str, row_num: int) -> str:
    """A blank row carrying the anchor row's height, format and cell styles.

    Cloning the styles is the point: `_new_row_xml` emits `<c>` with no `s=`,
    which in this form means no border, no wrap and the default row height —
    a visibly broken row in the delivered workbook.
    """
    attrs = ROW_ATTR_RE.sub("", anchor_attrs)
    cells = []
    for m in CELL_RE.finditer(anchor_xml):
        style = re.search(r'\ss="\d+"', m.group(3) or "")
        cells.append(f'<c r="{m.group(1)}{row_num}"'
                     f'{style.group(0) if style else ""}/>')
    return f'<row r="{row_num}"{attrs}>{"".join(cells)}</row>'


def insert_rows_xml(xml: str, insertions: dict[int, int]) -> str:
    """Insert blank styled rows into one worksheet's XML and shift its refs."""
    block = SHEETDATA_RE.search(xml)
    if not block:
        raise StructureError("sheet has no <sheetData> to insert into")

    rows = list(ROW_RE.finditer(block.group(2)))
    present = [int(m.group(1)) for m in rows]
    if present != sorted(present):
        raise StructureError(f"rows are not in ascending document order: "
                             f"{present[:10]}")
    missing = sorted(set(insertions) - set(present))
    if missing:
        raise StructureError(f"insertion anchors absent from the sheet: "
                             f"{missing}. An anchor must be an existing row, "
                             "so the inserted rows can clone its format")

    shift, out, cum = build_shift(insertions), [], 0
    for m in rows:
        old = int(m.group(1))
        new = old + cum
        body = m.group(3) or ""
        attrs = m.group(2) or ""
        renumbered = CELL_REF_RE.sub(
            lambda c: f"{c.group(1)}{c.group(2)}{new}{c.group(4)}", body)
        out.append(f'<row r="{new}"{attrs}>{renumbered}</row>'
                   if body else f'<row r="{new}"{attrs}/>')
        for i in range(insertions.get(old, 0)):
            out.append(_clone_row(body, attrs, new + 1 + i))
        cum += insertions.get(old, 0)

    patched = (xml[:block.start(2)] + "".join(out) + xml[block.end(2):])

    # Row-addressed references outside <sheetData>. Every one of these moves
    # with its rows; leaving any behind silently detaches a dropdown, a filter
    # or a colour scale from the data it belongs to.
    def sub_attr(pattern: str, text: str) -> str:
        return re.sub(
            pattern,
            lambda m: m.group(1) + shift_ref_list(m.group(2), shift) + m.group(3),
            text)

    for pat in (r'(<autoFilter ref=")([^"]*)(")',
                r'(<mergeCell ref=")([^"]*)(")',
                r'(<conditionalFormatting sqref=")([^"]*)(")',
                r'(<dataValidation\b[^>]*?\ssqref=")([^"]*)(")',
                r'(<xm:sqref>)([^<]*)(</xm:sqref>)',
                r'(<selection[^>]*?\ssqref=")([^"]*)(")'):
        patched = sub_attr(pat, patched)

    return _fix_dimension(patched)


def surgical_insert_rows(src: Path, out: Path, insertions: dict[int, int],
                         sheet: str, *, verify: bool = True) -> dict:
    """Insert blank styled rows into `sheet` of `src`, emitting to `out`.

    `insertions` maps an existing row number to how many rows to insert
    immediately after it. Values are written by a later `surgical_save` pass.
    """
    src, out = Path(src), Path(out)
    if not insertions:
        copy_unchanged(src, out)
        return {"inserted": 0, "differing": []}

    member = sheet_members(src).get(sheet)
    if member is None:
        raise StructureError(f"no zip member resolved for sheet {sheet!r}")
    with zipfile.ZipFile(src) as z:
        xml = z.read(member).decode("utf-8")
    patched = insert_rows_xml(xml, insertions)

    out.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(
            out, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = (patched.encode("utf-8") if info.filename == member
                    else zin.read(info.filename))
            zout.writestr(info, data)

    report = {"inserted": sum(insertions.values()),
              "anchors": len(insertions), "member": member}
    if verify:
        report.update(verify_structure(src, out, {member}))
    return report


# ------------------------------------------------------- surgical restyling
#
# `surgical_save` is a *value* patcher: `diff_cells` compares openpyxl values
# and `_cell_xml` copies the source cell's `s="…"` through untouched. That is
# deliberate — it is what keeps every style, and the x14 validation extension
# that rides alongside them, byte-for-byte intact.
#
# Alignment therefore has no channel at all on that path. `surgical_restyle`
# adds one, and adds it the same way: it never edits an existing `<xf>` (that
# would silently restyle every other cell sharing the id). It *derives* a new
# cellXfs entry from the one a cell already uses, appends it, and repoints only
# the named cells at it. Cells not named in the plan keep their id, so the
# blast radius is exactly the plan.

STYLES_MEMBER = "xl/styles.xml"
CELLXFS_RE = re.compile(r'(<cellXfs count=")(\d+)(">)(.*?)(</cellXfs>)', re.S)
XF_RE = re.compile(r"<xf\b(?:[^>]*?/>|.*?</xf>)", re.S)
ALIGNMENT_RE = re.compile(r"<alignment\b[^>]*?/>", re.S)


def _derive_xf(xf: str, override: dict) -> str:
    """`xf` with its `<alignment>` attributes overridden by `override`.

    Attributes the caller does not name are carried over from the source xf,
    so overriding `horizontal` never silently drops `wrapText`.
    """
    m = ALIGNMENT_RE.search(xf)
    attrs: dict[str, str] = {}
    if m:
        attrs = dict(re.findall(r'(\w+)="([^"]*)"', m.group(0)))
    attrs.update({k: str(v) for k, v in override.items()})
    align = "<alignment " + " ".join(
        f'{k}="{v}"' for k, v in attrs.items()) + "/>"

    if m:
        body = xf[:m.start()] + align + xf[m.end():]
    elif xf.endswith("/>"):
        body = xf[:-2] + ">" + align + "</xf>"
    else:
        body = xf.replace("</xf>", align + "</xf>")
    if "applyAlignment=" not in body:
        body = body.replace("<xf ", '<xf applyAlignment="1" ', 1)
    return body


def surgical_restyle(src: Path, out: Path, plan: dict, *,
                     verify: bool = True) -> dict:
    """Repoint the named cells at derived styles carrying `override`.

    `plan` is `{sheet name: (cells, override)}` where `cells` is an iterable of
    `(row, col)` 1-based pairs and `override` names `<alignment>` attributes,
    e.g. `{"horizontal": "left", "vertical": "top"}`.

    Every zip member but `xl/styles.xml` and the named sheets is copied
    byte-for-byte, so `sharedStrings.xml` — and with it the cell text — is
    provably untouched by this pass.
    """
    import openpyxl

    src, out = Path(src), Path(out)
    members = sheet_members(src)
    with zipfile.ZipFile(src) as z:
        styles = z.read(STYLES_MEMBER).decode("utf-8")
        sheet_xml = {}
        for name in plan:
            member = members.get(name)
            if member is None:
                raise StructureError(f"no zip member resolved for sheet {name!r}")
            sheet_xml[name] = (member, z.read(member).decode("utf-8"))

    cm = CELLXFS_RE.search(styles)
    if cm is None:
        raise StructureError("no <cellXfs> block in styles.xml")
    xfs = XF_RE.findall(cm.group(4))
    if len(xfs) != int(cm.group(2)):
        raise StructureError(
            f"cellXfs count={cm.group(2)} but {len(xfs)} <xf> parsed — refusing "
            "to append against a miscounted table")

    derived: dict[tuple[int, str], int] = {}   # (source id, override key) → new id
    appended: list[str] = []
    patched: dict[str, str] = {}
    repoints = 0

    for name, (member, xml) in sheet_xml.items():
        cells, override = plan[name]
        want = {(int(r), int(c)) for r, c in cells}
        key = repr(sorted(override.items()))

        def repoint(mo, _want=want, _key=key, _override=override):
            nonlocal repoints
            col, row, attrs = mo.group(1), int(mo.group(2)), mo.group(3)
            ci = openpyxl.utils.column_index_from_string(col)
            if (row, ci) not in _want:
                return mo.group(0)
            sm = re.search(r's="(\d+)"', attrs)
            sid = int(sm.group(1)) if sm else 0
            new = derived.get((sid, _key))
            if new is None:
                new = len(xfs) + len(appended)
                appended.append(_derive_xf(xfs[sid], _override))
                derived[(sid, _key)] = new
            repoints += 1
            if sm:
                attrs = attrs[:sm.start()] + f's="{new}"' + attrs[sm.end():]
            else:
                attrs = f' s="{new}"' + attrs
            # The pattern ends at a lookahead, so group(0) is exactly the
            # opening `<c r="…" …` — rebuild it and let the tail stand.
            return f'<c r="{col}{row}"{attrs}'

        patched[member] = re.sub(r'<c r="([A-Z]+)(\d+)"([^>]*?)(?=/>|>)',
                                 repoint, xml)

    if appended:
        total = len(xfs) + len(appended)
        styles = (styles[:cm.start()]
                  + f'<cellXfs count="{total}">'
                  + cm.group(4) + "".join(appended)
                  + "</cellXfs>" + styles[cm.end():])
        patched[STYLES_MEMBER] = styles

    out.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(
            out, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = (patched[info.filename].encode("utf-8")
                    if info.filename in patched else zin.read(info.filename))
            zout.writestr(info, data)

    report = {"cells_repointed": repoints, "xfs_appended": len(appended),
              "members_patched": sorted(patched)}
    if verify:
        # styles.xml is in the allowed set here — that widening IS this
        # channel. The invariants that matter are asserted separately below:
        # the x14/classic validation counts (inside verify_structure) and the
        # text store, which this pass must never touch.
        report.update(verify_structure(src, out, set(patched)))
        with zipfile.ZipFile(src) as a, zipfile.ZipFile(out) as b:
            for m in ("xl/sharedStrings.xml",):
                if m in a.namelist() and a.read(m) != b.read(m):
                    raise StructureError(
                        f"{m} changed during a restyle pass — alignment must "
                        "not touch cell text")
    return report
