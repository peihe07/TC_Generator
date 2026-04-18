"""inspect_workbook tool：預覽 .xlsx 結構但不寫 job_store。

Agent 用來快速回答「這個檔多大」、「有哪些 sheet」、「可能會 parse 到幾 row」
等問題，不會真的完整跑 parser。
"""

from __future__ import annotations

import os
import re
from pathlib import Path

from openpyxl import load_workbook

from .errors import ToolError
from .registry import SafetyLevel, ToolSpec, register_tool
from .schemas import INSPECT_WORKBOOK_SCHEMA


_TC_SHEET_NAMES = {
    "Test Case Specification&Result",
    "Test Case Specification & Result",
    "Test Case Specification",
}
_FILENAME_GROUP_RE = re.compile(r"_SWQT_(?P<group>[^_]+)_\d{8}", re.IGNORECASE)


def _detect_test_group(filename: str) -> str | None:
    match = _FILENAME_GROUP_RE.search(filename or "")
    return match.group("group") if match else None


def inspect_workbook_tool(*, path: str) -> dict:
    """輕量預覽：只讀 sheet 名稱 + 估算 row 數，不解析欄位內容。

    Returns:
        `{"path", "fileName", "fileSizeBytes", "sheets", "testGroup",
          "rowsEstimate", "hasTcSheet"}`
    """
    p = Path(path)
    if not p.is_file():
        raise ToolError(f"file not found: {path}", code="not_found")

    ext = p.suffix.lower()
    if ext not in {".xlsx", ".xlsm"}:
        raise ToolError(
            f"unsupported extension {ext!r}; need .xlsx or .xlsm",
            code="bad_request",
        )

    size = p.stat().st_size

    try:
        wb = load_workbook(str(p), read_only=True, data_only=True)
    except Exception as exc:
        raise ToolError(f"failed to open workbook: {exc}", code="unprocessable") from exc

    sheets = list(wb.sheetnames)
    tc_sheet_name = next((s for s in sheets if s in _TC_SHEET_NAMES), None)

    rows_estimate = 0
    if tc_sheet_name:
        ws = wb[tc_sheet_name]
        # Row 9 是 header；從 10 開始算到第一個空 req_id (col D)
        for row in ws.iter_rows(min_row=10, max_col=4, values_only=True):
            if not row or not row[-1]:
                continue
            rows_estimate += 1

    wb.close()

    return {
        "path": str(p),
        "fileName": p.name,
        "fileSizeBytes": size,
        "sheets": sheets,
        "testGroup": _detect_test_group(p.name),
        "rowsEstimate": rows_estimate,
        "hasTcSheet": tc_sheet_name is not None,
    }


register_tool(
    ToolSpec(
        name="inspect_workbook",
        func=inspect_workbook_tool,
        description=INSPECT_WORKBOOK_SCHEMA["description"],
        safety=SafetyLevel.READ_ONLY,
        input_schema=INSPECT_WORKBOOK_SCHEMA,
    )
)
