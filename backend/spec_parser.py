"""Spec document parser for supplementary context (RULES.md §14.5, §14.6).

Supports PDF, DOCX, XLSX formats. Extracts text segments keyed by PDM codes.
"""
import os

from openpyxl import load_workbook
from docx import Document

from spec_matcher import extract_pdm_codes


def detect_format(filepath: str) -> str | None:
    """Detect file format from extension."""
    ext = os.path.splitext(filepath)[1].lower()
    mapping = {".pdf": "pdf", ".docx": "docx", ".xlsx": "xlsx"}
    return mapping.get(ext)


def parse_pdf(filepath: str) -> dict:
    """
    Parse PDF file using pdfplumber.

    Returns:
        dict mapping PDM code -> {page, text}
    """
    import pdfplumber

    index = {}
    with pdfplumber.open(filepath) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            codes = extract_pdm_codes(text)
            for code in codes:
                if code not in index:
                    index[code] = {"page": page_num, "text": text}
    return index


def parse_docx(filepath: str) -> dict:
    """
    Parse DOCX file, split by heading styles.

    Returns:
        dict mapping PDM code -> {heading, text}
    """
    doc = Document(filepath)
    index = {}
    current_heading = ""

    for para in doc.paragraphs:
        if para.style.name.startswith("Heading"):
            current_heading = para.text
            continue

        text = para.text
        if not text:
            continue

        codes = extract_pdm_codes(text)
        for code in codes:
            if code not in index:
                index[code] = {"heading": current_heading, "text": text}

    return index


def parse_xlsx(filepath: str) -> dict:
    """
    Parse XLSX spec file, treat each row as a requirement entry.

    Returns:
        dict mapping PDM code -> {row, text}
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    index = {}

    row_num = 2  # Skip header
    while True:
        first_col = ws.cell(row=row_num, column=1).value
        if not first_col:
            break

        # Concatenate all columns for text extraction
        cells = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row_num, column=col).value
            if val:
                cells.append(str(val))
        text = " ".join(cells)

        codes = extract_pdm_codes(text)
        for code in codes:
            if code not in index:
                index[code] = {"row": row_num, "text": text}

        row_num += 1

    wb.close()
    return index


def merge_indexes(
    sys1_index: dict | None,
    slot_c_index: dict | None,
) -> dict:
    """
    Merge SYS1 index (Slot B) with Slot C index.

    SYS1 provides: nrl_id, source_id, outline (for Col N traceability).
    Slot C provides: full requirement text (for AI prompt context).

    Returns:
        dict mapping PDM code -> {nrl_id, source_id, outline, full_text}
    """
    if not sys1_index and not slot_c_index:
        return {}

    all_codes = set()
    if sys1_index:
        all_codes.update(sys1_index.keys())
    if slot_c_index:
        all_codes.update(slot_c_index.keys())

    merged = {}
    for code in all_codes:
        sys1 = sys1_index.get(code, {}) if sys1_index else {}
        slot_c = slot_c_index.get(code, {}) if slot_c_index else {}

        merged[code] = {
            "nrl_id": sys1.get("nrl_id"),
            "source_id": sys1.get("source_id"),
            "outline": sys1.get("outline"),
            "full_text": slot_c.get("text"),
        }

    return merged
