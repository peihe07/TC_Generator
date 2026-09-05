"""Excel parser for TC Specification xlsx files (RULES.md §10)."""
import os
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Column index mapping (1-based for openpyxl)
COL_MAP = {
    "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8,
    "I": 9, "J": 10, "K": 11, "L": 12, "M": 13, "N": 14,
    "O": 15, "P": 16, "Q": 17,
}

# Columns to read from each data row (legacy fixed layout — used as fallback
# when a field's header text cannot be located dynamically).
READ_COLUMNS = {
    "D": "req_id",
    "F": "tc_id",
    "G": "test_group",
    "H": "test_set",
    "I": "test_item",
    "J": "pre_conditions",
    "K": "input_test_data",
    "L": "test_procedure",
    "M": "expected_result",
    "N": "spec_reference",
    "P": "priority",
    "Q": "design_method",
}

# field_name -> default fallback column letter (inverse of READ_COLUMNS).
_FALLBACK_LETTERS = {field: letter for letter, field in READ_COLUMNS.items()}

# Header-keyword resolution. Each field is located by the stable English text in
# the header row, tolerant of inserted/shifted columns (e.g. an "Estimated Test
# Time" column pushing "Design Methods" from Q to R). `must` substrings must ALL
# be present; `must_not` substrings disambiguate twin columns (Polarion req-id
# vs the working req-id; TestRail tc-id vs the working tc-id). All lowercased.
HEADER_PATTERNS = {
    "req_id":          (["requirement or design id"], ["polarion"]),
    "tc_id":           (["test case id"], ["testrail"]),
    "test_group":      (["test group"], []),
    "test_set":        (["test set"], []),
    "test_item":       (["test item"], []),
    "pre_conditions":  (["pre-condition"], []),
    "input_test_data": (["input test data"], []),
    "test_procedure":  (["test procedure"], []),
    "expected_result": (["expected result"], []),
    "spec_reference":  (["specification reference"], []),
    "priority":        (["priority"], []),
    "design_method":   (["design", "method"], []),
}


def _resolve_columns(header_values: list) -> dict:
    """Map each field name -> 1-based column index using the header row text,
    falling back to the legacy fixed letter when no header matches."""
    lowered = [(i, str(v).lower()) for i, v in enumerate(header_values) if v]
    resolved = {}
    for field, (must, must_not) in HEADER_PATTERNS.items():
        col = None
        for i, h in lowered:
            if all(m in h for m in must) and not any(x in h for x in must_not):
                col = i + 1  # iter_rows index is 0-based; openpyxl cols are 1-based
                break
        if col is None:
            col = COL_MAP[_FALLBACK_LETTERS[field]]
        resolved[field] = col
    return resolved

HEADER_ROW = 9
DATA_START_ROW = 10
# 真實檔案中 sheet 名稱可能帶中文副標（如 "Product Document 記錄封面頁"），
# 因此用 prefix 比對。
#
# TC 分頁名有**兩個真實變體**（R-G48(b)）：母本一系帶中文副標，power／audio_mgmt
# 一系為 "&Result"。全 repo 145 本 FW036 工作簿實測 121 : 24
# （docs/reports/tc_sheetname_census_20260905.tsv）。兩者皆非誤植，不得擇一硬編。
# **本處為 backend 之唯一定義**，其餘模組一律自此 import（GC-03 §二-5-2）。
TC_SHEET_NAME = "Test Case Specification&Result"   # 僅用於「建立新分頁」時之預設名
TC_SHEET_NAME_CANDIDATES = (
    "Test Case Specification 測試用例規範",
    "Test Case Specification&Result",
    "Test Case Specification & Result",
)
TC_SHEET_PREFIX = "Test Case Specification"
PRODUCT_SHEET_PREFIX = "Product Document"


def resolve_tc_sheet(wb, preferred: str | None = None) -> str:
    """解出工作簿之 TC 分頁名。

    優先序（R-G48(b)）：feature.yaml 之 ``workbook.sheet``（由呼叫端傳入）→
    兩變體之候選集 → prefix 比對。皆不中則拋錯並列出實際 sheetnames ——
    **不得靜默回退到某個預設名**，否則會讀錯或寫錯分頁而看似成功。
    """
    names = list(wb.sheetnames)
    if preferred and preferred in names:
        return preferred
    for candidate in TC_SHEET_NAME_CANDIDATES:
        if candidate in names:
            return candidate
    hit = _find_sheet(wb, TC_SHEET_PREFIX)
    if hit:
        return hit
    raise KeyError(
        f"TC sheet not found (tried {preferred!r}, {TC_SHEET_NAME_CANDIDATES}, "
        f"prefix {TC_SHEET_PREFIX!r}); workbook sheets are: {names}"
    )


def _find_sheet(wb, prefix: str) -> str | None:
    """找出第一個名稱以 prefix 開頭的 sheet（case-insensitive、容忍前後空白）。"""
    p = prefix.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower().startswith(p):
            return name
    return None


def parse_test_group_from_filename(filename: str) -> str | None:
    """
    Extract Test Group from filename pattern: ..._SWQT_{TestGroup}...

    只要檔名包含 `_SWQT_<group>`（group 是下一個 `_` 之前、長度 ≥1、非空）
    就能抽出 group。group 後面接什麼（日期、`smoke`、`debug`、`拷貝`、副檔名）
    都不影響抽取。
    """
    basename = os.path.basename(filename)
    match = re.search(r"_SWQT_([^_./\\]+)", basename)
    return match.group(1) if match else None


def parse_tc_xlsx(filepath: str) -> dict:
    """
    Parse a TC Specification xlsx file.

    Returns dict with:
        - project: str (from Product Document sheet)
        - test_group: str (from filename)
        - row_count: int
        - rows: list[dict] (each row's data)
        - column_fill_status: dict[str, int] (count of non-empty cells per column)
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")

    wb = load_workbook(filepath, read_only=True, data_only=True)

    # Extract project name from Product Document sheet — 容忍中英雙語副標
    project = None
    product_sheet = _find_sheet(wb, PRODUCT_SHEET_PREFIX)
    if product_sheet:
        ws_pd = wb[product_sheet]
        project = ws_pd.cell(row=3, column=2).value

    # Extract test group from filename
    test_group = parse_test_group_from_filename(filepath)

    # Parse TC data rows
    rows = []
    column_fill_status = {}

    try:
        tc_sheet = resolve_tc_sheet(wb)
    except KeyError:
        tc_sheet = None          # 無 TC 分頁之簿仍可解出封面資訊，維持原有公開形狀
    if tc_sheet:
        ws_tc = wb[tc_sheet]

        # Resolve each field's column from the header row (dynamic, layout-safe).
        header_values = next(
            ws_tc.iter_rows(
                min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True
            ),
            (),
        )
        field_cols = _resolve_columns(list(header_values))
        max_col = max(field_cols.values())
        # Keyed by resolved column letter to preserve the public output shape.
        column_fill_status = {
            get_column_letter(c): 0 for c in field_cols.values()
        }
        req_idx = field_cols["req_id"] - 1

        for offset, values in enumerate(
            ws_tc.iter_rows(
                min_row=DATA_START_ROW,
                max_row=ws_tc.max_row,
                max_col=max_col,
                values_only=True,
            ),
            start=DATA_START_ROW,
        ):
            row_num = offset
            req_id = values[req_idx] if len(values) > req_idx else None
            if not req_id:
                continue

            row_data = {"row_num": row_num}
            for field_name, col_1based in field_cols.items():
                col_index = col_1based - 1
                value = values[col_index] if len(values) > col_index else None
                row_data[field_name] = value or ""
                if value:
                    column_fill_status[get_column_letter(col_1based)] += 1

            rows.append(row_data)

    wb.close()

    return {
        "project": project,
        "test_group": test_group,
        "row_count": len(rows),
        "rows": rows,
        "column_fill_status": column_fill_status,
    }
