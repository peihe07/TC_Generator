#!/usr/bin/env python3
"""Extract the Last Mode Table into a List Item -> behavior lookup for B7.

037 cites this spec as `... R1L-R (August 2 2021)_{n}`, where `{n}` is the
List Item number in the `Last Mode Table` sheet — NOT an outline number. The
file in `inputs/` carries a different release label (`R1 SR24 1A Post
DCR19344`) but is the same document; see ANOMALIES.md A-H03 for the ruling
and the 15-leaf verification table.

The sheet merges cells vertically: Operation is written once and applies to
every row beneath it until the next Operation, and Screen Display Status
behaves the same way within an Operation. Both are forward-filled here — a
row read in isolation would otherwise lose the trigger it belongs to.

Outputs (under --out):
  last_mode_items.json  {list_item: {operation, screen_display_status,
                                     specific_screen_element, behavior,
                                     logic_reference}}

Usage:
    python build_last_mode.py --last-mode "inputs/Last Mode Table*.xlsx" \
        --out data
"""
import argparse
import json
from pathlib import Path

import openpyxl

SHEET = "Last Mode Table"
HEADER_ROW = 2
FIRST_DATA_ROW = 3
# 0-based column indices, verified against the 2021-08-02 workbook header row
COLS = {
    "list_item": 0, "operation": 1, "screen_display_status": 2,
    "specific_screen_element": 3, "behavior": 4, "logic_reference": 5,
}
# Columns written once and inherited by the rows below them.
FORWARD_FILL = ("operation", "screen_display_status")


def build(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"{path}: no {SHEET!r} sheet; sheets={wb.sheetnames}")
    ws = wb[SHEET]

    header = [str(c or "").strip()
              for c in next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW,
                                         values_only=True))]
    if not header[0].startswith("List Item"):
        raise SystemExit(f"{path}: unexpected header row {HEADER_ROW}: "
                         f"{header[:6]}")

    items, carry = {}, {}
    for r in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        raw = r[COLS["list_item"]]
        row = {k: str(r[i]).strip() if i < len(r) and r[i] is not None else ""
               for k, i in COLS.items()}
        for k in FORWARD_FILL:
            if row[k]:
                carry[k] = row[k]
            else:
                row[k] = carry.get(k, "")
        try:
            n = int(str(raw).strip())
        except (TypeError, ValueError):
            continue  # spacer / note row
        row.pop("list_item")
        items[str(n)] = row
    wb.close()
    if not items:
        raise SystemExit(f"no List Item rows extracted from {path}")
    return items


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--last-mode", required=True,
                    help="Last Mode Table xlsx (not in feature.yaml paths)")
    ap.add_argument("--out", default="data")
    args = ap.parse_args()

    src = Path(args.last_mode)
    items = build(src)

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    (out / "last_mode_items.json").write_text(
        json.dumps(items, ensure_ascii=False, indent=2))

    home = sum(1 for v in items.values()
               if v["screen_display_status"].upper() == "HOME")
    print(f"last_mode_items: {len(items)} List Items "
          f"({home} with Screen Display Status = HOME) -> "
          f"{out / 'last_mode_items.json'}")


if __name__ == "__main__":
    main()
