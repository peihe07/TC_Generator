#!/usr/bin/env python3
"""Build remaining_leaves.json + sibling_map.json + row_segments.json.

Diff the 037 A03 Analysis Report (leaf Functional Requirements) against the
DONE region of FW036. Everything not covered there is the remaining work.

Home differs from Media in one structural way: the done region is not a
positional prefix. Arif's rows are INTERLEAVED with the blank-author regen
rows in three segments each, so the done region is detected by the Test Case
Author column (Z) being non-empty, never by a row threshold. The detected
segment boundaries are emitted to `row_segments.json` for the write-back step,
which rewrites regen segments in place and content-hashes the Arif segments.

Paths default to `feature.yaml` `paths.a03_report` / `paths.workbook`, and
the workbook sheet, header row and column letters come from the same file —
one source of constants for the whole pipeline. CLI paths are overrides.

Usage:
    python build_remaining.py --out data/
    python build_remaining.py --fw036 <other-036.xlsx> --out data/   # override
"""
import argparse
import hashlib
import json
import re
from collections import OrderedDict
from pathlib import Path

import openpyxl

from feature_config import load_feature_config, resolve_path

A03_SHEET = "Analysis Report"  # 037 layout, not described by feature.yaml

# Expected shape, verified against the 2026-07-20 workbook. A mismatch means
# the inputs changed underneath the plan — fail loud rather than regenerate
# against a moved target.
EXPECT_LEAVES = 140
EXPECT_REMAINING = 62
EXPECT_ARIF_ROWS = 144


def load_a03(path) -> "OrderedDict[str, dict]":
    """Return all leaf FRs from the 037 analysis report, in document order."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[A03_SHEET]
    leaves = OrderedDict()
    for r in ws.iter_rows(min_row=8, values_only=True):
        rid, cat = r[0], str(r[6] or "")
        if not rid or not str(rid).startswith("SWE1"):
            continue
        if cat != "Functional Requirement":
            continue
        src = str(r[2] or "")
        # Home Screen suffixes are outline numbers (4.5.6); Last Mode suffixes
        # are List Item numbers up to 3 digits (A-H03) — both parse here.
        m = re.search(r"_(\d{1,3}(?:\.\d+)*)$", src)
        leaves[str(rid)] = {
            "req_id": str(rid),
            "parent": str(rid).rsplit("-", 1)[0] if re.search(r"-\d+-\d+$", str(rid))
                      else str(rid),
            "title": str(r[3] or "").strip(),
            "description": str(r[4] or "").strip(),
            "hmi_source_id": src,
            "section": m.group(1) if m else "",
            "source_req_id": str(r[1] or ""),
            "frop": str(r[7] or "") if len(r) > 7 else "",
        }
    wb.close()
    return leaves


def load_fw036(path, cfg: dict) -> tuple[list[dict], list[dict]]:
    """Return (rows, segments) for the TC sheet.

    Each row is {row, req_id, author, content_hash}. Segments are contiguous
    runs of the same kind: {"kind": "ARIF"|"REGEN", "start": r, "end": r}.
    """
    col = cfg["col"]
    first_data_row = cfg["workbook"]["header_row"] + 1
    author_col, req_id_col = col["author"], col["req_id"]
    # Content columns hashed for the done-region invariant: req_id..remarks
    content_cols = range(req_id_col, col["remarks"] + 1)

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[cfg["workbook"]["sheet"]]
    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=first_data_row,
                                       values_only=True),
                          start=first_data_row):
        if not any(c is not None for c in r):
            continue
        cells = [str(r[c]) if c < len(r) and r[c] is not None else ""
                 for c in content_cols]
        rows.append({
            "row": i,
            "req_id": str(r[req_id_col] or ""),
            "author": str(r[author_col] or "").strip()
                      if len(r) > author_col else "",
            "content_hash": hashlib.sha256(
                "\x1f".join(cells).encode("utf-8")).hexdigest()[:16],
        })
    wb.close()

    segments = []
    for row in rows:
        kind = "ARIF" if row["author"] else "REGEN"
        if segments and segments[-1]["kind"] == kind \
                and segments[-1]["end"] == row["row"] - 1:
            segments[-1]["end"] = row["row"]
            segments[-1]["rows"] += 1
        else:
            segments.append({"kind": kind, "start": row["row"],
                             "end": row["row"], "rows": 1})
    return rows, segments


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--a03", help="override feature.yaml paths.a03_report")
    ap.add_argument("--fw036", help="override feature.yaml paths.workbook")
    ap.add_argument("--feature-dir", default=".")
    ap.add_argument("--out", default="data")
    ap.add_argument("--no-assert", action="store_true",
                    help="skip the expected-shape assertions (inputs changed)")
    args = ap.parse_args()

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    cfg = load_feature_config(args.feature_dir)
    leaves = load_a03(resolve_path(cfg, "a03_report", args.a03))
    rows, segments = load_fw036(resolve_path(cfg, "workbook", args.fw036), cfg)

    done = {r["req_id"] for r in rows if r["author"] and r["req_id"]}
    draft = {r["req_id"] for r in rows if not r["author"] and r["req_id"]}

    remaining = OrderedDict(
        (rid, info) for rid, info in leaves.items() if rid not in done
    )
    # Req IDs the workbook traces to but 037 does not define at all. This is an
    # upstream traceability break, not a generation input — record, never fix.
    orphans = sorted((done | draft) - set(leaves))
    uncovered = [rid for rid in remaining if rid not in draft]

    # sibling_map: parent -> ordered list of ALL its leaf sub-ids (done ones
    # included, so the generator sees full sibling context per docs.md §4.6)
    sibling_map = OrderedDict()
    for rid, info in leaves.items():
        sibling_map.setdefault(info["parent"], []).append(
            {"req_id": rid, "title": info["title"], "remaining": rid in remaining}
        )
    sibling_map = OrderedDict(
        (p, subs) for p, subs in sibling_map.items()
        if any(s["remaining"] for s in subs)
    )

    arif_rows = [r for r in rows if r["author"]]
    done_region = {
        "arif_row_count": len(arif_rows),
        "segments": segments,
        # Content-based invariant (RUNBOOK Step 4.2): the ordered sequence of
        # Arif row contents must survive write-back regardless of row shifts.
        "ordered_content_hash": hashlib.sha256(
            "\n".join(r["content_hash"] for r in arif_rows).encode()
        ).hexdigest(),
    }

    (out / "remaining_leaves.json").write_text(
        json.dumps(list(remaining.values()), ensure_ascii=False, indent=2))
    (out / "sibling_map.json").write_text(
        json.dumps(sibling_map, ensure_ascii=False, indent=2))
    (out / "row_segments.json").write_text(
        json.dumps(done_region, ensure_ascii=False, indent=2))

    print(f"037 leaves: {len(leaves)}  done: {len(done & set(leaves))}  "
          f"remaining: {len(remaining)} across {len(sibling_map)} parents")
    print("segments: " + "  ".join(
        f"{s['kind']} {s['start']}-{s['end']} ({s['rows']})" for s in segments))
    print(f"arif rows: {len(arif_rows)}  "
          f"ordered_content_hash: {done_region['ordered_content_hash'][:16]}…")
    if uncovered:
        print(f"not covered anywhere (no draft row either): {uncovered}")
    if orphans:
        print(f"ORPHAN req_ids in workbook but absent from 037: {orphans} "
              f"-> must be recorded in ANOMALIES.md")

    if not args.no_assert:
        problems = []
        if len(leaves) != EXPECT_LEAVES:
            problems.append(f"leaves {len(leaves)} != {EXPECT_LEAVES}")
        if len(remaining) != EXPECT_REMAINING:
            problems.append(f"remaining {len(remaining)} != {EXPECT_REMAINING}")
        if len(arif_rows) != EXPECT_ARIF_ROWS:
            problems.append(f"arif rows {len(arif_rows)} != {EXPECT_ARIF_ROWS}")
        if problems:
            raise SystemExit("input shape changed: " + "; ".join(problems)
                             + "\nre-verify the plan, then rerun --no-assert")


if __name__ == "__main__":
    main()
