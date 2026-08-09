#!/usr/bin/env python3
"""Step 4 — write the generated TCs back into the FW036 Home workbook.

Home's regen rows are INTERLEAVED between three frozen Arif segments, so the
Media strategy (freeze a row prefix, truncate, append) does not apply. Each
regen segment is rewritten in place and the frozen region is protected by a
CONTENT hash rather than by row numbers:

    the ordered sequence of the 144 Arif rows' contents (columns
    req_id..remarks) must hash identically before and after, no matter how far
    the rows moved.

The baseline is `data/row_segments.json` `ordered_content_hash`, produced by
build_remaining.py against the pristine workbook.

Segment membership is derived from the draft rows being discarded: each regen
segment's draft req_ids fix its upper bound in 037 document order, and every
remaining leaf is assigned to the first segment whose bound covers it. That
places the two leaves with no draft row at all (055-03, 066) correctly without
hard-coding them.

Column conventions are read from feature.yaml and verified against the done
region — Home differs from Media in two ways worth stating: the vehicle flag
columns S..Y are EMPTY on all 229 existing rows (Media writes 1), and column F
carries no formula (Media writes a TC ID formula). Only column B does.

Usage:
    python write_back.py                  # dry run, prints the diff summary
    python write_back.py --write          # produce output/<name>.xlsx
    python write_back.py --write --out custom.xlsx

Exit code 0 = ok, 1 = an invariant failed, 2 = bad invocation.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sys
import zipfile
from datetime import date
from pathlib import Path

import openpyxl

from feature_config import load_feature_config, resolve_path

HISTORY_SHEET = "ChangeHistory 修訂履歷"
PLACEHOLDER_BODY = "BLOCKED - see Remarks"
# Verified against all 229 existing rows: constant everywhere.
CONST_CELLS = {"tc_ref_id": "NEW", "functional_safety": "NA"}
ROW_NUMBER_FORMULA = "=ROW()-9"
# Columns S..Y (vehicle flags) and C/E/F are blank on every existing row.
# They are listed here so a reviewer sees the omission is deliberate.
INTENTIONALLY_BLANK = ("C (Polarion ID)", "E (TestRail ID)", "F (Test Case ID)",
                       "S..Y (vehicle flags)")


class WriteBackError(RuntimeError):
    pass


# --------------------------------------------------------------------- read

def load_rows(ws, cfg) -> list[dict]:
    col = cfg["col"]
    first = cfg["workbook"]["header_row"] + 1
    out = []
    for i in range(first, ws.max_row + 1):
        vals = [ws.cell(i, c + 1).value for c in range(col["req_id"],
                                                       col["remarks"] + 1)]
        if not any(v is not None for v in vals):
            continue
        out.append({
            "row": i,
            "req_id": str(ws.cell(i, col["req_id"] + 1).value or ""),
            "author": str(ws.cell(i, col["author"] + 1).value or "").strip(),
            "cells": ["" if v is None else str(v) for v in vals],
        })
    return out


def row_digest(cells: list[str]) -> str:
    return hashlib.sha256("\x1f".join(cells).encode("utf-8")).hexdigest()[:16]


def ordered_content_hash(rows: list[dict], author: str) -> str:
    """Same construction as build_remaining.py, over the done-region rows.

    build_remaining identifies them as "author is non-empty", which is only
    true of the pristine workbook: after write-back the regen rows carry
    PeiPYHsu. Selecting by the done-region author value instead agrees with
    build_remaining before the edit and stays correct after it.
    """
    done = [r for r in rows if r["author"] == author]
    return hashlib.sha256(
        "\n".join(row_digest(r["cells"]) for r in done).encode()).hexdigest()


def segments_of(rows: list[dict], author: str) -> list[dict]:
    segs: list[dict] = []
    for r in rows:
        kind = "ARIF" if r["author"] == author else "REGEN"
        if segs and segs[-1]["kind"] == kind and segs[-1]["end"] == r["row"] - 1:
            segs[-1]["end"] = r["row"]
            segs[-1]["ids"].append(r["req_id"])
        else:
            segs.append({"kind": kind, "start": r["row"], "end": r["row"],
                         "ids": [r["req_id"]]})
    return segs


# ------------------------------------------------------------------ planning

def collect_generated(generated: Path, leaf_order: list[str]) -> list[dict]:
    """Flatten generated/*.json into workbook rows, in 037 document order."""
    by_req: dict[str, list[dict]] = {}
    for path in sorted(generated.glob("*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        for tc in payload.get("tcs", []):
            by_req.setdefault(str(tc["req_id"]), []).append(tc)
    unknown = sorted(set(by_req) - set(leaf_order))
    if unknown:
        raise WriteBackError(
            f"generated TCs reference req_ids that are not remaining leaves: "
            f"{unknown}")
    rows = []
    for req_id in leaf_order:
        for tc in by_req.get(req_id, []):
            rows.append(tc)
    return rows


def assign_segments(leaf_order: list[str], regen_segs: list[dict]) -> dict[str, int]:
    """leaf req_id -> index into regen_segs.

    Each regen segment's draft req_ids give its upper bound in 037 order; a
    leaf belongs to the first segment whose bound covers it. Leaves past every
    bound (there are none today, but a re-scoped 037 could add some) fall to
    the last segment rather than being silently dropped.
    """
    idx = {rid: i for i, rid in enumerate(leaf_order)}
    bounds = []
    for seg in regen_segs:
        known = [idx[i] for i in seg["ids"] if i in idx]
        if not known:
            raise WriteBackError(
                f"regen segment {seg['start']}-{seg['end']} has no draft "
                "req_id that is still a remaining leaf; cannot place it")
        bounds.append(max(known))
    if bounds != sorted(bounds):
        raise WriteBackError(f"regen segment bounds are not monotonic: {bounds}")

    assignment = {}
    for rid in leaf_order:
        target = len(regen_segs) - 1
        for i, b in enumerate(bounds):
            if idx[rid] <= b:
                target = i
                break
        assignment[rid] = target
    return assignment


# ------------------------------------------------------------------- writing

def cell_values(tc: dict, cfg) -> dict[str, object]:
    col = cfg["col"]
    wb_cfg = cfg["write_back"]
    is_placeholder = tc.get("placeholder") is True
    values = {
        col["req_id"]: tc["req_id"],
        col["test_group"]: None,          # Profile §2: never populated
        col["test_set"]: None,            # Profile §2: never populated
        col["test_item"]: tc["test_item"] or None,
        col["pre_conditions"]: tc["pre_conditions"] or None,
        col["input_test_data"]: tc["input_test_data"] or None,
        col["test_procedure"]: tc["test_procedure"] or None,
        col["expected_result"]: tc["expected_result"] or None,
        col["spec_reference"]: tc["specification_reference"],
        col["tc_ref_id"]: CONST_CELLS["tc_ref_id"],
        col["priority"]: tc["priority"] or None,
        col["design_method"]: tc["design_method"] or None,
        col["functional_safety"]: CONST_CELLS["functional_safety"],
        col["author"]: wb_cfg["author_value"],
        col["remarks"]: tc.get("remarks") or None,
    }
    if is_placeholder:
        for key in ("priority", "design_method"):
            if values[col[key]] is not None:
                raise WriteBackError(
                    f"{tc['req_id']}: placeholder rows must leave {key} blank")
        for key in ("test_procedure", "expected_result"):
            if values[col[key]] != PLACEHOLDER_BODY:
                raise WriteBackError(
                    f"{tc['req_id']}: placeholder {key} must be "
                    f"{PLACEHOLDER_BODY!r}")
    return values


def rewrite_segments(ws, cfg, regen_segs: list[dict],
                     rows_by_seg: list[list[dict]]) -> list[dict]:
    """Rewrite each regen segment in place. Later segments first, so the row
    indices of the earlier ones are still valid when their turn comes."""
    report = []
    for i in range(len(regen_segs) - 1, -1, -1):
        seg, new_rows = regen_segs[i], rows_by_seg[i]
        old_len = seg["end"] - seg["start"] + 1
        ws.delete_rows(seg["start"], old_len)
        if new_rows:
            ws.insert_rows(seg["start"], len(new_rows))
        for offset, tc in enumerate(new_rows):
            r = seg["start"] + offset
            for column, value in cell_values(tc, cfg).items():
                ws.cell(r, column + 1).value = value
        report.append({"segment": i + 1, "start": seg["start"],
                       "old_rows": old_len, "new_rows": len(new_rows)})
    return list(reversed(report))


def reemit_row_numbers(ws, cfg) -> int:
    """Column B is the only formula column in this workbook (F is blank)."""
    col_b = 2
    first = cfg["workbook"]["header_row"] + 1
    req_col = cfg["col"]["req_id"] + 1
    n = 0
    for r in range(first, ws.max_row + 1):
        if ws.cell(r, req_col).value in (None, ""):
            continue
        ws.cell(r, col_b).value = ROW_NUMBER_FORMULA
        n += 1
    return n


def append_history(wb, cfg, when: str, report: list[dict], n_rows: int) -> str:
    ws = wb[HISTORY_SHEET]
    row = 5
    last = None
    while ws.cell(row, 1).value not in (None, ""):
        last = str(ws.cell(row, 1).value).strip()
        row += 1
    revision = chr(ord(last) + 1) if last and len(last) == 1 else "C"
    spans = "; ".join(
        f"segment {d['segment']} at row {d['start']}: {d['old_rows']} -> "
        f"{d['new_rows']} rows" for d in report)
    ws.cell(row, 1).value = revision
    ws.cell(row, 2).value = (
        f"Regenerated the three blank-author segments from FMWIFSM037A03 "
        f"remaining leaves ({spans}); {n_rows} rows total.\n"
        "The 144 rows authored by Arif are unchanged — verified by an ordered "
        "content hash over columns D..AG, not by row position.\n"
        "Leaves with no testable content of their own carry a placeholder row "
        "with the reason in Remarks (see ANOMALIES A-H01, A-H02, A-H20)."
    )
    ws.cell(row, 3).value = cfg["write_back"]["author_value"]
    ws.cell(row, 4).value = when
    return revision


# ------------------------------------------------------------- reproducible

_EPOCH = (1980, 1, 1, 0, 0, 0)
_DCTERMS_RE = re.compile(rb"(<dcterms:(created|modified)[^>]*>)[^<]*(</dcterms:\2>)")
_FIXED_STAMP = b"2000-01-01T00:00:00Z"


def normalize_for_reproducibility(path: Path) -> None:
    """Zero the zip timestamps and document dates so the SHA256 is stable."""
    src = zipfile.ZipFile(path)
    items = sorted(src.infolist(), key=lambda i: i.filename)
    tmp = path.with_suffix(".normalizing")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for info in items:
            data = src.read(info.filename)
            if info.filename == "docProps/core.xml":
                data = _DCTERMS_RE.sub(rb"\1" + _FIXED_STAMP + rb"\3", data)
            new = zipfile.ZipInfo(info.filename, date_time=_EPOCH)
            new.compress_type = zipfile.ZIP_DEFLATED
            new.external_attr = info.external_attr
            dst.writestr(new, data)
    src.close()
    tmp.replace(path)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ------------------------------------------------------------------- driver

def run(args) -> int:
    cfg = load_feature_config(args.feature_dir)
    data = Path(args.data)
    src = resolve_path(cfg, "workbook", args.workbook)
    sheet = cfg["workbook"]["sheet"]

    baseline = json.loads((data / "row_segments.json").read_text())
    leaves = json.loads((data / "remaining_leaves.json").read_text())
    leaf_order = [x["req_id"] for x in leaves]

    wb = openpyxl.load_workbook(src)
    ws = wb[sheet]

    before_rows = load_rows(ws, cfg)
    arif_author = cfg["done_region"]["author_value"]
    before_hash = ordered_content_hash(before_rows, arif_author)
    if before_hash != baseline["ordered_content_hash"]:
        raise WriteBackError(
            "the source workbook no longer matches data/row_segments.json "
            f"({before_hash[:16]} != {baseline['ordered_content_hash'][:16]}); "
            "rerun build_remaining.py against the workbook you intend to edit")

    segs = segments_of(before_rows, arif_author)
    regen = [s for s in segs if s["kind"] == "REGEN"]
    arif_before = [r for r in before_rows if r["author"] == arif_author]

    tcs = collect_generated(Path(args.generated), leaf_order)
    covered = {tc["req_id"] for tc in tcs}
    missing = [r for r in leaf_order if r not in covered]
    if missing:
        raise WriteBackError(f"{len(missing)} leaves have no generated row: "
                             f"{missing[:10]}")

    assignment = assign_segments(leaf_order, regen)
    rows_by_seg: list[list[dict]] = [[] for _ in regen]
    for tc in tcs:
        rows_by_seg[assignment[tc["req_id"]]].append(tc)

    report = rewrite_segments(ws, cfg, regen, rows_by_seg)
    numbered = reemit_row_numbers(ws, cfg)

    after_rows = load_rows(ws, cfg)
    after_hash = ordered_content_hash(after_rows, arif_author)
    arif_after = [r for r in after_rows if r["author"] == arif_author]
    if after_hash != before_hash:
        raise WriteBackError(
            "DONE REGION CHANGED: ordered content hash "
            f"{before_hash[:16]} -> {after_hash[:16]}")
    if len(arif_after) != len(arif_before):
        raise WriteBackError(
            f"Arif row count changed: {len(arif_before)} -> {len(arif_after)}")

    after_segs = segments_of(after_rows, arif_author)
    kinds_before = [s["kind"] for s in segs]
    kinds_after = [s["kind"] for s in after_segs]
    if kinds_before != kinds_after:
        raise WriteBackError(
            f"segment order invariant broken: {kinds_before} -> {kinds_after}")

    written = {r["req_id"] for r in after_rows
               if r["author"] != arif_author}
    if written != set(leaf_order):
        raise WriteBackError(
            f"regen req_ids != remaining leaves "
            f"(missing {sorted(set(leaf_order) - written)[:5]}, "
            f"extra {sorted(written - set(leaf_order))[:5]})")

    # ---- summary
    print(f"source      : {src.name}")
    print(f"done region : {len(arif_after)} Arif rows unchanged "
          f"(hash {after_hash[:16]}…)")
    print(f"segments    : " + "  ".join(
        f"{s['kind']} {s['start']}-{s['end']}" for s in after_segs))
    for d in report:
        delta = d["new_rows"] - d["old_rows"]
        print(f"  segment {d['segment']} @row {d['start']}: "
              f"{d['old_rows']} -> {d['new_rows']} rows ({delta:+d})")
    print(f"regen rows  : {len(tcs)} across {len(set(covered))} leaves "
          f"({sum(1 for t in tcs if t.get('placeholder'))} placeholder)")
    print(f"row numbers : {numbered} rows re-emitted with {ROW_NUMBER_FORMULA}")
    print(f"blank by convention: {', '.join(INTENTIONALLY_BLANK)}")
    print(f"sheet rows  : {before_rows[-1]['row']} -> {after_rows[-1]['row']}")

    if not args.write:
        print("\nDRY RUN — nothing written. Re-run with --write to produce the "
              "workbook.")
        return 0

    revision = append_history(wb, cfg, args.date or date.today().isoformat(),
                              report, len(tcs))
    out = Path(args.out) if args.out else Path("output") / src.name
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    normalize_for_reproducibility(out)
    digest = sha256_file(out)
    print(f"\nwrote       : {out}")
    print(f"ChangeHistory revision {revision} appended")
    print(f"SHA256      : {digest}")
    (out.parent / (out.stem + ".sha256")).write_text(
        f"{digest}  {out.name}\n", encoding="utf-8")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--write", action="store_true",
                    help="actually produce the workbook (default is a dry run)")
    ap.add_argument("--feature-dir", default=".")
    ap.add_argument("--data", default="data")
    ap.add_argument("--generated", default="generated")
    ap.add_argument("--workbook", help="override feature.yaml paths.workbook")
    ap.add_argument("--out", help="output path (default output/<source name>)")
    ap.add_argument("--date", help="ChangeHistory date (default today)")
    args = ap.parse_args()
    try:
        return run(args)
    except WriteBackError as exc:
        print(f"ABORTED: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
