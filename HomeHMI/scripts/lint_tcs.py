#!/usr/bin/env python3
"""Step 3 hard gate — lint generated Home HMI TCs before workbook write-back.

Every rule maps to a section of docs/runtime/ASPICE_SWE6_AI_Instruction.md, the
Home profile, or an entry in ANOMALIES.md; the authority is named in each
finding so a reviewer can jump straight to it.

Nothing is hard-coded that `feature.yaml` already defines: column indices,
input paths and the PU allow-list all come from there. The Design Method
whitelist comes from the workbook's own `下拉選單` sheet — feature.yaml only
points at the workbook, the sheet is the authority.

Three gates are deliberately DIFFERENT from Media, because Media's versions
would fail Arif's own compliant done-region rows:

  - no `trailing-period` gate. 28% of done-region lines end with a period and
    72% do not; it is not a convention, so it is not a rule.
  - `[...]` and `<...>` are NOT banned. In Home they are popup control tokens
    quoted from the Pop Up List (`<X>`, `[OK, X]`, `[Reorder]`); they are
    checked against the cited PU row instead of forbidden (A-H10). The check
    skips `test_item`, which quotes the RD's own notation verbatim.
  - the ER modal-verb check strips double-quoted spans first, so verbatim
    popup text such as "Widget cannot be moved here." passes (A-H08).

Usage:
    python lint_tcs.py generated/
    python lint_tcs.py generated/SWE1-HMI-HOME-020.json
    python lint_tcs.py generated/ --json-report lint_report.json

Exit code 0 = clean, 1 = at least one TC failed, 2 = bad invocation.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from feature_config import load_feature_config, resolve_path

REQUIRED_KEYS = [
    "req_id", "test_group", "test_set", "test_item", "pre_conditions",
    "input_test_data", "test_procedure", "expected_result",
    "specification_reference", "priority", "design_method",
]
# Home done-region convention: Input Test Data is blank, not "NA".
BLANK_ALLOWED_KEYS = {"input_test_data", "remarks", "test_group", "test_set"}

VALID_PRIORITIES = {"P0", "P1", "P2", "P3"}
DROPDOWN_SHEET = "下拉選單"
# A-H03: B7 cites the file that exists, not 037's `R1L-R` release label.
LAST_MODE_REFERENCE_PREFIX = (
    "Last Mode Table HMI Logic and Flow R1 SR24 1A (August 2 2021)_")

# §5.1 forbidden MAIN verbs — only at the head of a step (the main-verb slot);
# a mid-sentence purpose clause is fine.
FORBIDDEN_MAIN_VERBS = [
    "observe whether", "observe", "see if", "check whether",
    "confirm whether", "verify", "watch", "monitor", "inspect",
]
# §6: ER states observable facts, never the RD's modal voice.
ER_MODAL_VERBS = ["shall", "will", "should", "would"]

_STEP_NUM_RE = re.compile(r"^\s*(\d+)[.)]\s*(.*)$")
_DOUBLE_QUOTED_RE = re.compile(r'"[^"]*"')
_BRACKET_TOKEN_RE = re.compile(r"<[^>\n]{1,30}>|\[[^\]\n]{1,30}\]")
_PU_RE = re.compile(r"\bPU\d{3,4}\b")
_BR_TAG_RE = re.compile(r"<\s*br\s*/?\s*>", re.IGNORECASE)
_CITATION_RE = re.compile(r"as defined by PU\d{3,4} String/Popup Message")


@dataclass
class Finding:
    req_id: str
    rule: str
    message: str
    source: str = ""

    def format(self) -> str:
        where = f" [{self.source}]" if self.source else ""
        return f"  {self.req_id}{where}  {self.rule}: {self.message}"


# ------------------------------------------------------------- authorities

def load_design_methods(workbook: Path) -> set[str]:
    """The 9 dropdown strings, from the workbook sheet that defines them."""
    wb = openpyxl.load_workbook(workbook, read_only=True)
    if DROPDOWN_SHEET not in wb.sheetnames:
        raise SystemExit(f"{workbook.name}: no {DROPDOWN_SHEET!r} sheet")
    methods = {str(r[0]).strip() for r in wb[DROPDOWN_SHEET].iter_rows(
        values_only=True) if r and r[0]}
    wb.close()
    if not methods:
        raise SystemExit(f"{workbook.name}: {DROPDOWN_SHEET} sheet is empty")
    return methods


def load_popup_index(popup: Path) -> dict[str, dict]:
    wb = openpyxl.load_workbook(popup, read_only=True)
    ws = wb["Main"]
    idx = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        pid = str(r[0] or "").strip()
        if pid.startswith("PU"):
            idx[pid] = {"message": str(r[6] or ""),
                        "exit_conditions": str(r[3] or "")}
    wb.close()
    return idx


def load_outlines(data: Path) -> set[str]:
    path = data / "spec_id_to_outline.tsv"
    if not path.exists():
        raise SystemExit(f"missing {path}; run build_outline_map.py first")
    with path.open(encoding="utf-8") as f:
        rows = csv.reader((ln for ln in f if not ln.startswith("#")),
                          delimiter="\t")
        next(rows, None)
        return {r[1] for r in rows if len(r) > 1}


def load_leaves(data: Path) -> dict[str, dict]:
    path = data / "remaining_leaves.json"
    if not path.exists():
        raise SystemExit(f"missing {path}; run build_remaining.py first")
    return {x["req_id"]: x for x in json.loads(path.read_text())}


# ------------------------------------------------------------- helpers

def _norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


def split_steps(text: str) -> list[tuple[int, str]]:
    """[(declared_number, body)] for each non-blank numbered line."""
    steps = []
    for line in str(text or "").splitlines():
        if not line.strip():
            continue
        m = _STEP_NUM_RE.match(line)
        steps.append((int(m.group(1)), m.group(2).strip()) if m
                     else (0, line.strip()))
    return steps


def strip_quoted(text: str) -> str:
    """Drop double-quoted spans — quoted source text is not authored prose."""
    return _DOUBLE_QUOTED_RE.sub(" ", text)


# ------------------------------------------------------------- rules

def _check_keys(tc: dict, add) -> None:
    for key in REQUIRED_KEYS:
        if key not in tc:
            add("keys", f"missing key `{key}`")
            continue
        value = tc[key]
        if not isinstance(value, str):
            add("keys", f"`{key}` must be a string, got {type(value).__name__}")
        elif not value.strip() and key not in BLANK_ALLOWED_KEYS:
            add("keys", f"`{key}` is empty")
    if str(tc.get("input_test_data", "")).strip().upper() == "NA":
        add("blank-convention",
            "`input_test_data` is 'NA'; the Home done region leaves it blank")


def _check_blank_columns(tc: dict, add) -> None:
    """Profile §2: Test Group / Test Set are never written to the workbook."""
    for key in ("test_group", "test_set"):
        if str(tc.get(key, "")).strip():
            add("blank-column",
                f"`{key}` must be empty (Profile §2), got {tc[key]!r}")


def _check_dropdowns(tc: dict, methods: set[str], add) -> None:
    priority = str(tc.get("priority", "")).strip()
    if priority not in VALID_PRIORITIES:
        add("priority", f"{priority!r} not in {sorted(VALID_PRIORITIES)}")
    method = str(tc.get("design_method", "")).strip()
    if method not in methods:
        add("design-method",
            f"{method!r} is not one of the {len(methods)} {DROPDOWN_SHEET} "
            "strings (exact match required)")


def _check_spec_reference(tc: dict, leaves: dict, outlines: set[str],
                          template: str, add) -> None:
    ref = str(tc.get("specification_reference", "")).strip()
    req_id = str(tc.get("req_id", ""))
    leaf = leaves.get(req_id)
    home_prefix = template.split("{")[0]

    if ref.startswith(LAST_MODE_REFERENCE_PREFIX):
        suffix = ref[len(LAST_MODE_REFERENCE_PREFIX):]
        if not suffix.isdigit():
            add("spec-reference",
                f"Last Mode reference suffix {suffix!r} is not a List Item number")
        elif leaf and leaf.get("section") and suffix != leaf["section"]:
            add("spec-reference",
                f"List Item {suffix} != 037 List Item {leaf['section']}")
        return

    if not ref.startswith(home_prefix):
        add("spec-reference",
            f"does not match the Profile §3.5 format {template!r}: {ref!r}")
        return
    outline = ref[len(home_prefix):]
    if outline not in outlines:
        add("spec-reference",
            f"outline {outline!r} does not resolve through "
            "spec_id_to_outline.tsv")
    elif leaf and leaf.get("section") and outline != leaf["section"]:
        add("spec-reference",
            f"outline {outline} != the 037 section {leaf['section']} for {req_id}")


def _check_steps(tc: dict, add) -> None:
    proc = split_steps(tc.get("test_procedure", ""))
    er = split_steps(tc.get("expected_result", ""))

    if len(proc) < 2:
        add("step-count",
            f"test_procedure has {len(proc)} numbered step(s), need >= 2")
    if len(proc) != len(er):
        add("step-count",
            f"{len(proc)} procedure step(s) vs {len(er)} expected_result line(s); "
            "they must correspond 1:1")

    for label, steps in (("test_procedure", proc), ("expected_result", er)):
        for i, (declared, _) in enumerate(steps, 1):
            if declared != i:
                add("step-numbering",
                    f"{label} line {i} is numbered {declared or '(none)'}")
                break

    for i, (_, body) in enumerate(proc, 1):
        low = body.lower()
        for verb in FORBIDDEN_MAIN_VERBS:
            if low.startswith(verb):
                add("forbidden-verb",
                    f"procedure step {i} opens with forbidden main verb {verb!r} "
                    "(§5.1)")
                break

    # A-H08: quoted popup text is source material, not the author's voice.
    for i, (_, body) in enumerate(er, 1):
        low = strip_quoted(body).lower()
        for modal in ER_MODAL_VERBS:
            if re.search(rf"\b{modal}\b", low):
                add("er-modal",
                    f"expected_result step {i} contains modal verb {modal!r} "
                    "outside a quoted span (§6)")
                break


# Profile §3.1: Test Item carries the requirement's shall-sentence VERBATIM.
# The RD writes popup controls in its own notation (`OK and [X]`) which need
# not match the Pop Up List's (`<OK>` / `<X>`), and a requirement may mention
# a popup that a given TC does not exercise. Quoted requirement text is source
# material, not authored prose — same exemption as A-H08 grants quoted spans.
POPUP_SCOPE_KEYS = ("pre_conditions", "test_procedure", "expected_result",
                    "remarks")


def _check_popups(tc: dict, popups: dict, allow: set[str], add) -> None:
    blob = "\n".join(str(tc.get(k, "")) for k in POPUP_SCOPE_KEYS)
    cited = sorted(set(_PU_RE.findall(blob)))
    if not cited:
        return

    for pid in cited:
        if pid not in popups:
            add("popup-unknown", f"{pid} is not in the Pop Up List")
        elif allow and pid not in allow:
            add("popup-unknown",
                f"{pid} is not in feature.yaml lint.popup_ids {sorted(allow)}")

    if not _CITATION_RE.search(blob):
        add("popup-citation",
            f"cites {', '.join(cited)} but never in the Profile §3.4 form "
            "`as defined by PUxxxx String/Popup Message`")

    known = [popups[p] for p in cited if p in popups]
    if not known:
        return
    messages = " ".join(_norm_ws(p["message"]) for p in known)
    tokens_ok = set()
    for p in known:
        tokens_ok |= set(_BRACKET_TOKEN_RE.findall(
            p["message"] + " " + p["exit_conditions"]))

    # Quoted text on a PU-citing line must be that popup's wording verbatim.
    for line in str(tc.get("expected_result", "")).splitlines():
        if not _PU_RE.search(line):
            continue
        for span in _DOUBLE_QUOTED_RE.findall(line):
            inner = _norm_ws(span.strip('"'))
            if inner and inner not in messages:
                add("popup-verbatim",
                    f"quoted text {inner!r} is not verbatim in the "
                    f"String/Popup Message of {', '.join(cited)} (Profile §3.4)")

    # A-H10: bracket tokens are legitimate ONLY as cited popup controls.
    for token in set(_BRACKET_TOKEN_RE.findall(blob)):
        if token not in tokens_ok:
            add("popup-token",
                f"bracket token {token!r} is not a control of "
                f"{', '.join(cited)}; use \"...\" for ordinary UI labels")


def _check_br_tags(tc: dict, add) -> None:
    for key in REQUIRED_KEYS:
        if _BR_TAG_RE.search(str(tc.get(key, ""))):
            add("br-tag", f"`{key}` contains a literal <br>; use a real newline")


def lint_tc(tc: dict, ctx: dict, source: str = "") -> list[Finding]:
    findings: list[Finding] = []
    req_id = str(tc.get("req_id", "?"))

    def add(rule: str, message: str) -> None:
        findings.append(Finding(req_id, rule, message, source))

    if req_id not in ctx["leaves"]:
        add("unknown-req-id",
            "not a remaining leaf in remaining_leaves.json — a well-formed row "
            "pointing at a requirement that does not exist")

    _check_keys(tc, add)
    _check_blank_columns(tc, add)
    _check_dropdowns(tc, ctx["methods"], add)
    _check_spec_reference(tc, ctx["leaves"], ctx["outlines"],
                          ctx["template"], add)
    _check_steps(tc, add)
    _check_popups(tc, ctx["popups"], ctx["popup_allow"], add)
    _check_br_tags(tc, add)
    return findings


# ------------------------------------------------------------- driver

def collect_paths(target: Path) -> list[Path]:
    if target.is_dir():
        return sorted(target.glob("*.json"))
    return [target]


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("target", help="generated/ directory or a single json")
    ap.add_argument("--feature-dir", default=".")
    ap.add_argument("--data", default="data")
    ap.add_argument("--workbook", help="override feature.yaml paths.workbook")
    ap.add_argument("--popup", help="override feature.yaml paths.popup_list")
    ap.add_argument("--json-report")
    args = ap.parse_args()

    target = Path(args.target)
    if not target.exists():
        print(f"no such path: {target}", file=sys.stderr)
        return 2
    paths = collect_paths(target)
    if not paths:
        print(f"no json files under {target}", file=sys.stderr)
        return 2

    cfg = load_feature_config(args.feature_dir)
    data = Path(args.data)
    ctx = {
        "methods": load_design_methods(resolve_path(cfg, "workbook",
                                                    args.workbook)),
        "popups": load_popup_index(resolve_path(cfg, "popup_list", args.popup)),
        "popup_allow": set(cfg.get("lint", {}).get("popup_ids") or []),
        "outlines": load_outlines(data),
        "leaves": load_leaves(data),
        "template": cfg["spec_reference_template"],
    }

    findings: list[Finding] = []
    tc_count = 0
    seen_req_ids: Counter = Counter()
    for path in paths:
        payload = json.loads(path.read_text())
        for tc in payload.get("tcs", []):
            tc_count += 1
            seen_req_ids[str(tc.get("req_id", "?"))] += 1
            findings.extend(lint_tc(tc, ctx, source=path.name))

    by_rule = Counter(f.rule for f in findings)
    failed = {f.req_id for f in findings}

    print(f"linted {tc_count} TCs from {len(paths)} file(s) "
          f"against {len(ctx['methods'])} design methods, "
          f"{len(ctx['outlines'])} outlines, {len(ctx['leaves'])} leaves")
    if findings:
        print(f"\nFAIL — {len(findings)} finding(s) on {len(failed)} req_id(s):")
        for f in findings:
            print(f.format())
        print("\nby rule: " + ", ".join(f"{r}={n}" for r, n in
                                        by_rule.most_common()))
    else:
        print("PASS — no findings")

    covered = sorted(seen_req_ids)
    print(f"\ncoverage: {len(covered)} req_id(s), "
          f"{tc_count} TCs ({len(ctx['leaves'])} leaves remaining overall)")

    if args.json_report:
        Path(args.json_report).write_text(json.dumps({
            "tc_count": tc_count,
            "files": [p.name for p in paths],
            "req_ids": covered,
            "findings": [f.__dict__ for f in findings],
            "by_rule": dict(by_rule),
            "passed": not findings,
        }, ensure_ascii=False, indent=2))
        print(f"-> {args.json_report}")

    return 1 if findings else 0


if __name__ == "__main__":
    sys.exit(main())
