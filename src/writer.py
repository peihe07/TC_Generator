"""Excel writer for generated TC results (RULES.md §10.3)."""
import os
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Alignment

TC_SHEET_NAME = "Test Case Specification&Result"
FRAMEWORK_SHEET_NAME = "Test Case Framework"

# Column mapping: field name -> 1-based column index
WRITE_COLUMNS = {
    "tc_id": 6,           # F
    "test_group": 7,      # G
    "test_set": 8,        # H
    "pre_conditions": 10, # J
    "input_test_data": 11,# K
    "test_procedure": 12, # L
    "expected_result": 13,# M
    "spec_reference": 14, # N
    "priority": 16,       # P
    "design_method": 17,  # Q
}

WRAP_TEXT_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CLEARABLE_FIELDS = {"test_set", "spec_reference"}


def _merge_test_item_text(existing_text: str | None, rewrite: str) -> str:
    """Preserve the original test item text while replacing any older rewrite."""
    original = existing_text or ""
    if "\n\n" in original:
        original = original.split("\n\n", 1)[0]
    return f"{original}\n\n{rewrite}"


def build_output_path(input_path: str, output_dir: str | None = None) -> str:
    """Build output file path: {input_filename}_generated.xlsx"""
    dirname = output_dir or os.path.dirname(input_path)
    basename = os.path.splitext(os.path.basename(input_path))[0]
    return os.path.join(dirname, f"{basename}_generated.xlsx")


def write_generated_results(
    input_path: str,
    generated_rows: list[dict],
    output_path: str,
) -> None:
    """
    Write generated TC results back to xlsx.

    Preserves all original formatting. Only writes to specified columns.
    Col I (Test Item) gets rewrite appended, not overwritten.
    """
    wb = load_workbook(input_path)
    ws = wb[TC_SHEET_NAME]

    for row_data in generated_rows:
        row_num = row_data["row_num"]

        # Write standard generated columns (overwrite)
        for field, col_idx in WRITE_COLUMNS.items():
            if field not in row_data:
                continue

            value = row_data.get(field)
            cell = ws.cell(row=row_num, column=col_idx)
            if value is not None:
                cell.value = value
                cell.alignment = WRAP_TEXT_ALIGNMENT
            elif field in CLEARABLE_FIELDS:
                cell.value = None
                cell.alignment = WRAP_TEXT_ALIGNMENT

        # Col I (Test Item): append rewrite, preserve original
        rewrite = row_data.get("test_item_rewrite")
        if rewrite:
            cell_i = ws.cell(row=row_num, column=9)
            cell_i.value = _merge_test_item_text(cell_i.value, rewrite)
            cell_i.alignment = WRAP_TEXT_ALIGNMENT

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    wb.close()


def write_framework_sheet(
    input_path: str,
    framework_data: list[dict],
    output_path: str,
) -> None:
    """
    Populate the Test Case Framework sheet.

    Columns: A=Test Group, B=Test Set, C=Req count.
    """
    wb = load_workbook(input_path)

    if FRAMEWORK_SHEET_NAME not in wb.sheetnames:
        wb.create_sheet(FRAMEWORK_SHEET_NAME)
    ws = wb[FRAMEWORK_SHEET_NAME]

    # Write header
    ws.cell(row=1, column=1, value="Test Group")
    ws.cell(row=1, column=2, value="Test Set")
    ws.cell(row=1, column=3, value="Req Count")

    # Write data
    for i, entry in enumerate(framework_data):
        row = i + 2
        ws.cell(row=row, column=1, value=entry["test_group"])
        ws.cell(row=row, column=2, value=entry["test_set"])
        ws.cell(row=row, column=3, value=entry["req_count"])

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    wb.close()
