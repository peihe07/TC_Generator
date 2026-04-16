"""Excel parser for TC Specification xlsx files (RULES.md §10)."""
import os
import re

from openpyxl import load_workbook


# Column index mapping (1-based for openpyxl)
COL_MAP = {
    "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8,
    "I": 9, "J": 10, "K": 11, "L": 12, "M": 13, "N": 14,
    "O": 15, "P": 16, "Q": 17,
}

# Columns to read from each data row
READ_COLUMNS = {
    "D": "req_id",
    "F": "tc_id",
    "G": "test_group",
    "H": "test_set",
    "I": "test_item",
    "J": "pre_conditions",
    "K": "input_test_data",
    "L": "test_procedure",
    "M": "expected_result",
    "N": "spec_reference",
    "P": "priority",
    "Q": "design_method",
}

HEADER_ROW = 9
DATA_START_ROW = 10
TC_SHEET_NAME = "Test Case Specification&Result"
PRODUCT_SHEET_NAME = "Product Document"


def parse_test_group_from_filename(filename: str) -> str | None:
    """Extract Test Group from filename pattern: ..._SWQT_{TestGroup}_{date}.xlsx"""
    basename = os.path.basename(filename)
    match = re.search(r"_SWQT_([^_]+)_\d{8}\.xlsx$", basename)
    return match.group(1) if match else None


def parse_tc_xlsx(filepath: str) -> dict:
    """
    Parse a TC Specification xlsx file.

    Returns dict with:
        - project: str (from Product Document sheet)
        - test_group: str (from filename)
        - row_count: int
        - rows: list[dict] (each row's data)
        - column_fill_status: dict[str, int] (count of non-empty cells per column)
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")

    wb = load_workbook(filepath, read_only=True, data_only=True)

    # Extract project name from Product Document sheet
    project = None
    if PRODUCT_SHEET_NAME in wb.sheetnames:
        ws_pd = wb[PRODUCT_SHEET_NAME]
        project = ws_pd.cell(row=3, column=2).value

    # Extract test group from filename
    test_group = parse_test_group_from_filename(filepath)

    # Parse TC data rows
    rows = []
    column_fill_status = {col: 0 for col in READ_COLUMNS}

    if TC_SHEET_NAME in wb.sheetnames:
        ws_tc = wb[TC_SHEET_NAME]
        for row_num in range(DATA_START_ROW, ws_tc.max_row + 1):
            req_id = ws_tc.cell(row=row_num, column=COL_MAP["D"]).value
            if not req_id:
                continue

            row_data = {"row_num": row_num}
            for col_letter, field_name in READ_COLUMNS.items():
                value = ws_tc.cell(row=row_num, column=COL_MAP[col_letter]).value
                row_data[field_name] = value or ""
                if value:
                    column_fill_status[col_letter] += 1

            rows.append(row_data)

    wb.close()

    return {
        "project": project,
        "test_group": test_group,
        "row_count": len(rows),
        "rows": rows,
        "column_fill_status": column_fill_status,
    }
