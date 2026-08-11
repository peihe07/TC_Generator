#!/usr/bin/env python3
"""Step 4 (SXM) — write the generated TCs into the FW036 SXM workbook.

SXM is the first feature on a **BLANK** workbook, and the first on the
revision C layout. Both change what the write is and what has to be checked.

**No done region, no legacy region.** Home rewrites draft segments interleaved
between frozen rows; AMFM appends after 158 frozen rows. Here the source is a
scaffold copy of the blank form, so there is nothing to protect and no content
hash to hold. The invariant that replaces them is the opposite one: *the source
must contain no authored content at all*. A workbook that has quietly acquired
rows since the scaffold was taken would otherwise be appended to, and the
result would ship someone else's rows under our ChangeHistory entry.

**Template residue is the real hazard here.** The blank form ships three
sample rows: `D10`/`D11` hold `xxx`, and `F10`/`G10`/`S10` hold
`NR1L-AntiTheft-001` / `AntiTheft` / `NA` — a different feature's TC ID and
Test Group. Columns G and S are overwritten by our own values, so they clean
themselves; **column F is not written by this feature** (no ruled TC ID scheme
— canon §10.3 needs a ruled series and SXM has none), so `NR1L-AntiTheft-001`
would survive into a delivered row 10 as a stray identifier belonging to
another feature. Every declared column is cleared across the write range before
anything is written, and the residue check runs again afterwards.

**Revision C column positions are not the repo's usual ones** (FORMS.md): the
form inserted `Estimated Test Time` at Q, so design_method is R, functional
safety S, author AA and remarks AH — one column right of every A/B instance.
They are read from `feature.yaml`, never assumed.

Usage:
    python SXMHMI/scripts/write_back.py --feature-dir SXMHMI
    python SXMHMI/scripts/write_back.py --feature-dir SXMHMI --write

Exit code 0 = ok, 1 = an invariant failed, 2 = bad invocation.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import zipfile
from datetime import date
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent / "scripts"))
from feature_config import load_feature_config, resolve_path  # noqa: E402

HISTORY_SHEET = "ChangeHistory 修訂履歷"
# Verified against the blank form's own sample row: column S is "NA".
CONST_CELLS = {"functional_safety": "NA"}
ROW_NUMBER_FORMULA = '=IF(ISBLANK($D{row}),"",ROW()-9)'
# Sample content the blank form ships in its three template rows, keyed by the
# column it appears in. Scoped per column on purpose: a flat string set would
# have to include "NA" (the sample Functional Safety value), and that would
# then be tolerated in any column — including a half-written req_id.
TEMPLATE_RESIDUE = {
    "req_id": {"xxx"},
    "tc_id": {"NR1L-AntiTheft-001"},
    "test_group": {"AntiTheft"},
    "functional_safety": {"NA"},
}


def is_residue(column: str, value) -> bool:
    return value is not None and str(value).strip() in TEMPLATE_RESIDUE.get(
        column, ())
INTENTIONALLY_BLANK = ("C (Polarion ID)", "E (TestRail ID)",
                       "F (Test Case ID — no ruled series for SXM)",
                       "Q (Estimated Test Time)", "T..Z (vehicle flags)",
                       "AB..AG (execution columns)")


class WriteBackError(RuntimeError):
    pass


# --------------------------------------------------------------------- read

def data_rows(ws, cfg) -> list[dict]:
    """Rows at or below the header that carry anything in a declared column."""
    col = cfg["col"]
    first = cfg["workbook"]["header_row"] + 1
    out = []
    for r in range(first, ws.max_row + 1):
        vals = {k: ws.cell(r, i + 1).value for k, i in col.items()}
        if any(v not in (None, "") for v in vals.values()):
            out.append({"row": r, "cells": vals})
    return out


def assert_blank_source(rows: list[dict]) -> list[dict]:
    """The source must hold template residue only — never authored content.

    A row is authored if any declared column holds something that is not one of
    the form's own sample strings. Checking every declared column rather than
    just `req_id` is deliberate: a half-written row with an empty req_id and a
    filled procedure is exactly the state this must refuse.
    """
    authored = []
    for r in rows:
        real = {k: v for k, v in r["cells"].items()
                if v not in (None, "") and not is_residue(k, v)}
        if real:
            authored.append({"row": r["row"], "cells": real})
    if authored:
        raise WriteBackError(
            f"the source workbook already holds {len(authored)} authored "
            f"row(s) — this feature writes into a BLANK form and would be "
            f"appending to someone else's content. First: {authored[0]}")
    return rows


# --------------------------------------------------------------- Scope field

def find_scope_cell(ws, cfg) -> tuple[int, int]:
    """The header 範圍 Scope VALUE cell, located by its label text."""
    label = cfg["write_back"]["scope_label"]
    needle = label.split()[-1].casefold()
    hits = []
    for r in range(1, cfg["workbook"]["header_row"]):
        for c in range(1, 40):
            v = ws.cell(r, c).value
            if v is None:
                continue
            text = str(v)
            if needle in text.casefold() or label.split()[0] in text:
                hits.append((r, c))
    if len(hits) != 1:
        raise WriteBackError(
            f"expected exactly one {label!r} label above row "
            f"{cfg['workbook']['header_row']}, found {len(hits)}: {hits}")
    r, c = hits[0]
    return r, c + 1


def _merge_anchor(ws, row: int, col: int) -> tuple[int, int]:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col
    return row, col


def fix_scope(ws, cfg, expected: str) -> dict:
    """Name this deliverable's requirement report in the header Scope field.

    Home and AMFM both found this field naming the WRONG deliverable (A-H26,
    A-AM01). On a blank form it is simply empty, which is the same defect in a
    less obvious dress: an unnamed scope is not a controlled document's
    identity claim either.
    """
    row, col = _merge_anchor(ws, *find_scope_cell(ws, cfg))
    before = ws.cell(row, col).value
    before = "" if before is None else str(before)
    ws.cell(row, col).value = expected
    return {"cell": ws.cell(row, col).coordinate, "before": before,
            "after": expected, "changed": before != expected}


# ------------------------------------------------------------------ writing

def collect_generated(generated: Path, leaf_order: list[str]) -> list[dict]:
    """Flatten generated/*.json into workbook rows, in 037 document order."""
    by_req: dict[str, list[dict]] = {}
    for path in sorted(generated.glob("*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        for tc in payload.get("tcs", []):
            by_req.setdefault(str(tc["req_id"]), []).append(tc)
    unknown = sorted(set(by_req) - set(leaf_order))
    if unknown:
        raise WriteBackError(
            f"generated TCs reference req_ids that are not 037 leaves: {unknown}")
    return [tc for req_id in leaf_order for tc in by_req.get(req_id, [])]


def cell_values(tc: dict, cfg) -> dict[int, object]:
    col = cfg["col"]
    wb_cfg = cfg["write_back"]
    fill_sets = wb_cfg.get("fill_test_group_set", False)
    values = {
        col["req_id"]: tc["req_id"],
        col["test_group"]: (tc.get("test_group") or None) if fill_sets else None,
        col["test_set"]: (tc.get("test_set") or None) if fill_sets else None,
        col["test_item"]: tc["test_item"] or None,
        col["pre_conditions"]: tc["pre_conditions"] or None,
        col["input_test_data"]: tc["input_test_data"] or None,
        col["test_procedure"]: tc["test_procedure"] or None,
        col["expected_result"]: tc["expected_result"] or None,
        col["spec_reference"]: tc["specification_reference"],
        col["tc_ref_id"]: wb_cfg["tc_ref_id_value"],
        col["priority"]: tc["priority"] or None,
        col["design_method"]: tc["design_method"] or None,
        col["functional_safety"]: CONST_CELLS["functional_safety"],
        col["author"]: wb_cfg["author_value"],
        col["remarks"]: tc.get("remarks") or None,
    }
    if fill_sets and not (values[col["test_group"]] and values[col["test_set"]]):
        raise WriteBackError(
            f"{tc['req_id']}: fill_test_group_set is on but Test Group / Test "
            "Set is empty")
    return values


def clear_range(ws, cfg, first_row: int, last_row: int) -> int:
    """Blank every declared column across the write range before writing.

    Columns the feature does not write would otherwise keep the form's sample
    values — `F10` holds another feature's TC ID and nothing else in this
    script touches column F.
    """
    n = 0
    for r in range(first_row, last_row + 1):
        for idx in cfg["col"].values():
            if ws.cell(r, idx + 1).value not in (None, ""):
                ws.cell(r, idx + 1).value = None
                n += 1
    return n


def write_rows(ws, cfg, first_row: int, tcs: list[dict]) -> dict:
    """Write from `first_row`, growing the sheet only by what it lacks."""
    spare = max(ws.max_row - first_row + 1, 0)
    inserted = max(len(tcs) - spare, 0)
    if inserted:
        ws.insert_rows(ws.max_row + 1, inserted)
    cleared = clear_range(ws, cfg, first_row, first_row + len(tcs) - 1)
    for offset, tc in enumerate(tcs):
        r = first_row + offset
        for column, value in cell_values(tc, cfg).items():
            ws.cell(r, column + 1).value = value
    return {"first_row": first_row, "last_row": first_row + len(tcs) - 1,
            "rows": len(tcs), "template_rows_used": min(spare, len(tcs)),
            "inserted": inserted, "cleared_cells": cleared}


def reemit_row_numbers(ws, cfg) -> int:
    first = cfg["workbook"]["header_row"] + 1
    req_col = cfg["col"]["req_id"] + 1
    n = 0
    for r in range(first, ws.max_row + 1):
        if ws.cell(r, req_col).value in (None, ""):
            continue
        ws.cell(r, 2).value = ROW_NUMBER_FORMULA.format(row=r)
        n += 1
    return n


def residue_left(ws, cfg, first_row: int, written: set[str]) -> list[str]:
    """Sample strings still present in columns this feature does not write.

    Columns the write covers clean themselves, and one of them legitimately
    ends up holding a sample string: Functional Safety is `NA` on the form's
    sample row and `NA` on every row we write. Scanning those columns would
    report our own intended value as residue. The columns that matter are the
    ones nothing overwrites — `tc_id` above all, where the form leaves another
    feature's identifier.
    """
    found = []
    for r in range(first_row, ws.max_row + 1):
        for key, idx in cfg["col"].items():
            if key in written:
                continue
            v = ws.cell(r, idx + 1).value
            if is_residue(key, v):
                found.append(f"{ws.cell(r, idx + 1).coordinate} ({key}) = {v!r}")
    return found


def append_history(wb, cfg, when: str, plan: dict, n_leaves: int,
                   scope: dict | None = None) -> str:
    ws = wb[HISTORY_SHEET]
    row = 5
    last = None
    while ws.cell(row, 1).value not in (None, ""):
        last = str(ws.cell(row, 1).value).strip()
        row += 1
    revision = chr(ord(last) + 1) if last and len(last) == 1 else "D"
    note = (
        f"Added {plan['rows']} test cases covering all {n_leaves} leaves of "
        f"FM-WI-FSM-037-A03 (SXM), written into rows "
        f"{plan['first_row']}-{plan['last_row']} of the blank form.\n"
        "The form carried no authored rows before this revision; its three "
        "sample rows were cleared."
    )
    if scope and scope["changed"]:
        note += (f"\nSet the header 範圍 Scope field ({scope['cell']}) to "
                 f"{scope['after']}.")
    ws.cell(row, 1).value = revision
    ws.cell(row, 2).value = note
    ws.cell(row, 3).value = cfg["write_back"]["author_value"]
    ws.cell(row, 4).value = when
    return revision


# ------------------------------------------------------------- reproducible

_EPOCH = (1980, 1, 1, 0, 0, 0)
_DCTERMS_RE = re.compile(rb"(<dcterms:(created|modified)[^>]*>)[^<]*(</dcterms:\2>)")
_FIXED_STAMP = b"2000-01-01T00:00:00Z"


def normalize_for_reproducibility(path: Path) -> None:
    """Zero the zip timestamps and document dates so the SHA256 is stable."""
    src = zipfile.ZipFile(path)
    items = sorted(src.infolist(), key=lambda i: i.filename)
    tmp = path.with_suffix(".normalizing")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for info in items:
            data = src.read(info.filename)
            if info.filename == "docProps/core.xml":
                # \g<1> not \1: `\1` followed by a stamp starting "2" parses
                # as group 12 and silently destroys core.xml.
                data = _DCTERMS_RE.sub(rb"\g<1>" + _FIXED_STAMP + rb"\g<3>", data)
            new = zipfile.ZipInfo(info.filename, date_time=_EPOCH)
            new.compress_type = zipfile.ZIP_DEFLATED
            new.external_attr = info.external_attr
            dst.writestr(new, data)
    src.close()
    tmp.replace(path)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# -------------------------------------------------------------------- driver

def run(args) -> int:
    cfg = load_feature_config(args.feature_dir)
    root = Path(args.feature_dir)
    src = resolve_path(cfg, "workbook", args.workbook)
    stla = json.loads((root / args.data / "stla_to_cfts.json").read_text(
        encoding="utf-8"))
    leaf_order = list(stla)

    wb = openpyxl.load_workbook(src)
    ws = wb[cfg["workbook"]["sheet"]]

    before = data_rows(ws, cfg)
    assert_blank_source(before)

    scope_expected = resolve_path(cfg, cfg["write_back"]["scope_source"]).stem
    scope = fix_scope(ws, cfg, scope_expected)

    tcs = collect_generated(root / args.generated, leaf_order)
    covered = {tc["req_id"] for tc in tcs}
    missing = [r for r in leaf_order if r not in covered]
    if missing:
        raise WriteBackError(
            f"{len(missing)} leaves have no generated row: {missing[:10]}")

    first_row = cfg["workbook"]["header_row"] + 1
    plan = write_rows(ws, cfg, first_row, tcs)
    numbered = reemit_row_numbers(ws, cfg)

    # ---- invariants
    after = data_rows(ws, cfg)
    if len(after) != len(tcs):
        raise WriteBackError(
            f"{len(after)} data rows after the write, expected exactly "
            f"{len(tcs)} — a template row survived or a row was lost")
    if [r["row"] for r in after] != list(range(first_row, first_row + len(tcs))):
        raise WriteBackError("written rows are not contiguous from "
                             f"row {first_row}")
    written = [r["cells"]["req_id"] for r in after]
    if written != [tc["req_id"] for tc in tcs]:
        raise WriteBackError("row order does not match 037 document order")
    if set(written) != set(leaf_order):
        raise WriteBackError(
            "written req_ids != the 037 leaf set (missing "
            f"{sorted(set(leaf_order) - set(written))[:5]}, extra "
            f"{sorted(set(written) - set(leaf_order))[:5]})")
    written_cols = set(cell_values(tcs[0], cfg)) if tcs else set()
    written_keys = {k for k, i in cfg["col"].items() if i in written_cols}
    residue = residue_left(ws, cfg, first_row, written_keys)
    if residue:
        raise WriteBackError(
            f"blank-form sample content survived the write: {residue[:5]}")

    per_leaf: dict[str, int] = {}
    for tc in tcs:
        per_leaf[tc["req_id"]] = per_leaf.get(tc["req_id"], 0) + 1
    spread: dict[int, int] = {}
    for n in per_leaf.values():
        spread[n] = spread.get(n, 0) + 1
    prio: dict[str, int] = {}
    for tc in tcs:
        prio[tc["priority"]] = prio.get(tc["priority"], 0) + 1

    print(f"source        : {src.name}")
    print(f"source state  : BLANK — {len(before)} template row(s), 0 authored")
    print(f"rows written  : {plan['rows']} TCs into rows "
          f"{plan['first_row']}-{plan['last_row']} "
          f"({plan['template_rows_used']} template rows reused, "
          f"{plan['inserted']} inserted, {plan['cleared_cells']} residue "
          "cells cleared)")
    print(f"coverage      : {len(covered)} leaves == {len(leaf_order)} leaf set, "
          "exact match")
    print("TCs per leaf  : " + ", ".join(
        f"{k} TC x{v}" for k, v in sorted(spread.items())))
    print("priority      : " + ", ".join(
        f"{k}={v}" for k, v in sorted(prio.items())))
    print(f"row numbers   : {numbered} rows re-emitted with the column B formula")
    print(f"Scope {scope['cell']}     : {scope['before']!r} -> {scope['after']!r}")
    print(f"blank by convention: {', '.join(INTENTIONALLY_BLANK)}")

    if not args.write:
        print("\nDRY RUN — nothing written. Re-run with --write.")
        return 0

    revision = append_history(wb, cfg, args.date or date.today().isoformat(),
                              plan, len(leaf_order), scope)
    out = Path(args.out) if args.out else root / "output" / src.name
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    normalize_for_reproducibility(out)
    digest = sha256_file(out)
    (out.parent / (out.stem + ".sha256")).write_text(
        f"{digest}  {out.name}\n", encoding="utf-8")
    print(f"\nwrote         : {out}")
    print(f"ChangeHistory : revision {revision} appended")
    print(f"SHA256        : {digest}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--feature-dir", default=".")
    ap.add_argument("--data", default="data")
    ap.add_argument("--generated", default="generated")
    ap.add_argument("--workbook")
    ap.add_argument("--out")
    ap.add_argument("--date")
    ap.add_argument("--write", action="store_true")
    args = ap.parse_args()
    try:
        return run(args)
    except WriteBackError as exc:
        print(f"ABORTED: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
