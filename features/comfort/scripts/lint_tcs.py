#!/usr/bin/env python3
"""Comfort lint gate over generated/*.json — PASS/FAIL with measured values.

Written for Comfort, not copied from Privacy: Privacy's gate resolves
`specification_reference` against CFTS022 artifact ids, which Comfort does
not have. Comfort's authority is the SR24 outline set, so the reference gate
resolves against `data/layer3_map.tsv` — the 129 sections actually cited.

Authorities are READ, never hard-coded:
  design-method vocabulary  <- the workbook's own 下拉選單 sheet
  Test Group / tc_id format <- feature.yaml
  valid outlines            <- data/layer3_map.tsv
  clause text (token check) <- data/section_fulltext.tsv

Rulings encoded as gates rather than left to discipline:
  R-C1   spec_reference stem is the SR24 filename, never SR25
  R-C7   tc_id matches the frozen format and is unique and gap-free
  R-C22  ER carries no fabricated magnitude for a clause that gives none
  §3.4   `(-, +)` verbatim only in test_item / quoted fragments; procedure
         and non-quoting ER use "-" / "+"
  §3.2   every pre_condition line carries a source class
  §11    no trailing period on the four long fields; UI labels use "..."
  §4.6   axis="none" <=> duplicate_of is set

Exit 0 = clean, 1 = at least one finding.

Usage:
    python3 features/comfort/scripts/lint_tcs.py
"""

import csv
import difflib
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[3]
FEATURE = ROOT / "features" / "comfort"
GEN = FEATURE / "generated"
LAYER3 = FEATURE / "data" / "layer3_map.tsv"
FULLTEXT = FEATURE / "data" / "section_fulltext.tsv"

STEM = ("SYS1_HMI_Comfort_HMI_Logic_and_Flow_R1_SR24_Post_3A_CR24879_"
        "(September_25_2023)")
LONG_FIELDS = ["pre_conditions", "input_test_data", "test_procedure",
               "expected_result"]
# 81 §2 / R-C45 — `ext-verbatim` 為**外部文件**之逐字引用。另立一類而非借用
# `spec-verbatim`：後者之意義是「本 spec 該節之原話」，而 source-class-truthful
# 正是對照該節查證的。把外部句子標成 spec-verbatim，會使那道 gate 查一個
# 它查不到的東西而 FAIL —— 或更糟，查得到而讀者以為它出自本 spec。
# R-C45 —— 已認可之外部出處。**匯入**而非在此另抄一份，亦不以正規式重建：
# 兩份清單會分岔，一份不會（R-C43 之同一理由 —— 以身分界定，且身分只有一個）。
sys.path.insert(0, str(Path(__file__).resolve().parent))
import external_docs as _ext                                    # noqa: E402

EXTERNAL_REFS = {v for k, v in vars(_ext).items()
                 if k.startswith("EXT_") and isinstance(v, str)}

SOURCE_CLASSES = ("spec-verbatim", "spec-derived", "test-setup",
                  "ext-verbatim")
PRIORITIES = {"P0", "P1", "P2", "P3"}
MODALS = re.compile(r"\b(shall|must|should|will|would)\b", re.I)
# §10.1 — every TC object must carry all ten keys.
REQUIRED_KEYS = ("tc_title", "pre_conditions", "input_test_data",
                 "test_procedure", "expected_result", "specification_reference",
                 "design_method", "priority", "split_flag", "split_reason")
MIN_STEPS = 2                  # §10.5
# profile §5.1 / §5.2 — R-C24's [BLOCKED-SPEC] and R-C38's [BLOCKED-NON-HMI].
# Both produce an empty row; they differ in WHY, and the Remarks gate below
# is what keeps them from collapsing into one another.
BLOCKED_MARKERS = ("[BLOCKED-SPEC]", "[BLOCKED-NON-HMI]")
# R-C26 — a marker that grants a lint exemption must not be self-issuable.
# The whitelist is the named list in profile §5.1/§5.2; adding to it is a
# ruling, not an edit. Without this, a BLOCKED marker is an exemption anyone
# can take by typing it, which is the same as having no exemption condition
# at all.
MARKER_WHITELIST = {"[BLOCKED-SPEC]": {"NR1L-ComfortHMI-010",
                                       "NR1L-ComfortHMI-012",
                                       # 66 §1 — DR #43 ruled (一): 11.5 and
                                       # 12.6 delegate their whole content to
                                       # two DIFFERENT external documents.
                                       # 81 §2.3 — 11.5's row (-382) LEFT the
                                       # whitelist: the HMI Settings List was
                                       # found and carries the options, so the
                                       # row became a real TC. Removal is a
                                       # ruling too (R-C26), and 81 §2.3 is it.
                                       # 12.6's -383 stays: HMI Notes does not
                                       # exist in the customer tree.
                                       "NR1L-ComfortHMI-383"},
                    # 41 §1.2 — ruled together with R-C38 itself.
                    "[BLOCKED-NON-HMI]": {"NR1L-ComfortHMI-081"}}
OWNER_WINDOW = 60          # R-C27 — chars visible on the clipped first line
# 66 §3 / 67 §1 — rows whose Remarks carries a clause ambiguity rather than a
# marker. This is a REGISTER, not a permit: listing is not authorisation, and
# an author who meets 66 §3's condition writes the Remarks and registers the
# row in the same act, with no case-by-case ruling. Its job is identity
# reconciliation in BOTH directions (registered-but-absent is as wrong as
# present-but-unregistered), and the CONTENT limits below are enforced by the
# gate rather than by the register.
#
# The contrast with MARKER_WHITELIST is the point: that one EXEMPTS a row from
# lint, so it must be ruled (R-C26). This one exempts nothing — it reconciles.
# Same shape, opposite nature: one relaxes, one reconciles.
# 68 §2 — the value is a fragment that identifies THIS row's ambiguity, not a
# phrase every row happens to share. A shared phrase reconciles "was the
# sentence written" and nothing more; the strength of a reconciliation is
# exactly the distinguishability of what it reconciles.
AMBIGUITY_REMARKS = {
    # 14.12 三條 —— HVACP12 之 `the hard controls` 為全車集合語（DR #37）
    "NR1L-ComfortHMI-374": "whether radial popups apply",
    "NR1L-ComfortHMI-375": "whether vertical popups apply",
    "NR1L-ComfortHMI-376": "only the predominant control type",
}
# 67 §1 三 — content limits, hard, checked on every registered row.
REMARKS_FORBIDDEN = ("R-C", "DR #", "A-CF", "§")
REMARKS_MIN_CHARS = 40
# R-C38 — [BLOCKED-NON-HMI]'s first visible line must say what is missing,
# and must NOT name an owner: having no owner IS the classification. An
# "Owner:" here would be a [BLOCKED-SPEC] wearing the wrong marker.
NON_HMI_PHRASE = "Not an HMI-observable property"
# handoff 26 §4.1 — keys a TC object may carry WITHOUT landing in a column.
# `COLS` in write_back.py is a hand-kept list; if it loses an entry, nothing
# shouts. This gate makes the silence audible: every TC key must either map to
# a column or be named here. Adding to this list is a ruling, not an edit
# (same reason as MARKER_WHITELIST — a self-issuable whitelist is no whitelist).
# Extended to the doc level per 29 §5.1: four of the original eight entries
# (reasoning, keywords, duplicate_of, distinguishing_axis) live on the doc
# object, not the TC object, so naming them while scanning only TCs named
# nothing — they were never in the scanned layer.
NOT_IN_WORKBOOK = {
    "tc_title",             # canon §4.3 — derivation input, never a column
    "estimated_test_time",  # no column in revision C
    "split_flag", "split_reason",
    "reasoning", "keywords", "duplicate_of", "distinguishing_axis",
    # doc-level structural keys (29 §5.1)
    "assumptions", "batch", "outline", "parent", "source_clause", "tcs",
    "interface_axis_review",          # 36 §6 — R-C34's duty, recorded
    "emea_ics_review",                # 38 §1 — R-C36-1's per-TC answer
}
# 36 §6 — the four interface-type axes must each carry a non-empty answer.
# Correctness cannot be machine-checked; having been asked can.
IFACE_KEYS = ("observable_interface", "axis_9", "axis_12", "axis_13",
              "emea_ics")
# 36 §4 — a sibling whose counterpart section is not generated yet leaves
# duplicate_of empty. Legal while the counterpart is missing, a defect the
# day it lands. The table turns "remember to backfill" into a condition.
# Read from data/pending_sibling.tsv so the gate and the candidate generator
# share one table (37 §6). Only `sibling` verdicts are watched.
def _load_sibling_table() -> list:
    path = FEATURE / "data" / "pending_sibling.tsv"
    with path.open(encoding="utf-8") as fh:
        return list(csv.DictReader(fh, delimiter="\t"))


SIBLING_TABLE = _load_sibling_table()
PENDING_SIBLING = [(r["outline"], r["sibling_outline"])
                   for r in SIBLING_TABLE if r["verdict"] == "sibling"]
# 42 §1 — a `provisional` verdict was reached against a section that had no
# TCs. The day that section's Test Set is complete, the verdict must be
# looked at again and the flag cleared by hand. Re-confirmation MAY keep the
# verdict; what it may not do is stay silent.
TEST_SET_MAP = FEATURE / "data" / "test_set_map.tsv"


def _load_test_sets() -> dict:
    with TEST_SET_MAP.open(encoding="utf-8") as fh:
        return {r["outline"]: r["test_set"]
                for r in csv.DictReader(fh, delimiter="\t")}


SECTION_TEST_SET = _load_test_sets()
ANOMALY_ID = re.compile(r"\bA-CF\d+\b")
PROFILE = ROOT / "docs" / "runtime" / "profiles" / "FW036_R1L_Comfort_Profile.md"
# 35 §4 — a negated pre_condition ("the vehicle does NOT have X") covers
# whatever values the axis happens to have today, and axes gain values. The
# profile records the limit; this gate is the mechanism, because a record is
# not a mechanism. Adding a value without re-reviewing the negation users
# fails, and the failure names them.
AXIS_BLOCK = re.compile(r"```axis-values\n(.*?)```", re.S)
# 52 §3, criterion corrected by 54 §1 — `axis-type-reverse-test`.
#
# 52 §3 worded the question as "is there a TC whose FUNCTION is governed by
# this axis, and whose observable sits on the interface the axis removes?"
# The purpose it was given (35 §1.1) is the opposite case: "an interface-type
# axis matters because the function survives while ANOTHER TC's observable
# disappears". The wording is the narrower of the two, and 54 §1 rules the
# PURPOSE version to be the FAIL criterion:
#
#   FAIL  — any TC whose observable sits on an interface some axis value
#           removes, and which does not state that axis's value.
#   report— of those, how many also have an axis-governed function (the
#           wording version). Kept as a named line, not deleted: if the two
#           ever diverge again, the difference is the finding.
#
# Criteria are READ FROM THE PROFILE, never hard-coded (52 §3).
FN_AXIS_BLOCK = re.compile(r"```function-axis-reverse-test\n(.*?)```", re.S)
# 56 §4 — the fields that MAKE UP the declaration. `declared-at-tc-count`
# and `judged-at` are excluded on purpose: they record when, not what.
HASHED_FIELDS = ("axis", "function-keywords", "removed-interface-keywords",
                 "axis-pc-keywords", "judged-at-tc-count",
                 "judged-at-provenance")
# The profile's axis table, used only to catch a 功能型 axis with NO block —
# an omission is otherwise silent, which is the failure this gate exists for.
AXIS_TABLE_ROW = re.compile(
    r"^\|\s*([0-9]+|—)\s*\|\s*(.+?)\s*\|\s*\*{0,2}(功能型|介面型)\*{0,2}\s*\|",
    re.M)
# 43 §4 — one block per axis that uses a negated pre_condition, each carrying
# its own `negation:` string. Until 43 §4 this gate watched ONE hard-coded
# phrase (axis 13's), so the other four negations had no protection: 34 §4's
# reason is about what a negation covers, and that does not depend on which
# axis is negated.
# 60 §1 — module level so verify_no_tcid_gate.py reads THESE objects
# rather than a re-typed copy (the lesson from verify_provisional_gate).
# 62 §5 — req_id citations in prose. `123-45` and `123` are both legal keys
# (037 gives some sections a section-level leaf), so both forms are matched.
REQ_CITE = re.compile(r"`(\d{3}(?:-\d{2})?)`")


def _leaf_universe() -> set:
    recon = json.loads((FEATURE / "data" / "recon.json").read_text(
        encoding="utf-8"))
    return set(recon["leaves"])


TCID_LONG = re.compile(r"NR1L-ComfortHMI-\d+")
TCID_SHORT = re.compile(r"`-\d{3}`")
LEAF_UNIVERSE = _leaf_universe()
# 62 §1.1 (b) — generators declare stopped leaves as ("SWE1-HVAC-xxx-yy", why)
WITHHELD_DECL = re.compile(r'\("(SWE1-HVAC-\d+(?:-\d+)?)"')
FULLTEXT_BY_OUTLINE = {
    r["outline"]: r["full_text"] for r in csv.DictReader(
        (FEATURE / "data" / "section_fulltext.tsv").open(encoding="utf-8"),
        delimiter="\t")}

NEGATED_PC = re.compile(r"\bdoes not\b|\bis not\b|\bnot configured\b"
                        r"|\bnot present\b|\bnot currently\b")
# Negated pre_conditions that are NOT configuration axes: runtime state and
# test setup. Naming them is what lets the coverage check below FAIL on a NEW
# axis negation instead of quietly ignoring it (same shape as NOT_IN_WORKBOOK
# — a list nobody has to maintain is a list that protects nothing).
NON_AXIS_NEGATIONS = {
    "The lower screen is not in the stowed position":
        "13.2 之執行期狀態（同一台車兩種狀態），非配置軸",
    "The user is not in the climate section on the main head unit":
        "13.2 之執行期畫面位置，非配置軸",
    "The Seats tab is not currently shown":
        "test-setup 之起始狀態，非配置軸",
}
# §5.1's nine forbidden MAIN verbs, verbatim and nothing else.
# Authority: canon §5.1, via handoff 31 §1.
#
# `locate` was in this list for one batch and is now OUT (32 §2): canon §5.6
# uses it in a POSITIVE example — "Locate the phone and record its A2DP and
# HFP status shown in the list". It is a positioning action and defers no
# judgement to the tester, so the gate had mechanically forbidden a wording
# the canon recommends.
#
# Rule this cost us: an entry with no cited authority does not go in the
# list. A habit from hand-checking swayed one judgement; inside a gate it
# would have swayed every batch after it.
# Listed one-for-one against §5.1 so the two can be diffed by eye — nine
# entries, nine in the canon. `observe whether` is subsumed by `observe` as
# a matcher, but a list that silently carries eight where the canon says
# nine cannot be checked against its own authority.
FORBIDDEN_VERBS = ("observe whether", "observe", "see if", "check whether",
                   "confirm whether", "verify", "watch", "monitor", "inspect")
# Anchored at the step's own start — "1." then optional whitespace then the
# verb. §5.1 explicitly ALLOWS `verify` inside a purpose clause
# ("... to verify that ..."), so a substring test would fail legal usage, and
# a gate that fails legal usage teaches authors to route around it rather
# than to fix anything (31 §1).
# Longest-first, so the reported verb is the fullest phrase that matched
# ("check whether", not "check") rather than whichever came first in the list.
STEP_HEAD = re.compile(
    r"^\s*\d+\.\s*("
    + "|".join(sorted(FORBIDDEN_VERBS, key=len, reverse=True))
    + r")\b", re.I | re.M)
# A SAFETY NET, NOT THE CRITERION (31 §2). The criterion is §6 — the ER's
# subject must be something the system does — and it stays human-reviewed.
# rev1 shipped `is readable` and rev2 shipped `is recorded`: the same error
# changed words and survived, which is exactly why a word list cannot be
# promoted to a criterion. §9 item 10 may not cite this gate as its basis.
ER_SUBJECT_NET = ("is recorded", "is readable", "is noted", "can be read")
REASONING_SENTENCES = (2, 5)   # §10.4
# Chinese full stops are not followed by a space, so a lookahead for
# whitespace counts a whole paragraph as one sentence — which is how this
# gate first reported "1 sentence" for seven multi-sentence reasonings.
SENT_END = re.compile(r"[。！？]|[.!?](?=\s|$)")
# A magnitude that a clause without numbers cannot justify. Deliberately not
# a blanket digit ban: step numbering ("1.") and quoted spec values are legal.
FABRICATED_QTY = re.compile(r"\b\d+\s*(mm|cm|%|percent|seconds?|secs?|ms|"
                            r"levels?|steps?|degrees?)\b", re.I)


def load_authorities() -> dict:
    cfg = yaml.safe_load((FEATURE / "feature.yaml").read_text("utf-8"))
    wb_path = sorted(FEATURE.glob(cfg["paths"]["workbook"]))[0]
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    vocab = [str(c).strip() for r in wb["下拉選單"].iter_rows(values_only=True)
             for c in r if c and str(c).strip()]
    wb.close()
    outlines = {r["outline"] for r in
                csv.DictReader(LAYER3.open(encoding="utf-8"), delimiter="\t")}
    clauses = {r["outline"]: r["full_text"].replace("\\n", "\n") for r in
               csv.DictReader(FULLTEXT.open(encoding="utf-8"), delimiter="\t")}
    return {"vocab": vocab, "outlines": outlines, "clauses": clauses,
            "test_group": cfg["test_group"],
            "tc_id_re": re.compile(r"^NR1L-ComfortHMI-\d{3}$")}


RC42_QUAL_BLOCK = re.compile(r"```rc42-qualifiers\n(.*?)```", re.S)
RC42_EXC_BLOCK = re.compile(r"```rc42-disposition\n(.*?)```", re.S)
# 65 §1 / §4 — which leaves R-C42 unblocked. This was a tc-number boundary
# (">= 361") until 75 §4 appended two unrelated TCs at 384/385 and the gate
# claimed they were R-C42 unblocks: an inferred boundary is only a fact until
# the next append. It is now an IDENTITY — the leaves batch 16 produces, plus
# the two that stayed in batch 8 — derived from the generators themselves.
def _rc42_leaves() -> set:
    b16 = FEATURE / "scripts" / "gen_batch16.py"
    out = {"SWE1-HVAC-125-08", "SWE1-HVAC-126-02"}
    if b16.exists():
        out |= {f"SWE1-HVAC-{x}" for x in re.findall(
            r'\("(\d{3}(?:-\d{2})?)",', b16.read_text(encoding="utf-8"))}
    return out


RC42_LEAVES = _rc42_leaves()
PENDING_AXIS_BLOCK = re.compile(r"```pending-axis\n(.*?)```", re.S)

MIRROR_MAPS = (("ch16_mirror_map.tsv", "ch16_outline", "ch2_or_ch3_outline"),
               ("ch2_ch7_mirror_map.tsv", "ch7_outline", "ch2_outline"))
# 62 §1.1 (b) — "逐字相同" was asserted 18 times across the two mirror maps and
# never re-measured. R-C40's precondition reads off those labels, so a wrong
# `mirrored` is not a documentation slip: it decides whether a leaf stops.
# 40 characters is a proxy for "a shared clause"; per §5a the SEGMENT is
# printed beside it so the substance, not the number, is what gets read.
MIRROR_MIN_RUN = 40


def _longest_run(a: str, b: str) -> tuple:
    """Longest verbatim run shared by two texts.

    autojunk=False is REQUIRED. difflib's default heuristic drops characters
    appearing in >1% of a sequence longer than 200 elements — on section-length
    English text that is most of the alphabet, and it silently collapsed
    16.13↔2.13's shared 100-character sentence to a run of 1 (62 §1.1's first
    measurement). Every earlier ratio taken on long text is suspect for the
    same reason; see 上繳 41 §3.3.
    """
    sm = difflib.SequenceMatcher(None, a, b, autojunk=False)
    m = sm.find_longest_match(0, len(a), 0, len(b))
    return m.size, a[m.a:m.a + m.size]


def mirror_map_rows() -> list:
    out = []
    for name, ka, kb in MIRROR_MAPS:
        path = FEATURE / "data" / name
        with path.open(encoding="utf-8") as fh:
            for r in csv.DictReader(fh, delimiter="\t"):
                out.append((name, r[ka], r[kb], r["對應強度"]))
    return out


STOPWORDS = {
    "the", "a", "an", "is", "are", "of", "for", "on", "in", "with", "and",
    "or", "to", "that", "this", "such", "one", "those", "these", "its",
    "it", "be", "has", "have", "which", "at", "as", "by", "from", "not",
    # test vocabulary: words that belong to the ACT of testing, not to the
    # clause. A state clause is allowed to say "the vehicle under test".
    "vehicle", "vehicles", "under", "test", "system", "screen", "screens",
}


def two_part_state_words(docs: list) -> list:
    """71 §1 — weak measurement over the readable half of a two-part PC.

    A state clause is by definition a restatement of the verbatim fragment,
    so its content words should already appear in the cited section. Where
    they do not, either the clause introduced a fact the section does not
    carry (§8.4.1) or the restatement used a synonym. Both are worth seeing;
    only the first is a defect, and telling them apart needs a reader.

    Reported, never FAILed: the false-positive rate is measured this round
    (71 §1) before anyone decides whether it can carry a gate.
    """
    out = []
    for d in docs:
        for tc in d["tcs"]:
            for line in tc["pre_conditions"].split("\n"):
                if "[spec-verbatim]" not in line or " — " not in line:
                    continue
                cite = re.search(r"\(([\d.]+)\)\s*$", line.strip())
                src = cite.group(1) if cite else d["outline"]
                body = FULLTEXT_BY_OUTLINE.get(src, "").lower()
                tail = re.sub(r"\s*\([\d.]+\)\s*$", "",
                              line.split(" — ", 1)[1])
                words = [w for w in re.findall(r"[a-zA-Z][a-zA-Z/'-]+",
                                               tail.lower())
                         if w not in STOPWORDS and len(w) > 2]
                missing = [w for w in dict.fromkeys(words) if w not in body]
                out.append((tc["tc_id"], src, tail.strip(), missing))
    return out


def test_setup_longest(docs: list, n: int = 20) -> list:
    """71 §3 — the same question as spec_derived_shortest, asked backwards.

    `spec-derived` is suspicious when the shared run is SHORT (it claims a
    correspondence that may not exist). `test-setup` is suspicious when the
    run is LONG: it claims the line is ours, and a long verbatim overlap
    suggests the clause said it first. Labelling something self-written when
    the spec wrote it is the same class of untruth as labelling a paraphrase
    verbatim — only pointing the other way (R-C41's principle).
    """
    seen = {}
    for d in docs:
        for tc in d["tcs"]:
            for line in tc["pre_conditions"].split("\n"):
                if "[test-setup]" not in line:
                    continue
                cite = re.search(r"\(([\d.]+)\)\s*$", line.strip())
                src = cite.group(1) if cite else d["outline"]
                frag = re.sub(r"^\d+\.\s*\[test-setup\]\s*", "", line)
                frag = re.sub(r"\s*\([\d.]+\)\s*$", "", frag).strip()
                key = (src, frag)
                if key in seen:
                    seen[key][1] += 1
                    continue
                size, seg = _longest_run(frag, FULLTEXT_BY_OUTLINE.get(src, ""))
                seen[key] = [size, 1, tc["tc_id"], seg]
    rows = [(v[0], src, frag, v[1], v[2], v[3])
            for (src, frag), v in seen.items()]
    return sorted(rows, reverse=True)[:n]


def spec_derived_shortest(docs: list, n: int = 20) -> list:
    """70 §3 — a MEASUREMENT over `[spec-derived]` PC lines, never a FAIL.

    A paraphrase may legitimately share almost no wording with its clause, so
    a threshold here would produce mass false positives, and a gate that cries
    wolf is a gate nobody reads. What can be done honestly is to SORT: the
    lines least anchored in their cited section, printed every run, shortest
    first. If a line ever claims a correspondence that does not exist, this is
    where it will surface.

    Grouped by (section, wording) rather than printed per TC: the same
    exclusion line is written on hundreds of rows, and 20 copies of one claim
    is a list of one. The count travels with each entry so nothing is hidden
    by the grouping. (Deviation from 70 §3's literal "20 lines", reported.)
    """
    seen = {}
    for d in docs:
        for tc in d["tcs"]:
            for line in tc["pre_conditions"].split("\n"):
                if "[spec-derived]" not in line:
                    continue
                cite = re.search(r"\(([\d.]+)\)\s*$", line.strip())
                src = cite.group(1) if cite else d["outline"]
                frag = re.sub(r"^\d+\.\s*\[spec-derived\]\s*", "", line)
                frag = re.sub(r"\s*\([\d.]+\)\s*$", "", frag).strip()
                key = (src, frag)
                if key in seen:
                    seen[key][1] += 1
                    continue
                size, seg = _longest_run(frag, FULLTEXT_BY_OUTLINE.get(src, ""))
                seen[key] = [size, 1, tc["tc_id"], seg]
    rows = [(v[0], src, frag, v[1], v[2], v[3])
            for (src, frag), v in seen.items()]
    return sorted(rows)[:n]


def identical_tc_groups(docs: list) -> list:
    """61 §2 — corpus-wide scan for TCs identical in the three content fields.

    A MEASUREMENT, not a gate. Found by accident in round 61: four pairs had
    been recorded as equivalent while NINE more existed, because equivalence
    had only ever been looked for where a handoff pointed. It does not FAIL —
    037's decomposition granularity is upstream's fact and §8.2.2 forbids
    merging leaves — but it prints, so it cannot be unseen again.

    62 §2 — `pre_conditions` is REPORTED, never part of the key. Putting it
    in the key would hide the groups whose PCs differ slightly, and those are
    the ones most worth seeing (the same behaviour repeated across two sets
    of circumstances). So the key stays at three fields and each group gets a
    fourth datum: same / partial / no. `106-04` ≡ `110-06`'s PC identity was
    measured BY HAND in round 61; this is that hand-check mechanised.

    Returns [(members, pc_verdict)] where members is [(outline, tc_id,
    req_id)] and pc_verdict is "same" (character-identical), "partial" (same
    set of lines, different order — still one vehicle, one test) or "no".
    """
    ident = {}
    for d in docs:
        for tc in d["tcs"]:
            ident.setdefault(
                (tc["test_item"], tc["test_procedure"], tc["expected_result"]),
                []).append((d["outline"], tc["tc_id"], tc["req_id"],
                            tc["pre_conditions"]))
    out = []
    for members in ident.values():
        if len(members) < 2:
            continue
        pcs = [m[3] for m in members]
        if len(set(pcs)) == 1:
            verdict = "same"
        elif len({frozenset(
                l.split(". ", 1)[-1] for l in pc.split("\n") if l.strip())
                for pc in pcs}) == 1:
            verdict = "partial (same lines, different order)"
        else:
            verdict = "no"
        out.append(([m[:3] for m in members], verdict))
    return out


def lint(docs: list, auth: dict) -> list[tuple[str, str, str]]:
    """-> [(severity, gate, message)]"""
    out = []

    def bad(gate, msg):
        out.append(("FAIL", gate, msg))

    all_tcs = [(d, tc) for d in docs for tc in d["tcs"]]

    # R-C24: a BLOCKED row is exempt from the two gates that assume a
    # procedure exists. The exemption is REPORTED as its own line, never a
    # silent skip inside a condition — the precedent being the
    # `and n != "Comfort Widget"` that once hid a real naming defect behind a
    # green check (upstream 06 §2.1).
    blocked = [tc["tc_id"] for _, tc in all_tcs
               if any(m in tc.get("remarks", "") for m in BLOCKED_MARKERS)]

    # ---- id gates -------------------------------------------------------
    ids = [tc["tc_id"] for _, tc in all_tcs]
    for tid in ids:
        if not auth["tc_id_re"].match(tid):
            bad("tc-id-format", f"{tid!r} does not match NR1L-ComfortHMI-NNN (R-C7)")
    dup = [i for i, n in Counter(ids).items() if n > 1]
    if dup:
        bad("tc-id-unique", f"duplicate tc_ids {dup}")
    nums = sorted(int(i.rsplit("-", 1)[1]) for i in ids)
    if nums != list(range(1, len(nums) + 1)):
        bad("tc-id-sequence", f"tc_id numbers are not 1..{len(nums)} gap-free: {nums}")
    # §8.2.2 lets one leaf produce several TCs, so a repeated req_id is legal
    # — but ONLY when every row carrying it declares the split. An accidental
    # duplicate and a declared split look identical in the id column; the
    # split_flag is what tells them apart, so the gate reads it rather than
    # dropping the uniqueness check altogether.
    rids = [tc["req_id"] for _, tc in all_tcs]
    for rid, n in sorted(Counter(rids).items()):
        if n == 1:
            continue
        rows = [tc for _, tc in all_tcs if tc["req_id"] == rid]
        undeclared = [tc["tc_id"] for tc in rows if not tc.get("split_flag")]
        if undeclared:
            bad("req-id-unique",
                f"req_id {rid!r} appears on {n} TCs but {undeclared} do not "
                "set split_flag — a split must be declared on every row it "
                "produces (§8.2.2), otherwise this is a duplicate")
        elif not all(tc.get("split_reason") for tc in rows):
            bad("req-id-unique",
                f"req_id {rid!r} is split {n} ways but a row has an empty "
                "split_reason (§10.1)")

    for d, tc in all_tcs:
        w = tc["tc_id"]
        clause = auth["clauses"].get(d["outline"], "")

        # ---- §10.1 required keys ----------------------------------------
        missing = [k for k in REQUIRED_KEYS if k not in tc]
        if missing:
            bad("required-keys", f"{w}: missing §10.1 key(s) {missing}")

        # ---- §10.5 minimum two numbered procedure steps ------------------
        # proc-er-1to1 does NOT cover this: a single step against a single ER
        # line is 1:1 and passes. That is how TC-004 reached "25/25 PASS"
        # with a one-step procedure (handoff 20 §1.1).
        steps = len(re.findall(r"^\s*\d+\.", tc["test_procedure"], re.M))
        if w in blocked:
            # A BLOCKED row must be EMPTY, not merely short — a stray step
            # would mean the row is half-written rather than blocked.
            if tc["test_procedure"] or tc["expected_result"]:
                bad("blocked-row-empty",
                    f"{w}: carries a BLOCKED marker but "
                    "test_procedure/expected_result are not empty (R-C24)")
        elif steps < MIN_STEPS:
            bad("proc-min-steps",
                f"{w}: {steps} numbered step(s), §10.5 requires >= {MIN_STEPS} "
                "(Setup -> Verification)")

        # ---- spec reference (R-C1) --------------------------------------
        # R-C29 lets one TC cite several sections — its own, plus any section
        # a cross-section pre_condition draws its fact from (§10.7 "relied on
        # as setup"). Items are "; "-separated and each carries the full stem,
        # so the stem rule stays per-item rather than being checked once and
        # assumed for the rest.
        ref = tc["specification_reference"]
        refs = [r.strip() for r in ref.split(";") if r.strip()]
        if len(refs) != len(set(refs)):
            bad("spec-ref-outline",
                f"{w}: specification_reference repeats a section: {refs}")
        for item in refs:
            if item in EXTERNAL_REFS:      # R-C45 三 —— 已認可之外部出處
                continue
            if not item.startswith(STEM + "_"):
                bad("spec-ref-stem",
                    f"{w}: stem is not the ruled SR24 filename (R-C1), nor an "
                    f"approved external source (R-C45) — {item!r}")
                continue
            outline = item[len(STEM) + 1:]
            if outline not in auth["outlines"]:
                bad("spec-ref-outline",
                    f"{w}: {outline!r} is not one of the 129 cited sections")
        if refs and not refs[0].endswith("_" + d["outline"]):
            # The TC's own section leads; cited-for-setup sections follow.
            bad("spec-ref-outline",
                f"{w}: first specification_reference is not the TC's own "
                f"section {d['outline']!r} (R-C29: own section leads)")
        # R-C1 forbids SR25 **as the Comfort baseline**. 89 §3 approved citing
        # an external document whose own version happens to be SR25 (the HMI
        # Settings List), so the check is now per-segment and exempts approved
        # external sources by identity — not by "does the string contain SR25
        # anywhere", which conflated two different documents' version numbers.
        for item in refs:
            if item in EXTERNAL_REFS:
                continue
            if "SR25" in item:
                bad("spec-ref-sr25",
                    f"{w}: a Comfort-spec reference names SR25 (R-C1 forbids) "
                    f"— {item!r}")

        # ---- fixed columns ----------------------------------------------
        if tc["test_group"] != auth["test_group"]:
            bad("test-group", f"{w}: test_group {tc['test_group']!r} != feature.yaml")
        if tc["design_method"] not in auth["vocab"]:
            bad("design-method", f"{w}: {tc['design_method']!r} not in 下拉選單 vocabulary")
        if tc["priority"] not in PRIORITIES:
            bad("priority", f"{w}: {tc['priority']!r} not in {sorted(PRIORITIES)}")
        if tc["functional_safety"] != "NA":
            bad("functional-safety", f"{w}: S column must be 'NA' (profile §3.8)")
        if tc["estimated_test_time"] != "":
            bad("estimated-time", f"{w}: Q column must be blank (profile §3.7)")
        # ---- R-C26 marker whitelist -------------------------------------
        for mk, allowed in MARKER_WHITELIST.items():
            if mk in tc.get("remarks", "") and w not in allowed:
                bad("marker-whitelist",
                    f"{w}: carries {mk} but is not in profile §5.1's named "
                    "whitelist; an exemption-granting marker cannot be "
                    "self-issued (R-C26)")

        if w in blocked:
            mark = next((m for m in BLOCKED_MARKERS
                         if tc["remarks"].startswith(m)), None)
            if mark is None:
                bad("blocked-remarks",
                    f"{w}: one of {list(BLOCKED_MARKERS)} must be the leading "
                    "token of Remarks (R-C24 / R-C38)")
            if re.search(r"\bA-CF\d+\b|\bR-C\d+\b|§\d", tc["remarks"]):
                bad("blocked-remarks", f"{w}: Remarks is externally visible and "
                                       "must not carry an internal ruling id "
                                       "(AMFM R10-4)")
            # R-C27 — the Remarks column clips to one visible line, so what
            # the marker exists to point at must sit inside that line. For
            # [BLOCKED-SPEC] that is the owner; for [BLOCKED-NON-HMI] it is
            # the absence of one, which is why the two checks are opposites
            # rather than one shared check (R-C38: the missing owner is the
            # classification, so a lenient "either is fine" gate would let a
            # delegated leaf hide under the wrong marker).
            head = tc["remarks"][:OWNER_WINDOW]
            if mark == "[BLOCKED-NON-HMI]":
                if NON_HMI_PHRASE not in head:
                    bad("blocked-remarks",
                        f"{w}: {NON_HMI_PHRASE!r} must appear within the first "
                        f"{OWNER_WINDOW} characters of Remarks (R-C38); "
                        f"measured head = {head[:48]!r}")
                if "Owner:" in tc["remarks"]:
                    bad("blocked-remarks",
                        f"{w}: [BLOCKED-NON-HMI] must not name an owner "
                        "(R-C38) — a leaf with an owner is [BLOCKED-SPEC]")
            elif "Owner:" not in head:
                bad("blocked-remarks",
                    f"{w}: 'Owner:' must appear within the first "
                    f"{OWNER_WINDOW} characters of Remarks (R-C27); "
                    f"measured head = {head[:48]!r}")
        elif tc["remarks"] != "":
            # 66 §3 — a non-BLOCKED row MAY carry Remarks for exactly one
            # purpose: an ambiguity in the clause that leaves the tester
            # unable to decide at execution time. reasoning's reader is the
            # reviewer; Remarks' reader is the tester, and the tester is the
            # one who walks into it. Everything else stays empty, and the
            # externally-visible rules still hold (no internal ruling id, no
            # DR number, no A-CF id — profile §3.6).
            forbidden = [t for t in REMARKS_FORBIDDEN if t in tc["remarks"]]
            if forbidden:
                bad("remarks",
                    f"{w}: Remarks is externally visible and may not carry "
                    f"internal identifiers {forbidden} (profile §3.6)")
            elif tc["tc_id"] not in AMBIGUITY_REMARKS:
                bad("remarks",
                    f"{w}: carries Remarks but is not in the ambiguity "
                    f"register — 66 §3 says an ambiguity the tester cannot "
                    f"resolve MUST reach Remarks, and 67 §1 says registering "
                    f"it is part of writing it, not a permission to seek")
            elif len(tc["remarks"]) < REMARKS_MIN_CHARS:
                bad("remarks",
                    f"{w}: registered ambiguity text is {len(tc['remarks'])} "
                    f"characters — too short to tell a tester what cannot be "
                    f"determined (67 §1 三)")
            elif any(ord(ch) > 0x2FF for ch in tc["remarks"]):
                bad("remarks",
                    f"{w}: registered ambiguity text must be in the language "
                    f"the executor reads in this column (English); the row "
                    f"carries non-Latin characters (67 §1 三)")

        # ---- §11 formatting ---------------------------------------------
        for f in LONG_FIELDS:
            for ln in tc[f].split("\n"):
                if ln.rstrip().endswith("."):
                    bad("trailing-period", f"{w}.{f}: line ends with a period — {ln[:48]!r}")
            if re.search(r"\[[A-Za-z][^\]]*\]", tc[f]) and f != "pre_conditions":
                bad("ui-bracket", f"{w}.{f}: square-bracket label; use \"...\" (§11)")

        # ---- tc_title (§4.3) --------------------------------------------
        words = tc["tc_title"].split()
        if not 2 <= len(words) <= 14:
            bad("title-length", f"{w}: tc_title is {len(words)} words, need 2–14")
        if MODALS.search(tc["tc_title"]):
            bad("title-modal", f"{w}: tc_title contains a modal (§4.3)")

        # ---- test_item is the ONLY field allowed a modal (profile §3.1) --
        if not MODALS.search(tc["test_item"]):
            bad("item-modal", f"{w}: test_item carries no modal (profile §3.1)")
        if MODALS.search(tc["expected_result"]):
            bad("er-modal", f"{w}: expected_result contains a modal (§6)")

        # ---- pre_conditions source class (profile §3.2) ------------------
        for ln in [l for l in tc["pre_conditions"].split("\n") if l.strip()]:
            if not any(f"[{c}]" in ln for c in SOURCE_CLASSES):
                bad("source-class", f"{w}: pre_condition without source class — {ln[:52]!r}")

        # ---- procedure <-> ER 1:1 (§6) ----------------------------------
        np_ = len(re.findall(r"^\s*\d+\.", tc["test_procedure"], re.M))
        ne = len(re.findall(r"^\s*\d+\.", tc["expected_result"], re.M))
        if w in blocked:
            pass                      # exempt; reported on its own line
        elif np_ != ne or np_ == 0:
            bad("proc-er-1to1", f"{w}: {np_} procedure steps vs {ne} ER lines")

        # ---- §3.4 (-, +) placement (19 §3) -------------------------------
        if "(-, +)" in tc["test_procedure"]:
            bad("token-placement", f"{w}: '(-, +)' in test_procedure; use \"-\" / \"+\" (§3.4)")
        if "(-, +)" in tc["test_item"] and "(-, +)" not in clause:
            bad("token-source", f"{w}: '(-, +)' in test_item but absent from clause {d['outline']}")

        # ---- R-C22: no fabricated magnitude where the clause gives none --
        clause_has_num = bool(re.search(r"\d", clause))
        m = FABRICATED_QTY.search(tc["expected_result"])
        if m and not clause_has_num:
            bad("fabricated-qty", f"{w}: ER states {m.group(0)!r} but clause has no number (R-C22/§8.4.1)")

    # ---- §10.4 reasoning length / §10.6 duplicate_of encoding ------------
    for d in docs:
        n = len(SENT_END.findall(d.get("reasoning", "")))
        lo, hi = REASONING_SENTENCES
        if not lo <= n <= hi:
            bad("reasoning-sentences",
                f"{d['parent']}: reasoning has {n} sentence(s), §10.4 requires {lo}-{hi}")
        dup = d.get("duplicate_of", "")
        if dup and not re.fullmatch(r"\d+", dup):
            bad("duplicate-of-format",
                f"{d['parent']}: duplicate_of {dup!r} is not a digits-only row "
                "number (§10.6)")

    # ---- §4.6 sibling bookkeeping ---------------------------------------
    for d in docs:
        axis = d["distinguishing_axis"]["axis"]
        dup_of = d["duplicate_of"]
        if (axis == "none") != bool(dup_of):
            bad("sibling-axis", f"{d['parent']}: axis={axis!r} but duplicate_of={dup_of!r} "
                                "(§4.6 requires axis='none' <=> duplicate_of set)")

    # ---- §5.1 forbidden main verbs (31 §1) ------------------------------
    for _, tc in all_tcs:
        for m in STEP_HEAD.finditer(tc["test_procedure"]):
            step = tc["test_procedure"][m.start():].split("\n", 1)[0]
            bad("forbidden-verb",
                f"{tc['tc_id']}: step leads with the forbidden main verb "
                f"{m.group(1)!r} (§5.1) — {step[:64]!r}")

    # ---- ER subject safety net (31 §2) ----------------------------------
    for _, tc in all_tcs:
        for phrase in ER_SUBJECT_NET:
            if phrase in tc["expected_result"].lower():
                bad("er-subject-net",
                    f"{tc['tc_id']}: expected_result contains {phrase!r}, "
                    "which puts the observer in the subject position (§6). "
                    "This gate is a net, not the criterion — passing it does "
                    "not mean the ER subjects were checked")

    # ---- 36 §6 — R-C34's generation-time duty was discharged ------------
    for d in docs:
        rev = d.get("interface_axis_review")
        if not isinstance(rev, dict):
            bad("interface-axis-answered",
                f"{d['parent']} ({d['outline']}): no interface_axis_review; "
                "R-C34 requires naming the observable's interface and "
                "answering each interface-type axis")
            continue
        blank = [k for k in IFACE_KEYS if not str(rev.get(k, "")).strip()]
        if blank:
            bad("interface-axis-answered",
                f"{d['parent']} ({d['outline']}): interface_axis_review is "
                f"missing or empty for {blank}")

    # ---- 38 §1 / R-C36-1 — every EMEA exclusion carries a per-TC answer ---
    # The gate fails only on a MISSING answer. A "no" verdict is reported on
    # its own named line instead: removal is a ruling, not something lint may
    # force, and a red build would push the executor into removing it unilaterally.
    emea_no = []
    for _, tc in all_tcs:
        if "not an EMEA ICS" not in tc["pre_conditions"]:
            continue
        rev = tc.get("emea_ics_review")
        if not isinstance(rev, dict) or not str(rev.get("ch16_sentence", "")).strip():
            bad("emea-per-tc-answered",
                f"{tc['tc_id']}: carries the EMEA exclusion but has no "
                "per-TC judgement pointing at a ch16 sentence (R-C36-1); "
                "section-level `mirrored` is not an answer")
        elif rev.get("verdict") != "yes":
            emea_no.append((tc["tc_id"], rev.get("verdict"),
                            rev.get("ch16_outline")))

    # ---- 60 §1 — tc_id may not be used to NAME another row in prose ------
    # This is a PROHIBITION, not a correctness check. A "the cited tc_id
    # exists and matches its description" gate is mechanically possible for
    # the first half and useless (after a shift `-227` still exists, it just
    # points elsewhere) and unmechanisable for the second. R-C7 makes tc_id a
    # generator-assigned, shifting key; req_id is stable. So prose cites
    # req_id, full stop, and this gate needs no semantics to be reliable.
    #
    # TWO patterns, not one. 60 §1 specifies `NR1L-ComfortHMI-\d+` — measured
    # across the corpus that form appears TWICE, while the SHORT form
    # (`-233` in backticks, the house style) appears 132 times and is the
    # form that actually broke in 58 §2. A gate matching only the spelled-out
    # form would have passed on the very corpus that motivated it.
    for d in docs:
        prose = {"reasoning": d.get("reasoning", ""),
                 "distinguishing_axis.axis": d.get(
                     "distinguishing_axis", {}).get("axis", ""),
                 "distinguishing_axis.delta": d.get(
                     "distinguishing_axis", {}).get("delta", ""),
                 "assumptions": " ".join(d.get("assumptions", []) or [])}
        for tc in d["tcs"]:
            prose[f"{tc['tc_id']}.split_reason"] = tc.get("split_reason") or ""
            prose[f"{tc['tc_id']}.remarks"] = tc.get("remarks") or ""
        for field, text in prose.items():
            hits = TCID_LONG.findall(text) + TCID_SHORT.findall(text)
            if hits:
                bad("no-tcid-in-prose",
                    f"{d['outline']} ({field}): cites tc_id {sorted(set(hits))} "
                    f"in prose — profile §3.6 requires req_id, because tc_id "
                    f"moves (R-C7) and a moved citation still parses")

    # ---- 69 §2 (執行層自加，已報備) — pre_conditions lines are numbered ----
    # Found while reading a source-class-truthful FAIL message: five TCs had an
    # unnumbered line 2 while lines 3-5 were numbered, because a generator
    # joined one extra line by hand instead of through add_lines. Both ends
    # looked right; the middle was wrong. Cheap to check, and the numbering is
    # what the workbook's reader counts by.
    for d in docs:
        for tc in d["tcs"]:
            lines = [l for l in tc["pre_conditions"].split("\n") if l.strip()]
            got = [re.match(r"^(\d+)\.\s", l) for l in lines]
            if not all(got) or [int(m.group(1)) for m in got] != list(
                    range(1, len(lines) + 1)):
                shown = [l[:30] for l in lines]
                bad("pc-line-numbering",
                    f"{tc['tc_id']}: pre_conditions lines are not numbered "
                    f"1..n — {shown}")

    # ---- 69 §1.2 — a source class is a CLAIM about provenance -----------
    # 68 §4 declined to make "the PC must quote verbatim" a corpus-wide
    # obligation, and that still holds. This is the other thing: whatever the
    # author claims must be true. A line labelled [spec-verbatim] that is a
    # paraphrase leaves its reader unable to tell whether the sentence is the
    # spec's or ours — R-C41's principle, applied to provenance instead of to
    # method. Measured 69 §1: 150 of 155 such lines were paraphrases.
    #
    # Checked LINE BY LINE (69 §1.3): the label is per line, so a first-line
    # check would leave every later [spec-verbatim] unexamined for ever.
    def _norm_ws(t: str) -> str:
        return re.sub(r"\s+", " ", t).strip()

    for d in docs:
        for tc in d["tcs"]:
            for line in tc["pre_conditions"].split("\n"):
                if "[spec-verbatim]" not in line:
                    continue
                cite = re.search(r"\(([\d.]+)\)\s*$", line.strip())
                src = cite.group(1) if cite else d["outline"]
                frag = re.sub(r"^\d+\.\s*\[spec-verbatim\]\s*", "", line)
                frag = re.sub(r"\s*\([\d.]+\)\s*$", "", frag)
                # 70 §2 — two-part form: `<verbatim> — <readable state>`.
                # Only what precedes the em dash claims to be the clause's own
                # words; the state clause after it is ours, and §8.4.1 still
                # forbids it introducing facts the clause does not carry.
                frag = _norm_ws(frag.split(" — ")[0])
                body = _norm_ws(FULLTEXT_BY_OUTLINE.get(src, ""))
                if not body:
                    bad("source-class-truthful",
                        f"{tc['tc_id']}: a [spec-verbatim] line cites {src}, "
                        f"which has no full_text to be verbatim against")
                elif frag not in body:
                    bad("source-class-truthful",
                        f"{tc['tc_id']}: line labelled [spec-verbatim] is not "
                        f"a contiguous quotation of {src} — {frag[:56]!r}. "
                        f"Paraphrasing is allowed; calling it verbatim is not "
                        f"(69 §1.2: label [spec-derived] instead, or quote)")

    # ---- 67 §1 二 — the register reconciles in BOTH directions ----------
    # 44 §7.3 named this gap: a row listed in the register whose Remarks was
    # later emptied leaves the register claiming an ambiguity that no longer
    # reaches the tester. Same failure mode as moved-leaf-identity: the count
    # is right and the identity is not.
    # 68 §2 — two rows may not be reconciled against the same fragment.
    dupes = [v for v in set(AMBIGUITY_REMARKS.values())
             if list(AMBIGUITY_REMARKS.values()).count(v) > 1]
    if dupes:
        bad("ambiguity-register",
            f"the register reuses the fragment(s) {dupes} on more than one "
            f"row — a shared fragment only checks that the sentence exists, "
            f"never that it is about this row (68 §2)")
    by_id = {tc["tc_id"]: tc for d in docs for tc in d["tcs"]}
    for tc_id, phrase in sorted(AMBIGUITY_REMARKS.items()):
        tc = by_id.get(tc_id)
        if tc is None:
            bad("ambiguity-register",
                f"{tc_id} is in the ambiguity register but no such TC exists")
        elif phrase not in tc["remarks"]:
            bad("ambiguity-register",
                f"{tc_id} is registered as carrying a clause ambiguity, but "
                f"its Remarks does not contain {phrase!r} — the register "
                f"would go on asserting an ambiguity the tester never sees")

    # ---- 65 §3 — an R-C42 unblock must quote a QUALIFIED clause ---------
    # The judgement "this sentence is a condition, not a statement" is
    # [manual] and has no mechanical criterion. This does not mechanise it;
    # it makes the gap visible between "I read it as a condition" and "it
    # appears with a condition's syntax". The contrast case is 2.11's
    # "Adjusting … will alter the Front and Rear passengers" — an unqualified
    # statement, whose leaves stayed stopped.
    prof_txt = PROFILE.read_text(encoding="utf-8")
    qual_raw = RC42_QUAL_BLOCK.findall(prof_txt)
    markers = []
    for raw in qual_raw:
        for line in raw.strip().split("\n"):
            if line.startswith("markers:"):
                markers = [m.strip() for m in line.split(":", 1)[1].split("|")
                           if m.strip()]
    if not markers:
        bad("rc42-condition-marker",
            "profile carries no ```rc42-qualifiers``` block, so every R-C42 "
            "unblock would pass unexamined")
    exceptions = {}
    rc42_dispositions = []
    for raw in RC42_EXC_BLOCK.findall(prof_txt):
        f = {}
        for line in raw.strip().split("\n"):
            if ":" in line and not line.startswith(" "):
                k, v = line.split(":", 1)
                f[k.strip()] = v.strip()
        if f.get("req_id"):
            exceptions[f["req_id"]] = f
    for d in docs:
        for tc in d["tcs"]:
            if tc["req_id"] not in RC42_LEAVES:
                continue
            head = tc["pre_conditions"].split("\n")[0]
            cite = re.search(r"\(([\d.]+)\)\s*$", head)
            src = cite.group(1) if cite else d["outline"]
            text = FULLTEXT_BY_OUTLINE.get(src, "")
            frag = re.sub(r"^\d+\.\s*\[[a-z-]+\]\s*", "", head)
            frag = re.sub(r"\s*\([\d.]+\)\s*$", "", frag)
            size, seg = _longest_run(frag, text)
            key = tc["req_id"].replace("SWE1-HVAC-", "")
            if size < 15:
                bad("rc42-condition-marker",
                    f"{key}: its first pre_condition shares only {size} "
                    f"character(s) with {src} — an R-C42 unblock must QUOTE "
                    f"the clause it relies on")
                continue
            at = text.find(seg)
            before = text[max(0, at - 70):at]
            hit = any(re.search(rf"\b{re.escape(m)}\b", before)
                      or seg.lstrip().startswith(m) for m in markers)
            if hit:
                continue
            # 66 §2.2 — a miss no longer blocks; it DEMANDS a named
            # disposition. Blocking would stop a real condition for having
            # the wrong syntax, which is what happened on this gate's first
            # run (125-08 / 126-02). Silence still fails.
            exc = exceptions.get(key)
            if not exc or not (exc.get("condition")
                               or exc.get("not-a-condition")):
                bad("rc42-condition-marker",
                    f"{key} ({src}): the quoted fragment {seg[:40]!r} is not "
                    f"introduced by any listed qualifier and carries no named "
                    f"disposition — profile §3.2.2 requires `condition:` (it "
                    f"is a condition, with its verbatim fragment) or "
                    f"`not-a-condition:` (it is not; the leaf goes back to "
                    f"stopped). Silence is the one answer not available")
            else:
                rc42_dispositions.append(
                    (key, src, "condition" if exc.get("condition")
                     else "not-a-condition", exc.get("source", "—")))

    # ---- 65 §4 — MOVED_TO_BATCH16 is an identity, not just a count -------
    # The existing arithmetic (emitted + withheld + moved = framework.md)
    # only checks the NUMBER. A leaf declared moved that batch 16 never
    # produced still balances.
    declared_moved = set()
    for gen in sorted((FEATURE / "scripts").glob("gen_batch*.py")):
        text = gen.read_text(encoding="utf-8")
        m = re.search(r"MOVED_TO_BATCH16 = (\[[^\]]*\])", text)
        if m:
            declared_moved |= set(re.findall(r"'(SWE1-HVAC-[\d-]+)'", m.group(1))
                                  + re.findall(r'"(SWE1-HVAC-[\d-]+)"', m.group(1)))
    b16 = FEATURE / "scripts" / "gen_batch16.py"
    produced_16 = set()
    if b16.exists():
        b16_leaves = re.findall(r'\("(\d{3}(?:-\d{2})?)",', b16.read_text(
            encoding="utf-8"))
        produced_16 = {f"SWE1-HVAC-{x}" for x in b16_leaves}
    for req in sorted(declared_moved - produced_16):
        bad("moved-leaf-identity",
            f"{req} is declared in a MOVED_TO_BATCH16 list but gen_batch16 "
            f"does not produce it — the count still balances, which is "
            f"exactly why the count alone is not enough (65 §4)")
    for req in sorted(produced_16 - declared_moved):
        bad("moved-leaf-identity",
            f"{req} is produced by gen_batch16 but no generator declares it "
            f"moved — its原 batch's leaf arithmetic is silently short")

    # ---- R-C42 三 (64 §1) — a clause-local condition used in >=2 sections
    # must have a NAMED disposition. R-C42's wording says "register it as an
    # axis"; two of the four candidates cannot be registered without inventing
    # a value the spec never states (§8.4.1 — exactly what the three
    # conditions protect). So the gate demands a disposition, not a
    # registration: `registered: 第 N 軸` or `deferred: DR #x`. Silence FAILs.
    # The deviation and its reason are recorded in profile §3.2.1 for ruling.
    prof = PROFILE.read_text(encoding="utf-8")
    candidates = []
    for raw in PENDING_AXIS_BLOCK.findall(prof):
        f = {}
        for line in raw.strip().split("\n"):
            if ":" in line and not line.startswith(" "):
                k, v = line.split(":", 1)
                f[k.strip()] = v.strip()
        candidates.append(f)
    seen_sections = {}
    for f in candidates:
        name = f.get("condition", "?")
        pattern = f.get("pattern", "")
        declared = [x.strip() for x in f.get("sections", "").split("|")
                    if x.strip()]
        # Reach is measured over the PRE_CONDITIONS actually written, not the
        # section text: the condition's risk is that two TCs state it
        # differently, and only a written PC can do that. (First cut measured
        # section text and mis-declared three of four candidates — 9.2/9.4
        # never repeat 9.1's sentence, they say "in these variants".)
        hits = sorted({d["outline"] for d in docs for tc in d["tcs"]
                       if pattern and pattern.lower()
                       in tc["pre_conditions"].lower()},
                      key=lambda s2: [int(x) for x in s2.split(".")])
        seen_sections[name] = hits
        if set(hits) != set(declared):
            bad("axis-candidate-registered",
                f"pending-axis {name!r}: profile declares sections "
                f"{declared} but the corpus measures {hits} — a candidate's "
                f"reach must be measured, not remembered")
        if len(hits) >= 2 and not f.get("disposition"):
            bad("axis-candidate-registered",
                f"pending-axis {name!r} appears in {len(hits)} sections "
                f"{hits} with no disposition — R-C42 三: once the condition "
                f"is in two places, the risk of them diverging is real and "
                f"someone must own it")

    # ---- 63 §1 — an equivalence group's sections must be IN the table ----
    # sibling_candidates.py excludes same-Test-Set pairs by design (its
    # `group[a] == group[b]` branch). Part N merged ch11 and ch12 into one
    # set, so the corpus's strongest equivalences — two sections whose TCs are
    # character-identical — were never candidates and never judged. Merging
    # two chapters into one set silently removed them from sibling detection.
    # The scan finds them; this makes the table carry them.
    in_table = {frozenset((r["outline"], r["sibling_outline"]))
                for r in SIBLING_TABLE}
    for members, _pc in identical_tc_groups(docs):
        sections = frozenset(m[0] for m in members)
        if len(sections) > 1 and sections not in in_table:
            bad("equivalence-in-sibling-table",
                f"{sorted(sections)} produce character-identical TCs "
                f"({[m[2] for m in members]}) but the pair is absent from "
                f"pending_sibling.tsv — §4.6 owes it a verdict, and a pair "
                f"the candidate generator cannot reach is exactly the pair "
                f"nobody will notice")

    # ---- 62 §1.1 (b) — the mirror maps' "逐字相同" claims, re-measured ----
    mirror_measured = []
    for name, a, b, kind in mirror_map_rows():
        for o in (a, b):
            if o != "no-counterpart" and o not in FULLTEXT_BY_OUTLINE:
                bad("mirror-map-verified",
                    f"{name}: outline {o} is not in section_fulltext.tsv")
        if kind != "mirrored" or a == "no-counterpart" or b == "no-counterpart":
            continue
        if a not in FULLTEXT_BY_OUTLINE or b not in FULLTEXT_BY_OUTLINE:
            continue
        size, seg = _longest_run(FULLTEXT_BY_OUTLINE[a], FULLTEXT_BY_OUTLINE[b])
        mirror_measured.append((name, a, b, size, seg))
        if size < MIRROR_MIN_RUN:
            bad("mirror-map-verified",
                f"{name}: {a} ↔ {b} is labelled `mirrored`, but the longest "
                f"verbatim run they share is {size} character(s) "
                f"({seg[:40]!r}). R-C40 reads this label to decide whether a "
                f"leaf stops, so an unmeasured `mirrored` is not cosmetic")

    # ---- 62 §1.1 (b) — a leaf is either withheld or produced, never both ----
    produced = {tc["req_id"] for d in docs for tc in d["tcs"]}
    declared_withheld = set()
    for gen in sorted((FEATURE / "scripts").glob("gen_*.py")):
        declared_withheld |= set(WITHHELD_DECL.findall(
            gen.read_text(encoding="utf-8")))
    for req in sorted(declared_withheld & produced):
        bad("withheld-not-generated",
            f"{req} is declared in a generator's WITHHELD list AND produced "
            f"as a TC — a stopped leaf that is also delivered is the one "
            f"state stop-and-report may never be in")
    for req in sorted(declared_withheld - set(LEAF_UNIVERSE)):
        bad("withheld-not-generated",
            f"{req} is declared withheld but 037 carries no such leaf")

    # ---- 62 §5 — a req_id cited in prose must EXIST -------------------
    # 60 §1 removed the citation key that MOVES; this removes the failure it
    # left behind — a key that never pointed anywhere. Verified once by
    # verify_no_tcid_gate.py in round 60, which is exactly the kind of
    # one-shot check R-C41 says may not be quoted as "verified" later.
    #
    # The universe is 037's own leaf set (403, from recon.json), which already
    # includes section-level leaves (e.g. SWE1-HVAC-037) and every withheld
    # leaf — stopped leaves are cited on purpose and must not fail here.
    valid_reqs = {r.replace("SWE1-HVAC-", "") for r in LEAF_UNIVERSE}
    for d in docs:
        prose = {"reasoning": d.get("reasoning", ""),
                 "distinguishing_axis.delta": d.get(
                     "distinguishing_axis", {}).get("delta", ""),
                 "assumptions": " ".join(d.get("assumptions", []) or [])}
        for tc in d["tcs"]:
            prose[f"{tc['tc_id']}.split_reason"] = tc.get("split_reason") or ""
        for field, text in prose.items():
            cited = set(REQ_CITE.findall(text))
            missing = sorted(c for c in cited if c not in valid_reqs)
            if missing:
                bad("prose-reqid-exists",
                    f"{d['outline']} ({field}): cites req_id {missing}, "
                    f"which 037 does not carry — profile §3.6.1's key must "
                    f"point at a real leaf, or it points nowhere at all")

    # ---- 36 §4 — a pending sibling must be resolved once its section lands -
    generated = {d["outline"] for d in docs}
    for outline, sibling in sorted(PENDING_SIBLING):
        doc = next((d for d in docs if d["outline"] == outline), None)
        if doc is None or sibling not in generated:
            continue
        unresolved = [tc["tc_id"] for tc in doc["tcs"]
                      if not doc.get("duplicate_of")
                      and doc["distinguishing_axis"]["axis"] in
                      ("", "see per-TC titles")]
        if unresolved:
            bad("pending-sibling",
                f"{outline}'s sibling {sibling} is now generated, but "
                f"duplicate_of/distinguishing_axis is still unset for "
                f"{unresolved} — §4.6 判定須於對造節生成後回填")

    # ---- 42 §1 / 43 §1 — provisional verdicts owed a second look --------
    # TRIGGER CORRECTED (43 §1). 42 §1 fired when EITHER side's Test Set
    # completed — but `provisional` is caused by ONE side having no TCs, and
    # the side that completes is the side that already had them. The missing
    # evidence stayed missing, so the re-confirmation had nothing new to look
    # at: measured 632 rows due, 0 of them with both sides generated.
    #
    # 43 §1 names the discriminant outright: `provisional == true` AND both
    # sides generated <=> the side that was missing has landed. That is the
    # condition below. It is section-granular rather than Test-Set-granular,
    # so it fires the moment the evidence exists rather than at the coarser
    # set boundary — earlier, never later (上繳 32 §1.2).
    due = [r for r in SIBLING_TABLE
           if r.get("provisional") == "true"
           and r["outline"] in generated
           and r["sibling_outline"] in generated]
    if due:
        shown = ", ".join(f"{r['outline']}<->{r['sibling_outline']}"
                          f"[{r['verdict']}]" for r in due[:8])
        bad("provisional-sibling",
            f"{len(due)} provisional row(s) now have BOTH sides generated — "
            f"the side that was missing when the verdict was reached has "
            f"landed, so the verdict is owed a re-confirmation against TCs "
            f"rather than clauses (43 §1). Re-confirm — the verdict MAY "
            f"stand — then set provisional=false. First 8: {shown}"
            + (f" … and {len(due) - 8} more" if len(due) > 8 else ""))

    # ---- 35 §4 / 43 §4 — every negated axis, not just axis 13 ------------
    blocks = AXIS_BLOCK.findall(PROFILE.read_text(encoding="utf-8"))
    if not blocks:
        bad("axis-value-count",
            "profile carries no ```axis-values``` block; a negated "
            "pre_condition cannot be checked against the axis it negates")
    negations = []
    for raw in blocks:
        f = dict(l.split(":", 1) for l in raw.strip().split("\n")
                 if ":" in l and not l.lstrip().startswith("#"))
        axis = f.get("axis", "?").strip().split()[0]
        values = [v.strip() for v in f.get("values", "").split("|") if v.strip()]
        declared = f.get("value-count", "").strip()
        reviewed = f.get("negation-reviewed-at-value-count", "").strip()
        negation = f.get("negation", "").strip()
        listed = [v.strip() for v in f.get("negation-users", "").split(",")
                  if v.strip()]
        if not negation:
            bad("axis-value-count",
                f"axis {axis}: block carries no `negation:` field, so the "
                "gate cannot find the pre_conditions it protects (43 §4)")
            continue
        negations.append(negation)
        actual = [tc["tc_id"] for _, tc in all_tcs
                  if negation in tc["pre_conditions"]]
        if declared != str(len(values)):
            bad("axis-value-count",
                f"axis {axis}: profile declares value-count {declared!r} but "
                f"lists {len(values)} values {values}")
        elif reviewed != declared:
            bad("axis-value-count",
                f"axis {axis}: gained a value (now {declared}) but the negated "
                f"pre_condition was last reviewed at {reviewed!r}. "
                f"Re-review these {len(actual)} TCs and then bump "
                f"negation-reviewed-at-value-count: {sorted(actual)}")
        if sorted(listed) != sorted(actual):
            bad("axis-value-count",
                f"axis {axis}: negation-users list is stale — "
                f"missing {sorted(set(actual) - set(listed))}, "
                f"extra {sorted(set(listed) - set(actual))}")

    # ---- 52 §3 — axis-type-reverse-test ---------------------------------
    profile_text = PROFILE.read_text(encoding="utf-8")
    declared, live, vacuous, worded_hits = {}, 0, 0, []
    for raw in FN_AXIS_BLOCK.findall(profile_text):
        f = dict(l.split(":", 1) for l in raw.strip().split("\n")
                 if ":" in l and not l.lstrip().startswith("#"))
        axis = f.get("axis", "?").strip().split()[0]
        declared[axis] = f
        # 56 §4 — `declared-at-tc-count` records WHEN this declaration was
        # written, so it may only move when the declaration changes. The
        # content hash is the gate: recompute it, and a mismatch means the
        # block was edited without the timestamp following.
        payload = "\n".join(f"{k}={f.get(k, '').strip()}" for k in HASHED_FIELDS)
        want = hashlib.sha256(payload.encode()).hexdigest()[:12]
        got = f.get("content-sha", "").strip()
        if got != want:
            bad("axis-type-reverse-test",
                f"axis {axis}: content-sha is {got or 'missing'} but the "
                f"block hashes to {want}. The declaration changed without "
                f"`declared-at-tc-count` (currently "
                f"{f.get('declared-at-tc-count', '?').strip()}) following it "
                f"— update both together (56 §4)")

        def terms(key):
            v = f.get(key, "").strip()
            return [] if v in ("", "none") else [t.strip() for t in v.split("|")
                                                 if t.strip()]

        iface, fn = terms("removed-interface-keywords"), terms("function-keywords")
        pcs = terms("axis-pc-keywords")
        if not iface:
            vacuous += 1
            continue
        live += 1
        for _, tc in all_tcs:
            observable = f"{tc['test_procedure']}\n{tc['expected_result']}"
            if not any(t in observable for t in iface):
                continue
            # A TC that already states this axis's value is confined to one of
            # them; its observable cannot vanish unexpectedly.
            if any(t in tc["pre_conditions"] for t in pcs):
                continue
            # PURPOSE version (54 §1) — the function filter is NOT applied
            # here. A TC whose function belongs to some other section is
            # exactly the case this test exists for.
            subject = f"{tc['test_item']}\n{observable}"
            also_worded = bool(fn) and any(t in subject for t in fn)
            worded_hits.append((axis, tc["tc_id"])) if also_worded else None
            bad("axis-type-reverse-test",
                f"axis {axis}: {tc['tc_id']} observes "
                f"{[t for t in iface if t in observable]!r} — the interface "
                f"this axis's value removes — and states no value for that "
                f"axis, so it is false on the value that removes it. Either "
                f"add the axis pre_condition or re-decide the axis's type "
                f"(judged at {f.get('judged-at-tc-count', '?').strip()} TCs; "
                f"54 §1 purpose version)"
                + ("  [also matches the 52 §3 wording version]"
                   if also_worded else ""))
    for num, name, kind in AXIS_TABLE_ROW.findall(profile_text):
        if kind == "功能型" and num not in declared:
            bad("axis-type-reverse-test",
                f"axis {num} ({re.sub(chr(96) + '|[*]', '', name)[:40]}) is "
                "marked 功能型 but has no ```function-axis-reverse-test``` "
                "block, so its classification is never re-checked (52 §3). "
                "Declare it, using `removed-interface-keywords: none` if the "
                "axis removes no interface — an explicit `none` is a claim; "
                "a missing block is silence")
    fn_axis_report = (len(declared), live, vacuous, len(all_tcs),
                      worded_hits)

    # 43 §4 — the part that makes a NEW unprotected negation audible. Without
    # it, adding a negated pre_condition for an axis that has no block is
    # exactly as silent as axis 13's situation was before 34 §4.
    for _, tc in all_tcs:
        for line in tc["pre_conditions"].split("\n"):
            if not line.strip() or not NEGATED_PC.search(line):
                continue
            if any(n in line for n in negations):
                continue
            if any(k in line for k in NON_AXIS_NEGATIONS):
                continue
            bad("axis-value-count",
                f"{tc['tc_id']}: negated pre_condition matches no axis block "
                f"and is not named in NON_AXIS_NEGATIONS — its coverage "
                f"changes silently when that axis gains a value (43 §4): "
                f"{line.strip()[:96]!r}")

    # ---- handoff 26 §4.1 — every TC key lands in a column or is named ----
    from write_back import COLS
    columned = set(COLS.values())
    scanned = [(tc["tc_id"], tc) for _, tc in all_tcs]
    scanned += [(f"{d['parent']} (doc)", d) for d in docs]      # 29 §5.1
    for label, obj in scanned:
        stray = sorted(set(obj) - columned - NOT_IN_WORKBOOK)
        if stray:
            bad("json-key-coverage",
                f"{label}: key(s) {stray} neither map to a workbook "
                "column nor appear in NOT_IN_WORKBOOK. Either write_back.py's "
                "COLS lost an entry, or the key is deliberately not delivered "
                "and must be named (adding to the list is a ruling)")

    # ---- handoff 26 §4.2 — an anomaly id cited must actually be registered --
    # A-CF16 was used across two upstream packages while ANOMALIES.md never
    # carried it. Citing an id and the id existing are two different things,
    # and the failure is silent: nothing rejects a number that means nothing.
    registered = ANOMALY_ID.findall((FEATURE / "ANOMALIES.md").read_text("utf-8"))
    cited = {}
    for doc in sorted((FEATURE / "docs").rglob("*.md")):
        for aid in ANOMALY_ID.findall(doc.read_text(encoding="utf-8")):
            cited.setdefault(aid, doc.relative_to(FEATURE).as_posix())
    orphans = {a: p for a, p in sorted(cited.items()) if a not in registered}
    if orphans:
        bad("anomaly-id-registered",
            "anomaly id(s) cited in docs/ but absent from ANOMALIES.md: "
            + ", ".join(f"{a} (first seen {p})" for a, p in orphans.items()))

    # ---- handoff 26 §4.3 — the residue scan must reach max_row -------------
    # Verified by reading write_back.py's source, not by running it: a fixed
    # window (24-35) silently stops short of the sheet's real extent (59), so
    # residue past the window would never be looked at.
    wb_src = (FEATURE / "scripts" / "write_back.py").read_text("utf-8")
    if "ws.max_row" not in wb_src or re.search(r"range\(last, last \+ \d+\)", wb_src):
        bad("residue-scan-window",
            "write_back.py's post-write residue scan does not run to "
            "ws.max_row (a fixed-width window leaves the tail unchecked)")
    return (out, blocked, emea_no, fn_axis_report, mirror_measured,
            rc42_dispositions)


def main() -> int:
    auth = load_authorities()
    docs = [json.loads(p.read_text(encoding="utf-8"))
            for p in sorted(GEN.glob("*.json"))]
    n_tc = sum(len(d["tcs"]) for d in docs)
    (findings, blocked_ids, emea_no, fn_axis, mirror_rows,
     rc42_disp) = lint(docs, auth)

    gates = ["tc-id-format", "tc-id-unique", "tc-id-sequence", "req-id-unique",
             "spec-ref-stem", "spec-ref-outline", "spec-ref-sr25", "test-group",
             "design-method", "priority", "functional-safety", "estimated-time",
             "remarks", "trailing-period", "ui-bracket", "title-length",
             "title-modal", "item-modal", "er-modal", "source-class",
             "proc-er-1to1", "token-placement", "token-source",
             "fabricated-qty", "sibling-axis",
             # added 2026-08-15 after handoff 20 §1.1's coverage audit
             "required-keys", "proc-min-steps", "reasoning-sentences",
             "duplicate-of-format",
             # added 2026-08-15 with R-C24's BLOCKED-SPEC marker
             "blocked-row-empty", "blocked-remarks",
             # added 2026-08-15 with R-C26
             "marker-whitelist",
             # added 2026-08-15 per handoff 26 §4
             "json-key-coverage", "anomaly-id-registered",
             "residue-scan-window",
             # added 2026-08-15 per handoff 31 §1 / §2
             "forbidden-verb", "er-subject-net",
             # added 2026-08-15 per handoff 35 §4
             "axis-value-count",
             # added 2026-08-15 per handoff 36 §4 / §6
             "pending-sibling", "interface-axis-answered",
             # added 2026-08-15 per handoff 38 §1
             "emea-per-tc-answered",
             # added 2026-08-15 per handoff 42 §1
             "provisional-sibling",
             # added 2026-08-15 per handoff 52 §3
             "axis-type-reverse-test",
             # added 2026-08-16 per handoff 60 §1
             "no-tcid-in-prose",
             # added 2026-08-16 per handoff 62 §5
             "prose-reqid-exists",
             # added 2026-08-16 per handoff 62 §1.1 (b)
             "mirror-map-verified", "withheld-not-generated",
             # added 2026-08-16 per handoff 63 §1
             "equivalence-in-sibling-table",
             # added 2026-08-16 per handoff 64 §1 (R-C42 三)
             "axis-candidate-registered",
             # added 2026-08-16 per handoff 65 §3 / §4
             "rc42-condition-marker", "moved-leaf-identity",
             # added 2026-08-16 per handoff 67 §1
             "ambiguity-register",
             # added 2026-08-16 per handoff 69 §1.2
             "source-class-truthful",
             # added 2026-08-16 (執行層自加，見上繳 47 §2.2)
             "pc-line-numbering"]
    failed = {g for _, g, _ in findings}

    print(f"files: {len(docs)}   TCs: {n_tc}   "
          f"vocabulary: {len(auth['vocab'])} strings   "
          f"valid outlines: {len(auth['outlines'])}\n")
    print("gates:")
    for g in gates:
        # er-subject-net prints its own self-qualifying line below; a plain
        # "PASS — er-subject-net" alongside it would read as a second, equal
        # claim, which is the exact impression 31 §2 forbids.
        if g == "er-subject-net" and g not in failed:
            continue
        print(f"- {'**FAIL**' if g in failed else 'PASS'} — {g}")
    # R-C24 — the exemption is visible on every run, whether or not it fired.
    print(f"- PASS — rows exempted as BLOCKED, all markers "
          f"(proc-min-steps, proc-er-1to1): {sorted(blocked_ids) or 'none'}")
    for mk in BLOCKED_MARKERS:
        print(f"- PASS — marker whitelist (profile §5.1/§5.2) {mk}: "
              f"{sorted(MARKER_WHITELIST[mk])}")
    print("- PASS — mirror maps re-measured (62 §1.1 (b)): "
          f"{len(mirror_rows)} `mirrored` row(s), longest shared verbatim run "
          f"per row (threshold {MIRROR_MIN_RUN} chars produces candidates; "
          f"the segment is the substance, §5a)")
    for name, a, b, size, seg in mirror_rows:
        print(f"    · {a:8} ↔ {b:8} {size:4} chars  {seg[:52]!r}")
    print(f"- PASS — R-C42 unblocks whose quoted fragment does not match a "
          f"listed qualifier, and their named dispositions (66 §2.2): "
          f"{len(rc42_disp) or 'none'}")
    for key, src, kind, why in rc42_disp:
        print(f"    · {key} ({src}): {kind} — 出處 {why}")
    state_rows = two_part_state_words(docs)
    flagged = [r for r in state_rows if r[3]]
    print(f"- PASS — 兩段式狀態句之實詞（71 §1，量測不 FAIL）："
          f"{len(state_rows)} 行，其中 {len(flagged)} 行有實詞未見於所引節")
    for tc_id, src, tail, missing in state_rows:
        mark = f"未命中 {missing}" if missing else "全部命中"
        print(f"    · {tc_id} ({src}) {tail[:44]!r} — {mark}")

    longest = test_setup_longest(docs)
    print(f"- PASS — [test-setup] 之對應強度（71 §3，量測不 FAIL）："
          f"最長 {len(longest)} 種寫法（依 節×措辭 去重，×n 為其列數）—— "
          f"長者可能其實有出處而被標成自撰")
    for size, src, frag, count, tc_id, seg in longest:
        print(f"    · ({src:8}) run={size:3}/{len(frag):3} ×{count:3} "
              f"{seg[:22]!r:26} {frag[:46]!r}  e.g. {tc_id}")

    shortest = spec_derived_shortest(docs)
    print(f"- PASS — [spec-derived] 之對應強度（70 §3，量測不 FAIL）："
          f"最短 {len(shortest)} 種寫法（依 節×措辭 去重，×n 為其列數），"
          f"長度為其與所引節之最長共同連續字串")
    for size, src, frag, count, tc_id, seg in shortest:
        print(f"    · ({src:8}) run={size:3}/{len(frag):3} ×{count:3} "
              f"{seg[:16]!r:20} {frag[:56]!r}  e.g. {tc_id}")
    equivalent_groups = identical_tc_groups(docs)
    print(f"- PASS — identical-TC scan (61 §2, measurement): "
          f"{len(equivalent_groups)} group(s) of TCs whose test_item, "
          f"test_procedure and expected_result are character-identical — "
          f"037 decomposition artefacts, kept as separate rows (§8.2.2), "
          f"recorded in pending_sibling's equivalent_tc_pairs")
    for g, pc in equivalent_groups:
        print(f"    · {' ≡ '.join(f'{o}:{r}' for o, _, r in g)}"
              f"   pre_conditions: {pc}")
    if emea_no:
        print(f"- PASS — EMEA exclusions whose per-TC answer is NOT `yes` "
              f"(R-C36-1; over-strict, removal awaits a ruling): "
              f"{[(i, v, o) for i, v, o in emea_no]}")
    n_decl, n_live, n_vac, n_tc_seen, worded = fn_axis
    print(f"- PASS — axis-type-reverse-test re-ran on {n_tc_seen} TCs "
          f"(52 §3, criterion = 54 §1 purpose version): {n_decl} 功能型 axes "
          f"declared, {n_live} with a removed interface (live test), "
          f"{n_vac} declaring `none` (vacuous by claim, not by omission)")
    print(f"- PASS — of the purpose-version hits, those ALSO matching the "
          f"52 §3 wording version (function governed by the axis): "
          f"{worded or 'none'} — the two versions are reported separately so "
          f"a future divergence is visible rather than absorbed")
    print(f"- PASS — pending siblings awaiting their counterpart section "
          f"(36 §4 / 37 §6): {sorted(PENDING_SIBLING)}")
    print("- PASS — the pending-sibling table is produced by lexical overlap "
          "and is NOT a completeness proof (R-C37); run "
          "scripts/sibling_candidates.py when a Test Set completes")
    if "er-subject-net" not in failed:
        print("- PASS — er-subject-net (a safety net, not the criterion; "
              "the criterion is §6 and is human-reviewed)")
    print(f"- PASS — keys deliberately not in the workbook, TC + doc layer "
          f"(26 §4.1 / 29 §5.1): {sorted(NOT_IN_WORKBOOK)}")
    if findings:
        print(f"\n{len(findings)} finding(s):")
        for sev, g, msg in findings:
            print(f"  [{sev}] {g}: {msg}")
    print(f"\n{len(gates) - len(failed)} / {len(gates)} gates PASS; "
          f"{len(findings)} finding(s) across {n_tc} TCs")
    return 1 if findings else 0


if __name__ == "__main__":
    sys.exit(main())
