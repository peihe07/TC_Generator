#!/usr/bin/env python3
"""Write ALL current TCs into the prepared workbook — 25 §3 / 45 §2-§3.

45 §1 ruled the cadence: write back after EVERY batch, and write the WHOLE
corpus each time rather than appending. The reason is measured, not stylistic
— the pilot's 14 rows in `…_pilot.xlsx` are already stale (EMEA PCs removed
from 11 TCs, -019 gained a confining PC, -036 was split so 30 tc_ids shifted,
reasoning revised repeatedly). An append would leave the workbook a mixture of
generations, and nothing would say which row belonged to which.

Three stages, verified separately and never merged (25 §3):
  §3.1  pre-gates   — any failure stops before the splice
  §3.2  splice      — through backend/xlsx_surgical.py only (R18-3)
  §3.3  assertions  — read back FROM THE EMITTED FILE, not from memory

Written for Comfort. The four existing features' write_back.py are quarantined
and must not be used as a starting point (R20-5) — they predate the surgical
emit path and carry per-feature column policy that does not transfer.

What this does NOT do (25 §3.6): it does not copy anything to the customer
delivery path, does not touch the prepared file, does not modify ENTRY 001,
and does not run git. Excel's own four-point confirmation is Pei's (profile
§0.1) — a program-level check cannot stand in for it, so the script stops
after emitting and says so.

Usage:
    python3 features/comfort/scripts/write_back.py            # dry-run
    python3 features/comfort/scripts/write_back.py --write
"""

import argparse
import hashlib
import json
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))
from backend.xlsx_surgical import StructureError, surgical_save  # noqa: E402

FEATURE = ROOT / "features" / "comfort"
GEN = FEATURE / "generated"
# 78 §3 — the source is now the EXTENDED template (ENTRY 022). The 20260815
# prepared file remains untouched as ENTRY 001's object; it is no longer the
#母本 because its DV and B-column formulas stop at row 59.
SRC = FEATURE / "output" / ("FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT "
                            "STLA Test Case Specification & Result_SWQT_"
                            "Comfort_20260816_prepared_ext.xlsx")
# 45 §2 — a NEW file. The pilot file is neither overwritten nor deleted:
# DELIVERY ENTRY 002 is its identity record, and deleting it would leave the
# ledger pointing at nothing (the converse of R-C14 — a recorded identity may
# not lose its object).
# 45 §2 —— 每次寫回產一份**新檔**，不覆寫前一份。本次為 96 §1 之列序改版
# 併 97／98 之 31 條補產，故檔名以 `rowsort` 標其性質。
OUT = FEATURE / "output" / ("FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT "
                            "STLA Test Case Specification & Result_SWQT_"
                            "Comfort_20260817_rowsort.xlsx")
# 45 §3.4 said "ENTRY 003", which was TAKEN (the folder-attachment entry from
# 27 §3), so the second write-back became ENTRY 004. This is the third.
#
# 46 §1.2 reserves an ENTRY for the template extension. It is not written yet
# — the extension is Pei's Tier 3 work and has not happened — so the number is
# free and this write takes it. If the extension lands first, this constant
# moves; the one-shot gate below is what makes a collision loud rather than
# silent.
LEDGER_ENTRY = "ENTRY 035"
SHEET = "Test Case Specification 測試用例規範"
SRC_SHA = "6d53056e559bd0c13d26d38f16754536ede0230a5ce69c8596cce8e8b28b9d4c"
FIRST_ROW = 10

# profile §0 — revision C letters, verified against the header at recon.
COLS = {
    "D": "req_id", "F": "tc_id", "G": "test_group", "H": "test_set",
    "I": "test_item", "J": "pre_conditions", "K": "input_test_data",
    "L": "test_procedure", "M": "expected_result",
    "N": "specification_reference", "P": "priority", "R": "design_method",
    "S": "functional_safety", "AH": "remarks",
}
# Never written. B carries the template's own numbering formula; clearing or
# overwriting it removes the mechanism (profile §0.1).
NEVER_WRITE = ["B", "C", "E", "O", "Q", "T", "U", "V", "W", "X", "Y", "Z",
               "AB", "AC", "AD", "AE", "AF", "AG"]


def sha256(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------- 96 §1
# 列序與缺列之規則（下放包 96，Pei 裁定 2026-08-17）：
#
#     一、列序依 `Requirement or Design ID`（即 037 之 leaf id）遞增，
#         不依批次順序、不依 Test Set 順序。
#     二、未產出 TC 之 leaf 仍佔一列 —— 其 req_id 照填，其餘欄留空。
#
# 其目的：**工作簿由上而下即為 037 之 leaf 全集，其缺漏一眼可見。**
# 96 §3 記其代價之解除 —— 交付說明為此建立之一整套外部機制（79 §2.2 起）
# 本來就被既有規則解掉了：缺口在工作簿自身即可見，不需要一份文件去說明它。
#
# 037 之列序實測即 leaf id 之遞增（本檔 `verify_row_order_gates.py` 之第一向），
# 故「依 037」與「依 id 遞增」在本語料上是同一件事，不必二選一。
LEAF_KEY = re.compile(r"SWE1-HVAC-(\d+)(?:-(\d+))?$")


def leaf_sort_key(req_id: str) -> tuple:
    m = LEAF_KEY.match(req_id)
    return (int(m.group(1)), int(m.group(2) or 0))


def leaf_universe() -> list:
    """037 之 403 個 leaf，依 id 遞增。來源同 lint 之 LEAF_UNIVERSE。"""
    recon = json.loads((FEATURE / "data" / "recon.json").read_text(
        encoding="utf-8"))
    return sorted(recon["leaves"], key=leaf_sort_key)


def load_tcs() -> list:
    docs = [json.loads(p.read_text(encoding="utf-8"))
            for p in sorted(GEN.glob("*.json"))]
    tcs = [t for d in docs for t in d["tcs"]]
    return sorted(tcs, key=lambda t: int(t["tc_id"].rsplit("-", 1)[1]))


BLANK = "__BLANK__"


def row_plan(tcs: list) -> list:
    """96 §1 之列序：037 之 leaf 全集，逐 leaf 展開其 TC，無 TC 者一列留空。

    回傳之每一項為 dict：TC 列即該 TC 本身；留空列為
    `{"req_id": …, BLANK: True}`，其餘欄由 `cell_of()` 一律給空。
    同一 leaf 之多條 TC 依其 tc_id 遞增排列（其相對次序即生成之次序）。
    """
    by_leaf = {}
    for t in tcs:
        by_leaf.setdefault(t["req_id"], []).append(t)
    for v in by_leaf.values():
        v.sort(key=lambda t: int(t["tc_id"].rsplit("-", 1)[1]))
    plan = []
    for leaf in leaf_universe():
        if leaf in by_leaf:
            plan.extend(by_leaf[leaf])
        else:
            plan.append({"req_id": leaf, BLANK: True})
    return plan


# ---------------------------------------------------------------- 96 §4.1
# **F 欄之 tc_id 依列位重編，JSON 之 tc_id 不動**（Pei 裁定 2026-08-17）。
#
# 96 §4.1 要求「tc_id 須依新列序重新指派」，而 65 §1 立「既有編號不移動」——
# **兩條直接衝突**，其衝突至今未被解決，因為 96 §7 原定由 Pei 自行改工作簿。
#
# 裁定取**只改匯出側**，其依據為 §10.3 之所在：它列於
# 「## 10. Tool-Specific Output Contract (workbook export, not ASPICE rules)」
# 之下，且明寫「the generator handles assignment」—— **tc_id 是匯出側之物**。
# 故 F 欄寫列位編號（001…466，含留空列），語料之 tc_id 維持其生成時之值。
#
# **其代價須講明，不得靜默**：
#   一、**工作簿 F 欄與 JSON 之 tc_id 自此不同**。ENTRY 025 那種「逐格與
#       generated/*.json 比對」之複驗，其 F 欄一項須改以**列位**比對，
#       否則會報 465 格不符 —— 那不是內容變了，是兩側各自為政。
#   二、台帳與往返包引用 tc_id 時須講明是哪一側。**語料側為準**
#       （60 §1 之 `no-tcid-in-prose` gate 管的是語料）。
#   三、**留空列之 F 欄為空，且不佔一個編號** —— §10.3 要 tc_id 單調遞增，
#       既有之 `tc-id-sequence` gate 另要其連續無缺號；把一個號碼發給一個
#       沒有 TC 的列，會在 F 欄上開一個洞。故 F 為「**TC 於列序上之次序**」
#       （001–465），不是列位。兩者只差在留空列那一格，而那一格正是分野。
FMT_ROW_TCID = "NR1L-ComfortHMI-{n:03d}"


def cell_of(row: dict, field: str, row_no: int = 0) -> str:
    """留空列除 req_id 外一律空（96 §1 之二、其 `blank-row-shape`）。

    `tc_id` 為例外：其值取**列位**（96 §4.1），非語料之 tc_id。
    """
    if field == "tc_id":
        return "" if row.get(BLANK) else FMT_ROW_TCID.format(n=row_no)
    if row.get(BLANK):
        return row["req_id"] if field == "req_id" else ""
    return render(field, row[field])


# 83 §1 — the ledger check, without `--ignore-missing`.
#
# That flag conflated "this file is not there" with "this file need not be
# checked". Only one of those is something we know. A produced file deleted by
# accident and ENTRY 023's overwritten bytes looked identical under it: both
# simply never appeared in the count.
#
# So absence is now DECLARED, by path, in DELIVERY.sha256's `# absent:` block,
# and the check reports three numbers instead of one. Both directions fail:
#   - missing and NOT declared absent  -> FAIL (something vanished)
#   - declared absent but PRESENT      -> FAIL (the declaration is stale, or
#                                         the object came back and nobody said)
ABSENT_DECL = re.compile(r"^# absent: (.+?)\s+——", re.M)


def check_ledger(ledger: Path) -> tuple:
    """Return (ok, verified, absent_declared, problems)."""
    text = ledger.read_text(encoding="utf-8")
    declared_absent = set(ABSENT_DECL.findall(text))
    # A row may carry `#   archived : <new path>` on the line after it: the
    # bytes still exist, they moved. That is a different thing from absence,
    # and it is checked rather than excused — the digest is verified at the
    # new path.
    rows = []
    for m in re.finditer(r"^([0-9a-f]{64})  (.+)$", text, re.M):
        tail = text[m.end():m.end() + 400]
        moved = re.match(r"\n#   archived : (.+?)(?:  （|\n)", tail)
        rows.append((m.group(1), m.group(2),
                     moved.group(1).strip() if moved else None))
    verified, absent, problems = 0, 0, []
    for digest, rel, moved in rows:
        path = FEATURE / rel
        if not path.exists():
            if moved:
                mp = FEATURE / moved
                if not mp.exists():
                    problems.append(f"archived path missing: {moved}")
                elif sha256(mp) != digest:
                    problems.append(f"MISMATCH at archived path: {moved}")
                else:
                    verified += 1
                continue
            if rel in declared_absent:
                absent += 1
            else:
                problems.append(f"MISSING (not declared): {rel}")
            continue
        if rel in declared_absent:
            problems.append(f"declared absent but PRESENT: {rel}")
            continue
        if sha256(path) != digest:
            problems.append(f"MISMATCH: {rel}")
            continue
        verified += 1
    # a declaration with no ledger row is a dangling permit
    listed = {rel for _, rel, _ in rows}
    for rel in sorted(declared_absent - listed):
        problems.append(f"declared absent but no ledger row: {rel}")
    return (not problems), verified, absent, problems


# ------------------------------------------------------------ §3.1 pre-gates

def pre_gates(tcs: list) -> bool:
    print("## §3.1 前置 gate\n")
    ok = True

    def g(name, passed, note=""):
        nonlocal ok
        ok &= passed
        print(f"- {'PASS' if passed else '**FAIL**'} — {name}"
              + (f" — {note}" if note else ""))

    def shasum(args):
        return subprocess.run(["shasum", "-a", "256", "-c"] + args,
                              cwd=FEATURE, capture_output=True, text=True)

    # 87 §4 — was `n_ok == 8`, the number of baseline files on the day it was
    # written. `inputs/` went from 5 files to 8 and the count moved with it;
    # what the gate is for is "every listed file verifies", which is
    # `FAILED == 0` plus non-vacuity (R-C43 — the same defect as
    # `len(withheld) >= 20`, found the same way: the corpus grew).
    r = shasum(["BASELINE.sha256"])
    n_ok = r.stdout.count(": OK")
    n_bad = r.stdout.count("FAILED")
    g("BASELINE.sha256 逐檔全數 OK", n_bad == 0 and n_ok > 0,
      f"OK={n_ok}, FAILED={n_bad}")

    ok, n_ver, n_abs, problems = check_ledger(FEATURE / "DELIVERY.sha256")
    g("DELIVERY.sha256 逐列（不用 --ignore-missing）",
      ok and n_ver > 0,
      f"驗過 {n_ver}, 已知不存在 {n_abs}, 有問題 {len(problems)}"
      + (f" — {problems[:3]}" if problems else ""))
    # The one-shot guard, now keyed to THIS write's entry number. Once the
    # ledger carries it, --write can no longer run: that is the append-only
    # ledger working, not a defect.
    # Match the ENTRY HEADER, not the string anywhere. ENTRY 004's status text
    # names "ENTRY 005" as the template-extension entry to come, and a bare
    # substring search read that mention as the entry itself and blocked the
    # write. A ledger that discusses its own future entries is normal; a gate
    # that cannot tell a reference from a record is not.
    already = any(l.lstrip("# ").startswith(LEDGER_ENTRY + " ")
                  for l in (FEATURE / "DELIVERY.sha256")
                  .read_text("utf-8").splitlines())
    g(f"台帳尚無 {LEDGER_ENTRY}（一次性 gate）", not already,
      "present" if already else "absent")

    digest = sha256(SRC) if SRC.exists() else "(missing)"
    g("來源為 ENTRY 022 之擴充後母本（78 §3）", digest == SRC_SHA,
      f"measured {digest[:16]}…, expected {SRC_SHA[:16]}…")

    r = subprocess.run([sys.executable, str(FEATURE / "scripts" / "lint_tcs.py")],
                       capture_output=True, text=True)
    lint_ok = r.returncode == 0
    tail = [l for l in r.stdout.strip().split("\n") if "gates PASS" in l]
    g("lint 全數 PASS", lint_ok, tail[-1] if tail else "no summary line")

    # 45 §2 — the count is MEASURED, never pre-filled. What is asserted is
    # that it is non-zero, gap-free and matches what lint just counted.
    nums = [int(t["tc_id"].rsplit("-", 1)[1]) for t in tcs]
    contiguous = nums == list(range(1, len(nums) + 1))
    g("TC 數實測且 tc_id 連續無缺號", len(tcs) > 0 and contiguous,
      f"measured {len(tcs)} TCs, tc_id {nums[0]:03d}–{nums[-1]:03d}")
    print()
    return ok


# ------------------------------------------------------------ §3.2 splice

# 94 §2 — source class labels are an INTERNAL field of generated/*.json. They
# stay there (R-C28 Q1's evidence) and never reach the workbook: nobody ever
# asked whether `[spec-derived]` belongs in a cell the customer reads. The
# section reference in brackets DOES stay — that is the clause's provenance,
# not our vocabulary, and the reader needs it to locate the condition.
SOURCE_CLASS_IN_CELL = re.compile(
    r"\[(?:spec-verbatim|spec-derived|test-setup|ext-verbatim)\]\s*")


# Pei 2026-08-17 — 其自行編修之交付件（ENTRY 032）另有兩項，本層據以併入
# **寫入路徑**（JSON 不動，理由同 source class：內部依據與外部呈現分離）：
#   J 欄：連**節次括號**一併去除（94 §2.1 曾裁「保留」，Pei 之編修不保留）
#   N 欄：Comfort stem 去 `SYS1_` 前綴
# 後者之影響須明講：`specification_reference` 是 traceability 之字串比對
# 對象，工作簿內之 stem 自此與 R-C1 所定之基線檔名**差一個前綴**。
# JSON 內仍為全名，`spec-ref-stem` gate 驗的是 JSON，故未失效。
SECTION_BRACKET = re.compile(r"\s*\((\d+(?:\.\d+)*)\)\s*$", re.M)
SYS1_STEM = ("SYS1_HMI_Comfort_HMI_Logic_and_Flow_R1_SR24_Post_3A_"
             "CR24879_(September_25_2023)")


def render(field: str, value: str) -> str:
    """What goes in the cell, as opposed to what the JSON carries."""
    if field == "pre_conditions":
        return SECTION_BRACKET.sub("", SOURCE_CLASS_IN_CELL.sub("", value))
    if field == "specification_reference":
        return _compact_refs(value)
    return value


def _compact_refs(value: str) -> str:
    """Pei 之 N 欄寫法：stem 只寫一次，其節次以「、」相連；外部出處另段。

    JSON 內每一節各寫一次完整 stem（R-C29：本節在前，引用節在後），
    工作簿內重複 12 次同一個 600 字元之 stem 對讀者毫無用處。
    `SYS1_` 前綴一併去除（Comfort stem 與 CFTS043 皆然）。
    """
    stem_short = SYS1_STEM[len("SYS1_"):]
    outlines, others = [], []
    for seg in [x.strip() for x in value.split(";") if x.strip()]:
        seg = seg[len("SYS1_"):] if seg.startswith("SYS1_") else seg
        if seg.startswith(stem_short + "_"):
            outlines.append(seg[len(stem_short) + 1:])
        else:
            others.append(seg)
    parts = []
    if outlines:
        parts.append(f"{stem_short}_" + "、".join(outlines))
    parts.extend(others)
    return "; ".join(parts)


def tcid_sequence(plan: list) -> list:
    """F 欄之號碼：TC 列依列序遞增 1…N，留空列給 0（其 F 欄為空）。"""
    seq, n = [], 0
    for row in plan:
        if row.get(BLANK):
            seq.append(0)
        else:
            n += 1
            seq.append(n)
    return seq


def splice(plan: list) -> dict:
    wb = openpyxl.load_workbook(SRC)
    ws = wb[SHEET]
    seq = tcid_sequence(plan)
    for i, row in enumerate(plan):
        r = FIRST_ROW + i
        for col, field in COLS.items():
            v = cell_of(row, field, seq[i])
            ws[f"{col}{r}"] = v if v != "" else None
    return surgical_save(wb, SRC, OUT)


# ------------------------------------------------- §3.3 post-write assertions

def assertions(plan: list, report: dict) -> bool:
    """Every check reads the EMITTED file, never the in-memory workbook."""
    print("## §3.3 寫回後 assertion（自產出檔讀回）\n")
    ok = True

    def g(name, expected, actual, note=""):
        nonlocal ok
        p = expected == actual
        ok &= p
        print(f"- {'PASS' if p else '**FAIL**'} — {name}: "
              f"expected `{expected}`, measured `{actual}`"
              + (f" — {note}" if note else ""))

    import zipfile
    with zipfile.ZipFile(SRC) as a, zipfile.ZipFile(OUT) as b:
        src_names, out_names = set(a.namelist()), set(b.namelist())
    g("zip member 數與來源相同", len(src_names), len(out_names),
      f"symmetric difference: {sorted(src_names ^ out_names) or 'none'}")
    from backend.xlsx_surgical import sheet_members
    member = sheet_members(SRC)[SHEET]        # resolved, not hard-coded
    g("差異僅限目標 sheet 之 xml", [member], sorted(report.get("differing", [])))
    g("DV counts 與來源相同", "equal",
      "equal" if report.get("dv_counts") else "not reported",
      str(report.get("dv_counts")))

    wb = openpyxl.load_workbook(OUT)
    ws = wb[SHEET]
    tcs = [x for x in plan if not x.get(BLANK)]
    blanks_plan = [x for x in plan if x.get(BLANK)]
    seq = tcid_sequence(plan)
    mismatches = []
    for i, row in enumerate(plan):
        r = FIRST_ROW + i
        who = row.get("tc_id") or f"BLANK:{row['req_id']}"
        for col, field in COLS.items():
            got = ws[f"{col}{r}"].value
            got = "" if got is None else str(got)
            if got != cell_of(row, field, seq[i]):
                mismatches.append(f"{who}.{col}({field})")
    g("逐列全部寫入欄之值與 JSON（經 render 後）一致", [], mismatches,
      f"{len(plan)} rows x {len(COLS)} columns compared: {''.join(COLS)}")

    # ---- 96 §6 之三道 gate，讀回產出檔 --------------------------------------
    col_d = [ws[f"D{FIRST_ROW + i}"].value or "" for i in range(len(plan))]
    # **判準為「嚴格倒退」而非「未遞增」** —— 一個 leaf 拆出多條 TC 時，
    # 那幾列之 D 欄本來就相同（40 個 leaf 如此，最多一葉五列）。
    # 第一版寫 `<=`，被 `verify_row_order_gates.py` 之第一向當場抓出：
    # **乾淨之列序自己就會觸發**。一道對正確資料轉紅的檢查，比沒有還糟。
    first_bad = next((f"row{FIRST_ROW + i}: {col_d[i]} 在 {col_d[i - 1]} 之後"
                      for i in range(1, len(col_d))
                      if leaf_sort_key(col_d[i]) < leaf_sort_key(col_d[i - 1])),
                     None)
    g("row-order-by-reqid：D 欄自上而下為 037 之 leaf 序（96 §1 一）",
      None, first_bad, f"{len(plan)} 列讀回；不符即指名首個逆序之列")

    universe = leaf_universe()
    missing = sorted(set(universe) - set(col_d), key=leaf_sort_key)
    g("all-leaves-present：037 之 403 個 leaf 每一個皆於 D 欄出現至少一次",
      [], missing,
      f"037 leaf {len(universe)}；D 欄相異值 {len(set(col_d))}")

    bad_blank = []
    for i, row in enumerate(plan):
        if not row.get(BLANK):
            continue
        r = FIRST_ROW + i
        for col, field in COLS.items():
            if field == "req_id":
                continue
            if ws[f"{col}{r}"].value not in (None, ""):
                bad_blank.append(f"row{r}.{col}")
    g("blank-row-shape：留空列除 D 欄外各欄皆空（96 §1 二）", [], bad_blank,
      f"{len(blanks_plan)} 個留空列：" +
      ", ".join(x["req_id"] for x in blanks_plan))

    # 94 §2.3 — the gate this case says was missing: nothing ever asked
    # whether a cell the customer reads carries our own vocabulary.
    labelled = []
    for i, tc in enumerate(plan):
        cell = ws[f"J{FIRST_ROW + i}"].value or ""
        for m in SOURCE_CLASS_IN_CELL.finditer(str(cell)):
            labelled.append(f"{tc['tc_id']}:{m.group(0).strip()}")
    g("no-source-class-in-workbook：J 欄無 source class 標籤（94 §2.3）",
      [], labelled, f"{len(tcs)} 列之 J 欄逐格掃描")

    # 70 §1 — the same format invariant, re-checked ON THE WORKBOOK. Five TCs
    # carried an unnumbered second pre_condition line through ENTRY 016 and
    # 017: the defect existed in the JSON and in the workbook at the same
    # time, and nothing on either side said so. A format invariant is cheap
    # enough to assert twice, and the second assertion is the one that reads
    # what the customer will read.
    numbering = []
    for i, tc in enumerate(plan):
        if tc.get(BLANK):
            continue
        r = FIRST_ROW + i
        cell = ws[f"J{r}"].value or ""
        lines = [l for l in str(cell).split("\n") if l.strip()]
        got = [re.match(r"^(\d+)\.\s", l) for l in lines]
        if not all(got) or [int(m.group(1)) for m in got] != list(
                range(1, len(lines) + 1)):
            numbering.append(f"row{r}({tc['tc_id']})")
    g("J 欄 pre_conditions 之行號自 1 起連續（70 §1）", [], numbering,
      f"{len(tcs)} rows read back from the workbook")

    blanks = []
    for i in range(len(plan)):
        r = FIRST_ROW + i
        for col in ("Q",) + tuple("TUVWXYZ"):
            if ws[f"{col}{r}"].value not in (None, ""):
                blanks.append(f"row{r}.{col}")
    g("Q 與 T–Z 留白（profile §3.7／§3.9）", [], blanks)
    bad_s = [f"row{FIRST_ROW + i}" for i, x in enumerate(plan)
             if not x.get(BLANK) and ws[f"S{FIRST_ROW + i}"].value != "NA"]
    g("S 欄一律 NA（profile §3.8；留空列不適用）", [], bad_s)

    # B must still hold its formula, not a value openpyxl substituted. The
    # template carries the same formula well past the target range, so the
    # check runs on rows 10-35 — row 24+ renders empty only because D is
    # empty, and that is the mechanism, not residue.
    bad_b = []
    for r in range(FIRST_ROW, FIRST_ROW + len(plan) + 12):
        v = ws[f"B{r}"].value
        if v != f'=IF(ISBLANK($D{r}),"",ROW()-9)':
            bad_b.append(f"row{r}={v!r}")
    g(f"B 欄 row {FIRST_ROW}–{FIRST_ROW + len(plan) + 11} 之公式逐列原樣存在",
      [], bad_b,
      "96 §4.2 —— 留空列之 D 欄非空（照填 req_id），故 B 欄仍會給它一個編號；"
      "B 欄之編號自此等於**列數**，不等於 TC 數")

    # ---- assertion 11 (45 §3.3) — three marker classes, three rules -------
    # profile §5.1 / §5.2 / §5.2a. Each class's first visible line must carry
    # what that class exists to point at, and [BLOCKED-NON-HMI] must carry the
    # OPPOSITE of [BLOCKED-SPEC]: no owner at all.
    MARKER_RULE = {
        "[BLOCKED-SPEC]": ("Owner:", True),
        "[BLOCKED-NON-HMI]": ("Not an HMI-observable property", True),
        "[COVERED-BY]": ("[COVERED-BY]", True),
    }
    marked, bad_blk = {k: [] for k in MARKER_RULE}, []
    for i, t in enumerate(plan):
        if t.get(BLANK):
            continue
        rm_json = t["remarks"]
        mk = next((m for m in MARKER_RULE if rm_json.startswith(m)), None)
        if mk is None:
            continue
        marked[mk].append(t["tc_id"])
        r = FIRST_ROW + i
        for col in ("L", "M"):
            if ws[f"{col}{r}"].value not in (None,):
                bad_blk.append(f"{t['tc_id']}.{col}={ws[f'{col}{r}'].value!r}")
        rm = ws[f"AH{r}"].value or ""
        needle, must_have = MARKER_RULE[mk]
        if (needle in rm[:60]) is not must_have:
            bad_blk.append(f"{t['tc_id']}.AH[:60] {mk} rule: "
                           f"{needle!r} {'missing' if must_have else 'present'}")
        if mk == "[BLOCKED-NON-HMI]" and "Owner:" in rm:
            bad_blk.append(f"{t['tc_id']}.AH names an Owner under "
                           f"[BLOCKED-NON-HMI] — that is a [BLOCKED-SPEC]")
    # 80 §1 / R-C27-1 — the 60 is this check's own inspection window, not
    # a claim about what a reader can see. Measured visibility is ~11
    # characters; the reachable goal is that the row reads as MARKED.
    g("三類 marker 列之 L／M 為空且 Remarks 前綴（檢查窗口 60 字元）符合各自規則",
      [], bad_blk,
      "; ".join(f"{k} {v or 'none'}" for k, v in marked.items()))

    # ---- assertion 10 (45 §3.3) — A-CF19 measured, not assumed ------------
    # The anomaly is about PRESENTATION, so the check separates the two
    # questions it was always confusing: is the content complete (yes/no,
    # checkable) and is it visible (a number, reportable but not a pass/fail).
    # 對照 render 後之值：Pei 2026-08-17 之編修使 N 欄之 stem 只寫一次、
    # 節次以「、」相連且去 `SYS1_` 前綴。**節次一個不少**是這道 assertion
    # 現在要證的事 —— 故另驗其節次集合與 JSON 相同（下方 n_lost）。
    n_bad = [f"{t['tc_id']}"
             for i, t in enumerate(plan)
             if not t.get(BLANK) and (ws[f"N{FIRST_ROW + i}"].value or "")
             != render("specification_reference", t["specification_reference"])]
    n_lost = []
    for i, t in enumerate(plan):
        if t.get(BLANK):
            continue
        cell = ws[f"N{FIRST_ROW + i}"].value or ""
        want = set(re.findall(r"_(\d+(?:\.\d+)*)(?=[;、]|$)",
                              t["specification_reference"]))
        got = set(re.findall(r"[_、](\d+(?:\.\d+)*)(?=[;、]|$)", cell))
        if want - got:
            n_lost.append(f"{t['tc_id']}:{sorted(want - got)}")
    g("N 欄之節次一個不少（縮寫後仍涵蓋 JSON 所列之全部節次）", [], n_lost,
      f"{len(tcs)} cells compared（留空列不計）")
    g("N 欄 specification_reference 與 JSON（經 render 後）相同（A-CF19 之內容側）",
      [], n_bad, f"{len(tcs)} cells compared")
    lens = [(len(t["specification_reference"]), t["tc_id"]) for t in tcs]
    longest, longest_id = max(lens)
    multi = sum(1 for L, _ in lens if "; " in
                tcs[[i for i, t in enumerate(tcs)
                     if len(t["specification_reference"]) == L][0]]
                ["specification_reference"])
    width = ws.column_dimensions["N"].width
    rh = ws.row_dimensions[FIRST_ROW].height
    wrap = ws[f"N{FIRST_ROW}"].alignment.wrap_text
    per_line = int(width) if width else 0
    visible = 1 if (rh or 0) <= 15 else int((rh or 0) // 14)
    print(f"- MEASURED — A-CF19 之呈現側：N 欄最長 {longest} 字元"
          f"（{longest_id}）；欄寬 {width}；wrapText={wrap}；"
          f"列高 {rh} → 可見約 {visible} 行 ≈ {per_line * visible} 字元，"
          f"即最長者之 {100 * per_line * visible // max(longest, 1)}%。"
          f"**內容完整而僅首行可見** —— 這是 A-CF19 之實測，非 assertion："
          f"呈現屬 Tier 3，程式不得自行改列高（26 §2 之方向 3 已裁）")

    # handoff 26 §4.3 — scan to the sheet's real extent, not a fixed window.
    # The window used to be 12 rows wide while max_row is 59; residue past
    # row 35 would never have been looked at.
    last, end = FIRST_ROW + len(plan), ws.max_row
    residue = []
    for r in range(last, end + 1):
        for col in COLS:          # B excluded — its formula is template, not residue
            v = ws[f"{col}{r}"].value
            if v not in (None, ""):
                residue.append(f"row{r}.{col}={str(v)[:24]!r}")
    g(f"row {last} 起至 max_row 無殘留內容", [], residue,
      f"scanned rows {last}–{end} (ws.max_row={end})")
    # ---- assertion 13 (上繳 33 §9.3) — the template's own provisions must
    # actually REACH the rows we wrote. The B-formula check below caught this
    # only because its window happened to extend past row 59; the data
    # validations were never checked at all, and P/T-Z/AF stop at row 11.
    # A row outside a DV sqref looks written and is not deliverable: profile
    # §0.1's confirmation item 2 ("R 欄下拉可用且為九項") is false there.
    import re as _re
    import zipfile as _zip
    from backend.xlsx_surgical import sheet_members as _sm
    with _zip.ZipFile(OUT) as _z:
        _xml = _z.read(_sm(OUT)[SHEET]).decode("utf-8")

    def _cover(sqrefs):
        rows = set()
        for sq in sqrefs:
            for part in sq.split():
                m = _re.match(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?$", part)
                if m:
                    rows |= set(range(int(m.group(2)),
                                      int(m.group(4) or m.group(2)) + 1))
        return rows

    r_rows = _cover(_re.findall(r"<xm:sqref>([^<]+)</xm:sqref>", _xml))
    p_rows = _cover([m for m in _re.findall(
        r'<dataValidation[^>]*sqref="([^"]+)"', _xml) if m.startswith("P")])
    # (B is checked above through openpyxl — the XML carries SHARED formulas,
    # where only the master cell holds the <f> text, so a regex over the raw
    # XML under-counts. Left out deliberately rather than duplicated wrongly.)
    written = set(range(FIRST_ROW, FIRST_ROW + len(plan)))
    g("每一寫入列皆在 R 欄下拉（x14 DV）之涵蓋範圍內",
      [], sorted(written - r_rows)[:6] + (["…"] if len(written - r_rows) > 6 else []),
      f"{len(written - r_rows)} row(s) outside; DV covers rows "
      f"{min(r_rows)}–{max(r_rows)}")
    g("每一寫入列皆在 P 欄 DV 之涵蓋範圍內",
      [], sorted(written - p_rows)[:6] + (["…"] if len(written - p_rows) > 6 else []),
      f"{len(written - p_rows)} row(s) outside; DV covers rows "
      f"{min(p_rows)}–{max(p_rows)}" if p_rows else "no P DV found")

    # ---- assertion 12 (45 §3.3) — row count == TC count -------------------
    written = sum(1 for r in range(FIRST_ROW, end + 1)
                  if ws[f"D{r}"].value not in (None, ""))
    g("已寫入之列數等於 TC 數 ＋ 留空列數（96 §4.3）", len(plan), written,
      f"rows {FIRST_ROW}–{FIRST_ROW + len(plan) - 1}；"
      f"{len(tcs)} TC ＋ {len(blanks_plan)} 留空列")
    wb.close()
    print()
    return ok


# ------------------------------------------------- §3.4 歷史產出之歸檔（Pei 追加，2026-08-16）

# `output/` 只保留三份：ENTRY 001 之 prepared（台帳之對象）、現行母本、
# 本次產出。其餘產出檔搬入 `output/archive/`。
#
# 台帳為 append-only：既有之 checksum 行**一字不改**，只
#   (a) 在該行之後插入一行 `#   archived : …` 註記，
#   (b) 於文末之「歸檔後之可驗路徑」段落追加該檔於新路徑之 checksum 行。
# (b) 不可省 —— 只做 (a) 的話，舊路徑於 `shasum -c --ignore-missing` 下
# 會被**靜靜跳過**，而「仍可驗」就只是一句宣稱（R-C43）。
KEEP_IN_OUTPUT = {SRC.name, OUT.name,
                  "FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT STLA Test "
                  "Case Specification & Result_SWQT_Comfort_20260815_"
                  "prepared.xlsx"}
ARCHIVE_HEADER = "# 歸檔後之可驗路徑（2026-08-16）"


def archive_previous() -> list:
    """Move superseded outputs into output/archive/ and annotate the ledger."""
    out_dir = FEATURE / "output"
    arc = out_dir / "archive"
    arc.mkdir(exist_ok=True)
    moved = []
    for f in sorted(out_dir.glob("*.xlsx")):
        if f.name in KEEP_IN_OUTPUT:
            continue
        f.rename(arc / f.name)
        moved.append(f.name)
    if not moved:
        return moved

    led = FEATURE / "DELIVERY.sha256"
    before = led.read_text(encoding="utf-8").split("\n")
    out, annotated = [], []
    for line in before:
        out.append(line)
        m = re.match(r"^([0-9a-f]{64})  output/(.+)$", line)
        if m and m.group(2) in moved:
            out.append(f"#   archived : output/archive/{m.group(2)}"
                       "  （hash 與內容未變，見文末「歸檔後之可驗路徑」）")
            annotated.append(m.group(2))
    tail = [f"{sha256(arc / n)}  output/archive/{n}" for n in moved]
    if ARCHIVE_HEADER not in "\n".join(before):
        out += ["", "# " + "=" * 58, ARCHIVE_HEADER, "# " + "=" * 58]
    led.write_text("\n".join(out + tail) + "\n", encoding="utf-8")

    # append-only 之驗證：既有行必須原樣留存且順序不變
    after = led.read_text(encoding="utf-8").split("\n")
    kept = [l for l in after if l in before]
    assert [l for l in before if l] == [l for l in kept if l], \
        "DELIVERY.sha256 lost or reordered an existing line — append-only 破壞"
    print(f"\n## §3.4 歸檔\n\n- 搬入 output/archive/: {len(moved)} 檔")
    print(f"- 台帳增 archived 註記 {len(annotated)} 行、"
          f"新路徑 checksum {len(tail)} 行；既有行 0 處變動")
    return moved

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--verify-only", action="store_true",
                    help="re-run §3.3 against the already-emitted file. The "
                         "§3.1 pre-gates are one-shot — gate 2 requires the "
                         "ledger to NOT yet carry this write's entry, so once it "
                         "entry is appended --write can no longer run. That "
                         "is the append-only ledger working, not a defect.")
    args = ap.parse_args()

    tcs = load_tcs()
    plan = row_plan(tcs)          # 96 §1 —— 037 之 leaf 全集，缺列留空
    if args.verify_only:
        import zipfile
        with zipfile.ZipFile(SRC) as a, zipfile.ZipFile(OUT) as b:
            differing = sorted(m for m in a.namelist() if a.read(m) != b.read(m))
        from backend.xlsx_surgical import _dv_counts
        report = {"differing": differing, "dv_counts": _dv_counts(OUT)}
        ok = assertions(plan, report)
        print(f"output sha256: {sha256(OUT)}")
        return 0 if ok else 1

    if not pre_gates(tcs):
        print("STOPPED at §3.1 — a pre-gate failed; the splice was not run.",
              file=sys.stderr)
        return 1
    if not args.write:
        print("dry-run — pre-gates only. Re-run with --write to splice.")
        return 0

    print("## §3.2 splice\n")
    print(f"- source : {SRC.name}")
    print(f"- target : {OUT.name}")
    print(f"- rows   : {FIRST_ROW}–{FIRST_ROW + len(plan) - 1}"
          f"  （{len(tcs)} TC ＋ "
          f"{len(plan) - len(tcs)} 留空列，96 §1）")
    print(f"- 未寫入欄: {NEVER_WRITE}\n")
    try:
        report = splice(plan)
    except StructureError as exc:
        print(f"ABORTED (structure invariant): {exc}", file=sys.stderr)
        return 1
    print(f"- surgical report: {report}\n")

    ok = assertions(plan, report)
    print(f"output sha256: {sha256(OUT)}")
    if ok:
        archive_previous()
    else:
        print("\n## §3.4 歸檔\n\n- **未執行** —— assertion 未全數 PASS，"
              "前一份產出檔留在 output/ 以便逐項比對")
    print()
    print("NEXT: Excel 四項確認由 Pei 執行（profile §0.1）—— 無修復提示／"
          f"R 欄下拉九項可用／D5 Scope 正確／row {FIRST_ROW}–"
          f"{FIRST_ROW + len(tcs) - 1} 內容與編號正確。"
          "程式層檢查不能代替 Excel 自身之檔案完整性判定。")
    print("本腳本到此停下：未複製至客戶交付路徑、未動 prepared 檔、"
          "未改 ENTRY 001、未執行 git。")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
