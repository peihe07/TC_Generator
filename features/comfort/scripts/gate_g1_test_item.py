#!/usr/bin/env python3
"""G-1 — Test Item shape measurement against Home's done region (handoff 16 §2).

Profile §3.1 inherits "Test Item = a condensed requirement statement, modal
allowed in this column only" from Privacy/SXM as a STRUCTURAL clause. Handoff
15 §3.1 attaches a condition to that inheritance: measure the exemplar first,
because Comfort's named exemplar is `home`, not Privacy or SXM. If home's
column is shaped differently, pilot review would reject fourteen TCs as
style-divergence after they were written.

This script MEASURES. It does not adjust §3.1 — handoff 16 §2 is explicit
that adjusting the profile to fit the measurement is Tier 2 and forbidden
here. A mismatch stops §3.1 and goes back to the analysis layer.

PROVENANCE — read before trusting the numbers:
The file home's RECON.md measured (`…_Home_20260720.xlsx`, sha256
0e72b1ec…) is NOT in this repo, and neither is Home v2 (sha256 cfc007f3…,
tag fw036-home-regen-v2). The only artefact here carrying the 144-row done
region is forms/…_Home_20260809.xlsx (sha256 1895fb2a…), which FORMS.md
documents as the pre-A-H26 build with four editorial passes applied to
D5 / F / G / K / Z. Column I (Test Item) is NOT among those passes, which is
the argument that its content is unaffected — but that argument rests on
FORMS.md's own diff, which cannot be cross-checked without Home v2 present.
The verdict below is therefore reported as PROVISIONAL on that substitution.

Usage:
    python3 features/comfort/scripts/gate_g1_test_item.py
"""

import hashlib
import random
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
HOME_WB = (ROOT / "forms" / "FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT "
           "STLA Test Case Specification & Result_SWQT_Home_20260809.xlsx")
SHEET = "Test Case Specification&Result"
HEADER_ROW = 9
COL_TEST_ITEM = "I"
COL_AUTHOR = "Z"
COL_PROCEDURE = "L"
# FORMS.md: this copy's done region carries `ArifChen`, not feature.yaml's
# `Arif` — selecting on `Arif` would match 0 rows and silently measure an
# empty population, which is the failure mode that looks like a pass.
DONE_AUTHOR = "ArifChen"
EXPECTED_DONE_ROWS = 144

# handoff 16 §2: shall/will/should/would, case-insensitive, word boundary.
MODAL_RE = re.compile(r"\b(shall|will|should|would)\b", re.I)
SAMPLE_N = 10
SAMPLE_SEED = 20260815      # fixed so the "random 10" is reproducible


def col(letter: str) -> int:
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def sha256(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def main() -> None:
    if not HOME_WB.exists():
        sys.exit(f"G-1 cannot run: {HOME_WB} not found")
    digest = sha256(HOME_WB)

    wb = openpyxl.load_workbook(HOME_WB, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"G-1: sheet {SHEET!r} not in {wb.sheetnames}")
    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()

    i_item, i_auth, i_proc = col(COL_TEST_ITEM), col(COL_AUTHOR), col(COL_PROCEDURE)

    def cell(r, i):
        return str(r[i]).strip() if i < len(r) and r[i] is not None else ""

    done = [(n, cell(r, i_item), cell(r, i_proc))
            for n, r in enumerate(rows[HEADER_ROW:], start=HEADER_ROW + 1)
            if cell(r, i_auth) == DONE_AUTHOR]

    with_modal = [(n, t) for n, t, _ in done if MODAL_RE.search(t)]
    without = [(n, t) for n, t, _ in done if t and not MODAL_RE.search(t)]
    empty = [(n, t) for n, t, _ in done if not t]

    print("=" * 72)
    print("G-1 — Test Item 形態實測（handoff 16 §2）")
    print("=" * 72)
    print("\n## 量測條件（明載，不可省）\n")
    print(f"- 檔案      : {HOME_WB.relative_to(ROOT)}")
    print(f"- SHA256    : {digest}")
    print(f"- 工作表    : {SHEET!r}")
    print(f"- header 列 : {HEADER_ROW}")
    print(f"- 量測欄    : {COL_TEST_ITEM}（Test Item）")
    print(f"- 母體選取  : {COL_AUTHOR} 欄 == {DONE_AUTHOR!r}（done region 選擇器）")
    print(f"- 母體列數  : {len(done)}")
    print(f"- modal 判準: {MODAL_RE.pattern}（case-insensitive、詞界比對）")
    print(f"- 抽樣      : {SAMPLE_N} 列，seed={SAMPLE_SEED}（固定，可重現）")

    ok_pop = len(done) == EXPECTED_DONE_ROWS
    print("\n## Assertion\n")
    print(f"- {'PASS' if ok_pop else '**FAIL**'} — done region 母體列數: "
          f"expected `{EXPECTED_DONE_ROWS}`, measured `{len(done)}`")

    print("\n## 實測結果\n")
    print(f"- 含 modal   : **{len(with_modal)}** / {len(done)} "
          f"（{len(with_modal) / len(done):.1%}）" if done else "")
    print(f"- 不含 modal : **{len(without)}** / {len(done)}")
    print(f"- 空白       : {len(empty)} / {len(done)}")
    if with_modal:
        hits = {}
        for _, t in with_modal:
            for m in MODAL_RE.findall(t):
                hits[m.lower()] = hits.get(m.lower(), 0) + 1
        print(f"- modal 詞頻 : {hits}")

    lens = [len(t) for _, t, _ in done if t]
    if lens:
        print(f"- Test Item 長度: min {min(lens)}, 中位 "
              f"{sorted(lens)[len(lens) // 2]}, max {max(lens)}")

    print(f"\n## 隨機 {SAMPLE_N} 列全文（seed={SAMPLE_SEED}）—— 供分析層判讀\n")
    rnd = random.Random(SAMPLE_SEED)
    for n, t, _ in rnd.sample(done, min(SAMPLE_N, len(done))):
        mark = "MODAL" if MODAL_RE.search(t) else "no-modal"
        print(f"- row {n} [{mark}] {t!r}")

    if not ok_pop:
        sys.exit("\nG-1 FAILED: 母體列數不符，量測基礎不成立。")


if __name__ == "__main__":
    main()
