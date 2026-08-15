#!/usr/bin/env python3
"""Clear the template's residual sample rows — profile §0.1 / A-CF07.

The FM-WI-FSM-036-A01 blank template ships two sample rows (10–11). Under
BLANK, write-back appends from the first data row, so the residue would push
the first real TC down and misnumber column B. Profile §0.1 inherits
Privacy's R23-4 procedure verbatim:

  clear D10 / F10 / G10 / S10 / D11 — five cells, `s=` style kept in place
  DO NOT touch column B  — B10 is `=IF(ISBLANK($D10),"",ROW()-9)`, the
                           template's own numbering; clearing it removes it
  DO NOT delete rows     — deletion shifts the data-validation `sqref` and
                           the x14 dropdown anchored at R10

Writes through `backend/xlsx_surgical.py`, the sole permitted write path
(profile §6 / R18-3): zip members are copied byte-for-byte and only the
patched sheet's XML may differ. Its structure verification is ABORT-level,
never downgraded to a warning.

This script PREPARES the workbook. It does not confirm it — profile §0.1
(handoff 16 §1 ruling 3) reserves the Excel open-and-check to Pei, because a
program-level check cannot stand in for Excel's own file-integrity verdict.

Usage:
    python3 features/comfort/scripts/prepare_workbook.py            # dry-run
    python3 features/comfort/scripts/prepare_workbook.py --write
"""

import argparse
import hashlib
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))
from backend.xlsx_surgical import StructureError, surgical_save  # noqa: E402

FEATURE = ROOT / "features" / "comfort"
SRC = (FEATURE / "inputs" / "FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT "
       "STLA Test Case Specification & Result_SWQT_20260121.xlsx")
OUT = FEATURE / "output" / ("FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT "
                            "STLA Test Case Specification & Result_SWQT_"
                            "Comfort_20260815_prepared.xlsx")
SHEET = "Test Case Specification 測試用例規範"
SRC_SHA = "cd876c202c71e74b0eca92dd7b4454af1879ac9a700744d5fe448687f7a9287d"

# profile §0.1 — exactly these five, and nothing else.
CLEAR = ["D10", "F10", "G10", "S10", "D11"]
KEEP = ["B10", "B11"]          # formulas; asserted unchanged after the write


def sha256(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--write", action="store_true",
                    help="actually emit; without it, only report what would change")
    args = ap.parse_args()

    if not SRC.exists():
        sys.exit(f"source template not found: {SRC}")
    digest = sha256(SRC)
    print(f"source : {SRC.relative_to(ROOT)}")
    print(f"sha256 : {digest}")
    if digest != SRC_SHA:
        # R-C14: identity by content hash, never by name. A different template
        # would make every cell reference below a guess.
        sys.exit(f"ABORT: source sha256 does not match the ruled template "
                 f"({SRC_SHA[:16]}…). Refusing to patch an unidentified file.")
    print("gate   : PASS — matches the ruled blank template\n")

    wb = openpyxl.load_workbook(SRC)
    ws = wb[SHEET]

    before = {c: ws[c].value for c in CLEAR}
    keep_before = {c: ws[c].value for c in KEEP}
    print("cells to clear (profile §0.1):")
    for c in CLEAR:
        print(f"  {c:5} = {before[c]!r}")
    print("\ncells deliberately NOT touched:")
    for c in KEEP:
        print(f"  {c:5} = {keep_before[c]!r}")

    if not args.write:
        print("\ndry-run — nothing written. Re-run with --write to emit.")
        return 0

    for c in CLEAR:
        ws[c] = None

    try:
        report = surgical_save(wb, SRC, OUT)
    except StructureError as exc:
        print(f"\nABORTED (structure invariant): {exc}", file=sys.stderr)
        return 1

    print(f"\nwritten: {OUT.relative_to(ROOT)}")
    print(f"sha256 : {sha256(OUT)}")
    print(f"report : {report}")

    # ---- post-write verification, read back from the emitted file ----------
    chk = openpyxl.load_workbook(OUT)[SHEET]
    cleared = [c for c in CLEAR if chk[c].value is None]
    kept = [c for c in KEEP if chk[c].value == keep_before[c]]
    print("\nassertions:")
    print(f"- {'PASS' if len(cleared) == len(CLEAR) else '**FAIL**'} — five "
          f"cells cleared: expected `{CLEAR}`, measured `{cleared}`")
    print(f"- {'PASS' if len(kept) == len(KEEP) else '**FAIL**'} — column B "
          f"formulas intact: expected `{KEEP}`, measured `{kept}`")
    ok = len(cleared) == len(CLEAR) and len(kept) == len(KEEP)
    print("\nNEXT: profile §0.1 reserves the Excel open-and-check to Pei — "
          "no repair prompt, R-column dropdown usable with nine entries, "
          "D5 Scope correct, rows 10–11 clear with no residual row number. "
          "Phase 4 does not start before that confirmation.")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
