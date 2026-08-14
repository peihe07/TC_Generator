#!/usr/bin/env python3
"""Applicability verdict for the 17 in-baseline `substantive` sections (D-C10 前置).

Handoff 05 §5 / 06 §3. A-CF08 established that 17 SR24 sections carry
behavioural text the 037 never analysed. Classification answered "does this
read like a requirement"; this answers the different question D-C10 needs:
"should Comfort R1L-R verify it".

Three verdicts. `undetermined` is a legitimate — and required — outcome:
a section whose governing source is not in `inputs/` is undetermined, NOT
out_of_scope. Failing to read something is not evidence of its absence
(handoff 06 §3).

The machine-checkable half is re-derived on every run from the CFTS043
R1L-R scope tree view: whether each cited CFTS043 item sits in the workbook's
own `Scope=Yes` allow-list (sheet `工作表1`, 599 ForeignIDs), and its Radio /
EE Architecture / Market attributes. The judgement half — which CFTS043
section governs which SR24 clause — is declared here and cited per row, so a
reader can check the reasoning rather than take the verdict on trust.

MEASUREMENT ONLY. No TC is generated, nothing enters a coverage denominator,
nothing is marked BLOCKED, no RD item is filed, and R-C5 / R-C5-1 are
untouched (handoff 06 §3).

Usage:
    python3 features/comfort/scripts/applicability_substantive.py
"""

import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
FEATURE = ROOT / "features" / "comfort"
TREE = (FEATURE / "inputs" /
        "SYS1_CFTS043-HVAC Controls and Displays_Tree view_R1L-R scope.xlsx")
RECON = FEATURE / "data" / "recon.json"
OUT = FEATURE / "data" / "sr24_substantive_applicability.tsv"

# CFTS043 §1.3.5.1.22 "Alternate Rear Blower Control Softkeys" and its three
# sub-sections. SR24 §20's own title says "See CFTS043 for applicable
# vehicles", so this section IS the ruled answer to that pointer.
#
# NOTE ON THE TITLE MISMATCH: SR24 writes "LATAM Alternative Rear Blower";
# CFTS043 writes "Alternate Rear Blower" and never uses "Alternative" or
# LATAM for it. Searching the customer's own wording returns zero hits and
# would have produced a false out_of_scope — see A-CF11.
ARB_SECTION = "1.3.5.1.22 Alternate Rear Blower Control Softkeys"
ARB_ITEMS = {
    "20.1":   ["4803261", "4803262"],   # CRB1 — softkey availability vs vent mode
    "20.1.1": ["4803263"],              # CRB1.1 — pop-up on rear hard control
    "20.1.2": ["4803263", "4803265"],   # CRB1.2 — rear hard control while unavailable
    "20.1.3": ["4803263", "4803265"],   # CRB1.3 — unavailable + pop-up (PU1268)
    "20.2":   ["4803279", "4803281", "4803284", "4803285", "4803286"],  # CRB2 — Lock
    "20.3":   ["4803264", "4803275", "4803276", "4803278"],             # CRB3 — Select, 1-4
    "20.4":   ["4803266", "4803267", "4803271", "4803272"],             # CRB4 — Power
    "20.4.1": ["4803268", "4803272"],   # CRB4.1 — REAR ON re-activates
    "20.4.2": ["4803261", "4803263"],   # CRB4.2 — grey out vs front vent mode
    "20.4.3": ["4803262", "4803271"],   # CRB4.3 — OFF via HVAC mode hard control
}
PROXI_GATE = "$Indipendent_Rear_Fan$ = [Present] (CFTS043 4803260)"

# --------------------------------------------------------------- DR #6 / #7
#
# The Market Configuration Table (25PI3.5, sha256 ae4cf0b9…) was supplied to
# settle the screen-size and market questions. It does not: measured across
# all 8 sheets it carries ZERO occurrences of `R1L-R` and zero of any screen
# size. It is a country-keyed reference (149 EMEA / 37 APAC / 19 NAFTA /
# 19 LATAM) mapping countries to MARKET radio variants (ROW / ECE / US-CAN /
# ROW+ / CHN / JPN / MEX / KOR) — a different axis from radio MODEL, and it
# says nothing about which screens or markets THIS delivery covers.
#
# Per R-C13 the zero hits are an index-layer fact only, so a second and third
# path were run. The one that produced evidence was structural, in the ruled
# source itself: which sibling sections the 037 actually analysed.
#
#   ch 16  "ICS CLIMATE EMEA – CARRYOVER"      18 of 19 children cited, 99 leaves
#   ch 18  "10.25\" Home screen Comfort Widget"  1 of 4 children cited (18.1), 3 leaves
#   ch 19  "7\" Home screen Comfort Widget"      0 of 3 children cited, 0 leaves
#
# THE INFERENCE IS ONE-WAY, and the direction matters. Presence is evidence:
# the 037 could not have produced 99 leaves against an EMEA chapter, or 3
# against a 10.25"-specific one, if those dimensions were outside the
# delivery. Absence is NOT evidence: "the 037 did not cite it" was exactly
# the step that produced the R-C5 error (A-CF01), corrected by R-C5-1. So
# ch 19's silence buys nothing in either direction and 19.x stays
# undetermined.
STRUCTURAL = {
    "16.1": ("in_scope", "",
             "EMEA is within this delivery's market scope. Evidence is the "
             "ruled source's own behaviour: the 037 cites 18 of the 19 "
             "children of SR24 ch16 'ICS CLIMATE EMEA – CARRYOVER' "
             "(16.2–16.17), yielding 99 leaves = 25% of the feature; that is "
             "not producible if EMEA were out of delivery scope. 16.1 is the "
             "one uncited child, i.e. a coverage gap inside an in-scope "
             "chapter (A-CF08), not a scope question. "
             "NOT established by the Market Configuration Table: that file "
             "carries EMEA only as a 149-country geographic grouping for all "
             "R1 radios, with no R1L-R row and no delivery-scope statement. "
             "No contradicting source found, so R-C12 does not bite.",
             "EMEA market"),
    "18.2": ("in_scope", "",
             "10.25\" is within this delivery's screen configuration. "
             "Evidence: the 037 cites SR24 18.1 (ch18 = '10.25\" Home screen "
             "- Comfort Widget'), yielding leaves SWE1-HVAC-129-01/-02/-03. "
             "Note 18.1 and 19.1 carry the SAME clause text (W0.) — the 037 "
             "analysed the 10.25\" instance and not the 7\" one. Under either "
             "reading of that choice (10.25\" is the delivery's screen / the "
             "author de-duplicated a repeated clause) the analysed instance "
             "is the 10.25\" one, so this verdict is robust to both. "
             "18.2–18.4 are uncited children of that in-scope chapter — a "
             "coverage gap (A-CF08), not a scope question. "
             "NOT established by the Market Configuration Table (no screen "
             "size anywhere in it). No contradicting source, R-C12 not "
             "triggered.",
             "10.25\" display"),
    "18.3": None, "18.4": None,      # same basis as 18.2, filled below
    "19.1": ("undetermined", "DR #6 — 7\" screen configuration for R1LR ATL-H",
             "UNDETERMINED, and deliberately not out_of_scope. The Market "
             "Configuration Table carries no screen-size axis at all. The "
             "secondary candidate named in handoff 08 §3, 'Vehicle Category "
             "HMI Logic and Flow R1 SR24 Post 2A', WAS checked before use as "
             "that section requires: it states only which radios the DOCUMENT "
             "covers (R1 Low: 7\", 8.4\", 10.1\", 10.25\", 12.3\") — the same "
             "shape as SR24 §1.1, which handoff 06 §3 already ruled is not "
             "evidence of delivery scope — and contains zero 'R1L-R', zero "
             "'Atlantis', no configuration table. It fails verification and "
             "is NOT used. The 037 citing nothing in ch19 is NOT treated as "
             "evidence of exclusion (that inference is the R-C5 error, "
             "corrected by R-C5-1). Still needed: a source stating which "
             "screens THIS delivery ships.",
             "7\" display (unconfirmed for R1LR ATL-H)"),
    "19.2": None, "19.3": None,      # same basis as 19.1
}


def load_tree() -> tuple[dict, set]:
    if not TREE.exists():
        sys.exit(f"CFTS043 tree view not found: {TREE}")
    wb = openpyxl.load_workbook(TREE, read_only=True, data_only=True)
    allow = {str(r[1]).strip() for r in wb["工作表1"].iter_rows(values_only=True)
             if r[1] and str(r[0]).strip() == "Yes"}
    rows = list(wb["Basic Report"].iter_rows(values_only=True))
    hdr = [str(c) if c else "" for c in rows[0]]
    ix = {h: j for j, h in enumerate(hdr)}
    items: dict = {}
    for r in rows[1:]:
        fid = str(r[ix["ReqIF.ForeignID"]] or "").strip()
        if fid and fid not in items:
            items[fid] = {k: str(r[ix[k]] or "").strip() for k in
                          ("Scope", "Radio", "EE Architecture", "Market",
                           "Source Id")}
    wb.close()
    return items, allow


# R-C12 — a verdict whose source carries an unresolved internal contradiction
# is recorded undetermined, never in_scope: in_scope is a claim that WIDENS
# the verification scope and must stand on its own evidence, while
# undetermined can still converge either way once the contradiction is
# settled. The 20.x evidence below is kept in full (R-C12 requires it) — the
# downgrade does not retract it, and does not say the conclusion is false.
# Handoff 10 §2 — Pei rules DR #8 directly with RD, not through this pipeline.
# Per R15-2 an open PENDING means "awaiting a ruling", not "awaiting an
# external condition", so DR #8 leaves the PENDING list and the D-C10 blocker
# list. DEFERRED does NOT move the verdict: the contradiction is unresolved
# either way, so R-C12 still holds these at undetermined — deferring the
# question is not answering it.
RC12_PENDING = ("DEFERRED — Pei 直接向 RD 反應（2026-08-14）；原 DR #8 "
                "CFTS043 4803259 NOTE vs its own Radio attribute (A-CF12)")

# R-C16 §2 — a section that is in_scope but which the 037 never analysed is an
# in-scope COVERAGE GAP, not a TC work item. It goes to RD-1 for upstream 037
# analysis; until that lands it enters no coverage denominator, is not marked
# BLOCKED, and gets no tc_id. Recorded per row so the distinction survives
# into whatever reads this file next.
RC16_DISPOSITION = ("RD-1 coverage-gap item (R-C16) — 037 never analysed it; "
                    "NOT a TC work item, no tc_id, not in the coverage "
                    "denominator, not BLOCKED, pending upstream 037 analysis")


def main() -> None:
    items, allow = load_tree()
    out = ["outline\tscope_verdict\tbasis\tvariant_condition\tpending_on"
           "\tdisposition"]
    counts = {"in_scope": 0, "out_of_scope": 0, "undetermined": 0}

    for outline in sorted(ARB_ITEMS, key=lambda s: [int(x) for x in s.split(".")]):
        fids = ARB_ITEMS[outline]
        seen = [f for f in fids if f in items]
        missing = [f for f in fids if f not in items]
        in_allow = [f for f in seen if f in allow]
        radios = sorted({items[f]["Radio"] for f in seen})
        r1lr = all("R1L-R" in items[f]["Radio"] for f in seen) and seen
        if seen and len(in_allow) == len(seen) and r1lr:
            # Evidence retained verbatim; only the verdict moves, plus the
            # handoff 07 §3 correction to how the tree view is characterised.
            # It is a SYS.1 traceability index export, not an original source,
            # and per §8.6 an index must not outrank the source document it
            # indexes. So Scope=Yes is corroboration, not the "most direct
            # statement" the earlier package called it.
            basis = (f"[R-C12: downgraded from in_scope 2026-08-14; evidence "
                     f"below retained in full, not retracted] "
                     f"CFTS043 §{ARB_SECTION}; SR24 §20 title itself directs "
                     f"'See CFTS043 for applicable vehicles'. Items "
                     f"{'/'.join(seen)} all carry Scope=Yes in the tree view's "
                     f"own R1L-R allow-list (sheet 工作表1, {len(allow)} ids) "
                     f"and Radio includes R1L-R ({radios[0]}). Market=All. "
                     f"EE=Atlantis Mid, which does NOT gate scope here: the "
                     f"Scope=Yes set spans Atlantis High and Mid alike. "
                     f"LAYER (handoff 07 §3): the tree view is a SYS.1 "
                     f"traceability INDEX export, not an original source; per "
                     f"§8.6 it cannot outrank the CFTS043 main .doc it "
                     f"indexes. Scope=Yes is index-layer corroboration only. "
                     f"The unresolved contradiction is therefore internal to "
                     f"the main doc: item 4803259's prose NOTE ('only "
                     f"applicable to R1H starting on SR22') against the same "
                     f"item's Radio attribute (includes R1L-R). If forced to "
                     f"choose today, canon weight sits with the prose, i.e. "
                     f"toward out_of_scope — the opposite of the earlier "
                     f"provisional value.")
            verdict, pending = "undetermined", RC12_PENDING
        else:
            verdict, pending = "undetermined", RC12_PENDING
            basis = (f"CFTS043 §{ARB_SECTION} mapping incomplete — "
                     f"items not found in tree view: {missing or 'none'}; "
                     f"not in allow-list: {sorted(set(seen) - set(in_allow))}")
        counts[verdict] += 1
        # Undetermined: no disposition yet — R-C16 §2 applies to in_scope
        # sections, and these are not in_scope. If DR #8 ever resolves them
        # to in_scope, R-C16 makes them RD-1 items too (09 §3 says so
        # explicitly), but that is not today's state and is not recorded as
        # though it were.
        out.append("\t".join([outline, verdict, basis, PROXI_GATE, pending,
                              "—（verdict 未定，尚無處置）"]))

    base18, base19 = STRUCTURAL["18.2"], STRUCTURAL["19.1"]
    for outline in ("16.1", "18.2", "18.3", "18.4", "19.1", "19.2", "19.3"):
        spec = STRUCTURAL[outline]
        if spec is None:
            src, tag = (base18, "18.2") if outline.startswith("18.") else (base19, "19.1")
            spec = (src[0], src[1], src[2] + f" (same basis as {tag})", src[3])
        verdict, pending, basis, variant = spec
        counts[verdict] += 1
        disposition = (RC16_DISPOSITION if verdict == "in_scope"
                       else "—（verdict 未定，尚無處置）")
        out.append("\t".join([outline, verdict, " ".join(basis.split()),
                              variant, pending, disposition]))

    OUT.write_text("\n".join(out) + "\n", encoding="utf-8")
    total = sum(counts.values())
    print(f"{total} sections written to {OUT.relative_to(ROOT)}")
    for k, v in counts.items():
        print(f"  {k:<14} {v}")
    if total != 17:
        sys.exit(f"expected 17 substantive sections, wrote {total}")


if __name__ == "__main__":
    main()
