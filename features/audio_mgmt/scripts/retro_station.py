#!/usr/bin/env python3
"""R-AM18 retro-verification station, items 1-6 (package 30 section 2).

Item 7 depends on the re-exported pool DR-AM3 asks for and is NOT done here.
Package 30 gates the whole station on DR-AM3 returning; Pei ruled on
2026-08-27 to run the first six ahead of that, so item 7 is the only one
left and the station stays open.

Two kinds of change, and they are not the same operation:
  - append: the leaf keeps its behaviour anchor and gains the object that
    defines the value it reads. One id per line, R-AM2 / IN 10.7(a).
  - re-anchor: the anchor was wrong. The old id goes.

reasoning is not a workbook column, so item 6 touches the JSON only.

    python features/audio_mgmt/scripts/retro_station.py --write
"""

import argparse
import json
import shutil
import sys
from pathlib import Path

import openpyxl

FEATURE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(FEATURE.parent.parent))
sys.path.insert(0, str(FEATURE.parent.parent / "scripts"))
import write_back  # noqa: E402  applies the A-AM18 sparse-diff patch
from backend.xlsx_surgical import surgical_save  # noqa: E402
from feature_config import load_feature_config  # noqa: E402

APPEND = {  # leaf -> (id to append, why)
    "SWE1_AMM_287": ("CFTS019-4867782", "<Vent off> = -16 dB"),
    "SWE1_AMM_312": ("CFTS019-4867783", "<Vent Nav Off>"),
    "SWE1_AMM_313": ("CFTS019-4867783", "<Vent Nav Off>"),
    "SWE1_AMM_314": ("CFTS019-4867783", "<Vent Nav Off>"),
    "SWE1_AMM_315": ("CFTS019-4867783", "<Vent Nav Off>"),
    "SWE1_AMM_316": ("CFTS019-4867783", "<Vent Nav Off>"),
    "SWE1_AMM_317": ("CFTS019-4867783", "<Vent Nav Off>"),
}
REANCHOR = {  # leaf -> (old, new)
    "SWE1_AMM_264": ("CFTS019-4867598", "CFTS019-4867603"),
    "SWE1_AMM_268": ("CFTS019-4867647", "CFTS019-4867640"),
    "SWE1_AMM_174": ("CFTS019-4866662", "CFTS019-4866632"),
}
REASONING = {  # item 6, JSON only
    "SWE1_AMM_169": ("The anchor is the first of four verbatim parallel "
                     "instances and the leaf is an abstraction across the "
                     "sub-sections, so the position window is too wide to "
                     "discriminate here (package 30 section 2 item 6)."),
}
NOTE = {
    "append": ("R-AM18 retro station item {n}: {why} is defined at {new}, "
               "which is appended so the case cites the object defining the "
               "value it reads as well as the behaviour object."),
    "reanchor": ("R-AM18 retro station item {n}: re-anchored from {old} to "
                 "{new}."),
}


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--write", action="store_true")
    args = ap.parse_args()

    cfg = load_feature_config(FEATURE)
    src = FEATURE / "generated" / "SWQT_AudioMgmt_B1-B7.xlsx"
    wb = openpyxl.load_workbook(src)
    ws = wb[cfg["workbook"]["sheet"]]
    cols = write_back.resolve_columns(ws, cfg["workbook"]["header_row"], cfg)
    req_c, tc_c, spec_c = cols["req_id"], cols["tc_id"], cols["spec_reference"]

    plan, bad = [], []
    for r in range(cfg["workbook"]["header_row"] + 1, ws.max_row + 1):
        leaf = str(ws.cell(r, req_c).value or "").strip()
        cur = str(ws.cell(r, spec_c).value or "").strip()
        if not leaf:
            continue
        if leaf in APPEND:
            new_id, why = APPEND[leaf]
            if new_id in cur.split("\n"):
                bad.append(f"r{r} {leaf}: {new_id} already present")
                continue
            plan.append((r, leaf, cur, cur + "\n" + new_id, "append"))
        elif leaf in REANCHOR:
            old, new = REANCHOR[leaf]
            if cur != old:
                bad.append(f"r{r} {leaf}: expected {old!r}, found {cur!r}")
                continue
            plan.append((r, leaf, cur, new, "reanchor"))

    for r, leaf, cur, new, kind in plan:
        print(f"r{r:<4} {leaf:16} {kind:9} "
              f"{cur.replace(chr(10), '+')!r} -> {new.replace(chr(10), '+')!r}")
    print(f"\nrows to change: {len(plan)} "
          f"(append {sum(1 for p in plan if p[4] == 'append')}, "
          f"reanchor {sum(1 for p in plan if p[4] == 'reanchor')})")
    for b in bad:
        print("  refused:", b)
    want_re = len(REANCHOR)
    got_re = sum(1 for p in plan if p[4] == "reanchor")
    if bad or got_re != want_re:
        print("\nrefusing: preconditions not met")
        return 1
    if not args.write:
        print("\ndry run — nothing written. Pass --write to emit.")
        return 0

    # JSON first: spec_reference on both kinds, reasoning on all three.
    touched = {}
    for path in sorted((FEATURE / "generated").glob("B?.json")):
        data = json.loads(path.read_text(encoding="utf-8"))
        n = 0
        for tc in data["tcs"]:
            leaf, cur = tc["req_id"], tc["spec_reference"]
            if leaf in APPEND and APPEND[leaf][0] not in cur.split("\n"):
                new_id, why = APPEND[leaf]
                tc["spec_reference"] = cur + "\n" + new_id
                item = 1 if leaf == "SWE1_AMM_287" else 2
                tc["reasoning"] = tc["reasoning"].rstrip() + " " + NOTE[
                    "append"].format(n=item, why=why, new=new_id)
                n += 1
            elif leaf in REANCHOR and cur == REANCHOR[leaf][0]:
                old, new = REANCHOR[leaf]
                tc["spec_reference"] = new
                item = {"SWE1_AMM_264": 3, "SWE1_AMM_268": 4,
                        "SWE1_AMM_174": 5}[leaf]
                tc["reasoning"] = tc["reasoning"].rstrip() + " " + NOTE[
                    "reanchor"].format(n=item, old=old, new=new)
                n += 1
            if leaf in REASONING and REASONING[leaf] not in tc["reasoning"]:
                tc["reasoning"] = tc["reasoning"].rstrip() + " " \
                    + REASONING[leaf]
                n += 1
        if n:
            path.write_text(json.dumps(data, ensure_ascii=False, indent=2)
                            + "\n", encoding="utf-8")
            touched[path.name] = n
    print(f"\njson updated: {touched}")

    for r, _, _, new, _ in plan:
        ws.cell(r, spec_c).value = new

    stage = src.with_name(src.stem + "_stage.xlsx")
    tmp = src.with_name(src.stem + "_tmp.xlsx")
    shutil.copy2(src, stage)
    report = surgical_save(wb, stage, tmp, verify=True)
    cf = write_back.check_conditional_formatting(stage, tmp)
    stage.unlink()
    tmp.replace(src)

    print(f"\nwrote    : {src.name}")
    print(f"  sha256 : {write_back.sha256_file(src)}")
    print(f"  members: {report['members']}, patched {report['members_patched']}")
    print(f"  dv counts unchanged (classic, x14): {report['dv_counts']}")
    print(f"  conditionalFormatting unchanged  : {cf}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
