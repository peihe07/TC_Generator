#!/usr/bin/env python3
"""Route 2 for B5, with the three gates package 18 section 3 requires.

A negative conclusion — "no object matches this leaf" — is only allowed
after all three, which is the A-AM11 improvement list made mandatory:

  1. word-form expansion: singular and plural, and the obvious synonyms.
     146 was missed by searching "remaining channel" where the spec writes
     "remaining audio channels".
  2. untruncated reading: the whole clause, never a window. 020 was rejected
     on a row cut at 135 characters whose next clause was the match.
  3. self-citation scan: the leaf's own SWE.1 description, for a CFTS019-nnn
     it may name outright. 024's anchor was sitting there.

A fourth check reports objects whose exported Description is an image or
wrapper reference: no text search can reach those, so a miss there means
nothing (024 again).

Pool basis is expanded v2, 891 ids (A-AM12).

Usage:
    python features/audio_mgmt/scripts/route2_b5.py --grade B
    python features/audio_mgmt/scripts/route2_b5.py --leaf SWE1_AMM_168
    python features/audio_mgmt/scripts/route2_b5.py --scan 4866591-4866602
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl

FEATURE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(FEATURE.parent.parent / "scripts"))
from feature_config import load_feature_config, resolve_path  # noqa: E402

HANDOFF = FEATURE / "docs" / "handoff" / "18_B5_anchor_candidates.md"
IMAGEY = re.compile(r"\(image:|WrapperResource|\.rtf\b", re.I)
ATTRS = re.compile(r"\[(Artifact|State|ECU|Market|Model|Radio|EE)[^\]]*\]")


def pool() -> dict[str, dict]:
    """ObjectID -> exported row, expanded pool v2 (A-AM12)."""
    cfg = load_feature_config(FEATURE)
    out = {}
    for key in ("sys1_export", "sys1_export_part2"):
        wb = openpyxl.load_workbook(resolve_path(cfg, key), read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c or "").strip().lower() for c in rows[0]]
        oid_i = max(range(len(header)), key=lambda i: sum(
            1 for r in rows[1:]
            if re.fullmatch(r"\s*48\d{5}\s*", str(r[i] or ""))))
        desc_i = next(i for i, h in enumerate(header) if h == "description")
        for r in rows[1:]:
            for oid in re.findall(r"\b(48\d{5})\b", str(r[oid_i] or "")):
                out[oid] = {"desc": " ".join(str(r[desc_i] or "").split()),
                            "src": key}
    return out


def blocks() -> dict[str, str]:
    text = (FEATURE / "data" / "cfts019_text.txt").read_text(encoding="utf-8")
    return {m.group(1): " ".join(m.group(2).split()) for m in re.finditer(
        r"^(48\d{5}): (\[.*?)(?=^\d{7}: \[|\Z)", text, re.M | re.S)}


def body(text: str) -> str:
    return ATTRS.sub("", text).strip()


def swe_rows() -> dict[str, dict]:
    cfg = load_feature_config(FEATURE)
    wb = openpyxl.load_workbook(resolve_path(cfg, "a03_report"), read_only=True)
    out = {}
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if row[0]:
                out.setdefault(str(row[0]).strip(), {
                    "source_id": str(row[1] or ""), "title": str(row[2] or ""),
                    "desc": " ".join(str(row[3] or "").split())})
    return out


def forms(word: str) -> set[str]:
    """Gate 1: word-form expansion."""
    w = word.lower()
    out = {w}
    out.add(w + "s")
    if w.endswith("s"):
        out.add(w[:-1])
    if w.endswith("y"):
        out.add(w[:-1] + "ies")
    return out


def search(terms: list[str], corpus: dict[str, str], p: dict) -> list[tuple]:
    """Every object whose text carries every term in some form."""
    hits = []
    for oid, text in corpus.items():
        low = body(text).lower()
        if all(any(f in low for f in forms(t)) for t in terms):
            hits.append((oid, oid in p, body(text)))
    return sorted(hits)


def grades() -> dict[str, list[tuple[str, str]]]:
    t = HANDOFF.read_text(encoding="utf-8")
    a = t[t.index("## 一、A 級"):t.index("## 二、B 級")]
    b = t[t.index("## 二、B 級"):t.index("## 三、C 級")]
    pat = r"(SWE1_AMM_\d+)\s*\|?\s*(?:→|CFTS019-)\s*(?:CFTS019-)?(48\d{5})"
    alt = r"(\d{3})→(48\d{5})"
    out = {"A": re.findall(r"\|\s*(SWE1_AMM_\d+)\s*\|\s*CFTS019-(48\d{5})", a),
           "B": re.findall(r"\|\s*(SWE1_AMM_\d+)\s*\|\s*(?:CFTS019-)?(48\d{5})", b)
                + [(f"SWE1_AMM_{n}", o) for n, o in re.findall(alt, b)]}
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--grade", choices=["A", "B"])
    ap.add_argument("--leaf")
    ap.add_argument("--terms", nargs="*", help="gate 1 search, all terms required")
    ap.add_argument("--scan", help="read an ObjectID range, e.g. 4866591-4866602")
    args = ap.parse_args()

    p, blk, swe = pool(), blocks(), swe_rows()

    if args.scan:
        lo, hi = (int(x) for x in args.scan.split("-"))
        for oid in sorted(blk):
            if lo <= int(oid) <= hi and body(blk[oid]):
                print(f"{oid} {'pool' if oid in p else 'OUT '} | {body(blk[oid])}")
        return 0

    if args.terms:
        hits = search(args.terms, blk, p)
        print(f"terms {args.terms} (gate 1 expands each form): {len(hits)} hits")
        for oid, inpool, text in hits:
            print(f"  {oid} {'pool' if inpool else 'OUT '} | {text[:400]}")
        return 0

    g = grades()
    pairs = ([(s, o) for gr in g.values() for s, o in gr if s == args.leaf]
             if args.leaf else g[args.grade])
    for sid, oid in pairs:
        rec = swe.get(sid, {})
        row = p.get(oid)
        cited = re.findall(r"CFTS019-(\d+)", rec.get("desc", ""))
        print("=" * 78)
        flag = "" if row else "   << NOT IN POOL >>"
        print(f"{sid}  candidate CFTS019-{oid}{flag}")
        print(f"  LEAF : {rec.get('title', '')}")
        print(f"         {rec.get('desc', '')}")            # gate 2: no cut
        if cited:                                            # gate 3
            print(f"  SELF-CITES: {cited} <- the leaf names an ObjectID")
        if row:
            note = "  [image/wrapper — no text search can reach it]" if \
                IMAGEY.search(row["desc"]) else ""
            print(f"  POOL : {row['desc']}{note}")           # gate 2: no cut
        elif oid in blk:
            print(f"  TEXT : {body(blk[oid])}")
    print("=" * 78)
    print(f"{len(pairs)} pairs, read whole. Pool basis expanded v2, "
          f"{len(p)} ids.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
