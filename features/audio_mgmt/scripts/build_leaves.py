#!/usr/bin/env python3
"""Build the full SWE.1 leaf set for audio_mgmt across all four sheets.

The SWE.1 report splits its 318 rows over `Audio-Management Part 01..04`.
`scripts/recon.py` surveys a single sheet (`paths_meta.a03_sheet`), so recon
sees only the representative sheet; this file is the authority for the leaf
set, mirroring the vehicle_setting arrangement (R-VS4: recon takes one 037 as
representative, `data/leaves.tsv` holds the full set).

Usage:
    python features/audio_mgmt/scripts/build_leaves.py
"""

import sys
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT.parent.parent / "scripts"))

from feature_config import load_feature_config, resolve_path  # noqa: E402

# Header row 1, data from row 2; column order verified 2026-08-26.
COLS = {
    "swe_id": 0,
    "source_id": 1,
    "title": 2,
    "description": 3,
    "status": 4,
    "categorization": 6,
    "sub_categorization": 7,
    "priority": 16,
    "verification_criteria": 17,
    "verification_method": 18,
}
OUT_COLS = ["swe_id", "source_id", "sheet", "title", "status",
            "categorization", "sub_categorization", "priority"]


def clean(v: object) -> str:
    """Flatten a cell to one TSV-safe line."""
    if v is None:
        return ""
    return " ".join(str(v).split()).replace("\t", " ")


def main() -> None:
    cfg = load_feature_config(ROOT)
    wb = openpyxl.load_workbook(resolve_path(cfg, "a03_report"), read_only=True)

    rows = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        part = sheet.replace("Audio-Management ", "")
        for row in ws.iter_rows(min_row=2, values_only=True):
            swe_id = clean(row[COLS["swe_id"]])
            if not swe_id:
                continue
            rows.append({"sheet": part,
                         **{k: clean(row[i]) for k, i in COLS.items()}})

    out = ROOT / "data" / "leaves.tsv"
    out.parent.mkdir(exist_ok=True)
    with out.open("w", encoding="utf-8") as fh:
        fh.write("\t".join(OUT_COLS) + "\n")
        for r in rows:
            fh.write("\t".join(r[c] for c in OUT_COLS) + "\n")

    ids = [r["swe_id"] for r in rows]
    dupes = {k: v for k, v in Counter(ids).items() if v > 1}
    src_dupes = {k: v for k, v in Counter(r["source_id"] for r in rows).items()
                 if v > 1}

    print(f"wrote {out.relative_to(ROOT.parent.parent)}")
    print(f"  rows           {len(rows)}")
    print(f"  unique SWE ID  {len(set(ids))}")
    print(f"  per sheet      "
          + ", ".join(f"{k}={v}" for k, v in
                      Counter(r['sheet'] for r in rows).items()))
    print(f"  SWE ID dupes   {dupes or 'none'}")
    print(f"  source dupes   {src_dupes or 'none'}")
    for r in rows:
        if r["swe_id"] in dupes:
            print(f"    {r['swe_id']}  {r['source_id']}  [{r['sheet']}]  "
                  f"{r['title'][:52]}")


if __name__ == "__main__":
    main()
