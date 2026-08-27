#!/usr/bin/env python3
"""Assemble the generation context for an audio_mgmt batch.

For every leaf in the batch this pulls together, from ruled sources only:
  - the SWE.1 row (Requirement Description is the verbatim source for the
    upper half of test_item per package 03 section 3.2 / IN R-S4)
  - the CFTS019 anchor text, looked up by ObjectID in the full-text PDF
  - the sibling axis the leaf sits on (package 03 section 5)

Anchors are read from the package 03 section 4 table and never recomputed:
the execution layer may not re-anchor (package 03 section 3.4). Leaves whose
anchor falls outside the R-AM2 pool are carried with `anchor_in_pool: false`
so the generator can hold them back (A-AM03).

Usage:
    python features/audio_mgmt/scripts/build_batch_context.py --batch B1
"""

import argparse
import json
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT.parent.parent / "scripts"))

from feature_config import load_feature_config, resolve_path  # noqa: E402

HANDOFF = {"B1": ROOT / "docs" / "handoff" / "03_batch_B1_handoff.md",
           "B2": ROOT / "docs" / "handoff" / "08_B2_final_anchors.md",
           "B3": ROOT / "docs" / "handoff" / "12_B3_final_anchors.md",
           "B4": ROOT / "docs" / "handoff" / "13_B4_anchor_candidates.md",
           "B5": ROOT / "docs" / "handoff" / "18_B5_anchor_candidates.md",
           "B6": ROOT / "docs" / "handoff" / "22_B6_anchor_candidates.md",
           "B7": ROOT / "docs" / "handoff" / "26_B7_final_batch.md"}

# B7, the final batch: package 26's ruled table, verified by route 2.
B7_ANCHORS = {"174": "4866662", "177": "4866674", "178": "4866677", "188": "4866714", "221": "4866489", "242": "4867027", "243": "4867028", "244": "4867029", "245": "4867162", "246": "4867426", "247": "4867457", "297": "4866141", "298": "4866142", "299": "4866143", "300": "4866144", "301": "4866145", "302": "4866146", "303": "4866147"}
B7_SET = {"174": "Power and Persistence", "177": "Power and Persistence", "178": "Power and Persistence", "188": "Power and Persistence", "221": "Power and Persistence", "242": "Logistic Mode", "243": "Logistic Mode", "244": "Logistic Mode", "245": "Power and Persistence", "246": "Power and Persistence", "247": "Power and Persistence", "297": "Power and Persistence", "298": "Power and Persistence", "299": "Power and Persistence", "300": "Power and Persistence", "301": "Power and Persistence", "302": "Power and Persistence", "303": "Power and Persistence"}

# B6: package 22's B table with package 24's rulings applied throughout —
# the eight returned leaves, the three cluster resolutions, D-B6-01's 131
# correction, and A-AM14-b's Surround re-ordering. 140 ships PENDING against
# DR-AM10: its description says restore at initialisation while the object
# its position fixes says store, and neither route may reconcile that.
B6_ANCHORS = {"052": "4866100", "057": "4866117", "062": "4866124", "092": "4866251", "093": "4866254", "094": "4866256", "095": "4866257", "096": "4866259", "097": "4866260", "098": "4867602", "099": "4867601", "100": "4866264", "101": "4866261", "102": "4866266", "103": "4866267", "104": "4866268", "105": "4866271", "110": "4866295", "111": "4866296", "112": "4866297", "113": "4866298", "115": "4866300", "116": "4866301", "117": "4866304", "118": "4866305", "120": "4866310", "121": "4866311", "131": "4866466", "140": None, "161": "4866530", "170": "4866607", "171": "4866616", "172": "4866620", "173": "4866629", "248": "4867484", "249": "4867486", "250": "4867487", "251": "4867507", "252": "4867508", "253": "4867510", "254": "4867517", "255": "4867518", "258": "4867577", "261": "4867581", "265": "4866263", "267": "4867639", "268": "4867647", "269": "4867641", "270": "4867643", "271": "4867648"}
B6_SET = {"052": "Power and Persistence", "057": "Power and Persistence", "062": "Power and Persistence", "092": "Surround and Fade", "093": "Surround and Fade", "094": "Surround and Fade", "095": "Surround and Fade", "096": "Surround and Fade", "097": "Surround and Fade", "098": "Surround and Fade", "099": "Surround and Fade", "100": "Surround and Fade", "101": "Surround and Fade", "102": "Surround and Fade", "103": "Surround and Fade", "104": "Surround and Fade", "105": "Surround and Fade", "110": "Surround and Fade", "111": "Surround and Fade", "112": "Surround and Fade", "113": "Surround and Fade", "115": "Surround and Fade", "116": "Surround and Fade", "117": "Surround and Fade", "118": "Surround and Fade", "120": "Surround and Fade", "121": "Surround and Fade", "131": "Power and Persistence", "140": "Power and Persistence", "161": "Power and Persistence", "170": "Power and Persistence", "171": "Power and Persistence", "172": "Power and Persistence", "173": "Power and Persistence", "248": "Audio Processing", "249": "Audio Processing", "250": "Audio Processing", "251": "Audio Processing", "252": "Audio Processing", "253": "Audio Processing", "254": "Audio Processing", "255": "Audio Processing", "258": "Audio Processing", "261": "Audio Processing", "265": "Audio Processing", "267": "Audio Processing", "268": "Audio Processing", "269": "Audio Processing", "270": "Audio Processing", "271": "Audio Processing"}

# B5: package 18's A and B tables, with package 20's rulings applied. 293 is
# held — package 20 section 2.3 authorised writing it only if route 2 found
# an in-pool anchor, and 4866193 is out of pool, so it goes back as a single
# item rather than being written on one route.
# 293 was held pending a ruling; the ruling landed with the B5 delivery
# review — 4866193, out of pool, partial coverage — so nothing is held now.
B5_HELD: set[str] = set()
B5_SET = {"011": "Tones and Alerts", "012": "Tones and Alerts", "016": "Audio Processing", "017": "Tones and Alerts", "018": "Tones and Alerts", "019": "Tones and Alerts", "021": "Tones and Alerts", "022": "Tones and Alerts", "023": "Tones and Alerts", "025": "Audio Processing", "029": "Audio Processing", "033": "Tones and Alerts", "034": "Tones and Alerts", "035": "Audio Processing", "036": "Tones and Alerts", "037": "Tones and Alerts", "038": "Tones and Alerts", "039": "Tones and Alerts", "040": "Tones and Alerts", "041": "Audio Processing", "042": "Audio Processing", "043": "Audio Processing", "045": "Audio Processing", "046": "Audio Processing", "047": "Audio Processing", "048": "Audio Processing", "049": "Audio Processing", "080": "Tones and Alerts", "106": "Tones and Alerts", "107": "Tones and Alerts", "109": "Audio Processing", "125": "Audio Processing", "126": "Audio Processing", "127": "Tones and Alerts", "128": "Tones and Alerts", "160": "Audio Processing", "168": "Audio Processing", "222": "Audio Processing", "279": "Tones and Alerts", "280": "Tones and Alerts", "281": "Tones and Alerts", "282": "Tones and Alerts", "283": "Tones and Alerts", "284": "Tones and Alerts", "285": "Tones and Alerts", "292": "Tones and Alerts", "293": "Tones and Alerts", "294": "Tones and Alerts", "304": "Tones and Alerts", "305": "Tones and Alerts"}
B5_RULED = {
    "SWE1_AMM_021": ["4865986"], "SWE1_AMM_023": ["4865986"],
    "SWE1_AMM_040": ["4865982"], "SWE1_AMM_292": ["4866173"],
    "SWE1_AMM_294": ["4866173"], "SWE1_AMM_025": ["4866311"],
    "SWE1_AMM_168": ["4866594"], "SWE1_AMM_281": ["4865984"],
    "SWE1_AMM_304": ["4866200"], "SWE1_AMM_305": ["4866201"],
    "SWE1_AMM_293": ["4866193"],
}

# B4 runs under the R-AM20 green channel, so its leaf set is assembled from
# package 13's A and B tables plus route 2's own resolutions rather than from
# a single ruled table. Held back: 002 and 122, where the two routes
# disagree — R-AM15 bars a single route from settling an anchor, and the
# green channel does not lift that.
# Package 16 ruled both, adopting route 2's proposals, so nothing is held.
B4_HELD: set[str] = set()
B4_ROUTE2 = {
    # Route 2's own resolutions, ruled at package 16 section 2.
    "SWE1_AMM_145": ["4866497"],
    "SWE1_AMM_155": ["4866513"],
    "SWE1_AMM_002": ["4865913"],
    "SWE1_AMM_122": ["4866444"],
    # Package 16 section 3 overturned all three of route 2's PENDINGs. Each
    # failure had a findable cause: 020's rejection rested on a truncated
    # read, 024's anchor is named in the leaf's own description and its
    # export row is an embedded table with no extractable text, and 146 was
    # searched for "remaining channel" where the spec writes "remaining
    # audio channels".
    "SWE1_AMM_020": ["4865981", "4866286"],
    "SWE1_AMM_024": ["4866001"],
    "SWE1_AMM_146": ["4866498"],
}
B4_SET = {  # test set per leaf, from package 13's section headings
    **{k: "Volume Control" for k in
       ("194", "196", "197", "220", "262", "272", "273", "274",
        "264", "266", "306", "307", "308", "024", "146")},
    **{k: "Audio Sources" for k in
       ("001", "005", "006", "007", "010", "108", "148", "149", "151", "152",
        "162", "163", "164", "165", "175", "176", "210", "214", "217", "256",
        "257", "311", "003", "009", "013", "202", "204", "207", "228", "263",
        "020", "145", "155", "002", "122")},
}

# Package 03 section 5. Each leaf carries the axes it belongs to so the
# generator can differentiate tc_title tokens within a family.
SIBLING_AXES = {
    "Ent->Ent transition": ["SWE1_AMM_205", "SWE1_AMM_206", "SWE1_AMM_208",
                            "SWE1_AMM_209", "SWE1_AMM_212"],
    "Ent->Info transition": ["SWE1_AMM_156", "SWE1_AMM_224"],
    "Info1->Info2 transition": ["SWE1_AMM_157", "SWE1_AMM_241"],
    "activation/deactivation sequence": [
        "SWE1_AMM_132", "SWE1_AMM_133", "SWE1_AMM_134", "SWE1_AMM_135",
        "SWE1_AMM_136", "SWE1_AMM_137", "SWE1_AMM_138",
        "SWE1_AMM_142", "SWE1_AMM_143", "SWE1_AMM_144"],
    "SOS mute/restore near-duplicate pair": [
        "SWE1_AMM_198", "SWE1_AMM_199", "SWE1_AMM_218", "SWE1_AMM_219"],
    "queue determination, same text different anchor": [
        "SWE1_AMM_130", "SWE1_AMM_139"],
    "boundary value candidates (25ms/50ms)": [
        "SWE1_AMM_275", "SWE1_AMM_276", "SWE1_AMM_277", "SWE1_AMM_278"],
    # Package 08 section 3. The amplifier-type pairs carry the same clause
    # text at two anchors, so the distinguishing token has to name the
    # amplifier; 234/235 pair across batches with B1's 226/227.
    "1.3.3.17 non-amplifier / booster": [
        "SWE1_AMM_229", "SWE1_AMM_230", "SWE1_AMM_231", "SWE1_AMM_232",
        "SWE1_AMM_234", "SWE1_AMM_235"],
    "1.3.3.18 CAN amplifier": [
        "SWE1_AMM_236", "SWE1_AMM_237", "SWE1_AMM_238", "SWE1_AMM_239"],
    "navigation fade-out chain": [
        "SWE1_AMM_312", "SWE1_AMM_313", "SWE1_AMM_314", "SWE1_AMM_315",
        "SWE1_AMM_316", "SWE1_AMM_317"],
    "ENTMuted indication, same text different anchor": [
        "SWE1_AMM_058", "SWE1_AMM_066"],
    "VSIM mute group": ["SWE1_AMM_179", "SWE1_AMM_180", "SWE1_AMM_181",
                        "SWE1_AMM_182"],
    "VSIM unmute group": ["SWE1_AMM_184"],
    "source selection shared anchor (R-AM16)": [
        "SWE1_AMM_031", "SWE1_AMM_032"],
    "TLM mute without ICS node": ["SWE1_AMM_068", "SWE1_AMM_069",
                                  "SWE1_AMM_078", "SWE1_AMM_079"],
    # Package 12 section 4.
    "VSIM mute group (B3 remainder)": ["SWE1_AMM_183"],
    "VSIM unmute group (B3)": ["SWE1_AMM_185", "SWE1_AMM_186",
                               "SWE1_AMM_187"],
    "Fade/Balance HMI update, same text different anchor": [
        "SWE1_AMM_114", "SWE1_AMM_119"],
    "market enumeration pair, must ship together": [
        "SWE1_AMM_081", "SWE1_AMM_082"],
    "SCV setup chain": ["SWE1_AMM_088", "SWE1_AMM_089", "SWE1_AMM_090",
                        "SWE1_AMM_091"],
    "reverse mute chain": ["SWE1_AMM_288", "SWE1_AMM_289", "SWE1_AMM_290",
                           "SWE1_AMM_291"],
    "information volume store and recall": ["SWE1_AMM_147", "SWE1_AMM_158"],
}

ROW_RE = re.compile(
    r"^\|\s*(SWE1_AMM_\d+)\s*\|\s*(SYS-RA-AMM-\d+)\s*\|\s*([^|]+?)\s*\|"
    r"\s*([^|]+?)\s*\|\s*CFTS019-(\d+)\s*\|\s*([^|]+?)\s*\|", re.M)
# Package 08 tabulates leaf, anchor, pool and coverage, under a heading per
# test set. A leaf may carry more than one anchor (SWE1_AMM_032 under R-AM16),
# so the anchor cell is parsed as a list rather than a single id.
# The anchor cell may hold one id, two joined by a return mark (R-AM16 shared
# anchors and dual citations), or an em dash where the leaf has no anchor at
# all and ships as PENDING. The pool cell likewise carries one mark or two.
ROW_RE_B2 = re.compile(
    r"^\|\s*(SWE1_AMM_\d+)\s*\|\s*([^|]*?)\s*\|\s*([^|]*?)\s*\|"
    r"\s*([^|]+?)\s*\|", re.M)
# Headings differ per package: "### Audio Arbitration（13 葉）" in 08,
# "### Mute Requests 後 13" in 12. Take the leading words either way.
SET_RE = re.compile(r"^### ([A-Za-z][A-Za-z /]+?)\s*(?:（|後|前|$)", re.M)


def parse_handoff(batch: str) -> list[dict]:
    """Read the batch's leaf/anchor table out of its handoff package."""
    text = HANDOFF[batch].read_text(encoding="utf-8")
    if batch == "B1":
        return [{"swe_id": m.group(1), "source_id": m.group(2),
                 "title": m.group(3), "test_set": m.group(4),
                 "anchors": [m.group(5)], "coverage": "完整",
                 "anchor_note": m.group(6)}
                for m in ROW_RE.finditer(text)]

    if batch == "B7":
        # 221 is held: package 26 anchors it at 4866489, but route 2 finds
        # 4866893 matching word for word and sitting between its neighbours'
        # anchors, where 4866489 fails both tests. R-AM15 bars one route from
        # settling an anchor, so it returns rather than shipping wrong.
        return [{"swe_id": f"SWE1_AMM_{k}", "anchors": [v],
                 "anchor_in_pool_ruled": None, "coverage": "完整",
                 "test_set": B7_SET[k]}
                for k, v in sorted(B7_ANCHORS.items())
                if k != "221"]

    if batch == "B6":
        return [{"swe_id": f"SWE1_AMM_{k}", "anchors": [v] if v else [],
                 "anchor_in_pool_ruled": None,
                 "coverage": "完整" if v else "PENDING: DR-AM10",
                 "test_set": B6_SET[k]}
                for k, v in sorted(B6_ANCHORS.items())]

    if batch in ("B4", "B5"):
        held = B4_HELD if batch == "B4" else B5_HELD
        ruled = B4_ROUTE2 if batch == "B4" else B5_RULED
        a = text[text.index("## 一、A 級"):text.index("## 二、B 級")]
        b = text[text.index("## 二、B 級"):text.index("## 三、C 級")]
        pat = r"\|\s*(SWE1_AMM_\d+)\s*\|\s*(?:CFTS019-)?(48\d{5})\s*\|\s*([✓✗])"
        rows = {m.group(1): ([m.group(2)], m.group(3) == "✓")
                for m in re.finditer(pat, a + b)}
        # Package 18 writes its B table as a prose shorthand, 034->4866058,
        # rather than as table rows; take both forms.
        rows.update({f"SWE1_AMM_{n}": ([o], None)
                     for n, o in re.findall(r"\b(\d{3})→(48\d{5})\b", b)})
        rows.update({k: (v, bool(v)) for k, v in ruled.items()})
        sets = B4_SET if batch == "B4" else B5_SET
        return [{"swe_id": k, "anchors": v[0], "anchor_in_pool_ruled": v[1],
                 "coverage": "完整" if v[0] else "PENDING: DR-AM1",
                 "test_set": sets.get(k.split("_")[-1], "")}
                for k, v in sorted(rows.items()) if k not in held]

    head = "## 一、定案錨表" if "## 一、定案錨表" in text else "## 二、定案錨表"
    tail = ("## 二、逐案處置" if "## 二、逐案處置" in text
            else "## 三、逐案裁定明細")
    section = text[text.index(head):text.index(tail)]
    # Walk the section so each row picks up the test set heading above it.
    bounds = [(m.start(), m.group(1)) for m in SET_RE.finditer(section)]
    out = []
    for m in ROW_RE_B2.finditer(section):
        test_set = next((name for pos, name in reversed(bounds)
                         if pos < m.start()), "")
        out.append({
            "swe_id": m.group(1),
            "anchors": re.findall(r"CFTS019-(\d+)", m.group(2)),
            "anchor_in_pool_ruled": "✓" in m.group(3),
            "coverage": " ".join(m.group(4).split()),
            "test_set": test_set.strip()})
    return out


def swe_rows(cfg: dict) -> dict[str, list[dict]]:
    """Every SWE.1 row across all four sheets, keyed by SWE ID.

    A list per key: SWE1_AMM_076 legitimately has two rows (R-AM6).
    """
    cols = {"swe_id": 0, "source_id": 1, "title": 2, "description": 3,
            "status": 4, "categorization": 6, "sub_categorization": 7,
            "priority": 16, "verification_criteria": 17,
            "verification_method": 18}
    wb = openpyxl.load_workbook(resolve_path(cfg, "a03_report"), read_only=True)
    out: dict[str, list[dict]] = {}
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            rec = {k: ("" if row[i] is None else str(row[i]).strip())
                   for k, i in cols.items()}
            rec["sheet"] = sheet.replace("Audio-Management ", "")
            out.setdefault(rec["swe_id"], []).append(rec)
    return out


def spec_blocks(cfg: dict) -> dict[str, str]:
    """ObjectID -> its clause text from the CFTS019 full-text PDF."""
    pdf = resolve_path(cfg, "spec_pdf")
    cache = ROOT / "data" / "cfts019_text.txt"
    if not cache.exists():
        subprocess.run(["pdftotext", str(pdf), str(cache)], check=True)
    text = cache.read_text(encoding="utf-8")
    return {m.group(1): m.group(2).strip() for m in re.finditer(
        r"^(48\d{5}): (\[.*?)(?=^\d{7}: \[|\Z)", text, re.M | re.S)}


def anchor_pool(cfg: dict) -> set[str]:
    """The R-AM2 anchor pool: ObjectIDs the two Basic Reports export.

    Expanded pool v2, 891 ids. Two things this gets wrong if done naively:

    - A-AM12: 58 cells hold more than one ObjectID. Testing each cell with a
      fullmatch sees only the lone ones and drops 86 ids, which is what
      produced the false out-of-pool findings across B1-B4.
    - Membership is read from the ObjectID column alone. Scanning every cell
      picks up prose references such as {CFTS019-5129} in comment columns and
      counts 897 — but a mention is not an exported object.
    """
    pool: set[str] = set()
    for key in ("sys1_export", "sys1_export_part2"):
        wb = openpyxl.load_workbook(resolve_path(cfg, key), read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        oid_i = max(range(len(rows[0])), key=lambda i: sum(
            1 for r in rows[1:]
            if re.fullmatch(r"\s*48\d{5}\s*", str(r[i] or ""))))
        for row in rows[1:]:
            pool.update(re.findall(r"\b(48\d{5})\b", str(row[oid_i] or "")))
    return pool


def axes_for(swe_id: str) -> list[str]:
    return [name for name, ids in SIBLING_AXES.items() if swe_id in ids]


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--batch", default="B1")
    args = ap.parse_args()

    cfg = load_feature_config(ROOT)
    leaves = parse_handoff(args.batch)
    swe = swe_rows(cfg)
    blocks = spec_blocks(cfg)
    pool = anchor_pool(cfg)

    out = []
    for leaf in leaves:
        rows = swe.get(leaf["swe_id"], [])
        # Match on Source Requirement ID: it disambiguates the 076 collision.
        want = leaf.get("source_id")
        if want:
            row = next((r for r in rows if r["source_id"] == want), rows[0])
        elif len(rows) > 1:
            # SWE1_AMM_076 collides (R-AM6). Package 06 states this batch
            # carries 076b = SYS-RA-AMM-246, the Part 02 row.
            row = next((r for r in rows if r["source_id"] == "SYS-RA-AMM-246"),
                       rows[-1])
        else:
            row = rows[0] if rows else None
        if row is None:
            raise SystemExit(f"{leaf['swe_id']} not found in SWE.1")
        anchors = leaf["anchors"]
        # A leaf ruled to ship without an anchor (PENDING: DR-AM1) still gets
        # a row — the completeness invariant runs both ways — so it is carried
        # here with an empty anchor list rather than dropped.
        out.append({**leaf,
                    "source_id": leaf.get("source_id") or row["source_id"],
                    "title": leaf.get("title") or row["title"],
                    "anchor": anchors[0] if anchors else "",
                    "anchor_in_pool": bool(anchors) and all(a in pool
                                                            for a in anchors),
                    "swe": row,
                    "spec_text": "\n\n".join(
                        f"[{a}] {blocks.get(a, '')}" for a in anchors),
                    "sibling_axes": axes_for(leaf["swe_id"])})

    dest = ROOT / "batches" / f"{args.batch}_context.json"
    dest.parent.mkdir(exist_ok=True)
    dest.write_text(json.dumps(out, ensure_ascii=False, indent=2),
                    encoding="utf-8")

    held = [l["swe_id"] for l in out if not l["anchor_in_pool"]]
    no_text = [l["swe_id"] for l in out if not l["spec_text"]]
    print(f"wrote {dest.relative_to(ROOT.parent.parent)}  ({len(out)} leaves)")
    print(f"  anchor in R-AM2 pool   {len(out) - len(held)}/{len(out)}")
    print(f"  out-of-pool, carried per R-AM2. {held}")
    print(f"  no spec text found     {no_text or 'none'}")
    print(f"  test sets              "
          + ", ".join(sorted({l['test_set'] for l in out})))


if __name__ == "__main__":
    main()
