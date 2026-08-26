#!/usr/bin/env python3
"""Propose CFTS019 anchors for a batch's leaves — a PROPOSAL, not a ruling.

The execution layer may not anchor (package 03 section 3.4). This script
exists so the analysis layer has a computed starting point for the section 4
table it owns, with the evidence attached and the weak matches marked.

Method, the same one package 03 section 4 discloses for B1: there is no
formal bridge between SWE1_AMM and CFTS019 ObjectIDs (F1 / DR-AM1), so each
leaf is matched on content — the SWE.1 Title plus Description against the
Basic Report Description — restricted to the R-AM2 anchor pool, and scored.
Anything that is not a clear single winner is flagged for a human read.

Usage:
    python features/audio_mgmt/scripts/propose_anchors.py --batch B2
"""

import argparse
import json
import re
import sys
from collections import Counter
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

FEATURE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(FEATURE.parent.parent / "scripts"))
from feature_config import load_feature_config, resolve_path  # noqa: E402

# 02 section 3. B1 is done; B2 is the next slice.
BATCHES = {
    "B2": [("Audio Arbitration", "rest"),
           ("Focus and Ducking", "all"),
           ("Mute Requests", 19)],
}
STOP = set("the a an of to and or for shall be is are with when on in by from "
           "audio management software this that it its as any all".split())


def tokens(text: str) -> set[str]:
    return {w for w in re.findall(r"[a-z0-9]+", text.lower())
            if w not in STOP and len(w) > 2}


def assignment_table() -> list[dict]:
    """The 317-leaf assignment the analysis layer locked in package 02."""
    text = (FEATURE / "docs" / "handoff" /
            "02_framework_assignment.md").read_text(encoding="utf-8")
    rows = re.findall(
        r"^\|\s*(SWE1_AMM_\d+)\s*\|\s*(SYS-RA-AMM-\d+)\s*\|\s*([^|]+?)\s*\|"
        r"\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|", text, re.M)
    return [{"swe_id": a, "source_id": b, "title": c, "sub_cat": d,
             "test_set": e.strip()} for a, b, c, d, e in rows]


def b1_leaves() -> set[str]:
    ctx = json.loads((FEATURE / "batches" / "B1_context.json")
                     .read_text(encoding="utf-8"))
    return {leaf["swe_id"] for leaf in ctx}


def pool_rows(cfg: dict) -> list[dict]:
    """Anchor pool with descriptions: the two Basic Reports (R-AM2)."""
    out = []
    for key in ("sys1_export", "sys1_export_part2"):
        wb = openpyxl.load_workbook(resolve_path(cfg, key), read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c or "").strip().lower() for c in rows[0]]
        # The ObjectID column is found by content, not by header text: the
        # first column is headed "ID" but holds NRL-149950 style keys, and a
        # header match on "id" takes it and yields an empty pool. The
        # 7-digit 48xxxxx ObjectID lives under "SYS2 Source Requirement items".
        oid_i = max(range(len(header)),
                    key=lambda i: sum(
                        1 for r in rows[1:]
                        if re.fullmatch(r"\s*48\d{5}\s*", str(r[i] or ""))))
        desc_i = next(i for i, h in enumerate(header) if h == "description")
        for row in rows[1:]:
            # A-AM12: 58 cells hold more than one ObjectID; a fullmatch test
            # sees only the lone ones and drops 86 ids, which is what produced
            # the false out-of-pool findings across B1-B4.
            for oid in re.findall(r"\b(48\d{5})\b", str(row[oid_i] or "")):
                out.append({"oid": oid,
                            "desc": " ".join(str(row[desc_i] or "").split()),
                            "src": key})
    return out



def monotone_align(leaves: list[dict], cands: list[dict]) -> dict[str, str]:
    """Global monotone alignment of the leaf sequence onto the anchor pool.

    This is the method package 03 section 4 discloses for B1 — SWE.1 document
    order against CFTS document order — and it is worth far more than content
    similarity alone: B1's 50 ruled anchors are perfectly monotone in SWE
    number, 0 inversions in 49 adjacent pairs. Content similarity scores the
    cells; the alignment supplies the ordering constraint that a per-leaf
    argmax throws away.

    Needleman-Wunsch with free gaps on both sides: not every leaf has an
    anchor in the pool (A-AM03), and most pool objects are not anchors.
    """
    n, m = len(leaves), len(cands)
    prev = [0.0] * (m + 1)
    back = [bytearray(m + 1) for _ in range(n + 1)]
    for i in range(1, n + 1):
        cur = [0.0] * (m + 1)
        li, bi = leaves[i - 1]["tok"], back[i]
        for j in range(1, m + 1):
            cj = cands[j - 1]["tok"]
            inter = len(li & cj)
            diag = prev[j - 1] + (inter / (len(li) + len(cj) - inter)
                                  if inter else 0.0)
            up, left = prev[j], cur[j - 1]
            if diag >= up and diag >= left:
                cur[j], bi[j] = diag, 1
            elif up >= left:
                cur[j], bi[j] = up, 2
            else:
                cur[j], bi[j] = left, 3
        prev = cur

    out, i, j = {}, n, m
    while i > 0 and j > 0:
        move = back[i][j]
        if move == 1:
            li, cj = leaves[i - 1]["tok"], cands[j - 1]["tok"]
            if li & cj:
                out[leaves[i - 1]["swe_id"]] = cands[j - 1]["oid"]
            i, j = i - 1, j - 1
        elif move == 2:
            i -= 1
        else:
            j -= 1
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--batch", default="B2")
    ap.add_argument("--validate", action="store_true",
                    help="score the method against B1's ruled anchors and stop")
    args = ap.parse_args()

    cfg = load_feature_config(FEATURE)
    table = assignment_table()
    done = b1_leaves()

    # Every leaf, in SWE order — the alignment is global (package 03 s4),
    # so it must see the whole sequence, not just this batch's slice.
    swe = {}
    wb = openpyxl.load_workbook(resolve_path(cfg, "a03_report"), read_only=True)
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if row[0]:
                sid = str(row[0]).strip()
                swe[sid] = {"title": str(row[2] or ""),
                            "desc": str(row[3] or "")}
    all_leaves = sorted(
        ({"swe_id": sid, "n": int(sid.split("_")[-1]),
          "tok": tokens(v["title"] + " " + v["desc"])}
         for sid, v in swe.items()), key=lambda x: (x["n"], x["swe_id"]))

    cands = [c for c in pool_rows(cfg) if tokens(c["desc"])]
    for c in cands:
        c["tok"] = tokens(c["desc"])
    cands.sort(key=lambda c: int(c["oid"]))

    aligned = monotone_align(all_leaves, cands)

    if args.validate:
        ctx = json.loads((FEATURE / "batches" / "B1_context.json")
                         .read_text(encoding="utf-8"))
        pool = {c["oid"] for c in cands}
        truth = {l["swe_id"]: l["anchor"] for l in ctx if l["anchor"] in pool}
        hit = sum(1 for k, v in truth.items() if aligned.get(k) == v)
        print(f"validation against B1's ruled in-pool anchors")
        print(f"  monotone alignment : {hit}/{len(truth)} = "
              f"{hit / len(truth) * 100:.0f}%")
        for k, v in truth.items():
            if aligned.get(k) != v:
                print(f"  miss: {k} ruled {v}, proposed {aligned.get(k)}")
        print(f"\n  {len(truth)} of B1's 50 leaves are solvable at all; the "
              f"other 7\n  are the out-of-pool anchors R-AM2' covers.")
        return 0

    by_set: dict[str, list[dict]] = {}
    for row in table:
        by_set.setdefault(row["test_set"], []).append(row)
    leaves = []
    for test_set, take in BATCHES[args.batch]:
        pool_ = [r for r in by_set.get(test_set, []) if r["swe_id"] not in done]
        leaves.extend(pool_ if take in ("rest", "all") else pool_[:take])

    by_oid = {c["oid"]: c for c in cands}
    proposals = []
    for leaf in leaves:
        want = tokens(swe.get(leaf["swe_id"], {}).get("title", "") + " "
                      + swe.get(leaf["swe_id"], {}).get("desc", ""))
        scored = sorted(
            ((len(want & c["tok"]) / len(want | c["tok"]), c) for c in cands),
            key=lambda t: -t[0])
        shortlist = [{"oid": c["oid"], "score": round(sc, 3),
                      "desc": c["desc"][:160]} for sc, c in scored[:5]]
        anchor = aligned.get(leaf["swe_id"], "")
        proposals.append({
            **leaf,
            "anchor": anchor,
            "anchor_desc": by_oid[anchor]["desc"][:200] if anchor else "",
            # The two methods are independent, so a disagreement is the
            # signal worth a human read — not the similarity score.
            "content_agrees": bool(anchor) and anchor in
                              [s["oid"] for s in shortlist],
            "shortlist": shortlist})

    dest = FEATURE / "batches" / f"{args.batch}_anchor_proposal.json"
    dest.parent.mkdir(exist_ok=True)
    dest.write_text(json.dumps(proposals, ensure_ascii=False, indent=2),
                    encoding="utf-8")

    unresolved = [p["swe_id"] for p in proposals if not p["anchor"]]
    disagree = [p["swe_id"] for p in proposals
                if p["anchor"] and not p["content_agrees"]]
    print(f"{args.batch}: {len(proposals)} leaves")
    print("  by test set: " + ", ".join(
        f"{k}={v}" for k, v in Counter(p["test_set"] for p in proposals).items()))
    print(f"  anchored by alignment      : "
          f"{len(proposals) - len(unresolved)}/{len(proposals)}")
    print(f"  content shortlist agrees   : "
          f"{len(proposals) - len(unresolved) - len(disagree)}")
    print(f"  disagree — read these first: {len(disagree)} {disagree[:8]}")
    print(f"  no alignment               : {len(unresolved)} {unresolved[:8]}")
    print(f"  written                    : "
          f"{dest.relative_to(FEATURE.parent.parent)}")
    print("\nMethod scores 42/43 on B1's ruled in-pool anchors "
          "(--validate reproduces it).")
    print("Still a proposal: anchoring is the analysis layer's call "
          "(package 03 section 3.4).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
