#!/usr/bin/env python3
"""R-AM24 / Comfort 96 section 1: put the delivery rows in req_id order.

The ruling — row order follows the leaf id, ascending — has existed since
Comfort 96 and is declared in features/user_profiles/feature.yaml as
`row_order`. It was never lifted out of that one feature, so every other
write_back.py in this repo, audio_mgmt's included, appends in generation
order. This puts audio_mgmt right and declares the parameter so the next
port cannot drop it silently again.

Pei ruled option (b) on 2026-08-27: rows are sorted AND tc_id is renumbered,
so that column D and column F both ascend. That renumbers nearly every TC,
so this script emits the old -> new map for the documents that name them.

Row content is rebuilt from generated/B*.json, which is verified cell-for-
cell against the workbook first — a rebuild that starts from a source that
has drifted would silently overwrite the drift.

    python features/audio_mgmt/scripts/reorder_rows.py --write
"""

import argparse
import importlib.util
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl

FEATURE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(FEATURE.parent.parent))
sys.path.insert(0, str(FEATURE.parent.parent / "scripts"))
_spec = importlib.util.spec_from_file_location(
    "am_write_back", FEATURE / "scripts" / "write_back.py")
wbm = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(wbm)
from backend.xlsx_surgical import surgical_save  # noqa: E402
from feature_config import load_feature_config  # noqa: E402

BATCHES = ("B1", "B2", "B3", "B4", "B5", "B6", "B7")


def leaf_n(req_id: str) -> int:
    return int(re.search(r"(\d+)$", req_id).group(1))


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--write", action="store_true")
    args = ap.parse_args()

    cfg = load_feature_config(FEATURE)
    src = FEATURE / "generated" / "SWQT_AudioMgmt_B1-B7.xlsx"
    header_row = cfg["workbook"]["header_row"]
    fmt = cfg["write_back"]["tc_id_format"]

    tcs = []
    for b in BATCHES:
        data = json.loads((FEATURE / "generated" / f"{b}.json")
                          .read_text(encoding="utf-8"))
        for tc in data["tcs"]:
            tcs.append(dict(tc, _batch=b))

    wb = openpyxl.load_workbook(src)
    ws = wb[cfg["workbook"]["sheet"]]
    cols = wbm.resolve_columns(ws, header_row, cfg)

    # Gate: the JSON must be what is in the workbook today, or the rebuild
    # would quietly discard whatever the workbook has that the JSON lacks.
    drift = []
    for i, tc in enumerate(tcs):
        r = header_row + 1 + i
        for k in wbm.DELIVERY_KEYS:
            if (ws.cell(r, cols[k]).value or "") != (tc.get(k, "") or ""):
                drift.append((r, tc["req_id"], k))
    if drift:
        print(f"refusing: {len(drift)} cell(s) differ between the JSON and "
              f"the workbook; reconcile first")
        for d in drift[:10]:
            print("  ", d)
        return 1
    old_tc_id = {id(tc): str(ws.cell(header_row + 1 + i, cols["tc_id"]).value)
                 for i, tc in enumerate(tcs)}

    # Stable: leaves that carry several cases keep the order they were
    # authored in, which is the order the sibling distinction was reasoned in.
    ordered = sorted(tcs, key=lambda t: leaf_n(t["req_id"]))
    mapping = []
    for n, tc in enumerate(ordered, start=1):
        new = fmt.format(n=n)
        mapping.append((old_tc_id[id(tc)], new, tc["req_id"], tc["_batch"]))

    moved = sum(1 for o, n, _, _ in mapping if o != n)
    nums = [leaf_n(t["req_id"]) for t in ordered]
    print(f"rows            : {len(ordered)}")
    print(f"req_id ascending: {all(a <= b for a, b in zip(nums, nums[1:]))}")
    print(f"tc_id renumbered: {moved} of {len(mapping)} change")
    print("first 5:", [(o, n, q) for o, n, q, _ in mapping[:5]])
    print("last  5:", [(o, n, q) for o, n, q, _ in mapping[-5:]])

    dest = FEATURE / "data" / "tc_id_remap.tsv"
    if args.write:
        dest.write_text("old_tc_id\tnew_tc_id\treq_id\tbatch\n" + "".join(
            f"{o}\t{n}\t{q}\t{b}\n" for o, n, q, b in mapping),
            encoding="utf-8")
        print(f"remap written   : {dest.relative_to(FEATURE.parent.parent)}")
    if not args.write:
        print("\ndry run — nothing written. Pass --write to emit.")
        return 0

    wb_cfg = cfg["write_back"]
    for i, tc in enumerate(ordered):
        r = header_row + 1 + i
        for k in wbm.DELIVERY_KEYS:
            ws.cell(r, cols[k]).value = tc.get(k, "")
        ws.cell(r, cols["tc_id"]).value = fmt.format(n=i + 1)
        ws.cell(r, cols["tc_ref_id"]).value = wb_cfg["tc_ref_id_value"]
        ws.cell(r, cols["author"]).value = wb_cfg["author_value"]

    stage = src.with_name(src.stem + "_stage.xlsx")
    tmp = src.with_name(src.stem + "_tmp.xlsx")
    shutil.copy2(src, stage)
    report = surgical_save(wb, stage, tmp, verify=True)
    cf = wbm.check_conditional_formatting(stage, tmp)
    stage.unlink()
    tmp.replace(src)

    print(f"\nwrote    : {src.name}")
    print(f"  sha256 : {wbm.sha256_file(src)}")
    print(f"  members: {report['members']}, patched {report['members_patched']}")
    print(f"  dv counts unchanged (classic, x14): {report['dv_counts']}")
    print(f"  conditionalFormatting unchanged  : {cf}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
