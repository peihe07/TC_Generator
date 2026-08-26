#!/usr/bin/env python3
"""Route 2 of R-AM15 for B3, read against the Basic Report exports.

Package 10 ran route 1 against the CFTS019 full text, so route 2 takes the
other artefact: the two Basic Report exports that form the R-AM2 anchor
pool. Different corpus, different failure modes — the full text carries the
objects the exports drop (A-AM03), and the exports carry the SYS2 review
columns the full text does not.

Prints the leaf beside the candidate's exported Description so the
comparison is made by reading. No score here settles anything: R-AM15 bars
single-route algorithm output from being a basis.

Usage:
    python features/audio_mgmt/scripts/route2_b3.py --grade B
    python features/audio_mgmt/scripts/route2_b3.py --oid 4866826
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl

FEATURE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(FEATURE.parent.parent / "scripts"))
from feature_config import load_feature_config, resolve_path  # noqa: E402

HANDOFFS = {"B3": "10_B3_anchor_candidates.md",
            "B4": "13_B4_anchor_candidates.md"}


def pool() -> dict[str, dict]:
    """ObjectID -> exported row. The Basic Report's own words, not the PDF's."""
    cfg = load_feature_config(FEATURE)
    out = {}
    for key in ("sys1_export", "sys1_export_part2"):
        wb = openpyxl.load_workbook(resolve_path(cfg, key), read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c or "").strip().lower() for c in rows[0]]
        # The ObjectID column is found by content: column A is headed "ID"
        # but holds NRL-nnnnnn keys.
        oid_i = max(range(len(header)), key=lambda i: sum(
            1 for r in rows[1:]
            if re.fullmatch(r"\s*48\d{5}\s*", str(r[i] or ""))))
        desc_i = next(i for i, h in enumerate(header) if h == "description")
        cat_i = next((i for i, h in enumerate(header) if "category" in h
                      and "sub" not in h), None)
        for r in rows[1:]:
            # A-AM12: take every id in the cell, not only a lone one.
            for oid in re.findall(r"\b(48\d{5})\b", str(r[oid_i] or "")):
                out[oid] = {"desc": " ".join(str(r[desc_i] or "").split()),
                            "cat": str(r[cat_i] or "") if cat_i else "",
                            "src": key}
    return out


def swe_rows() -> dict[str, list[dict]]:
    cfg = load_feature_config(FEATURE)
    wb = openpyxl.load_workbook(resolve_path(cfg, "a03_report"), read_only=True)
    out: dict[str, list[dict]] = {}
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if row[0]:
                out.setdefault(str(row[0]).strip(), []).append(
                    {"source_id": str(row[1] or ""), "title": str(row[2] or ""),
                     "desc": " ".join(str(row[3] or "").split())})
    return out


def grades(batch: str) -> dict[str, list[tuple[str, str]]]:
    t = (FEATURE / "docs" / "handoff" / HANDOFFS[batch]).read_text(
        encoding="utf-8")
    a = t[t.index("## 一、A 級"):t.index("## 二、B 級")]
    b = t[t.index("## 二、B 級"):t.index("## 三、C 級")]
    # Package 10 writes the anchor as CFTS019-nnnnnnn, package 13 as the bare
    # id. Accept either rather than assuming one house style holds.
    pat = r"\|\s*(SWE1_AMM_\d+)\s*\|\s*(?:CFTS019-)?(48\d{5})"
    return {"A": re.findall(pat, a), "B": re.findall(pat, b)}


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--batch", default="B4", choices=["B3", "B4"])
    ap.add_argument("--grade", choices=["A", "B"])
    ap.add_argument("--leaf")
    ap.add_argument("--oid", help="print one pool row and stop")
    args = ap.parse_args()

    p = pool()
    if args.oid:
        row = p.get(args.oid)
        print(f"CFTS019-{args.oid}: "
              + (f"[{row['cat']}] {row['desc'][:600]}" if row
                 else "NOT IN THE R-AM2 POOL (export omits it)"))
        return 0

    swe = swe_rows()
    g = grades(args.batch)
    pairs = ([(args.leaf, o) for gr in g.values() for s, o in gr if s == args.leaf]
             if args.leaf else g[args.grade])
    for sid, oid in pairs:
        recs = swe.get(sid, [{}])
        rec = recs[0]
        row = p.get(oid)
        print("=" * 78)
        print(f"{sid}  candidate CFTS019-{oid}"
              + ("" if row else "   << NOT IN POOL >>"))
        print(f"  LEAF : {rec.get('title', '')}")
        print(f"         {rec.get('desc', '')[:260]}")
        if row:
            print(f"  POOL : [{row['cat']}] {row['desc'][:300]}")
    print("=" * 78)
    print(f"{len(pairs)} pairs read against the export, not the PDF.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
