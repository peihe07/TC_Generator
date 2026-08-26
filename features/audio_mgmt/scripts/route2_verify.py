#!/usr/bin/env python3
"""Route 2 of R-AM15: independent full-text corroboration of a candidate anchor.

R-AM15 requires two independent routes to agree before an anchor may be
written. Route 1 is the analysis layer's targeted search against the Basic
Report. Route 2 must not lean on it, so this reads the **CFTS019 full text**
instead — a different corpus, and the one that carries the objects the Basic
Report exports omit.

This tool does not decide. It lays the leaf text and the candidate's clause
text side by side, with the strongest full-text alternatives, so the
comparison is made by reading. Single-route algorithm output is explicitly
not a basis for a ruling (R-AM15), so no score here is a verdict.

Usage:
    python features/audio_mgmt/scripts/route2_verify.py --grade B
    python features/audio_mgmt/scripts/route2_verify.py --leaf SWE1_AMM_310
"""

import argparse
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

FEATURE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(FEATURE.parent.parent / "scripts"))
from feature_config import load_feature_config, resolve_path  # noqa: E402

HANDOFF = FEATURE / "docs" / "handoff" / "06_B2_anchor_candidates.md"
STOP = set("the a an of to and or for shall be is are with when on in by from "
           "audio management software this that it its as any all not if then "
           "hu becomes active".split())


def tokens(text: str) -> set[str]:
    return {w for w in re.findall(r"[a-z0-9]+", text.lower())
            if w not in STOP and len(w) > 2}


def full_text_blocks() -> dict[str, str]:
    """ObjectID -> clause text, from the full-text PDF."""
    cfg = load_feature_config(FEATURE)
    cache = FEATURE / "data" / "cfts019_text.txt"
    if not cache.exists():
        subprocess.run(["pdftotext", str(resolve_path(cfg, "spec_pdf")),
                        str(cache)], check=True)
    text = cache.read_text(encoding="utf-8")
    return {m.group(1): " ".join(m.group(2).split())
            for m in re.finditer(r"^(48\d{5}): (\[.*?)(?=^\d{7}: \[|\Z)",
                                 text, re.M | re.S)}


def candidates_from_handoff() -> dict[str, list[tuple[str, str]]]:
    """A/B/C grades and their proposed anchors, read from package 06."""
    text = HANDOFF.read_text(encoding="utf-8")
    out: dict[str, list[tuple[str, str]]] = {"A": [], "B": [], "C": []}
    sec_a = text.split("## 二、A 級")[1].split("## 三、B 級")[0]
    for m in re.finditer(r"\|\s*(SWE1_AMM_\d+)\s*\|\s*CFTS019-(\d+)\s*\|", sec_a):
        out["A"].append((m.group(1), m.group(2)))
    sec_b = text.split("## 三、B 級")[1].split("## 四、C 級")[0]
    for m in re.finditer(r"\|\s*(\d{3})\s*\|\s*(48\d{5})\s*", sec_b):
        out["B"].append((f"SWE1_AMM_{m.group(1)}", m.group(2)))
    sec_c = text.split("## 四、C 級")[1].split("## 五、")[0]
    for m in re.finditer(r"\|\s*(SWE1_AMM_\d+)\s*\|", sec_c):
        out["C"].append((m.group(1), ""))
    return out


def swe_rows() -> dict[str, dict]:
    cfg = load_feature_config(FEATURE)
    wb = openpyxl.load_workbook(resolve_path(cfg, "a03_report"), read_only=True)
    out = {}
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if row[0]:
                out.setdefault(str(row[0]).strip(), []).append(
                    {"source_id": str(row[1] or ""), "title": str(row[2] or ""),
                     "desc": " ".join(str(row[3] or "").split())})
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--grade", choices=["A", "B", "C"])
    ap.add_argument("--leaf")
    ap.add_argument("--top", type=int, default=3)
    args = ap.parse_args()

    blocks = full_text_blocks()
    swe = swe_rows()
    grades = candidates_from_handoff()

    if args.leaf:
        pairs = [(g, sid, oid) for g, rows in grades.items()
                 for sid, oid in rows if sid == args.leaf]
    else:
        pairs = [(args.grade, sid, oid) for sid, oid in grades[args.grade]]

    scored_blocks = [(oid, tokens(t)) for oid, t in blocks.items()]
    for grade, sid, oid in pairs:
        recs = swe.get(sid, [{}])
        rec = recs[-1] if sid == "SWE1_AMM_076" else recs[0]
        want = tokens(rec.get("title", "") + " " + rec.get("desc", ""))
        alts = sorted(
            ((len(want & tk) / len(want | tk) if tk else 0, o)
             for o, tk in scored_blocks), key=lambda t: -t[0])[:args.top]
        print("=" * 78)
        print(f"[{grade}] {sid}  proposed CFTS019-{oid or '(none)'}")
        print(f"  LEAF : {rec.get('title', '')}")
        print(f"         {rec.get('desc', '')[:300]}")
        if oid:
            body = re.sub(r"^\[.*?\]\s*", "", blocks.get(oid, "(not in full text)"))
            body = re.sub(r"\[(Artifact|State|ECU|Market|Model|Radio|EE)[^\]]*\]",
                          "", body).strip()
            print(f"  ANCHOR full text:")
            print(f"         {body[:340]}")
        print(f"  full-text alternatives: "
              + ", ".join(f"{o}({s:.2f})" for s, o in alts))
    print("=" * 78)
    print(f"{len(pairs)} leaf/anchor pairs. Read, do not tally: R-AM15 bars a "
          f"score from settling an anchor.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
