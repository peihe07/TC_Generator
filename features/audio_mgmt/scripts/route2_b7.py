#!/usr/bin/env python3
"""Route 2 for B7, the final batch (package 26 section 6).

Same three gates as route2_b5 — word forms, untruncated reading, self-citation
— plus a fourth this feature has been settling anchors with by hand:

  4. position: where the candidate sits relative to the anchors already
     delivered for the leaf's SYS-RA neighbours. 221 was held on exactly this
     test (A-AM16) and route2_b5 had no column for it, so the reasoning lived
     only in prose. A test that decides cases belongs in the tool.

The window is built from delivered anchors, not from candidates, so it cannot
be talked into agreeing with the batch under review.

This tool does not decide. R-AM15 bars a single route — and a score — from
settling an anchor; the output is laid out to be read.

Usage:
    python features/audio_mgmt/scripts/route2_b7.py
    python features/audio_mgmt/scripts/route2_b7.py --leaf SWE1_AMM_221
    python features/audio_mgmt/scripts/route2_b7.py --scan 4866888-4866896
"""

import argparse
import csv
import glob
import json
import re
import sys
from pathlib import Path

import openpyxl

FEATURE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(FEATURE.parent.parent / "scripts"))
from feature_config import load_feature_config, resolve_path  # noqa: E402

HANDOFF = FEATURE / "docs" / "handoff" / "26_B7_final_batch.md"
IMAGEY = re.compile(r"\(image:|WrapperResource|\.rtf\b", re.I)
ATTRS = re.compile(r"\[(Artifact|State|ECU|Market|Model|Radio|EE)[^\]]*\]")


def pool() -> dict[str, dict]:
    cfg = load_feature_config(FEATURE)
    out = {}
    for key in ("sys1_export", "sys1_export_part2"):
        wb = openpyxl.load_workbook(resolve_path(cfg, key), read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c or "").strip().lower() for c in rows[0]]
        oid_i = max(range(len(header)), key=lambda i: sum(
            1 for r in rows[1:]
            if re.fullmatch(r"\s*48\d{5}\s*", str(r[i] or ""))))
        desc_i = next(i for i, h in enumerate(header) if h == "description")
        for r in rows[1:]:
            for oid in re.findall(r"\b(48\d{5})\b", str(r[oid_i] or "")):
                out[oid] = {"desc": " ".join(str(r[desc_i] or "").split()),
                            "src": key}
    return out


def blocks() -> dict[str, str]:
    text = (FEATURE / "data" / "cfts019_text.txt").read_text(encoding="utf-8")
    return {m.group(1): " ".join(m.group(2).split()) for m in re.finditer(
        r"^(48\d{5}): (\[.*?)(?=^\d{7}: \[|\Z)", text, re.M | re.S)}


def body(text: str) -> str:
    return ATTRS.sub("", text).strip()


def swe_rows() -> dict[str, dict]:
    cfg = load_feature_config(FEATURE)
    wb = openpyxl.load_workbook(resolve_path(cfg, "a03_report"), read_only=True)
    out = {}
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if row[0]:
                out.setdefault(str(row[0]).strip(), {
                    "source_id": str(row[1] or ""), "title": str(row[2] or ""),
                    "desc": " ".join(str(row[3] or "").split())})
    return out


def sysra() -> dict[str, int]:
    """swe_id -> SYS-RA ordinal, from leaves.tsv."""
    out = {}
    with (FEATURE / "data" / "leaves.tsv").open(encoding="utf-8") as fh:
        for r in csv.DictReader(fh, delimiter="\t"):
            m = re.search(r"(\d+)\s*$", r["source_id"] or "")
            if m:
                out[r["swe_id"]] = int(m.group(1))
    return out


def delivered() -> dict[str, str]:
    """swe_id -> anchor, from every generated batch. Delivered only —
    candidates never enter the window (that is the point of gate 4)."""
    out = {}
    for f in sorted(glob.glob(str(FEATURE / "generated" / "B*.json"))):
        for t in json.loads(Path(f).read_text("utf-8"))["tcs"]:
            ref = t["spec_reference"]
            m = re.fullmatch(r"CFTS019-(48\d{5})", ref)
            if m:
                out[t["req_id"]] = m.group(1)
    return out


def window(sid: str, ra: dict[str, int], dl: dict[str, str]):
    """Nearest delivered neighbour below and above, by SYS-RA order."""
    if sid not in ra:
        return None, None
    mine = ra[sid]
    below = [(ra[s], s, dl[s]) for s in dl if s in ra and ra[s] < mine]
    above = [(ra[s], s, dl[s]) for s in dl if s in ra and ra[s] > mine]
    return (max(below) if below else None), (min(above) if above else None)


def anchors_from_handoff() -> list[tuple[str, str, str]]:
    """(swe_id, ObjectID, pool mark) from package 26 section 2's two tables."""
    t = HANDOFF.read_text(encoding="utf-8")
    return [(m.group(1), m.group(2), m.group(3)) for m in re.finditer(
        r"\|\s*(SWE1_AMM_\d+)\s*\|\s*CFTS019-(48\d{5})\s*\|\s*([✓✗])", t)]


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--leaf")
    ap.add_argument("--scan", help="read an ObjectID range, e.g. 4866888-4866896")
    ap.add_argument("--terms", nargs="*", help="gate 1 search, all terms required")
    args = ap.parse_args()

    p, blk, swe = pool(), blocks(), swe_rows()
    ra, dl = sysra(), delivered()

    if args.scan:
        lo, hi = (int(x) for x in args.scan.split("-"))
        for oid in sorted(blk):
            if lo <= int(oid) <= hi and body(blk[oid]):
                print(f"{oid} {'pool' if oid in p else 'OUT '} | {body(blk[oid])}")
        return 0

    if args.terms:
        def forms(w):
            w = w.lower()
            return {w, w + "s"} | ({w[:-1]} if w.endswith("s") else set())
        hits = [(o, o in p, body(t)) for o, t in blk.items()
                if all(any(f in body(t).lower() for f in forms(x))
                       for x in args.terms)]
        print(f"terms {args.terms}: {len(hits)} hits")
        for o, inp, t in sorted(hits):
            print(f"  {o} {'pool' if inp else 'OUT '} | {t[:400]}")
        return 0

    pairs = anchors_from_handoff()
    if args.leaf:
        pairs = [x for x in pairs if x[0] == args.leaf]
    flagged, untestable = [], []
    for sid, oid, mark in pairs:
        rec = swe.get(sid, {})
        row = p.get(oid)
        cited = re.findall(r"CFTS019-(\d+)", rec.get("desc", ""))
        lo, hi = window(sid, ra, dl)
        print("=" * 78)
        note = "" if row else "   << NOT IN POOL >>"
        print(f"{sid}  package 26 anchor CFTS019-{oid}  (handoff pool mark {mark})"
              f"{note}")
        print(f"  SYS-RA: {ra.get(sid, '?')}")
        print(f"  LEAF  : {rec.get('title', '')}")
        print(f"          {rec.get('desc', '')}")          # gate 2: never cut
        if cited:                                           # gate 3
            print(f"  SELF-CITES: {cited} <- the leaf names an ObjectID")
        if row:
            im = "  [image/wrapper — no text search can reach it]" if \
                IMAGEY.search(row["desc"]) else ""
            print(f"  POOL  : {row['desc']}{im}")           # gate 2: never cut
        if oid in blk:
            print(f"  TEXT  : {body(blk[oid])}")
        # gate 4 —— position
        lo_s = f"{lo[1]}@{lo[2]} (SYS-RA {lo[0]})" if lo else "(none below)"
        hi_s = f"{hi[1]}@{hi[2]} (SYS-RA {hi[0]})" if hi else "(none above)"
        # A window whose own bounds are inverted tests nothing — it must not
        # be reported as "outside". 297's neighbours run 4867712 then 4866142;
        # any candidate at all would "fail" there. Reporting an untestable
        # case as a finding is how a gate starts producing noise.
        inverted = bool(lo and hi) and int(lo[2]) >= int(hi[2])
        inside = bool(lo and hi) and not inverted and int(lo[2]) < int(oid) < int(hi[2])
        verdict = ("inside" if inside else
                   "**window INVERTED — untestable**" if inverted else
                   "**OUTSIDE**" if (lo and hi) else "window open")
        print(f"  WINDOW: {lo_s}  <  CFTS019-{oid}  <  {hi_s}   -> {verdict}")
        if inverted:
            untestable.append((sid, oid, lo[2], hi[2]))
        elif lo and hi and not inside:
            flagged.append((sid, oid, lo[2], hi[2]))
    print("=" * 78)
    print(f"{len(pairs)} pairs, read whole. Pool basis {len(p)} ids; "
          f"delivered anchors in window basis: {len(dl)}.")
    print(f"gate 4 — outside a valid window: {len(flagged)}"
          + ("" if flagged else " (none)"))
    for sid, oid, a, b in flagged:
        print(f"  {sid}: {oid} not in ({a}, {b})")
    print(f"gate 4 — untestable, window own bounds inverted: {len(untestable)}"
          + ("" if untestable else " (none)"))
    for sid, oid, a, b in untestable:
        print(f"  {sid}: neighbours run {a} then {b} — no candidate can pass")
    print("R-AM15: no single route, and no count above, settles an anchor.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
