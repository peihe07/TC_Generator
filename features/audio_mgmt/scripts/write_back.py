#!/usr/bin/env python3
"""Write a generated audio_mgmt batch back into the FW036 workbook.

Ported from `features/time_management/scripts/write_back.py`. Two things
carry over unchanged because they are what keep the delivered file valid:

1. **Never `openpyxl.save`.** openpyxl drops `xmlns:x14` / `xmlns:xm` on
   save, which destroys the data validation the controlled form carries.
   Cells are edited on an in-memory workbook purely so the diff can be
   computed; the file itself is written by `backend.xlsx_surgical.
   surgical_save`, which patches the sheet XML inside a byte-for-byte copy.
2. **Structure is verified, not assumed.** `verify_structure` compares the
   zip member set and the classic and x14 `dataValidation` counts before and
   after, and raises rather than warns.

`verify_structure` does not count `<conditionalFormatting>`, and it cannot
catch a loss there on its own: the target sheet is expected to differ, so a
dropped element inside it passes. `check_conditional_formatting` below adds
that count locally. It is deliberately not pushed into the shared module —
that module backs fifteen features and turning a new hard gate on for all of
them is not this batch's call to make (A-AM06).

Usage:
    python features/audio_mgmt/scripts/write_back.py            # dry run
    python features/audio_mgmt/scripts/write_back.py --write
"""

import argparse
import hashlib
import json
import re
import sys
import zipfile
from pathlib import Path

import openpyxl
import yaml

FEATURE = Path(__file__).resolve().parents[1]
REPO_ROOT = next(p for p in FEATURE.resolve().parents
                 if (p / "pyproject.toml").is_file())
sys.path.insert(0, str(REPO_ROOT))
import backend.xlsx_surgical as _surgical  # noqa: E402
from backend.xlsx_surgical import (  # noqa: E402
    StructureError, surgical_save, verify_structure)


def _sparse_diff_cells(original, mutated):
    """diff_cells over the cells that exist, not over the sheet's extent.

    A-AM18: the shared diff_cells walks 1..max_row x 1..max_column with
    `ws.cell(r, c)`, and on a writable worksheet that call *creates* the cell
    rather than reading it. This form's template reports max_row 1411 against
    377 rows of data, so each pass instantiated some 35,000 empty cells in
    both workbooks; the second load never finished inside ten minutes.

    openpyxl already stores cells sparsely in `ws._cells`, so taking the
    union of the two sheets' populated coordinates gives the same answer:
    a coordinate absent from both holds None on both sides and cannot be a
    change. Equivalence is not assumed — rebuilding the B1-B6 workbook with
    this in place reproduces the delivered file byte for byte.

    Filed as a recommendation against backend/xlsx_surgical.py, which fifteen
    features share; patched here rather than there for the same reason
    A-AM06 was.
    """
    changes = {}
    for name in mutated.sheetnames:
        if name not in original.sheetnames:
            raise StructureError(f"sheet {name!r} is new; surgical emit only "
                                 "patches sheets that exist in the source")
        old, new = original[name], mutated[name]
        coords = set(old._cells) | set(new._cells)
        sheet_changes = {}
        for coord in coords:
            a = old._cells[coord].value if coord in old._cells else None
            b = new._cells[coord].value if coord in new._cells else None
            if a != b:
                sheet_changes[coord] = b
        if sheet_changes:
            changes[name] = sheet_changes
    return changes


_surgical.diff_cells = _sparse_diff_cells

CF_RE = re.compile(r"<conditionalFormatting[ >]")
# The ten delivery keys plus reasoning; reasoning stays in the JSON and is
# never written to a delivery column.
DELIVERY_KEYS = ("req_id", "test_group", "test_set", "test_item",
                 "pre_conditions", "input_test_data", "test_procedure",
                 "expected_result", "spec_reference", "priority",
                 "design_method", "remarks")


class WriteBackError(RuntimeError):
    pass


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def cf_counts(path: Path) -> dict[str, int]:
    """Per-sheet `<conditionalFormatting>` count."""
    out = {}
    with zipfile.ZipFile(path) as z:
        for member in z.namelist():
            if member.startswith("xl/worksheets/sheet"):
                out[member] = len(CF_RE.findall(z.read(member).decode("utf-8")))
    return out


def check_conditional_formatting(src: Path, out: Path) -> dict[str, int]:
    before, after = cf_counts(src), cf_counts(out)
    bad = {m: (before[m], after.get(m)) for m in before
           if before[m] != after.get(m)}
    if bad:
        raise StructureError(
            f"conditionalFormatting counts changed (before, after): {bad}. "
            "The banding and the flagged-cell rules are part of the "
            "controlled form and a write-back may not drop them")
    return {m: c for m, c in sorted(after.items()) if c}


def norm(v) -> str:
    return " ".join(str(v).split()).lower() if v is not None else ""


def resolve_columns(ws, header_row: int, cfg: dict) -> dict[str, int]:
    """Resolve columns by header text; feature.yaml letters are only a prior.

    A workbook whose header has shifted must fail loudly here rather than
    write into whatever column the letter happens to name now.
    """
    wanted = {
        "req_id": "requirement or design id",
        "tc_id": "test case id",
        "test_group": "test group",
        "test_set": "test set",
        "test_item": "test item",
        "pre_conditions": "pre-conditions",
        "input_test_data": "input test data",
        "test_procedure": "test procedure",
        "expected_result": "expected result",
        "spec_reference": "specification reference",
        "tc_ref_id": "test case reference id",
        "priority": "test case priority",
        "design_method": "test case design methods",
        "author": "test case author",
        "remarks": "remarks",
    }
    header = {}
    for idx, cell in enumerate(next(ws.iter_rows(min_row=header_row,
                                                 max_row=header_row)), 1):
        text = norm(cell.value)
        if not text:
            continue
        # Headers are bilingual, English line first.
        header.setdefault(text.split("\n")[0].strip(), idx)

    cols, missing = {}, []
    for key, label in wanted.items():
        # Two pairs of look-alike columns sit next to each other in this form:
        # C "Requirement or Design ID (Polarion)" beside D "Requirement or
        # Design ID", and E "Test Case ID (TestRail)" beside F "Test Case ID".
        # A plain prefix match takes whichever comes first and silently writes
        # the trace into the wrong column, so a candidate whose label is
        # followed by a qualifier in parentheses loses to the bare one.
        cands = [(text, i) for text, i in header.items()
                 if text.startswith(label)]
        cands.sort(key=lambda ti: ti[0][len(label):].lstrip().startswith("("))
        hit = cands[0][1] if cands else None
        if hit is None:
            missing.append(f"{key} ({label!r})")
        else:
            cols[key] = hit
    if missing:
        raise WriteBackError("header text did not resolve: " + ", ".join(missing))

    declared = cfg["workbook"]["columns"]
    drift = {k: (declared[k], openpyxl.utils.get_column_letter(cols[k]))
             for k in declared if k in cols
             and declared[k] != openpyxl.utils.get_column_letter(cols[k])}
    if drift:
        raise WriteBackError(
            f"feature.yaml column letters disagree with the header: {drift}")
    return cols


def first_free_row(ws, cols: dict, header_row: int) -> int:
    row = header_row + 1
    while ws.cell(row=row, column=cols["req_id"]).value not in (None, ""):
        row += 1
    return row


def next_tc_seq(ws, cols: dict, header_row: int) -> int:
    """Continue the tc_id sequence rather than restarting it per batch.

    Deriving the next number from the rows already in the workbook is the
    only way that holds across batches: taking it from the batch's own
    offset gave B2 the same NR1L-AMM-001..066 that B1 had already used.
    """
    seq = 0
    row = header_row + 1
    while ws.cell(row=row, column=cols["req_id"]).value not in (None, ""):
        val = str(ws.cell(row=row, column=cols["tc_id"]).value or "")
        m = re.search(r"(\d+)\s*$", val)
        if m:
            seq = max(seq, int(m.group(1)))
        row += 1
    return seq + 1


def write_rows(ws, cols: dict, tcs: list[dict], cfg: dict,
               start_row: int, start_seq: int) -> dict:
    wb_cfg = cfg["write_back"]
    fmt = wb_cfg.get("tc_id_format")
    if not fmt or "{n" not in fmt:
        raise WriteBackError(
            "feature.yaml write_back.tc_id_format missing or without a {n} "
            "field — tc_id is a delivery value and is never improvised here")

    for offset, tc in enumerate(tcs):
        row = start_row + offset
        for key in DELIVERY_KEYS:
            ws.cell(row=row, column=cols[key]).value = tc.get(key, "")
        ws.cell(row=row, column=cols["tc_id"]).value = fmt.format(
            n=start_seq + offset)
        ws.cell(row=row, column=cols["tc_ref_id"]).value = wb_cfg["tc_ref_id_value"]
        ws.cell(row=row, column=cols["author"]).value = wb_cfg["author_value"]
    return {"rows": len(tcs), "first_row": start_row,
            "last_row": start_row + len(tcs) - 1, "tc_id_format": fmt,
            "first_tc_id": fmt.format(n=start_seq),
            "last_tc_id": fmt.format(n=start_seq + len(tcs) - 1)}


def check_written_back(out: Path, sheet: str, cols: dict, plan: dict,
                       tcs: list[dict]) -> None:
    """Traceability and completeness, read back off the delivered file."""
    ws = openpyxl.load_workbook(out, data_only=False)[sheet]
    written = {}
    for offset in range(plan["rows"]):
        row = plan["first_row"] + offset
        written[row] = {k: ws.cell(row=row, column=cols[k]).value
                        for k in ("req_id", "test_item", "spec_reference")}
    if len(written) != len(tcs):
        raise WriteBackError(f"expected {len(tcs)} rows, read {len(written)}")
    for (row, got), tc in zip(sorted(written.items()), tcs):
        if got["req_id"] != tc["req_id"]:
            raise WriteBackError(
                f"row {row}: req_id is {got['req_id']!r}, expected "
                f"{tc['req_id']!r} — traceability broken")
        if not got["spec_reference"]:
            raise WriteBackError(f"row {row}: spec_reference is empty")
    leaves_in = {tc["req_id"] for tc in tcs}
    leaves_out = {g["req_id"] for g in written.values()}
    if leaves_in != leaves_out:
        raise WriteBackError(
            f"leaf set changed: missing {sorted(leaves_in - leaves_out)}, "
            f"extra {sorted(leaves_out - leaves_in)}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--batch", default="B1")
    ap.add_argument("--write", action="store_true",
                    help="actually write; without it this is a dry run")
    ap.add_argument("--out", default=None)
    # A batch writes into the previous batch's output, not into a fresh copy
    # of the master: the delivered workbook is cumulative, and starting from
    # the master again would silently drop every row already written.
    ap.add_argument("--source", default=None,
                    help="workbook to append into (default: paths.workbook)")
    args = ap.parse_args()

    cfg = yaml.safe_load((FEATURE / "feature.yaml").read_text(encoding="utf-8"))
    sheet = cfg["workbook"]["sheet"]
    header_row = int(cfg["workbook"]["header_row"])
    src = (Path(args.source) if args.source
           else next(FEATURE.glob(cfg["paths"]["workbook"])))
    if not src.is_file():
        raise WriteBackError(f"source workbook not found: {src}")
    data = json.loads((FEATURE / "generated" / f"{args.batch}.json")
                      .read_text(encoding="utf-8"))
    tcs = data["tcs"]

    wb = openpyxl.load_workbook(src)
    ws = wb[sheet]
    cols = resolve_columns(ws, header_row, cfg)
    start = first_free_row(ws, cols, header_row)
    seq = next_tc_seq(ws, cols, header_row)
    plan = write_rows(ws, cols, tcs, cfg, start, seq)

    print(f"source     : {src.name}")
    print(f"  sha256   : {sha256_file(src)}")
    print(f"sheet      : {sheet!r}, header row {header_row}")
    print(f"columns    : resolved from header text, {len(cols)} fields")
    print(f"batch      : {args.batch}, {plan['rows']} TCs -> rows "
          f"{plan['first_row']}-{plan['last_row']}")
    print(f"tc_id      : {plan['first_tc_id']} .. {plan['last_tc_id']}")

    if not args.write:
        print("\ndry run — nothing written. Pass --write to emit.")
        return 0

    out = Path(args.out) if args.out else (
        FEATURE / "generated" / f"{src.stem}_{args.batch}.xlsx")
    report = surgical_save(wb, src, out, verify=True)
    cf = check_conditional_formatting(src, out)
    check_written_back(out, sheet, cols, plan, tcs)

    shown = out.resolve()
    shown = (shown.relative_to(REPO_ROOT)
             if shown.is_relative_to(REPO_ROOT) else shown)
    print(f"\nwrote      : {shown}")
    print(f"  sha256   : {sha256_file(out)}")
    print(f"  members  : {report['members']}, patched "
          f"{report['members_patched']}")
    print(f"  dv counts unchanged (classic, x14): {report['dv_counts']}")
    print(f"  conditionalFormatting unchanged  : {cf}")
    print("  traceability and completeness verified on the written file")
    return 0


if __name__ == "__main__":
    sys.exit(main())
