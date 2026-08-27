#!/usr/bin/env python3
"""R-AM23 / D-CLOSE-01: downgrade the DR-AM1 PENDING anchors to NA.

Pei ruled that SWE1_AMM_026 and SWE1_AMM_076a ship with `NA` in
spec_reference rather than a PENDING string, so that the workbook clears
the IN section 8.4.3 shipping gate. The TC bodies, the tc_ids and the row
order are untouched — this rewrites one cell per row and nothing else.

Only the two DR-AM1 rows are in scope. SWE1_AMM_140's PENDING is DR-AM10
and was not ruled on, so it is left alone and reported.

    python features/audio_mgmt/scripts/downgrade_pending.py --write
"""

import argparse
import json
import shutil
import sys
from pathlib import Path

import openpyxl

FEATURE = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(FEATURE.parent.parent))
sys.path.insert(0, str(FEATURE.parent.parent / "scripts"))
import write_back  # noqa: E402,F401  applies the A-AM18 sparse-diff patch
from backend.xlsx_surgical import surgical_save  # noqa: E402
from feature_config import load_feature_config  # noqa: E402

# Two passes, ruled separately, and the note differs because the reason
# differs. Pass 1 (026, 076a) is an anchor that does not exist; pass 2 (140)
# is an anchor that may well exist but cannot be identified while the leaf
# contradicts its own source. Both end at NA in the cell, so the reasoning
# has to carry the distinction the cell can no longer make.
PASSES = {
    "dr-am1": {
        "doc": "B3.json",
        "targets": ("SWE1_AMM_026", "SWE1_AMM_076"),
        "old": "PENDING: DR-AM1 no CFTS019 object found for this leaf",
        "note": ("No CFTS019 object was found for this leaf after all four "
                 "gates; the anchor is downgraded to NA under R-AM23 so the "
                 "workbook carries no PENDING value at the section 8.4.3 "
                 "gate. DR-AM1 remains open and the real anchor is to be "
                 "filled in when it is answered."),
    },
    "dr-am10": {
        "doc": "B6.json",
        "targets": ("SWE1_AMM_140",),
        "old": ("PENDING: DR-AM10 anchor withheld pending upstream "
                "confirmation of this leaf's source"),
        "note": ("Downgraded to NA under R-AM23 so the workbook carries no "
                 "PENDING value at the section 8.4.3 gate. NA here does not "
                 "mean the anchor does not exist: position fixes it at "
                 "4866489, and the reason it is not written is that 4866489 "
                 "says store where the leaf says restore at initialisation. "
                 "DR-AM10 stays open on that contradiction and the anchor is "
                 "to be filled in when upstream resolves it."),
    },
    # The three rows above also restate the anchor situation in remarks, and
    # after the downgrade that column contradicts the cell next to it: N says
    # NA while AH still says PENDING about the very same anchor. The DR is
    # still open, so the trace stays — what goes is the blocking prefix.
    # Only these three. The other 133 PENDING remarks are DR-AM4/5/6/9 notes
    # about DBC signal names and undefined values, a different class, and
    # rewriting them here would be a ruling nobody made.
    "remarks": {
        "doc": None,
        "col": "remarks",
        "targets": ("SWE1_AMM_026", "SWE1_AMM_076", "SWE1_AMM_140"),
        "old": None,
        "new": {
            "SWE1_AMM_026": ("DR-AM1 open: no CFTS019 object found for this "
                             "leaf after all four gates; spec_reference "
                             "downgraded to NA under R-AM23"),
            "SWE1_AMM_076": ("DR-AM1 open: no CFTS019 object found for this "
                             "leaf after all four gates; spec_reference "
                             "downgraded to NA under R-AM23"),
            "SWE1_AMM_140": ("DR-AM10 open: the leaf describes restoration at "
                             "initialisation while the object at its document "
                             "position (4866489) states storage; "
                             "spec_reference downgraded to NA under R-AM23, "
                             "which records that no anchor is written here, "
                             "not that none exists"),
        },
    },
}
NEW = "NA"


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--pass", dest="pass_", choices=sorted(PASSES),
                    required=True)
    args = ap.parse_args()
    spec = PASSES[args.pass_]
    TARGETS, OLD, NOTE = spec["targets"], spec["old"], spec.get("note")

    cfg = load_feature_config(FEATURE)
    sheet = cfg["workbook"]["sheet"]
    src = FEATURE / "generated" / "SWQT_AudioMgmt_B1-B7.xlsx"
    doc = (FEATURE / "generated" / spec["doc"]) if spec["doc"] else None

    if spec["doc"]:
        data = json.loads(doc.read_text(encoding="utf-8"))
        json_hits = [tc for tc in data["tcs"]
                     if tc["req_id"] in TARGETS and tc["spec_reference"] == OLD]
    else:
        data, json_hits = None, None

    wb = openpyxl.load_workbook(src)
    ws = wb[sheet]
    cols = write_back.resolve_columns(ws, cfg["workbook"]["header_row"], cfg)
    req_c, tc_c = cols["req_id"], cols["tc_id"]
    spec_c = cols[spec.get("col", "spec_reference")]

    hits, other = [], []
    for r in range(cfg["workbook"]["header_row"] + 1, ws.max_row + 1):
        val = ws.cell(r, spec_c).value
        if not val or "PENDING" not in str(val):
            continue
        leaf = str(ws.cell(r, req_c).value or "").strip()
        want_old = OLD if OLD else str(val).strip()
        if str(val).strip() == want_old and leaf in TARGETS:
            hits.append((r, leaf, str(ws.cell(r, tc_c).value)))
        else:
            other.append((r, leaf, str(ws.cell(r, tc_c).value), str(val)))

    if json_hits is not None:
        print(f"json rows to change     : {len(json_hits)} "
              f"{[t['req_id'] for t in json_hits]}")
    print(f"workbook rows to change : {len(hits)} {hits}")
    print(f"left as is (not ruled)  : {len(other)} row(s)")
    for r, leaf, tid, val in other[:3]:
        print(f"  r{r} {leaf} {tid} -> {val[:70]}")
    want = len(TARGETS)
    if len(hits) != want or (json_hits is not None and len(json_hits) != want):
        print(f"\nrefusing: expected exactly {want} row(s) in each of json "
              f"and workbook")
        return 1
    if not args.write:
        print("\ndry run — nothing written. Pass --write to emit.")
        return 0

    if json_hits is not None:
        for tc in json_hits:
            tc["spec_reference"] = NEW
            tc["reasoning"] = tc["reasoning"].rstrip() + " " + NOTE
        doc.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n",
                       encoding="utf-8")

    for r, leaf, _ in hits:
        ws.cell(r, spec_c).value = (spec["new"][leaf] if spec.get("new")
                                    else NEW)

    backup = src.with_name("SWQT_AudioMgmt_B1-B7_pre_R-AM23.xlsx")
    if not backup.exists():
        shutil.copy2(src, backup)
    # surgical_save patches a byte-for-byte copy of the file the workbook was
    # loaded from, so the source is the current delivery file, not the
    # pre-R-AM23 backup — a second pass patched against the backup would
    # silently undo the first.
    tmp = src.with_name(src.stem + "_tmp.xlsx")
    stage = src.with_name(src.stem + "_stage.xlsx")
    shutil.copy2(src, stage)
    report = surgical_save(wb, stage, tmp, verify=True)
    cf = write_back.check_conditional_formatting(stage, tmp)
    stage.unlink()
    tmp.replace(src)

    print(f"\nbackup   : {backup.name} (pre-R-AM23, kept)")
    print(f"wrote    : {src.name}")
    print(f"  sha256 : {write_back.sha256_file(src)}")
    print(f"  members: {report['members']}, patched {report['members_patched']}")
    print(f"  dv counts unchanged (classic, x14): {report['dv_counts']}")
    print(f"  conditionalFormatting unchanged  : {cf}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
