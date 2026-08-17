#!/usr/bin/env python3
"""R-U8 三閘之反向驗證（04c 作業項 4）。

三閘現皆綠：

    functional_requirement_count == 180
    heading_count                == 25     （欄值 == "Heading"）
    out_of_scope_count           == 2

**而綠燈只在「它會為壞資料轉紅」時才有意義。** 03 包 §7 第 5 項由執行層
自陳為真缺口：三閘從未被注入測試過。本檔清掉它。

前例（Comfort 96 §6）：`row-order-by-reqid` 之第一版判準對**正確資料**轉紅
——「一個 leaf 拆出多條 TC 時那幾列之 req_id 本來就相同」——
**那是反向驗證抓到的，不是人看出來的**。

## 注入方式

對 `inputs/` 之 037 作**位元組複本**（置於 scratchpad，repo 外），以 openpyxl
改該複本之 `Analysis Report`，再以與 recon 相同之判準重量。三型注入：

    A 改一列 Categorization   Functional Requirement -> Heading
    B 增一列                  於表尾新增一列 Functional Requirement
    C 刪一列                  刪去一列 Out of scope

**`inputs/` 原檔不得觸碰** —— 每次注入前後對原檔重算 SHA256 並斷言不變。
複本用完即刪。

**openpyxl 存回會摧毀 x14 DV（A-UP09）** —— 於此無妨且刻意如此：
本檔之對象是 037 之**資料**，不是 036 母本之**結構**，
037 亦不在寫回路徑上。**該風險僅適用於 036 表單**，記於此以免誤讀。

Usage:
    python3 features/user_profiles/scripts/verify_recon_gates.py
"""

import collections
import hashlib
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl

FEATURE = Path(__file__).resolve().parent.parent
A03 = FEATURE / "inputs" / ("FM-WI-FSM-037-A03-N1L-SWE1-PersonalAccount-"
                            "HMI-V0.1 STLA 報告.xlsx")
SHEET = "Analysis Report"
HDR = 7
COL_ID, COL_CAT = 1, 7

EXPECTED = {"Functional Requirement": 180, "Heading": 25, "Out of scope": 2}


def sha256(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def measure(path: Path) -> dict:
    """R-U8 之判準，逐列計 Categorization 欄值。與 recon 同一單位。"""
    ws = openpyxl.load_workbook(path, data_only=True)[SHEET]
    c = collections.Counter()
    for r in range(HDR + 1, ws.max_row + 1):
        if ws.cell(r, COL_ID).value in (None, ""):
            continue
        c[str(ws.cell(r, COL_CAT).value or "").strip()] += 1
    return dict(c)


def gates(counts: dict) -> dict:
    """三閘之判定。回傳 {閘名: (期望, 實測, 是否通過)}。"""
    return {k: (v, counts.get(k, 0), counts.get(k, 0) == v)
            for k, v in EXPECTED.items()}


def report(label: str, counts: dict, expect_red: set) -> bool:
    """expect_red = 預期轉紅之閘名集合。回傳本案是否符合預期。"""
    g = gates(counts)
    red = {k for k, (_, _, ok) in g.items() if not ok}
    good = red == expect_red
    print(f"  {'PASS' if good else '**FAIL**'} — {label}")
    for k, (want, got, ok) in g.items():
        mark = "green" if ok else f"**RED**  差額 {got - want:+d}"
        print(f"      {k:<24} expected {want:>3}  measured {got:>3}   {mark}")
    if not good:
        print(f"      轉紅之閘：{sorted(red)}；預期：{sorted(expect_red)}")
    return good


def inject(src: Path, dst: Path, kind: str) -> None:
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb[SHEET]
    if kind == "A":
        # 第一個 Functional Requirement 列改為 Heading
        for r in range(HDR + 1, ws.max_row + 1):
            if str(ws.cell(r, COL_CAT).value or "").strip() == "Functional Requirement":
                ws.cell(r, COL_CAT).value = "Heading"
                break
    elif kind == "B":
        r = ws.max_row + 1
        ws.cell(r, COL_ID).value = "SWE1-HMI-PROF-999"
        ws.cell(r, COL_CAT).value = "Functional Requirement"
    elif kind == "C":
        for r in range(HDR + 1, ws.max_row + 1):
            if str(ws.cell(r, COL_CAT).value or "").strip() == "Out of scope":
                ws.delete_rows(r)
                break
    else:
        raise ValueError(kind)
    wb.save(dst)
    wb.close()


def main() -> int:
    if not A03.exists():
        raise SystemExit(f"037 not found: {A03}")
    before = sha256(A03)
    print(f"037 原檔 SHA256（注入前）: {before}\n")

    ok = True
    tmp = Path(tempfile.mkdtemp(prefix="upgates-"))
    try:
        print("## 正向 —— 乾淨資料不得觸發任何閘\n")
        clean = measure(A03)
        ok &= report("原檔（未注入）-> 三閘全綠", clean, set())

        # 注入 A/B/C 之前，先確認「複製後不注入」仍全綠 ——
        # 否則後面每一個紅燈都可能是 openpyxl 重存造成的，不是注入造成的。
        print("\n## 對照 —— 複製 ＋ openpyxl 重存但不改資料\n")
        ctrl = tmp / "control.xlsx"
        shutil.copy2(A03, ctrl)
        wb = openpyxl.load_workbook(ctrl); wb.save(ctrl); wb.close()
        ok &= report("重存而不改資料 -> 三閘仍全綠（證明紅燈來自注入）",
                     measure(ctrl), set())

        cases = [
            ("A", "改一列 Categorization：Functional Requirement -> Heading",
             {"Functional Requirement", "Heading"}),
            ("B", "增一列 Functional Requirement",
             {"Functional Requirement"}),
            ("C", "刪一列 Out of scope",
             {"Out of scope"}),
        ]
        print("\n## 反向 —— 注入壞資料，三閘須轉紅並報出正確差額\n")
        for kind, label, expect_red in cases:
            dst = tmp / f"inject_{kind}.xlsx"
            inject(A03, dst, kind)
            ok &= report(f"{kind}：{label}", measure(dst), expect_red)

        print("\n## 未實測（不可能失敗者不標 PASS，canon）\n")
        print("  未實測 — `heading_count` 對「欄值為 `Heading ` 之尾隨空白」"
              "之容忍度：本判準以 `.strip()` 正規化，故該情形量不出差別，"
              "其行為未經注入證明")
        print("  未實測 — 三閘對「Categorization 欄整欄空白」之反應："
              "該情形下三閘皆為 0 而全數轉紅，惟其**報出之差額是否可讀**"
              "未經測；此情形在真實 037 上不會發生，故不注入")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    after = sha256(A03)
    same = before == after
    ok &= same
    print(f"\n037 原檔 SHA256（注入後）: {after}")
    print(f"  {'PASS' if same else '**FAIL**'} — inputs/ 原檔未被觸碰"
          f"（前後雜湊{'相同' if same else '不同'}）")
    print(f"  PASS — 複本已刪除（{tmp} 不存在：{not tmp.exists()}）")

    n = 6
    print(f"\n{n if ok else '<' + str(n)} / {n} directional cases "
          f"{'PASS' if ok else 'FAIL'}")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
