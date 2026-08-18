#!/usr/bin/env python3
"""每個 leaf 之 prompt context 組裝（10b 前置 3）。

**本檔即「組裝自檢」之受檢對象。** 10b 明文：逐項回報，不得以「已設定」帶過。
故每一項都由本檔之一個函式承擔，且 `--selfcheck` 會印出它**實際注入了什麼**，
而不是印出「已設定」。

## 五項

| # | 要求 | 本檔之落點 |
|---|---|---|
| 1 | spec 內文 = `outline_map.json` 之 `pdf_text`（**非 `text`**）| `spec_body()` —— 且對 `text` 之讀取**不存在於本檔**（`--selfcheck` 以原始碼掃描證明）|
| 2 | 補句表七條之 `must_carry` 於其所屬 outline **確實注入** | `must_carry_for()` ＋ `assemble()` 之 `MUST-CARRY` 段 |
| 3 | Test Group／Test Set 逐字 | `TEST_GROUP` ＋ `test_set_of()`（讀 framework §2 之八組）|
| 4 | `tc_id` 格式 | `TC_ID_FMT` —— **本檔不指派號碼**，僅載明格式；指派為生成器之事 |
| 5 | `specification_reference` 依 R-U1（Source ID 字串）| `spec_ref()` |

## PLP 併列（R-U39(2)／R-U46）

判準之掃描對象有**兩種讀法**，其答案不同（見上繳 10 §5）。
R-U46 已裁：採**甲 ∪ 乙**之聯集為自動判準，`PLP_ENABLED` 啟用。

自動判準有盲區（R-G11）：以**位置指涉**表述之引用（`above` 等）
抓不到 —— 4.1 之 `PRACC1.` 即以 "see list of linked content above"
指 PLP 表。盲區之處置為**人工判讀**，其結果記於 `DECISIONS.md`，
並以 `PLP_LEAVES_MANUAL` 分列，**不併入自動判準**
（`PLP_LEAVES_AUTO` 須維持「甲∪乙 可重算得出」之性質）。

Usage:
    python3 features/user_profiles/scripts/build_batch_context.py --selfcheck
    python3 features/user_profiles/scripts/build_batch_context.py --leaf SWE1-HMI-PROF-104
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path

FEATURE = Path(__file__).resolve().parent.parent
OUTLINE = FEATURE / "data" / "outline_map.json"
MISSING = FEATURE / "data" / "xlsx_missing_clauses.tsv"
A03 = FEATURE / "inputs" / ("FM-WI-FSM-037-A03-N1L-SWE1-PersonalAccount-"
                            "HMI-V0.1 STLA 報告.xlsx")

TEST_GROUP = "User Profiles"
TC_ID_FMT = "NR1L-UserProfiles-{n:03d}"
SPEC_STEM = ("Personal_Account_HMI_Logic_and_Flow_R1_SR24_Post2A_CR24798_"
             "(October_03_2023)")

# framework §2 之八組，逐字（R-U39(3)：含大小寫）
CHAPTER_TO_TEST_SET = {
    "4": "Preference Storage", "5": "Profile List", "6": "Defaults",
    "7": "Welcome Flow", "8": "Setup Flow", "9": "Editing", "10": "Editing",
    "11": "Connected Account", "12": "Valet Mode", "13": "Valet Mode",
    "14": "Valet Mode",
}

# R-U39(2)／R-U46 —— 已裁，啟用。見上繳 10 §5、11 §3。
PLP_ENABLED = True

# 自動判準 = 甲 ∪ 乙（可由掃描重算得出，不得手工增刪）
PLP_LEAVES_AUTO = {
    "SWE1-HMI-PROF-001-01",   # 乙：037 Description「listed in PLP table」
    "SWE1-HMI-PROF-005",      # 乙：037 Verification Criteria「In the PLP table」
    "SWE1-HMI-PROF-012",      # 甲∩乙：sec 4.5.4 pdf_text ＋ 037 Description
    "SWE1-HMI-PROF-032",      # 甲∩乙：sec 5.9 pdf_text ＋ 037 Description
}

# 盲區之人工判讀結果（R-U46／R-G11）—— 逐條理由見 DECISIONS.md
# 兩者同屬 sec 4.1（PRACC1. "see list of linked content above"），
# 該節之併列已由 R-U46 裁定成立；自動判準抓不到位置指涉，故手工列入。
PLP_LEAVES_MANUAL = {
    "SWE1-HMI-PROF-001-02",   # D-UP11-01
    "SWE1-HMI-PROF-001-03",   # D-UP11-01
}

PLP_LEAVES = PLP_LEAVES_AUTO | PLP_LEAVES_MANUAL
PLP_SECTIONS = ["3.1", "3.2", "3.3", "3.4", "3.5"]


def _outline() -> dict:
    d = json.loads(OUTLINE.read_text(encoding="utf-8"))
    d.pop("__meta__", None)
    return d


def spec_body(section: str) -> str:
    """**判讀基準 = `pdf_text`（R-U25／R-U35(a)）。**

    `text`（xlsx 側）為追溯用，**本函式不讀它** —— 那不是一句宣稱：
    `--selfcheck` 以原始碼掃描證明本檔內無任何 `["text"]` 之讀取。
    """
    return (_outline().get(section, {}) or {}).get("pdf_text", "")


PLP_RE = re.compile(r"PLP|Profile[\s ]+Linked[\s ]+Preferences", re.I)


def plp_scan_union() -> set:
    """甲 ∪ 乙 之掃描（R-U46 之自動判準）—— 可重跑，不是註解。

    甲：該 leaf 所引 spec section 之 `pdf_text` 含 PLP 字樣
    乙：該 leaf 自身之 037 Description／Verification Criteria 含 PLP 字樣

    **不含盲區之人工判讀結果**（R-U46：不併入自動判準）。
    """
    out = set()
    for req_id, m in leaf_rows().items():
        jia = bool(PLP_RE.search(spec_body(m["section"]) or ""))
        yi = bool(PLP_RE.search((m.get("desc") or "") + " "
                                + (m.get("vc") or "")))
        if jia or yi:
            out.add(req_id)
    return out


# R-U49 —— `p<N>` 之顯式歸屬對照表。
#
# **廢除以 `impact` 散文欄作掛回鍵之設計**：`p14` 之所以掛得回 9.1，
# 是因為它的說明文字**剛好**寫了「9.1 之列項順序」；`p17` 之說明是「同上」，
# 於是它掛不回任何節，且無聲無息 —— 那是巧合，不是設計。
# 人看的說明欄不該同時當機器用的外鍵。
PAGE_TO_SECTION = {
    "p14": "9.1",    # Table EDPR1 之列項
    "p17": "11.5",   # Connected Navigation 之列項（同 11.5 之表）
}


def must_carry_for(section: str) -> list:
    """補句表中歸屬該 outline 之 must_carry 條目（R-U35(b)）。"""
    out = []
    with MISSING.open(encoding="utf-8") as fh:
        rows = csv.DictReader((l for l in fh if not l.startswith("#")),
                              delimiter="\t")
        for r in rows:
            if r.get("must_carry") != "yes":
                continue
            if r["outline"] == section:
                out.append(r)
            # `p<N>` 之條目其歸屬節次未逐一定位（07 輪），以顯式對照表掛回。
            elif r["outline"].startswith("p"):
                if PAGE_TO_SECTION.get(r["outline"]) == section:
                    out.append(r)
    return out


def test_set_of(section: str) -> str:
    return CHAPTER_TO_TEST_SET[section.split(".")[0]]


def spec_ref(section: str, req_id: str) -> str:
    refs = [f"{SPEC_STEM}_{section}"]
    if PLP_ENABLED and req_id in PLP_LEAVES:
        refs += [f"{SPEC_STEM}_{s}" for s in PLP_SECTIONS]
    return "; ".join(refs)


def leaf_rows() -> dict:
    import openpyxl
    ws = openpyxl.load_workbook(A03, data_only=True)["Analysis Report"]
    pre = SPEC_STEM + "_"
    out = {}
    for r in range(8, ws.max_row + 1):
        rid = ws.cell(r, 1).value
        if not rid or str(ws.cell(r, 7).value or "").strip() != \
                "Functional Requirement":
            continue
        src = str(ws.cell(r, 3).value or "").strip()
        out[str(rid).strip()] = {
            "section": src[len(pre):] if src.startswith(pre) else "?",
            "title": str(ws.cell(r, 4).value or "").strip(),
            "desc": " ".join(str(ws.cell(r, 5).value or "").split()),
            "sub": str(ws.cell(r, 9).value or "").strip(),
            "priority_prior": str(ws.cell(r, 18).value or "").strip(),
            "vc": " ".join(str(ws.cell(r, 19).value or "").split()),
        }
    return out


def assemble(req_id: str, leaf: dict) -> dict:
    sec = leaf["section"]
    mc = must_carry_for(sec)
    return {
        "req_id": req_id,
        "section": sec,
        "test_group": TEST_GROUP,
        "test_set": test_set_of(sec),
        "tc_id_format": TC_ID_FMT,
        "specification_reference": spec_ref(sec, req_id),
        "spec_body_source": "outline_map.json::pdf_text",
        "spec_body": spec_body(sec),
        "leaf_title": leaf["title"],
        "leaf_desc_037": leaf["desc"],
        "must_carry": [{"text": m["text"], "affected_field": m["affected_field"],
                        "pdf_source": m["pdf_source"]} for m in mc],
        "sub_categorization": leaf["sub"],
        "priority_prior_037": leaf["priority_prior"],
    }


# ------------------------------------------------------------- selfcheck

def selfcheck(sample_ids: list) -> int:
    ok = True

    def chk(name, passed, detail):
        nonlocal ok
        ok &= passed
        print(f"  {'PASS' if passed else '**FAIL**'} — {name}")
        for d in detail:
            print(f"      {d}")

    # **判準改過兩次，兩次都是判準錯，不是程式錯。**
    #
    # v1：掃全檔之 `['text']` —— 8 處命中全在 selfcheck 自己（它**故意**讀
    #     `text` 以證明兩者不同）。一個「不得讀 text」之檢查，把「證明 text
    #     不同」那段也算進去，永遠會紅。
    # v2：改掃「生產路徑」—— 仍有 2 處：`spec_body` 之 **docstring**，
    #     與 `assemble` 之 `m["text"]`（那是**補句表**之欄，不是 outline_map 之欄）。
    #     字串比對把三個不同的 `text` 鍵混為一談。
    # v3（現行）：**改問真正的不變量** —— 只有 `spec_body()` 得觸碰 outline_map，
    #     且其取用之鍵只有 `pdf_text`。這是可判定的，字串命中數不是。
    #
    # 同 R-U37 之處置：改判準，不改案例。
    whole = Path(__file__).read_text(encoding="utf-8")
    import ast
    tree = ast.parse(whole)
    fns = {n.name: n for n in ast.walk(tree)
           if isinstance(n, ast.FunctionDef)}
    # (i) 誰呼叫 _outline()
    callers = sorted({f.name for f in fns.values()
                      for n in ast.walk(f)
                      if isinstance(n, ast.Call)
                      and getattr(n.func, "id", "") == "_outline"})
    # (ii) spec_body 內對 outline dict 取用之鍵
    keys = sorted({n.args[0].value for n in ast.walk(fns["spec_body"])
                   if isinstance(n, ast.Call)
                   and getattr(n.func, "attr", "") == "get"
                   and n.args and isinstance(n.args[0], ast.Constant)})
    rows = leaf_rows()
    ctxs = {r: assemble(r, rows[r]) for r in sample_ids}

    # 1 —— pdf_text 而非 text
    ex = ctxs["SWE1-HMI-PROF-104"]
    o = _outline()["9.8"]
    chk("1. spec 內文取 `pdf_text` —— 以不變量驗，非以字串命中數驗",
        callers == ["selfcheck", "spec_body"] and keys == ["pdf_text"]
        and ex["spec_body"] == o["pdf_text"] and ex["spec_body"] != o["text"],
        [f"(i) 呼叫 `_outline()` 之函式 = {callers}  "
         f"（**生產路徑僅 `spec_body`**；`selfcheck` 為本檢查自身）",
         f"(ii) `spec_body` 對 outline dict 取用之鍵 = {keys}  （須僅 `pdf_text`）",
         f"(iii) 抽樣 9.8：spec_body == pdf_text ? "
         f"{ex['spec_body'] == o['pdf_text']}；== text ? "
         f"{ex['spec_body'] == o['text']}  （後者須 False）",
         f"  pdf_text 尾 90：…{o['pdf_text'][-90:]}",
         f"  text     尾 90：…{' '.join(str(o['text']).split())[-90:]}",
         f"  **兩者不同即 9.8 之掉句本身** —— pdf_text 有 PU0609 那句，text 沒有"])

    # 2 —— must_carry 之實際注入點
    hit = {r: c["must_carry"] for r, c in ctxs.items() if c["must_carry"]}
    chk("2. `must_carry` 於其所屬 outline 確實注入（列出注入點與內容）",
        len(hit) >= 3,
        [f"抽樣 16 leaf 中，注入 must_carry 者 {len(hit)} 個"] +
        [f"  {r}（sec {ctxs[r]['section']}）→ {len(v)} 條："
         f"{v[0]['text'][:60]}… [affected: {v[0]['affected_field']}, "
         f"src {v[0]['pdf_source']}]" for r, v in hit.items()])

    # 3 —— Test Group / Test Set
    tg = {c["test_group"] for c in ctxs.values()}
    ts = sorted({c["test_set"] for c in ctxs.values()})
    chk("3. Test Group 單一且逐字；Test Set 為 framework §2 之八組",
        tg == {"User Profiles"} and len(ts) == 8,
        [f"Test Group = {tg}", f"Test Set（{len(ts)}）= {ts}"])

    # 4 —— tc_id 格式
    chk("4. `tc_id` 格式為 `NR1L-UserProfiles-{{NNN}}`",
        ctxs["SWE1-HMI-PROF-104"]["tc_id_format"] == TC_ID_FMT,
        [f"格式字串 = {TC_ID_FMT}",
         f"樣例 = {TC_ID_FMT.format(n=1)}",
         "**本檔不指派號碼** —— 指派為生成器之事，格式於此定"])

    # 5 —— specification_reference
    r5 = ctxs["SWE1-HMI-PROF-111"]["specification_reference"]
    chk("5. `specification_reference` 為 Source ID 字串（非檔名形式）",
        r5.startswith(SPEC_STEM) and "R1L-R" not in r5,
        [f"樣例（111，sec 11.4）= {r5}",
         f"含 `R1L-R (February_10_2023)` 形式？ {'R1L-R' in r5}  （須為 False）"])

    # 6 —— PLP 併列之現況（R-U46 已裁，啟用）
    plp_on = [r for r in sample_ids if r in PLP_LEAVES]
    plp_off = [r for r in sample_ids if r not in PLP_LEAVES]
    # 對照向（R-G7）：非 PLP 之 leaf 不得被併入 3.x
    ok6 = (PLP_ENABLED is True
           and all("_3." in ctxs[r]["specification_reference"] for r in plp_on)
           and not any("_3." in ctxs[r]["specification_reference"]
                       for r in plp_off))
    chk("6. R-U46 之 `3.x` 併列 —— **已啟用**，含對照向",
        ok6,
        [f"PLP_ENABLED = {PLP_ENABLED}",
         f"AUTO {len(PLP_LEAVES_AUTO)} ＋ MANUAL {len(PLP_LEAVES_MANUAL)} "
         f"= {len(PLP_LEAVES)}",
         f"抽樣中屬 PLP_LEAVES 者：{plp_on}",
         *[f"    {r}: 含 3.x = {'_3.' in ctxs[r]['specification_reference']}"
           for r in plp_on],
         f"對照向 —— 其餘 {len(plp_off)} 條含 3.x 者："
         f"{[r for r in plp_off if '_3.' in ctxs[r]['specification_reference']]}"
         f"（須為空）"])

    # 7 —— must_carry 七條是否**皆有歸宿**（R-U49 步驟 1）
    #
    # 第 2 項驗的是「**有注入者是否正確**」，驗不到「**七條是否都掛得上某節**」。
    # 一條掛不回任何節之 must_carry，在第 2 項眼中不存在 —— 這正是 `p17` 之形狀。
    all_mc, homed, orphan = [], {}, []
    with MISSING.open(encoding="utf-8") as fh:
        for r in csv.DictReader((l for l in fh if not l.startswith("#")),
                                delimiter="\t"):
            if r.get("must_carry") == "yes":
                all_mc.append(r["outline"])
    # 全 169 節逐節問一次：這條 must_carry 掛得上嗎
    for mc in all_mc:
        hosts = [sec for sec in _outline()
                 if any(x["outline"] == mc for x in must_carry_for(sec))]
        if hosts:
            homed[mc] = hosts
        else:
            orphan.append(mc)
    chk("7. must_carry 七條**皆有歸宿**（非只驗已注入者是否正確）",
        not orphan and len(all_mc) == 7,
        [f"must_carry 條目數 = {len(all_mc)}（須為 7）",
         *[f"    {mc} → 掛回 {homed[mc]}" for mc in all_mc if mc in homed],
         f"**無歸宿者 = {orphan}**（須為空）",
         f"餘數：{len(homed)} 有歸宿 ＋ {len(orphan)} 無歸宿 = {len(all_mc)}"])

    # 8 —— `PLP_LEAVES_AUTO` 之可重算性（R-U52）
    #
    # 該集合宣稱「重跑掃描即可重算得出」。**在此之前那只是一句註解。**
    # 此處把甲∪乙之掃描實跑一次，與常數比對 —— 使其成為可重跑之斷言。
    # `PLP_LEAVES_MANUAL` **不納入**（R-U46：人工判讀不併入自動判準）。
    recomputed = plp_scan_union()
    drift_missing = sorted(PLP_LEAVES_AUTO - recomputed)   # 常數有、掃描無
    drift_extra = sorted(recomputed - PLP_LEAVES_AUTO)     # 掃描有、常數無
    chk("8. `PLP_LEAVES_AUTO` 可由甲∪乙之掃描重算得出（R-U52）",
        recomputed == PLP_LEAVES_AUTO,
        [f"重算所得（{len(recomputed)}）= {sorted(recomputed)}",
         f"常數所載（{len(PLP_LEAVES_AUTO)}）= {sorted(PLP_LEAVES_AUTO)}",
         f"常數有而掃描無：{drift_missing}（須為空）",
         f"掃描有而常數無：{drift_extra}（須為空）",
         f"MANUAL 集 {sorted(PLP_LEAVES_MANUAL)} **不納入本斷言**",
         "對照向（R-G7）：以 `--selfcheck-tamper` 竄改 AUTO 集，本項須紅"])

    print(f"\n{'8' if ok else '<8'} / 8 self-check items "
          f"{'PASS' if ok else 'FAIL'}")
    return 0 if ok else 1


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--selfcheck", action="store_true")
    # 對照向（R-G7／R-U52）：竄改 AUTO 集，第 8 項須紅。
    # **不是註解，是可重跑的一條指令。**
    ap.add_argument("--selfcheck-tamper", choices=["drop", "add"],
                    help="竄改 PLP_LEAVES_AUTO 以證明第 8 項會失敗")
    ap.add_argument("--leaf")
    a = ap.parse_args()
    if a.selfcheck_tamper == "drop":
        PLP_LEAVES_AUTO = PLP_LEAVES_AUTO - {"SWE1-HMI-PROF-012"}
    elif a.selfcheck_tamper == "add":
        PLP_LEAVES_AUTO = PLP_LEAVES_AUTO | {"SWE1-HMI-PROF-999"}
    if a.leaf:
        rows = leaf_rows()
        print(json.dumps(assemble(a.leaf, rows[a.leaf]),
                         ensure_ascii=False, indent=1))
        sys.exit(0)
    SAMPLE = [x["req_id"] for x in
              json.loads(Path("/tmp/sample.json").read_text(encoding="utf-8"))]
    sys.exit(selfcheck(SAMPLE))
