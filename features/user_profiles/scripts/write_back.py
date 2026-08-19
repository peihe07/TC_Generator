#!/usr/bin/env python3
"""寫回工作簿（42 包作業 2）。**本輪不得產出交付件** —— 見 §未決 1 之閘。

## 三條硬約束（41 包 §二，本檔逐條落實）

1. **不呼叫 `diff_cells()`。** 走「直接給定 edits → `patch_sheet_xml` →
   逐 member 複寫」。實測：`diff_cells` 對本母本之 TC 分頁逾 100 秒未完成，
   封裝路徑 0.04 秒。**我方本就知道要寫哪些格 —— 求差異是在解一個我們沒有的問題。**
2. **母本永不被寫入。** 讀 `inputs/…_ext.xlsx`（SHA 6372fb6b…，
   `BASELINE.sha256` 保護），寫 `output/` 之新檔。
3. **寫回後必跑 `verify_dv_integrity`**，四項全綠，任一不綠即不留下檔案。

## 未決 1 之閘（42 包 §一）

T:Z 七個車型欄之填值**無依據**：填了是編造，不填則該七欄不在 DV 涵蓋內。
**故 `--write` 在 `--vehicle-columns` 未給定時拒絕執行**，
且該拒絕**不是可用旗標關掉的**（沒有 `--force`）。
`--self-test` 寫的是 scratchpad 之暫存檔，不落 `output/`，不入台帳 ——
**那是驗證本程式，不是產出交付件**。

## 參數化之四項（42 包 §一之處置）

| 參數 | 預設 | 依據 |
|---|---|---|
| `vehicle_columns` | `None` → **T:Z 不寫入** | 未決 1，待 Pei |
| `author_value` | `None` → **AA 不寫入** | Comfort 交付件實測 AA 為空（未決 3 之同類） |
| `tc_ref_id_value` | `None` → **O 不寫入** | 同上，Comfort 實測 O 為空 |
| `row_order` | `"req_id"` | Comfort 96 §1（Pei 裁定）：列序依 leaf id 遞增 |

**`row_order` 之預設與我 41 輪之設計草案不同**（草案寫 `tc_id` 序）——
理由與其代價見上繳 42 §2.3，**已具名待確認**。

Usage:
    python3 scripts/write_back.py                      # dry-run（預設）
    python3 scripts/write_back.py --self-test          # 方向性案例（scratchpad）
    python3 scripts/write_back.py --write --vehicle-columns '{"T":"1"}'
"""

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from backend.xlsx_surgical import (patch_sheet_xml,            # noqa: E402
                                   sheet_members, col_to_idx)
import verify_dv_integrity as DV                               # noqa: E402

FEATURE = Path(__file__).resolve().parent.parent
SRC = FEATURE / "inputs" / ("FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT "
                            "STLA Test Case Specification & Result_SWQT_"
                            "20260817_ext.xlsx")
SRC_SHA = "6372fb6be02f48dc3a3e091a60d2e2b3cf26d8704c27e25d79b7c9516fb825b2"
SHEET = "Test Case Specification 測試用例規範"
FIRST_ROW = 10
CAPACITY_LAST_ROW = 1411          # B 欄公式與各 DV 之實測上界

# feature.yaml §workbook.columns（rev C；本輪以 header 探測複驗）
COLS = {
    "D": "req_id", "F": "tc_id", "G": "test_group", "H": "test_set",
    "I": "test_item", "J": "pre_conditions", "K": "input_test_data",
    "L": "test_procedure", "M": "expected_result",
    "N": "specification_reference", "P": "priority", "R": "design_method",
    "S": "functional_safety", "AH": "remarks",
}
# 一律不寫。**逐欄之理由見上繳 42 §2.2** —— 此處只留一句最要緊的：
# B 帶母本自己的序號公式（shared formula），寫入即破壞其共用主格。
NEVER_WRITE = ["B", "C", "E", "Q", "T", "U", "V", "W", "X", "Y", "Z",
               "AB", "AC", "AD", "AE", "AF", "AG"]
# 條件寫入 —— 其值由參數給定，未給定即不寫
OPTIONAL = {"O": "tc_ref_id_value", "AA": "author_value"}

LEAF_KEY = re.compile(r"SWE1-HMI-PROF-(\d+)(?:-(\d+))?(?:-([a-z]+))?$")

# `applied: false` 之項目所對應之欄 —— WB-0 據此驗「宣告未生效者確實沒被寫」
APPLIED_COL = {"author_value": "AA", "tc_ref_id_value": "O"}


def yaml_params() -> dict:
    """自 `feature.yaml` 之 `write_back` 段取**生效中**之參數（G-C，43 包 §一）。

    `<name>: {value, applied, why}` 之形；**只有 `applied: true` 者回傳其值**。
    未採用者回 `None`，其宣告仍留在 yaml 內（它是一段有來歷的決定，不刪）。

    **這是 G-C 之機械化**：42 輪之前，`author_value: "PeiPYHsu"` 躺在 yaml 裡
    而沒有任何程式讀它 —— 它看起來像是決定過的。現在它被讀了，
    且其「是否生效」是一個要填的欄位，不是一個沒人問的預設。
    """
    import yaml
    d = yaml.safe_load((FEATURE / "feature.yaml").read_text(encoding="utf-8"))
    wb = (d or {}).get("write_back", {})
    out = {}
    for k in ("author_value", "tc_ref_id_value", "vehicle_columns",
              "row_order"):
        v = wb.get(k)
        if isinstance(v, dict):
            out[k] = v.get("value") if v.get("applied") else None
        else:
            out[k] = v                       # 舊式純量，視為生效
    if out.get("row_order") is None:
        out["row_order"] = "req_id"
    return out


def sha256(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def records() -> list:
    out = []
    for p in sorted((FEATURE / "generated").glob("*.json")):
        d = json.loads(p.read_text(encoding="utf-8"))
        for t in d["tcs"]:
            out.append(t)
    return out


def _leaf_sort(tc: dict):
    m = LEAF_KEY.search(tc["req_id"])
    if not m:
        return (999, 999, "", tc["tc_id"])
    # 同一 leaf 之多條（`-neg` 等）以 tc_id 續其後
    return (int(m.group(1)), int(m.group(2) or 0), m.group(3) or "",
            tc["tc_id"])


def order(tcs: list, row_order: str) -> list:
    if row_order == "req_id":
        return sorted(tcs, key=_leaf_sort)
    if row_order == "tc_id":
        return sorted(tcs, key=lambda t: t["tc_id"])
    raise SystemExit(f"未知之 row_order：{row_order}")


def build_edits(tcs: list, *, vehicle_columns=None, author_value=None,
                tc_ref_id_value=None, row_order="req_id") -> dict:
    """`{(row, col_idx): value}` —— **不經 `diff_cells`**。"""
    plan = order(tcs, row_order)
    if FIRST_ROW + len(plan) - 1 > CAPACITY_LAST_ROW:
        raise SystemExit(f"{len(plan)} 條超出範本容量（末列 "
                         f"{CAPACITY_LAST_ROW}）")
    opts = {"tc_ref_id_value": tc_ref_id_value, "author_value": author_value}
    edits = {}
    for i, tc in enumerate(plan):
        r = FIRST_ROW + i
        for col, field in COLS.items():
            edits[(r, col_to_idx(col))] = tc.get(field, "")
        for col, key in OPTIONAL.items():
            if opts[key] is not None:
                edits[(r, col_to_idx(col))] = opts[key]
        for col, val in (vehicle_columns or {}).items():
            edits[(r, col_to_idx(col))] = val
    return edits


def splice(src: Path, out: Path, edits: dict) -> None:
    import zipfile
    member = sheet_members(src)[SHEET]
    with zipfile.ZipFile(src) as zin:
        patched = patch_sheet_xml(zin.read(member).decode("utf-8"), edits)
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = (patched.encode("utf-8") if info.filename == member
                        else zin.read(info.filename))
                zout.writestr(info, data)


# ─────────────────────────────────────────────── 寫回後之檢查（六項）

def verify(src: Path, out: Path, n_rows: int, *, params=None) -> list:
    """違規清單（空 ＝ 綠）。**任一不綠即不留下檔案。**"""
    import zipfile
    bad = []

    member = sheet_members(out)[SHEET]
    with zipfile.ZipFile(out) as z:
        xml = z.read(member).decode("utf-8")

    # WB-0（G-C，43 包）—— `feature.yaml` 宣告 `applied: false` 之項，
    # 其對應欄須**確實沒有被寫**。宣告與生效若分岔，這裡會叫。
    #
    # **以 XML 掃，不以 openpyxl 讀。** 首版用 `load_workbook(read_only=True)`
    # 再逐格 `ws["AA10"]` —— read-only 之工作表不支援隨機存取，
    # 十個方向性案例遂由數秒變成逾兩分鐘。**檢查沒錯，取值途徑錯了。**
    params = yaml_params() if params is None else params
    for key, col in APPLIED_COL.items():
        if params.get(key) is not None:
            continue
        filled = [int(m.group(1)) for m in re.finditer(
            rf'<c r="{col}(\d+)"[^/>]*>(?:(?!</c>).)*?<(?:v|is)\b', xml,
            re.S) if FIRST_ROW <= int(m.group(1)) < FIRST_ROW + n_rows]
        if filled:
            bad.append(f"WB-0 `{key}` 於 feature.yaml 為 applied:false，"
                       f"而 {col} 欄有 {len(filled)} 列被寫入（首列 "
                       f"{min(filled)}）—— 宣告與生效分岔")

    # WB-1～WB-4 —— DV 完整性（`verify_dv_integrity` 之四項）
    bad += [f"WB-DV {x}" for x in DV.verify(src, out)]

    # WB-5 —— 每一寫入列須落在 R 欄 x14 DV 與 P 欄 DV 之涵蓋範圍內
    #         （Comfort write_back §3.3 之同型；**與 DV 完整性互補**：
    #          前者問「DV 還在嗎」，本項問「它蓋得到我寫的列嗎」）
    def cover(sqrefs):
        rows = set()
        for sq in sqrefs:
            for part in sq.split():
                m = re.match(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?$", part)
                if m:
                    rows |= set(range(int(m.group(2)),
                                      int(m.group(4) or m.group(2)) + 1))
        return rows

    written = set(range(FIRST_ROW, FIRST_ROW + n_rows))
    r_rows = cover(re.findall(r"<xm:sqref>([^<]+)</xm:sqref>", xml))
    p_rows = cover([m for m in re.findall(
        r'<dataValidation[^>]*sqref="([^"]+)"', xml) if m.startswith("P")])
    for name, got in (("R 欄 design_method（x14）", r_rows),
                      ("P 欄 priority", p_rows)):
        miss = sorted(written - got)
        if miss:
            bad.append(f"WB-5 {len(miss)} 列不在 {name} 之 DV 涵蓋範圍內"
                       f"（首列 {miss[0]}）")

    # WB-6 —— 換行字元：**LF only**（未決 5 之自裁，寫成可測形式）
    #
    # **判準改過一次（首跑即誤報）。** v1 掃整份 sheet XML，
    # 而母本之 XML 宣告行本身就以 `?>\r\n<worksheet` 結尾 ——
    # 那是**檔案之序列化格式**，不是儲存格內容。v1 遂對一次正確之寫回轉紅。
    # v2 只掃 `<t>` 內之文字（即儲存格值）。
    for txt in re.findall(r"<t[^>]*>(.*?)</t>", xml, re.S):
        if "&#13;" in txt or "\r" in txt:
            bad.append("WB-6 儲存格內容出現 CR —— 多行欄位須為 LF（`\\n`）")
            break
    return bad


# ─────────────────────────────────────────────── 產出（受未決 1 之閘）

def decision_gate() -> list:
    """產出之前置閘 —— **問「決定過了嗎」，不問「值是不是空的」**。

    ## 判準改過一次（48 包）

    v1 為「`vehicle_columns is None` 即拒絕產出」。該判準寫於 42 輪，
    當時 T:Z 確實未決。**44 輪已將其定為留空**（Comfort 交付件 466 列全空、
    其 `NEVER_WRITE` 明列該七欄、母本該區 DV 自帶 `allowBlank="1"`）——
    **而 v1 分不出「還沒決定」與「決定了留空」**：兩者之 value 都是 `None`。

    v2 改問 `feature.yaml`：該項須存在、須有明確之 `applied`、
    且 **`why` 不得空白**（G-C 之必填要求）。
    **「決定了留空」與「還沒想好」在 yaml 裡長得一樣，只有 `why` 分得開** ——
    本閘即讀那一欄。

    **仍然沒有 `--force`。** 可以用旗標關掉的閘不是閘。
    """
    import yaml
    d = yaml.safe_load((FEATURE / "feature.yaml").read_text(encoding="utf-8"))
    wb = (d or {}).get("write_back", {})
    bad = []
    for key in ("vehicle_columns", "author_value", "tc_ref_id_value"):
        v = wb.get(key)
        if not isinstance(v, dict) or "applied" not in v:
            bad.append(f"WB-G `{key}` 未以 {{value, applied, why}} 之形記於 "
                       f"feature.yaml —— **尚未決定，不得產出**")
        elif not str(v.get("why") or "").strip():
            bad.append(f"WB-G `{key}` 之 `why` 空白 —— "
                       f"**「決定了不做」與「還沒想好」無從分辨**（G-C）")
    return bad


def write(out: Path, *, vehicle_columns=None, **kw) -> Path:
    gate = decision_gate()
    if gate:
        raise SystemExit("拒絕產出：\n  " + "\n  ".join(gate))
    if sha256(SRC) != SRC_SHA:
        raise SystemExit(f"來源母本之 SHA 不符：{sha256(SRC)[:16]}…")
    tcs = records()
    edits = build_edits(tcs, vehicle_columns=vehicle_columns, **kw)
    out.parent.mkdir(parents=True, exist_ok=True)
    tmp = out.with_suffix(".tmp.xlsx")
    splice(SRC, tmp, edits)
    bad = verify(SRC, tmp, len(tcs),
                 params={"author_value": kw.get("author_value"),
                         "tc_ref_id_value": kw.get("tc_ref_id_value")})
    if bad:
        tmp.unlink()
        for b in bad:
            print(f"  {b}")
        raise SystemExit(f"寫回後檢查 {len(bad)} 項不綠 —— **未留下檔案**")
    tmp.rename(out)
    return out


# ─────────────────────────────────────────────── 實跑探針（43 包作業 5）

def pre_gates() -> list:
    """**寫回前之二閘**（41 輪設計草案 §5.4 之第一段）。

    - `audit_assignment`：號碼指派表自生成器重算，與產物相符
      —— **不符即代表產物與生成器分岔**，那種情形下寫回是把分岔封進交付件
    - `lint_tcs`：P 欄值 ∈ DV、R 欄值 ∈ 下拉九條，及其餘 20 項
    """
    import audit_assignment as A
    import lint_tcs as L
    bad = []
    bad += [f"assignment: {x}" for x in A.audit()]
    tcs = [t for _d, t in L.corpus()] if hasattr(L, "corpus") else None
    if tcs is None:
        import json as _j
        tcs = []
        for q in sorted((FEATURE / "generated").glob("*.json")):
            tcs += _j.loads(q.read_text(encoding="utf-8"))["tcs"]
    for t in tcs:
        bad += [f"lint: {x}" for x in L.gate_tc(t)]
    bad += [f"lint: {x}" for x in L.gate_corpus(tcs)]
    return bad


def probe(out: Path, *, src: Path = None, row_order=None) -> int:
    """**以 scratchpad 之母本複本實跑一次，三段接點逐段回報。**

    與 `--write` 之別：**不落 `output/`、不入台帳、不視為交付件**。
    故其不受未決 1 之閘拘束 —— 該閘守的是交付件，不是驗證。
    """
    if FEATURE / "output" in out.parents or out.parent == FEATURE / "output":
        raise SystemExit("probe 之產物不得落 `output/` —— 那是產出物之位置，"
                         "而 probe 不是產出物（43 包作業 5）")
    prm = yaml_params()
    src = src or SRC
    ro = row_order or prm["row_order"]
    tcs = records()

    print("## 第一段 —— 寫回前之二閘\n")
    g1 = pre_gates()
    print(f"  audit_assignment ＋ lint_tcs：違規 {len(g1)}")
    for x in g1[:5]:
        print(f"    {x}")
    if g1:
        raise SystemExit("寫回前之閘不綠 —— 不進行寫回")

    print(f"\n## 第二段 —— 封裝（`patch_sheet_xml`，不經 `diff_cells`）\n")
    print(f"  來源 {src.name}")
    print(f"  來源 SHA {'相符' if sha256(src) == SRC_SHA else '**不符**'}")
    edits = build_edits(tcs, vehicle_columns=prm["vehicle_columns"],
                        author_value=prm["author_value"],
                        tc_ref_id_value=prm["tc_ref_id_value"],
                        row_order=ro)
    out.parent.mkdir(parents=True, exist_ok=True)
    splice(src, out, edits)
    print(f"  {len(tcs)} 條 → row {FIRST_ROW}–{FIRST_ROW + len(tcs) - 1}；"
          f"edits {len(edits)} 格；列序 {ro}")
    print(f"  產物 {out}")

    print(f"\n## 第三段 —— 寫回後之六項檢查\n")
    bad = verify(src, out, len(tcs), params=prm)
    names = ["WB-0 applied:false 之欄確實未寫",
             "WB-DV DV-1 zip member 集合", "WB-DV DV-2 x14 節點數",
             "WB-DV DV-3 xm:sqref 範圍", "WB-DV DV-4 legacy 節點數",
             "WB-5 寫入列在 R／P 欄 DV 涵蓋內", "WB-6 儲存格內容無 CR"]
    for n in names:
        hit = [x for x in bad if x.split(" ")[0] in n or n.startswith(
            " ".join(x.split(" ")[:2]))]
        print(f"  {'**紅**' if hit else '綠'} — {n}")
    print(f"\n違規 {len(bad)}")
    for x in bad:
        print(f"  {x}")
    print(f"\n**本產物不落 `output/`、不入 `DELIVERY.sha256`、不視為交付件。**")
    return 1 if bad else 0


def dry_run(row_order="req_id") -> int:
    tcs = records()
    plan = order(tcs, row_order)
    edits = build_edits(tcs, row_order=row_order)
    print(f"來源 {SRC.name}")
    print(f"  SHA {'相符' if sha256(SRC) == SRC_SHA else '**不符**'}"
          f"（{SRC_SHA[:16]}…）")
    print(f"\nTC {len(tcs)} 條 → row {FIRST_ROW}–{FIRST_ROW + len(tcs) - 1}"
          f"（容量末列 {CAPACITY_LAST_ROW}，餘裕 "
          f"{CAPACITY_LAST_ROW - FIRST_ROW - len(tcs) + 1} 列）")
    print(f"列序：{row_order}；首三列 "
          f"{[t['req_id'] for t in plan[:3]]} … 末列 {plan[-1]['req_id']}")
    print(f"\n寫入欄 {len(COLS)}：{' '.join(COLS)}")
    print(f"不寫入欄 {len(NEVER_WRITE)}：{' '.join(NEVER_WRITE)}")
    print(f"條件欄（未給值即不寫）：{' '.join(OPTIONAL)}")
    print(f"\nedits 共 {len(edits)} 格")
    print("\n**未產出檔案** —— 未決 1 未定前，`--write` 拒絕執行。")
    return 0


def self_test() -> int:
    """方向性案例 —— 全程寫 scratchpad，**不落 `output/`、不入台帳**。"""
    scratch = DV.SCRATCH.parent / "writeback"
    scratch.mkdir(parents=True, exist_ok=True)
    tcs = records()
    ok, cases = True, []

    def case(name, fn, expect_red, note=""):
        nonlocal ok
        cases.append(name)
        bad = fn()
        good = bool(bad) == expect_red
        ok &= good
        print(f"  {'PASS' if good else '**FAIL**'} — {name}："
              f"{'紅' if bad else '綠'}，期望 {'紅' if expect_red else '綠'}")
        if note:
            print(f"      note {note}")
        for b in bad[:2]:
            print(f"      └ {b}")

    # ① 正常寫回（T:Z 不寫）→ 六項全綠
    out = scratch / "probe_default.xlsx"
    splice(SRC, out, build_edits(tcs))
    case("預設參數（T:Z／O／AA 皆不寫）→ 綠", lambda: verify(SRC, out, len(tcs)),
         False, f"{len(tcs)} 列，row {FIRST_ROW}–{FIRST_ROW + len(tcs) - 1}")

    # ② 給定 vehicle_columns → 仍綠（參數化未破壞任何檢查）
    out2 = scratch / "probe_vehicles.xlsx"
    splice(SRC, out2, build_edits(tcs, vehicle_columns={"T": "1", "U": "0"}))
    case("給定 `vehicle_columns` → 綠（參數化不改變任何檢查）",
         lambda: verify(SRC, out2, len(tcs)), False)

    # ③ 注入向 —— 寫到 DV 涵蓋範圍之外
    out3 = scratch / "probe_beyond_dv.xlsx"
    ed = {(CAPACITY_LAST_ROW + 5, col_to_idx("R")): "功能測試 (Functional "
          "based ; no specific technique)"}
    splice(SRC, out3, ed)
    case("**注入：寫到 DV 涵蓋範圍之外（row 1416）→ 紅**",
         lambda: verify(SRC, out3, CAPACITY_LAST_ROW + 5 - FIRST_ROW + 1),
         True)

    # ④ 注入向 —— 以 openpyxl 存回（A-UP09 之形態）
    out4 = scratch / "probe_openpyxl.xlsx"
    import openpyxl
    wb = openpyxl.load_workbook(SRC)
    wb.save(out4)
    wb.close()
    case("**注入：以 openpyxl 存回 → 紅**（DV 完整性）",
         lambda: verify(SRC, out4, len(tcs)), True)

    # ⑤ 注入向 —— 多行欄位含 CRLF
    out5 = scratch / "probe_crlf.xlsx"
    splice(SRC, out5, {(FIRST_ROW, col_to_idx("J")): "1. a\r\n2. b"})
    case("**注入：多行欄位以 CRLF 寫入 → 紅**（未決 5 之 LF 自裁）",
         lambda: verify(SRC, out5, 1), True)

    # ⑥ 產出之閘（v2，48 包）—— 問「決定過了嗎」，不問「值是不是空的」
    case("**產出閘：`feature.yaml` 三項皆已決定且具名 `why` → 綠**",
         decision_gate, False)

    def undecided():
        import yaml as _y
        orig = _y.safe_load
        def fake(txt):
            d = orig(txt)
            if isinstance(d, dict) and "write_back" in d:
                d["write_back"]["vehicle_columns"]["why"] = "   "
            return d
        _y.safe_load = fake
        try:
            return decision_gate()
        finally:
            _y.safe_load = orig
    case("**注入：`vehicle_columns` 之 `why` 被清空 → 拒絕產出**"
         "（『決定了留空』與『還沒想好』無從分辨）", undecided, True)

    def not_a_mapping():
        import yaml as _y
        orig = _y.safe_load
        def fake(txt):
            d = orig(txt)
            if isinstance(d, dict) and "write_back" in d:
                d["write_back"]["vehicle_columns"] = None   # 42 輪之舊式純量
            return d
        _y.safe_load = fake
        try:
            return decision_gate()
        finally:
            _y.safe_load = orig
    case("**注入：`vehicle_columns` 退回舊式純量（無 applied）→ 拒絕產出**",
         not_a_mapping, True)

    # ⑦ 護欄 —— 列序兩種皆為全排列且不重不漏
    def order_total():
        bad = []
        for ro in ("req_id", "tc_id"):
            p = order(tcs, ro)
            if len(p) != len(tcs) or {t["tc_id"] for t in p} != \
                    {t["tc_id"] for t in tcs}:
                bad.append(f"row_order={ro} 非全排列")
        return bad
    case("**護欄**：`row_order` 兩種皆為全排列（不重不漏）→ 綠",
         order_total, False)

    # ⑧ WB-0（G-C）—— yaml 說 applied:false，而該欄被寫入
    out8 = scratch / "probe_applied_split.xlsx"
    splice(SRC, out8, build_edits(tcs, author_value="PeiPYHsu"))
    case("**WB-0 注入：`author_value` 為 applied:false 而 AA 欄被寫 → 紅**",
         lambda: verify(SRC, out8, len(tcs),
                        params={"author_value": None,
                                "tc_ref_id_value": None}), True,
         "宣告與生效分岔即轉紅（G-C 之機械化）")

    # ⑨ WB-0 之另一半 —— 宣告生效時，同一份產物須為綠
    case("WB-0 護欄：`author_value` 改為 applied:true → 綠（同一份產物）",
         lambda: verify(SRC, out8, len(tcs),
                        params={"author_value": "PeiPYHsu",
                                "tc_ref_id_value": None}), False)

    # ⑩ probe 不得落 `output/`
    def probe_refuses_output():
        try:
            probe(FEATURE / "output" / "must_not_exist.xlsx")
        except SystemExit as e:
            return [str(e).splitlines()[0]]
        return []
    case("**probe 之產物落 `output/` → 拒絕**", probe_refuses_output, True)
    assert not (FEATURE / "output" / "must_not_exist.xlsx").exists()

    n = len(cases)
    print(f"\n{n if ok else '<' + str(n)} / {n} directional cases "
          f"{'PASS' if ok else 'FAIL'}")
    print(f"\n（全程寫 scratchpad `{scratch}`；"
          f"`output/` 未產生任何檔案，台帳未記帳）")
    return 0 if ok else 1


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--self-test", action="store_true")
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--probe", action="store_true",
                    help="以 scratchpad 之複本實跑，不落 output/")
    ap.add_argument("--src", default=None)
    ap.add_argument("--vehicle-columns", default=None,
                    help='JSON，如 \'{"T":"1","U":"0"}\'')
    ap.add_argument("--row-order", default="req_id",
                    choices=["req_id", "tc_id"])
    ap.add_argument("--out", default=None)
    a = ap.parse_args()
    if a.self_test:
        sys.exit(self_test())
    if a.probe:
        if not a.out:
            ap.error("--probe 須給 --out（scratchpad 之路徑）")
        sys.exit(probe(Path(a.out),
                       src=Path(a.src) if a.src else None,
                       row_order=a.row_order))
    if not a.write:
        sys.exit(dry_run(a.row_order))
    vc = json.loads(a.vehicle_columns) if a.vehicle_columns else None
    out = Path(a.out) if a.out else (FEATURE / "output" / (
        "FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT STLA Test Case "
        "Specification & Result_SWQT_UserProfiles_full.xlsx"))
    print("寫出", write(out, vehicle_columns=vc, row_order=a.row_order))
