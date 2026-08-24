"""W-VF46 —— VF230 之委派判定（V17 §5；**不生成 TC**）。

**W-VF50（V18 §5）—— 比對範圍自 Comfort 一個擴及全部候選 feature。**
候選之全集以 `features/` 下之**實有目錄**為準（R-VF53 二：不以列舉為準
—— 列舉必不完整）。每一候選回報三態：已比對／無 037 素材無法比對／
有素材而本輪未比對。**不得以「推定無關」代替比對。**

**沿用 Part 1 之委派判準，不新設**（V17 §5 第 1 項）：
  判準為「**同一實體功能於 Comfort 037 全集是否有對應**」，
  於 **Layer 3（簇）層級**判定（Part 1 之 `delegation_lookup.tsv` 之 `basis`
  逐字為「同一實體功能：…｜Layer 3 層級，未逐 leaf 收斂」／
  「Comfort 037 全集無同一實體功能之對應」）。

**委派之效果依 Part 1 之 R-VS59（Pei 2026-08-23）** ——
**委派不免除產出 TC 之義務**；其作用為「該 leaf 之畫面層內容取自 Comfort 素材」，
非「故不寫」。R-VS7(a) 之原效果已撤回。

依 R-VF21／R-VF28 附三錨點；**鑑別錨點為「簇名含 Comfort 而實不應委派」者**。

輸出：docs/reports/vf230_delegation.tsv ＋ data/_vf230_delegation.json
"""
import csv
import glob
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
FEATURES = Path("/Users/peihe/Work_Projects/TC_Generator/features")


def norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


def feature_titles(feat: str) -> tuple[set[str], str]:
    """某 feature 之 037 Requirement Title 全集（正規化）與其狀態。

    回 (titles, state)；state ∈ {"已比對", "無 037 素材，無法比對"}。
    """
    inp = FEATURES / feat / "inputs"
    g = sorted(x for x in inp.glob("*.xlsx") if "037" in x.name) if inp.is_dir() else []
    if not g:
        return set(), "無 037 素材，無法比對"
    wb = openpyxl.load_workbook(g[0], read_only=True, data_only=True)
    out = set()
    for nm in wb.sheetnames:
        rs = list(wb[nm].iter_rows(values_only=True))
        i = next((j for j, r in enumerate(rs)
                  if any("requirement description" in norm(str(v or "")).replace(" ", " ")
                         for v in r)), None)
        if i is None:
            continue
        ti = next((j for j, v in enumerate(rs[i]) if "title" in str(v or "").lower()), None)
        if ti is None:
            continue
        for r in rs[i + 1:]:
            if r[0] and r[ti]:
                out.add(norm(str(r[ti])))
        break
    wb.close()
    return out, "已比對"


# Comfort 所擁有之實體功能域（自其 037 之簇名歸納，逐字列出供覆核）
COMFORT_DOMAIN = ("hvac", "climate", "atc", "mtc", "defrost", "fan speed",
                  "sync", "recirculation", "air distribution", "temperature control",
                  "heated seat", "vented seat", "heated steering")


def main() -> None:
    cands = sorted(d.name for d in FEATURES.iterdir()
                   if d.is_dir() and d.name != "vehicle_setting")
    per_feat, states = {}, {}
    for f in cands:
        t, st = feature_titles(f)
        per_feat[f], states[f] = t, st
    ct = per_feat.get("comfort", set())
    lv = list(csv.DictReader(
        (ROOT / "data" / "vf230_leaves.tsv").open(encoding="utf-8"), delimiter="\t"))
    enum = json.loads((ROOT / "data" / "_wvf35_enum.json").read_text(encoding="utf-8"))
    ts_of = {c: t for t, cs in enum["final"].items() for c in cs}

    by_cluster: dict[str, list] = defaultdict(list)
    for r in lv:
        by_cluster[r["title"]].append(r)

    rows = []
    for title, leaves in sorted(by_cluster.items()):
        n = norm(title)
        # 逐 feature 比對，取首個命中者為委派標的
        tgt = next((f for f in cands if n in per_feat[f]), None)
        exact = tgt is not None
        # **詞界比對，非子字串**（R-VF52：誤報以判準精確化消除，不以白名單）。
        # 首版以 `d in n` 比對，致 `sync` 誤配 `Enhanced Display
        # Synchronization`（6 leaf）—— 逐條比對後確認二者為不同功能：
        # VF230 為 SRT PROXI 配置之顯示同步，Comfort 為空調雙區溫度同步。
        domain = [d for d in COMFORT_DOMAIN
                  if re.search(rf"\b{re.escape(d)}\b", n)]
        if exact:
            dele, basis = "yes", f"同一實體功能：`{tgt}` 之 037 有同名簇 `{title}`"
        elif domain:
            dele, basis = "pending", ("簇名落於 Comfort 之實體功能域（"
                                      + "／".join(domain) + "）而 Comfort 037 無同名簇"
                                      " —— 須逐條比對條文，本輪標 pending")
        else:
            dele, basis = "no", ("已比對之 "
                                 + str(sum(1 for f in cands if states[f] == "已比對"))
                                 + " 個 feature 之 037 全集皆無同一實體功能之對應")
        for lf in leaves:
            rows.append({"leaf_id": lf["swe_id"],
                         "test_set": ts_of.get(title, ""),
                         "layer3": title.replace("\\n", " "),
                         "delegate": dele,
                         "target_feature": (tgt or "comfort") if dele != "no" else "",
                         "basis": basis,
                         "screen_source": (tgt or "comfort") if dele == "yes" else ""})

    # --- 錨點（R-VF21／R-VF28）---
    def dele_of(t):
        return next((r["delegate"] for r in rows if r["layer3"] == t), None)

    # 必不命中：`Speed Unit` 為單位顯示，Comfort 無此功能
    if dele_of("Speed Unit") != "no":
        raise SystemExit("必不命中錨點不符：`Speed Unit` 應判 no，停")
    # 鑑別：簇名含 `Comfort` 而實為車輛設定，不應委派
    disc = [t for t in by_cluster if "comfort" in norm(t)]
    bad = [t for t in disc if dele_of(t) != "no"]
    if bad:
        raise SystemExit(f"鑑別錨點不符：簇名含 Comfort 而被判委派 —— {bad}，停")
    print(f"鑑別錨點：簇名含 `Comfort` 者 {len(disc)} 個"
          f"（{'／'.join(disc)}），皆判 no ✅")
    # --- W-VF50(4)：逐 feature 之鑑別錨點 ---
    # 取「簇名與該 feature 之名近似而實不應委派」者。
    # 一條以 feature 名為鍵之規則會誤配之。
    DISC = {
        "comfort": ["Auto On Driver Comfort - 3 Option",
                    "Auto On Driver Comfort - 2 Option"],
        "power": ["Power Tailgate", "Max Power Level", "Charge Power Level",
                  "Power Unit", "Power Side Step", "Engine Off Power Delay"],
        "power_moding": ["SWITCH 1 Power Mode", "SWITCH 2 Power Mode",
                         "SWITCH 3 Power Mode", "SWITCH 4 Power Mode"],
        "user_profiles": [],
        "sxm": [],
    }
    for feat, cl in DISC.items():
        if states.get(feat) != "已比對":
            continue
        present = [c for c in cl if c in by_cluster]
        if not present:
            print(f"鑑別錨點（{feat}）：**不存在** —— VF230 無簇名與之近似者")
            continue
        bad = [c for c in present if dele_of(c) != "no"]
        if bad:
            raise SystemExit(f"鑑別錨點不符（{feat}）：{bad} 被判委派，停")
        print(f"鑑別錨點（{feat}）：{len(present)} 個近名簇"
              f"（{'／'.join(present[:3])}{'…' if len(present) > 3 else ''}）"
              " 皆判 no ✅")

    # 必命中：037 有同名簇者（若不存在則具名）
    hit = [t for t in by_cluster if norm(t) in ct]
    if hit:
        print(f"必命中錨點：Comfort 037 有同名簇者 {len(hit)} 個 —— {hit[:3]}")
    else:
        print("必命中錨點：**不存在** —— VF230 之 106 簇無一與 Comfort 037 同名")
    if len(rows) != 627:
        raise SystemExit(f"R-VF16：應為 627 列，實得 {len(rows)}")

    p = ROOT / "docs" / "reports" / "vf230_delegation.tsv"
    with p.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0]), delimiter="\t",
                           lineterminator="\n")
        w.writeheader()
        w.writerows(rows)

    dist = Counter(r["delegate"] for r in rows)
    per = defaultdict(Counter)
    for r in rows:
        per[r["test_set"]][r["delegate"]] += 1
    (ROOT / "data" / "_vf230_delegation.json").write_text(json.dumps({
        "dist": dict(dist), "per_test_set": {k: dict(v) for k, v in per.items()},
        "candidates": cands, "states": states,
        "titles_per_feature": {f: len(per_feat[f]) for f in cands},
        "vf230_clusters": len(by_cluster),
        "exact_hits": hit, "name_contains_comfort": disc},
        ensure_ascii=False, indent=2), encoding="utf-8")

    print("\n候選 feature（features/ 之實有目錄，扣 vehicle_setting 自身）："
          f"{len(cands)}")
    for f in cands:
        print(f"  {f:16} {states[f]:22}"
              + (f"037 簇 {len(per_feat[f]):4}；與 VF230 同名 "
                 f"{len({t for t in by_cluster if norm(t) in per_feat[f]})}"
                 if states[f] == "已比對" else ""))
    print(f"\nVF230 之簇 {len(by_cluster)}")
    print(f"委派分布（627 leaf）：{dict(dist)}")
    for t, c in sorted(per.items(), key=lambda x: -sum(x[1].values())):
        print(f"  {sum(c.values()):4}  {t:24} {dict(c)}")
    print(f"wrote {p.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
