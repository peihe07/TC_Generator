#!/usr/bin/env python3
"""VS-SL-03 §2 —— 沙盒寫回稿（Revise 稿）。

**只寫 `features/<slug>/sandbox/vssl/`。不動 `output/`、不動 `delivered/`、不動 `inputs/`。**

來源現本：
  vf230  `features/vehicle_setting/inputs/…_SWQT_VF230_20260819.xlsx`
         —— 本 feature **無 `output/`**（且 `output/` 於其 `.gitignore` 內），
            故依 `inputs/` 之現本複製；此為與包內「自 output/ 複製」之差異，已於上繳報明。
  BL     `features/bed_lowering/output/…_SWQT_BedLowering_20260827.xlsx`
  VC     `features/vehicle_category/output/…_VehicleCategory_20260827_working.xlsx`

vf230 之改動（依 v3 報告逐列施作）：
  1. 非 NAFTA 19 列移除（先存移除清單）—— **儲存格改動全部套完後才刪列**，避免列號位移
  2. PROXI 形制改寫（R-VS86）；NEG_CONTRA 3 列改 `= 0 (Absent)`
  3. 分支 (2) 補 PROXI 行；分支 (3) 寫 `PENDING`
  4. 有 `path_proposed` 者補導覽路徑與控件句（Options 依 VS-SL-02 §2.4 正規化）
  5. `Test procedure` 與 `Expected Result` 之步數逐列須相等（硬規，落檔前 assert）
"""

from __future__ import annotations

import csv
import re
import shutil
import sys
import warnings
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import vs_sl01_dryrun as v1  # noqa: E402

warnings.filterwarnings("ignore")
ROOT = v1.ROOT
SHEET = v1.SHEET
HDR = v1.HEADER_ROW

COL_PRE, COL_PROC, COL_ER = 10, 12, 13

BOOKS = {
    "vf230": (v1.VF230, "features/vehicle_setting/sandbox/vssl/vf230_vssl.xlsx",
              "features/vehicle_setting/reports/vf230_settings_dryrun_v3.tsv"),
    "bl": (v1.BL, "features/bed_lowering/sandbox/vssl/bl_vssl.xlsx",
           "features/bed_lowering/reports/bl_settings_dryrun_v3.tsv"),
    "vc": (v1.VC, "features/vehicle_category/sandbox/vssl/vc_vssl.xlsx",
           "features/vehicle_category/reports/vc_settings_dryrun_v3.tsv"),
}

GENERIC_OPEN = re.compile(r"Open the Vehicle Settings menu.*$", re.I)
MENU_SHOWN = re.compile(r"The Vehicle Settings menu is displayed.*$", re.I)
CHECK_LINE = re.compile(r"check that the .+ setting is (not )?displayed", re.I)
SHOWN_LINE = re.compile(r'^(The )?".+" customer setting is (not )?displayed', re.I)
# 設定項名之引號正規化（§2 硬規「雙引號 lint 過」）。名之逐詞須為大寫起首或數字，
# 泛稱句（`the setting named in the test data`）因此不受影響。
_NAME = r"(?:[A-Z][A-Za-z0-9\-/&]*)(?:\s+(?:[A-Z0-9][A-Za-z0-9\-/&]*|with|and|or|for))*"
UNQUOTED_NAME = re.compile(rf'(?<![\w"])(?P<lead>[Tt]he\s+)(?P<name>{_NAME})(?P<tail>\s+setting\b)')


def quote_setting_names(text: str) -> str:
    """`the SWITCH 5 Power Mode setting` → `the "SWITCH 5 Power Mode" setting`。"""
    def sub(m):
        return f'{m.group("lead")}"{m.group("name")}"{m.group("tail")}'
    return "\n".join(UNQUOTED_NAME.sub(sub, ln) for ln in text.split("\n"))


def lines(text: str) -> list[str]:
    """去序號之行清單。"""
    return [re.sub(r"^\s*\d+\.\s*", "", x).strip()
            for x in str(text or "").split("\n") if x.strip()]


def numbered(items: list[str]) -> str:
    return "\n".join(f"{i}. {x}" for i, x in enumerate(items, 1))


def new_pre(original: str, proposed: str) -> str:
    """Pre-Condition：保留非 PROXI 行之原序，PROXI 行以提議取代。"""
    keep = [x for x in lines(original) if not x.upper().startswith("PROXI")]
    add = [x.strip() for x in proposed.split("｜")[0].split(" ; ") if x.strip()]
    return numbered(keep + add)


def is_real_path(path: str) -> bool:
    """真路徑之判準：以 `Settings > ` 起首且至少三節。"""
    nodes = [x.strip() for x in (path or "").split(" > ")]
    return len(nodes) >= 3 and nodes[0] == "Settings"


def nav_steps(path: str) -> list[str]:
    """`Settings > A > B > item` → 導覽步驟（不含末項之檢查句）。"""
    nodes = [x.strip() for x in path.split(" > ")]
    out = ['Press "Settings" on Menu Bar']
    out += [f'Select "{n}"' for n in nodes[1:-1]]
    return out


def nav_expected(path: str) -> list[str]:
    nodes = [x.strip() for x in path.split(" > ")]
    return ["The Settings screen is displayed"] + \
           [f'The "{n}" page is displayed' for n in nodes[1:-1]]


def rewrite_pair(proc: str, er: str, path: str, control: str) -> tuple[str, str]:
    """同時改寫 Procedure 與 Expected Result。

    兩欄原本逐列 1:1 對齊（實測 457 列 0 筆不一致），故**於同一索引插入同長度之
    導覽段**，長度相等由構造保證，不靠事後修補。
    """
    prc, exp = lines(proc), lines(er)
    if len(prc) != len(exp):
        raise AssertionError(f"來源本 Procedure {len(prc)} 步 vs Expected {len(exp)} 步")

    nav_p, nav_e = nav_steps(path), nav_expected(path)
    # **控件句只套在原有之行**（其中指名該設定項者）；導覽段為本層新造，不得再套，
    # 否則會生出「The Settings screen is displayed as 3 radio buttons …」之誤句。
    prc = [_with_control(x, control) for x in prc]
    exp = [_with_control(x, control, expected=True) for x in exp]

    idx = next((i for i, x in enumerate(prc) if GENERIC_OPEN.search(x)), None)
    if idx is not None:
        # 泛稱之開啟句 → 以導覽段取代（Expected 之對應句同索引取代）
        prc = prc[:idx] + nav_p + prc[idx + 1:]
        exp = exp[:idx] + nav_e + exp[idx + 1:]
    else:
        idx = next((i for i, x in enumerate(prc) if CHECK_LINE.search(x)), len(prc))
        prc = prc[:idx] + nav_p + prc[idx:]
        exp = exp[:idx] + nav_e + exp[idx:]
    if len(prc) != len(exp):
        raise AssertionError("導覽段插入後步數不等")
    return numbered(prc), numbered(exp)


def _with_control(line: str, control: str, expected: bool = False) -> str:
    """在「顯示」之肯定句尾補控件句。

    否定句不動；**已載明形態者（`is displayed as/with …`）亦不動** ——
    例 `… is displayed as IGNITION` 為既有之值敘述，覆寫將毀其語意。
    """
    if not control:
        return line
    if re.search(r"\bis not displayed\b", line, re.I):
        return line
    if re.search(r"\bis displayed\s+(as|with)\b", line, re.I):
        return line
    if not re.search(r"\bis displayed\b", line, re.I):
        return line
    joiner = "as" if expected else "with"
    return re.sub(r"\bis displayed\b", f"is displayed {joiner} {control}", line,
                  count=1, flags=re.I)


def apply_vf230(plan: list[dict], ws, log: list[str]) -> list[dict]:
    removed = []
    for p in plan:
        r = int(p["row"])
        if "NON_NAFTA" in p["flags"]:
            removed.append({"row": r, "req_id": ws.cell(r, 4).value,
                            "setting": p["setting"], "reason": "R-VS84(4) 非 NAFTA"})
            continue
        _apply_row(ws, r, p, log)
    return removed


def _apply_row(ws, r: int, p: dict, log: list[str]) -> None:
    pre_old = ws.cell(r, COL_PRE).value
    if p["proxi_proposed"]:
        pre_new = new_pre(pre_old, p["proxi_proposed"])
        if pre_new != (pre_old or ""):
            ws.cell(r, COL_PRE).value = pre_new
            log.append(f"r{r} Pre")
    # 只有真路徑才改 Proc/ER。BL／VC 之 `path_proposed` 欄為說明字串
    # （「（本 feature 不套 path 段）」），非路徑 —— 誤判會塞進假的導覽步驟。
    if is_real_path(p["path_proposed"]):
        proc_new, er_new = rewrite_pair(ws.cell(r, COL_PROC).value,
                                        ws.cell(r, COL_ER).value,
                                        p["path_proposed"], p["control_proposed"])
        ws.cell(r, COL_PROC).value = proc_new
        ws.cell(r, COL_ER).value = er_new
        log.append(f"r{r} Proc/ER")
    for c in (COL_PROC, COL_ER):
        cur = ws.cell(r, c).value
        if cur:
            fixed = quote_setting_names(str(cur))
            if fixed != cur:
                ws.cell(r, c).value = fixed
                log.append(f"r{r} quote")


def run(tag: str) -> dict:
    src, dst, rep = BOOKS[tag]
    out = ROOT / dst
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(ROOT / src, out)
    wb = openpyxl.load_workbook(out)
    ws = wb[SHEET]
    plan = list(csv.DictReader(open(ROOT / rep), delimiter="\t"))
    log: list[str] = []

    removed = apply_vf230(plan, ws, log) if tag == "vf230" else []
    if tag != "vf230":
        for p in plan:
            _apply_row(ws, int(p["row"]), p, log)

    for x in sorted((d["row"] for d in removed), reverse=True):
        ws.delete_rows(x, 1)

    wb.save(out)
    return {"tag": tag, "path": dst, "rows_before": len(plan),
            "rows_after": len(plan) - len(removed), "cells": len(log),
            "removed": removed}


def main() -> int:
    stats = []
    for tag in ("vf230", "bl", "vc"):
        s = run(tag)
        stats.append(s)
        print(f"{tag:6} {s['rows_before']} → {s['rows_after']} 列；"
              f"改動 {s['cells']} 處；移除 {len(s['removed'])} 列")
        if s["removed"]:
            p = ROOT / "features/vehicle_setting/reports/vf230_removed_non_nafta.tsv"
            with open(p, "w", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=["row", "req_id", "setting", "reason"],
                                   delimiter="\t")
                w.writeheader()
                w.writerows(s["removed"])
            print(f"       移除清單 → {p.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
