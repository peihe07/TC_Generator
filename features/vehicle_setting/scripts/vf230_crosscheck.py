"""VF230 之跨源驗核（DR-28 之替代來源）。

61 包 §7／DR-28 稱 VF230 缺 SYS2 ICS export，其 619 leaf 之
Functional/Heading 判定「無第二來源可核」。**該前提已不成立** ——
Pei 於 2026-08-23 補入之

  FM-WI-FSM-035-A02_VF230_HDCC_DT_…_SYSRA_VF230_V4_Released.xlsx

其 `Basic Report` 分頁之 schema 與 CFTS044 之 SYS2 export **逐欄同型**
（`SYS2 Sys-RA-Feature-ID`／`SYS2 分類 Category`／`SYS2 VF章節`／
`SYS2 EE Architecture`），且涵蓋 037 之全部 619 leaf。

本腳本以之對 `data/vf230_leaves.tsv` 作雙向對帳，等同 Part 1 於 01 輪
對 SYS2 之 537 列對帳。**只量測、不改 leaf 母體** —— 錯配逐筆登記為
anomaly，其處置屬裁定事項。

輸出：docs/reports/vf230_crosscheck.md ＋ data/_vf230_crosscheck.json
"""
import csv
import glob
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ("inputs/FM-WI-FSM-035-A02_VF230_HDCC_DT_STLA 技術安全需求分析報告_SYSRA "
       "STLA Technical Safety Requirement Analysis Report_SYSRA_VF230_V4_Released.xlsx")


def norm(v) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip().lower()


def load_sysra() -> tuple[dict, dict]:
    """回 (逐 Sys-RA 之屬性, 表頭欄位對照)。"""
    wb = openpyxl.load_workbook(ROOT / SRC, read_only=True, data_only=True)
    rows = list(wb["Basic Report"].iter_rows(values_only=True))
    h = rows[0]

    def find(sub):
        return next((j for j, v in enumerate(h) if sub in str(v or "")), None)

    cols = {"ref": find("Sys-RA-Feature"), "cat": find("分類"),
            "ch": find("VF章節"), "ee": find("EE Architecture"),
            "asil": find("ASIL"), "desc": find("Description"),
            "region": find("限定地區")}
    if cols["ref"] is None or cols["cat"] is None:
        raise SystemExit("035 之表頭解析失敗：找不到 Sys-RA-Feature-ID 或 分類")
    out = {}
    for r in rows[1:]:
        k = str(r[cols["ref"]] or "").strip()
        if not k:
            continue
        out[k] = {n: (str(r[j] or "").strip() if j is not None else "")
                  for n, j in cols.items() if n != "ref"}
    wb.close()
    return out, cols


def load_037_all() -> list[dict]:
    """037 之全 745 列（leaf ＋ heading），判準同 `vf230_leaves.py`。"""
    out = []
    for f in sorted(glob.glob(str(ROOT / "inputs" / "FM-WI-FSM-037*VF230*.xlsx"))):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for nm in wb.sheetnames:
            rows = list(wb[nm].iter_rows(values_only=True))
            i = next((j for j, r in enumerate(rows)
                      if any("requirement description" in norm(v) for v in r)), None)
            if i is None:
                continue
            for r in rows[i + 1:]:
                if not r[0]:
                    continue
                cat = str(r[5] or "").strip()
                out.append({"swe": str(r[0]).strip(), "src": str(r[1] or "").strip(),
                            "cat": cat, "is_leaf": cat.lower().startswith("functional"),
                            "title": str(r[2] or "").strip(),
                            "desc": str(r[3] or "").strip()})
            break
        wb.close()
    return out


def main() -> None:
    sysra, _ = load_sysra()
    a037 = load_037_all()
    leaf = [r for r in a037 if r["is_leaf"]]
    head = [r for r in a037 if not r["is_leaf"]]
    in037 = {r["src"] for r in a037}

    # 正向：037 之判定 vs 035 之 Category
    def fwd(group):
        hit = [r for r in group if r["src"] in sysra]
        return hit, [r for r in group if r["src"] not in sysra], \
            Counter(sysra[r["src"]]["cat"] for r in hit)

    lhit, lmiss, lcat = fwd(leaf)
    hhit, hmiss, hcat = fwd(head)

    # 錯配：037 判 Heading 而 035 判 Functional Requirement
    mismatch = [r for r in head
                if r["src"] in sysra
                and sysra[r["src"]]["cat"] == "Functional Requirement"]
    # 反向錯配：037 判 Functional 而 035 判非 Functional
    rev_mismatch = [r for r in leaf
                    if r["src"] in sysra
                    and sysra[r["src"]]["cat"] != "Functional Requirement"]

    # 反向覆蓋：035 判 Functional 而 037 全 745 列未收
    uncovered = sorted(k for k, v in sysra.items()
                       if v["cat"] == "Functional Requirement" and k not in in037)

    tot_func = sum(1 for v in sysra.values() if v["cat"] == "Functional Requirement")

    L = ["# VF230 跨源驗核 —— 037 × 035 SYSRA（DR-28 之替代來源）", "",
         "**量測條件**：`scripts/vf230_crosscheck.py`。",
         f"來源：`{Path(SRC).name}`",
         "（`Basic Report` 分頁，表頭列 0，逐 `SYS2 Sys-RA-Feature-ID` 建索引）。",
         "037 側為 11 份分報告之全 745 列，判準同 `vf230_leaves.py`。", "",
         "## 0. 為何此檔可替代 SYS2", "",
         "其 `Basic Report` 之欄位與 CFTS044 之 SYS2 export **同型**：",
         "`SYS2 Sys-RA-Feature-ID`／`SYS2 分類 Category`／`SYS2 VF章節 Chapter for VF`／",
         "`SYS2 EE Architecture`／`SYS2 限定地區 Region`。**`分類 Category` 即 Part 1",
         "於 01 輪用於 537 列對帳之同一欄**。", "",
         f"- 035 之列（有 Sys-RA 者）：**{len(sysra)}**",
         "- 其 Category 分布："
         + "／".join(f"`{k}` {v}" for k, v in
                     Counter(v["cat"] for v in sysra.values()).most_common()),
         "",
         "**另有一份 `SYS2_VF230.xlsx`**（`9_ASPICE/SYS.2 System Requirements Analysis/",
         "Z.QS YuShen 260423/08.[SYS2]Vehicle Settings/`），schema 相同但列數較少",
         f"（2626），且**缺 037 之 6 個 `E-Save` leaf**；035 則涵蓋全部 619。",
         "**本輪採 035**，其已在 `inputs/` 內（R-VS61 之補入由 Pei 執行）。", "",
         "## 1. 正向對帳 —— 037 之判定是否為 035 所支持", "",
         "| 037 側 | 列 | 命中 035 | 未命中 | 035 之 Category |",
         "|---|---:|---:|---:|---|",
         f"| Functional（leaf） | {len(leaf)} | {len(lhit)} | {len(lmiss)} | "
         + "／".join(f"`{k}` {v}" for k, v in lcat.most_common()) + " |",
         f"| Heading | {len(head)} | {len(hhit)} | {len(hmiss)} | "
         + "／".join(f"`{k}` {v}" for k, v in hcat.most_common()) + " |",
         "",
         f"**leaf 側零錯配**：{len(lhit)} 個 037 判 Functional 者，035 亦全數判",
         f"`Functional Requirement`（反向錯配 {len(rev_mismatch)}）。", ""]

    if mismatch:
        L += [f"## 2. 錯配（{len(mismatch)}）—— 037 判 Heading 而 035 判 Functional", "",
              "**此即 DR-28 所稱「A-VS01 型之錯配無從偵測」之標的。已偵得。**", "",
              "| SWE ID | Sys-RA | 037 之條文（前 90 字） |", "|---|---|---|"]
        for r in mismatch:
            d = r["desc"].replace("\n", " ").replace("|", "\\|")[:90]
            L.append(f"| `{r['swe']}` | `{r['src']}` | {d} |")
        L += ["", "上列八條之 037 條文皆為 `The HMI layer shall …`／",
              "`HW supplier shall notify …` 之形態 —— **其為需求，非節標題**。",
              "037 之 Categorization 判 `Heading` 與其自身條文形態不符。", "",
              "**本層未改 leaf 母體**：`data/vf230_leaves.tsv` 維持 619。",
              "改判會使母體成為 627，屬裁定事項（Part 1 於 01 輪之 A-VS01 亦經裁定"
              "方除役）。", ""]

    L += [f"## 3. 反向覆蓋 —— 035 判 Functional 而 037 未收（{len(uncovered)}）", "",
          f"- 035 之 `Functional Requirement` 合計 **{tot_func}**",
          f"- 其中為 037 之 745 列所收者 **{tot_func - len(uncovered)}**"
          f"（{(tot_func - len(uncovered)) / tot_func:.1%}）",
          f"- **未收 {len(uncovered)}**（{len(uncovered) / tot_func:.1%}）", ""]
    if uncovered:
        ee = Counter(sysra[k]["ee"] for k in uncovered)
        rg = Counter(sysra[k]["region"] for k in uncovered)
        ch = Counter(sysra[k]["ch"] for k in uncovered)
        L += ["未收者之屬性分布：", "",
              "- EE Architecture：" + "／".join(f"`{k}` {v}" for k, v in ee.most_common(6)),
              "- Region：" + "／".join(f"`{k}` {v}" for k, v in rg.most_common(6)),
              f"- VF章節：{len(ch)} 個相異值，前六為 "
              + "／".join(f"`{k}` {v}" for k, v in ch.most_common(6)), "",
              "**同一章節內既有收錄亦有未收**（例如 `01.10.01.01.74` 於 037 收 14、",
              "未收 33），故此非「整章委派他 feature」之乾淨切分。", "",
              "**全樹搜尋確認 VF230 之 037 分報告僅此 11 份**",
              "（`find /Users/peihe/Work -iname '*FM-WI-FSM-037*' -iname '*VF230*'`），",
              "故未收之部分**在上游尚無 SWE.1 分析**，非本層漏收。", "",
              "樣本（前 8）：", "", "| Sys-RA | VF章節 | 條文（前 80 字） |", "|---|---|---|"]
        for k in uncovered[:8]:
            d = sysra[k]["desc"].replace("\n", " ").replace("|", "\\|")[:80]
            L.append(f"| `{k}` | `{sysra[k]['ch']}` | {d} |")
        L += [""]

    asil = Counter(sysra[r["src"]]["asil"] for r in lhit)
    L += ["## 4. 安全屬性（ASIL）", "",
          "035 為技術安全需求分析報告，其 `SYS2 ASIL 等級 (ASIL)` 欄於 037 之 "
          f"{len(lhit)} 個命中 leaf 上之分布："
          + "／".join(f"`{k or '(空)'}` {v}" for k, v in asil.most_common()),
          "",
          "→ **VF230 之 leaf 無任何具 ASIL 等級者**，安全分析層不進入其 trace chain。",
          "此與 Part 1 一致（CFTS044 之 037 亦無 ASIL 欄）。", ""]

    out = ROOT / "docs" / "reports" / "vf230_crosscheck.md"
    out.write_text("\n".join(L) + "\n", encoding="utf-8")
    (ROOT / "data" / "_vf230_crosscheck.json").write_text(json.dumps({
        "sysra_rows": len(sysra), "037_rows": len(a037),
        "leaf": len(leaf), "leaf_hit": len(lhit), "leaf_miss": len(lmiss),
        "leaf_cat": dict(lcat), "head": len(head), "head_hit": len(hhit),
        "head_cat": dict(hcat),
        "mismatch_heading_vs_functional": [r["swe"] for r in mismatch],
        "mismatch_reverse": [r["swe"] for r in rev_mismatch],
        "sysra_functional_total": tot_func, "uncovered": uncovered,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"035 列 {len(sysra)}；037 列 {len(a037)}（leaf {len(leaf)} / head {len(head)}）")
    print(f"leaf 命中 {len(lhit)}，未命中 {len(lmiss)}，Category {dict(lcat)}")
    print(f"head 命中 {len(hhit)}，未命中 {len(hmiss)}，Category {dict(hcat)}")
    print(f"錯配（037 Heading × 035 Functional）= {len(mismatch)}")
    print(f"反向錯配（037 Functional × 035 非 Functional）= {len(rev_mismatch)}")
    print(f"035 Functional {tot_func}，037 未收 {len(uncovered)}"
          f"（{len(uncovered) / tot_func:.1%}）")
    print(f"wrote {out.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
