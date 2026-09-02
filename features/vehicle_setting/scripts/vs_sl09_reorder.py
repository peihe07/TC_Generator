#!/usr/bin/env python3
"""VS-SL-09 —— R-G42 一(a) 重排＋F 欄重賦（vf230＋BL；VC 不動）。

  §1  資料列依修後 `req_key(D)` **穩定排序**，整列搬移（全欄同動）
  §2  F 欄依新列序重賦；對照表重出為端到端（`old_F` = SL-06 之前之原值）

**B 欄之處置（兩本不同，見上繳 §5）**：
  BL   B 欄為列相對之自參照公式 `=IF(ISBLANK($D<row>),"",ROW()-9)` ——
       逐字搬移會指向錯列，故**依新列位重寫**（樣式不變、語義不變）。
  vf230 B 欄為字面序號（244…），依包內「整列搬移」隨列走，**不重編**。
"""

from __future__ import annotations

import csv
import hashlib
import re
import shutil
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[3] / "scripts"))
import vs_sl01_dryrun as v1  # noqa: E402
import vs_sl07_assign as A  # noqa: E402
from lint_delivery_spec import req_key  # noqa: E402

ROOT = v1.ROOT
COL_NO, COL_REQ, COL_E, COL_TC = 2, 4, 5, 6
BL_FORMULA = '=IF(ISBLANK($D{row}),"",ROW()-9)'
TARGETS = {"vehicle_setting": "VS", "bed_lowering": "BLM"}
MAP_COLS = ["row", "req_id", "old_E", "old_F", "new_F", "status", "note"]
MANIFEST_COLS = ["filename", "sha256", "source_path", "round", "status", "note"]


def sha256_of(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def row_hash(vals: list, skip: set[int]) -> str:
    """列內容之指紋（1-based 欄號於 `skip` 者不計）。"""
    body = "\x1f".join("" if v is None else str(v)
                       for i, v in enumerate(vals, 1) if i not in skip)
    return hashlib.sha256(body.encode("utf-8")).hexdigest()


def run(slug: str) -> dict:
    cfg = A.BOOKS[slug]
    book = ROOT / cfg["book"]
    bak = book.with_suffix(".presl09.bak.xlsx")
    if not bak.exists():                                   # 不覆寫既有備份
        shutil.copy2(book, bak)

    ws_ro, g_ro = v1.cells(book)
    rows = v1.data_rows(ws_ro, g_ro)                       # data_only 判列
    wbk = openpyxl.load_workbook(book)
    ws = wbk[v1.SHEET]
    ncol = ws.max_column

    payload = [[ws.cell(r, c).value for c in range(1, ncol + 1)] for r in rows]
    # 內容比對排除 F 欄（本包重賦），BL 另排除 B 欄（其自參照公式依新列位重寫）。
    skip = {COL_TC} | ({COL_NO} if slug == "bed_lowering" else set())
    before = sorted(row_hash(p, skip) for p in payload)
    old_f_by_cur = {str(p[COL_TC - 1] or "").strip(): None for p in payload}

    # 端到端之 old_F：自 SL-07 對照表以其 new_F 反查
    for m in csv.DictReader(open(ROOT / cfg["map"]), delimiter="\t"):
        if m["new_F"] in old_f_by_cur:
            old_f_by_cur[m["new_F"]] = m["old_F"]

    order = sorted(range(len(payload)),
                   key=lambda i: req_key(str(payload[i][COL_REQ - 1] or "").strip()))
    payload = [payload[i] for i in order]                  # sorted() 為穩定排序

    abbr = TARGETS[slug]
    recs = []
    for n, (r, vals) in enumerate(zip(rows, payload), 1):
        cur_f = str(vals[COL_TC - 1] or "").strip()
        new_f = f"NR1L-{abbr}-{n:03d}"
        vals[COL_TC - 1] = new_f
        if slug == "bed_lowering":
            vals[COL_NO - 1] = BL_FORMULA.format(row=r)    # 自參照公式改寫至新列
        for c, v in enumerate(vals, 1):
            ws.cell(r, c).value = v
        recs.append({"row": r, "req_id": str(vals[COL_REQ - 1] or "").strip(),
                     "old_E": str(vals[COL_E - 1] or "").strip(),
                     "old_F": old_f_by_cur.get(cur_f) or "",
                     "new_F": new_f,
                     "status": "assigned" if slug == "vehicle_setting" else "renamed",
                     "note": "VS-SL-09 重排後依新列序重賦（R-G42 一(a)）"})
    wbk.save(book)

    after = sorted(row_hash([ws.cell(r, c).value for c in range(1, ncol + 1)], skip)
                   for r in rows)
    # B 欄之改寫另立判準：每列須為其自身列位之公式
    b_ok = (all(ws.cell(r, COL_NO).value == BL_FORMULA.format(row=r) for r in rows)
            if slug == "bed_lowering" else True)

    if cfg["removed"]:
        for x in csv.DictReader(open(ROOT / cfg["removed"]), delimiter="\t"):
            recs.append({"row": x["row"], "req_id": x["req_id"], "old_E": "",
                         "old_F": "", "new_F": "", "status": "RETIRED",
                         "note": f"{x['reason']}（VS-SL-03 移除，未入轉正本）"})

    with open(ROOT / cfg["map"], "w", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MAP_COLS, delimiter="\t")
        w.writeheader()
        w.writerows(recs)

    sha = sha256_of(book)
    man = book.parent / "MANIFEST.tsv"
    mrows = list(csv.DictReader(open(man), delimiter="\t"))
    for m in mrows:
        if m["filename"] == book.name:
            m["sha256"], m["round"] = sha, "VS-SL-09"
            m["note"] = "R-G42 一(a) 重排＋F 欄重賦；出貨 gate 未過"
    with open(man, "w", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MANIFEST_COLS, delimiter="\t")
        w.writeheader()
        w.writerows(mrows)

    D = [str(p[COL_REQ - 1] or "").strip() for p in payload]
    K = [req_key(x) for x in D]
    breaks = sum(1 for i in range(1, len(K)) if K[i] < K[i - 1])
    ids = [r["new_F"] for r in recs if r["status"] != "RETIRED"]
    return {"slug": slug, "rows": len(rows), "breaks": breaks, "sha": sha,
            "map_rows": len(recs), "ids": ids,
            "content_same": before == after, "b_ok": b_ok}


def main() -> int:
    for slug in TARGETS:
        s = run(slug)
        nums = [int(i.rsplit("-", 1)[1]) for i in s["ids"]]
        assert s["breaks"] == 0, f"{slug}：一(a) 逆序 {s['breaks']} 處，非 0"
        assert s["content_same"], f"{slug}：重排前後列內容集合不同"
        assert s["b_ok"], f"{slug}：B 欄自參照公式未對齊其列位"
        assert len(set(s["ids"])) == len(s["ids"]), f"{slug}：F 有重號"
        assert nums == list(range(1, len(nums) + 1)), f"{slug}：號非連號"
        print(f"{slug:<18} {s['rows']:>4} 列  一(a) 逆序 {s['breaks']}  "
              f"F {s['ids'][0]}…{s['ids'][-1]}  對照表 {s['map_rows']:>4}  "
              f"內容集合相同 {s['content_same']}  B 欄對齊 {s['b_ok']}")
        print(f"{'':<18} SHA {s['sha'][:32]}…")
    return 0


if __name__ == "__main__":
    sys.exit(main())
