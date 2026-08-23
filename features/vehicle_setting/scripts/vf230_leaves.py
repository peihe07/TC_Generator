"""VF230 leaf 母體建置（61 包 §5.2 之工單；W 號待改，見 A-VS 撞號登記）。

11 份 037 分報告逐份解析，**每份分別報列數與 leaf 數，不先合併**；
合併後之總數自各份重算（canon §5a）。

leaf 判準沿用 recon.py 之既有實作：Categorization 正規化後
以 "functional" 起首者為 leaf，其餘為 heading／other。
不引入 ID 後綴啟發式（R-C3 明禁）。

**R-VF16（Pei 裁定 2026-08-23）—— 母體為 627，非 619**：
037 判 `Heading` 而 035 SYSRA 判 `Functional Requirement` 之 8 列
（A-VS132）計入可測 leaf。該 8 列以 `source_disagreement=1` 標記，
**不得靜默併入** —— 本層之 Categorization 於該 8 列刻意偏離 037，
須可分辨。其餘 619 列標 `0`。

分頁以**版面**定位（掃描各分頁尋找含 `Requirement Description` 之表頭列），
不以分頁名定位：11 份中 2 份之分頁名為 `Sheet1` 而非 `Analysis Report`，
其版面與其餘 9 份逐欄相同（W-103 實測）。

輸出：
  data/vf230_leaves.tsv          欄位與 data/leaves.tsv 同（swe_id/family/src_ref/title/desc）
  data/_vf230_w103.json          逐份統計與 Categorization 分布
"""
import glob
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent


def norm(v) -> str:
    """表頭比對用之正規化：小寫、空白摺疊。"""
    return re.sub(r"\s+", " ", str(v or "")).strip().lower()


def family_of(path: Path) -> str:
    """自檔名取族群標籤（`SWRA` 之後、副檔名之前之逐字片段）。"""
    stem = path.stem
    i = stem.rfind("SWRA")
    return stem[i + len("SWRA"):].lstrip("_ ").strip() if i >= 0 else stem


def parse_one(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet, rows, hdr = None, None, None
    for name in wb.sheetnames:
        cand = list(wb[name].iter_rows(values_only=True))
        i = next((j for j, r in enumerate(cand)
                  if any("requirement description" in norm(v) for v in r)), None)
        if i is not None:
            sheet, rows, hdr = name, cand, i
            break
    if hdr is None:
        raise SystemExit(f"{path.name}: 各分頁皆無表頭列"
                         f"（無 'Requirement Description' 儲存格）：{wb.sheetnames}")
    header = rows[hdr]

    def find(*need, forbid=()):
        hits = [j for j, v in enumerate(header)
                if all(t in norm(v) for t in need)
                and not any(t in norm(v) for t in forbid)]
        return hits[0] if len(hits) == 1 else None

    cat_i = find("categorization", forbid=("sub",))
    src_i = find("source requirement id")
    title_i = find("requirement title")
    desc_i = find("requirement description")
    if None in (cat_i, src_i, title_i, desc_i):
        raise SystemExit(f"{path.name}: 表頭欄解析失敗 "
                         f"cat={cat_i} src={src_i} title={title_i} desc={desc_i}")

    fam = family_of(path)
    data_rows, leaves, headings = 0, [], []
    cat_dist: Counter = Counter()
    for r in rows[hdr + 1:]:
        if not r[0]:
            continue
        data_rows += 1
        rid = str(r[0]).strip()
        cat = str(r[cat_i] or "").strip()
        cat_dist[cat or "(blank)"] += 1
        rec = {"swe_id": rid, "family": fam, "disagree": "0",
               "src_ref": str(r[src_i] or "").strip(),
               "title": str(r[title_i] or "").strip(),
               "desc": str(r[desc_i] or "").strip()}
        (leaves if cat.lower().startswith("functional") else headings).append(rec)

    wb.close()
    return {"file": path.name, "family": fam, "sheet": sheet, "header_row": hdr,
            "data_rows": data_rows, "leaves": leaves, "headings": headings,
            "cat_dist": dict(cat_dist),
            "cat_col": cat_i, "src_col": src_i}


def main() -> None:
    files = sorted(Path(p) for p in glob.glob(str(ROOT / "inputs" / "FM-WI-FSM-037*VF230*.xlsx")))
    if len(files) != 11:
        raise SystemExit(f"預期 11 份 037，實得 {len(files)} 份")

    per_file = [parse_one(p) for p in files]

    # --- R-VF16：8 列之偏離，錨點先行（R-VF11）---
    dis = json.loads((ROOT / "data" / "_vf230_crosscheck.json")
                     .read_text(encoding="utf-8"))
    must_hit = set(dis["mismatch_heading_vs_functional"])       # 必命中，8
    if len(must_hit) != 8:
        raise SystemExit(f"R-VF16 之偏離集應為 8，實得 {len(must_hit)}")
    all_head = {r["swe_id"] for d in per_file for r in d["headings"]}
    must_miss = all_head - must_hit                            # 必不命中，118
    print(f"R-VF11 錨點：必命中 {len(must_hit)}／必不命中 {len(must_miss)}")
    if must_hit & must_miss:
        raise SystemExit("錨點集相交，判準有誤")
    if len(must_miss) != 118:
        raise SystemExit(f"必不命中錨點應為 118，實得 {len(must_miss)}")

    promoted = []
    for d in per_file:
        keep = []
        for rec in d["headings"]:
            if rec["swe_id"] in must_hit:
                rec = rec | {"disagree": "1"}
                d["leaves"].append(rec)
                promoted.append(rec["swe_id"])
            else:
                keep.append(rec)
        d["headings"] = keep
    print(f"R-VF16 提列 {len(promoted)} 列為 leaf")
    if set(promoted) != must_hit:
        raise SystemExit("錨點實測不符：提列集與必命中集不等，停")

    print(f"{'family':46} {'sheet':>16} {'rows':>5} {'leaf':>5} {'head':>5} {'other':>5}")
    for d in per_file:
        other = d["data_rows"] - len(d["leaves"]) - len(d["headings"])
        print(f"{d['family'][:46]:46} {d['sheet'][:16]:>16} {d['data_rows']:5} "
              f"{len(d['leaves']):5} {len(d['headings']):5} {other:5}")

    # canon §5a —— 總數自各份重算，不取合併後之 len()
    tot_rows = sum(d["data_rows"] for d in per_file)
    tot_leaf = sum(len(d["leaves"]) for d in per_file)
    tot_head = sum(len(d["headings"]) for d in per_file)
    print(f"{'TOTAL (自各份重算)':>60} {tot_rows:5} {tot_leaf:5} {tot_head:5}")

    all_leaves = [rec for d in per_file for rec in d["leaves"]]
    assert len(all_leaves) == tot_leaf, "合併後筆數與逐份加總不符"

    ids = Counter(r["swe_id"] for r in all_leaves)
    dupes = {k: v for k, v in ids.items() if v > 1}
    print(f"\n相異 swe_id = {len(ids)}；跨份重複 = {len(dupes)}")
    if dupes:
        for k, v in sorted(dupes.items())[:20]:
            print(f"  DUP {k} x{v}")

    cat_all: Counter = Counter()
    for d in per_file:
        cat_all.update(d["cat_dist"])
    print("\nCategorization 分布（全體）：")
    for k, v in cat_all.most_common():
        print(f"  {v:5}  {k}")

    out = ROOT / "data" / "vf230_leaves.tsv"
    cols = ["swe_id", "family", "src_ref", "title", "desc", "disagree"]
    with out.open("w", encoding="utf-8") as fh:
        fh.write("\t".join(cols) + "\n")
        for r in all_leaves:
            fh.write("\t".join(str(r.get(c, "")).replace("\t", " ")
                                 .replace("\n", "\\n") for c in cols) + "\n")
    print(f"\nwrote {out.relative_to(ROOT)}  ({len(all_leaves)} leaf)")

    meta = {"files": [{k: d[k] for k in
                       ("file", "family", "sheet", "header_row", "data_rows", "cat_dist")}
                      | {"leaf": len(d["leaves"]), "heading": len(d["headings"])}
                      for d in per_file],
            "total_rows": tot_rows, "total_leaf": tot_leaf,
            "total_heading": tot_head, "distinct_swe_id": len(ids),
            "source_disagreement": sorted(must_hit),
            "cross_file_dupes": dupes}
    (ROOT / "data" / "_vf230_w103.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
