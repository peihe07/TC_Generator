"""W-112（63 包 §7）—— Comfort 素材之畫面層對照表（R-VS59(2) 之前置）。

對每個 `delegate ∈ {yes, pending}` 之 leaf，自 Comfort 素材查其畫面層內容，
產 `docs/reports/screen_source.tsv`：

    leaf_id / comfort_leaf_ids / 畫面層內容之逐字節錄 / 來源檔與列 / 查無者標 PENDING

`delegation_lookup.tsv` 之 `comfort_leaf_ids` 為 R-VS7(a)′ 之**功能群層級**指名
（Layer 3 群，非單一 leaf id）；本檔以該群名回查 Comfort 037 之
`Requirement Title` 與 `Requirement Description`。

**必列**：查得／查無兩數。**查無者即 R-VS59(4) 之 PENDING 標的。**
"""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

import openpyxl

FEAT = Path(__file__).resolve().parents[1]
COMFORT = FEAT.parent / "comfort" / "inputs"
REPORT = COMFORT / ("FM-WI-FSM-037-A03-N1L-SWE1-Comfort-HMI-V0.1 STLA 報告.xlsx")
SETTINGS = COMFORT / "HMI Settings List R1 SR25 Post R1L-R (Feb 13 2026).xlsx"
POPUP = COMFORT / "Pop Up List HMI R1 SR24 Post 2A (Dec 15, 2023).xlsx"

# 本 feature 之 Layer 3 → Comfort 037 之檢索詞（實體功能名）
TERMS = {
    "LeftFrontHeatedSeat": ["heated seat", "seat heat"],
    "RightFrontHeatedSeat": ["heated seat", "seat heat"],
    "OneStageHeatedSeat": ["heated seat", "seat heat"],
    "TwoStagesHeatedSeat": ["heated seat", "seat heat"],
    "ThreeStagesHeatedSeat": ["heated seat", "seat heat"],
    "LeftFrontVentedSeat": ["vented seat", "ventilated seat", "seat vent"],
    "RightFrontVentedSeat": ["vented seat", "ventilated seat", "seat vent"],
    "TwoStagesVentedSeatsManagement": ["vented seat", "ventilated seat", "seat vent"],
    "ThreeStagesVentedSeatsManagement": ["vented seat", "ventilated seat", "seat vent"],
    "HeatedSteeringWheel": ["heated steering", "steering wheel heat"],
    "HeatedSteeringWheelManagement": ["heated steering", "steering wheel heat"],
    "SwitchLHD/RHDConfiguration": ["driver side", "left hand drive", "right hand drive"],
    "Stop-StartSystem": ["stop", "start"],
    "StopStartSystemBehavior": ["stop", "start"],
    "ThirdRowHeadrestDump": ["head restraint", "headrest"],
    "ScreenOFF": ["screen off", "display off"],
    "FeaturesEnableCriteria": [],
    "PHEVFeatures": [],
}


def comfort_rows() -> list[dict]:
    ws = openpyxl.load_workbook(REPORT, read_only=True, data_only=True)["Analysis Report"]
    out = []
    for i, row in enumerate(ws.iter_rows(min_row=8, values_only=True), start=8):
        if not row[0]:
            continue
        out.append({"row": i, "id": str(row[0]).strip(),
                    "title": str(row[3] or "").strip(),
                    "desc": str(row[4] or "").strip(),
                    "cat": str(row[6] or "").strip()})
    return out


def sheet_hits(path: Path, terms: list[str], limit: int = 2) -> list[str]:
    """於 HMI 清單／彈窗清單中找含檢索詞之列（僅取前 `limit` 列作佐證）。"""
    if not terms or not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for name in wb.sheetnames:
        for i, row in enumerate(wb[name].iter_rows(values_only=True), start=1):
            text = " ".join(str(c) for c in row if c is not None)
            low = text.casefold()
            if any(t in low for t in terms):
                out.append(f"{path.name}／{name}:{i}｜{text[:110]}")
                if len(out) >= limit:
                    return out
    return out


def main() -> None:
    lookup = list(csv.DictReader(
        (FEAT / "docs/reports/delegation_lookup.tsv").open(encoding="utf-8"), delimiter="\t"))
    crows = comfort_rows()
    func = [r for r in crows if r["cat"].casefold().startswith("functional")]

    out, found, missing = [], 0, 0
    cache: dict[str, list[str]] = {}
    for r in lookup:
        if r["delegate"] not in ("yes", "pending"):
            continue
        l3 = r["layer3"]
        terms = TERMS.get(l3, [])
        hits = [c for c in func
                if any(t in (c["title"] + " " + c["desc"]).casefold() for t in terms)]
        if l3 not in cache:
            cache[l3] = sheet_hits(SETTINGS, terms) + sheet_hits(POPUP, terms)
        extra = cache[l3]
        if hits:
            found += 1
            ids = ";".join(sorted({h["id"] for h in hits})[:8])
            quote = " ⏐ ".join(f"{h['id']}：{h['desc'][:150]}" for h in hits[:2])
            src = ";".join(f"{REPORT.name}／Analysis Report:{h['row']}" for h in hits[:2])
            if extra:
                src += ";" + ";".join(extra)
            status = "found"
        else:
            missing += 1
            ids, quote, status = "", "PENDING", "PENDING"
            src = ";".join(extra)
        out.append({"leaf_id": r["leaf_id"], "layer3": l3, "delegate": r["delegate"],
                    "status": status, "comfort_leaf_ids": ids,
                    "screen_text": quote, "source": src,
                    "comfort_hit_count": len(hits)})

    p = FEAT / "docs/reports/screen_source.tsv"
    with p.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(out[0]), delimiter="\t", lineterminator="\n")
        w.writeheader()
        w.writerows(out)

    print(f"標的（`delegate ∈ {{yes, pending}}`）：{len(out)}")
    print(f"**查得 {found}／查無 {missing}**")
    import collections
    print("\n逐 Layer 3：")
    by = collections.defaultdict(lambda: [0, 0])
    for r in out:
        by[r["layer3"]][0 if r["status"] == "found" else 1] += 1
    for k, (a, b) in sorted(by.items()):
        print(f"  {k:34s} 查得 {a:3d}／查無 {b:3d}")


if __name__ == "__main__":
    main()
