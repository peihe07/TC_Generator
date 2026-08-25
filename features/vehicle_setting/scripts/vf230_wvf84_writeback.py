"""VF230：036 之**實寫**（W-VF84；授權見 R-VF124，其二條件已成就）。

**其與 `writeback_036.py` 之別**：後者之 `BOOK` 寫死且指向 **CFTS044**
（A-VF30；`CROSSLINE.md` 已登錄其跨線風險）。本檔之目標取自 `feature.yaml`
之 `paths.workbook_vf230`，且**只寫 repo 內副本**（R-VF112）。

**四道閘，任一不成即停**：
  R-VF114  B 欄起始號由實測導出（二本 CFTS044 末號取最大 +1），與裁定值交叉驗證
  R-VF115  **寫入前實測目標之資料列；非 0 即停手回報**
  R-VF113  x14 之保全：openpyxl 存檔後補回 `<extLst>`，**補回後複驗 x14dv 之數**
  W-VF84   寫入後**重讀該檔**與預覽表逐欄比對（438 × 14），差異即停

**R-VF116**：實寫後之 Excel 確認為 Pei 之動作，**本檔於完成時明示提示，不靜默結束**。
"""
from __future__ import annotations

import csv
import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl

FEAT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(FEAT / "scripts"))
import vf230_wvf77_dryrun as DRY          # noqa: E402  單一權威：欄映射與列之來源

SHEET, HEADER_ROW, FIRST_DATA_ROW = DRY.SHEET, DRY.HEADER_ROW, DRY.FIRST_DATA_ROW
COL_IDX = {c: openpyxl.utils.column_index_from_string(c) for c in DRY.COLS}


def probe_x14(p: Path) -> int:
    z = zipfile.ZipFile(p)
    n = sum(len(re.findall(r"x14:dataValidation[ >]",
                           z.read(f).decode("utf-8", errors="replace")))
            for f in z.namelist()
            if f.startswith("xl/worksheets/") and f.endswith(".xml"))
    z.close()
    return n


def restore_extlst(src: Path, dst: Path) -> None:
    """R-VF113：openpyxl 存檔會移除 x14；自原檔補回 `<extLst>`。"""
    # **命名空間須一併補回** —— `<extLst>` 內用 `x14:`／`xm:` 前綴，
    # 而 openpyxl 重寫根元素時**丟棄了其 `xmlns:` 宣告**；
    # 只補 `<extLst>` 而不補宣告，其 XML 為 `unbound prefix` 而不合法。
    # **首版即如此，於 W-VF86.1 之空跑被閘 4 之 XML 解析攔下。**
    # **其亦更正本層先前之宣稱**：上繳 V45 §三稱替代路徑「已驗其形」，
    # 而其時只驗了 `x14dv` 之元素數與 openpyxl 可讀（其解析較寬容），
    # **未驗 XML 之合法性** —— 該宣稱過強。本函式現以嚴格 parser 自驗。
    import xml.etree.ElementTree as ET

    z = zipfile.ZipFile(src)
    ext, nsdecl = {}, {}
    for n in z.namelist():
        if not (n.startswith("xl/worksheets/") and n.endswith(".xml")):
            continue
        x = z.read(n).decode("utf-8", errors="replace")
        m = re.search(r"<extLst>.*?</extLst>", x, re.S)
        if m:
            ext[n] = m.group(0)
            root = re.search(r"<worksheet\b[^>]*>", x)
            nsdecl[n] = dict(re.findall(r'(xmlns:[A-Za-z0-9_]+)="([^"]+)"',
                                        root.group(0) if root else ""))
    z.close()
    if not ext:
        return
    tmp = dst.with_suffix(".tmp.xlsx")
    zin, zout = zipfile.ZipFile(dst), zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED)
    for it in zin.infolist():
        data = zin.read(it.filename)
        if it.filename in ext:
            x = data.decode("utf-8")
            if "<extLst>" not in x:
                m = re.search(r"<worksheet\b[^>]*>", x)
                if m:
                    tag = m.group(0)
                    add = "".join(f' {k}="{v}"' for k, v in nsdecl[it.filename].items()
                                  if k not in tag)
                    if add:
                        x = x.replace(tag, tag[:-1] + add + ">", 1)
                x = x.replace("</worksheet>", ext[it.filename] + "</worksheet>")
                try:
                    ET.fromstring(x)
                except ET.ParseError as e:
                    raise SystemExit(f"x14 補回後之 XML 不合法（{it.filename}）：{e}"
                                     " —— 停，不寫出")
                data = x.encode("utf-8")
        zout.writestr(it, data)
    zin.close()
    zout.close()
    tmp.replace(dst)


def main() -> None:
    book = DRY.book()
    print(f"=== W-VF84 實寫（授權 R-VF124；目標為 **repo 內副本**）===")
    print(f"目標：{book}")

    b_start, seen = DRY.cfts044_last_b()
    print(f"\n[閘 1／R-VF114] B 欄起始號：repo {seen.get('repo')}／交付 "
          f"{seen.get('delivery')} → **{b_start}**（裁定 {DRY.B_START_RULED}）")
    if b_start != DRY.B_START_RULED:
        raise SystemExit("B 欄起始號之實測與裁定不符，停")

    rows = DRY.rows_from(DRY.sources(), b_start)
    print(f"          來源 {len(rows)} 列，B {rows[0]['B']}–{rows[-1]['B']}")

    # ---- 閘 2（R-VF115）：分辨「公式」與「資料」----
    # **首版以非 `data_only` 讀取，將 237 格公式計為「非空」而停手**（上繳 V57 §三）。
    # 其停手為正 —— 它揭露了 B 欄之公式模型（`R-VF129` 因而撤銷 `R-VF83`／`R-VF114`）。
    # **`R-VF132` 明令覆寫該 237 格公式**，故本閘須能分辨二者。
    # **其修正有裁定為據，非逕行放寬**（R-VF130 之判準：
    # 「修正後本次所攔者是否仍會被攔？」答否 —— **而該放寬經 `R-VF132` 裁定**）。
    # **二者皆量，僅「資料」使其停。**
    wb = openpyxl.load_workbook(book)
    ws = wb[SHEET]
    n_formula = n_data = 0
    for r in ws.iter_rows(min_row=FIRST_DATA_ROW):
        vals = [c.value for c in r if c.value not in (None, "")]
        if not vals:
            continue
        if all(isinstance(v, str) and v.lstrip().startswith("=") for v in vals):
            n_formula += 1
        else:
            n_data += 1
    print(f"\n[閘 2／R-VF115] 目標之列：**資料 {n_data}**｜公式 {n_formula}")
    if n_data != 0:
        raise SystemExit(f"目標有 {n_data} 列**資料** —— 停手回報。"
                         "**有人在本層之外寫過，或本檔已寫入。**")
    print(f"          （{n_formula} 格公式將依 R-VF132 被覆寫；"
          f"**其自動編號機制自此失效**）")

    x14_before = probe_x14(book)
    print(f"[閘 3／R-VF113] 寫入前 x14dv：{x14_before}")

    overwritten = 0
    for i, r in enumerate(rows):
        row_no = FIRST_DATA_ROW + i
        for c, v in r.items():
            if c.startswith("_") or c not in COL_IDX:
                continue
            if not str(v).strip():
                continue
            cell = ws.cell(row=row_no, column=COL_IDX[c])
            if (c == "B" and isinstance(cell.value, str)
                    and cell.value.lstrip().startswith("=")):
                overwritten += 1
            cell.value = v
    print(f"          B 欄覆寫之公式格數：**{overwritten}**（預期 {n_formula}）")
    if overwritten != n_formula:
        raise SystemExit(f"B 欄覆寫之格數（{overwritten}）與實測之公式格數"
                         f"（{n_formula}）不符，停")
    wb.save(book)
    wb.close()
    x14_saved = probe_x14(book)
    print(f"          存檔後 x14dv：{x14_saved}")
    restore_extlst(BACKUP, book)
    x14_after = probe_x14(book)
    print(f"          補回後 x14dv：**{x14_after}**（原 {x14_before}）")
    if x14_after != x14_before:
        raise SystemExit("x14 之保全失敗（補回後之數與原檔不符），停")

    # [閘 4] 重讀比對
    wb2 = openpyxl.load_workbook(book, read_only=True, data_only=True)
    ws2 = wb2[SHEET]
    got = {}
    for row in ws2.iter_rows(min_row=FIRST_DATA_ROW, values_only=False):
        b = row[1].value
        if b in (None, ""):
            continue
        got[int(b)] = {c: ("" if row[COL_IDX[c] - 1].value is None
                           else str(row[COL_IDX[c] - 1].value)) for c in DRY.COLS}
    wb2.close()
    diff = []
    for r in rows:
        g = got.get(r["B"])
        if g is None:
            diff.append((r["B"], "整列缺"))
            continue
        for c in DRY.COLS:
            if str(r[c]).strip() != g[c].strip():
                diff.append((r["B"], c))
    print(f"\n[閘 4] 重讀比對 {len(rows)} 列 × {len(DRY.COLS)} 欄："
          f"**差異 {len(diff)}**")
    for d in diff[:6]:
        print(f"          B{d[0]} 欄 {d[1]}")
    if diff:
        raise SystemExit("寫入後之重讀比對有差異，停")

    print(f"\n寫入 {len(rows)} 列，B {rows[0]['B']}–{rows[-1]['B']}。四閘全過。")
    print("\n" + "=" * 66)
    print("**R-VF116：以下為 Pei 之動作，本層不得代之**")
    print("  1. 以 **Excel** 開啟下列檔案，確認**下拉選單可用**：")
    print(f"     {book}")
    print("     （openpyxl 會移除 x14，本檔已補回其 XML —— **已驗其形，未驗其用**）")
    print("  2. 確認後，**逐字具名檔名**複製至交付路徑：")
    print("     /Users/peihe/Work/…/Vehicle Settings/VF230_V1_R5/")
    print("     **不得以萬用字元或目錄複製** —— CFTS044 二本現已分岔"
          "（237／243 列），誤覆蓋會使 Part 1 之 238–243 六列消失")
    print("=" * 66)


if __name__ == "__main__":
    BACKUP = Path(str(FEAT / "inputs" / "_vf230_036_prewrite_backup.xlsx"))
    _b = DRY.book()
    if not BACKUP.exists():
        shutil.copy2(_b, BACKUP)
        print(f"已備份寫入前之本 → {BACKUP.name}")
    main()
