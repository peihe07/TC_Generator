#!/usr/bin/env python3
"""VS-SL-03 §2 之硬規 lint（對沙盒稿）。

三項：`test_item` 括號下半、行尾句號、設定項名之雙引號。
列母體取自 v3 報告（非以 `max_row` 掃，避免讀到公式列）。
"""

from __future__ import annotations

import csv
import re
import sys
import warnings
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import vs_sl01_dryrun as v1  # noqa: E402

warnings.filterwarnings("ignore")
ROOT = v1.ROOT

BOOKS = {
    "vf230": ("features/vehicle_setting/sandbox/vssl/vf230_vssl.xlsx",
              "features/vehicle_setting/reports/vf230_settings_dryrun_v3.tsv"),
    "bl": ("features/bed_lowering/sandbox/vssl/bl_vssl.xlsx",
           "features/bed_lowering/reports/bl_settings_dryrun_v3.tsv"),
    "vc": ("features/vehicle_category/sandbox/vssl/vc_vssl.xlsx",
           "features/vehicle_category/reports/vc_settings_dryrun_v3.tsv"),
}

PAREN_TAIL = re.compile(r"\([^)]{5,}\)\s*$", re.S)
# 設定項名之樣式：逐詞須為大寫起首或數字（`with`／`and`／`or`／`for` 除外），
# 以免把 `the setting named in the test data` 之類的泛稱句誤判為未加引號之名。
NAME = r"(?:[A-Z][A-Za-z0-9\-/&]*)(?:\s+(?:[A-Z0-9][A-Za-z0-9\-/&]*|with|and|or|for))*"
UNQUOTED = re.compile(rf"\bthe\s+(?!\")({NAME})\s+setting\b")
UNQUOTED_ER = re.compile(rf"^The\s+(?!\")({NAME})\s+setting\b")


def rows_of(rep: str, removed: set[int]) -> list[int]:
    return [int(r["row"]) for r in csv.DictReader(open(ROOT / rep), delimiter="\t")
            if int(r["row"]) not in removed]


def check(tag: str, removed: set[int]) -> dict:
    book, rep = BOOKS[tag]
    ws = openpyxl.load_workbook(ROOT / book)[v1.SHEET]
    g = lambda r, c: ("" if ws.cell(r, c).value is None else str(ws.cell(r, c).value))
    # 刪列後列號位移：以 D 欄比對回原列
    by_req = {}
    for r in range(v1.HEADER_ROW + 1, ws.max_row + 1):
        if g(r, 4).strip():
            by_req.setdefault(g(r, 4).strip(), r)
    plan = list(csv.DictReader(open(ROOT / rep), delimiter="\t"))

    no_paren, dot, noq = [], [], []
    for p in plan:
        src = int(p["row"])
        if src in removed:
            continue
        req = None
        for k, vv in by_req.items():
            pass
        r = _locate(ws, g, src, removed)
        if r is None:
            continue
        if not PAREN_TAIL.search(g(r, 9).strip()):
            no_paren.append(r)
        for c in (10, 12, 13):
            if any(x.strip().endswith(".") for x in g(r, c).split("\n") if x.strip()):
                dot.append(r)
                break
        for c in (12, 13):
            for ln in g(r, c).split("\n"):
                if UNQUOTED.search(ln) or UNQUOTED_ER.search(ln):
                    noq.append(r)
                    break
            else:
                continue
            break
    return {"tag": tag, "rows": len(plan) - len(removed & {int(p['row']) for p in plan}),
            "no_paren": no_paren, "dot": dot, "noq": noq}


def _locate(ws, g, src_row: int, removed: set[int]) -> int | None:
    """刪列後之新列號 = 原列號 − 其前被刪之列數。"""
    shift = sum(1 for x in removed if x < src_row)
    r = src_row - shift
    return r if g(r, 2).strip() or g(r, 4).strip() else None


def main() -> int:
    rm = set()
    p = ROOT / "features/vehicle_setting/reports/vf230_removed_non_nafta.tsv"
    if p.exists():
        rm = {int(x["row"]) for x in csv.DictReader(open(p), delimiter="\t")}
    ok = True
    for tag in BOOKS:
        removed = rm if tag == "vf230" else set()
        res = check(tag, removed)
        bad = len(res["no_paren"]) + len(res["dot"]) + len(res["noq"])
        ok &= bad == 0
        print(f"{tag:6} 列 {res['rows']:4} | test_item 無括號下半 {len(res['no_paren']):4} "
              f"| 尾句號 {len(res['dot']):4} | 設定項未加雙引號 {len(res['noq']):4} "
              f"| {'PASS' if bad == 0 else 'FAIL'}")
        for k in ("no_paren", "dot", "noq"):
            if res[k]:
                print(f"       {k}: {res[k][:8]}{' …' if len(res[k]) > 8 else ''}")
    print("\n總判：", "PASS" if ok else "FAIL")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
