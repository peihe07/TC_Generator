#!/usr/bin/env python3
"""Vehicle Settings 設定項查找配方（下放包 VS-SL-01 §2 任務 1）。

輸入一個設定項顯示名，輸出四段 JSON：`path` / `control` / `proxi` / `coopen`。

來源（皆唯讀，**本檔不呼叫 openpyxl save**）：
  Settings List  `features/vehicle_setting/inputs/HMI Settings List R1 SR25 Post R1L-R (Feb 13 2026).xlsx`
                 分頁 `Settings`（A/B/C 路徑、D 控件、E 選項、G Notes）
                 分頁 `Brand-Specific Names`（B 欄 = Jeep / Chrysler / Ram / Dodge）
  总控表         `forms/R1L FIP 总控表 V1.1.0.xlsx` 分頁 `FeatureSet(Gen4-5)`
                 只讀 E 欄（Atlantis／DT）——R-VS{live+3}，`new HW (637)` 與 PNet 不展開
  raw 值         `features/vehicle_setting/data/_vf230_proxi_values.json`

市場變體一律 NAFTA（R-VS{live}）：Settings 同名多列時取 Notes 含 NAFTA 或
無市場標記之列；Notes 標 EMEA／MASAH／MASPN／LATAM 者不取。

PROXI 前置之來源優先序（配方細則）：
  (1) 需求原文之 `$var$ = [label]` 條件式  → 主來源
  (2) 总控表 Atlantis 條件式              → 補充（AND 複合、co-open、原文未載之附加條件）
  (3) 二者皆無                            → `PENDING: DR-{n}`，不猜值
`Always false` 者 (2) 為空，只依 (1)（R-VS{live+1}）。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

SETTINGS_XLSX = ("features/vehicle_setting/inputs/"
                 "HMI Settings List R1 SR25 Post R1L-R (Feb 13 2026).xlsx")
ATLANTIS_XLSX = "forms/R1L FIP 总控表 V1.1.0.xlsx"
ATLANTIS_SHEET = "FeatureSet(Gen4-5)"
PROXI_VALUES = "features/vehicle_setting/data/_vf230_proxi_values.json"

# Notes 欄之市場標記；NAFTA 取，其餘不取（R-VS{live}(1)）
MARKETS = ("NAFTA", "EMEA", "MASAH", "MASPN", "LATAM")
NON_NAFTA_MARKETS = tuple(m for m in MARKETS if m != "NAFTA")

CATEGORY_RE = re.compile(r"^\s*\d+\.\s+(\S.*)$", re.S)
LEVEL1_RE = re.compile(r"^\s*\d+\s*$")
LEVEL2_RE = re.compile(r"^\s*\d+\.\d+(\s+or\s+\d+\.\d+)?\s*$")
LEVEL3_RE = re.compile(r"^\s*(\d+\.\d+\.\d+)\s*(.*)$", re.S)
# 需求原文之條件式：`Param = [label]`、`Param is [label]`
REQ_COND_RE = re.compile(r'[$"”\'\s(]([A-Za-z][A-Za-z0-9_ ]{2,60}?)["”\']?\s*(?:=|is)\s*\[([^\]]+)\]')
# 既有 TC 之 PROXI 舊形制：`PROXI $Param$ is set to "label"`
PROXI_OLD_RE = re.compile(r'PROXI\s+\$([^$]+)\$\s+is set to\s+"([^"]*)"')


def _clean(value) -> str:
    """儲存格轉字串；None 視為空字串，去除首尾空白與不斷行空格。"""
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def _norm(name: str) -> str:
    """比名用之正規化：去 `*`、去括號註、小寫、壓空白。"""
    s = re.sub(r"\([^)]*\)", " ", name.replace("*", " "))
    s = re.sub(r"[^0-9a-z]+", " ", s.lower())
    return " ".join(s.split())


def split_options(text: str) -> list[str]:
    """E 欄選項切分。分隔符實測不一致（`/`、`,`、` / `），切後 strip，保留原大小寫。"""
    if not text:
        return []
    body = text.replace("\n", " / ")
    parts = re.split(r"[/,]", body) if ("/" in body or "," in body) else [body]
    return [p.strip() for p in parts if p.strip()]


# --------------------------------------------------------------------------
# Settings List
# --------------------------------------------------------------------------

def load_settings(root: Path) -> list[dict]:
    """讀 `Settings` 分頁，攤平為設定項清單，逐列帶其 A/B/C 路徑。"""
    ws = openpyxl.load_workbook(root / SETTINGS_XLSX, data_only=True)["Settings"]
    items: list[dict] = []
    category = ""
    parent = ""
    for r in range(4, ws.max_row + 1):
        a, b, c = (_clean(ws.cell(r, i).value) for i in (1, 2, 3))
        template, options, notes = (_clean(ws.cell(r, i).value) for i in (4, 5, 7))
        if not any((a, b, c, template, options, notes)):
            continue

        m = CATEGORY_RE.match(a)
        if m and not LEVEL1_RE.match(a):
            # 分類標題列：`4. Safety & Driving Assistance`。括號註為顯示條件，不入路徑。
            category = re.sub(r"\s*\([^)]*\)\s*$", "", m.group(1)).strip()
            parent = ""
            continue

        if LEVEL1_RE.match(a) and b:
            # 第一層項目。D 欄為 `>` 者為純容器；**D 欄為空者亦可能是容器**
            # （實測 `21. Aux Switches` 之 `Aux 1`–`Aux 6` r565–r588 即此形），
            # 故兩者皆令其名成為後續 `N.M` 列之 parent；只有 `>` 者不自成一項。
            name = b
            parent = name.replace("*", "").strip() if template in (">", "") else ""
            if template == ">":
                continue
            items.append(_item(r, category, "", name, template, options, notes))
            continue

        if LEVEL2_RE.match(b) and c:
            items.append(_item(r, category, parent, c, template, options, notes))
            continue

        m3 = LEVEL3_RE.match(c)
        if m3 and m3.group(2):
            items.append(_item(r, category, parent, m3.group(2).strip(),
                               template, options, notes))
            continue

        if items and not any((a, b, c)) and (template or options or notes):
            # 續列（同一項之另一段 Notes／Graphics），併回前一項
            items[-1]["notes_extra"].append(notes)
    return items


def _item(row: int, category: str, parent: str, name: str,
          template: str, options: str, notes: str) -> dict:
    return {
        "row": row, "category": category, "parent": parent,
        "name": name.strip(), "name_plain": name.replace("*", "").strip(),
        "brand_specific": "*" in name,
        "template": template, "options_raw": options,
        "options": split_options(options),
        "notes": notes, "notes_extra": [],
        "market": _market_of(notes),
    }


def _market_of(notes: str) -> str:
    """Notes 欄之市場標記；無標記回 `""`（依 R-VS{live}(1) 亦屬可取）。"""
    for m in MARKETS:
        if re.search(rf"\b{m}\b", notes, re.I):
            return m.upper()
    return ""


def load_brand_names(root: Path) -> dict[str, str]:
    """`Brand-Specific Names` B 欄（Jeep / Chrysler / Ram / Dodge）—— DT／HDCC27 = RAM。"""
    ws = openpyxl.load_workbook(root / SETTINGS_XLSX,
                                data_only=True)["Brand-Specific Names"]
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        jeep = _clean(ws.cell(r, 2).value)
        if jeep:
            out.setdefault(_norm(jeep), jeep)
    return out


def pick_nafta(rows: list[dict]) -> tuple[dict | None, str]:
    """同名多列時依 R-VS{live}(1) 取列。回 (選中列, 說明)。"""
    if not rows:
        return None, "NO_ROW"
    nafta = [x for x in rows if x["market"] == "NAFTA"]
    if nafta:
        return nafta[0], f"NAFTA 標記列 r{nafta[0]['row']}"
    plain = [x for x in rows if not x["market"]]
    if len(plain) == 1:
        return plain[0], f"無市場標記單列 r{plain[0]['row']}"
    if plain:
        return plain[0], ("VARIANT_UNRESOLVED：無市場標記者 "
                          f"{len(plain)} 列 {[x['row'] for x in plain]}，取首列")
    return None, ("VARIANT_UNRESOLVED：僅有非 NAFTA 列 "
                  f"{[(x['row'], x['market']) for x in rows]}")


# --------------------------------------------------------------------------
# 总控表 Atlantis
# --------------------------------------------------------------------------

def load_atlantis(root: Path) -> list[dict]:
    """讀 `FeatureSet(Gen4-5)` 之 A/B/C/D/E 欄；只取 E（Atlantis／DT）。"""
    ws = openpyxl.load_workbook(root / ATLANTIS_XLSX, data_only=True)[ATLANTIS_SHEET]
    out = []
    for r in range(4, ws.max_row + 1):
        no = _clean(ws.cell(r, 1).value)
        if not no:
            continue
        expr = _clean(ws.cell(r, 5).value)
        out.append({
            "no": no, "row": r,
            "request": _clean(ws.cell(r, 2).value),
            "feature_type": _clean(ws.cell(r, 3).value),
            "name": _clean(ws.cell(r, 4).value),
            "expr": expr, "parsed": parse_expr(expr),
        })
    return out


def parse_expr(expr: str) -> dict:
    """总控表條件式之解析。實測四類皆須處理，另有無法歸類者記 `OTHER`。"""
    text = " ".join(expr.split())
    low = text.lower()
    if not text:
        return {"kind": "EMPTY", "terms": [], "returns": None}
    if low.startswith("always true"):
        return {"kind": "ALWAYS_TRUE", "terms": [], "returns": True}
    if low.startswith("always false"):
        return {"kind": "ALWAYS_FALSE", "terms": [], "returns": False}

    terms = _parse_terms(text)
    m = re.search(r"return value (?:is|true|false)\s*(\d+)?", text, re.I)
    ret = m.group(1) if m and m.group(1) else None
    if ret is not None:
        return {"kind": "NON_BOOL", "terms": terms, "returns": ret}
    if not re.match(r"\s*if\b", text, re.I):
        return {"kind": "OTHER", "terms": terms, "returns": None}
    kind = "IF_AND" if len(terms) > 1 else "IF_SIMPLE"
    return {"kind": kind, "terms": terms, "returns": True}


def _dedup_terms(terms: list[dict]) -> list[dict]:
    """同 param 合併其 label（保序），避免同一參數重複成列。"""
    out: dict[str, dict] = {}
    for t in terms:
        key = _norm(t["param"])
        if key not in out:
            out[key] = {"param": t["param"], "labels": list(t["labels"])}
        else:
            for l in t["labels"]:
                if l not in out[key]["labels"]:
                    out[key]["labels"].append(l)
    return list(out.values())


def _parse_terms(text: str) -> list[dict]:
    """抽出 `"Param" is [A] or [B]` 之項。引號實測有 `"`／`”`／無引號三型。"""
    terms = []
    pat = re.compile(r'["”\']?([A-Za-z][A-Za-z0-9_ /()\-]{2,60}?)["”\']?\s+is\s+((?:\[[^\]]+\]\s*(?:or\s*)?)+)')
    for m in pat.finditer(text):
        labels = re.findall(r"\[([^\]]+)\]", m.group(2))
        param = re.sub(r"^(?:if|and|or|when)\s+", "", m.group(1).strip(), flags=re.I)
        terms.append({"param": param.strip(),
                      "labels": [x.strip() for x in labels]})
    return _dedup_terms(terms)


# --------------------------------------------------------------------------
# raw 值
# --------------------------------------------------------------------------

def load_proxi_values(root: Path) -> dict[str, dict[str, str]]:
    return json.loads((root / PROXI_VALUES).read_text(encoding="utf-8"))


def resolve_raw(values: dict, param: str, label: str) -> tuple[str | None, str | None]:
    """label → raw。查無回 (None, None)，由呼叫端記 RAW_MISSING（不猜值）。"""
    table = values.get(param) or _fuzzy_param(values, param)
    if not table:
        return None, None
    want = _norm(label)
    for raw, lab in table.items():
        if _norm(lab) == want:
            return raw, lab
    if want in ("set", "present", "true"):
        for raw, lab in table.items():
            if _norm(lab) in ("present", "set"):
                return raw, lab
    if want in ("not set", "absent", "false"):
        for raw, lab in table.items():
            if _norm(lab) in ("absent", "not set"):
                return raw, lab
    return None, None


def _fuzzy_param(values: dict, param: str) -> dict | None:
    """參數名以正規化比對；仍查無則回 None（同 R-13，不以語意相近之名代入）。"""
    want = _norm(param)
    for k, v in values.items():
        if _norm(k) == want:
            return v
    return None


def coopen_of(atlantis: list[dict], params: list[str], self_name: str) -> list[str]:
    """同一 param／node 同時開啟之其他設定項。"""
    wanted = {_norm(p) for p in params}
    out = []
    for rec in atlantis:
        if rec["name"] == self_name:
            continue
        for t in rec["parsed"]["terms"]:
            if _norm(t["param"]) in wanted:
                out.append(rec["name"])
                break
    return sorted(set(out))


# --------------------------------------------------------------------------
# 查找主體
# --------------------------------------------------------------------------

class Lookup:
    """設定項查找器。建構時載入三來源，之後 `query()` 逐名查。"""

    def __init__(self, root: Path):
        self.root = root
        self.settings = load_settings(root)
        self.brands = load_brand_names(root)
        self.atlantis = load_atlantis(root)
        self.values = load_proxi_values(root)
        self._by_name: dict[str, list[dict]] = {}
        for it in self.settings:
            self._by_name.setdefault(_norm(it["name_plain"]), []).append(it)
        self._atl_by_name: dict[str, list[dict]] = {}
        for rec in self.atlantis:
            self._atl_by_name.setdefault(_norm(rec["name"]), []).append(rec)

    # -- 各段 ------------------------------------------------------------
    def hmi_rows(self, name: str) -> list[dict]:
        return self._by_name.get(_norm(name), [])

    def atlantis_rows(self, name: str) -> list[dict]:
        return self._atl_by_name.get(_norm(name), [])

    def query(self, name: str, req_text: str = "", dr_no: str = "DR-49") -> dict:
        """回 path／control／proxi／coopen 四段，另附 flags 與 evidence。"""
        flags: list[str] = []
        rows = self.hmi_rows(name)
        item, note = pick_nafta(rows)
        if note.startswith("VARIANT_UNRESOLVED") or (rows and item is None):
            flags.append("VARIANT_UNRESOLVED")

        path = None
        control = None
        if item:
            path = ["Settings", item["category"]] + \
                   ([item["parent"]] if item["parent"] else []) + [item["name_plain"]]
            control = {"template": item["template"], "options": item["options"]}
            if item["brand_specific"] and _norm(item["name_plain"]) not in self.brands:
                flags.append("BRAND_NAME_UNVERIFIED")

        proxi, source, pflags = self.proxi_of(name, req_text, dr_no)
        flags.extend(pflags)
        params = [p["param"] for p in proxi if p.get("param")]

        return {
            "name": name,
            "path": path,
            "control": control,
            "proxi": proxi,
            "proxi_source": source,
            "coopen": coopen_of(self.atlantis, params, name) if params else [],
            "flags": sorted(set(flags)),
            "evidence": {
                "settings_rows": [x["row"] for x in rows],
                "settings_pick": note,
                "atlantis_no": [r["no"] for r in self.atlantis_rows(name)],
            },
        }

    def proxi_of(self, name: str, req_text: str,
                 dr_no: str) -> tuple[list[dict], str, list[str]]:
        """依來源優先序求 PROXI 前置清單。"""
        flags: list[str] = []
        atl = self.atlantis_rows(name)
        always_false = any(r["parsed"]["kind"] == "ALWAYS_FALSE" for r in atl)
        if always_false:
            flags.append("ALWAYS_FALSE")

        # (1) 需求原文
        terms = [{"param": re.sub(r"^(?:if|and|or|when)\s+", "", p.strip(), flags=re.I).strip(),
                   "labels": [l.strip()]}
                  for p, l in REQ_COND_RE.findall(req_text or "")]
        terms = _dedup_terms(terms)
        source = "req" if terms else ""

        # (2) 总控表；`Always false` 者不提供條件（R-VS{live+1}）
        if not terms and atl and not always_false:
            for rec in atl:
                if rec["parsed"]["terms"]:
                    terms = rec["parsed"]["terms"]
                    source = f"atlantis No.{rec['no']}"
                    break
        elif terms and atl and not always_false:
            known = {_norm(t["param"]) for t in terms}
            for rec in atl:
                extra = [t for t in rec["parsed"]["terms"] if _norm(t["param"]) not in known]
                if extra:
                    terms = terms + extra
                    source = f"req + atlantis No.{rec['no']}"
                    break

        if not terms:
            return [{"pending": f"PENDING: {dr_no}"}], "none", flags + ["PROXI_PENDING"]

        out = []
        for t in terms:
            for label in t["labels"]:
                raw, lab = resolve_raw(self.values, t["param"], label)
                if raw is None:
                    flags.append("RAW_MISSING")
                out.append({"param": t["param"], "raw": raw,
                            "label": lab or label})
            if len(t["labels"]) > 1:
                flags.append("OR_VALUE")
        return out, source, flags


def format_proxi(proxi: list[dict]) -> str:
    """R-VS{live+2}：`PROXI <Param> = <raw> (<label>)`；無 `$`、無 `is set to`。"""
    parts = []
    for p in proxi:
        if p.get("pending"):
            parts.append(p["pending"])
        elif p["raw"] is None:
            parts.append(f"PROXI {p['param']} = ? ({p['label']})")
        else:
            parts.append(f"PROXI {p['param']} = {p['raw']} ({p['label']})")
    return " ; ".join(parts)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--root", default=".")
    ap.add_argument("--name", help="設定項顯示名")
    ap.add_argument("--req", default="", help="需求原文（供來源優先序 (1)）")
    ap.add_argument("--list-settings", action="store_true", help="列出攤平後之設定項")
    args = ap.parse_args()

    lk = Lookup(Path(args.root).resolve())
    if args.list_settings:
        for it in lk.settings:
            print(f"{it['row']}\t{it['category']}\t{it['parent']}\t{it['name']}"
                  f"\t{it['template']}\t{'|'.join(it['options'])}\t{it['market']}")
        return 0
    if not args.name:
        ap.error("需 --name 或 --list-settings")
    res = lk.query(args.name, args.req)
    res["proxi_text"] = format_proxi(res["proxi"])
    print(json.dumps(res, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
