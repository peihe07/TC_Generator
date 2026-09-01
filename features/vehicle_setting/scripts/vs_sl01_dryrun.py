#!/usr/bin/env python3
"""VS-SL-01 §2 任務 2／3 之產生器：別名表草稿與三份 dry-run 報告。

**唯讀**：只讀工作簿與 forms，不呼叫 openpyxl save，不動 `delivered/`。
輸出：
  features/vehicle_setting/data/settings_alias.tsv
  features/vehicle_setting/reports/vf230_settings_dryrun.tsv ＋ .md
  features/bed_lowering/reports/bl_settings_dryrun.tsv
  features/vehicle_category/reports/vc_settings_dryrun.tsv
"""

from __future__ import annotations

import re
import sys
import warnings
from collections import Counter, OrderedDict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from settings_lookup import Lookup, _norm, format_proxi  # noqa: E402

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[3]
DR_NO = "DR-49"
SHEET = "Test Case Specification 測試用例規範"
HEADER_ROW = 9

VF230 = ("features/vehicle_setting/inputs/FM-WI-FSM-036-A01 STLA 測試用例規範與結果_"
         "SWQT STLA Test Case Specification & Result_SWQT_VF230_20260819.xlsx")
BL = ("features/bed_lowering/output/FM-WI-FSM-036-A01 STLA 測試用例規範與結果_"
      "SWQT STLA Test Case Specification & Result_SWQT_BedLowering_20260827.xlsx")
VC = ("features/vehicle_category/output/FM-WI-FSM-036-A01 STLA 測試用例規範與結果_"
      "SWQT STLA Test Case Specification & Result_SWQT_VehicleCategory_20260827_working.xlsx")

SETTING_RE = re.compile(r'"([^"]+)"\s+customer setting', re.I)
PROXI_OLD_RE = re.compile(r'PROXI\s+\$([^$]+)\$\s+is set to\s+"([^"]*)"')
STEP_RE = re.compile(r'(?:Press|Select|Open|Tap)\s+"([^"]+)"')
NEG_RE = re.compile(r'\$?([A-Za-z][A-Za-z0-9_]*)\$?\s*!=\s*((?:\[[^\]]+\][,\s]*(?:or\s*)?)+)')

def is_negative(item: str) -> bool:
    """負向 TC：需求原文為 `!=` 或 `shall not display`。"""
    return bool(NEG_RE.search(item)) or "shall not display" in item


def propose_proxi(lk, res: dict, item: str, proxi_now: str) -> tuple[list[dict], list[str]]:
    """依 §4 定 PROXI 提議：負向取 raw 0 (Absent)；OR 列舉取本列之值，兄弟另立。

    回 (proxi 清單, 附註)。不猜值 —— Absent 標籤查無者維持原提議並記 RAW_MISSING。
    """
    proxi, notes = res["proxi"], []
    if not proxi or proxi[0].get("pending"):
        return proxi, notes

    if is_negative(item):
        out = []
        for p in proxi:
            raw, lab = None, None
            table = lk.values.get(p["param"])
            if table:
                for k, v in table.items():
                    if v.strip().lower() in ("absent", "not set"):
                        raw, lab = k, v
                        break
            out.append({"param": p["param"], "raw": raw, "label": lab or "Absent"})
            if raw is None:
                notes.append(f"RAW_MISSING：{p['param']} 無 Absent 標籤")
        seen, dedup = set(), []
        for p in out:
            if p["param"] not in seen:
                seen.add(p["param"])
                dedup.append(p)
        return dedup, notes + ["負向 TC 依 §4 取 Absent"]

    by_param: dict[str, list[dict]] = {}
    for p in proxi:
        by_param.setdefault(p["param"], []).append(p)
    out = []
    for param, plist in by_param.items():
        if len(plist) == 1:
            out.append(plist[0])
            continue
        cur = [l for pp, l in PROXI_OLD_RE.findall(proxi_now) if pp.strip() == param]
        pick = next((x for x in plist if cur and x["label"] in cur[0]), plist[0])
        out.append(pick)
        sib = "／".join(x["raw"] or "?" for x in plist)
        notes.append(f"EP 兄弟：{param} raw {sib} 各一條")
    return out, notes


COLUMNS = ["row", "tc_id", "setting", "alias_status", "path_now", "path_proposed",
           "control_proposed", "proxi_now", "proxi_proposed", "flags"]


def cells(path: Path):
    ws = openpyxl.load_workbook(path, data_only=True)[SHEET]
    return ws, (lambda r, c: "" if ws.cell(r, c).value is None else str(ws.cell(r, c).value))


def data_rows(ws, g) -> list[int]:
    """資料列 = 表頭之下、B 或 D 欄非空者。"""
    return [r for r in range(HEADER_ROW + 1, ws.max_row + 1)
            if g(r, 2).strip() or g(r, 4).strip()]


def setting_of(g, r: str) -> str:
    text = "\n".join(g(r, c) for c in (9, 10, 11, 12, 13))
    m = SETTING_RE.search(text)
    return m.group(1) if m else ""


GENERIC_OPEN_RE = re.compile(r"Open the Vehicle Settings menu", re.I)


def path_now_of(g, r) -> str:
    """現行 procedure 之導覽路徑。實測多數列只有泛稱之開啟句，無逐層路徑。"""
    steps = [s for s in STEP_RE.findall(g(r, 12)) if "customer setting" not in s]
    if steps:
        return " > ".join(steps)
    if GENERIC_OPEN_RE.search(g(r, 12)):
        return "（無逐層路徑：Open the Vehicle Settings menu）"
    return ""


def proxi_now_of(g, r) -> str:
    hits = PROXI_OLD_RE.findall(g(r, 10))
    return " ; ".join(f'PROXI ${p}$ is set to "{l}"' for p, l in hits)


# --------------------------------------------------------------------------
# 任務 2：別名表
# --------------------------------------------------------------------------

def collect_names(g, rows) -> "OrderedDict[str, int]":
    names: OrderedDict[str, int] = OrderedDict()
    for r in rows:
        text = "\n".join(g(r, c) for c in (9, 10, 11, 12, 13))
        for m in SETTING_RE.findall(text):
            names[m] = names.get(m, 0) + 1
    return names


def _tight(s: str) -> str:
    """空白不敏感之比對鍵（`Park Sense` ↔ `ParkSense`），只用於 manual 候選。"""
    return _norm(s).replace(" ", "")


def build_alias(lk: Lookup, names) -> list[dict]:
    hmi_tight: dict[str, list[dict]] = {}
    for it in lk.settings:
        hmi_tight.setdefault(_tight(it["name_plain"]), []).append(it)
    fip_tight: dict[str, list[dict]] = {}
    for rec in lk.atlantis:
        fip_tight.setdefault(_tight(rec["name"]), []).append(rec)

    out = []
    for name in names:
        hmi_rows = lk.hmi_rows(name)
        fip_rows = lk.atlantis_rows(name)
        hmi_name = hmi_rows[0]["name_plain"] if hmi_rows else ""
        fip_name = fip_rows[0]["name"] if fip_rows else ""
        ev = []
        if hmi_rows:
            ev.append(f"Settings r{','.join(str(x['row']) for x in hmi_rows)}")
        if fip_rows:
            ev.append(f"总控表 No.{','.join(x['no'] for x in fip_rows)}")

        if hmi_rows or fip_rows:
            match = "exact"
        else:
            cand_h = hmi_tight.get(_tight(name), [])
            cand_f = fip_tight.get(_tight(name), [])
            if cand_h or cand_f:
                match = "manual"
                hmi_name = cand_h[0]["name_plain"] if cand_h else ""
                fip_name = cand_f[0]["name"] if cand_f else ""
                if cand_h:
                    ev.append(f"Settings r{','.join(str(x['row']) for x in cand_h)}（空白不敏感）")
                if cand_f:
                    ev.append(f"总控表 No.{','.join(x['no'] for x in cand_f)}（空白不敏感）")
            else:
                match = "UNRESOLVED"
                ev.append(f"母體無對應，開 {DR_NO}")
        out.append({"tc_name": name, "hmi_name": hmi_name, "fip_name": fip_name,
                    "match_type": match, "evidence": "；".join(ev)})
    return out


# --------------------------------------------------------------------------
# 任務 3：dry-run
# --------------------------------------------------------------------------

def dryrun_vf230(lk: Lookup, alias_by_tc: dict) -> list[dict]:
    ws, g = cells(ROOT / VF230)
    rows = data_rows(ws, g)
    out = []
    for r in rows:
        item = g(r, 9)
        name = setting_of(g, r)
        alias = alias_by_tc.get(name, {"match_type": "UNRESOLVED", "hmi_name": ""})
        target = alias.get("hmi_name") or name
        res = lk.query(target, item, DR_NO) if target else {
            "path": None, "control": None, "proxi": [], "flags": ["ALIAS_UNRESOLVED"]}

        flags = set(res["flags"])
        if alias["match_type"] == "UNRESOLVED":
            flags.add("ALIAS_UNRESOLVED")
        elif alias["match_type"] == "manual":
            flags.add("ALIAS_MANUAL")
        pn = proxi_now_of(g, r)
        if any(re.search(r"\bor\b", l, re.I) for _, l in PROXI_OLD_RE.findall(pn)):
            flags.add("OR_VALUE")
        if "non-NAFTA" in item or "non-NAFTA" in g(r, 4):
            flags.add("NON_NAFTA")
        if _neg_contra(item, pn):
            flags.add("NEG_CONTRA")
        if path_now_of(g, r).startswith("（無逐層路徑"):
            flags.add("PATH_ABSENT")
        proposed, pnotes = propose_proxi(lk, res, item, pn)
        if any(n.startswith("RAW_MISSING") for n in pnotes):
            flags.add("RAW_MISSING")
        if any(n.startswith("EP 兄弟") for n in pnotes):
            flags.add("EP_SIBLING")

        out.append({
            "row": r, "tc_id": g(r, 6) or "(F 欄空)", "setting": name,
            "alias_status": alias["match_type"],
            "path_now": path_now_of(g, r),
            "path_proposed": " > ".join(res["path"]) if res["path"] else "",
            "control_proposed": (f"{res['control']['template']}"
                                 f" [{' / '.join(res['control']['options'])}]"
                                 if res["control"] else ""),
            "proxi_now": pn,
            "proxi_proposed": (format_proxi(proposed) + ("｜" + "；".join(pnotes) if pnotes else "")
                               if proposed else ""),
            "flags": ";".join(sorted(flags)),
        })
    return out


def _neg_contra(item: str, proxi_now: str) -> bool:
    """`!=` 之 TC，其 Pre 卻把 PROXI 設為列舉內之正值。"""
    m = NEG_RE.search(item)
    if not m:
        return False
    excluded = {_norm(x) for x in re.findall(r"\[([^\]]+)\]", m.group(2))}
    return any(_norm(l) in excluded for _, l in PROXI_OLD_RE.findall(proxi_now))


PROSE_MAP = {
    # BedLowering
    "the vehicle is equipped with the air suspension system":
        ("总控表 No.36 Bed Lowering Mode", "PROXI CAN node 27 (ASM/ASCM) = 1 (Present)", "BODY_TYPE_CONFLICT"),
    "the vehicle configuration is either dt or dj/d2":
        ("Body_Types", "PROXI Body_Types = 7 (Type 7 - DT) 或 4 (Type 4 - DJ) 或 1 (Type 1 - D2)", "BODY_TYPE_CONFLICT"),
    "the vehicle is a dt configuration":
        ("Body_Types", "PROXI Body_Types = 7 (Type 7 - DT)", "BODY_TYPE_CONFLICT"),
    "the vehicle is a dj/d2 configuration":
        ("Body_Types", "PROXI Body_Types = 4 (Type 4 - DJ) 或 1 (Type 1 - D2)", ""),
    # VehicleCategory
    "the vehicle is equipped with the vehicle category feature":
        ("", f"PENDING: {DR_NO}", "NO_MAPPING"),
    "the glove box feature is activated with a known 4-digit pin":
        ("Glove_Box_Soft_Button", "PROXI Glove_Box_Soft_Button = 1 (Present)", ""),
    "the glove box feature is not activated":
        ("Glove_Box_Soft_Button", "PROXI Glove_Box_Soft_Button = 1 (Present)", ""),
    "the vehicle is equipped with an electrochromic controls item":
        ("总控表 No.24 Mirror Dimmer",
         "PROXI EC_Mirror = 1 (Present) ; PROXI EC_Mirror_Hard_Button_Present = 0 (Absent)", ""),
    "the vehicle is equipped with a headrest fold controls item":
        ("总控表 No.18 Headrest Fold", f"PENDING: {DR_NO}", "ALWAYS_FALSE"),
    "the vehicle is equipped with the camera app":
        ("总控表 No.196 Cam App", f"PENDING: {DR_NO}", "ALWAYS_FALSE"),
}


def dryrun_prose(path: str, tag: str) -> list[dict]:
    """BL／VC：入口不走 Settings List，只套 proxi 段（VS-SL-01 §5）。"""
    ws, g = cells(ROOT / path)
    out = []
    for r in data_rows(ws, g):
        pre = g(r, 10)
        lines = [re.sub(r"^\s*\d+\.\s*", "", x).strip() for x in pre.split("\n")]
        lines = [x for x in lines if x]
        proposed, flags, src = [], set(), []
        for ln in lines:
            key = _norm(ln).replace(" ", " ")
            hit = PROSE_MAP.get(" ".join(ln.lower().split()).rstrip("."))
            if hit:
                ref, prop, flag = hit
                proposed.append(prop)
                if ref:
                    src.append(ref)
                if flag:
                    flags.add(flag)
                flags.add("PROSE_PRECOND")
            elif re.search(r"\bequipped with\b|\bconfiguration is\b", ln, re.I):
                flags.add("PROSE_PRECOND")
                flags.add("NO_MAPPING")
                proposed.append(f"NO_MAPPING: {ln}")
        if tag == "BL":
            flags.add("BODY_TYPE_CONFLICT")
        out.append({
            "row": r, "tc_id": g(r, 6) or "(F 欄空)", "setting": g(r, 8) or g(r, 7),
            "alias_status": "n/a（不走 Settings List）",
            "path_now": path_now_of(g, r), "path_proposed": "（本 feature 不套 path 段）",
            "control_proposed": "", "proxi_now": proxi_now_of(g, r) or "（無 PROXI）",
            "proxi_proposed": " ; ".join(dict.fromkeys(proposed)),
            "flags": ";".join(sorted(flags)) + (f"｜來源 {'；'.join(dict.fromkeys(src))}" if src else ""),
        })
    return out


def write_tsv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    body = ["\t".join(COLUMNS)]
    body += ["\t".join(str(r[c]).replace("\t", " ").replace("\n", " | ") for c in COLUMNS)
             for r in rows]
    path.write_text("\n".join(body) + "\n", encoding="utf-8")


def main() -> int:
    lk = Lookup(ROOT)
    ws, g = cells(ROOT / VF230)
    names = collect_names(g, data_rows(ws, g))
    alias = build_alias(lk, names)
    alias_by_tc = {a["tc_name"]: a for a in alias}

    ap = ROOT / "features/vehicle_setting/data/settings_alias.tsv"
    cols = ["tc_name", "hmi_name", "fip_name", "match_type", "evidence"]
    ap.write_text("\n".join(["\t".join(cols)] +
                            ["\t".join(a[c] for c in cols) for a in alias]) + "\n",
                  encoding="utf-8")

    vf = dryrun_vf230(lk, alias_by_tc)
    write_tsv(ROOT / "features/vehicle_setting/reports/vf230_settings_dryrun.tsv", vf)
    bl = dryrun_prose(BL, "BL")
    write_tsv(ROOT / "features/bed_lowering/reports/bl_settings_dryrun.tsv", bl)
    vc = dryrun_prose(VC, "VC")
    write_tsv(ROOT / "features/vehicle_category/reports/vc_settings_dryrun.tsv", vc)

    fc = Counter(f for r in vf for f in r["flags"].split(";") if f)
    print(f"alias rows={len(alias)} "
          f"{dict(Counter(a['match_type'] for a in alias))}")
    print(f"vf230 rows={len(vf)}  bl rows={len(bl)}  vc rows={len(vc)}")
    print("vf230 flags:", dict(fc))
    return 0


if __name__ == "__main__":
    sys.exit(main())
