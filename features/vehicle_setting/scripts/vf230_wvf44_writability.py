"""W-VF44 —— VF230 之 writability 分級（V16 §5；**不生成 TC**）。

**沿用 Part 1 之分級定義與 blocker 分類碼，不新設分類**（V16 §5 第 1 項）：

  R-VS47 ＋ R-VS71   W0 無未解值／W1 部分可寫（扣 PENDING 後 ≥2 步且驗證
                     目標非該未解值）／W2 僅二類 —— (a) 條文無可測內容、
                     (b) 與他 leaf 之可測內容不可分辨
  blocker 分類碼      B4-preamble ／ B5-signal-absent ／ B6-value-absent
                     （Part 1 之 B1／B2／B3 為委派與畫面層之碼，
                      VF230 尚無委派判定，本輪不用，亦不新設）

**值域來源鏈（V16 §5 第 2 項，依 R-VF13）—— 取用順序明列**：

  1. LID 之 `Atlantis High` 欄組（R-VS67）
  2. 基線 DBC 之 `VAL_`（R-VS39／A-VS137 之補收路徑）
  3. PROXI 表（`inputs/PROXI_HDCC27_R3_20250424.xlsx` 之 `Format` 分頁）
     —— **R-VS49 於 Part 1 只涉四指定參數；VF230 引用 46 個相異 PROXI 參數**，
        其中 **35 個於該表可查得、11 個無**。無者之值域**無來源** → 標 `PENDING`。
        （首版之量測為「僅 4 個」，係掃描範圍過窄所致：只掃各分頁前 400 列、
         前 6 欄且要求逐字相等。改掃 `Format` 分頁 800 列後得 35。**已更正。**）
  4. 037 之 `Verification Criteria`／`Verification Method`（**R-VF13**）
     —— **`not clear` 之列不得為源**（R-VF13 第 4 項）；
        其處置依 R-VF15，於 TC 書寫時逐字轉錄至 Remarks

**前三者為既有鏈，第 4 為 R-VF13 所增；順序即上列，先命中者為準。**

依 R-VF21／R-VF28 附三錨點（以內容定錨）。

輸出：docs/reports/vf230_writability.tsv ＋ data/_vf230_writability.json
"""
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))
import importlib.util

_spec = importlib.util.spec_from_file_location(
    "wd", ROOT / "scripts" / "writability_driver.py")
wd = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(wd)          # 只 import，不執行其 run()／--write

import openpyxl


def vcvm() -> dict[str, dict]:
    """037 之 VC/VM 欄：swe_id -> {vc, vm}。"""
    import glob
    out = {}
    for f in sorted(glob.glob(str(ROOT / "inputs" / "FM-WI-FSM-037*VF230*.xlsx"))):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for nm in wb.sheetnames:
            rs = list(wb[nm].iter_rows(values_only=True))
            i = next((j for j, r in enumerate(rs)
                      if any("requirement description" in
                             re.sub(r"\s+", " ", str(v or "")).strip().lower()
                             for v in r)), None)
            if i is None:
                continue
            h = rs[i]

            def col(sub):
                return next((j for j, v in enumerate(h)
                             if sub in str(v or "").lower()), None)

            ivc, ivm = col("verification criteria"), col("verification method")
            for r in rs[i + 1:]:
                if r[0]:
                    out[str(r[0]).strip()] = {
                        "vc": str(r[ivc] or "").strip() if ivc is not None else "",
                        "vm": str(r[ivm] or "").strip() if ivm is not None else ""}
            break
        wb.close()
    return out


PROXI_REF = re.compile(
    r"retrieve the ([A-Za-z][A-Za-z0-9_]{2,})\s+PROXI configuration")


def pnorm(x: str) -> str:
    """PROXI 參數名之比對用正規化 —— 去非英數字並轉小寫。

    表內有 `FOA _Presence`（名中多一空格）等排版瑕疵，
    其與條文之 `FOA_Presence` 為同一參數。
    """
    return re.sub(r"[^a-z0-9]", "", str(x or "").lower())


def proxi_known() -> set[str]:
    """PROXI 表所載之參數名（`Format` 分頁前六欄），**以正規化形回傳**。

    Part 1 之 R-VS49 只涉四指定參數；VF230 引用者達 46 個，
    故此處以**表內實有者**為準，不以該四者為全集。

    **W-VF70 之二處修正**：
      (1) 首版寫死 `max_row=800`，而該分頁實有 **1060 列** ——
          **801 列以後之參數全部看不見**。DR-34 所列之 11 個「無來源」參數中，
          **9 個在 800 列之後**，其「無來源」為抽取式之截斷所造，非資料所缺。
          改為讀全表，不設上限。
      (2) 名以 `pnorm()` 比對，吸收表內之排版瑕疵（`FOA _Presence`）。
          其一併回收 1 個偽陽。
    """
    wb = openpyxl.load_workbook(ROOT / "inputs" / "PROXI_HDCC27_R3_20250424.xlsx",
                                read_only=True, data_only=True)
    out = set()
    for r in wb["Format"].iter_rows(values_only=True):
        for v in r[:6]:
            sv = str(v or "").strip()
            if re.fullmatch(r"[A-Za-z][A-Za-z0-9_ ]{2,}", sv):
                out.add(pnorm(sv))
    wb.close()
    return out


# ---- W-VF70：新增之二 blocker 碼（R-VF80 二）----
# B8：條文之訊號引用**只到 message 一層**（有 `TELEMATIC_*`／`IPC_*` 而無 `.訊號名`）。
#     首版之 `B5-signal-absent` 只認「全無訊號引用」，本形態有引用而不完整。
SIG_MSG_ONLY = re.compile(r"\b(TELEMATIC_\w+|IPC_\w+)\b(?!\.)")
# B9：PROXI 型而條文未帶值，且該參數於 PROXI 表**亦無值域**。
PROXI_CLAUSE_VALUE = [
    re.compile(r"If .{0,80}?=\s*\[([^\]]+)\]", re.I),
    re.compile(r"receives the value\s*(?:as\s*)?\[?([A-Za-z0-9_ ]+?)\]?\s*via signal",
               re.I),
]


def proxi_values() -> dict[str, dict[str, str]]:
    """PROXI 表之**值域**（`Table` 欄），鍵為 `pnorm()` 之名。"""
    src = ROOT / "data" / "_vf230_proxi_values.json"
    raw = json.loads(src.read_text(encoding="utf-8"))
    return {pnorm(k): v for k, v in raw.items()}


# R-VF75 一：無可測內容之判準。**已知集合，非全集**（R-VF71 三）。
NOTE_ONLY = re.compile(r"^\s*Note[:\uff1a]|managed in CFTS|not HMI setting", re.I)

VALUE_ASSIGN = re.compile(
    r"\b([A-Z][A-Z0-9_]{2,}(?:\.[A-Za-z]\w{2,})?)\s*==?\s*[\"'\[]?([A-Za-z0-9_ ]{1,40})")


def vcvm_domain(rec: dict) -> dict[str, set[str]]:
    """自 VC/VM 取 (訊號, 值)；`not clear` 之列一律排除（R-VF13 第 4 項）。"""
    txt = (rec.get("vc", "") + "\n" + rec.get("vm", "")) if rec else ""
    if not txt.strip() or "not clear" in txt.lower():
        return {}
    d: dict[str, set[str]] = defaultdict(set)
    for m in VALUE_ASSIGN.finditer(txt):
        sig, val = m.group(1).split(".")[-1], m.group(2).strip()
        if val:
            d[sig].add(val)
    return dict(d)


def main() -> None:
    leaves = list(csv.DictReader(
        (ROOT / "data" / "vf230_leaves.tsv").open(encoding="utf-8"), delimiter="\t"))
    names = json.loads((ROOT / "data" / "_wvf35_enum.json").read_text(encoding="utf-8"))
    ts_of = {c: t for t, cs in names["final"].items() for c in cs}
    vv = vcvm()

    high = wd.dbc_value_backfill(wd.bus_domain())   # 鏈 2 已併入鏈 1 之補收
    mid = wd.lid_column_domain()                    # 鏈 1：LID 欄組
    in_dbc = wd.dbc_signals()
    px = proxi_known()                              # 鏈 3：PROXI 表所實有者（正規化名）
    pv = proxi_values()                             # 鏈 3：其值域

    rows, chain_used = [], Counter()
    for lf in leaves:
        leaf, text = lf["swe_id"], re.sub(r"\s+", " ", lf["desc"])
        rec = vv.get(leaf, {})
        note, blocker, detail = {}, "", ""

        # --- W2(a) 之第二路徑（R-VF75 一，A-VF21）---
        # R-VS71 之 W2(a)「條文無可測內容」於首版**從未被實作** ——
        # 其 W2 僅由 B4-preamble／B5／B6 三路徑產生。
        # `E-Save-095` 之全文為「Note: … managed in CFTS 088」，
        # 無觸發、無可觀察之結果，而首版判其 W0 且其在選池內。
        if NOTE_ONLY.search(text) or len(text.strip()) < 120:
            grade, blocker = "W2", "B7-no-testable-content"
            detail = ("條文無可測內容（R-VS71 之 W2(a)）—— "
                      "註記式或過短，無觸發亦無可觀察之結果")
            rows.append({
                "leaf_id": leaf, "test_set": ts_of.get(lf["title"], ""),
                "layer3": lf["title"].replace("\\n", " "),
                "src_ref": lf["src_ref"], "writable": grade,
                "blocker_class": blocker, "blocker_detail": detail,
                "value_source": "", "dr_id": "",
                "vcvm_not_clear": "0", "disagree": lf.get("disagree", "0")})
            continue

        # W2(a) 之第一路徑：適用性前言（B4-preamble，Part 1 之既有碼）
        if any(rx.search(text) for rx in wd.PREAMBLE):
            grade, blocker = "W2", "B4-preamble"
            detail = "唯一來源條文為適用性前言，無可測之功能行為"
        else:
            # B8（R-VF80 二）：有 message 而無訊號名 —— 其 TC 之訊號無從得知
            if not wd.SIG_REF.search(text) and SIG_MSG_ONLY.search(text):
                mo = SIG_MSG_ONLY.search(text)
                rows.append({
                    "leaf_id": leaf, "test_set": ts_of.get(lf["title"], ""),
                    "layer3": lf["title"].replace("\\n", " "),
                    "src_ref": lf["src_ref"], "writable": "W2",
                    "blocker_class": "B8-signal-incomplete",
                    "blocker_detail": (f"訊號引用只到 message `{mo.group(1)}` 一層，"
                                       "條文未指出其下何訊號（DR-37）"),
                    "value_source": "", "dr_id": "DR-37",
                    "vcvm_not_clear": "1" if "not clear" in
                                      (rec.get("vc", "") + rec.get("vm", "")).lower()
                                      else "0", "disagree": lf.get("disagree", "0")})
                continue

            sigs = {m.group(2) for m in wd.SIG_REF.finditer(text)}
            verdicts, srcs = {}, {}
            for sg in sigs:
                # 值域來源鏈，先命中者為準
                if sg in mid and mid[sg]:
                    srcs[sg] = "1-LID"
                elif sg in high and high[sg]:
                    srcs[sg] = "2-DBC"
                elif sg in vcvm_domain(rec):
                    srcs[sg] = "4-VCVM"
                else:
                    srcs[sg] = ""
                chain_used[srcs[sg] or "(無來源)"] += 1
                v_ok = bool(srcs[sg])
                verdicts[sg] = ("PASS" if sg in in_dbc and v_ok else
                                "B6" if sg in in_dbc else
                                "B5" if not v_ok else "WARN")
            if any(v == "B5" for v in verdicts.values()):
                grade, blocker = "W2", "B5-signal-absent"
                detail = "斷言目標訊號不存在於基線 DBC 且無逐字來源（L-VS2 FAIL）"
            elif any(v == "B6" for v in verdicts.values()):
                grade, blocker = "W2", "B6-value-absent"
                detail = "訊號名有來源而其值域無來源（R-VS57(4)）"
            else:
                # --- 鏈 3：PROXI 參數之值域 ---
                # 252 leaf 之可測內容立於 PROXI 配置之取得，而非訊號斷言。
                # 其參數若不在 PROXI 表內，即為未解值。
                pref = {m.group(1) for m in PROXI_REF.finditer(text)}
                # B9（R-VF80 二）：條文未帶值，且該參數於 PROXI 表亦無值域
                if pref and not any(rx.search(text) for rx in PROXI_CLAUSE_VALUE):
                    noval = sorted(p for p in pref
                                   if not pv.get(pnorm(p)))
                    if noval:
                        grade, blocker = "W2", "B9-proxi-value-absent"
                        detail = ("條文未帶值，且該 PROXI 參數於表內亦無值域："
                                  + "／".join(noval[:3]))
                        rows.append({
                            "leaf_id": leaf, "test_set": ts_of.get(lf["title"], ""),
                            "layer3": lf["title"].replace("\\n", " "),
                            "src_ref": lf["src_ref"], "writable": grade,
                            "blocker_class": blocker, "blocker_detail": detail,
                            "value_source": "", "dr_id": "",
                            "vcvm_not_clear": "1" if "not clear" in
                                      (rec.get("vc", "") + rec.get("vm", "")).lower()
                                      else "0",
                            "disagree": lf.get("disagree", "0")})
                        continue
                unres = sorted(p for p in pref if pnorm(p) not in px)
                if unres:
                    # R-VS47／R-VS71：未解值非驗證目標本身（其為前提條件，
                    # 條文之可測結果為顯示／啟用行為）→ **W1**，標 PENDING
                    grade, blocker = "W1", ""
                    detail = ("PROXI 參數之值域無來源："
                              + "／".join(unres[:4])
                              + (f" 等 {len(unres)}" if len(unres) > 4 else ""))
                    note = {f"PROXI:{u}": "PENDING" for u in unres}
                    for u in unres:
                        chain_used["3-PROXI(未解)"] += 1
                else:
                    grade = "W0"
                    for u in pref:
                        chain_used["3-PROXI"] += 1
            if not note:
                note = {sg: srcs[sg] for sg in sorted(sigs)}

        rows.append({
            "leaf_id": leaf, "test_set": ts_of.get(lf["title"], ""),
            "layer3": lf["title"].replace("\\n", " "),
            "src_ref": lf["src_ref"], "writable": grade,
            "blocker_class": blocker, "blocker_detail": detail,
            "value_source": ";".join(f"{k}={v}" for k, v in note.items() if v),
            "dr_id": "DR-34" if grade == "W1" else "",
            "vcvm_not_clear": "1" if "not clear" in
                              (rec.get("vc", "") + rec.get("vm", "")).lower() else "0",
            "disagree": lf.get("disagree", "0")})

    # --- 錨點（R-VF21／R-VF28，以內容定錨）---
    by_id = {r["leaf_id"]: r for r in rows}
    if len(rows) != 627:
        raise SystemExit(f"R-VF16：應為 627 列，實得 {len(rows)}")
    nc = [r for r in rows if r["vcvm_not_clear"] == "1"]
    if len(nc) != 90:
        raise SystemExit(f"錨點：`not clear` 之 leaf 應為 90（W-120 實測），實得 {len(nc)}")
    if any(r["value_source"].count("4-VCVM") for r in nc):
        raise SystemExit("R-VF13 第 4 項：`not clear` 之列被用為值域來源，停")
    # 鑑別錨點：LID／DBC 皆無而 VC-VM 有者（A-VS118 於 VF230 之對應形態）
    disc = [r for r in rows if "4-VCVM" in r["value_source"]]
    # --- PROXI 鏈之三錨點（R-VF21／R-VF28）---
    px_all = proxi_known()
    # **W-VF70：本組三錨點之二須更換，其更換之理由須逐字記**——
    # 舊「必不命中」為 `AUX_Switch_Types`，舊「鑑別」之不在表側為
    # `Blindspot_Trailer_Detection`。二者實**皆在表內**（列 911／列 810），
    # **其舊日之「不在表內」全來自 `max_row=800` 之截斷** ——
    # 即該二錨點之通過，是為它們本該攔下的缺陷背書。
    # 真正表內所無者，全池 11 個未解參數中僅 `Greeting_Light` 一個。
    if "heatedseats" not in px_all:
        raise SystemExit("PROXI 必命中錨點不符：`Heated_Seats`（R-VS49 明列）"
                         "不在 PROXI 表內，停")
    if "greetinglight" in px_all:
        raise SystemExit("PROXI 必不命中錨點不符：`Greeting_Light` 不應在表內，停")
    # 鑑別：`Blind_Spot_Monitoring`（列 ≤800）與 `Blindspot_Trailer_Detection`
    # （列 810）二者名近**且皆在表內** —— 其鑑別點已改為
    # 「讀全表方能二者皆見」：任一於截斷式下缺席即為失效之證。
    if not ("blindspotmonitoring" in px_all
            and "blindspottrailerdetection" in px_all):
        raise SystemExit("PROXI 鑑別錨點不符：近名之二參數未皆被讀入"
                         "（讀全表之證），停")
    # 排版瑕疵之錨：表內為 `FOA _Presence`（名中多一空格），
    # 條文為 `FOA_Presence` —— `pnorm()` 須吸收之
    if "foapresence" not in px_all:
        raise SystemExit("PROXI 正規化錨點不符：`FOA _Presence` 未被 pnorm 吸收，停")
    print("PROXI 錨點：必命中 `Heated_Seats` ✅／必不命中 `Greeting_Light` ✅／"
          "鑑別（讀全表）`Blind_Spot_Monitoring` ＋ `Blindspot_Trailer_Detection` ✅／"
          "正規化 `FOA _Presence` ✅")

    p = ROOT / "docs" / "reports" / "vf230_writability.tsv"
    cols = list(rows[0])
    with p.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols, delimiter="\t", lineterminator="\n")
        w.writeheader()
        w.writerows(rows)

    dist = Counter(r["writable"] for r in rows)
    blk = Counter(r["blocker_class"] for r in rows if r["blocker_class"])
    per_ts = defaultdict(Counter)
    for r in rows:
        per_ts[r["test_set"]][r["writable"]] += 1

    (ROOT / "data" / "_vf230_writability.json").write_text(json.dumps({
        "dist": dict(dist), "blocker": dict(blk),
        "per_test_set": {k: dict(v) for k, v in per_ts.items()},
        "chain_used": dict(chain_used),
        "vcvm_only": [r["leaf_id"] for r in disc],
        "not_clear": len(nc), "total": len(rows)},
        ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"627 leaf；分級 {dict(dist)}")
    print(f"blocker {dict(blk)}")
    print(f"值域來源鏈使用 {dict(chain_used)}")
    print(f"`not clear` 之 leaf {len(nc)}（皆未用為來源 ✅）")
    print(f"鑑別錨點（LID／DBC 皆無而 VC-VM 有）：{len(disc)} 個"
          + (f" 例 {disc[0]['leaf_id']}" if disc else " —— **不存在**"))
    print("逐 Test Set：")
    for t, c in sorted(per_ts.items(), key=lambda x: -sum(x[1].values())):
        print(f"  {sum(c.values()):4}  {t:24} {dict(c)}")
    print(f"wrote {p.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
