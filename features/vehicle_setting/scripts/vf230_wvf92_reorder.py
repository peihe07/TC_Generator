"""VF230：457 列之**全表重排**（W-VF92.2／.3；裁定 `R-G32`）。

**排序鍵 = D 欄（Requirement or Design ID）整字串升冪，stable sort。**
B 欄（No.#）依重排後之位置全部重寫，**自實測起點連續**。

---

**其與既有寫回腳本之別 —— 本檔不以 openpyxl 存檔**：

`vf230_wvf84_writeback.py`／`vf230_wvf91_append.py` 皆以 openpyxl 存檔
再補回 `<extLst>`（`R-VF113`）。**本輪不循其例** —— V72 明令
「openpyxl 不得存檔（x14 命名空間剝離之既知缺陷）」，
故本檔為 **XML 手術式**：只動目標分頁之 `<sheetData>` 內列 10–466 之
`<c>` 元素，**其餘 part 逐位元不觸**（含 `extLst`／`dataValidations`／
`conditionalFormatting`／`drawing`）。**x14 遂無被剝離之機會，不需補回。**

**其與 `backend/xlsx_surgical.surgical_save` 之別**（R18-3 所指之單一寫出路徑）：
該模組之模型為「openpyxl 改**格值** → 對源 XML 逐格套用」。
**本輪之操作非改值，而是列之置換** —— 逐格套用則須重寫 445 列 × 16 欄之
長字串，且**其不搬 `<c>` 之 `s`（樣式索引）**，樣式與內容遂脫鉤。
**故本檔自行搬移 `<c>` 區塊（樣式隨內容），而其結構不變式改用該模組之
`verify_structure` —— 檢查仍走共用之權威實作，不另造判準。**
**本檔無任何 openpyxl 存檔呼叫**（`tests/test_single_write_path.py` 之掃描不命中）。

**`<row>` 之屬性留在原位，不隨內容移動** —— 實測列 10–246 有
`ht="14" customHeight="1"`（樣板容量之範圍），列 247–466 無。
其為**位置之屬性**（列高），非內容之屬性；隨內容移動則使 19 條補入列
之「無列高」散入全表。**留在原位即維持現況之外觀，本輪不改之。**
其為既有之不一致（樣板容量 237 列之第五種表現），**具名而不順手改**。

---

**B 欄之起點 = 244，不自 001 起 —— 本檔於此不從 V72 §2.2 之字面**：

V72 §2.2-2 令「B 欄……自 001 起連續」。**其與 `R-VF83` 直接衝突**：

    R-VF83（Pei 裁定 2026-08-23）
      「VF230 之 036 workbook，B 欄「No.#/序號」自 238 起連續遞增，
        **不自 1 起**。」「續號為跨 workbook 之連續序號」
    實測起點 244（Pei 裁定 2026-08-24，CFTS044 交付本末號 243 + 1）

**自 001 起將使 VF230 之 B 1–243 與 CFTS044 之 B 1–243 撞號**，
而 `R-VF83` 配套 2 明文「CFTS044 之 B 欄為凍結欄」。
**故本檔取實測之現行起點（244）重寫 B，並於上繳具名此一不從。**
若 Pei 裁「確自 001 起」，本檔改 `B_START` 之來源即可重跑（重排本身不變）。

---

**四閘＋一新閘，任一不成即停**：

  閘 0  前置實測：資料列數、B 連號、D 之尾碼格式（三位零填 → 字串序即數值序）
  閘 3  **part 級對帳**：`<dataValidation`／`x14:dataValidation`／
        `<conditionalFormatting` 之數、分頁數、drawing/chart rel 數，前後一致
  閘 4  **重讀比對 457 × 16** —— 以「舊 B → 內容」為基準，
        重排後之每一列須與其舊內容逐欄相同（**重排不得改動任何值**）
  新閘  **D 欄逐列非降序**，違者即停
  對帳  資料列 457；選池 574 之四分（另見 `vf230_wvf91_recon.py`）
"""
from __future__ import annotations

import collections
import re
import shutil
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string as ci

FEAT = Path(__file__).resolve().parents[1]
REPO = FEAT.parents[1]
sys.path.insert(0, str(REPO))
sys.path.insert(0, str(FEAT / "scripts"))
from backend import xlsx_surgical as SURG   # noqa: E402  共用之結構不變式
import vf230_wvf77_dryrun as DRY          # noqa: E402  單一權威：欄映射與分頁名
import vf230_wvf84_writeback as WB        # noqa: E402  復用 probe_x14

SHEET, FIRST = DRY.SHEET, DRY.FIRST_DATA_ROW
BACKUP = FEAT / "inputs/_vf230_036_reorder_backup.xlsx"
REMAP = FEAT / "data/vf230_id_remap.tsv"

ROW_RE = re.compile(r"<row ([^>]*)>(.*?)</row>", re.S)
CELL_R_RE = re.compile(r'(<c r=")([A-Z]+)\d+(")')


# ---------------------------------------------------------------- 讀取層

def sheet_part(book: Path) -> str:
    """由分頁名解出其 `xl/worksheets/sheetN.xml` —— **不寫死 sheet6**。"""
    z = zipfile.ZipFile(book)
    wbx = z.read("xl/workbook.xml").decode("utf-8")
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    z.close()
    m = re.search(r'<sheet [^>]*name="%s"[^>]*r:id="([^"]+)"' % re.escape(SHEET), wbx)
    if not m:
        raise SystemExit(f"workbook.xml 內找不到分頁 {SHEET!r}，停")
    t = re.search(r'<Relationship [^>]*Id="%s"[^>]*Target="([^"]+)"' % m.group(1), rels)
    if not t:
        t = re.search(r'<Relationship [^>]*Target="([^"]+)"[^>]*Id="%s"' % m.group(1),
                      rels)
    if not t:
        raise SystemExit(f"rels 內找不到 {m.group(1)}，停")
    return t.group(1).lstrip("/")


def snapshot(book: Path) -> dict[int, dict[str, str]]:
    """讀全部資料列，回 {B: {欄: 值}} —— 閘 4 之基準（不經本檔之 XML 路徑）。"""
    wb = openpyxl.load_workbook(book, read_only=True, data_only=True)
    ws = wb[SHEET]
    out = {}
    for row in ws.iter_rows(min_row=FIRST):
        b, d = row[1].value, row[3].value
        if b in (None, "") or d in (None, ""):
            continue
        out[int(b)] = {c: ("" if row[ci(c) - 1].value is None
                           else str(row[ci(c) - 1].value)) for c in DRY.COLS}
    wb.close()
    return out


def parts_profile(book: Path) -> dict[str, int]:
    """閘 3 之 part 級指紋 —— 前後須逐項一致。"""
    z = zipfile.ZipFile(book)
    names = z.namelist()
    ws_xml = "".join(z.read(n).decode("utf-8", errors="replace")
                     for n in names
                     if n.startswith("xl/worksheets/") and n.endswith(".xml"))
    rels = "".join(z.read(n).decode("utf-8", errors="replace")
                   for n in names if n.endswith(".rels"))
    prof = {
        "dataValidation": ws_xml.count("<dataValidation"),
        "x14:dataValidation": len(re.findall(r"x14:dataValidation[ >]", ws_xml)),
        "conditionalFormatting": ws_xml.count("<conditionalFormatting"),
        "extLst": ws_xml.count("<extLst"),
        "sheets": len([n for n in names
                       if n.startswith("xl/worksheets/") and n.endswith(".xml")]),
        "drawing_rel": len(re.findall(r"/drawing\b", rels)),
        "chart_rel": len(re.findall(r"/chart\b", rels)),
        "parts": len(names),
    }
    z.close()
    return prof


# ---------------------------------------------------------------- 重排層

def plan(book: Path) -> tuple[list[dict], int]:
    """回（重排後之列序，B 起點）。**stable sort，鍵為 D 之整字串。**"""
    wb = openpyxl.load_workbook(book, read_only=True, data_only=True)
    ws = wb[SHEET]
    rows = []
    for row in ws.iter_rows(min_row=FIRST):
        b, d = row[1].value, row[3].value
        if b in (None, "") or d in (None, ""):
            continue
        rows.append({"row": row[0].row, "B": int(b), "D": str(d)})
    wb.close()

    # ---- 閘 0 ----
    if not rows:
        raise SystemExit("目標無資料列，停")
    bs = sorted(r["B"] for r in rows)
    if bs != list(range(bs[0], bs[-1] + 1)):
        raise SystemExit("B 欄非連號 —— 重寫之基準不可靠，停")
    span = [r["row"] for r in rows]
    if span != list(range(span[0], span[-1] + 1)):
        raise SystemExit("資料列不連續（中有空列）—— 本檔之整段搬移前提不成立，停")
    pad = collections.Counter()
    for r in rows:
        m = re.search(r"-(\d+)$", r["D"])
        pad[len(m.group(1)) if m else "無數字尾碼"] += 1
    fams = collections.defaultdict(set)
    for r in rows:
        m = re.search(r"-(\d+)$", r["D"])
        if m:
            fams[r["D"][:m.start()]].add(len(m.group(1)))
    mixed = {k: v for k, v in fams.items() if len(v) > 1}
    print(f"[閘 0] 資料列 **{len(rows)}**（列 {span[0]}–{span[-1]}）｜"
          f"B {bs[0]}–{bs[-1]} 連號｜相異 D {len(set(r['D'] for r in rows))}")
    print(f"       D 之數字尾碼位數：{dict(pad)}")
    print(f"       同一 family 內位數不一者：**{len(mixed)}**"
          f"{'（→ 字串序即數值序）' if not mixed else ' ⚠'}")
    if mixed:
        raise SystemExit(f"family 內尾碼位數不一：{list(mixed)[:3]} —— "
                         "字串序不等於數值序，停，回報")

    order = sorted(rows, key=lambda r: r["D"])          # Python 之 sort 為 stable
    return order, bs[0]


def surgery(book: Path, order: list[dict], b_start: int, part: str) -> None:
    """只改目標分頁之 `<sheetData>`；其餘 part 逐位元原樣寫出。"""
    z = zipfile.ZipFile(book)
    x = z.read(part).decode("utf-8")
    others = [(it, z.read(it.filename)) for it in z.infolist() if it.filename != part]
    it_part = next(it for it in z.infolist() if it.filename == part)
    z.close()

    head, sd, tail = x.partition("<sheetData>")
    if not sd:
        raise SystemExit(f"{part} 內無 <sheetData>，停")
    body, close, rest = tail.partition("</sheetData>")

    cells, attrs = {}, {}
    n_match = 0
    for m in ROW_RE.finditer(body):
        n_match += 1
        rn = int(re.search(r'r="(\d+)"', m.group(1)).group(1))
        attrs[rn] = m.group(0)[:m.group(0).index(">") + 1]
        cells[rn] = m.group(2)
    if n_match != len(ROW_RE.findall(body)):
        raise SystemExit("列之抽取不自洽，停")

    first = order[0] and min(r["row"] for r in order)
    moved = {}
    for i, r in enumerate(order):
        target = first + i
        inner = cells[r["row"]]
        inner = CELL_R_RE.sub(lambda mm: f"{mm.group(1)}{mm.group(2)}{target}"
                              f"{mm.group(3)}", inner)
        newb = b_start + i
        bcell = re.compile(r'(<c r="B%d"[^>]*>)(.*?)(</c>)' % target, re.S)
        if not bcell.search(inner):
            raise SystemExit(f"列 {r['row']} 無 B 格，停")
        inner = bcell.sub(lambda mm: f"{mm.group(1)}<v>{newb}</v>{mm.group(3)}",
                          inner, count=1)
        moved[target] = inner

    out = []
    for rn in sorted(cells):
        out.append(attrs[rn] + moved.get(rn, cells[rn]) + "</row>")
    newbody = "".join(out)
    newx = head + "<sheetData>" + newbody + "</sheetData>" + rest
    try:
        ET.fromstring(newx)
    except ET.ParseError as e:
        raise SystemExit(f"改後 XML 不合法：{e} —— 停，不寫出")

    tmp = book.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zo:
        for it, data in others:
            zo.writestr(it, data)
        zo.writestr(it_part, newx.encode("utf-8"))
    tmp.replace(book)


# ---------------------------------------------------------------- 主流程

def main() -> None:
    write = "--write" in sys.argv
    book = DRY.book()
    part = sheet_part(book)
    print("=== W-VF92.2／.3 全表重排（R-G32；目標為 **repo 內副本**）===")
    print(f"目標：{book.name[:58]}…\n分頁 part：{part}"
          f"{'' if write else '   【空跑，未加 --write】'}\n")

    before = snapshot(book)
    prof_before = parts_profile(book)
    order, b_start = plan(book)
    print(f"       B 之起點取 **{b_start}**（實測現行起點；R-VF83「不自 1 起」，"
          f"**不從 V72 §2.2 之「自 001 起」**，見檔頭）")

    def fam(d: str) -> str:
        return re.sub(r"-\d+$", "", d)

    def segs(seq) -> int:
        n, prev = 0, object()
        for v in seq:
            if v != prev:
                n += 1
            prev = v
        return n

    cur = sorted(before)
    d_by_b = {b: before[b]["D"] for b in cur}
    print(f"\n[量測] family 相異 **{len(set(fam(v) for v in d_by_b.values()))}**")
    print(f"       重排前之 family 段數 **{segs([fam(d_by_b[b]) for b in cur])}**")
    print(f"       重排後之 family 段數 **{segs([fam(r['D']) for r in order])}**"
          "（= family 數即每 family 恰一段）")

    # ---- 對照表（**重排使 B 全變，無此表則舊 B 之引用全部斷鏈**）----
    lines = ["old_b\tnew_b\treq_id"]
    for i, r in enumerate(order):
        lines.append(f"{r['B']}\t{b_start + i}\t{r['D']}")
    if write:
        REMAP.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"\n[對照] {REMAP.name}：**{len(order)} 列**"
          f"{'（已寫出）' if write else '（空跑，未寫出）'}｜"
          f"B 位置不變者 {sum(1 for i, r in enumerate(order) if r['B'] == b_start + i)}")

    if not write:
        print("\n**空跑結束，未動任何檔。** 加 --write 施行。")
        return

    shutil.copy2(book, BACKUP)
    print(f"\n[備份] 重排前之本 → {BACKUP.name}")
    surgery(book, order, b_start, part)

    # ---- 閘 3：part 級對帳 ----
    # **共用之權威實作優先** —— `verify_structure` 之三項：
    # zip member 集合相同、逐分頁之 dv 數（classic／x14）相同、
    # **且相異之 part 只許為所改之該一分頁**。其違即 raise，不降為 warning。
    rep = SURG.verify_structure(BACKUP, book, {part})
    print("\n[閘 3a] xlsx_surgical.verify_structure：**通過**")
    print(f"        part 總數 {rep['members']}｜"
          f"**內容相異之 part {len(rep['differing'])}**：{rep['differing']}")
    prof_after = parts_profile(book)
    print("\n[閘 3b] part 級對帳（前 → 後，本檔之獨立實作）")
    bad = []
    for k in prof_before:
        ok = prof_before[k] == prof_after[k]
        print(f"       {k:24s} {prof_before[k]:>4} → {prof_after[k]:>4}  "
              f"{'一致' if ok else '**不一致**'}")
        if not ok:
            bad.append(k)
    print(f"       probe_x14（獨立實作）{WB.probe_x14(book)}")
    if bad:
        raise SystemExit(f"part 級對帳不一致：{bad} —— 停")

    # ---- 閘 4：重讀比對 457 × 16 ----
    after = snapshot(book)
    diff = []
    for i, r in enumerate(order):
        newb = b_start + i
        got = after.get(newb)
        if got is None:
            diff.append((newb, "整列缺"))
            continue
        old = before[r["B"]]
        for c in DRY.COLS:
            if c == "B":
                continue
            if old[c].strip() != got[c].strip():
                diff.append((newb, c))
    print(f"\n[閘 4] 重讀比對 **{len(after)} 列 × {len(DRY.COLS)} 欄**"
          f"（B 除外，其為本輪所改）：差異 **{len(diff)}**")
    for d in diff[:6]:
        print(f"       新 B{d[0]} 欄 {d[1]}")
    if diff or len(after) != len(before):
        raise SystemExit("重讀比對有差異，停")

    # ---- 新閘：D 欄非降序 ----
    seq = [after[b]["D"] for b in sorted(after)]
    viol = [(i, seq[i], seq[i + 1]) for i in range(len(seq) - 1) if seq[i] > seq[i + 1]]
    print(f"[新閘] D 欄逐列非降序：違者 **{len(viol)}**")
    if viol:
        for v in viol[:5]:
            print(f"       位置 {v[0]}：{v[1]} > {v[2]}")
        raise SystemExit("D 欄非非降序，停")
    print(f"[新閘] B 欄連號：{min(after)}–{max(after)}｜"
          f"連號 {sorted(after) == list(range(min(after), max(after) + 1))}")

    print("\n" + "=" * 66)
    print("**R-VF112／R-VF116：以下為 Pei 之動作，本層不得代之**")
    print("  1. 以 **Excel** 開啟並確認下拉與版面（重排後之新本）")
    print("  2. **逐字具名檔名**複製至交付路徑 VF230_V1_R5/")
    print("=" * 66)


if __name__ == "__main__":
    main()
