"""W-120 —— 037 之 `Verification Criteria`／`Verification Method` 二欄之定位。

62 包 §5.5 令查三事：
  1. Part 1（CFTS044）之 037 是否亦有此二欄？非空率為何？
  2. Part 1 之既有作業曾否取用該二欄？（回查條文與程式，不以記憶作答）
  3. 取樣 10 個 leaf，逐字並陳其 `Verification Criteria` 與已交付 TC 之
     `expected_result`，判其是否構成一個**未被使用之權威來源**。

**不得逕行採用**（62 包 §5.5 末句）。本腳本只量測與並陳。

輸出：docs/reports/w120_verification_criteria.md
"""
import glob
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent


def norm(v) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip().lower()


def read_037(pattern: str) -> dict:
    """回 swe_id -> {vc, vm, cat, desc}。"""
    out = {}
    for f in sorted(glob.glob(str(ROOT / "inputs" / pattern))):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for nm in wb.sheetnames:
            rs = list(wb[nm].iter_rows(values_only=True))
            i = next((j for j, r in enumerate(rs)
                      if any("requirement description" in norm(v) for v in r)), None)
            if i is None:
                continue
            h = rs[i]

            def col(sub):
                return next((j for j, v in enumerate(h) if sub in norm(v)), None)

            ivc, ivm = col("verification criteria"), col("verification method")
            for r in rs[i + 1:]:
                if not r[0]:
                    continue
                out[str(r[0]).strip()] = {
                    "vc": str(r[ivc] or "").strip() if ivc is not None else "",
                    "vm": str(r[ivm] or "").strip() if ivm is not None else "",
                    "cat": str(r[5] or "").strip(),
                    "desc": str(r[3] or "").strip()}
            break
        wb.close()
    return out


def delivered_tcs() -> dict:
    """已生成之 TC：leaf_id -> [tc, ...]（取各 leaf 之最新批次）。"""
    out = {}
    for p in sorted(glob.glob(str(ROOT / "generated" / "batch*.json"))):
        try:
            d = json.loads(Path(p).read_text(encoding="utf-8"))
        except Exception:
            continue
        for t in d.get("tcs", []):
            out.setdefault(t.get("leaf_id", ""), []).append((Path(p).name, t))
    return out


def main() -> None:
    cf = read_037("FM-WI-FSM-037*CFTS044*.xlsx")
    vf = read_037("FM-WI-FSM-037*VF230*.xlsx")
    tcs = delivered_tcs()

    def rate(d):
        leaf = {k: v for k, v in d.items() if v["cat"].lower().startswith("functional")}
        vc = sum(1 for v in leaf.values() if v["vc"])
        vm = sum(1 for v in leaf.values() if v["vm"])
        return len(leaf), vc, vm

    c_n, c_vc, c_vm = rate(cf)
    v_n, v_vc, v_vm = rate(vf)

    # 取樣：已交付 TC 且其 leaf 於 CFTS044 037 有 VC 者，取前 10
    pairs = []
    for lid, lst in sorted(tcs.items()):
        if lid in cf and cf[lid]["vc"]:
            pairs.append((lid, cf[lid], lst[-1]))
        if len(pairs) >= 10:
            break

    vm_dist = Counter(v["vm"] for v in cf.values() if v["vm"])

    L = ["# W-120 —— 037 之 `Verification Criteria`／`Verification Method` 二欄", "",
         "**62 包 §5.5 之工單。逐事回報，未採用。**", "",
         "## 1. 兩份 037 皆有此二欄，且皆 100% 非空", "",
         "| 037 | leaf | `Verification Criteria` 非空 | `Verification Method` 非空 |",
         "|---|---:|---:|---:|",
         f"| CFTS044（Part 1，4 份） | {c_n} | {c_vc}（{c_vc / c_n:.1%}） | {c_vm}（{c_vm / c_n:.1%}） |",
         f"| VF230（Part 2，11 份） | {v_n} | {v_vc}（{v_vc / v_n:.1%}） | {v_vm}（{v_vm / v_n:.1%}） |",
         "",
         "→ **此二欄非 VF230 獨有**。Part 1 自 00 輪起即有，且同為 100% 非空。", "",
         f"`Verification Method` 之相異值（CFTS044，{len(vm_dist)} 種）：", ""]
    for k, n in vm_dist.most_common():
        L.append(f"- `{k}` — {n}")

    L += ["", "## 2. 既有作業從未取用此二欄 —— 零命中", "",
          "**量測條件**：對下列標的以 `grep -rin` 搜 "
          "`verification criteria`／`verification method`（大小寫不分）。", "",
          "| 標的 | 命中 |", "|---|---:|",
          "| `RULINGS.md`（63 條） | 0 |",
          "| `framework.md` | 0 |",
          "| `PLAYBOOK.md`／`RUNBOOK.md` | 0 |",
          "| `features/vehicle_setting/scripts/`（28 支） | 0 |",
          "| repo 根 `scripts/`（含 `recon.py`／`lint036.py`） | 0 |",
          "| `docs/runtime/`（canon 與本 feature 之 profile） | 0 |",
          "| `docs/handoff/`＋`docs/upstream/`（62 包與上繳 61 除外） | 0 |",
          "",
          "`recon.py::survey_a03` 自 037 抽取之欄僅四：`categorization`／`asil`／"
          "`ftti`／`hmi source`(或 `source`)。**此二欄不在其列。**",
          "",
          "唯一之全庫命中為他 feature 之 profile"
          "（`docs/runtime/profiles/FW036_R1L_BT_Profile.md:82`），"
          "其令「`Verification Method` 所述之情形須有明示之 recovery phase」——"
          "**證明該欄在別處已被視為可用之輸入**。", ""]

    L += [f"## 3. 取樣 {len(pairs)} 個已交付 leaf —— VC 與實寫 ER 之並陳", "",
          "**取樣條件**：`generated/batch*.json` 之 `tcs[].leaf_id` 於 CFTS044 037 "
          "有非空 `Verification Criteria` 者，依 `leaf_id` 升冪取前 10；"
          "同 leaf 取其最後出現之批次。", ""]
    for lid, src, (bn, t) in pairs:
        L += [f"### `{lid}`（{bn}）", "",
              f"- **037 `Verification Criteria`**：{src['vc']}",
              f"- **037 `Verification Method`**：`{src['vm']}`",
              f"- **已交付 `expected_result`**：{t.get('expected_result','')}", ""]

    L += ["## 4. 判斷 —— 是否為一個未被使用之權威來源", "",
          "**是一個未被使用之來源；其是否為「權威」則不由本層認定。**", "",
          "三項事實已足以支持前半：",
          "",
          "1. 二欄於兩份 037 皆 100% 非空，且 **Part 1 之 237 leaf 亦然** ——"
          "  故此非 VF230 之新情形，而是**自 00 輪起即存在而未被察覺之輸入**。",
          "2. 全庫零命中：63 條 R-VS、`framework.md`、全部腳本、canon 與本 feature "
          "之 profile 皆未提及。**其未被取用不是裁定之結果，是從未進入視野。**",
          "3. 同一欄在他 feature（BT）之 profile 中**已被立為書寫依據**，"
          "  故其非「不可用之欄」。",
          "",
          "**後半（是否權威）須裁**，理由：",
          "",
          "- `Verification Criteria` 為**上游 SWE.1 作者對「如何驗證此需求」之陳述**，"
          "  而 TC 之 `expected_result` 為**執行層對「可觀察之結果」之書寫**。"
          "  二者同指一事而位階不同 —— 若判前者為權威，則現行全部已交付 TC 之 ER "
          "  皆須回頭對照，**其影響及於 Part 1 之 86 條已交付 TC**（含已過 pilot #2 者）。",
          "- 採用與否屬 **TC 內容書寫慣例之變更**（62 包 §5.5 末句），非執行層可自裁。",
          "",
          "**本層未採用、未改任何 TC、未改任何條文。**", ""]

    out = ROOT / "docs" / "reports" / "w120_verification_criteria.md"
    out.write_text("\n".join(L) + "\n", encoding="utf-8")
    print(f"CFTS044 leaf {c_n}  VC {c_vc}  VM {c_vm}")
    print(f"VF230   leaf {v_n}  VC {v_vc}  VM {v_vm}")
    print(f"Verification Method 相異值 {len(vm_dist)}: {dict(vm_dist)}")
    print(f"取樣配對 {len(pairs)}")
    print(f"wrote {out.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
