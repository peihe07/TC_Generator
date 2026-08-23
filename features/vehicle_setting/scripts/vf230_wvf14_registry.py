"""W-VF14 —— R-VF13 之回溯掃描（**存查性；不觸發任何變更**）。

V06 §5.1 依 R-VF14 修訂本工單之目的：

  - 仍執行掃描（兩 feature 分報）
  - **不產出「應轉之新分級」之建議**
  - 產出為**存查表**：哪些 leaf 之值域在 VC/VM 欄有解、其現行分級為何、
    是否已交付
  - **不觸發任何變更、不改分級、不改 TC**

其價值：日後若上游質疑某 leaf 之分級，可證本層知其存在且依 R-VF14 未改。

「已交付」之三種候選判準（W-VF16，未裁）逐一並列，不擇一。

輸出：docs/reports/wvf14_vcrit_registry.md ＋ data/_wvf14_registry.json
"""
import csv
import glob
import json
import os
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

# 值域形態：`SIGNAL = "值"` 或 `SIGNAL == "值"`（VC/VM 欄之實際書寫）
VALUE_ASSIGN = re.compile(
    r"\b([A-Z][A-Z0-9_]{2,}(?:\.[A-Za-z]\w{2,})?)\s*==?\s*[\"'\[]?([A-Za-z0-9_ ]{1,40})[\"'\]]?")


def norm(v) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip().lower()


def read_vcrit(pattern: str) -> dict:
    """037 之 VC/VM 欄：swe_id -> {vc, vm, cat}。"""
    out = {}
    for f in sorted(glob.glob(str(ROOT / "inputs" / pattern))):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for nm in wb.sheetnames:
            rs = list(wb[nm].iter_rows(values_only=True))
            i = next((j for j, r in enumerate(rs)
                      if any("requirement description" in norm(v) for v in r)), None)
            if i is None:
                continue
            h = rs[i]

            def col(sub):
                return next((j for j, v in enumerate(h) if sub in norm(v)), None)

            ivc, ivm = col("verification criteria"), col("verification method")
            for r in rs[i + 1:]:
                if not r[0]:
                    continue
                out[str(r[0]).strip()] = {
                    "vc": str(r[ivc] or "").strip() if ivc is not None else "",
                    "vm": str(r[ivm] or "").strip() if ivm is not None else "",
                    "cat": str(r[5] or "").strip()}
            break
        wb.close()
    return out


def grades() -> dict:
    """Part 1 之現行分級（`docs/reports/writability.tsv`）。"""
    p = ROOT / "docs" / "reports" / "writability.tsv"
    if not p.exists():
        return {}
    out = {}
    with p.open(encoding="utf-8") as fh:
        for r in csv.DictReader(fh, delimiter="\t"):
            out[r.get("leaf_id", "")] = r
    return out


def delivered() -> tuple[set, set, set]:
    """三種候選判準各自之「已交付」集（W-VF16，未裁）。

    (a) 已寫回交付路徑並 tag —— 實測交付路徑 036 之本 pipeline 產出欄
        （F/G/J/K/O/P/R/AA/AH）全數 0 filled，故為**空集**。
    (b) 已生成之批次（各 batch 取最高版次）。
    (c) 已於 RD-1 送出之項次 —— RD-1 之標的為 `Heated Seat`(88)＋
        `Vented Seat`(72) 共 160 leaf，以 layer3 欄辨識。
    """
    a: set = set()
    b: set = set()
    best = {}
    for p in sorted(glob.glob(str(ROOT / "generated" / "batch*.json"))):
        m = re.match(r"batch(\d+)(?:_v(\d+))?\.json", os.path.basename(p))
        if not m:
            continue
        n, v = int(m.group(1)), int(m.group(2) or 1)
        if n not in best or v > best[n][0]:
            best[n] = (v, p)
    for _, p in best.values():
        d = json.loads(Path(p).read_text(encoding="utf-8"))
        b.update(t["leaf_id"] for t in d.get("tcs", []))
    # RD-1 之標的以 layer3 之**族名**辨識：`*HeatedSeat*`／`*VentedSeat*`。
    # 首版以全等比對四個字串，命中 0 —— 實際值為
    # `ThreeStagesHeatedSeat`／`LeftFrontVentedSeat` 等 11 種。
    # `HeatedSteeringWheel*` **不在 RD-1 範圍**（其標的為 FL/FR_HS_RQ 等，
    # 非方向盤），故須明排除。
    c = {k for k, r in grades().items()
         if re.search(r"(HeatedSeat|VentedSeat)", r.get("layer3", ""))
         and "SteeringWheel" not in r.get("layer3", "")}
    return a, b, c


def solvable(rec: dict) -> list[tuple[str, str]]:
    """VC/VM 欄內之 (訊號, 值) 對。`not clear` 之列排除（R-VF13 第 4 項）。"""
    txt = rec["vc"] + "\n" + rec["vm"]
    if "not clear" in txt.lower():
        return []
    seen, out = set(), []
    for m in VALUE_ASSIGN.finditer(txt):
        sig, val = m.group(1), m.group(2).strip()
        if not val or sig in seen or len(val) < 2:
            continue
        seen.add(sig)
        out.append((sig, val))
    return out


def main() -> None:
    da, db, dc = delivered()
    g = grades()
    rows = []
    for tag, pat in (("CFTS044", "FM-WI-FSM-037*CFTS044*.xlsx"),
                     ("VF230", "FM-WI-FSM-037*VF230*.xlsx")):
        src = read_vcrit(pat)
        for sid, rec in sorted(src.items()):
            pairs = solvable(rec)
            if not pairs:
                continue
            gr = g.get(sid, {})
            rows.append({
                "feature": tag, "leaf": sid,
                "pairs": pairs,
                "writable": gr.get("writable", "—（無分級紀錄）"),
                "blocker": gr.get("blocker_class", ""),
                "dr": gr.get("dr_id", ""),
                "delivered_a": sid in da, "delivered_b": sid in db,
                "delivered_c": sid in dc})

    L = ["# W-VF14 —— R-VF13 之回溯掃描（**存查表；未觸發任何變更**）", "",
         "**V06 §5.1 依 R-VF14 修訂本工單為存查性。**",
         "本表**不含「應轉之新分級」之建議**，不改分級、不改 TC、不改條文。", "",
         "其價值：日後若上游質疑某 leaf 之分級，可證本層知其存在且依 R-VF14 未改。", "",
         "## 0. 量測條件", "",
         "- 值域形態：VC/VM 欄內之 `SIGNAL = \"值\"` 或 `SIGNAL == \"值\"`",
         "- **`not clear` 之列一律排除**（R-VF13 第 4 項）",
         "- 現行分級取自 `docs/reports/writability.tsv`（Part 1）；"
         "VF230 尚無分級，一律記 `—`",
         "- 「已交付」之三判準（W-VF16，**未裁**）並列，不擇一：", "",
         "  | 判準 | 定義 | 集合大小 |", "  |---|---|---:|",
         f"  | (a) | 已寫回交付路徑並 tag | **{len(da)}** —— 036 之本 pipeline "
         f"產出欄全數 0 filled |",
         f"  | (b) | 已生成之批次（各批取最高版次） | {len(db)} |",
         f"  | (c) | 已於 RD-1 送出之項次（Heated／Vented Seat） | {len(dc)} |", "",
         f"## 1. 存查表（{len(rows)} 列）", "",
         "| feature | leaf | VC/VM 內之 (訊號, 值) | 現行分級 | blocker | DR | 已交付 (a)/(b)/(c) |",
         "|---|---|---|---|---|---|---|"]
    for r in rows:
        pr = "／".join(f"`{s}` = `{v}`" for s, v in r["pairs"][:3])
        d = "／".join("是" if r[k] else "否"
                      for k in ("delivered_a", "delivered_b", "delivered_c"))
        L.append(f"| {r['feature']} | `{r['leaf']}` | {pr} | {r['writable']} | "
                 f"{r['blocker'] or '—'} | {r['dr'] or '—'} | {d} |")

    w2 = [r for r in rows if str(r["writable"]).upper().startswith("W2")]
    none_deliv = [r for r in rows
                  if not (r["delivered_a"] or r["delivered_b"] or r["delivered_c"])]
    L += ["", "## 2. 小結（陳述事實，不作建議）", "",
          f"- 存查列合計 **{len(rows)}**",
          f"- 其中現行分級為 **W2** 者 **{len(w2)}**",
          f"- 其中**三判準皆判「未交付」**者 **{len(none_deliv)}** ——",
          "  **對此類，R-VF14 之排除效力於任一判準下皆不成立。**", "",
          "**本層未依本表改動任何 leaf 之分級。**", ""]
    if w2:
        L += ["現行 W2 且三判準皆未交付者：", ""]
        L += [f"- `{r['leaf']}`（{r['feature']}，blocker `{r['blocker'] or '—'}`）"
              for r in w2 if r in none_deliv]
        L += [""]

    out = ROOT / "docs" / "reports" / "wvf14_vcrit_registry.md"
    out.write_text("\n".join(L) + "\n", encoding="utf-8")
    (ROOT / "data" / "_wvf14_registry.json").write_text(
        json.dumps({"rows": rows, "delivered_a": sorted(da),
                    "delivered_b_n": len(db), "delivered_c_n": len(dc)},
                   ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"存查列 {len(rows)}（CFTS044 "
          f"{sum(1 for r in rows if r['feature'] == 'CFTS044')} / VF230 "
          f"{sum(1 for r in rows if r['feature'] == 'VF230')}）")
    print(f"現行 W2 {len(w2)}；三判準皆未交付 {len(none_deliv)}")
    print(f"已交付集：(a) {len(da)}  (b) {len(db)}  (c) {len(dc)}")
    print(f"wrote {out.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
