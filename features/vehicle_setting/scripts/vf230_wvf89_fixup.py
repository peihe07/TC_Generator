"""W-VF89.1：`sqref` 四處延伸至 1000 ＋ 列 247–447 之樣式補齊。

**二者同源**：樣板之容量為 **237 列**（列 10–246），而本次寫入 **438 列**（列 10–447）。
其表現有三：
  B 欄公式  237 格（已由 `R-VF132` 覆寫為硬值）
  `sqref`   P/R 至列 132、T–Z 至 132、AF 至 13（`R-VF138`／`R-VF139`）
  **儲存格樣式  至列 246 —— 列 247–447 之 201 列無 `wrap_text`**
            → **其多行欄之換行於 Excel 顯示為空白**，
              即「編號 `2.` `3.` 前面多一個空格」之成因。

**`Q10:Q11` 不延伸** —— Q 為 `Estimated Test Time (mins)`，
其被納入 Priority 之 `dataValidation`（選項 `P0,P1,P2,P3`）為樣板之既有瑕疵；
**延伸之則將一個錯誤之驗證自 2 列擴大至 991 列**。維持原狀並具名。

**樣式之來源為列 10** —— 實測列 10 與列 100 之樣式相同，故其為標準資料列樣式。
**列 10–246 不動**（其樣式為樣板既有），只補列 247–447。
"""
from __future__ import annotations

import sys
import zipfile
import xml.etree.ElementTree as ET
from copy import copy
from pathlib import Path

import openpyxl

FEAT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(FEAT / "scripts"))
import vf230_wvf77_dryrun as DRY          # noqa: E402
import vf230_wvf84_writeback as WB        # noqa: E402  復用其 restore_extlst 與 probe_x14

LIMIT = 1000
STYLE_SRC_ROW, FIRST, LAST = 10, 10, 447
# **`R-VF139` 明文為 P／R／T–Z／AF 四處，本表即以此為限。**
#
# 【已撤回之處置，記之以防再犯】本腳本前一版曾將 `H10:H132` 列為「第五處下拉」
# 而一併延伸至 1000。**其前提為假** —— `H10:H132` 之母元素為
# `<conditionalFormatting>`（colorScale），**不是 `dataValidation`**。
# 其誤因為抽取時只取 `sqref="..."` 字串而不問母元素為何
# （同一抽取所得之 `L9` 實為 `<selection activeCell="L9">`）。
# **已還原為 `H10:H132`，其與交付本 CFTS044 一致。**
#
# 【未納入者，其為刻意】
#   `Q10:Q11`     其被納入 Priority 之下拉而 Q 為 Estimated Test Time——樣板既有瑕疵，
#                 延伸之則將一個錯誤之驗證自 2 列擴大至 991 列。
#   `autoFilter`  `ref="A9:AH132"` 而資料至列 447（樣板容量之第四處）。
#                 **其為另一元素，不在 `R-VF139` 之射程**；且交付本 CFTS044 亦如此。
#                 其修正屬 Excel 內之一個動作，已列入交付說明 §六。
SUBS = [
        ('sqref="P10:P132 Q10:Q11"', f'sqref="P10:P{LIMIT} Q10:Q11"', "P（Priority）"),
        ('sqref="T10:Z132"', f'sqref="T10:Z{LIMIT}"', "T–Z（車型適用）"),
        ('sqref="AF10:AF13"', f'sqref="AF10:AF{LIMIT}"', "AF（測試結果）"),
        ("<xm:sqref>R10:R132</xm:sqref>",
         f"<xm:sqref>R10:R{LIMIT}</xm:sqref>", "R（Design，x14）")]


def fix_styles(book: Path) -> int:
    wb = openpyxl.load_workbook(book)
    ws = wb[DRY.SHEET]
    n = 0
    for col in range(1, 35):
        src = ws.cell(row=STYLE_SRC_ROW, column=col)
        # **範圍為列 10–447，非僅 247–447** ——
        # 樣板於列 10–246 內亦有缺（實測 L181、M179–181、M190–192 共 7 格，
        # **其值皆為多行而無 `wrap_text`**，即使用者所報「編號前多空格」之同一形態）。
        # **只補缺者，不動已有 `wrap_text` 之格** —— 其樣式為樣板既有。
        for rn in range(FIRST, LAST + 1):
            dst = ws.cell(row=rn, column=col)
            if src.alignment.wrap_text is None:
                continue
            if dst.alignment.wrap_text is None or dst.font.name != src.font.name:
                dst._style = copy(src._style)
                n += 1
    wb.save(book)
    wb.close()
    return n


def fix_sqref(book: Path) -> list:
    tmp = book.with_suffix(".tmp.xlsx")
    zin, zout = zipfile.ZipFile(book), zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED)
    done = []
    for it in zin.infolist():
        data = zin.read(it.filename)
        if it.filename.startswith("xl/worksheets/") and it.filename.endswith(".xml"):
            x = data.decode("utf-8")
            # **冪等**：已為 LIMIT 之形式者視為已完成 ——
            # 本檔須可重跑（樣式與 `sqref` 之修正可能分次施行）。
            for old, new, tag in SUBS:
                if old in x:
                    x = x.replace(old, new, 1)
                    done.append(tag)
                elif new in x:
                    done.append(tag + "（已為 %d，未動）" % LIMIT)
            try:
                ET.fromstring(x)
            except ET.ParseError as e:
                zin.close(); zout.close(); tmp.unlink(missing_ok=True)
                raise SystemExit(f"改後 XML 不合法（{it.filename}）：{e} —— 停，不寫出")
            data = x.encode("utf-8")
        zout.writestr(it, data)
    zin.close(); zout.close()
    if len(done) != 5:
        tmp.unlink(missing_ok=True)
        raise SystemExit(f"`sqref` 五處僅命中 {len(done)}：{done} —— 停")
    tmp.replace(book)
    return done


def main() -> None:
    book = DRY.book()
    print(f"目標：{book.name[:58]}…")
    # **順序關鍵**：`fix_styles` 以 openpyxl 存檔，**其again移除 x14 之 `extLst`** ——
    # 故須先備份原檔為 extLst 之來源，改樣式後補回，方可改 x14 之 `sqref`。
    # **首版未為之，`fix_sqref` 遂只命中 3／4**（x14 之 `<xm:sqref>` 已不存在）。
    import shutil, tempfile
    src_bak = Path(tempfile.mkdtemp()) / "extlst_src.xlsx"
    shutil.copy2(book, src_bak)
    x14_before = WB.probe_x14(book)

    n = fix_styles(book)
    print(f"\n[樣式] 列 10–447 補齊之儲存格（只補缺者）：**{n}**")
    print(f"[x14]  改樣式前 {x14_before} → 存檔後 {WB.probe_x14(book)}")
    WB.restore_extlst(src_bak, book)
    x14_after = WB.probe_x14(book)
    print(f"       補回後 **{x14_after}**（原 {x14_before}）")
    if x14_after != x14_before:
        raise SystemExit("x14 之保全失敗，停")
    shutil.rmtree(src_bak.parent, ignore_errors=True)

    done = fix_sqref(book)
    print(f"[sqref] **五處**皆改至 {LIMIT}：")
    for d in done:
        print(f"          {d}")
    print("        **Q10:Q11 未動**（樣板既有瑕疵，延伸之則擴大之）")

    idx = openpyxl.utils.column_index_from_string
    rows = DRY.rows_from(DRY.sources(), DRY.cfts044_last_b()[0])
    wb = openpyxl.load_workbook(book, read_only=True, data_only=True)
    ws = wb[DRY.SHEET]
    got = {}
    for row in ws.iter_rows(min_row=FIRST, values_only=False):
        b = row[1].value
        if b in (None, ""):
            continue
        got[int(b)] = {c: ("" if row[idx(c) - 1].value is None
                           else str(row[idx(c) - 1].value)) for c in DRY.COLS}
    wb.close()
    diff = [(r["B"], c) for r in rows for c in DRY.COLS
            if str(r[c]).strip() != got.get(r["B"], {}).get(c, "\x00").strip()]
    print(f"\n[閘 4] 重讀比對 {len(rows)} 列 × {len(DRY.COLS)} 欄：**差異 {len(diff)}**")
    if diff:
        for d in diff[:5]:
            print(f"        B{d[0]} 欄 {d[1]}")
        raise SystemExit("改後之重讀比對有差異，停")
    print("        ✅ 資料未因樣式與 sqref 之修改而變動")


if __name__ == "__main__":
    main()
