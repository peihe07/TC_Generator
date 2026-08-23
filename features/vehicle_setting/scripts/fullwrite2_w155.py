"""W-155（82 包 §4）—— 全量寫回，依 R-VS70 外科式 emit、R-VS72 一 leaf 一列。

與 W-144 之別（三項）：
  (1) 來源集含 `batch21_probe.json`（A-VS160(b) 修正後）
  (2) **held_out 7 之理由入 AH**（82 包 §4 W-155(3)）
  (3) 寫前後之 raw XML 七項計數逐項比對，任一下降即中止並還原；
      重讀後逐列比對十六欄，不符亦中止並還原
"""
from __future__ import annotations

import hashlib
import json
import re
import shutil
import sys
import zipfile
from pathlib import Path

FEAT = Path(__file__).resolve().parents[1]
REPO = FEAT.parents[1]
sys.path.insert(0, str(REPO / "backend"))
sys.path.insert(0, str(FEAT / "scripts"))

from writeback_036 import BOOK, COL_IDX, FIRST_DATA_ROW, SHEET, latest_batches  # noqa
from fullwrite_w144 import build_rows                                          # noqa
from completeness_w154 import classify                                         # noqa
from xlsx_surgical import surgical_save                                        # noqa

# 還原之來源 —— 取**寫前之最新備份**（55 輪為 fullwrite3）
BACKUP = BOOK.parent / "REF/036_pre_fullwrite3_20260823.xlsx"


def seven(path: Path) -> dict:
    """raw XML 七項計數 —— 任一下降即結構受損。"""
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        sheets = [n for n in names if re.match(r"xl/worksheets/sheet\d+\.xml$", n)]
        blob = b"".join(z.read(n) for n in sheets)
        rels = b"".join(z.read(n) for n in names
                        if n.startswith("xl/worksheets/_rels/"))
        ss = z.read("xl/sharedStrings.xml") if "xl/sharedStrings.xml" in names else b""
    return {"dataValidation": blob.count(b"<dataValidation "),
            "x14:dataValidation": blob.count(b"<x14:dataValidation"),
            "conditionalFormatting": blob.count(b"<conditionalFormatting"),
            "sheets": len(sheets),
            "drawing_chart_rel": rels.count(b"drawing") + rels.count(b"chart"),
            "sharedStrings": ss.count(b"<si>"),
            "members": len(names)}


def held_out_ah() -> dict[str, str]:
    """held_out 之理由 → AH（82 包 §4 W-155(3)）。"""
    r = classify()
    return {l: "HELD OUT: " + re.sub(r"\s+", " ", r["held"][l]["reason"]).strip()
            for l in r["buckets"]["held_out"]}


def main() -> int:
    import openpyxl

    rows, stat = build_rows()
    ho = held_out_ah()
    for row in rows:
        if row["D"] in ho and not row["I"]:
            row["AH"] = ho[row["D"]]
    print(f"列數 **{len(rows)}**（已生成 {stat['generated']}／"
          f"未生成 {stat['not_generated']}，其中 held_out {len(ho)}）")

    before_hash = hashlib.sha256(BOOK.read_bytes()).hexdigest()
    before = seven(BOOK)
    print(f"寫前 sha256：{before_hash}")
    print("寫前七項：", before)

    wb = openpyxl.load_workbook(BOOK)
    ws = wb[SHEET]
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        for c in range(COL_IDX["B"], COL_IDX["AH"] + 1):
            ws.cell(row=r, column=c).value = None
    for i, row in enumerate(rows):
        for k, idx in COL_IDX.items():
            if k in row:
                ws.cell(row=FIRST_DATA_ROW + i, column=idx).value = row[k] or None

    out = BOOK.with_suffix(".surgical.xlsx")
    surgical_save(wb, BOOK, out)                       # R-VS70；其自身亦驗結構

    after = seven(out)
    drop = {k: (before[k], after[k]) for k in before if after[k] < before[k]}
    if drop:
        out.unlink(missing_ok=True)
        print("**七項有下降，中止（母本未動）**：", drop)
        return 1
    print("寫後七項：", after, "—— 無下降")

    shutil.move(out, BOOK)

    # (5) 重讀逐列比對十六欄
    ws2 = openpyxl.load_workbook(BOOK)[SHEET]
    bad = []
    for i, row in enumerate(rows):
        for k, idx in COL_IDX.items():
            if k not in row:
                continue
            got = ws2.cell(row=FIRST_DATA_ROW + i, column=idx).value
            want = row[k] or None
            if (str(got) if got is not None else None) != (
                    str(want) if want is not None else None):
                bad.append((FIRST_DATA_ROW + i, k, repr(want)[:60], repr(got)[:60]))
    if bad:
        shutil.copy(BACKUP, BOOK)
        print(f"**重讀比對不符 {len(bad)} 格，已自 {BACKUP.name} 還原**")
        for b in bad[:10]:
            print("   ", b)
        return 1
    print(f"重讀逐列比對十六欄：**{len(rows)} 列 × 16 欄，0 項不符**")

    after_hash = hashlib.sha256(BOOK.read_bytes()).hexdigest()
    print(f"寫後 sha256：{after_hash}")
    json.dump({"rows": len(rows), "generated": stat["generated"],
               "not_generated": stat["not_generated"], "held_out_in_ah": len(ho),
               "sha256_before": before_hash, "sha256_after": after_hash,
               "seven_before": before, "seven_after": after},
              (FEAT / "data/_w155_fullwrite2.json").open("w"),
              ensure_ascii=False, indent=1)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
