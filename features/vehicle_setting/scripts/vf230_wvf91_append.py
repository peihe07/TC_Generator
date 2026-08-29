"""VF230：19 條之**追加寫入**（W-VF91.1／.3；裁定 R-VF142）。

**其與 `vf230_wvf84_writeback.py` 之別**：後者之閘 2 要求「目標資料列為 0」——
**追加寫入必然被其攔下，而該攔下為正**（目標現有 438 列資料）。
**故本檔另立四閘，不修改該檔之閘** —— `R-VF130`：
一個閘若其保守性正是他人所需，不得為遷就本次之需要而放寬之。

**四閘，任一不成即停**：
  閘 1  B 欄續號起點**由實測導出**（讀目標之 B 欄末號），
        **不採信任何推算之值**（R-VF140 明令；R-VF114 之精神）
  閘 2  **待寫之列必須為空** —— 其若有資料即停手回報
  閘 3  x14 之保全：openpyxl 存檔後補回 `<extLst>`，補回後複驗 x14dv 之數
  閘 4  寫入後**重讀全檔**逐欄比對 —— **既有 438 列與新增 19 列一併驗**，
        既有列之任一格相異即停（**追加不得改動既有列**）

**R-VF116**：寫入後之 Excel 確認與複製至交付路徑為 Pei 之動作，本檔明示提示。
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string as ci

FEAT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(FEAT / "scripts"))
import vf230_wvf77_dryrun as DRY          # noqa: E402  單一權威：欄映射與列之來源
import vf230_wvf84_writeback as WB        # noqa: E402  復用 probe_x14／restore_extlst

SHEET, FIRST = DRY.SHEET, DRY.FIRST_DATA_ROW
SRC = FEAT / "generated/vf230_backfill.json"
BACKUP = FEAT / "inputs/_vf230_036_append_backup.xlsx"


def snapshot(book: Path) -> dict[int, dict[str, str]]:
    """讀目標之全部資料列，回 {B: {欄: 值}}。**閘 4 之既有列基準。**"""
    wb = openpyxl.load_workbook(book, read_only=True, data_only=True)
    ws = wb[SHEET]
    out = {}
    for row in ws.iter_rows(min_row=FIRST):
        b = row[1].value
        if b in (None, "") or row[3].value in (None, ""):
            continue
        out[int(b)] = {c: ("" if row[ci(c) - 1].value is None
                           else str(row[ci(c) - 1].value)) for c in DRY.COLS}
    wb.close()
    return out


def main() -> None:
    book = DRY.book()
    print("=== W-VF91 追加寫入（裁定 R-VF142；目標為 **repo 內副本**）===")
    print(f"目標：{book}")

    before = snapshot(book)
    if not before:
        raise SystemExit("目標無任何資料列 —— 追加之前提不成立，停")
    b_last = max(before)
    row_last = FIRST + len(before) - 1
    print(f"\n[閘 1] 目標現況（**實測**）：資料 {len(before)} 列｜"
          f"B {min(before)}–{b_last}｜末列 {row_last}")
    if sorted(before) != list(range(min(before), b_last + 1)):
        raise SystemExit("目標之 B 欄非連號 —— 續號之起點不可靠，停")
    b_start = b_last + 1
    print(f"       續號起點 **{b_start}**（= 實測末號 {b_last} + 1，不採信推算值）")

    rows = DRY.rows_from([SRC], b_start)
    print(f"       來源 {SRC.name}：{len(rows)} 列，B {rows[0]['B']}–{rows[-1]['B']}")

    # ---- 閘 2：待寫之列必須為空 ----
    wb = openpyxl.load_workbook(book)
    ws = wb[SHEET]
    first_new = row_last + 1
    busy = []
    for i in range(len(rows)):
        rn = first_new + i
        for c in DRY.COLS:
            v = ws.cell(row=rn, column=ci(c)).value
            if v not in (None, "") and not (isinstance(v, str)
                                            and v.lstrip().startswith("=")):
                busy.append(f"{c}{rn}={v!r}")
    print(f"\n[閘 2] 待寫之列 {first_new}–{first_new + len(rows) - 1}："
          f"**非空之格 {len(busy)}**")
    if busy:
        raise SystemExit(f"待寫之列已有資料：{busy[:5]} —— 停手回報")

    shutil.copy2(book, BACKUP)
    print(f"       寫入前備份 → {BACKUP.name}")

    x14_before = WB.probe_x14(book)
    print(f"[閘 3] 寫入前 x14dv：{x14_before}")

    for i, r in enumerate(rows):
        rn = first_new + i
        for c in DRY.COLS:
            v = r.get(c, "")
            if not str(v).strip():
                continue
            ws.cell(row=rn, column=ci(c)).value = v
        # 樣式取自首個資料列 —— **`wrap_text` 缺席即為使用者所見之「編號前多空格」**
        # （W-VF89 之同一形態，其成因為樣板容量 237 列）。
        for col in range(1, 35):
            src = ws.cell(row=FIRST, column=col)
            if src.alignment.wrap_text is None:
                continue
            dst = ws.cell(row=rn, column=col)
            if dst.alignment.wrap_text is None or dst.font.name != src.font.name:
                from copy import copy
                dst._style = copy(src._style)
    wb.save(book)
    wb.close()
    print(f"       存檔後 x14dv：{WB.probe_x14(book)}")
    WB.restore_extlst(BACKUP, book)
    x14_after = WB.probe_x14(book)
    print(f"       補回後 x14dv：**{x14_after}**（原 {x14_before}）")
    if x14_after != x14_before:
        raise SystemExit("x14 之保全失敗（補回後之數與原檔不符），停")

    # ---- 閘 4：既有列與新增列一併重讀比對 ----
    after = snapshot(book)
    diff_old = [(b, c) for b, v in before.items() for c in DRY.COLS
                if b not in after or after[b][c].strip() != v[c].strip()]
    diff_new = []
    for r in rows:
        g = after.get(r["B"])
        if g is None:
            diff_new.append((r["B"], "整列缺"))
            continue
        for c in DRY.COLS:
            if str(r[c]).strip() != g[c].strip():
                diff_new.append((r["B"], c))
    print(f"\n[閘 4] 重讀比對 {len(after)} 列 × {len(DRY.COLS)} 欄")
    print(f"       **既有 {len(before)} 列之差異：{len(diff_old)}**（追加不得改動既有列）")
    print(f"       **新增 {len(rows)} 列之差異：{len(diff_new)}**")
    for d in (diff_old + diff_new)[:6]:
        print(f"       B{d[0]} 欄 {d[1]}")
    if diff_old or diff_new:
        raise SystemExit("重讀比對有差異，停")

    print(f"\n追加 {len(rows)} 列，B {rows[0]['B']}–{rows[-1]['B']}；"
          f"合計 **{len(after)}** 列。四閘全過。")
    print("\n" + "=" * 66)
    print("**R-VF112／R-VF116：以下為 Pei 之動作，本層不得代之**")
    print("  1. 以 **Excel** 開啟並確認下拉與版面")
    print("  2. **逐字具名檔名**複製至交付路徑 VF230_V1_R5/")
    print("     **不得以萬用字元或目錄複製**")
    print("=" * 66)


if __name__ == "__main__":
    main()
