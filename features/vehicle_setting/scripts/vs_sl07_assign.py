#!/usr/bin/env python3
"""VS-SL-07 §2 —— 賦號與前綴正名（對 `output/` 三本轉正本施作）。

  vf230  438 列 F 欄賦 `NR1L-VS-001` 起連號
  BL／VC `newR1L-{ABBR}-{nnn}` → `NR1L-{ABBR}-{nnn}`，數字部不動

**賦號之序與包內字面不同，理由（實測，見上繳 §5）**：
包內令「依 D 欄升冪」。惟
  (1) 本簿 D 欄**非非遞減**（438 列，首處逆序於第 1 列）；
  (2) R-G42 之官方比較器 `scripts/lint_delivery_spec.py:req_key()` 對本簿之 D 值
      **直接拋 TypeError** —— `SWE1-VC-6AuxSwitches-002` 之鍵為
      `('SWE',1,'VC',6,'AuxSwitches',2)`，`SWE1-VC-AutoDoorLocks-015` 之鍵為
      `('SWE',1,'VC','AutoDoorLocks',15)`，第 4 位 int 與 str 不可比。
故「D 欄升冪」在本批資料上**未定義**。本層改依**列序**（即文件序）賦號 ——
與現行 `delivered/` 之 ICS 交付本形制一致（其 F 欄逐列 `NR1L-ICS-001`…）。
**排序若須改，重跑本腳本即可（備份 `.presl07.bak` 保留原稿）。**
"""

from __future__ import annotations

import csv
import hashlib
import re
import shutil
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import vs_sl01_dryrun as v1  # noqa: E402

ROOT = v1.ROOT
COL_TC = 6                                     # F 欄
OLD_PREFIX = re.compile(r"^newR1L-")

BOOKS = {
    "vehicle_setting": {
        "book": "features/vehicle_setting/output/FM-WI-FSM-036-A01 STLA 測試用例規範與結果_"
                "SWQT STLA Test Case Specification & Result_SWQT_VF230_20260902.xlsx",
        "map": "features/vehicle_setting/reports/testrail_id_map_vehicle_setting.tsv",
        "mode": "assign", "abbr": "VS",
        "removed": "features/vehicle_setting/reports/vf230_removed_non_nafta.tsv"},
    "bed_lowering": {
        "book": "features/bed_lowering/output/FM-WI-FSM-036-A01 STLA 測試用例規範與結果_"
                "SWQT STLA Test Case Specification & Result_SWQT_BedLowering_20260902.xlsx",
        "map": "features/bed_lowering/reports/testrail_id_map_bed_lowering.tsv",
        "mode": "rename", "abbr": "BLM", "removed": None},
    "vehicle_category": {
        "book": "features/vehicle_category/output/FM-WI-FSM-036-A01 STLA 測試用例規範與結果_"
                "SWQT STLA Test Case Specification & Result_SWQT_VehicleCategory_20260902_working.xlsx",
        "map": "features/vehicle_category/reports/testrail_id_map_vehicle_category.tsv",
        "mode": "rename", "abbr": "VC", "removed": None},
}
MAP_COLS = ["row", "req_id", "old_E", "old_F", "new_F", "status", "note"]
MANIFEST_COLS = ["filename", "sha256", "source_path", "round", "status", "note"]


def sha256_of(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def run(slug: str) -> dict:
    cfg = BOOKS[slug]
    book = ROOT / cfg["book"]
    bak = book.with_suffix(".presl07.bak.xlsx")
    if not bak.exists():                        # 不覆寫既有備份
        shutil.copy2(book, bak)

    # 資料列一律以 data_only 讀取判定 —— 直接讀公式儲存格會把公式字串
    # 當成有值，BL／VC 遂由 151／126 膨脹為 1,402（本輪曾犯，見上繳 §5）。
    ws_ro, g_ro = v1.cells(book)
    rows = v1.data_rows(ws_ro, g_ro)

    wbk = openpyxl.load_workbook(book)
    ws = wbk[v1.SHEET]
    recs, changed = [], 0
    for n, r in enumerate(rows, 1):
        old = str(ws.cell(r, COL_TC).value or "").strip()
        req = str(ws.cell(r, 4).value or "").strip()
        old_e = str(ws.cell(r, 5).value or "").strip()
        if cfg["mode"] == "assign":
            new = f"NR1L-{cfg['abbr']}-{n:03d}"
            status, note = "assigned", "依列序（文件序）連號；見上繳 §5 之排序說明"
        else:
            new = OLD_PREFIX.sub("NR1L-", old)
            status = "renamed" if new != old else "kept"
            note = "R-VS95：前綴正名，數字部不動" if status == "renamed" else ""
        if new != old:
            ws.cell(r, COL_TC).value = new
            changed += 1
        recs.append({"row": r, "req_id": req, "old_E": old_e, "old_F": old,
                     "new_F": new, "status": status, "note": note})
    wbk.save(book)

    retired = 0
    if cfg["removed"]:
        for x in csv.DictReader(open(ROOT / cfg["removed"]), delimiter="\t"):
            recs.append({"row": x["row"], "req_id": x["req_id"], "old_E": "",
                         "old_F": "", "new_F": "", "status": "RETIRED",
                         "note": f"{x['reason']}（VS-SL-03 移除，未入轉正本，不賦號）"})
            retired += 1

    with open(ROOT / cfg["map"], "w", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MAP_COLS, delimiter="\t")
        w.writeheader()
        w.writerows(recs)

    sha = sha256_of(book)
    man = book.parent / "MANIFEST.tsv"
    mrows = list(csv.DictReader(open(man), delimiter="\t")) if man.exists() else []
    for m in mrows:
        if m["filename"] == book.name:
            m["sha256"] = sha
            m["round"] = "VS-SL-07"
            m["note"] = "賦號／正名後（R-VS94／R-VS95）；出貨 gate 未過"
    with open(man, "w", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MANIFEST_COLS, delimiter="\t")
        w.writeheader()
        w.writerows(mrows)

    return {"slug": slug, "rows": len(rows), "changed": changed, "retired": retired,
            "map_rows": len(recs), "sha": sha,
            "ids": [r["new_F"] for r in recs if r["status"] != "RETIRED"]}


def main() -> int:
    for slug, cfg in BOOKS.items():
        s = run(slug)
        ids = [i for i in s["ids"] if i]
        assert len(set(ids)) == len(ids), f"{slug}：F 欄有重號"
        assert all(re.fullmatch(r"NR1L-[A-Za-z]+-\d{3}", i) for i in ids), \
            f"{slug}：有 F 值不合 R-G42 形制"
        assert s["map_rows"] == s["rows"] + s["retired"], f"{slug}：對照表列數不符"
        if cfg["mode"] == "assign":
            nums = [int(i.rsplit("-", 1)[1]) for i in ids]
            assert nums == list(range(1, len(nums) + 1)), f"{slug}：號非自 001 起連號"
        print(f"{slug:<18} {s['rows']:>4} 列  改 {s['changed']:>4}  RETIRED {s['retired']:>2}  "
              f"對照表 {s['map_rows']:>4}  {ids[0]}…{ids[-1]}")
        print(f"{'':<18} SHA {s['sha'][:32]}…")
    return 0


if __name__ == "__main__":
    sys.exit(main())
