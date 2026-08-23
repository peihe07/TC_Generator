"""W-120（67 包）—— 036 工作簿之寫回器 ＋ **dry-run**。

母本：`.../CFTS044/FM-WI-FSM-036-A01 …_CFTS044_Vehicle Controls_20260819.xlsx`
分頁 `Test Case Specification 測試用例規範`；**表頭列 9；資料列 10 起**。

**本檔預設只做 dry-run。實寫須 `--write` 且待 Pei 核可（67 包之閘）。**

欄位對映（66 包 §3）：只寫下列 16 欄，其餘一格不動。
`reasoning` 不入工作簿；E／O／Q／S／T–Z／AB–AG 留空。
"""
from __future__ import annotations

import collections
import csv
import json
import re
import sys
from pathlib import Path

FEAT = Path(__file__).resolve().parents[1]
BOOK = Path("/Users/peihe/Work/02_Project_R1LR/10_Reviewing/00_TestCase/ASW-R2/"
            "Vehicle Settings/CFTS044/FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT "
            "STLA Test Case Specification & Result_SWQT_CFTS044_Vehicle Controls_20260819.xlsx")
SHEET = "Test Case Specification 測試用例規範"
HEADER_ROW, FIRST_DATA_ROW = 9, 10

COLS = {  # 欄字母 → (欄名, 取值來源)
    "B": "No.#", "C": "Requirement or Design", "D": "Requirement or Design ID",
    "F": "Test Case ID", "G": "Test Group", "H": "Test Set", "I": "Test Item",
    "J": "Pre-Conditions", "K": "Input Test Data", "L": "Test procedure",
    "M": "Expected Result", "N": "Specification Reference", "P": "Test Case Priority",
    "R": "Test Case Design", "AA": "Test Case Author", "AH": "Remarks",
}
CONTROLLED = {
    "功能測試 (Functional based ; no specific technique)", "狀態轉換 (State Transition Testing)",
    "決策表 (Decision Table Testing)", "等價劃分 (Equivalence Partitioning, EP)",
    "邊界值分析 (Boundary Value Analysis, BVA)",
    "組合測試 (Combinatorial Testing ; Pairwise / t-wise)",
    "情境 / 用例 (Scenario / Use Case Testing)", "負向測試 (Negative / Invalid)",
    "基礎故障注入 (Fault Injection Lite)",
}
PROJECT, ABBR = "NR1L", "VS"
# **D-4（47 輪，Pei 2026-08-23）**：`AA` 欄之作者姓名。46 輪以 `<AUTHOR>` 佔位。
AUTHOR = "PeiPYHsu"


def latest_batches() -> list[Path]:
    groups: dict[str, list] = collections.defaultdict(list)
    for f in (FEAT / "generated").glob("batch*.json"):
        m = re.match(r"(batch\d+)(?:_v(\d+))?\.json$", f.name)
        if m:
            groups[m.group(1)].append((int(m.group(2) or 1), f))
    return [max(v)[1] for k, v in sorted(groups.items())]


def req_title() -> dict[str, str]:
    """C 欄之來源 —— 037 之 `Requirement Title`（本 repo 已解為 leaf → title）。"""
    p = FEAT / "data/leaves.tsv"
    out = {}
    with p.open(encoding="utf-8") as f:
        for r in csv.DictReader(f, delimiter="\t"):
            key = r.get("swe_id") or r.get("leaf_id") or r.get("id")
            title = r.get("title") or r.get("requirement_title") or ""
            if key:
                out[key] = title
    return out


def rows_from_json(files: list[Path]) -> list[dict]:
    titles = req_title()
    disclosure = {}
    p = FEAT / "docs/reports/delivery_disclosure.md"
    if p.exists():
        for line in p.read_text(encoding="utf-8").split("\n"):
            m = re.match(r"\| `batch\d+\w*` \| `([^`]+)` \| ([^|]*)\|", line)
            if m:
                disclosure[m.group(1)] = m.group(2).strip()
    out, seq = [], 0
    for f in files:
        for tc in json.loads(f.read_text(encoding="utf-8"))["tcs"]:
            seq += 1
            leaf = tc["leaf_id"]
            pending = "PENDING" in tc["expected_result"]
            out.append({
                "B": seq, "C": titles.get(leaf, ""), "D": leaf,
                "F": f"{PROJECT}-{ABBR}-{seq:03d}", "G": "Vehicle Setting",
                "H": tc["test_set"], "I": tc["test_item"], "J": tc["pre_conditions"],
                "K": tc["input_test_data"], "L": tc["test_procedure"],
                "M": tc["expected_result"], "N": tc["specification_reference"],
                "P": tc["priority"], "R": tc["design_method"], "AA": AUTHOR,
                # **48 輪 D-3**：AH 之來源改取 `remarks`（其已載 BLOCKED／IMPL_GAP
                # 之逐條註記）；原以 `dr_dependent` 推導者為 42 輪之權宜。
                "AH": str(tc.get("remarks", "")).strip(),
            })
    return out


def dry_run(rows: list[dict], label: str) -> dict:
    n = len(rows)
    nonempty = {k: sum(1 for r in rows if str(r[k]).strip()) for k in COLS}
    nlines = collections.Counter(len(str(r["N"]).split("\n")) for r in rows)
    two_part = sum(1 for r in rows if "\n\n" in str(r["I"]))
    paren = sum(1 for r in rows if str(r["I"]).rstrip().endswith(")"))
    k_bad = sum(1 for r in rows if str(r["K"]).strip() != "NA")
    n_bad = sum(1 for r in rows if re.search(r"[,;] *CFTS", str(r["N"])))
    p_cnt = collections.Counter(r["P"] for r in rows)
    r_bad = sum(1 for r in rows if r["R"] not in CONTROLLED)
    ah = sum(1 for r in rows if str(r["AH"]).strip())
    return {"label": label, "rows": n, "nonempty": nonempty, "n_lines": dict(nlines),
            "I_two_part": two_part, "I_paren": paren, "K_not_NA": k_bad,
            "N_comma_joined": n_bad, "P": dict(p_cnt), "R_out_of_domain": r_bad,
            "AH_nonempty": ah}


COL_IDX = {  # 欄字母 → openpyxl 之 1-based 欄號
    "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8, "I": 9, "J": 10,
    "K": 11, "L": 12, "M": 13, "N": 14, "O": 15, "P": 16, "Q": 17, "R": 18,
    "S": 19, "T": 20, "U": 21, "V": 22, "W": 23, "X": 24, "Y": 25, "Z": 26,
    "AA": 27, "AB": 28, "AC": 29, "AD": 30, "AE": 31, "AF": 32, "AG": 33, "AH": 34,
}


def write_back(rows: list[dict]) -> tuple[int, int]:
    """W-137（48 輪）—— 實寫。

    (1) 先清空現有資料列之 **B–AH**（R-VS1：效力 BLANK、全欄重生）
    (2) 自列 10 起 append，依 66 包 §3 之欄位對映
    **列 1–9、其他分頁、B–AH 以外之欄一格不動。**
    """
    import openpyxl
    wb = openpyxl.load_workbook(BOOK)
    ws = wb[SHEET]
    cleared = 0
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        touched = False
        for c in range(COL_IDX["B"], COL_IDX["AH"] + 1):
            if ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).value = None
                touched = True
        cleared += touched
    for i, row in enumerate(rows):
        r = FIRST_DATA_ROW + i
        for k, idx in COL_IDX.items():
            if k in row:
                ws.cell(row=r, column=idx).value = row[k]
    wb.save(BOOK)
    return cleared, len(rows)


def verify(rows: list[dict]) -> list[str]:
    """W-137(3)：重讀，逐列比對 JSON 與工作簿之十六欄。"""
    import openpyxl
    ws = openpyxl.load_workbook(BOOK, data_only=True)[SHEET]
    bad = []
    for i, row in enumerate(rows):
        r = FIRST_DATA_ROW + i
        for k in COLS:
            got = ws.cell(row=r, column=COL_IDX[k]).value
            want = row[k]
            if (str(got) if got is not None else "") != (str(want) if want is not None else ""):
                bad.append(f"列 {r} 欄 {k}：工作簿 {got!r} ／ JSON {want!r}")
    return bad


def main() -> None:
    files = latest_batches()
    rows = rows_from_json(files)
    if "--write" in sys.argv:
        import hashlib
        h = hashlib.sha256(BOOK.read_bytes()).hexdigest()
        expect = "ebe5a65f30a0d4bcf9e46b51a43145ce222027ac49ad523fe5c2d2b6566a5089"
        print(f"實寫前之母本 sha256：{h}")
        if h != expect:
            raise SystemExit(f"⚠ 母本雜湊不等於 {expect} —— 依 W-136(2) 中止")
        cleared, n = write_back(rows)
        print(f"清空 {cleared} 列之 B–AH；自列 {FIRST_DATA_ROW} 起寫入 {n} 列")
        bad = verify(rows)
        print(f"重讀比對：不符 **{len(bad)}** 處")
        for b in bad[:20]:
            print("   ", b)
        if bad:
            raise SystemExit("⚠ 重讀比對不符 —— 依 W-137(3) 中止，請自 "
                             "REF/036_pre_writeback_20260823.xlsx 還原")
        print(f"實寫後之 sha256：{hashlib.sha256(BOOK.read_bytes()).hexdigest()}")
        return
    subj = dry_run(rows, "正常輸入")

    # ── R-VS54 錨點：刻意違規之 JSON（四處植入）──────────────────
    bad = [dict(r) for r in rows]
    bad[0]["K"] = "None"                                   # K 欄非 NA
    # N 欄逗號串接 —— 須挑**多值列**植入，否則單值列改不出違規
    multi = next((i for i, r in enumerate(rows) if "\n" in str(r["N"])), 1)
    bad[multi]["N"] = str(bad[multi]["N"]).replace("\n", ", ")
    bad[2]["I"] = str(bad[2]["I"]).replace("\n\n", " ")    # I 欄無空行
    bad[3]["R"] = "State Transition"                        # R 欄純英文
    anch = dry_run(bad, "刻意違規之錨點")

    out = ["# writeback dry-run（W-120，42 輪）", "",
           f"母本：`{BOOK.name}`", f"分頁：`{SHEET}`；表頭列 {HEADER_ROW}；"
           f"資料列 {FIRST_DATA_ROW} 起", "",
           "**本輪不實寫**（67 包之閘：dry-run 通過 ＋ Pei 核可 ＋ 母本備份）。", "",
           "## (1) 將寫入之列數", "",
           f"| 項 | 值 |", "|---|---:|",
           f"| 將寫入 | **{subj['rows']}** |", f"| 對照（交付累計） | 139 |",
           f"| 差 | **{subj['rows'] - 139}** |", "",
           "## (2) 逐欄非空數（16 欄）", "",
           "| 欄 | 欄名 | 非空 | 應為 |", "|---|---|---:|---:|"]
    for k, name in COLS.items():
        out.append(f"| `{k}` | {name} | {subj['nonempty'][k]} | "
                   f"{subj['rows'] if k not in ('AH',) else 21} |")
    out += ["", "## (3) N 欄之多值列 —— 行數分布", "", "| 行數 | 列數 |", "|---:|---:|"]
    for k in sorted(subj["n_lines"]):
        out.append(f"| {k} | {subj['n_lines'][k]} |")
    out += ["", "## (4)–(7) 判準檢查（正常輸入 vs 錨點並列）", "",
            "| 判準 | 應為 | 正常輸入 | 錨點（刻意違規） | 判 |", "|---|---|---:|---:|---|",
            f"| (4) I 欄含空行 | {subj['rows']} | {subj['I_two_part']} | "
            f"{anch['I_two_part']} | "
            f"{'PASS，可失敗' if subj['I_two_part'] == subj['rows'] and anch['I_two_part'] < subj['rows'] else '⚠'} |",
            f"| (4) I 欄括號收尾 | {subj['rows']} | {subj['I_paren']} | {anch['I_paren']} | "
            f"{'PASS' if subj['I_paren'] == subj['rows'] else '⚠'} |",
            f"| (5) K 欄非 `NA` | 0 | {subj['K_not_NA']} | {anch['K_not_NA']} | "
            f"{'PASS，可失敗' if subj['K_not_NA'] == 0 and anch['K_not_NA'] > 0 else '⚠'} |",
            f"| (6) R 欄不在受控 9 值 | 0 | {subj['R_out_of_domain']} | "
            f"{anch['R_out_of_domain']} | "
            f"{'PASS，可失敗' if subj['R_out_of_domain'] == 0 and anch['R_out_of_domain'] > 0 else '⚠'} |",
            f"| （附）N 欄逗號串接 | 0 | {subj['N_comma_joined']} | "
            f"{anch['N_comma_joined']} | "
            f"{'PASS，可失敗' if subj['N_comma_joined'] == 0 and anch['N_comma_joined'] > 0 else '⚠'} |",
            f"| (7) AH 欄非空 | 21 | {subj['AH_nonempty']} | {anch['AH_nonempty']} | "
            f"{'PASS' if subj['AH_nonempty'] == 21 else '⚠ 見上繳 37 §2'} |", "",
            "## (6) P 欄之三級計數", "", "| 級 | 列數 |", "|---|---:|"]
    for k in sorted(subj["P"]):
        out.append(f"| {k} | {subj['P'][k]} |")
    out += ["", "## (8) 母本現有列將被清空之欄範圍", ""]
    try:
        import openpyxl
        ws = openpyxl.load_workbook(BOOK, read_only=True, data_only=True)[SHEET]
        cnt, tot = collections.Counter(), 0
        for i, row in enumerate(ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True),
                                start=FIRST_DATA_ROW):
            if not any(c is not None for c in row):
                continue
            tot += 1
            for k in COLS:
                idx = (ord(k[-1]) - 65) + (26 if len(k) == 2 else 0)
                if idx < len(row) and row[idx] is not None and str(row[idx]).strip():
                    cnt[k] += 1
        out += [f"母本現有資料列：**{tot}**（清空範圍 B–AH）", "",
                "| 欄 | 清空前非空 |", "|---|---:|"]
        for k in COLS:
            out.append(f"| `{k}` | {cnt[k]} |")
    except Exception as e:                     # noqa: BLE001
        out.append(f"**母本讀取失敗**：{e}")
    (FEAT / "docs/reports/writeback_dryrun.md").write_text("\n".join(out), encoding="utf-8")
    print("\n".join(out[:4]))
    print(f"\n將寫入 {subj['rows']} 列；I 兩段 {subj['I_two_part']}／"
          f"K 非 NA {subj['K_not_NA']}／R 出域 {subj['R_out_of_domain']}／"
          f"AH 非空 {subj['AH_nonempty']}")
    print(f"錨點：I 兩段 {anch['I_two_part']}／K 非 NA {anch['K_not_NA']}／"
          f"R 出域 {anch['R_out_of_domain']}／N 逗號 {anch['N_comma_joined']}")


if __name__ == "__main__":
    main()
