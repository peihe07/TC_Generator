#!/usr/bin/env python3
"""VS-SL-05 —— 送抽前修正：Pre 矛盾 52 列＋R4 否定式代表值 7 列＋v5 報告。

依 `down/20260902_VS-SL-04_review.md` 之第二、三節。**BL／VC 沙盒稿不動。**

  一、Pre 含 `The Vehicle Settings menu is open` **且** Procedure 已插本輪之
      `Press "Settings" on Menu Bar` 導覽段者，刪該 Pre 行並重編號。
      Pre 斷言之狀態隨後由步驟建立，且屬步驟可控狀態，依 IN §4.4 不得為 Pre。
      **未插導覽之列不動**（原稿無導覽段時該 Pre 行自洽）。

  二、R4：否定式條件取規格值表內之代表值（FO §8.3 三層檢驗所本之 EP 代表值選取，非造值）。
      No.149 `Country_Code is not [Australia]` —— 值表 15 個值中**無 Australia**，
             取 `2 (United States of America)`（R-VS84 NAFTA 之自然代表）
      No.268 `Trailer_Light_Check is NOT [Absent]` —— 取 `1 (Type 1 (Radio))`，
             兄弟值 `2 (Type 2)`／`3 (Type 3)` 註於提議之註記段
"""

from __future__ import annotations

import csv
import re
import shutil
import sys
import warnings
from collections import Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import vs_sl01_dryrun as v1  # noqa: E402
import vs_sl03_writeback as wb  # noqa: E402

warnings.filterwarnings("ignore")
ROOT = v1.ROOT
V4 = "features/vehicle_setting/reports/vf230_settings_dryrun_v4.tsv"
V5 = "features/vehicle_setting/reports/vf230_settings_dryrun_v5.tsv"
RESOLUTION = "features/vehicle_setting/reports/_v4_branch3_resolution.tsv"
PRE_FIX = "features/vehicle_setting/reports/_v5_pre_fix.tsv"
SANDBOX = "features/vehicle_setting/sandbox/vssl/vf230_vssl.xlsx"
BACKUP = "features/vehicle_setting/sandbox/vssl/vf230_vssl.v4.bak.xlsx"

PRE_LINE = "The Vehicle Settings menu is open"
NAV_STEP = 'Press "Settings" on Menu Bar'
BRANCH_R4 = "(2e) 否定式 EP 代表值"

# 設定名 → (提議, 註記, 證據)。二者皆經本層複驗值表。
R4 = {
    "Engine Off Power Delay": (
        "PROXI Country_Code = 2 (United States of America)",
        "EP 代表值；`Country_Code` 值表 15 個值中無 Australia，故其補集即全表",
        "总控表 No.149 `If Country_Code is not [Australia]`；"
        "proxi_values['Country_Code']['2'] = 'United States of America'"),
    "Automatic Trailer Light Check": (
        "PROXI Trailer_Light_Check = 1 (Type 1 (Radio))",
        "EP 代表值；兄弟值 2 (Type 2)／3 (Type 3)",
        "总控表 No.268 `If Trailer_Light_Check is NOT [Absent])`；"
        "proxi_values['Trailer_Light_Check'] = {0: Absent, 1: Type 1 (Radio), "
        "2: Type 2, 3: Type 3}"),
}


def update_resolution() -> list[str]:
    """`_v4_branch3_resolution.tsv` 就地改判 R4。回受影響之 row 清單。

    **冪等**：已改判為 R4 之列視同命中，故本腳本可重跑（審閱端得以重現）。
    """
    p = ROOT / RESOLUTION
    rows = list(csv.DictReader(open(p), delimiter="\t"))
    hit = []
    for r in rows:
        if r["subcase"] not in ("NEGATED_CONDITION", "NEGATED_EP_REPRESENTATIVE"):
            continue
        proposal, note, ev = R4[r["setting"]]
        r["resolution"] = "R4"
        r["subcase"] = "NEGATED_EP_REPRESENTATIVE"
        r["proxi_added"] = proposal
        r["evidence"] = f"{ev}；{note}（審閱 VS-SL-04 §2）"
        hit.append(r["row"])
    with open(p, "w", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0]), delimiter="\t")
        w.writeheader()
        w.writerows(rows)
    return hit


def build_v5(r4_rows: set[str]) -> list[dict]:
    res = {r["row"]: r for r in csv.DictReader(open(ROOT / RESOLUTION), delimiter="\t")}
    out = []
    for row in csv.DictReader(open(ROOT / V4), delimiter="\t"):
        if row["row"] in r4_rows:
            r = res[row["row"]]
            note = R4[row["setting"]][1]
            flags = {f for f in row["flags"].split("｜")[0].split(";") if f}
            flags.discard("PROXI_PENDING")
            flags.discard("NEGATED_CONDITION")
            flags.add("NEGATED_EP_REPRESENTATIVE")
            row["proxi_proposed"] = f"{r['proxi_added']}｜{note}"
            row["flags"] = ";".join(sorted(flags)) + f"｜分支 {BRANCH_R4}"
        out.append(row)
    return out


def fix_pre(path: Path) -> list[dict]:
    """刪 Pre 之 `menu is open` 行並重編號；只動已插導覽之列。"""
    wbk = openpyxl.load_workbook(path)
    ws = wbk[v1.SHEET]
    g = lambda r, c: ("" if ws.cell(r, c).value is None else str(ws.cell(r, c).value))
    hits = []
    for r in range(v1.HEADER_ROW + 1, ws.max_row + 1):
        if not (g(r, 2).strip() or g(r, 4).strip()):
            continue
        pre, proc = g(r, 10), g(r, 12)
        if PRE_LINE not in pre or NAV_STEP not in proc:
            continue
        lines = wb.lines(pre)
        idx = [i for i, x in enumerate(lines) if x == PRE_LINE]
        if not idx:
            continue
        kept = [x for i, x in enumerate(lines) if i not in set(idx)]
        ws.cell(r, 10).value = wb.numbered(kept)
        hits.append({"row": r, "tc_id": g(r, 6) or "(F 欄空)",
                     "req_id": g(r, 4), "removed_line_no": idx[0] + 1,
                     "pre_lines_before": len(lines), "pre_lines_after": len(kept)})
    wbk.save(path)
    return hits


def main() -> int:
    r4 = update_resolution()
    assert len(r4) == 7, f"R4 為 {len(r4)} 列，非 7"

    rows = build_v5(set(r4))
    v1.write_tsv(ROOT / V5, rows)

    src, bak = ROOT / SANDBOX, ROOT / BACKUP
    if src.exists() and not bak.exists():          # 不覆寫既有備份
        shutil.copy2(src, bak)
    wb.BOOKS["vf230"] = (v1.VF230, SANDBOX, V5)
    s = wb.run("vf230")

    hits = fix_pre(src)
    with open(ROOT / PRE_FIX, "w", encoding="utf-8") as f:
        w = csv.DictWriter(f, delimiter="\t", fieldnames=[
            "row", "tc_id", "req_id", "removed_line_no",
            "pre_lines_before", "pre_lines_after"])
        w.writeheader()
        w.writerows(hits)

    assert all(h["pre_lines_before"] - h["pre_lines_after"] == 1 for h in hits), \
        "有列刪除之行數不為 1"
    pend = sum(1 for r in rows if "PROXI_PENDING" in r["flags"])
    assert pend == 25, f"PENDING 為 {pend}，非 25"

    br = Counter(r["flags"].split("｜分支 ")[1] for r in rows)
    print(f"R4 改判 {len(r4)} 列（assert = 7 PASS）")
    print(f"沙盒稿重生：{s['rows_before']} → {s['rows_after']} 列；改動 {s['cells']} 處")
    print(f"Pre 矛盾修正：{len(hits)} 列（每列刪 1 行，assert PASS）")
    print(f"v5 PENDING：{pend}（assert = 25 PASS）")
    print("v5 新分支:", {k: v for k, v in br.items() if not k.startswith("(2) 总控表")})
    return 0


if __name__ == "__main__":
    sys.exit(main())
