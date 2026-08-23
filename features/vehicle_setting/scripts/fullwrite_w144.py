"""W-144（77 包 §2）—— 全量寫回 237 leaf（R-VS72），**依 R-VS70 以外科式 emit**。

R-VS72：一 leaf 一列（拆分者多列），依 R-VS4 之 Test Set ＋ reqid 升冪排列。
  已生成者：十六欄照 66 包 §3 之對映
  未生成者：C／D／G／H／N 照填；I/J/K/L/M/P/R 留空；
            AH 記其阻塞類別與所待之 DR

R-VS70：**不得以 openpyxl 存檔** —— openpyxl 僅作計算層，
emit 走 `backend/xlsx_surgical.surgical_save`（原樣重打包、逐項驗結構）。
"""
from __future__ import annotations

import collections
import csv
import json
import re
import sys
from pathlib import Path

FEAT = Path(__file__).resolve().parents[1]
REPO = FEAT.parents[1]
sys.path.insert(0, str(REPO / "backend"))
sys.path.insert(0, str(FEAT / "scripts"))

from writeback_036 import (BOOK, COL_IDX, COLS, FIRST_DATA_ROW, SHEET,  # noqa: E402
                           AUTHOR, PROJECT, ABBR, latest_batches, req_title)
from xlsx_surgical import surgical_save                                  # noqa: E402

TESTSET_ORDER = {"Common Features": 0, "Heated Seat": 1, "Vented Seat": 2,
                 "Heated Steering Wheel": 3}


def build_rows() -> tuple[list[dict], dict]:
    titles = req_title()
    l2r = {r["swe_id"]: r for r in csv.DictReader(
        (FEAT / "data/leaf_to_reqid.tsv").open(encoding="utf-8"), delimiter="\t")}
    gen = {r["leaf_id"]: r for r in csv.DictReader(
        (FEAT / "docs/reports/generatable.tsv").open(encoding="utf-8"), delimiter="\t")}
    wr = {r["leaf_id"]: r for r in csv.DictReader(
        (FEAT / "docs/reports/writability.tsv").open(encoding="utf-8"), delimiter="\t")}

    by_leaf: dict[str, list[dict]] = collections.defaultdict(list)
    for f in latest_batches():
        for tc in json.loads(f.read_text(encoding="utf-8"))["tcs"]:
            by_leaf[tc["leaf_id"]].append(tc)

    def reqid(leaf: str) -> str:
        m = re.findall(r"\d{7}", (l2r.get(leaf, {}).get("reqid_list") or ""))
        return m[0] if m else "9999999"

    ordered = sorted(gen, key=lambda l: (TESTSET_ORDER.get(gen[l]["layer2"], 9),
                                         reqid(l), l))
    rows, seq, stat = [], 0, collections.Counter()
    for leaf in ordered:
        g, w = gen[leaf], wr.get(leaf, {})
        spec = (l2r.get(leaf, {}).get("reqid_list") or "").replace(";", "\n")
        base = {"C": titles.get(leaf, ""), "D": leaf, "G": "Vehicle Setting",
                "H": g["layer2"], "N": spec}
        tcs = by_leaf.get(leaf, [])
        if tcs:
            for tc in tcs:
                seq += 1
                stat["generated"] += 1
                rows.append({**base, "B": seq, "F": f"{PROJECT}-{ABBR}-{seq:03d}",
                             "I": tc["test_item"], "J": tc["pre_conditions"],
                             "K": tc["input_test_data"], "L": tc["test_procedure"],
                             "M": tc["expected_result"], "P": tc["priority"],
                             "R": tc["design_method"], "AA": AUTHOR,
                             "AH": str(tc.get("remarks", "")).strip()})
        else:
            seq += 1
            stat["not_generated"] += 1
            blk = w.get("blocker_class") or ""
            why = (f"{blk} — {w.get('driver_reason', '')}" if blk
                   else (w.get("driver_reason") or "尚未生成"))
            dr = w.get("dr_dependent") or ""
            rows.append({**base, "B": seq, "F": f"{PROJECT}-{ABBR}-{seq:03d}",
                         "I": "", "J": "", "K": "", "L": "", "M": "", "P": "", "R": "",
                         "AA": AUTHOR,
                         "AH": f"NOT GENERATED: {why}" + (f"；見 {dr}" if dr else "")})
    return rows, stat


def main() -> None:
    import hashlib
    import openpyxl

    rows, stat = build_rows()
    print(f"列數 **{len(rows)}**（已生成 {stat['generated']}／"
          f"未生成 {stat['not_generated']}）；母體 leaf "
          f"{stat['generated'] and len({r['D'] for r in rows})}")

    before = hashlib.sha256(BOOK.read_bytes()).hexdigest()
    print(f"寫前 sha256：{before}")

    wb = openpyxl.load_workbook(BOOK)
    ws = wb[SHEET]
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        for c in range(COL_IDX["B"], COL_IDX["AH"] + 1):
            ws.cell(row=r, column=c).value = None
    for i, row in enumerate(rows):
        for k, idx in COL_IDX.items():
            if k in row:
                ws.cell(row=FIRST_DATA_ROW + i, column=idx).value = row[k] or None

    out = BOOK.with_suffix(".surgical.xlsx")
    report = surgical_save(wb, BOOK, out)          # R-VS70：外科式 emit ＋ 驗結構
    print("surgical_save 報告：", {k: v for k, v in report.items()
                                   if k in ("patched", "cells", "members")})
    import shutil
    shutil.move(out, BOOK)
    after = hashlib.sha256(BOOK.read_bytes()).hexdigest()
    print(f"寫後 sha256：{after}")
    json.dump({"rows": len(rows), "generated": stat["generated"],
               "not_generated": stat["not_generated"],
               "sha256_before": before, "sha256_after": after},
              (FEAT / "data/_w144_fullwrite.json").open("w"), ensure_ascii=False, indent=1)


if __name__ == "__main__":
    main()
