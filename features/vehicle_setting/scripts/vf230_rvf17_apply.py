"""R-VF17 —— A-VS118 之 4 leaf 由 W2 轉 W0（V07 §2）。

**本腳本為唯一施行點，且刻意設計為可重跑、可稽核、範圍受限。**

R-VF17 之三項限制：
  1. 只及於該 4 leaf（`HeatedSteeringWheelManagement-029/030/033/034`）
  2. 值域來源須逐字記錄（037 之 `Verification Method` 欄、reqid、值域）
  3. 不擴及其他 leaf；R-VF14(4) 之「不作全面回溯重跑」維持

依 **R-VF11** 附錨點，錨點不符即停：
  必命中   該 4 leaf 於施行前皆為 `W2` ／ `B6-value-absent`
  必不命中 `HeatedSteeringWheelManagement-023`（同 layer3 而非本條標的）
           施行前後分級不變

**durability 警告**：`writability.tsv`／`generatable.tsv` 由
`scripts/writability_driver.py --write` 產生，而該 driver 之
`value_sourced()` **尚未認 037 之 VC/VM 為值域來源**。故 driver 若重跑，
本變更會被回復為 W2。**本層未改 driver**（其為 Part 1 之產物且併行線
正在修改）。此風險已於上繳 V07 具名。
"""
import csv
import json
import re
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WRIT = ROOT / "docs" / "reports" / "writability.tsv"
GEN = ROOT / "docs" / "reports" / "generatable.tsv"

TARGET = {
    "SWE1-VC-HeatedSteeringWheelManagement-029": "4859496",
    "SWE1-VC-HeatedSteeringWheelManagement-030": "4859497",
    "SWE1-VC-HeatedSteeringWheelManagement-033": "4859500",
    "SWE1-VC-HeatedSteeringWheelManagement-034": "4859501"}
# 必不命中錨點須**在被掃描之檔內**，否則其值恆為 None，
# 「前後不變」與「檔內無此列」不可分辨 —— 首版誤選 -023（不在
# writability.tsv 內），與 W-VF18 首版之錯同型（A-VF4）。
ANCHOR_MISS = "SWE1-VC-HeatedSteeringWheelManagement-031"


def norm(v) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip().lower()


def vm_source() -> dict[str, str]:
    """自 037 逐字取該 4 leaf 之 `Verification Method` 內含 `HSW_Cmd_Tlm` 之行。"""
    out = {}
    import glob
    for f in sorted(glob.glob(str(ROOT / "inputs" / "FM-WI-FSM-037*CFTS044*.xlsx"))):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        for nm in wb.sheetnames:
            rs = list(wb[nm].iter_rows(values_only=True))
            i = next((j for j, r in enumerate(rs)
                      if any("requirement description" in norm(v) for v in r)), None)
            if i is None:
                continue
            h = rs[i]
            ivm = next((j for j, v in enumerate(h)
                        if "verification method" in norm(v)), None)
            for r in rs[i + 1:]:
                sid = str(r[0] or "").strip()
                if sid in TARGET and ivm is not None:
                    line = next((ln.strip() for ln in str(r[ivm] or "").splitlines()
                                 if "HSW_Cmd_Tlm" in ln), "")
                    out[sid] = line
            break
        wb.close()
    return out


def read_tsv(p: Path) -> tuple[list[str], list[dict]]:
    with p.open(encoding="utf-8") as fh:
        rd = csv.DictReader(fh, delimiter="\t")
        return list(rd.fieldnames or []), list(rd)


def write_tsv(p: Path, cols: list[str], rows: list[dict]) -> None:
    with p.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols, delimiter="\t",
                           lineterminator="\n", extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def main() -> None:
    src = vm_source()
    missing = set(TARGET) - set(src)
    if missing:
        raise SystemExit(f"037 之 Verification Method 取不到來源行：{missing}")
    for sid, line in src.items():
        if "HSW_Cmd_Tlm" not in line:
            raise SystemExit(f"{sid} 之來源行不含 HSW_Cmd_Tlm：{line!r}")

    wcols, wrows = read_tsv(WRIT)
    gcols, grows = read_tsv(GEN)
    by_w = {r["leaf_id"]: r for r in wrows}
    by_g = {r["leaf_id"]: r for r in grows}

    # --- R-VF11 錨點，先於施行 ---
    pre = {k: (by_w[k]["writable"], by_w[k].get("blocker_class", ""))
           for k in TARGET if k in by_w}
    if len(pre) != 4:
        raise SystemExit(f"4 leaf 未全在 writability.tsv：{sorted(set(TARGET)-set(pre))}")
    already = all(v[0] == "W0" for v in pre.values())
    if not already:
        for k, (w, b) in pre.items():
            if (w, b) != ("W2", "B6-value-absent"):
                raise SystemExit(f"必命中錨點不符：{k} 為 {w}/{b}，非 W2/B6-value-absent，停")
    if ANCHOR_MISS not in by_w:
        raise SystemExit(f"必不命中錨點 {ANCHOR_MISS} 不在 writability.tsv 內，"
                         "其值恆為 None，不構成錨點，停")
    anchor_miss_pre = by_w[ANCHOR_MISS]["writable"]
    print(f"R-VF11 錨點（前）：4 leaf {sorted({v[0] for v in pre.values()})}"
          f"／必不命中 {ANCHOR_MISS} = {anchor_miss_pre}")

    if already:
        print("4 leaf 已為 W0，無須再施行（本腳本可重跑）。")
        return

    shutil.copy(WRIT, WRIT.with_name("writability_pre_rvf17.tsv"))
    shutil.copy(GEN, GEN.with_name("generatable_pre_rvf17.tsv"))

    for sid, reqid in TARGET.items():
        note = (f"R-VF17：值域來源為 037 之 `Verification Method` 欄，"
                f"reqid {reqid}，逐字 `{src[sid]}`，解出 HSW_Cmd_Tlm 值域 {{ON, OFF}}")
        r = by_w[sid]
        r["writable"] = "W0"
        r["blocker_class"] = ""
        r["blocker_detail"] = ""
        prev = r.get("evidence_note", "")
        r["evidence_note"] = (prev + " ｜ " if prev else "") + note
        if sid in by_g:
            by_g[sid]["writable"] = "W0"

    if by_w[ANCHOR_MISS]["writable"] != anchor_miss_pre:
        raise SystemExit("必不命中錨點於施行後改變，停")

    write_tsv(WRIT, wcols, wrows)
    write_tsv(GEN, gcols, grows)

    post = {k: by_w[k]["writable"] for k in TARGET}
    print(f"R-VF11 錨點（後）：4 leaf {sorted(set(post.values()))}"
          f"／必不命中 {ANCHOR_MISS} = {by_w[ANCHOR_MISS]['writable']}")
    (ROOT / "data" / "_rvf17_apply.json").write_text(json.dumps({
        "targets": TARGET, "source_lines": src,
        "pre": {k: list(v) for k, v in pre.items()}, "post": post,
        "anchor_miss": {"leaf": ANCHOR_MISS, "pre": anchor_miss_pre,
                        "post": by_w[ANCHOR_MISS]["writable"]}},
        ensure_ascii=False, indent=2), encoding="utf-8")
    for sid in sorted(TARGET):
        print(f"  {sid}  W2 -> W0   來源 reqid {TARGET[sid]}")
    print("快照：writability_pre_rvf17.tsv ／ generatable_pre_rvf17.tsv")


if __name__ == "__main__":
    main()
