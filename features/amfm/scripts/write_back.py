#!/usr/bin/env python3
"""Step 4 (AMFM) — write the generated TCs into the FW036 Radio workbook.

AMFM is NOT Home's shape. Home rewrites three draft segments interleaved
between frozen rows; this workbook has **no draft rows at all** — 158 rows, all
authored by Wilson, tracing the superseded `SWE-RAD-*` family, and zero rows
covering the 102 ruled leaves (A-AM02). So the operation is an append after a
frozen prefix, and the invariant is correspondingly simpler and stricter:

    the 158 legacy rows must come out byte-identical, in the same order, in the
    same row positions, and nothing may be written above row 168.

R4 option (i) froze that region rather than adopting it, so it is excluded from
coverage and traceability claims — but it is still customer content in the file
we ship, which is why it is protected by an ordered CONTENT hash over columns
D..AG rather than by row numbers alone.

Two header-region corrections travel with the write, both outside the hashed
data range:

  - 範圍 Scope names the superseded SWRA-A02 report (A-AM01 / RULINGS C1),
    corrected to the 037-A03 filename by label text, exactly as Home A-H26.
  - ChangeHistory gains a revision row describing the append.

Column conventions differ from Home in ways worth stating rather than
discovering: column B carries a per-row formula (`=IF(ISBLANK($D{r}),"",
ROW()-9)`, not Home's bare `ROW()-9`), column F carries the legacy TC ID scheme
`newR1L-Radio-nnn` for which NEW ROWS HAVE NO RULED SCHEME (left blank and
reported, never invented), and columns G/H ARE filled on new rows per R7-Q1/Q2,
unlike Home where the profile bans them.

Usage:
    python features/amfm/scripts/write_back.py --feature-dir features/amfm
    python features/amfm/scripts/write_back.py --feature-dir features/amfm --init-baseline
    python features/amfm/scripts/write_back.py --feature-dir features/amfm --write

Exit code 0 = ok, 1 = an invariant failed, 2 = bad invocation.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import zipfile
from datetime import date
from pathlib import Path

import openpyxl

# Repo root = the nearest ancestor carrying pyproject.toml. Resolved by
# marker rather than by parent count: the feature directory moved from the
# repo root to features/ on 2026-08-11, and a hard-coded depth broke.
REPO_ROOT = next(p for p in Path(__file__).resolve().parents
                 if (p / "pyproject.toml").is_file())
sys.path.insert(0, str(REPO_ROOT / "scripts"))
from feature_config import load_feature_config, resolve_path  # noqa: E402

HISTORY_SHEET = "ChangeHistory 修訂履歷"
PLACEHOLDER_BODY = "BLOCKED - see Remarks"
# Verified against all 158 legacy rows: column R is "NA" everywhere.
CONST_CELLS = {"functional_safety": "NA"}
# Per-row formula: the legacy rows reference their own row in $D, so this
# cannot be a constant string the way Home's `=ROW()-9` is.
ROW_NUMBER_FORMULA = '=IF(ISBLANK($D{row}),"",ROW()-9)'
# Blank on all 158 legacy rows — listed so a reviewer sees the omission is
# deliberate rather than missed.
INTENTIONALLY_BLANK = ("C (Polarion ID)", "E (TestRail ID)",
                       "S..Y (vehicle flags)", "AA..AF (execution columns)")


class WriteBackError(RuntimeError):
    pass


# --------------------------------------------------------------------- read

def load_rows(ws, cfg) -> list[dict]:
    col = cfg["col"]
    first = cfg["workbook"]["header_row"] + 1
    out = []
    for i in range(first, ws.max_row + 1):
        vals = [ws.cell(i, c + 1).value for c in range(col["req_id"],
                                                       col["remarks"] + 1)]
        if not any(v is not None for v in vals):
            continue
        out.append({
            "row": i,
            "req_id": str(ws.cell(i, col["req_id"] + 1).value or ""),
            "author": str(ws.cell(i, col["author"] + 1).value or "").strip(),
            "cells": ["" if v is None else str(v) for v in vals],
        })
    return out


def row_digest(cells: list[str]) -> str:
    return hashlib.sha256("\x1f".join(cells).encode("utf-8")).hexdigest()[:16]


def ordered_content_hash(rows: list[dict], author: str) -> str:
    """Ordered hash of the legacy rows' contents, selected by author value.

    Selecting by the legacy author (not by "author is non-empty") keeps the
    same rows in scope after the write, when the appended rows carry PeiPYHsu.
    """
    legacy = [r for r in rows if r["author"] == author]
    return hashlib.sha256(
        "\n".join(row_digest(r["cells"]) for r in legacy).encode()).hexdigest()


def segments_of(rows: list[dict], author: str) -> list[dict]:
    """Contiguous runs of LEGACY / REGEN rows, in sheet order."""
    segs: list[dict] = []
    for r in rows:
        kind = "LEGACY" if r["author"] == author else "REGEN"
        if segs and segs[-1]["kind"] == kind and segs[-1]["end"] == r["row"] - 1:
            segs[-1]["end"] = r["row"]
            segs[-1]["ids"].append(r["req_id"])
        else:
            segs.append({"kind": kind, "start": r["row"], "end": r["row"],
                         "ids": [r["req_id"]]})
    return segs


# --------------------------------------------------------------- Scope field

def find_scope_cell(ws, cfg) -> tuple[int, int]:
    """Locate the header 範圍 Scope VALUE cell — (row, column), 1-based.

    Found by label text, not by coordinate: the same template puts the block at
    different offsets per feature. Only rows above the data header are searched,
    so column N's header (which also contains "Reference") cannot be hit.
    """
    label = cfg["write_back"]["scope_label"]
    needle = label.split()[-1].casefold()
    hits = []
    for r in range(1, cfg["workbook"]["header_row"]):
        for c in range(1, 21):
            v = ws.cell(r, c).value
            if v is None:
                continue
            text = str(v)
            if needle in text.casefold() or label.split()[0] in text:
                hits.append((r, c))
    if len(hits) != 1:
        raise WriteBackError(
            f"expected exactly one {label!r} label above row "
            f"{cfg['workbook']['header_row']}, found {len(hits)}: {hits}")
    r, c = hits[0]
    return r, c + 1


def _merge_anchor(ws, row: int, col: int) -> tuple[int, int]:
    """openpyxl refuses a write to a merged range's non-anchor member."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col
    return row, col


def fix_scope(ws, cfg, expected: str) -> dict:
    """A-AM01 / RULINGS C1 — the Scope field names the superseded report."""
    row, col = find_scope_cell(ws, cfg)
    row, col = _merge_anchor(ws, row, col)
    before = ws.cell(row, col).value
    before = "" if before is None else str(before)
    ws.cell(row, col).value = expected
    return {"cell": ws.cell(row, col).coordinate, "before": before,
            "after": expected, "changed": before != expected}


# ------------------------------------------------------------------ planning

def collect_generated(generated: Path, leaf_order: list[str]) -> list[dict]:
    """generated/*.json -> workbook rows, in 037 document order."""
    by_req: dict[str, list[dict]] = {}
    for path in sorted(generated.glob("*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        for tc in payload.get("tcs", []):
            by_req.setdefault(str(tc["req_id"]), []).append(tc)
    unknown = sorted(set(by_req) - set(leaf_order))
    if unknown:
        raise WriteBackError(
            f"generated TCs reference req_ids that are not 037 leaves: {unknown}")
    rows = []
    for req_id in leaf_order:
        rows.extend(by_req.get(req_id, []))
    return rows


# ------------------------------------------------------------------- writing

def cell_values(tc: dict, cfg, tc_id: str | None = None) -> dict[int, object]:
    col = cfg["col"]
    wb_cfg = cfg["write_back"]
    fill_sets = wb_cfg.get("fill_test_group_set", False)
    is_placeholder = tc.get("placeholder") is True
    values = {
        col["req_id"]: tc["req_id"],
        # R7-Q1 / R7-Q2 — filled here, unlike Home where the profile bans them.
        col["test_group"]: (tc.get("test_group") or None) if fill_sets else None,
        col["test_set"]: (tc.get("test_set") or None) if fill_sets else None,
        col["test_item"]: tc["test_item"] or None,
        col["pre_conditions"]: tc["pre_conditions"] or None,
        col["input_test_data"]: tc["input_test_data"] or None,
        col["test_procedure"]: tc["test_procedure"] or None,
        col["expected_result"]: tc["expected_result"] or None,
        col["spec_reference"]: tc["specification_reference"],
        col["tc_ref_id"]: wb_cfg["tc_ref_id_value"],
        col["priority"]: tc["priority"] or None,
        col["design_method"]: tc["design_method"] or None,
        col["functional_safety"]: CONST_CELLS["functional_safety"],
        col["author"]: wb_cfg["author_value"],
        col["remarks"]: tc.get("remarks") or None,
    }
    # R13 — the author-side TC ID (canon §10.3). Only written when the feature
    # declares both the column and the series, so a feature without a ruled
    # scheme leaves the cell alone rather than inventing one.
    if tc_id is not None and "tc_id" in col:
        values[col["tc_id"]] = tc_id
    if fill_sets and not (values[col["test_group"]] and values[col["test_set"]]):
        raise WriteBackError(
            f"{tc['req_id']}: fill_test_group_set is on but Test Group / Test "
            "Set is empty (R7-Q1/Q2)")
    if is_placeholder:
        for key in ("priority", "design_method"):
            if values[col[key]] is not None:
                raise WriteBackError(
                    f"{tc['req_id']}: placeholder rows must leave {key} blank")
        for key in ("test_procedure", "expected_result"):
            if values[col[key]] != PLACEHOLDER_BODY:
                raise WriteBackError(
                    f"{tc['req_id']}: placeholder {key} must be "
                    f"{PLACEHOLDER_BODY!r}")
    return values


def assign_tc_ids(cfg, tcs: list[dict], taken: set[str]) -> list[str | None]:
    """`newR1L-AMFM-001` upward, in 037 document order (R13, canon §10.3).

    Monotonic within its own {project}-{abbr} series, which is what §10.3 asks
    for — the frozen `newR1L-Radio-*` series is a different group and is not
    continued, so deleting the legacy region later cannot leave a hole in this
    one. A collision with an id already in the sheet aborts rather than
    producing two rows with the same TC ID.
    """
    fmt = cfg["write_back"].get("tc_id_format")
    if not fmt or "tc_id" not in cfg["col"]:
        return [None] * len(tcs)
    ids = [fmt.format(n=i) for i in range(1, len(tcs) + 1)]
    clash = sorted(set(ids) & taken)
    if clash:
        raise WriteBackError(
            f"generated TC IDs collide with ids already in the workbook: "
            f"{clash[:5]}")
    return ids


def append_rows(ws, cfg, first_row: int, tcs: list[dict],
                tc_ids: list[str | None] | None = None) -> dict:
    """Write the generated rows starting at `first_row`, inserting as needed.

    The template already carries the column B formula on a tail of empty rows;
    those are consumed first and the remainder is inserted, so the sheet grows
    by exactly the number of rows the tail could not absorb.
    """
    spare = max(ws.max_row - first_row + 1, 0)
    inserted = max(len(tcs) - spare, 0)
    if inserted:
        ws.insert_rows(ws.max_row + 1, inserted)
    ids = tc_ids or [None] * len(tcs)
    for offset, (tc, tc_id) in enumerate(zip(tcs, ids)):
        r = first_row + offset
        for column, value in cell_values(tc, cfg, tc_id).items():
            ws.cell(r, column + 1).value = value
    return {"first_row": first_row, "rows": len(tcs),
            "template_rows_used": min(spare, len(tcs)), "inserted": inserted,
            "tc_id_first": ids[0], "tc_id_last": ids[-1] if ids else None}


def reemit_row_numbers(ws, cfg) -> int:
    """Column B, the only formula column, references its own row in $D."""
    first = cfg["workbook"]["header_row"] + 1
    req_col = cfg["col"]["req_id"] + 1
    n = 0
    for r in range(first, ws.max_row + 1):
        if ws.cell(r, req_col).value in (None, ""):
            continue
        ws.cell(r, 2).value = ROW_NUMBER_FORMULA.format(row=r)
        n += 1
    return n


def append_history(wb, cfg, when: str, plan: dict, n_leaves: int,
                   scope: dict | None = None) -> str:
    ws = wb[HISTORY_SHEET]
    row = 5
    last = None
    while ws.cell(row, 1).value not in (None, ""):
        last = str(ws.cell(row, 1).value).strip()
        row += 1
    revision = chr(ord(last) + 1) if last and len(last) == 1 else "C"
    note = (
        f"Added {plan['rows']} test cases covering the {n_leaves} leaves of "
        f"FM-WI-FSM-037-A03, appended from row {plan['first_row']}.\n"
        "The 158 existing rows are unchanged — verified by an ordered content "
        "hash over columns D..AG, not by row position."
    )
    if scope and scope["changed"]:
        note += (f"\nCorrected the header 範圍 Scope field ({scope['cell']}), "
                 f"which named the superseded requirement report: "
                 f"{scope['before']} -> {scope['after']}.")
    ws.cell(row, 1).value = revision
    ws.cell(row, 2).value = note
    ws.cell(row, 3).value = cfg["write_back"]["author_value"]
    ws.cell(row, 4).value = when
    return revision


# ------------------------------------------------------------- reproducible

_EPOCH = (1980, 1, 1, 0, 0, 0)
_DCTERMS_RE = re.compile(rb"(<dcterms:(created|modified)[^>]*>)[^<]*(</dcterms:\2>)")
_FIXED_STAMP = b"2000-01-01T00:00:00Z"


def normalize_for_reproducibility(path: Path) -> None:
    """Zero the zip timestamps and document dates so the SHA256 is stable."""
    src = zipfile.ZipFile(path)
    items = sorted(src.infolist(), key=lambda i: i.filename)
    tmp = path.with_suffix(".normalizing")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for info in items:
            data = src.read(info.filename)
            if info.filename == "docProps/core.xml":
                # \g<1> not \1: `\1` followed by the stamp's leading "2" would
                # parse as group 12 and silently corrupt core.xml.
                data = _DCTERMS_RE.sub(rb"\g<1>" + _FIXED_STAMP + rb"\g<3>", data)
            new = zipfile.ZipInfo(info.filename, date_time=_EPOCH)
            new.compress_type = zipfile.ZIP_DEFLATED
            new.external_attr = info.external_attr
            dst.writestr(new, data)
    src.close()
    tmp.replace(path)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ------------------------------------------------------------------- driver

def run(args) -> int:
    cfg = load_feature_config(args.feature_dir)
    root = Path(args.feature_dir)
    data = root / args.data
    src = resolve_path(cfg, "workbook", args.workbook)
    legacy_author = cfg["done_region"]["author_value"]

    stla = json.loads((data / "stla_to_cfts.json").read_text(encoding="utf-8"))
    leaf_order = list(stla)

    wb = openpyxl.load_workbook(src)
    ws = wb[cfg["workbook"]["sheet"]]

    before_rows = load_rows(ws, cfg)
    before_hash = ordered_content_hash(before_rows, legacy_author)
    legacy_before = [r for r in before_rows if r["author"] == legacy_author]
    if len(legacy_before) != len(before_rows):
        raise WriteBackError(
            f"{len(before_rows) - len(legacy_before)} rows in the source are "
            f"not authored by {legacy_author!r}; this workbook is expected to "
            "hold the frozen legacy region only (A-AM02)")

    baseline_path = data / "legacy_baseline.json"
    if args.init_baseline:
        baseline_path.write_text(json.dumps(
            {"ordered_content_hash": before_hash, "rows": len(legacy_before),
             "first_row": legacy_before[0]["row"],
             "last_row": legacy_before[-1]["row"], "source": src.name},
            ensure_ascii=False, indent=1), encoding="utf-8")
        print(f"baseline written: {baseline_path} "
              f"({len(legacy_before)} rows, hash {before_hash[:16]}…)")
        return 0
    if baseline_path.exists():
        baseline = json.loads(baseline_path.read_text(encoding="utf-8"))
        if baseline["ordered_content_hash"] != before_hash:
            raise WriteBackError(
                "the source workbook no longer matches "
                f"data/legacy_baseline.json ({before_hash[:16]} != "
                f"{baseline['ordered_content_hash'][:16]}); the legacy region "
                "changed upstream — re-run with --init-baseline only after "
                "that change is understood and ruled")
    else:
        print(f"WARNING: no {baseline_path.name}; run --init-baseline against "
              "the pristine workbook so a later upstream edit is detectable",
              file=sys.stderr)

    scope_expected = resolve_path(cfg, cfg["write_back"]["scope_source"]).stem
    scope = fix_scope(ws, cfg, scope_expected)

    tcs = collect_generated(root / args.generated, leaf_order)
    covered = {tc["req_id"] for tc in tcs}
    missing = [r for r in leaf_order if r not in covered]
    if missing:
        raise WriteBackError(
            f"{len(missing)} leaves have no generated row: {missing[:10]}")

    first_row = legacy_before[-1]["row"] + 1
    taken = set()
    if "tc_id" in cfg["col"]:
        taken = {str(ws.cell(r["row"], cfg["col"]["tc_id"] + 1).value)
                 for r in before_rows}
    tc_ids = assign_tc_ids(cfg, tcs, taken)
    plan = append_rows(ws, cfg, first_row, tcs, tc_ids)
    numbered = reemit_row_numbers(ws, cfg)

    # ---- invariants
    after_rows = load_rows(ws, cfg)
    after_hash = ordered_content_hash(after_rows, legacy_author)
    legacy_after = [r for r in after_rows if r["author"] == legacy_author]
    if after_hash != before_hash:
        raise WriteBackError("LEGACY REGION CHANGED: ordered content hash "
                             f"{before_hash[:16]} -> {after_hash[:16]}")
    if len(legacy_after) != len(legacy_before):
        raise WriteBackError(f"legacy row count changed: {len(legacy_before)} "
                             f"-> {len(legacy_after)}")
    moved = [(b["row"], a["row"]) for b, a in zip(legacy_before, legacy_after)
             if b["row"] != a["row"]]
    if moved:
        raise WriteBackError(f"legacy rows moved: {moved[:5]}")
    after_segs = segments_of(after_rows, legacy_author)
    if [s["kind"] for s in after_segs] != ["LEGACY", "REGEN"]:
        raise WriteBackError(
            "segment order invariant broken: expected LEGACY then REGEN, got "
            f"{[s['kind'] for s in after_segs]}")
    written = {r["req_id"] for r in after_rows if r["author"] != legacy_author}
    if written != set(leaf_order):
        raise WriteBackError(
            "regen req_ids != the 037 leaf set (missing "
            f"{sorted(set(leaf_order) - written)[:5]}, extra "
            f"{sorted(written - set(leaf_order))[:5]})")

    # ---- dry-run summary: every field the canon §6 checklist requires
    placeholders = [tc["req_id"] for tc in tcs if tc.get("placeholder")]
    total_before = before_rows[-1]["row"]
    total_after = after_rows[-1]["row"]
    print(f"source        : {src.name}")
    print(f"segments      : " + "  ".join(
        f"{s['kind']} {s['start']}-{s['end']} ({s['end'] - s['start'] + 1} rows)"
        for s in after_segs))
    print(f"  LEGACY       : {len(legacy_before)} -> {len(legacy_after)} rows "
          f"(+0), rows {legacy_after[0]['row']}-{legacy_after[-1]['row']} "
          "unmoved")
    print(f"  REGEN        : 0 -> {plan['rows']} rows (+{plan['rows']}), "
          f"appended from row {plan['first_row']}")
    print(f"  arithmetic   : {len(legacy_before)} + {plan['rows']} = "
          f"{len(legacy_after) + plan['rows']} data rows; last sheet row "
          f"{total_before} -> {total_after} "
          f"({plan['template_rows_used']} template rows reused, "
          f"{plan['inserted']} inserted)")
    print(f"legacy region : hash {after_hash[:16]}… unchanged, order unchanged")
    print(f"coverage      : {len(written)} regen req_ids == "
          f"{len(leaf_order)} leaf set, exact match")
    print(f"regen rows    : {len(tcs)} TCs across {len(covered)} leaves")
    print(f"placeholders  : {placeholders if placeholders else 'none'}")
    print(f"row numbers   : {numbered} rows re-emitted with the column B formula")
    print(f"TC IDs (F)    : {plan['tc_id_first']} … {plan['tc_id_last']} "
          f"(R13; the frozen newR1L-Radio-001…158 series is not continued)")
    print(f"TC ref ID (O) : {cfg['write_back']['tc_ref_id_value']!r} "
          "(R13; matches the legacy value, no dropdown defines this column)")
    print(f"blank by convention: {', '.join(INTENTIONALLY_BLANK)}")
    print("note          : the priority dropdown covers P10:P167 only, so the "
          "appended rows carry no data validation; values are lint-checked "
          "against P0..P3 — not extended here, flag if the form should match")
    if scope["changed"]:
        print(f"Scope {scope['cell']}    : {scope['before']!r}\n"
              f"             -> {scope['after']!r}  (A-AM01)")
    else:
        print(f"Scope {scope['cell']}    : already correct")

    if not args.write:
        print("\nDRY RUN — nothing written. Re-run with --write to produce the "
              "workbook.")
        return 0

    revision = append_history(wb, cfg, args.date or date.today().isoformat(),
                              plan, len(leaf_order), scope)
    out = Path(args.out) if args.out else Path("output") / src.name
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    normalize_for_reproducibility(out)
    digest = sha256_file(out)
    print(f"\nwrote         : {out}")
    print(f"ChangeHistory revision {revision} appended")
    print(f"SHA256        : {digest}")
    (out.parent / (out.stem + ".sha256")).write_text(
        f"{digest}  {out.name}\n", encoding="utf-8")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--write", action="store_true",
                    help="actually produce the workbook (default is a dry run)")
    ap.add_argument("--init-baseline", action="store_true",
                    help="record the legacy region hash and exit")
    ap.add_argument("--feature-dir", default=".")
    ap.add_argument("--data", default="data")
    ap.add_argument("--generated", default="generated")
    ap.add_argument("--workbook", help="override feature.yaml paths.workbook")
    ap.add_argument("--out", help="output path (default output/<source name>)")
    ap.add_argument("--date", help="ChangeHistory date (default today)")
    args = ap.parse_args()
    try:
        return run(args)
    except WriteBackError as exc:
        print(f"ABORTED: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
