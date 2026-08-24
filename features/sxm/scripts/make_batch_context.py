#!/usr/bin/env python3
"""Step 2 (SXM) — assemble the generation context for one batch.

Ported from the AMFM copy. Same contract, three SXM-specific facts:

1. **Clause text comes from the HYBRID map**, so `spec_paragraph` is the
   ReqIF text and `section` is the docx printed number — the builder does not
   re-read either source, it reads `data/stla_to_cfts.json`.
2. **The leaf prefix is derived, not hard-coded.** AMFM's `SWE-RA-RAD-` would
   silently expand every batch row into ids that match no leaf.
3. **Markers travel with the leaf.** A-SX03 `(add)` leaves, A-SX07's leaf 154
   and the A-SX04 twin rows each carry their instruction in the context, for
   the same reason the R11 cite-form instruction does: the generator cannot
   infer any of them from the clause text alone.

Usage:
    python features/sxm/scripts/make_batch_context.py --feature-dir features/sxm \
        --batches-md docs/batches-sxm.md --batch "B1 (pilot) — Instant Replay"
"""
from __future__ import annotations

import argparse
import difflib
import json
import re
import sys
from pathlib import Path

import docx
import openpyxl

# Repo root = the nearest ancestor carrying pyproject.toml. Resolved by
# marker rather than by parent count: the feature directory moved from the
# repo root to features/ on 2026-08-11, and a hard-coded depth broke.
REPO_ROOT = next(p for p in Path(__file__).resolve().parents
                 if (p / "pyproject.toml").is_file())
sys.path.insert(0, str(REPO_ROOT / "scripts"))
from feature_config import load_feature_config, resolve_path  # noqa: E402

BATCH_ROW = re.compile(r"^\|\s*([^|]+?)\s*\|\s*([^|]*?)\s*\|\s*(\d+)\s*\|"
                       r"\s*([^|]+?)\s*\|\s*([^|]*?)\s*\|")
ANOMALY = re.compile(r"A-AM\d\d")
# A batch cites a spec worksheet as `[[table:NAME]]` in its context note, so
# membership and its evidence stay in the one batch table.
TABLE_CITE = re.compile(r"\[\[table:([A-Za-z0-9_.-]+)\]\]")
HEADING_NUM = re.compile(r"^(\d+(?:\.\d+)*)\s")
REQ_PREFIX = "SWE-RA-SXM-"
# Above this the 037 title is a verbatim quote of the CFTS clause (the id tail
# and whitespace account for the gap); below it the two texts genuinely differ.
WORDING_VERBATIM = 0.95


def _fold(s: str) -> str:
    """Compare wording, not typography or the trailing STLA id tag."""
    s = re.sub(r"[({（]\s*\d{7}\s*[)}）]", " ", str(s or ""))
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9$ ]", " ", s.lower())).strip()


class ContextError(RuntimeError):
    pass


def parse_leaf_selector(spec: str) -> list[str]:
    out = []
    for part in str(spec).split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = (int(x) for x in part.split("-", 1))
            out.extend(f"{REQ_PREFIX}{n:03d}" for n in range(a, b + 1))
        else:
            out.append(f"{REQ_PREFIX}{int(part):03d}")
    return out


def load_batches(md: Path) -> dict:
    if not md.exists():
        raise ContextError(f"missing {md}")
    batches = {}
    for line in md.read_text(encoding="utf-8").splitlines():
        m = BATCH_ROW.match(line)
        if not m or m.group(3) == "n":
            continue
        name, secs, count, ids, note = m.groups()
        req_ids = parse_leaf_selector(ids)
        if len(req_ids) != int(count):
            raise ContextError(
                f"{name}: table says {count} leaves but lists {len(req_ids)}")
        # The table's first cell is `B1 (pilot) — Instant Replay`: a batch
        # label AND its Test Set. Column H takes the Test Set only — writing
        # the batch label there would ship "B1 (pilot) — …" to the customer.
        # Batches ≠ Test Sets (framework Part IV): B1/B2 are both Instant
        # Replay, B8/B9 are both Browse.
        test_set = name.split("—")[-1].strip() if "—" in name else name
        batches[name] = {
            "batch_label": name,
            "test_set": test_set, "req_ids": req_ids,
            "declared_sections": [s.strip() for s in re.split(r"[,;]", secs)
                                  if s.strip()],
            "context_note": note.strip(),
            "anomalies": sorted(set(ANOMALY.findall(note))),
            "spec_tables": sorted(set(TABLE_CITE.findall(note))),
        }
    if not batches:
        raise ContextError(f"no batch rows parsed from {md}")
    return batches


def section_texts(docx_path: Path) -> dict[str, dict]:
    """{section number: {title, level, text, children, subtree_chars}}.

    `text` is the section's OWN body — the paragraphs between its heading and
    the next heading of any level. That is what a leaf bracketing into this
    section is governed by: anything under a subsection belongs to a leaf that
    brackets into the subsection instead.

    The distinction is not cosmetic. §1.3 "HU Analog Tuner" runs to §1.4, so a
    same-or-higher-level span hands 82k characters — the entire analog tuner
    chapter — to the two leaves whose requirement is the section's own
    two-paragraph AM-presence gate. `children` and `subtree_chars` are kept so
    a batch that genuinely needs the subtree can ask for it knowingly.
    """
    doc = docx.Document(docx_path)
    paras = doc.paragraphs
    heads = []
    for i, p in enumerate(paras):
        if not p.style.name.startswith("Heading"):
            continue
        m = HEADING_NUM.match(p.text.strip())
        if not m:
            continue
        try:
            level = int(p.style.name.split()[-1])
        except ValueError:
            level = m.group(1).count(".") + 1
        heads.append((i, m.group(1), p.text.strip(), level))

    out = {}
    for k, (i, num, title, level) in enumerate(heads):
        own_end = heads[k + 1][0] if k + 1 < len(heads) else len(paras)
        subtree_end = len(paras)
        children = []
        for j, nnum, _, nlevel in heads[k + 1:]:
            if nlevel <= level:
                subtree_end = j
                break
            if nlevel == level + 1:
                children.append(nnum)
        body = [p.text.strip() for p in paras[i + 1:own_end] if p.text.strip()]
        subtree = [p.text.strip() for p in paras[i + 1:subtree_end] if p.text.strip()]
        out[num] = {"title": title, "level": level, "text": "\n".join(body),
                    "children": children,
                    "subtree_chars": len("\n".join(subtree))}
    return out


def load_spec_tables(cfg, root: Path, wanted: list[str]) -> dict:
    """Inject the spec worksheets a batch cites, as rows rather than prose.

    Mode D's spec is not only the CFTS docx: several clauses delegate their
    detail to a companion workbook ("See 'CIP_Radio_Tables*', 'SEEK
    Cancel_Stop Transitions' worksheet"). Those tables ARE the requirement for
    a State Transition TC — a seek batch generated without the cancel/stop
    matrix has no source for which event cancels and which stops, and would
    have to guess.

    Declared per batch in feature.yaml `spec_tables`; a batch that names a
    sheet which is not in the file fails loud rather than generating without
    it. Merged header rows are forward-filled, because the event columns are
    grouped under a banner row and a lone column label loses its group.
    """
    out = {}
    for name in wanted:
        spec = (cfg.get("spec_tables") or {}).get(name)
        if spec is None:
            raise ContextError(
                f"batch cites spec table {name!r} but feature.yaml "
                "spec_tables has no such entry")
        path = resolve_path(cfg, spec["source"])
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = spec["sheet"]
        if sheet not in wb.sheetnames:
            wb.close()
            raise ContextError(
                f"{path.name} has no worksheet {sheet!r}; present: "
                f"{wb.sheetnames}")
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        wb.close()
        # Some worksheets stack several tables (the market configuration sheet
        # opens with a per-market feature grid, then the tuner configuration
        # table below it). `first_row` points at the wanted table's own header
        # instead of forcing a second sheet or a hand-copied extract.
        first = spec.get("first_row", 0)
        if first >= len(rows):
            raise ContextError(
                f"{path.name}/{sheet}: first_row {first} is past the last row "
                f"({len(rows)})")
        rows = rows[first:]
        hdr_rows = spec.get("header_rows", 1)
        banner, header = [], []
        for i in range(hdr_rows):
            cells = [str(v).strip().replace("\n", " ") if v else ""
                     for v in rows[i]]
            if i < hdr_rows - 1:
                run = ""
                filled = []
                for c in cells:
                    run = c or run
                    filled.append(run)
                banner = filled
            else:
                header = cells
        records = []
        for r in rows[hdr_rows:]:
            label = str(r[0]).strip() if r[0] else ""
            if not label:
                continue
            cells = {}
            for j, v in enumerate(r):
                if j == 0 or v in (None, ""):
                    continue
                key = header[j] if j < len(header) else f"col{j}"
                if banner and j < len(banner) and banner[j]:
                    key = f"{banner[j]} / {key}"
                cells[key] = str(v).replace("\n", " ").strip()
            records.append({"state": label, "events": cells})
        out[name] = {"source": path.name, "sheet": sheet,
                     "row_label": header[0] if header else "",
                     "rows": records}
    return out


def leaf_citations(citations: dict, req_id: str) -> list[dict]:
    """Cross-document citations made by this leaf's own clause.

    A leaf whose clause says `HU shall play the rejection tone {See CFTS019-718}`
    borrows an outcome it does not define. R11 rules this CITE-FORM, not
    absorption: the outcome is asserted where the citing clause puts it, but
    anchored to the token, and the cited document's own rule surface stays out
    of scope. The instruction travels with the leaf because the two failure
    modes are opposite — silently dropping the outcome under-covers the citing
    clause, and silently adopting it claims verification of another delivery's
    requirement.
    """
    out = []
    for token, e in sorted(citations.items()):
        if req_id not in e.get("req_ids", []):
            continue
        item = {"token": token, "doc": e["doc"], "status": e["status"],
                "cited_in": e["citing_clauses"][0]["context"],
                "handling": "cite-form (R11) — reference, do not absorb",
                "instruction": (
                    f"Cite {token} verbatim as a second citation in "
                    "specification_reference, after this leaf's own clause. "
                    "The ER may assert the borrowed outcome, anchored to the "
                    f"citation — the observable outcome, then `as defined by "
                    f"{token}`. Do NOT test that document's own "
                    f"rule surface (which conditions qualify, the specification "
                    f"of the behaviour itself): that is {e['doc']}'s delivery. "
                    "Verify only that the outcome occurs in the citing "
                    "clause's scenario."),
                }
        if e.get("resolved_text"):
            item |= {"resolved_clause": e["resolved_clause"],
                     "resolved_section": e.get("resolved_section"),
                     "referenced_text": e["resolved_text"],
                     "ruling": e.get("ruling")}
        else:
            item["candidates"] = e.get("candidates", [])
        out.append(item)
    return out


# A-SX04 twin rows: SXM leaf -> the analog-chapter clause id its Remarks cites.
# Clause ids, never AMFM TC ids (profile §3.6): TC ids shift on any AMFM regen.
TWIN_ROWS = {
    "020": 4872413, "024": 4872442, "037": 4872475, "108": 4872429,
    "110": 4872430, "132": 4872493, "140": 4872499, "142": 4872500,
    "143": 4872501, "148": 4872506, "149": 4872508,
}
ADD_LEAVES = {"080", "083", "110", "148", "149", "154", "155", "156", "157",
              "158", "182"}


def test_set_index(cfg) -> dict[str, str]:
    """req_id -> framework Part IV Test Set, expanded from feature.yaml.

    Column H belongs to the leaf, not to the batch that happens to generate
    it. B1 carries leaf 154, whose Set is Traffic & Weather; B5 and B13 each
    span two Sets outright. A batch-derived value is right only for batches
    that happen to hold one Set.
    """
    out = {}
    for name, spec in (cfg.get("test_sets") or {}).items():
        for rid in parse_leaf_selector(spec):
            if rid in out and out[rid] != name:
                raise ContextError(
                    f"{rid} is claimed by two Test Sets: {out[rid]!r} and "
                    f"{name!r} — framework Part IV must partition the leaves")
            out[rid] = name
    return out


def leaf_markers(req_id: str, entry: dict) -> list[dict]:
    """The per-leaf instructions no clause text can convey (profile §5).

    Each marker is an upstream fact about the LEAF, not about the requirement:
    how it entered the 037, whether its title and its id agree, and whether the
    same text exists under another id in the analog chapter. A generator given
    only the clause would produce a plausible TC and miss all three.
    """
    tail = req_id.rsplit("-", 1)[-1]
    out = []
    if tail in ADD_LEAVES:
        out.append({
            "marker": "[A-SX03]",
            "fact": "This leaf entered the 037 outside the release process: "
                    "its title carries a trailing `(add)` after the id, and "
                    "Release Version and Requirement Status are both empty.",
            "instruction": "Generate normally. Put `[A-SX03]` in reasoning so "
                           "a later upstream withdrawal is a grep, not an "
                           "audit. Nothing about the TC content changes.",
        })
    if tail == "154":
        out.append({
            "marker": "[A-SX07]",
            "fact": "The 037 title reproduces clause 4872961 (leaf 153's "
                    "clause, the 'Jump' button entry path) while the declared "
                    "id is 4872962 (the Browse entry path). Ruled reading: the "
                    "id is right and the title was copy-pasted.",
            "instruction": "Content follows clause 4872962 — the Browse entry "
                           "path — not the 037 title. Put `[A-SX07]` in "
                           "reasoning alongside `[A-SX03]`, and note the "
                           "title/clause divergence per §8.6.",
        })
    if tail in TWIN_ROWS:
        out.append({
            "marker": "A-SX04 twin",
            "fact": f"This clause is a >=0.95 text twin of CFTS024-"
                    f"{TWIN_ROWS[tail]} in the analog chapter, which the AM/FM "
                    "deliverable already covers under its own leaf.",
            "instruction": "Generate normally against THIS clause in SAT "
                           "context — nothing is cross-cited. Remarks carries "
                           f"exactly: `Analog-chapter twin: CFTS024-"
                           f"{TWIN_ROWS[tail]} (covered in the AM/FM "
                           "deliverable)` — clause id only, no AMFM TC id, no "
                           "internal anomaly id (Remarks is external-facing).",
        })
    return out


def build(cfg, batch: dict, stla_map: dict, sections: dict,
          exemplars: dict, spec_docs: dict, spec_tables: dict | None = None,
          citations: dict | None = None, unalloc_index: dict | None = None,
          set_index: dict | None = None) -> dict:
    primary = spec_docs.get("primary")
    leaves, needed, blocked = [], set(), []
    for rid in batch["req_ids"]:
        e = stla_map.get(rid)
        if e is None:
            raise ContextError(f"{rid} is in the batch table but not the "
                               "STLA map — rerun build_stla_map.py")
        entry = {"req_id": rid, "stla_id": e["stla_id"], "doc": e["doc"],
                 "section": e["section"], "section_title": e["section_title"],
                 "requirement_text": e["title"],
                 "analysis_note": e["description"],
                 "source_components": e["source_components"],
                 "spec_paragraph": e.get("spec_paragraph"),
                 "spec_paragraph_metadata": e.get("spec_paragraph_metadata"),
                 "spec_reference": cfg["spec_reference_template"].format(
                     doc=e["doc"], stla_id=e["stla_id"])}
        # The 037 title is supposed to quote the CFTS clause verbatim. Where it
        # does not, §8.6 says the source spec wins — so the divergence has to
        # reach the generator rather than be averaged away by having both texts
        # side by side with no comment.
        if e.get("spec_paragraph"):
            ratio = difflib.SequenceMatcher(
                None, _fold(e["title"]), _fold(e["spec_paragraph"])).ratio()
            entry["wording_agreement"] = round(ratio, 3)
            if ratio < WORDING_VERBATIM:
                entry["wording_note"] = (
                    "037 title and CFTS clause differ; §8.6 — the source spec "
                    "is authority for wording, the 037 title for scope")
        refs = leaf_citations(citations or {}, rid)
        if refs:
            entry["cross_references"] = refs
        entry["test_set"] = (set_index or {}).get(rid)
        if set_index and not entry["test_set"]:
            raise ContextError(
                f"{rid} belongs to no Test Set in feature.yaml test_sets — "
                "framework Part IV must cover every leaf")
        marks = leaf_markers(rid, e)
        if marks:
            entry["markers"] = marks
        if e["section"]:
            needed.add((e["doc"], e["section"]))
        else:
            reason = ("clause not found in the owning document"
                      if e["doc"] else "no document allocated")
            entry["section_text_absent"] = reason
            blocked.append(rid)
        leaves.append(entry)

    # Siblings: every other leaf bracketing into a section this batch uses.
    # They are the split axis made visible, not extra targets. Keyed by
    # (doc, section): CFTS011 §1.5.5.1 and CFTS024 §1.5.5.1 are different
    # sections of different documents, and matching on the number alone would
    # hand a batch the siblings of a document it is not testing.
    in_batch = set(batch["req_ids"])
    siblings = [{"req_id": rid, "doc": e["doc"], "section": e["section"],
                 "requirement_text": e["title"],
                 "not_a_target": "reference only — this leaf belongs to "
                                 "another batch; do NOT generate a TC for it"}
                for rid, e in stla_map.items()
                if (e["doc"], e["section"]) in needed and rid not in in_batch]

    spec: dict[str, dict] = {}
    missing_sections = []
    for doc, num in sorted(needed):
        available = sections.get(doc) or {}
        if num in available:
            spec.setdefault(doc, {})[num] = available[num]
        else:
            missing_sections.append(f"{doc} §{num}")
    if missing_sections:
        raise ContextError(
            f"sections {missing_sections} are cited by the bracket map but "
            "absent from their document — the map and the spec files have "
            "diverged")

    # R10-2 decision material, carried with the batch rather than left in a
    # side file: the test runs per clause AT GENERATION, and a generator that
    # has to open another artefact to see what is unallocated will not run it.
    unallocated = {}
    for doc, secs in sorted(spec.items()):
        for num in secs:
            entry_ = (unalloc_index or {}).get(num)
            if entry_ and entry_.get("unallocated"):
                unallocated[num] = {
                    "section_title": entry_["section_title"],
                    "claimed": entry_["claimed"],
                    "unallocated": entry_["unallocated"],
                }

    return {
        "feature": cfg["feature"],
        "batch": batch.get("batch_label", batch["test_set"]),
        "test_set": batch["test_set"],
        "spec_mode": cfg["spec_mode"],
        "leaves": leaves,
        "spec_sections": spec,
        "siblings": siblings,
        "siblings_note": (
            "Siblings are the OTHER leaves that share a section with this "
            "batch — they exist to make the split axis visible (§8.3) and to "
            "show what a neighbouring leaf already owns, so this batch does "
            "not absorb it (§8.2.1). They are NOT generation targets: this "
            "batch produces test cases for the `leaves` list only."),
        "spec_tables": spec_tables or {},
        "unallocated_clauses": unallocated,
        "absorption_test": (
            "R10-2 (A-SX08): absorb an unallocated clause iff (a) it is in the "
            "same spec section as the leaf AND (b) it elaborates the leaf's "
            "cited clause. On absorption: `[A-SX08]` in assumptions naming the "
            "absorbed id, AND that id added to specification_reference. "
            "Failing either condition: record a coverage hole in reasoning and "
            "leave it to RD-1 — never absorb silently. Whole-section gaps "
            "cannot pass (a). Read each clause's scope tags before deciding: "
            "ECU / Market / Radio / EE Architecture decide whether the clause "
            "is even this configuration's behaviour."),
        "exemplars": exemplars.get("exemplars", []),
        "fingerprint": _fingerprint("features/sxm"),
        "exemplar_basis": exemplars.get("basis"),
        "column_conventions": {
            "test_group": {"value": cfg["test_group"],
                           "note": "framework Part IV Layer 1 — every generated "
                                   "row carries this; the workbook is BLANK, so "
                                   "there is no local precedent to match"},
            "test_set": {"value": "per leaf — see each leaf's `test_set`",
                         "note": "framework Part IV Layer 2 — the capability "
                                 "name, NOT the batch label. Batches are a "
                                 "generation unit; B1 and B2 are both Instant "
                                 "Replay. **Each leaf carries its own Set** — "
                                 "this batch is not uniform: leaf 154 is "
                                 "Traffic & Weather. Layer 3 (CFTS024 section "
                                 "numbers) is framework-internal, never "
                                 "written."},
            "author": {"value": cfg["write_back"]["author_value"],
                       "note": "every row is generated here — BLANK workbook, "
                               "no other author to preserve"},
            "tc_ref_id": {"value": cfg["write_back"]["tc_ref_id_value"]},
            "spec_reference": {
                "template": cfg["spec_reference_template"],
                "note": "profile §3.5 — the id comes from data/stla_to_cfts.json "
                        "(HYBRID: ReqIF clause, docx printed section), never "
                        "guessed. Cite-form second tokens per R11 where the "
                        "leaf's cross_references say so."},
            "estimated_test_time": {
                "value": None,
                "note": "column Q, revision C only — LEFT BLANK. Not a "
                        "convention: no fill policy is ruled (A-SX05), and "
                        "estimating without a source is fabrication."},
        },
        "context_note": batch["context_note"],
        "anomalies": batch["anomalies"],
        "blocked_section_text": blocked,
    }


def run(args) -> int:
    root = Path(args.feature_dir)
    cfg = load_feature_config(args.feature_dir)
    batches = load_batches(root / args.batches_md)

    if args.list:
        for name, b in batches.items():
            print(f"{len(b['req_ids']):3}  {name:22} "
                  f"{', '.join(b['declared_sections'])}")
        return 0

    wanted = list(args.batch or [])
    if args.pilot:
        wanted = [n for n in batches if n in ("Tuner Availability", "Tune")]
    if not wanted:
        raise ContextError("give --batch NAME (repeatable), --pilot, or --list")

    stla_path = root / "data" / "stla_to_cfts.json"
    if not stla_path.exists():
        raise ContextError("data/stla_to_cfts.json missing — run "
                           "build_stla_map.py first")
    stla_map = json.loads(stla_path.read_text(encoding="utf-8"))
    ex_path = root / "data" / "exemplars.json"
    exemplars = (json.loads(ex_path.read_text(encoding="utf-8"))
                 if ex_path.exists() else {})
    if not exemplars:
        print("WARNING: no data/exemplars.json — generating without style "
              "anchors", file=sys.stderr)
    spec_docs = cfg.get("spec_docs") or {}
    # Section text per document. Keyed by doc because section NUMBERS collide
    # across the three CFTS files (CFTS011 shares 30 numbers with CFTS024,
    # CFTS004 shares 38): a flat map would silently hand a leaf the identically
    # numbered section of the wrong document, which reads as perfectly normal
    # spec text.
    sections = {spec_docs.get("primary"): section_texts(
        resolve_path(cfg, "cfts_doc"))}
    for doc, key in (spec_docs.get("docs") or {}).items():
        if doc in (spec_docs.get("external") or {}) and cfg["paths"].get(key):
            sections[doc] = section_texts(resolve_path(cfg, key))
    ua_path = root / "data" / "unallocated_clauses.json"
    unalloc = (json.loads(ua_path.read_text(encoding="utf-8"))
               if ua_path.exists() else {})
    set_index = test_set_index(cfg)
    cite_path = root / "data" / "cross_doc_citations.json"
    citations = (json.loads(cite_path.read_text(encoding="utf-8"))
                 if cite_path.exists() else {})
    if not citations:
        print("WARNING: no data/cross_doc_citations.json — leaves citing "
              "another CFTS will reach the generator unflagged; run "
              "build_stla_map.py", file=sys.stderr)

    out_dir = root / "batches"
    out_dir.mkdir(exist_ok=True)
    for name in wanted:
        if name not in batches:
            raise ContextError(f"unknown batch {name!r}; --list to see them")
        tables = load_spec_tables(cfg, root, batches[name]["spec_tables"])
        ctx = build(cfg, batches[name], stla_map, sections, exemplars,
                    spec_docs, tables, citations, unalloc, set_index)
        slug = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
        path = out_dir / f"{slug}.json"
        path.write_text(json.dumps(ctx, ensure_ascii=False, indent=1),
                        encoding="utf-8")
        by_doc = ctx["spec_sections"]
        n_sections = sum(len(v) for v in by_doc.values())
        chars = sum(len(s["text"]) for v in by_doc.values() for s in v.values())
        tbl = ctx.get("spec_tables") or {}
        diverged = [l["req_id"] for l in ctx["leaves"] if "wording_note" in l]
        cross = [f"{l['req_id']}:{r['token']}" for l in ctx["leaves"]
                 for r in l.get("cross_references", [])]
        if cross:
            print(f"  cross-document citations (R11 cite-form, multi-cite + "
                  f"anchored ER): {cross}", file=sys.stderr)
        if diverged:
            print(f"  037 title diverges from the CFTS clause: {diverged} "
                  "(§8.6 — spec wins on wording)", file=sys.stderr)
        print(f"{name:22} {len(ctx['leaves'])} leaves, "
              f"{n_sections} sections in {len(by_doc)} doc(s) ({chars} chars), "
              f"{len(ctx['siblings'])} siblings, "
              f"{len(ctx['exemplars'])} exemplars"
              + (", tables " + "+".join(
                  f"{k}({len(v['rows'])} rows)" for k, v in tbl.items())
                 if tbl else "")
              + (f", {len(ctx['blocked_section_text'])} without section text"
                 if ctx["blocked_section_text"] else "")
              + f" -> {path}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--feature-dir", default=".")
    ap.add_argument("--batches-md", default="docs/batches-amfm.md")
    ap.add_argument("--batch", action="append", help="batch name (repeatable)")
    ap.add_argument("--pilot", action="store_true",
                    help="the pilot pair proposed in batches-amfm.md")
    ap.add_argument("--list", action="store_true")
    args = ap.parse_args()
    try:
        return run(args)
    except ContextError as exc:
        print(f"ABORTED: {exc}", file=sys.stderr)
        return 1



# --- R-G19：prompt／exemplar 指紋（26 包 §D-5）------------------------------

def _fingerprint(feature_dir: str) -> dict:
    """本批實際使用之 prompt 模板與 exemplar 集之 sha（R-G19@`bd206972`）。

    指紋之計算集中於 `scripts/prompt_fingerprint.py`，各 feature 只呼叫。
    **取不到時回 `{"error": ...}` 而非省略該鍵** —— 少一個鍵與指紋相符
    在 manifest 上長得不一樣，而少一個鍵與「沒有指紋」長得一樣（G-D）。
    """
    import importlib.util
    root = Path(__file__).resolve().parents[3]
    spec = importlib.util.spec_from_file_location(
        "tc_prompt_fingerprint", root / "scripts" / "prompt_fingerprint.py")
    if spec is None or spec.loader is None:
        return {"error": "scripts/prompt_fingerprint.py 不可載入"}
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.fingerprint(root, feature_dir)

if __name__ == "__main__":
    sys.exit(main())
