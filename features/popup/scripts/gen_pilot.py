#!/usr/bin/env python3
"""Popup pilot 批（下放包 02 §六-5）—— 全量一批。

**R-G3**：全程不呼叫 `openpyxl` 之 `save()`；工作簿以
`backend/xlsx_surgical.surgical_save` 之 zip 層外科寫入。
母本 R10:R1411 之設計方法下拉為 **x14 擴充**，openpyxl 讀不到，
`save()` 會靜默刪除它（A-POP5）——故本路徑不是偏好，是唯一合法路徑。

**R-G25**：輸出落 `features/popup/sandbox/<tag>/`。

值一律逐字取自 `forms/Pop Up List HMI R1 (26PI).xlsx`（R-POP6），
引用時併記 PU id 與欄名（profile §2）。

**下放包 03（R-POP13／14／15）之六件修正**，逐件落於本檔之語料：

- **F1** 末步驟一律 `Read <對象> ... and check that <可觀察結果>`（IN §5.5）
- **F2** Procedure 之按壓標的一律 `"..."`；PU 控制記法（`<OK>`／`<Trks>`）
  **只保留於 ER 引文段與 test_item**；反引號等 Markdown 記號不進交付欄（IN §11）
- **F3** 刪 `The vehicle is stationary with the ignition in RUN` 類環境穩定性
  前提（IN §8.5）—— popup 關閉行為之規格側無運動狀態觸發
- **F4** timeout 值單一欄位歸屬（IN §4.5）：`input_test_data` 一律 `NA`，
  值內聯於 Procedure／ER
- **F5／F6** reasoning 之 anomaly 號 live 查台帳後書寫；理由以裁定為準
- **R-POP13** TC ID 前綴改 `NR1L-Popup-{NNN}`（取自 `feature.yaml`）
- **R-POP14** `-002-05` 照 GP4-4 規格原句生成，**不引 PU**、不落 PENDING
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[3]
FEAT = ROOT / "features/popup"
sys.path.insert(0, str(ROOT))
from backend.xlsx_surgical import surgical_save  # noqa: E402

SPEC = "SYS1_HMI_Core_HMI_Logic_and_Flow_R1_SR24_Post_2A_(February_2_2023)"
GROUP, TEST_SET = "Popup", "Pop-up Close"
DESIGN_METHOD = "狀態轉換 (State Transition Testing)"

# 每條 TC 之 PU 引用 —— `pu` 欄之值於落檔前逐格對 Pop Up List 原檔複驗
# （verify_pu_quotes()），不符即停：靜態轉錄與其來源分家是 G-F 之標的。
TCS = [
    {
        "req_id": "SWE1-POP-002-01",
        "pu": {"id": "PU0942",
               "Timeout (sec)": "5",
               "String/Popup Message": "<X>\nPage added! [Reorder]",
               "Module": "Home Screen"},
        # 上半 = 037 E 欄 verbatim；下半 = 作者生成之情境標籤（IN §4.3.1）
        "test_item": [
            "Pop-ups can be closed after time-out "
            "(timeout is defined in Pop-up List document)",
            "(Time-out closure of PU0942, 5 s)",
        ],
        "pre_conditions": [
            "The Home Screen is displayed",
        ],
        "input_test_data": "NA",
        "steps": [
            "Enter the Home Screen page-management view and add a home screen page",
            "Leave the screen and all buttons untouched for 10 seconds after the "
            "pop-up appears",
            "Read the pop-up display status and the elapsed time and check that "
            "the pop-up has closed by itself 5 seconds after it appeared, "
            "matching the Time-out defined by PU0942 Timeout (sec)",
        ],
        "expected": [
            'The pop-up is displayed showing "<X>" and "Page added! [Reorder]", '
            "as defined by PU0942 String/Popup Message",
            "No user interaction is registered during the 10-second window",
            "The pop-up closes by itself 5 seconds after it appeared, matching the "
            "Time-out defined by PU0942 Timeout (sec), and the Home Screen is shown "
            "without the pop-up",
        ],
        "spec_sections": ["5.6"],
        "reasoning":
            "驗證目標單一：GP4 第 1 途徑（time-out 自動關閉）。"
            "PU 之選定理由 —— (a) 類 240 列中取 PU0942：Timeout (sec) 為純數值 `5`"
            "（非 `5 seconds` 之帶單位寫法，量測不需再解析），觸發路徑純屬 HMI"
            "（新增 home screen 頁），台架上不需車輛運動、外接裝置或雲端連線即可重現。"
            "步驟 2 之觀察窗設為 10 秒（2× 標稱值）而非剛好 5 秒 —— 只看到 5 秒時仍在，"
            "分不出「還沒關」與「不會關」。timeout 值不落 PENDING：DR-POP1 已由 "
            "R-POP6 結案，值有來源可引。",
    },
    {
        "req_id": "SWE1-POP-002-02",
        "pu": {"id": "PU0215",
               "Exit Conditions": "<Trks>",
               "String/Popup Message": "Tracks List",
               "Description": "Display when user is in Media screen and presses Track list button",
               "Module": "Media"},
        "test_item": [
            "Pop-ups can be closed after pressing button that opened pop-up again "
            "(eg. Tracks popups)",
            "(Second press of the opening control <Trks>, PU0215)",
        ],
        "pre_conditions": [
            "A media source containing at least one track list is available",
            "The Media screen is displayed",
        ],
        "input_test_data": "NA",
        "steps": [
            'On the Media screen, press the Track list button "Trks"',
            'Press "Trks" a second time while the pop-up is displayed',
            "Read the pop-up display status immediately after the second press "
            "and check that the pop-up is no longer displayed",
        ],
        "expected": [
            'The "Tracks List" pop-up is displayed, as defined by '
            "PU0215 String/Popup Message",
            'The second press lands on the same "<Trks>" button that opened '
            "the pop-up",
            "The pop-up is closed and the Media screen is shown without the pop-up",
        ],
        "spec_sections": ["5.5", "5.6"],
        "reasoning":
            "驗證目標單一：GP4 第 2 途徑（再按開啟鍵關閉）。"
            "spec_reference 併列 `_5.5`＋`_5.6`（R-POP8）—— 5.5（GP3）與 5.6 第 2 途徑"
            "為同一行為之兩處敘述，037 K8 逐字 `Duplicated feature of SWE1-POP-002-02`。"
            "PU0215 為全表唯一 Exit Conditions 即開啟鍵本身者（`<Trks>`），"
            "且 037 E 欄自舉 `eg. Tracks popups`，來源與例證對得上。"
            "input_test_data = `NA`：`<Trks>` 屬互動資料，依 IN §4.5 歸 Procedure，"
            "不重複列於本欄。"
            "**device 軸不拆，理由為規格側無此分支**（R-POP12，2026-08-27）："
            "SYS1 5.6 逐字 `pressing the button a second time`，未區分按鍵型別；"
            "037 S11 之 `a physical hard button or a specific UI button on the "
            "screen` 是 VC 對「button」之列舉性註解，非規格分支，037 本身亦未拆為"
            "兩個 sub-id。依 IN §8.2（RD 為需求單位之權威）與 §8.4.2"
            "（規則定義地測試），判一條 TC。"
            "**本條不主張「軸為真而缺實例」** —— 該說蘊含「欠一條待補件」，"
            "為 R-POP12 所否決。037 VC 之措辭記 Remarks 供 RD-1 順帶確認，不阻斷。",
    },
    {
        "req_id": "SWE1-POP-002-03",
        "pu": {"id": "PU0580",
               "Exit Conditions": "Timeout, Touch outside of popup, X",
               "Timeout (sec)": "5 seconds",
               "String/Popup Message": '"Welcome [username]", "X"',
               "Module": "Personal Account/Driver Profiles"},
        "test_item": [
            "Pop-ups can be closed when touching screen outside of pop-up",
            "(Touch outside the pop-up bounds, PU0580)",
        ],
        "pre_conditions": [
            "At least two driver Profiles exist on the head unit",
            "The All Profiles tab is reachable from the current screen",
        ],
        "input_test_data": "NA",
        "steps": [
            "Open the All Profiles tab and switch manually to another Profile",
            "Within 5 seconds of the pop-up appearing, tap an area of the screen "
            "outside the pop-up window frame",
            "Read the pop-up display status immediately after the tap and check "
            "that the pop-up is no longer displayed, before the 5-second Time-out "
            "defined by PU0580 Timeout (sec) has elapsed",
        ],
        "expected": [
            'The pop-up is displayed showing "Welcome [username]", "X", '
            "as defined by PU0580 String/Popup Message",
            "The tap lands outside the pop-up window frame and activates no "
            "pop-up control",
            "The pop-up is closed, and the closure occurs before the 5-second "
            "Time-out defined by PU0580 Timeout (sec) elapses",
        ],
        "spec_sections": ["5.6"],
        "reasoning":
            "驗證目標單一：GP4 第 3 途徑（點 popup 外關閉）。"
            "037 K12 自陳該機制 `default to disable, requester should call the API "
            "to enable`，故受測 popup 必須是**明載已啟用**者 —— (b) 類 102 列即以 "
            "`Exit Conditions` 含 `outside` 為判準，PU0580 在列。"
            "選它而非其他 101 列：其 Exit Conditions 僅 `Timeout, Touch outside of "
            "popup, X` 三項，無雲端往返、無 keyboard，台架以切換 Profile 即可重現。"
            "步驟 2 之「5 秒內」與 ER 3 之「早於 time-out」為必要設計 —— "
            "PU0580 同時具 time-out，不設時窗則觀察到的關閉分不出是哪一條途徑造成。"
            "input_test_data = `NA`（R-POP15 F4）：`5 s` 已內聯於 Procedure 步驟 2／3 "
            "與 ER 3。"
            "pre_conditions 之「兩個 Profile 存在」「All Profiles tab 可達」**保留** —— "
            "那是本 TC 觸發該 popup 所必需之規格前提，非 IN §8.5 之環境穩定性前提"
            "（後者已依 R-POP15 F3 刪除）。",
    },
    {
        "req_id": "SWE1-POP-002-04",
        "pu": {"id": "PU0949",
               "Exit Conditions": "<OK>\n<X>",
               "Timeout (sec)": "3",
               "String/Popup Message": "No Phone is Connected.        <OK>    <X>",
               "Description": 'This popup will be displayed if user presses the grayed out "Make a Call" button on Shortcuts when there is no connected phone.',
               "Module": "Home Screen"},
        "test_item": [
            "Pop-ups can be closed after making a selection inside the pop-up "
            "if applicable",
            "(Selection <OK> inside the pop-up, PU0949)",
        ],
        "pre_conditions": [
            "No mobile phone is paired with the head unit",
            "The Home Screen Shortcuts view is displayed",
        ],
        "input_test_data": "NA",
        "steps": [
            'On the Home Screen Shortcuts view, press the grayed out "Make a Call" '
            "button",
            'Within 3 seconds of the pop-up appearing, press "OK" inside the pop-up',
            "Read the pop-up display status immediately after the press and check "
            "that the pop-up is no longer displayed, before the 3-second Time-out "
            "defined by PU0949 Timeout (sec) has elapsed",
        ],
        "expected": [
            "The pop-up is displayed showing "
            '"No Phone is Connected.        <OK>    <X>", as defined by '
            "PU0949 String/Popup Message",
            'The "<OK>" press is registered while the pop-up is still displayed',
            "The pop-up is closed, and the closure occurs before the 3-second "
            "Time-out defined by PU0949 Timeout (sec) elapses",
        ],
        "spec_sections": ["5.6"],
        "reasoning":
            "驗證目標單一：GP4 第 4 途徑（選擇後關閉）。"
            "受測對象取 `<OK>` 而非 `<X>`：037 S11 之 VC 逐字舉例 "
            "`tap an option or press a \"Confirm\"/\"Cancel\" button inside the pop-up`，"
            "`<X>` 是關閉鍵不是選項，測 `<X>` 測不到本 leaf 的命題。"
            "PU0949 之 Exit Conditions 逐字含 `<OK>` 與 `<X>` 兩者，故「選 `<OK>` 會關」"
            "有來源明載，非推定。時窗設計同 -002-03：PU0949 之 time-out 為 3 秒，"
            "不設時窗則關閉原因不可辨。"
            "input_test_data = `NA`（R-POP15 F4）：`3 s` 已內聯於 Procedure 步驟 2／3 "
            "與 ER 3。",
    },
]

# -002-05 —— **本輪生成**（R-POP14，A-POP8 採乙案改良）。
# 下放包 02 曾因 search keyboard 於 Pop Up List 查無對應列而停下（A-POP8）；
# 分析層裁定 GP4-4 之 `e.g in the search keyboard` 是**規格自己的舉例**，
# 不是向 Pop Up List 之委派（對照 GP4-1 逐字
# `timeout is defined in Pop-up List document` 才是委派），故不適用 R-POP6
# 之值引用規則 —— 照規格原句生成、**不引 PU**、不落 PENDING、
# 不列舉 search keyboard 以外之實例（IN §8.4.1）。
# `pu` 為 None：`verify_pu_quotes()` 與 `render()` 皆須容得下無 PU 之條。
TCS.append({
    "req_id": "SWE1-POP-002-05",
    "pu": None,
    # 上半 = 037 E 欄 verbatim（含其原有之直引號與連字號寫法）
    "test_item": [
        "Exceptions when the popup allows the user to perform more than 1 task- "
        "e.g in the search keyboard only X button 'close', any other buttons "
        "do not close the popup",
        "(Non-closing control inside a multi-task pop-up; no PU cited)",
    ],
    # F3 後只餘規格觸發前提。兩項皆自 GP4-4 與 037 S11 之 Precondition 逐字所本，
    # 不引 PU、不指名具體 popup。
    "pre_conditions": [
        "A pop-up that allows the user to perform more than 1 task is available, "
        "the search keyboard being the example given in the requirement",
        "That pop-up is displayed",
    ],
    "input_test_data": "NA",
    "steps": [
        "Trigger the pop-up that allows the user to perform more than 1 task",
        'Inside the displayed pop-up, press a button other than the "X" button',
        "Read the pop-up display status immediately after the press and check "
        "that the pop-up is still displayed",
    ],
    "expected": [
        "The pop-up is displayed and allows the user to perform more than 1 task",
        'The button pressed is a button other than the "X" button of that pop-up',
        "The pop-up is not closed by that press and remains displayed",
    ],
    "spec_sections": ["5.6"],
    "reasoning":
        "驗證目標單一：GP4-4 之**例外分支** —— multi-task popup 內，"
        "非 `X` 之按鍵按下後 popup **不關閉**。037 E 欄逐字 "
        "`Exceptions when the popup allows the user to perform more than 1 task- "
        "e.g in the search keyboard only X button 'close', any other buttons "
        "do not close the popup`；037 S11 之 Action／Expected Result 逐字 "
        "`Tap the button (or interactive component) that is not suppose close "
        "pop-up` ／ `The pop-up should not be closed`，命題與本條一致。"
        "**不引 PU**（R-POP14）：`e.g in the search keyboard` 是規格自載之舉例，"
        "非向 Pop Up List 之委派 —— 對照 GP4-1 逐字 "
        "`timeout is defined in Pop-up List document` 才是委派。"
        "故本條之 `pu_citation` 為 null，`spec_reference` 單行 `_5.6`。"
        "**不代入 PU0022／PU0023／PU0861**：那須先認定「該列即 search keyboard」，"
        "而 Pop Up List 三 sheet 全欄實測 `search keyboard` 連續詞組命中 0、"
        "同列兼含 `keyboard` 與 `search` 之 PU 列 0（A-POP8 之量測），"
        "該認定為來源沒有承載之推定（IN §8.4.1）。"
        "**只測「不關閉」一個命題，不併測「按 X 會關」** —— 後者屬 GP4-4 前半，"
        "已由 `-002-04` 之 `<OK>` 覆蓋其「選擇後關閉」面；併入會使本條有兩個"
        "驗證目標。"
        "design_method 取狀態轉換而非負向測試：受測者是同一台 popup 狀態機在"
        "特定輸入下**不發生轉移**，與 -002-01～04 同一機、同一族，"
        "四條同法可對讀；負向測試曾列為候補，記此以備 Pei 改裁。"
        "multi-task popup 之完整例外清單另開 **DR-POP4**（2026-08-27 live 查 "
        "DATA_REQUESTS.md：已登記、未送出），不阻斷本條。"})


def cell(row, i):
    return "" if i >= len(row) or row[i] is None else str(row[i]).strip()


def verify_pu_quotes(src: Path) -> list[str]:
    """每一個 PU 引文對原檔逐格複驗（G-F）。不符即回報，由呼叫端停。"""
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    rows = list(wb["Main"].iter_rows(values_only=True))
    wb.close()
    header = [str(v).replace("\n", " ").strip() if v else "" for v in rows[1]]
    index = {h: i for i, h in enumerate(header)}
    by_id = {cell(r, 0): r for r in rows[2:] if re.match(r"^PU\d", cell(r, 0))}

    bad = []
    for tc in TCS:
        pu = tc["pu"]
        if pu is None:
            continue          # -002-05 不引 PU（R-POP14），無引文可驗
        row = by_id.get(pu["id"])
        if row is None:
            bad.append(f"{pu['id']}：Main 查無此 id")
            continue
        for field, quoted in pu.items():
            if field == "id":
                continue
            i = index.get(field)
            if i is None:
                bad.append(f"{pu['id']}：Main 無欄 {field!r}")
                continue
            actual = cell(row, i)
            if actual != quoted:
                bad.append(f"{pu['id']} {field}：引文 {quoted!r} ≠ 原檔 {actual!r}")
    return bad


def render(tc: dict, n: int, cfg: dict) -> dict:
    return {
        "tc_id": cfg["tc_id_format"].format(n=n),
        "req_id": tc["req_id"],
        "test_group": GROUP,
        "test_set": TEST_SET,
        "test_item": "\n".join(tc["test_item"]),
        "pre_conditions": "\n".join(f"{i}. {s}" for i, s
                                    in enumerate(tc["pre_conditions"], 1)),
        "input_test_data": tc["input_test_data"],
        "test_procedure": "\n".join(f"{i}. {s}" for i, s
                                    in enumerate(tc["steps"], 1)),
        "expected_result": "\n".join(f"{i}. {s}" for i, s
                                     in enumerate(tc["expected"], 1)),
        "spec_reference": "\n".join(f"{SPEC}_{s}" for s in tc["spec_sections"]),
        "tc_ref_id": cfg["write_back"]["tc_ref_id_value"],
        "priority": "P1",
        "design_method": DESIGN_METHOD,
        "functional_safety": "NA",
        "author": cfg["write_back"]["author_value"],
        "pu_citation": tc["pu"],   # -002-05 為 None（R-POP14：不引 PU）
        "reasoning": tc["reasoning"],
    }


DELIVERY_FIELDS = ("test_item", "pre_conditions", "input_test_data",
                   "test_procedure", "expected_result")


def audit(tcs: list[dict]) -> None:
    """下放包 03 §八之預期數字，自語料逐項實測後印出。

    **不在此判 PASS／FAIL** —— 數字與預期不符時由人停下（§九：不符停下
    不調和）。工具只負責把數字量出來，且量測條件與 §八 之「量測條件」
    欄逐項對應，量法與宣稱分家就等於沒量。
    """
    def items(v):
        return [x for x in str(v).split("\n") if x.strip()]

    rows = [
        ("TC 總數", len(tcs), 5),
        ("PENDING 佔位", sum("PENDING:" in str(v) for t in tcs
                             for v in t.values() if isinstance(v, str)), 0),
        ("`newR1L` 殘留", sum("newR1L" in str(t) for t in tcs), 0),
        ("`PROJ` 殘留", sum("PROJ" in str(t) for t in tcs), 0),
        ("tc_id 含 `POP-`", sum("POP-" in t["tc_id"] for t in tcs), 0),
        ("反引號殘留（五交付欄逐 item）",
         sum("`" in it for t in tcs for f in DELIVERY_FIELDS
             for it in items(t[f])), 0),
        ("input_test_data == `NA`",
         sum(t["input_test_data"] == "NA" for t in tcs), 5),
        ("pre_conditions 含 `ignition in RUN`",
         sum("ignition in run" in t["pre_conditions"].lower() for t in tcs), 0),
        ("Final Step 含 `check that`",
         sum("check that" in items(t["test_procedure"])[-1] for t in tcs), 5),
        ("spec_reference 兩行者", sum(len(items(t["spec_reference"])) == 2
                                       for t in tcs), 1),
        ("pu_citation 為 null 者",
         sum(t["pu_citation"] is None for t in tcs), 1),
    ]
    print("\n§八 預期數字（自 generated/ 語料實測）")
    print("| 項 | 實測 | 預期 | 判 |")
    print("|---|---|---|---|")
    for name, got, want in rows:
        print(f"| {name} | {got} | {want} | {'相符' if got == want else '**不符**'} |")


def main() -> int:
    cfg = yaml.safe_load((FEAT / "feature.yaml").read_text(encoding="utf-8"))
    src_list = sorted(FEAT.glob(cfg["paths"]["popup_list"]))
    if len(src_list) != 1:
        sys.exit(f"paths.popup_list 命中 {len(src_list)} 檔")
    bad = verify_pu_quotes(src_list[0])
    if bad:
        print("FAIL（G-F：引文與原檔不符，停）", file=sys.stderr)
        for b in bad:
            print("  " + b, file=sys.stderr)
        return 2
    quoted = sum(len(t["pu"]) - 1 for t in TCS if t["pu"])
    no_pu = [t["req_id"] for t in TCS if not t["pu"]]
    print(f"PU 引文複驗：{quoted} 格全數相符"
          f"（不引 PU 者 {len(no_pu)} 條：{'、'.join(no_pu) or '—'}）")

    tcs = [render(tc, i, cfg) for i, tc in enumerate(TCS, 1)]
    out = FEAT / "generated/pilot_01.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(tcs, ensure_ascii=False, indent=2) + "\n",
                   encoding="utf-8")
    print(f"寫入 {out.relative_to(ROOT)}：{len(tcs)} 條")

    # --- 工作簿 ---------------------------------------------------------
    wb_src = sorted(FEAT.glob(cfg["paths"]["workbook"]))[0]
    col = cfg["workbook"]["columns"]
    sheet = cfg["workbook"]["sheet"]
    first = cfg["workbook"]["first_data_row"]

    wb = openpyxl.load_workbook(wb_src)          # **不 save()**（R-G3）
    ws = wb[sheet]
    def idx(letter):
        n = 0
        for ch in letter:
            n = n * 26 + (ord(ch) - 64)
        return n
    for k, tc in enumerate(tcs):
        r = first + k
        ws.cell(r, idx("F")).value = tc["tc_id"]
        for key in ("req_id", "test_group", "test_set", "test_item",
                    "pre_conditions", "input_test_data", "test_procedure",
                    "expected_result", "spec_reference", "tc_ref_id",
                    "priority", "design_method", "functional_safety", "author"):
            ws.cell(r, idx(col[key])).value = tc[key]

    audit(tcs)

    dst = FEAT / "sandbox/pilot01" / wb_src.name
    report = surgical_save(wb, wb_src, dst)
    print(f"工作簿 → {dst.relative_to(ROOT)}")
    print("  patched:", report["sheets_patched"])
    for k, v in report.items():
        if k not in ("sheets_patched", "members_patched"):
            print(f"  {k}: {v}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
