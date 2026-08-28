#!/usr/bin/env python3
"""Popup 交付候選簿（下放包 05 §二）。

**產出方式為「位元複製」而非重寫**：`sandbox/pilot01/` 之工作簿已由
`gen_pilot.py` 經 `surgical_save` 產出且經 lint 全綠、x14 DV 實測存活，
交付候選與它**不應有任何一格之差**。故本檔以 `shutil.copy2` 複製，
再以 **sha256 相等**證明其無損 —— 這比再跑一次 `surgical_save` 強：
後者只能證明「這次也沒壞」，前者證明「與已驗過的那一份是同一份」。

R-G3 之禁令針對的是 `openpyxl.save()` 之**改寫**（A-POP5：會靜默刪除
x14 下拉）。本檔不改任何一格，故不觸發該路徑；落檔後仍以 `zipfile`
直讀複驗 x14 DV，不以「應該沒事」代替量測。

**R-POP22**：Q 欄（Estimated Test Time）留空，不寫入。
**§三**：E 欄（TestRail）留空 —— 五條全為 NEW，無舊 ID 可映。
**§八 升級條件**：`output/` 已有同名檔即停下回報，**不覆寫**。
"""
from __future__ import annotations

import hashlib
import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[3]
FEAT = ROOT / "features/popup"
DELIVERY_NAME = ("FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT "
                 "STLA Test Case Specification & Result_SWQT_Popup_20260828.xlsx")

EXPECTED_F = [f"NR1L-Popup-{n:03d}" for n in range(1, 6)]
EXPECTED_D = [f"SWE1-POP-002-{n:02d}" for n in range(1, 6)]
DESIGN_METHOD = "狀態轉換 (State Transition Testing)"


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def x14_dv(path: Path) -> list[tuple[str, str, str]]:
    """`zipfile` 直讀取回 (sheet, f, sqref) —— openpyxl 讀不到 x14 擴充。"""
    out = []
    with zipfile.ZipFile(path) as z:
        for name in sorted(n for n in z.namelist()
                           if n.startswith("xl/worksheets/sheet")):
            xml = z.read(name).decode("utf-8", "replace")
            for f, sq in zip(re.findall(r"<xm:f>([^<]*)</xm:f>", xml),
                             re.findall(r"<xm:sqref>([^<]*)</xm:sqref>", xml)):
                out.append((name, f, sq))
    return out


def main() -> int:
    cfg = yaml.safe_load((FEAT / "feature.yaml").read_text(encoding="utf-8"))
    src = sorted(FEAT.glob("sandbox/pilot01/*.xlsx"))
    if len(src) != 1:
        sys.exit(f"sandbox/pilot01/ 命中 {len(src)} 檔，預期 1")
    src = src[0]

    out_dir = FEAT / "output"
    existing = sorted(out_dir.glob("*")) if out_dir.is_dir() else []
    print(f"list_directory {out_dir.relative_to(ROOT)}/："
          f"{[p.name for p in existing] or '（目錄不存在或為空）'}")
    dst = out_dir / DELIVERY_NAME
    if dst.exists():
        print(f"FAIL（§八 升級條件）：同名檔已存在，不覆寫 —— {dst}",
              file=sys.stderr)
        return 2

    out_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    s_src, s_dst = sha256_of(src), sha256_of(dst)
    print(f"複製 {src.relative_to(ROOT)}\n  → {dst.relative_to(ROOT)}")
    print(f"  sha256 來源 {s_src}")
    print(f"  sha256 產出 {s_dst}")
    if s_src != s_dst:
        print("FAIL：複製後 sha256 不相等", file=sys.stderr)
        return 2
    print("  無損：兩者相等")

    # --- 逐欄回讀複驗（自產出檔，不自語料）-------------------------------
    wbk = cfg["workbook"]
    wb = openpyxl.load_workbook(dst)          # **不 save()**（R-G3）
    ws = wb[wbk["sheet"]]
    first = wbk["first_data_row"]
    rows = range(first, first + 5)

    def col(letter, r):
        return ws[f"{letter}{r}"].value

    dropdown = [c[0].value for c in wb["下拉選單"]["A1:A9"]]

    checks = [
        ("交付簿資料列（F 欄非空）", sum(bool(col("F", r)) for r in rows), 5),
        ("F 欄值相符", sum(col("F", r) == e for r, e in zip(rows, EXPECTED_F)), 5),
        ("D 欄值相符", sum(col("D", r) == e for r, e in zip(rows, EXPECTED_D)), 5),
        ("G 欄 = Popup", sum(col("G", r) == "Popup" for r in rows), 5),
        ("H 欄 = Pop-up Close", sum(col("H", r) == "Pop-up Close" for r in rows), 5),
        ("P 欄 = P1", sum(col("P", r) == "P1" for r in rows), 5),
        ("Q 欄非空（R-POP22）", sum(col("Q", r) not in (None, "") for r in rows), 0),
        ("E 欄非空（§三）", sum(col("E", r) not in (None, "") for r in rows), 0),
        ("C 欄非空", sum(col("C", r) not in (None, "") for r in rows), 0),
        ("R 欄值屬下拉 9 字串", sum(col("R", r) in dropdown for r in rows), 5),
        ("O 欄 = NEW", sum(col("O", r) == "NEW" for r in rows), 5),
        ("AA 欄（Author）= PeiPYHsu",
         sum(col(wbk["columns"]["author"], r) == "PeiPYHsu" for r in rows), 5),
        ("spec_reference 兩行者",
         sum(len(str(col("N", r) or "").split("\n")) == 2 for r in rows), 1),
    ]

    pending = sum(1 for sheet in wb.worksheets for row in sheet.iter_rows()
                  for c in row if c.value and "PENDING:" in str(c.value))
    checks.append(("PENDING 佔位（全簿全欄）", pending, 0))

    dv = x14_dv(dst)
    checks.append(("x14 DV 個數", len(dv), 1))

    print("\n§六 預期數字（自交付簿回讀實測）")
    print("| 項 | 實測 | 預期 | 判 |")
    print("|---|---|---|---|")
    bad = 0
    for name, got, want in checks:
        ok = got == want
        bad += not ok
        print(f"| {name} | {got} | {want} | {'相符' if ok else '**不符**'} |")

    print("\nx14 DV（`zipfile` 直讀）")
    for sheet, f, sq in dv:
        print(f"  {sheet}: f={f!r}  sqref={sq!r}")

    print("\nB 欄（No.# 序號）—— 母本公式，非本包寫入")
    for r in rows:
        print(f"  B{r} = {col('B', r)!r}")

    print("\n下拉選單!$A$1:$A$9 逐字")
    for i, v in enumerate(dropdown, 1):
        print(f"  A{i} = {v!r}")

    print(f"\n交付簿 sha256：{s_dst}")
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())
