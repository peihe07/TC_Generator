#!/usr/bin/env python3
"""04 包步驟 4 — 列出 SYS1 匯出中未被任何 leaf 引用之 outline 並機器分類。

分類欄以規則產生（可重跑），不得以人讀結論填寫。規則依序求值，先命中先取：

  chapter_node      outline 無 `.`（第 1 層），即章節層節點本身
  image_placeholder Description 含 "Please refer to the diagram"
  assumptions       所屬章之 Description 逐字為 "Assumptions"
  other             以上皆非 —— **帶實質文字而未被引用者會落在這裡**

輸出 data/uncited_sections.tsv。不改動 scripts/recon.py 或任何共用腳本。

用法:
    python scripts/classify_uncited_sections.py --feature .
"""

import argparse
import re
from pathlib import Path

import openpyxl
import yaml

IMAGE_MARKER = "Please refer to the diagram"


def sort_key(outline: str):
    return [int(p) for p in outline.split(".")]


def classify(outline: str, desc: str, chapter_desc: str) -> str:
    if "." not in outline:
        return "chapter_node"
    if IMAGE_MARKER in desc:
        return "image_placeholder"
    if chapter_desc.strip() == "Assumptions":
        return "assumptions"
    return "other"


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--feature", default=".")
    args = ap.parse_args()
    fd = Path(args.feature).resolve()
    cfg = yaml.safe_load((fd / "feature.yaml").read_text(encoding="utf-8"))

    s = openpyxl.load_workbook(fd / cfg["paths"]["sys1_export"],
                               data_only=True)["Basic Report"]
    outlines = {}
    for r in range(2, s.max_row + 1):
        if not s.cell(r, 1).value:
            continue
        outlines[str(s.cell(r, 3).value).strip()] = {
            "polarion_id": str(s.cell(r, 1).value).strip(),
            "desc": str(s.cell(r, 4).value or ""),
        }

    a = openpyxl.load_workbook(fd / cfg["paths"]["a03_report"],
                               data_only=True)["Analysis Report"]
    cited = {str(a.cell(r, 3).value).split("_")[-1].strip()
             for r in range(8, a.max_row + 1)
             if str(a.cell(r, 7).value).strip() == "Functional Requirement"}

    uncited = sorted((o for o in outlines if o not in cited), key=sort_key)
    out = fd / "data" / "uncited_sections.tsv"
    out.parent.mkdir(exist_ok=True)
    with out.open("w", encoding="utf-8") as fh:
        fh.write("outline_number\tpolarion_id\tlevel\ttitle\tclassification\n")
        for o in uncited:
            ch = o.split(".")[0]
            d = outlines[o]["desc"]
            title = re.sub(r"\s+", " ", d.replace("_x000D_", " ")).strip()[:80]
            fh.write(f"{o}\t{outlines[o]['polarion_id']}\t{o.count('.') + 1}\t"
                     f"{title}\t{classify(o, d, outlines[ch]['desc'])}\n")

    # 餘數驗證（R-G10）：全集 − 引用 − 未引用 須為 0
    assert len(outlines) - len(cited) - len(uncited) == 0, "餘數非零"
    print(f"outline 全集 {len(outlines)}；被引用 {len(cited)}；"
          f"未引用 {len(uncited)}；餘數 {len(outlines) - len(cited) - len(uncited)}")
    print(f"written: {out}")


if __name__ == "__main__":
    main()
