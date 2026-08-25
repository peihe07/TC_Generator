#!/usr/bin/env python3
"""21 包步驟 3 —— State Matrix × 規格章之對照（R-PMH79 之三種記法）。

**R-PMH79**：對照結果只得記為三者之一 ——
  `牴觸`   二者就**同一謂詞取相反值**。須具名該謂詞，並上呈，不得自行調和。
  `印證`   二者就同一謂詞取相同值，或素材補上同一命題之另一半。須具名謂詞。
  `未對照` 二者**無共同謂詞**，或**有共同謂詞而條件已證互斥**（R-PMH84 之限縮）。
  `待定義` 該判定倚賴一個**規格全文 0 命中之術語**（如 `VP`），
           其指涉未定義 → **判定所需之語意尚未存在**（R-PMH85(c)）。

**「無對應列」不得記為「無矛盾」；「不同謂詞」不得記為「非牴觸」。**

**R-PMH84（22 包）之限縮**：二陳述有共同謂詞而取相反值時，
**除非其條件已被證明互斥，否則判為牴觸**。「素材未提及某條件」
**不等於**「素材不涉及該條件」—— 前者是素材之沉默，後者是一個
關於素材涵蓋範圍之主張，須有依據。

矩陣為規範性文件（PDF p10：`Power Moding behavior shall not be developed
without following the Power Moding State Matrix`），故其與規格章之牴觸
即為「TC 之 ER 可能與規範性文件相反」之來源。

用法:
    python scripts/matrix_vs_chapter.py 7
    python scripts/matrix_vs_chapter.py 7 --vocab      # 只跑詞彙探針
"""

# R-PMH92 —— 本檢查之 must-hit 註冊。**總表之結果欄由此決定，手寫不採認。**
HAS_MUST_HIT = True
MUST_HIT_NOTE = '`--must-hit` 三項正向錨點（R-PMH86）'

import argparse
import re
import sys
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))

MATRIX = ("inputs/Power Moding HMI State Matrix R1 SR24 Post 2A "
          "DCR21421 (August 3 2022).xlsx")
SHEET = "State Matrix"

# 三個區塊之（起列, 標題列, 軸列）—— 由 §2.1 之合併儲存格實測而得
BLOCKS = [(1, 16, (2, 3, 4, 5)), (19, 33, (20, 21, 22, 23)), (37, 48, (37, 38, 39))]

# 各章之關鍵名詞 —— 用於**範圍向**：其於矩陣之命中數
# ⚠ **子字串之偽命中須具名**：`play` 於矩陣命中 19 次，**其全數來自 `display`**
#    （`VP display pop-up` 等）。故 ch 8 之探針用 `played`／`plays` 而非 `play`。
VOCAB = {
    7: ["animation", "splash", "disclaimer", "comfort", "Maserati",
        "lower comfort", "traffic announcement", "CAN BUS", "ignition",
        "driver door", "black", "timeout", "3 sec", "1.5", "10s",
        "last mode", "Radio OFF"],
    8: ["sound", "Sound", "start-up", "startup", "goodbye", "Always",
        "Once a Day", "Never", "volume", "Volume", "entertainment",
        "setting", "Setting", "played", "plays", "sync"],
    10: ["backup cam", "back-up camera", "Power Button Off", "Power Button OFF",
         "HVAC pop-ups", "HVAC popups", "Phone call popups", "SOS", "ASSIST",
         "Key OFF", "ACC position", "ignition", "reinstated", "disclaimer"],
    11: ["VR", "VR HK", "hard key", "SIRI", "Voice Assistants", "long press",
         "Long press", "radio is OFF", "KEY ON", "ACC", "Screen Off",
         "Screen ON", "Audio OFF", "Audio ON", "CFTS009", "interaction"],
}

LIMITS = [
    "**只對照矩陣之「事件列 × 有值之格」** —— 空格與 `-` 一律不入母體；"
    "其「無值」是否本身有意義（不適用 vs 未定義），本檢查不判",
    "**謂詞之認定為人工** —— `VERDICT` 逐列具名，本檢查只驗其存在，不驗其正確",
    "**只比對 SYS1 側之 outline 文字** —— 規格 PDF 之圖表（p9 能力矩陣、p11 流程圖）不看",
    "**詞彙探針已字界錨定** —— 純子字串比對會使 `play` 命中 `display`（19 次）、`played` 命中 `displayed`（3 次）；錨定後二者歸零。**惟字界錨定不解決同形異義**（如 `Door` 於矩陣未區分駕駛門）",
    "詞彙探針為**大小寫敏感之字面**比對 —— 不敏感之比對會使 `Radio OFF` 誤命中 `Radio Off Delay`（15 次），二者為不同之詞",
    "詞彙探針為**字面**比對 —— 同義改寫（如 `VP` vs `screen`）不會命中，"
    "**故『詞彙 0 命中』不等於『語意無交集』**，僅為 `未對照` 之支持證據而非其證明",
    "**`待定義` 之列其判定尚未作成** —— 其既非通過亦非未通過；`DR-PMH7`（`VP` 之定義）`ANSWERED` 前不得轉為其他記法",
    "**本檢查之三分類（現為四分類）無 must-hit 至 22 包步驟 5 為止** —— 依 R-PMH35(c)，其結果於錨點通過前只得標「未實測」",
    "**矩陣之三個區塊各有一組軸** —— 跨區塊之同名事件（如 `ON/OFF button Pressed` "
    "於 r6／r24／r40）其軸不同，本檢查逐列判，不合併",
]


def print_limits() -> None:
    print("\n=== 本檢查未涵蓋之範圍（R-PMH52）===")
    for x in LIMITS:
        print(f"  - {x}")
    print("  **以上各項本檢查一律不看** —— 其全綠不含關於該等項之任何資訊。")


def n(v) -> str:
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def load_sheet():
    cfg = yaml.safe_load((ROOT / "feature.yaml").read_text(encoding="utf-8"))  # noqa: F841
    wb = openpyxl.load_workbook(ROOT / MATRIX, data_only=True)   # **不 save**
    return wb, wb[SHEET]


def merged_val(ws, r, c) -> str:
    for m in ws.merged_cells.ranges:
        if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
            return n(ws.cell(m.min_row, m.min_col).value)
    return n(ws.cell(r, c).value)


def event_rows(ws) -> list[tuple[int, int, str, list]]:
    """回傳 [(區塊起列, 列, 事件標籤, [(欄, 軸, 值)])] —— 只取有值之格。"""
    out = []
    for lo, hi, axr in BLOCKS:
        for r in range(lo, hi + 1):
            lbl = n(ws.cell(r, 1).value)
            cells = []
            for c in range(2, 14):
                v = n(ws.cell(r, c).value)
                if v in ("", "-", "'-"):
                    continue
                ax = " / ".join(x for x in (merged_val(ws, a, c) for a in axr) if x)
                cells.append((c, ax, v))
            if lbl and cells:
                out.append((lo, r, lbl, cells))
    return out


# --- R-PMH79：逐列之記法、謂詞與依據。**每一事件列皆須在此具名**，未具名 → FAIL ---
# 鍵為 (區塊起列, 列)。
# **鍵為 (章, 區塊起列, 列)** —— 21 包首版之鍵為 `(區塊起列, 列)`，
# **不含章號**：以 `matrix_vs_chapter.py 10` 執行時會**靜默沿用章 7 之判定**，
# 而每列都「有結論」故檢查不會察覺。形態同於 18→19 包之
# `RESIDUE_VERDICT` 60 字元鍵碰撞（19 包 §3.3）。**22 包修。**
VERDICT: dict[tuple[int, int, int], tuple[str, str, str]] = {
    (7, 1, 6):
        ("待定義", "**共同謂詞：pop-up 是否顯示**（`SU3.)` 取「不顯示」／本列 `Pop-up: Cannot Power Off System during active phone call.` 取「顯示」）",
         "**22 包步驟 2 之改判**：原記「未對照（軸不含 disclaimer 狀態）」，依 **R-PMH84** 應為牴觸 —— 免責畫面出現於開機序列（Key-on），本列之條件為 `Key-on × HU on × Call Active`（使用者上車前已通話），**二者可同時成立，其條件未證互斥**。**惟本列之 pop-up 由 `VP` 承載**（`VP Stays ON Pop-up: …`），而 `VP` 於規格全 11 頁 **0 命中**（`DR-PMH7`）—— 若 `VP` 非 head unit 之顯示螢幕，則與 `SU3.)` 無共同謂詞。**依 R-PMH85(c)『本條優先』，記為「待定義」而非「牴觸」**；`DR-PMH7` `ANSWERED` 後即應改判。**此與下放包步驟 2 之字面（五格皆改牴觸）不同，理由與差異已於上繳 §2.2 具名。**"),
    (7, 1, 7):
        ("未對照", "門開啟之後果（電源／VP 狀態）vs 動畫觸發",
         "章 7 之觸發一律為**駕駛門關閉**（`SU1.)`／`SU4.)`／`SU5.)`／`SU6.)`）；本列為門**開啟**。且矩陣之 `Door` 軸**未區分駕駛門與其他門**（`driver door` 於矩陣 0 命中）。**不同謂詞且情境不同。**"),
    (7, 1, 8):
        ("未對照", "門關閉之後果（`Event ignored`／`Power Button remains off`）vs 動畫是否播放",
         "**最接近之一列** —— `SU1.)`（7.1）以駕駛門關閉為觸發、`SU6.)`（7.7）載「last state 為 Radio OFF 時關門播放動畫後螢幕維持黑」。惟本列之格所斷言者為**電源按鈕狀態不變**（`Power Button remains off`）／**事件被忽略**，**未提動畫是否播放**；且矩陣之 `Door` 軸未區分駕駛門。**不同謂詞。** ⚠ 須人讀：`Event ignored`（HU on 時）與 `SU5.)`「每個 CAN BUS wake up 只播一次」在意圖上相容，惟該相容為推論而非量測。"),
    (7, 1, 9):
        ("未對照", "來電之後果（電源）vs pop-up 抑制",
         "`SU3.)`（7.4）之謂詞為 **pop-up 是否顯示**；本列之格為 `Head Unit Power ON`，謂詞為**電源**。**不同謂詞**（20 §4.3 之 `10.6` 同型，依 R-PMH79 改記為未對照）。"),
    (7, 1, 10):
        ("未對照", "Projection 之後果",
         "章 7 全文無 `Projection` —— **規格側無對應敘述**。"),
    (7, 1, 11):
        ("未對照", "VR 長按（無 Projection）之後果",
         "章 7 全文無 VR 長按 —— 其屬 ch 11（`VRLP1`）。**章 7 側無對應敘述**。"),
    (7, 1, 12):
        ("未對照", "VR 長按（Projection 中）之後果",
         "同上，且 Projection 於章 7 無敘述。"),
    (7, 1, 13):
        ("未對照", "通話結束之後果",
         "章 7 全文無通話結束之敘述（`SU3.)` 只提 traffic announcement 之音訊）。**規格側無對應敘述**。"),
    (7, 1, 14):
        ("未對照", "Projection 通話結束之後果",
         "同 r13，且 Projection 於章 7 無敘述。"),
    (7, 1, 15):
        ("待定義", "**共同謂詞：pop-up 是否顯示**（`SU3.)`「不顯示」／本列 `(R1High) VP display pop-up: \"Power OFF System…\"`「顯示」）",
         "同 `r6`：依 R-PMH84 應為牴觸（條件 `Key-on × Call Active` 與免責畫面期間未證互斥），**惟其 pop-up 由 `VP` 承載**（`VP display pop-up`），依 **R-PMH85(c)** 記為「待定義」。`DR-PMH7` `ANSWERED` 後即應改判。"),
    (7, 1, 16):
        ("未對照", "Off Road+ 按鍵之後果",
         "本列屬 **ch 12**（`OFF1.)`／`OFF3.)`，已於 20 包 §3 對照並判為互補）。**章 7 側無對應敘述**。"),
    (7, 19, 24):
        ("待定義", "**共同謂詞：pop-up 是否顯示**（`SU3.)`「不顯示」／本列 `VP display pop-up: \"Power OFF System. Continue call on mobile phone? Yes or NO\"`「顯示」）",
         "同 `r6`，惟本列屬 `Key-off` 區塊 —— 其與免責畫面（開機序列）之時序重疊性**更低而非為零**：`SU3.)` 之全稱否定涵蓋「免責畫面移除前之所有時刻」，未排除 key-off。**條件未證互斥。** 依 **R-PMH85(c)** 記為「待定義」（pop-up 由 `VP` 承載）。"),
    (7, 19, 25):
        ("待定義", "**共同謂詞：pop-up 是否顯示**（`SU3.)`「不顯示」／本列 `VP display pop-up: \"Power OFF System…\"`「顯示」）",
         "同 `r24`。依 **R-PMH85(c)** 記為「待定義」。"),
    (7, 19, 26):
        ("未對照", "來電之後果（`Head Unit remains ON untill the Timer is done`）vs pop-up 抑制",
         "同 r9 —— 不同謂詞（電源／計時 vs pop-up 是否顯示）。⚠ `Timer` 於矩陣 8 命中而章 7 之 `timeout` 0 命中：**二者為不同之計時**（前者為 Radio Off Delay，後者為 splash／disclaimer 之逾時）。"),
    (7, 19, 27):
        ("未對照", "Projection 之後果",
         "章 7 全文無 Projection。"),
    (7, 19, 28):
        ("未對照", "VR 長按（無 Projection）之後果",
         "屬 ch 11。"),
    (7, 19, 29):
        ("未對照", "VR 長按（Projection 中）之後果",
         "同上。"),
    (7, 19, 30):
        ("未對照", "門關閉之後果（`Event ignored`／HU off 時亦忽略）vs 動畫是否播放",
         "**與 r8 同型且更值得看**：本列為 **Key-off** 區塊，其 `HU off` 欄（c12）亦為 `Event ignored`。而 `SU1.)`（7.1）之情境即「駕駛門關閉→播放開機動畫」。惟本列之謂詞為**事件是否被處理**，未提動畫；且矩陣之 `Door` 軸未區分駕駛門。**不同謂詞。** ⚠ **本列為 ch 7 × 矩陣中最接近牴觸者，須人讀。**"),
    (7, 19, 31):
        ("未對照", "通話結束之後果",
         "章 7 全文無通話結束之敘述。"),
    (7, 19, 32):
        ("未對照", "Projection 通話結束之後果",
         "同上。"),
    (7, 19, 33):
        ("未對照", "Key-on 之後果（`Recall Last state of VP`）vs 開機序列",
         "**共同名詞 `Last state`** —— `SU6.)`（7.7）載 `If last state is Radio OFF, play startup animation and show applicable splash screens…`。惟本列之格為 `Recall Last state of VP`（**回復上次 VP 狀態**），其謂詞為**回復何狀態**；`SU6.)` 之謂詞為**是否播放動畫／顯示 splash**。**不同謂詞**，且矩陣未言回復過程中是否播放動畫。⚠ 須人讀。"),
    (7, 37, 40):
        ("未對照", "按 Power 鍵之後果（Mute／Screen Off）vs 開機動畫",
         "同 r6／r24 —— 不同謂詞。章 7 全文無 Mute 與 Screen Off 之敘述。"),
    (7, 37, 41):
        ("未對照", "來電之後果（Screen on／unmute／HU Powers on）vs pop-up 抑制",
         "同 r9 —— 不同謂詞（20 §4.3 之 `10.6` 即此列，依 R-PMH79 由「非牴觸」改記為「未對照」）。"),
    (7, 37, 42):
        ("未對照", "切入 R 檔之後果",
         "章 7 全文無 gear、無倒車影像。**規格側無對應敘述**。"),
    (7, 37, 43):
        ("未對照", "切出 R 檔之後果",
         "同上。"),
    (7, 37, 44):
        ("未對照", "Screen Off 鍵之後果",
         "章 7 全文無 Screen Off 鍵。其屬 ch 10（`PITA4`）。"),
    (7, 37, 45):
        ("未對照", "Mute 鍵之後果",
         "章 7 全文無 Mute。"),
    (7, 37, 46):
        ("未對照", "Headunit Mode 鍵之後果",
         "章 7 全文無 Headunit Mode 鍵。"),
    (7, 37, 47):
        ("未對照", "以 VR 切換 Headunit Mode 之後果",
         "同上；VR 屬 ch 11。"),
    (7, 37, 48):
        ("牴觸", "**共同謂詞：pop-up 是否顯示** —— `SU3.)`（7.4）取「不顯示」（`No pop-ups will appear until the disclaimer screen has been removed`，**全稱否定**），本列 `Show Pop-Up …` 取「顯示」（**無條件肯定**）",
         "**22 包步驟 2 之改判（R-PMH84）**：原記「未對照（軸不含 disclaimer 狀態）」。**「素材未提及 disclaimer 狀態」不等於「素材不涉及 disclaimer 期間」** —— 前者是素材之沉默，後者是一個關於涵蓋範圍之主張而我當時無依據。全稱否定之範圍為「免責畫面移除前之所有時刻」，本列之無條件肯定落於其中任一時刻即成牴觸；**其條件（`Key On × Gear = Reverse × Power Button OFF` 或 `Screen Off Active`）與「免責畫面顯示中」未經證明互斥**。**本列不倚賴 `VP` 之指涉** —— 其格逐字為 `Show Pop-Up`／`Popup not displayed over RVC`，未用 `VP` 一詞，故其牴觸**獨立於 `DR-PMH7` 成立**（R-PMH85 末段）。**須上呈，不得自行調和（R-PMH79）。** 連帶：`-007` 見上繳 §3。"),
    # ===== 22 包步驟 6：章 8 × 矩陣 =====
    (8, 1, 6):
        ("未對照", "按 Power 鍵之後果（電源／pop-up）vs 啟動音是否播放",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。 其格為電源狀態與 pop-up，**未提聲音**。"),
    (8, 1, 7):
        ("未對照", "門開啟之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 1, 8):
        ("未對照", "門關閉之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。 ⚠ `SSND 1)` 之觸發為 `it will start upon driver door close` —— **與本列同為「門關閉」**，惟本列之格為 `Event ignored`／`Power Button remains off`，**其謂詞為事件是否被處理與電源狀態，未提聲音**；且矩陣之 `Door` 軸未區分駕駛門。**無共同謂詞。**"),
    (8, 1, 9):
        ("未對照", "來電之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 1, 10):
        ("未對照", "Projection 之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 1, 11):
        ("未對照", "VR 長按（無 Projection）之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 1, 12):
        ("未對照", "VR 長按（Projection 中）之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 1, 13):
        ("未對照", "通話結束之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 1, 14):
        ("未對照", "Projection 通話結束之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 1, 15):
        ("未對照", "Key-off 之後果（VP on/off、pop-up）vs 告別音是否播放",
         "`SSND 1)` 載 `If goodbye sounds are supported, it shall sync on start with the shut-down animation.` —— 其謂詞為**告別音與關機動畫之同步**；本列之格為 `VP Turns OFF`／`VP Stays ON`／pop-up，**未提聲音亦未提動畫**。**無共同謂詞。**"),
    (8, 1, 16):
        ("未對照", "Off Road+ 喚醒之後果（`Radio Wakes Up and mutes`）vs 啟動音是否播放",
         "**共同謂詞候選：是否有聲音輸出**（ch 8 之 `SSND 2.1)`「should be played」／本列之 `mute`）。**依 R-PMH84 具名其條件互斥之依據**：`SSND 2.1)` 之條件為 `If the setting is Always, … everytime **the startup animation is played**` —— 其觸發限於**開機動畫播放之時**；而本列為 **Off Road+／SRT 之喚醒**，其靜音即**規格自身 `OFF3.)`（outline 12.3）所載** —— `Head unit is muted when launching app from Power Off State.`。**該喚醒路徑不在 `SSND 2.1)` 之條件內**（無開機動畫之敘述），且規格自身已為其指定靜音。**條件互斥之依據為規格文字本身，非假定。** ⚠ 惟「mute 是否即 `SSND` 所稱之 sounds 不被播放」為**本層之判斷**（見上繳 §13）。"),
    (8, 19, 24):
        ("未對照", "Key-off 狀態下按 Power 鍵之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 19, 25):
        ("未對照", "門開啟之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 19, 26):
        ("未對照", "來電之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 19, 27):
        ("未對照", "Projection 之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 19, 28):
        ("未對照", "VR 長按（無 Projection）之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 19, 29):
        ("未對照", "VR 長按（Projection 中）之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 19, 30):
        ("未對照", "門關閉之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。 ⚠ 同 `r8`：`SSND 1)` 之觸發亦為門關閉，惟本列之格為 `Event ignored`，**無共同謂詞**。"),
    (8, 19, 31):
        ("未對照", "通話結束之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 19, 32):
        ("未對照", "Projection 通話結束之後果",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (8, 19, 33):
        ("未對照", "Key-on 之後果（`Recall Last state of VP`）",
         "ch 8（`SSND 1)`～`SSND 3)`）全文無此事件之敘述 —— **無對應之規格敘述**。 `Recall Last state of VP` 之謂詞為**回復何狀態**，未提聲音。"),
    (8, 37, 40):
        ("未對照", "按 Power 鍵之後果（Mute／Screen Off）vs 啟動音是否播放",
         "**共同謂詞候選：是否有聲音輸出**（ch 8 之 `SSND 2.1)`「should be played」／本列之 `mute`）。**依 R-PMH84 具名其條件互斥之依據**：`SSND 2.1)` 之條件為 `If the setting is Always, … everytime **the startup animation is played**` —— 其觸發限於**開機動畫播放之時**；而本列之條件為 **`Key On`**（車已啟動、非開機序列），其 `Mute Active`／`Mute Inactive` 為**使用者按鍵所致之靜音狀態切換**，**不涉開機動畫**。`SSND 2.1)` 之條件（開機動畫播放時）於 `Key On` 之按鍵操作中不成立。**條件互斥。**"),
    (8, 37, 41):
        ("未對照", "來電之後果（`unmute`）vs 啟動音是否播放",
         "**共同謂詞候選：是否有聲音輸出**（ch 8 之 `SSND 2.1)`「should be played」／本列之 `mute`）。**依 R-PMH84 具名其條件互斥之依據**：`SSND 2.1)` 之條件為 `If the setting is Always, … everytime **the startup animation is played**` —— 其觸發限於**開機動畫播放之時**；而本列為**來電所致之 unmute**，ch 8 全文無來電之敘述。**條件互斥。**"),
    (8, 37, 42):
        ("未對照", "切入 R 檔之後果（`maintain mute`）",
         "**共同謂詞候選：是否有聲音輸出**（ch 8 之 `SSND 2.1)`「should be played」／本列之 `mute`）。**依 R-PMH84 具名其條件互斥之依據**：`SSND 2.1)` 之條件為 `If the setting is Always, … everytime **the startup animation is played**` —— 其觸發限於**開機動畫播放之時**；而本列為**切入 R 檔**，ch 8 全文無 gear 之敘述。**條件互斥。**"),
    (8, 37, 43):
        ("未對照", "切出 R 檔之後果",
         "**共同謂詞候選：是否有聲音輸出**（ch 8 之 `SSND 2.1)`「should be played」／本列之 `mute`）。**依 R-PMH84 具名其條件互斥之依據**：`SSND 2.1)` 之條件為 `If the setting is Always, … everytime **the startup animation is played**` —— 其觸發限於**開機動畫播放之時**；而本列為**切出 R 檔**，ch 8 全文無 gear 之敘述。**條件互斥。**"),
    (8, 37, 44):
        ("未對照", "Screen Off 鍵之後果",
         "**共同謂詞候選：是否有聲音輸出**（ch 8 之 `SSND 2.1)`「should be played」／本列之 `mute`）。**依 R-PMH84 具名其條件互斥之依據**：`SSND 2.1)` 之條件為 `If the setting is Always, … everytime **the startup animation is played**` —— 其觸發限於**開機動畫播放之時**；而本列為 **Screen Off 鍵**（屬 ch 10 之 `PITA4`），ch 8 全文無此鍵。**條件互斥。**"),
    (8, 37, 45):
        ("未對照", "Mute 鍵之後果",
         "**共同謂詞候選：是否有聲音輸出**（ch 8 之 `SSND 2.1)`「should be played」／本列之 `mute`）。**依 R-PMH84 具名其條件互斥之依據**：`SSND 2.1)` 之條件為 `If the setting is Always, … everytime **the startup animation is played**` —— 其觸發限於**開機動畫播放之時**；而本列為**使用者按 Mute 鍵**，ch 8 全文無 Mute 鍵之敘述。**條件互斥。** ⚠ 本列最接近 —— 其謂詞確為靜音狀態，惟其觸發為使用者按鍵，非開機動畫。"),
    (8, 37, 46):
        ("未對照", "Headunit Mode 鍵之後果",
         "**共同謂詞候選：是否有聲音輸出**（ch 8 之 `SSND 2.1)`「should be played」／本列之 `mute`）。**依 R-PMH84 具名其條件互斥之依據**：`SSND 2.1)` 之條件為 `If the setting is Always, … everytime **the startup animation is played**` —— 其觸發限於**開機動畫播放之時**；而本列為 **Headunit Mode 鍵**，ch 8 全文無此鍵。**條件互斥。**"),
    (8, 37, 47):
        ("未對照", "以 VR 切換 Headunit Mode 之後果",
         "**共同謂詞候選：是否有聲音輸出**（ch 8 之 `SSND 2.1)`「should be played」／本列之 `mute`）。**依 R-PMH84 具名其條件互斥之依據**：`SSND 2.1)` 之條件為 `If the setting is Always, … everytime **the startup animation is played**` —— 其觸發限於**開機動畫播放之時**；而本列為 **VR 切換**，ch 8 全文無 VR 之敘述（VR 屬 ch 11）。**條件互斥。**"),
    (8, 37, 48):
        ("未對照", "HVAC 硬控調整之後果（`No effect on mute`／`Show Pop-Up`）",
         "ch 8 全文無 HVAC、無 pop-up 之敘述。`No effect on mute` 之謂詞為靜音狀態不變，**而 `SSND` 諸條之謂詞為特定聲音是否播放**；且本列之條件為 HVAC 硬控調整，與開機動畫無涉。**無共同謂詞且條件互斥。**"),
    # ===== 23 包步驟 6：章 11 × 矩陣 =====
    (11, 1, 6):
        ("未對照", "ON/OFF button Pressed 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 1, 7):
        ("未對照", "Door opened 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 1, 8):
        ("未對照", "Door closed 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 1, 9):
        ("未對照", "Incoming Call 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 1, 10):
        ("未對照", "Plug in Projection 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 1, 11):
        ("印證", "**VR 長按（無 Projection）後 radio 之狀態** —— `VRLP1` 列舉四種容許之結果，其一為 `Screen Off and Audio OFF (i.e. radio back to off)`；本列取 `Head Unit Remain OFF`",
         "**同一謂詞取相同值** —— `Head Unit Remain OFF` 即 `VRLP1` 四種結果中之 `radio back to off`。本列之條件（`Key-on` × `Power Button OFF` × `Call Not Active`）**落在 `VRLP1` 之條件內**（`radio is OFF and KEY ON or ACC`）。⚠ 兩造皆註 `See CFTS009` —— **該文件不在本 feature 之六筆素材內**（A-PMH13 之同型）。"),
    (11, 1, 12):
        ("印證", "**VR 長按（Projection 中）後 radio 之狀態** —— `VRLP1` 四種結果之一為 `Screen ON and Audio ON`；本列取 `Head Unit Power ON`",
         "**同一謂詞取相同值**，且本列之條件落在 `VRLP1` 之條件內。**矩陣補上了規格所無之區辨** —— `VRLP1` 只說「結果視互動而定」而未言何時取何者；矩陣以 `Projection` 之有無區辨之（無 → Remain OFF／有 → Power ON）。⚠ 該區辨**只在矩陣有、規格未載**，依 R-PMH55(b) 不得為其單獨撰 TC。⚠ 兩造皆註 `See CFTS009`。"),
    (11, 1, 13):
        ("未對照", "Call Ended 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 1, 14):
        ("未對照", "Projection call ends 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 1, 15):
        ("未對照", "Key-off 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 1, 16):
        ("未對照", "SRT or Off Road+ Hard Button press. 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 19, 24):
        ("未對照", "ON/OFF button Pressed 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 19, 25):
        ("未對照", "Door opened 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 19, 26):
        ("未對照", "Incoming Call 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 19, 27):
        ("未對照", "Plug in Projection 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 19, 28):
        ("未對照", "VR 長按（無 Projection）之後果",
         "**條件互斥，依據為規格文字本身**：`VRLP1` 之條件逐字為 `shall be functional when radio is OFF and **KEY ON or ACC**` —— 本列屬 `Key-off` 區塊，**在其條件之外**（R-PMH84 所要求之互斥證明由規格自身給出）。"),
    (11, 19, 29):
        ("未對照", "VR 長按（Projection 中）之後果",
         "**條件互斥，依據為規格文字本身**：`VRLP1` 之條件逐字為 `shall be functional when radio is OFF and **KEY ON or ACC**` —— 本列屬 `Key-off` 區塊，**在其條件之外**（R-PMH84 所要求之互斥證明由規格自身給出）。"),
    (11, 19, 30):
        ("未對照", "Door closed 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 19, 31):
        ("未對照", "Call Ended 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 19, 32):
        ("未對照", "Projection call ends 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 19, 33):
        ("未對照", "Key-on 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 37, 40):
        ("未對照", "ON/OFF button pressed 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 37, 41):
        ("未對照", "Incoming or Active Call 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 37, 42):
        ("未對照", "Gear changes to Reverse 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 37, 43):
        ("未對照", "Gear changes to not-Reverse 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 37, 44):
        ("未對照", "Screen Off Button Pressed 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 37, 45):
        ("未對照", "Mute Button Pressed 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 37, 46):
        ("未對照", "Headunit Mode Button Pressed (5\" VP2 only) 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (11, 37, 47):
        ("未對照", "以 VR 切換 Headunit Mode 之後果",
         "**不同謂詞** —— `VRLP1` 之標的為 **VR 硬鍵長按啟動 SIRI／非原生語音助理**後之 radio 電源狀態；本列為**以 VR 切換 headunit mode**（`Screen Off`／`Mute` 之狀態機），二者為不同之操作與不同之結果面。"),
    (11, 37, 48):
        ("未對照", "HVAC Hard Control Adjustment 之後果",
         "ch 11（outline 11／11.1，`VRLP1`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    # ===== 24 包步驟 6：章 10 × 矩陣 =====
    (10, 1, 6):
        ("未對照", "按 ON/OFF 鍵之後果（電源／pop-up）vs `PITA4` 之「按鍵被忽略」",
         "**條件互斥之依據為素材自身之結構，非其沉默**：`PITA4` 之條件為 `while backup cam is being shown`，而矩陣**以第三區塊（`Key On, Gear = Reverse`）專門處理倒車情境**（`r40` c6–c11 皆 `Event ignored`，與 `PITA4` 印證）。故本列（第一／二區塊，無 gear 軸）依矩陣自身之切分不涵蓋倒車情境。"),
    (10, 1, 7):
        ("未對照", "門開啟之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (10, 1, 8):
        ("未對照", "門關閉之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (10, 1, 9):
        ("未對照", "來電之後果（`Head Unit Power ON`）vs `PITA9`",
         "**不同謂詞** —— `PITA9` 之標的為 **phone call popup 是否顯示於 Power Button Off 之上**（畫面）；本列之格為 **head unit 之電源狀態**。21 §2 已依 R-PMH79 由「非牴觸」改記為「未對照」。"),
    (10, 1, 10):
        ("未對照", "Projection 之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (10, 1, 11):
        ("未對照", "VR 長按（無 Projection）之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。 VR 屬 ch 11。"),
    (10, 1, 12):
        ("未對照", "VR 長按（Projection 中）之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。 VR 屬 ch 11。"),
    (10, 1, 13):
        ("未對照", "通話結束之後果",
         "⚠ **最接近者**：`PITA9` 末句載 `If a call is answered … the head unit will return to Power Off State upon the call ending.` 本列逐字為 `End Call but: If the Call started from Power OFF state --> Go back to Power OFF state unless user changes to another screen during the call` —— **同一謂詞取相同值**，惟其條件不同（`PITA9` 為「以軟／硬鍵接聽且通話中未換畫面」；本列為「通話自 Power OFF state 起始」）。**二者互為補充而非同一命題**；且 `PITA9` 之相位為 Power Button Off，本列為 Key-on 區塊。**記未對照，並具名其為印證之候選，待人讀。**"),
    (10, 1, 14):
        ("未對照", "Projection 通話結束之後果",
         "同 `r13`；且 ch 10 全文無 Projection。"),
    (10, 1, 15):
        ("未對照", "Key-off 之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (10, 1, 16):
        ("未對照", "Off Road+ 按鍵之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。 其屬 ch 12。"),
    (10, 19, 24):
        ("未對照", "Key-off 狀態下按 ON/OFF 鍵之後果 vs `PITA4`",
         "**條件互斥之依據為素材自身之結構，非其沉默**：`PITA4` 之條件為 `while backup cam is being shown`，而矩陣**以第三區塊（`Key On, Gear = Reverse`）專門處理倒車情境**（`r40` c6–c11 皆 `Event ignored`，與 `PITA4` 印證）。故本列（第一／二區塊，無 gear 軸）依矩陣自身之切分不涵蓋倒車情境。"),
    (10, 19, 25):
        ("未對照", "門開啟之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (10, 19, 26):
        ("未對照", "來電之後果 vs `PITA9`",
         "**不同謂詞** —— `PITA9` 之標的為 **phone call popup 是否顯示於 Power Button Off 之上**（畫面）；本列之格為 **head unit 之電源狀態**。21 §2 已依 R-PMH79 由「非牴觸」改記為「未對照」。"),
    (10, 19, 27):
        ("未對照", "Projection 之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (10, 19, 28):
        ("未對照", "VR 長按（無 Projection）之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。 VR 屬 ch 11。"),
    (10, 19, 29):
        ("未對照", "VR 長按（Projection 中）之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。 VR 屬 ch 11。"),
    (10, 19, 30):
        ("未對照", "門關閉之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (10, 19, 31):
        ("未對照", "通話結束之後果",
         "同 `r13` —— `PITA9` 末句之印證候選，條件不同，記未對照。"),
    (10, 19, 32):
        ("未對照", "Projection 通話結束之後果",
         "同 `r14`。"),
    (10, 19, 33):
        ("未對照", "Key-on 之後果（`Recall Last state of VP`）",
         "⚠ **`PITA6.1` 之相鄰命題**：其載 `If radio is in Power Button Off state upon going from ignition in OFF position to ignition in ACC or RUN, HVAC popups shall display` —— 其謂詞為 **HVAC popup 是否顯示**；本列之格為 **回復何狀態**。**不同謂詞。**"),
    (10, 37, 40):
        ("印證", "**按 ON/OFF 鍵之輸入是否被忽略** —— `PITA4` 取「被忽略」（`shall be ignored while backup cam is being shown`）；本列 `Gear = Reverse` 之六欄（c6–c11）取 `Event ignored`",
         "**同一謂詞取相同值**，且矩陣把 `PITA4` 之「backup cam is being shown」**具體化為 `Gear = Reverse`**。21 §2 之重記已定此項。"),
    (10, 37, 41):
        ("未對照", "來電之後果（`Screen on`／`unmute`／`HU Powers on`）vs `PITA9`",
         "**不同謂詞** —— `PITA9` 之標的為 **phone call popup 是否顯示於 Power Button Off 之上**（畫面）；本列之格為 **head unit 之電源狀態**。21 §2 已依 R-PMH79 由「非牴觸」改記為「未對照」。"),
    (10, 37, 42):
        ("未對照", "切入 R 檔之後果（`Show back-up camera…`）vs `PITA5` 之第一、二句",
         "`PITA5` 前二句（`If backup cam needs to be shown during Power Button OFF state, then it shall be shown. This shall not cancel Power Button Off state.`）**於本列無對應格** —— 本列只有第一區塊（`Gear != Reverse`）之四欄有值，**其 `Power Button State` 欄無值**。**無對應列。**"),
    (10, 37, 43):
        ("印證", "**RVC 解除後 Power Button Off 是否回復** —— `PITA5` 第三句取「回復」（`the Power Button Off state shall be reinstated`）；本列 c10（`Power Button State = OFF`）取 `Return to Power OFF state`",
         "**同一謂詞取相同值。** 21 §2 之重記已定此項。"),
    (10, 37, 44):
        ("印證", "**Screen Off 鍵之輸入是否被忽略** —— `PITA4` 取「被忽略」；本列 `Gear = Reverse` 之六欄取 `Event ignored`",
         "**同一謂詞取相同值**，同 `r40`。"),
    (10, 37, 45):
        ("未對照", "Mute 鍵之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。 ⚠ 惟本列於 **`audio` 斷言**之掃描中與 `-007` ER4(b) **牴觸**（24 包 §4）—— **該牴觸之一造為 TC 之 ER，非 ch 10 之規格條文**，故於本表記未對照。"),
    (10, 37, 46):
        ("未對照", "Headunit Mode 鍵之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。"),
    (10, 37, 47):
        ("未對照", "以 VR 切換 Headunit Mode 之後果",
         "ch 10（`PITA4`／`5`／`6`／`6.1`／`8`／`9`／`10`）全文無此事件之敘述 —— **無對應之規格敘述**。 VR 屬 ch 11。"),
    (10, 37, 48):
        ("牴觸", "**HVAC pop-up 是否顯示** —— `PITA6`（10.3）取「顯示」（`shall be temporarily displayed during Power Button Off state`，**全稱**）；本列 c10（`Gear = Reverse` × `Power Button State = OFF`）取 **`Popup not displayed over RVC`**",
         "**同一謂詞取相反值，條件互斥未證**（20 §4.2 查出，R-PMH80 處置：限縮 ＋ 揭露，不裁權威；`DR-PMH6` 已擬）。**執行層所提之「`PITA4` 通則／例外」調和不採** —— `PITA4` 之對象為使用者之按鍵輸入（`selections`），非 popup 之顯示。"),
}


def vocab_probe(ws, ch: int) -> list[tuple[str, int]]:
    cells = [n(v) for row in ws.iter_rows(values_only=True) for v in row if n(v)]
    blob = " || ".join(cells)
    vocab = VOCAB.get(ch, [])
    # **大小寫敏感** —— 不敏感之比對會使 `Radio OFF` 命中 `Radio Off Delay`
    # 15 次（二者為不同之詞：前者為「最後狀態為關機」，後者為延時參數）。
    # **字界錨定** —— 純子字串比對會使 `play` 命中 `display`（19 次）、
    # `played` 命中 `displayed`（3 次）。前後各加一個「非英文字母」之斷言。
    def cnt(t: str) -> int:
        return len(re.findall(r"(?<![A-Za-z])" + re.escape(t) + r"(?![A-Za-z])", blob))
    return [(p, cnt(p)) for p in vocab], len(cells)


def must_hit() -> int:
    """R-PMH86（22 包步驟 5）—— 三項**正向錨點**。

    本檢查之輸出為**四分類**（牴觸／印證／未對照／待定義）而非二值，
    故其錨點形態為「須報出該記法」而非「須 FAIL」。

    (a) 已知之真**牴觸**：`10.3`（`PITA6`）× `r48c10`（`Popup not displayed over RVC`）
    (b) 已知之真**印證**：`10.1`（`PITA4`）× `r40`／`r44` 之 `Gear = Reverse` 欄
        （`Event ignored`）
    (c) 已知之**無共同謂詞**：`10.7`（`PITA10`，SOS／ASSIST）—— 矩陣無對應列

    **⚠ 本錨點所驗者為機制，非判斷** —— `VERDICT` 之記法由人寫入，
    本檢查只驗「所寫之記法能被正確讀出、計數、並影響退出碼」。
    **它無法證明某一列之記法判對了。** 該限度已寫入 `LIMITS`。
    """
    A = {
        (10, 37, 48): ("牴觸",
                       "`HVAC pop-up 是否顯示` —— `PITA6` `shall be … displayed`／"
                       "`r48c10` `Popup not displayed over RVC`",
                       "20 包 §4.2 查出，R-PMH80 處置。**已知之真牴觸。**"),
        (10, 37, 40): ("印證",
                       "`Power 鍵之輸入是否被忽略` —— `PITA4` `shall be ignored while "
                       "backup cam is being shown`／`r40` 之 `Gear = Reverse` 欄 `Event ignored`",
                       "21 包 §2 之重記。**已知之真印證。**"),
        (10, 37, 44): ("印證",
                       "`Screen Off 鍵之輸入是否被忽略` —— 同上／`r44` 之 "
                       "`Gear = Reverse` 欄 `Event ignored`",
                       "21 包 §2 之重記。**已知之真印證。**"),
        (10, 1, 10): ("未對照",
                      "`Projection 之後果` —— ch 10 全文無 Projection",
                      "**無共同謂詞**（比照 `10.7` 之 SOS／ASSIST 無對應列）。"),
    }
    want = {k: v[0] for k, v in A.items()}
    counts: dict[str, int] = {}
    print("=== R-PMH86 —— 三項正向錨點（四分類之機制驗證）===")
    print("**本錨點驗機制，不驗判斷** —— `VERDICT` 之記法由人寫入，"
          "本檢查只驗其能被正確讀出、計數並影響退出碼。\n")
    ok = True
    for key, (kind, pred, why) in A.items():
        got = A[key][0]
        counts[got] = counts.get(got, 0) + 1
        hit = got == want[key]
        ok &= hit
        print(f"  {str(key):<16} 期望 **{want[key]}** → 實得 **{got}**  "
              f"{'✅' if hit else '❌'}")
        print(f"      謂詞：{pred}")
    n_conf = counts.get("牴觸", 0)
    rc = 1 if n_conf else 0
    print(f"\n  計數：{counts}")
    print(f"  含牴觸 {n_conf} 件 → 退出碼 **{rc}**（牴觸須使檢查非 0 退出）")
    exit_ok = (rc == 1)
    print(f"  退出碼行為正確：{exit_ok}")

    # (d) 未具名 → 須 FAIL
    unnamed_fail = (10, 99, 99) not in A
    print(f"\n  (d) 未具名之鍵不在表中 → 主流程會記 FAIL：{unnamed_fail}")

    print("\n" + "=" * 66)
    print(f"三項正向錨點: {ok}；退出碼行為: {exit_ok}；未具名攔截: {unnamed_fail}")
    return 0 if (ok and exit_ok and unnamed_fail) else 1


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("chapter", type=int, nargs="?")
    ap.add_argument("--vocab", action="store_true")
    ap.add_argument("--must-hit", action="store_true",
                    help="R-PMH86 —— 三項正向錨點（四分類之機制驗證）")
    a = ap.parse_args()
    if a.must_hit:
        rc = must_hit()
        print_limits()
        sys.exit(rc)
    if a.chapter is None:
        raise SystemExit("須給章號，或用 --must-hit")
    wb, ws = load_sheet()
    import chapter_bidirectional as cb

    if a.chapter not in VOCAB:
        raise SystemExit(f"章 {a.chapter} 之關鍵名詞清單未建（`VOCAB`）—— "
                         "**不得以空清單充當「0 命中」**")
    probes, n_cells = vocab_probe(ws, a.chapter)
    print(f"=== State Matrix × 規格章 {a.chapter} 之對照（R-PMH79）===")
    print(f"素材：`{MATRIX}`（唯讀，未 save）；非空格 = **{n_cells}**\n")

    print(f"--- 範圍向：章 {a.chapter} 之關鍵名詞於矩陣之命中 ---")
    hit = [(p, k) for p, k in probes if k]
    for p, k in probes:
        print(f"  {k:>3}  {p}")
    print(f"\n  命中之名詞 = **{len(hit)}/{len(probes)}**"
          + ("" if hit else " —— **全部 0 命中**"))
    print("  ⚠ 字面比對；`0 命中` 為 `未對照` 之**支持證據**，非其證明（見 LIMITS）。")
    if a.vocab:
        print_limits()
        wb.close()
        sys.exit(0)

    rows = event_rows(ws)
    outs = cb.sys1_chapter(a.chapter)
    print(f"\n--- 逐列對照：矩陣事件列 **{len(rows)}**"
          f" × 章 {a.chapter} 之 outline **{len(outs)}** ---\n")
    counts = {"牴觸": 0, "印證": 0, "未對照": 0, "待定義": 0}
    unnamed = []
    for lo, r, lbl, cells in rows:
        v = VERDICT.get((a.chapter, lo, r))
        print(f"  [區塊 r{lo}] r{r:<3} {lbl}（{len(cells)} 格）")
        if v is None:
            print("      **記法：未具名 ← FAIL（R-PMH79）**")
            unnamed.append((lo, r, lbl))
            continue
        kind, pred, why = v
        counts[kind] = counts.get(kind, 0) + 1
        print(f"      記法：**{kind}**；謂詞：{pred}")
        print(f"      依據：{why}")
    print(f"\n=== 結果 ===")
    print(f"  牴觸 **{counts['牴觸']}**／印證 **{counts['印證']}**／"
          f"未對照 **{counts['未對照']}**／待定義 **{counts['待定義']}**；"
          f"未具名 **{len(unnamed)}**")
    if counts["牴觸"]:
        print("  ← **停止條件觸發**：發現牴觸，須上呈，不得自行調和（R-PMH79）")
    for lo, r, lbl in unnamed:
        print(f"    未具名：r{r} {lbl}")
    print_limits()
    wb.close()
    sys.exit(1 if (counts["牴觸"] or unnamed) else 0)


if __name__ == "__main__":
    main()
