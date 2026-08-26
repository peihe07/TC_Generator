#!/usr/bin/env python3
"""本 feature 之首次寫回（38 包步驟 5／6）。

**其目標為 repo 內部之工作副本，非交付路徑** —— 交付路徑之複製屬 Pei（R-G5）。

三段，各自獨立驗證，不合併：
  §1 前置閘  —— `check_write_back` 之三項（R-PMH22），任一失敗即中止；
  §2 外科寫入 —— **只經 `backend/xlsx_surgical.surgical_save`**（R-G3：
                 **openpyxl 之 `save()` 絕不呼叫** —— 其會摧毀 x14 DV）；
  §3 讀回斷言 —— **自產出之檔案讀回**，非自記憶體。

**四項不變量**（38 §四步驟 6）於寫回前後各測一次並比對：
分頁數、DV 組數（含 x14）、`last_capacity_row`、B 欄公式。
前三項由 `verify_structure` 承載（其對 zip member 全集、classic/x14 DV 計數、
以及「差異僅限目標分頁」逐項為之）；`last_capacity_row` 與 B 欄公式由本檔另測。

用法：
    python3 scripts/write_back.py              # 乾跑（不產檔）
    python3 scripts/write_back.py --self-test  # 步驟 5(b)：故意失敗須被攔下
    python3 scripts/write_back.py --write      # 真寫回
"""
import argparse
import hashlib
import json
import sys
from pathlib import Path

import openpyxl

FEATURE = Path(__file__).resolve().parent.parent
ROOT = FEATURE.parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(FEATURE / "scripts"))
from backend.xlsx_surgical import StructureError, surgical_save  # noqa: E402
import check_write_back as cwb  # noqa: E402

SHEET = "Test Case Specification 測試用例規範"
FIRST_ROW = 10
SRC = FEATURE / "inputs" / ("FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT "
                            "STLA Test Case Specification & Result_SWQT_20260817_ext.xlsx")
SRC_SHA = "6372fb6be02f48dc3a3e091a60d2e2b3cf26d8704c27e25d79b7c9516fb825b2"
# 39 包步驟 4 —— **第二次寫回**：自**母本**重新產生，不在第一次之產出上疊改。
# 其檔名以 `rev2` 標其性質；**第一次之產出不覆寫、不刪除**（其為 38 包上繳所載之對象）。
OUT = FEATURE / "output" / ("FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT "
                            "STLA Test Case Specification & Result_SWQT_"
                            "PowerModing_20260826_writeback_rev2.xlsx")

# profile §0.1（rev C 34 欄）—— 38 §四步驟 6 所明定者
COLS = {"D": "leaf_id", "F": "tc_id", "G": None, "H": "test_set",
        "I": "test_item", "J": "pre_conditions", "K": None,
        "L": "test_procedure", "M": "expected_result",
        "N": "specification_reference", "P": "priority",
        "R": "design_method", "S": None, "AA": None, "AH": None}
CONST = {"G": "Disclaimer screen", "K": "NA", "S": "NA", "AA": "PeiPYHsu"}
# **絕不寫入**：`B` 為母本自身之編號公式（清之即毀其機制）；`Q` 之 DV 為
# priority 之 `"P0,P1,P2,P3"`，任何分鐘數皆會被 Excel 擋下（profile §3.6）；
# `T`–`Z` 留白（profile §3.8）；`D3`／`D4`／`D5` 留空（R-PMH27）。
NEVER = ["B", "C", "E", "O", "Q", "T", "U", "V", "W", "X", "Y", "Z",
         "AB", "AC", "AD", "AE", "AF", "AG"]


def sha256(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def load_final() -> list:
    """自 `generated/final/` 讀（R-PMH143 —— 其為指派後之檔）。"""
    out = []
    for p in sorted((FEATURE / "generated" / "final").glob("batch*.json")):
        d = json.loads(p.read_text(encoding="utf-8"))
        if d.get("tc_id_status") != "final":
            raise SystemExit(f"{p.name} 之 tc_id_status 非 final —— 拒絕寫回")
        out.extend(d["tcs"])
    out.sort(key=lambda t: t["tc_id"])
    return out


def invariants(path: Path) -> dict:
    """`last_capacity_row` 與 B 欄公式 —— `verify_structure` 不涵蓋者。"""
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET]
    last = max(r for r in range(FIRST_ROW, ws.max_row + 1)
               if ws.cell(r, 2).value is not None)
    b = {r: ws.cell(r, 2).value for r in (FIRST_ROW, last, (FIRST_ROW + last) // 2)}
    n = len(wb.sheetnames)
    wb.close()
    return {"sheets": n, "last_capacity_row": last, "b_formula": b}


def cell_value(tc: dict, letter: str) -> str:
    if letter in CONST:
        return CONST[letter]
    if letter == "AH":
        return tc.get("blocked_reason", "") or ""
    return str(tc[COLS[letter]])


def build(tcs: list):
    wb = openpyxl.load_workbook(SRC)
    ws = wb[SHEET]
    for i, tc in enumerate(tcs):
        r = FIRST_ROW + i
        for letter in COLS:
            ws[f"{letter}{r}"] = cell_value(tc, letter)
    return wb


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--self-test", action="store_true")
    a = ap.parse_args()

    got = sha256(SRC)
    if got != SRC_SHA:
        raise SystemExit(f"母本 SHA256 不符（R-PMH7）—— 期望 {SRC_SHA}，實測 {got}")
    print(f"母本 SHA256 == R-PMH7 ✅  {got[:16]}…")

    tcs = load_final()
    print(f"待寫回 = **{len(tcs)}** 條（自 generated/final/）")

    if a.self_test:
        print("\n=== 步驟 5(b) —— 故意失敗須攔下寫回 ===")
        ok = True
        try:
            cwb.check_start_row_source(44, cwb.load_cfg(FEATURE), FEATURE)
            print("  (b) 起始列取自 outline_map 之 44 → **未被攔下** ❌"); ok = False
        except cwb.CheckFailed as e:
            print(f"  (b) 起始列 44 → **被攔下** ✅  {str(e)[:78]}…")
        prov = FEATURE / "generated" / "batch01.json"
        try:
            cwb.check_tc_id_not_provisional(prov)
            print("  (d) provisional 批次 → **未被攔下** ❌"); ok = False
        except cwb.CheckFailed as e:
            print(f"  (d) provisional → **被攔下** ✅  {str(e)[:78]}…")
        try:
            cwb.check_row_count_delta(0, 50, len(tcs))
            print("  (c) 列數差 → **未被攔下** ❌"); ok = False
        except cwb.CheckFailed as e:
            print(f"  (c) 列數 0+{len(tcs)} != 50 → **被攔下** ✅  {str(e)[:60]}…")
        print(f"\n三項故意失敗全被攔下：{ok}")
        return 0 if ok else 1

    print("\n=== §1 前置閘（R-PMH22，check_write_back 之三項）===")
    for line in cwb.run(FEATURE, len(tcs)):
        print(f"  {line}")
    for p in sorted((FEATURE / "generated" / "final").glob("batch*.json")):
        print(f"  {cwb.check_tc_id_not_provisional(p)}  ({p.name})")

    before = invariants(SRC)
    print(f"\n寫回前不變量：{before}")
    if not a.write:
        print("\n（乾跑 —— 未產檔。加 --write 方寫入。）")
        return 0

    print("\n=== §2 外科寫入（surgical_save；openpyxl.save 未被呼叫）===")
    wb = build(tcs)
    report = surgical_save(wb, SRC, OUT)
    print(f"  patched sheets = {report['sheets_patched']}")
    print(f"  zip members = {report['members']}；differing = {report['differing']}")

    print("\n=== §3 讀回斷言（自產出之檔案）===")
    after = invariants(OUT)
    print(f"寫回後不變量：{after}")
    same = all(before[k] == after[k] for k in before)
    print(f"四項不變量之前三項（分頁數／last_capacity_row／B 欄公式）相同：{same}")
    wb2 = openpyxl.load_workbook(OUT)
    ws2 = wb2[SHEET]
    n = sum(1 for r in range(FIRST_ROW, FIRST_ROW + len(tcs) + 5)
            if ws2.cell(r, 4).value)
    ids = [ws2.cell(FIRST_ROW + i, 6).value for i in range(len(tcs))]
    blank_q = all(ws2.cell(FIRST_ROW + i, 17).value in (None, "") for i in range(len(tcs)))
    d345 = [ws2[f"D{r}"].value for r in (3, 4, 5)]
    wb2.close()
    print(f"  D 欄非空列 = {n}（應為 {len(tcs)}）")
    print(f"  F 欄 tc_id 首末 = {ids[0]} … {ids[-1]}；相異 = {len(set(ids))}")
    print(f"  Q 欄留白 = {blank_q}；D3/D4/D5 = {d345}")
    print(f"\n產出：{OUT.relative_to(ROOT)}\n  SHA256 = {sha256(OUT)}")
    ok = same and n == len(tcs) and len(set(ids)) == len(tcs) and blank_q \
        and all(v is None for v in d345)
    print(f"\n**寫回驗證：{'全部通過 ✅' if ok else '❌ 有未通過項'}**")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
