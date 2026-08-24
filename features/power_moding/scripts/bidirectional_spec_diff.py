#!/usr/bin/env python3
"""R-PMH51 —— PDF ↔ SYS1 匯出之**雙向**句級比對。

01 包只驗方向一（SYS1→PDF），**看不見漏句** —— 漏句不會顯示為「不符」，
只顯示為「沒有這一則」。本檔補方向二。

  方向一  SYS1 之每一句是否出現於 PDF   —— 抓「SYS1 多出／改寫」
  方向二  PDF 之每一句是否出現於 SYS1   —— **抓漏句**（本輪新增）

比對單位為**句**（以句號後之空白切分），正規化：去 `_x000D_`、
摺疊空白、統一彎引號與省略號。命中判定為「正規化後之句為對方全文之子字串」。

輸出：docs/reports/bidirectional_spec_diff.md
"""
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PDF_TXT = ROOT / "sandbox" / "spec.txt"
SYS1 = ROOT / "inputs" / "SYS1_HMI_Power_Moding_HMI_Logic_and_Flow_R1_SR24_2A.xlsx"

# 句子最短長度 —— 短於此者為標題／片語，不作句級比對（另計）
MIN = 25


def norm(t: str) -> str:
    t = str(t).replace("_x000D_", " ")
    t = (t.replace("‘", "'").replace("’", "'")
          .replace("“", '"').replace("”", '"')
          .replace("…", "...").replace("–", "-").replace("—", "-"))
    return re.sub(r"\s+", " ", t).strip()


def sentences(t: str) -> list[str]:
    t = norm(t)
    return [s.strip() for s in re.split(r"(?<=[.!?])\s+", t) if s.strip()]


def main() -> None:
    pdf_pages = PDF_TXT.read_text(errors="replace").split(chr(12))[:11]
    pdf_norm = [norm(p) for p in pdf_pages]
    pdf_all = norm(" ".join(pdf_pages))

    ws = openpyxl.load_workbook(SYS1, data_only=True)["Basic Report"]
    rows = {}
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value:
            rows[str(ws.cell(r, 3).value).strip()] = str(ws.cell(r, 4).value or "")
    sys_all = norm(" ".join(rows.values()))

    key = lambda x: [int(p) for p in x.split(".")]
    L = ["# 雙向規格比對 —— PDF ↔ SYS1 匯出（R-PMH51）", "",
         f"- 產出：`scripts/bidirectional_spec_diff.py`",
         f"- PDF：11 頁，正規化後 {len(pdf_all):,} 字元",
         f"- SYS1 `Basic Report`：{len(rows)} 則，正規化後 {len(sys_all):,} 字元",
         f"- 句長門檻：{MIN} 字元（短於此者不作句級比對，另計）", "",
         "## 方向一 —— SYS1 之每一句是否出現於 PDF（01 包已做，本輪複算）", "",
         "| outline | 句數 | 命中 | 未命中 |", "|---|---:|---:|---:|"]
    d1_miss = {}
    for o in sorted(rows, key=key):
        ss = [s for s in sentences(rows[o]) if len(s) >= MIN]
        if not ss:
            continue
        miss = [s for s in ss if s not in pdf_all]
        if miss:
            d1_miss[o] = miss
        L.append(f"| {o} | {len(ss)} | {len(ss)-len(miss)} | **{len(miss)}** |"
                 if miss else f"| {o} | {len(ss)} | {len(ss)} | 0 |")
    L += ["", f"**方向一未命中之 outline：{sorted(d1_miss, key=key) or '無'}**", ""]

    # --- 方向二：PDF → SYS1（本輪新增，抓漏句）---
    L += ["## 方向二 —— **PDF 之每一句是否出現於 SYS1**（本輪新增，抓漏句）", "",
          "| PDF 頁 | 句數 | 命中 | **未命中（漏句候選）** |", "|---|---:|---:|---:|"]
    d2 = {}
    for i, p in enumerate(pdf_norm, 1):
        ss = [s for s in sentences(p) if len(s) >= MIN]
        if not ss:
            continue
        miss = [s for s in ss if s not in sys_all]
        if miss:
            d2[i] = miss
        L.append(f"| p{i} | {len(ss)} | {len(ss)-len(miss)} | **{len(miss)}** |")
    tot2 = sum(len(v) for v in d2.values())
    L += ["", f"**方向二未命中合計：{tot2} 句**（分布於 p{sorted(d2)}）", ""]

    L += ["## 方向二之未命中逐句（漏句候選）", ""]
    for pg in sorted(d2):
        L.append(f"### p{pg}")
        L.append("")
        for s in d2[pg]:
            L.append(f"- `{s}`")
        L.append("")

    out = ROOT / "docs" / "reports" / "bidirectional_spec_diff.md"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(L) + "\n", encoding="utf-8")
    print(f"方向一未命中 outline：{sorted(d1_miss, key=key) or '無'}")
    print(f"方向二未命中：{tot2} 句，分布 p{sorted(d2)}")
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
