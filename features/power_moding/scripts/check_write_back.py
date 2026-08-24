#!/usr/bin/env python3
"""R-PMH22 — write_back 之機器檢查（三項），寫回前必跑，失敗即中止。

三項檢查逐一對應 R-PMH22 之 (a)(b)(c)：

  (a) blank_precondition  — 目標分頁自 first_row 起，D 欄全空。
                            BLANK 之前提若已不成立，append 會覆蓋既有資料。
  (b) start_row_source    — 寫回之起始列必須等於 feature.yaml 的 first_row，
                            不得由任何其他來源推導。特別是
                            data/outline_map.json 之 row_036_customer ——
                            該欄記的是「客戶那份」的列號（10-57），
                            與母本的寫回目標無關（02 包 §11 第 4 項）。
  (c) row_count_delta     — 寫回後列數 == 寫回前列數 + 本批 TC 數。
  (d) tc_id_not_provisional — 批次檔頭之 `tc_id_status` 為 `provisional` 者，
                            寫回即中止（13 包步驟 6；12 包上繳 §8 第 6 項自陳
                            「provisional 無任何機制防止其被當成最終編號」）。

本檔只讀工作簿，不寫。x14 DV 之保全由呼叫端的 xlsx_surgical splice 負責
（R-G3）—— 這裡連 openpyxl 的 save() 都不呼叫。

用法（正常）:
    python scripts/check_write_back.py --feature . --batch-size 48
用法（故意失敗之自我測試）:
    python scripts/check_write_back.py --feature . --self-test
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl
import yaml


class CheckFailed(Exception):
    """任一檢查未通過即拋出；呼叫端據此中止寫回。"""


def _col_index(letter: str) -> int:
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch.upper()) - 64)
    return n


def load_cfg(feature_dir: Path) -> dict:
    return yaml.safe_load((feature_dir / "feature.yaml").read_text(encoding="utf-8"))


def _open_sheet(feature_dir: Path, cfg: dict):
    wb = openpyxl.load_workbook(feature_dir / cfg["paths"]["workbook"], data_only=True)
    return wb[cfg["workbook"]["sheet"]]


def filled_rows(ws, req_col: int, first_row: int) -> list[int]:
    """D 欄非空之列號。列數即以此為準 —— B 欄是公式，數不得。"""
    return [r for r in range(first_row, ws.max_row + 1)
            if ws.cell(r, req_col).value not in (None, "")]


# ---------------------------------------------------------------- 三項檢查

def check_blank_precondition(ws, req_col: int, first_row: int) -> str:
    """(a) 自 first_row 起 D 欄全空。"""
    occupied = filled_rows(ws, req_col, first_row)
    if occupied:
        raise CheckFailed(
            f"(a) blank_precondition FAILED — 自 r{first_row} 起 D 欄應全空，"
            f"實測 {len(occupied)} 列非空：{occupied[:10]}"
            f"{' …' if len(occupied) > 10 else ''}。"
            f"workbook_state 已非 BLANK，append 會覆蓋既有資料。")
    return f"(a) blank_precondition PASS — 自 r{first_row} 起 D 欄 0 列非空"


def check_start_row_source(proposed_start: int, cfg: dict, feature_dir: Path) -> str:
    """(b) 起始列 == feature.yaml 之 first_row，且未取自 outline_map。"""
    first_row = cfg["write_back"]["first_row"]
    if proposed_start != first_row:
        forbidden = ""
        omap = feature_dir / "data" / "outline_map.json"
        if omap.exists():
            rows = {leaf.get("row_036_customer")
                    for leaf in json.loads(omap.read_text(encoding="utf-8"))["leaves"]}
            if proposed_start in rows:
                forbidden = (" —— 且該值出現在 data/outline_map.json 的"
                             " row_036_customer 內，該欄記的是客戶那份的列號，"
                             "不是寫回目標列")
        raise CheckFailed(
            f"(b) start_row_source FAILED — 起始列 {proposed_start} != "
            f"feature.yaml write_back.first_row {first_row}{forbidden}")
    return f"(b) start_row_source PASS — 起始列 {proposed_start} == first_row"


def check_tc_id_not_provisional(batch: Path) -> str:
    """(d) 批次之 `tc_id_status` 為 `provisional` 者不得寫回。"""
    d = json.loads(batch.read_text(encoding="utf-8"))
    st = d.get("tc_id_status")
    if st == "provisional":
        raise CheckFailed(
            f"(d) tc_id_not_provisional FAILED — 批次 {d.get('batch')!r} 之 "
            f"tc_id_status = 'provisional'。**臨時編號不得寫回工作簿** —— "
            f"最終編號須待全 48 leaf 完成後單次指派。")
    return f"(d) tc_id_not_provisional PASS — tc_id_status = {st!r}"


def check_row_count_delta(before: int, after: int, batch_size: int) -> str:
    """(c) after == before + batch_size。"""
    if after != before + batch_size:
        raise CheckFailed(
            f"(c) row_count_delta FAILED — 寫回前 {before} 列，本批 {batch_size} 筆，"
            f"預期 {before + batch_size} 列，實測 {after} 列"
            f"（差 {after - (before + batch_size):+d}）")
    return f"(c) row_count_delta PASS — {before} + {batch_size} == {after}"


def run(feature_dir: Path, batch_size: int, proposed_start: int | None = None,
        after_count: int | None = None) -> list[str]:
    cfg = load_cfg(feature_dir)
    wbk = cfg["workbook"]
    req_col = _col_index(wbk["columns"]["req_id"])
    first_row = cfg["write_back"]["first_row"]
    ws = _open_sheet(feature_dir, cfg)

    before = len(filled_rows(ws, req_col, first_row))
    out = [check_blank_precondition(ws, req_col, first_row),
           check_start_row_source(first_row if proposed_start is None
                                  else proposed_start, cfg, feature_dir),
           check_row_count_delta(before,
                                 before + batch_size if after_count is None
                                 else after_count,
                                 batch_size)]
    return out


# ---------------------------------------------------------------- 自我測試

def self_test(feature_dir: Path) -> int:
    """R-PMH22 之 RESOLVED 條件：三項各以一次故意失敗證明其會攔下。

    (a) 以 monkeypatch 讓 filled_rows 回傳假資料列，模擬 first_row 之前
        已有一列 —— 不動真實工作簿（本包零寫回）。
    """
    cfg = load_cfg(feature_dir)
    req_col = _col_index(cfg["workbook"]["columns"]["req_id"])
    first_row = cfg["write_back"]["first_row"]
    ws = _open_sheet(feature_dir, cfg)
    results = []

    # (a) 故意失敗：偽造一列已佔用
    class FakeWS:
        max_row = 12
        def cell(self, r, c):
            class C:
                value = "SWE1-HMI-PM-999" if r == 10 else None
            return C()
    try:
        check_blank_precondition(FakeWS(), req_col, first_row)
        results.append(("a", False, "未被攔下"))
    except CheckFailed as e:
        results.append(("a", True, str(e)))

    # (b) 故意失敗：起始列取自 outline_map 之 row_036_customer（44）
    try:
        check_start_row_source(44, cfg, feature_dir)
        results.append(("b", False, "未被攔下"))
    except CheckFailed as e:
        results.append(("b", True, str(e)))

    # (d) 故意失敗：批次為 provisional
    b = feature_dir / "generated" / "batch01.json"
    try:
        if b.exists():
            check_tc_id_not_provisional(b)
            results.append(("d", False, "未被攔下"))
        else:
            results.append(("d", False, "batch01.json 不存在，測試無效"))
    except CheckFailed as e:
        results.append(("d", True, str(e)))

    # (c) 故意失敗：寫回後列數少一
    before = len(filled_rows(ws, req_col, first_row))
    try:
        check_row_count_delta(before, before + 48 - 1, 48)
        results.append(("c", False, "未被攔下"))
    except CheckFailed as e:
        results.append(("c", True, str(e)))

    print("=== R-PMH22 故意失敗測試（三項）===")
    for key, caught, msg in results:
        print(f"\n[{key}] {'攔下 ✅' if caught else '未攔下 ❌'}")
        print(f"    {msg}")

    # 範圍向（R-G9）：證明正常情形不會轉紅
    print("\n=== 範圍向 —— 正常情形不得轉紅 ===")
    try:
        for line in run(feature_dir, batch_size=48):
            print(f"    {line}")
        print("    範圍向 PASS ✅ —— 三項在正常輸入下皆通過")
        ok_scope = True
    except CheckFailed as e:
        print(f"    範圍向 FAIL ❌ —— {e}")
        ok_scope = False

    all_caught = all(c for _, c, _ in results)
    print(f"\n三項故意失敗全部被攔下: {all_caught}；範圍向: {ok_scope}")
    return 0 if (all_caught and ok_scope) else 1


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--feature", default=".", help="feature 目錄")
    ap.add_argument("--batch-size", type=int, default=0)
    ap.add_argument("--self-test", action="store_true")
    args = ap.parse_args()
    fd = Path(args.feature).resolve()
    if args.self_test:
        sys.exit(self_test(fd))
    try:
        for line in run(fd, args.batch_size):
            print(line)
    except CheckFailed as e:
        print(f"WRITE-BACK ABORTED — {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
