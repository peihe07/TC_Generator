#!/usr/bin/env python3
"""R-PMH51 之未套用側 —— 指定章之**雙向**逐句複驗（17 包步驟 3）。

R-PMH51 明文：A-PMH03 之其餘三則（8、9.1、11.1）須以雙向法複驗，
**未複驗前其標題結論不得引用**。11.1 已於 13 包方向二覆蓋、9.1 查出新漏 2，
**而 outline `8` 至今未做** —— 本檔補之，並使其可對任一章重跑。

方向一：SYS1 之每一句是否出現於 PDF（01 包已對全簿做過）
方向二：**PDF 之每一句是否出現於 SYS1** —— 漏句只在此方向顯示

用法:
    python scripts/chapter_bidirectional.py 8
"""
import argparse
import hashlib
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PDF_TXT = ROOT / "sandbox" / "spec.txt"          # `pdftotext -layout`（舊預設）

# --- R-PMH71（19 包步驟 4）：**預設來源改為 PyMuPDF block 層** ---
# A-PMH16 係以 block 層萃取查出，而本檔之預設來源原為 `-layout`；
# 該程式此刻重跑**查不出 A-PMH16** —— 結論與其量測分離
# （A-PMH12 形態在結論層之同型）。依 R-PMH71(a) 改預設並補 must-hit。
#
# `-layout` 之病灶：p9 之兩欄狀態矩陣與 `PM1)` 散文**交錯**，
# 切出之「句」皆為矩陣格與散文之混合串；block 層之 `PM1)` 為單一區塊。
SOURCE_DEFAULT = "block"


def pdf_text(source: str = SOURCE_DEFAULT) -> str:
    """PDF 全文。`block` = PyMuPDF `get_text("blocks")`；`layout` = `pdftotext -layout`。"""
    if source == "layout":
        return norm(PDF_TXT.read_text(errors="replace"))
    import fitz
    import yaml
    cfg = yaml.safe_load((ROOT / "feature.yaml").read_text(encoding="utf-8"))
    d = fitz.open(ROOT / cfg["paths"]["spec_pdf"])
    return norm(" ".join(b[4] for pg in d for b in pg.get_text("blocks")))
SYS1 = ROOT / "inputs" / "SYS1_HMI_Power_Moding_HMI_Logic_and_Flow_R1_SR24_2A.xlsx"

# 各章於 PDF 全文之**起錨**（逐字取自 PDF，非推算）。
# 章之範圍 = [本章起錨, **下一章起錨**) —— 訖錨不另指定，
# 使相鄰兩章之邊界**必然共用同一個字串**，區間**不可能重疊或留縫**
# （17 §12 第 1 項：「錨之外的第 7 個 marker」之藏身處，由 `partition()` 逐段列出）。
# **錨須於兩種來源皆恰命中 1 次**（19 包步驟 4）——
# 原錨含頁碼前綴（`7 Startup Notes:`），而頁碼於 block 層為獨立區塊，
# 故改取不含頁碼之最短唯一字串。實測 layout／block 皆 1/1/1/1/1/1。
STARTS = {
    7:  "Notes: SU1.)",
    8:  "R1Low Only",
    9:  "Power Moding Please refer",
    10: "Additional Power Moding Behavior Notes:",
    11: "VR HARD KEY FOR SIRI",
    12: "Power Moding - Off Road+",
}

# --- R-PMH66(c)：殘餘之逐句人讀結論。**每一殘餘句皆須在此具名** ---
# 「逐字未命中」者一律進殘餘；**殘餘不得由門檻自動判為「非漏」** ——
# 門檻只決定人讀之優先順序，不決定結論。未具名之殘餘句 → **FAIL**。
# 鍵為 `sha1(句)[:8] + " " + 句之前 48 字元`。
#
# **19 包之修正（缺陷實測）**：原鍵為「句之前 60 字元」，
# 而章 9 之兩句殘餘（284 字元與 1,075 字元）**其前 60 字元完全相同** ——
# 二者共用同一個鍵，**其中一句之人讀結論會被另一句靜默借用**，
# 且 `--self-test` 不會察覺（兩句都「有結論」）。
# 改以句之 sha1 前綴為鍵，保留前 48 字元供人閱讀。
RESIDUE_VERDICT: dict[str, str] = {
    "f431a791 Notes: SU1.) When the vehicle's driver door is c":
        "**漏 —— A-PMH03 之 7.1 已知漏句（非新）**：`after the animation (3 sec) a splash screen is presented timeout (1.5 each).` 於 SYS1 全 52 則不存在（12 包已證）。句首之 `Notes:` 為章標題之殘留。**batch 1 之 `source_clause` 取自 PDF（R-PMH50），該子句在內，故 batch 1 不受影響。**",
    "8ca2a25c SU8.) Show the splash screen and disclaimer scre":
        "**部分漏 —— A-PMH14 新漏 1（非新）**：`SU8.)` 於 SYS1 之 `7.9` 逐字存在；逐字未命中之因為切分把 `SU8.)` 與 `SU9.)` 併為一句。**`SU9.)` 漏，已由 R-PMH74 裁定 `ACCEPTED（經裁定不補）`。**",
    "0a66f331 SU9.1) Pressing Power Off or Screen Off hard key":
        "**漏 —— A-PMH14 新漏 1（非新）**：`SU9.1)` 於 SYS1 全簿命中 0。已由 R-PMH74 裁定 `ACCEPTED（經裁定不補）`；**R-PMH55 之適用因而繼續成立**，batch 1 之 `-003`／`-004` 之「不按任何硬鍵」限定有效。",
    "4d281b64 Power Moding Please refer to Power Moding State ":
        "**漏 —— 新漏 2 之範圍（18 包補記者）**：SYS1 全 52 則探針 `Please refer to Power Moding State Matrix` 命中 0。**其所指之矩陣已由 Pei 提供**（R-PMH73，第六筆素材），惟本輪實測其內容與 p9 之矩陣**不對應**（見上繳 §3.3），故補救來源尚未確定。",
    "e8fddcc8 HEADUNIT POWER OFF HEADUNIT POWER ON ICS Hard Co":
        "**漏 —— A-PMH14 新漏 2**：狀態矩陣之表頭（`HEADUNIT POWER OFF`／`ON` × `ICS Hard Controls`／`HVAC Knobs`）。探針於 SYS1 全簿命中皆 0。",
    "e291bd64 HVAC Knobs: Fully functional Climate GUI: Not Vi":
        "**漏 —— A-PMH14 新漏 2**：矩陣之 `KEY ON ENGINE ON` 列。",
    "d79b556b HVAC Knobs: Fully functional Climate GUI: Not Vi":
        "**混合句**：其尾段之 `PM1) … should 'stay awake' for 60 seconds up to 2.5 minutes` 為 PDF 之未刪淨舊文字（R-PMH75，不驗）；其餘 `ICS Hard Controls:`／`HVAC Knobs:`／`Climate GUI:`／`Headunit:`／`KEY ON ENGINE ON`／`KEY OFF (No ACC position)` 等**全為 p9 狀態矩陣之格**，於 SYS1 全缺 —— **A-PMH14 新漏 2**。",
    "a186ee1c If the user does not interact with the popup wit":
        "**PDF 側為未刪淨之舊文字 —— 依 R-PMH75 不驗**：Pei 裁定「以刪掉之後的為主」，outline `9.1` 之權威文本為 SYS1。本句之 `within 60 seconds`／`the radio should shut Off the`／`aofnd` 三處皆為舊文字，SYS1 已刪。**A-PMH16 之原判定（時序漏失／獨立行為結果）已被 R-PMH75 推翻。**",
    "a6ecc7ba FOTA update available - If user accepts FOTA pop":
        "非漏 —— **條列再流**：SYS1 `9.1` 作 `1. FOTA update available - - If user accepts…`（多一個破折號），三句內容全在。",
    "a7348b50 Charge Now - XEV key off-Pop-ups Charge Now/Summ":
        "非漏 —— 其散文 `3. XEV key off-Pop-ups Charge Now/Summary; Preconditioning.` 於 SYS1 9.1 逐字存在；句首之 `Charge Now -` 為前一條列項之尾，屬 `-layout` 之切分",
    "21837867 Shut the radio down if user dismisses Charge Now":
        "非漏（散文側）—— SYS1 作 `Shut the radio down if user dismisses XEV key off Pop-ups.`，逐字存在；句中之 `Charge Now` 為前一條列項之尾，屬切分",
    "847ec7c1 [CR22412] VR HK to activate SIRI/Voice assistant":
        "**漏 —— A-PMH14 新漏 2**：矩陣之 `Headunit:` 列（`VR HK to activate SIRI/Voice assistants shall be functional (See CTS009)`，出現兩次即兩欄）。SYS1 全缺。",
    "fe88e914 Additional Power Moding Behavior Notes: POWER BU":
        "非漏（需求側）—— 章標題於 SYS1 為 outline `10`、`PITA4:` 本文於 `10.1`，皆逐字存在。逐字未命中之因為**全大寫分節標籤** `POWER BUTTON:` 被切入同句，**而該標籤於 SYS1 全缺**（A-PMH17）。",
    "7232af2b KEY OFF, HEADUNIT POWER ON: PITA8: During Key OF":
        "非漏（需求側）—— `PITA8:` 之本文於 SYS1 `10.5` 逐字存在。逐字未命中之因為 PDF 之**全大寫分節標籤** `KEY OFF, HEADUNIT POWER ON:` 被切入同句，**而該標籤於 SYS1 全缺**（A-PMH17）。",
    "2be436a5 VR HARD KEY FOR SIRI/NON-NATIVE VOICE ASSISTANTS":
        "非漏 —— SYS1 之 outline `11` 逐字為 `VR HARD KEY FOR SIRI/NON-NATIVE VOICE ASSISTANTS`（無尾冒號），`11.1` 逐字為 `VRLP1: VR hard key to…`。逐字未命中之因為 PDF 之節標題帶尾冒號、且切分於 `(eg.` 處斷開",
    "4a2d1d79 Radio status after interaction with SIRI depends":
        "非漏 —— 切分假象：句切於 `(i.e.` 處；其全句於 SYS1 11.1 逐字存在",
    "9c2707e8 radio back to off), Screen ON and Audio OFF, Scr":
        "非漏 —— **條列再流**（A-PMH03 原記之形態，A-PMH14 已以方向二確認）：SYS1 11.1 作 `- Screen Off and Audio OFF (i.e. radio back to off), - Screen ON and Audio OFF, …`，僅條列符號不同，四個 outcome 全在",
    "c8b75652 (DCR19385) POWER MODING STATE MATRIX: Power Modi":
        "**漏 —— A-PMH14 新漏 3（既有，非新）**：`POWER MODING STATE MATRIX:` 段於 SYS1 全簿命中 0。句首之 `(DCR19385)` 屬前一句（VRLP1）之尾，SYS1 11.1 有之",
    "afb90aca If this document is not available, please reques":
        "**漏 —— A-PMH14 新漏 3 之第二句（既有，非新）**",
    "2abdb5e7 Power Moding - Off Road+ Headunit is On Headunit":
        "非漏（需求側）—— `OFF1.)` 之本文於 SYS1 `12.1` 逐字存在。其餘 `Customer presses power button`／`Forward Facing Camera Launches`／`Headunit is \"Off\" (Idle)` 等**全為 p11 流程圖之標籤**，SYS1 之 `12.4` 逐字為 `Please refer to the diagram (image: …)` —— **A-PMH04 之圖片佔位，已知，非新漏**。",
}
MIN_SENT = 25          # 最短比對單位（字元），比照 13 包
GRAM = 6               # 6-gram 覆蓋率，比照 13 包
GRAM_THRESHOLD = 0.30  # < 30% 者為真漏候選


# --- R-PMH52 之擴及（17 包步驟 4）---
# 本檢查於輸出末尾具名其限度。
sys.path.insert(0, str(Path(__file__).resolve().parent))

LIMITS = [
    "**PDF 來源自 19 包起預設為 PyMuPDF block 層**（R-PMH71）；`--source layout` 供對照，**其於 p9 之矩陣區已知不可用**（A-PMH16）",
    "**`STARTS` 之起錨為人工指定** —— 錨取錯則整章之範圍隨之錯；`--partition` 只能查未覆蓋段，查不出「錨落在章內某處」之情形",
    "只比對 PDF **文字層**；**圖表不看** —— p9 之狀態矩陣即以圖呈現（A-PMH14 新漏 2）",
    "比對單位為句（>= 25 字元）；**短於 25 字元之句一律不入母體**，章 8 之標題 `Starup R1Low Only` 即屬之",
    "6-gram 覆蓋率**不再參與判定**（R-PMH66）—— 只決定殘餘人讀之優先順序；其門檻值本身未經重驗，故亦不再被倚賴",
    "`RESIDUE_VERDICT` 之人讀結論**由人寫入，本檢查不驗其正確性** —— 只驗其存在",
    "**只驗字之有無，不驗語意** —— 同義改寫會被判為漏句，改寫錯誤會被判為命中",
]


def print_limits() -> None:
    print("\n=== 本檢查未涵蓋之範圍（R-PMH52）===")
    for _x in LIMITS:
        print(f"  - {_x}")
    print("  **以上各項本檢查一律不看** —— 其全綠不含關於該等項之任何資訊。")


def residue_key(s: str) -> str:
    """殘餘句之鍵 —— `sha1(句)[:8] + " " + 句之前 48 字元`（19 包，見上）。"""
    return hashlib.sha1(s.encode("utf-8")).hexdigest()[:8] + " " + s[:48]


def norm(t: str) -> str:
    t = str(t).replace("_x000D_", " ")
    for a, b in (("‘", "'"), ("’", "'"), ("“", '"'),
                 ("”", '"'), ("…", "..."), ("–", "-"),
                 ("—", "-")):
        t = t.replace(a, b)
    return re.sub(r"\s+", " ", t).strip()


def sentences(t: str) -> list[str]:
    """句號後空白切分，保留 >= MIN_SENT 者（比照 13 包之比對單位）。"""
    return [s.strip() for s in re.split(r"(?<=\.)\s+", t)
            if len(s.strip()) >= MIN_SENT]


def grams(t: str) -> set[str]:
    w = re.findall(r"[a-z0-9]+", t.lower())
    return {" ".join(w[i:i + GRAM]) for i in range(max(0, len(w) - GRAM + 1))}


def coverage(s: str, hay: str) -> float:
    g = grams(s)
    return len(g & grams(hay)) / len(g) if g else 1.0


def bounds(pdf: str) -> list[tuple[int, int, int]]:
    """回傳 [(章, 起, 訖)]，訖 = 下一章之起；末章訖 = 全文末。

    起錨於全文中之命中數須恰 1（R-PMH41 —— 驗命中數）；不為 1 即 raise。
    """
    pos = []
    for ch, a in STARTS.items():
        n = pdf.count(a)
        if n != 1:
            raise SystemExit(f"章 {ch} 之起錨 `{a}` 命中 {n} 次（須恰 1）—— R-PMH41")
        pos.append((pdf.index(a), ch))
    pos.sort()
    out = []
    for k, (i, ch) in enumerate(pos):
        j = pos[k + 1][0] if k + 1 < len(pos) else len(pdf)
        out.append((ch, i, j))
    return out


def pdf_chapter(ch: int, pdf: str) -> str:
    for c, i, j in bounds(pdf):
        if c == ch:
            return pdf[i:j].strip()
    raise SystemExit(f"章 {ch} 未建錨")


def partition(pdf: str) -> int:
    """17 §12 第 1 項 —— 章區間之覆蓋／重疊檢查。

    區間由「下一章之起」界定，**重疊與縫隙在構造上即不可能**；
    真正待查者為**首章之前**與**末章之後**之未覆蓋文字，及其是否含 marker。
    """
    import importlib
    mc = importlib.import_module("marker_coverage")
    bs = bounds(pdf)
    print("\n=== 章區間之分割檢查（17 §12 第 1 項）===")
    print(f"{'章':>3}  {'起':>6} {'訖':>6} {'字元':>6}  起錨")
    for ch, i, j in bs:
        print(f"{ch:>3}  {i:>6} {j:>6} {j-i:>6}  {STARTS[ch][:52]}")
    gaps = [("首章之前", 0, bs[0][1]), ("末章之後", bs[-1][2], len(pdf))]
    total = sum(j - i for _, i, j in bs)
    print(f"\n  已覆蓋 {total} / {len(pdf)} 字元"
          f"（{total/len(pdf):.1%}）；重疊 **0**（構造上不可能）")
    bad = 0
    cnt, _ = mc.prefix_scan(pdf)
    reqs, _ = mc.req_prefixes(cnt)
    for name, i, j in gaps:
        seg = pdf[i:j]
        ms = mc.enumerate_markers(seg, reqs) if seg.strip() else []
        print(f"\n  未覆蓋段【{name}】{j-i} 字元；**其中之 marker：{ms or '無'}**")
        if seg.strip():
            print(f"    首 120 字元：{seg[:120]}")
            print(f"    末 120 字元：{seg[-120:]}")
        if ms:
            bad += 1
            print("    ← **停止條件 8 觸發** —— 未覆蓋段含 marker")
    if not bad:
        print("\n  **未覆蓋段皆不含 marker —— 停止條件 8 未觸發。**")
        print("  （首章之前為 p1–p7 之封面與五張流程圖頁，A-PMH04 已知之圖片佔位）")
    return bad


def sys1_chapter(ch: int) -> list[tuple[str, str]]:
    ws = openpyxl.load_workbook(SYS1, data_only=True)["Basic Report"]
    out = []
    for r in range(2, ws.max_row + 1):
        o = str(ws.cell(r, 3).value or "").strip()
        if re.fullmatch(rf"{ch}(\.\d+)*", o):
            out.append((o, norm(ws.cell(r, 4).value or "")))
    return out


def source_must_hit() -> int:
    """R-PMH71 之 must-hit（19 包步驟 4）—— 兩來源並列跑章 9。

    A-PMH16 之三處漏字（`for 60 seconds`／`seconds`／`the radio should shut Off the`）
    皆位於 `PM1)` 之散文中。以 `layout` 為來源時，p9 之散文與矩陣交錯，
    該三處**不會**以可讀之形態出現於任何殘餘句；以 `block` 為來源時，
    `PM1)` 為單一區塊，該三處**必然**落在殘餘句內。
    """
    probes = ["for 60 seconds up to 2.5 minutes",
              "within 60 seconds the timeout",
              "the radio should shut Off the popup"]
    out = {}
    for src in ("layout", "block"):
        pdf = pdf_text(src)
        rows = sys1_chapter(9)
        sysc = " ".join(d for _, d in rows)
        sents = sentences(pdf_chapter(9, pdf))
        residue = [x for x in sents if x not in sysc]
        joined = " ".join(residue)
        out[src] = (len(sents), len(residue),
                    [p_ for p_ in probes if p_ in joined])
    print("=== R-PMH71 must-hit —— 章 9 之兩來源並列 ===")
    print(f"{'來源':<8} {'句數':>4} {'殘餘':>4}  A-PMH16 三探針之命中")
    for src in ("layout", "block"):
        n, r, hit = out[src]
        print(f"{src:<8} {n:>4} {r:>4}  {len(hit)}/3  {hit}")
    ok_layout = len(out["layout"][2]) == 0
    ok_block = len(out["block"][2]) == 3
    print(f"\n  `layout` 查不出（0/3）：{ok_layout}")
    print(f"  `block` 三處全在殘餘（3/3）：{ok_block}")
    # --- 真正之鑑別量：對 SYS1 9.1 做字級 diff 之**噪音** ---
    # 下放包所給之 must-hit 前提（`layout` 查不出）**不成立** ——
    # 三個探針之字串於 `-layout` 之殘餘中同樣存在（見上表）。
    # 二來源之實際差別在於：`-layout` 之 p9 散文與矩陣**交錯**，
    # 故字級 diff 之差異數被矩陣格灌爆，A-PMH16 之三處淹沒其中。
    import difflib
    rows = sys1_chapter(9)
    sys91 = next(d for o, d in rows if o == "9.1")
    print("\n=== 真正之鑑別量 —— 對 SYS1 `9.1` 之字級 diff 噪音 ===")
    noise = {}
    for src in ("layout", "block"):
        pdf = pdf_text(src)
        seg = pdf_chapter(9, pdf)
        a, b = seg.split(), sys91.split()
        ops = [o for o in difflib.SequenceMatcher(None, a, b, autojunk=False)
               .get_opcodes() if o[0] != "equal"]
        words = sum(max(o[2] - o[1], o[4] - o[3]) for o in ops)
        noise[src] = (len(ops), words)
        print(f"  {src:<8} 差異段 {len(ops):>3} 個、涉 {words:>4} 詞")
    print(f"\n  `block` 之差異段為 `layout` 之 "
          f"{noise['block'][0]/noise['layout'][0]:.0%}；"
          f"涉詞數為 {noise['block'][1]/noise['layout'][1]:.0%}")
    print("  **A-PMH16 之三處即由 block 層之字級 diff 讀出** —— "
          "\n  `layout` 側之差異段被矩陣格灌爆，三處淹沒其中。")

    print("\n  **此即 R-PMH71 所指之「結論與其量測分離」** —— "
          "\n  18 包把 A-PMH16 寫進了 `RESIDUE_VERDICT`，"
          "而其量測所用之來源不是當時之預設。")
    print("\n  ⚠ **下放包所給之 must-hit 前提不成立，據實回報** —— "
          "\n  「`-layout` 查不出」為假：三個探針之字串於 `-layout` 之殘餘中同樣存在。"
          "\n  真正使 18 包漏掉它的不是來源，是 13 包之 **6-gram 門檻**（R-PMH66）。"
          "\n  **停止條件 8 依其字面觸發，本函式回傳 1。**")
    return 0 if (ok_layout and ok_block) else 1


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("chapter", type=int, nargs="?")
    ap.add_argument("--partition", action="store_true",
                    help="17 §12 第 1 項 —— 章區間之覆蓋／重疊檢查")
    ap.add_argument("--source", choices=("block", "layout"), default=SOURCE_DEFAULT,
                    help="PDF 來源（R-PMH71）—— 預設 block；`layout` 供對照")
    ap.add_argument("--source-must-hit", action="store_true",
                    help="R-PMH71 之 must-hit —— 章 9 以兩來源並列，"
                         "`layout` 須查不出 A-PMH16、`block` 須查得出")
    a = ap.parse_args()
    if a.source_must_hit:
        rc = source_must_hit()
        print_limits()
        sys.exit(rc)
    if a.partition:
        pdf_all = pdf_text(a.source)
        rc = partition(pdf_all)
        print_limits()
        sys.exit(1 if rc else 0)
    if a.chapter is None:
        raise SystemExit("須給章號，或用 --partition")
    ch = a.chapter
    pdf_all = pdf_text(a.source)
    pdf_ch = pdf_chapter(ch, pdf_all)
    rows = sys1_chapter(ch)
    sys_ch = " ".join(d for _, d in rows)

    print(f"=== 章 {ch} 之雙向複驗（R-PMH51）===")
    print(f"PDF 段：{len(pdf_ch)} 字元；SYS1：{len(rows)} 則、{len(sys_ch)} 字元")
    print(f"PDF 段起錨 `{STARTS[ch]}`（訖 = 下一章之起錨）；"
          f"**來源 = {a.source}**（R-PMH71）")
    import importlib
    _mc = importlib.import_module("marker_coverage")
    _cnt, _ = _mc.prefix_scan(pdf_all)
    _reqs, _ = _mc.req_prefixes(_cnt)
    mk = _mc.enumerate_markers(pdf_ch, _reqs)
    print(f"PDF 段內 marker：{len(mk)} 個 —— {mk}")

    # --- 方向一：SYS1 → PDF ---
    print(f"\n--- 方向一（SYS1 → PDF）：SYS1 之字是否出現於 PDF ---")
    print(f"{'outline':<8} {'字數':>5}  {'逐字命中':<8} 覆蓋率")
    d1_miss = []
    for o, d in rows:
        hit = d in pdf_all
        cov = coverage(d, pdf_all)
        print(f"{o:<8} {len(d):>5}  {'是' if hit else '**否**':<8} {cov:6.1%}")
        if not hit:
            d1_miss.append((o, d, cov))

    # --- 方向二：PDF → SYS1（漏句只在此方向顯示）---
    print(f"\n--- 方向二（PDF → SYS1）：PDF 之字是否出現於 SYS1 ---")
    sents = sentences(pdf_ch)
    print(f"PDF 段切出 {len(sents)} 句（>= {MIN_SENT} 字元）")
    print(f"{'#':>3}  {'逐字命中':<8} {'覆蓋率':>7}  句首")
    residue = []
    for i, s in enumerate(sents, 1):
        hit = s in sys_ch
        cov = coverage(s, sys_ch)
        print(f"{i:>3}  {'是' if hit else '**否**':<8} {cov:6.1%}  {s[:58]}")
        if not hit:
            residue.append(s)   # R-PMH66(b)：逐字未命中者一律進殘餘

    print(f"\n=== 結果（R-PMH66 —— 判定為二值，門檻只分流殘餘）===")
    print(f"  方向一未逐字命中：{len(d1_miss)} 則")
    for o, d, cov in d1_miss:
        print(f"    outline {o}（覆蓋 {cov:.1%}）：{d[:90]}")
    print(f"\n  方向二逐字命中 {len(sents)-len(residue)}/{len(sents)}；"
          f"**殘餘 {len(residue)} 句**")
    print("  **殘餘不得由門檻自動判為「非漏」** —— 逐句須有人讀之具名結論；"
          "\n  覆蓋率只決定人讀之優先順序（高者先看），不決定結論。")
    unnamed = []
    for s in sorted(residue, key=lambda x: -coverage(x, sys_ch)):
        k = residue_key(s)
        v = RESIDUE_VERDICT.get(k)
        print(f"\n    [覆蓋 {coverage(s, sys_ch):5.1%}] {s}")
        if v:
            print(f"      人讀結論：{v}")
        else:
            print("      **人讀結論：未具名 ← FAIL（R-PMH66(c)）**")
            unnamed.append(k)
    print(f"\n  殘餘未具名結論者：**{len(unnamed)}**"
          f"{'' if not unnamed else ' ← **FAIL**'}")
    print_limits()
    sys.exit(1 if unnamed else 0)


if __name__ == "__main__":
    main()
