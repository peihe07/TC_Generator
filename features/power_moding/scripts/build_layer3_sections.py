#!/usr/bin/env python3
"""05 包步驟 4 —— Layer 3 表之機器產出。

Layer 3 為**規格章節分群**，取規格自身之 section id（canon §4.1.1）——
`outline_number` 與其所屬「章」皆取自 SYS1 匯出之 `Outline Number`，
**不自創標籤**。`FROP` 取自 037，`pdf_page` 取自 `data/outline_map.json`。

輸出 data/layer3_sections.tsv（48 列 ＋ 表頭）。
本檔不擬 Layer 2 名稱、不定 granularity —— 該提案屬分析層。
"""
import argparse
import csv
import json
import re
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parent.parent

# R-PMH72 —— 保留於台帳但排除於交付之 leaf
EXCLUDED = {"SWE1-HMI-PM-028": "EXCLUDED-BY-R-PMH72"}


def main() -> None:
    # 20 包步驟 6（19 §14 第 4 項）—— **重跑不得靜默覆寫現值**。
    # `--out` 使重跑之產物落於暫存檔，供與現值逐列 diff。
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default=None,
                    help="輸出路徑（預設 data/layer3_sections.tsv）；"
                         "給定他值即只產出不覆寫，供 diff")
    args = ap.parse_args()
    cfg = yaml.safe_load((ROOT / "feature.yaml").read_text(encoding="utf-8"))
    a = openpyxl.load_workbook(ROOT / cfg["paths"]["a03_report"],
                               data_only=True)["Analysis Report"]
    s = openpyxl.load_workbook(ROOT / cfg["paths"]["sys1_export"],
                               data_only=True)["Basic Report"]
    # TSV 之欄分隔為 tab、列分隔為換行 —— 標題內含實體換行與 `_x000D_` 字面量，
    # 未正規化即寫出會把一列拆成多列（首版即如此，見上繳 §5.1）。
    def flat(v: str, n: int = 120) -> str:
        return re.sub(r"\s+", " ", str(v or "").replace("_x000D_", " ")).strip()[:n]

    outline_title = {}
    for r in range(2, s.max_row + 1):
        if s.cell(r, 1).value:
            outline_title[str(s.cell(r, 3).value).strip()] = flat(s.cell(r, 4).value)
    omap = {l["swe_req_id"]: l for l in json.loads(
        (ROOT / "data" / "outline_map.json").read_text(encoding="utf-8"))["leaves"]}

    rows, missing = [], []
    for r in range(8, a.max_row + 1):
        if str(a.cell(r, 7).value).strip() != "Functional Requirement":
            continue
        sid = str(a.cell(r, 1).value).strip()
        outline = str(a.cell(r, 3).value).split("_")[-1].strip()
        if outline not in outline_title:          # 停止條件 9
            missing.append((sid, outline))
        ch = outline.split(".")[0]
        rows.append({
            "swe_requirement_id": sid,
            "outline_number": outline,
            "chapter": ch,
            "chapter_title": outline_title.get(ch, ""),
            "section_title": outline_title.get(outline, ""),
            "frop": flat(a.cell(r, 8).value, 60),
            "pdf_page": omap.get(sid, {}).get("pdf_page", ""),
            # R-PMH72（Pei 2026-08-24「DR-PMH1 拿掉」）：該 leaf **不寫入交付
            # 工作簿、不產出 TC**，惟其列保留於本內部台帳（G-D：「不做」與
            # 「沒發現」須在紙上分得開）。有 TC 之 leaf 因而為 47。
            "excluded_by": EXCLUDED.get(sid, ""),
        })

    print(f"leaf {len(rows)}；對應到規格自身 section id 者 {len(rows) - len(missing)}"
          f"/{len(rows)}")
    if missing:
        raise SystemExit(f"停止條件 9：{len(missing)} 個 leaf 無 section id -> {missing}")

    out = Path(args.out) if args.out else ROOT / "data" / "layer3_sections.tsv"
    with out.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0]), delimiter="\t")
        w.writeheader()
        w.writerows(rows)
    # 寫出後複驗結構：列數與欄數（R-G7-1 之對照向 —— 不以「寫出成功」為通過）
    with out.open(encoding="utf-8") as fh:
        back = list(csv.reader(fh, delimiter="\t"))
    ncol = len(back[0])
    bad = [i for i, r in enumerate(back) if len(r) != ncol]
    if len(back) != len(rows) + 1 or bad:
        raise SystemExit(f"TSV 結構自檢失敗：列數 {len(back)}（期望 {len(rows)+1}）、"
                         f"欄數異常之列 {bad[:5]}")
    print(f"wrote {out} —— 回讀 {len(back)-1} 列 × {ncol} 欄，結構自檢通過")


if __name__ == "__main__":
    main()
