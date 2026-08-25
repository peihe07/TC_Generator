#!/usr/bin/env python3
"""R-PMH54／R-PMH57 —— 規格覆蓋以 **marker 枚舉**為權威判準。

自 PDF 枚舉全部需求標記，逐一檢查其是否出現於 SYS1 匯出之 `Description` 全文。
**無門檻、無取樣、無相似度參數** —— 結果為二值（在／不在）。

**R-PMH57（15 包）**：marker 之前綴清單**不得人工列舉**，
須先以反向掃描產生候選前綴，逐一判定為「需求 marker」或
「交叉參照／偽命中」，再據該判定枚舉。

  人工列舉之六前綴（14 包）遺漏 `DS`（`DS4.1)`，對應 outline 7.5.1），
  致全集被算成 30 而非 31。**該錯為分析層與執行層各自算出之同一值** ——
  二者用了同一份人工清單；**先算後比只能抓「算法不同而結果不同」，
  抓不到「前提相同而前提本身錯」**。

與 13 包之句級雙向 diff 之分工（R-PMH54）：
    marker 枚舉 → **需求單位是否存在**（本檔，權威）
    句級 diff  → 已存在之單位其內容是否完整（`bidirectional_spec_diff.py`，輔助）

用法:
    python scripts/marker_coverage.py
    python scripts/marker_coverage.py --prefix-scan   # 反向掃描與判定表
    python scripts/marker_coverage.py --self-test
"""

# R-PMH92 —— 本檢查之 must-hit 註冊。**總表之結果欄由此決定，手寫不採認。**
HAS_MUST_HIT = True
MUST_HIT_NOTE = '`--self-test` 之 must-hit A／B／C／D'

import argparse
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PDF_TXT = ROOT / "sandbox" / "spec.txt"
SYS1 = ROOT / "inputs" / "SYS1_HMI_Power_Moding_HMI_Logic_and_Flow_R1_SR24_2A.xlsx"

# --- R-PMH57 步驟一：反向掃描之候選形態（不含任何前綴名）---
CANDIDATE = re.compile(r"\b([A-Za-z][A-Za-z_ ]{0,8}?)\s?(\d+(?:\.\d+)?)\s?([.):])")

# --- R-PMH57 步驟二：候選前綴之逐項判定。**每一候選皆須在此具名** ---
# 三值：`req` 需求 marker／`xref` 交叉參照（他文件之編號）／`noise` 一般文句之偽命中。
# 未在此表者 → **FAIL**，不得沉默略過（新規格版本引入新前綴時即由此攔下）。
VERDICT: dict[str, tuple[str, str]] = {
    "SU":      ("req",   "Start Up，章 7 之需求 marker"),
    "SSND":    ("req",   "Start Up Sounds，章 8"),
    "PM":      ("req",   "Power Moding，章 9"),
    "PITA":    ("req",   "Power ITA，章 10"),
    "VRLP":    ("req",   "Voice Recognition Long Press，章 11"),
    "OFF":     ("req",   "Power Off，章 12"),
    "DS":      ("req",   "章 7 之需求 marker；其編號 4.1 與 SU4.) 呈父子，"
                         "前綴由 SU 變 DS，極可能為規格原文筆誤（15 包 §2.3，"
                         "依 R-PMH26 只登記不開 DR，一律照原文處理）"),
    "DCR":     ("xref",  "變更申請單號，非本規格之需求"),
    "CR":      ("xref",  "同上（`CR19385)` 與 `DCR19385)` 同一單）"),
    "CFTS":    ("xref",  "他規格文件編號（`CFTS009`）"),
    "CTS":     ("xref",  "同上（`See CTS009)`，`CFTS009` 之另一寫法）"),
    "High":    ("noise", "一般文句：`High` + 版面尺寸數字"),
    "Low":     ("noise", "一般文句：`Low` + 版面尺寸數字"),
    "and":     ("noise", "一般文句：`Last and 1.` 之類"),
    "sec":     ("noise", "一般文句：`sec 1.` 之類"),
    "a":       ("noise", "一般文句：`with a 1.`"),
    "the":     ("noise", "一般文句：`of the 10.`"),
    "of":      ("noise", "一般文句：`of 10.`"),
    "to":      ("noise", "一般文句：`up to 2.`"),
    "expires": ("noise", "一般文句：`expires 3.`"),
    # 以下二者**不出現於 `pdftotext -layout` 之萃取**，只出現於 PyMuPDF 之
    # 萃取（16 包步驟 4）—— p9 流程圖之標籤，兩種萃取器之閱讀順序不同，
    # 致相鄰標籤被併成不同字串。**候選集合本身是萃取相依的。**
    "Loading": ("noise", "p9 流程圖標籤：`System Loading` 後接 `1.5 sec timeout`，"
                         "二者為圖上相鄰之獨立標籤，非編號"),
    "each":    ("noise", "p9 流程圖標籤：`…with a 1.5 timeout each` 後接 "
                         "`1.5 sec timeout`，同上"),
}

# marker 所屬之章（依 PDF 之編排；供分章列表，不參與判定）
CHAPTER = {"SU": 7, "DS": 7, "SSND": 8, "PM": 9, "PITA": 10, "VRLP": 11, "OFF": 12}


# --- R-PMH52 之擴及（17 包步驟 4）---
# R-PMH52 之措詞為「**任何** lint 之輸出須具名其未涵蓋之範圍」，
# 而 16 包 §5.3 查出該條實際只施行於 `lint_batch.py` —— 單向套用。
# 本檢查自此於輸出末尾具名其限度。**內容須為本檢查之限度，
# 不得寫成一般性免責。**
LIMITS = [
    "marker 之**內容**不看 —— 只驗其標記字串是否出現於 SYS1；該 marker 之需求文句是否完整（如 7.1 之漏句）屬 `bidirectional_spec_diff.py`",
    "`VERDICT` 之判定值仍為人工 —— must-hit C 只攔「未判定」，**不攔「判錯」**；R-PMH61 之語氣檢查為部分補救，且其窗口只取 marker 之後",
    "**候選集合隨萃取器而變**（16 §4.2：PyMuPDF 多出 `Loading`／`each`）—— 本檢查之候選以 `sandbox/spec.txt` 為準",
    "候選形態 `CANDIDATE` 仍是一個正規式 —— 若規格改用 `[A-3]` 或 `Req 4.1 -` 之類形態，反向掃描一樣看不見",
    "SYS1 側只讀 `Basic Report` 之 `Description` 欄；**其餘欄不看**",
]


def print_limits() -> None:
    print("\n=== 本檢查未涵蓋之範圍（R-PMH52）===")
    for _x in LIMITS:
        print(f"  - {_x}")
    print("  **以上各項本檢查一律不看** —— 其全綠不含關於該等項之任何資訊。")


def norm(t: str) -> str:
    t = str(t).replace("_x000D_", " ")
    for a, b in (("‘", "'"), ("’", "'"),
                 ("“", '"'), ("”", '"'), ("…", "...")):
        t = t.replace(a, b)
    return re.sub(r"\s+", " ", t).strip()


def canon_marker(m: str) -> str:
    """比對用之正規形：去空白（`SSND 1)` → `SSND1)`）。"""
    return re.sub(r"\s+", "", m)


def prefix_scan(pdf: str) -> tuple[Counter, dict[str, list[str]]]:
    """R-PMH57 步驟一 —— 反向掃描，回傳候選前綴之分布與樣例。

    候選之前綴取其**末一個字詞**（`cycle SU9.)` 之前綴為 `SU`，非 `cycle SU`）。
    """
    cnt: Counter = Counter()
    ex: dict[str, list[str]] = defaultdict(list)
    for m in CANDIDATE.finditer(pdf):
        p = re.split(r"\s+", m.group(1).strip())[-1]
        cnt[p] += 1
        if len(ex[p]) < 4:
            ex[p].append(m.group(0).strip())
    return cnt, ex


def req_prefixes(cnt: Counter, extra_drop: tuple[str, ...] = ()) -> tuple[list[str], list[str]]:
    """判定表套用於掃描結果，回傳（需求前綴, 未判定之候選）。"""
    unjudged = [p for p in cnt if p not in VERDICT]
    reqs = [p for p in cnt
            if VERDICT.get(p, ("", ""))[0] == "req" and p not in extra_drop]
    # 長者優先，使 `SSND` 不被 `SU` 之類前綴搶匹配（此處無此情形，仍固定其序）
    return sorted(reqs, key=lambda s: (-len(s), s)), sorted(unjudged)


def enumerate_markers(pdf: str, prefixes: list[str]) -> list[str]:
    """依判定所得之前綴枚舉 marker，保持 PDF 出現順序、去重。

    形態：`前綴` + 空白? + `數字[.數字]` + 選擇性之尾點 + `)` 或 `:`。
    尾點（`SU1.)`）與小數（`SU9.1)`）皆須收得 —— 14 包首版即漏 `SU1.)`。
    """
    pat = re.compile(r"\b(" + "|".join(prefixes) + r")\s*\d+(?:\.\d+)?\.?[):]")
    seen, order = set(), []
    for m in pat.finditer(pdf):
        c = canon_marker(m.group(0))
        if c not in seen:
            seen.add(c)
            order.append(m.group(0))
    return order


def load(extra_drop: tuple[str, ...] = ()) -> tuple[list[str], str, Counter, dict, list[str]]:
    pdf = norm(PDF_TXT.read_text(errors="replace"))
    cnt, ex = prefix_scan(pdf)
    prefixes, unjudged = req_prefixes(cnt, extra_drop)
    order = enumerate_markers(pdf, prefixes)
    ws = openpyxl.load_workbook(SYS1, data_only=True)["Basic Report"]
    sysall = norm(" ".join(str(ws.cell(r, 4).value or "")
                           for r in range(2, ws.max_row + 1)))
    return order, sysall, cnt, ex, unjudged


def verify_extraction(other: Path) -> int:
    """R-PMH60 —— 兩份獨立萃取之等同性，以 **marker 集合逐項相等**驗之。

    **不以字元數、行數或位元組數** —— R-PMH21 已裁定抽取字元數不得作為
    完整性、正確性或版本一致性之判準，而「兩份萃取是否為同一份文件」
    即版本一致性之問題。marker 為需求單位之標記，**不受正規化策略影響**。
    """
    other = Path(other).resolve()
    a = norm(PDF_TXT.read_text(errors="replace"))
    b = norm(other.read_text(errors="replace"))
    ca, _ = prefix_scan(a)
    cb, _ = prefix_scan(b)
    pa, ua = req_prefixes(ca)
    pb, ub = req_prefixes(cb)
    ma = [canon_marker(x) for x in enumerate_markers(a, pa)]
    mb = [canon_marker(x) for x in enumerate_markers(b, pb)]
    sa, sb = set(ma), set(mb)
    print("\n=== 兩份萃取之等同性（R-PMH60）===")
    print(f"  A：`{PDF_TXT.relative_to(ROOT)}`")
    print(f"  B：`{other}`")
    print(f"\n  marker 全集：A = **{len(sa)}**；B = **{len(sb)}**")
    only_a, only_b = sorted(sa - sb), sorted(sb - sa)
    for ch in sorted({CHAPTER.get(re.match(r'[A-Za-z]+', m).group(0), 0)
                      for m in sa | sb}):
        na = sum(1 for m in sa if CHAPTER.get(re.match(r'[A-Za-z]+', m).group(0), 0) == ch)
        nb = sum(1 for m in sb if CHAPTER.get(re.match(r'[A-Za-z]+', m).group(0), 0) == ch)
        print(f"    章 {ch:>2}：A = {na:>2}；B = {nb:>2}"
              + ("" if na == nb else "   ← **不等**"))
    ok = sa == sb
    if ok:
        print("\n  **逐項相等 —— 等同性成立**（二者為同一份規格之同一版本）")
    else:
        print(f"\n  **不相等** —— 只在 A：{only_a}；只在 B：{only_b}")
    if ua or ub:
        print(f"  ⚠ 未判定之候選前綴：A {ua}；B {ub}")
    # 字元數只**列出**，明載其不作為判準
    print(f"\n  （字元數 A = {len(a)}／B = {len(b)}，"
          f"差 {abs(len(a)-len(b))} —— **依 R-PMH21／R-PMH60 不作為判準**，"
          "列出僅為記錄）")
    return 0 if (ok and not ua and not ub) else 1


def print_verdict_table(cnt: Counter, ex: dict) -> None:
    print("\n=== 反向掃描之前綴判定表（R-PMH57）===")
    print(f"候選形態 = `{CANDIDATE.pattern}`；候選前綴 = {len(cnt)} 種\n")
    print(f"{'前綴':<10} {'次數':>4}  {'判定':<6} {'樣例':<34} 依據")
    order = sorted(cnt, key=lambda p: (
        {"req": 0, "xref": 1, "noise": 2}.get(VERDICT.get(p, ("?", ""))[0], -1),
        -cnt[p], p))
    for p in order:
        v, why = VERDICT.get(p, ("**未判定**", "← 須逐項判定，不得沉默略過"))
        print(f"{p:<10} {cnt[p]:>4}  {v:<6} {' '.join(ex[p])[:34]:<34} {why}")


# --- R-PMH61（16 包）：`VERDICT` **誤判**之偵測 ---
# 反向掃描（R-PMH57）保證「沒有候選被漏看」，must-hit C 保證「沒有候選未判定」。
# **二者皆不攔「判錯」** —— 把真需求前綴標為 `noise`，該章一樣靜默消失。
# 故對判為 `noise`／`xref` 者，檢查其**鄰近文句是否具需求語氣**，
# 命中者升為「須人讀確認」並具名，**不自行改判**。
MODAL = re.compile(r"\b(shall|should|must|will|needs? to|is required to)\b", re.I)
# 祈使句：marker 之後（容許一個 `If …,` 前置子句）以原形動詞起首
IMPERATIVE = re.compile(
    r"^\s*(?:if\b[^,]{0,120},\s*)?"
    r"(do not|don't|show|display|play|jump|set|go|present|keep|turn|enable|"
    r"disable|remove|hide|wait|use|select|ensure|start|stop|mute|unmute)\b", re.I)
TONE_WINDOW = 180   # marker 之後取之字元數（其所引領之句）


def tone_scan(pdf: str, prefixes: list[str],
              side: str = "after") -> dict[str, list[tuple[str, str]]]:
    """回傳 {前綴: [(命中之 marker 文字, 語氣證據)]}。

    `side="after"`（預設）：窗口取 marker **之後** `TONE_WINDOW` 字元 ——
    需求 marker 之作用是引領其需求文句，故證據在其後而非其前。

    `side="before"`（17 包步驟 6，16 §10 第 1 項）：取其**前** —— 供對照。
    **窗口方向本身是一個假設**：若某前綴之 marker 恆出現於需求句**尾**，
    其語氣證據會落在 `after` 窗之外而本檢查看不見。
    **只回報兩窗之旗標差異，不改現行窗口之選擇。**
    """
    hits: dict[str, list[tuple[str, str]]] = {}
    for p_ in prefixes:
        pat = re.compile(r"\b" + re.escape(p_) + r"\s*\d+(?:\.\d+)?\.?[):]")
        for m in pat.finditer(pdf):
            after = (pdf[m.end():m.end() + TONE_WINDOW] if side == "after"
                     else pdf[max(0, m.start() - TONE_WINDOW):m.start()])
            ev = ""
            if (im := IMPERATIVE.match(after)):
                ev = f"祈使句起首 `{im.group(1)}`"
            elif (mo := MODAL.search(after)):
                ev = f"情態動詞 `{mo.group(1)}`"
            if ev:
                hits.setdefault(p_, []).append((m.group(0), ev))
    return hits


def print_tone_report(pdf: str, cnt: Counter, extra_req_drop: tuple[str, ...] = (),
                      side: str = "after") -> list[str]:
    """對判為 `noise`／`xref` 者做語氣檢查，回傳須人讀之前綴清單。"""
    targets = [p_ for p_ in cnt
               if VERDICT.get(p_, ("", ""))[0] in {"noise", "xref"}
               or p_ in extra_req_drop]
    targets.sort()
    hits = tone_scan(pdf, targets, side)
    lbl = "其後" if side == "after" else "其前"
    print(f"\n=== 非需求前綴之需求語氣檢查（R-PMH61，窗口取 marker {lbl}）===")
    print(f"受檢前綴 = {len(targets)} 個（判為 `noise`／`xref` 者）\n")
    print(f"{'前綴':<10} {'判定':<6} {'語氣命中':>6}  證據")
    flagged = []
    for p_ in targets:
        v = VERDICT.get(p_, ("(測試替身)", ""))[0]
        h = hits.get(p_, [])
        if h:
            flagged.append(p_)
        ev = "；".join(f"{a} → {b}" for a, b in h[:2]) if h else "—"
        print(f"{p_:<10} {v:<6} {len(h):>6}  {ev[:80]}")
    print(f"\n**須人讀確認之前綴**：{flagged or '無'}")
    print("  （只具名，**不自行改判** —— 判定值之變更須另有依據）")
    return flagged


def print_window_compare(pdf: str, cnt: Counter) -> bool:
    """17 包步驟 6 —— 兩窗口之旗標集合對照。**只回報，不改判準。**"""
    a = set(print_tone_report(pdf, cnt, side="after"))
    b = set(print_tone_report(pdf, cnt, side="before"))
    print("\n=== 兩窗口之旗標對照（17 包步驟 6）===")
    print(f"  取其後（現行）：{sorted(a) or '無'}")
    print(f"  取其前（對照）：{sorted(b) or '無'}")
    print(f"  只在其後：{sorted(a - b) or '無'}")
    print(f"  **只在其前：{sorted(b - a) or '無'}**"
          "  ← 此集合非空即表示現行窗口看不見之語氣證據確實存在")
    print("  **只回報，不改現行窗口之選擇**（16 §10 第 1 項之自提項）。")
    return a == b


def coverage(markers: list[str], sysall: str) -> list[tuple[str, bool]]:
    """回傳 [(marker, 是否命中於 SYS1)]。比對亦去空白，避免 `SSND 1)` 之空白差異。"""
    sys_c = re.sub(r"\s+", "", sysall)
    return [(m, canon_marker(m) in sys_c) for m in markers]


def report(markers, sysall, label="") -> tuple[int, int]:
    res = coverage(markers, sysall)
    miss = [m for m, ok in res if not ok]
    print(f"\n=== marker 覆蓋（R-PMH54）{label} ===")
    print(f"PDF marker 全集 = **{len(res)}**；SYS1 缺 = **{len(miss)}**\n")
    by_ch: dict[int, list] = {}
    for m, ok in res:
        pre = re.match(r"[A-Za-z]+", m).group(0)
        by_ch.setdefault(CHAPTER.get(pre, 0), []).append((m, ok))
    print(f"{'章':>3}  {'PDF marker':<62} {'缺':>3}")
    for ch in sorted(by_ch):
        items = by_ch[ch]
        ms = " ".join(m for m, _ in items)
        n = sum(1 for _, ok in items if not ok)
        print(f"{ch:>3}  {ms[:62]:<62} {n:>3}")
        if len(ms) > 62:
            print(f"     {ms[62:]}")
    print(f"\n缺漏清單：{miss or '無'}")
    return len(res), len(miss)


def self_test() -> int:
    markers, sysall, cnt, ex, unjudged = load()
    print_verdict_table(cnt, ex)
    ok_judged = not unjudged
    print(f"\n  全部候選前綴皆已判定：{ok_judged}"
          f"{'' if ok_judged else '  ← 未判定：' + str(unjudged)}")

    print("\n=== 範圍向（R-G9）—— 現行素材 ===")
    total, miss = report(markers, sysall, "（範圍向）")
    ok_scope = (total, miss) == (31, 2)
    print(f"\n  與分析層 15 包 §2.2 之 31／2 相符：{ok_scope}")

    # must-hit A：自 SYS1 側移除一個**已知存在**之 marker（測試替身）
    victim = next(m for m, ok in coverage(markers, sysall) if ok)
    tampered = re.sub(re.escape(victim), "XXXX", sysall, count=99)
    print(f"\n=== must-hit A —— 自 SYS1 側移除已知存在之 marker `{victim}` ===")
    _, m2 = report(markers, tampered, "（must-hit A）")
    caught = m2 == miss + 1 and victim in [
        m for m, ok in coverage(markers, tampered) if not ok]
    print(f"\n  缺漏由 {miss} 增為 {m2}，且 `{victim}` 在缺漏清單內：{caught}")

    # must-hit B（R-PMH57，15 包步驟 2）——
    # 自前綴清單移除 `DS`，須使全集降為 30 且 `DS4.1)` 靜默消失。
    print("\n=== must-hit B —— 自前綴清單移除 `DS`（14 包之人工列舉狀態）===")
    m_drop, sysall2, _, _, _ = load(extra_drop=("DS",))
    t3, m3 = report(m_drop, sysall2, "（must-hit B）")
    ds_gone = not any(x.startswith("DS") for x in m_drop)
    caught_b = (t3 == 30) and ds_gone
    print(f"\n  全集降為 {t3}（期望 30）：{t3 == 30}；"
          f"`DS4.1)` 自全集消失：{ds_gone}")
    print("  **且缺漏數不變** —— 故若僅比對缺漏數，此錯不會被察覺；"
          f"分母 {t3} vs {total} 才是證據。")

    # must-hit C —— 判定表缺一項時須攔下（模擬新規格引入未判定之前綴）
    print("\n=== must-hit C —— 自判定表移除 `OFF` 之判定（模擬未判定之新前綴）===")
    saved = VERDICT.pop("OFF")
    _, _, cnt3, _, unj3 = load()
    VERDICT["OFF"] = saved
    caught_c = "OFF" in unj3
    print(f"  未判定清單 = {unj3}；被攔下：{caught_c}")

    # --- R-PMH61：語氣檢查與其 must-hit ---
    pdf = norm(PDF_TXT.read_text(errors="replace"))
    flagged = print_tone_report(pdf, cnt)

    same_window = print_window_compare(pdf, cnt)

    print("\n=== must-hit D —— 將 `SU` 之判定由 `req` 改為 `noise`（測試替身）===")
    print("  其鄰近文句必含需求語氣，**故須被升為須人讀並攔下**；"
          "\n  攔不下者，本檢查對真正之誤判亦無效（R-PMH61）。")
    saved = VERDICT["SU"]
    VERDICT["SU"] = ("noise", "（測試替身 —— must-hit D）")
    flagged_d = print_tone_report(pdf, cnt)
    VERDICT["SU"] = saved
    caught_d = "SU" in flagged_d
    print(f"\n  `SU` 被升為須人讀：{caught_d}")

    print("\n" + "=" * 70)
    print(f"前綴全判定: {ok_judged}；範圍向 31／2: {ok_scope}；"
          f"must-hit A: {caught}；must-hit B: {caught_b}；"
          f"must-hit C: {caught_c}；must-hit D: {caught_d}")
    print(f"兩窗口旗標集合相同: {same_window}"
          "（不同不構成 FAIL —— 步驟 6 明令只回報）")
    return 0 if all((ok_judged, ok_scope, caught, caught_b,
                     caught_c, caught_d)) else 1


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--self-test", action="store_true")
    ap.add_argument("--prefix-scan", action="store_true")
    ap.add_argument("--tone-scan", action="store_true",
                    help="R-PMH61 —— 非需求前綴之需求語氣檢查")
    ap.add_argument("--window-compare", action="store_true",
                    help="17 包步驟 6 —— 語氣窗口取其前／其後之旗標對照")
    ap.add_argument("--verify-extraction", metavar="TEXT_FILE",
                    help="R-PMH60 —— 與另一份萃取比對 marker 集合")
    a = ap.parse_args()
    if a.self_test:
        rc = self_test()
        print_limits()
        sys.exit(rc)
    if a.verify_extraction:
        rc = verify_extraction(Path(a.verify_extraction))
        print_limits()
        sys.exit(rc)
    markers, sysall, cnt, ex, unjudged = load()
    if a.prefix_scan:
        print_verdict_table(cnt, ex)
        print_limits()
        sys.exit(1 if unjudged else 0)
    if a.window_compare:
        print_window_compare(norm(PDF_TXT.read_text(errors="replace")), cnt)
        print_limits()
        sys.exit(0)
    if a.tone_scan:
        print_tone_report(norm(PDF_TXT.read_text(errors="replace")), cnt)
        print_limits()
        sys.exit(1 if unjudged else 0)
    total, miss = report(markers, sysall)
    print_limits()
    if unjudged:
        print(f"\n⚠ 未判定之候選前綴：{unjudged}  ← R-PMH57，須逐項判定")
    sys.exit(0 if (miss == 0 and not unjudged) else 1)


if __name__ == "__main__":
    main()
