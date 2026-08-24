#!/usr/bin/env python3
"""29 包分析層：依下放包 28 §二 A／B／C 產生 `plan.json`。

基底 `features/power/sandbox/b28/pm_28.xlsx`（① 括號收斂之產物）。
輸出僅為計畫；不碰任何 xlsx。
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "scripts"))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import lint036                                              # noqa: E402
import rules                                                # noqa: E402

BASE = ROOT / "features/power/sandbox/b28/pm_28.xlsx"
SHEET = "Test Case Specification&Result"
FIRST_ROW = 10
COLS = {"test_item": "I", "pre": "J", "input": "K", "proc": "L", "er": "M"}
OUT = Path(__file__).resolve().parent / "plan.json"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def lines_of(text: str) -> list[str]:
    return str(text or "").split("\n")


def transform_row(cells: dict[str, str], audit: dict) -> dict[str, str]:
    """回傳該列各欄之新值（未改動者維持原值）。"""
    new = dict(cells)

    # --- B／C：proc 步驟改寫 + ER 1:1 同步 --------------------------------
    proc_lines, er_lines = lines_of(cells["proc"]), lines_of(cells["er"])
    proc_idx = [i for i, ln in enumerate(proc_lines) if rules.NUMBERED.match(ln)]
    er_idx = [i for i, ln in enumerate(er_lines) if rules.NUMBERED.match(ln)]
    if len(proc_idx) == len(er_idx):
        for step, (pi, ei) in enumerate(zip(proc_idx, er_idx)):
            p_pre, p_body = rules.split_body(proc_lines[pi])
            e_pre, e_body = rules.split_body(er_lines[ei])
            hit = rules.rewrite_proc(p_body.strip())
            if not hit:
                continue
            new_p, new_e, rule = hit
            proc_lines[pi], er_lines[ei] = p_pre + new_p, e_pre + new_e
            audit["steps"].append({
                "row": audit["row"], "step": step + 1, "rule": rule,
                "proc_old": p_body.strip(), "proc_new": new_p,
                "er_old": e_body.strip(), "er_new": new_e})
    else:
        audit["e_misaligned"] = [len(proc_idx), len(er_idx)]
    new["proc"], new["er"] = "\n".join(proc_lines), "\n".join(er_lines)

    # --- B-1：pre 之設定宣告 ----------------------------------------------
    pre_lines = lines_of(cells["pre"])
    for i, ln in enumerate(pre_lines):
        if not rules.NUMBERED.match(ln):
            continue
        prefix, body = rules.split_body(ln)
        hit = rules.rewrite_pre(body.strip())
        if not hit:
            continue
        pre_lines[i] = prefix + hit[0]
        audit["pre"].append({"row": audit["row"], "line": i + 1,
                             "old": body.strip(), "new": hit[0]})
    new["pre"] = "\n".join(pre_lines)

    # --- A：TLM→HU（四欄全欄 + test_item 之括號下半）----------------------
    for key in ("pre", "input", "proc", "er"):
        new[key] = rules.tlm_to_hu(new[key])
    item_lines = lines_of(new["test_item"])
    for i, ln in enumerate(item_lines):
        if lint036.RE_PAREN_LINE.match(ln.strip()):
            item_lines[i] = rules.tlm_to_hu(ln)
    new["test_item"] = "\n".join(item_lines)
    return new


def main() -> None:
    wb = openpyxl.load_workbook(BASE)
    ws = wb[SHEET]
    plan = {"sheet": SHEET, "base": str(BASE.relative_to(ROOT)),
            "base_sha256": sha256(BASE), "cells": {}, "audit": []}
    dup_alerts, unfit, press_alerts = [], [], []

    for row in range(FIRST_ROW, ws.max_row + 1):
        cells = {k: str(ws[f"{c}{row}"].value or "")
                 for k, c in COLS.items()}
        if not any(v.strip() for v in cells.values()):
            continue
        audit = {"row": row, "tc": str(ws[f"F{row}"].value or ""),
                 "steps": [], "pre": []}
        new = transform_row(cells, audit)

        # 同列改寫後**新生**之逐字重複步驟／前提（基底已重複者不計）。
        # 不自行處置：去重等於裁定「兩條前提是同一件事」，超出下放包授權。
        for key in ("proc", "pre"):
            def dup_set(text: str) -> set[str]:
                bodies = [rules.split_body(ln)[1].strip()
                          for ln in lines_of(text)
                          if rules.NUMBERED.match(ln)]
                return {b for b in bodies if bodies.count(b) > 1}
            born = dup_set(new[key]) - dup_set(cells[key])
            if born:
                dup_alerts.append({"row": row, "field": key,
                                   "dups": sorted(born)})

        # C 與 B2 各自套句式後，同列可能出現兩次電源鍵按壓（語意瑕疵）
        presses = [ln for ln in lines_of(new["proc"])
                   if rules.NUMBERED.match(ln) and "Press the HU power button" in ln]
        if len(presses) > 1:
            press_alerts.append({"row": row, "tc": audit["tc"],
                                 "steps": [rules.split_body(l)[1].strip()
                                           for l in presses]})

        changed = {COLS[k]: new[k] for k in COLS if new[k] != cells[k]}
        if changed:
            plan["cells"][str(row)] = changed
        if audit["steps"] or audit["pre"] or audit.get("e_misaligned"):
            plan["audit"].append(audit)

        # 表外殘留（四欄）：仍含目標變數之編號行，逐行收集
        for key in ("pre", "proc", "er"):
            for ln in lines_of(new[key]):
                if not rules.NUMBERED.match(ln):
                    continue
                body = rules.split_body(ln)[1].strip()
                if body.startswith("PENDING:"):      # 本包新置之佔位，非殘留
                    continue
                body = rules.DOLLAR.sub("", body)
                m = rules.RE_TARGET.search(body)
                if m:
                    unfit.append({"row": row, "field": key, "var": m.group(0),
                                  "line": rules.split_body(ln)[1].strip()})

    plan["dup_alerts"] = dup_alerts
    plan["press_alerts"] = press_alerts
    plan["unfit"] = unfit
    OUT.write_text(json.dumps(plan, ensure_ascii=False, indent=1),
                   encoding="utf-8")
    n_steps = sum(len(a["steps"]) for a in plan["audit"])
    n_pre = sum(len(a["pre"]) for a in plan["audit"])
    print(f"基底 sha256 {plan['base_sha256'][:20]}")
    print(f"改動列 {len(plan['cells'])}｜proc/er 步 {n_steps}｜pre 行 {n_pre}")
    print(f"重複告警 {len(dup_alerts)}｜雙重按壓告警 {len(press_alerts)}"
          f"｜表外殘留行 {len(unfit)}")


if __name__ == "__main__":
    main()
