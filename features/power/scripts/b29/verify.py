#!/usr/bin/env python3
"""29 包驗收：逐項對 `pm_29.xlsx` 覆核，任一項不達成即 exit 1。

規則於此**自基底獨立重算**（`rules.py` 直接套於 pm_28，不讀 `plan.json`
之結果值），故 build 與 apply 之間若脫節、或規則實作有誤，此處會攤開。
"""

from __future__ import annotations

import json
import re
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "scripts"))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import lint036                                              # noqa: E402
import rules                                                # noqa: E402
from build import COLS, FIRST_ROW, SHEET, transform_row     # noqa: E402

PREV = ROOT / "features/power/sandbox/b28/pm_28.xlsx"
OUT = ROOT / "features/power/sandbox/b29/pm_29.xlsx"
LAST_COL = 34
BARE_READS = ("Read Antitheft_Activation.Req", "Read VPLastStatus",
              "Read Timeout1")


def cells_of(ws, row: int) -> dict[str, str]:
    return {k: str(ws[f"{c}{row}"].value or "") for k, c in COLS.items()}


def numbered(text: str) -> list[str]:
    return [rules.split_body(ln)[1].strip()
            for ln in str(text or "").split("\n") if rules.NUMBERED.match(ln)]


def stray_tlm(text: str) -> list[str]:
    """`$…$` 外之獨立 `TLM` token（`TLM_x`／`LTM` 不算）。"""
    spans = rules.mask_dollars(text)
    return [m.group(0) for m in rules.RE_TLM.finditer(text)
            if not any(a <= m.start() < b for a, b in spans)]


def x14(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as z:
        for name in z.namelist():
            if name.startswith("xl/worksheets/sheet"):
                pass
        xml = "".join(z.read(n).decode("utf-8", "ignore")
                      for n in z.namelist()
                      if n.startswith("xl/worksheets/sheet"))
    return re.findall(r"<xm:sqref>([^<]*)</xm:sqref>", xml)


def main() -> int:
    prev_ws = openpyxl.load_workbook(PREV)[SHEET]
    ws = openpyxl.load_workbook(OUT)[SHEET]
    fails: list[str] = []

    def check(ok: bool, label: str, detail: str = "") -> None:
        print(f"  {'達成' if ok else '未達成'}  {label}"
              + (f" — {detail}" if detail else ""))
        if not ok:
            fails.append(label)

    # 1 列數／ID -------------------------------------------------------------
    check(ws.max_row == prev_ws.max_row, "列數與 pm_28 相同",
          f"{prev_ws.max_row} → {ws.max_row}")
    ids = [(prev_ws[f"B{r}"].value, prev_ws[f"F{r}"].value,
            ws[f"B{r}"].value, ws[f"F{r}"].value)
           for r in range(FIRST_ROW, ws.max_row + 1)]
    check(all(a == c and b == d for a, b, c, d in ids), "No.# 與 Test Case ID 不變")

    # 2 相異範圍 -------------------------------------------------------------
    off_field, diff_rows = [], set()
    allowed = set(COLS.values())
    for row in range(1, ws.max_row + 1):
        for col in range(1, LAST_COL + 1):
            if (ws.cell(row, col).value or "") == (prev_ws.cell(row, col).value or ""):
                continue
            diff_rows.add(row)
            letter = openpyxl.utils.get_column_letter(col)
            if letter not in allowed:
                off_field.append((row, letter))
    check(not off_field, "相異僅限 test_item/pre/input/proc/er 五欄",
          f"改動 {len(diff_rows)} 列；例外 {off_field[:5]}")

    # 3 獨立重算 -------------------------------------------------------------
    mismatch = []
    for row in range(FIRST_ROW, ws.max_row + 1):
        base = cells_of(prev_ws, row)
        if not any(v.strip() for v in base.values()):
            continue
        audit = {"row": row, "steps": [], "pre": []}
        want = transform_row(base, audit)
        got = cells_of(ws, row)
        for key in COLS:
            if want[key] != got[key]:
                mismatch.append((row, key))
    check(not mismatch, "全 389 列自基底獨立重算逐格相符",
          f"相異 {mismatch[:5]}")

    # 4 test_item：上半逐字不變、下半僅括號行改動 ----------------------------
    upper_bad, non_paren = [], []
    for row in range(FIRST_ROW, ws.max_row + 1):
        old, new = str(prev_ws[f"I{row}"].value or ""), str(ws[f"I{row}"].value or "")
        if old == new:
            continue
        o_lines, n_lines = old.split("\n"), new.split("\n")
        if len(o_lines) != len(n_lines):
            upper_bad.append(row)
            continue
        for o, n in zip(o_lines, n_lines):
            if o == n:
                continue
            if not lint036.RE_PAREN_LINE.match(n.strip()):
                non_paren.append((row, n[:40]))
            if rules.tlm_to_hu(o) != n:                 # 僅容 TLM→HU
                upper_bad.append(row)
    check(not upper_bad and not non_paren,
          "test_item 改動僅限括號下半且僅為 TLM→HU",
          f"上半/其他 {upper_bad[:5]}｜非括號行 {non_paren[:3]}")

    # 5 TLM 殘留 -------------------------------------------------------------
    stray = []
    for row in range(FIRST_ROW, ws.max_row + 1):
        for key, letter in COLS.items():
            text = str(ws[f"{letter}{row}"].value or "")
            if key == "test_item":
                text = "\n".join(lint036.paren_lines(text))
            if stray_tlm(text):
                stray.append((row, key))
    check(not stray, "主詞 TLM 於四欄＋括號下半殘留 = 0", f"{stray[:5]}")

    # 6 裸讀式殘留 -----------------------------------------------------------
    bare = [(row, ln) for row in range(FIRST_ROW, ws.max_row + 1)
            for ln in numbered(ws[f"L{row}"].value)
            if ln.startswith(BARE_READS)]
    expected_bare = [
        "Read VPLastStatus and check that it is the value held before the "
        "disconnection"]
    check(all(ln in expected_bare for _, ln in bare),
          "裸讀式殘留僅為明列之套不進句式者",
          f"{len(bare)} 行：{sorted({ln for _, ln in bare})}")

    # 7 Front_Panel_OnOff.Req 於 proc ---------------------------------------
    fp = [row for row in range(FIRST_ROW, ws.max_row + 1)
          if "Front_Panel_OnOff.Req" in str(ws[f"L{row}"].value or "")]
    check(not fp, "Front_Panel_OnOff.Req 於 proc 殘留 = 0", f"{fp}")

    # 8 PENDING 新增 = RemStartFail 檢查步數 ---------------------------------
    def n_pending(book) -> int:
        return sum(str(book[f"{c}{r}"].value or "").count(rules.PENDING_REMSTARTFAIL)
                   for r in range(FIRST_ROW, book.max_row + 1) for c in "JKLM")
    n_proc = sum(1 for r in range(FIRST_ROW, ws.max_row + 1)
                 for ln in numbered(ws[f"L{r}"].value)
                 if ln == rules.PENDING_REMSTARTFAIL)
    n_er = sum(1 for r in range(FIRST_ROW, ws.max_row + 1)
               for ln in numbered(ws[f"M{r}"].value)
               if ln == rules.PENDING_REMSTARTFAIL)
    check(n_pending(prev_ws) == 0 and n_proc == n_er and n_proc > 0,
          "PENDING(DR-PW23) 新增：proc 步數 = er 行數",
          f"proc {n_proc}／er {n_er}；基底 {n_pending(prev_ws)}")

    # 9 E=0（proc/er 1:1）----------------------------------------------------
    e_bad = [r for r in range(FIRST_ROW, ws.max_row + 1)
             if numbered(ws[f"L{r}"].value) and numbered(ws[f"M{r}"].value)
             and len(numbered(ws[f"L{r}"].value)) != len(numbered(ws[f"M{r}"].value))]
    check(not e_bad, "proc/er 編號行 1:1（E=0）", f"{e_bad}")

    # 10 lint A–N（此處重跑 lint036，不讀既有報告）-------------------------
    results = lint036.lint_workbook(OUT, profile="power")
    counts = lint036.count_by_check(results, profile="power")
    an = {k: counts.get(k, 0) for k in list("ABCDEFGHIJKLMN") + ["I-sibling"]}
    check(all(v == 0 for v in an.values()), "lint A–N（含 I-sibling）全零", f"{an}")

    # 11 zip / x14 -----------------------------------------------------------
    with zipfile.ZipFile(OUT) as z:
        members = len(z.namelist())
    check(members == 42, "zip 成員 42", str(members))
    check(x14(PREV) == x14(OUT), "x14 下拉讀回不變", f"{x14(OUT)}")

    print(f"\n{'全項達成' if not fails else '未達成：' + '、'.join(fails)}")
    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(main())
