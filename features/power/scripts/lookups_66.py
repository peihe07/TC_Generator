"""66 包 §H 第 4 步 —— §1 表中「執行層查」各項之逐一查證。

判準：R-P368 三段鏈（規格名 → LID → `MESSAGE.Signal` → forms DBC `SG_`）；
PROXI 參數走 R-P368(c)（PROXI `Format` 分頁 `Parameter Name`）；
HMI 設定條目走 R-P375(b)（`HMI Settings List` `Settings` 分頁）。

**命中即候選，非認定**（R-P375(d)）；查無者記「未解得（止於段 n）」，
**不得記查無**（R-G13 / R-P368(d)）。

用法：
    python features/power/scripts/lookups_66.py
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
OUT = ROOT / "features/power/data/lookups_66.md"

# §1 表所令查者
PROXI_PARAMS = [
    "Brand_Configuration_2", "SDARS_Presence", "Audio_Brand",
    "VC_VEH_BRAND", "VC_VEH_LINE", "VC_SpecialPKG", "Car_Shape_Configuration",
    "Number_of_Doors", "Rear_View_Camera", "Switch_Off_Time",
    "Ecall_Button_Variant", "TBM_Present", "Country_Code",
]
SPEC_SIGNALS = [
    "Themed_Sound", "Door_Ajar_Status", "VC_BODY_STYLE", "Radio_Theme",
    "ICSPowerButton", "Telematic_Power", "PowerMode",
]
HMI_SETTINGS = ["Welcome Onboard Sound", "Startup Animation Selection"]
DBC_PROBES = ["DIS_CENTERSTACK", "Chime", "Door_Ajar", "DCSD"]


def proxi_index() -> dict[str, list[int]]:
    wb = openpyxl.load_workbook(ROOT / "forms/PROXI_HDCC27_R3_20250424.xlsx",
                                data_only=True, read_only=True)
    idx: dict[str, list[int]] = {}
    for i, row in enumerate(wb["Format"].iter_rows(values_only=True), start=1):
        for v in row:
            if isinstance(v, str) and v.strip():
                idx.setdefault(v.strip().lower(), []).append(i)
    wb.close()
    return idx


def lid_index() -> dict[str, tuple[int, str]]:
    wb = openpyxl.load_workbook(
        ROOT / "forms/Logical Identifiers and CAN Mapping v1_78.xlsx",
        data_only=True, read_only=True)
    idx: dict[str, tuple[int, str]] = {}
    for i, row in enumerate(wb["CAN Mapping"].iter_rows(min_row=4, values_only=True),
                            start=4):
        lid = row[0]
        if isinstance(lid, str) and lid.strip():
            ah = row[25] if len(row) > 25 and isinstance(row[25], str) else ""
            idx[lid.strip().lower()] = (i, ah.replace("\n", " / "))
    wb.close()
    return idx


def hmi_index() -> dict[str, list[int]]:
    wb = openpyxl.load_workbook(
        ROOT / "forms/HMI Settings List R1 SR25 Post R1L-R (Feb 13 2026).xlsx",
        data_only=True, read_only=True)
    idx: dict[str, list[int]] = {}
    for i, row in enumerate(wb["Settings"].iter_rows(values_only=True), start=1):
        for v in row:
            if isinstance(v, str) and v.strip():
                idx.setdefault(v.strip().lower(), []).append(i)
    wb.close()
    return idx


def dbc_sgs() -> dict[str, set[str]]:
    out = {}
    for tag, p in (("BHCAN2", "forms/PDT27_E2A_R1_BHCAN2.dbc"),
                   ("FDCAN8", "forms/PDT27_E2A_R1_FDCAN8.dbc")):
        t = (ROOT / p).read_text(encoding="cp1252", errors="replace")
        cur, pairs = None, set()
        for line in t.splitlines():
            m = re.match(r"^BO_ \d+ (\w+):", line)
            if m:
                cur = m.group(1)
                continue
            m = re.match(r"^\s*SG_\s+(\w+)\s*:", line)
            if m and cur:
                pairs.add(f"{cur}.{m.group(1)}")
        out[tag] = pairs
    return out


def main() -> None:
    proxi, lid, hmi, sgs = proxi_index(), lid_index(), hmi_index(), dbc_sgs()
    allsg = sgs["BHCAN2"] | sgs["FDCAN8"]
    md = ["# 66 包 §H 第 4 步 —— 「執行層查」逐項結果", "",
          "> 判準：R-P368 三段鏈；PROXI 走 (c)；HMI 設定條目走 R-P375(b)。",
          "> **命中即候選，非認定**（R-P375(d)）；查無者記「未解得」，"
          "**不記查無**（R-G13 / R-P368(d)）。", ""]

    md += ["## 1. PROXI `Format` 參數（R-P368(c)）", "",
           "| 參數名 | `Format` 列 | 結果 |", "|---|---|---|"]
    for p in PROXI_PARAMS:
        rows = proxi.get(p.lower(), [])
        md.append(f"| `{p}` | {'、'.join(f'r{r}' for r in rows[:4]) or '—'} | "
                  f"{'**查得**' if rows else '**未解得**（`Format` 之 `Parameter Name` 無此名）'} |")

    md += ["", "## 2. 規格 `$X$` 之三段鏈（段 1 LID → 段 2 → 段 3 DBC）", "",
           "| 規格名 | 段 1（LID 列）| 段 2（`MESSAGE.Signal`）| 段 3 | 結果 |",
           "|---|---|---|---|---|"]
    for s in SPEC_SIGNALS:
        got = lid.get(s.lower())
        if not got:
            md.append(f"| `${s}$` | — | — | — | **未解得（止於段 1）** |")
            continue
        r, ah = got
        cands = [v.strip() for v in ah.split(" / ") if "." in v]
        ok = [c for c in cands if c in allsg]
        md.append(f"| `${s}$` | r{r} | {'、'.join(cands) or '—'} | "
                  f"{'、'.join(ok) or '—'} | "
                  f"{'**解得**' if ok else '**未解得（止於段 2）**'} |")

    md += ["", "## 3. HMI Settings List `Settings` 分頁", "",
           "| 條目 | `Settings` 列 | 結果 |", "|---|---|---|"]
    for h in HMI_SETTINGS:
        rows = hmi.get(h.lower(), [])
        md.append(f"| `{h}` | {'、'.join(f'r{r}' for r in rows[:4]) or '—'} | "
                  f"{'**查得**' if rows else '**未解得**'} |")

    md += ["", "## 4. DBC 訊息／訊號探查（§1 表所指者）", "",
           "| 探查 | BHCAN2 命中 | FDCAN8 命中 |", "|---|---|---|"]
    for d in DBC_PROBES:
        b = sorted(x for x in sgs["BHCAN2"] if d.lower() in x.lower())[:3]
        f = sorted(x for x in sgs["FDCAN8"] if d.lower() in x.lower())[:3]
        md.append(f"| `{d}` | {'、'.join(f'`{x}`' for x in b) or '**0**'} | "
                  f"{'、'.join(f'`{x}`' for x in f) or '**0**'} |")

    OUT.write_text("\n".join(md) + "\n")
    print("\n".join(md))
    print(f"\n→ {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
