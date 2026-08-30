"""62 包 B —— 十一名依 R-P375(a) 於 forms/ 全部參考檔重查。

R-P368 之段 1 原限於 LID `CAN Mapping`，R-P375 擴為 forms/ 全部參考檔。
`.Req` 類為 HMI 設定值、`.Info` 類致能狀態另走 UI / PROXI 路徑。

**命中即「候選」，非認定**（R-P375(d)）—— 逐筆記檔／分頁／列，
比對依據須載明（R-P368(b)），語意跳接仍不許。

用法：
    python features/power/scripts/forms_requery_62.py
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
OUT = ROOT / "features/power/data/forms_requery_62.tsv"

FILES = {
    "LID": ("forms/Logical Identifiers and CAN Mapping v1_78.xlsx", None),
    "HMI": ("forms/HMI Settings List R1 SR25 Post R1L-R (Feb 13 2026).xlsx", None),
    "PROXI": ("forms/PROXI_HDCC27_R3_20250424.xlsx", None),
    "SR26": ("forms/SR26 Default Settings and PNet ECU Configuration v1_0.xlsx", None),
    "SR24": ("forms/SR24 R1 Market Configuration Table v1.6.xlsx", None),
}

# 十一名（58 包 B4′ 之「未解得（止於段 1）」）
NAMES = [
    "Phone_Call.Info", "PhoneCall.Info", "Auto_SwitchOn_Setting.Req",
    "Antitheft_Activation.Req", "Antitheft_Result.Info", "RemStartFail",
    "SwitchOff_Timeout_Setting.Req", "SwitchOffSetting.Req",
    "Rear_Camera_Enable.Info", "Front_Panel_OnOff.Req",
    "Audio_Data_Exchange.Info",
]


def norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def key_tokens(name: str) -> list[str]:
    """規格原名去 `.Info` / `.Req` 後之詞素（R-P368(b) 允許之前後綴差異）。"""
    base = name.replace(".Info", "").replace(".Req", "")
    return [t for t in re.split(r"[_\s]+", base) if len(t) > 2]


def cells(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for j, v in enumerate(row, start=1):
                if isinstance(v, str) and v.strip():
                    yield sn, i, j, v
    wb.close()


def main() -> None:
    index = {}
    for tag, (rel, _) in FILES.items():
        index[tag] = list(cells(ROOT / rel))
        print(f"{tag:6s} 讀入 {len(index[tag]):>7d} 個非空字串格")

    rows = ["spec_name\tfile\tsheet\trow\tcol\tvalue\tmatched_tokens"]
    summary = {}
    for name in NAMES:
        toks = [t.lower() for t in key_tokens(name)]
        hits = []
        for tag in index:
            for sn, i, j, v in index[tag]:
                nv = norm(v)
                got = [t for t in toks if norm(t) in nv]
                # 全部詞素同格命中方計 —— 單一詞素（如 `Phone`）過寬
                if len(got) == len(toks) and toks:
                    hits.append((tag, sn, i, j, v.strip()[:70], "+".join(got)))
        # 去重：同檔同分頁同列只留第一格
        seen, uniq = set(), []
        for h in hits:
            k = (h[0], h[1], h[2])
            if k in seen:
                continue
            seen.add(k)
            uniq.append(h)
        summary[name] = len(uniq)
        for h in uniq[:12]:
            rows.append("\t".join([name, *map(str, h)]))
        print(f"\n### {name}  命中 {len(uniq)}")
        for h in uniq[:6]:
            print(f"   {h[0]:6s} {h[1][:26]:26s} r{h[2]:<5d} c{h[3]:<3d} {h[4]}")

    OUT.write_text("\n".join(rows) + "\n")
    print(f"\n→ {OUT.relative_to(ROOT)}")
    print("\n命中彙總：", {k: v for k, v in summary.items()})


if __name__ == "__main__":
    main()
