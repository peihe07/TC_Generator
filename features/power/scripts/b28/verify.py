#!/usr/bin/env python3
"""28 包驗收：逐項對 `pm_28.xlsx` 覆核，任一項不達成即 exit 1。

V2 通則於此**獨立重算**（不讀 `plan.json` 之 `I` 值），故 build 與 apply
之間若脫節、或 V2 規則實作有誤，此處會攤開。
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))

import lint036                                              # noqa: E402
from xlsx_surgical import build_shift                       # noqa: E402

HERE = Path(__file__).resolve().parent
B19 = ROOT / "features/power/sandbox/b19/pm_19.xlsx"
PREV = ROOT / "features/power/sandbox/b27/pm_27.xlsx"
OUT = ROOT / "features/power/sandbox/b28/pm_28.xlsx"
FIRST, LAST_COL = 10, 34
A_ROWS = {11, 12, 23, 179, 180}
RESTORED_SRC = 109
NUM = re.compile(r"^\s*\d+[.)]")
DENUM = re.compile(r"^\s*\d+[.)]\s*")
OBS_HEAD = re.compile(r"^Read\b")
OBS_MID = re.compile(r"check that", re.I)
TRIGGER_VERB = re.compile(
    r"^(Send|Select|Set|Wait|Let|Keep|Run|Place|End"
    r"|Bring|Power up|Reconnect|Open|Issue)\b")


def bodies(text) -> list[str]:
    return [DENUM.sub("", ln).strip()
            for ln in str(text or "").split("\n") if NUM.match(ln)]


def has_observation(body: str) -> bool:
    return bool(OBS_HEAD.match(body) or OBS_MID.search(body))


def lower_first(s: str) -> str:
    if not s:
        return s
    head = s.split(maxsplit=1)[0]
    if not (head[:1].isupper() and head[1:].isalpha() and head[1:].islower()):
        return s
    return s[:1].lower() + s[1:]


def bare(paren: str) -> str:
    """去掉消歧前綴（`<前綴> — `），還原為裸形態。"""
    inner = paren[1:-1]
    return f"({inner.split(' — ', 1)[-1]})" if " — " in inner else paren


def main() -> int:
    plan = json.loads((HERE / "plan.json").read_text(encoding="utf-8"))
    sheet = plan["sheet"]
    src = openpyxl.load_workbook(B19)[sheet]
    prev = openpyxl.load_workbook(PREV)[sheet]
    ws = openpyxl.load_workbook(OUT)[sheet]
    shift = build_shift({int(k): v for k, v in plan["insertions"].items()})
    setup_of = {a["row"]: a["setup"] for a in plan["audit_b"]}
    fails: list[str] = []

    def check(ok: bool, label: str, detail: str = "") -> None:
        print(f"  {'達成' if ok else '未達成'}  {label}"
              + (f" — {detail}" if detail else ""))
        if not ok:
            fails.append(label)

    def paren_of(row: int, book=None) -> str:
        book = book or ws
        return "\n".join(lint036.paren_lines(str(book[f"I{row}"].value or "")))

    # --- 相異範圍 -----------------------------------------------------------
    check(ws.max_row == prev.max_row, "列數與 pm_27 相同",
          f"{prev.max_row} → {ws.max_row}")
    off_field, diff_rows = [], set()
    for row in range(1, ws.max_row + 1):
        for col in range(1, LAST_COL + 1):
            if (ws.cell(row, col).value or "") == (prev.cell(row, col).value or ""):
                continue
            diff_rows.add(row)
            letter = openpyxl.utils.get_column_letter(col)
            if letter != "I":
                off_field.append((row, letter))
    check(not off_field, "相異僅限 test_item（I）欄", f"例外 {off_field[:5]}")

    # --- 不該動者零變動 -----------------------------------------------------
    a_rows = [shift(r) + i for r in sorted(A_ROWS)
              for i in range(len(next(s["variants"] for s in plan["splits"]
                                      if s["src_row"] == r)))]
    check(all(ws[f"I{r}"].value == prev[f"I{r}"].value for r in a_rows),
          "A 型 14 列括號零變動", f"{len(a_rows)} 列")
    restored = shift(RESTORED_SRC)
    check(ws[f"I{restored}"].value == src[f"I{RESTORED_SRC}"].value,
          f"§八-1 row {restored} 仍與 b19 原列逐字同")

    # --- V2 通則獨立重算 ----------------------------------------------------
    bad, tracks, over20, merged_same = [], Counter(), [], 0
    for split in plan["splits"]:
        s_row = split["src_row"]
        if s_row in A_ROWS:
            continue
        setup = setup_of[s_row]
        anchor = shift(s_row)
        src_proc = bodies(src[f"L{s_row}"].value)
        src_er = bodies(src[f"M{s_row}"].value)
        if len(split["variants"]) == 1:            # §八-1 還原列，不套通則
            tracks["原列逐字"] += 1
            continue
        for offset, variant in enumerate(split["variants"]):
            row = anchor + offset
            steps = bodies(variant["L"])[setup:]
            observe = steps[-1]
            er = src_er[src_proc.index(observe)]
            own = [s for s in steps[:-1] if TRIGGER_VERB.match(s)]
            lead = [s for s in src_proc[:setup] if TRIGGER_VERB.match(s)]
            trigger = own or ([lead[-1]] if lead else None)
            want_t = (f"({' / '.join(lower_first(d) for d in trigger)} "
                      f"-> {er})") if trigger else None
            want_o = f"({lower_first(observe)} -> {er})"
            # trigger 式若含 lint check C 之 hedge 詞即落回規則 3（§四 要求
            # A–N 全零；`Successfully` 為訊號值但 C 對括號不設引號豁免）。
            if want_t and lint036.RE_C.search(want_t):
                want_t = None
            got = bare(paren_of(row))
            if got == want_t:
                tracks["trigger -> ER"] += 1
            elif got == want_o:
                tracks["觀察步 -> ER（規則 3／同列退路）"] += 1
            else:
                bad.append((row, got, want_t or want_o))
            if len(got.split()) > 20:
                over20.append(row)
            if len(steps) > 1 and paren_of(row) == paren_of(row, prev):
                merged_same += 1
    check(not bad, "全本 B 型面向列括號合 V2 通則（獨立重算）",
          f"例外 {bad[:3]}")
    check(merged_same == 16, "16 列縮併列括號與 pm_27 逐字同（形態 B 為 V2 特例）",
          f"{merged_same}/16")

    # --- 收斂達成：不得再有 ER-only 括號 ------------------------------------
    er_only = [r for r in range(FIRST, ws.max_row + 1)
               if (pa := paren_of(r)) and "->" not in pa
               and r != restored]
    check(not er_only, "全本無 ER-only 括號（四軌收斂）",
          f"例外 {er_only[:8]}")
    print(f"  （參考）軌數分布 {dict(tracks)}；逾 20 詞 {len(over20)} 列"
          f" —— 依延伸判斷免上限")

    # --- 不變式 -------------------------------------------------------------
    ids = [str(ws[f"F{r}"].value or "").strip()
           for r in range(FIRST, ws.max_row + 1)]
    nums = [int(i.rsplit("-", 1)[1]) for i in ids if i]
    check(nums == list(range(1, len(nums) + 1)), "Test Case ID 連續無跳號",
          f"001–{max(nums):03d}／{len(nums)} 列")
    check([ws[f"B{r}"].value for r in range(FIRST, ws.max_row + 1)]
          == list(range(1, ws.max_row - FIRST + 2)), "No.# 連續重編")
    e_bad = [r for r in range(FIRST, ws.max_row + 1)
             if len(bodies(ws[f"L{r}"].value)) != len(bodies(ws[f"M{r}"].value))]
    check(not e_bad, "proc↔er 編號數逐列相等", f"例外 {e_bad[:5]}")
    no_obs = [r for r in range(FIRST, ws.max_row + 1)
              if bodies(ws[f"L{r}"].value)
              and not any(has_observation(b) for b in bodies(ws[f"L{r}"].value))]
    check(not no_obs, "全本無「無觀察步」之列", f"例外 {no_obs[:8]}")
    groups: dict[tuple[str, str], list[int]] = defaultdict(list)
    for row in range(FIRST, ws.max_row + 1):
        req = str(ws[f"D{row}"].value or "").strip()
        if req and (pa := paren_of(row)):
            groups[(req, pa)].append(row)
    dups = {k: v for k, v in groups.items() if len(v) > 1}
    check(not dups, "括號下半同源不逐字相同", f"{list(dups.values())[:3]}")

    print(f"\n{'全項達成' if not fails else f'未達成 {len(fails)} 項'}")
    return 0 if not fails else 1


if __name__ == "__main__":
    sys.exit(main())
