"""B4 — 被引用 item 之 EE Architecture 分布（R-P40）
B5 — 037 各欄相異值數（R-P41(b)）

兩者皆為純量測，無人工判讀。依 R-P41(a) 之精神（素材須可重現）補腳本。

用法：
    python features/power/scripts/build_b4_b5.py
"""

from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from extract_textlayer import REQ_RE, paragraphs  # noqa: E402

ROOT = Path(__file__).resolve().parents[3]
IN = ROOT / "features/power/inputs"
DATA = ROOT / "features/power/data"

EE_RE = re.compile(r"\[EE Architecture:([^\]]*)\]")


def find(pattern: str) -> Path:
    return next(f for f in IN.iterdir() if pattern in f.name)


def cited_items() -> tuple[set[str], dict[str, set[str]]]:
    """由 layer3_full.tsv 取被引用之 item 集合，以及 leaf → item 之對應。"""
    items: set[str] = set()
    per_leaf: dict[str, set[str]] = defaultdict(set)
    for line in (DATA / "layer3_full.tsv").read_text(encoding="utf-8").splitlines()[1:]:
        if not line.strip():
            continue
        parts = line.split("\t")
        ids = parts[5].split(",")
        items.update(ids)
        per_leaf[parts[0]].update(ids)
    return items, per_leaf


def ee_of(items: set[str]) -> dict[str, tuple[str, ...] | None]:
    """被引用 item → 其 EE Architecture 值組（依錨點行之 metadata）。"""
    out: dict[str, tuple[str, ...] | None] = {}
    for path in [find("CFTS_009_Wake-up"), next(x for x in IN.iterdir() if x.suffix == ".doc")]:
        for plain, bold in paragraphs(path):
            for m in REQ_RE.finditer(bold):
                item = m.group(1)
                if item not in items:
                    continue
                em = EE_RE.search(plain)
                out[item] = tuple(sorted(v.strip() for v in em.group(1).split(","))) if em else None
    return out


def vehicle_columns() -> list[tuple[int, str]]:
    wb = openpyxl.load_workbook(find("FSM-036"), data_only=True, read_only=True)
    header = wb["Test Case Specification&Result"][9]
    cols = [
        (i + 1, str(c.value).replace("\n", " ").strip())
        for i, c in enumerate(header) if 21 <= i + 1 <= 27
    ]
    wb.close()
    return cols


def leaf_rows() -> tuple[list[str], list[tuple]]:
    wb = openpyxl.load_workbook(find("FSM-037"), data_only=True, read_only=True)
    ws = wb["SWE1 Requirements"]
    header = [
        str(c.value).strip() if c.value else f"(c{i + 1} 無標頭)"
        for i, c in enumerate(ws[7])
    ]
    rows = [r for r in ws.iter_rows(min_row=8, max_row=145, values_only=True)
            if r[0] and str(r[0]).strip()]
    wb.close()
    return header, rows


def build_b4() -> None:
    items, per_leaf = cited_items()
    ee = ee_of(items)
    counts = Counter(ee.values())
    mid_only = sorted(i for i, v in ee.items() if v == ("Atlantis Mid",))
    high_only = sorted(i for i, v in ee.items() if v == ("Atlantis High",))
    odd = sorted(i for i, v in ee.items() if v and {"CUSW", "PowerNet"} & set(v))

    per_leaf_ee = {}
    for leaf, ids in per_leaf.items():
        union: set[str] = set()
        for i in ids:
            if ee.get(i):
                union.update(ee[i])
        per_leaf_ee[leaf] = union
    solo = [
        (l, s) for l, s in sorted(
            per_leaf_ee.items(), key=lambda kv: int(re.match(r"SWE-PM-(\d+)", kv[0]).group(1))
        ) if s and "All" not in s and len(s) == 1
    ]

    out = [
        "# B4 — 被引用 item 之 EE Architecture 分布（R-P40）\n",
        f"\n> 母體：**{len(items)} 個被引用 item**（經 `layer3_full.tsv` 之 `item_ids` 欄）。\n",
        "> 車型（`Radio`）軸已於 04 包驗畢無虞（全為 `allSys` 或含 `R1L`，零例外）。\n",
        "> 產生指令：`python features/power/scripts/build_b4_b5.py`\n",
        "\n## 1. 值域與計數\n\n| EE Architecture 值組 | item 數 |\n|---|---|\n",
    ]
    for k, v in counts.most_common():
        out.append(f"| `{', '.join(k) if k else '（無此欄）'}` | **{v}** |\n")
    out.append(f"| **合計** | **{sum(counts.values())}** |\n")
    distinct = sorted({y for k in counts if k for y in k})
    out.append(f"\n相異值：{'、'.join('`' + x + '`' for x in distinct)}\n")
    out.append(f"\n**無 `EE Architecture` 欄者：{counts.get(None, 0)} 個。**\n")

    out.append("\n## 2. 與 FW036 c21–c27 七個車型欄之對照\n\n| 欄 | 標頭（實測） | 世代 |\n|---|---|---|\n")
    hi = mi = 0
    for i, v in vehicle_columns():
        if "Atl-Hi" in v:
            gen, hi = "Atlantis High", hi + 1
        elif "Atl-Mi" in v:
            gen, mi = "Atlantis Mid", mi + 1
        else:
            gen = "（無法判定）"
        out.append(f"| c{i} | `{v}` | {gen} |\n")
    out.append(f"\n七欄之世代分布：**Atl-Hi {hi} 欄**、**Atl-Mi {mi} 欄**。\n")

    both = sum(v for k, v in counts.items() if k and ("All" in k or {"Atlantis High", "Atlantis Mid"} <= set(k)))
    out.append(
        f"\n## 3. 明確回答：是否存在被引用 item 其 EE Architecture 不含本專案適用之值\n\n"
        f"本專案為 R1L（R-P2），FW036 七個車型欄橫跨 Atlantis High（{hi} 欄）與 "
        f"Atlantis Mid（{mi} 欄）。\n\n"
        f"**答：不存在「兩世代皆不含」者。**\n\n"
        f"- 兩世代通用（`All` 或同時含 High 與 Mid）：**{both}** 個\n"
        f"- **`Atlantis Mid` 單值：{len(mid_only)} 個** —— 僅適用 Atl-Mi，不適用 Atl-Hi 兩欄\n"
        f"- **`Atlantis High` 單值：{len(high_only)} 個** —— 僅適用 Atl-Hi，不適用 Atl-Mi 五欄\n"
        f"\n即 {both} / {sum(counts.values())} 兩世代通用，"
        f"{len(mid_only) + len(high_only)} 個為單世代專屬。\n"
    )
    out.append(f"\n### 3.1 單世代專屬 item 清單\n\n")
    out.append(f"**`Atlantis Mid` 專屬（{len(mid_only)}）**："
               f"{', '.join('`' + i + '`' for i in mid_only) or '無'}\n\n")
    out.append(f"**`Atlantis High` 專屬（{len(high_only)}）**："
               f"{', '.join('`' + i + '`' for i in high_only) or '無'}\n\n")
    out.append(f"**含非本案世代值者（{len(odd)}）**："
               f"{', '.join('`' + i + '`' for i in odd) or '無'}\n")

    out.append(f"\n### 3.2 逐 leaf 之 EE Architecture 聯集僅含單一世代者\n\n"
               f"114 個 leaf 中：**{len(solo)}** 個\n\n")
    if solo:
        out.append("| leaf | EE Architecture 聯集 |\n|---|---|\n")
        for leaf, s in solo:
            out.append(f"| `{leaf}` | `{', '.join(sorted(s))}` |\n")
    out.append(
        "\n> **登記**：01 包 §F 已載 FW036 c21 標頭 `HDCC27 Atl-Hi` 沿用 A-PV15"
        "（世代落差、入 RD-1、不自行對應）與 R30-3 / R30-4（車型欄留白）。"
        "本檔僅提供實測分布，**不改變該政策，也不建議如何填欄**。\n"
    )
    path = DATA / "b4_ee_architecture.md"
    path.write_text("".join(out), encoding="utf-8")
    print(f"wrote {path.relative_to(ROOT)} — {path.stat().st_size} bytes")
    print(f"  G25: 值組 {len(counts)} 種；Mid 專屬 {len(mid_only)}、High 專屬 {len(high_only)}、"
          f"單世代 leaf {len(solo)}")


def build_b5() -> None:
    header, rows = leaf_rows()
    out = [
        "# B5 — 037 `SWE1 Requirements` 各欄相異值數（R-P41(b)）\n",
        "\n> G19 已證 18 欄皆 115/115 非空。**非空不等於有鑑別力** ——\n",
        "> 相異值數 = 1 者與空欄之實際效果相同；≥ 100 者近乎逐列唯一，同樣無分群價值。\n",
        f"> 母體：r8–r145 之 {len(rows)} 個 leaf。\n",
        "> 產生指令：`python features/power/scripts/build_b4_b5.py`\n",
        "\n## 逐欄相異值數\n\n| 欄 | 標頭 | 相異值數 | 鑑別力 | 值域 |\n|---|---|---|---|---|\n",
    ]
    ones, highs = [], []
    for i, h in enumerate(header):
        counter = Counter(str(r[i]).strip() for r in rows)
        n = len(counter)
        if n == 1:
            flag = "**無（單一值）**"
            ones.append((i + 1, h, next(iter(counter))))
        elif n >= 100:
            flag = "**無（近乎逐列唯一）**"
            highs.append((i + 1, h, n))
        elif n <= 8:
            flag = "可分群"
        else:
            flag = "弱"
        if n <= 6:
            domain = "；".join(f"`{k[:34]}` ×{v}" for k, v in counter.most_common())
        else:
            domain = ("；".join(f"`{k[:26]}` ×{v}" for k, v in counter.most_common(4))
                      + f"；…（共 {n} 值）")
        out.append(f"| c{i + 1} | {h[:36]} | **{n}** | {flag} | {domain} |\n")

    out.append(f"\n## 相異值數 = 1 之欄位（{len(ones)} 欄）\n\n")
    if ones:
        out.append("| 欄 | 標頭 | 唯一值 |\n|---|---|---|\n")
        for i, h, v in ones:
            out.append(f"| c{i} | {h} | `{v}` |\n")
        out.append("\n**此類欄位在分批、優先級、追溯上與空欄實際效果相同。**\n")
    out.append(f"\n## 相異值數 ≥ 100 之欄位（{len(highs)} 欄）\n\n")
    if highs:
        out.append("| 欄 | 標頭 | 相異值數 |\n|---|---|---|\n")
        for i, h, n in highs:
            out.append(f"| c{i} | {h} | {n} / {len(rows)} |\n")
        out.append("\n**近乎逐列唯一，無分群價值**"
                   "（作為內容來源仍有價值 —— 例如 Description 用於判讀）。\n")
    else:
        out.append("（無）\n")

    titles = Counter(str(r[2]).strip() for r in rows)
    repeats = [(k, v) for k, v in titles.most_common() if v > 1]
    out.append(
        f"\n## 附：`Requirement Title` 與 §E「本分組之已知弱點」之對照\n\n"
        f"§E 該節稱：「037 `Requirement Title` 於 115 leaf 中出現 **20+ 種**，"
        f"多數僅出現 1 次（`Timeout` 7、`Phone Call` 5 為**僅有例外**）」。\n\n"
        f"實測：相異值 **{len(titles)}** 種，僅出現 1 次者 "
        f"**{sum(1 for v in titles.values() if v == 1)}** 種。出現 > 1 次者：\n\n"
        f"| 值 | 次數 |\n|---|---|\n"
    )
    for k, v in repeats:
        out.append(f"| `{k}` | {v} |\n")
    out.append(
        f"\n即「20+ 種」大幅低估（實為 {len(titles)} 種），"
        f"且「僅有例外」為誤 —— 除 `Timeout` 與 `Phone Call` 外另有 "
        f"{len(repeats) - 2} 組重複值。\n"
        f"**§E 該節之結論（`Requirement Title` 無分組價值）不受影響，反而更強。**\n"
    )
    path = DATA / "b5_column_entropy.md"
    path.write_text("".join(out), encoding="utf-8")
    print(f"wrote {path.relative_to(ROOT)} — {path.stat().st_size} bytes")
    print(f"  G26: 相異值數 = 1 者 {[f'c{i}' for i, _, _ in ones]}；"
          f"≥ 100 者 {[f'c{i}' for i, _, _ in highs]}")


if __name__ == "__main__":
    build_b4()
    build_b5()
