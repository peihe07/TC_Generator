#!/usr/bin/env python3
"""驗收：lint036 全項 + 檢查 P（三件組殘留）+ x14 下拉讀回。"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "scripts"))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import lint036                                     # noqa: E402
import edits as E                                  # noqa: E402

SANDBOX = ROOT / "features/power/sandbox/b02"


def read_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next(s for s in wb.sheetnames if s.startswith(lint036.TC_SHEET_PREFIX))
    ws = wb[sheet]
    header = lint036.find_header_row(ws)
    raw = list(ws.iter_rows(min_row=1, values_only=True))
    columns = lint036.build_column_map(list(raw[header - 1]))
    out = []
    for offset, row in enumerate(raw[header:], start=header + 1):
        row = row or ()
        fields = {k: (lint036.cell_text(row[i]) if i < len(row) else "")
                  for k, i in columns.items()}
        if not any(fields[k].strip() for k in ("test_item", "proc", "er")):
            continue
        fields["row"] = offset
        out.append(fields)
    wb.close()
    return out


def check_p(rows: list[dict]) -> list[tuple[int, str, str]]:
    """P：舊式 CAN 記法之殘留（R-1，範圍依 R-6）。

    施用於作者生成之內容 —— 四欄 ＋ test_item 括號下半；
    test_item 上半之 verbatim 保留來源原文，排除在外。
    lint036 已內建同名檢查，此處保留為獨立第二實作以資交叉核對。
    """
    hits = []
    for r in rows:
        for field in lint036.P_FIELDS:
            for m in lint036.RE_P_LEGACY_CAN.finditer(r[field]):
                hits.append((r["row"], field, m.group(0)))
        for line in lint036.paren_lines(r["test_item"]):
            for m in lint036.RE_P_LEGACY_CAN.finditer(line):
                hits.append((r["row"], "test_item(括號下半)", m.group(0)))
    return hits


def counts(path: Path) -> dict[str, int]:
    return lint036.count_by_check(lint036.lint_workbook(path))


def main() -> None:
    src, out = SANDBOX / "pm_work.xlsx", SANDBOX / "pm_remediated.xlsx"
    before, after = counts(src), counts(out)
    print(f"{'檢查':<10}{'前':>7}{'後':>7}   判定")
    for key in lint036.CHECK_ORDER:
        mark = "" if before[key] == after[key] else "  <-- 變動"
        print(f"{key:<10}{before[key]:>7}{after[key]:>7}{mark}")

    p_before, p_after = check_p(read_rows(src)), check_p(read_rows(out))
    print(f"\nP（舊式記法殘留，R-6 範圍）  前 {len(p_before)}  後 {len(p_after)}")
    for row, field, token in p_after[:10]:
        print(f"   殘留 row {row} {field}: {token}")

    # x14 下拉讀回
    import zipfile
    for label, path in (("src", src), ("out", out)):
        with zipfile.ZipFile(path) as z:
            x14 = sum(len(lint036.re.findall(r"<x14:dataValidation[ >]",
                                             z.read(m).decode("utf-8")))
                      for m in z.namelist() if m.startswith("xl/worksheets/sheet"))
        print(f"x14 dataValidation ({label}): {x14}")


if __name__ == "__main__":
    main()
