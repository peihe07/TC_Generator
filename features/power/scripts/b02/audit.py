#!/usr/bin/env python3
"""抽驗與零變動證明（下放包 02 驗收要求）。"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "scripts"))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import lint036                                     # noqa: E402
import edits as E                                  # noqa: E402

SANDBOX = ROOT / "features/power/sandbox/b02"
TARGET_COLUMNS = {"I", "J", "K", "L", "M"}          # test_item/pre/input/proc/er


def all_cells(path: Path) -> dict[tuple[str, str], str]:
    """全工作簿每一格之值，供零變動證明。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    for name in wb.sheetnames:
        for row in wb[name].iter_rows():
            for cell in row:
                if cell.value is not None:
                    out[(name, cell.coordinate)] = str(cell.value)
    wb.close()
    return out


def main() -> None:
    src, out = SANDBOX / "pm_work.xlsx", SANDBOX / "pm_remediated.xlsx"
    before, after = all_cells(src), all_cells(out)

    changed = {k for k in before | after.keys()
               if before.get(k) != after.get(k)}
    columns = {"".join(c for c in coord if c.isalpha()) for _, coord in changed}
    sheets = {sheet for sheet, _ in changed}
    print(f"變動格數 {len(changed)}｜涉及 sheet {sheets}｜涉及欄 {sorted(columns)}")
    print(f"非目標欄變動：{sorted(columns - TARGET_COLUMNS) or '無'}")
    print(f"新增/刪除之格：{len(set(after) ^ set(before))}")

    payload = json.loads((SANDBOX / "edits.json").read_text(encoding="utf-8"))
    planned = {(int(r), c) for r, fields in payload["changes"].items()
               for c in fields}
    print(f"計畫變動 {len(planned)} 格｜實際 {len(changed)} 格｜"
          f"相符：{len(planned) == len(changed)}")


if __name__ == "__main__":
    main()
