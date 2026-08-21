#!/usr/bin/env python3
"""將 edits.json 寫回 PM 工作副本。

寫入路徑僅 `backend/xlsx_surgical.surgical_save()`（R16 / R-G3）；
全域不呼叫 `Workbook.save()`。來源交付檔全程未觸碰。
"""

from __future__ import annotations

import hashlib
import json
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))

import lint036                                     # noqa: E402
from xlsx_surgical import surgical_save            # noqa: E402

SANDBOX = ROOT / "features/power/sandbox/b02"
SRC = SANDBOX / "pm_work.xlsx"
OUT = SANDBOX / "pm_remediated.xlsx"

# lint036 之欄位鍵 → 工作簿欄字母（0-based 索引 8..12）
COLUMN_LETTER = {"test_item": "I", "pre": "J", "input": "K",
                 "proc": "L", "er": "M"}


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def main() -> None:
    payload = json.loads((SANDBOX / "edits.json").read_text(encoding="utf-8"))
    changes = {int(k): v for k, v in payload["changes"].items()}

    work = SANDBOX / ".work.xlsx"
    shutil.copyfile(SRC, work)
    assert sha256(work) == sha256(SRC), "工作副本非位元組相同"

    wb = openpyxl.load_workbook(work)
    sheet = next(s for s in wb.sheetnames
                 if s.startswith(lint036.TC_SHEET_PREFIX))
    ws = wb[sheet]

    written = 0
    for row, fields in changes.items():
        for key, value in fields.items():
            ws[f"{COLUMN_LETTER[key]}{row}"] = value
            written += 1

    report = surgical_save(wb, work, OUT)           # 不呼叫 wb.save()
    work.unlink()

    print(f"寫入 {written} 格／{len(changes)} 列 → {OUT.name}")
    print("patched sheets:", report["sheets_patched"])
    print("zip members:", report["members"], "differing:", report["differing"])
    print("data-validation counts:", report["dv_counts"])
    print("src sha256:", sha256(SRC))
    print("out sha256:", sha256(OUT))


if __name__ == "__main__":
    main()
