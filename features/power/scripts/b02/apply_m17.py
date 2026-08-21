#!/usr/bin/env python3
"""M17-PM 寫回：以批 1 工作副本為基底，只改 proc 欄 20 格。

寫入路徑僅 `surgical_save()`；不呼叫 `Workbook.save()`。
"""

from __future__ import annotations

import hashlib
import json
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import lint036                                     # noqa: E402
import m17_edits as M                              # noqa: E402
import verify as V                                 # noqa: E402
from xlsx_surgical import surgical_save            # noqa: E402

SANDBOX = ROOT / "features/power/sandbox/b02"
SRC = SANDBOX / "pm_remediated.xlsx"               # 批 1 之工作副本
OUT = SANDBOX / "pm_batch2.xlsx"
PROC_COLUMN = "L"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def main() -> None:
    rows = {r["row"]: r for r in V.read_rows(SRC)}
    changes = {row: M.rewrite_proc(rows[row]["proc"], row)
               for row in sorted(M.REPLACEMENTS)}

    work = SANDBOX / ".work_m17.xlsx"
    shutil.copyfile(SRC, work)
    assert sha256(work) == sha256(SRC), "工作副本非位元組相同"

    wb = openpyxl.load_workbook(work)
    ws = wb[next(s for s in wb.sheetnames
                 if s.startswith(lint036.TC_SHEET_PREFIX))]
    for row, value in changes.items():
        ws[f"{PROC_COLUMN}{row}"] = value

    report = surgical_save(wb, work, OUT)
    work.unlink()

    (SANDBOX / "m17_edits.json").write_text(
        json.dumps({"changes": {str(k): v for k, v in changes.items()},
                    "divergence": {str(k): v
                                   for k, v in M.TABLE_DIVERGENCE.items()}},
                   ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"寫入 {len(changes)} 格（proc 欄）→ {OUT.name}")
    print("patched:", report["sheets_patched"], "| members:", report["members"],
          "| differing:", report["differing"])
    print("dv counts:", report["dv_counts"])
    print("base sha256:", sha256(SRC))
    print("out  sha256:", sha256(OUT))


if __name__ == "__main__":
    main()
