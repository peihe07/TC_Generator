#!/usr/bin/env python3
"""19 包寫回：row 72 更正回復 ＋ 軌 C 補 rows 271–274 ＋ row 291 標 PENDING。

- §一 撤銷 17 包 §五之代入：`PowerModeSts_Telematic` 為 CFTS009-4941562
  之規格原文訊號名，非筆誤。依 **R-13** 保留原文名稱（DBC 無對應，
  比照 R-1 v3(d) 不加 `$`），另開 DR-PW21 向上游查其 message 與 VAL_。
- §二 補寫附件 G／H 所遺漏之 rows 271–274（`SWE-PM-104`）。
- §三-1 row 291 之二擇一改標 `PENDING: DR-PW22`，其餘步驟不動。

改動範圍限 `pre`／`input`／`proc`／`er` 四欄；
`test_item`／`spec_reference` 零變動。
`surgical_save` 唯一路徑；交付本唯讀不覆寫。
"""

from __future__ import annotations

import hashlib
import json
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))

import lint036                                         # noqa: E402
from xlsx_surgical import surgical_save                # noqa: E402

SANDBOX = ROOT / "features/power/sandbox/b19"
SRC = ROOT / "features/power/sandbox/b18/pm_18.xlsx"
OUT = SANDBOX / "pm_19.xlsx"
COLUMN = {"pre": "J", "input": "K", "proc": "L", "er": "M"}


def sha256(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def main() -> None:
    edits = {int(k): v for k, v in
             json.loads((SANDBOX / "edits.json").read_text(encoding="utf-8")).items()}
    work = SANDBOX / ".work.xlsx"
    shutil.copyfile(SRC, work)
    assert sha256(work) == sha256(SRC)

    wb = openpyxl.load_workbook(work)
    ws = wb[next(s for s in wb.sheetnames
                 if s.startswith(lint036.TC_SHEET_PREFIX))]
    n = 0
    for row, fields in sorted(edits.items()):
        for key, value in fields.items():
            ws[f"{COLUMN[key]}{row}"] = value
            n += 1

    report = surgical_save(wb, work, OUT)
    work.unlink()

    print(f"寫入 {n} 格／{len(edits)} 列 → {OUT.name}")
    print("members:", report["members"], "| differing:", report["differing"])
    print("dv:", report["dv_counts"])
    print("out sha256:", sha256(OUT)[:20])


if __name__ == "__main__":
    main()
