#!/usr/bin/env python3
"""25 包 §四 驗收：逐項對 `pm_25.xlsx` 覆核，任一項不達成即 exit 1。

覆核對象為輸出本身（非 plan），因此 plan 與寫入之間若脫節，此處會攤開。
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))

import lint036                                              # noqa: E402
from xlsx_surgical import build_shift                       # noqa: E402

HERE = Path(__file__).resolve().parent
BASE = ROOT / "features/power/sandbox/b19/pm_19.xlsx"
OUT = ROOT / "features/power/sandbox/b25/pm_25.xlsx"
FIRST = 10
NUM = re.compile(r"^\s*\d+[.)]")
DENUM = re.compile(r"^\s*\d+[.)]\s*")
A_ROWS = {11, 12, 23, 179, 180}
# 179b／180b 依下放包 §一 於 PRE 增一行；其餘面向列 PRE 須與原列逐字同。
PRE_EXTRA = "An incoming phone call is active on the HU"


def steps(text) -> list[str]:
    return [ln for ln in str(text or "").split("\n") if NUM.match(ln)]


def bodies(text) -> list[str]:
    return [DENUM.sub("", ln).strip() for ln in steps(text)]


def main() -> int:
    plan = json.loads((HERE / "plan.json").read_text(encoding="utf-8"))
    sheet = plan["sheet"]
    src = openpyxl.load_workbook(BASE)[sheet]
    ws = openpyxl.load_workbook(OUT)[sheet]
    insertions = {int(k): v for k, v in plan["insertions"].items()}
    shift = build_shift(insertions)
    fails: list[str] = []

    def check(ok: bool, label: str, detail: str = "") -> None:
        print(f"  {'達成' if ok else '未達成'}  {label}"
              + (f" — {detail}" if detail else ""))
        if not ok:
            fails.append(label)

    # --- 列數 ---------------------------------------------------------------
    src_rows = src.max_row - FIRST + 1
    out_rows = ws.max_row - FIRST + 1
    inserted = sum(insertions.values())
    check(out_rows == src_rows + inserted, "全本列數",
          f"{src_rows} + {inserted} = {out_rows}")

    # --- 面向列逐列存在，原列為首面向 ---------------------------------------
    a_seen = b_seen = 0
    for split in plan["splits"]:
        anchor = shift(split["src_row"])
        for offset, variant in enumerate(split["variants"]):
            row = anchor + offset
            for field in ("I", "J", "L", "M"):
                if str(ws[f"{field}{row}"].value or "") != variant[field]:
                    fails.append(f"面向列內容 row {row} 欄 {field}")
            if split["src_row"] in A_ROWS:
                a_seen += 1
            else:
                b_seen += 1
    check(a_seen == 14, "A 型面向列", f"{a_seen}/14")
    check(b_seen == 144, "B 型面向列", f"{b_seen}/144")
    check(not [f for f in fails if f.startswith("面向列內容")],
          "面向列四欄與 plan 逐字相符")

    # --- PRE 逐字、setup 段逐字 ---------------------------------------------
    pre_bad, setup_bad = [], []
    for split in plan["splits"]:
        s_row = split["src_row"]
        anchor = shift(s_row)
        src_pre = bodies(src[f"J{s_row}"].value)
        src_proc = bodies(src[f"L{s_row}"].value)
        for offset in range(len(split["variants"])):
            row = anchor + offset
            out_pre = bodies(ws[f"J{row}"].value)
            allowed = (src_pre, [*src_pre[:-1], PRE_EXTRA, src_pre[-1]])
            if out_pre not in allowed:
                pre_bad.append(row)
            if s_row not in A_ROWS:
                # setup 段 = 面向列 PROC 除末步；須為原列前綴。
                head = bodies(ws[f"L{row}"].value)[:-1]
                if head != src_proc[:len(head)]:
                    setup_bad.append(row)
    check(not pre_bad, "面向列 PRE 與原列逐字同（179b／180b 增一行）",
          f"例外 {pre_bad[:5]}")
    check(not setup_bad, "B 型 setup 段與原列逐字同", f"例外 {setup_bad[:5]}")

    # --- ID 連續無跳號無重複 -------------------------------------------------
    ids, stubs, nos = [], [], []
    for row in range(FIRST, ws.max_row + 1):
        nos.append(ws[f"B{row}"].value)
        tc = str(ws[f"F{row}"].value or "").strip()
        (ids if tc else stubs).append(row if not tc else tc)
    nums = [int(i.rsplit("-", 1)[1]) for i in ids]
    check(nums == list(range(1, len(nums) + 1)), "Test Case ID 連續無跳號",
          f"001–{max(nums):03d}／{len(nums)} 列")
    check(len(set(ids)) == len(ids), "Test Case ID 無重複")
    check(nos == list(range(1, out_rows + 1)), "No.# 連續重編",
          f"1–{out_rows}")
    check(stubs == [shift(230)], "存根列略過不給 ID", f"{stubs}")

    # --- proc↔er 編號數逐列相等 ---------------------------------------------
    e_bad = [row for row in range(FIRST, ws.max_row + 1)
             if len(steps(ws[f"L{row}"].value)) !=
             len(steps(ws[f"M{row}"].value))]
    check(not e_bad, "proc↔er 編號數逐列相等", f"例外 {e_bad[:5]}")

    # --- 括號下半：同 Requirement ID 下不得逐字相同 -------------------------
    groups: dict[tuple[str, str], list[int]] = defaultdict(list)
    for row in range(FIRST, ws.max_row + 1):
        req = str(ws[f"D{row}"].value or "").strip()
        paren = "\n".join(lint036.paren_lines(str(ws[f"I{row}"].value or "")))
        if req and paren:
            groups[(req, paren)].append(row)
    dups = {k: v for k, v in groups.items() if len(v) > 1}
    check(not dups, "括號下半同源不逐字相同", f"{list(dups.values())[:3]}")

    # --- 未拆列零變動 -------------------------------------------------------
    split_rows = {s["src_row"] for s in plan["splits"]}
    drift = []
    for s_row in range(FIRST, src.max_row + 1):
        if s_row in split_rows:
            continue
        row = shift(s_row)
        for field in ("I", "J", "K", "L", "M", "N", "P", "R"):
            if (src[f"{field}{s_row}"].value or "") != \
                    (ws[f"{field}{row}"].value or ""):
                drift.append((s_row, row, field))
    check(not drift, "未拆原列非 ID 欄零變動", f"例外 {drift[:5]}")

    print(f"\n{'全項達成' if not fails else f'未達成 {len(fails)} 項'}")
    return 0 if not fails else 1


if __name__ == "__main__":
    sys.exit(main())
