"""03 包 §D 全表自驗（G0–G16）。

G0 為前置閘（R-P14(a)）：不通過則其後一律不執行、不回報。
G6 已依 R-P18 拆為 G6a（錨點鏈第一段）/ G6b（第三段），原編號不再使用。
G11 已依 R-P14(b) 移除。

本腳本**不重算 §E**（03 包 §E 禁區）；§E 之 leaf 分布沿用 02 包狀態。
主章節僅用於 G14 之次章節判定與 B1 之歷史紀錄欄，不產生新的 §E 數字。

用法：
    python features/power/scripts/verify_gates_03.py
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from extract_textlayer import SEC_RE, paragraphs  # noqa: E402

ROOT = Path(__file__).resolve().parents[3]
IN = ROOT / "features/power/inputs"
DATA = ROOT / "features/power/data"

LEDGER = {
    "ce93174794d0d43c03d25dcd577c2811b85a8ebb2fd754a5201e5d6979297eda": "FW036-A01 交付標的",
    "2284abf5e6c17e4d1a082cf70d676b6551e16e023387c642f3fd1959c3478c8d": "037 需求母體",
    "6af7bfd314a28b3925636b642dc80c87918b50136fb8fa264563b3dc117eb592": "SYS2 CFTS009",
    "f318b14623fcbf972c6d0428408a6cf5297cbe7d0a710c7fafb8a0e05f538e61": "SYS2 CFTS010",
    "eb3eb0861363fc1d85c6127564355a7414ecd37eac0800336bb0d2e9b2911a26": "CFTS009 規格本文",
    "47c402a01b1a2e3a537797843b968ad621fc2bbc6f7a416a33dfe490247ea505": "CFTS010 規格本文",
    "cb6bf7d81030abc8ce47a444b4cf90b6bf527816fde7887b9e9b872c22338ae4": "SYS3 SYSAD",
    "a29fe63963192b804e20ed2fc6278dc9c434cbbb36b26bd77627cc0ea92949bb": "WrapperResource O829（49 包 R-P319 登記）",
    "dede965f228429c6e95aa7f7c4de08f0a52f1fd28b12ef10da3fc6db8638a9cc": "WrapperResource O1584（49 包 R-P319 登記）",
}

PM_RE = re.compile(r"Sys-RA-PM-\d{4}")
PD_RE = re.compile(r"Sys-RA-PD[_-]\d+")

# §C 讀取座標（SYS2 CFTS009 依 R-P18 訂正為 r2–r339）
COORDS = {
    "037 SWE1 Requirements": ("SWE1 Requirements", 7, 8, 145),
    "037 SYS2 Traceability": ("SYS2 Traceability", 1, 2, 34),
    "037 Excluded NRLs": ("Excluded NRLs (HW-only)", 1, 2, 27),
    "SYS2 CFTS009 Basic Report": ("Basic Report", 1, 2, 339),
    "SYS2 CFTS010 Basic Report": ("Basic Report", 1, 2, 74),
    "FW036 TC Spec&Result": ("Test Case Specification&Result", 9, 10, 221),
}


def find(pattern: str) -> Path:
    return next(f for f in IN.iterdir() if pattern in f.name)


def rows_of(path: Path, sheet: str, first: int, last: int) -> list[tuple]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    out = [
        r for r in ws.iter_rows(min_row=first, max_row=last, values_only=True)
        if any(v is not None and str(v).strip() for v in r)
    ]
    max_row = ws.max_row
    wb.close()
    return out, max_row


def report(gate: str, value, expected=None) -> None:
    if expected is None:
        print(f"{gate:5} {value}")
    else:
        ok = "PASS" if value == expected else "**MISMATCH**"
        print(f"{gate:5} 實測 {value}  期望 {expected}  {ok}")


def main() -> None:
    # ---- G0 前置閘 ----
    seen = {
        hashlib.sha256(f.read_bytes()).hexdigest(): f.name
        for f in sorted(IN.iterdir()) if not f.name.startswith(".")
    }
    matched = len(set(seen) & set(LEDGER))
    report("G0", f"{matched} / {len(LEDGER)}", f"{len(LEDGER)} / {len(LEDGER)}")
    if set(seen) != set(LEDGER):
        raise SystemExit("G0 未通過，依 R-P14(a) 停止，不執行亦不回報 G1 以後")

    f037, f036 = find("FSM-037"), find("FSM-036")
    s9f, s10f = find("SYS2_CFTS_009"), find("SYS2_CFTS_010")

    # ---- G1 / G2 ----
    leaf_rows, _ = rows_of(f037, "SWE1 Requirements", 8, 145)
    leaves = [
        (str(r[0]).strip(), str(r[1] or ""), str(r[5] or "").strip())
        for r in leaf_rows if r[0] and str(r[0]).strip()
    ]
    nums = sorted(int(m.group(1)) for m in (re.match(r"SWE-PM-(\d+)$", x[0]) for x in leaves) if m)
    report("G1", f"{len(leaves)}，連續={nums == list(range(1, len(nums) + 1))}", "115，連續=True")
    report("G2", dict(Counter(x[2] for x in leaves)), {"Functional Requirement": 115})

    # ---- 錨點鏈 ----
    chapters = json.loads((DATA / "item_to_chapter.json").read_text(encoding="utf-8"))
    tables = {"009": chapters["cfts009"], "010": chapters["cfts010"]}

    def sys2(path: Path, last: int) -> dict[str, list[str]]:
        rows, _ = rows_of(path, "Basic Report", 2, last)
        return {
            str(r[1] or "").strip(): re.findall(r"\d{6,8}", str(r[4] or ""))
            for r in rows if str(r[1] or "").strip()
        }

    s9, s10 = sys2(s9f, 339), sys2(s10f, 74)

    hits = {}
    for lid, src, _ in leaves:
        found = []
        for tok in PM_RE.findall(src):
            found += [("009", *tables["009"][i]) for i in s9.get(tok, []) if i in tables["009"]]
        for tok in PD_RE.findall(src):
            found += [("010", *tables["010"][i]) for i in s10.get(tok, []) if i in tables["010"]]
        if found:
            hits[lid] = found

    # ---- G3 / G4 / G5 / G5b ----
    failed = [lid for lid, _, _ in leaves if lid not in hits]
    report("G3", f"{len(hits)} / {len(leaves)}，失敗={failed}", "114 / 115，失敗=['SWE-PM-089']")
    dom = {"009": [], "010": [], "none": [], "both": []}
    for lid, src, _ in leaves:
        a, b = bool(PM_RE.search(src)), bool(PD_RE.search(src))
        dom["both" if a and b else "009" if a else "010" if b else "none"].append(lid)
    report("G4", f"{len(dom['009'])} / {len(dom['010'])} / {len(dom['none'])}"
                 f"，兩者皆有={len(dom['both'])}", "111 / 3 / 1，兩者皆有=0")
    report("G5", dom["010"], ["SWE-PM-071", "SWE-PM-072", "SWE-PM-073"])
    g5b = {lid: sorted({(d, n) for d, n, _ in hits.get(lid, [])}) for lid in dom["010"]}
    report("G5b", g5b, {"SWE-PM-071": [("010", "1.7.1.1.1")],
                        "SWE-PM-072": [("010", "1.7.1.1.1")],
                        "SWE-PM-073": [("010", "1.7.2")]})

    # ---- G6a / G6b（R-P18）----
    sec_ids = {}
    for tag, path in [("009", find("CFTS_009_Wake-up")),
                      ("010", next(x for x in IN.iterdir() if x.suffix == ".doc"))]:
        sec_ids[tag] = {
            m.group(3): (m.group(1), m.group(2))
            for plain, _ in paragraphs(path) if (m := SEC_RE.match(plain))
        }

    s9_rows, _ = rows_of(s9f, "Basic Report", 2, 339)
    with_tok = [(str(r[1] or "").strip() or str(r[0]), re.findall(r"\d{6,8}", str(r[4] or "")))
                for r in s9_rows]
    tok_rows = [x for x in with_tok if x[1]]
    report("G6a", f"{len(tok_rows)} / {len(s9_rows)}（r2–r339）")

    def resolvable(item: str) -> bool:
        return (item in tables["009"] or item in tables["010"]
                or item in sec_ids["009"] or item in sec_ids["010"])

    all_tok = [t for _, ts in tok_rows for t in ts]
    full_rows = [k for k, ts in tok_rows if all(resolvable(t) for t in ts)]
    bad = [(k, [t for t in ts if not resolvable(t)]) for k, ts in tok_rows
           if not all(resolvable(t) for t in ts)]
    report("G6b", f"列層 {len(full_rows)} / {len(tok_rows)}；"
                  f"token 層 {sum(1 for t in all_tok if resolvable(t))} / {len(all_tok)}；未解析={bad}")

    # ---- G7 ----
    s10_rows, _ = rows_of(s10f, "Basic Report", 2, 74)
    g7 = sum(1 for r in s10_rows if re.findall(r"\d{6,8}", str(r[4] or "")))
    report("G7", f"{g7} / {len(s10_rows)}", "73 / 73")

    # ---- G8 / G9 / G16 ----
    for gate, tag, path, expected in [
        ("G8", "009", find("CFTS_009_Wake-up"), "904 / 196"),
        ("G9", "010", next(x for x in IN.iterdir() if x.suffix == ".doc"), "148 / 92"),
        ("G16", "sys3", find("SYS3_"), None),
    ]:
        key = {"009": "cfts009", "010": "cfts010"}.get(tag)
        n_req = len(tables[tag]) if key else 0
        n_sec = len(sec_ids[tag]) if tag in sec_ids else len(
            {m.group(3) for plain, _ in paragraphs(path) if (m := SEC_RE.match(plain))}
        )
        if tag == "sys3":
            report("G16", f"章節錨點 {n_sec}（§C rule 1 於本文件不匹配，見 B3）")
        else:
            report(gate, f"{n_req} / {n_sec}", expected)

    # ---- G10 ----
    wb = openpyxl.load_workbook(f036, data_only=True, read_only=True)
    ws = wb["Test Case Specification&Result"]
    filled = sum(
        1 for row in ws.iter_rows(min_row=10, max_row=221, min_col=2, max_col=35, values_only=True)
        for v in row if v is not None and str(v).strip()
    )
    wb.close()
    report("G10", f"非空 {filled} → {'BLANK' if filled == 0 else 'NOT BLANK'}", "非空 0 → BLANK")

    # ---- G12 / G15 ----
    print("G12   §C 讀取座標實測：")
    for label, (sheet, header, first, last) in COORDS.items():
        path = {"037": f037, "SYS2 CFTS009": s9f, "SYS2 CFTS010": s10f, "FW036": f036}[
            "037" if label.startswith("037") else
            "SYS2 CFTS009" if "CFTS009" in label else
            "SYS2 CFTS010" if "CFTS010" in label else "FW036"
        ]
        rows, max_row = rows_of(path, sheet, first, last)
        span = last - first + 1
        print(f"        {label:28} 表頭 r{header}；r{first}–r{last} = {span} 列，"
              f"非空 {len(rows)}；sheet max_row={max_row}")
    print("G15   （含於 G12：037 兩分頁）")

    # ---- G13 / G14（R-P22）----
    main_ch, multi = {}, []
    for lid, found in hits.items():
        freq = Counter((d, n) for d, n, _ in found)
        main_ch[lid] = max(
            freq.items(), key=lambda kv: (kv[1], tuple(-int(x) for x in kv[0][1].split(".")))
        )[0]
        if len(freq) > 1:
            multi.append(lid)
    report("G13", f"{len(multi)}", "11")

    main_set = set(main_ch.values())
    titles = {(d, n): t for found in hits.values() for d, n, t in found}
    dropped = defaultdict(list)
    for lid, found in hits.items():
        for key in {(d, n) for d, n, _ in found} - {main_ch[lid]}:
            dropped[key].append(lid)
    uncovered = sorted(k for k in dropped if k not in main_set)
    report("G14", f"被丟棄次章節 {len(dropped)} 個；**未被任何 leaf 主章節覆蓋 {len(uncovered)} 個**")
    for key in uncovered:
        print(f"        CFTS{key[0]} §{key[1]} — {titles[key][:58]}"
              f"（來自 {', '.join(sorted(dropped[key]))}）")

    print(f"\n跨多章節 leaf：{sorted(multi)}")


if __name__ == "__main__":
    main()
