#!/usr/bin/env python3
"""A5：spec_reference 家族遷移（R-2(a)）。

來源鏈（全程可查證，無推測）：
  工作簿 `req_id`（SWE-PM-nnn）
    → `data/layer3_full.tsv` 之 `tokens`（Sys-RA-PM-nnnn）
    → SYS2 Polarion 匯出之 `Source Requirement ID`（7 位 ObjectID）
    → SYS2 `Document ID`（CFTS009／CFTS010）

**不由章節號反推**：ObjectID 一律取該 leaf 於 SYS2 所引之全集，
不自多個錨點中挑選其一 —— 挑選即為推測。

**輸出格式（Pei 裁定，2026-08-21）**：一個 ObjectID 一行，
每行皆為完整之 `CFTS{nnn}-{ObjectID}`；不留文件名與章節號、不用頓號。
⚠ 此與 canon §10.7 現行條文（「多個 ObjectID 以 `, ` 續列且文件前綴
僅敘明一次」）相反，該條文須依本裁定更新。
"""

from __future__ import annotations

import collections
import csv
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
INPUTS = ROOT / "features/power/inputs"
LAYER3 = ROOT / "features/power/data/layer3_full.tsv"

SPLIT = re.compile(r"[\n,;]+")


def load_sys2() -> dict[str, tuple[list[str], str]]:
    """SYS2 匯出 → {Sys-RA token: ([ObjectID…], Document ID)}。

    `Source Requirement ID` 單格可含多個以換行分隔之 ObjectID。
    """
    out: dict[str, tuple[list[str], str]] = {}
    for path in sorted(INPUTS.glob("SYS2_CFTS_*.xlsx")):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not (row and row[1]):
                continue
            ids = [x.strip() for x in SPLIT.split(str(row[4] or "")) if x.strip()]
            out[str(row[1]).strip()] = (ids, str(row[5] or "").strip())
        wb.close()
    return out


def load_leaf_map() -> dict[str, dict]:
    """leaf → {doc, ids}；ids 為該 leaf 於 SYS2 所引之 ObjectID 全集。"""
    sys2 = load_sys2()
    acc = collections.defaultdict(lambda: {"docs": set(), "ids": set()})
    for rec in csv.DictReader(LAYER3.open(encoding="utf-8"), delimiter="\t"):
        for token in (t.strip() for t in rec["tokens"].split(",") if t.strip()):
            if token not in sys2:                      # 查無者不猜，交呼叫端標 PENDING
                continue
            ids, doc = sys2[token]
            acc[rec["leaf"]]["ids"].update(ids)
            if doc:
                acc[rec["leaf"]]["docs"].add(doc)
    return {leaf: {"docs": sorted(v["docs"]), "ids": sorted(v["ids"])}
            for leaf, v in acc.items()}


def build(req_id: str, existing: str, leaf_map: dict[str, dict]) -> str | None:  # noqa: ARG001
    """回傳新的 spec_reference；無法解析回 None。"""
    entry = leaf_map.get(req_id.strip())
    if not entry or not entry["ids"] or len(entry["docs"]) != 1:
        return None
    doc = entry["docs"][0]
    return "\n".join(f"{doc}-{oid}" for oid in entry["ids"])
