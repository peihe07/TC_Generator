#!/usr/bin/env python3
"""10a 寫回：`surgical_save` 唯一路徑；交付本唯讀不覆寫。"""

from __future__ import annotations

import hashlib
import json
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))

import lint036                                         # noqa: E402
from xlsx_surgical import surgical_save                # noqa: E402

SANDBOX = ROOT / "features/power/sandbox/b10"
SRC, OUT = SANDBOX / "pm_base.xlsx", SANDBOX / "pm_10a.xlsx"
COLUMN = {"test_item": "I", "pre": "J", "input": "K", "proc": "L",
          "er": "M", "spec": "N"}


def sha256(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def main() -> None:
    changes = {int(k): v for k, v in
               json.loads((SANDBOX / "edits.json").read_text(encoding="utf-8"))["changes"].items()}
    work = SANDBOX / ".work.xlsx"
    shutil.copyfile(SRC, work)
    assert sha256(work) == sha256(SRC)

    wb = openpyxl.load_workbook(work)
    ws = wb[next(s for s in wb.sheetnames if s.startswith(lint036.TC_SHEET_PREFIX))]
    n = 0
    for row, fields in changes.items():
        for key, value in fields.items():
            ws[f"{COLUMN[key]}{row}"] = value
            n += 1
    report = surgical_save(wb, work, OUT)
    work.unlink()
    print(f"寫入 {n} 格／{len(changes)} 列 → {OUT.name}")
    print("members:", report["members"], "| differing:", report["differing"])
    print("dv:", report["dv_counts"])
    print("out sha256:", sha256(OUT)[:20])


if __name__ == "__main__":
    main()
