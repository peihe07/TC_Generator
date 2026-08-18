"""G103 —— layer3 之 token 層完整性（R-P144）。

G99 驗 `source_anchor` 等於 `layer3_full.tsv` 之 `item_ids`，
而 layer3 本身之正確性係 03–06 包所驗，非該閘所驗。
**若某 leaf 之 `Source Requirement ID` 於 layer3 建表時漏了一個
`Sys-RA-*` token，G94 全綠、G99 亦全綠，而該錨點從頭到尾無人看過**
（19 §八(甲)1）。

G103 自 **037 之 `Source Requirement ID` 欄獨立重算**錨點鏈
（token → SYS2 → item id），與 layer3 之 `item_ids` 逐 leaf 比對。

**R-P144(a)：重算不得讀取 layer3 之任何中間產物。**
本檔只讀三份**素材**：
  `FSM-037`（`SWE1 Requirements` 分頁，A 欄 leaf、B 欄 Source Requirement ID）
  `SYS2_CFTS_009` / `SYS2_CFTS_010`（`Basic Report` 分頁）
**不讀** `item_to_chapter.json`、`leaf_main_chapter.json`、`layer3_full.tsv`
之任何內容作為輸入 —— `layer3_full.tsv` 僅作為**比對對象**讀入，非輸入。

§C 之正則於本檔獨立宣告（與 `build_layer3.py` 相同之四條），
使二者為真正之獨立實作而非共用同一段程式碼。

用法：
    python features/power/scripts/verify_layer3.py
    python features/power/scripts/verify_layer3.py --self-test
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
IN = ROOT / "features/power/inputs"
DATA = ROOT / "features/power/data"
GENERATED = ROOT / "features/power/generated"
LAYER3 = DATA / "layer3_full.tsv"

# §C rule 1 / 2 —— Sys-RA token 之兩種形態（本檔獨立宣告，不自他處 import）
PM_RE = re.compile(r"Sys-RA-PM-\d{4}")
PD_RE = re.compile(r"Sys-RA-PD[_-]\d+")
# §C rule 3 —— SYS2 之 item id
ITEM_RE = re.compile(r"\d{6,8}")
SYS2_LAST = {"009": 339, "010": 74}


def find(pattern: str) -> Path:
    return next(f for f in IN.iterdir() if pattern in f.name)


def sys2_map(path: Path, last_row: int) -> dict[str, list[str]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Basic Report"]
    out: dict[str, list[str]] = {}
    for r in ws.iter_rows(min_row=2, max_row=last_row, values_only=True):
        key = str(r[1] or "").strip()
        if key:
            out[key] = ITEM_RE.findall(str(r[4] or ""))
    wb.close()
    return out


def leaf_sources() -> list[tuple[str, str]]:
    wb = openpyxl.load_workbook(find("FSM-037"), data_only=True, read_only=True)
    ws = wb["SWE1 Requirements"]
    rows = [(str(r[0]).strip(), str(r[1] or ""))
            for r in ws.iter_rows(min_row=8, max_row=145, values_only=True)
            if r[0] and str(r[0]).strip()]
    wb.close()
    return rows


def recompute(sources: list[tuple[str, str]]) -> dict[str, dict]:
    """037 → token → SYS2 → item id，逐 leaf。"""
    sys2 = {"009": sys2_map(find("SYS2_CFTS_009"), SYS2_LAST["009"]),
            "010": sys2_map(find("SYS2_CFTS_010"), SYS2_LAST["010"])}
    out = {}
    for leaf, src in sources:
        tokens = PM_RE.findall(src) + PD_RE.findall(src)
        items, unresolved = set(), []
        for tok in tokens:
            hit = False
            for cfts in ("009", "010"):
                if tok in sys2[cfts]:
                    items.update(sys2[cfts][tok])
                    hit = True
            if not hit:
                unresolved.append(tok)
        out[leaf] = {"tokens": tokens, "items": items, "unresolved": unresolved}
    return out


def layer3_items() -> dict[str, set[str]]:
    """**比對對象**（非輸入）：layer3 之 `item_ids` 逐 leaf 聯集。"""
    out: dict[str, set[str]] = {}
    rows = LAYER3.read_text(encoding="utf-8").splitlines()
    header = rows[0].split("\t")
    i_leaf, i_ids = header.index("leaf"), header.index("item_ids")
    for row in rows[1:]:
        if not row.strip():
            continue
        c = row.split("\t")
        out.setdefault(c[i_leaf], set()).update(
            x.strip() for x in c[i_ids].split(",") if x.strip())
    return out


def scoped_leaves() -> list[str]:
    """已產出 TC 之 leaf —— 本閘之比對範圍（其餘 leaf 尚未撰寫）。"""
    import json
    out = []
    for p in sorted(GENERATED.glob("*.json")):
        for leaf in json.loads(p.read_text(encoding="utf-8")).get("leaves", []):
            out.append(leaf["parent"])
    return sorted(set(out))


def compare(leaf: str, recomputed: dict, layer3: dict[str, set[str]]) -> dict:
    got = layer3.get(leaf, set())
    exp = recomputed[leaf]["items"]
    return {"leaf": leaf, "tokens": recomputed[leaf]["tokens"],
            "unresolved": recomputed[leaf]["unresolved"],
            "recomputed": sorted(exp), "layer3": sorted(got),
            "missing_in_layer3": sorted(exp - got),
            "extra_in_layer3": sorted(got - exp),
            # **不可解析之 token 亦判 FAIL。** 初版僅比對 item 集合，
            # 而 037 若引用了 SYS2 未載之 token，該 token 不產生任何 item，
            # 集合仍相等 —— 錨點形同消失而閘門全綠。
            # 此形態由本檔第四個 fixture 當場暴露（其期望值原本寫錯），
            # 已改為併入判定，見上繳 §四。
            "ok": exp == got and not recomputed[leaf]["unresolved"]}


def self_test() -> int:
    """R-P144(c) —— 以刻意自 037 某列刪去一個 token 之 fixture 證明會失敗。"""
    failures = 0
    sources = leaf_sources()
    layer3 = layer3_items()
    target = "SWE-PM-057"
    src = next(s for lf, s in sources if lf == target)

    def case(label: str, synth_src: str, want_ok: bool) -> None:
        nonlocal failures
        rec = recompute([(target, synth_src)])
        r = compare(target, rec, layer3)
        ok = r["ok"] == want_ok
        failures += not ok
        print(f"  [{'PASS' if ok else '**FAIL**'}] G103 {label}")
        print(f"          期望 {'相等' if want_ok else 'FAIL'}；"
              f"實際 {'相等' if r['ok'] else 'FAIL'}"
              f"（layer3 缺 {r['missing_in_layer3']}，layer3 多 {r['extra_in_layer3']}）")

    case("應相等 —— 037 原值", src, True)
    tokens = PM_RE.findall(src) + PD_RE.findall(src)
    assert len(tokens) >= 2, "本 leaf 之 token 少於 2，fixture 無法建構"
    case(f"應 FAIL —— 刻意刪去一個 token（`{tokens[0]}`）",
         src.replace(tokens[0], "", 1), False)
    case("應 FAIL —— 037 該列為空", "", False)
    # 初版此案之期望值寫成「應 FAIL —— 多一個不存在之 token」而實測為「相等」——
    # 因不可解析之 token 不產生任何 item，集合自然相等。
    # **那不是閘門的瑕疵，是我的期望值寫錯**，且它暴露了一個真的漏洞：
    # 錨點形同消失而閘門全綠。已將 unresolved 併入判定，本案改為：
    case("應 FAIL —— 037 引用 SYS2 未載之 token（錨點形同消失）",
         src + "\nSys-RA-PM-9999", False)
    print(f"\n  G103 fixture 全數如期：{'是' if not failures else '否'}")
    return 1 if failures else 0


def main() -> None:
    if "--self-test" in sys.argv:
        raise SystemExit(self_test())

    sources = leaf_sources()
    rec = recompute(sources)
    layer3 = layer3_items()
    scope = scoped_leaves()
    results = [compare(lf, rec, layer3) for lf in scope]
    bad = [r for r in results if not r["ok"]]

    lines = ["# G103 —— layer3 之 token 層完整性（R-P144）\n",
             "\n> 自 **037 之 `Source Requirement ID` 欄獨立重算** token → SYS2 → item id，\n",
             "> 與 `layer3_full.tsv` 之 `item_ids` 逐 leaf 比對。\n",
             "> **重算未讀取 layer3 之任何中間產物**（R-P144(a)）——\n",
             "> 只讀 037 與二份 SYS2 匯出；`layer3_full.tsv` 僅作比對對象。\n",
             "> §C 之正則於 `verify_layer3.py` 獨立宣告，非自 `build_layer3.py` import。\n",
             "\n## 比對範圍：已產出 TC 之 11 leaf\n\n"
             "| leaf | token 數 | 重算 item 數 | layer3 item 數 | layer3 缺 | layer3 多 | 判定 |\n"
             "|---|---|---|---|---|---|---|\n"]
    for r in results:
        lines.append(f"| `{r['leaf']}` | {len(r['tokens'])} | {len(r['recomputed'])} | "
                     f"{len(r['layer3'])} | {'、'.join(r['missing_in_layer3']) or '—'} | "
                     f"{'、'.join(r['extra_in_layer3']) or '—'} | "
                     f"{'**相等**' if r['ok'] else '**FAIL**'} |\n")
    unresolved = {r["leaf"]: r["unresolved"] for r in results if r["unresolved"]}
    lines.append(f"\n**{len(results) - len(bad)} / {len(results)} 相等。**\n")
    lines.append(f"\n未能解析至任何 SYS2 列之 token："
                 f"{unresolved or '（無）'}\n")
    (DATA / "g103_layer3.md").write_text("".join(lines), encoding="utf-8")

    for r in results:
        print(f"  {r['leaf']:12} tokens {len(r['tokens']):2}  重算 {len(r['recomputed']):2}  "
              f"layer3 {len(r['layer3']):2}  "
              f"{'相等' if r['ok'] else '**FAIL** 缺=' + str(r['missing_in_layer3']) + ' 多=' + str(r['extra_in_layer3'])}")
    print(f"\nG103：{len(results) - len(bad)} / {len(results)} 相等")
    if unresolved:
        print(f"未解析之 token：{unresolved}")
    raise SystemExit(1 if bad else 0)


if __name__ == "__main__":
    main()
