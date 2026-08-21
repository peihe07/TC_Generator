#!/usr/bin/env python3
"""16 包驗收：對 pm_16.xlsx 跑本包之九項驗收，並與 pm_10a5b.xlsx 逐格比對。

驗收範圍分兩級印出：
  scope  —— 軌 A＋軌 B 共 253 列（本包所改）
  full   —— 全 283 列（含軌 C 之 30 列，未改，供對照）
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "scripts"))
import lint036                                         # noqa: E402

BEFORE = ROOT / "features/power/sandbox/b10/pm_10a5b.xlsx"
AFTER = Path(sys.argv[1]) if len(sys.argv) > 1 else \
    ROOT / "features/power/sandbox/b16/pm_16.xlsx"
COLUMN = {"test_item": "I", "pre": "J", "input": "K", "proc": "L",
          "er": "M", "spec": "N"}
FOUR = ("pre", "input", "proc", "er")
TRACK_C = set(range(124, 128)) | {149, 181, 233, 234} | set(range(265, 283)) \
    | {289, 290, 291, 293}
BLANK_ROW = 230
TOOL_LINE = "LIN and CAN tool is available on HU"

NUMBERED = re.compile(r"^\d+\. ")
PREDICATE = re.compile(r"\b(is|are|was|were|reads?|holds?|has|have)\b")
TRIPLET = re.compile(r"\b[A-Za-z0-9_]+\s+in\s+[A-Z][A-Z0-9_]{2,}\s+on\s+[A-Za-z0-9-]+\b")


def load(path: Path) -> dict[int, dict[str, str]]:
    wb = openpyxl.load_workbook(path)
    ws = wb[next(s for s in wb.sheetnames
                 if s.startswith(lint036.TC_SHEET_PREFIX))]
    rows = {}
    for r in range(10, 294):
        if r == BLANK_ROW:
            continue
        rows[r] = {k: (ws[f"{c}{r}"].value or "") for k, c in COLUMN.items()}
    return rows


def lines(text: str) -> list[str]:
    return [ln for ln in text.split("\n") if ln.strip()]


def pre_multi_condition(line: str) -> bool:
    """R-9(a)：一行載兩個以上獨立條件。"""
    body = NUMBERED.sub("", line).replace(TOOL_LINE, "")
    if " and " not in body and ", " not in body:
        return False
    return len(PREDICATE.findall(body)) >= 2


def read_observations(line: str) -> int:
    """一個 Read 步驟所讀之對象數。"""
    body = NUMBERED.sub("", line)
    if not body.startswith("Read "):
        return 0
    obj = body[5:].split(" and check")[0]
    return 1 + obj.count(", ") + obj.count(" and ")


def audit(rows: dict[int, dict[str, str]], scope: set[int]) -> dict[str, list]:
    f = {k: [] for k in (
        "input_not_na", "listed_in_input", "triplet", "send_can",
        "pre_unnumbered", "pre_multi", "pre_first_is_tool", "pre_last_not_tool",
        "step_multi_obs", "read_without_value", "nbsp", "proc_er_mismatch")}
    for r in sorted(scope):
        d = rows[r]
        joined = "\n".join(d[k] for k in FOUR)
        if d["input"].strip() != "NA":
            f["input_not_na"].append(r)
        if "listed in Input Test Data" in joined:
            f["listed_in_input"].append(r)
        if TRIPLET.search(joined):
            f["triplet"].append(r)
        if "Send CAN:" in joined:
            f["send_can"].append(r)
        for key in ("test_item",) + FOUR:
            if "\xa0" in d[key] or "　" in d[key]:
                f["nbsp"].append((r, key))
        pre = lines(d["pre"])
        for ln in pre:
            if not NUMBERED.match(ln):
                f["pre_unnumbered"].append((r, ln[:60]))
            elif pre_multi_condition(ln):
                f["pre_multi"].append((r, ln[:70]))
        if pre:
            if TOOL_LINE in pre[0]:
                f["pre_first_is_tool"].append(r)
            if TOOL_LINE not in pre[-1]:
                f["pre_last_not_tool"].append((r, pre[-1][:60]))
        for ln in lines(d["proc"]):
            if read_observations(ln) >= 2:
                f["step_multi_obs"].append((r, ln[:70]))
            body = NUMBERED.sub("", ln)
            if body.startswith("Read ") and "check that" not in body:
                f["read_without_value"].append((r, ln[:70]))
        n_proc = len([x for x in lines(d["proc"]) if NUMBERED.match(x)])
        n_er = len([x for x in lines(d["er"]) if NUMBERED.match(x)])
        if n_proc != n_er:
            f["proc_er_mismatch"].append((r, n_proc, n_er))
    return f


def strip_invisible(text: str) -> str:
    text = text.replace("\xa0", " ").replace("　", " ")
    return "\n".join(re.sub(r"[ \t]+$", "", ln) for ln in text.split("\n"))


def main() -> None:
    before, after = load(BEFORE), load(AFTER)
    scope = set(after) - TRACK_C
    print(f"改動範圍 {len(scope)} 列／全表 {len(after)} 列")

    for label, sel in (("scope", scope), ("full", set(after))):
        print(f"\n=== 驗收（{label}）===")
        for name, hits in audit(after, sel).items():
            mark = "OK " if not hits else "FAIL"
            extra = "" if not hits else f"  {hits[:4]}{' …' if len(hits) > 4 else ''}"
            print(f"  [{mark}] {name}: {len(hits)}{extra}")

    print("\n=== 逐格 diff ===")
    changed = {k: 0 for k in COLUMN}
    item_content_changed = []
    for r in after:
        for key in COLUMN:
            if before[r][key] != after[r][key]:
                changed[key] += 1
                if key == "test_item" and \
                        strip_invisible(before[r][key]) != after[r][key]:
                    item_content_changed.append(r)
    print("  變動格數:", {k: v for k, v in changed.items() if v})
    print("  test_item 非不可見字元之變動:", len(item_content_changed))
    print("  spec_reference 變動:", changed["spec"])
    off_scope = [r for r in TRACK_C if r in after
                 and any(before[r][k] != after[r][k] for k in FOUR)]
    print("  軌 C 四欄變動列:", off_scope)

    print("\n=== 下拉（x14）讀回 ===")
    wb = openpyxl.load_workbook(AFTER)
    for name in wb.sheetnames:
        dvs = wb[name].data_validations.dataValidation
        if dvs:
            print(f"  {name}: {len(dvs)} 個 DV")


if __name__ == "__main__":
    main()
