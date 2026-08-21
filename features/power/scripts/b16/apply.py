#!/usr/bin/env python3
"""16 包寫回：軌 A（56 列，附件 A-E 逐字）＋ 軌 B（197 列，執行層改寫）。

- 改動範圍限 `pre`／`input`／`proc`／`er` 四欄。
- `test_item` 僅施 R-10(a) 不可見字元正規化（NBSP／全形空格／行尾空白），
  引號、方括號、句號、破折號一律不動（R-10(c)）。
- `spec_reference` 零變動（已經錨鏈驗證為正確）。
- 軌 C 之 30 列不動。
- `surgical_save` 唯一路徑；交付本唯讀不覆寫。
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))

import lint036                                         # noqa: E402
from xlsx_surgical import surgical_save                # noqa: E402

SANDBOX = ROOT / "features/power/sandbox/b16"
SRC = ROOT / "features/power/sandbox/b10/pm_10a5b.xlsx"
OUT = SANDBOX / "pm_16.xlsx"
COLUMN = {"test_item": "I", "pre": "J", "input": "K", "proc": "L", "er": "M"}
FOUR = ("pre", "input", "proc", "er")

TRACK_C = set(range(124, 128)) | {149, 181, 233, 234} | set(range(265, 283)) \
    | {289, 290, 291, 293}


def sha256(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def normalise_invisible(text: str) -> str:
    """R-10(a)：不可見字元不構成內容，全欄位適用（含 verbatim 上半）。"""
    text = text.replace("\xa0", " ").replace("　", " ")
    lines = [re.sub(r"[ \t]+$", "", line) for line in text.split("\n")]
    return "\n".join(lines)


def main() -> None:
    edits = {int(k): v for k, v in
             json.loads((SANDBOX / "edits.json").read_text(encoding="utf-8")).items()}
    assert not (set(edits) & TRACK_C), "軌 C 之列不得出現於改動清單"

    work = SANDBOX / ".work.xlsx"
    shutil.copyfile(SRC, work)
    assert sha256(work) == sha256(SRC)

    wb = openpyxl.load_workbook(work)
    ws = wb[next(s for s in wb.sheetnames
                 if s.startswith(lint036.TC_SHEET_PREFIX))]

    n_four = 0
    for row, fields in sorted(edits.items()):
        for key in FOUR:
            ws[f"{COLUMN[key]}{row}"] = fields[key]
            n_four += 1

    n_item = 0
    for row in range(10, 294):
        cell = ws[f"{COLUMN['test_item']}{row}"]
        if not isinstance(cell.value, str):
            continue
        fixed = normalise_invisible(cell.value)
        if fixed != cell.value:
            cell.value = fixed
            n_item += 1

    report = surgical_save(wb, work, OUT)
    work.unlink()
    print(f"四欄寫入 {n_four} 格／{len(edits)} 列；test_item 正規化 {n_item} 格")
    print("members:", report["members"], "| differing:", report["differing"])
    print("dv:", report["dv_counts"])
    print("out sha256:", sha256(OUT)[:20])


if __name__ == "__main__":
    main()
