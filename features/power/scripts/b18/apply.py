#!/usr/bin/env python3
"""18 包寫回：軌 C 套用（附件 G／H 逐字）。

基底為 `pm_17.xlsx`（= pm_16 ＋ 17 包 §五 之 `PowerModeSts_Telematic` 改寫），
因本包硬性第 4 項即該裁定。

⚠ 附件 G／H 合計僅 26 列，軌 C 之 rows 271–274 未獲附件 ——
該四列不動，見上繳 18 §一。

改動範圍限 `pre`／`input`／`proc`／`er` 四欄；
`test_item`／`spec_reference` 零變動。
`surgical_save` 唯一路徑；交付本唯讀不覆寫。
"""

from __future__ import annotations

import hashlib
import json
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))

import lint036                                         # noqa: E402
from xlsx_surgical import surgical_save                # noqa: E402

SANDBOX = ROOT / "features/power/sandbox/b18"
SRC = ROOT / "features/power/sandbox/b17/pm_17.xlsx"
OUT = SANDBOX / "pm_18.xlsx"
COLUMN = {"pre": "J", "input": "K", "proc": "L", "er": "M"}


def sha256(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def main() -> None:
    edits = {int(k): v for k, v in
             json.loads((SANDBOX / "edits.json").read_text(encoding="utf-8")).items()}
    work = SANDBOX / ".work.xlsx"
    shutil.copyfile(SRC, work)
    assert sha256(work) == sha256(SRC)

    wb = openpyxl.load_workbook(work)
    ws = wb[next(s for s in wb.sheetnames
                 if s.startswith(lint036.TC_SHEET_PREFIX))]
    n = 0
    for row, fields in sorted(edits.items()):
        for key, value in fields.items():
            ws[f"{COLUMN[key]}{row}"] = value
            n += 1

    report = surgical_save(wb, work, OUT)
    work.unlink()

    print(f"寫入 {n} 格／{len(edits)} 列 → {OUT.name}")
    print("members:", report["members"], "| differing:", report["differing"])
    print("dv:", report["dv_counts"])
    print("out sha256:", sha256(OUT)[:20])


if __name__ == "__main__":
    main()
