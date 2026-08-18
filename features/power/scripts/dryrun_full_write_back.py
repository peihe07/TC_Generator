"""G200 / G201 —— **全量** dry-run 寫回（R-P287）。

41 §K 第 3 項：G66 / G71 / G72 迄今僅有 **10 列之 dry-run 證據**（16 包），
而實際寫回為 **115 列**。本檔以現行全量實測三閘之行為。

**嚴格遵守 16 §I 與 R-P287**：

  - 僅對 `features/power/sandbox/` 之副本為之；
    **客戶樹與 `inputs/` 之原始檔一律不觸碰**（唯讀來源，位元組複製）
  - 寫回路徑為 `backend/xlsx_surgical.py` 之 `surgical_save()`；
    **本檔全域無 `Workbook.save()` 之呼叫**
  - **本包只做 dry-run 與評估，不對客戶樹寫回**（R-P287）
  - 列序依 `(SWE-PM ID, split_index)`（R-P113 / R-P115），
    **不因 dry-run 結果自行調整**（§I）

驗四項（R-P287(a)–(d)）：
  (a) 全量之三閘行為
  (b) DV（含 x14）／合併儲存格／條件式格式／分頁清單／欄寬與凍結窗格（G95）
  (c) B 欄序號與列序逐列
  (d) `SWE-PM-089` 之**留白列**（R-P141）於全量下之落位

用法：
    python features/power/scripts/dryrun_full_write_back.py
"""

from __future__ import annotations

import glob
import json
import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
FEATURE = ROOT / "features/power"
SANDBOX = FEATURE / "sandbox"
DATA = FEATURE / "data"
sys.path.insert(0, str(Path(__file__).resolve().parent))

from dryrun_write_back import (HEADER_ROW, FIRST_DATA_ROW, NS_MAIN,  # noqa: E402
                               load_cfg, row_values, dv_snapshot,
                               structure_snapshot, sha256)

BLANK_LEAF = "SWE-PM-089"          # R-P141：留白列，僅 req_id 有值


def ordered_tcs() -> list[dict]:
    """依 `(SWE-PM ID, split_index)` 排序（R-P113 / R-P115）。"""
    tcs = []
    for f in sorted(glob.glob(str(FEATURE / "generated/*.json"))):
        tcs += json.loads(Path(f).read_text(encoding="utf-8"))["tcs"]

    def key(t: dict) -> tuple[int, int]:
        m = re.match(r"SWE-PM-(\d+)", t["req_id"])
        return (int(m.group(1)), int(t.get("split_index") or 1))
    return sorted(tcs, key=key)


def pane_and_widths(path: Path) -> dict:
    """G95 —— 凍結窗格與欄寬。"""
    import xml.etree.ElementTree as ET
    out = {"panes": [], "cols": []}
    with zipfile.ZipFile(path) as z:
        for m in sorted(x for x in z.namelist() if x.startswith("xl/worksheets/sheet")):
            root = ET.fromstring(z.read(m))
            for pane in root.iter(NS_MAIN + "pane"):
                out["panes"].append((m, pane.get("topLeftCell"), pane.get("state")))
            for c in root.iter(NS_MAIN + "col"):
                out["cols"].append((m, c.get("min"), c.get("max"), c.get("width")))
    return out


def main() -> None:
    cfg = load_cfg()
    src_cfg = cfg["workbook"]
    # 來源路徑與 16 包之 dry-run 同源（`cfg["paths"]["workbook"]`，相對於 feature 根）
    src = FEATURE / cfg["paths"]["workbook"]
    if not src.exists():
        raise SystemExit(f"來源工作簿不存在：{src}")

    SANDBOX.mkdir(parents=True, exist_ok=True)
    work = SANDBOX / "full_dryrun_src.xlsx"
    out = SANDBOX / "full_dryrun_out.xlsx"
    shutil.copyfile(src, work)                       # 位元組複製，來源唯讀
    assert sha256(work) == sha256(src), "來源副本非位元組相同"

    tcs = ordered_tcs()
    # (d) 留白列 —— 僅 `req_id` 有值（R-P141）
    blank = {"req_id": BLANK_LEAF, "tc_id": "", "tc_title": "", "test_item": "",
             "pre_conditions": "", "input_test_data": "", "test_procedure": "",
             "expected_result": "", "specification_reference": "", "priority": "",
             "design_method": "", "split_flag": "", "split_reason": "",
             "functional_safety": "", "estimated_test_time": "", "remarks": "",
             "split_index": 1}
    # 依 SWE-PM ID 序插入 089 之位置
    pos = next((i for i, t in enumerate(tcs)
                if int(re.match(r"SWE-PM-(\d+)", t["req_id"]).group(1)) > 89), len(tcs))
    rows = tcs[:pos] + [blank] + tcs[pos:]

    before_dv, before_st = dv_snapshot(work), structure_snapshot(work)
    before_pw = pane_and_widths(work)

    sys.path.insert(0, str(ROOT / "backend"))
    from xlsx_surgical import surgical_save
    wb = openpyxl.load_workbook(work)
    ws = wb[src_cfg["sheet"]]
    for i, tc in enumerate(rows):
        r = FIRST_DATA_ROW + i
        for letter, value in row_values(tc, cfg, r).items():
            ws[f"{letter}{r}"] = value
    surgical_save(wb, work, out)                     # **不呼叫 wb.save()**

    after_dv, after_st = dv_snapshot(out), structure_snapshot(out)
    after_pw = pane_and_widths(out)

    # (c) B 欄與列序逐列
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2[src_cfg["sheet"]]
    colB, col_req = "B", cfg["workbook"]["columns"]["req_id"]
    bad_b, bad_order = [], []
    for i, tc in enumerate(rows):
        r = FIRST_DATA_ROW + i
        if ws2[f"{colB}{r}"].value != r - HEADER_ROW:
            bad_b.append((r, ws2[f"{colB}{r}"].value, r - HEADER_ROW))
        if str(ws2[f"{col_req}{r}"].value or "") != tc["req_id"]:
            bad_order.append((r, ws2[f"{col_req}{r}"].value, tc["req_id"]))

    # (d) 留白列之落位
    br = FIRST_DATA_ROW + pos
    blank_only_req = all(
        (ws2[f"{c}{br}"].value in (None, "")) for c in
        [v for k, v in cfg["workbook"]["columns"].items() if k != "req_id"]
        if re.fullmatch(r"[A-Z]+", str(c))) if True else False
    prev_req = str(ws2[f"{col_req}{br - 1}"].value or "")
    next_req = str(ws2[f"{col_req}{br + 1}"].value or "")

    checks = {
        "DV（含 x14）未變": before_dv == after_dv,
        "合併儲存格未變": before_st["merges"] == after_st["merges"],
        "條件式格式未變": before_st["cf"] == after_st["cf"],
        "分頁清單未變": before_st["sheets"] == after_st["sheets"],
        "壓縮成員清單未變": before_st["members"] == after_st["members"],
        "凍結窗格未變（G95）": before_pw["panes"] == after_pw["panes"],
        "欄寬未變（G95）": before_pw["cols"] == after_pw["cols"],
        "B 欄序號逐列相符": not bad_b,
        "列序依 (SWE-PM ID, split_index)": not bad_order,
        f"留白列落於 {BLANK_LEAF}，僅 req_id 有值": blank_only_req,
    }
    ok = all(checks.values())

    md = ["# G200 / G201 —— 全量 dry-run 寫回（R-P287）\n",
          "\n> **僅對沙箱副本**；來源為位元組複製之唯讀副本。\n",
          "> 寫回路徑為 `surgical_save()`，**全域無 `Workbook.save()`**。\n",
          "> **本包不對客戶樹寫回**（R-P287）。\n",
          f"\n## 一、規模\n\n| 項 | 值 |\n|---|---|\n"
          f"| TC 列數 | **{len(tcs)}** |\n"
          f"| ＋ 留白列（`{BLANK_LEAF}`，R-P141） | 1 |\n"
          f"| **合計寫入列** | **{len(rows)}** |\n"
          f"| 首列 / 末列 | {FIRST_DATA_ROW} / {FIRST_DATA_ROW + len(rows) - 1} |\n"
          f"| 16 包之 dry-run 規模 | 10 列 |\n",
          "\n## 二、逐項驗證\n\n| 項 | 結果 |\n|---|---|\n"]
    for k, v in checks.items():
        md.append(f"| {k} | {'**PASS**' if v else '**FAIL**'} |\n")
    md.append(f"\n## 三、留白列之落位（R-P287(d)）\n\n"
              f"- 落於第 **{br}** 列（第 {pos + 1} 個資料列）\n"
              f"- 前一列之 `req_id`：`{prev_req}`\n"
              f"- 後一列之 `req_id`：`{next_req}`\n"
              f"- 僅 `req_id` 有值：{'**是**' if blank_only_req else '**否**'}\n")
    if bad_b:
        md.append(f"\n## B 欄不符（{len(bad_b)}）\n\n| 列 | 實際 | 應為 |\n|---|---|---|\n")
        for r, got, want in bad_b[:20]:
            md.append(f"| {r} | {got} | {want} |\n")
    if bad_order:
        md.append(f"\n## 列序不符（{len(bad_order)}）\n\n| 列 | 實際 | 應為 |\n|---|---|---|\n")
        for r, got, want in bad_order[:20]:
            md.append(f"| {r} | {got} | {want} |\n")
    md.append(f"\n## 四、DV 條數\n\n寫回前 **{len(before_dv)}**、寫回後 **{len(after_dv)}**"
              f"（其中 x14：{sum(1 for x in before_dv if x[0] == 'x14')} → "
              f"{sum(1 for x in after_dv if x[0] == 'x14')}）。\n")

    (DATA / "g200_full_dryrun.md").write_text("".join(md), encoding="utf-8")
    print(f"wrote {(DATA / 'g200_full_dryrun.md').relative_to(ROOT)}")
    print(f"寫入 {len(rows)} 列（TC {len(tcs)} ＋ 留白 1）")
    for k, v in checks.items():
        print(f"  [{'PASS' if v else '**FAIL**'}] {k}")
    print(f"\nDV {len(before_dv)} → {len(after_dv)}；"
          f"x14 {sum(1 for x in before_dv if x[0]=='x14')} → "
          f"{sum(1 for x in after_dv if x[0]=='x14')}")
    raise SystemExit(0 if ok else 1)


if __name__ == "__main__":
    main()
