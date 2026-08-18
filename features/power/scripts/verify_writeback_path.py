"""B2 / B3 —— 寫回路徑之閘門（R-P119 / R-P120）。

**G89** —— `backend/xlsx_surgical.py` 之 `verify_structure()` 至今無人驗過。
它是本專案**唯一授權之寫回路徑**，而 16 包對它的驗證只是「它沒拋例外」。
本節為其三項檢查各建**刻意弄壞**之案例，證明它確實會拋 `StructureError`，
另建正常案例證明它不誤拋。

**G90** —— append 邊界保護。dry-run 只驗過 BLANK 工作簿、十列、單次寫回。
真實寫回時簿內可能已有他人之列（Home 之 Arif 144 列即為先例）。
本節以**合成之非 BLANK 副本**驗：既有列逐格不變、新列不覆蓋、
B 欄序號銜接、以及刻意重疊時**須失敗**。

全程僅對 `features/power/sandbox/` 之副本為之；
**無任何 `Workbook.save()` 呼叫**（R16 / R-G3）。

用法：
    python features/power/scripts/verify_writeback_path.py
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[3]
FEATURE = ROOT / "features/power"
SANDBOX = FEATURE / "sandbox"
DATA = FEATURE / "data"

sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(Path(__file__).resolve().parent))
from xlsx_surgical import StructureError, surgical_save, verify_structure  # noqa: E402
from dryrun_write_back import (FIRST_DATA_ROW, HEADER_ROW,  # noqa: E402
                               load_cfg, row_values, sha256)

EXISTING_ROWS = 5          # 合成之「他人既有列」

NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
NS_X14 = "{http://schemas.microsoft.com/office/spreadsheetml/2009/9/main}"
NS_XM = "{http://schemas.microsoft.com/office/excel/2006/main}"


def clone_zip(src: Path, out: Path, *, drop=None, mutate=None) -> None:
    """位元組複製，可選擇丟掉一個 member 或改寫一個 member（供刻意弄壞）。"""
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            if drop and info.filename == drop:
                continue
            data = zin.read(info.filename)
            if mutate and info.filename == mutate[0]:
                data = mutate[1](data)
            zout.writestr(info, data)


# ---------------------------------------------------------------- G95


def sheet_structure(path: Path) -> dict:
    """G95 / R-P130 —— **目標分頁內**之結構屬性快照。

    `verify_structure()` 之允許相異清單涵蓋目標分頁，故該分頁內之
    `<mergeCell>` / `<conditionalFormatting>` 變動**不會被它攔下**
    （17 §七(乙)7）。本函式補該缺口，涵蓋五類：
    合併儲存格、條件式格式、DV（含 x14）、欄寬、凍結窗格。
    """
    import xml.etree.ElementTree as ET
    snap: dict[str, list] = {"merges": [], "cf": [], "dv": [],
                             "cols": [], "panes": []}
    with zipfile.ZipFile(path) as z:
        for member in sorted(m for m in z.namelist()
                             if m.startswith("xl/worksheets/sheet")):
            root = ET.fromstring(z.read(member))
            snap["merges"] += [(member, x.get("ref"))
                               for x in root.iter(NS_MAIN + "mergeCell")]
            snap["cf"] += [(member, x.get("sqref"),
                            tuple(r.get("type") for r in x.iter(NS_MAIN + "cfRule")))
                           for x in root.iter(NS_MAIN + "conditionalFormatting")]
            for dv in root.iter(NS_MAIN + "dataValidation"):
                snap["dv"].append((member, "main", dv.get("sqref"), dv.get("type")))
            for dv in root.iter(NS_X14 + "dataValidation"):
                sq = dv.find(NS_XM + "sqref")
                snap["dv"].append((member, "x14",
                                   (sq.text if sq is not None else ""), dv.get("type")))
            snap["cols"] += [(member, c.get("min"), c.get("max"), c.get("width"))
                             for c in root.iter(NS_MAIN + "col")]
            snap["panes"] += [(member, p.get("topLeftCell"), p.get("state"),
                               p.get("xSplit"), p.get("ySplit"))
                              for p in root.iter(NS_MAIN + "pane")]
    return snap


def g95(src: Path, out_ok: Path) -> list[dict]:
    """正常寫回不得改動五類屬性；刻意改動其一須被偵出。"""
    base = sheet_structure(src)
    cases = [{"case": "正常：dry-run 寫回後（不得改動）",
              "same": sheet_structure(out_ok) == base,
              "changed": [k for k in base
                          if sheet_structure(out_ok)[k] != base[k]]}]

    # **每一類屬性須改在確實含有該屬性之分頁上** ——
    # 首次實作時一律改於 `_target_sheet()`（首個含 DV 之分頁 = sheet5），
    # 而條件式格式與凍結窗格位於資料分頁 sheet6，致該二案未觸發而誤判為
    # 「本閘攔不住」。此為 fixture 建構之誤，非閘門之弱點（見上繳 §三）。
    breaks = [
        (b"<mergeCell ", "弄壞：刪去一個 `<mergeCell>`",
         lambda b: re.sub(rb"<mergeCell [^>]*/>", b"", b, count=1)),
        (b"<conditionalFormatting", "弄壞：改動 `<conditionalFormatting>` sqref",
         lambda b: re.sub(rb'(<conditionalFormatting sqref=")[^"]*(")',
                          rb"\g<1>ZZ1:ZZ2\g<2>", b, count=1)),
        (b"<dataValidation", "弄壞：抹去一條 `<dataValidation>`",
         lambda b: re.sub(rb"<dataValidation[ >].*?</dataValidation>", b"",
                          b, count=1, flags=re.S)),
        (b"<col ", "弄壞：改動一個 `<col>` 寬度",
         lambda b: re.sub(rb'(<col [^>]*width=")[^"]*(")', rb"\g<1>99\g<2>",
                          b, count=1)),
        (b"<pane ", "弄壞：改動 `<pane>` topLeftCell",
         lambda b: re.sub(rb'(<pane [^>]*topLeftCell=")[^"]*(")', rb"\g<1>Z99\g<2>",
                          b, count=1)),
    ]
    for i, (marker, label, mut) in enumerate(breaks):
        try:
            target = _sheet_containing(src, marker)
        except RuntimeError:
            # 本工作簿無此屬性 —— 依 R-P119(c) 之標準標「未實測」而非 PASS。
            cases.append({"case": f"{label} —— **未實測**",
                          "untested": True,
                          "why": f"本工作簿之十個分頁皆無 {marker.decode().strip()} 元素",
                          "same": None, "changed": []})
            continue
        broken = SANDBOX / f"g95_{i}.xlsx"
        clone_zip(src, broken, mutate=(target, mut))
        snap = sheet_structure(broken)
        cases.append({"case": f"{label}（於 `{target.rsplit('/', 1)[-1]}`）",
                      "untested": False,
                      "same": snap == base,
                      "changed": [k for k in base if snap[k] != base[k]]})
    return cases


# ---------------------------------------------------------------- G89

def g89(src: Path, patched: set[str]) -> list[dict]:
    cases = []

    def run(label: str, build) -> None:
        out = SANDBOX / f"g89_{len(cases)}.xlsx"
        build(out)
        try:
            verify_structure(src, out, patched)
            cases.append({"case": label, "raised": False, "msg": ""})
        except StructureError as e:
            cases.append({"case": label, "raised": True, "msg": str(e)[:110]})

    # 正常案例 —— 位元組複製，不得誤拋
    run("正常：位元組複製（不得誤拋）",
        lambda out: clone_zip(src, out))

    # 檢查 1 —— zip member 集合
    run("弄壞 1：刪去一個 zip member（`xl/calcChain.xml` 或次末者）",
        lambda out: clone_zip(src, out, drop=_droppable(src)))

    # 檢查 2 —— 逐分頁 DV 計數
    run("弄壞 2：抹去目標分頁之一條 `dataValidation`",
        lambda out: clone_zip(src, out, mutate=(
            _target_sheet(src),
            lambda b: re.sub(rb"<dataValidation[ >].*?</dataValidation>", b"",
                             b, count=1, flags=re.S))))

    # 檢查 3 —— 相異 part 集合（改動一個未被寫入之 member）
    run("弄壞 3：改動未被寫入之 `xl/styles.xml`",
        lambda out: clone_zip(src, out, mutate=(
            "xl/styles.xml", lambda b: b.replace(b"<styleSheet", b"<styleSheet ", 1))))
    return cases


def _droppable(src: Path) -> str:
    with zipfile.ZipFile(src) as z:
        names = [n for n in z.namelist()
                 if n.startswith("xl/") and "worksheets" not in n
                 and not n.endswith(".rels") and n != "xl/workbook.xml"]
    return names[-1]


def _sheet_containing(src: Path, marker: bytes) -> str:
    """回傳首個含該標記之分頁 XML —— 各屬性所在之分頁不同，須逐類尋找。"""
    with zipfile.ZipFile(src) as z:
        for n in sorted(z.namelist()):
            if n.startswith("xl/worksheets/sheet") and marker in z.read(n):
                return n
    raise RuntimeError(f"找不到含 {marker!r} 之分頁")


def _target_sheet(src: Path) -> str:
    with zipfile.ZipFile(src) as z:
        for n in sorted(z.namelist()):
            if n.startswith("xl/worksheets/sheet") and b"<dataValidation" in z.read(n):
                return n
    raise RuntimeError("找不到含 dataValidation 之分頁 —— 弄壞 2 無法建構")


# ---------------------------------------------------------------- G90

def cell_snapshot(path: Path, sheet: str, rows: range) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    snap = {}
    for row in ws.iter_rows(min_row=rows.start, max_row=rows.stop - 1):
        for c in row:
            if c.value is not None:
                snap[c.coordinate] = str(c.value)
    wb.close()
    return snap


def make_non_blank(src: Path, out: Path, cfg: dict, tcs: list[dict]) -> None:
    """合成「他人既有列」—— 以 surgical_save 寫入前 EXISTING_ROWS 列。"""
    wb = openpyxl.load_workbook(src)
    ws = wb[cfg["workbook"]["sheet"]]
    for i, tc in enumerate(tcs[:EXISTING_ROWS]):
        row = FIRST_DATA_ROW + i
        for letter, value in row_values(tc, cfg, row).items():
            ws[f"{letter}{row}"] = value
        ws[f"{cfg['workbook']['columns']['author']}{row}"] = "SomeoneElse"
    surgical_save(wb, src, out)


def append_rows(base: Path, out: Path, cfg: dict, tcs: list[dict],
                start_row: int) -> dict:
    wb = openpyxl.load_workbook(base)
    ws = wb[cfg["workbook"]["sheet"]]
    for i, tc in enumerate(tcs):
        row = start_row + i
        for letter, value in row_values(tc, cfg, row).items():
            ws[f"{letter}{row}"] = value
    return surgical_save(wb, base, out)


def g90(src: Path, cfg: dict, tcs: list[dict]) -> dict:
    sheet = cfg["workbook"]["sheet"]
    base = SANDBOX / "g90_base.xlsx"
    make_non_blank(src, base, cfg, tcs)
    existing = range(FIRST_DATA_ROW, FIRST_DATA_ROW + EXISTING_ROWS)
    before = cell_snapshot(base, sheet, existing)
    before_hash = hashlib.sha256(
        json.dumps(before, sort_keys=True, ensure_ascii=False).encode()).hexdigest()

    # (a)(b)(c) —— 自既有列之後起始
    ok_start = FIRST_DATA_ROW + EXISTING_ROWS
    out_ok = SANDBOX / "g90_append.xlsx"
    append_rows(base, out_ok, cfg, tcs, ok_start)
    after = cell_snapshot(out_ok, sheet, existing)
    after_hash = hashlib.sha256(
        json.dumps(after, sort_keys=True, ensure_ascii=False).encode()).hexdigest()
    changed = sorted(k for k in set(before) | set(after)
                     if before.get(k) != after.get(k))

    b_col = "B"
    b_seq = []
    wb = openpyxl.load_workbook(out_ok, data_only=True, read_only=True)
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW,
                            max_row=ok_start + len(tcs) - 1,
                            min_col=2, max_col=2):
        b_seq.append(row[0].value)
    wb.close()

    # (d) —— 刻意重疊：自既有列之中段起始
    bad_start = FIRST_DATA_ROW + EXISTING_ROWS - 2
    out_bad = SANDBOX / "g90_overlap.xlsx"
    append_rows(base, out_bad, cfg, tcs, bad_start)
    after_bad = cell_snapshot(out_bad, sheet, existing)
    overlap_changed = sorted(k for k in set(before) | set(after_bad)
                             if before.get(k) != after_bad.get(k))

    return {
        "existing_rows": EXISTING_ROWS,
        "existing_cells": len(before),
        "before_hash": before_hash, "after_hash": after_hash,
        "a_existing_unchanged": before == after,
        "b_new_start_row": ok_start,
        "b_no_overwrite": not changed,
        "c_b_sequence": b_seq,
        "c_b_continuous": b_seq == list(range(1, len(b_seq) + 1)),
        "d_overlap_start_row": bad_start,
        "d_overlap_detected": bool(overlap_changed),
        "d_overlap_changed_cells": overlap_changed[:8],
        "d_overlap_changed_count": len(overlap_changed),
    }


def main() -> None:
    cfg = load_cfg()
    src = FEATURE / cfg["paths"]["workbook"]
    SANDBOX.mkdir(parents=True, exist_ok=True)
    (SANDBOX / ".gitignore").write_text("*\n", encoding="utf-8")
    src_before = sha256(src)

    batch = json.loads((FEATURE / "generated/batch_001_power_down.json")
                       .read_text(encoding="utf-8"))
    tcs = sorted(batch["tcs"], key=lambda t: int(t["tc_id"].split("-")[-1]))

    patched = {_target_sheet(src)}
    cases = g89(src, patched)
    boundary = g90(src, cfg, tcs)
    struct = g95(src, SANDBOX / "g90_append.xlsx")

    result = {"src_sha256_before": src_before, "src_sha256_after": sha256(src),
              "src_untouched": src_before == sha256(src),
              "g89": cases, "g90": boundary, "g95": struct, "tc_count": len(tcs)}
    (DATA / "b2b3_writeback_path.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"來源未被觸碰：{result['src_untouched']}\n")
    print("G89 —— verify_structure 之三項檢查")
    for c in cases:
        expect = "不得拋" if c["case"].startswith("正常") else "須拋"
        ok = (not c["raised"]) if expect == "不得拋" else c["raised"]
        print(f"  [{'PASS' if ok else '**FAIL**'}] {c['case']}（{expect}）"
              f" → raised={c['raised']}")
        if c["msg"]:
            print(f"        {c['msg']}")
    print(f"\nG90 —— append 邊界（既有 {boundary['existing_rows']} 列，"
          f"{boundary['existing_cells']} 格）")
    print(f"  (a) 既有列逐格不變：{boundary['a_existing_unchanged']}"
          f"（hash {boundary['before_hash'][:12]} → {boundary['after_hash'][:12]}）")
    print(f"  (b) 新列自 r{boundary['b_new_start_row']} 起，無覆蓋："
          f"{boundary['b_no_overwrite']}")
    print(f"  (c) B 欄序號連續：{boundary['c_b_continuous']} —— {boundary['c_b_sequence']}")
    print(f"  (d) 刻意自 r{boundary['d_overlap_start_row']} 起（重疊）→ "
          f"既有列被改動 {boundary['d_overlap_changed_count']} 格："
          f"{boundary['d_overlap_detected']}")
    print("\nG95 —— 目標分頁內之結構屬性（merges / cf / DV / cols / panes）")
    for c in struct:
        if c.get("untested"):
            print(f"  [未實測] {c['case']} —— {c['why']}")
            continue
        expect = "須相同" if c["case"].startswith("正常") else "須偵出差異"
        ok = c["same"] if expect == "須相同" else not c["same"]
        print(f"  [{'PASS' if ok else '**FAIL**'}] {c['case']}（{expect}）"
              f" → same={c['same']} 差異類別={c['changed']}")


if __name__ == "__main__":
    main()
