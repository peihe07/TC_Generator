"""B4 —— 寫回（R-P310）。

**授權**：R-P309（Pei 逐字「授權啊」），其效力範圍限於**授權當時之 260 條內容**
加上 R-P308 之 6 條 `axis` 指派。

**嚴守 R-P310 之四段程序與 §I**：
  - 寫入路徑為 `xlsx_surgical.py` 之 `surgical_save()`（唯一授權之路徑，R16 / R-G3）；
    **全域無 `Workbook.save()`**
  - **不得對 `inputs/` 之原始檔寫入** —— 其為唯讀來源，位元組複製後寫副本
  - 交付副本置於 **`output/`**（非客戶樹、非 `inputs/`）
  - 最終 `tc_id` 於此刻依 R-P113(c) 統一指派，序為 `(SWE-PM ID, split_index)`（R-P115）
  - `SWE-PM-089` 之留白列依 R-P141 保留，僅填 `req_id`
  - **不送達客戶目錄、不執行任何 git 操作**

用法：
    python features/power/scripts/write_back_47.py
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[3]
FEATURE = ROOT / "features/power"
OUTPUT = ROOT / "output"
DATA = FEATURE / "data"
sys.path.insert(0, str(Path(__file__).resolve().parent))

from dryrun_write_back import (HEADER_ROW, FIRST_DATA_ROW, load_cfg,  # noqa: E402
                               row_values, dv_snapshot, structure_snapshot, sha256)
from dryrun_full_write_back import ordered_tcs, pane_and_widths, BLANK_LEAF  # noqa: E402

PREFIX = "NR1L-PowerManagement-"


def main() -> None:
    cfg = load_cfg()
    src = FEATURE / cfg["paths"]["workbook"]
    OUTPUT.mkdir(parents=True, exist_ok=True)
    out = OUTPUT / src.name

    src_sha = sha256(src)
    work = OUTPUT / f".work_{src.name}"
    shutil.copyfile(src, work)                       # 位元組複製；來源唯讀
    assert sha256(work) == src_sha, "副本非位元組相同"

    tcs = ordered_tcs()                              # (SWE-PM ID, split_index) 序
    blank = {"req_id": BLANK_LEAF, "tc_id": "", "tc_title": "", "test_item": "",
             "pre_conditions": "", "input_test_data": "", "test_procedure": "",
             "expected_result": "", "specification_reference": "", "priority": "",
             "design_method": "", "split_flag": "", "split_reason": "",
             "functional_safety": "", "estimated_test_time": "", "remarks": "",
             "split_index": 1}
    pos = next((i for i, t in enumerate(tcs)
                if int(re.match(r"SWE-PM-(\d+)", t["req_id"]).group(1)) > 89), len(tcs))
    rows = tcs[:pos] + [blank] + tcs[pos:]

    # ── 最終 tc_id 之指派（R-P113(c) / R-P115），於此刻為之 ──
    mapping = []
    n = 0
    for t in rows:
        if t is blank:
            continue
        n += 1
        final = f"{PREFIX}{n:03d}"
        mapping.append((t["tc_id"], final, t["req_id"], t.get("split_index")))
        t["_final_tc_id"] = final

    before_dv, before_st = dv_snapshot(work), structure_snapshot(work)
    before_pw = pane_and_widths(work)

    sys.path.insert(0, str(ROOT / "backend"))
    from xlsx_surgical import surgical_save
    wb = openpyxl.load_workbook(work)
    ws = wb[cfg["workbook"]["sheet"]]
    tc_col = cfg["workbook"]["columns"]["tc_id"]
    for i, t in enumerate(rows):
        r = FIRST_DATA_ROW + i
        for letter, value in row_values(t, cfg, r).items():
            ws[f"{letter}{r}"] = value
        if t is not blank:
            ws[f"{tc_col}{r}"] = t["_final_tc_id"]   # 最終號覆寫臨時號
    surgical_save(wb, work, out)                     # **不呼叫 wb.save()**
    work.unlink()

    after_dv, after_st = dv_snapshot(out), structure_snapshot(out)
    after_pw = pane_and_widths(out)

    # ── 寫入後驗證（R-P310(三)）──
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2[cfg["workbook"]["sheet"]]
    req_col = cfg["workbook"]["columns"]["req_id"]
    bad_b, bad_order, bad_id = [], [], []
    for i, t in enumerate(rows):
        r = FIRST_DATA_ROW + i
        if ws2[f"B{r}"].value != r - HEADER_ROW:
            bad_b.append((r, ws2[f"B{r}"].value))
        if str(ws2[f"{req_col}{r}"].value or "") != t["req_id"]:
            bad_order.append((r, ws2[f"{req_col}{r}"].value, t["req_id"]))
        want = "" if t is blank else t["_final_tc_id"]
        if str(ws2[f"{tc_col}{r}"].value or "") != want:
            bad_id.append((r, ws2[f"{tc_col}{r}"].value, want))

    checks = {
        "DV（含 x14）未變": before_dv == after_dv,
        "合併儲存格未變": before_st["merges"] == after_st["merges"],
        "條件式格式未變": before_st["cf"] == after_st["cf"],
        "分頁清單未變": before_st["sheets"] == after_st["sheets"],
        "壓縮成員清單未變": before_st["members"] == after_st["members"],
        "凍結窗格未變（G95）": before_pw["panes"] == after_pw["panes"],
        "欄寬未變（G95）": before_pw["cols"] == after_pw["cols"],
        "B 欄序號逐列相符": not bad_b,
        "列序依 (SWE-PM ID, split_index)": not bad_order,
        "最終 tc_id 逐列相符（001–260 連號）": not bad_id,
        "**來源原始檔未被改動**": sha256(src) == src_sha,
    }
    ok = all(checks.values())

    md = ["# B4 —— 寫回紀錄（R-P310）\n",
          "\n> **授權**：R-P309（Pei 逐字「授權啊」）。\n",
          "> 寫入路徑為 `surgical_save()`；**全域無 `Workbook.save()`**。\n",
          "> **未對 `inputs/` 之原始檔寫入**；**未送達客戶目錄**；**未執行任何 git 操作**。\n",
          f"\n## 一、寫入對象\n\n| 項 | 值 |\n|---|---|\n"
          f"| 來源（唯讀） | `{src.relative_to(ROOT)}` |\n"
          f"| 來源 SHA256（寫入前） | `{src_sha}` |\n"
          f"| 來源 SHA256（寫入後） | `{sha256(src)}` |\n"
          f"| **交付副本** | `{out.relative_to(ROOT)}` |\n"
          f"| 副本 SHA256 | `{sha256(out)}` |\n"
          f"| 副本大小 | {out.stat().st_size:,} bytes |\n"
          f"| 寫入列數 | **{len(rows)}**（TC {len(tcs)} ＋ 留白 1） |\n"
          f"| 列範圍 | {FIRST_DATA_ROW} – {FIRST_DATA_ROW + len(rows) - 1} |\n",
          "\n## 二、寫入後驗證\n\n| 項 | 結果 |\n|---|---|\n"]
    for k, v in checks.items():
        md.append(f"| {k} | {'**PASS**' if v else '**FAIL**'} |\n")
    md.append(f"\n## 三、XML 層 diff —— 相異之 part\n\n")
    diff_parts = sorted(set(before_st["members"]) ^ set(after_st["members"]))
    md.append(f"壓縮成員之增減：{diff_parts or '**無**'}\n\n")
    md.append(f"DV 條數：{len(before_dv)} → {len(after_dv)}"
              f"（x14：{sum(1 for x in before_dv if x[0]=='x14')} → "
              f"{sum(1 for x in after_dv if x[0]=='x14')}）\n")
    (DATA / "writeback_47.md").write_text("".join(md), encoding="utf-8")

    # 最終號對照
    with (DATA / "final_tc_id_map_47.tsv").open("w", encoding="utf-8") as f:
        f.write("provisional_tc_id\tfinal_tc_id\treq_id\tsplit_index\n")
        for prov, fin, req, si in mapping:
            f.write(f"{prov}\t{fin}\t{req}\t{si}\n")

    print(f"交付副本：{out.relative_to(ROOT)}")
    print(f"寫入 {len(rows)} 列（TC {len(tcs)} ＋ 留白 1）")
    for k, v in checks.items():
        print(f"  [{'PASS' if v else '**FAIL**'}] {k}")
    print(f"\n最終號對照 {len(mapping)} 列 → data/final_tc_id_map_47.tsv")
    raise SystemExit(0 if ok else 1)


if __name__ == "__main__":
    main()
