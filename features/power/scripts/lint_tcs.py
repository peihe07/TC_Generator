"""Power lint —— R-P42 錨點層範圍上界之執行期閘門（R-P52）。

R-P42 規範 Phase 4 撰寫 TC 之行為：某 leaf 之 TC 範圍上界為其
`Source Requirement ID` 所實際引用之需求錨點；未被引用之錨點一律不測。
本檔把該條文變成可機械檢查之閘門。

閘門 `check_rp42_unreferenced_anchor`：

  (a) 任一 TC 之 `specification_reference` 命中黑名單錨點 id → FAIL
  (b) 任一 TC 之內容（`pre_conditions` / `test_procedure` / `expected_result`
      / `test_item` / `input_test_data`）逐字引用黑名單錨點之**特徵字串** → FAIL

**黑名單為腳本導出，非人工維護**（R-P52(c)）——
讀 `data/unreferenced_anchors.tsv`，該檔由 `build_blacklist.py` 自
`layer3_full.tsv` 與兩份 CFTS 之全部錨點導出，leaf 增修時自動跟隨。

### 特徵字串之抽取方式（執行層自裁，見 07 上繳包 §三）

對每個黑名單錨點取其內文，切為句子，保留長度 ≥ MIN_FINGERPRINT 之句子，
再**扣除任何也出現於被引用錨點內文中的句子** —— 剩下者即為該錨點獨有之
特徵字串。扣除這一步是必要的：兩個錨點常共用樣板句（例如
「In the following "Ignition Working Conditions": ...」），
若不扣除，合法引用被引用錨點之 TC 會被誤判。

### 08 包新增四閘（R-P60）

  G37 R-P1  —— `SWE-PM-089` 不得有 TC，其來源欄不得被填補
  G38 R-P2  —— tc_id 格式、唯一、單調遞增、無跳號
  G39 R-P8  —— priority 值域 P0–P3，且不得與 037 `Priority` 呈一對一映射
  G40 R-P35 —— Test Set 須為五個定版值之一、逐 leaf 符 `leaf_testset.tsv`、
                全體分布 63 / 24 / 16 / 8 / 3

**權威來源之一項登記**：`feature.yaml` 自 scaffold 以來未更新 ——
其 `spec_mode: "A"` 與 R-P9 / R-P3′（= D）不符、`test_group: "Power"`
與 R-P2（workbook Test Group = "Power Management"）不符，且無 tc_id 格式欄。
本檔之權威值因此直接取自裁決條文並於此註明，**未逕改 feature.yaml**（見 A-PW36）。

用法：
    python features/power/scripts/lint_tcs.py                # 檢查 generated/*.json
    python features/power/scripts/lint_tcs.py --self-test    # 以合成 fixture 驗全部閘門
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from extract_textlayer import REQ_RE, SEC_RE, paragraphs  # noqa: E402

ROOT = Path(__file__).resolve().parents[3]
IN = ROOT / "features/power/inputs"
DATA = ROOT / "features/power/data"
GENERATED = ROOT / "features/power/generated"

# 特徵字串之最短長度。過短會與樣板句碰撞，過長會漏抓。
MIN_FINGERPRINT = 40

# 受檢之 TC 內容欄位（R-P52(b)）
CONTENT_FIELDS = (
    "test_item", "pre_conditions", "input_test_data",
    "test_procedure", "expected_result",
)

# `\b` 在底線前不成立（`_` 為 word 字元），而 spec_reference 慣以底線串接，
# 故以「前後皆非數字」界定，勿用 \b —— 此為 G33 fixture (a) 首次執行時抓到之 bug。
ANCHOR_ID_RE = re.compile(r"(?<!\d)\d{6,8}(?!\d)")
SENTENCE_SPLIT_RE = re.compile(r"(?<=[.;:])\s+|\n+")

# ── 08 包新增之權威值（取自裁決條文，非 feature.yaml —— 見 docstring / A-PW36）──
BLOCKED_LEAF = "SWE-PM-089"                                   # R-P1
TC_ID_RE = re.compile(r"^NR1L-PowerManagement-(\d{3})$")       # R-P2
VALID_PRIORITIES = {"P0", "P1", "P2", "P3"}                   # R-P8 / §10.2
LOCKED_DISTRIBUTION = {                                       # R-P35
    "Power State": 63, "Startup Display": 24,
    "Branding and Theme": 16, "Timeout Settings": 8, "Power Down": 3,
}
LOCKED_TOTAL = sum(LOCKED_DISTRIBUTION.values())              # 114

# G45 / §10.7：specification_reference 須非空，每項形如 {spec_filename}_{section_id}。
# **檔名部分得含空白** —— §10.7 之範例即為
# `Menu Bar and App Drawer HMI Logic and Flow R1 SR24 3A (September 11 2023)_2.5`，
# 而本 feature 之 CFTS010 檔名亦含空格。
# 初版寫 `^\S+_...` 會把每一條合法 TC 都判 FAIL —— 由 R-P71 之真實檔案 lint 當場抓到。
SPEC_REF_ITEM_RE = re.compile(r"^.+_\d+(?:\.\d+)*$")

# G46 / R-P69(c)：feature.yaml 須與裁決條文一致
YAML_EXPECTED = {
    "spec_mode": ("D", "R-P9 / R-P3′"),
    "test_group": ("Power Management", "R-P2（workbook Test Group）"),
    "tc_id_format": ("NR1L-PowerManagement-{NNN}", "R-P2"),
    # R-P77 —— 兩項可疑值之訂正一併納入一致性檢查
    "author_value": ("", "R-P77(a)：填佔位值會靜默匹配零列"),
    "fill_test_group_set": ("true", "R-P77(b)：本 workbook 為 BLANK"),
}

# G50 / §11：四個長欄位無 trailing period；UI 標籤用雙引號。
#
# **方括號之兩項豁免依 profile 條款，非硬編碼**（R-P82 / G59）——
# 見 `docs/runtime/profiles/FW036_R1L_Power_Profile.md`：
#   §3.2 訊號值記法 `[Nh]`（逐字引自規格，如 `[1h]` / `[0h]`）
#   §3.1 行首之 source-class 標記（`[spec-derived]` 等，§3.2 既有慣例）
# 二者之正則於下方定義，並於 PROFILE_PATH 存在時方生效 ——
# profile 不存在即無豁免依據，方括號一律違規。
PROFILE_PATH = ROOT / "docs/runtime/profiles/FW036_R1L_Power_Profile.md"

# ── profile 條款之可機械檢查值（R-P91 / G71 / G72）──
PROFILE_TEST_GROUP = "Power Management"                       # profile §2
DESIGN_METHODS = {                                            # profile §3.3
    "功能測試 (Functional based ; no specific technique)",
    "狀態轉換 (State Transition Testing)",
    "決策表 (Decision Table Testing)",
    "等價劃分 (Equivalence Partitioning, EP)",
    "邊界值分析 (Boundary Value Analysis, BVA)",
    "組合測試 (Combinatorial Testing ; Pairwise / t-wise)",
    "情境 / 用例 (Scenario / Use Case Testing)",
    "負向測試 (Negative / Invalid)",
    "基礎故障注入 (Fault Injection Lite)",
}
SPEC_STEMS = {                                                # profile §3.4
    "R1LR_Atl-H_25PI3.5_Activation and Configuration_CFTS_009_"
    "Wake-up and Power-up_SR26_20250909-1658",
    "R1LR_Atl-H_25PI3.5_Activation and Configuration_CFTS_010_"
    "Power Down _SR26_20250909-1658",
}
# G71：`workbook.columns` 之各鍵應命中之 r9 標頭關鍵字
COLUMN_HEADER_KEYWORD = {
    "req_id": "Requirement or Design ID", "tc_id": "Test Case ID",
    "test_group": "Test Group", "test_set": "Test Set", "test_item": "Test Item",
    "pre_conditions": "Pre-Conditions", "input_test_data": "Input Test Data",
    "test_procedure": "Test procedure", "expected_result": "Expected Result",
    "spec_reference": "Specification Reference", "tc_ref_id": "Test Case Reference ID",
    "estimated_time": "Estimated Test Time", "priority": "Test Case Priority",
    "design_method": "Test Case Design", "functional_safety": "Functional Safety",
    "author": "Test Case Author", "remarks": "Remarks",
}
PROFILE_EXEMPTS_BRACKETS = PROFILE_PATH.exists()
# R-P131（18 包）：`remarks` 補入。A-PW88 —— 該欄原不在任何格式閘門之視野內，
# 致 17 包 G92（`015` 之 DR-PW8 標記）之 PASS 為判準空洞。
LONG_FIELDS = ("pre_conditions", "input_test_data", "test_procedure",
               "expected_result", "remarks")
SIGNAL_BRACKET_RE = re.compile(r"\[[0-9A-Fa-f]+h\]")      # [1h] / [0h] — 訊號值
# 行首之 source-class 標記（§3.2）為既有慣例 —— Comfort 之已交付 TC 即以
# `1. [spec-derived] …` 書寫。§11 之方括號禁令不及於此，故先行剝除再檢查。
SOURCE_CLASS_RE = re.compile(r"^\s*\d+\.\s*\[[a-z-]+\]\s*")
BAD_BRACKET_RE = re.compile(r"\[[^\]]*\]")
BAD_QUOTE_RE = re.compile(r"(?<![A-Za-z])'[^']{2,}'|<[A-Za-z][^>]*>")

# G51 / §4.4：Pre-Condition 為狀態／環境，不得含動作。
#
# **R-P83：判準改以經驗基礎導出。** 11 包 B4 自 Comfort + Privacy 之已交付
# `test_procedure`（依定義即為動作）取行首動詞聯集，得 20 個（出現 ≥ 3 次）；
# 再以其已交付 `pre_conditions` 1823 行量偽陽性 —— **人工與經驗清單皆為 0**。
# 因二者偽陽性同為零，取**聯集**以最大化涵蓋而無已量測之代價。
#
# 09 包之人工清單有 12 個動詞未涵蓋（adjust / change / count / do / move /
# note / operate / put / toggle / touch / turn / wait），皆實際出現於已交付
# procedure；另 13 個人工列舉者未見於該語料，**非誤列**，僅為未經佐證。
PRECOND_ACTION_VERBS = sorted({
    # 09 包人工列舉（保留）
    "insert", "press", "connect", "disconnect", "check", "confirm", "verify",
    "open", "select", "start", "send", "set", "launch", "navigate", "enter",
    "tap", "click", "read", "record", "compare", "inject", "trigger",
    # 11 包 B4 自已交付 procedure 導出之補充
    "adjust", "change", "count", "do", "move", "note", "operate", "put",
    "toggle", "touch", "turn", "wait",
})
# G64 / §4.4 + §8.5（R-P88）：系統預設與環境穩定性前提。
#
# **詞彙之經驗基礎為 canon 原文本身**（R-P88 禁憑印象列舉）——
# 取自 `docs/runtime/ASPICE_SWE6_AI_Instruction.md`：
#   §4.4 Forbidden 之逐字範例：`HU is powered on.`（system defaults）
#   §8.5 逐字：「Environmental stability conditions owned by other RDs do NOT
#              belong in Pre-Condition — testers naturally ensure the
#              environment is stable before execution.」
# 由該二處導出兩類形態：
#   (1) 供電／開機之系統預設 —— powered / power(ed) on / supplied
#   (2) 環境穩定性 —— stable / steady / normal operating / properly connected
ENV_STABILITY_RE = re.compile(
    r"\b(?:power(?:ed)?\s+(?:on|up|from|by)|is\s+powered|stable|steady|"
    r"normal\s+operating|properly\s+(?:connected|configured)|"
    r"functioning\s+normally|works?\s+normally)\b",
    re.I,
)

# G65 / §4.5（R-P89）：input_test_data 不得與 procedure / pre_conditions 重複。
# 詞（≥3 字元）**或任何數字串** —— 初版之 `{3,}` 會把 `25` 這類兩位數值排除，
# 導致「值未重複而僅措詞重複」被誤判為跨欄重複（12 包真實 lint 當場抓到）。
DATA_TOKEN_RE = re.compile(r"[A-Za-z_.$][A-Za-z0-9_.$]{2,}|\d+")

# G50 之表格偵測（§11 / R-P93）：`|` 分隔之 Markdown 表格或 HTML <table>
TABLE_RE = re.compile(r"\|[^|\n]*\|[^|\n]*\||<\s*(?:table|tr|td|th)\b", re.I)

# G81 / R-P107：個案型誤讀關鍵詞黑名單，掃**全部**欄位。
#
# **本閘為個案型** —— 其價值在於證明 `006` 之時序誤讀確已清除，
# **不宣稱可攔下未來之其他誤讀**（R-P107 明訂）。
# 關鍵詞取自該次誤讀之實際字串，非憑印象列舉。
MISREAD_TERMS = [
    ("after boot completes", "R-P103：規格載 while the boot is still completing"),
    ("after the boot completes", "同上"),
    ("開機完成後", "同上"),
    ("follow the injected order", "R-P108：4942338 未載按注入順序處理"),
    ("按注入順序", "同上"),
]
# 掃描對象：TC 之全部欄位（含 13 欄以外者）＋ leaf 之 reasoning / reasoning_note。
# **`source_clause` 為規格原文，不納入黑名單掃描** —— 原文若含該詞即為事實。
TC_SCAN_SKIP = {"source_clause"}
# 查證記錄本身須引述被否定之措詞。引號內之出現視為**引述**，不判違。
QUOTED_SPAN_RE = re.compile(r"「[^」]*」|`[^`]*`|“[^”]*”")

# G82 / R-P109：ER 所斷言之行為，其規格依據須完整出現於 source_clause。
# 判準：ER 之**專有標的**（訊號名、全大寫識別子、_Time 類參數、數值）
# 須於 source_clause 出現。一般英文詞不納入 —— 措詞本就不會逐字相同。
ER_PROPER_RE = re.compile(
    r"\b(?:[A-Z][A-Za-z]*(?:_[A-Za-z0-9]+)+|[A-Z]{3,}(?:\.[A-Za-z0-9_]+)?|\d+)\b")
# 非規格標的之常見誤命中（測試環境用語與本 TC 自身編號）
ER_PROPER_SKIP = {"TLM", "NA", "AUD", "LVL", "ICS", "HVAC", "ACN", "CAN", "LIN"}


# G77 / §5.2B + §5.5（R-P101）：`test_procedure` 之末步須含驗證意圖措詞。
#
# **語料實測（14 包 B4，R-P101 令以已交付 `test_procedure` 末步為語料）**——
# Comfort + Privacy 已交付共 **472** 條之末步：
#   `check` **0** / `verify` **0** / `confirm` **0** / `ensure` **0** / `observe` **0**
#   `read`  **243（51.5%）**；無任何驗證措詞者 **232（49.2%）**
#
# **即 §5.2B 之措詞在已交付實務中 0 / 472 attested。**
# 已交付件之末步慣例為「Read <具體可觀察標的>」——
# 以「所讀之標的」滿足 §5.5「Final Step 自身即揭示所檢查者」，不另加子句。
# 本閘依 R-P101 之明令仍要求 §5.2B 措詞（其判準明確、可機械判定，
# 與 G73 之「無法與合法回讀區分」性質不同），
# 惟**須明載：Power 之末步慣例將因此與 Comfort / Privacy 分歧**（A-PW67）。
FINAL_STEP_INTENT_RE = re.compile(
    r"\b(?:check(?:s|ed)?\s+(?:that|whether)|to\s+check|and\s+check|"
    r"verif(?:y|ies|ied)\s+that|to\s+verify|and\s+verify|"
    r"confirm(?:s|ed)?\s+that|to\s+confirm|and\s+confirm)\b", re.I)


# G73 / §6（R-P96）：ER 行不得為 procedure 動作之複述。
#
# **判準之經驗基礎**（R-P96 禁憑印象列舉，比照 R-P83 / R-P88）——
# 13 包 B3 自 Comfort + Privacy 之已交付件取 proc 與 ER 皆 1:1 對齊者，
# 得 **1076 組 (步驟, ER 行)** 語料（含輕量字尾正規化），量測
# `overlap = |ER 實詞 ∩ proc 實詞| / |ER 實詞|`：
#   P50 0.500 / P75 0.667 / P80 0.750 / P90 1.000 / 最大 1.000
# 門檻取 **P50 = 0.50**：本閘為非阻斷之待裁類，recall 重於 precision ——
# 取 P80（0.75）會漏掉 R-P96 所舉之 `004` ER2（overlap 0.60）。
#   overlap = 1.00 者 **120 組（11.2%）**
#
# **實測結論（決定本閘之型別）：R-P96 之判準無法機械化為阻斷閘。**
#   tier 1（動作述語 ＋ overlap ≥ P50）於已交付語料觸發 **69 / 1076（6.4%）**
#   tier 2（overlap = 1.00）觸發 **120 / 1076（11.2%）**
# 其形態為「Select the rear Feet mode → The rear Feet mode is selected」，
# 即 §6 所稱之「prove condition established」狀態回讀 —— 可失敗、具判讀價值。
# 更甚者，已交付 Privacy 之 ER 含「The output volume is read」、
# 「The state of the speed controlled volume is recorded」，
# **與 R-P96 所舉之五例同形**。
# 故 G73 **全部列為待人工裁決類**（比照 R-P76 之 R-P42(b)），不使 exit=1。
ER_OVERLAP_P50 = 0.50
# 功能詞（標準停用詞，非本判準之發明）
ER_STOPWORDS = {
    "the", "a", "an", "is", "are", "was", "were", "be", "been", "and", "or",
    "of", "to", "in", "on", "at", "for", "from", "with", "by", "that", "this",
    "it", "its", "as", "into", "until", "while", "then", "again", "still",
    "no", "not", "all", "any", "each", "every", "both", "same", "other",
    "shall", "should", "will", "can", "may", "must", "has", "have", "had",
    "there", "which", "when", "if", "so", "than", "up", "out", "over",
}
ER_WORD_RE = re.compile(r"[A-Za-z][A-Za-z0-9_.$'-]*")
# 執行者動作之述語形態：`is/are <verb>ed`、`<verb>s`、`<verb>ing`
ER_ACTOR_RE = re.compile(
    r"\b(?:is|are)\s+(?:" + "|".join(v + "(?:e?d)?" for v in PRECOND_ACTION_VERBS) + r")\b"
    r"|\b(?:" + "|".join(v + "s" for v in PRECOND_ACTION_VERBS) + r")\b",
    re.I,
)

# G74 / §8.4.1（R-P97）：時間量測之 ER 不得寫數值相等。
# 形態取自 R-P97 所引之兩處實例（`001` / `004` 之 ER3）：
#   「The recorded time **equals** SplashScreen_Time」
# 即「時間類名詞 … equals/is equal to/matches/is exactly … 時間類識別子」。
TIME_TOKEN_RE = re.compile(
    r"\b(?:elapsed\s+time|recorded\s+time|measured\s+time|duration|timer|"
    r"\w*_?Time\b|\d+\s*(?:ms|milliseconds?|s|seconds?|min|minutes?))\b", re.I)
TIME_EQUALITY_RE = re.compile(
    r"\b(?:equals?|is\s+equal\s+to|are\s+equal\s+to|matches|is\s+exactly)\b", re.I)

PRECOND_ACTION_RE = re.compile(
    r"^\s*\d+\.\s*(?:\[[^\]]*\]\s*)?(" + "|".join(PRECOND_ACTION_VERBS) + r")\b",
    re.I,
)


def find(pattern: str) -> Path:
    return next(f for f in IN.iterdir() if pattern in f.name)


def load_blacklist() -> dict[str, tuple[str, str, str]]:
    """anchor_id -> (cfts, chapter_num, chapter_title)。"""
    path = DATA / "unreferenced_anchors.tsv"
    if not path.exists():
        raise SystemExit(
            f"缺 {path.relative_to(ROOT)} —— 先跑 build_blacklist.py（R-P52(c)：黑名單須腳本導出）"
        )
    out = {}
    for line in path.read_text(encoding="utf-8").splitlines()[1:]:
        if line.strip():
            anchor, cfts, num, title, _touched, _tier = line.split("\t")
            out[anchor] = (cfts, num, title)
    return out


def anchor_bodies() -> dict[str, list[str]]:
    """anchor_id -> 其後之內文段落（至下一錨點或章節錨點止）。"""
    bodies: dict[str, list[str]] = {}
    for path in [find("CFTS_009_Wake-up"),
                 next(x for x in IN.iterdir() if x.suffix == ".doc")]:
        current = None
        for plain, bold in paragraphs(path):
            if SEC_RE.match(plain):
                current = None
                continue
            found = REQ_RE.findall(bold)
            if found:
                current = found[0]
                bodies.setdefault(current, [])
            elif current and plain.strip():
                bodies[current].append(plain.strip())
    return bodies


def sentences(text: str) -> list[str]:
    return [s.strip() for s in SENTENCE_SPLIT_RE.split(text) if len(s.strip()) >= MIN_FINGERPRINT]


def build_fingerprints(blacklist: dict, bodies: dict[str, list[str]]) -> dict[str, list[str]]:
    """黑名單錨點獨有之特徵句 —— 扣除任何也出現於被引用錨點內文者。"""
    cited_sentences: set[str] = set()
    for anchor, lines in bodies.items():
        if anchor not in blacklist:
            for line in lines:
                cited_sentences.update(sentences(line))

    out: dict[str, list[str]] = {}
    for anchor in blacklist:
        unique = []
        for line in bodies.get(anchor, []):
            for s in sentences(line):
                if s not in cited_sentences:
                    unique.append(s)
        if unique:
            out[anchor] = unique
    return out


def check_rp42_unreferenced_anchor(
    tcs: list[dict], blacklist: dict, fingerprints: dict[str, list[str]]
) -> list[dict]:
    """回傳 findings；空 list = PASS。"""
    findings = []
    for tc in tcs:
        tc_id = tc.get("tc_id") or tc.get("req_id") or "(無 id)"

        # (a) specification_reference 命中黑名單錨點 id
        for anchor in ANCHOR_ID_RE.findall(str(tc.get("specification_reference", ""))):
            if anchor in blacklist:
                cfts, num, title = blacklist[anchor]
                findings.append({
                    "rule": "R-P42(a)", "tc_id": tc_id, "anchor": anchor,
                    "detail": f"specification_reference 命中未被引用之錨點 "
                              f"`{anchor}`（CFTS{cfts} §{num} {title}）",
                })

        # (b) 內容逐字引用黑名單錨點之特徵字串
        content = "\n".join(str(tc.get(f, "")) for f in CONTENT_FIELDS)
        for anchor, prints in fingerprints.items():
            for fp in prints:
                if fp in content:
                    cfts, num, title = blacklist[anchor]
                    findings.append({
                        "rule": "R-P42(b)", "tc_id": tc_id, "anchor": anchor,
                        "detail": f"內容逐字引用未被引用之錨點 `{anchor}`"
                                  f"（CFTS{cfts} §{num} {title}）之特徵字串："
                                  f"「{fp[:70]}…」",
                    })
                    break
    return findings


def leaf_of(tc: dict) -> str:
    """自 req_id / parent 取 leaf id（`SWE-PM-057-02` → `SWE-PM-057`）。"""
    raw = str(tc.get("req_id") or tc.get("parent") or "")
    m = re.match(r"(SWE-PM-\d+)", raw)
    return m.group(1) if m else raw


def load_leaf_testset() -> dict[str, str]:
    path = DATA / "leaf_testset.tsv"
    return {
        line.split("\t")[0]: line.split("\t")[1]
        for line in path.read_text(encoding="utf-8").splitlines()[1:] if line.strip()
    }


def load_037_priority() -> dict[str, str]:
    """037 `Priority` 欄 —— 僅供 G39 之「不得一對一映射」比對（R-P8）。"""
    import openpyxl
    wb = openpyxl.load_workbook(find("FSM-037"), data_only=True, read_only=True)
    ws = wb["SWE1 Requirements"]
    out = {
        str(r[0]).strip(): str(r[15] or "").strip()
        for r in ws.iter_rows(min_row=8, max_row=145, values_only=True)
        if r[0] and str(r[0]).strip()
    }
    wb.close()
    return out


def check_rp1_blocked_leaf(tcs: list[dict]) -> list[dict]:
    """G37 / R-P1：`SWE-PM-089` 不得有 TC，其來源欄不得被填補。"""
    findings = []
    for tc in tcs:
        tc_id = tc.get("tc_id") or tc.get("req_id") or "(無 id)"
        if leaf_of(tc) == BLOCKED_LEAF:
            findings.append({
                "rule": "R-P1", "tc_id": tc_id,
                "detail": f"{BLOCKED_LEAF} 之 TC 存在 —— R-P1 裁定該 leaf 留空待 DR-PW1",
            })
        elif BLOCKED_LEAF in str(tc.get("specification_reference", "")):
            findings.append({
                "rule": "R-P1", "tc_id": tc_id,
                "detail": f"specification_reference 引用 {BLOCKED_LEAF}，其來源尚未裁定",
            })
    return findings


def check_rp2_tc_id(tcs: list[dict]) -> list[dict]:
    """G38 / R-P2：tc_id 格式、唯一、單調遞增、無跳號。"""
    findings, seen, numbers = [], {}, []
    previous = None
    for tc in tcs:
        tc_id = str(tc.get("tc_id", ""))
        m = TC_ID_RE.match(tc_id)
        if not m:
            findings.append({
                "rule": "R-P2", "tc_id": tc_id or "(空)",
                "detail": f"tc_id 不符 R-P2 之格式 NR1L-PowerManagement-{{NNN}}",
            })
            continue
        n = int(m.group(1))
        if tc_id in seen:
            findings.append({"rule": "R-P2", "tc_id": tc_id,
                             "detail": f"tc_id 重複（前次見於 {seen[tc_id]}）"})
        seen[tc_id] = tc.get("_file", "?")
        if previous is not None and n <= previous:
            findings.append({"rule": "R-P2", "tc_id": tc_id,
                             "detail": f"tc_id 未單調遞增（前一個為 {previous:03d}）"})
        previous = n
        numbers.append(n)
    if numbers:
        expected = set(range(min(numbers), min(numbers) + len(set(numbers))))
        gaps = sorted(expected - set(numbers))
        if gaps:
            findings.append({
                "rule": "R-P2", "tc_id": "(全體)",
                "detail": f"tc_id 有跳號：{', '.join(f'{g:03d}' for g in gaps[:10])}",
            })
    return findings


def check_rp8_priority(tcs: list[dict], leaf_priority: dict[str, str]) -> list[dict]:
    """G39 / R-P8：priority 值域；且不得與 037 `Priority` 呈一對一映射。"""
    findings = []
    for tc in tcs:
        p = str(tc.get("priority", "")).strip()
        if p not in VALID_PRIORITIES:
            findings.append({
                "rule": "R-P8", "tc_id": tc.get("tc_id", "(無 id)"),
                "detail": f"priority `{p}` 不在 P0–P3 之內",
            })

    # 一對一映射偵測：037 之每個 Priority 值是否恰對應單一 TC priority
    mapping: dict[str, set] = defaultdict(set)
    for tc in tcs:
        src = leaf_priority.get(leaf_of(tc))
        p = str(tc.get("priority", "")).strip()
        if src and p in VALID_PRIORITIES:
            mapping[src].add(p)
    live = {k: v for k, v in mapping.items() if v}
    if len(live) >= 2 and all(len(v) == 1 for v in live.values()):
        targets = [next(iter(v)) for v in live.values()]
        if len(set(targets)) == len(targets):
            findings.append({
                "rule": "R-P8", "tc_id": "(全體)",
                "detail": "priority 與 037 `Priority` 欄呈一對一映射（"
                          + "、".join(f"{k}→{next(iter(v))}" for k, v in live.items())
                          + "）—— R-P8 裁定 037 Priority 不具映射權威",
            })
    return findings


def check_rp35_test_set(tcs: list[dict], leaf_testset: dict[str, str]) -> list[dict]:
    """G40 / R-P35：Test Set 值域、逐 leaf 相符、全體分布。"""
    findings = []
    for tc in tcs:
        ts = str(tc.get("test_set", "")).strip()
        tc_id = tc.get("tc_id", "(無 id)")
        if ts not in LOCKED_DISTRIBUTION:
            findings.append({"rule": "R-P35", "tc_id": tc_id,
                             "detail": f"Test Set `{ts}` 不在五個定版值之內"})
            continue
        expected = leaf_testset.get(leaf_of(tc))
        if expected and ts != expected:
            findings.append({
                "rule": "R-P35", "tc_id": tc_id,
                "detail": f"{leaf_of(tc)} 之 Test Set 為 `{ts}`，"
                          f"與 leaf_testset.tsv 之 `{expected}` 不符",
            })

    # R-P68：改為子集檢查 —— 每批皆驗，不再只在 114 齊備時判定
    covered = Counter()
    for tc in tcs:
        ts = str(tc.get("test_set", "")).strip()
        if ts in LOCKED_DISTRIBUTION:
            covered[ts] = len({leaf_of(t) for t in tcs
                               if str(t.get("test_set", "")).strip() == ts})
    for name, want in LOCKED_DISTRIBUTION.items():
        got = covered.get(name, 0)
        if got > want:                                        # (a) 上界
            findings.append({
                "rule": "R-P35", "tc_id": "(全體)",
                "detail": f"Test Set `{name}` 已產出 {got} 個 leaf，超過定版之 {want}",
            })
    leaves = {leaf_of(tc) for tc in tcs if leaf_of(tc) in leaf_testset}
    if len(leaves) == LOCKED_TOTAL:                           # (c) 齊備時驗相等
        for name, want in LOCKED_DISTRIBUTION.items():
            got = covered.get(name, 0)
            if got != want:
                findings.append({
                    "rule": "R-P35", "tc_id": "(全體)",
                    "detail": f"114 leaf 齊備，Test Set `{name}` 涵蓋 {got} 個，定版為 {want}",
                })
    return findings


def check_s107_spec_reference(tcs: list[dict]) -> list[dict]:
    """G45 / §10.7（R-P66）：specification_reference 非空且每項符形態。

    形態 `{spec_filename}_{section_id}` —— 檔名任意非空白，section_id 為
    點分隔之數字串（`1.7.2`、`1.7.1.1.1`）。多項以 `; ` 分隔。
    """
    findings = []
    for tc in tcs:
        tc_id = tc.get("tc_id") or tc.get("req_id") or "(無 id)"
        raw = str(tc.get("specification_reference", "")).strip()
        if not raw:
            findings.append({
                "rule": "§10.7", "tc_id": tc_id,
                "detail": "specification_reference 為空 —— §10.7 要求必填",
            })
            continue
        for item in [x.strip() for x in raw.split(";") if x.strip()]:
            if not SPEC_REF_ITEM_RE.match(item):
                findings.append({
                    "rule": "§10.7", "tc_id": tc_id,
                    "detail": f"specification_reference 之項目 `{item}` "
                              f"不符 {{spec_filename}}_{{section_id}} 形態",
                })
    return findings


def check_feature_yaml_consistency(text: str | None = None) -> list[dict]:
    """G46 / R-P69(c)：feature.yaml 與裁決條文一致 —— 偵測未來漂移。

    lint 之權威值仍取自裁決條文（R-P69(b)）；本閘只驗二者未分歧。
    `text` 供合成 fixture 注入（09 §I：不得以 repo 現況作為 fixture 對照）。
    """
    if text is None:
        path = ROOT / "features/power/feature.yaml"
        if not path.exists():
            return [{"rule": "R-P69", "tc_id": "(feature.yaml)", "detail": "檔案不存在"}]
        text = path.read_text(encoding="utf-8")
    findings = []
    for key, (want, source) in YAML_EXPECTED.items():
        m = re.search(rf'^\s*{re.escape(key)}\s*:\s*(?:"([^"]*)"|([^"#\n]*?))\s*(?:#.*)?$',
                      text, re.M)
        got = (m.group(1) if m and m.group(1) is not None else
               (m.group(2).strip() if m else None))
        if got != want:
            findings.append({
                "rule": "R-P69", "tc_id": "(feature.yaml)",
                "detail": f"`{key}` 為 {got!r}，裁決條文（{source}）定為 {want!r}",
            })
    if "<" in text and ">" in re.sub(r"#.*", "", text):
        for m in re.finditer(r"^\s*(\w+)\s*:\s*\"?([^\"\n]*<[^>]+>[^\"\n]*)\"?",
                             text, re.M):
            findings.append({
                "rule": "R-P69", "tc_id": "(feature.yaml)",
                "detail": f"`{m.group(1)}` 仍為 scaffold placeholder：{m.group(2).strip()}",
            })
    return findings


def check_s11_formatting(tcs: list[dict]) -> list[dict]:
    """G50 / §11（R-P75）：trailing period、UI 標籤引號、方括號。"""
    findings = []
    for tc in tcs:
        tc_id = tc.get("tc_id") or tc.get("req_id") or "(無 id)"
        for field in LONG_FIELDS:
            for line in str(tc.get(field, "")).split("\n"):
                stripped = line.rstrip()
                if stripped.endswith((".", "。")):
                    findings.append({
                        "rule": "§11", "tc_id": tc_id,
                        "detail": f"`{field}` 之行以句點結尾：「{stripped[-46:]}」",
                    })
                residue = stripped
                if PROFILE_EXEMPTS_BRACKETS:      # profile §3.1 / §3.2
                    residue = SIGNAL_BRACKET_RE.sub("", SOURCE_CLASS_RE.sub("", residue))
                if BAD_BRACKET_RE.search(residue):
                    findings.append({
                        "rule": "§11", "tc_id": tc_id,
                        "detail": f"`{field}` 含方括號（非訊號值）：「{stripped[:46]}」",
                    })
                if TABLE_RE.search(stripped):
                    findings.append({
                        "rule": "§11", "tc_id": tc_id,
                        "detail": f"`{field}` 含表格（Markdown `|` 或 HTML）："
                                  f"「{stripped[:46]}」—— §11 禁 HTML / Markdown 表格",
                    })
                if BAD_QUOTE_RE.search(stripped):
                    findings.append({
                        "rule": "§11", "tc_id": tc_id,
                        "detail": f"`{field}` 之 UI 標籤未用雙引號：「{stripped[:46]}」",
                    })
    return findings


def check_s44_precondition(tcs: list[dict]) -> list[dict]:
    """G51 / §4.4（R-P75）：Pre-Condition 為狀態／環境，不得含動作。"""
    findings = []
    for tc in tcs:
        tc_id = tc.get("tc_id") or tc.get("req_id") or "(無 id)"
        for line in str(tc.get("pre_conditions", "")).split("\n"):
            m = PRECOND_ACTION_RE.match(line)
            if m:
                findings.append({
                    "rule": "§4.4", "tc_id": tc_id,
                    "detail": f"pre_conditions 以動作動詞 `{m.group(1)}` 起始："
                              f"「{line.strip()[:52]}」—— §4.4 僅容許狀態／環境",
                })
    return findings


def run_all_gates(tcs: list[dict], blacklist: dict, fingerprints: dict,
                  leaf_testset: dict, leaf_priority: dict) -> tuple[list, list]:
    """回傳 (阻斷類, 待人工裁決類)。

    R-P76：R-P42(b) 之觸發**不使 exit=1**，另列一節輸出待人工裁決；
    其餘閘門皆為阻斷類。R-P67 明令其不得自動判 FAIL。
    """
    all_findings = (
        check_rp42_unreferenced_anchor(tcs, blacklist, fingerprints)
        + check_rp1_blocked_leaf(tcs)
        + check_rp2_tc_id(tcs)
        + check_rp8_priority(tcs, leaf_priority)
        + check_rp35_test_set(tcs, leaf_testset)
        + check_s107_spec_reference(tcs)
        + check_s11_formatting(tcs)
        + check_s44_precondition(tcs)
        + check_s6_proc_er_parity(tcs)
        + check_s6_er_restatement(tcs)
        + check_s841_time_equality(tcs)
        + check_s52b_final_step_intent(tcs)
        + check_s44_env_stability(tcs)
        + check_s45_data_ownership(tcs)
        + check_profile_clauses(tcs)
        + check_feature_yaml_consistency()
        + check_workbook_columns()
    )
    ADJUDICATE_RULES = {"R-P42(b)", "R-P96(a)", "R-P96(b)", "R-P142"}
    blocking = [f for f in all_findings if f["rule"] not in ADJUDICATE_RULES]
    adjudicate = [f for f in all_findings if f["rule"] in ADJUDICATE_RULES]
    return blocking, adjudicate


def check_s6_proc_er_parity(tcs: list[dict]) -> list[dict]:
    """G63 / §6（R-P87）：procedure 之編號步數 = expected_result 之編號行數。"""
    findings = []
    num = re.compile(r"^\s*\d+\.")
    for tc in tcs:
        tc_id = tc.get("tc_id") or tc.get("req_id") or "(無 id)"
        steps = sum(1 for l in str(tc.get("test_procedure", "")).split("\n") if num.match(l))
        ers = sum(1 for l in str(tc.get("expected_result", "")).split("\n") if num.match(l))
        if steps != ers:
            findings.append({
                "rule": "§6", "tc_id": tc_id,
                "detail": f"procedure {steps} 步 ≠ expected_result {ers} 行 —— §6 要求 1:1",
            })
    return findings


def _stem(w: str) -> str:
    """輕量字尾正規化 —— 使 `start` 與 `starts`、`record` 與 `recorded` 對齊。
    僅去 -ing / -ed / -es / -s，長度 > 4 者方去，不做詞形還原。"""
    for suf in ("ing", "ed", "es", "s"):
        if w.endswith(suf) and len(w) - len(suf) >= 3:
            return w[: -len(suf)]
    return w

def _er_words(line: str) -> set[str]:
    """取實詞（去編號、去 source-class 標記、去停用詞、輕量字尾正規化）。"""
    body = re.sub(r"^\s*\d+[.)]\s*(?:\[[a-z-]+\]\s*)?", "", line).strip()
    ws = {w.lower() for w in ER_WORD_RE.findall(body)} - ER_STOPWORDS
    return {_stem(w) for w in ws}


def check_s6_er_restatement(tcs: list[dict]) -> list[dict]:
    """G73 / §6（R-P96）：ER 行不得為 procedure 動作之複述。

    僅對 1:1 對齊者逐行比對（不對齊者由 G63 攔下）。
    (a) 阻斷類：ER 行述及執行者自身之動作，且 overlap ≥ P90。
    (b) 待人工裁決類（`rule` 為 `R-P96(b)`）：overlap = 1.00 且無新增標的。
    """
    findings = []
    num = re.compile(r"^\s*\d+\.")
    for tc in tcs:
        tc_id = tc.get("tc_id") or tc.get("req_id") or "(無 id)"
        steps = [l for l in str(tc.get("test_procedure", "")).split("\n") if num.match(l)]
        ers = [l for l in str(tc.get("expected_result", "")).split("\n") if num.match(l)]
        if len(steps) != len(ers):
            continue
        for i, (p_line, e_line) in enumerate(zip(steps, ers), 1):
            # R-P142（20 包）：**末步 ER 行不再以 overlap 判定。**
            # R-P133（19 包）曾令先剝除 R-P101 所強制之驗證意圖子句再計 overlap；
            # 19 §六實測：剝除使 overlap 由 1.00 降至 0.80，**而判定 12 → 12，
            # 無一條脫離** —— 即 overlap 對末步為**錯誤工具**，
            # R-P133 後段已由 R-P142 撤回（剝除邏輯隨之移除）。
            #
            # 根因為 R-P96 之判準有誤：其載「若某 ER 行在『該步驟被執行』之外
            # 不含任何額外資訊，即為複述」，**而回讀含有額外資訊 —— 即該值本身**。
            # 新判準為**可證偽性**：該 ER 行能否在其 procedure 步驟成功執行之
            # 情形下仍然失敗。`The elapsed time is recorded` 記了即成立 → 複述；
            # `TLM_Status.Info reads "Standby"` 讀到他值即 fail → 非複述。
            #
            # **可證偽性無法機械化**，故末步一律入 R-P76 之待人工裁決類（永久）；
            # 依 R-P142(d) 該裁決不得省略 ——「沉默不算裁決」（R-P118(e)）同樣適用。
            # 非末步之 ER 行維持 overlap 判定不變（R-P142(c)）。
            if i == len(steps):
                findings.append({
                    "rule": "R-P142", "tc_id": tc_id,
                    "detail": f"末步 ER 第 {i} 行須以**可證偽性**人工裁決"
                              f"（該行能否在其步驟成功執行時仍然失敗）："
                              f"「{e_line.strip()[:56]}」",
                })
                continue
            pw, ew = _er_words(p_line), _er_words(e_line)
            if not ew:
                continue
            overlap = len(ew & pw) / len(ew)
            actor = ER_ACTOR_RE.search(e_line)
            if actor and overlap >= ER_OVERLAP_P50:
                findings.append({
                    "rule": "R-P96(a)", "tc_id": tc_id,
                    "detail": f"ER 第 {i} 行為 procedure 動作之複述"
                              f"（述語 `{actor.group(0)}`，overlap {overlap:.2f}）："
                              f"「{e_line.strip()[:56]}」—— R-P96",
                })
            elif overlap >= 1.0:
                findings.append({
                    "rule": "R-P96(b)", "tc_id": tc_id,
                    "detail": f"ER 第 {i} 行之實詞全數見於 procedure 步驟，"
                              f"無新增可觀察標的：「{e_line.strip()[:56]}」"
                              f"—— 已交付語料同型 101 組，須人工判別",
                })
    return findings


def check_s52b_final_step_intent(tcs: list[dict]) -> list[dict]:
    """G77 / §5.2B + §5.5（R-P101）：末步須含驗證意圖措詞，且 ≤ 18 字。"""
    findings = []
    num = re.compile(r"^\s*\d+\.")
    for tc in tcs:
        tc_id = tc.get("tc_id") or tc.get("req_id") or "(無 id)"
        steps = [l for l in str(tc.get("test_procedure", "")).split("\n") if num.match(l)]
        if not steps:
            continue
        body = re.sub(r"^\s*\d+\.\s*(?:\[[a-z-]+\]\s*)?", "", steps[-1]).strip()
        if not FINAL_STEP_INTENT_RE.search(body):
            findings.append({
                "rule": "§5.2B", "tc_id": tc_id,
                "detail": f"Final Step 無驗證意圖措詞（check that / to verify / …）："
                          f"「{body[:60]}」—— R-P101",
            })
        elif len(body.split()) > 18:
            findings.append({
                "rule": "§5.2B", "tc_id": tc_id,
                "detail": f"Final Step {len(body.split())} 字，逾 §5.2B 之 18 字上限："
                          f"「{body[:60]}」",
            })
    return findings


def check_misread_terms(batch: dict) -> list[dict]:
    """G81 / R-P107：誤讀關鍵詞掃全部欄位（個案型）。"""
    findings = []
    for tc in batch.get("tcs", []):
        tc_id = tc.get("tc_id") or "(無 id)"
        for field, value in tc.items():
            if field in TC_SCAN_SKIP:
                continue
            text = json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else str(value)
            text = QUOTED_SPAN_RE.sub(" ", text)
            for term, why in MISREAD_TERMS:
                if term.lower() in text.lower():
                    findings.append({
                        "rule": "R-P107", "tc_id": tc_id,
                        "detail": f"欄位 `{field}` 殘留誤讀關鍵詞「{term}」—— {why}",
                    })
    for leaf in batch.get("leaves", []):
        for field in ("reasoning", "reasoning_note"):
            text = QUOTED_SPAN_RE.sub(" ", str(leaf.get(field, "")))
            for term, why in MISREAD_TERMS:
                if term.lower() in text.lower():
                    findings.append({
                        "rule": "R-P107", "tc_id": leaf.get("parent", "?"),
                        "detail": f"leaf `{field}` 殘留誤讀關鍵詞「{term}」—— {why}",
                    })
    return findings


def check_er_clause_coverage(batch: dict) -> list[dict]:
    """G82 / R-P109：ER 之專有標的須出現於該 leaf 之 `source_clause`。"""
    findings = []
    clause = {l.get("parent"): str(l.get("source_clause", "")) for l in batch.get("leaves", [])}
    for tc in batch.get("tcs", []):
        src = clause.get(tc.get("req_id"), "")
        tc_id = tc.get("tc_id") or "(無 id)"
        er = "\n".join(re.sub(r"^\s*\d+\.\s*", "", ln)
                       for ln in str(tc.get("expected_result", "")).split("\n"))
        for tok in sorted(set(ER_PROPER_RE.findall(er))):
            if tok in ER_PROPER_SKIP:
                continue
            if tok not in src:
                findings.append({
                    "rule": "R-P109", "tc_id": tc_id,
                    "detail": f"ER 之標的 `{tok}` 未見於該 leaf 之 `source_clause` —— "
                              f"截斷不得遮蔽 ER 所斷言之內容",
                })
    return findings


def check_source_clause(batch: dict) -> list[dict]:
    """G79 / R-P104：`leaves` 陣列每筆須附非空之 `source_clause`。"""
    findings = []
    for leaf in batch.get("leaves", []):
        if not str(leaf.get("source_clause", "")).strip():
            findings.append({
                "rule": "R-P104", "tc_id": leaf.get("parent", "(無 parent)"),
                "detail": "leaf 缺 `source_clause`（規格原文子句）—— 無之則技術覆核"
                          "僅能檢視 TC 之自我一致性",
            })
    return findings


def check_s841_time_equality(tcs: list[dict]) -> list[dict]:
    """G74 / §8.4.1（R-P97）：時間量測之 ER 不得寫數值相等。"""
    findings = []
    for tc in tcs:
        tc_id = tc.get("tc_id") or tc.get("req_id") or "(無 id)"
        for line in str(tc.get("expected_result", "")).split("\n"):
            if not line.strip():
                continue
            eq = TIME_EQUALITY_RE.search(line)
            if eq and TIME_TOKEN_RE.search(line):
                findings.append({
                    "rule": "§8.4.1", "tc_id": tc_id,
                    "detail": f"時間量測之 ER 寫數值相等（`{eq.group(0)}`）："
                              f"「{line.strip()[:56]}」—— 嚴格執行必然 fail，"
                              f"應改為行為描述（R-P97）",
                })
    return findings


def check_s44_env_stability(tcs: list[dict]) -> list[dict]:
    """G64 / §4.4 + §8.5（R-P88）：pre_conditions 不得含系統預設與環境穩定性前提。"""
    findings = []
    for tc in tcs:
        tc_id = tc.get("tc_id") or tc.get("req_id") or "(無 id)"
        for line in str(tc.get("pre_conditions", "")).split("\n"):
            m = ENV_STABILITY_RE.search(line)
            if m:
                findings.append({
                    "rule": "§8.5", "tc_id": tc_id,
                    "detail": f"pre_conditions 含系統預設／環境穩定性前提 "
                              f"`{m.group(0)}`：「{line.strip()[:56]}」",
                })
    return findings


def check_s45_data_ownership(tcs: list[dict]) -> list[dict]:
    """G65 / §4.5（R-P89）：input_test_data 不得與 procedure / pre_conditions 重複。

    判準：`input_test_data` 之每一行，若其**全部**長度 ≥ 3 之 token
    皆出現於 procedure 或 pre_conditions，即為跨欄重複。
    """
    findings = []
    for tc in tcs:
        tc_id = tc.get("tc_id") or tc.get("req_id") or "(無 id)"
        data = str(tc.get("input_test_data", "")).strip()
        if not data or data.upper() == "NA":
            continue
        other = (str(tc.get("test_procedure", "")) + "\n"
                 + str(tc.get("pre_conditions", ""))).lower()
        for line in data.split("\n"):
            toks = [t.lower() for t in DATA_TOKEN_RE.findall(line)]
            if toks and all(t in other for t in toks):
                findings.append({
                    "rule": "§4.5", "tc_id": tc_id,
                    "detail": f"input_test_data 之「{line.strip()[:50]}」"
                              f"全部 token 亦見於 procedure / pre_conditions —— §4.5 禁跨欄重複",
                })
    return findings


def check_profile_clauses(tcs: list[dict]) -> list[dict]:
    """G72 / R-P91：profile 之可機械檢查條款。

    §2   test_group 值
    §3.3 design_method 須為 `下拉選單!A1:A9` 九詞條之一
    §3.4 spec_reference 之檔名須為本 feature 兩份規格之一
    §3.7 functional_safety 一律 `NA`
    """
    findings = []
    for tc in tcs:
        tc_id = tc.get("tc_id") or tc.get("req_id") or "(無 id)"
        if str(tc.get("test_group", "")).strip() != PROFILE_TEST_GROUP:
            findings.append({"rule": "profile §2", "tc_id": tc_id,
                             "detail": f"test_group `{tc.get('test_group')}` "
                                       f"≠ `{PROFILE_TEST_GROUP}`"})
        dm = str(tc.get("design_method", "")).strip()
        if dm and dm not in DESIGN_METHODS:
            findings.append({"rule": "profile §3.3", "tc_id": tc_id,
                             "detail": f"design_method `{dm[:40]}` 不在 `下拉選單!A1:A9` 九詞條內"})
        fs = str(tc.get("functional_safety", "")).strip()
        if fs and fs != "NA":
            findings.append({"rule": "profile §3.7", "tc_id": tc_id,
                             "detail": f"functional_safety `{fs}` ≠ `NA`"})
        for item in [x.strip() for x in str(tc.get("specification_reference", "")).split(";") if x.strip()]:
            stem = item.rsplit("_", 1)[0]
            if stem not in SPEC_STEMS:
                findings.append({"rule": "profile §3.4", "tc_id": tc_id,
                                 "detail": f"spec_reference 之檔名 `{stem[:44]}…` "
                                           f"非本 feature 兩份規格之一"})
    return findings


def check_workbook_columns() -> list[dict]:
    """G71 / R-P91：`feature.yaml` 之 `workbook.columns` 須與 FW036 r9 實測標頭相符。

    A-PW40 係靠人工盤點查出，本閘使其成為機械檢查 —— 寫回不可逆。
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    yaml_text = (ROOT / "features/power/feature.yaml").read_text(encoding="utf-8")
    mapping = {}
    for m in re.finditer(r'^\s{4}(\w+):\s*"([A-Z]{1,2})"', yaml_text, re.M):
        mapping[m.group(1)] = m.group(2)
    wb = openpyxl.load_workbook(find("FSM-036"), data_only=True, read_only=True)
    ws = wb["Test Case Specification&Result"]
    headers = {get_column_letter(i + 1): (str(c.value).replace("\n", " ").strip()
                                          if c.value else "")
               for i, c in enumerate(ws[9])}
    wb.close()
    findings = []
    for key, col in mapping.items():
        want = COLUMN_HEADER_KEYWORD.get(key)
        if not want:
            continue
        got = headers.get(col, "")
        if want.lower() not in got.lower():
            findings.append({"rule": "R-P91/G71", "tc_id": "(feature.yaml)",
                             "detail": f"`workbook.columns.{key}` = `{col}`，"
                                       f"該欄之 r9 標頭為 `{got[:40]}`，未含 `{want}`"})
    return findings


def check_b_column_numbering(rows_written: int, b_filled: int) -> list[dict]:
    """G66 / R-P90：寫回後 B 欄非空列數須等於 TC 列數。

    Power 之範本 B 欄無自動編號公式（11 包 B1 實測），故須明寫。
    本包僅實作，寫回時方能實測。
    """
    if b_filled != rows_written:
        return [{"rule": "R-P90", "tc_id": "(全體)",
                 "detail": f"寫回後 B 欄非空 {b_filled} 列 ≠ TC {rows_written} 列 —— "
                           f"Power 之範本無編號公式，B 欄須明寫"}]
    return []


def load_tcs(directory: Path) -> list[dict]:
    tcs = []
    for path in sorted(directory.glob("*.json")):
        data = json.loads(path.read_text(encoding="utf-8"))
        for tc in data.get("tcs", []):
            tc = dict(tc)
            tc.setdefault("_file", path.name)
            tcs.append(tc)
    return tcs


def load_batches(directory: Path) -> list[dict]:
    """批次層（`leaves` ＋ `tcs`）—— G79 / G81 / G82 以整批為單位檢查。"""
    return [json.loads(p.read_text(encoding="utf-8"))
            for p in sorted(directory.glob("*.json"))]


def run_batch_gates(batches: list[dict]) -> list[dict]:
    """批次層閘門：G79（R-P104）、G81（R-P107）、G82（R-P109）。"""
    findings = []
    for b in batches:
        findings += check_source_clause(b)
        findings += check_misread_terms(b)
        findings += check_er_clause_coverage(b)
    return findings


def make_tc(n: int, leaf: str, test_set: str, **over) -> dict:
    """合成一條合法 TC（不讀 generated/，08 §I）。"""
    tc = {
        "req_id": f"{leaf}-01",
        "tc_id": f"NR1L-PowerManagement-{n:03d}",
        "test_group": "Power Management",
        "test_set": test_set,
        "test_item": "Verify the power state transition is applied.",
        "pre_conditions": "1. [spec-derived] Vehicle in Ignition On",
        "input_test_data": "NA",
        "test_procedure": "1. Set the power mode signal to the target state\n"
                          "2. Read the reported state and check that it matches",
        "expected_result": "1. The power mode signal carries the target state\n"
                           "2. The head unit applies the corresponding policy",
        "specification_reference": "CFTS009_1.6.2.1.1_4941357",
        "priority": "P2",
    }
    tc.update(over)
    return tc


def synthetic_full_set(leaf_testset: dict[str, str]) -> list[dict]:
    """依 leaf_testset.tsv 合成 114 條 —— 用於 G40 之分布檢查。"""
    return [
        make_tc(i + 1, leaf, ts)
        for i, (leaf, ts) in enumerate(sorted(
            leaf_testset.items(), key=lambda kv: int(re.match(r"SWE-PM-(\d+)", kv[0]).group(1))
        ))
    ]


def self_test(blacklist: dict, fingerprints: dict) -> int:
    """以合成 fixture 驗全部閘門 —— 不以 repo 現況為對照（07 / 08 §I）。"""
    leaf_testset = load_leaf_testset()
    leaf_priority = load_037_priority()
    blacklisted = next(a for a in fingerprints)
    fingerprint = fingerprints[blacklisted][0]
    cited_anchor = next(a for a in anchor_bodies() if a not in blacklist)
    full = synthetic_full_set(leaf_testset)

    cases: list[tuple[str, str, list[dict], list[str]]] = [
        # ── G33 / R-P42（07 包，沿用）──
        ("G33", "應 PASS —— 只引用被引用之錨點",
         [make_tc(1, "SWE-PM-001", "Power State",
                  specification_reference=f"CFTS009_1.6.2.1.1_{cited_anchor}")], []),
        ("G33", "應因 (a) FAIL —— specification_reference 命中黑名單",
         [make_tc(1, "SWE-PM-001", "Power State",
                  specification_reference=f"CFTS009_1.6.2.1.1_{blacklisted}")], ["R-P42(a)"]),
        ("G33", "應因 (b) FAIL —— 內容逐字引用黑名單特徵字串",
         # 特徵字串後接續文字 —— 否則其自帶之句尾句點會同時觸發 §11，
         # 使本 fixture 之期望值失焦（二閘皆觸發為正確行為，但此處只驗 (b)）
         [make_tc(1, "SWE-PM-001", "Power State",
                  test_procedure=f"1. {fingerprint} on the bench\n"
                                 f"2. Read the reported state",
                  expected_result="1. The step completes\n2. The state is read")],
         ["R-P42(b)"]),

        # ── G37 / R-P1 ──
        ("G37", "應 PASS —— 未觸及 SWE-PM-089",
         [make_tc(1, "SWE-PM-001", "Power State")], []),
        ("G37", "應 FAIL —— 存在 SWE-PM-089 之 TC",
         [make_tc(1, "SWE-PM-089", "Power State")], ["R-P1"]),
        # 本例同時違反 §10.7（`SWE-PM-089_source_pending` 之 section_id 非數字串），
        # 二閘皆觸發為正確行為，故期望值列出兩者。
        ("G37", "應 FAIL —— specification_reference 引用 SWE-PM-089",
         [make_tc(1, "SWE-PM-001", "Power State",
                  specification_reference="SWE-PM-089_source_pending")],
         ["R-P1", "§10.7"]),

        # ── G38 / R-P2 ──
        ("G38", "應 PASS —— tc_id 格式正確且連號",
         [make_tc(1, "SWE-PM-001", "Power State"),
          make_tc(2, "SWE-PM-002", "Power State")], []),
        ("G38", "應 FAIL —— tc_id 格式錯誤",
         [make_tc(1, "SWE-PM-001", "Power State", tc_id="NR1L-Power-001")], ["R-P2"]),
        ("G38", "應 FAIL —— tc_id 重複",
         [make_tc(1, "SWE-PM-001", "Power State"),
          make_tc(1, "SWE-PM-002", "Power State")], ["R-P2"]),
        ("G38", "應 FAIL —— tc_id 跳號",
         [make_tc(1, "SWE-PM-001", "Power State"),
          make_tc(3, "SWE-PM-002", "Power State")], ["R-P2"]),

        # ── G39 / R-P8 ──
        ("G39", "應 PASS —— priority 為 P0–P3 且非一對一映射",
         [make_tc(1, "SWE-PM-001", "Power State", priority="P1"),
          make_tc(2, "SWE-PM-057", "Timeout Settings", priority="P1"),
          make_tc(3, "SWE-PM-002", "Power State", priority="P2")], []),
        ("G39", "應 FAIL —— priority 為 037 之 `High`",
         [make_tc(1, "SWE-PM-001", "Power State", priority="High")], ["R-P8"]),
        ("G39", "應 FAIL —— 與 037 Priority 呈一對一映射",
         None, ["R-P8"]),   # 於下方依實際 037 值組出

        # ── G45 / §10.7（R-P66）──
        ("G45", "應 PASS —— spec_reference 非空且形態正確",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  specification_reference="CFTS010_1.7.1.1.1; CFTS010_1.7.2")], []),
        ("G45", "應 FAIL —— spec_reference 為空",
         [make_tc(1, "SWE-PM-071", "Power Down", specification_reference="")], ["§10.7"]),
        ("G45", "應 PASS —— 檔名含空格與括號（§10.7 之範例形態）",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  specification_reference=(
                      "Menu Bar and App Drawer HMI Logic and Flow R1 SR24 3A "
                      "(September 11 2023)_2.5"))], []),
        ("G45", "應 FAIL —— spec_reference 形態不符",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  specification_reference="see the power down chapter")], ["§10.7"]),

        # ── G50 / §11（R-P75）──
        ("G50", "應 PASS —— 無句點、UI 標籤用雙引號、方括號僅用於訊號值",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  pre_conditions="1. A LIN simulation tool is available on the bench",
                  test_procedure='1. Press "Power" on the front panel\n'
                                 '2. Read STATUS_LIN.Batt_ST_Crit and check that it is [1h]',
                  expected_result="1. The front panel accepts the press\n"
                                  "2. The unit enters the critical handling state")], []),
        ("G50", "應 FAIL —— test_procedure 之行以句點結尾",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  test_procedure="1. Set the power mode signal\n"
                                 "2. Read the reported state and check that it changed.",
                  expected_result="1. The signal is set\n2. The reported state changed")],
         ["§11"]),
        ("G50", "應 FAIL —— UI 標籤用方括號",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  test_procedure='1. Press [Power] on the front panel\n'
                                 '2. Read the reported state',
                  expected_result="1. The panel accepts the press\n2. The state is read")],
         ["§11"]),

        # ── G51 / §4.4（R-P75）──
        ("G51", "應 PASS —— pre_conditions 僅述狀態與環境",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  pre_conditions="1. A LIN simulation tool is available on the bench\n"
                                 "2. An event injection tool is available")], []),
        ("G51", "應 FAIL —— pre_conditions 含動作",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  pre_conditions="1. Connect the LIN simulation tool to the bench")],
         ["§4.4"]),

        # ── G63 / §6（R-P87）──
        ("G63", "應 PASS —— procedure 3 步、ER 3 行",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  test_procedure="1. Start the boot sequence\n2. Record the elapsed time\n"
                                 "3. Compare the time and check that the screen changed",
                  expected_result="1. The boot sequence starts\n2. The time is recorded\n"
                                  "3. The screen change matches the configured time")], []),
        ("G63", "應 FAIL —— procedure 3 步而 ER 2 行",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  test_procedure="1. Start the boot sequence\n2. Record the elapsed time\n"
                                 "3. Compare the time and check that the screen changed",
                  expected_result="1. The boot sequence starts\n2. The screen changes")],
         ["§6"]),

        # ── G64 / §4.4 + §8.5（R-P88）──
        ("G64", "應 PASS —— pre_conditions 為 hardware / peripheral",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  pre_conditions="1. A suspend-resume boot sequence is available on the bench",
                  test_procedure="1. Start the boot sequence\n2. Read the display",
                  expected_result="1. The boot sequence starts\n2. The display changes")], []),
        ("G64", "應 FAIL —— pre_conditions 含「powered from a stable supply」",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  pre_conditions="1. The TLM is powered from a stable supply",
                  test_procedure="1. Start the boot sequence\n2. Read the display",
                  expected_result="1. The boot sequence starts\n2. The display changes")],
         ["§8.5"]),

        # ── G65 / §4.5（R-P89）──
        ("G65", "應 PASS —— input_test_data 為獨立資料集，未見於他欄",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  input_test_data="Event burst: 20 events at 100 ms intervals",
                  pre_conditions="1. An event injection tool is connected to the bench",
                  test_procedure="1. Start the boot sequence\n"
                                 "2. Inject the event burst listed in Input Test Data",
                  expected_result="1. The boot sequence starts\n2. The events reach the TLM")], []),
        ("G65", "應 FAIL —— input_test_data 之值逐字重複於 procedure",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  input_test_data="Boot target status: Standby",
                  pre_conditions="1. A boot sequence is available on the bench",
                  test_procedure="1. Set the boot target status to Standby\n2. Read the display",
                  expected_result="1. The status is Standby\n2. The display changes")],
         ["§4.5"]),

        # ── G50 表格檢查（R-P93）──
        ("G50", "應 FAIL —— expected_result 含 Markdown 表格",
         [make_tc(1, "SWE-PM-071", "Power Down",
                  test_procedure="1. Start the boot sequence\n2. Read the display",
                  expected_result="1. The boot sequence starts\n"
                                  "2. | state | value | result |")], ["§11"]),

        # ── G40 / R-P35（R-P68 之子集檢查）──
        ("G40", "應 PASS —— 完整 114 條，分布符定版",
         full, []),
        ("G40", "應 FAIL —— Test Set 非五個定版值之一",
         [make_tc(1, "SWE-PM-001", "Power Management")], ["R-P35"]),
        ("G40", "應 FAIL —— 某 leaf 之歸屬與 leaf_testset.tsv 不符",
         [make_tc(1, "SWE-PM-057", "Power State")], ["R-P35"]),
        ("G40", "應 FAIL —— 完整 114 條但分布不符定版",
         None, ["R-P35"]),  # 於下方由 full 改一條組出
        ("G40", "應 PASS —— 部分批次（3 條 Power Down），未達上界",
         [make_tc(1, "SWE-PM-071", "Power Down"),
          make_tc(2, "SWE-PM-072", "Power Down"),
          make_tc(3, "SWE-PM-073", "Power Down")], []),
        ("G40", "應 FAIL —— 整批一致地歸錯，超過該 Test Set 之上界（R-P68(a)）",
         None, ["R-P35"]),  # 於下方組出：4 個 leaf 全記 Power Down
    ]

    # G39 一對一映射之違規 fixture：037 之每個 Priority 值對應唯一之 TC priority
    priority_map = {"High": "P1", "Medium": "P3"}
    mapped = [
        make_tc(i + 1, leaf, leaf_testset[leaf],
                priority=priority_map.get(leaf_priority.get(leaf, ""), "P2"))
        for i, leaf in enumerate(sorted(
            leaf_testset, key=lambda x: int(re.match(r"SWE-PM-(\d+)", x).group(1)))[:20])
    ]
    # 分布違規 fixture：把一條 Power State 改記為 Timeout Settings
    skewed = [dict(tc) for tc in full]
    for tc in skewed:
        if tc["test_set"] == "Power State":
            tc["test_set"] = "Timeout Settings"
            break

    # R-P68(a) 之上界違規：4 個 leaf 全記 Power Down（定版僅 3）——
    # 原讀法（僅在 114 齊備時比對）攔不下此情形
    over_cap = [make_tc(i + 1, leaf, "Power Down")
                for i, leaf in enumerate(
                    ["SWE-PM-071", "SWE-PM-072", "SWE-PM-073", "SWE-PM-093"])]

    placeholders = {"G39": mapped}
    seen_g40 = 0
    filled = []
    for gate, label, tcs, expected in cases:
        if tcs is None:
            if gate == "G39":
                tcs = mapped
            else:
                seen_g40 += 1
                tcs = skewed if seen_g40 == 1 else over_cap
        filled.append((gate, label, tcs, expected))

    print("lint fixture 測試（全為合成 TC，未讀 generated/）")
    print(f"  黑名單錨點樣本：{blacklisted}")
    print(f"  特徵字串樣本  ：「{fingerprint[:64]}…」")
    print(f"  被引用錨點樣本：{cited_anchor}\n")

    failures = 0
    for gate, label, tcs, expected in filled:
        findings = (
            check_rp42_unreferenced_anchor(tcs, blacklist, fingerprints)
            + check_rp1_blocked_leaf(tcs)
            + check_rp2_tc_id(tcs)
            + check_rp8_priority(tcs, leaf_priority)
            + check_rp35_test_set(tcs, leaf_testset)
            + check_s107_spec_reference(tcs)
            + check_s11_formatting(tcs)
            + check_s44_precondition(tcs)
            + check_s6_proc_er_parity(tcs)
            + check_s44_env_stability(tcs)
            + check_s45_data_ownership(tcs)
        )
        got = sorted({f["rule"] for f in findings})
        ok = got == sorted(expected)
        failures += not ok
        print(f"  [{'PASS' if ok else '**FAIL**'}] {gate} {label}")
        print(f"          期望 {expected or '（無）'}；實際 {got or '（無）'}")
        for f in findings[:2]:
            print(f"          → {f['rule']} {f['tc_id']}: {f['detail'][:92]}")
    # G71 / G72 / R-P91 —— profile 條款
    bad_profile = [make_tc(1, "SWE-PM-071", "Power Down",
                           test_group="Power", design_method="Ad-hoc",
                           functional_safety="Yes",
                           specification_reference="Some_Other_Spec_1.2")]
    good_profile = [make_tc(1, "SWE-PM-071", "Power Down",
                            design_method="狀態轉換 (State Transition Testing)",
                            specification_reference=(
                                "R1LR_Atl-H_25PI3.5_Activation and Configuration_"
                                "CFTS_010_Power Down _SR26_20250909-1658_1.7.2"))]
    for label, tcs_, want in [("應 PASS —— 全數符 profile 條款", good_profile, 0),
                              ("應 FAIL —— test_group / design_method / "
                               "functional_safety / spec 檔名皆違", bad_profile, 4)]:
        got = check_profile_clauses(tcs_)
        ok = (len(got) == 0) if want == 0 else (len(got) >= want)
        failures += not ok
        print(f"\n  [{'PASS' if ok else '**FAIL**'}] G72 {label}")
        print(f"          期望 {'0' if want == 0 else f'≥{want}'} 項；實際 {len(got)} 項")
        for f in got[:4]:
            print(f"          → {f['rule']} {f['detail'][:80]}")

    col_findings = check_workbook_columns()
    ok71 = not col_findings
    failures += not ok71
    print(f"\n  [{'PASS' if ok71 else '**FAIL**'}] G71 workbook.columns 對 r9 實測標頭")
    print(f"          期望 0 項；實際 {len(col_findings)} 項")
    for f in col_findings[:3]:
        print(f"          → {f['detail'][:88]}")

    # G66 / R-P90 —— B 欄序號，寫回時方能實測；本包僅驗閘門邏輯
    ok66 = (not check_b_column_numbering(10, 10)) and bool(check_b_column_numbering(10, 0))
    failures += not ok66
    print(f"\n  [{'PASS' if ok66 else '**FAIL**'}] G66 B 欄非空列數 = TC 列數"
          f"（合成；寫回時方能實測）")
    print(f"          相等時 findings 0；B 欄全空時 findings "
          f"{len(check_b_column_numbering(10, 0))}")

    # G73 / R-P96 ＋ R-P142 —— ER 複述偵測。
    # 依 G46 / G71–G74 之既有慣例，**不併入 per-fixture 聚合**。
    # **R-P142（20 包）改判準後，末步 ER 行一律入待裁類**，
    # 故「應 PASS」之案改為「非末步 0 項 R-P96、末步 1 項 R-P142」。
    er_ok = [make_tc(1, "SWE-PM-071", "Power Down",
                     test_procedure="1. Start the suspend-resume boot sequence\n"
                                    "2. Read the TLM display through SplashScreen_Time "
                                    "to check that no splash screen is shown",
                     expected_result="1. The TLM display stays blank while the boot "
                                     "sequence runs\n"
                                     "2. No splash screen is shown through SplashScreen_Time")]
    # 複述置於**非末步**（三步），使 overlap 判定仍適用
    er_bad = [make_tc(1, "SWE-PM-071", "Power Down",
                      test_procedure="1. Start the suspend-resume boot sequence\n"
                                     "2. Record the elapsed time from boot start\n"
                                     "3. Read the TLM display to check that the splash "
                                     "screen is loaded",
                      expected_result="1. The boot sequence starts\n"
                                      "2. The elapsed time is recorded from boot start\n"
                                      "3. The splash screen is loaded")]
    for label, tcs_, want_rp96, want_rp142 in [
            ("應 PASS —— 非末步無複述（末步依 R-P142 入待裁）", er_ok, 0, 1),
            ("應 FAIL —— 非末步之 ER 複述 procedure 動作", er_bad, 2, 1)]:
        got = check_s6_er_restatement(tcs_)
        n96 = len([f for f in got if f["rule"].startswith("R-P96")])
        n142 = len([f for f in got if f["rule"] == "R-P142"])
        ok = (n96 == want_rp96 if want_rp96 == 0 else n96 >= want_rp96) and n142 == want_rp142
        failures += not ok
        print(f"\n  [{'PASS' if ok else '**FAIL**'}] G73 {label}")
        print(f"          期望 R-P96 {'0' if want_rp96 == 0 else f'≥{want_rp96}'} 項、"
              f"R-P142 {want_rp142} 項；實際 {n96} / {n142}")
        for f in got[:3]:
            print(f"          → {f['rule']} {f['detail'][:76]}")

    # G74 / R-P97 —— 時間量測之 ER 不得寫數值相等
    t74_ok = [make_tc(1, "SWE-PM-071", "Power Down",
                      expected_result="1. The TLM display stays blank\n"
                                      "2. No splash screen is shown through "
                                      "SplashScreen_Time")]
    t74_bad = [make_tc(1, "SWE-PM-071", "Power Down",
                       expected_result="1. The TLM display stays blank\n"
                                       "2. The recorded time equals SplashScreen_Time")]
    for label, tcs_, want in [("應 PASS —— 以行為描述取代數值相等", t74_ok, 0),
                              ("應 FAIL —— ER 寫「equals SplashScreen_Time」", t74_bad, 1)]:
        got = check_s841_time_equality(tcs_)
        ok = (len(got) == 0) if want == 0 else (len(got) >= want)
        failures += not ok
        print(f"\n  [{'PASS' if ok else '**FAIL**'}] G74 {label}")
        print(f"          期望 {'0' if want == 0 else f'≥{want}'} 項；實際 {len(got)} 項")
        for f in got[:2]:
            print(f"          → {f['rule']} {f['detail'][:84]}")

    # G77 / R-P101 —— Final Step 驗證意圖。
    fs_ok = [make_tc(1, "SWE-PM-071", "Power Down",
                     test_procedure="1. Start the suspend-resume boot sequence\n"
                                    "2. Read the TLM display before and after "
                                    "SplashScreen_Time to check that the splash "
                                    "screen is loaded")]
    fs_bad = [make_tc(1, "SWE-PM-071", "Power Down",
                      test_procedure="1. Start the suspend-resume boot sequence\n"
                                     "2. Read the TLM display through SplashScreen_Time")]
    fs_long = [make_tc(1, "SWE-PM-071", "Power Down",
                       test_procedure="1. Start the suspend-resume boot sequence\n"
                                      "2. Read the TLM display carefully and slowly "
                                      "before and after SplashScreen_Time has fully "
                                      "elapsed to check that the splash screen is "
                                      "loaded and shown")]
    for label, tcs_, want in [("應 PASS —— 末步含 `to check that ...`", fs_ok, 0),
                              ("應 FAIL —— 末步無 check target", fs_bad, 1),
                              ("應 FAIL —— 末步逾 §5.2B 之 18 字上限", fs_long, 1)]:
        got = check_s52b_final_step_intent(tcs_)
        ok = (len(got) == 0) if want == 0 else (len(got) >= want)
        failures += not ok
        print(f"\n  [{'PASS' if ok else '**FAIL**'}] G77 {label}")
        print(f"          期望 {'0' if want == 0 else f'≥{want}'} 項；實際 {len(got)} 項")
        for f in got[:2]:
            print(f"          → {f['rule']} {f['detail'][:84]}")

    # G79 / R-P104 —— leaves 之 source_clause 必附
    for label, batch, want in [
            ("應 PASS —— 三 leaf 皆有 source_clause",
             {"leaves": [{"parent": f"SWE-PM-07{i}", "source_clause": "spec text"}
                         for i in (1, 2, 3)]}, 0),
            ("應 FAIL —— 一 leaf 之 source_clause 為空",
             {"leaves": [{"parent": "SWE-PM-071", "source_clause": "spec text"},
                         {"parent": "SWE-PM-072", "source_clause": ""}]}, 1)]:
        got = check_source_clause(batch)
        ok = (len(got) == 0) if want == 0 else (len(got) >= want)
        failures += not ok
        print(f"\n  [{'PASS' if ok else '**FAIL**'}] G79 {label}")
        print(f"          期望 {'0' if want == 0 else f'≥{want}'} 項；實際 {len(got)} 項")

    # G81 / R-P107 —— 誤讀關鍵詞全欄掃描（個案型）。
    mis_ok = {"tcs": [make_tc(1, "SWE-PM-072", "Power Down",
                              test_item="Buffered events processed as soon as possible",
                              split_reason="本條驗處理面：開機仍在進行時即處理")],
              "leaves": [{"parent": "SWE-PM-072", "source_clause": "while the boot is still completing"}]}
    mis_bad = {"tcs": [make_tc(1, "SWE-PM-072", "Power Down",
                               test_item="Buffered events processed after boot completes",
                               split_reason="本條驗處理面：緩衝之事件於開機完成後處理")],
               "leaves": [{"parent": "SWE-PM-072", "source_clause": "while the boot is still completing"}]}
    mis_quoted = {"tcs": [make_tc(1, "SWE-PM-072", "Power Down",
                                  split_reason="查證記錄：規格未載「開機完成後」處理")],
                  "leaves": []}
    for label, batch, want in [("應 PASS —— 無誤讀關鍵詞", mis_ok, 0),
                               ("應 FAIL —— test_item 與 split_reason 殘留", mis_bad, 2),
                               ("應 PASS —— 引號內之引述不判違", mis_quoted, 0)]:
        got = check_misread_terms(batch)
        ok = (len(got) == 0) if want == 0 else (len(got) >= want)
        failures += not ok
        print(f"\n  [{'PASS' if ok else '**FAIL**'}] G81 {label}")
        print(f"          期望 {'0' if want == 0 else f'≥{want}'} 項；實際 {len(got)} 項")
        for f in got[:2]:
            print(f"          → {f['detail'][:80]}")

    # G82 / R-P109 —— ER 之標的須見於 source_clause
    cov_ok = {"tcs": [make_tc(1, "SWE-PM-073", "Power Down",
                              req_id="SWE-PM-073",
                              expected_result="1. AUD_LVL carries the updated level\n"
                                              "2. The TLM is muted")],
              "leaves": [{"parent": "SWE-PM-073",
                          "source_clause": "send the AUD_LVL signal with the updated volume "
                                           "level. If Ecall mode is not active, TLM shall be muted."}]}
    cov_bad = {"tcs": [make_tc(1, "SWE-PM-073", "Power Down",
                               req_id="SWE-PM-073",
                               expected_result="1. AUD_LVL carries the updated level\n"
                                               "2. The TLM is muted")],
               "leaves": [{"parent": "SWE-PM-073",
                           "source_clause": "the TLM shall immediately reduce the maximum "
                                            "volume level to 20 ..."}]}
    for label, batch, want in [("應 PASS —— ER 之標的皆見於 source_clause", cov_ok, 0),
                               ("應 FAIL —— `...` 遮蔽 AUD_LVL 條款", cov_bad, 1)]:
        got = check_er_clause_coverage(batch)
        ok = (len(got) == 0) if want == 0 else (len(got) >= want)
        failures += not ok
        print(f"\n  [{'PASS' if ok else '**FAIL**'}] G82 {label}")
        print(f"          期望 {'0' if want == 0 else f'≥{want}'} 項；實際 {len(got)} 項")
        for f in got[:2]:
            print(f"          → {f['detail'][:80]}")

    # G96 / R-P131 —— `remarks` 已入 LONG_FIELDS，§11 之規則及於該欄。
    for label, remarks, want in [
            ("應 PASS —— remarks 無句點、無方括號",
             "DR-PW8 (High) —— 待 voltage out of range 之門檻值", 0),
            ("應 FAIL —— remarks 以句點結尾", "待門檻值。", 1),
            ("應 FAIL —— remarks 以方括號標 UI 標籤", "見 [Power] 分頁", 1)]:
        got = check_s11_formatting([make_tc(1, "SWE-PM-071", "Power Down",
                                            remarks=remarks)])
        ok = (len(got) == 0) if want == 0 else (len(got) >= want)
        failures += not ok
        print(f"\n  [{'PASS' if ok else '**FAIL**'}] G96 {label}")
        print(f"          期望 {'0' if want == 0 else f'≥{want}'} 項；實際 {len(got)} 項")
        for f in got[:2]:
            print(f"          → {f['detail'][:80]}")

    # G54 / R-P76 —— findings 分流：R-P42(b) 觸發時「待裁決」節有內容而 exit 仍為 0
    triage_tc = [make_tc(1, "SWE-PM-071", "Power Down",
                         design_method="狀態轉換 (State Transition Testing)",
                         specification_reference=(
                             "R1LR_Atl-H_25PI3.5_Activation and Configuration_"
                             "CFTS_010_Power Down _SR26_20250909-1658_1.7.2"),
                         test_procedure=f"1. {fingerprint} on the bench\n"
                                        f"2. Read the reported state to check that it changed",
                         expected_result="1. The step completes\n"
                                         "2. The reported state differs from its initial value")]
    blocking, adjudicate = run_all_gates(
        triage_tc, blacklist, fingerprints, leaf_testset, leaf_priority)
    ok = len(adjudicate) >= 1 and not blocking
    failures += not ok
    print(f"\n  [{'PASS' if ok else '**FAIL**'}] G54 findings 分流"
          f"（R-P42(b) 觸發，exit 仍為 0）")
    print(f"          阻斷類 {len(blocking)} 項（期望 0）；"
          f"待人工裁決類 {len(adjudicate)} 項（期望 ≥ 1）")
    for f in adjudicate[:1]:
        print(f"          → {f['rule']} {f['tc_id']}: {f['detail'][:88]}")

    # G46 / R-P69(c) —— 與 tcs 無關，獨立驗證。
    # 違規案例以**合成 yaml** 注入，不以 repo 現況為對照（09 §I）。
    GOOD_YAML = (
        'feature: "Power"\n'
        'test_group: "Power Management"\n'
        'spec_mode: "D"\n'
        'tc_id_format: "NR1L-PowerManagement-{NNN}"\n'
        'author_value: ""\n'
        'fill_test_group_set: true\n'
    )
    BAD_YAML = (
        'feature: "Power"\n'
        'test_group: "Power"\n'            # R-P2 定為 Power Management
        'spec_mode: "A"\n'                 # R-P9 / R-P3′ 定為 D
        'workbook: "inputs/<FW036 xlsx>"\n'  # 殘留 scaffold placeholder
    )
    for label, text, should_fail in [
        ("應 PASS —— 合成之一致 yaml", GOOD_YAML, False),
        ("應 FAIL —— 合成之漂移 yaml（test_group / spec_mode / placeholder）",
         BAD_YAML, True),
        ("應 PASS —— repo 現況之 feature.yaml", None, False),
    ]:
        yaml_findings = check_feature_yaml_consistency(text)
        ok = bool(yaml_findings) == should_fail
        failures += not ok
        print(f"\n  [{'PASS' if ok else '**FAIL**'}] G46 {label}")
        print(f"          期望 {'FAIL' if should_fail else '（無）'}；"
              f"實際 {len(yaml_findings)} 項")
        for f in yaml_findings[:3]:
            print(f"          → {f['rule']} {f['detail'][:88]}")

    print(f"\n  全部 {len(filled)} 個 TC fixture ＋ G46 皆如期："
          f"{'是' if failures == 0 else '否'}")
    return 1 if failures else 0


def main() -> None:
    blacklist = load_blacklist()
    fingerprints = build_fingerprints(blacklist, anchor_bodies())
    print(f"R-P42 黑名單 {len(blacklist)} 個錨點；"
          f"其中 {len(fingerprints)} 個具獨有特徵字串"
          f"（共 {sum(len(v) for v in fingerprints.values())} 句，"
          f"最短 {MIN_FINGERPRINT} 字元）\n")

    if "--self-test" in sys.argv:
        raise SystemExit(self_test(blacklist, fingerprints))

    if not GENERATED.exists() or not any(GENERATED.glob("*.json")):
        print(f"{GENERATED.relative_to(ROOT)} 無 TC —— Phase 4 尚未開始，閘門就位待用。")
        print("以 --self-test 驗證閘門本身。")
        raise SystemExit(0)

    tcs = load_tcs(GENERATED)
    blocking, adjudicate = run_all_gates(
        tcs, blacklist, fingerprints, load_leaf_testset(), load_037_priority())
    blocking += run_batch_gates(load_batches(GENERATED))

    print(f"檢查 {len(tcs)} 個 TC")
    print(f"\n【阻斷類】{'PASS' if not blocking else f'**{len(blocking)} 項 FAIL**'}")
    for f in blocking:
        print(f"  {f['rule']} {f['tc_id']}: {f['detail']}")

    print(f"\n【待人工裁決類 —— R-P42(b)，依 R-P67 不自動判 FAIL】"
          f"{'（無觸發）' if not adjudicate else f'{len(adjudicate)} 項待裁'}")
    for f in adjudicate:
        print(f"  {f['rule']} {f['tc_id']}: {f['detail']}")
    if adjudicate:
        print("  → 須逐條人工裁決（真違規 / 偽陽性 ＋ 依據）並登記；本節不影響 exit code")

    raise SystemExit(1 if blocking else 0)


if __name__ == "__main__":
    main()
