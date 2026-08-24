#!/usr/bin/env python3
"""25 包 §四 驗收：逐項對 `pm_25.xlsx` 覆核，任一項不達成即 exit 1。

覆核對象為輸出本身（非 plan），因此 plan 與寫入之間若脫節，此處會攤開。
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))

import lint036                                              # noqa: E402
from xlsx_surgical import build_shift                       # noqa: E402

HERE = Path(__file__).resolve().parent
BASE = ROOT / "features/power/sandbox/b19/pm_19.xlsx"
OUT = ROOT / "features/power/sandbox/b26/pm_26.xlsx"
PREV = ROOT / "features/power/sandbox/b25/pm_25.xlsx"
PREV_PLAN = ROOT / "features/power/scripts/b25/plan.json"
FIRST = 10
NUM = re.compile(r"^\s*\d+[.)]")
DENUM = re.compile(r"^\s*\d+[.)]\s*")
A_ROWS = {11, 12, 23, 179, 180}
# 179b／180b 依 26 包 §二 換掉 PRE 第 1 行；其餘面向列 PRE 須與原列逐字同。
PRE_STATE = ("The HU is in FULL OPERATION mode due to an active incoming "
             "phone call")
# 重掃之觀察判準較 §一 之切分判準寬一格：既有未拆列有
# `Attempt … and check that it is rejected` 之合併步（如新 rows 24／29），
# 觀察在句中而非句首。切分判準沿 §一 之句首錨定（已驗 30 列無誤判）。
OBS_HEAD = re.compile(r"^Read\b")
OBS_MID = re.compile(r"check that", re.I)


def has_observation(body: str) -> bool:
    return bool(OBS_HEAD.match(body) or OBS_MID.search(body))


def steps(text) -> list[str]:
    return [ln for ln in str(text or "").split("\n") if NUM.match(ln)]


def bodies(text) -> list[str]:
    return [DENUM.sub("", ln).strip() for ln in steps(text)]


def main() -> int:
    plan = json.loads((HERE / "plan.json").read_text(encoding="utf-8"))
    sheet = plan["sheet"]
    src = openpyxl.load_workbook(BASE)[sheet]
    ws = openpyxl.load_workbook(OUT)[sheet]
    insertions = {int(k): v for k, v in plan["insertions"].items()}
    shift = build_shift(insertions)
    fails: list[str] = []

    def check(ok: bool, label: str, detail: str = "") -> None:
        print(f"  {'達成' if ok else '未達成'}  {label}"
              + (f" — {detail}" if detail else ""))
        if not ok:
            fails.append(label)

    # --- 列數 ---------------------------------------------------------------
    src_rows = src.max_row - FIRST + 1
    out_rows = ws.max_row - FIRST + 1
    inserted = sum(insertions.values())
    check(out_rows == src_rows + inserted, "全本列數",
          f"{src_rows} + {inserted} = {out_rows}")

    # --- 面向列逐列存在，原列為首面向 ---------------------------------------
    a_seen = b_seen = 0
    for split in plan["splits"]:
        anchor = shift(split["src_row"])
        for offset, variant in enumerate(split["variants"]):
            row = anchor + offset
            for field in ("I", "J", "L", "M"):
                if str(ws[f"{field}{row}"].value or "") != variant[field]:
                    fails.append(f"面向列內容 row {row} 欄 {field}")
            if split["src_row"] in A_ROWS:
                a_seen += 1
            else:
                b_seen += 1
    check(a_seen == 14, "A 型面向列", f"{a_seen}/14")
    check(b_seen == 127, "B 型面向列", f"{b_seen}/127")
    check(not [f for f in fails if f.startswith("面向列內容")],
          "面向列四欄與 plan 逐字相符")

    # --- PRE 逐字、setup 段逐字、面向步取自原列 -----------------------------
    setup_of = {a["row"]: a["setup"] for a in plan["audit_b"]}
    pre_bad, setup_bad, body_bad = [], [], []
    for split in plan["splits"]:
        s_row = split["src_row"]
        anchor = shift(s_row)
        src_pre = bodies(src[f"J{s_row}"].value)
        src_proc = bodies(src[f"L{s_row}"].value)
        for offset in range(len(split["variants"])):
            row = anchor + offset
            out_pre = bodies(ws[f"J{row}"].value)
            allowed = (src_pre, [PRE_STATE, *src_pre[1:]])
            if out_pre not in allowed:
                pre_bad.append(row)
            if s_row in A_ROWS:
                continue
            setup = setup_of[s_row]
            out_proc = bodies(ws[f"L{row}"].value)
            if out_proc[:setup] != src_proc[:setup]:
                setup_bad.append(row)
            # 面向本體之各步亦須逐字取自原列（順序保持）。
            rest, at = out_proc[setup:], setup
            for body in rest:
                if body not in src_proc[at:]:
                    body_bad.append((row, body))
                    break
                at = src_proc.index(body, at) + 1
    check(not pre_bad, "面向列 PRE 與原列逐字同（179b／180b 換狀態行）",
          f"例外 {pre_bad[:5]}")
    check(not setup_bad, "B 型 setup 段與原列逐字同", f"例外 {setup_bad[:5]}")
    check(not body_bad, "B 型面向步逐字取自原列且順序保持",
          f"例外 {body_bad[:3]}")

    # --- §四：規則 2 v2 全表重掃 —— 不得有無觀察步之列 ----------------------
    no_obs = [row for row in range(FIRST, ws.max_row + 1)
              if bodies(ws[f"L{row}"].value)
              and not any(has_observation(b)
                          for b in bodies(ws[f"L{row}"].value))]
    check(not no_obs, "全本無「無觀察步」之列（規則 2 v2 重掃）",
          f"例外 {no_obs[:8]}")

    # --- ID 連續無跳號無重複 -------------------------------------------------
    ids, stubs, nos = [], [], []
    for row in range(FIRST, ws.max_row + 1):
        nos.append(ws[f"B{row}"].value)
        tc = str(ws[f"F{row}"].value or "").strip()
        (ids if tc else stubs).append(row if not tc else tc)
    nums = [int(i.rsplit("-", 1)[1]) for i in ids]
    check(nums == list(range(1, len(nums) + 1)), "Test Case ID 連續無跳號",
          f"001–{max(nums):03d}／{len(nums)} 列")
    check(len(set(ids)) == len(ids), "Test Case ID 無重複")
    check(nos == list(range(1, out_rows + 1)), "No.# 連續重編",
          f"1–{out_rows}")
    check(stubs == [shift(230)], "存根列略過不給 ID", f"{stubs}")

    # --- proc↔er 編號數逐列相等 ---------------------------------------------
    e_bad = [row for row in range(FIRST, ws.max_row + 1)
             if len(steps(ws[f"L{row}"].value)) !=
             len(steps(ws[f"M{row}"].value))]
    check(not e_bad, "proc↔er 編號數逐列相等", f"例外 {e_bad[:5]}")

    # --- 括號下半：同 Requirement ID 下不得逐字相同 -------------------------
    groups: dict[tuple[str, str], list[int]] = defaultdict(list)
    for row in range(FIRST, ws.max_row + 1):
        req = str(ws[f"D{row}"].value or "").strip()
        paren = "\n".join(lint036.paren_lines(str(ws[f"I{row}"].value or "")))
        if req and paren:
            groups[(req, paren)].append(row)
    dups = {k: v for k, v in groups.items() if len(v) > 1}
    check(not dups, "括號下半同源不逐字相同", f"{list(dups.values())[:3]}")

    # --- 未拆列零變動 -------------------------------------------------------
    split_rows = {s["src_row"] for s in plan["splits"]}
    drift = []
    for s_row in range(FIRST, src.max_row + 1):
        if s_row in split_rows:
            continue
        row = shift(s_row)
        for field in ("I", "J", "K", "L", "M", "N", "P", "R"):
            if (src[f"{field}{s_row}"].value or "") != \
                    (ws[f"{field}{row}"].value or ""):
                drift.append((s_row, row, field))
    check(not drift, "未拆原列非 ID 欄零變動", f"例外 {drift[:5]}")

    # --- 與 25 包之等價性（硬閘門）------------------------------------------
    # 本包實際基底為 b19（`xlsx_surgical` 無刪列段），故須逐格證明輸出即
    # 「pm_25 縮併 17 列 + 4 列 PRE 修正 + ID／No.# 重排」，別無他變。
    prev_plan = json.loads(PREV_PLAN.read_text(encoding="utf-8"))
    prev = openpyxl.load_workbook(PREV)[sheet]
    prev_shift = build_shift({int(k): v
                              for k, v in prev_plan["insertions"].items()})
    prev_setup = {a["row"]: a["setup"] for a in prev_plan["audit_b"]}
    carried = [c for c in "ACDEGHKNOPQRS" if c not in "BF"]
    dropped: list[int] = []
    eq_bad: list[tuple] = []

    def compare(row: int, prev_row: int, *, fields: tuple[str, ...]) -> None:
        for field in fields:
            if (ws[f"{field}{row}"].value or "") != \
                    (prev[f"{field}{prev_row}"].value or ""):
                eq_bad.append((row, prev_row, field))

    matched: set[int] = set()
    for split in plan["splits"]:
        s_row = split["src_row"]
        anchor, prev_anchor = shift(s_row), prev_shift(s_row)
        setup = prev_setup.get(s_row, 0)
        cum = 0
        for offset, variant in enumerate(split["variants"]):
            row = anchor + offset
            if s_row in A_ROWS:
                prev_row = prev_anchor + offset
                span = [prev_row]
            else:
                cum += len(bodies(variant["L"])) - setup
                prev_row = prev_anchor + cum - 1
                span = list(range(prev_anchor + cum
                                  - (len(bodies(variant["L"])) - setup),
                                  prev_anchor + cum))
            matched.add(prev_row)
            dropped.extend(r for r in span if r != prev_row)
            # 未縮併之欄位：test_item（括號）與各承載欄逐字相同。
            compare(row, prev_row, fields=("I", *carried))
            # PRE：僅 179b／180b 得異動。
            if not (s_row in (179, 180) and offset == 1):
                compare(row, prev_row, fields=("J",))
            # 縮併列之 PROC／ER = 被刪列與存留列步驟序列之串接。
            want_l = [b for r in span for b in bodies(prev[f"L{r}"].value)[setup:]]
            want_m = [b for r in span for b in bodies(prev[f"M{r}"].value)[setup:]]
            if bodies(variant["L"])[setup:] != want_l or \
                    bodies(variant["M"])[setup:] != want_m:
                eq_bad.append((row, prev_row, "L/M 串接"))

    for s_row in range(FIRST, src.max_row + 1):
        if s_row in {sp["src_row"] for sp in plan["splits"]}:
            continue
        compare(shift(s_row), prev_shift(s_row), fields=("I", "J", "L", "M",
                                                         *carried))

    check(not eq_bad, "與 25 包逐格等價（除縮併／PRE／ID／No.#）",
          f"例外 {eq_bad[:5]}")
    check(len(dropped) == 17, "被刪列恰 17 列", f"{len(dropped)} 列 {dropped}")
    dropped_bad = [r for r in dropped
                   if has_observation(bodies(prev[f"L{r}"].value)[-1])]
    check(not dropped_bad, "被刪列末步皆為驅動步（無觀察）",
          f"例外 {dropped_bad[:5]}")

    print(f"\n{'全項達成' if not fails else f'未達成 {len(fails)} 項'}")
    return 0 if not fails else 1


if __name__ == "__main__":
    sys.exit(main())
