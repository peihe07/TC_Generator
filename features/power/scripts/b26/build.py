#!/usr/bin/env python3
"""26 包分析層：拆分缺陷修正 → `plan.json`。

26 包 §一／§二。25 包之 407 列有 17 列「setup + 驅動步」而無觀察步 ——
該列驗不到任何東西，且 lint 抓不到（`proc↔er` 編號數相符，E=0 照樣成立）。

- **§一 規則 2 v2**：setup 段之後以**觀察步**為錨切面向 —— 每一觀察步
  （`Read`／`Check that` 起首）為一面向之終點，其前之連續驅動步
  （`Send`／`Select`／`Set`／`Wait`／`Let`／`Keep`／`Run`／`Place`／`End`
  起首）併入該面向、不自成一列。面向數 = setup 段後之觀察步數。
  B 型 144 → **127** 列，全本 407 → **390** 列。
- **§二 179b／180b PRE 互斥修正**：25 包依下放包 §一 增行後，PRE 同時
  聲明 IDLE 與通話進行中。改為 PRE 第 1 行換成
  `The HU is in FULL OPERATION mode due to an active incoming phone call`，
  並不再增通話行（狀態行已含通話前提）。

規則 1（setup 段判定，含 25 包之末步修補）與括號消歧沿用 25 包，未動。
括號取該面向之**觀察步**（= 該面向末步），故未縮併之面向列括號與 25 包
逐字相同 —— 本包輸出即「25 包 407 列縮併 17 列 + 4 列 PRE 修正」。
"""

from __future__ import annotations
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "scripts"))

import lint036                                          # noqa: E402

BASE = ROOT / "features/power/sandbox/b19/pm_19.xlsx"
OUT = Path(__file__).resolve().parent / "plan.json"
SHEET = "Test Case Specification&Result"

COL = {"no": "B", "req": "D", "tc": "F", "item": "I",
       "pre": "J", "proc": "L", "er": "M"}

NUM = re.compile(r"^\s*\d+[.)]")
DENUM = re.compile(r"^\s*\d+[.)]\s*")
POWERSTS = "Read the signal $STATUS_TELEMATIC.PowerSts_Telematic$"
LEAD_VERB = re.compile(r"^(Send|Bring|Power up|Reconnect)\b")
# §一 規則 2 v2：面向以觀察步為錨，其前之連續驅動步併入同一面向。
OBSERVE_VERB = re.compile(r"^(Read|Check that)\b")
DRIVE_VERB = re.compile(
    r"^(Send|Select|Set|Wait|Let|Keep|Run|Place|End)\b")
PAREN_WORD_LIMIT = 20

B_ROWS = [10, 17, 21, 24, 26, 28, 29, 30, 32, 39, 45, 97, 102, 109,
          124, 125, 126, 127, 157, 158, 159, 162, 170, 188, 189, 190,
          194, 197, 204, 285]

# 下放包 §二 之分析層預核值（= 該列 ER 行數）。實測不一致時以演算法為準。
B_TABLE = {10: 5, 17: 5, 21: 6, 24: 7, 26: 7, 28: 5, 29: 6, 30: 7, 32: 5,
           39: 5, 45: 7, 97: 5, 102: 5, 109: 5, 124: 8, 125: 8, 126: 8,
           127: 8, 157: 6, 158: 6, 159: 6, 162: 6, 170: 5, 188: 6, 189: 8,
           190: 5, 194: 8, 197: 5, 204: 5, 285: 6}


# --- 取用原列 ---------------------------------------------------------------

def steps(text: str) -> list[str]:
    """編號行清單，逐字（不去編號）。"""
    return [ln for ln in str(text).split("\n") if NUM.match(ln)]


def denum(line: str) -> str:
    return DENUM.sub("", line).strip()


def renumber(lines: list[str]) -> str:
    return "\n".join(f"{i}. {denum(ln)}" for i, ln in enumerate(lines, 1))


def item_upper(text: str) -> str:
    """test_item 上半，逐字（沿 lint036 之括號行定義）。"""
    return lint036.upper_half(str(text))


def item_paren(text: str) -> str:
    lines = lint036.paren_lines(str(text))
    return lines[0] if lines else ""


def compose_item(upper: str, paren: str) -> str:
    """上半 + 空行 + 括號行。`upper_half` 會留下原列之空行，須先剪掉，
    否則組出三個換行、與原表形態不符。"""
    return f"{upper.rstrip()}\n\n{paren}"


def lower_first(s: str) -> str:
    """句首降格。首詞為縮寫或識別字時不動 —— `PROXI` 不可壓成 `pROXI`，
    `AUD_LVL`／`SwitchOff_Timeout_Setting.Req` 同理。"""
    if not s:
        return s
    head = s.split(maxsplit=1)[0]
    if not (head[:1].isupper() and head[1:].isalpha() and head[1:].islower()):
        return s
    return s[:1].lower() + s[1:]


# --- §二 規則 1：setup 段判定（含修補） --------------------------------------

def setup_length(proc_steps: list[str]) -> tuple[int, str]:
    """回傳 (setup 步數, 判定依據)。

    §二 規則 1 三分支，加一修補：`PowerSts` 讀取步為 PROC **末步**時
    屬觀察步，不併入 setup。
    """
    hits = [i for i, ln in enumerate(proc_steps) if POWERSTS in ln]
    if hits and hits[0] < len(proc_steps) - 1:
        return hits[0] + 1, "powersts"
    lead = 0
    for ln in proc_steps:
        if LEAD_VERB.match(denum(ln)):
            lead += 1
        else:
            break
    if lead:
        return lead, "lead-verb" + ("+repair" if hits else "")
    return 1, "first-step" + ("+repair" if hits else "")


def aspect_groups(procs: list[str], setup: int, row: int) -> list[list[int]]:
    """§一 規則 2 v2：setup 段後以觀察步為錨切面向，回傳各面向之步索引。

    尾端若有無觀察步可歸之驅動步，或面向內出現既非驅動亦非觀察之步，
    即中止 —— 兩者都表示動詞清單漏了東西，不可默默放行。
    """
    groups: list[list[int]] = []
    current: list[int] = []
    for k in range(setup, len(procs)):
        body = denum(procs[k])
        if not (OBSERVE_VERB.match(body) or DRIVE_VERB.match(body)):
            raise ValueError(f"row {row} 第 {k + 1} 步既非驅動亦非觀察："
                             f"{body!r}")
        current.append(k)
        if OBSERVE_VERB.match(body):
            groups.append(current)
            current = []
    if current:
        raise ValueError(f"row {row} 尾端 {len(current)} 個驅動步無觀察步可歸："
                         f"{[denum(procs[k]) for k in current]}")
    if not groups:
        raise ValueError(f"row {row}: setup={setup} 之後無觀察步")
    return groups


# --- 括號下半 ---------------------------------------------------------------

def aspect_paren(proc_step: str, er_step: str) -> str:
    """括號下半，取該面向之觀察步（= 該面向末步）；逾 20 詞時取 ER 行。

    取觀察步而非整段，是本包輸出與 25 包逐字銜接之所繫：未縮併之面向
    其觀察步即唯一步，括號因此與 25 包相同。
    """
    full = f"({lower_first(denum(proc_step))} -> {denum(er_step)})"
    if len(full.split()) <= PAREN_WORD_LIMIT:
        return full
    return f"({denum(er_step)})"


def _pre_lines(v: dict) -> list[str]:
    return [denum(ln) for ln in steps(v.get("J", ""))]


def pick_setup_head(v: dict, _group: list[dict]) -> str:
    """消歧候選 1：該列 setup 首步（= PROC 第 1 步）。"""
    return denum(v.get("setup_head", ""))


def pick_pre_discriminator(v: dict, group: list[dict]) -> str:
    """消歧候選 2：該列 PRE 中未被同組全體共有之首行。

    原表既有之區分形態即此 —— rows 189／194 之括號前綴
    `The TLM is in BODY ON mode` 就是該列 PRE 首行。
    """
    common = set(_pre_lines(group[0]))
    for other in group[1:]:
        common &= set(_pre_lines(other))
    for line in _pre_lines(v):
        if line not in common:
            return line
    return ""


DISCRIMINATORS = (pick_setup_head, pick_pre_discriminator)


def disambiguate(variants: list[dict]) -> tuple[int, list[list[str]]]:
    """同一 Requirement ID 下括號逐字相同者，加前綴消歧。

    §四 要求任兩同源面向列括號不得逐字相同，而 lint I-sibling 之同源判準
    為同一 Requirement ID。候選前綴依序試（setup 首步 → 具區分性之 PRE 行），
    取第一個能解開該組者；沿用原列既有之 `—` 分隔形態，不引入新文字。
    未拆之原列（`frozen`）只參與碰撞判定，不被改寫。
    """
    base = {id(v): item_paren(v["I"]) for v in variants}
    fixed: set[int] = set()

    def collisions() -> list[list[dict]]:
        groups: dict[tuple[str, str], list[dict]] = defaultdict(list)
        for v in variants:
            groups[(v["req"], item_paren(v["I"]))].append(v)
        return [m for m in groups.values() if len(m) > 1]

    for pick in DISCRIMINATORS:
        for group in collisions():
            movable = [v for v in group if not v.get("frozen")]
            if not movable:
                continue
            for v in movable:
                prefix = lower_first(pick(v, group))
                inner = base[id(v)][1:-1]
                v["I"] = compose_item(
                    v["upper"],
                    f"({prefix} — {inner})" if prefix else base[id(v)])
                fixed.add(id(v))

    residual = [[f"{v['src_row']}" for v in g] for g in collisions()]
    return len(fixed), residual


# --- A 型 -------------------------------------------------------------------

IGNITION = [(5, "Ignition_Pre_Start"), (6, "Ignition_Start"),
            (7, "Ignition_Cranking"), (8, "Ignition_On_EngOn")]

AUDIO = [
    ("SDCARD",
     "1. Select SDCARD as the audio active source\n"
     "2. Read the played audio source and check that it is the SDCARD",
     "1. The SDCARD is selected as the audio active source\n"
     "2. The TLM plays the SDCARD as the audio active source",
     "(select SDCARD -> SDCARD is played)"),
    ("BT",
     "1. Select BT Music streaming as the audio active source\n"
     "2. Read the played audio source and check that it is the BT Music streaming",
     "1. The BT Music streaming is selected as the audio active source\n"
     "2. The TLM plays the BT Music streaming as the audio active source",
     "(select BT Music streaming -> BT Music streaming is played)"),
    ("call",
     "1. Place a phone call\n"
     "2. Read the played audio source and check that it is the phone call",
     "1. The phone call is established\n"
     "2. The TLM plays the phone call as the audio active source",
     "(place a phone call -> the phone call is played)"),
]

CALL_TRANSITION = [
    ("a", None,
     "1. Let the bench place an incoming phone call to the HU\n"
     "2. Read the HU mode and check that it is FULL OPERATION",
     "1. The incoming phone call is presented to the HU\n"
     "2. The HU transitions from IDLE to FULL OPERATION",
     "incoming call -> FULL OPERATION"),
    # §二：b 列之 PRE 第 1 行換為含通話前提之狀態行，不再增行。
    ("b", "The HU is in FULL OPERATION mode due to an active incoming "
          "phone call",
     "1. Let the phone call become inactive\n"
     "2. Read the HU mode and check that it is IDLE",
     "1. The phone call becomes inactive\n"
     "2. The HU transitions back to IDLE",
     "call becomes inactive -> IDLE"),
]


def pre_replace_state(pre: str, state: str) -> str:
    """PRE 第 1 行（狀態行）換為 `state`，其餘行逐字不動。"""
    lines = [denum(ln) for ln in steps(pre)]
    lines[0] = state
    return "\n".join(f"{i}. {ln}" for i, ln in enumerate(lines, 1))


def build_a(src: dict[int, dict]) -> list[dict]:
    out: list[dict] = []

    r = src[11]
    for value, label in IGNITION:
        out.append({
            "src_row": 11, "req": r["req"], "upper": item_upper(r["I"]),
            "J": r["J"],
            "L": (f"1. Send the signal $STATUS_BH_BCM1.OperationalModeSts$ "
                  f"= {value} ({label})\n"
                  f"2. Read the signal $STATUS_TELEMATIC.PowerSts_Telematic$ "
                  f"and check that it is 4 (Full_Operation)"),
            "M": (f"1. The signal $STATUS_BH_BCM1.OperationalModeSts$ "
                  f"= {value} ({label}) is registered without a bus error\n"
                  f"2. The signal value $STATUS_TELEMATIC.PowerSts_Telematic$ "
                  f"= 4 (Full_Operation) is received"),
            "I": compose_item(item_upper(r["I"]),
                              f"(send OperationalModeSts = {value} {label} "
                              f"-> Full-Operation is kept)"),
        })

    for row in (12, 23):
        r = src[row]
        for _tag, proc, er, paren in AUDIO:
            out.append({"src_row": row, "req": r["req"],
                        "upper": item_upper(r["I"]), "J": r["J"],
                        "L": proc, "M": er,
                        "I": compose_item(item_upper(r["I"]), paren)})

    for row in (179, 180):
        r = src[row]
        # 原列括號之區分前綴（display 為 main screen／projection call UI），
        # 保留之以維持 §四 sibling 可分。
        prefix = item_paren(r["I"])[1:].split(" — ", 1)[0]
        for _tag, state, proc, er, tail in CALL_TRANSITION:
            out.append({
                "src_row": row, "req": r["req"], "upper": item_upper(r["I"]),
                "J": pre_replace_state(r["J"], state) if state else r["J"],
                "L": proc, "M": er,
                "I": compose_item(item_upper(r["I"]), f"({prefix} — {tail})"),
            })
    return out


# --- B 型 -------------------------------------------------------------------

def build_b(src: dict[int, dict]) -> tuple[list[dict], list[dict]]:
    out: list[dict] = []
    audit: list[dict] = []
    for row in B_ROWS:
        r = src[row]
        procs, ers = steps(r["L"]), steps(r["M"])
        if len(procs) != len(ers):
            raise ValueError(f"row {row}: proc {len(procs)} != er {len(ers)}")
        s, basis = setup_length(procs)
        groups = aspect_groups(procs, s, row)
        audit.append({"row": row, "req": r["req"], "proc_steps": len(procs),
                      "er_lines": len(ers), "setup": s, "basis": basis,
                      "aspects": len(groups), "aspects_v1": len(procs) - s,
                      "merged": len(procs) - s - len(groups),
                      "table": B_TABLE[row],
                      "delta": len(groups) - B_TABLE[row]})
        for group in groups:
            last = group[-1]
            out.append({
                "src_row": row, "req": r["req"], "upper": item_upper(r["I"]),
                "setup_head": procs[0],
                "J": r["J"],
                "L": renumber(procs[:s] + [procs[k] for k in group]),
                "M": renumber(ers[:s] + [ers[k] for k in group]),
                "I": compose_item(item_upper(r["I"]),
                                  aspect_paren(procs[last], ers[last])),
            })
    return out, audit


# --- 主流程 -----------------------------------------------------------------

def read_source() -> dict[int, dict]:
    wb = openpyxl.load_workbook(BASE)
    ws = wb[SHEET]
    src: dict[int, dict] = {}
    for row in range(10, ws.max_row + 1):
        src[row] = {"req": str(ws[f"{COL['req']}{row}"].value or ""),
                    "I": ws[f"{COL['item']}{row}"].value or "",
                    "J": ws[f"{COL['pre']}{row}"].value or "",
                    "L": ws[f"{COL['proc']}{row}"].value or "",
                    "M": ws[f"{COL['er']}{row}"].value or ""}
    return src


def main() -> None:
    src = read_source()
    a_variants = build_a(src)
    b_variants, audit = build_b(src)

    for v in a_variants:
        v.setdefault("setup_head", "1. " + denum(steps(v["L"])[0]))

    # 消歧須看全本同 Requirement ID 之列，含未拆之原列。
    split_rows = {v["src_row"] for v in a_variants + b_variants}
    untouched = [{"req": r["req"], "upper": item_upper(r["I"]),
                  "I": r["I"], "src_row": row, "frozen": True,
                  "setup_head": ""}
                 for row, r in src.items()
                 if row not in split_rows and str(r["I"]).strip()]
    fixed, residual = disambiguate(a_variants + b_variants + untouched)
    frozen_changed = [v for v in untouched if v.get("frozen")
                      and v["I"] != src[v["src_row"]]["I"]]

    by_src: dict[int, list[dict]] = defaultdict(list)
    for v in a_variants + b_variants:
        by_src[v["src_row"]].append(v)

    plan = {
        "base": str(BASE.relative_to(ROOT)),
        "sheet": SHEET,
        "insertions": {str(row): len(vs) - 1
                       for row, vs in sorted(by_src.items())},
        "splits": [{"src_row": row,
                    "variants": [{k: v[k] for k in ("I", "J", "L", "M")}
                                 for v in vs]}
                   for row, vs in sorted(by_src.items())],
        "audit_b": audit,
        "paren_disambiguated": fixed,
        "paren_residual_collisions": residual,
        "frozen_rows_touched": [v["src_row"] for v in frozen_changed],
    }
    OUT.write_text(json.dumps(plan, ensure_ascii=False, indent=1),
                   encoding="utf-8")

    a_rows = len(a_variants)
    b_rows = len(b_variants)
    print(f"A 型 {len({v['src_row'] for v in a_variants})} 原列 → {a_rows} 列")
    merged = sum(a["merged"] for a in audit)
    print(f"B 型 {len(B_ROWS)} 原列 → {b_rows} 列"
          f"（25 包 v1 {sum(a['aspects_v1'] for a in audit)}，縮併 {merged}；"
          f"預核表 {sum(B_TABLE.values())}）")
    print(f"插入列數 {sum(plan['insertions'].values())}"
          f"（A {a_rows - 5} + B {b_rows - len(B_ROWS)}）")
    print(f"括號消歧 {fixed} 列；殘餘碰撞 {len(residual)} 組；"
          f"未拆原列被動改動 {len(frozen_changed)} 列")
    print(f"縮併列數 {merged}（下放包 §一 之 17 列清單）")
    print("→", OUT.relative_to(ROOT))


if __name__ == "__main__":
    main()
