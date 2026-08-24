#!/usr/bin/env python3
"""26 包執行層：拆分缺陷修正套用 ＋ 全本 ID 重排。

26 包 §四。基底 `features/power/sandbox/b19/pm_19.xlsx`（b4dd5ca0…）。

兩段式寫出，兩段都不經 `openpyxl.save`：

1. `surgical_insert_rows` —— 純結構插列（123 列），逐列複製錨列之列高與
   逐格樣式，並搬移 `<sheetData>` 外之列位參照（autoFilter／
   conditionalFormatting／dataValidation／x14 `xm:sqref`）。
2. `surgical_save` —— 寫入面向列四欄（`test_item`／`pre`／`proc`／`er`）、
   面向列其餘各欄自錨列逐字複製，最後全本重排 `Test Case ID`（F）與
   `No.#`（B）。

基底仍取 b19（25 包之基底），而非下放包 §四 所寫之 `sandbox/b25/pm_25.xlsx`：
縮併 17 列須**刪列**，而 `xlsx_surgical` 只有插列段、無刪列段。以修正後之
規則 2 v2 自 b19 重出（插 106 列而非 123 列），結果與「b25 縮併 17 列 +
4 列 PRE 修正」逐字等價 —— 該等價性由 `verify.py` 作硬性閘門逐格比對，
不成立即 exit 1。如此免去一個只此一用的刪列工具，寫出仍為既有二段一路。

沿用 25 包之 Pei 裁定：規則 2（v2）為準 + 修補規則 1；row 230 存根列
略過不給 ID、B 欄一併重編。
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))

import lint036                                              # noqa: E402
from xlsx_surgical import (                                 # noqa: E402
    StructureError, build_shift, sheet_members, surgical_insert_rows,
    surgical_save,
)

HERE = Path(__file__).resolve().parent
PLAN = HERE / "plan.json"
BASE = ROOT / "features/power/sandbox/b19/pm_19.xlsx"
BASE_SHA = "b4dd5ca0c0f02394117e52d4c8b342743d1ccef236d5b2ca392f8ba16f9871ca"
SANDBOX = ROOT / "features/power/sandbox/b26"
OUT = SANDBOX / "pm_26.xlsx"

FIRST_DATA_ROW = 10
LAST_COL = 34
ID_TEMPLATE = "NR1L-PowerManagement-{:03d}"
COL_NO, COL_TC = "B", "F"
WRITTEN_FIELDS = ("I", "J", "L", "M")


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def x14_dropdowns(path: Path, member: str) -> list[str]:
    """讀回 x14 下拉之 `xm:sqref`，寫回前後須逐項對得上。"""
    with zipfile.ZipFile(path) as z:
        xml = z.read(member).decode("utf-8")
    return re.findall(r"<xm:sqref>([^<]*)</xm:sqref>", xml)


def snapshot(ws, row: int) -> dict[int, object]:
    return {c: ws.cell(row, c).value for c in range(1, LAST_COL + 1)}


def is_stub(ws, row: int) -> bool:
    """存根列：無 Test Case ID 且無 test_item。row 230（SWE-PM-089）。"""
    return not str(ws[f"{COL_TC}{row}"].value or "").strip() and \
        not str(ws["I%d" % row].value or "").strip()


def main() -> None:
    plan = json.loads(PLAN.read_text(encoding="utf-8"))
    sheet = plan["sheet"]

    if sha256(BASE) != BASE_SHA:
        raise StructureError(f"基底 sha256 不符：{sha256(BASE)[:20]}")

    insertions = {int(k): v for k, v in plan["insertions"].items()}
    shift = build_shift(insertions)
    member = sheet_members(BASE)[sheet]
    dv_before = x14_dropdowns(BASE, member)

    SANDBOX.mkdir(parents=True, exist_ok=True)
    expanded = SANDBOX / ".expanded.xlsx"
    ins_report = surgical_insert_rows(BASE, expanded, insertions, sheet)

    wb = openpyxl.load_workbook(expanded)
    ws = wb[sheet]

    # --- 面向列四欄寫入；其餘各欄自錨列逐字複製 -----------------------------
    cells = 0
    for split in plan["splits"]:
        anchor = shift(split["src_row"])
        source = snapshot(ws, anchor)
        for offset, variant in enumerate(split["variants"]):
            row = anchor + offset
            if offset:                          # 新插入列：先補齊非拆分欄
                for col, value in source.items():
                    ws.cell(row, col).value = value
            for field in WRITTEN_FIELDS:
                ws[f"{field}{row}"] = variant[field]
                cells += 1

    # --- §三 全本重排：Test Case ID 連續，No.# 一併重編 ---------------------
    seq = 0
    stub_rows: list[int] = []
    for order, row in enumerate(range(FIRST_DATA_ROW, ws.max_row + 1), 1):
        ws[f"{COL_NO}{row}"] = order
        if is_stub(ws, row):
            stub_rows.append(row)
            continue
        seq += 1
        ws[f"{COL_TC}{row}"] = ID_TEMPLATE.format(seq)

    report = surgical_save(wb, expanded, OUT)
    expanded.unlink()

    dv_after = x14_dropdowns(OUT, member)
    if len(dv_before) != len(dv_after):
        raise StructureError(f"x14 讀回不符：{dv_before} → {dv_after}")

    rows_total = ws.max_row - FIRST_DATA_ROW + 1
    print(f"插入 {ins_report['inserted']} 列／{ins_report['anchors']} 個錨列")
    print(f"面向四欄寫入 {cells} 格／{len(plan['splits'])} 個原列")
    print(f"資料列 {rows_total}；Test Case ID 001–{seq:03d}；"
          f"存根列略過 {stub_rows}")
    print(f"x14 讀回 {dv_before} → {dv_after}")
    print("zip 成員", report["members"], "| 差異成員", report["differing"])
    print("out sha256:", sha256(OUT)[:20])


if __name__ == "__main__":
    main()
