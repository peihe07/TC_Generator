"""Layer 3 全集（R-P24）。

R-P24：一個 leaf 得對應多個 Layer 3 章節，Layer 3 記全集不擇一；
Layer 2（Test Set）仍為單值。故本檔逐 (leaf, 章節) 一列，
一個 leaf 可出現多列。

輸出 features/power/data/layer3_full.tsv，欄位：

    leaf            SWE-Requirement ID
    cfts            009 / 010
    chapter_num     章節號
    chapter_title   章節標題
    hit_count       該 leaf 命中此章節之次數（= 解析到此章節之 item 數）
    item_ids        命中之 Polarion item id，以 , 分隔
    tokens          命中此章節之 Sys-RA token，以 , 分隔

排序：leaf 之數字序，其次章節號之數字序 —— 使輸出可重現、可 diff。

用法：
    python features/power/scripts/build_layer3.py
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
IN = ROOT / "features/power/inputs"
DATA = ROOT / "features/power/data"

PM_RE = re.compile(r"Sys-RA-PM-\d{4}")
PD_RE = re.compile(r"Sys-RA-PD[_-]\d+")
SYS2_LAST = {"009": 339, "010": 74}


def find(pattern: str) -> Path:
    return next(f for f in IN.iterdir() if pattern in f.name)


def sys2_map(path: Path, last_row: int) -> dict[str, list[str]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Basic Report"]
    out = {}
    for r in ws.iter_rows(min_row=2, max_row=last_row, values_only=True):
        key = str(r[1] or "").strip()
        if key:
            out[key] = re.findall(r"\d{6,8}", str(r[4] or ""))
    wb.close()
    return out


def chapter_sort_key(num: str) -> tuple[int, ...]:
    return tuple(int(x) for x in num.split("."))


def main() -> None:
    chapters = json.loads((DATA / "item_to_chapter.json").read_text(encoding="utf-8"))
    tables = {"009": chapters["cfts009"], "010": chapters["cfts010"]}
    sys2 = {
        "009": sys2_map(find("SYS2_CFTS_009"), SYS2_LAST["009"]),
        "010": sys2_map(find("SYS2_CFTS_010"), SYS2_LAST["010"]),
    }

    wb = openpyxl.load_workbook(find("FSM-037"), data_only=True, read_only=True)
    ws = wb["SWE1 Requirements"]
    leaves = [
        (str(r[0]).strip(), str(r[1] or ""))
        for r in ws.iter_rows(min_row=8, max_row=145, values_only=True)
        if r[0] and str(r[0]).strip()
    ]
    wb.close()

    rows = []
    for leaf_id, src in leaves:
        # (cfts, chapter_num) -> {title, items, tokens}
        by_chapter: dict[tuple[str, str], dict] = {}
        for tok in PM_RE.findall(src) + PD_RE.findall(src):
            domain = "009" if tok.startswith("Sys-RA-PM") else "010"
            for item in sys2[domain].get(tok, []):
                if item not in tables[domain]:
                    continue
                num, title = tables[domain][item]
                entry = by_chapter.setdefault(
                    (domain, num), {"title": title, "items": [], "tokens": []}
                )
                entry["items"].append(item)
                if tok not in entry["tokens"]:
                    entry["tokens"].append(tok)
        for (domain, num), entry in by_chapter.items():
            rows.append((
                leaf_id, domain, num, entry["title"],
                len(entry["items"]), ",".join(entry["items"]), ",".join(entry["tokens"]),
            ))

    def leaf_key(leaf_id: str) -> int:
        m = re.match(r"SWE-PM-(\d+)$", leaf_id)
        return int(m.group(1)) if m else 10**9

    rows.sort(key=lambda r: (leaf_key(r[0]), r[1], chapter_sort_key(r[2])))

    header = "leaf\tcfts\tchapter_num\tchapter_title\thit_count\titem_ids\ttokens"
    path = DATA / "layer3_full.tsv"
    path.write_text(
        header + "\n" + "\n".join("\t".join(str(c) for c in r) for r in rows) + "\n",
        encoding="utf-8",
    )

    leaves_with_rows = {r[0] for r in rows}
    distinct_chapters = {(r[1], r[2]) for r in rows}
    multi = sorted(
        {lid for lid in leaves_with_rows if sum(1 for r in rows if r[0] == lid) > 1},
        key=leaf_key,
    )
    print(f"wrote {path.relative_to(ROOT)} — {len(rows)} 列")
    print(f"  涵蓋 leaf 數: {len(leaves_with_rows)} / {len(leaves)}")
    print(f"  G13b 相異章節總數: {len(distinct_chapters)}")
    print(f"  多章節 leaf 數: {len(multi)}  {multi}")
    print(f"  每 leaf 章節數分布: "
          f"{dict(sorted({n: sum(1 for l in leaves_with_rows if sum(1 for r in rows if r[0] == l) == n) for n in {sum(1 for r in rows if r[0] == l) for l in leaves_with_rows}}.items()))}")


if __name__ == "__main__":
    main()
