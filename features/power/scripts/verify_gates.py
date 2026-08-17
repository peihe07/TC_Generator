"""§D 閃點自驗（03 包版）與 §E leaf 分布重算。

前置：先跑 features/power/scripts/extract_textlayer.py 產出
      features/power/data/item_to_chapter.json

用法：
    python features/power/scripts/verify_gates.py
"""

from __future__ import annotations

import hashlib
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
IN = ROOT / "features/power/inputs"
DATA = ROOT / "features/power/data"

# §B 台帳：原始檔 SHA256 全 64 碼（G0 前置閘）
LEDGER = {
    "ce93174794d0d43c03d25dcd577c2811b85a8ebb2fd754a5201e5d6979297eda": "FW036-A01 交付標的",
    "2284abf5e6c17e4d1a082cf70d676b6551e16e023387c642f3fd1959c3478c8d": "037 需求母體",
    "6af7bfd314a28b3925636b642dc80c87918b50136fb8fa264563b3dc117eb592": "SYS2 CFTS009",
    "f318b14623fcbf972c6d0428408a6cf5297cbe7d0a710c7fafb8a0e05f538e61": "SYS2 CFTS010",
    "eb3eb0861363fc1d85c6127564355a7414ecd37eac0800336bb0d2e9b2911a26": "CFTS009 規格本文",
    "47c402a01b1a2e3a537797843b968ad621fc2bbc6f7a416a33dfe490247ea505": "CFTS010 規格本文",
    "cb6bf7d81030abc8ce47a444b4cf90b6bf527816fde7887b9e9b872c22338ae4": "SYS3 SYSAD",
}

# §C rule 4：037 Source Requirement ID 之兩域 token（區分大小寫）
PM_RE = re.compile(r"Sys-RA-PM-\d{4}")
PD_RE = re.compile(r"Sys-RA-PD[_-]\d+")

# §E Layer 3 章節 → Test Set 對應
TEST_SETS = {
    "Power State": 64,
    "Startup Display": 24,
    "Branding and Theme": 16,
    "Timeout Settings": 7,
    "Power Down": 3,
}


def find(pattern: str) -> Path:
    return next(f for f in IN.iterdir() if pattern in f.name)


def gate0() -> bool:
    """G0 素材身分前置閘：七份原始檔 SHA256 全數登記。"""
    seen = {}
    for f in sorted(IN.iterdir()):
        if f.name.startswith("."):
            continue
        seen[hashlib.sha256(f.read_bytes()).hexdigest()] = f.name
    missing = set(LEDGER) - set(seen)
    extra = set(seen) - set(LEDGER)
    print(f"G0 素材身分: {len(set(seen) & set(LEDGER))} / {len(LEDGER)}")
    for h in missing:
        print(f"   缺: {LEDGER[h]} ({h[:16]})")
    for h in extra:
        print(f"   多: {seen[h]} ({h[:16]})")
    return not missing and not extra


def leaves() -> list[tuple[str, str, str]]:
    """037 SWE1 Requirements r8-r145 之 leaf：(id, source_req_id, categorization)。"""
    wb = openpyxl.load_workbook(find("FSM-037"), data_only=True, read_only=True)
    ws = wb["SWE1 Requirements"]
    rows = [
        r for r in ws.iter_rows(min_row=8, max_row=145, values_only=True)
        if r[0] and str(r[0]).strip()
    ]
    wb.close()
    return [(str(r[0]).strip(), str(r[1] or ""), str(r[5] or "").strip()) for r in rows]


def sys2_map(path: Path, last_row: int) -> dict[str, list[str]]:
    """SYS2 Basic Report：Sys-RA-Feature-ID → [Polarion item id]。"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Basic Report"]
    out = {}
    for r in ws.iter_rows(min_row=2, max_row=last_row, values_only=True):
        key = str(r[1] or "").strip()
        if key:
            out[key] = re.findall(r"\d{6,8}", str(r[4] or ""))
    wb.close()
    return out


def bucket(domain: str, num: str) -> str:
    """§E Layer 3 章節號 → Test Set 名稱。無對應者回傳 '未歸類 …'。"""
    if domain == "010":
        return "Power Down" if num.startswith(("1.7.1", "1.7.2")) else f"未歸類 010 §{num}"
    if num.startswith("1.6.2.1."):
        parts = num.split(".")
        k = int(parts[4]) if len(parts) > 4 else 0
        if 1 <= k <= 15:
            return "Power State"
        if k == 16:
            return "Startup Display"
        return f"未歸類 009 §{num}"
    prefixes = [
        ("1.3.5", "Startup Display"), ("1.9.8", "Startup Display"),
        ("1.9.9", "Startup Display"), ("1.9.10", "Startup Display"),
        ("1.9.15", "Branding and Theme"), ("1.9.16", "Branding and Theme"),
        ("1.9.17", "Branding and Theme"),
        ("1.6.3", "Timeout Settings"), ("1.6.4", "Timeout Settings"),
        ("1.6.7", "Timeout Settings"),
        ("1.7.1", "Power State"), ("1.8.1", "Power State"),
        ("1.9.3", "Power State"), ("1.9.4", "Power State"),
        ("1.9.5", "Power State"), ("1.9.12", "Power State"),
    ]
    for prefix, name in prefixes:
        if num == prefix or num.startswith(prefix + "."):
            return name
    return f"未歸類 009 §{num}"


def main() -> None:
    if not gate0():
        raise SystemExit("G0 未通過，依 R-P14(a) 停止，不執行亦不回報 G1 以後")

    rows = leaves()
    ids = [x[0] for x in rows]
    nums = sorted(int(m.group(1)) for m in (re.match(r"SWE-PM-(\d+)$", i) for i in ids) if m)
    print(f"G1 leaf 數: {len(rows)} | 連續無斷點: {nums == list(range(1, len(nums) + 1))}")
    print(f"G2 Categorization: {dict(Counter(x[2] for x in rows))}")

    # G4 / G5：域歸屬純由 037 Source Requirement ID 判定
    dom = {"009": [], "010": [], "none": [], "both": []}
    for lid, src, _ in rows:
        a, b = bool(PM_RE.search(src)), bool(PD_RE.search(src))
        dom["both" if a and b else "009" if a else "010" if b else "none"].append(lid)
    print(f"G4 域分布: 009={len(dom['009'])} 010={len(dom['010'])} "
          f"皆無={len(dom['none'])} 兩者皆有={len(dom['both'])}")
    print(f"G5 需 CFTS010 之 leaf: {dom['010']}")

    # G3 / G5b：完整錨點鏈
    chapters = json.loads((DATA / "item_to_chapter.json").read_text(encoding="utf-8"))
    s9 = sys2_map(find("SYS2_CFTS_009"), 339)
    s10 = sys2_map(find("SYS2_CFTS_010"), 74)
    resolved, failed = {}, []
    for lid, src, _ in rows:
        hits = []
        for tok in PM_RE.findall(src):
            hits += [("009", *chapters["cfts009"][i]) for i in s9.get(tok, []) if i in chapters["cfts009"]]
        for tok in PD_RE.findall(src):
            hits += [("010", *chapters["cfts010"][i]) for i in s10.get(tok, []) if i in chapters["cfts010"]]
        (resolved.setdefault(lid, hits) if hits else failed.append(lid))
    print(f"G3 leaf → CFTS 章節解析成功: {len(resolved)} / {len(rows)} | 失敗: {failed}")
    for lid in dom["010"]:
        print(f"G5b {lid}: {sorted({(d, n) for d, n, _ in resolved.get(lid, [])})}")

    # §E 重算：每 leaf 只計主章節（出現最多者，同數取章節號最深）
    main_ch = {}
    for lid, hits in resolved.items():
        c = Counter((d, n) for d, n, _ in hits)
        main_ch[lid] = max(
            c.items(),
            key=lambda kv: (kv[1], tuple(-int(x) for x in kv[0][1].split("."))),
        )[0]
    (DATA / "leaf_main_chapter.json").write_text(
        json.dumps({k: list(v) for k, v in main_ch.items()}, ensure_ascii=False, indent=1),
        encoding="utf-8",
    )

    counts = Counter(bucket(d, n) for d, n in main_ch.values())
    print("\n§E 重算:")
    for name, expected in TEST_SETS.items():
        got = counts.get(name, 0)
        print(f"  {name:20} 實測 {got:>3}  §E {expected:>3}  "
              f"{'PASS' if got == expected else '**MISMATCH**'}")
    for name, n in sorted(counts.items()):
        if name.startswith("未歸類"):
            who = [l for l, (d, c) in main_ch.items() if bucket(d, c) == name]
            print(f"  {name}: {n}  {who}")
    print(f"  合計 {sum(counts.values())}")


if __name__ == "__main__":
    main()
