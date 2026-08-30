"""55 包 B1 —— 基底確認與缺陷家族重量（R-DD26 verify before act）。

附件 `55_review_findings.md` 之列號與列數全部量自 `delivered/pm_29.xlsx`
（389 條 / 390 列，27 包交付副本）。現行 corpus 為
`features/power/generated/batch_00{1..7}.json`（283 條），二者相差兩代。
本腳本以同一組偵測器同時掃描兩者，產出
`features/power/data/findings_remeasure_55.tsv`。

偵測器先在 pm_29 上校準（其計數須逼近附件 §2 之列數），
再套用於現行 corpus；差額即 51–54 包之間形態變化之量。

用法：
    python features/power/scripts/remeasure_55.py
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
POWER = ROOT / "features/power"
PM29 = POWER / "delivered/pm_29.xlsx"
SHEET = "Test Case Specification&Result"
OUT = POWER / "data/findings_remeasure_55.tsv"

# 附件 §2 之列數（量自 pm_29），作為偵測器之校準標的
ATTACH = {
    "A_lower": 290, "A_upper": 290, "A1": 38, "B": 226, "C1": 69, "C2": 40,
    "B_shape": 226, "C3": 66, "D": 25, "E": 16, "F": 12, "G": 15, "H": 100, "I": 15,
}

# R-P353 白名單四類之判別特徵
SIGNAL_RE = re.compile(r"\$[A-Za-z0-9_]+\.[A-Za-z0-9_]+\$")
UI_QUOTED_RE = re.compile(r'"[^"]+"')
AUDIO_RE = re.compile(r"\bAUD_LVL\b|source indicator|speaker", re.I)
LOG_RE = re.compile(r"\blog\b|\btrace\b|counter", re.I)

# 家族偵測用之片語
READ_RE = re.compile(r"^\s*\d*\.?\s*(?:Read|Check that)\b(.*)$", re.I)
HU_STATE_PRE_RE = re.compile(r"The HU is in\s+(.+?)\s+(?:state|mode)\b", re.I)
# B_shape：形態無關 —— 任何以電源狀態作前置之句（現行 corpus 已改用
# `TLM_Status.Info and $Telematic_Power$ read "<State>"` / `The TLM is in <X> status`）
STATES = ("Full-Operation", "FULL OPERATION", "Full_Operation", "Timed", "Idle",
          "IDLE", "Standby", "Partial_Operation", "Partial Operation", "Sleep",
          "Bench", "INIT", "operative")
B_SHAPE_RE = re.compile(
    r"(?:The (?:HU|TLM) is in|TLM_Status\.Info|\$Telematic_Power\$|PowerSts_Telematic)",
    re.I)
FUNCTIONALITY_RE = re.compile(r"functionality is (?:not )?available", re.I)
ABSTRACT_VERB_RE = re.compile(
    r"^\s*\d*\.?\s*(?:Let\b|Bring the HU|Attempt\b|Issue a Network Sleep"
    r"|Apply a manual time adjustment)", re.I | re.M)
SET_INTERNAL_RE = re.compile(r"Set\s+[A-Za-z0-9_]+\.(?:Info|Req)\s+to\b")
PRE_INTERNAL_RE = re.compile(r"[A-Za-z0-9_]+\.(?:Info|Req)\b|RemStartFail")
SPEC_LEAK_RE = re.compile(r"\bproper\b|\bas defined\b|\bnormal\w*\b", re.I)
STEP_CTRL = (
    "The previous internal state was", "before the disconnection",
    "The boot of the HU is not ended", "has not yet been shown",
    "has already played the startup sound", "has already ended",
)
HU_MODE_READ_RE = re.compile(r"Read the HU (?:mode|state)", re.I)
OBJECTID_RE = re.compile(r"\b49\d{5}\b")
NON_OBSERVABLE = (
    "functionality", "reaction", "behavior", "network state", "HU mode",
    "screen sequence", "main CPU", "CAN micro", "functionality state",
)


def whitelisted(fragment: str) -> bool:
    """R-P353 四類白名單：訊號 / 具名 UI（引號）/ 可量測音訊 / log。"""
    return bool(
        SIGNAL_RE.search(fragment)
        or UI_QUOTED_RE.search(fragment)
        or AUDIO_RE.search(fragment)
        or LOG_RE.search(fragment)
    )


AND_CHECK_RE = re.compile(r"\band check that\b", re.I)


def family_a_lower(tc: dict) -> bool:
    """A 下界：以 Read / Check that 起首之句，其受詞非白名單四類。"""
    for field in ("test_procedure", "expected_result"):
        for line in (tc[field] or "").splitlines():
            m = READ_RE.match(line)
            if m and not whitelisted(m.group(1)):
                return True
    return False


def family_a_upper(tc: dict) -> bool:
    """A 上界：併計「<動作> and check that …」之非白名單句。"""
    if family_a_lower(tc):
        return True
    for field in ("test_procedure", "expected_result"):
        for line in (tc[field] or "").splitlines():
            if AND_CHECK_RE.search(line) and not whitelisted(line):
                return True
    return False


DETECTORS = {
    "A_lower": family_a_lower,
    "A_upper": family_a_upper,
    "A1": lambda tc: bool(
        FUNCTIONALITY_RE.search(tc["test_procedure"] or "")
        or FUNCTIONALITY_RE.search(tc["expected_result"] or "")
    ),
    "B": lambda tc: bool(HU_STATE_PRE_RE.search(tc["pre_conditions"] or "")),
    "B_shape": lambda tc: any(
        B_SHAPE_RE.search(ln) and any(st in ln for st in STATES)
        for ln in (tc["pre_conditions"] or "").splitlines()),
    "C1": lambda tc: bool(ABSTRACT_VERB_RE.search(tc["test_procedure"] or "")),
    "C2": lambda tc: bool(SET_INTERNAL_RE.search(tc["test_procedure"] or "")),
    "C3": lambda tc: bool(PRE_INTERNAL_RE.search(tc["pre_conditions"] or "")),
    "D": lambda tc: bool(
        SPEC_LEAK_RE.search(tc["test_procedure"] or "")
        or SPEC_LEAK_RE.search(tc["expected_result"] or "")
    ),
    "E": lambda tc: any(s in (tc["pre_conditions"] or "") for s in STEP_CTRL),
    "G": lambda tc: bool(
        HU_MODE_READ_RE.search(tc["test_procedure"] or "")
        or HU_MODE_READ_RE.search(tc["expected_result"] or "")
    ),
    "H": lambda tc: len(OBJECTID_RE.findall(tc["specification_reference"] or "")) >= 6,
    "I": lambda tc: "PENDING" in " ".join(
        (tc[f] or "") for f in ("test_item", "pre_conditions", "input_test_data",
                                "test_procedure", "expected_result", "remarks")
    ),
}

FOUR_COLS = ("test_item", "pre_conditions", "test_procedure", "expected_result")


def dup_pairs(tcs: list[dict]) -> list[tuple[str, str]]:
    """F：四欄逐字相同之重複對。"""
    seen: dict[tuple, str] = {}
    pairs = []
    for tc in tcs:
        key = tuple((tc[c] or "").strip() for c in FOUR_COLS)
        if key in seen:
            pairs.append((seen[key], tc["row"]))
        else:
            seen[key] = tc["row"]
    return pairs


def load_pm29() -> list[dict]:
    ws = openpyxl.load_workbook(PM29, data_only=True)[SHEET]
    cols = {"req_id": 4, "tc_id": 6, "test_item": 9, "pre_conditions": 10,
            "input_test_data": 11, "test_procedure": 12, "expected_result": 13,
            "specification_reference": 14, "remarks": 34}
    tcs = []
    for r in range(10, ws.max_row + 1):
        row = {k: ws.cell(r, c).value for k, c in cols.items()}
        if all(v in (None, "") for v in row.values()):
            continue
        row["row"] = str(ws.cell(r, 2).value or r)
        tcs.append({k: (v if isinstance(v, str) else ("" if v is None else str(v)))
                    for k, v in row.items()})
    return tcs


def load_current() -> list[dict]:
    tcs = []
    for path in sorted((POWER / "generated").glob("batch_*.json")):
        for tc in json.loads(path.read_text())["tcs"]:
            tc = dict(tc)
            tc.setdefault("remarks", "")
            tc["row"] = tc["tc_id"]
            tcs.append(tc)
    return tcs


def main() -> None:
    pm29, cur = load_pm29(), load_current()
    print(f"pm_29        : {len(pm29)} 列（{PM29.relative_to(ROOT)}）")
    print(f"現行 corpus  : {len(cur)} 條（features/power/generated/batch_*.json）")

    lines = ["family\tpm_29_rows\tcurrent_rows\tdelta\tattachment\tdetector_gap"]
    for fam, fn in DETECTORS.items():
        a = [t["row"] for t in pm29 if fn(t)]
        b = [t["row"] for t in cur if fn(t)]
        gap = len(a) - ATTACH[fam]
        lines.append(f"{fam}\t{len(a)}\t{len(b)}\t{len(b) - len(a)}\t"
                     f"{ATTACH[fam]}\t{gap:+d}")
    fa, fb = dup_pairs(pm29), dup_pairs(cur)
    lines.append(f"F\t{len(fa)}\t{len(fb)}\t{len(fb) - len(fa)}\t{ATTACH['F']}\t"
                 f"{len(fa) - ATTACH['F']:+d}")

    # R-P357 之四欄鍵在現行 corpus 上會誤判：其以
    # `Input Test Data` 帶參數之型別，四欄逐字相同而測項不同。
    idx = {t["row"]: t for t in cur}
    real = [(a, b) for a, b in fb
            if (idx[a]["input_test_data"] or "").strip()
            == (idx[b]["input_test_data"] or "").strip()]
    itd = [(a, b) for a, b in fb if (a, b) not in real]
    same_req = [(a, b) for a, b in real if idx[a]["req_id"] == idx[b]["req_id"]]
    lines.append(f"F_true\t—\t{len(real)}\t—\t—\t含 ITD 亦同者")
    lines.append(f"F_itd_falsepos\t—\t{len(itd)}\t—\t—\tITD 不同，R-P357 四欄鍵誤判")
    lines.append(f"F_deletable\t—\t{len(same_req)}\t—\t—\tR-P357(a) 同 req_id 可刪者")

    OUT.write_text("\n".join(lines) + "\n")
    print("\n".join(lines))
    print(f"\n→ {OUT.relative_to(ROOT)}")
    print("\npm_29 重複對:", fa)
    print(f"現行 corpus 重複對 {len(fb)} = 真重複 {len(real)} ＋ ITD 誤判 {len(itd)}；"
          f"R-P357(a) 可刪者 {len(same_req)}")


if __name__ == "__main__":
    main()
