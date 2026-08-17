"""B3 —— 首批 dry-run 寫回（R-P114）。

**唯一目的**：使 G66 / G71 / G72 三閘取得**真實**證據，並首度實測
R-G3 所載之「openpyxl `save()` 破壞 x14 dataValidation」缺陷。

嚴格遵守 16 §I：

  - 僅對 `features/power/sandbox/` 之副本為之；
    **客戶樹與 `inputs/` 之原始檔一律不觸碰**（唯讀來源，位元組複製）
  - 寫回路徑為 `backend/xlsx_surgical.py` 之 `surgical_save()`；
    **本檔全域無 `Workbook.save()` 之呼叫**
  - dry-run 之 tc_id 為臨時編號 001–010（R-P113(b)/(e)），
    **不代表最終指派**；011 起之補測條依 R-P114(e) 不納入本次 dry-run
  - 沙箱不入版控（`.gitignore`）

用法：
    python features/power/scripts/dryrun_write_back.py
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[3]
FEATURE = ROOT / "features/power"
SANDBOX = FEATURE / "sandbox"
DATA = FEATURE / "data"

NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
HEADER_ROW = 9
FIRST_DATA_ROW = 10
DRYRUN_MAX_TC = 10          # R-P114(e)

# feature.yaml 之 workbook.columns 鍵 → TC 之欄位名（None = 常數或留白）
FIELD_OF = {
    "req_id": "req_id", "tc_id": "tc_id", "test_group": "test_group",
    "test_set": "test_set", "test_item": "test_item",
    "pre_conditions": "pre_conditions", "input_test_data": "input_test_data",
    "test_procedure": "test_procedure", "expected_result": "expected_result",
    "spec_reference": "specification_reference", "priority": "priority",
    "design_method": "design_method", "functional_safety": "functional_safety",
}
# profile §3.6 / §3.8：estimated_time 與車型欄留白（R-P54 / R-P81）
BLANK_COLUMNS = ["estimated_time"]


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load_cfg() -> dict:
    return yaml.safe_load((FEATURE / "feature.yaml").read_text(encoding="utf-8"))


def row_values(tc: dict, cfg: dict, row: int) -> dict[str, object]:
    """回傳 {欄字母: 值}。B 欄為序號（Power 之範本無編號公式 —— R-P90）。"""
    cols = cfg["workbook"]["columns"]
    wb_cfg = cfg["write_back"]
    out: dict[str, object] = {"B": row - HEADER_ROW}
    for key, letter in cols.items():
        if key in BLANK_COLUMNS:
            continue
        if key == "tc_ref_id":
            out[letter] = wb_cfg["tc_ref_id_value"]
        elif key == "author":
            out[letter] = wb_cfg["author_value"]
        elif key == "remarks":
            out[letter] = tc.get("remarks", "") or ""
        elif key in FIELD_OF:
            out[letter] = tc.get(FIELD_OF[key], "")
    if not cfg["write_back"].get("fill_test_group_set"):
        out.pop(cols["test_group"], None)
        out.pop(cols["test_set"], None)
    return out


def write_rows(src: Path, out: Path, tcs: list[dict], cfg: dict,
               *, shift: int = 0, blank_b: bool = False) -> dict:
    """以 surgical_save 寫入。`shift` / `blank_b` 供「刻意寫錯」之失敗證明。"""
    sys.path.insert(0, str(ROOT / "backend"))
    from xlsx_surgical import col_to_idx, idx_to_col, surgical_save

    wb = openpyxl.load_workbook(src)
    ws = wb[cfg["workbook"]["sheet"]]
    for i, tc in enumerate(tcs):
        row = FIRST_DATA_ROW + i
        for letter, value in row_values(tc, cfg, row).items():
            if blank_b and letter == "B":
                continue
            target = idx_to_col(col_to_idx(letter) + shift) if shift else letter
            ws[f"{target}{row}"] = value
    return surgical_save(wb, src, out)      # **不呼叫 wb.save()**


def dv_snapshot(path: Path) -> list[tuple[str, str, str]]:
    """(ns, sqref, type) —— 逐條 DV，含 x14。以 zipfile 直讀，不經 openpyxl。"""
    import xml.etree.ElementTree as ET
    X14 = "{http://schemas.microsoft.com/office/spreadsheetml/2009/9/main}"
    XM = "{http://schemas.microsoft.com/office/excel/2006/main}"
    rows = []
    with zipfile.ZipFile(path) as z:
        for member in sorted(m for m in z.namelist()
                             if m.startswith("xl/worksheets/sheet")):
            root = ET.fromstring(z.read(member))
            for dv in root.iter(NS_MAIN + "dataValidation"):
                rows.append(("main", dv.get("sqref") or "", dv.get("type") or ""))
            for dv in root.iter(X14 + "dataValidation"):
                sq = dv.find(XM + "sqref")
                rows.append(("x14", (sq.text or "") if sq is not None else "",
                             dv.get("type") or ""))
    return rows


def structure_snapshot(path: Path) -> dict:
    import xml.etree.ElementTree as ET
    with zipfile.ZipFile(path) as z:
        members = sorted(z.namelist())
        wbx = ET.fromstring(z.read("xl/workbook.xml"))
        sheets = [(s.get("name"), s.get("state") or "visible")
                  for s in wbx.iter(NS_MAIN + "sheet")]
        merges, cf = [], []
        for m in members:
            if m.startswith("xl/worksheets/sheet"):
                root = ET.fromstring(z.read(m))
                merges += [(m, x.get("ref")) for x in root.iter(NS_MAIN + "mergeCell")]
                cf += [(m, x.get("sqref")) for x in root.iter(NS_MAIN + "conditionalFormatting")]
    return {"members": members, "sheets": sheets, "merges": merges, "cf": cf}


def measure(path: Path, tcs: list[dict], cfg: dict) -> dict:
    """G66 / G71 / G72 之**工作簿層**實測（唯讀）。"""
    cols = cfg["workbook"]["columns"]
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[cfg["workbook"]["sheet"]]
    grid = {}
    for row in ws.iter_rows(min_row=HEADER_ROW,
                            max_row=FIRST_DATA_ROW + len(tcs) - 1):
        for c in row:
            if c.value not in (None, ""):
                grid[c.coordinate] = c.value
    wb.close()

    b_filled = sum(1 for r in range(FIRST_DATA_ROW, FIRST_DATA_ROW + len(tcs))
                   if f"B{r}" in grid)
    g66 = {"b_filled": b_filled, "rows": len(tcs), "pass": b_filled == len(tcs)}

    # G71 —— 逐欄回報實際落點：該欄之 r9 標頭與所寫之值是否對得上
    g71 = []
    for key, letter in cols.items():
        header = str(grid.get(f"{letter}{HEADER_ROW}", "")).replace("\n", " / ")
        cell = f"{letter}{FIRST_DATA_ROW}"
        written = grid.get(cell, "")
        # `remarks` 依 profile 為選填 —— 本批十條皆無備註，故其空白為預期值，
        # 非欄位錯位。首次執行時本行缺此判別，致 AI10 被誤報 FAIL（見上繳 §一）。
        expect_blank = (key in BLANK_COLUMNS or key.startswith("vehicle")
                        or key == "remarks")
        ok = (written == "") if expect_blank else (written != "")
        g71.append({"key": key, "col": letter, "cell": cell, "header": header,
                    "written": str(written)[:44], "expect_blank": expect_blank,
                    "pass": ok})

    # G72 —— profile §2 test_group / §3.3 design_method / §3.4 spec 檔名 / §3.7
    g72 = []
    for i in range(len(tcs)):
        r = FIRST_DATA_ROW + i
        g72.append({
            "row": r,
            "test_group": grid.get(f"{cols['test_group']}{r}"),
            "design_method": grid.get(f"{cols['design_method']}{r}"),
            "spec_reference": str(grid.get(f"{cols['spec_reference']}{r}", ""))[:60],
            "functional_safety": grid.get(f"{cols['functional_safety']}{r}"),
        })
    return {"g66": g66, "g71": g71, "g72": g72, "grid_cells": len(grid)}


def main() -> None:
    cfg = load_cfg()
    src = FEATURE / cfg["paths"]["workbook"]
    SANDBOX.mkdir(parents=True, exist_ok=True)
    (SANDBOX / ".gitignore").write_text("*\n", encoding="utf-8")

    base = SANDBOX / "base.xlsx"
    shutil.copy(src, base)                       # 位元組複製，來源唯讀
    assert sha256(base) == sha256(src), "沙箱副本與來源不同 —— 複製有誤"

    batch = json.loads((FEATURE / "generated/batch_001_power_down.json")
                       .read_text(encoding="utf-8"))
    tcs = sorted(batch["tcs"], key=lambda t: int(t["tc_id"].split("-")[-1]))
    tcs = [t for t in tcs
           if int(t["tc_id"].split("-")[-1]) <= DRYRUN_MAX_TC]      # R-P114(e)

    dv_before = dv_snapshot(base)
    st_before = structure_snapshot(base)

    out = SANDBOX / "dryrun.xlsx"
    report = write_rows(base, out, tcs, cfg)

    dv_after = dv_snapshot(out)
    st_after = structure_snapshot(out)
    m = measure(out, tcs, cfg)

    # 失敗證明（R-P114 之驗證條件：須確認三閘確實可能失敗）
    fail_b = write_rows(base, SANDBOX / "fail_b.xlsx", tcs, cfg, blank_b=True)
    m_fail_b = measure(SANDBOX / "fail_b.xlsx", tcs, cfg)
    fail_shift = write_rows(base, SANDBOX / "fail_shift.xlsx", tcs, cfg, shift=1)
    m_fail_shift = measure(SANDBOX / "fail_shift.xlsx", tcs, cfg)

    result = {
        "src_sha256": sha256(src), "base_sha256": sha256(base),
        "out_sha256": sha256(out),
        "src_untouched": sha256(src) == sha256(base),
        "surgical_report": report,
        "dv_before": dv_before, "dv_after": dv_after,
        "dv_identical": dv_before == dv_after,
        "structure_identical": {
            "members": st_before["members"] == st_after["members"],
            "sheets": st_before["sheets"] == st_after["sheets"],
            "merges": st_before["merges"] == st_after["merges"],
            "cf": st_before["cf"] == st_after["cf"],
        },
        "measure": m,
        "fail_proof": {
            "blank_b": m_fail_b["g66"],
            "shift_one": [x for x in m_fail_shift["g71"] if not x["pass"]],
            "shift_one_g72": m_fail_shift["g72"][0],
        },
        "tc_count": len(tcs),
    }
    (DATA / "b3_dryrun.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"來源 {src.name}")
    print(f"  src SHA256  {result['src_sha256']}")
    print(f"  來源未被觸碰：{result['src_untouched']}")
    print(f"  dry-run 寫入 {len(tcs)} 條（R-P114(e)：僅 001–010）")
    print(f"  surgical_save：patched {report['members_patched']}，"
          f"differing {report['differing']}")
    print(f"\nG86 DV 存活：{len(dv_before)} 條 → {len(dv_after)} 條，"
          f"逐字相同 = **{result['dv_identical']}**")
    for a, b in zip(dv_before, dv_after):
        print(f"   {a[0]:5} {a[1]:24} {a[2]:8}  →  {b[0]:5} {b[1]:24} {b[2]:8}"
              f"  {'同' if a == b else '**異**'}")
    print(f"\nG87 結構：{result['structure_identical']}")
    print(f"G66 {m['g66']}")
    print("G71 逐欄落點：")
    for x in m["g71"]:
        print(f"   {x['key']:18} {x['cell']:5} 標頭「{x['header'][:34]}」"
              f" 值「{x['written'][:30]}」 {'PASS' if x['pass'] else '**FAIL**'}")
    print(f"G72 首列：{m['g72'][0]}")
    print(f"\n失敗證明 —— B 欄留空：{m_fail_b['g66']}")
    print(f"失敗證明 —— 欄位右移一格：G71 FAIL {len(result['fail_proof']['shift_one'])} 欄；"
          f"G72 首列 {result['fail_proof']['shift_one_g72']}")


if __name__ == "__main__":
    main()
