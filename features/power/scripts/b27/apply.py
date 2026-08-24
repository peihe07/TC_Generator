#!/usr/bin/env python3
"""27 包執行層：`test_item` 括號兩案修正套用。

26 包上繳 §八-1／§八-2 之 Pei 裁定。基底 `features/power/sandbox/b26/pm_26.xlsx`
（sha256 `0181f6de…`）。

本包**不動列數**：面向切分、setup 判定、消歧候選一律沿 26 包，插列數與
Test Case ID 皆不變。故無須插列段，`surgical_save` 單段即足 ——
寫入範圍限 `test_item`（I 欄）一欄，其餘各欄零變動。
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))

from xlsx_surgical import (                                 # noqa: E402
    StructureError, build_shift, sheet_members, surgical_save,
)

HERE = Path(__file__).resolve().parent
PLAN = HERE / "plan.json"
BASE = ROOT / "features/power/sandbox/b26/pm_26.xlsx"
BASE_SHA = "0181f6de79097db6891a59e7e4e54bcee8f89036951d55951ca753ae6ab29fc1"
SANDBOX = ROOT / "features/power/sandbox/b27"
OUT = SANDBOX / "pm_27.xlsx"
FIELD = "I"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def x14_dropdowns(path: Path, member: str) -> list[str]:
    with zipfile.ZipFile(path) as z:
        xml = z.read(member).decode("utf-8")
    return re.findall(r"<xm:sqref>([^<]*)</xm:sqref>", xml)


def main() -> None:
    plan = json.loads(PLAN.read_text(encoding="utf-8"))
    sheet = plan["sheet"]
    if sha256(BASE) != BASE_SHA:
        raise StructureError(f"基底 sha256 不符：{sha256(BASE)[:20]}")

    shift = build_shift({int(k): v for k, v in plan["insertions"].items()})
    member = sheet_members(BASE)[sheet]
    dv_before = x14_dropdowns(BASE, member)

    SANDBOX.mkdir(parents=True, exist_ok=True)
    work = SANDBOX / ".work.xlsx"
    shutil.copyfile(BASE, work)
    wb = openpyxl.load_workbook(work)
    ws = wb[sheet]

    changed: list[int] = []
    for split in plan["splits"]:
        anchor = shift(split["src_row"])
        for offset, variant in enumerate(split["variants"]):
            row = anchor + offset
            if ws[f"{FIELD}{row}"].value != variant[FIELD]:
                ws[f"{FIELD}{row}"] = variant[FIELD]
                changed.append(row)

    report = surgical_save(wb, work, OUT)
    work.unlink()

    dv_after = x14_dropdowns(OUT, member)
    if dv_before != dv_after:
        raise StructureError(f"x14 讀回不符：{dv_before} → {dv_after}")

    print(f"{FIELD} 欄改寫 {len(changed)} 列：{changed}")
    print(f"x14 讀回 {dv_before} → {dv_after}")
    print("zip 成員", report["members"], "| 差異成員", report["differing"])
    print("out sha256:", sha256(OUT)[:20])


if __name__ == "__main__":
    main()
