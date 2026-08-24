#!/usr/bin/env python3
"""27 包驗收：逐項對 `pm_27.xlsx` 覆核，任一項不達成即 exit 1。

本包只動 `test_item`，故驗收重心在「該動的動了、不該動的一格沒動」：
與 `pm_26` 逐格比對，證明相異僅限 I 欄之 17 列。
"""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))

import lint036                                              # noqa: E402
from xlsx_surgical import build_shift                       # noqa: E402

HERE = Path(__file__).resolve().parent
B19 = ROOT / "features/power/sandbox/b19/pm_19.xlsx"
PREV = ROOT / "features/power/sandbox/b26/pm_26.xlsx"
OUT = ROOT / "features/power/sandbox/b27/pm_27.xlsx"
FIRST, LAST_COL = 10, 34
RESTORED_SRC = 109          # §八-1：面向數為 1，test_item 逐字還原
NUM = re.compile(r"^\s*\d+[.)]")
DENUM = re.compile(r"^\s*\d+[.)]\s*")
OBS_HEAD = re.compile(r"^Read\b")
OBS_MID = re.compile(r"check that", re.I)
A_ROWS = {11, 12, 23, 179, 180}


def bodies(text) -> list[str]:
    return [DENUM.sub("", ln).strip()
            for ln in str(text or "").split("\n") if NUM.match(ln)]


def has_observation(body: str) -> bool:
    return bool(OBS_HEAD.match(body) or OBS_MID.search(body))


def lower_first(s: str) -> str:
    if not s:
        return s
    head = s.split(maxsplit=1)[0]
    if not (head[:1].isupper() and head[1:].isalpha() and head[1:].islower()):
        return s
    return s[:1].lower() + s[1:]


def main() -> int:
    plan = json.loads((HERE / "plan.json").read_text(encoding="utf-8"))
    sheet = plan["sheet"]
    src = openpyxl.load_workbook(B19)[sheet]
    prev = openpyxl.load_workbook(PREV)[sheet]
    ws = openpyxl.load_workbook(OUT)[sheet]
    shift = build_shift({int(k): v for k, v in plan["insertions"].items()})
    setup_of = {a["row"]: a["setup"] for a in plan["audit_b"]}
    fails: list[str] = []

    def check(ok: bool, label: str, detail: str = "") -> None:
        print(f"  {'達成' if ok else '未達成'}  {label}"
              + (f" — {detail}" if detail else ""))
        if not ok:
            fails.append(label)

    # --- 與 pm_26 之相異範圍 ------------------------------------------------
    check(ws.max_row == prev.max_row, "列數與 pm_26 相同",
          f"{prev.max_row} → {ws.max_row}")
    diff_rows, off_field = set(), []
    for row in range(1, ws.max_row + 1):
        for col in range(1, LAST_COL + 1):
            a = ws.cell(row, col).value or ""
            b = prev.cell(row, col).value or ""
            if a == b:
                continue
            letter = openpyxl.utils.get_column_letter(col)
            diff_rows.add(row)
            if letter != "I":
                off_field.append((row, letter))
    check(not off_field, "相異僅限 test_item（I）欄", f"例外 {off_field[:5]}")
    check(len(diff_rows) == 17, "相異列恰 17 列",
          f"{len(diff_rows)} 列 {sorted(diff_rows)}")

    # --- §八-1：row 109 逐字還原 -------------------------------------------
    restored = shift(RESTORED_SRC)
    check(ws[f"I{restored}"].value == src[f"I{RESTORED_SRC}"].value,
          f"§八-1 row {restored} test_item 與 b19 row {RESTORED_SRC} 逐字同")
    same = all((ws.cell(restored, c).value or "")
               == (src.cell(RESTORED_SRC, c).value or "")
               for c in range(1, LAST_COL + 1) if c not in (2, 6))
    check(same, f"§八-1 row {restored} 除 No.#／TC ID 外全欄與原列逐字同")

    # --- §八-2：16 列縮併列括號為形態 B --------------------------------------
    form_b, other_bad = [], []
    for split in plan["splits"]:
        s_row = split["src_row"]
        if s_row in A_ROWS:
            continue
        setup = setup_of[s_row]
        anchor = shift(s_row)
        src_proc = bodies(src[f"L{s_row}"].value)
        src_er = bodies(src[f"M{s_row}"].value)
        for offset, variant in enumerate(split["variants"]):
            row = anchor + offset
            steps = bodies(variant["L"])[setup:]
            paren = "\n".join(lint036.paren_lines(
                str(ws[f"I{row}"].value or "")))
            if len(steps) <= 1 or s_row == RESTORED_SRC:
                # 非縮併列（及已還原之 row 109）括號須與 pm_26 逐字相同。
                if s_row != RESTORED_SRC and \
                        ws[f"I{row}"].value != prev[f"I{row}"].value:
                    other_bad.append(row)
                continue
            drives, obs = steps[:-1], steps[-1]
            er = src_er[src_proc.index(obs)]
            want = (f"({' / '.join(lower_first(d) for d in drives)} "
                    f"-> {er})")
            # 碰撞列會由既有消歧器加 `<前綴> — ` ，故形態 B 為裸式或帶前綴式。
            prefixed = paren.endswith(f" — {want[1:-1]})")
            form_b.append((row, paren == want or prefixed,
                           len(want.split()), prefixed))
    bad_b = [r for r, ok, _, _ in form_b if not ok]
    check(len(form_b) == 16, "§八-2 縮併列 16 列", f"{len(form_b)} 列")
    check(not bad_b, "§八-2 括號為形態 B（驅動步 -> ER，逐字取自原列）",
          f"例外 {bad_b}")
    check(not other_bad, "非縮併面向列括號與 pm_26 逐字相同",
          f"例外 {other_bad[:5]}")
    pref = [r for r, _, _, q in form_b if q]
    print(f"  （參考）縮併列中經消歧加前綴者 {len(pref)} 列 {pref}")
    over = [(r, n) for r, _, n, _ in form_b if n > 20]
    print(f"  （參考）縮併列括號逾 20 詞 {len(over)} 列 {over} —— 依裁定免上限")

    # --- 不變式：ID／No.#／E／觀察步／sibling --------------------------------
    ids = [str(ws[f"F{r}"].value or "").strip()
           for r in range(FIRST, ws.max_row + 1)]
    nums = [int(i.rsplit("-", 1)[1]) for i in ids if i]
    check(nums == list(range(1, len(nums) + 1)),
          "Test Case ID 連續無跳號", f"001–{max(nums):03d}／{len(nums)} 列")
    check([ws[f"B{r}"].value for r in range(FIRST, ws.max_row + 1)]
          == list(range(1, ws.max_row - FIRST + 2)), "No.# 連續重編")
    e_bad = [r for r in range(FIRST, ws.max_row + 1)
             if len(bodies(ws[f"L{r}"].value))
             != len(bodies(ws[f"M{r}"].value))]
    check(not e_bad, "proc↔er 編號數逐列相等", f"例外 {e_bad[:5]}")
    no_obs = [r for r in range(FIRST, ws.max_row + 1)
              if bodies(ws[f"L{r}"].value)
              and not any(has_observation(b)
                          for b in bodies(ws[f"L{r}"].value))]
    check(not no_obs, "全本無「無觀察步」之列", f"例外 {no_obs[:8]}")
    groups: dict[tuple[str, str], list[int]] = defaultdict(list)
    for row in range(FIRST, ws.max_row + 1):
        req = str(ws[f"D{row}"].value or "").strip()
        paren = "\n".join(lint036.paren_lines(str(ws[f"I{row}"].value or "")))
        if req and paren:
            groups[(req, paren)].append(row)
    dups = {k: v for k, v in groups.items() if len(v) > 1}
    check(not dups, "括號下半同源不逐字相同", f"{list(dups.values())[:3]}")

    print(f"\n{'全項達成' if not fails else f'未達成 {len(fails)} 項'}")
    return 0 if not fails else 1


if __name__ == "__main__":
    sys.exit(main())
