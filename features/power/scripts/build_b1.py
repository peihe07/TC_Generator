"""B1 — 跨多章節 leaf 清單（03 包 §B1，R-P15(b) 之裁定素材）。

依 R-P15(b)，本腳本**不做任何指派**，也不輸出建議歸屬。
它只把裁定所需的事實攤開：每個 token 的解析結果、相異章節集合、
以及 02 包當時所採規則（次數最多、同數取最深）指派過的結果 —— 後者
是歷史紀錄，不是建議。

錨點鏈採 §C 字面讀法（rule 3 之 item id 僅比對需求錨點）。

前置：features/power/data/item_to_chapter.json（由 extract_textlayer.py 產出）

用法：
    python features/power/scripts/build_b1.py
"""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
IN = ROOT / "features/power/inputs"
DATA = ROOT / "features/power/data"

PM_RE = re.compile(r"Sys-RA-PM-\d{4}")
PD_RE = re.compile(r"Sys-RA-PD[_-]\d+")

# §C 讀取座標（SYS2 CFTS009 已依 R-P18 訂正為 r2–r339）
SYS2_LAST = {"009": 339, "010": 74}


def find(pattern: str) -> Path:
    return next(f for f in IN.iterdir() if pattern in f.name)


def sys2_map(path: Path, last_row: int) -> dict[str, list[str]]:
    """SYS2 Basic Report：Sys-RA-Feature-ID → [Polarion item id]。"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Basic Report"]
    out = {}
    for r in ws.iter_rows(min_row=2, max_row=last_row, values_only=True):
        key = str(r[1] or "").strip()
        if key:
            out[key] = re.findall(r"\d{6,8}", str(r[4] or ""))
    wb.close()
    return out


def leaves() -> list[tuple[str, str, str, str]]:
    """037 SWE1 Requirements r8–r145：(leaf id, source req id, title, description)。"""
    wb = openpyxl.load_workbook(find("FSM-037"), data_only=True, read_only=True)
    ws = wb["SWE1 Requirements"]
    rows = [
        (str(r[0]).strip(), str(r[1] or ""), str(r[2] or "").strip(), str(r[3] or "").strip())
        for r in ws.iter_rows(min_row=8, max_row=145, values_only=True)
        if r[0] and str(r[0]).strip()
    ]
    wb.close()
    return rows


def main() -> None:
    chapters = json.loads((DATA / "item_to_chapter.json").read_text(encoding="utf-8"))
    sys2 = {
        "009": sys2_map(find("SYS2_CFTS_009"), SYS2_LAST["009"]),
        "010": sys2_map(find("SYS2_CFTS_010"), SYS2_LAST["010"]),
    }
    tables = {"009": chapters["cfts009"], "010": chapters["cfts010"]}

    out = [
        "# B1 — 跨多章節 leaf 清單（R-P15(b) 之裁定素材）\n",
        "\n> 依 R-P15(b)：本檔**不含任何建議歸屬**。指派為 Pei 之裁決。\n",
        "> 產生指令：`python features/power/scripts/build_b1.py`\n",
        "> 錨點鏈採 §C 字面讀法（rule 3 之 item id 僅比對需求錨點）。\n",
        "> 已驗：改採含章節錨點 id 之延伸讀法，11 個 leaf 之名單不變。\n",
    ]

    count = 0
    for leaf_id, src, title, desc in leaves():
        tokens = PM_RE.findall(src) + PD_RE.findall(src)
        rows = []
        chapter_hits = []
        for tok in tokens:
            domain = "009" if tok.startswith("Sys-RA-PM") else "010"
            items = sys2[domain].get(tok, [])
            for item in items:
                if item in tables[domain]:
                    num, chapter_title = tables[domain][item]
                    rows.append((tok, domain, item, f"§{num} — {chapter_title}"))
                    chapter_hits.append((domain, num))
                else:
                    rows.append((tok, domain, item, "**未解析**"))
            if not items:
                rows.append((tok, domain, "（SYS2 無此 Feature-ID 或欄內無 item）", "—"))

        distinct = sorted(set(chapter_hits))
        if len(distinct) < 2:
            continue

        count += 1
        freq = Counter(chapter_hits)
        legacy = max(
            freq.items(),
            key=lambda kv: (kv[1], tuple(-int(x) for x in kv[0][1].split("."))),
        )[0]

        out.append(f"\n---\n\n## {count}. `{leaf_id}`\n")
        out.append(f"\n**Requirement Title**：{title}\n")
        out.append(
            f"\n**Source Requirement ID 欄之完整 token 清單**（{len(tokens)} 個）：\n\n"
            f"```\n{src.strip()}\n```\n"
        )
        out.append(
            "\n**每個 token 之解析結果**：\n\n"
            "| token | 域 | Polarion item id | 解析到之 (章節號, 章節標題) |\n|---|---|---|---|\n"
        )
        for tok, domain, item, chapter in rows:
            out.append(f"| `{tok}` | CFTS{domain} | `{item}` | {chapter} |\n")
        out.append(f"\n**相異章節集合**（{len(distinct)} 個）：\n\n")
        for domain, num in distinct:
            out.append(f"- CFTS{domain} §{num}（出現 {freq[(domain, num)]} 次）\n")
        out.append(
            f"\n**02 包所採規則（次數最多、同數取最深）指派之結果**："
            f"CFTS{legacy[0]} §{legacy[1]}\n"
        )
        out.append(f"\n**Requirement Description 前 300 字元**：\n\n```\n{desc[:300]}\n```\n")

    path = DATA / "multi_chapter_leaves.md"
    path.write_text("".join(out), encoding="utf-8")
    print(f"wrote {path.relative_to(ROOT)} — {count} 個 leaf，{path.stat().st_size} bytes")


if __name__ == "__main__":
    main()
