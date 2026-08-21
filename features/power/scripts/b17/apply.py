#!/usr/bin/env python3
"""17 包寫回：覆核裁決之唯一內容項 —— §五 `PowerModeSts_Telematic` 全案改寫。

DBC 無 `PowerModeSts_Telematic`；Pei 裁定一律採
`$STATUS_TELEMATIC.PowerSts_Telematic$`（`PowerModeSts` 不使用）。
036 之四個作者欄中該名稱僅出現於 row 72（proc／er 各一行），
`test_item` 上半之出現屬 verbatim，不動。

其餘各節（§一/二/三/四/六/七）為追認或另立包，無內容改動。
`surgical_save` 唯一路徑；交付本唯讀不覆寫。
"""

from __future__ import annotations

import hashlib
import json
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(ROOT / "scripts"))

import lint036                                         # noqa: E402
from xlsx_surgical import surgical_save                # noqa: E402

SANDBOX = ROOT / "features/power/sandbox/b17"
SRC = ROOT / "features/power/sandbox/b16/pm_16.xlsx"
OUT = SANDBOX / "pm_17.xlsx"
COLUMN = {"pre": "J", "input": "K", "proc": "L", "er": "M"}


def sha256(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def main() -> None:
    edits = {int(k): v for k, v in
             json.loads((SANDBOX / "edits.json").read_text(encoding="utf-8")).items()}
    work = SANDBOX / ".work.xlsx"
    shutil.copyfile(SRC, work)
    assert sha256(work) == sha256(SRC)

    wb = openpyxl.load_workbook(work)
    ws = wb[next(s for s in wb.sheetnames
                 if s.startswith(lint036.TC_SHEET_PREFIX))]
    n = 0
    for row, fields in sorted(edits.items()):
        for key, value in fields.items():
            ws[f"{COLUMN[key]}{row}"] = value
            n += 1

    report = surgical_save(wb, work, OUT)
    work.unlink()

    left = [r for r in range(10, 294)
            if any("PowerModeSts" in (ws[f"{c}{r}"].value or "")
                   for c in COLUMN.values())]
    print(f"寫入 {n} 格／{len(edits)} 列 → {OUT.name}")
    print("四欄殘留 PowerModeSts 之列:", left)
    print("members:", report["members"], "| differing:", report["differing"])
    print("dv:", report["dv_counts"])
    print("out sha256:", sha256(OUT)[:20])


if __name__ == "__main__":
    main()
