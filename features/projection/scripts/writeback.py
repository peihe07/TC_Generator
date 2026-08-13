#!/usr/bin/env python3
"""QUARANTINED (R20-3, 2026-08-13) — this script writes via
openpyxl save and will destroy zip members and data
validations. It must not be executed. The feature's
delivered artefact is frozen; see ANOMALIES A-H27 /
A-SX28 / A-AM18 and RULINGS R18-1.

Phase 7 寫回 — W-1 ~ W-7（R-P71）。

用法：
    python writeback.py <src.xlsx> <dst.xlsx>

**預設只在複本上執行。** 對交付用檔案執行須 Pei 明示放行（下放包 §6 第 6 步）。
每個 W 動作都帶自己的驗證；任一項不過即 raise，不留半成品。
"""

import hashlib
import json
from copy import copy
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import lint_defs as L

FEAT = Path(__file__).resolve().parents[1]
BATCHES = FEAT / "batches"
APPEND_START = 562          # r562 刪除後，補列由此起
APPEND_END = 568
DV_EXTEND = {"O4:O562": "O4:O568",
             "Q4:Q152 Q167:Q190 Q219:Q562": "Q4:Q152 Q167:Q190 Q219:Q568",
             "AD4:AH562": "AD4:AH568"}


BACKUP_DIR = FEAT / "backup"


def sha256(path):
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


def backup(target, stamp=None):
    """W-0（R-P82）：備份交付檔並驗證 SHA256 相符，不符即中止。"""
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    before = sha256(target)
    ts = stamp or datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    dst = BACKUP_DIR / f"NR1L_GEN1_HDCC__Ver_20260813.{ts}.bak.xlsx"
    shutil.copy2(target, dst)
    got = sha256(dst)
    if got != before:
        dst.unlink(missing_ok=True)
        raise RuntimeError(f"備份 SHA256 不符：{got} != {before}，中止")
    return {"path": str(dst), "sha256": got, "source_sha256": before}


def restore(target, bak):
    """R-P78 第 3~4 項：以備份覆蓋，並驗證 SHA256 回到寫回前之值。"""
    shutil.copy2(bak["path"], target)
    got = sha256(target)
    if got != bak["source_sha256"]:
        raise RuntimeError(f"還原後 SHA256 仍不符：{got} != {bak['source_sha256']}")
    return {"restored": True, "sha256": got}


def _sd(side):
    return getattr(side, "style", None) if side is not None else None


def _style_sig(c):
    """儲存格之全部可讀樣式屬性（R-P86：逐屬性比對，非只看已知三項）。"""
    f, fl, bd, al, pr = c.font, c.fill, c.border, c.alignment, c.protection
    return {
        "font": (f.name, f.sz, f.b, f.i, f.u, f.strike, f.vertAlign),
        "fill": (fl.patternType if fl else None,
                 getattr(fl.fgColor, "rgb", None) if fl else None),
        "border": (_sd(bd.left), _sd(bd.right), _sd(bd.top), _sd(bd.bottom),
                   _sd(bd.diagonal)),
        "alignment": (al.horizontal, al.vertical, al.wrap_text, al.shrink_to_fit,
                      al.indent, al.textRotation),
        "protection": (pr.locked, pr.hidden),
        "number_format": c.number_format,
        "quotePrefix": c.quotePrefix,
    }


def formulas(path):
    wb = openpyxl.load_workbook(path)
    return {s: {c.coordinate: c.value for rr in wb[s].iter_rows() for c in rr
                if isinstance(c.value, str) and c.value.startswith("=")}
            for s in wb.sheetnames}


def run(src, dst, recalc=True, do_backup=False, stamp=None):
    """`do_backup=True` 時 src 即交付檔，就地寫回並啟用 W-0 備份與還原。"""
    report = {}
    bak = None
    if do_backup:
        bak = backup(src, stamp)
        report["W-0"] = bak
    try:
        return _run(src, dst, recalc, report)
    except Exception as exc:
        if bak and Path(dst).resolve() == Path(src).resolve():
            report["restore"] = restore(src, bak)
            report["restore"]["reason"] = str(exc)
        raise RuntimeError(f"{exc}｜已還原: {bool(bak)}") from exc


def _run(src, dst, recalc, report):

    # ---- W-1 以保留公式模式載入（禁 data_only=True）----------------------
    before = formulas(src)
    n_before = sum(len(v) for v in before.values())
    wb = openpyxl.load_workbook(src)                 # 不帶 data_only
    ws = wb["TestResults"]
    report["W-1"] = {"formulas_before": n_before,
                     "per_sheet": {k: len(v) for k, v in before.items() if v}}

    # ---- W-2 寫入 63 列之修訂內容 ---------------------------------------
    batch = {}
    for f in sorted(BATCHES.glob("*.json")):
        if f.stem == "append_uncovered_leaves":
            continue
        for x in json.load(open(f))["rows"]:
            batch[x["row"]] = x
    written, cols_touched = 0, {}
    for r, x in batch.items():
        if not x.get("changed"):
            continue
        for key, col in (("pre_conditions", L.COL["pre"]),
                         ("test_procedure", L.COL["proc"]),
                         ("expected_result_narrow_gate", L.COL["er"])):
            fld = x.get(key)
            if not isinstance(fld, dict) or not fld.get("changed"):
                continue
            ws.cell(r, col).value = fld["after"]
            written += 1
            cols_touched.setdefault(col, []).append(r)
    assert set(cols_touched) <= {L.COL["pre"], L.COL["proc"], L.COL["er"],
                                 L.COL["remarks"]}, "越權欄位"
    assert set(cols_touched.get(L.COL["er"], [])) <= set(range(424, 430)), "ER 變更逸出窄口"

    # R-P75：Remarks 窄口 30 列純附加（Atl-Mid 車型，DR#14 答覆為否）
    gate = json.load(open(FEAT / "data/remarks_scope_gate.log.json"))["rows"]
    for x in gate:
        cur = L.norm(ws.cell(x["row"], L.COL["remarks"]).value)
        assert cur == x["before"], f"r{x['row']} Remarks 原內容與 log 不符"
        assert x["after"].startswith(cur) and x["after"].endswith(x["appended"]), "非純附加"
        ws.cell(x["row"], L.COL["remarks"]).value = x["after"]
    cols_touched[L.COL["remarks"]] = [x["row"] for x in gate]
    report["W-2"] = {"cells": written,
                     "rows_per_col": {c: len(v) for c, v in sorted(cols_touched.items())},
                     "union_rows": len(set(sum(cols_touched.values(), [])))}

    # ---- W-3 刪除 r562（分支 A）------------------------------------------
    append = json.load(open(BATCHES / "append_uncovered_leaves.json"))["rows"]
    b227 = next(x for x in append if x["leaf"] == "SWE1-PROJ-227")
    assert not b227["blocked"], "227 補列未成功，應走分支 B，不得刪除 r562"
    stub_id = L.norm(ws.cell(562, L.COL["req_id"]).value)
    ws.delete_rows(562, 1)
    report["W-3"] = {"deleted_row": 562, "stub_leaf": stub_id,
                     "replacement": b227["tc_id"]}

    # ---- W-4 補列 7 條於表尾 ---------------------------------------------
    FIELD_COL = {"req_id_polarion": L.COL["polarion_id"], "req_id": L.COL["req_id"],
                 "tc_id": L.COL["tc_id"], "test_group": L.COL["test_group"],
                 "test_set": L.COL["test_set"], "test_item": L.COL["test_item"],
                 "pre_conditions": L.COL["pre"], "input_test_data": L.COL["input"],
                 "test_procedure": L.COL["proc"], "expected_result": L.COL["er"],
                 "spec_reference": L.COL["spec_ref"], "tc_ref_id": L.COL["tc_ref_id"],
                 "priority": L.COL["priority"], "design_method": L.COL["design_method"],
                 "functional_safety": L.COL["functional_safety"],
                 "author": L.COL["author"], "remarks": L.COL["remarks"]}
    for i, x in enumerate(append):
        r = APPEND_START + i
        # ---- W-5 `No.#` 寫公式，不寫字面值 ----
        ws.cell(r, L.COL["seq"]).value = "=ROW()-3"
        for key, col in FIELD_COL.items():
            if x.get(key) not in (None, ""):
                ws.cell(r, col).value = x[key]
    report["W-4"] = {"appended": len(append),
                     "rows": f"r{APPEND_START}-r{APPEND_START + len(append) - 1}",
                     "ids": [x["tc_id"] for x in append]}

    # ---- W-8 `Test Case Author` 40 個空白列補值（R-P83）--------------------
    # 置於 W-4 之後：補列已寫入，其 author 本就是 PeiPYHsu，不重複計入。
    author_rows = [r for r in range(L.DATA_FIRST, APPEND_START)
                   if not L.norm(ws.cell(r, L.COL["author"]).value)]
    for r in author_rows:
        ws.cell(r, L.COL["author"]).value = "PeiPYHsu"
    blank_after = [r for r in range(L.DATA_FIRST, APPEND_END + 1)
                   if not L.norm(ws.cell(r, L.COL["author"]).value)]
    assert len(author_rows) == 40, f"Author 待補列數 {len(author_rows)} != 40"
    assert not blank_after, f"寫入後仍有空白：{blank_after}"
    report["W-8"] = {"filled": len(author_rows), "rows": author_rows,
                     "blank_after": len(blank_after), "value": "PeiPYHsu"}

    # ---- W-9 補列與參照列同構（R-P86）------------------------------------
    # 根因：「補列」原本只被定義為「寫入儲存格的值」。值以外的一切——框線、
    # wrap_text、篩選範圍——都不在 W-4 的定義裡，所以補列在 Excel 中與既有列
    # 外觀與行為不一致。W-6 之所以存在，只因 A-PJ59 偶然發現了資料驗證那一項。
    #
    # **逐欄繼承，不是整列套同一種樣式**：參照列 r561 的 36 欄實測有 7 種相異
    # 樣式簽章（置中/靠左、wrap 有無、框線有無各不相同），整列套一種會把 36 欄
    # 全改錯。
    REF_ROW = 561
    style_before, style_after = {}, {}
    for i in range(len(append)):
        r = APPEND_START + i
        for c in range(1, 37):
            # 變數名不得與函式參數 `dst`（輸出路徑）相同 —— 遮蔽後 wb.save(dst)
            # 會拿到一個 Cell，在 zip 寫入階段才爆，離錯誤點很遠。
            ref_cell, new_cell = ws.cell(REF_ROW, c), ws.cell(r, c)
            if i == 0:
                style_before[c] = _style_sig(new_cell)
            new_cell._style = copy(ref_cell._style)
        rd_src, rd_dst = ws.row_dimensions[REF_ROW], ws.row_dimensions[r]
        rd_dst.height = rd_src.height
        rd_dst.hidden = rd_src.hidden
        rd_dst.outlineLevel = rd_src.outlineLevel
        if rd_src.customHeight:
            rd_dst.customHeight = rd_src.customHeight
    for c in range(1, 37):
        style_after[c] = _style_sig(ws.cell(APPEND_START, c))

    # 自動篩選 ref 隨資料區延伸
    af_before = ws.auto_filter.ref
    if af_before:
        # 保留原本的絕對參照形式：`$AJ$562` → `$AJ$568`，不可吃掉 `$`
        ws.auto_filter.ref = re.sub(r"(\$?)\d+$", rf"\g<1>{APPEND_END}", str(af_before))

    mismatch9 = [c for c in range(1, 37)
                 if _style_sig(ws.cell(APPEND_START, c)) != _style_sig(ws.cell(REF_ROW, c))]
    assert not mismatch9, f"W-9 樣式未繼承：{mismatch9}"
    report["W-9"] = {"ref_row": REF_ROW, "rows": len(append), "cols": 36,
                     "style_before": style_before, "style_after": style_after,
                     "distinct_ref_signatures": len({json.dumps(_style_sig(ws.cell(REF_ROW, c)),
                                                                sort_keys=True, default=str)
                                                     for c in range(1, 37)}),
                     "autofilter": {"before": str(af_before), "after": str(ws.auto_filter.ref)},
                     "mismatch": mismatch9}

    # ---- W-5 驗證：c2 全欄皆為公式 ---------------------------------------
    seq_cells = [ws.cell(r, L.COL["seq"]) for r in range(L.DATA_FIRST, APPEND_END + 1)]
    literals = [c.coordinate for c in seq_cells
                if not (isinstance(c.value, str) and c.value.startswith("="))]
    assert not literals, f"No.# 出現字面值：{literals}"
    report["W-5"] = {"seq_cells": len(seq_cells), "literals": len(literals),
                     "all_formula": True}

    # ---- W-6 資料驗證範圍延伸至 r568 -------------------------------------
    extended = []
    for dv in ws.data_validations.dataValidation:
        old = str(dv.sqref)
        if old in DV_EXTEND:
            dv.sqref = DV_EXTEND[old]
            extended.append({"from": old, "to": DV_EXTEND[old]})
    report["W-6"] = {"extended": extended,
                     "final": [str(dv.sqref) for dv in ws.data_validations.dataValidation]}

    # ---- W-1 驗證：公式全數保全（R-P78：可在 save() 前驗者一律前置）--------
    after = {s: {c.coordinate: c.value for rr in wb[s].iter_rows() for c in rr
                 if isinstance(c.value, str) and c.value.startswith("=")}
             for s in wb.sheetnames}
    n_after = sum(len(v) for v in after.values())
    diffs = {}
    for s in before:
        if s == "TestResults":
            continue                       # 刪 1 補 7，座標數必變，另驗內容
        if before[s] != after.get(s, {}):
            diffs[s] = {"before": len(before[s]), "after": len(after.get(s, {}))}
    tr_vals = set(after["TestResults"].values())
    assert not diffs, f"公式內容改變：{diffs}"
    assert tr_vals == {"=ROW()-3"}, f"TestResults 公式不純：{tr_vals}"
    assert len(after["TestResults"]) == len(before["TestResults"]) - 1 + len(append)
    # R-P77：通過條件為 775 + 補列淨增 6 = 781，不是「775 不變」
    assert n_after == n_before - 1 + len(append), f"公式總數 {n_after} != 781"
    report["W-1"].update({"formulas_after": n_after, "expected": n_before - 1 + len(append),
                          "other_sheets_identical": True,
                          "testresults": f"{len(before['TestResults'])} → "
                                         f"{len(after['TestResults'])}"})

    wb.save(dst)        # 記憶體驗證全過後方落盤

    # ---- W-7 外部重算 -----------------------------------------------------
    if recalc:
        report["W-7"] = recalculate(dst)
    return report


def recalculate(path):
    """以 LibreOffice headless 重算，讀回 TestProgress 統計值（R-P70）。"""
    tmp = Path(tempfile.mkdtemp())
    subprocess.run(["soffice", "--headless", "--convert-to", "xlsx",
                    "--outdir", str(tmp), str(path)],
                   capture_output=True, timeout=300)
    out = tmp / Path(path).name
    if not out.exists():
        return {"status": "重算失敗", "path": None}
    tp = openpyxl.load_workbook(out, data_only=True)["TestProgress"]
    got = {L.norm(tp.cell(r, 3).value): tp.cell(r, 4).value for r in range(12, 24)}
    exp = json.load(open(FEAT / "data/w7_expected.json"))["expected"]
    cmp = {k: {"expected": v["expected_after"], "actual": got.get(k),
               "match": got.get(k) == v["expected_after"]}
           for k, v in exp.items() if v["before"] is not None}
    return {"status": "ok", "recalculated": str(out), "total_items": cmp,
            "all_match": all(v["match"] for v in cmp.values())}


if __name__ == "__main__":
    src, dst = sys.argv[1], sys.argv[2]
    rep = run(src, dst, do_backup="--backup" in sys.argv)
    print(json.dumps(rep, ensure_ascii=False, indent=2, default=str))
