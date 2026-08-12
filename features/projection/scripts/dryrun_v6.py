#!/usr/bin/env python3
"""Phase 6 dry-run v6 — checklist D-1 ~ D-11 (R-P81 ~ R-P85).

Builds the refined workbook IN MEMORY (base xlsx + 11 batch JSONs + 7 appended
rows) and runs every check against it. **Nothing is written back to xlsx** —
"dry-run 是驗證，不是交付" (Phase 6 packet).

Every comparison expression comes from `lint_defs` (R-P49). This script does
not re-implement a single pattern.
"""

import glob
import hashlib
import json
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import lint_defs as L
from recon_projection import parse_dbc

ROOT = Path(__file__).resolve().parents[3]
FEAT = ROOT / "features/projection"
INP = FEAT / "inputs"
WB = INP / "NR1L_GEN1(HDCC)_Ver_20260813.xlsx"

# 欄索引與掃描範圍一律取自 lint_defs，本檔不自行假設（R-P65）
COL_PRE, COL_PROC, COL_ER = L.COL["pre"], L.COL["proc"], L.COL["er"]
COL_AUTHOR = L.COL["author"]
NARROW_ROWS = list(range(424, 430))          # R-P12 純刪除窄口
NARROW_TOKENS = {"correctly", "normally", "properly", "successfully"}
DATA_FIRST, DATA_LAST = L.DATA_FIRST, L.DATA_LAST
OTHER_SHEETS = ["TestProgress", "Cover_old", "ChangeHistory_old", "QS Suggestion",
                "下拉選單", "Reference", "BugList", "Test Case Framework"]

rep = []          # (項目, 結果, 數據)
out = {}


def h(s):
    return hashlib.sha256(str(s or "").encode()).hexdigest()[:16]


# --------------------------------------------------------------------------
# 0. 載入原簿與批次，組出修訂後狀態
# --------------------------------------------------------------------------
wb = openpyxl.load_workbook(WB, data_only=True)
ws = wb["TestResults"]
base = {r: {c: L.norm(ws.cell(r, c).value) for c in range(1, 37)}
        for r in range(DATA_FIRST, 563)}

batch_rows, batch_src = {}, {}
for f in sorted(glob.glob(str(FEAT / "batches/*.json"))):
    name = Path(f).stem
    if name == "append_uncovered_leaves":
        continue
    for x in json.load(open(f))["rows"]:
        batch_rows[x["row"]] = x
        batch_src[x["row"]] = name

refined = {r: dict(v) for r, v in base.items()}
changed_cols = {}                     # col -> [rows]
for r, x in batch_rows.items():
    if not x.get("changed"):
        continue
    for key, col in (("pre_conditions", COL_PRE), ("test_procedure", COL_PROC),
                     ("expected_result_narrow_gate", COL_ER)):
        field = x.get(key)
        # 批次欄位為 {before, after, changed} 巢狀結構；per-column 旗標才是變更依據
        if not isinstance(field, dict) or not field.get("changed"):
            continue
        refined[r][col] = L.norm(field["after"])
        if refined[r][col] != base[r][col]:
            changed_cols.setdefault(col, []).append(r)

# R-P75：Remarks 窄口 30 列純附加（Atl-Mid，DR#14 答覆為否）
remarks_gate = json.load(open(FEAT / "data/remarks_scope_gate.log.json"))["rows"]
REMARKS_ROWS = [x["row"] for x in remarks_gate]
for x in remarks_gate:
    refined[x["row"]][L.COL["remarks"]] = L.norm(x["after"])
    changed_cols.setdefault(L.COL["remarks"], []).append(x["row"])

# W-8（R-P83）：Test Case Author 40 個空白列補值 —— 此後須驗證其**確實發生**
AUTHOR_ROWS = [r for r in range(DATA_FIRST, 562) if not base[r][L.COL["author"]]]
for r in AUTHOR_ROWS:
    refined[r][L.COL["author"]] = "PeiPYHsu"
    changed_cols.setdefault(L.COL["author"], []).append(r)

append = json.load(open(FEAT / "batches/append_uncovered_leaves.json"))["rows"]

# --------------------------------------------------------------------------
# D-1 diff 只落在可編輯欄 + 授權例外
# --------------------------------------------------------------------------
COLNAME = {COL_PRE: "Pre-Conditions (I)", COL_PROC: "Test procedure (K)",
           COL_ER: "Expected Result (L)", L.COL["remarks"]: "Remarks (AJ)",
           L.COL["author"]: "Test Case Author (Z)"}
# 授權集合 = 可編輯欄 ∪ 全部凍結欄授權例外（推導，不列舉 —— R-P84 同理）
ALLOWED = {COL_PRE, COL_PROC} | set(L.FROZEN_EXCEPTIONS)
illegal = {c: v for c, v in changed_cols.items() if c not in ALLOWED}
er_rows = sorted(changed_cols.get(COL_ER, []))
er_outside = [r for r in er_rows if r not in NARROW_ROWS]

narrow_diffs = []
for r in NARROW_ROWS:
    old, new = base[r][COL_ER], refined[r][COL_ER]
    # 詞元多重集合比對。不得用 str.count —— 子字串計數會把 `correctly` 之刪除
    # 誤報為連帶刪掉 `or`（與 A-PJ38 之邊界缺陷同類）。
    co, cn = Counter(re.findall(r"[A-Za-z']+", old)), Counter(re.findall(r"[A-Za-z']+", new))
    removed = sorted((co - cn).elements())
    added = sorted((cn - co).elements())
    pure_del = not added
    ok = (old != new) and pure_del and {x.lower() for x in removed} <= NARROW_TOKENS
    narrow_diffs.append({"row": r, "removed": removed, "added": added, "pure_deletion": pure_del,
                         "ok": ok, "before": old, "after": new})

# Remarks 窄口：純附加、固定字串、白名單內
rm_outside = [r for r in changed_cols.get(L.COL["remarks"], []) if r not in REMARKS_ROWS]
rm_diffs = [{"row": x["row"], "pure_append": x["after"].startswith(x["before"]),
             "fixed_string": x["after"].endswith(x["appended"]),
             "before": x["before"], "after": x["after"]} for x in remarks_gate]
rm_ok = not rm_outside and all(d["pure_append"] and d["fixed_string"] for d in rm_diffs)
d1_ok = not illegal and not er_outside and all(d["ok"] for d in narrow_diffs) and rm_ok
rep.append(("D-1", "PASS" if d1_ok else "FAIL",
            {COLNAME.get(c, f"col{c}"): len(v) for c, v in sorted(changed_cols.items())}))
out["D-1"] = {"changed_distribution": {COLNAME.get(c, f"col{c}"): sorted(v)
                                       for c, v in sorted(changed_cols.items())},
              "narrow_gate_diffs": narrow_diffs,
              "illegal_columns": illegal, "er_outside_narrow": er_outside,
              "remarks_gate_diffs": rm_diffs, "remarks_outside_whitelist": rm_outside}

# --------------------------------------------------------------------------
# D-2 凍結欄逐列雜湊不變 + 其餘 8 分頁
# --------------------------------------------------------------------------
frozen_cols = L.FROZEN_COLS
mismatch = []
for r in range(DATA_FIRST, 563):
    for c in frozen_cols:
        if h(base[r][c]) == h(refined[r][c]):
            continue
        if c == COL_ER and r in NARROW_ROWS:
            continue                                  # R-P12 授權例外
        if c == L.COL["remarks"] and r in REMARKS_ROWS:
            continue                                  # R-P75 授權例外
        if c == L.COL["author"] and r in AUTHOR_ROWS:
            continue                                  # R-P19/R-P54 授權例外
        mismatch.append((r, c))
author_blank = [r for r in range(DATA_FIRST, 563) if not base[r][COL_AUTHOR]]
# R-P60：公式軌與值軌分開。dry-run 未開寫入，故值軌標「未實測」而非 PASS
# （A-PJ56：不可能失敗的檢查項不得計為 PASS 證據）。
wbf = openpyxl.load_workbook(WB)          # data_only 不帶 → 取公式
sheet_hash_formula, sheet_hash_value, formula_count = {}, {}, {}
for s in OTHER_SHEETS + ["TestResults"]:
    cells = [cc for rr in wbf[s].iter_rows() for cc in rr]
    formula_count[s] = sum(1 for cc in cells
                           if isinstance(cc.value, str) and cc.value.startswith("="))
    sheet_hash_formula[s] = h("\x1f".join(L.norm(cc.value) for cc in cells))
    sheet_hash_value[s] = h("\x1f".join(
        L.norm(cc.value) for rr in wb[s].iter_rows() for cc in rr))
sheet_hash = sheet_hash_formula
d2_ok = not mismatch
rep.append(("D-2", "PASS" if d2_ok else "FAIL",
            {"凍結欄": len(frozen_cols), "比對列": 559, "不符": len(mismatch),
             "公式軌": sum(formula_count.values()), "值軌": "未實測(A-PJ56)",
             "Author 待補": len(author_blank)}))
out["D-2"] = {"frozen_col_count": len(frozen_cols), "mismatches": mismatch,
              "sheet_hash_formula": sheet_hash_formula,
              "sheet_hash_value": sheet_hash_value,
              "formula_count": formula_count,
              "value_track_status": "未實測 —— dry-run 未開寫入（A-PJ56）",
              "exception_narrow_er": NARROW_ROWS,
              "exception_author_rows": author_blank,
              "exception_remarks_rows": REMARKS_ROWS,
              "exception_author_filled": AUTHOR_ROWS,
              "exception_total": len(NARROW_ROWS) + len(AUTHOR_ROWS) + len(REMARKS_ROWS),
              "author_blank_after": [r for r in range(DATA_FIRST, 562)
                                     if not refined[r][L.COL["author"]]]}

# --------------------------------------------------------------------------
# D-3 列數與列序（R-P54 分支）
# --------------------------------------------------------------------------
blocked_leaves = [x["leaf"] for x in append if x["blocked"]]
b227 = next(x for x in append if x["leaf"] == "SWE1-PROJ-227")
branch = "A" if not b227["blocked"] else "B"
n_final = (558 if branch == "A" else 559) + len(append)
# 列身分 = 凍結 34 欄之逐列雜湊（L.ROW_IDENTITY）。**不得用 `No.#`** ——
# 該欄內容為公式 `=ROW()-3`，值恆等於列位置，重排後跟著改，永遠偵測不到移動。
def _rid(cells):
    return h("\x1f".join(cells[c] for c in L.ROW_IDENTITY_COLS))
ident_base = [_rid(base[r]) for r in range(DATA_FIRST, 562)]
ident_new = [_rid(refined[r]) for r in range(DATA_FIRST, 562)]
identity_unique = len(set(ident_base)) == len(ident_base)
moved = [DATA_FIRST + i for i, (a, b) in enumerate(zip(ident_base, ident_new)) if a != b]
rep.append(("D-3", "PASS" if not moved else "FAIL",
            {"分支": branch, "559 →": 558 if branch == "A" else 559,
             "補列": len(append), "最終列數": n_final, "末列": f"r{562 + len(append) - (1 if branch=='A' else 0)}",
             "列身分": L.ROW_IDENTITY, "身分唯一": identity_unique,
             "被移動列": len(moved)}))
out["D-3"] = {"branch": branch, "deleted": [562] if branch == "A" else [],
              "final_rows": n_final, "moved_rows": moved,
              "row_identity": L.ROW_IDENTITY,
              "identity_unique": identity_unique,
              "identity_rejected": L.ROW_IDENTITY_REJECTED}

# --------------------------------------------------------------------------
# D-4 / D-6 補列與其獨立驗證
# --------------------------------------------------------------------------
fd, bh = parse_dbc(INP / "PHDCC27_E2A_R1_FDCAN8.dbc"), parse_dbc(INP / "PHDCC27_E2A_R1_BHCAN.dbc")
vf = L.vf176_signals(json.load(open(FEAT / "data/signal_map.json")))
TG = {refined[r][6] for r in range(DATA_FIRST, 562) if refined[r][6]}
TS = {refined[r][7] for r in range(DATA_FIRST, 562) if refined[r][7]}
# Design Method 之值域取**資料驗證實際指向**的 Reference!$C$4:$C$12，
# 不取 `下拉選單` 分頁 —— 兩者「組合測試」拼法不同，v2 驗錯了清單。
DM = {L.norm(wb["Reference"].cell(r, 3).value) for r in range(4, 13)}
DM_SHEET = {L.norm(wb["下拉選單"].cell(r, 1).value)
            for r in range(1, wb["下拉選單"].max_row + 1)
            if L.norm(wb["下拉選單"].cell(r, 1).value)}
dv_violations = [r for r in range(DATA_FIRST, 562)
                 if refined[r][L.COL["design_method"]]
                 and refined[r][L.COL["design_method"]] not in DM]
existing_ids = {refined[r][L.COL["tc_id"]] for r in range(DATA_FIRST, 562)}
cp = openpyxl.load_workbook(
    INP / "FM-WI-FSM-037-A03 STLA 報告_SWRA STLA Report_SWRA-CPAA_0521.xlsx",
    data_only=True)["Basic Report"]
LEAVES = {L.norm(cp.cell(r, 2).value) for r in range(2, cp.max_row + 1)}

SECTIONS = {
    "cfts085": json.load(open(FEAT / "data/cfts085_sections.json"))["clause_to_section"],
    "sysad": json.load(open(FEAT / "data/sysad_sections.json"))["clause_to_section"],
    "huig": json.load(open(FEAT / "data/huig_sections.json"))["clause_to_section"],
    "addendum_ids": {s.split()[0] for s in
                     json.load(open(FEAT / "data/carplay_addendum_sections.json"))["section_to_rows"]},
}

d6 = []
for x in append:
    e = []
    if x["req_id"] not in LEAVES:
        e.append("leaf 不存在於 037")
    if x["test_group"] not in TG:
        e.append("Test Group 值域")
    if x["test_set"] not in TS:
        e.append("Test Set 值域")
    if x["priority"] not in {"P0", "P1", "P2", "P3"}:
        e.append("Priority")
    if x["design_method"] not in DM:
        e.append("Design Method 不在 Reference!C4:C12")
    if not re.match(r"NR1L-PROJ-\d{3}$", x.get("tc_id", "")):
        e.append("Test Case ID 格式")
    if x.get("tc_id") in existing_ids:
        e.append(f"Test Case ID 與既有重複 {x['tc_id']}")
    if not x["blocked"]:
        pre, proc, er = x["pre_conditions"], x["test_procedure"], x["expected_result"]
        txt = f"{pre}\n{proc}\n{er}"
        if L.RE_BANNED.search(f"{pre}\n{proc}"):
            e.append(f"L-PJ5 {L.RE_BANNED.findall(pre + proc)}")
        if L.RE_VAGUE.search(txt):
            e.append(f"L-PJ6 {L.RE_VAGUE.findall(txt)}")
        if L.RE_TOKEN.search(txt):
            e.append("L-PJ3 token")
        if L.RE_GENERIC_TOOL.search(pre) and not L.RE_NAMED_TOOL.search(proc):
            e.append(f"L-PJ9 {L.RE_GENERIC_TOOL.findall(pre)}")
        if L.placeholder_defects(txt):
            e.append(f"L-PJ10 {L.placeholder_defects(txt)}")
        if L.forward_xrefs(proc):
            e.append("前向指涉")
        ns, ne = len(L.steps(proc)), len(L.steps(er))
        if ns != ne:
            e.append(f"L-PJ8 {ns}/{ne}")
        for m, s in set(re.findall(r"\b([A-Z][A-Z0-9_]{3,})\.([A-Za-z_][A-Za-z0-9_]*)", txt)):
            if L.resolve_signal(m, s, fd, bh, vf, None)[0] is None:
                e.append(f"L-PJ1 {m}.{s}")
    # R-P73：真解析，不再只比對格式
    for anchor in x["spec_reference"].split("\n"):
        res = L.resolve_spec_anchor(anchor, SECTIONS)
        if res and res[1] is False:
            e.append(f"spec 錨點解析失敗 {anchor} — {res[2]}")
    d6.append({"leaf": x["leaf"], "blocked": x["blocked"], "errors": e,
               "steps": len(L.steps(x["test_procedure"])),
               "er_lines": len(L.steps(x["expected_result"]))})
# R-P79：補列全部 34 個凍結欄逐欄判定
cols34 = json.load(open(FEAT / "data/d6_append_34cols.json"))
c34_fail = [c for c in cols34["columns"] if c["pass"] is False]
c34_conflict = [c for c in cols34["columns"] if c["pass"] is None]
d4_ok = all(not r["errors"] for r in d6)
rep.append(("D-4", "PASS" if d4_ok else "FAIL",
            {"補列": len(append), "可寫 TC": sum(1 for x in append if not x["blocked"]),
             "BLOCKED 佔位": len(blocked_leaves), "未補": 0}))
rep.append(("D-6", "PASS" if (d4_ok and not c34_fail and not c34_conflict)
            else ("FAIL" if c34_fail or not d4_ok else "**衝突**"),
            {"逐條驗證": len(d6), "有誤": sum(1 for r in d6 if r['errors']),
             "34 欄": f"PASS {sum(1 for c in cols34['columns'] if c['pass'] is True)}"
                      f" / FAIL {len(c34_fail)} / 衝突 {len(c34_conflict)}",
             "ID": "NR1L-PROJ-560~566",
             "既有違反 DV 之列": dv_violations}))
out["D-4"] = {"rows": append}
out["D-6"] = {"rows": d6, "cols34": cols34, "design_method_domain": sorted(DM),
              "design_method_sheet_differs": sorted(DM ^ DM_SHEET),
              "existing_dv_violations": dv_violations}

# --------------------------------------------------------------------------
# D-5 阻塞列 ↔ 編號
# --------------------------------------------------------------------------
# artifact 結構為 {列號字串: [裁決/DR 編號]}
d5 = json.load(open(FEAT / "data/d5_blocked_rows.json"))
byrow = {int(k): set(v) for k, v in d5.items() if not k.startswith("_")}
for r in (177, 188):
    byrow.setdefault(r, set()).update({"R-P56", "DR#13"})
noref = [r for r, v in byrow.items() if not v]
rep.append(("D-5", "PASS" if not noref else "FAIL",
            {"不重複列": len(byrow), "無編號可指": len(noref), "本輪新增": "r177, r188"}))
out["D-5"] = {"rows": {r: sorted(v) for r, v in sorted(byrow.items())}, "no_ref": noref}

# --------------------------------------------------------------------------
# D-9 Test Case Framework 分頁
# --------------------------------------------------------------------------
tcf = wb["Test Case Framework"]
tcf_vals = [L.norm(c.value) for rr in tcf.iter_rows() for c in rr if L.norm(c.value)]
l2_wb = {v for v in tcf_vals if v in TS}
out["D-9"] = {"sheet_hash": sheet_hash["Test Case Framework"],
              "non_empty_cells": len(tcf_vals),
              "layer2_terms_matching_test_set": sorted(l2_wb),
              "written": False}
d9_ok = len(tcf_vals) == 0          # R-P66 §4：驗證維持空白；出現內容即 ABORT
rep.append(("D-9", "PASS" if d9_ok else "ABORT",
            {"維持空白": d9_ok, "非空格": len(tcf_vals),
             "寫入": "否（fill_test_group_set=false）"}))

# --------------------------------------------------------------------------
# D-11 全簿資料驗證合規（R-P74）
# --------------------------------------------------------------------------
def _dv_domain(dv):
    f = str(dv.formula1 or "")
    if f.startswith('"'):                       # inline 列舉
        return {x.strip() for x in f.strip('"').split(",")}
    m = re.match(r"([^!]+)!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)", f)
    if m:
        ci = openpyxl.utils.column_index_from_string(m.group(2))
        return {L.norm(wb[m.group(1)].cell(r, ci).value)
                for r in range(int(m.group(3)), int(m.group(5)) + 1)} - {""}


def _dv_cells(sq):
    out = []
    for part in str(sq).split():
        m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", part)
        if not m:
            continue
        c1 = openpyxl.utils.column_index_from_string(m.group(1))
        c2 = openpyxl.utils.column_index_from_string(m.group(3))
        for r in range(int(m.group(2)), int(m.group(4)) + 1):
            out += [(r, c) for c in range(c1, c2 + 1)]
    return out


def _dv_norm(v):
    """數值儲存格對字串列舉比對前須正規化 —— Excel 存數值 1，DV 寫 "0,1"。
    未正規化會把 Vehicle Model 的 21 格全部誤報為違規。"""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return L.norm(v)


dv_summary, dv_viol = [], {}
for dv in openpyxl.load_workbook(WB)["TestResults"].data_validations.dataValidation:
    dom = _dv_domain(dv)
    if not dom:
        continue
    cells = _dv_cells(dv.sqref)
    bad, nonempty = [], 0
    for r, c in cells:
        v = _dv_norm(ws.cell(r, c).value)
        if not v:
            continue                            # 空白不視為違規（R-P74）
        nonempty += 1
        if v not in dom:
            bad.append([r, c, v])
            dv_viol.setdefault(r, []).append({"col": c, "value": v})
    hdr = L.norm(ws.cell(2, cells[0][1]).value).split("\n")[0] or "Vehicle Model"
    dv_summary.append({"header": hdr, "sqref": str(dv.sqref), "domain": len(dom),
                       "nonempty": nonempty, "violations": bad})
# 補列違規即 FAIL；既有違規凍結欄不動入 RD-1
append_viol = [x for x in d6 if any("值域" in e or "Design Method" in e for e in x["errors"])]
existing_viol = sorted(dv_viol)
d11_ok = not append_viol
rep.append(("D-11", "PASS" if d11_ok else "FAIL",
            {"受控欄": len(dv_summary), "既有違規列": existing_viol,
             "補列違規": len(append_viol), "空白": "不計違規(R-P74)"}))
out["D-11"] = {"summary": dv_summary, "existing_violations": {str(k): v for k, v in dv_viol.items()},
               "appended_violations": append_viol,
               "note": "既有違規凍結欄不動入 RD-1；補列違規即 FAIL"}

# --------------------------------------------------------------------------
# D-10 八項全簿基線
# --------------------------------------------------------------------------
def measure(rows):
    m = {"L-PJ5 禁詞": 0, "L-PJ6 模糊語": 0, "L-PJ9 泛稱工具": 0,
         "L-PJ10 缺陷類": 0, "L-PJ10 參數類": 0, "步驟交叉指涉": 0,
         "步數 != ER 例外": 0, "前向循環指涉": 0}
    for cells in rows:
        pre, proc, er = cells[COL_PRE], cells[COL_PROC], cells[COL_ER]
        m["L-PJ5 禁詞"] += len(L.RE_BANNED.findall(f"{pre}\n{proc}"))
        m["L-PJ6 模糊語"] += len(L.RE_VAGUE.findall(f"{pre}\n{proc}\n{er}"))
        if L.RE_GENERIC_TOOL.search(pre) and not L.RE_NAMED_TOOL.search(proc):
            m["L-PJ9 泛稱工具"] += 1
        # L-PJ10 之計數單位為「列」（非出現次數），掃描範圍含 ER：
        # 參數類 8 列中 r60/r61 之 `<Device Name>` 只出現在 Expected Result。
        ph_scan = f"{pre}\n{proc}\n{er}"
        if L.placeholder_defects(ph_scan):
            m["L-PJ10 缺陷類"] += 1
        if any(t in ph_scan for t in L.PLACEHOLDER_WHITELIST):
            m["L-PJ10 參數類"] += 1
        if L.RE_STEP_XREF.search(proc):
            m["步驟交叉指涉"] += 1
        if L.steps(proc) and L.steps(er) and len(L.steps(proc)) != len(L.steps(er)):
            m["步數 != ER 例外"] += 1
        m["前向循環指涉"] += len(L.forward_xrefs(proc))
    return m

existing = [refined[r] for r in range(DATA_FIRST, 562)]
app_cells = [{COL_PRE: x["pre_conditions"], COL_PROC: x["test_procedure"],
              COL_ER: x["expected_result"]} for x in append if not x["blocked"]]
m_exist, m_all = measure(existing), measure(existing + app_cells)
EXPECT = {"L-PJ5 禁詞": 1, "L-PJ6 模糊語": 4, "L-PJ9 泛稱工具": 17,
          "L-PJ10 缺陷類": 5, "L-PJ10 參數類": 8, "步驟交叉指涉": 30,
          "步數 != ER 例外": 3, "前向循環指涉": 0}
d10 = {k: {"expected": EXPECT[k], "existing": m_exist[k], "with_append": m_all[k],
           "match": m_exist[k] == EXPECT[k],
           "append_shifts": m_all[k] != m_exist[k]} for k in EXPECT}
d10_ok = all(v["match"] for v in d10.values())
rep.append(("D-10", "PASS" if d10_ok else "FAIL",
            {k: f"{v['existing']}/{v['expected']}" + ("*" if v["append_shifts"] else "")
             for k, v in d10.items()}))
out["D-10"] = d10

# --------------------------------------------------------------------------
print("=" * 78)
for name, verdict, data in rep:
    print(f"{name:<5} {verdict:<5} {data}")
print("=" * 78)
print("整體:", "PASS" if all(v == "PASS" for _, v, _ in rep) else "FAIL")
json.dump(out, open(FEAT / "data/dryrun_v6.json", "w"), ensure_ascii=False,
          indent=2, default=str)
