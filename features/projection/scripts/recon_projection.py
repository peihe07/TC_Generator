#!/usr/bin/env python3
"""Phase 0/1 recon for the Projection feature (workbook_state = FULL_REFINE).

The shared scripts/recon.py assumes an FM-WI-FSM-036 form instance (header
row 9, done-region-by-author, BLANK/PARTIAL/FULL states). Projection's base
workbook is NOT that form — it is the NR1L_GEN1(HDCC) execution workbook
(header row 2, data from row 4, seven vehicle-model columns and five
build-result columns). So the survey is written here rather than forced
through the shared script, per canon "copy+yaml, no shared library".

This script DETECTS only. It writes data/recon.json and prints a report;
it never edits the workbook and never rules on anything.

Usage:
    .venv/bin/python features/projection/scripts/recon_projection.py
"""

import hashlib
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
FEAT = ROOT / "features" / "projection"
INP = FEAT / "inputs"

WORKBOOK = INP / "NR1L_GEN1(HDCC)_Ver_20260813.xlsx"
SWRA = INP / "FM-WI-FSM-037-A03 STLA 報告_SWRA STLA Report_SWRA-CPAA_0521.xlsx"
SWRA_MD = INP / "SWE1_PROJ_FM-WI-FSM-037-A03 STLA 報告_SWRA STLA Report_SWRA_MD20260324.xlsx"

SHEET = "TestResults"
HEADER_ROW = 2
DATA_START = 4

# Column indices verified against the header row by check_headers(), not
# assumed. Letters are a prior; the header text is the authority.
COLS = {
    "seq": 2,               # B  No.#
    "req_id_polarion": 3,   # C
    "req_id": 4,            # D  Requirement or Design ID
    "tc_id": 5,             # E  Test Case ID
    "test_group": 6,        # F
    "test_set": 7,          # G
    "test_item": 8,         # H
    "pre_conditions": 9,    # I
    "input_test_data": 10,  # J
    "test_procedure": 11,   # K
    "expected_result": 12,  # L
    "spec_reference": 13,   # M
    "tc_ref_id": 14,        # N
    "priority": 15,         # O
    "estimated_test_time": 16,  # P
    "design_method": 17,    # Q
    "functional_safety": 18,    # R
    "author": 26,           # Z
    "remarks": 36,          # AJ
}

HEADER_TOKENS = {
    "req_id": "requirement or design id",
    "tc_id": "test case id",
    "test_group": "test group",
    "test_set": "test set",
    "test_item": "test item",
    "pre_conditions": "pre-conditions",
    "input_test_data": "input test data",
    "test_procedure": "test procedure",
    "expected_result": "expected result",
    "spec_reference": "specification reference",
    "tc_ref_id": "test case reference id",
    "priority": "test case priority",
    "estimated_test_time": "estimated test time",
    "design_method": "test case design methods",
    "functional_safety": "functional safety",
    "author": "test case author",
    "remarks": "remarks",
}

# Build-result columns; row 3 carries the build label, row 2 the group header.
BUILD_COLS = {30: "00.01.01.02", 31: "00.01.01.03", 32: "Daily Build 20260805",
              33: "00.01.01.04", 34: "00.01.01.05"}
EXEC_BUILD_COL = 33   # 00.01.01.04 — the build the 下放包 §4 reports

# canon §5.1 banned procedure verbs. Scanned on the procedure column only —
# they are procedure defects; the same word in an Expected Result is a
# statement of outcome, not an unexecutable instruction.
BANNED_VERBS = ["observe", "check whether", "confirm whether", "see if",
                "watch", "monitor", "inspect"]
# Vague language is counted per ROW over procedure + expected result together:
# a row carrying "correctly" in both columns is ONE defective row, not two.
VAGUE_WORDS = ["correctly", "a while", "normally", "properly", "successfully"]
VAGUE_FIELDS = ["test_procedure", "expected_result"]

# Tool dependency is read off the procedure column: it measures what the
# tester is instructed to operate, not what the requirement mentions.
# `CAN` is matched CASE-SENSITIVELY — a case-insensitive \bCAN\b matches the
# English modal verb "can" and inflates the count from 39 to 86.
TOOL_PATTERNS = {
    "CAN tool": (r"\bCAN\b", 0),
    "PCTS": (r"PCTS", re.I),
    "ATS": (r"\bATS\b", 0),
    "testapp": (r"test\s*app", re.I),
    "logcat": (r"logcat", re.I),
    "adb": (r"\badb\b", re.I),
}

# A row is "完全空白" when every TC-content column is empty. Seq / Polarion id
# / req id / Test Group / Test Item may still carry values — the row exists as
# a traceability stub with no test case in it.
CONTENT_FIELDS = ["test_set", "pre_conditions", "test_procedure",
                  "expected_result", "spec_reference", "tc_ref_id",
                  "priority", "design_method", "functional_safety"]


def norm(v) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip()


def lower(v) -> str:
    return norm(v).lower()


def sha256(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def check_headers(ws) -> dict:
    """Confirm every mapped column's header text matches its expected token."""
    out = {}
    for field, token in HEADER_TOKENS.items():
        col = COLS[field]
        actual = lower(ws.cell(HEADER_ROW, col).value)
        out[field] = {"col": col, "header": norm(ws.cell(HEADER_ROW, col).value),
                      "match": token in actual}
    return out


def load_rows(ws):
    """Every physical row from DATA_START to the last row carrying any value."""
    rows = []
    last = ws.max_row
    for r in range(DATA_START, last + 1):
        vals = {k: ws.cell(r, c).value for k, c in COLS.items()}
        vals["_row"] = r
        vals["_builds"] = {lbl: ws.cell(r, c).value for c, lbl in BUILD_COLS.items()}
        rows.append(vals)
    # Trim the trailing all-empty tail: a row counts as present if ANY mapped
    # cell or any build cell carries a value.
    def occupied(v):
        return any(norm(v[k]) for k in COLS) or any(norm(x) for x in v["_builds"].values())
    while rows and not occupied(rows[-1]):
        rows.pop()
    return rows


def survey_workbook():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]
    headers = check_headers(ws)
    rows = load_rows(ws)

    fill = {k: sum(1 for r in rows if norm(r[k])) for k in COLS if not k.startswith("_")}
    content_blank = [r["_row"] for r in rows
                     if not any(norm(r[k]) for k in CONTENT_FIELDS)]

    # $token$ inventory across every text-bearing field
    text_fields = ["test_item", "pre_conditions", "input_test_data",
                   "test_procedure", "expected_result", "remarks"]
    tok_rows = Counter()
    tok_hits = Counter()
    rows_with_token = 0
    for r in rows:
        blob = " ".join(norm(r[f]) for f in text_fields)
        found = set(re.findall(r"\$[^$\s]+\$", blob))
        if found:
            rows_with_token += 1
        for t in found:
            tok_rows[t] += 1
        for t in re.findall(r"\$[^$\s]+\$", blob):
            tok_hits[t] += 1

    # Tool dependency, reported under BOTH scopes. No single scope reproduces
    # the 下放包 §4 table: the procedure column alone gives CAN=39 (matching)
    # but adb=0, while procedure+pre-conditions gives adb=2 (matching) but
    # CAN=40. The two adb rows carry the instruction in Pre-Conditions, not in
    # Procedure. Reporting both is the honest form; picking one silently is not.
    tools = {}
    for scope, fields in (("procedure", ["test_procedure"]),
                          ("procedure+pre_conditions", ["test_procedure", "pre_conditions"])):
        tools[scope] = {}
        for name, (pat, flags) in TOOL_PATTERNS.items():
            rx = re.compile(pat, flags)
            tools[scope][name] = sum(
                1 for r in rows if rx.search(" ".join(norm(r[f]) for f in fields)))

    # environment blockers — exact Remarks strings
    remarks = Counter(norm(r["remarks"]) for r in rows if norm(r["remarks"]))
    blockers = {}
    for probe in ["Need to test in real car env", "Not in ASW-R1 Release Scope",
                  "Section 35", "only 8.4"]:
        blockers[probe] = sum(1 for r in rows if probe.lower() in lower(r["remarks"]))

    # quality defects in the procedure column
    banned = Counter()
    for v in BANNED_VERBS:
        banned[v] = sum(1 for r in rows if v in lower(r["test_procedure"]))
    vague = Counter()
    for v in VAGUE_WORDS:
        vague[v] = sum(1 for r in rows
                       if v in " ".join(lower(r[f]) for f in VAGUE_FIELDS))

    groups = Counter(norm(r["test_group"]) for r in rows if norm(r["test_group"]))
    sets_ = Counter(norm(r["test_set"]) for r in rows if norm(r["test_set"]))
    sets_blank = sum(1 for r in rows if not norm(r["test_set"]))

    exec_status = Counter()
    for r in rows:
        v = norm(r["_builds"][BUILD_COLS[EXEC_BUILD_COL]])
        exec_status[v or "(未執行)"] += 1

    # PCTS frozen region (R-P6): rows whose text mentions PCTS
    pcts_rows = [r["_row"] for r in rows
                 if re.search(r"PCTS", " ".join(norm(r[f]) for f in text_fields), re.I)]

    return {
        "sha256": sha256(WORKBOOK),
        "sheets": wb.sheetnames,
        "sheet": SHEET,
        "header_row": HEADER_ROW,
        "data_start_row": DATA_START,
        "data_rows": len(rows),
        "last_row": rows[-1]["_row"] if rows else None,
        "headers": headers,
        "fill": fill,
        "content_blank_rows": content_blank,
        "tokens": {"distinct": len(tok_rows), "rows_with_any": rows_with_token,
                   "by_token_rows": dict(tok_rows.most_common()),
                   "by_token_hits": dict(tok_hits.most_common())},
        "tools": tools,
        "blockers": blockers,
        "remarks_top": dict(remarks.most_common(12)),
        "banned_verbs": {k: v for k, v in banned.items() if v},
        "banned_verbs_total": sum(banned.values()),
        "vague_words": {k: v for k, v in vague.items() if v},
        "vague_words_total": sum(vague.values()),
        "test_groups": dict(groups.most_common()),
        "test_sets": dict(sets_.most_common()),
        "test_sets_blank": sets_blank,
        "exec_00_01_01_04": dict(exec_status.most_common()),
        "pcts_rows": pcts_rows,
        "req_ids": sorted({norm(r["req_id"]) for r in rows if norm(r["req_id"])}),
    }


def survey_swra():
    wb = openpyxl.load_workbook(SWRA, data_only=True)
    ws = wb["Basic Report"]
    leaves, headings = [], []
    src = Counter()
    hmi_src_filled = 0
    for r in range(2, ws.max_row + 1):
        rid = norm(ws.cell(r, 2).value)          # B  SEW1 SWE-Requirement ID
        typ = lower(ws.cell(r, 22).value)        # V  Type
        if typ == "heading" or not rid:
            if norm(ws.cell(r, 1).value):
                headings.append(norm(ws.cell(r, 1).value))
            continue
        leaves.append(rid)
        s = norm(ws.cell(r, 3).value)            # C  SWE1 Source Requirement ID
        if s:
            # Family = the id with its trailing numeric segment(s) removed, so
            # SYS-RA-PROJ-071 and SYS-RA-PROJ-198 collapse to SYS-RA-PROJ.
            src[re.sub(r"-\d+(-\d+)?$", "", s)] += 1
        if norm(ws.cell(r, 4).value):            # D  SWE1 HMI Source ID
            hmi_src_filled += 1
    return {
        "sha256": sha256(SWRA),
        "leaves": leaves,
        "leaf_count": len(leaves),
        "headings": len(headings),
        "hmi_source_id_filled": hmi_src_filled,
        "source_families": dict(src.most_common()),
    }


def parse_dbc(path: Path):
    """message -> [signals] and signal -> {value: label}. DBCs are ISO-8859."""
    txt = path.read_text(encoding="latin-1")
    msgs, cur = {}, None
    for line in txt.splitlines():
        m = re.match(r"^BO_ (\d+) ([^:]+):", line)
        if m:
            cur = m.group(2).strip()
            msgs[cur] = []
            continue
        m = re.match(r"^\s+SG_ ([A-Za-z0-9_]+)", line)
        if m and cur:
            msgs[cur].append(m.group(1))
    vals = {}
    for m in re.finditer(r"^VAL_ (\d+) ([A-Za-z0-9_]+) (.*?);\s*$", txt, re.M):
        pairs = re.findall(r'(-?\d+)\s+"([^"]*)"', m.group(3))
        vals.setdefault(m.group(2), {}).update({int(k): v for k, v in pairs})
    return msgs, vals


def survey_signals():
    """Verify every §5 signal claim against the DBCs, PROXI and mapping table.

    Nothing here is asserted from the 下放包 — each row is looked up and the
    lookup result is what lands in recon.json.
    """
    fd_m, fd_v = parse_dbc(INP / "PHDCC27_E2A_R1_FDCAN8.dbc")
    bh_m, bh_v = parse_dbc(INP / "PHDCC27_E2A_R1_BHCAN.dbc")

    def resolve(msg, sig):
        for bus, (m, v) in (("FD", (fd_m, fd_v)), ("CAN-B", (bh_m, bh_v))):
            if msg in m and sig in m[msg]:
                return {"bus": bus, "found": True, "values": v.get(sig, {})}
        return {"bus": None, "found": False, "values": {}}

    claims = {
        "BCM_FD_27.DAY_LGT_MD_DISP": ("BCM_FD_27", "DAY_LGT_MD_DISP"),
        "BCM_FD_27.DAY_LGT_MODE_DISP": ("BCM_FD_27", "DAY_LGT_MODE_DISP"),
        "TELEMATIC_FD_4.CurrentSource": ("TELEMATIC_FD_4", "CurrentSource"),
        "STATUS_TELEMATIC.CurrentSource": ("STATUS_TELEMATIC", "CurrentSource"),
        "STATUS_BH_BCM1.LowFuelWarningSts": ("STATUS_BH_BCM1", "LowFuelWarningSts"),
        "BCM_FD_14.Command_02Sts": ("BCM_FD_14", "Command_02Sts"),
        "HYBRID_DISPLAY.EstimatedRange": ("HYBRID_DISPLAY", "EstimatedRange"),
        "IPC_FD_5.Est_Range_Disp": ("IPC_FD_5", "Est_Range_Disp"),
        "HCP_CHARGING_STAT.Est_Range_FullCharge": ("HCP_CHARGING_STAT",
                                                   "Est_Range_FullCharge"),
    }

    # PROXI parameter table (Format sheet, column F = Parameter Name,
    # column I = the value table as free text).
    pwb = openpyxl.load_workbook(INP / "PROXI_HDCC27_R3_20250424.xlsx", data_only=True)
    pfs = pwb["Format"]
    proxi = {}
    for r in range(3, pfs.max_row + 1):
        name = norm(pfs.cell(r, 6).value)
        if name:
            proxi[name] = {"group": norm(pfs.cell(r, 1).value),
                           "coding": norm(pfs.cell(r, 8).value),
                           "table": norm(pfs.cell(r, 9).value)}

    # Mapping table, Atlantis High column block.
    mwb = openpyxl.load_workbook(INP / "Logical Identifiers and CAN Mapping v1_76.xlsx",
                                 data_only=True)
    # LID keys are lowercased: the mapping table writes VC_VEH_LINE in upper
    # case while the workbook tokens use three different casings of the same id.
    lids = {}
    for sheet, sig_col in (("CAN Mapping", 26), ("Proxi & Configuration", 16)):
        ws = mwb[sheet]
        for r in range(4, ws.max_row + 1):
            lid = lower(ws.cell(r, 1).value)
            if lid:
                lids.setdefault(lid, {"sheet": sheet,
                                      "signal": norm(ws.cell(r, sig_col).value),
                                      "can": norm(ws.cell(r, sig_col + 1).value),
                                      "format": norm(ws.cell(r, sig_col + 2).value)})

    return {
        "dbc": {
            "FDCAN8": {"messages": len(fd_m), "signals": sum(len(v) for v in fd_m.values()),
                       "value_tables": len(fd_v)},
            "BHCAN": {"messages": len(bh_m), "signals": sum(len(v) for v in bh_m.values()),
                      "value_tables": len(bh_v)},
            "message_overlap": len(set(fd_m) & set(bh_m)),
        },
        "signal_claims": {k: resolve(*v) for k, v in claims.items()},
        "proxi": {"parameter_rows": pfs.max_row - 2, "distinct_parameters": len(proxi),
                  "probed": {k: proxi.get(k) for k in
                             ["Radio_Display_Type", "Vehicle_Line_Configuration",
                              "Brand_Configuration_2", "Projection_Mode",
                              "Projection_Mode_Selection", "GPS_Presence",
                              "NAV_Presence", "Nav_Repetition",
                              "WiFi_2_BT_BLE_External_Antenna_Presence"]}},
        "mapping_lids": {k: lids.get(k.lower()) for k in
                         ["Day_Night_Mode", "VC_Veh_Brand", "VC_Veh_Line",
                          "HUModeStatus", "FuelLvlLow", "Screen_Size",
                          "Head_Unit_Screen_Size", "Est_Range_BEV"]},
    }


def survey_md_vs_cpaa():
    """A-PJ01: the two 037 sets, compared leaf by leaf."""
    cp = openpyxl.load_workbook(SWRA, data_only=True)["Basic Report"]
    md = openpyxl.load_workbook(SWRA_MD, data_only=True)["Analysis Report"]
    CP = {}
    for r in range(2, cp.max_row + 1):
        rid = norm(cp.cell(r, 2).value)
        if not rid or lower(cp.cell(r, 22).value) == "heading":
            continue
        CP[rid] = (norm(cp.cell(r, 6).value), norm(cp.cell(r, 20).value))
    MD = {}
    for r in range(9, md.max_row + 1):          # MD header row is 8
        rid = norm(md.cell(r, 1).value)
        if rid:
            MD[rid] = (norm(md.cell(r, 5).value), norm(md.cell(r, 35).value))
    inter = set(CP) & set(MD)
    qa = openpyxl.load_workbook(SWRA_MD, data_only=True)["QA"]
    qa_status = Counter(norm(qa.cell(r, 4).value)
                        for r in range(2, qa.max_row + 1) if norm(qa.cell(r, 1).value))
    return {
        "cpaa_leaves": len(CP), "md_leaves": len(MD), "intersection": len(inter),
        "cpaa_only": sorted(set(CP) - set(MD)), "md_only": sorted(set(MD) - set(CP)),
        "description_differs": sum(1 for k in inter if CP[k][0] != MD[k][0]),
        "verification_criteria_differs": sum(1 for k in inter if CP[k][1] != MD[k][1]),
        "cpaa_vc_empty": sum(1 for v in CP.values() if not v[1]),
        "md_vc_empty": sum(1 for v in MD.values() if not v[1]),
        "qa_rows": sum(qa_status.values()), "qa_status": dict(qa_status),
        "proj_133_in_cpaa": "SWE1-PROJ-133" in CP,
        "proj_133_in_md": "SWE1-PROJ-133" in MD,
    }


def main():
    wbk = survey_workbook()
    swra = survey_swra()
    signals = survey_signals()
    md = survey_md_vs_cpaa()

    leaf_set = set(swra["leaves"])
    wb_ids = set(wbk["req_ids"])
    covered = sorted(leaf_set & wb_ids)
    uncovered = sorted(leaf_set - wb_ids)
    overflow = sorted(wb_ids - leaf_set)

    out = {
        "workbook": wbk,
        "swra": swra,
        "signals": signals,
        "md_vs_cpaa": md,
        "coverage": {
            "leaves": swra["leaf_count"],
            "covered": len(covered),
            "uncovered": uncovered,
            "overflow": overflow,
        },
        "inputs_sha256": {p.name: sha256(p) for p in sorted(INP.iterdir()) if p.is_file()},
    }
    (FEAT / "data").mkdir(exist_ok=True)
    (FEAT / "data" / "recon.json").write_text(
        json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"data rows            {wbk['data_rows']}  (rows {DATA_START}-{wbk['last_row']})")
    print(f"header match         {sum(1 for h in wbk['headers'].values() if h['match'])}"
          f"/{len(wbk['headers'])}")
    print(f"leaves               {swra['leaf_count']}  headings {swra['headings']}")
    print(f"covered              {len(covered)}")
    print(f"uncovered            {len(uncovered)}  {uncovered}")
    print(f"overflow             {len(overflow)}  {overflow[:10]}")
    print(f"HMI Source ID filled {swra['hmi_source_id_filled']}/{swra['leaf_count']}")
    print(f"source families      {swra['source_families']}")
    print("fill:", {k: v for k, v in wbk["fill"].items()})
    print(f"content-blank rows   {len(wbk['content_blank_rows'])} {wbk['content_blank_rows']}")
    print("tokens:", wbk["tokens"]["distinct"], "distinct /",
          wbk["tokens"]["rows_with_any"], "rows")
    print("  ", wbk["tokens"]["by_token_rows"])
    print("tools:", wbk["tools"])
    print("blockers:", wbk["blockers"])
    print("banned verbs:", wbk["banned_verbs"], "total", wbk["banned_verbs_total"])
    print("vague:", wbk["vague_words"], "total", wbk["vague_words_total"])
    print("groups:", wbk["test_groups"])
    print("sets:", wbk["test_sets"], "blank", wbk["test_sets_blank"])
    print("exec 00.01.01.04:", wbk["exec_00_01_01_04"])
    print("PCTS rows:", len(wbk["pcts_rows"]))
    print("dbc:", signals["dbc"])
    for k, v in signals["signal_claims"].items():
        print(f"  {'OK ' if v['found'] else 'MISS'} {k:42s} {v['bus'] or '-':6s}"
              f" vals={list(v['values'].items())[:4]}")
    print("proxi params:", signals["proxi"]["distinct_parameters"],
          "missing:", [k for k, v in signals["proxi"]["probed"].items() if v is None])
    print("mapping LIDs:", {k: (v or {}).get("signal") for k, v in
                            signals["mapping_lids"].items()})
    print("md vs cpaa:", {k: v for k, v in md.items() if k not in ("cpaa_only", "md_only")})


if __name__ == "__main__":
    main()
