#!/usr/bin/env python3
"""Write the pilot batch into the 036 workbook, zip-surgically.

R-BLM3 / R-G1: the master's R-column design_method dropdown is an x14
extension. `openpyxl.save()` drops it while row contents stay correct and
lint stays green -- so openpyxl is used here ONLY as the calculation layer,
and `backend.xlsx_surgical.surgical_save` does the emit: it diffs the mutated
workbook against a fresh read of the source and patches the SOURCE sheet XML
as text, copying every other zip member byte for byte.

`verify_structure` (called inside `surgical_save`) raises rather than warns
on: a changed zip member set, changed classic/x14 data-validation counts, or
any member differing that was not a patch target.

After writing, `--verify` reads the 13 rows back out of the delivered file
and compares every field against `pilot_tcs.json`. A write that cannot be
read back is not a write.
"""
import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))
sys.path.insert(0, str(Path(__file__).resolve().parents[3] / "scripts"))

import openpyxl                                                    # noqa: E402
from backend.xlsx_surgical import surgical_save                    # noqa: E402
from feature_config import load_feature_config, resolve_path       # noqa: E402

FEAT = Path(__file__).resolve().parents[1]
DEFAULT_BATCH = FEAT / "batches" / "pilot" / "pilot_tcs.json"

# 工作簿常數欄。功能安全欄全案填 "NA"（本 feature 無 FuSa 需求，
# 037 之 Categorization 亦無安全相關標記）。
# R-BLM16(2)(3)：兩欄一律留空，隨交付多數。
# 原填 "NA" / "1.0" 係上繳 07 §二-5 之造值 —— 全案四本 775 列實測，
# AB 欄無一列填過，S 欄僅 privacy 之 11 列填 NA（764/775 為空）。
# 空字典保留於此而非刪除，使「兩欄刻意不寫」這件事在程式裡看得見。
CONST: dict[str, str] = {}


def cell_values(tc: dict, cfg: dict, tc_id: str, test_set: str) -> dict[int, object]:
    col = cfg["col"]
    wb_cfg = cfg["write_back"]
    fill = wb_cfg.get("fill_test_group_set", False)
    v = {
        col["req_id"]: tc["req_id"],
        col["tc_id"]: tc_id,
        col["test_group"]: cfg["test_group"] if fill else None,
        # test_set 取**批次層**之值，不取 tc.get("test_set")。
        # `pilot_tcs.json` 之 test_set 只存在於批次層，逐條 TC 無此鍵 ——
        # 原寫法 `tc.get("test_set")` 靜默回 None，H 欄 13 列全空，
        # 而 round-trip 驗證用同一個 cell_values 產生期望值，
        # 於是「沒寫」與「該寫什麼」兩邊一致，比對全綠。
        # 抓到它的是讀工作簿的交付 lint（gate A-empty），不是 round-trip。
        col["test_set"]: test_set if fill else None,
        col["test_item"]: tc["tc_title"],
        col["pre_conditions"]: tc["pre_conditions"],
        col["input_test_data"]: tc["input_test_data"],
        col["test_procedure"]: tc["test_procedure"],
        col["expected_result"]: tc["expected_result"],
        col["spec_reference"]: tc["specification_reference"],
        col["tc_ref_id"]: wb_cfg["tc_ref_id_value"],
        col["priority"]: tc["priority"],
        col["design_method"]: tc["design_method"],
        col["author"]: wb_cfg["author_value"],
    }
    return {k: val for k, val in v.items() if val is not None}


def first_free_row(ws, cfg: dict) -> int:
    """First row at or below the header whose req_id / tc_id are both empty."""
    col = cfg["col"]
    r = cfg["workbook"]["header_row"] + 1
    while True:
        a = ws.cell(r, col["req_id"] + 1).value
        b = ws.cell(r, col["tc_id"] + 1).value
        if not (a or b):
            return r
        r += 1


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(FEAT / "workbook" / "bed_lowering_01.xlsx"))
    ap.add_argument("--verify", action="store_true")
    ap.add_argument("--batch", default=str(DEFAULT_BATCH))
    ap.add_argument("--src", help="來源工作簿；預設取 feature.yaml paths.workbook")
    # IN §8.4.3：含 PENDING 之工作簿不得出貨。故帶 PENDING 之列不寫回，
    # 停在 batch json 待 DR 回覆。此旗標使該過濾**顯式且可回報**，
    # 而不是靠寫回時忘記處理。
    ap.add_argument("--skip-pending", action="store_true")
    ap.add_argument("--start-id", type=int, default=1)
    # 覆寫既有列（修訂既寫回之批次）。不給則自首個空列往下追加。
    ap.add_argument("--start-row", type=int)
    a = ap.parse_args()

    cfg = load_feature_config(FEAT)
    src = Path(a.src) if a.src else resolve_path(cfg, "workbook")
    batch = json.loads(Path(a.batch).read_text(encoding="utf-8"))
    all_tcs = batch["tcs"]
    test_set = batch["test_set"]
    if a.skip_pending:
        held = [t["req_id"] for t in all_tcs if t.get("has_pending")]
        tcs = [t for t in all_tcs if not t.get("has_pending")]
        print(f"IN §8.4.3 保留不寫回 {len(held)} 條: {held}")
    else:
        tcs = all_tcs
    fmt = cfg["tc_id_format"]

    wb = openpyxl.load_workbook(src)
    ws = wb[cfg["workbook"]["sheet"]]
    start = a.start_row or first_free_row(ws, cfg)
    print(f"來源 {src.name}")
    print(f"起始列 {start}"
          + ("（--start-row 指定，覆寫既有列）" if a.start_row
             else f"（首個空列，表頭第 {cfg['workbook']['header_row']} 列）"))

    written = []
    for i, tc in enumerate(tcs):
        tc_id = fmt.format(n=a.start_id + i)
        row = start + i
        for cidx, val in cell_values(tc, cfg, tc_id, test_set).items():
            ws.cell(row, cidx + 1).value = val
        written.append((row, tc_id, tc["req_id"]))

    out = Path(a.out)
    report = surgical_save(wb, src, out)
    print(f"\n寫入 {out}")
    print(f"  sheets_patched  {report['sheets_patched']}")
    print(f"  zip members     {report['members']}")
    print(f"  differing       {report['differing']}")
    print(f"  dv_counts       {report['dv_counts']}")

    print("\n列對映：")
    for row, tc_id, rid in written:
        print(f"  列 {row:>3}  {tc_id}  {rid}")

    if a.verify:
        print("\n--- round-trip 讀回比對 ---")
        rb = openpyxl.load_workbook(out, data_only=True)
        rws = rb[cfg["workbook"]["sheet"]]
        col = cfg["col"]
        bad = 0
        for i, tc in enumerate(tcs):
            row = start + i
            want = cell_values(tc, cfg, fmt.format(n=a.start_id + i), test_set)
            for cidx, val in want.items():
                got = rws.cell(row, cidx + 1).value
                if str(got if got is not None else "") != str(val):
                    bad += 1
                    name = [k for k, v in col.items() if v == cidx][0]
                    print(f"  **差異** 列{row} {name}: 寫入{val!r} 讀回{got!r}")
        print(f"  比對 {len(tcs)} 列 × {len(want)} 欄，差異 {bad}")
        if bad:
            return 1
        print("  round-trip PASS —— 差異 0")
    return 0


if __name__ == "__main__":
    sys.exit(main())
