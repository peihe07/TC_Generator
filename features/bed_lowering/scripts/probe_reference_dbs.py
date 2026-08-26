#!/usr/bin/env python3
"""Open every bound reference database and report what is actually inside.

Upstream 03 §5.2 disclosed that the four were hashed but never opened, so
"the signal will be findable once bound" was an expectation, not a
measurement. This closes that: a sha256 proves which bytes, never that the
bytes can be parsed.
"""
import re
import sys
from pathlib import Path

import yaml

ROOT = Path(__file__).resolve().parents[3]
CFG = Path(__file__).resolve().parents[1] / "feature.yaml"

# DBC grammar, per the format's own line syntax:
#   BO_ <id> <Name>: <dlc> <transmitter>
#   SG_ <Name> : <start>|<len>@<order><sign> (<factor>,<offset>) ...
#   VAL_ <msgid> <signal> <raw> "<label>" ... ;
BO = re.compile(r"^BO_\s+(\d+)\s+([A-Za-z0-9_]+)\s*:", re.M)
SG = re.compile(r"^\s*SG_\s+([A-Za-z0-9_]+)\s*(?:[Mm]\d*\s*)?:", re.M)
VAL = re.compile(r"^VAL_\s+(\d+)\s+([A-Za-z0-9_]+)\s+(.*?);", re.M | re.S)


def probe_dbc(path: Path) -> dict:
    raw = path.read_bytes()
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise SystemExit(f"{path.name}: undecodable")
    msgs = BO.findall(text)
    sigs = SG.findall(text)
    vals = VAL.findall(text)
    return {"encoding": enc, "lines": text.count("\n") + 1,
            "messages": len(msgs), "signals": len(sigs),
            "distinct_signals": len(set(sigs)), "VAL_ tables": len(vals),
            "text": text, "msg_names": [m[1] for m in msgs],
            "sig_names": sigs}


def probe_xlsx(path: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    info = {"sheets": wb.sheetnames, "sheet_rows": {}}
    for name in wb.sheetnames:
        n = sum(1 for _ in wb[name].iter_rows(values_only=True))
        info["sheet_rows"][name] = n
    wb.close()
    return info


def main() -> int:
    cfg = yaml.safe_load(CFG.read_text(encoding="utf-8"))
    ref = cfg["reference"]
    for key in ("dbc_b", "dbc_fd"):
        p = ROOT / ref[key]["file"]
        r = probe_dbc(p)
        print(f"## {key} — {p.name}")
        print(f"   encoding {r['encoding']}, {r['lines']} lines")
        print(f"   BO_ messages      {r['messages']}")
        print(f"   SG_ signals       {r['signals']} ({r['distinct_signals']} distinct)")
        print(f"   VAL_ tables       {r['VAL_ tables']}")
        print()
    for key in ("lid", "proxi"):
        p = ROOT / ref[key]["file"]
        r = probe_xlsx(p)
        print(f"## {key} — {p.name}")
        for s in r["sheets"]:
            print(f"   sheet {s!r:<46} rows {r['sheet_rows'][s]}")
        print()
    return 0


if __name__ == "__main__":
    sys.exit(main())
