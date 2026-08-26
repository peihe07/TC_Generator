#!/usr/bin/env python3
"""Delivery lint for bed_lowering — reads the WORKBOOK, not the batch JSON.

This is deliberately a different program from `selfcheck_pilot.py`. That one
reads `pilot_tcs.json` and answers "did I author this correctly". This one
opens the delivered xlsx and answers "is what is actually in the customer's
file correct". Upstream 04 §八-2 recorded the gap between the two for three
rounds; this closes it.

Ported from `features/amfm/scripts/lint_tcs.py` (same shape: read workbook via
openpyxl read_only, gate list returning (gate, message), --json-report).
Feature-specific gates: N-column constant (R-BLM5), and the forbidden-verb
gate matches ANYWHERE in the line rather than at line start -- upstream 05
found two `and observe that` occurrences that a line-start pattern missed.

Exit 0 clean / 1 findings.
"""
import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[3] / "scripts"))

import openpyxl                                                    # noqa: E402
from feature_config import load_feature_config, resolve_path       # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from make_batch_context import _spec_reference_constant             # noqa: E402

FEAT = Path(__file__).resolve().parents[1]

FORBIDDEN = re.compile(
    r"\b(observe|observes|observing|see if|check whether|confirm whether|"
    r"watch|monitor|inspect)\b", re.I)
VERIFY_MAIN = re.compile(r"^\s*\d+\.\s*verify\b", re.I)
MODALS = re.compile(r"\b(shall|will|should|would)\b", re.I)
CJK = re.compile(r"[一-鿿]")
PRIORITIES = {"P0", "P1", "P2", "P3"}
METHODS = {"Negative / Invalid", "Fault Injection", "State Transition",
           "Decision Table", "Equivalence Partitioning",
           "Boundary Value Analysis", "Combinatorial", "Scenario / Use Case",
           "Functional Based"}
TEXT_COLS = ["pre_conditions", "input_test_data", "test_procedure",
             "expected_result"]


def rows_of(ws, cfg):
    col = cfg["col"]
    r = cfg["workbook"]["header_row"] + 1
    out = []
    while True:
        rid = ws.cell(r, col["req_id"] + 1).value
        tcid = ws.cell(r, col["tc_id"] + 1).value
        if not (rid or tcid):
            break
        out.append((r, {k: ws.cell(r, v + 1).value for k, v in col.items()}))
        r += 1
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook")
    ap.add_argument("--json-report")
    a = ap.parse_args()

    cfg = load_feature_config(FEAT)
    wbp = Path(a.workbook) if a.workbook else resolve_path(cfg, "workbook")
    wb = openpyxl.load_workbook(wbp, read_only=True, data_only=True)
    ws = wb[cfg["workbook"]["sheet"]]
    rows = rows_of(ws, cfg)
    # 取 037 `HMI Source ID` 欄之原值，不取任何宣告。
    # 原寫法讀 `recon_assertions["spec_reference_stem"]`，該鍵已由
    # R-BLM15(1) 作廢 —— 改讀另一個宣告只會把同一個錯誤搬家，
    # 故與 adapter 共用同一支「讀來源欄並斷言唯一」之函式。
    const = _spec_reference_constant()

    findings: list[tuple[str, str]] = []

    def bad(gate, msg):
        findings.append((gate, msg))

    lowers: dict[str, list[str]] = {}
    for r, c in rows:
        tag = f"row{r} {c['req_id']}"

        # A. 十鍵齊備（工作簿側之對應欄非空）
        for k in ["req_id", "tc_id", "test_group", "test_set", "test_item",
                  "pre_conditions", "test_procedure", "expected_result",
                  "spec_reference", "priority", "design_method", "author"]:
            if not str(c.get(k) or "").strip():
                bad("A-empty", f"{tag}: {k} is empty")

        # B. test_item 兩段式 + 下半語言
        ti = str(c.get("test_item") or "")
        parts = ti.split("\n")
        if len(parts) < 2:
            bad("B-testitem", f"{tag}: test_item has no bracketed lower half")
        else:
            lower = parts[-1].strip()
            if not (lower.startswith("(") and lower.endswith(")")):
                bad("B-testitem", f"{tag}: lower half not wrapped in ( ): {lower!r}")
            if CJK.search(lower):
                bad("B-lang", f"{tag}: lower half contains CJK: {lower!r}")
            head = str(c["req_id"]).rsplit("-", 1)[0]
            lowers.setdefault(head, []).append(lower)

        # C. 尾句號 / 行首尾空白 / 方括號 / 單引號
        for k in TEXT_COLS:
            v = str(c.get(k) or "")
            for line in v.split("\n"):
                if not line.strip():
                    continue
                if line != line.strip():
                    bad("C-space", f"{tag}/{k}: leading or trailing space: {line!r}")
                if line.rstrip().endswith((".", "。")):
                    bad("C-period", f"{tag}/{k}: trailing period: {line!r}")
                if "[" in line or "]" in line:
                    bad("C-bracket", f"{tag}/{k}: square bracket: {line!r}")
                if re.search(r"(?<!\w)'[^']+'(?!\w)", line):
                    bad("C-quote", f"{tag}/{k}: single-quoted label: {line!r}")

        # D. 禁用主動詞 —— 全行比對，不限行首（上繳 05 之教訓）
        for line in str(c.get("test_procedure") or "").split("\n"):
            if FORBIDDEN.search(line) or VERIFY_MAIN.match(line):
                bad("D-verb", f"{tag}: forbidden verb: {line!r}")

        # E. ER 情態動詞 + 1:1
        proc = [l for l in str(c.get("test_procedure") or "").split("\n") if l.strip()]
        er = [l for l in str(c.get("expected_result") or "").split("\n") if l.strip()]
        if len(proc) != len(er):
            bad("E-pairing", f"{tag}: procedure {len(proc)} vs ER {len(er)}")
        if len(proc) < 2:
            bad("E-minsteps", f"{tag}: fewer than 2 procedure steps")
        for line in er:
            if MODALS.search(line):
                bad("E-modal", f"{tag}: modal verb in ER: {line!r}")

        # F. N 欄常數（R-BLM5）
        if str(c.get("spec_reference") or "").strip() != const:
            bad("F-specref", f"{tag}: spec_reference != ruled constant")

        # G. 值域
        if str(c.get("priority") or "") not in PRIORITIES:
            bad("G-priority", f"{tag}: priority {c.get('priority')!r}")
        if str(c.get("design_method") or "") not in METHODS:
            bad("G-method", f"{tag}: design_method {c.get('design_method')!r}")
        if str(c.get("test_group") or "") != cfg["test_group"]:
            bad("G-group", f"{tag}: test_group {c.get('test_group')!r}")

    # H. sibling 下半互異
    for head, ls in lowers.items():
        if len(set(ls)) != len(ls):
            bad("H-sibling", f"{head}: duplicate lower halves")

    print(f"workbook {wbp.name}")
    print(f"rows     {len(rows)}")
    print(f"gates    A-empty B-testitem B-lang C-space C-period C-bracket "
          f"C-quote D-verb E-pairing E-minsteps E-modal F-specref "
          f"G-priority G-method G-group H-sibling")
    if findings:
        print(f"\nFINDINGS {len(findings)}")
        for g, m in findings:
            print(f"  [{g}] {m}")
    else:
        print("\nclean — 0 findings")

    if a.json_report:
        Path(a.json_report).write_text(json.dumps(
            {"workbook": str(wbp), "rows": len(rows),
             "findings": [{"gate": g, "message": m} for g, m in findings]},
            ensure_ascii=False, indent=1), encoding="utf-8")
    return 1 if findings else 0


if __name__ == "__main__":
    sys.exit(main())
