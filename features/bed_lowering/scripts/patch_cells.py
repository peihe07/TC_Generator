#!/usr/bin/env python3
"""Surgically clear or set individual cells in an already-written workbook.

R-BLM16(2)(3) needs two whole columns emptied across rows that are already
delivered. Re-running write_back would rebuild every cell from the batch
json; this touches only the named cells, so a mistake elsewhere in the batch
cannot ride along with the fix.

Emit goes through `backend.xlsx_surgical.surgical_save` for the same reason
write_back does: openpyxl's save() would drop the x14 dropdown.
"""
import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))
sys.path.insert(0, str(Path(__file__).resolve().parents[3] / "scripts"))

import openpyxl                                                    # noqa: E402
from backend.xlsx_surgical import surgical_save                    # noqa: E402
from feature_config import load_feature_config                     # noqa: E402

FEAT = Path(__file__).resolve().parents[1]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--rows", required=True, help="例 10-28")
    ap.add_argument("--clear", default="", help="逗號分隔之 feature.yaml 欄名")
    # 上繳 07 §四-4 具名之缺口：本工具原本只能清空。
    # --set-from 自 batch json 取指名 req_id 之指名欄，覆寫其既有列。
    # 仍只碰指名儲存格 —— 不重跑 write_back，故 batch 內他處之改動
    # 不會搭順風車混進本次修訂。
    ap.add_argument("--set-from", help="batch json 路徑")
    ap.add_argument("--req-ids", default="", help="逗號分隔之 req_id")
    ap.add_argument("--fields", default="", help="逗號分隔之 feature.yaml 欄名")
    a = ap.parse_args()

    cfg = load_feature_config(FEAT)
    col = cfg["col"]
    lo, hi = (int(x) for x in a.rows.split("-"))
    names = [n.strip() for n in a.clear.split(",") if n.strip()]
    for n in names:
        if n not in col:
            sys.exit(f"unknown column name {n!r}; known: {sorted(col)}")

    wb = openpyxl.load_workbook(Path(a.src))
    ws = wb[cfg["workbook"]["sheet"]]

    setlog = []
    if a.set_from:
        import json as _json
        batch = _json.loads(Path(a.set_from).read_text(encoding="utf-8"))
        by = {t["req_id"]: t for t in batch["tcs"]}
        want = [x.strip() for x in a.req_ids.split(",") if x.strip()]
        flds = [x.strip() for x in a.fields.split(",") if x.strip()]
        KEY = {"test_procedure": "test_procedure", "expected_result": "expected_result",
               "pre_conditions": "pre_conditions", "test_item": "tc_title",
               "input_test_data": "input_test_data"}
        # 以 req_id 定位既有列，不靠列號推算 —— 列號會隨批次順序而錯
        rmap = {}
        for r in range(cfg["workbook"]["header_row"] + 1, ws.max_row + 1):
            v = ws.cell(r, col["req_id"] + 1).value
            if v:
                rmap[str(v)] = r
        for rid in want:
            if rid not in rmap:
                sys.exit(f"{rid} not found in the workbook")
            if rid not in by:
                sys.exit(f"{rid} not in {a.set_from}")
            for fl in flds:
                old = ws.cell(rmap[rid], col[fl] + 1).value
                new = by[rid][KEY[fl]]
                ws.cell(rmap[rid], col[fl] + 1).value = new
                setlog.append((rmap[rid], rid, fl, old != new))

    before = {}
    for n in names:
        vals = [ws.cell(r, col[n] + 1).value for r in range(lo, hi + 1)]
        before[n] = vals
        for r in range(lo, hi + 1):
            ws.cell(r, col[n] + 1).value = None

    report = surgical_save(wb, Path(a.src), Path(a.out))
    for row, rid, fl, changed in setlog:
        print(f"  set 列{row} {rid} {fl}: {'已改' if changed else '無變化'}")
    print(f"來源 {Path(a.src).name} -> {Path(a.out).name}")
    print(f"列範圍 {lo}-{hi}（{hi - lo + 1} 列）")
    for n in names:
        seen = sorted({str(v) for v in before[n] if v not in (None, "")})
        print(f"  清空 {n}（欄 {cfg['workbook']['columns'][n]}）"
              f" 原值 {seen or '(已為空)'}")
    print(f"  sheets_patched {report['sheets_patched']}")
    print(f"  zip members {report['members']}  differing {report['differing']}")
    print(f"  dv_counts {report['dv_counts']}")

    rb = openpyxl.load_workbook(Path(a.out), data_only=True)
    rws2 = rb[cfg["workbook"]["sheet"]]
    bad2 = 0
    for row, rid, fl, _ in setlog:
        import json as _json
        batch = _json.loads(Path(a.set_from).read_text(encoding="utf-8"))
        by = {t["req_id"]: t for t in batch["tcs"]}
        KEY = {"test_procedure": "test_procedure", "expected_result": "expected_result",
               "pre_conditions": "pre_conditions", "test_item": "tc_title",
               "input_test_data": "input_test_data"}
        if rws2.cell(row, col[fl] + 1).value != by[rid][KEY[fl]]:
            bad2 += 1
            print(f"  **差異** 列{row} {rid} {fl} 讀回不符")
    if setlog:
        print(f"  set round-trip：{len(setlog)} 格，不符 {bad2}")
    rws = rb[cfg["workbook"]["sheet"]]
    bad = 0
    for n in names:
        for r in range(lo, hi + 1):
            if rws.cell(r, col[n] + 1).value not in (None, ""):
                bad += 1
                print(f"  **差異** 列{r} {n} 未清空")
    print(f"  round-trip 讀回：{len(names) * (hi - lo + 1)} 格，未清空 {bad}")
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())
