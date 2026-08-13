#!/usr/bin/env python3
"""P7 write-back safety probe for the FM-WI-FSM-036-A01 template family.

Answers one question with evidence, not opinion: what does a write path cost
the workbook? Compares two strategies against the untouched original —

1. `openpyxl` load → mutate → save (the path that produced AMFM v1)
2. `backend.xlsx_surgical.surgical_save` — rewrite only the target sheet's
   XML, copy every other zip member byte-for-byte (the path R16-1 mandates)

The surgical arm calls the PRODUCTION module, not a parallel implementation:
a probe that tests its own copy of the logic proves nothing about the writer
that ships.

Measured 2026-08-13:
- Privacy blank template (A-PV09): openpyxl loses 2 x14 DV groups, 5 printer
  settings, the VML comment layer, sharedStrings; JPEG re-encoded to PNG
- AMFM customer original (R16 §2): openpyxl loses 21 members including the
  whole `xl/diagrams/` SmartArt set, adds 10, and takes x14 DV to zero

Both are LOSSLESS under the surgical path.

Usage:
    python features/privacy/scripts/xlsx_roundtrip_probe.py \
        --workbook <path.xlsx> [--sheet-name <sheet>] [--tmp <dir>]

Exit code 0 = surgical LOSSLESS, 1 = it is not (R16 §4 stop condition 2:
never lower the standard to make this pass).
"""

import argparse
import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl

REPO_ROOT = next(p for p in Path(__file__).resolve().parents
                 if (p / "pyproject.toml").is_file())
sys.path.insert(0, str(REPO_ROOT))
from backend.xlsx_surgical import sheet_members, surgical_save  # noqa: E402

X14_SQREF = re.compile(r"<xm:sqref>([^<]+)</xm:sqref>")
X14_ELEM = re.compile(r"<x14:dataValidation[ >]")
CLASSIC_DV = re.compile(r'<dataValidation [^>]*sqref="([^"]+)"')

PROBE_TEXT = "round-trip probe"


def probe(path: Path, sheet_xml: str) -> dict:
    """Structural fingerprint: DV coverage (both flavours) + zip inventory."""
    with zipfile.ZipFile(path) as z:
        xml = z.read(sheet_xml).decode("utf-8")
        return {
            "x14_groups": X14_SQREF.findall(xml),
            "x14_elements": len(X14_ELEM.findall(xml)),
            "classic_dv": CLASSIC_DV.findall(xml),
            "members": set(z.namelist()),
        }


def pick_target(ws) -> str:
    """First empty cell on the sheet's data area — probing must not overwrite
    customer content, and an overwrite would also mask a no-op write."""
    merged = {c for rng in ws.merged_cells.ranges for c in rng.cells}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 400)):
        for cell in row:
            # MergedCell has no writable value; skip the whole merged range.
            if cell.value is None and (cell.row, cell.column) not in merged:
                return cell.coordinate
    raise SystemExit("no empty cell found to probe with")


def report(tag: str, base: dict, after: dict) -> bool:
    lost_groups = set(base["x14_groups"]) - set(after["x14_groups"])
    lost_dv = set(base["classic_dv"]) - set(after["classic_dv"])
    dropped = sorted(base["members"] - after["members"])
    added = sorted(after["members"] - base["members"])
    elem_delta = (base["x14_elements"], after["x14_elements"])
    ok = not (lost_groups or lost_dv or dropped or added
              or elem_delta[0] != elem_delta[1])
    print(f"\n--- {tag}: {'LOSSLESS' if ok else 'LOSSY'}")
    print(f"  x14 DV elements  : {elem_delta[0]} -> {elem_delta[1]}")
    print(f"  x14 DV groups lost: {sorted(lost_groups) or 'none'}")
    print(f"  classic DV lost  : {sorted(lost_dv) or 'none'}")
    print(f"  zip members lost : {dropped or 'none'}")
    print(f"  zip members added: {added or 'none'}")
    return ok


def compare_delivered(source: Path, delivered: Path, sheet_name: str | None) -> int:
    """R16-3 retrospective: measure an ALREADY DELIVERED file against the
    customer original. Same fingerprint as the forward probe, so a past
    delivery and a candidate delivery are judged by one definition."""
    wb = openpyxl.load_workbook(source)
    name = sheet_name or next(
        (n for n in wb.sheetnames if "Test Case Specification" in n), wb.sheetnames[0])
    sheet_xml = sheet_members(source)[name]
    base = probe(source, sheet_xml)
    try:
        after = probe(delivered, sheet_xml)
    except KeyError:
        print(f"  {sheet_xml} absent from the delivered file — structure "
              "diverged beyond comparison")
        return 1

    print(f"source    : {source.name}  ({source.stat().st_size:,} B, "
          f"{len(base['members'])} members)")
    print(f"delivered : {delivered.name}  ({delivered.stat().st_size:,} B, "
          f"{len(after['members'])} members)")
    print(f"sheet     : {name!r} -> {sheet_xml}")
    ok = report("delivered vs customer original", base, after)
    print(f"  classic DV count : {len(base['classic_dv'])} -> "
          f"{len(after['classic_dv'])}")
    return 0 if ok else 1


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", required=True, type=Path)
    ap.add_argument("--sheet-name", help="default: the sheet carrying the TC table")
    ap.add_argument("--compare", type=Path,
                    help="R16-3 mode: measure this already-delivered file "
                         "against --workbook instead of running the two "
                         "write strategies")
    ap.add_argument("--tmp", type=Path, default=Path("features/privacy/data/roundtrip"))
    args = ap.parse_args()

    if args.compare:
        return compare_delivered(args.workbook, args.compare, args.sheet_name)

    args.tmp.mkdir(parents=True, exist_ok=True)
    original = args.tmp / "original.xlsx"
    shutil.copy(args.workbook, original)

    wb = openpyxl.load_workbook(original)
    sheet_name = args.sheet_name or next(
        (n for n in wb.sheetnames if "Test Case Specification" in n), wb.sheetnames[0])
    sheet_xml = sheet_members(original)[sheet_name]
    target = pick_target(wb[sheet_name])
    print(f"workbook : {args.workbook.name}")
    print(f"sheet    : {sheet_name!r} -> {sheet_xml}")
    print(f"probe cell: {target} (empty in the source)")

    base = probe(original, sheet_xml)
    print(f"baseline : {len(base['members'])} zip members, "
          f"x14 DV {base['x14_elements']} element(s) in {len(base['x14_groups'])} "
          f"group(s), classic DV {base['classic_dv']}")

    # --- strategy 1: openpyxl load/save
    via_openpyxl = args.tmp / "via_openpyxl.xlsx"
    wb1 = openpyxl.load_workbook(original)
    wb1[sheet_name][target] = PROBE_TEXT
    wb1.save(via_openpyxl)
    ok_openpyxl = report("openpyxl load/save", base, probe(via_openpyxl, sheet_xml))

    # --- strategy 2: the production surgical path
    via_surgical = args.tmp / "via_surgical.xlsx"
    wb2 = openpyxl.load_workbook(original)
    wb2[sheet_name][target] = PROBE_TEXT
    surgical_save(wb2, original, via_surgical)
    ok_surgical = report("zip-level surgical splice", base,
                         probe(via_surgical, sheet_xml))

    readback = openpyxl.load_workbook(via_surgical)[sheet_name][target].value
    applied = readback == PROBE_TEXT
    print(f"\nwrite landed: {target} = {readback!r} "
          f"({'OK' if applied else 'NOT WRITTEN — probe proved nothing'})")

    if ok_surgical and applied:
        print("\nverdict: surgical path is LOSSLESS and the write landed"
              + ("; openpyxl path is LOSSY" if not ok_openpyxl else ""))
        return 0
    print("\nverdict: FAILED — do not lower the standard (R16 §4.2)")
    return 1


if __name__ == "__main__":
    sys.exit(main())
