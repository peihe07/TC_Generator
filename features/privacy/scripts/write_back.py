#!/usr/bin/env python3
"""Step 4 (Privacy) — write the generated TCs into the FW036 workbook.

Built on `backend.xlsx_surgical` from the first line (R20-5). The four existing
feature `write_back.py` scripts are quarantined (R20-3) and were **not** copied:
each of them ends in `wb.save(out)`, and copying one would have inherited the
openpyxl save path that cost AMFM v1 twenty-one zip members and both x14
dropdowns. openpyxl is used here only to compute cell values; the file is
emitted by splicing the target sheet's XML into a byte-for-byte copy of the
input (R18-3 rule 1).

Privacy's shape is simpler than AMFM's and stricter than it looks:

- **BLANK workbook** — no legacy region, no done region, nothing to freeze.
  The invariant is therefore not "the frozen rows must not move" but "nothing
  above row 10 may change and every non-target sheet must come out identical".
- **Rows 10–11 were cleared** of the template's residual sample under R23-4, so
  the first generated TC lands on row 10 and column B's formula
  `=IF(ISBLANK($D{r}),"",ROW()-9)` numbers it 1.
- **One BLOCKED row** (-008, R34-3). It is written, not skipped: a leaf missing
  from the deliverable leaves an unexplained hole in the traceability table,
  and the marker in Remarks is what makes the hole auditable.

Column policy, all ruled, none inferred:

| column | value | ruling |
|---|---|---|
| G Test Group | `Privacy` | framework Part VI |
| H Test Set | per framework Part VI | R25-1 |
| Q Estimated Test Time | blank (`UNRULED_BLANK`) | profile §3.7 |
| S Functional Safety | `NA` | R30-3 |
| T–Z Vehicle Model | blank | R30-4 |
| AA Author | `PeiPYHsu` | feature.yaml |
| AH Remarks | empty except the BLOCKED row | R34-3 |

Usage:
    python features/privacy/scripts/write_back.py --feature-dir features/privacy
    python features/privacy/scripts/write_back.py --feature-dir features/privacy --write

Exit code 0 = ok, 1 = an invariant failed, 2 = bad invocation.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

import openpyxl
import yaml

REPO_ROOT = next(p for p in Path(__file__).resolve().parents
                 if (p / "pyproject.toml").is_file())
sys.path.insert(0, str(REPO_ROOT))
from backend.xlsx_surgical import (  # noqa: E402
    StructureError, surgical_save, verify_structure)

FIRST_DATA_ROW = 10
CONST_FUNCTIONAL_SAFETY = "NA"          # R30-3
PLACEHOLDER_BODY = "BLOCKED - see Remarks"
ROW_NUMBER_FORMULA = '=IF(ISBLANK($D{row}),"",ROW()-9)'
# Columns deliberately left blank, listed so a reviewer sees the omission is a
# decision rather than an oversight.
BLANK_BY_DECISION = {
    "C (Polarion ID)": "no Polarion export for this feature",
    "E (TestRail ID)": "assigned downstream",
    "O (Test Case Reference ID)": "feature.yaml write_back.tc_ref_id_value",
    "Q (Estimated Test Time)": "UNRULED_BLANK — profile §3.7",
    "T–Z (Vehicle Model)": "R30-4 — AMFM precedent is 0/158",
}


class WriteBackError(RuntimeError):
    pass


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ------------------------------------------------------------------ columns

def resolve_columns(ws, header_row: int) -> dict[str, int]:
    """Header TEXT decides the column, not feature.yaml's letters.

    feature.yaml still carries pre-revision-C letters (A-PV13); recon reports
    the drift rather than silently honouring either side. Resolving here from
    the instance in front of us is the same discipline recon.py applies.
    """
    want = {
        "req_id": ("requirement or design id", "polarion"),
        "tc_id": ("test case id", "testrail"),
        "test_group": ("test group", None),
        "test_set": ("test set", None),
        "test_item": ("test item", None),
        "pre_conditions": ("pre-condition", None),
        "input_test_data": ("input test data", None),
        "test_procedure": ("test procedure", None),
        "expected_result": ("expected result", None),
        "spec_reference": ("specification reference", None),
        "tc_ref_id": ("test case reference id", None),
        "priority": ("test case priority", None),
        "design_method": ("test case design", None),
        "functional_safety": ("functional safety", None),
        "author": ("test case author", None),
        "remarks": ("remarks", None),
    }
    header = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    norm = [re.sub(r"\s+", " ", str(v or "")).strip().lower() for v in header]
    out: dict[str, int] = {}
    for key, (need, forbid) in want.items():
        hits = [i + 1 for i, h in enumerate(norm)
                if need in h and not (forbid and forbid in h)]
        if len(hits) != 1:
            raise WriteBackError(
                f"column {key!r}: {len(hits)} header matches for {need!r} "
                f"— resolve by header text, never by guessing a letter")
        out[key] = hits[0]
    return out


def vehicle_columns(ws, header_row: int) -> list[int]:
    """T–Z under the merged `Vehicle Model 車型` banner, found by the banner."""
    banner_row = header_row - 1
    for rng in ws.merged_cells.ranges:
        anchor = ws.cell(rng.min_row, rng.min_col).value
        if rng.min_row == banner_row and "Vehicle Model" in str(anchor or ""):
            return list(range(rng.min_col, rng.max_col + 1))
    raise WriteBackError("the `Vehicle Model 車型` banner was not found on row "
                         f"{banner_row}; do not guess T–Z")


# ------------------------------------------------------------------ payload

def load_tcs(feature_dir: Path, generated: str) -> list[dict]:
    """generated/*.json -> rows in leaf order, TC order within a leaf."""
    rows: list[dict] = []
    for path in sorted((feature_dir / generated).glob("*.json")):
        obj = json.loads(path.read_text(encoding="utf-8"))
        for tc in obj["tcs"]:
            rows.append(tc)
    if not rows:
        raise WriteBackError(f"no TCs under {feature_dir / generated}")
    return rows


def write_rows(ws, cols: dict[str, int], veh: list[int], rows: list[dict],
               cfg: dict) -> dict:
    wb_cfg = cfg["write_back"]
    fmt = wb_cfg["tc_id_format"]
    n_blocked = 0
    for offset, tc in enumerate(rows):
        r = FIRST_DATA_ROW + offset
        placeholder = bool(tc.get("placeholder"))
        n_blocked += placeholder

        ws.cell(r, cols["req_id"]).value = tc["req_id"]
        ws.cell(r, cols["tc_id"]).value = fmt.format(n=offset + 1)
        ws.cell(r, cols["test_group"]).value = tc["test_group"]
        ws.cell(r, cols["test_set"]).value = tc["test_set"]
        ws.cell(r, cols["test_item"]).value = tc["test_item"]
        for key in ("pre_conditions", "input_test_data",
                    "test_procedure", "expected_result"):
            ws.cell(r, cols[key]).value = tc[key]
        ws.cell(r, cols["spec_reference"]).value = tc["specification_reference"]
        ws.cell(r, cols["tc_ref_id"]).value = wb_cfg["tc_ref_id_value"]
        # A BLOCKED row carries no priority and no design method: there is
        # nothing to prioritise and nothing was designed (profile §5).
        ws.cell(r, cols["priority"]).value = tc["priority"] or None
        ws.cell(r, cols["design_method"]).value = tc["design_method"] or None
        ws.cell(r, cols["functional_safety"]).value = CONST_FUNCTIONAL_SAFETY
        ws.cell(r, cols["author"]).value = wb_cfg["author_value"]
        ws.cell(r, cols["remarks"]).value = tc["remarks"] or None
        for c in veh:                      # R30-4 — explicit, not "left alone"
            ws.cell(r, c).value = None
        ws.cell(r, 2).value = ROW_NUMBER_FORMULA.format(row=r)
    return {"rows": len(rows), "blocked": n_blocked,
            "first_row": FIRST_DATA_ROW, "last_row": FIRST_DATA_ROW + len(rows) - 1}


# ---------------------------------------------------------------- invariants

def check_header_untouched(src: Path, out: Path, sheet: str, header_row: int) -> None:
    """Nothing above the first data row may change.

    A BLANK workbook has no frozen data region, so this is the equivalent
    guarantee: the header block carries the Scope field, the form id and the
    document control area, and none of them is this script's business.
    """
    a = openpyxl.load_workbook(src)[sheet]
    b = openpyxl.load_workbook(out)[sheet]
    drift = []
    for r in range(1, header_row + 1):
        for c in range(1, max(a.max_column, b.max_column) + 1):
            if a.cell(r, c).value != b.cell(r, c).value:
                drift.append(f"{a.cell(r, c).coordinate}: "
                             f"{a.cell(r, c).value!r} -> {b.cell(r, c).value!r}")
    if drift:
        raise WriteBackError("the header block changed: " + "; ".join(drift[:5]))


def check_other_sheets(src: Path, out: Path, sheet: str) -> None:
    a, b = openpyxl.load_workbook(src), openpyxl.load_workbook(out)
    if a.sheetnames != b.sheetnames:
        raise WriteBackError(f"sheet list changed: {a.sheetnames} -> {b.sheetnames}")
    for name in a.sheetnames:
        if name == sheet:
            continue
        sa, sb = a[name], b[name]
        for r in range(1, max(sa.max_row, sb.max_row) + 1):
            for c in range(1, max(sa.max_column, sb.max_column) + 1):
                if sa.cell(r, c).value != sb.cell(r, c).value:
                    raise WriteBackError(
                        f"sheet {name!r} changed at {sa.cell(r, c).coordinate}")


# -------------------------------------------------------------------- driver

def run(args) -> int:
    feature_dir = Path(args.feature_dir)
    cfg = yaml.safe_load((feature_dir / "feature.yaml").read_text(encoding="utf-8"))
    sheet = cfg["workbook"]["sheet"]
    header_row = cfg["workbook"]["header_row"]

    src = Path(args.source) if args.source else next(
        (feature_dir / "output").glob("*_Privacy_*.xlsx"))
    if not src.is_file():
        raise WriteBackError(f"input baseline not found: {src}")

    rows = load_tcs(feature_dir, args.generated)
    wb = openpyxl.load_workbook(src)
    ws = wb[sheet]
    cols = resolve_columns(ws, header_row)
    veh = vehicle_columns(ws, header_row)
    plan = write_rows(ws, cols, veh, rows, cfg)

    print(f"source        : {src.name}")
    print(f"  SHA256      : {sha256_file(src)}")
    print(f"sheet         : {sheet!r}, header row {header_row}")
    print(f"columns       : resolved from header text — "
          + ", ".join(f"{k}={openpyxl.utils.get_column_letter(v)}"
                      for k, v in cols.items()))
    print(f"vehicle block : "
          + "".join(openpyxl.utils.get_column_letter(c) for c in veh)
          + " (from the merged `Vehicle Model 車型` banner)")
    print(f"rows          : {plan['rows']} TCs at rows "
          f"{plan['first_row']}-{plan['last_row']}, "
          f"{plan['blocked']} BLOCKED")
    print(f"tc ids        : {cfg['write_back']['tc_id_format'].format(n=1)} … "
          f"{cfg['write_back']['tc_id_format'].format(n=plan['rows'])} "
          "(sequential, no gap — the BLOCKED row takes its id)")
    print(f"column S      : {CONST_FUNCTIONAL_SAFETY!r} on every row (R30-3)")
    print(f"blank by decision: "
          + "; ".join(f"{k} — {v}" for k, v in BLANK_BY_DECISION.items()))

    if not args.write:
        print("\nDRY RUN — nothing written. Re-run with --write.")
        return 0

    out = Path(args.out) if args.out else src.with_name(
        src.stem + "_regen-v1.xlsx")
    report = surgical_save(wb, src, out)
    check_header_untouched(src, out, sheet, header_row)
    check_other_sheets(src, out, sheet)
    verify_structure(src, out, set(report["members_patched"]))

    print(f"\nwrote         : {out}")
    print(f"SHA256        : {sha256_file(out)}")
    print(f"bytes         : {out.stat().st_size:,}")
    print(f"structure     : {report['members']} zip members, "
          f"differing: {', '.join(report['differing'])}")
    print(f"  DV          : " + ", ".join(
        f"{m.split('/')[-1]} classic={c} x14={x}"
        for m, (c, x) in report["dv_counts"].items()))
    print("  header block unchanged; all non-target sheets identical")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--feature-dir", default="features/privacy")
    ap.add_argument("--generated", default="generated")
    ap.add_argument("--source", help="input baseline (default: the prepared "
                                     "workbook in output/)")
    ap.add_argument("--out")
    ap.add_argument("--write", action="store_true")
    args = ap.parse_args()
    try:
        return run(args)
    except (WriteBackError, StructureError) as exc:
        print(f"ABORTED: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
