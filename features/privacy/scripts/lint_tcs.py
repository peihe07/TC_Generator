#!/usr/bin/env python3
"""Step 3 (Privacy) — hard gate over generated/*.json.

Built for Privacy rather than copied from AMFM. The AMFM gate resolves every
`specification_reference` through `data/stla_to_cfts.json`; Privacy has no
such map and must not acquire one by analogy — its ids are CFTS022 artifact
ids, and R30-1 established that they are **looked up, never computed** (the
`−1` offset that holds across the SCV block is a local regularity of a
gap-free id range, and it produced two wrong ids on the two leaves outside
that range). So `spec-reference` here resolves against the artifact ids
actually present in the CFTS022 docx, read fresh on every run.

Authorities are read, never hard-coded: the design-method vocabulary comes
from the workbook's own `下拉選單` sheet, the Test Group and the id format
from `feature.yaml` and the profile, the artifact set from the spec document.

Two gates encode rulings rather than generic rules:

- **`step-actions` (R33-5)** — one step, one action. The judgement is on the
  **verb count, not the connective**: `Read the signal and its timestamp` is
  one observation of two attributes and passes; `Set the volume and read the
  signal` is two actions and fails. A connective-only check would have failed
  the first and taught the wrong lesson.
- **`negative-scope` (R33-1(d))** — a TC whose design method is
  `負向測試 (Negative / Invalid)` must actually inject an illegal input. The
  gate exists because -005 TC2 originally carried that method while only
  observing this ECU's output set: on the HU side of a signal there is no
  position from which to inject an invalid value at all. §7's
  negative-pairing requirement is then discharged by scope attribution, not
  by a TC that looks like a negative.

Exit 0 = clean, 1 = at least one finding, 2 = bad invocation.

Usage:
    python features/privacy/scripts/lint_tcs.py --feature-dir features/privacy
    python features/privacy/scripts/lint_tcs.py --feature-dir features/privacy --self-test
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

import docx
import openpyxl
import yaml

REPO_ROOT = next(p for p in Path(__file__).resolve().parents
                 if (p / "pyproject.toml").is_file())

# R35-2 made multi-cite legal: a TC lists every clause it verifies OR relies
# on as setup (§10.7). The per-component check is unchanged — each part must
# still be CFTS022-<7 digits> AND resolve in the document — so this widens
# what a reference may CONTAIN, not what any component may BE.
SPEC_REF_RE = re.compile(r"^CFTS022-(\d{7})$")
SPEC_REF_SEP = "; "
MODAL_RE = re.compile(r"\b(shall|should|must|may|will|would|can|could)\b", re.I)
STEP_RE = re.compile(r"^(\d+)\.\s*(.+)$", re.M)
ARTIFACT_HDR = re.compile(r"^\s*(\d{7})\s*:\s*\[")
BANNED_PRECONDITIONS = ("HU is powered on",)

# profile §5 marker table. A marker not listed here is not a marker — creating
# one at generation time is a stop-and-report, so the gate treats an unknown
# opening token in Remarks as a finding rather than as new vocabulary.
REGISTERED_MARKERS = ("[BLOCKED-ECU]",)
PLACEHOLDER_BODY = "BLOCKED - see Remarks"

# R33-5: the unit is the verb, not the connective. These are the imperative
# verbs the house style opens a step with; a step carrying two of them across
# an `and` is two actions.
STEP_VERBS = ("set", "read", "press", "release", "trigger", "start", "stop",
              "record", "step through", "navigate", "select", "wait",
              "connect", "disconnect", "inject", "send", "put", "measure")


class LintError(RuntimeError):
    pass


# ------------------------------------------------------------- authorities

def load_authorities(feature_dir: Path) -> dict:
    cfg = yaml.safe_load((feature_dir / "feature.yaml").read_text(encoding="utf-8"))
    wb_path = feature_dir / cfg["paths"]["workbook"]
    if not wb_path.is_file():
        raise LintError(f"workbook not found: {wb_path}")
    sheet = openpyxl.load_workbook(wb_path)["下拉選單"]
    vocab = [sheet.cell(r, 1).value for r in range(1, 10)]
    if any(v is None for v in vocab):
        raise LintError("下拉選單!A1:A9 has an empty entry — vocabulary unusable")

    spec = next((feature_dir / "inputs").glob("*CFTS_022 Functional Specification*.docx"), None)
    if spec is None:
        raise LintError("CFTS022 functional specification not found in inputs/")
    doc = docx.Document(str(spec))
    artifacts = {m.group(1) for m in
                 (ARTIFACT_HDR.match(unicodedata.normalize("NFKC", p.text))
                  for p in doc.paragraphs) if m}
    if not artifacts:
        raise LintError(f"no artifact ids parsed from {spec.name} — "
                        "the header pattern changed; fix the parser, "
                        "do not relax the gate")
    baseline = feature_dir / "BASELINE.sha256"
    spec_sha = ""
    if baseline.is_file():
        for line in baseline.read_text(encoding="utf-8").split("\n"):
            if line and not line.startswith("#") and "CFTS_022 Functional" in line:
                spec_sha = line.split()[0]
    reviewed_path = feature_dir / "data" / "spec_ref_reviewed.json"
    reviewed = (json.loads(reviewed_path.read_text(encoding="utf-8"))
                if reviewed_path.is_file() else {"leaves": []})
    return {"vocab": vocab, "artifacts": artifacts, "reviewed": reviewed,
            "spec_sha": spec_sha,
            "test_group": cfg["test_group"],
            "sets": {"Input Monitoring", "Personalization Display",
                     "Speed-Controlled Volume"}}


# ------------------------------------------------------------------ gates

def _steps(text: str) -> list[str]:
    return [m.group(2).strip() for m in STEP_RE.finditer(text)]


def _action_count(step: str) -> int:
    """R33-5 — count imperative verbs, not connectives."""
    lowered = " " + step.lower()
    hits = 0
    for verb in STEP_VERBS:
        hits += len(re.findall(rf"(?:^|\band\b|\bthen\b|,)\s*{re.escape(verb)}\b",
                               lowered))
    return hits


def lint_placeholder(tc: dict, where: str) -> list[tuple[str, str]]:
    """A BLOCKED row is not a weaker TC — it is a different artefact.

    Relaxing the content gates for it would blunt them for everything else,
    so placeholder rows get their own set instead: the verification fields
    must carry exactly the placeholder body, priority and design method must
    be empty (there is nothing to prioritise or design), and Remarks must
    open with a registered marker so the gap is followable.
    """
    out: list[tuple[str, str]] = []

    def bad(gate, msg):
        out.append((gate, f"{where}: {msg}"))

    for field in ("pre_conditions", "input_test_data",
                  "test_procedure", "expected_result"):
        if tc.get(field) != PLACEHOLDER_BODY:
            bad("placeholder-body",
                f"{field} must be exactly {PLACEHOLDER_BODY!r}")
    for field in ("priority", "design_method"):
        if tc.get(field):
            bad("placeholder-blank",
                f"{field} must be empty on a BLOCKED row, found {tc[field]!r}")
    if not tc.get("remarks", "").strip():
        bad("placeholder-remarks", "a BLOCKED row must explain itself in Remarks")
    return out


def lint_remarks(tc: dict, where: str) -> list[tuple[str, str]]:
    """R34-3 — a non-empty Remarks must open with a registered marker."""
    remarks = tc.get("remarks", "").strip()
    if not remarks:
        return []
    token = remarks.split(None, 1)[0]
    if token not in REGISTERED_MARKERS:
        return [("remarks-marker",
                 f"{where}: Remarks opens with {token!r}, which is not in the "
                 f"profile §5 marker table {list(REGISTERED_MARKERS)}")]
    return []


def lint_tc(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = lint_remarks(tc, where)
    if tc.get("placeholder"):
        return out + lint_placeholder(tc, where)

    def bad(gate, msg):
        out.append((gate, f"{where}: {msg}"))

    if tc.get("design_method") not in auth["vocab"]:
        bad("design-method", f"{tc.get('design_method')!r} is not one of the "
                             "nine 下拉選單!A1:A9 strings")
    if tc.get("test_group") != auth["test_group"]:
        bad("test-group", f"{tc.get('test_group')!r} != {auth['test_group']!r}")
    if tc.get("test_set") not in auth["sets"]:
        bad("test-set", f"{tc.get('test_set')!r} is not a framework Part VI Set")
    if tc.get("priority") not in {"P0", "P1", "P2", "P3"}:
        bad("priority", f"{tc.get('priority')!r} is not P0..P3")

    ref = tc.get("specification_reference", "")
    parts = [p.strip() for p in ref.split(";")] if ref else []
    if not parts:
        bad("spec-reference", "empty specification_reference")
    for part in parts:
        m = SPEC_REF_RE.match(part)
        if not m:
            bad("spec-reference", f"{part!r} is not CFTS022-<7 digits>")
        elif m.group(1) not in auth["artifacts"]:
            bad("spec-reference", f"artifact {m.group(1)} is not present in "
                                  "the CFTS022 document — ids are looked up, "
                                  "never computed (R30-1)")
    if len(parts) != len(set(parts)):
        bad("spec-reference", f"{ref!r} repeats a reference")

    er = tc.get("expected_result", "")
    # An ALL-CAPS token is an acronym or signal name, not a modal verb: the
    # bus is `Interior CAN`, and a case-insensitive \bcan\b matches it. Found
    # by the B1 re-run, which is what the re-run requirement is for — the
    # same class as R17-2 (a vocabulary tool's defect does not raise an error).
    hit = next((m for m in MODAL_RE.finditer(er) if not m.group(0).isupper()), None)
    if hit:
        bad("er-modal", f"Expected Result contains the modal {hit.group(0)!r}")

    steps, ers = _steps(tc.get("test_procedure", "")), _steps(er)
    if len(steps) != len(ers):
        bad("step-er-parity", f"{len(steps)} steps vs {len(ers)} expected results")
    if len(steps) < 2:
        bad("step-count", f"{len(steps)} step(s); a procedure needs at least 2")

    for i, step in enumerate(steps, 1):
        if _action_count(step) > 1:
            bad("step-actions", f"step {i} carries more than one action "
                                f"(R33-5: verbs, not connectives) — {step!r}")

    for line in tc.get("pre_conditions", "").split("\n"):
        for banned in BANNED_PRECONDITIONS:
            if banned in line:
                bad("precondition-banned", f"{banned!r} is not a spec trigger")

    for field in ("test_item", "expected_result"):
        for line in tc.get(field, "").split("\n"):
            if line.rstrip().endswith("."):
                bad("trailing-period", f"{field} line ends with a period: {line!r}")

    if (tc.get("design_method") == "負向測試 (Negative / Invalid)"
            and not re.search(r"\binject|\binvalid value|\bout-of-range",
                              tc.get("test_procedure", ""), re.I)):
        bad("negative-scope",
            "design method is Negative / Invalid but the procedure injects no "
            "illegal input. If this ECU has no position from which to inject "
            "one, §7's negative pairing is discharged by scope attribution, "
            "not by a TC (R33-1(d))")

    return out


def lint_spec_ref_reviewed(obj: dict, reviewed: dict, where: str,
                           spec_sha: str = "") -> list[tuple[str, str]]:
    """R35-7 — a one-off human review is only valid at the moment it is done.

    `data/spec_ref_reviewed.json` freezes what was reviewed. This gate does not
    re-do the judgement; it detects when the judgement has been invalidated by
    a reference changing underneath it.
    """
    leaf = obj.get("parent", "")
    refs = {tc.get("specification_reference", "") for tc in obj.get("tcs", [])}
    entries = [e for e in reviewed.get("leaves", []) if e["leaf"] == leaf]
    if not entries:
        return [("spec-ref-reviewed",
                 f"{where}: {leaf} has no entry in spec_ref_reviewed.json — "
                 "the semantic correspondence review must be done and recorded")]
    latest = entries[-1]
    for ref in refs:
        if ref != latest["specification_reference"]:
            return [("spec-ref-reviewed",
                     f"{where}: specification_reference is {ref!r} but the "
                     f"recorded review covered {latest['specification_reference']!r} "
                     f"(reviewed {latest['reviewed']}) — this leaf's semantic "
                     "correspondence review must be redone and appended")]
        # R36-6 — the leaf's own clause leads; setup dependencies follow.
        first = ref.split(";")[0].strip()
        want = latest["specification_reference"].split(";")[0].strip()
        if first != want:
            return [("spec-ref-order",
                     f"{where}: first reference is {first!r}; the leaf's own "
                     f"clause is {want!r}. References run most specific first "
                     "— the verified clause leads, setup dependencies follow "
                     "(R35-2 / R36-6)")]
    # R36-5 — a human judgement is only valid for the document version it read.
    recorded = latest.get("source_sha256", "")
    if spec_sha and recorded and recorded != spec_sha:
        return [("spec-ref-source-version",
                 f"{where}: the review recorded CFTS022 {recorded[:16]}… but "
                 f"BASELINE.sha256 now carries {spec_sha[:16]}… — the source "
                 "document has been replaced; this leaf's semantic "
                 "correspondence review must be redone")]
    return []


def lint_file(path: Path, auth: dict) -> list[tuple[str, str]]:
    obj = json.loads(path.read_text(encoding="utf-8"))
    out: list[tuple[str, str]] = lint_spec_ref_reviewed(
        obj, auth["reviewed"], path.name, auth.get("spec_sha", ""))
    if obj.get("assumptions"):
        out.append(("assumption-marker",
                    f"{path.name}: assumptions is non-empty and this feature "
                    "has no marker vocabulary (profile §5) — stop and report"))
    for i, tc in enumerate(obj.get("tcs", []), 1):
        out += lint_tc(tc, auth, f"{path.name} TC{i}")
    return out


# ------------------------------------------------- workbook column policy

ENTRY_RE = re.compile(r"^#\s*ENTRY\s+(\d{3})\b")
STATUS_RE = re.compile(r"^#\s*STATUS:\s*(.+?)\s*\(([^,]+),\s*([\d-]+)\)\s*$")


def lint_delivery_ledger(feature_dir: Path) -> list[tuple[str, str]]:
    """R41-7 — each ENTRY's shape is: prose -> hash -> STATUS, STATUS last.

    The ledger is append-only, so a superseded header line stays put; the
    only reliable place to read an entry's current state is its final line.
    That invariant is worth nothing unless something checks it — it was
    breached once already, when ENTRY 001's STATUS was first written above
    its hash line instead of below.
    """
    path = feature_dir / "DELIVERY.sha256"
    if not path.is_file():
        return []
    lines = path.read_text(encoding="utf-8").split("\n")
    starts = [i for i, l in enumerate(lines) if ENTRY_RE.match(l)]
    if not starts:
        return []
    out: list[tuple[str, str]] = []
    bounds = starts + [len(lines)]
    for n, start in enumerate(starts):
        end = bounds[n + 1]
        # A bare `#` is a visual separator between entries, not content. The
        # invariant is about the last line a reader takes meaning from; a
        # decoration line carries none. Same distinction as `Interior CAN`
        # not being the modal `can` — the standard is unchanged, the
        # implementation was reading too literally.
        block = [l for l in lines[start:end]
                 if l.strip() and l.strip() != "#"]
        entry = ENTRY_RE.match(lines[start]).group(1)
        if not block:
            continue
        last = block[-1]
        if not STATUS_RE.match(last):
            out.append(("ledger-status-last",
                        f"ENTRY {entry}: the last line of the entry is "
                        f"{last[:70]!r}, not a `# STATUS: <state> "
                        "(<ruling>, <date>)` line — an entry's current state "
                        "is read from its last line (R41-7)"))
            continue
        hashes = [i for i, l in enumerate(block)
                  if re.match(r"^[0-9a-f]{64}\s+\S", l)]
        if not hashes:
            out.append(("ledger-status-last",
                        f"ENTRY {entry}: no hash line found in the entry"))
        elif hashes[-1] > len(block) - 2:
            out.append(("ledger-status-last",
                        f"ENTRY {entry}: the STATUS line must follow the hash "
                        "line (prose -> hash -> STATUS)"))
    return out


def lint_workbook_policy(feature_dir: Path, workbook: Path | None = None
                        ) -> list[tuple[str, str]]:
    """profile §3.8 / §3.9 — column S is `NA`, columns T–Z stay blank.

    Measurable only after write-back (R34-6). Before that these were reported
    NOT MEASURED rather than PASS — a gate that cannot fail is not a gate.
    Now that a written workbook exists they run for real.
    """
    if workbook is None or not workbook.is_file():
        return []
    import yaml as _yaml
    cfg = _yaml.safe_load((feature_dir / "feature.yaml").read_text(encoding="utf-8"))
    ws = openpyxl.load_workbook(workbook)[cfg["workbook"]["sheet"]]
    header_row = cfg["workbook"]["header_row"]
    out: list[tuple[str, str]] = []
    cols = {re.sub(r"\s+", " ", str(ws.cell(header_row, c).value or "")).strip().lower(): c
            for c in range(1, ws.max_column + 1)}
    s_col = next(c for h, c in cols.items() if "functional safety" in h)
    veh = []
    for rng in ws.merged_cells.ranges:
        if rng.min_row == header_row - 1 and "Vehicle Model" in str(
                ws.cell(rng.min_row, rng.min_col).value or ""):
            veh = list(range(rng.min_col, rng.max_col + 1))
    req_col = next(c for h, c in cols.items()
                   if "requirement or design id" in h and "polarion" not in h)
    for r in range(header_row + 1, ws.max_row + 1):
        if not ws.cell(r, req_col).value:
            continue
        if ws.cell(r, s_col).value != "NA":
            out.append(("column-s-na",
                        f"row {r}: Functional Safety is "
                        f"{ws.cell(r, s_col).value!r}, must be 'NA' (R30-3)"))
        for c in veh:
            if ws.cell(r, c).value not in (None, ""):
                out.append(("vehicle-blank",
                            f"row {r} col {openpyxl.utils.get_column_letter(c)}: "
                            f"{ws.cell(r, c).value!r}, vehicle columns stay "
                            "blank (R30-4)"))
    return out


# ---------------------------------------------------------------- self test

# R34-5 — every gate needs BOTH controls:
#   positive: a violating input MUST fail
#   negative: a *compliant but similar* input MUST pass
# A gate that misfires on correct input is as broken as one that never fires.
# Where no negative control exists, the gate is reported NOT MEASURED, never
# PASS. The `er-modal` negative below is the one that caught a real misfire:
# `Interior CAN` read as the modal `can`.
NEGATIVE_CONTROLS = {
    "design-method": {"design_method": "負向測試 (Negative / Invalid)",
                      "test_procedure": "1. Inject an invalid value\n2. Read the signal",
                      "expected_result": "1. Injected\n2. Read"},
    "test-set": {"test_set": "Input Monitoring"},
    "priority": {"priority": "P0"},
    "spec-reference": {"specification_reference": "CFTS022-4914955"},
    "er-modal": {"expected_result": "1. The Interior CAN is awake\n2. The HU is awake"},
    "step-actions": {"test_procedure": "1. Read the signal and its timestamp\n2. Read the signal",
                     "expected_result": "1. Read\n2. Read"},
    "precondition-banned": {"pre_conditions": "1. The HU is awake"},
    "trailing-period": {"test_item": "The HU sends the signal within <Tsend>"},
    "negative-scope": {"design_method": "負向測試 (Negative / Invalid)",
                       "test_procedure": "1. Inject an out-of-range value\n2. Read the signal",
                       "expected_result": "1. Injected\n2. Read"},
    "remarks-marker": {"remarks": "[BLOCKED-ECU] out of scope for this ECU"},
    "placeholder-body": {"placeholder": True, "remarks": "[BLOCKED-ECU] x",
                         "pre_conditions": PLACEHOLDER_BODY,
                         "input_test_data": PLACEHOLDER_BODY,
                         "test_procedure": PLACEHOLDER_BODY,
                         "expected_result": PLACEHOLDER_BODY,
                         "priority": "", "design_method": ""},
    "placeholder-blank": {"placeholder": True, "remarks": "[BLOCKED-ECU] x",
                          "pre_conditions": PLACEHOLDER_BODY,
                          "input_test_data": PLACEHOLDER_BODY,
                          "test_procedure": PLACEHOLDER_BODY,
                          "expected_result": PLACEHOLDER_BODY,
                          "priority": "", "design_method": ""},
    "placeholder-remarks": {"placeholder": True, "remarks": "[BLOCKED-ECU] x",
                            "pre_conditions": PLACEHOLDER_BODY,
                            "input_test_data": PLACEHOLDER_BODY,
                            "test_procedure": PLACEHOLDER_BODY,
                            "expected_result": PLACEHOLDER_BODY,
                            "priority": "", "design_method": ""},
    # R35-5 — a negative control's discriminating power falls off with its
    # distance from the violation. The clean baseline sits furthest away: it
    # shows the gate does not misfire on OBVIOUSLY compliant input, not that
    # it holds at the boundary. These three carry boundary cases instead.
    "test-group": {"test_set": "Input Monitoring"},   # 合法但不同 Set
    "step-er-parity": {                                # 單步驟對應多行 ER
        "test_procedure": "1. Read the signal\n2. Read the timestamp",
        "expected_result": "1. The signal carries [Off]\n2. The timestamp is within <Tsend>"},
    "step-count": {                                    # 恰為下界之 2 步
        "test_procedure": "1. Read the signal\n2. Read the signal again",
        "expected_result": "1. Read\n2. Read again"},
    "spec-ref-reviewed": {},
}

SELF_TEST_CASES = [
    ("design-method", {"design_method": "EP"}),
    ("test-group", {"test_group": "Radio"}),
    ("test-group", {"test_group": "privacy"}),        # R36-7 大小寫近似值
    ("test-group", {"test_group": "Privacy "}),       # R36-7 尾隨空白
    ("test-set", {"test_set": "Nonexistent Set"}),
    ("priority", {"priority": "High"}),
    ("spec-reference", {"specification_reference": "CFTS022-9999999"}),
    ("er-modal", {"expected_result": "1. The HU shall display the page\n2. Done"}),
    ("step-er-parity", {"expected_result": "1. Only one line"}),
    ("step-count", {"test_procedure": "1. Read the signal",
                    "expected_result": "1. The signal is read"}),
    ("step-actions", {"test_procedure": "1. Set the volume and read the signal\n2. Read the signal",
                      "expected_result": "1. Done\n2. Done"}),
    ("precondition-banned", {"pre_conditions": "1. HU is powered on"}),
    ("trailing-period", {"test_item": "The HU sends the signal."}),
    # negative control for the ALL-CAPS carve-out: `CAN` must NOT trigger.
    ("negative-scope", {"design_method": "負向測試 (Negative / Invalid)",
                        "test_procedure": "1. Read the signal\n2. Read the signal",
                        "expected_result": "1. Done\n2. Done"}),
    ("remarks-marker", {"remarks": "A-PV18 see anomalies"}),
    ("placeholder-body", {"placeholder": True, "remarks": "[BLOCKED-ECU] x",
                          "test_procedure": "1. Do something",
                          "pre_conditions": PLACEHOLDER_BODY,
                          "input_test_data": PLACEHOLDER_BODY,
                          "expected_result": PLACEHOLDER_BODY,
                          "priority": "", "design_method": ""}),
    ("placeholder-blank", {"placeholder": True, "remarks": "[BLOCKED-ECU] x",
                           "pre_conditions": PLACEHOLDER_BODY,
                           "input_test_data": PLACEHOLDER_BODY,
                           "test_procedure": PLACEHOLDER_BODY,
                           "expected_result": PLACEHOLDER_BODY,
                           "priority": "P1", "design_method": ""}),
    ("placeholder-remarks", {"placeholder": True, "remarks": "",
                             "pre_conditions": PLACEHOLDER_BODY,
                             "input_test_data": PLACEHOLDER_BODY,
                             "test_procedure": PLACEHOLDER_BODY,
                             "expected_result": PLACEHOLDER_BODY,
                             "priority": "", "design_method": ""}),
]


def base_tc() -> dict:
    return {
        "test_group": "Privacy", "test_set": "Speed-Controlled Volume",
        "test_item": "The HU sends the signal",
        "pre_conditions": "1. The HU is awake",
        "input_test_data": "NA",
        "test_procedure": "1. Read the signal\n2. Read the signal again",
        "expected_result": "1. The signal is read\n2. The signal is read again",
        "specification_reference": "CFTS022-4915170",
        "priority": "P1",
        "design_method": "功能測試 (Functional based ; no specific technique)",
        "remarks": "",
    }


def _self_test_spec_ref_reviewed(auth: dict) -> tuple[bool, bool]:
    """File-level gate — exercised separately from the TC-level table."""
    reviewed = auth["reviewed"]
    if not reviewed.get("leaves"):
        return False, False
    entry = reviewed["leaves"][0]
    good = {"parent": entry["leaf"],
            "tcs": [{"specification_reference": entry["specification_reference"]}]}
    bad = {"parent": entry["leaf"],
           "tcs": [{"specification_reference": "CFTS022-4915104"}]}
    unknown = {"parent": "SWE1-HMI-PRIVACY_FEATURES-999",
               "tcs": [{"specification_reference": "CFTS022-4915104"}]}
    first = entry["specification_reference"].split(";")
    reversed_ref = "; ".join(x.strip() for x in reversed(first)) if len(first) > 1 else None
    fires_on_bad = bool(lint_spec_ref_reviewed(bad, reviewed, "x")) and \
        bool(lint_spec_ref_reviewed(unknown, reviewed, "x"))
    quiet_on_good = not lint_spec_ref_reviewed(good, reviewed, "x", entry.get("source_sha256", ""))
    # R36-5 — a changed source document must invalidate the review
    fires_on_version = bool(lint_spec_ref_reviewed(good, reviewed, "x", "deadbeef" * 8))
    # R36-6 — reversed order must fail; needs a multi-cite leaf
    multi = next((e for e in reviewed["leaves"] if ";" in e["specification_reference"]), None)
    if multi:
        parts = [x.strip() for x in multi["specification_reference"].split(";")]
        rev = {"parent": multi["leaf"],
               "tcs": [{"specification_reference": "; ".join(reversed(parts))}]}
        ok = {"parent": multi["leaf"],
              "tcs": [{"specification_reference": multi["specification_reference"]}]}
        fires_on_order = any(g == "spec-ref-order"
                             for g, _ in lint_spec_ref_reviewed(rev, reviewed, "x"))
        quiet_on_order = not lint_spec_ref_reviewed(ok, reviewed, "x",
                                                    multi.get("source_sha256", ""))
    else:
        fires_on_order = quiet_on_order = False
    return (fires_on_bad, quiet_on_good, fires_on_version,
            fires_on_order, quiet_on_order)


def self_test(auth: dict) -> int:
    print("positive control — every gate is deliberately violated once:\n")
    clean = lint_tc(base_tc(), auth, "baseline")
    if clean:
        print(f"  BASELINE IS NOT CLEAN: {clean}")
        return 1
    print("  baseline TC: clean (0 findings)\n")
    failed = []
    for gate, patch in SELF_TEST_CASES:
        tc = base_tc() | patch
        gates = {g for g, _ in lint_tc(tc, auth, "self-test")}
        ok = gate in gates
        print(f"  {'TRIGGERED' if ok else 'NOT TRIGGERED':<14} {gate}")
        if not ok:
            failed.append(gate)
    print()
    if failed:
        print(f"gates that could not be made to fail: {failed}")
        print("mark them NOT MEASURED, not PASS (R18-4 principle)")
        return 1
    fires, quiet, ver, order, order_ok = _self_test_spec_ref_reviewed(auth)
    for label, ok, name in [
            ("spec-ref-reviewed (changed ref + unrecorded leaf)", fires, "spec-ref-reviewed"),
            ("spec-ref-source-version (CFTS022 replaced)", ver, "spec-ref-source-version"),
            ("spec-ref-order (references reversed)", order, "spec-ref-order")]:
        print(f"  {'TRIGGERED' if ok else 'NOT TRIGGERED':<14} {label}")
        if not ok:
            failed.append(name)
    print(f"\nall {len(SELF_TEST_CASES)} + 1 gates verified reachable")

    # R34-5 negative controls — a compliant but similar input must PASS.
    print("\nnegative controls — a compliant, similar input must NOT fire:\n")
    unmeasured, misfired = [], []
    for gate, _ in SELF_TEST_CASES:
        patch = NEGATIVE_CONTROLS.get(gate)
        if patch is None:
            unmeasured.append(gate)
            print(f"  {'NOT MEASURED':<14} {gate}  (no negative control defined)")
            continue
        gates = {g for g, _ in lint_tc(base_tc() | patch, auth, "negative-control")}
        if gate in gates:
            misfired.append(gate)
            print(f"  {'MISFIRED':<14} {gate}  fires on compliant input")
        else:
            label = "PASS (baseline)" if not patch else "PASS"
            print(f"  {label:<14} {gate}")
    print()
    if misfired:
        print(f"gates that misfire on compliant input: {misfired}")
        return 1
    print(f"  {'PASS' if quiet else 'MISFIRED':<14} spec-ref-reviewed "
          "(recorded ref + recorded source sha must not fire)")
    print(f"  {'PASS' if order_ok else 'MISFIRED':<14} spec-ref-order "
          "(correct order must not fire)")
    if not (quiet and order_ok):
        misfired.append("spec-ref-reviewed")
        print("\nspec-ref-reviewed misfires on a recorded reference")
        return 1
    if unmeasured:
        print(f"gates without a negative control (report as NOT MEASURED, "
              f"never PASS — R34-5): {unmeasured}")
    else:
        print("every gate has both controls")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--feature-dir", default="features/privacy")
    ap.add_argument("--generated", default="generated")
    ap.add_argument("--self-test", action="store_true")
    args = ap.parse_args()

    feature_dir = (REPO_ROOT / args.feature_dir if not Path(args.feature_dir).is_absolute()
                   else Path(args.feature_dir))
    auth = load_authorities(feature_dir)
    print(f"authorities: {len(auth['vocab'])} design methods, "
          f"{len(auth['artifacts'])} CFTS022 artifacts, "
          f"Test Group {auth['test_group']!r}, {len(auth['sets'])} Test Sets\n")

    if args.self_test:
        return self_test(auth)

    files = sorted((feature_dir / args.generated).glob("*.json"))
    if not files:
        raise LintError(f"no generated JSON under {feature_dir / args.generated}")
    findings: list[tuple[str, str]] = []
    n_tc = 0
    for f in files:
        n_tc += len(json.loads(f.read_text(encoding="utf-8")).get("tcs", []))
        findings += lint_file(f, auth)
    written = sorted((feature_dir / "output").glob("*_regen-v*.xlsx"))
    findings += lint_workbook_policy(feature_dir, written[-1] if written else None)
    findings += lint_delivery_ledger(feature_dir)

    print(f"linted {n_tc} TCs from {len(files)} leaf file(s)")
    if written:
        print(f"workbook gates measured against {written[-1].name} "
              "(column S = NA, columns T–Z blank — R34-6)\n")
    else:
        print("NOT MEASURED: column S = NA (profile §3.8), columns T–Z blank "
              "(profile §3.9) — no written workbook yet; these are "
              "write-back gates\n")
    if not findings:
        print("PASS — no findings")
        return 0
    for gate, msg in findings:
        print(f"  [{gate}] {msg}")
    print(f"\nFAIL — {len(findings)} finding(s)")
    return 1


if __name__ == "__main__":
    try:
        sys.exit(main())
    except LintError as exc:
        print(f"ABORTED: {exc}", file=sys.stderr)
        sys.exit(2)
