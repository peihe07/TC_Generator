#!/usr/bin/env python3
"""036 master column mapping + workbook_state probe (handoff 01 steps 9-10).

READ-ONLY. The master's R-column design_method dropdown is an x14 extension
that openpyxl discards on save (R-G1), so this script never writes the
workbook — it only loads it.

Column mapping is derived from THIS file's header row, not inherited from
another feature's table. workbook_state follows canon §2 step by step.

Two baselines are printed, each named (handoff 03 §2.1). Printing only one
of them lets a reader draw the opposite conclusion about A-DM7:

  template-declared  — docs/fw036/templates/feature.yaml, i.e. what the
                       scaffold writes into a brand-new feature. Its
                       mismatches ARE A-DM7.
  effective-declared — features/display/feature.yaml as it now stands, i.e.
                       this feature's corrected map. A clean run here is a
                       re-verification of the correction, NOT evidence that
                       the template is right.

Header cells are printed with repr(): the master's headers embed newlines
('No.#\n序號'), so a whitespace-normalised print would show them as clean
(handoff 03 §2.2 — same defect class as A-DM5).
"""
import re
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]
WB = ROOT / "inputs" / ("FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT "
                        "STLA Test Case Specification & Result_SWQT_20260817_ext.xlsx")

# feature.yaml key -> the header label expected to sit above that column.
# Labels are the workbook's own English/Chinese header text; matching is by
# substring on the whitespace-normalised header, case-insensitive.
EXPECT = {
    "req_id": "requirement or design id",
    "test_group": "test group",
    "test_set": "test set",
    "test_item": "test item",
    "pre_conditions": "pre-condition",
    "input_test_data": "input",
    "test_procedure": "procedure",
    "expected_result": "expected result",
    "spec_reference": "spec",
    "tc_ref_id": "test case reference id",
    "priority": "test case priority",
    "design_method": "design methods",
    "functional_safety": "functional safety",
    "author": "test case author",
    "remarks": "remark",
}


def norm(s):
    return " ".join(str(s or "").split())


def main():
    cfg = yaml.safe_load((ROOT / "feature.yaml").read_text(encoding="utf-8"))
    tpl_path = ROOT.parents[1] / "docs" / "fw036" / "templates" / "feature.yaml"
    tpl = yaml.safe_load(tpl_path.read_text(encoding="utf-8")
                         .replace("{FEATURE}", "Display"))
    sheet = cfg["workbook"]["sheet"]
    hdr_row = cfg["workbook"]["header_row"]
    declared = cfg["workbook"]["columns"]

    wb = openpyxl.load_workbook(WB, data_only=True)
    print("# 036 master probe — READ-ONLY (never saved: x14 DV, R-G1)")
    print(f"file={WB.name}")
    print(f"sheets: {wb.sheetnames}")
    print(f"feature.yaml DECLARES sheet={sheet!r} header_row={hdr_row}")

    # The scaffold template's sheet name is the rev A/B one and is absent from
    # this master. Derive the effective sheet and header row from the file
    # itself rather than from another feature's feature.yaml (handoff 01 §9:
    # no inherited column table).
    def find_header(w):
        for r in range(1, min(30, w.max_row) + 1):
            labels = [norm(w.cell(r, c).value).lower()
                      for c in range(1, w.max_column + 1)]
            if any("test item" in x for x in labels) and \
               any("expected result" in x for x in labels):
                return r
        return None

    eff_sheet, eff_hdr = None, None
    for name in wb.sheetnames:
        r = find_header(wb[name])
        if r:
            eff_sheet, eff_hdr = name, r
            break
    print(f"EFFECTIVE (derived from the master): sheet={eff_sheet!r} "
          f"header_row={eff_hdr}")
    print("derivation: first sheet carrying a row within r1-r30 whose cells "
          "include both 'test item' and 'expected result'")
    if eff_sheet is None:
        raise SystemExit("no TC header row found in any sheet")
    if eff_sheet != sheet:
        print(f"DELTA: declared sheet {sheet!r} is ABSENT from this master "
              f"(scaffold template carries the rev A/B name)")
    if eff_hdr != hdr_row:
        print(f"DELTA: declared header_row {hdr_row} != effective {eff_hdr}")
    sheet, hdr_row = eff_sheet, eff_hdr
    ws = wb[sheet]
    print(f"dims: max_row={ws.max_row} max_col={ws.max_column}")

    hdr, hdr_raw = {}, {}
    print()
    print("## header row content — repr() of the RAW cell value")
    print("   (newlines are the master's own; nothing is normalised here)")
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(hdr_row, c).value
        v = norm(raw)
        letter = openpyxl.utils.get_column_letter(c)
        if v:
            hdr[letter] = v          # normalised, used for matching only
            hdr_raw[letter] = raw
            print(f"  {letter}: {raw!r}")

    def compare(label, declared):
        print()
        print(f"## column mapping — {label} vs header-derived")
        print("| key | declared | header text at declared col | expected label"
              " | verdict |")
        print("|---|---|---|---|---|")
        ok, bad = 0, []
        for key, letter in declared.items():
            text = hdr.get(letter, "")
            want = EXPECT[key]
            good = want in text.lower()
            ok += good
            if not good:
                bad.append((key, letter, text, want))
            print(f"| {key} | {letter} | {text or '（空）'} | {want} | "
                  f"{'MATCH' if good else 'MISMATCH'} |")
        print(f"\nmatch count ({label}): {ok}/{len(declared)}")
        for key, letter, text, want in bad:
            cands = [l for l, x in hdr.items() if want in x.lower()]
            print(f"  MISMATCH {key} @ {letter}: header={text!r}; "
                  f"columns whose header contains {want!r}: {cands or 'none'}")
        return ok, bad

    print()
    print("matching method: the expected label, whitespace-normalised and "
          "lower-cased, as a substring of the whitespace-normalised header "
          "cell at the declared column")
    print(f"template-declared source: {tpl_path}")
    print(f"effective-declared source: {ROOT / 'feature.yaml'}")
    t_ok, t_bad = compare("template-declared", tpl["workbook"]["columns"])
    e_ok, e_bad = compare("effective-declared", declared)

    print()
    print("## 兩基準之並列結論")
    print(f"  template-declared : {t_ok}/{len(tpl['workbook']['columns'])} "
          f"— 不符 {len(t_bad)} 鍵 {[k for k, *_ in t_bad] or '無'}")
    print(f"  effective-declared: {e_ok}/{len(declared)} "
          f"— 不符 {len(e_bad)} 鍵 {[k for k, *_ in e_bad] or '無'}")
    print("  template 之分頁名 "
          f"{tpl['workbook']['sheet']!r} 於本母本 "
          f"{'存在' if tpl['workbook']['sheet'] in wb.sheetnames else '不存在'}")
    print("  -> effective 之全綠是「更正已生效」之複驗，"
          "不得讀為「模板無誤」。模板之不符即 A-DM7 之內容。")

    print()
    print("## header-derived column map (candidates per key)")
    print("| key | expected label | candidate columns | template | effective |")
    print("|---|---|---|---|---|")
    for key, want in EXPECT.items():
        cands = [l for l, x in hdr.items() if want in x.lower()]
        print(f"| {key} | {want} | {','.join(cands) or 'none'} | "
              f"{tpl['workbook']['columns'][key]} | {declared[key]} |")

    # ---------------- workbook_state (canon §2)
    print()
    print("## workbook_state — canon §2")
    ti = declared["test_item"]
    tc = declared["tc_ref_id"]
    au = declared["author"]
    pr = declared["test_procedure"]
    filled, qualifying = [], []
    for r in range(hdr_row + 1, ws.max_row + 1):
        v_ti = norm(ws[f"{ti}{r}"].value)
        v_tc = norm(ws[f"{tc}{r}"].value)
        if v_ti or v_tc:
            filled.append(r)
        v_au = norm(ws[f"{au}{r}"].value)
        v_pr = norm(ws[f"{pr}{r}"].value)
        steps = len(re.findall(r"(?m)^\s*\d+[\.\)]", v_pr))
        if v_au and steps >= 2:
            qualifying.append(r)
    print(f"data rows scanned: r{hdr_row + 1}–r{ws.max_row} "
          f"({ws.max_row - hdr_row})")
    print(f"step 1 filled rows (Test Item or TC ID non-empty): "
          f"{len(filled)} -> {filled if filled else '（無）'}")
    print(f"step 2 qualifying done rows (author AND >=2 numbered steps): "
          f"{len(qualifying)} -> {qualifying if qualifying else '（無）'}")
    state = "BLANK" if not filled else "SEE canon §2 step 3 (segmentation)"
    print(f"step 3 -> workbook_state = {state}")
    print("note: 'content is non-placeholder' (step 2 third clause) is 未實測 "
          "—— with zero qualifying rows there is nothing to inspect; it is "
          "not asserted as PASS.")
    wb.close()


if __name__ == "__main__":
    main()
