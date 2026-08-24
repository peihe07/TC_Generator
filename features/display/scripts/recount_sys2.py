#!/usr/bin/env python3
"""SYS2 CFTS_020 independent recount (handoff 01 step 7).

Measurement conditions:
  - engine: openpyxl, data_only=True, READ-ONLY mode
  - data row: column A ('ID') non-blank after str().strip(), from r2
  - the Category cross-tab is printed BOTH case-normalised and verbatim,
    so the 8 case-variant rows are visible either way
"""
import re
from collections import Counter
from pathlib import Path

import openpyxl

WB = Path(__file__).resolve().parents[1] / "inputs" / \
    "SYS2_CFTS_020_DISP_TCH_ICS_20260616_All_HW_System_Accepted & Released.xlsx"


def norm(s):
    return " ".join(str(s or "").split())


def main():
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)
    print("# SYS2 recount — measurement conditions")
    print("engine=openpyxl data_only=True | mode=READ-ONLY")
    print("data row rule=column A non-blank after str().strip(), from r2")
    print(f"file={WB.name}")
    print("sheets:", wb.sheetnames)

    ws = wb["Basic Report"]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    hdr = [norm(h) for h in grid[0]]
    print(f"Basic Report dims: {len(grid)} rows x {len(hdr)} cols")

    def col(name):
        """Column index by exact whitespace-normalised header, else by prefix.

        The SW/HW column's real header carries a bilingual parenthetical tail
        ('SYS2 SW/HW/System (如果是HW+SW，就選System) (...)'), so the short
        name used in the handoff only matches as a prefix.
        """
        hits = [i for i, h in enumerate(hdr) if h == name]
        if not hits:
            hits = [i for i, h in enumerate(hdr) if h.startswith(name)]
        if len(hits) != 1:
            raise SystemExit(f"header {name!r}: {len(hits)} matches")
        return hits[0]

    a = 0
    data = [(i + 1, r) for i, r in enumerate(grid)
            if i >= 1 and str(r[a] or "").strip() != ""]
    print(f"data rows (A non-blank, r2+): {len(data)}")

    fid = col("SYS2 Sys-RA-Feature-ID")
    cat = col("SYS2 分類 Category")
    swhw = col("SYS2 SW/HW/System")
    grp_hits = [i for i, h in enumerate(hdr) if h == "SYS2 Grouping"]
    melco = col("SYS2 Melco ID")

    dm = sum(1 for _, r in data
             if re.fullmatch(r"SYS-RA-DM-\d+", str(r[fid] or "").strip()))
    ra = sum(1 for _, r in data
             if re.fullmatch(r"SYS2-RA-\d+", str(r[fid] or "").strip()))
    disp = sum(1 for _, r in data if "DISP" in str(r[fid] or ""))
    print(f"Sys-RA-Feature-ID ^SYS-RA-DM-\\d+$ : {dm}")
    print(f"Sys-RA-Feature-ID ^SYS2-RA-\\d+$  : {ra}")
    print(f"other                            : {len(data) - dm - ra}")
    print(f"ids containing 'DISP'            : {disp}")
    if grp_hits:
        g = grp_hits[0]
        blank = sum(1 for _, r in data if str(r[g] or "").strip() == "")
        print(f"SYS2 Grouping blank              : {blank}/{len(data)}")

    def seg(r):
        v = str(r[fid] or "").strip()
        if re.fullmatch(r"SYS-RA-DM-\d+", v):
            return "SYS-RA-DM-*"
        if re.fullmatch(r"SYS2-RA-\d+", v):
            return "SYS2-RA-*"
        return "other"

    for label, key in (("CASE-NORMALISED (lower)", lambda v: norm(v).lower()),
                       ("VERBATIM (case-sensitive)", norm)):
        print()
        print(f"## Category x id-segment — {label}")
        tab = Counter((key(r[cat]), seg(r)) for _, r in data)
        cats = sorted({c for c, _ in tab})
        print("| Category | SYS-RA-DM-* | SYS2-RA-* | other | total |")
        print("|---|---|---|---|---|")
        for c in cats:
            a1, a2, a3 = (tab[(c, "SYS-RA-DM-*")], tab[(c, "SYS2-RA-*")],
                          tab[(c, "other")])
            print(f"| {c} | {a1} | {a2} | {a3} | {a1 + a2 + a3} |")

    print()
    print("## case-variant rows (verbatim Category differs from its lower form)")
    lower_groups = {}
    for rn, r in data:
        lower_groups.setdefault(norm(r[cat]).lower(), Counter())[norm(r[cat])] += 1
    variants = {k: v for k, v in lower_groups.items() if len(v) > 1}
    total_variant_rows = 0
    for k, v in variants.items():
        minor = v.most_common()[1:]
        n = sum(c for _, c in minor)
        total_variant_rows += n
        print(f"  {k!r}: {dict(v)}  -> minority rows: {n}")
        for rn, r in data:
            if norm(r[cat]) in dict(minor):
                print(f"     r{rn} {r[fid]} {norm(r[cat])!r}")
    print(f"  total rows a verbatim gate would miscount: {total_variant_rows}")

    print()
    print("## SYS2 SW/HW/System distribution (verbatim)")
    for k, n in Counter(norm(r[swhw]) for _, r in data).most_common():
        print(f"  {k!r}: {n}")
    print("  SW rows:")
    for rn, r in data:
        if norm(r[swhw]) == "SW":
            print(f"     r{rn} {r[fid]} | {norm(r[cat])} | "
                  f"{norm(r[col('Description')])[:90]}")

    print()
    toks = set()
    for _, r in data:
        for t in re.split(r"[,\s;]+", str(r[melco] or "")):
            if t.strip():
                toks.add(t.strip())
    print(f"## SYS2 Melco ID distinct tokens: {len(toks)}")
    excluded = ["PSCFTS020-1-45-1", "PSCFTS020-1-45-2", "PSCFTS020-1-45-3",
                "PSCFTS020-1-45-4", "PSCFTS020-1-56-9", "PSCFTS020-1-56-10",
                "PSCFTS020-1-2-7", "PSCFTS020-1-2-8"]
    hit = sum(1 for e in excluded if e in toks)
    print(f"037 Excluded-NRL values found in Melco ID tokens: {hit}/8")
    for e in excluded:
        print(f"   {e}: {'HIT' if e in toks else 'MISS'}")


if __name__ == "__main__":
    main()
