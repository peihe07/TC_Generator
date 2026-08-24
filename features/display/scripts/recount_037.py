#!/usr/bin/env python3
"""037 A03 SWRA independent recount (handoff 01 step 6).

Measurement conditions are declared in the output, not assumed:
  - engine: openpyxl, data_only=True
  - mode:   NON-read-only full scan (read-only max_row reports the declared
            sheet dimension, which this file inflates with residual styling)
  - empty:  a row counts as data iff at least one cell in A..max_column has
            str(value).strip() != ""
  - case:   string comparison is case-SENSITIVE
"""
import re
import sys
from pathlib import Path

import openpyxl

WB = Path(__file__).resolve().parents[1] / "inputs" / \
    "Display_Management_FM-WI-FSM-037-A03_STLA_Report_SWRA.xlsx"


def nonempty_rows(ws):
    """Row indices whose cells are not all blank."""
    out = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                out.append(r)
                break
    return out


def main():
    wb_ro = openpyxl.load_workbook(WB, read_only=True, data_only=True)
    ro_dims = {ws.title: (ws.max_row, ws.max_column) for ws in wb_ro.worksheets}
    wb_ro.close()

    wb = openpyxl.load_workbook(WB, read_only=False, data_only=True)
    print("# 037 recount — measurement conditions")
    print("engine=openpyxl data_only=True | mode=NON-read-only full scan")
    print("empty-row rule=all cells in A..max_column blank after str().strip()")
    print("string compare=case-SENSITIVE")
    print(f"file={WB.name}")
    print()
    print("| sheet | read_only max_row | non-read-only max_row | non-empty rows | row indices |")
    print("|---|---|---|---|---|")
    rows = {}
    for ws in wb.worksheets:
        ne = nonempty_rows(ws)
        rows[ws.title] = ne
        idx = ",".join(str(i) for i in ne) if ne else "—"
        print(f"| {ws.title} | {ro_dims[ws.title][0]} | {ws.max_row} | {len(ne)} | {idx} |")

    ws = wb["SWE1 Requirements"]
    # The first non-empty row is a bilingual title banner, not the header.
    # Locate the header by the column label itself rather than by a row
    # number taken from the handoff.
    hdr_row = next(r for r in rows["SWE1 Requirements"]
                   if any(str(ws.cell(r, c).value or "").strip()
                          == "SWE-Requirement ID"
                          for c in range(1, ws.max_column + 1)))
    hdr = [ws.cell(hdr_row, c).value for c in range(1, ws.max_column + 1)]
    data_rows = [r for r in rows["SWE1 Requirements"] if r > hdr_row]
    print()
    print(f"header row (first non-empty) = r{hdr_row}; data rows = "
          f"r{data_rows[0]}–r{data_rows[-1]} ({len(data_rows)})")
    print("columns (RAW, repr — note irregular whitespace):")
    for i, h in enumerate(hdr, 1):
        if h is not None:
            print(f"   c{i}: {h!r}")

    def norm(s):
        return " ".join(str(s or "").split())

    def col(name):
        """Column index by whitespace-normalised header label.

        The raw headers carry trailing and doubled spaces
        ('SWE-Requirement ID ', 'Requirement  Title'), so an exact-string
        lookup misses every one of them.
        """
        hits = [i for i, h in enumerate(hdr, 1) if norm(h) == name]
        if len(hits) != 1:
            raise SystemExit(f"header {name!r}: {len(hits)} matches, expected 1")
        return hits[0]

    id_col = col("SWE-Requirement ID")
    src_col = col("Source Requirement ID")
    cat_col = col("Categorization")
    sub_col = col("Sub Categorization")
    title_col = col("Requirement Title")

    hit_id = sum(bool(re.fullmatch(r"SWE-DM-\d{3}",
                 str(ws.cell(r, id_col).value or "").strip())) for r in data_rows)
    hit_src = sum(bool(re.fullmatch(r"SYS-DISP-\d{3}",
                  str(ws.cell(r, src_col).value or "").strip())) for r in data_rows)
    cats = [str(ws.cell(r, cat_col).value or "").strip() for r in data_rows]
    subs = [str(ws.cell(r, sub_col).value or "").strip() for r in data_rows]
    print()
    print(f"SWE-Requirement ID matches ^SWE-DM-\\d{{3}}$ : {hit_id}/{len(data_rows)}")
    print(f"Source Requirement ID matches ^SYS-DISP-\\d{{3}}$ : {hit_src}/{len(data_rows)}")
    print(f"Categorization distinct: {sorted(set(cats))} "
          f"(Functional Requirement {cats.count('Functional Requirement')}/{len(data_rows)})")
    print(f"Sub Categorization distinct count: {len(set(subs))}")
    print()
    print("| ID | Sub Categorization | Requirement Title |")
    print("|---|---|---|")
    for r in data_rows:
        print(f"| {ws.cell(r, id_col).value} | {ws.cell(r, sub_col).value} "
              f"| {ws.cell(r, title_col).value} |")

    tr = wb["SYS2 Traceability"]
    tr_rows = rows["SYS2 Traceability"]
    tr_hdr = [tr.cell(tr_rows[0], c).value for c in range(1, tr.max_column + 1)]
    print()
    print(f"SYS2 Traceability header r{tr_rows[0]} (RAW): "
          + " / ".join(repr(h) for h in tr_hdr if h is not None))
    nrl_hits = [i for i, h in enumerate(tr_hdr, 1)
                if norm(h) == "Source NRL ID(s)"]
    if len(nrl_hits) != 1:
        raise SystemExit(f"Source NRL ID(s): {len(nrl_hits)} matches")
    nrl_col = nrl_hits[0]
    tr_data = tr_rows[1:]
    empty_nrl = sum(1 for r in tr_data
                    if str(tr.cell(r, nrl_col).value or "").strip() == "")
    print(f"Source NRL ID(s) EMPTY: {empty_nrl}/{len(tr_data)}")
    for r in tr_data:
        print("   r%d: %s" % (r, " | ".join(
            str(tr.cell(r, c).value) for c in range(1, tr.max_column + 1))))

    ex = wb["Excluded NRLs (HW-only)"]
    ex_rows = rows["Excluded NRLs (HW-only)"]
    ex_hdr = [ex.cell(ex_rows[0], c).value for c in range(1, ex.max_column + 1)]
    print()
    print(f"Excluded NRLs header r{ex_rows[0]} (RAW): "
          + " / ".join(repr(h) for h in ex_hdr if h is not None))
    for r in ex_rows[1:]:
        print("   r%d: %s" % (r, " | ".join(
            str(ex.cell(r, c).value) for c in range(1, ex.max_column + 1))))
    wb.close()


if __name__ == "__main__":
    sys.exit(main())
