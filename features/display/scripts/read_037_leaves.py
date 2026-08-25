#!/usr/bin/env python3
"""037 leaf close-read (handoff 06 step 8) — four rounds overdue.

Prints each of the 8 leaves in full: title, sub-categorisation, the whole
Requirement Description verbatim, the concrete values it does or does not
carry, the external documents it cites, and which SYS2 rows the EXISTING
anchors tie it to.

Produces no TC and makes no scope ruling (handoff 06 step 8).
"""
import csv
import re
from pathlib import Path

import openpyxl

FEAT = Path(__file__).resolve().parents[1]
F037 = FEAT / "inputs" / \
    "Display_Management_FM-WI-FSM-037-A03_STLA_Report_SWRA.xlsx"

# a number with a unit, i.e. the shape of a threshold
VALUE_RX = re.compile(
    r"\b\d+(?:\.\d+)?\s*(?:ms|s\b|sec|second|min|minute|%|°|deg|C\b|Hz|fps)",
    re.I)
# `SD\.\w+` used to be in this list and matched "SD.The" out of
# "...from DCSD.The software..." — the source runs two sentences together
# with no space after the period, so a dotted-prefix pattern fires on the
# next sentence's first word. Every id form here is now anchored on a word
# boundary and requires a digit, which that shape cannot satisfy.
DOC_RX = re.compile(r"\{?\b(CFTS\d+(?:-\d+)*|SYS-RA-[A-Z]+-\d+|"
                    r"SYS-DISP-\d+|PSCFTS\d+(?:-\d+)*)\b\}?")
SIGNAL_RX = re.compile(r"\$([A-Za-z0-9_]+)\$")


def norm(s):
    return " ".join(str(s or "").split())


def main():
    wb = openpyxl.load_workbook(F037, data_only=True)
    ws = wb["SWE1 Requirements"]
    hdr = {norm(ws.cell(7, c).value): c for c in range(1, ws.max_column + 1)}
    tr = wb["SYS2 Traceability"]

    cov = {}
    cov_path = FEAT / "data" / "coverage_sys2_vs_swe_dm.tsv"
    if cov_path.exists():
        with cov_path.open(encoding="utf-8") as fh:
            for row in csv.DictReader(fh, delimiter="\t"):
                for lf in filter(None, row["candidate_leaf"].split(",")):
                    cov.setdefault(lf, []).append(row["sys2_row"])

    print("# 037 八條 leaf 之逐條精讀")
    print("來源：`SWE1 Requirements` r8–r15，`Requirement Description` 欄全文")
    print("引擎 openpyxl / data_only=True / 非唯讀；欄名以空白正規化後比對")
    print("**本輸出不產出 TC、不作範圍裁定**（下放包 06 步驟 8）")

    for r in range(8, 16):
        rid = norm(ws.cell(r, hdr["SWE-Requirement ID"]).value)
        desc = str(ws.cell(r, hdr["Requirement Description"]).value or "")
        print(f"\n\n## {rid}　（`SWE1 Requirements` r{r}）")
        print(f"- Sub Categorization：{norm(ws.cell(r, hdr['Sub Categorization']).value)}")
        print(f"- Requirement Title：{norm(ws.cell(r, hdr['Requirement Title']).value)}")
        print(f"- Source Requirement ID：{norm(ws.cell(r, hdr['Source Requirement ID']).value)}")
        print(f"- Priority：{norm(ws.cell(r, hdr['Priority']).value)}"
              f"　Verification Method：{norm(ws.cell(r, hdr['Verification Method']).value)}")
        tr_row = r - 6            # r8 -> Traceability r2
        print(f"- `SYS2 Traceability` r{tr_row}："
              f"{norm(tr.cell(tr_row, 1).value)} / NRL="
              f"{norm(tr.cell(tr_row, 2).value) or '（空）'} / "
              f"{norm(tr.cell(tr_row, 3).value)} / "
              f"{norm(tr.cell(tr_row, 4).value)}")
        print(f"\n### Requirement Description（逐字全文）\n")
        print(f"> {norm(desc)}")

        vals = VALUE_RX.findall(desc)
        docs = sorted(set(DOC_RX.findall(desc)))
        sigs = sorted(set(SIGNAL_RX.findall(desc)))
        print(f"\n### 本條所含之具體值\n")
        print(f"- 數值＋單位（門檻之形態）：**{len(VALUE_RX.findall(desc))}** 處"
              f"{'　→ ' + str(vals) if vals else '　—— 無'}")
        print(f"- `$Signal$` token：{sigs or '無'}")
        print(f"- 外部文件／id 引用：{docs or '無'}")
        # `[a-z]\.[A-Z]` missed SWE-DM-002's "from DCSD.The software" — the
        # character before the period is an upper-case D. Require a
        # following lower-case letter instead, so a real sentence start
        # is matched without firing on an all-caps abbreviation.
        runon = len(re.findall(r"[A-Za-z]\.[A-Z][a-z]", desc))
        print(f"- 句號後缺空格之接續（`x.Y`）：{runon} 處"
              f"{'　—— 任何以句號斷句之實作會把兩句併為一句' if runon else ''}")
        states = sorted(set(re.findall(r"\bDISPLAY_[A-Z]+\b", desc)))
        if states:
            print(f"- 037 自用之狀態名：{states}"
                  f"　（SYS2／DBC 側為 `DISP_ON`／`DISP_OFF`，**逐字不等**）")
        print(f"- 現有錨連到之 SYS2 列（`candidate_leaf`，"
              f"anchor_kind 見 coverage TSV）："
              f"{('r' + ', r'.join(cov[rid])) if rid in cov else '（無）'}")

        print(f"\n### `Verification Criteria` 欄（參考輸入，非權威 —— R-DM8）\n")
        vc = norm(ws.cell(r, hdr["Verification Criteria"]).value)
        print(f"> {vc or '（空）'}")
    wb.close()


if __name__ == "__main__":
    main()
