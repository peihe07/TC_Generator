#!/usr/bin/env python3
"""R-DM13 anchored coverage cross-reference: SYS2 FR rows vs SWE-DM leaves.

Replaces the bag-of-words version retracted by R-DM13 (see
data/coverage_sys2_vs_swe_dm.RETRACTED.tsv). Nothing here uses similarity,
fuzziness or scoring: every anchor is a VERBATIM comparison, and a row with
no verbatim anchor is recorded as having none (handoff 03 §七第 10 條).

Population: SYS2 `Basic Report` rows whose Category normalises to
'functional requirement' (80).

Anchors recorded per row (all of them, in their own columns):
  signals  — `$NAME$` tokens in Description
  values   — `[VALUE]` tokens in Description
  heading  — the nearest preceding Category=='heading' row (positional)
  melco    — Melco ID tokens that appear on 037's Excluded NRLs sheet

`anchor_kind` names the highest-priority anchor PRESENT, in the R-DM13
order: signal > value > heading > melco > none.

`candidate_leaf` is produced by the heading anchor and, since R-DM22, by
the glossary anchor. Those are the only anchors whose counterpart exists on
the 037 side: 037 carries no
signal-layer information (R-DM14), and Melco hits mark 037's HW exclusions,
never a leaf. The test is verbatim substring containment of a leaf phrase in
the heading text, where a leaf phrase is:

    Requirement Title split on ' - ' and ' & ', plus Sub Categorization,
    whitespace-normalised, lower-cased, segments shorter than
    MIN_PHRASE characters dropped

MIN_PHRASE is the one tunable in this file. It does not make the match
fuzzy — a phrase either occurs verbatim or it does not — it only stops
generic fragments from matching. Every produced candidate carries the
matched phrase verbatim in `note`, so any reader can check it by eye.

Per R-DM12 the result column is named candidate_leaf, never 對應/mapping,
and must not be cited without `anchor_kind` alongside it.
"""
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SYS2 = ROOT / "inputs" / ("SYS2_CFTS_020_DISP_TCH_ICS_20260616_All_HW_System"
                          "_Accepted & Released.xlsx")
F037 = ROOT / "inputs" / \
    "Display_Management_FM-WI-FSM-037-A03_STLA_Report_SWRA.xlsx"

MIN_PHRASE = 8

# R-DM22: a closed, per-entry-sourced abbreviation table is not similarity.
# glossary.tsv is built by scripts/build_glossary.py, which only records a
# pair when the abbreviation and its expansion stand side by side in one
# sentence of one source. Matching here is VERBATIM SUBSTRING, CASE
# SENSITIVE (R-DM22(b)), against the row's Description and its heading
# ancestor, and expansions of fewer than two words are refused (R-DM22(a)).
GLOSSARY_TSV = ROOT / "data" / "glossary.tsv"

# R-DM25, applied to this side too (handoff 08 step 4). Upstream 07 §11.6
# guessed that SYS2's Description, being prose, would rarely use
# underscores — a guess is not a measurement, so the normalised pass runs
# here as well and both counts are reported.
SEP_NORM = re.compile(r"[ _]+")


def sep_norm(s):
    return SEP_NORM.sub(" ", s)

SIGNAL = re.compile(r"\$([A-Za-z0-9_]+)\$")
# R-DM18 (supersedes R-DM16): extract with the wide form, then DROP any
# token containing ':'. The colon is the export's own verbatim marker for
# Polarion metadata ([Artifact Type:...], [State:...], [Market:...],
# [Radio:...], [EE Architecture:...]); it is not a spec value. Literal test,
# no similarity.
#
# Three of the surviving tokens are document/protocol names rather than
# values and are emitted separately as kind=document (R-DM18).
VALUE = re.compile(r"\[([^\]]+)\]")
# R-DM16's definition, KEPT for the record per R-TM13. ITS DEFINITION IS
# REPEALED — the column exists so the retracted numbers stay auditable, and
# must not be used as a value source.
VALUE_NARROW = re.compile(r"\[([A-Za-z0-9_%\s]+)\]")
DOCUMENT_TOKENS = {
    "DCSD_and_HU_LVDS_Backchannel_Protocol",
    "DCSD* and HU CAN and LVDS Backchannel Message Sequence Charts",
    "SD.xxxxx DCSD LVDS VIDEO COMMUNICATION INTERFACE",
}
# Tokens can contain commas ([Radio:R1M, VP5R120, R1H]). Serialising with a
# comma and splitting it back is what produced the retracted "44 distinct"
# figure in upstream 04 §0.2 — the extraction was never wrong, the
# aggregation was. Use a separator that cannot occur in the data.
SEP = " ¦ "


def norm(s):
    return " ".join(str(s or "").split())


def load_glossary():
    """abbrev -> expansion, for usable (>= 2 word) entries only."""
    out = {}
    if not GLOSSARY_TSV.exists():
        return out
    with GLOSSARY_TSV.open(encoding="utf-8") as fh:
        head = fh.readline().rstrip("\n").split("\t")
        for line in fh:
            d = dict(zip(head, line.rstrip("\n").split("\t")))
            if d.get("usable") == "Y" and len(d["expansion"].split()) >= 2:
                out[d["abbrev"]] = d["expansion"]
    return out


def leaves():
    wb = openpyxl.load_workbook(F037, data_only=True)
    ws = wb["SWE1 Requirements"]
    hdr = {norm(ws.cell(7, c).value): c for c in range(1, ws.max_column + 1)}
    out = []
    for r in range(8, 16):
        title = norm(ws.cell(r, hdr["Requirement Title"]).value)
        sub = norm(ws.cell(r, hdr["Sub Categorization"]).value)
        phrases = [s.strip() for s in re.split(r" - | & ", title)]
        phrases.append(sub)
        phrases = [p for p in phrases if len(p) >= MIN_PHRASE]
        out.append({"id": norm(ws.cell(r, hdr["SWE-Requirement ID"]).value),
                    "title": title, "sub": sub, "phrases": phrases})
    ex = wb["Excluded NRLs (HW-only)"]
    excluded = {norm(ex.cell(r, 1).value) for r in range(2, 10)}
    wb.close()
    return out, excluded


def sys2_rows():
    wb = openpyxl.load_workbook(SYS2, read_only=True, data_only=True)
    grid = [list(r) for r in wb["Basic Report"].iter_rows(values_only=True)]
    wb.close()
    head = [norm(h) for h in grid[0]]

    def col(name):
        hits = [i for i, h in enumerate(head) if h == name] or \
               [i for i, h in enumerate(head) if h.startswith(name)]
        assert len(hits) == 1, (name, hits)
        return hits[0]

    rows = [(i + 1, r) for i, r in enumerate(grid)
            if i >= 1 and str(r[0] or "").strip() != ""]
    return rows, col


def main():
    lv, excluded = leaves()
    gloss = load_glossary()
    rows, col = sys2_rows()
    c_fid, c_cat, c_desc = (col("SYS2 Sys-RA-Feature-ID"),
                            col("SYS2 分類 Category"), col("Description"))
    c_swhw, c_melco = col("SYS2 SW/HW/System"), col("SYS2 Melco ID")
    c_l1, c_l2, c_l3 = (col("SYS2 功能(一階) Function (Level 1)"),
                        col("SYS2 功能(二階) Function (Level 2)"),
                        col("SYS2 功能(三階) Function (Level 3)"))

    # positional heading ancestor, same rule as sys2_heading_tree.py
    ancestor, cur = {}, None
    for rn, r in rows:
        if norm(r[c_cat]).lower() == "heading":
            cur = (rn, norm(r[c_fid]), norm(r[c_desc]))
        ancestor[rn] = cur

    pop = [(rn, r) for rn, r in rows
           if norm(r[c_cat]).lower() == "functional requirement"]

    print("# R-DM13 anchored coverage cross-reference")
    print(f"population (Category=='functional requirement', case-normalised): "
          f"{len(pop)}")
    print(f"anchors are verbatim only; MIN_PHRASE={MIN_PHRASE} chars")
    print("candidate_leaf is a CANDIDATE (R-DM12): cite it only together "
          "with anchor_kind")
    print()
    print(f"glossary（R-DM22，usable=Y）: {len(gloss)} 條 -> "
          f"{sorted(gloss)}")
    print()
    print("## 037 leaf phrases used for the heading anchor")
    for lf in lv:
        print(f"  {lf['id']}: {lf['phrases']}")

    out_rows = []
    for rn, r in pop:
        desc = norm(r[c_desc])
        sigs = sorted(set(SIGNAL.findall(desc)))
        wide = {v.strip() for v in VALUE.findall(desc) if v.strip()}
        kept = {v for v in wide if ":" not in v}
        vals = sorted(kept - DOCUMENT_TOKENS)
        docs = sorted(kept & DOCUMENT_TOKENS)
        vals_n = sorted({v.strip() for v in VALUE_NARROW.findall(desc)
                         if v.strip()})
        anc = ancestor[rn]
        anc_txt = anc[2] if anc else ""
        mel = sorted({t.strip() for t in re.split(r"[,\s;]+", norm(r[c_melco]))
                      if t.strip()} & excluded)

        hits = []
        for lf in lv:
            m = [p for p in lf["phrases"] if p.lower() in anc_txt.lower()]
            if m:
                hits.append((lf["id"], m))

        # --- glossary anchor (R-DM22), case-sensitive verbatim substring
        gl_hits, gl_norm_hits = [], []
        fields = (("heading", anc_txt), ("description", desc))
        for lf in lv:
            got, got_norm = [], []
            for ab, exp in gloss.items():
                if not any(re.search(rf"\b{re.escape(ab)}\b", p)
                           for p in lf["phrases"]):
                    continue
                # (i) the expansion itself, (ii) each leaf phrase with the
                # abbreviation substituted
                cands = [exp] + [re.sub(rf"\b{re.escape(ab)}\b", exp, p_)
                                 for p_ in lf["phrases"]
                                 if re.search(rf"\b{re.escape(ab)}\b", p_)]
                for c in dict.fromkeys(cands):
                    for field, txt in fields:
                        if c in txt:                             # strict
                            got.append(f"{ab}->{c!r} @{field}")
                        elif sep_norm(c) in sep_norm(txt):       # R-DM25
                            got_norm.append(f"{ab}->{c!r} @{field} (norm)")
            if got:
                gl_hits.append((lf["id"], sorted(set(got))))
            if got_norm:
                gl_norm_hits.append((lf["id"], sorted(set(got_norm))))

        # R-DM26: heading drops to next-to-last. It is present on 80/80
        # rows, so as a high-priority anchor it masked every anchor below it
        # — that is why glossary_phrase never surfaced here (upstream 06
        # §3.2). Its largest node also holds 48 of the 80 FR rows.
        #   signal > value > glossary_phrase > glossary_phrase_norm
        #          > melco > heading > none
        if sigs:
            kind = "signal"
        elif vals:
            kind = "value"
        elif gl_hits:
            kind = "glossary_phrase"
        elif gl_norm_hits:
            kind = "glossary_phrase_norm"
        elif mel:
            kind = "melco"
        elif anc_txt:
            kind = "heading"
        else:
            kind = "none"

        cand_parts, notes = [], []
        if hits:
            cand_parts += [h[0] for h in hits]
            notes.append("heading 錨逐字含 leaf 片語：" + "；".join(
                f"{i}←{'|'.join(repr(x) for x in m)}" for i, m in hits))
        else:
            notes.append("heading 錨無逐字 leaf 片語"
                         if anc_txt else "無 heading 祖先")
        if gl_hits:
            cand_parts += [h[0] for h in gl_hits]
            notes.append("glossary 錨（R-DM22，嚴格，區分大小寫）：" + "；".join(
                f"{i}←{'|'.join(m)}" for i, m in gl_hits))
        else:
            notes.append("glossary 錨（嚴格）無命中")
        if gl_norm_hits:
            cand_parts += [h[0] for h in gl_norm_hits]
            notes.append("glossary 錨（R-DM25 正規化後才成立）：" + "；".join(
                f"{i}←{'|'.join(m)}" for i, m in gl_norm_hits))
        cand = SEP.join(sorted(set(cand_parts)))
        note = "；".join(notes)
        if mel:
            note += f"；Melco 命中 037 Excluded（HW 排除項）{mel}"

        out_rows.append({
            "sys2_row": rn, "sys_ra_id": norm(r[c_fid]),
            "category": norm(r[c_cat]), "swhw": norm(r[c_swhw]),
            "heading_ancestor": f"r{anc[0]} {anc[2]}" if anc else "",
            "signals": SEP.join(sigs), "values": SEP.join(vals),
            "documents": SEP.join(docs),
            "values_narrow_REPEALED": SEP.join(vals_n),
            "melco": SEP.join(mel), "anchor_kind": kind,
            "candidate_leaf": cand,
            # R-DM23: an empty candidate_leaf here means (3) — every anchor
            # was applied and none reached, i.e. NOT reached, not verified
            # absent. It is never (1) 查無 and never (2) 未追查.
            "empty_semantics": "" if cand else
                "(3) 方法之界線：heading／glossary 兩錨皆已施用而未接上；"
                "非「已查證不存在」，亦非「未追查」",
            "candidate_from": SEP.join(
                (["heading"] if hits else []) +
                (["glossary_phrase"] if gl_hits else []) +
                (["glossary_phrase_norm"] if gl_norm_hits else [])),
            "note": note,
            "func_l1": norm(r[c_l1]), "func_l2": norm(r[c_l2]),
            "func_l3": norm(r[c_l3]),
        })

    cols = ["sys2_row", "sys_ra_id", "category", "swhw", "heading_ancestor",
            "signals", "values", "documents", "values_narrow_REPEALED",
            "melco", "anchor_kind", "candidate_leaf", "candidate_from",
            "empty_semantics", "note", "func_l1", "func_l2", "func_l3"]
    out = ROOT / "data" / "coverage_sys2_vs_swe_dm.tsv"
    with out.open("w", encoding="utf-8") as fh:
        fh.write("\t".join(cols) + "\n")
        for d in out_rows:
            fh.write("\t".join(str(d[c]) for c in cols) + "\n")

    print()
    print("## candidate_from 分布（哪一種錨產生了候選）")
    from collections import Counter as _C2
    print("  heading only      : "
          f"{sum(1 for d in out_rows if d['candidate_from'] == 'heading')}")
    print("  glossary only（嚴格）: "
          f"{sum(1 for d in out_rows if d['candidate_from'] == 'glossary_phrase')}")
    print("  glossary_norm only（R-DM25 正規化後才成立）: "
          f"{sum(1 for d in out_rows if d['candidate_from'] == 'glossary_phrase_norm')}")
    print("  多錨並存          : "
          f"{sum(1 for d in out_rows if SEP in d['candidate_from'])}")
    print("  無候選            : "
          f"{sum(1 for d in out_rows if not d['candidate_from'])}")

    print()
    print("## anchor_kind 分布（最高優先之現存錨；R-DM26 新序：")
    print("##   signal > value > glossary_phrase > glossary_phrase_norm "
          "> melco > heading > none）")
    from collections import Counter
    for k, n in Counter(d["anchor_kind"] for d in out_rows).most_common():
        print(f"  {k}: {n}")
    print()
    print("## 各錨之存在數（非互斥，逐列獨立計）")
    print(f"  含 $signal$        : {sum(1 for d in out_rows if d['signals'])}")
    print(f"  含 [value]（R-DM18） : "
          f"{sum(1 for d in out_rows if d['values'])}")
    print(f"  有 heading 祖先    : "
          f"{sum(1 for d in out_rows if d['heading_ancestor'])}")
    print(f"  Melco 命中 Excluded: {sum(1 for d in out_rows if d['melco'])}")
    from collections import Counter as _C
    sig = sorted({s for d in out_rows for s in d["signals"].split(SEP) if s})
    # Two counts, because they are different questions and the handoff's
    # table gives the first: occurrences = every match in every row;
    # rows = how many FR rows carry the token at least once.
    wide_all, val_occ, val_row, doc_occ, doc_row = set(), _C(), _C(), _C(), _C()
    for rn, r in pop:
        all_toks = [v.strip() for v in VALUE.findall(norm(r[c_desc]))
                    if v.strip()]
        wide_all |= set(all_toks)
        for v in all_toks:
            if ":" in v:
                continue
            (doc_occ if v in DOCUMENT_TOKENS else val_occ)[v] += 1
        for v in set(all_toks):
            if ":" in v:
                continue
            (doc_row if v in DOCUMENT_TOKENS else val_row)[v] += 1
    print(f"  相異訊號名 {len(sig)}: {sig}")
    print(f"  [VALUE] 擷取（R-DM18）——")
    print(f"    寬式 \\[([^\\]]+)\\] 相異        : {len(wide_all)}")
    print(f"    其中含 ':'（Polarion metadata）: "
          f"{sum(1 for v in wide_all if ':' in v)}")
    print(f"    不含 ':'                      : "
          f"{sum(1 for v in wide_all if ':' not in v)}")
    print(f"      -> 值 token (kind=value)    : {len(val_occ)}")
    print(f"         {'出現次數':>8} {'列數':>6}  token")
    for v, n in sorted(val_occ.items(), key=lambda x: (-x[1], x[0])):
        print(f"         {n:>8} {val_row[v]:>6}  {v}")
    print(f"      -> 文件／協定名 (kind=document): {len(doc_occ)}")
    for v, n in sorted(doc_occ.items(), key=lambda x: (-x[1], x[0])):
        print(f"         {n:>8} {doc_row[v]:>6}  {v}")
    rows_v = sum(1 for d in out_rows if d["values"] or d["documents"])
    print(f"    至少含一個不含 ':' token 之 FR 列: {rows_v}")
    print(f"  values_narrow_REPEALED 欄：R-DM16 之定義，已由 R-DM18 廢止，"
          f"保留供稽核（R-TM13），不得作為值域來源")

    print()
    print("## candidate_leaf 分布（候選，非裁定）")
    cnt = Counter()
    for d in out_rows:
        if d["candidate_leaf"]:
            for x in d["candidate_leaf"].split(SEP):
                cnt[x] += 1
    for lf in lv:
        print(f"  {lf['id']} ({lf['sub']}): {cnt.get(lf['id'], 0)}")
    print(f"  有候選之列: {sum(1 for d in out_rows if d['candidate_leaf'])}")
    print(f"  無候選之列: {sum(1 for d in out_rows if not d['candidate_leaf'])}"
          f"　—— R-DM23 語意別 **(3) 方法之界線**（兩錨皆施用而未接上），"
          f"非查無、非未追查")

    print()
    print("| sys2_row | sys_ra_id | swhw | heading_ancestor | anchor_kind "
          "| candidate_leaf | note |")
    print("|---|---|---|---|---|---|---|")
    for d in out_rows:
        print(f"| r{d['sys2_row']} | {d['sys_ra_id']} | {d['swhw']} "
              f"| {d['heading_ancestor']} | {d['anchor_kind']} "
              f"| {d['candidate_leaf'] or '（無）'} | {d['note']} |")
    print(f"\nwrote {out}")


if __name__ == "__main__":
    main()
