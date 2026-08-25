#!/usr/bin/env python3
"""R-DM20 PROXI candidate survey (handoff 05 step 5).

Population: every data row of LID `Proxi & Configuration` (446 rows).
For each, three questions are answered by VERBATIM test only — no
similarity, no semantic nearness, no human inference (停止條件 14):

  A. does the PROXI Format sheet define this parameter?
     -> R-G13(2): the LID's left column is a Logical Identifier, NOT the
        PROXI parameter name. The PROXI-side name lives in the LID's
        `Atlantis & Atlantis High` group `Signal Name` column (c16, whose
        `CAN` column reads PROXI), e.g. LID `ACN_Hardwired` -> PROXI
        `ACN_Hardwire`. Looking up by the Logical Identifier alone is the
        same error as querying a DBC with `ICSPowerButton` instead of
        `Radio_btn0` (upstream 04). Keys are tried in order — Atlantis
        `Signal Name`, `Object Text`, then the Logical Identifier — and the
        key that matched is recorded in `lookup_key`.
  B. does the LID row name CFTS_020 as its Primary CFTS Usage?
     -> `CFTS020` as a verbatim substring of column 27
  C. can it be tied to a SWE-DM leaf?
     -> a leaf phrase (R-DM13's phrase set: Requirement Title split on
        ' - ' and ' & ', plus Sub Categorization, >= 8 chars) occurring
        verbatim in the Logical Identifier, Function, or Object Text

anchor_kind names the highest-priority anchor present:
  leaf_phrase > cfts_usage > proxi_param > none
related_leaf is populated ONLY by C, because A and B tie a parameter to a
sheet or a document, never to a leaf.

The `keyword_note` column carries a disclosure-only keyword scan
(DISP/DCSD/RVC/Camera/Display/Touch/Screen). It is NOT an anchor and never
feeds related_leaf; it exists so a reader can see what a verbatim test
cannot reach.

Nothing here is written to any TC field (R-DM20 末段).
"""
import re
from pathlib import Path

import openpyxl

from tsv_meta import write_meta

ROOT = Path(__file__).resolve().parents[3]
FEAT = Path(__file__).resolve().parents[1]
LID = ROOT / "forms" / "Logical Identifiers and CAN Mapping v1_78.xlsx"
PROXI = ROOT / "forms" / "PROXI_HDCC27_R3_20250424.xlsx"
F037 = FEAT / "inputs" / \
    "Display_Management_FM-WI-FSM-037-A03_STLA_Report_SWRA.xlsx"

MIN_PHRASE = 8
GLOSSARY_TSV = FEAT / "data" / "glossary.tsv"

# R-DM25: underscore <-> space, applied to BOTH sides before comparing.
# This is not "relaxing one more level" (which R-DM22(c) still forbids) —
# it is a declared, symmetric, reversible character-class normalisation of
# the same kind as case folding. Only this one transformation is allowed;
# hyphens, dots and camelCase are out of scope until separately ruled.
# Strict and normalised hits are counted separately and never merged, and a
# candidate that exists only after normalisation is marked
# `glossary_phrase_norm` (R-DM25(b)(c)).
SEP_NORM = re.compile(r"[ _]+")


def sep_norm(s):
    """Collapse runs of underscore/space to a single space. R-DM25."""
    return SEP_NORM.sub(" ", s)
KEYWORDS = ["DISP", "DCSD", "RVC", "Camera", "Display", "Touch", "Screen"]
SEP = " ¦ "


def norm(s):
    return " ".join(str(s or "").split())


def load_glossary():
    out = {}
    if GLOSSARY_TSV.exists():
        with GLOSSARY_TSV.open(encoding="utf-8") as fh:
            head = fh.readline().rstrip("\n").split("\t")
            for line in fh:
                d = dict(zip(head, line.rstrip("\n").split("\t")))
                if d.get("usable") == "Y" and len(d["expansion"].split()) >= 2:
                    out[d["abbrev"]] = d["expansion"]
    return out


def leaf_phrases():
    wb = openpyxl.load_workbook(F037, data_only=True)
    ws = wb["SWE1 Requirements"]
    hdr = {norm(ws.cell(7, c).value): c for c in range(1, ws.max_column + 1)}
    out = []
    for r in range(8, 16):
        title = norm(ws.cell(r, hdr["Requirement Title"]).value)
        sub = norm(ws.cell(r, hdr["Sub Categorization"]).value)
        ph = [s.strip() for s in re.split(r" - | & ", title)] + [sub]
        out.append((norm(ws.cell(r, hdr["SWE-Requirement ID"]).value),
                    [p for p in ph if len(p) >= MIN_PHRASE]))
    wb.close()
    return out


def atl_name_early(row, idx):
    """The Atlantis signal-name cell, whitespace-normalised, for matching."""
    return " ".join(str(row[idx] or "").split()) if len(row) > idx else ""


def main():
    wb = openpyxl.load_workbook(PROXI, read_only=True, data_only=True)
    pgrid = [list(r) for r in wb["Format"].iter_rows(values_only=True)]
    wb.close()
    phdr = [norm(v) for v in pgrid[1]]
    print("# R-DM20 PROXI candidate survey")
    print(f"PROXI `Format`: {len(pgrid)} 列 × {len(phdr)} 欄；r1 標題、"
          f"r2 欄名、資料自 r3 起")
    print(f"  欄名: {[h for h in phdr if h]}")
    pname_i, ptable_i = phdr.index("Parameter Name"), phdr.index("Table")
    pgroup_i = phdr.index("Parameter Group")
    prows = {}
    for i, r in enumerate(pgrid[2:], start=3):
        k = norm(r[pname_i])
        if k:
            prows.setdefault(k, []).append(
                (i, norm(r[ptable_i]), norm(r[pgroup_i])))
    print(f"  `Parameter Name` 非空之資料列: "
          f"{sum(len(v) for v in prows.values())}；相異參數名 {len(prows)}")

    wb = openpyxl.load_workbook(LID, read_only=True, data_only=True)
    lgrid = [list(r) for r in wb["Proxi & Configuration"].iter_rows(
        values_only=True)]
    wb.close()
    lhdr = [norm(v) for v in lgrid[2]]
    lgroups = {norm(v): i for i, v in enumerate(lgrid[1]) if v is not None}
    cfts_i = lhdr.index("Primary CFTS Usage")
    atl_i = lgroups["Atlantis & Atlantis High"]      # Signal Name column
    print(f"  `Atlantis & Atlantis High` 欄組起於 c{atl_i + 1}；"
          f"其 r3 欄名 {lhdr[atl_i:atl_i + 3]}")
    data = [(i + 1, r) for i, r in enumerate(lgrid) if i >= 3 and norm(r[0])]
    print(f"\nLID `Proxi & Configuration`: 資料列 {len(data)}（r4 起）")
    print(f"  `Primary CFTS Usage` 非空: "
          f"{sum(1 for _, r in data if norm(r[cfts_i]))}/{len(data)} "
          f"—— **該欄多數為空，故其未載 CFTS020 不構成「與本 feature 無關」"
          f"之證據**")

    leaves = leaf_phrases()
    gloss = load_glossary()
    print(f"\nglossary（R-DM22，usable=Y）: {len(gloss)} 條")
    kw = re.compile("|".join(re.escape(k) for k in KEYWORDS), re.I)

    cols = ["lid_row", "logical_identifier", "function",
            "atlantis_signal_name", "lookup_key", "proxi_row",
            "proxi_values", "related_leaf", "anchor_kind",
            "empty_semantics", "note", "keyword_note"]
    out, counts = [], {}
    for rn, r in data:
        lid_name = norm(r[0])
        func = norm(r[1])
        objt = norm(r[2]) if len(r) > 2 else ""
        hay = " | ".join([lid_name, func, objt]).lower()

        hits = []
        for lid_leaf, phs in leaves:
            m = [p for p in phs if p.lower() in hay]
            if m:
                hits.append((lid_leaf, m))

        # --- glossary anchor (R-DM22), case-sensitive verbatim substring
        hay_cs = " | ".join([lid_name, func, objt, atl_name_early(r, atl_i)])
        hay_norm = sep_norm(hay_cs)
        gl_hits, gl_norm_hits = [], []
        for lid_leaf, phs in leaves:
            got, got_norm = [], []
            for ab, exp in gloss.items():
                if not any(re.search(rf"\b{re.escape(ab)}\b", p) for p in phs):
                    continue
                cands = [exp] + [re.sub(rf"\b{re.escape(ab)}\b", exp, p_)
                                 for p_ in phs
                                 if re.search(rf"\b{re.escape(ab)}\b", p_)]
                for c in dict.fromkeys(cands):
                    if c in hay_cs:                       # strict
                        got.append(f"{ab}->{c!r}")
                    elif sep_norm(c) in hay_norm:         # only after R-DM25
                        got_norm.append(f"{ab}->{c!r} (norm)")
            if got:
                gl_hits.append((lid_leaf, sorted(set(got))))
            if got_norm:
                gl_norm_hits.append((lid_leaf, sorted(set(got_norm))))

        atl_raw = str(r[atl_i] or "") if len(r) > atl_i else ""
        atl_name = norm(atl_raw)
        # One cell may hold several names, newline-separated — the same shape
        # as CAN Mapping's multi-valued Signal Name. Collapsing them with
        # norm() produced a single unmatchable string ("Rear_View_Camera
        # Rear_View_Camera_Soft_Button"), so split first and try each.
        atl_keys = [norm(x.split(":")[0])
                    for x in atl_raw.splitlines() if x.strip()]
        cands = ([(k, "Atlantis Signal Name") for k in atl_keys]
                 + [(objt, "Object Text"), (lid_name, "Logical Identifier")])
        # A multi-valued cell can resolve to SEVERAL PROXI parameters
        # (RVC_SK_PRSNT -> Rear_View_Camera AND
        # Rear_View_Camera_Soft_Button). Taking only the first would hide
        # one of them, so every matching key is kept.
        pr, keys_hit, seen = [], [], set()
        for key, label in cands:
            if key and key in prows:
                for x in prows[key]:
                    if x[0] not in seen:
                        seen.add(x[0])
                        pr.append(x)
                keys_hit.append(f"{label}={key}")
        lookup_key = SEP.join(keys_hit)
        cfts = "CFTS020" in norm(r[cfts_i]).replace(" ", "")

        if hits:
            kind = "leaf_phrase"
        elif gl_hits:
            kind = "glossary_phrase"
        elif gl_norm_hits:
            kind = "glossary_phrase_norm"
        elif cfts:
            kind = "cfts_usage"
        elif pr:
            kind = "proxi_param"
        else:
            kind = "none"
        counts[kind] = counts.get(kind, 0) + 1

        note = []
        if hits:
            note.append("leaf 片語逐字命中：" + "；".join(
                f"{i}←{'|'.join(repr(x) for x in m)}" for i, m in hits))
        else:
            note.append("無 leaf 片語逐字命中")
        if gl_hits:
            note.append("glossary 錨命中（嚴格）：" + "；".join(
                f"{i}←{'|'.join(m)}" for i, m in gl_hits))
        if gl_norm_hits:
            note.append("glossary 錨命中（R-DM25 正規化後才成立）：" + "；".join(
                f"{i}←{'|'.join(m)}" for i, m in gl_norm_hits))
        if cfts:
            note.append(f"Primary CFTS Usage 逐字含 CFTS020"
                        f"（{norm(r[cfts_i])}）")
        if pr:
            note.append(f"PROXI Format 定義於 r{','.join(str(x[0]) for x in pr)}"
                        f"（Parameter Group: {pr[0][2] or '（空）'}；"
                        f"查詢鍵 {lookup_key}）")
        else:
            tried = [k for k in (atl_keys + [objt, lid_name]) if k]
            note.append("PROXI Format `Parameter Name` 逐字查無；"
                        f"已試之鍵 {tried}")

        khits = sorted({m.group(0) for m in kw.finditer(
            " ".join([lid_name, func, objt]))})
        out.append(dict(zip(cols, [
            rn, lid_name, func, atl_name, lookup_key,
            SEP.join(str(x[0]) for x in pr),
            SEP.join(x[1] for x in pr if x[1]),
            # R-DM33: supply-side matching stopped. Three attempts
            # (keyword adjacency, heading, Used-by-contains-ETM) all failed;
            # the last measured 8.0% inside vs 9.1% outside. The column is
            # kept for shape but is no longer filled from this side — a leaf
            # that needs a pre-condition looks the parameter up instead.
            # The anchors still RUN, and what they would have found stays in
            # `note`, so nothing measured is thrown away.
            "",
            kind,
            # R-DM23: `related_leaf` empty on this sheet is (2) — only the
            # three A-DM16 starting points were pursued; the rest were never
            # investigated. It is NOT a finding that they are unrelated.
            # R-DM33: every row is (2) now — the column is not filled from
            # this side at all, so no row can be (1) or (3) here.
            ("(2) 未追查：R-DM33 起本欄不由供給側填寫；"
             "leaf 需要前置條件時再逐一查 PROXI"
             + ("" if kind != "none" else "。且 PROXI Format 亦未查得其定義")),
            "；".join(note),
            ("僅揭露，非錨：" + SEP.join(khits)) if khits else ""])))

    p = FEAT / "data" / "proxi_candidates.tsv"
    with p.open("w", encoding="utf-8") as fh:
        # R-DM30: header first, no comment lines. Provenance goes to the
        # sidecar — a comment line here would be read AS the header by a
        # plain csv.DictReader (A-DM23).
        fh.write("\t".join(cols) + "\n")
        for d in out:
            fh.write("\t".join(str(d[c]).replace("\t", " ") for c in cols) + "\n")

    import hashlib

    def _sha(path):
        return hashlib.sha256(Path(path).read_bytes()).hexdigest()

    write_meta(
        p, cols, len(out),
        generated_by="features/display/scripts/proxi_candidates.py",
        # R-G15/R-G23: this index is only as current as the files it was
        # built from. Recording their sha256 here does not make it expire on
        # its own — verify_reference_binding.py is what notices — but it
        # makes the tie visible instead of implied.
        inputs=[{"file": str(LID.relative_to(ROOT)), "sha256": _sha(LID)},
                {"file": str(PROXI.relative_to(ROOT)), "sha256": _sha(PROXI)},
                {"file": str(F037.relative_to(FEAT)), "sha256": _sha(F037)}],
        rulings=["R-DM12", "R-DM13", "R-DM20", "R-DM22", "R-DM23", "R-DM25",
                 "R-DM33"],
        measurement_conditions=(
            "母體＝LID `Proxi & Configuration` 之資料列；R-DM25 正規化＝"
            "比對前對兩側同時施加 re.sub(r'[ _]+', ' ', s)，僅此一項"
            "（連字號、點號、駝峰切分不在範圍內）"),
        notes=("R-DM33：本檔保留為索引（177 列之值域已查得），"
               "`related_leaf` 欄不再由供給側填寫，全欄語意為 R-DM23 之 "
               "(2) 未追查。錨仍照跑，其結果留在 `note` 欄供 Phase 2 參考。"
               "anchor_kind=glossary_phrase 者為嚴格比對即成立；"
               "=glossary_phrase_norm 者為正規化後才成立，兩者不合併計數。"))

    print(f"\n## anchor_kind 分布")
    for k in ["leaf_phrase", "glossary_phrase", "glossary_phrase_norm",
              "cfts_usage", "proxi_param", "none"]:
        print(f"  {k}: {counts.get(k, 0)}")
    print(f"  合計 {len(out)}")
    print(f"\n  R-DM33：`related_leaf` 自本輪起不由供給側填寫，"
          f"故全 {len(out)} 列皆空，語意一律為 R-DM23 之 **(2) 未追查**。"
          f"\n  錨仍照跑，其結果留在 `note` 欄"
          f"（本輪 anchor_kind 非 none 者 "
          f"{sum(1 for d in out if d['anchor_kind'] != 'none')} 列）"
          f"，供 Phase 2 逐 leaf 查前置條件時參考。")
    print(f"\n  於 PROXI Format 逐字查得定義者: "
          f"{sum(1 for d in out if d['proxi_row'])}/{len(out)}")
    print(f"  keyword 命中（僅揭露）: "
          f"{sum(1 for d in out if d['keyword_note'])}")

    print("\n## A-DM16 之三個起點 —— 逐字查其 PROXI 列與值域")
    for name in ["DCSD_cfg", "DSP_SK_PRSNT", "RVC_SK_PRSNT"]:
        d = next((x for x in out if x["logical_identifier"] == name), None)
        print(f"\n### {name}")
        if d is None:
            print("  LID `Proxi & Configuration` 逐字查無此 Logical Identifier")
            continue
        print(f"  LID r{d['lid_row']} | Function: {d['function'] or '（空）'}")
        print(f"  Atlantis Signal Name: {d['atlantis_signal_name'] or '（空）'}"
              f" | 命中之查詢鍵: {d['lookup_key'] or '（無）'}")
        print(f"  PROXI Format 列: {d['proxi_row'] or '（查無）'}")
        print(f"  值域: {d['proxi_values'] or '（無）'}")
        print(f"  anchor_kind: {d['anchor_kind']}")
        print(f"  note: {d['note']}")

    print("\n## anchor_kind != none 之全列")
    print("| lid_row | logical_identifier | function | proxi_row "
          "| proxi_values | related_leaf | anchor_kind |")
    print("|---|---|---|---|---|---|---|")
    for d in out:
        if d["anchor_kind"] != "none":
            print(f"| r{d['lid_row']} | {d['logical_identifier']} "
                  f"| {d['function'] or '—'} | {d['proxi_row'] or '—'} "
                  f"| {d['proxi_values'] or '—'} "
                  f"| {d['related_leaf'] or '（無）'} | {d['anchor_kind']} |")
    print(f"\nwrote {p}")


if __name__ == "__main__":
    main()
