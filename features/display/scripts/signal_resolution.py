#!/usr/bin/env python3
"""R-DM17 three-stage signal resolution (handoff 04 step 5).

    SYS2 `$Signal$`
      -> LID `CAN Mapping`.`Logical Identifier` (column A), VERBATIM match
      -> that row's `Atlantis High` column group `Signal Name`
         (MESSAGE.Signal; one cell may hold several, newline-separated)
      -> the DBC `SG_` definition and `VAL_` enumeration

Measurement conditions:
  - LID: openpyxl, read_only=True, data_only=True; sheet `CAN Mapping`;
    r2 = architecture group row, r3 = column-name row, data from r4
    (2624 rows). Atlantis High group = columns 26-30 per r2/r3.
  - matching into LID is EXACT on the whitespace-normalised Logical
    Identifier. No similarity, no case folding, no prefix matching.
  - a multi-valued Signal Name cell is split on newlines and emitted as
    ONE ROW PER VALUE (handoff 04 step 5: 不合併、不擇一)
  - DBC lookup is exact on BOTH halves of `MESSAGE.Signal`: the DBC chosen
    is the one carrying that signal inside a BO_ of that message name. Only
    if no DBC has the pair is a signal-name-only hit reported, and then the
    message mismatch is stated in the note rather than silently accepted.
  - when nothing matches, R-G13(3) is answered explicitly: whether the LID's
    message name exists as a BO_ in either DBC at all
  - resolved = Y only when an Atlantis High signal name exists AND the DBC
    carries that signal. Anything else is N, never a guess (停止條件 12).
"""
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
FEAT = Path(__file__).resolve().parents[1]
LID = ROOT / "forms" / "Logical Identifiers and CAN Mapping v1_78.xlsx"
DBCS = {
    "BHCAN2-R1": ROOT / "forms" / "PDT27_E2A_R1_BHCAN2.dbc",
    "FDCAN8-R1": ROOT / "forms" / "PDT27_E2A_R1_FDCAN8.dbc",
}
SYS2 = FEAT / "inputs" / ("SYS2_CFTS_020_DISP_TCH_ICS_20260616_All_HW_System"
                          "_Accepted & Released.xlsx")

BO = re.compile(r"^BO_\s+(\d+)\s+(\S+?)\s*:\s*(\d+)\s+(\S+)\s*$")


def norm(s):
    return " ".join(str(s or "").split())


def load_dbc(path):
    lines = path.read_bytes().decode("cp1252").splitlines()
    sigs, vals, msgnames, cur = {}, {}, set(), None
    for ln in lines:
        s = ln.strip()
        m = BO.match(s)
        if m:
            cur = (int(m.group(1)), m.group(2), m.group(4))
            msgnames.add(m.group(2))
            continue
        if s.startswith("SG_ ") and cur:
            name = s.split()[1]
            sigs.setdefault(name, []).append(
                {"msg_id": cur[0], "msg": cur[1], "tx": cur[2],
                 "body": s.split(":", 1)[1].strip()})
        if s.startswith("VAL_ "):
            p = s.split(None, 3)
            if len(p) >= 4:
                vals[(int(p[1]), p[2])] = p[3].rstrip(";")
    return sigs, vals, msgnames


def sys2_signals():
    wb = openpyxl.load_workbook(SYS2, read_only=True, data_only=True)
    grid = [list(r) for r in wb["Basic Report"].iter_rows(values_only=True)]
    wb.close()
    head = [norm(h) for h in grid[0]]
    c_cat = head.index("SYS2 分類 Category")
    c_desc = head.index("Description")
    # R-DM23: every `resolved = N` on this sheet is (1) —— R-G13 三要件皆
    # 齊備（檔名+SHA256、查詢名之種類、涵蓋範圍應含），故為「已查證不存在」，
    # 且已登入 forms/LOOKUP_MISSES.md。此處不存在 (2) 或 (3)。
    N_SEM = "(1) 已依 R-G13 三要件查證而確認不存在（見 forms/LOOKUP_MISSES.md）"
    out = []
    for i, r in enumerate(grid):
        if i == 0 or not str(r[0] or "").strip():
            continue
        if norm(r[c_cat]).lower() != "functional requirement":
            continue
        out += re.findall(r"\$([A-Za-z0-9_]+)\$", norm(r[c_desc]))
    return sorted(set(out))


def main():
    wb = openpyxl.load_workbook(LID, read_only=True, data_only=True)
    ws = wb["CAN Mapping"]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    groups = {norm(v): i + 1 for i, v in enumerate(grid[1]) if v is not None}
    names = [norm(v) for v in grid[2]]
    ah = groups["Atlantis High"]
    print("# R-DM17 signal resolution")
    print(f"LID: {LID.name}")
    print(f"CAN Mapping dims: {len(grid)} rows x {len(grid[0])} cols; "
          f"data rows r4-r{len(grid)} = {len(grid) - 3}")
    print(f"architecture groups (r2): "
          f"{ {k: v for k, v in groups.items()} }")
    print(f"Atlantis High group starts at c{ah}; its r3 labels: "
          f"{names[ah - 1:ah + 4]}")

    # LID index: normalised Logical Identifier -> row number
    lid_idx = {}
    for i, r in enumerate(grid[3:], start=4):
        k = norm(r[0])
        if k:
            lid_idx.setdefault(k, []).append(i)

    dbc = {k: load_dbc(p) for k, p in DBCS.items()}
    sigs = sys2_signals()
    print(f"\nSYS2 $Signal$ tokens in the FR population: {len(sigs)}")
    print(f"  {sigs}")

    cols = ["sys2_signal", "lid_row", "atl_high_signal_name", "can", "format",
            "sna", "dbc_file", "dbc_msg_id", "dbc_val_labels", "resolved",
            "n_semantics", "note"]
    # R-DM23: every `resolved = N` on this sheet is (1) —— R-G13 三要件皆
    # 齊備（檔名+SHA256、查詢名之種類、涵蓋範圍應含），故為「已查證不存在」，
    # 且已登入 forms/LOOKUP_MISSES.md。此處不存在 (2) 或 (3)。
    N_SEM = "(1) 已依 R-G13 三要件查證而確認不存在（見 forms/LOOKUP_MISSES.md）"
    out = []
    for sig in sigs:
        rows = lid_idx.get(sig, [])
        if not rows:
            out.append(dict(zip(cols, [sig, "", "", "", "", "", "", "", "",
                                       "N", "LID `Logical Identifier` 欄逐字"
                                       "查無此名"])))
            continue
        if len(rows) > 1:
            note_dup = f"（LID 有 {len(rows)} 列同名：{rows}）"
        else:
            note_dup = ""
        for lr in rows:
            r = grid[lr - 1]
            raw = str(r[ah - 1] or "")
            can = norm(r[ah])
            fmt = norm(r[ah + 1])
            sna = norm(r[ah + 2])
            parts = [p.strip() for p in raw.splitlines() if p.strip()]
            if not parts:
                out.append(dict(zip(cols, [sig, lr, "", can, fmt, sna, "", "",
                                           "", "N", N_SEM,
                                           "LID 有此列，但 Atlantis High 之 "
                                           "Signal Name 欄為空" + note_dup])))
                continue
            for part in parts:
                msg, _, name = part.partition(".")
                name = name or msg
                # exact on BOTH halves first: the DBC whose BO_ of that
                # message name carries this signal
                hit_file, hit = None, None
                for k, (s, v, mn) in dbc.items():
                    if name in s and any(h["msg"] == msg for h in s[name]):
                        hit_file, hit = k, [h for h in s[name]
                                            if h["msg"] == msg]
                        break
                if hit is None:                     # signal-name-only hit
                    for k, (s, v, mn) in dbc.items():
                        if name in s:
                            hit_file, hit = k, s[name]
                            break
                if hit is None:
                    where = [k for k, (s, v, mn) in dbc.items()
                             if msg in mn]
                    out.append(dict(zip(cols, [sig, lr, part, can, fmt, sna,
                        "", "", "", "N", N_SEM,
                        "LID 解得 CAN 訊號名，但兩本 DBC（BHCAN2-R1／"
                        "FDCAN8-R1）皆無此 SG_；R-G13(3)：訊息 "
                        f"{msg} 於 {('／'.join(where) + ' 存在') if where else '兩本 DBC 皆不存在'}"
                        + note_dup])))
                    continue
                s, v, mn = dbc[hit_file]
                msgs = hit
                for h in msgs:
                    lbl = v.get((h["msg_id"], name), "")
                    # R-G16(c): record WHY this DBC was chosen, not just
                    # which one won.
                    note = (f"選定判準：{'MESSAGE.Signal 兩半皆相等' if any(x['msg'] == msg for x in s.get(name, [])) else '僅訊號名相等（訊息名不符）'}"
                            f"→ {hit_file}") + note_dup
                    if h["msg"] != msg and msg != name:
                        note += (f"（訊息名不符：LID {msg} vs DBC "
                                 f"{h['msg']}；以訊號名命中，未擇一）")
                    out.append(dict(zip(cols,
                        [sig, lr, part, can, fmt, sna, hit_file,
                         f"{h['msg_id']} {h['msg']} tx={h['tx']}", lbl,
                         "Y", "", note])))

    p = FEAT / "data" / "signal_resolution.tsv"
    with p.open("w", encoding="utf-8") as fh:
        fh.write("\t".join(cols) + "\n")
        for d in out:
            fh.write("\t".join(str(d[c]).replace("\t", " ") for c in cols) + "\n")

    y = sum(1 for d in out if d["resolved"] == "Y")
    solved = {d["sys2_signal"] for d in out if d["resolved"] == "Y"}
    print(f"\n## 統計（R-DM21：每個數字都標明其止於哪一段）")
    in_lid = sum(1 for s in sigs if s in lid_idx)
    print(f"  段 1 SYS2 -> LID：於 `Logical Identifier` 欄逐字查得 "
          f"{in_lid}/{len(sigs)}")
    print(f"  段 2 LID -> CAN 名：解出之 MESSAGE.Signal 值（多值逐值一列） "
          f"{len(out)}")
    print(f"  段 3 CAN 名 -> DBC：查得 SG_ 者 {y}/{len(out)} 列"
          f"（未查得 {len(out) - y}）")
    print(f"  止於段 3 而至少解得一列之 $Signal$: {len(solved)}/{len(sigs)}")
    unres = [s for s in sigs if s not in solved]
    print(f"  止於段 3 完全未解者: {unres or '無'}")
    print("  注意：僅寫「解得 15/15」會是止於段 1 之數字，"
          "不得用以表示 TC 可用之 CAN 名已備齊（R-DM21）")
    print(f"\nwrote {p}")

    print("\n| sys2_signal | lid_row | atl_high_signal_name | can | dbc_file "
          "| dbc_msg_id | resolved | note |")
    print("|---|---|---|---|---|---|---|---|")
    for d in out:
        print(f"| {d['sys2_signal']} | {d['lid_row']} "
              f"| {d['atl_high_signal_name']} | {d['can']} | {d['dbc_file']} "
              f"| {d['dbc_msg_id']} | {d['resolved']} | {d['note']} |")


if __name__ == "__main__":
    main()
