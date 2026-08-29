#!/usr/bin/env python3
"""T42b／T42c —— `Error_Code_List.xlsx` 之欄位全覽與錯誤碼台帳（下放包 29）。

- **T42b**（R-SU26(a)）：9 分頁逐頁之欄序、標頭、非空列數、值型態摘要
  （**含實例** —— PLAYBOOK (29)：`unique 120` 說明不了它是什麼）、用途標記。
- **T42c**：`Error Code List` 分頁之全部碼 → `ERROR_CODES.md`。
  閉合檢查：台帳碼數 = 分頁碼數，不符即 `sys.exit`（PLAYBOOK (31)）。

**用途標記為陳報，不是裁定** —— R-SU26(b) 之 `已用／不用／未定` 三態中，
本輪一律標 `未定（本輪陳報）`，由分析層裁。

Usage: python3 scripts/error_codes.py b   # 全覽（印至 stdout）
       python3 scripts/error_codes.py c   # 台帳（寫 ERROR_CODES.md）
"""
import re
import sys
import warnings
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from write_back_036 import FEAT  # noqa: E402

warnings.filterwarnings("ignore")
SRC = FEAT / "inputs" / "Error_Code_List.xlsx"
SHEET = "Error Code List"
LEDGER = FEAT / "ERROR_CODES.md"

# R-SU35(a) 之階段 → Test Set 對照。**只抄條文所載者**；
# 條文未載之階段一律填 `—（R-SU35(a) 未載）`，不由執行層推定。
STAGE_MAP = {
    "Precondition": "`USB Update`",
    "Package Header check & unpack": "`Integrity Verification`",
    "Rollback Protection": "`Update Agent`",
    "Security check": "`Integrity Verification`",
    "Install ( M-CPU )": "`Interruption Handling`／`Update Agent`",
    "Install ( M-CPU: Redbend )": "`Interruption Handling`／`Update Agent`",
    "Install ( V-CPU )": "`Interruption Handling`／`Update Agent`",
    "Install ( SXM )": "**不用**（非本 feature 範圍）",
}
NOT_ERROR = {"458760", "458763"}          # R-SU35(c)3：不作失敗判準
GEN1 = re.compile(r"\*?Not support at GEN1")


def _txt(v):
    return re.sub(r"\s+", " ", str(v)).strip() if v not in (None, "") else ""


def load():
    return openpyxl.load_workbook(SRC, read_only=True, data_only=True)


def parse_codes():
    """回傳 [(階段, 碼, Description, Root cause, Recovery, Contact)]。"""
    ws = load()[SHEET]
    stage, out, stages = None, [], []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] in (None, ""):
            continue
        head = _txt(r[0])
        if all(c in (None, "") for c in r[1:]):
            stage = head
            stages.append(head)
        elif r[1] not in (None, ""):
            out.append((stage, head, _txt(r[1]), _txt(r[2]), _txt(r[3]), _txt(r[4])))
    return out, stages


def t42b():
    wb = load()
    print("## T42b —— `Error_Code_List.xlsx` 欄位全覽（R-SU26(a)）\n")
    print(f"- 來源：`inputs/Error_Code_List.xlsx`｜sha256 `4625753c…`｜"
          f"**{len(wb.sheetnames)} 分頁**")
    print("- **用途標記一律為陳報（`未定（本輪陳報）`），不是裁定** —— R-SU26(b) 由分析層裁\n")
    print("| # | 分頁 | 非空列 | 欄數 | 標頭（欄序） | 用途（陳報） |")
    print("|---:|---|---:|---:|---|---|")
    ROLE = {
        "Error Code List": "**已用** —— R-SU35：負向路徑之錯誤碼定義",
        "ProvideSW_final": "疑為台架作業記錄（下放包 29 T42b 之預判）",
        "Flash Status": "疑為台架作業記錄（同上）",
        "Flash Record": "疑為台架作業記錄（同上）",
        "MD_IMAGE": "疑為台架作業記錄（同上）",
        "R1L_Need_Machine": "疑為台架作業記錄（同上）",
        "PROD_Parameter_Compare": "疑為台架作業記錄（同上）",
        "Model Code": "**⚠ 不在下放包 29 T42b 之預判清單內**",
        "Issue Mapping Version": "**⚠ 不在下放包 29 T42b 之預判清單內**",
    }
    detail = []
    for n, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        rows = [r for r in ws.iter_rows(values_only=True)
                if any(c not in (None, "") for c in r)]
        hdr = [_txt(c) for c in rows[0]] if rows else []
        ncol = max((len([c for c in r if c not in (None, "")]) for r in rows), default=0)
        print(f"| {n} | `{name}` | {len(rows)} | {len(hdr)} | "
              + "／".join(f"`{h}`" for h in hdr[:6] if h) + " | "
              + ROLE.get(name, "—") + " |")
        detail.append((name, rows, hdr))

    print("\n### 逐頁之值型態摘要（**含實例** —— PLAYBOOK (29)）\n")
    for name, rows, hdr in detail:
        print(f"\n#### `{name}`（非空 {len(rows)} 列）\n")
        print("| 欄 | 標頭 | 非空 | unique | 型態 | **實例（首個非空值）** |")
        print("|---:|---|---:|---:|---|---|")
        body = rows[1:]
        for i, h in enumerate(hdr):
            vals = [_txt(r[i]) for r in body if i < len(r)]
            ne = [v for v in vals if v]
            if not h and not ne:
                continue
            uniq = len(set(ne))
            kind = ("**常數欄**" if uniq == 1 else
                    "枚舉" if 0 < uniq <= 6 else
                    "數字狀" if ne and all(v.replace("-", "").isdigit() for v in ne) else
                    "自由文字")
            ex = ne[0][:56] + ("…" if ne and len(ne[0]) > 56 else "") if ne else "**(全空)**"
            print(f"| {i} | `{h or '(無標頭)'}` | {len(ne)} | {uniq} | {kind} | `{ex}` |")
    return 0


def t42c():
    codes, stages = parse_codes()
    ws_n = sum(1 for r in load()[SHEET].iter_rows(min_row=2, values_only=True)
               if r[0] not in (None, "") and r[1] not in (None, "")
               and not all(c in (None, "") for c in r[1:]))
    if len(codes) != ws_n:
        sys.exit(f"T42c：台帳碼數 {len(codes)} ≠ 分頁碼數 {ws_n} —— 停並回報")

    def ts_of(stage):
        """R-SU35(a) 之對照；階段標題含尾註（如 `*This function supports only…`），故用前綴比對。"""
        k = next((k for k in STAGE_MAP if stage.startswith(k)), None)
        return STAGE_MAP[k] if k else None

    unmapped = sorted({s for s, *_ in codes if ts_of(s) is None})
    lines = []
    A = lines.append
    A("# ERROR_CODES — `Error_Code_List.xlsx` 錯誤碼台帳（SW Update）")
    A("")
    A("來源：`inputs/Error_Code_List.xlsx` 分頁 `Error Code List`"
      "｜sha256 `4625753c…`｜**Pei 裁可用**（R-SU35，2026-08-28）。")
    A("")
    A("**本表為引用來源，不是錨點來源** —— `specification_reference` 仍走 CFTS057"
      "（R-SU35(b)2）；本表之引用記於 `reasoning`，"
      "格式 `Error_Code_List.xlsx <分頁> <碼>`。")
    A("")
    A("**碼值與 Description 一律 verbatim**（R-SU35(b)1），"
      "含原文之拼寫殘留（如 `335890` 之 `sessiion`）—— **不改正**，同 D-4 之處置。")
    A("")
    A("> ⚠ **碼有了，看碼的地方還沒有**（R-SU35(b)3）——")
    A("> 錯誤碼於 HU 上之呈現途徑為 **DR-SU2 v2(a)** 之未解項；")
    A("> 未答前，觀測步驟一律掛 `PENDING: DR-SU2`。二者不得混同。")
    A("")
    A("> ⚠ **本表為 USB／SWDL 路徑**（R-SU35(c)1）——")
    A("> **不得引用以充當 Wi-Fi FOTA session 之觀測面**（DR-SU2 v2(b) 未解）。")
    A("")
    A("---")
    A("")
    A("## 閉合檢查")
    A("")
    A("| 項 | 值 |")
    A("|---|---:|")
    A(f"| 分頁之碼列數 | **{ws_n}** |")
    A(f"| 本台帳之碼數 | **{len(codes)}** |")
    A(f"| 階段標題數 | **{len(stages)}** |")
    A(f"| 相符 | {'✅' if len(codes) == ws_n else '❌'} |")
    A("")
    A("### 階段別碼數")
    A("")
    A("| 階段（分頁原文） | 碼數 | **Test Set 候選（R-SU35(a)）** |")
    A("|---|---:|---|")
    for s in stages:
        n = sum(1 for c in codes if c[0] == s)
        A(f"| `{s}` | {n} | {ts_of(s) or '**—（R-SU35(a) 未載）**'} |")
    A("")
    if unmapped:
        A("> ### ⚠ **R-SU35(a) 之對照表未涵蓋下列階段**")
        A(">")
        for s in unmapped:
            n = sum(1 for c in codes if c[0] == s)
            A(f"> - `{s}` —— **{n} 碼**")
        A(">")
        A("> **執行層不推定其歸屬** —— 依 R-SU20(d)，"
          "「階段名與組名字面相近」是循環，不是依據。待分析層補裁。")
        A("")
    A("---")
    A("")
    A("## 逐碼")
    A("")
    A("| 碼 | Description（verbatim） | 階段 | 平台限定 | Test Set 候選 | 註 |")
    A("|---|---|---|---|---|---|")
    for stage, code, desc, root, rec, contact in codes:
        ts = ts_of(stage) or "**—（未載）**"
        plat = "**`*Not support at GEN1`**" if GEN1.search(desc) else "—"
        note = "**不作失敗判準**（R-SU35(c)3）" if code in NOT_ERROR else ""
        A(f"| `{code}` | {desc} | `{stage[:34]}` | {plat} | {ts} | {note} |")
    A("")
    A("---")
    A("")
    A("## 引用時之拘束（R-SU35 摘）")
    A("")
    A("1. ER 得寫「對應之 error code 被報告」並具體列碼"
      "（如 `error code 335890 is reported`）；**碼值 verbatim，不得自造或改寫**。")
    A("2. `specification_reference` **不列本表**；引用記於 `reasoning`。")
    A("3. 讀碼位置未定前，該觀測步驟掛 `PENDING: DR-SU2`。")
    A("4. 平台限定之碼（`*Not support at GEN1`，**"
      f"{sum(1 for c in codes if GEN1.search(c[2]))} 碼**）引用時須連同限定一併記。")
    A("5. `458760`／`458763`（`Not an actual error`）**不作失敗判準**。")
    A("6. **正向路徑不因本表而有觀測面**（R-SU35(d)）——"
      "本表解的是「失敗時看什麼」，不是「成功進行中看什麼」。")
    LEDGER.write_text("\n".join(lines) + "\n")

    print("## T42c —— `ERROR_CODES.md` 產出\n")
    print(f"- 碼數 **{len(codes)}**｜階段 **{len(stages)}**｜閉合檢查 **✅**"
          f"（台帳 {len(codes)} = 分頁 {ws_n}）")
    print(f"- 平台限定 `*Not support at GEN1`：**"
          f"{sum(1 for c in codes if GEN1.search(c[2]))} 碼**")
    print(f"- `Not an actual error`：**{sum(1 for c in codes if c[1] in NOT_ERROR)} 碼**")
    if unmapped:
        print(f"\n### ⚠ R-SU35(a) 未涵蓋之階段（**{len(unmapped)}** 個，"
              f"共 {sum(1 for c in codes if c[0] in unmapped)} 碼）\n")
        for s in unmapped:
            print(f"- `{s}` —— {sum(1 for c in codes if c[0] == s)} 碼")
        print("\n**執行層不推定其歸屬**（R-SU20(d)：字面相近是循環，不是依據）。")
    return 0


if __name__ == "__main__":
    want = set(sys.argv[1:]) or {"b", "c"}
    if "b" in want:
        t42b()
    if "c" in want:
        t42c()
