#!/usr/bin/env python3
"""T38a／T38c —— `C`／`E` 二欄之陳報、難類涵蓋盤點（下放包 25 §五）。

**陳報事實，不裁定**（下放包 25 §五 T38a）。

Usage: python3 scripts/ce_columns.py 38a 38c
"""

import re
import sys
import warnings
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from corpus_v2 import _rows_desc                                   # noqa: E402
from anchor_table import C_ID                                      # noqa: E402
from observability import classify, RE_EXTERNAL                    # noqa: E402
from stratified_gt import GT_A1, GT_B                              # noqa: E402
from verif_columns import _load, _txt, C_VC, NUM                   # noqa: E402
from layer2_close import SETS                                      # noqa: E402
from write_back_036 import MASTER, SHEET_NAME, FEAT                # noqa: E402

warnings.filterwarnings("ignore")
ROOT = FEAT.parent.parent
VC = ROOT / "features/vehicle_category"


def t38a():
    print("## T38a —— `C`／`E` 二欄之陳報（陳報事實，不裁定）\n")

    # ── (i) vehicle_category 之 C 欄填值來源
    print("### (i) `vehicle_category` 之 `C` 欄填值來源之追查\n")
    a037 = openpyxl.load_workbook(next(VC.glob("inputs/*037*.xlsx")),
                                  read_only=True, data_only=True)["Analysis Report"]
    hdr = [_txt(c) for c in next(a037.iter_rows(min_row=7, max_row=7, values_only=True))]
    print(f"`vehicle_category` 之 037：**{len(hdr)} 欄 rev D**，"
          f"其欄 2 為 **`{hdr[2]}`**（本 feature 之 037 為 **18 欄舊版面，無該欄**）。\n")
    r037 = {}
    for r in a037.iter_rows(min_row=8, values_only=True):
        if r[0] not in (None, ""):
            r037[_txt(r[0])] = dict(hmi_src=_txt(r[2]), sub=_txt(r[8]), cat=_txt(r[6]))

    sys1 = openpyxl.load_workbook(next(VC.glob("inputs/SYS1_*.xlsx")),
                                  read_only=True, data_only=True)["Basic Report"]
    srows = list(sys1.iter_rows(values_only=True))
    sh = [_txt(c) for c in srows[0]]
    i_id, i_src = sh.index("ID"), next(i for i, h in enumerate(sh) if "Source ID" in h)
    by_src = {_txt(r[i_src]): _txt(r[i_id]) for r in srows[1:] if r[0] not in (None, "")}

    book = next(VC.glob("output/*.xlsx"))
    ws = openpyxl.load_workbook(book, read_only=True, data_only=True)["Test Case Specification 測試用例規範"]
    tcs = [r for r in ws.iter_rows(min_row=10, values_only=True)
           if len(r) > 5 and r[3] not in (None, "")]

    print("**逐列比對（前 12 列）**：037 之 `HMI Source ID` → SYS1 之 `SYSRE_HMI_Source ID`"
          " → SYS1 之 `ID` → 036 之 `C`\n")
    print("| 036 `D`（037 列 id） | 037 `HMI Source ID`（尾） | SYS1 `ID` | 036 `C` | 相符 |")
    print("|---|---|---|---|:--:|")
    ok = miss = 0
    seen = set()
    for r in tcs:
        leaf = _txt(r[3])
        base = re.sub(r"-\d+$", "", leaf)
        src = r037.get(leaf, r037.get(base, {})).get("hmi_src", "")
        want = by_src.get(src, "")
        got = _txt(r[2])
        good = want and want == got
        ok += bool(good)
        miss += not good
        if len(seen) < 12:
            seen.add(leaf)
            print(f"| `{leaf}` | …`{src[-14:] or '(無)'}` | `{want or '—'}` | `{got}` | "
                  f"{'✅' if good else '❌'} |")
    print(f"\n**全 {len(tcs)} 列之比對：相符 {ok}／不符 {miss}** —— "
          f"{'**鏈路成立**' if miss == 0 else '**鏈路不完全成立**'}\n")
    print("即 `C` 之取值路徑為：\n")
    print("```\n037 之 `HMI Source ID`  ──match──▶  SYS1 `SYSRE_HMI_Source ID`\n"
          "                                          │\n"
          "                                          ▼\n"
          "                                    SYS1 `ID`（NRL-…）\n"
          "                                          │\n"
          "                                          ▼\n"
          "                                     036 之 `C` 欄\n```\n")

    # 純 Service 列
    subs = Counter(v["sub"] for v in r037.values())
    print(f"**`vehicle_category` 之 037 之 `Sub Categorization` 分佈**："
          + "／".join(f"`{k or '(空)'}` {v}" for k, v in subs.most_common()) + "\n")
    svc = [k for k, v in r037.items() if v["sub"] == "Service"]
    print(f"純 Service 列：**{len(svc)}** 列"
          + ("（**該 feature 無 Service 列** —— 故「純 Service 列怎麼填 `C`」"
             "**在該 feature 無實例可循**）" if not svc else ""))
    if svc:
        nohmi = [k for k in svc if not r037[k]["hmi_src"]]
        print(f"其中無 `HMI Source ID` 者：**{len(nohmi)}** 列")

    # ── (ii) E 欄
    print("\n### (ii) `E` 欄之標頭原文與已交付簿之填值\n")
    m = openpyxl.load_workbook(FEAT / "inputs" / MASTER, read_only=True, data_only=True)[SHEET_NAME]
    h9 = list(next(m.iter_rows(min_row=9, max_row=9, values_only=True)))
    print(f"- 標頭原文（`E9`）：`{_txt(h9[4])}`\n")
    print("| feature | 簿 | 資料列 | `E` 非空 |")
    print("|---|---|---:|---:|")
    tot = ne = 0
    for b in sorted(ROOT.glob("features/*/delivered/*.xlsx")) + sorted(ROOT.glob("features/*/output/*.xlsx")):
        try:
            w = openpyxl.load_workbook(b, read_only=True, data_only=True)
            s = next((w[n] for n in w.sheetnames if n.startswith("Test Case Specification")), None)
            if s is None:
                continue
            rs = [r for r in s.iter_rows(min_row=10, values_only=True)
                  if len(r) > 4 and r[3] not in (None, "")]
            if not rs:
                continue
            c = sum(1 for r in rs if r[4] not in (None, ""))
            tot += len(rs)
            ne += c
            f = b.parts[b.parts.index("features") + 1]
            print(f"| {f} | `{b.parent.name}/{b.name[:12]}…` | {len(rs)} | "
                  f"{'**' + str(c) + '**' if c else 0} |")
        except Exception:                                          # noqa: BLE001
            pass
    print(f"| **合計** | | **{tot}** | **{ne}** |")
    print(f"\n**`E` 欄於全部已交付／產出簿**{'皆為空' if ne == 0 else f'有 {ne} 格填值'}**。**")

    # ── (iii) DV／條件式格式
    print("\n### (iii) `C`／`E` 於 036 母本之 DV 與條件式格式\n")
    wb = openpyxl.load_workbook(FEAT / "inputs" / MASTER, data_only=True)
    ws2 = wb[SHEET_NAME]
    print("| 欄 | 標準 DV | x14 DV | 條件式格式 |")
    print("|---|---|---|---|")
    import zipfile
    raw = zipfile.ZipFile(FEAT / "inputs" / MASTER).read("xl/worksheets/sheet6.xml").decode("utf8", "replace")
    for L in ("C", "E"):
        dv = [str(d.sqref) for d in ws2.data_validations.dataValidation
              if re.search(rf"\b{L}\d", str(d.sqref))]
        x14 = re.findall(rf"<xm:sqref>[^<]*{L}\d[^<]*</xm:sqref>", raw)
        cf = re.findall(rf'<conditionalFormatting sqref="[^"]*{L}\d[^"]*"', raw)
        print(f"| `{L}` | {dv or '**無**'} | {x14 or '**無**'} | {cf or '**無**'} |")
    print(f"\n全簿之 `<conditionalFormatting` 計數：**"
          f"{len(re.findall(r'<conditionalFormatting', raw))}**（sheet6）")


def t38c():
    by, d, owner = _load()
    internal = {i for i in d if classify(d[i])[0]}
    has_ext_vc = lambda i: any(r.search(_txt(by[i][C_VC])) for r in RE_EXTERNAL.values())
    g105 = {i for i in internal if not has_ext_vc(i)}
    gt = {f"SWE1-FOTA-{k}" for k in GT_A1} | {f"SWE1-FOTA-{k}" for k in GT_B}

    print("\n\n## T38c —— 難類涵蓋之全案盤點（R-SU31(b) 之落地）\n")
    print("> 用途：分析層據以排後續批次之順序，並使**每批之難類涵蓋可見**。\n")
    print("| Test Set | 總列數 | 126 內部列 | **105 列** | 105 佔比 | 含 GT 之列 |")
    print("|---|---:|---:|---:|---:|---:|")
    rows = []
    for name, _ in SETS:
        ids = [i for i in d if owner[i] == name]
        n, ni, n5, ng = (len(ids), sum(1 for i in ids if i in internal),
                         sum(1 for i in ids if i in g105), sum(1 for i in ids if i in gt))
        rows.append((name, n, ni, n5, ng))
    for name, n, ni, n5, ng in sorted(rows, key=lambda x: -(x[3] / x[1] if x[1] else 0)):
        print(f"| `{name}` | {n} | {ni} | {'**' + str(n5) + '**' if n5 else 0} | "
              f"**{n5/n*100:.0f}%** | {ng} |")
    T = [sum(x[i] for x in rows) for i in (1, 2, 3, 4)]
    print(f"| **合計** | **{T[0]}** | **{T[1]}** | **{T[2]}** | "
          f"**{T[2]/T[0]*100:.0f}%** | **{T[3]}** |")

    hi = [r for r in rows if r[1] and r[3] / r[1] >= 0.6]
    lo = [r for r in rows if r[3] == 0]
    print(f"\n- **105 佔比 ≥60% 之組**：**{len(hi)}** —— "
          + "、".join(f"`{r[0]}`（{r[3]}/{r[1]}）" for r in
                      sorted(hi, key=lambda x: -(x[3] / x[1]))))
    print(f"- **105 列為 0 之組**：**{len(lo)}** —— "
          + "、".join(f"`{r[0]}`（{r[1]} 列）" for r in lo))
    print(f"- 含 GT 之列合計 **{T[3]}**（GT-A1 28 + GT-B 4 = 32；"
          f"{'相符' if T[3] == 32 else '**不符**'}）")
    return rows


if __name__ == "__main__":
    want = set(sys.argv[1:]) or {"38a", "38c"}
    if "38a" in want:
        t38a()
    if "38c" in want:
        t38c()
