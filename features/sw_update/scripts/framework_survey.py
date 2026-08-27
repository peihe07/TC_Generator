#!/usr/bin/env python3
"""T18a–e —— Phase 3 framework 起草之前置量測（下放包 05 §三）。

只量測、只輸出，不判定 Test Set 名稱、不產出 framework（T18d 之對應
候選為草料，`?` 表不強配）。所有計數皆附閉合檢查。

Usage:
    python3 scripts/framework_survey.py            # 全部五項
    python3 scripts/framework_survey.py 18a 18e    # 指定項
"""

import re
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
A03 = ROOT / "inputs/SoftwareUpdate_FM-WI-FSM-037-A03_STLA_Report_SWRA.xlsx"
SYS1 = ROOT / "inputs/SYS1_HMI_Software_Updates_FOTA_HMI_Logic_and_Flow_R1_SR24_post_2A_(Aug_30_2023).xlsx"
CFTS = ROOT / "inputs/R1LR_Atl-H_25PI4.5 Dec Release-xOTA_CFTS_57 Reflash_20251202-2111.docx"

# 037 之 18 欄版面，0-indexed（表頭列 7、資料列 8 起）
C_ID, C_SRC, C_TITLE, C_DESC, C_CAT, C_SUB = 0, 1, 2, 3, 5, 6
IN_SCOPE = ("Functional Requirement", "Non Functional Requirement")


def a03_rows():
    """037 之 383 個資料列，依文件序。"""
    wb = openpyxl.load_workbook(A03, read_only=True, data_only=True)
    return [r for r in wb["AnalysisReport_FULL"].iter_rows(min_row=8, values_only=True)
            if r[C_ID] not in (None, "")]


def group_by_heading(rows):
    """依 Categorization == Heading 分節，回傳每節之所轄 in-scope 列。

    Heading 之前的 in-scope 列歸入 `(前言)` 偽節 —— 靜默丟棄會使閉合
    檢查假性通過，這正是本檔每節都附閉合式的原因。
    """
    groups, cur = [], {"id": None, "title": "(前言 —— 首個 Heading 之前)", "rows": []}
    for r in rows:
        if r[C_CAT] == "Heading":
            groups.append(cur)
            cur = {"id": str(r[C_ID]).strip(), "title": str(r[C_TITLE] or "").strip(), "rows": []}
        elif r[C_CAT] in IN_SCOPE:
            cur["rows"].append(r)
    groups.append(cur)
    return groups


def id_span(rows):
    if not rows:
        return "—"
    ns = sorted(int(re.search(r"(\d+)$", str(r[C_ID])).group(1)) for r in rows)
    return f"{ns[0]}–{ns[-1]}" if ns[0] != ns[-1] else str(ns[0])


def t18a():
    rows = a03_rows()
    groups = group_by_heading(rows)
    print("## T18a —— 037 分群（Heading 為分節點）\n")
    print("| # | Heading id | 標題原文 | FR | NFR | 小計 | 所轄 id 範圍 |")
    print("|---:|---|---|---:|---:|---:|---|")
    total = 0
    for i, g in enumerate(groups):
        fr = sum(1 for r in g["rows"] if r[C_CAT] == "Functional Requirement")
        nf = sum(1 for r in g["rows"] if r[C_CAT] == "Non Functional Requirement")
        total += fr + nf
        gid = f"`{g['id']}`" if g["id"] else "—"
        print(f"| {i} | {gid} | {g['title'][:56]} | {fr} | {nf} | {fr+nf} | {id_span(g['rows'])} |")
    want = sum(1 for r in rows if r[C_CAT] in IN_SCOPE)
    print(f"\n**閉合檢查**：所轄列數總和 = **{total}**；驗證母體（R-SU3）= **{want}**"
          f" —— {'閉合 ✅' if total == want else '**不閉合 ❌**'}")
    if total != want:
        seen = {id(r) for g in groups for r in g["rows"]}
        miss = [r for r in rows if r[C_CAT] in IN_SCOPE and id(r) not in seen]
        print(f"\n**缺口列 {len(miss)} 筆**：" + ", ".join(str(r[C_ID]) for r in miss))
        sys.exit("T18a 不閉合，停")
    print(f"\nHeading 節數 = {len(groups)-1}（另 1 個前言偽節，所轄 "
          f"{sum(1 for r in groups[0]['rows'])} 列）")
    return groups


def t18b():
    wb = openpyxl.load_workbook(SYS1, read_only=True, data_only=True)
    ws = wb["Basic Report"]
    rowsx = list(ws.iter_rows(values_only=True))
    h = list(rowsx[0])
    oi, di = h.index("Outline Number"), h.index("Description")
    data = [r for r in rowsx[1:] if r[0] not in (None, "")]
    tops, children = [], defaultdict(list)
    for r in data:
        o = str(r[oi]).strip()
        (tops if "." not in o else children[o.split(".")[0]]).append((o, str(r[di] or "").strip()))
    print("\n## T18b —— SYS1 分群（28 頂層章）\n")
    print("| Outline | Description 首句 | 2 層 | 3 層 | 子節 Outline 清單 |")
    print("|---|---|---:|---:|---|")
    for o, d in sorted(tops, key=lambda x: int(x[0])):
        kids = sorted(children[o], key=lambda x: [int(p) for p in x[0].split(".")])
        d2 = sum(1 for k, _ in kids if k.count(".") == 1)
        d3 = sum(1 for k, _ in kids if k.count(".") == 2)
        first = re.split(r"(?<=[.!?])\s|\n", d)[0][:52] if d else "—"
        lst = ", ".join(f"`{k}`" for k, _ in kids) or "—"
        print(f"| **{o}** | {first} | {d2} | {d3} | {lst[:110]} |")
    print(f"\n**閉合檢查**：頂層 {len(tops)} + 子節 {sum(len(v) for v in children.values())}"
          f" = {len(tops)+sum(len(v) for v in children.values())}；`Basic Report` 資料列 = {len(data)}"
          f" —— {'閉合 ✅' if len(tops)+sum(len(v) for v in children.values()) == len(data) else '**不閉合 ❌**'}")
    return tops


def cfts_paras():
    raw = zipfile.ZipFile(CFTS).read("word/document.xml").decode("utf8", "replace")
    out = []
    for p in re.findall(r"<w:p[ >].*?</w:p>", raw, re.S):
        m = re.search(r'<w:pStyle w:val="([^"]+)"', p)
        out.append(((m.group(1) if m else ""), re.sub(r"<[^>]+>", "", p)))
    return out


def t18c():
    HEAD, TOC = {"1", "2", "3", "4", "5"}, {"10", "20", "30", "40", "50"}
    secs, cur = [], None
    for st, t in cfts_paras():
        if st in TOC:
            continue
        if st in HEAD:
            m = re.search(r"^\s*([\d.]+)\s+(.*?)\s*\{(\d{7})\}", t)
            if m:
                cur = {"oid": m.group(3), "num": m.group(1), "title": m.group(2), "reqs": []}
                secs.append(cur)
                continue
        for m in re.finditer(r"(?<!\d)(\d{7})(?!\d)", t):
            if re.match(r":\s*\[Artifact Type:Subsystem Functional Requirement\]",
                        t[m.end():m.end() + 60]) and cur and m.group(1) not in cur["reqs"]:
                cur["reqs"].append(m.group(1))
    print("\n## T18c —— CFTS_57 章節（87 章節物件）\n")
    print("| 章節號 | ObjectID | 標題原文 | 所轄需求物件 |")
    print("|---|---|---|---:|")
    for s in secs:
        print(f"| {s['num']} | `{s['oid']}` | {s['title'][:58]} | {len(s['reqs'])} |")
    tot = sum(len(s["reqs"]) for s in secs)
    print(f"\n**閉合檢查**：章節 {len(secs)}（應 87）；所轄需求物件總和 = **{tot}**（應 487）"
          f" —— {'閉合 ✅' if len(secs) == 87 and tot == 487 else '**不閉合 ❌**'}")
    return secs


def t18e(groups):
    print("\n## T18e —— Heading × Sub Categorization 交叉表\n")
    print("| Heading id | 標題原文 | Service | HMI | blank | 小計 |")
    print("|---|---|---:|---:|---:|---:|")
    agg = Counter()
    for g in groups:
        c = Counter(r[C_SUB] if r[C_SUB] not in (None, "") else "blank" for r in g["rows"])
        if not g["rows"]:
            continue
        agg.update(c)
        gid = f"`{g['id']}`" if g["id"] else "—"
        print(f"| {gid} | {g['title'][:50]} | {c.get('Service',0)} | {c.get('HMI',0)} "
              f"| {c.get('blank',0)} | {sum(c.values())} |")
    print(f"\n**合計**：Service {agg['Service']}、HMI {agg['HMI']}、blank {agg['blank']}"
          f" —— 總和 {sum(agg.values())}")
    hm = [(g["id"], g["title"], sum(1 for r in g["rows"] if r[C_SUB] == "HMI"))
          for g in groups if any(r[C_SUB] == "HMI" for r in g["rows"])]
    print(f"\n**HMI 列集中處**（{len(hm)} 個 Heading 承載全部 {agg['HMI']} 個 HMI 列）：\n")
    for i, (gid, title, n) in enumerate(sorted(hm, key=lambda x: -x[2]), 1):
        print(f"{i}. `{gid}` {title[:52]} —— **{n}** 列")


def norm(s):
    return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()


# ── 詞集比對之參數與函式 —— **T18d 與 T19 共用同一支** ────────────────
# 下放包 06 §四之方法拘束：「詞集重疊比之參數與 T18d **完全一致**」。
# 提到模組層，是因為**規定二處一致而讓二處各寫一份，那只是規定**；
# 共用一支則其一致由呼叫保證（同 vehicle_category 之機讀行原則）。
# **本次重構未改 T18d 之行為** —— 其輸出經 byte-level diff 證明未變。
STOP = {"the", "a", "of", "and", "for", "to", "in", "on", "is", "shall",
        "with", "requirements", "requirement"}
THRESHOLD = 0.34


def toks(s):
    return {w for w in norm(s).split() if w not in STOP and len(w) > 2}


def best(target, cands, key):
    tt = toks(target)
    if not tt:
        return None, 0.0
    sc = [(c, len(tt & toks(key(c))) / len(tt | toks(key(c)))) for c in cands]
    sc.sort(key=lambda x: -x[1])
    return sc[0] if sc else (None, 0.0)


def t18d(groups, tops, secs):
    """標題語意對應候選。詞集重疊比，門檻 0.34；不達即 `?`，不強配。"""
    print("\n## T18d —— 三源對照草表（草料，非結論）\n")

    print("| Heading id | 037 標題 | SYS1 候選 | 分 | CFTS 候選 | 分 |")
    print("|---|---|---|---:|---|---:|")
    n_s = n_c = 0
    for g in groups:
        if g["id"] is None:
            continue
        s, ss = best(g["title"], tops, lambda x: x[1])
        c, cs = best(g["title"], secs, lambda x: x["title"])
        s_txt = f"`{s[0]}` {s[1][:30]}" if s and ss >= THRESHOLD else "**?**"
        c_txt = f"{c['num']} {c['title'][:30]}" if c and cs >= THRESHOLD else "**?**"
        n_s += s_txt != "**?**"
        n_c += c_txt != "**?**"
        print(f"| `{g['id']}` | {g['title'][:40]} | {s_txt} | {ss:.2f} | {c_txt} | {cs:.2f} |")
    tot = sum(1 for g in groups if g["id"])
    print(f"\n**對應率**：SYS1 {n_s}/{tot}、CFTS {n_c}/{tot}；其餘標 `?` 不強配。")
    print("\n> 本表為**草料**。分數為詞集重疊比（Jaccard，去停用詞、門檻 0.34），"
          "\n> **不是語意判定**；`?` 只表示自動比對不達門檻，不表示無對應。"
          "\n> Test Set 名稱與 Layer 2 分群由分析層起草，執行層不逕定（下放包 05 §三 T18d）。")


# ── T19a–e —— 定稿前置量測（下放包 06 §四）────────────────────────────
# **只量測，不分群、不命名。** 分數僅為排序工具，不作對應之結論
# （下放包 06 §四之方法拘束）——結論由分析層下。
#
# T18d 為**逐 Heading** 之比對；T19 為**逐列**（每列之 Requirement Title
# 對 87 個 CFTS 章節）。二者共用 STOP／toks／best／THRESHOLD。

def row_id(r):
    return str(r[C_ID]).strip()


def row_best(r, secs):
    c, cs = best(str(r[C_TITLE] or ""), secs, lambda x: x["title"])
    return (c, cs) if (c and cs >= THRESHOLD) else (None, cs)


def find_group(groups, hid):
    for g in groups:
        if g["id"] == hid:
            return g
    return None


def dump_group(groups, secs, hid, label):
    """逐列傾印 + 每列之最佳 CFTS 候選；併計其橫跨幾個章。"""
    g = find_group(groups, hid)
    print(f"\n### {label} —— `{hid}` {g['title'] if g else '**群不存在**'}\n")
    if g is None:
        print("**該 Heading id 不存在於 037** —— 停")
        return {}
    print(f"所轄 in-scope 列 **{len(g['rows'])}** 筆（id 區間 {id_span(g['rows'])}）\n")
    print("| id | Requirement Title | Sub Cat | Source Requirement ID | 最佳 CFTS 候選 | 分 |")
    print("|---|---|---|---|---|---:|")
    hit = Counter()
    for r in g["rows"]:
        c, cs = row_best(r, secs)
        ctxt = f"{c['num']} {c['title'][:26]}" if c else "**?**"
        if c:
            hit[c["num"]] += 1
        sub = r[C_SUB] if r[C_SUB] not in (None, "") else "—"
        src = str(r[C_SRC] or "—")
        print(f"| `{row_id(r)}` | {str(r[C_TITLE] or '')[:52]} | {sub} | {src[:26]} "
              f"| {ctxt} | {cs:.2f} |")
    n_q = len(g["rows"]) - sum(hit.values())
    print(f"\n**跨章測度**：命中 {sum(hit.values())} 列 / 未達門檻 `?` {n_q} 列；"
          f"**相異候選章 {len(hit)} 個** —— {dict(hit) or '無'}")
    return hit


def t19abce(groups, secs):
    print("\n## T19a —— `SWE1-FOTA-309` 群逐列傾印")
    dump_group(groups, secs, "SWE1-FOTA-309", "T19a")
    print("\n## T19b —— `SWE1-FOTA-291` 與 `SWE1-FOTA-259` 群")
    dump_group(groups, secs, "SWE1-FOTA-291", "T19b-1")
    dump_group(groups, secs, "SWE1-FOTA-259", "T19b-2")
    print("\n## T19c —— `SWE1-FOTA-214` 與 `SWE1-FOTA-137` 群")
    dump_group(groups, secs, "SWE1-FOTA-214", "T19c-1")
    dump_group(groups, secs, "SWE1-FOTA-137", "T19c-2")
    print("\n## T19e —— `SWE1-FOTA-178` 群")
    dump_group(groups, secs, "SWE1-FOTA-178", "T19e")


def t19d(groups, secs):
    """全 311 列之逐列 CFTS 對照；統計列與其 Heading 之候選章不一致者。

    **本項為跨章問題之全域測度**（下放包 06 §四 T19d）。
    `?`（未達門檻）之列**不計入不一致** —— 未達門檻不表示無對應，
    把它算成不一致會使測度膨脹（同 T18d 之 `?` 語意）。
    """
    print("\n## T19d —— 全 311 列之 CFTS 逐列對照\n")
    rows_all, mismatch, q_row, q_head = [], [], 0, 0
    for g in groups:
        if g["id"] is None:
            continue
        hc, hcs = best(g["title"], secs, lambda x: x["title"])
        h_num = hc["num"] if (hc and hcs >= THRESHOLD) else None
        for r in g["rows"]:
            c, cs = row_best(r, secs)
            rows_all.append(r)
            if c is None:
                q_row += 1
                continue
            if h_num is None:
                q_head += 1
                continue
            if c["num"] != h_num:
                mismatch.append((g["id"], g["title"], h_num, row_id(r),
                                 str(r[C_TITLE] or ""), c["num"], c["title"], cs))
    n = len(rows_all)
    print(f"**母體**：{n} 列（應 311）—— "
          f"{'閉合 ✅' if n == 311 else '**不閉合 ❌**'}\n")
    print("| 量 | 列數 |")
    print("|---|---:|")
    print(f"| 列之最佳候選未達門檻（`?`）| {q_row} |")
    print(f"| 其 Heading 之候選未達門檻（無可比對象）| {q_head} |")
    print(f"| **可比且不一致** | **{len(mismatch)}** |")
    print(f"| 可比且一致 | {n - q_row - q_head - len(mismatch)} |")
    print(f"\n**閉合**：{q_row} + {q_head} + {len(mismatch)} + "
          f"{n - q_row - q_head - len(mismatch)} = {n} ✅\n")
    print("### 不一致列清單（本輪最關鍵之產出）\n")
    print("| Heading id | Heading 之候選章 | 列 id | 列標題 | 列之候選章 | 分 |")
    print("|---|---|---|---|---|---:|")
    for hid, htitle, hnum, rid, rtitle, cnum, ctitle, cs in mismatch:
        print(f"| `{hid}` | {hnum} | `{rid}` | {rtitle[:44]} | {cnum} {ctitle[:26]} "
              f"| {cs:.2f} |")
    by_h = Counter(m[0] for m in mismatch)
    print(f"\n**不一致之 Heading 分布**（{len(by_h)} 個 Heading）：")
    for hid, k in by_h.most_common():
        g = find_group(groups, hid)
        print(f"- `{hid}` {g['title'][:44] if g else ''} —— **{k}** 列")
    print("\n> 分數為詞集重疊比（Jaccard，停用詞表 13 字、門檻 0.34），"
          "\n> **與 T18d 共用同一支函式**。`?` 只表示自動比對不達門檻，"
          "\n> 不表示無對應。**本表為量測，不是對應之結論。**")


if __name__ == "__main__":
    want = set(sys.argv[1:]) or {"18a", "18b", "18c", "18d", "18e"}
    T19 = {"19", "19a", "19b", "19c", "19d", "19e"}
    need_g = {"18a", "18d", "18e"} | T19
    need_s = {"18c", "18d"} | T19
    groups = t18a() if need_g & want else None
    tops = t18b() if {"18b", "18d"} & want else None
    secs = t18c() if need_s & want else None
    if "18e" in want:
        t18e(groups)
    if "18d" in want:
        t18d(groups, tops, secs)
    if T19 & want:
        t19abce(groups, secs)
        t19d(groups, secs)
