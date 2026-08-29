#!/usr/bin/env python3
"""T36d —— pilot 批 v3（下放包 23 §五）之工作簿產出，供 lint 用。

沿革：
- **v1**（`pilot01`，下放包 19 §四）：5 個 TC 皆含「Read <內部服務> 收到的 X」類
  步驟，台架上不可執行（上繳包 18 §7.1）；PENDING 說明為中文（違 R-14）。
- **v2**（`pilot02`，下放包 20 §四）：5 個 TC 全面改寫（R-SU25 可觀測面）+
  PENDING 英文化。lint K=0／T=0／U=3。
- **v3**（`pilot03`，下放包 23 §四）：**TC 內容逐字沿 v2 不動**，
  僅補齊寫回欄集 —— `S`（functional_safety）填 **`NA`**（他 feature 5/6 之實務）、
  **`T`–`Z` 七個車型旗標留空**（他 feature 6/6 之實務）。
- **v4**（`pilot04`，T43b／下放包 30 §六）：**R-SU36 之時間解析度** ——
  三處錄影步驟改 `as continuous video capture`（明文排除定時截圖）；
  **TC-4 之 `Record every SW Update screen …` 依 R-SU36(c) 改寫**
  （`every` 是一個宣稱不是一個動作），其 ER 二行對應改。
  **TC 之驗證單元不變**，改動限於觀測手段之具體化。

前一版之產出一律保留，**不覆寫**。

**本檔為 lint 之受檢物，不是交付本**：輸出落 `sandbox/pilot01/`（R-G25），
`inputs/` 之母本一字不動。TC 內容**逐字取自下放包 19 §四**，
執行層不改寫（T32b：「不得為使其通過而改寫」）。

寫入路徑沿 `scripts/write_back_036.py::_set_row`（R-SU2 之 XML 外科式，
上繳包 16 §5.2 已以實寫探針驗其保全 `<row>` 屬性、儲存格數與 `s=` 樣式）。

TC ID 依 **R-SU24**：`{project}-SU-{NNN}`，`{project}` 取自 036 母本
**D2 實測值**（T32a），不推定。

未寫之欄：`T`–`Z`（車型適用旗標，**留空**，下放包 23 §四）、`AH`（remarks）。
"""
import re
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from write_back_036 import _set_row, MASTER, SHEET_NAME, HEADER_ROW, FEAT  # noqa: E402

TAG = "pilot04"                    # v1–v3 為 pilot01–03，**不覆寫**
OUT = FEAT / "sandbox" / TAG / MASTER
TEST_GROUP, TEST_SET = "SW Update", "Silent Update"
AUTHOR = "PeiPYHsu"
FN = "功能測試 (Functional based ; no specific technique)"
NEG = "負向測試 (Negative / Invalid)"

TCS = [
 dict(req="SWE1-FOTA-175", spec="CFTS057-4907475", dm=FN, prio="P1",
  item=["When the update type is identified as Silent Update, the WiFi Update Service shall automatically execute the update in background mode.",
        "(Silent update runs in background with no HMI interaction)"],
  pre=["1. The head unit is connected to a Wi-Fi network with internet access",
       "2. An update package with update type Silent Update is staged on the OTA Server for this head unit"],
  proc=["1. Read the software version shown on the head unit and record it as Version_initial",
        "2. Trigger an update availability check to the OTA Server",
        "3. Record the head unit screen content as continuous video capture until the update finishes",
        "4. Read the software version shown on the head unit and record it as Version_after",
        "5. Check that Version_after differs from Version_initial and that no SW Update prompt or progress notification appears in the recorded screen content"],
  er=["1. Version_initial is recorded",
      "2. The update availability check completes and an update is reported as available",
      "3. The head unit screen content until the update finishes is recorded as continuous video capture",
      "4. Version_after is recorded",
      "5. Version_after differs from Version_initial; the recorded screen content contains no SW Update prompt and no progress notification"]),
 dict(req="SWE1-FOTA-176", spec="CFTS057-4907476", dm=FN, prio="P1",
  item=["During a Silent Update session, the WiFi Update Service shall not trigger the SW Update HMI for update progress notifications.",
        "(No progress notification during a silent session)"],
  pre=["1. The head unit is connected to a Wi-Fi network with internet access",
       "2. An update package with update type Silent Update is staged on the OTA Server for this head unit"],
  proc=["1. Read the software version shown on the head unit and record it as Version_initial",
        "2. Trigger an update availability check to the OTA Server",
        "3. Record the head unit screen content as continuous video capture until the software version changes",
        "4. Check that no update progress notification appears anywhere in the recorded screen content"],
  er=["1. Version_initial is recorded",
      "2. The update availability check completes and an update is reported as available",
      "3. The head unit screen content until the software version changes is recorded as continuous video capture",
      "4. The recorded screen content contains no update progress notification at any point of the session"]),
 dict(req="SWE1-FOTA-176", spec="CFTS057-4907477", dm=FN, prio="P1",
  item=["During a Silent Update session, the WiFi Update Service shall allow user notification only when required to satisfy safety-related requirements.",
        "(Safety-required notification is permitted during a silent session)"],
  pre=["1. The head unit is connected to a Wi-Fi network with internet access",
       "2. An update package with update type Silent Update is staged on the OTA Server for this head unit",
       "3. PENDING: DR-SU1 list of safety-related notification conditions applicable during a silent session"],
  proc=["1. Trigger an update availability check to the OTA Server",
        "2. Record the head unit screen content continuously from the start of the session",
        "3. PENDING: DR-SU1 step to bring one safety-related condition into effect",
        "4. Check that the safety-related notification is displayed on the head unit while the session continues"],
  er=["1. The update availability check completes and an update is reported as available",
      "2. The head unit screen content from the start of the session is recorded",
      "3. PENDING: DR-SU1 observable state showing the safety-related condition is in effect",
      "4. The safety-related notification is displayed on the head unit and the session continues"]),
 dict(req="SWE1-FOTA-177", spec="CFTS057-4907478", dm=NEG, prio="P2",
  item=["If the SW Update HMI is available, the assigned update service shall not present the user with options to opt out of or defer the update.",
        "(No opt-out or defer option offered when HMI is available)"],
  pre=["1. The head unit is connected to a Wi-Fi network with internet access",
       "2. An update package with update type Silent Update is staged on the OTA Server for this head unit",
       "3. The SW Update HMI is available on the head unit"],
  proc=["1. Trigger an update availability check to the OTA Server",
        "2. Record the head unit screen content as continuous video capture until the update finishes",
        "3. Check that no opt-out control and no defer control appear in the recorded screen content"],
  er=["1. The update availability check completes and an update is reported as available",
      "2. The head unit screen content until the update finishes is recorded as continuous video capture",
      "3. The recorded screen content contains no opt-out control and no defer control"]),
 dict(req="SWE1-FOTA-183", spec="CFTS057-4907485", dm=FN, prio="P2",
  item=["When the update completes, the OTA client will display a success notification and what's new details.",
        "(Completion notification with What's New shown after a silent update)"],
  pre=["1. The head unit is connected to a Wi-Fi network with internet access",
       "2. An update package with update type Silent Update is staged on the OTA Server for this head unit"],
  proc=["1. Read the software version shown on the head unit and record it as Version_initial",
        "2. Trigger an update availability check to the OTA Server",
        "3. Read the software version shown on the head unit until it differs from Version_initial",
        "4. Check that the head unit displays the update success notification together with the What's New details"],
  er=["1. Version_initial is recorded",
      "2. The update availability check completes and an update is reported as available",
      "3. The software version shown on the head unit differs from Version_initial",
      "4. The head unit displays the update success notification and the What's New details of the deployed package"]),
]


def project_name():
    """T32a —— 036 母本之專案名稱，實測 D2（標籤 C2）。**不推定**。"""
    import openpyxl, warnings
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(FEAT / "inputs" / MASTER, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    label = ws["C2"].value
    val = ws["D2"].value
    return str(val).strip(), str(label).strip()


def main():
    proj, label = project_name()
    src = FEAT / "inputs" / MASTER
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(src) as z:
        wbxml = z.read("xl/workbook.xml").decode("utf-8")
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        rid = re.search(r'<sheet[^>]*name="%s"[^>]*r:id="([^"]+)"'
                        % re.escape(SHEET_NAME), wbxml).group(1)
        tgt = re.search(r'Id="%s"[^>]*Target="([^"]+)"' % re.escape(rid), rels).group(1)
        sheet_path = "xl/" + tgt.lstrip("/")
        sx = z.read(sheet_path).decode("utf-8")

    print("## T32a —— 036 母本之專案名稱（實測）\n")
    print(f"| 儲存格 | 值 |")
    print(f"|---|---|")
    print(f"| `C2`（標籤） | `{label}` |")
    print(f"| **`D2`（值）** | **`{proj}`** |")
    print(f"\n工作表：`{SHEET_NAME}`｜表頭列 {HEADER_ROW}"
          f"｜TC ID 欄為 **F**（`F9` = `Test Case ID 測試用例ID`）\n")

    rows = []
    for n, t in enumerate(TCS, 1):
        tcid = f"{proj}-SU-{n:03d}"
        vals = {"D": t["req"], "F": tcid, "G": TEST_GROUP, "H": TEST_SET,
                "I": "\n".join(t["item"]), "J": "\n".join(t["pre"]),
                "K": "NA", "L": "\n".join(t["proc"]), "M": "\n".join(t["er"]),
                "N": t["spec"], "O": "NEW", "P": t["prio"], "R": t["dm"],
                "S": "NA",              # 下放包 23 §四：從他 feature 5/6 之實務
                "AA": AUTHOR}
        # `T`–`Z`（車型適用旗標）**留空**：他 feature 6/6 一律留空（下放包 23 §四）
        sx = _set_row(sx, HEADER_ROW + n, vals)
        rows.append((HEADER_ROW + n, tcid, t["req"], t["spec"]))

    with zipfile.ZipFile(src) as zin, \
            zipfile.ZipFile(OUT, "w", zipfile.ZIP_DEFLATED) as zo:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_path:
                data = sx.encode("utf-8")
            zo.writestr(item, data)

    print(f"## pilot 批之產出（受檢物，非交付本）\n")
    print(f"- 輸出：`sandbox/{TAG}/…_ext.xlsx`（R-G25）｜母本未動\n")
    print("| 列 | TC ID | 037 列 | spec_reference |")
    print("|---|---|---|---|")
    for r, tid, req, sp in rows:
        print(f"| {r} | `{tid}` | `{req}` | `{sp}` |")
    print(f"\n- `S`（functional_safety）填 **`NA`**；`T`–`Z`（車型旗標）**留空**；`AH`（remarks）未寫。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
