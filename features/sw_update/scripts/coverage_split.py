#!/usr/bin/env python3
"""T35a–d —— 105 列 × Verification Method 交叉表、說明欄抽樣、036 寫回欄集查核。

Usage: python3 scripts/coverage_split.py 35a 35b 35c 35d
"""

import random
import re
import sys
import warnings
from collections import Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from corpus_v2 import _rows_desc                                    # noqa: E402
from anchor_table import A03, C_ID, C_TITLE, C_CAT, IN_SCOPE        # noqa: E402
from framework_survey import a03_rows                               # noqa: E402
from observability import classify, RE_EXTERNAL                     # noqa: E402
from verif_columns import _load, _txt, C_VC, C_VM, NUM              # noqa: E402
from write_back_036 import MASTER, SHEET_NAME, FEAT                 # noqa: E402

warnings.filterwarnings("ignore")
SEED_B = 35
ROOT = FEAT.parent.parent


def groups():
    """四組（下放包 22 §五 T35a）。"""
    by, d, owner = _load()
    has_ext_vc = lambda i: any(r.search(_txt(by[i][C_VC])) for r in RE_EXTERNAL.values())
    internal = [i for i in sorted(d, key=NUM) if classify(d[i])[0]]
    g105 = [i for i in internal if not has_ext_vc(i)]
    g21 = [i for i in internal if has_ext_vc(i)]
    g185 = [i for i in sorted(d, key=NUM) if i not in set(internal)]
    g32 = [i for i in sorted(d, key=NUM) if "HMI Validation" in _txt(by[i][C_VM])]
    return by, d, owner, g105, g21, g185, g32


def _vm_kind(v):
    """`Verification Method` 之三分：僅 Integration Test／含 System Test／其他。"""
    v = _txt(v)
    if v == "Integration Test":
        return "**僅 `Integration Test`**"
    if v.startswith("Unit Test") or v.startswith("Unit Testing"):
        return "含 `Unit/Integration/System`"
    return "其他"


def t35a():
    by, d, owner, g105, g21, g185, g32 = groups()
    G = [("**105 列**（內部列且 VC 亦無外部面）", g105),
         ("21 列（內部列但 VC 有外部面）", g21),
         ("185 列（非內部列）", g185),
         ("32 列（`HMI Validation Testing`）", g32)]
    KINDS = ["**僅 `Integration Test`**", "含 `Unit/Integration/System`", "其他"]

    print("## T35a —— 四組 × `Verification Method` 之交叉表（本輪核心）\n")
    print("> **只出表，不作歸屬之判斷**（下放包 22 §五 T35a）。\n")
    print("| 組 | 列數 | " + " | ".join(KINDS) + " |")
    print("|---|---:|" + "---:|" * len(KINDS))
    rates = {}
    for lbl, g in G:
        c = Counter(_vm_kind(by[i][C_VM]) for i in g)
        rates[lbl] = c[KINDS[0]] / len(g) * 100 if g else 0
        print(f"| {lbl} | {len(g)} | "
              + " | ".join(f"{c[k]}（{c[k]/len(g)*100:.0f}%）" if g else "—" for k in KINDS)
              + " |")

    print(f"\n### 「僅 `Integration Test`」之佔比並列\n")
    print("| 組 | 佔比 |")
    print("|---|---:|")
    for lbl, _ in G:
        print(f"| {lbl} | **{rates[lbl]:.0f}%** |")
    base = rates["185 列（非內部列）"]
    top = rates["**105 列**（內部列且 VC 亦無外部面）"]
    print(f"\n**105 列之佔比 {top:.0f}% vs 非內部列 {base:.0f}%** —— "
          f"{'105 列**低於**對照組' if top < base else '105 列**高於**對照組'}"
          f"（差 {top-base:+.0f} 個百分點）。")
    print(f"\n> **下放包 22 §四之丙路，其前提為「105 列多數標 `Integration Test`」。**"
          f" 實測 {top:.0f}%，**{'不成立' if top < 50 else '成立'}** —— "
          f"該 105 列並非集中於整合測試層級。**執行層只陳報此數，不裁其歸屬。**\n")

    print("### `Verification Method` 之完整值分佈（四組）\n")
    for lbl, g in G:
        c = Counter(_txt(by[i][C_VM]) for i in g)
        print(f"\n**{lbl}**（{len(g)} 列）\n")
        print("| 值 | 列數 |")
        print("|---|---:|")
        for k, v in c.most_common(6):
            print(f"| `{k[:64]}` | {v} |")
        if len(c) > 6:
            print(f"| （其餘 {len(c)-6} 種） | {sum(v for _, v in c.most_common()[6:])} |")
    return g105, g21, g185, g32


def t35b():
    rows = [r for r in a03_rows() if r[C_CAT] in IN_SCOPE]
    COLS = {8: "Description/Action for Feasibility",
            10: "Description/Action for Impact",
            12: "Description/Action for Risk Factor",
            14: "Description/Action for Reusable"}
    print("\n\n## T35b —— 四個說明欄之抽樣傾印（R-SU28(b)）\n")
    print(f"取樣：自 311 母體以 `random.Random({SEED_B}).sample(rows, 8)` 各抽 8 列。\n")
    for c, name in COLS.items():
        vals = [_txt(r[c]) for r in rows]
        uniq = Counter(vals)
        print(f"\n---\n\n### 欄 {c} — `{name}`\n")
        print(f"- 非空 **{sum(1 for v in vals if v)}／311**｜unique **{len(uniq)}**\n")
        print("**unique 值前 5 名**：\n")
        print("| 次數 | 值（前 90 字元） |")
        print("|---:|---|")
        for v, n in uniq.most_common(5):
            print(f"| {n} | {v[:90] or '(空)'} |")
        rng = random.Random(SEED_B)
        pick = rng.sample(rows, 8)
        print(f"\n**抽樣 8 列之全文**：\n")
        for r in pick:
            print(f"- `{str(r[C_ID]).strip()}` — {_txt(r[C_TITLE])[:44]}")
            print(f"  > {_txt(r[c]) or '**(空)**'}\n")


def t35c():
    print("\n\n## T35c —— 036 母本之寫回欄集查核\n")
    wb = openpyxl.load_workbook(FEAT / "inputs" / MASTER, data_only=True)
    ws = wb[SHEET_NAME]
    hdr = {c.column_letter: _txt(c.value).replace("\n", " ／ ")
           for c in next(ws.iter_rows(min_row=9, max_row=9))}
    dvs = []
    for dv in ws.data_validations.dataValidation:
        dvs.append((str(dv.sqref), str(dv.formula1)))
    print("### (i)(ii) 標頭原文與 DV\n")
    print("| 欄 | 標頭原文 | 所屬 DV |")
    print("|---|---|---|")
    for L in ["S", "T", "U", "V", "W", "X", "Y", "Z"]:
        d = next((f"`{s}` = {f}" for s, f in dvs
                  if re.match(rf"^{L}\d+:|:{L}\d+$|^{L}\d+$", s)
                  or (":" in s and s.split(":")[0][0] <= L <= s.split(":")[1][0])), "**無**")
        print(f"| `{L}` | {hdr.get(L,'')[:46]} | {d} |")

    print("\n### 他 feature 之既有交付本於該八欄之填值（如實回報，不代裁）\n")
    books = sorted(ROOT.glob("features/*/delivered/*.xlsx")) + \
        sorted(ROOT.glob("features/*/output/*.xlsx"))
    if not books:
        print("**查無任何他 feature 之已交付簿**（`features/*/delivered/*.xlsx`"
              "／`features/*/output/*.xlsx` 皆空）。")
        print("\n> `output/` 於 `.gitignore` 內，故 repo 內不必然存在；"
              "**此為「查無」，非「已查得為空」** —— 二者不同（PLAYBOOK §7(7)）。")
        return
    print("| feature | 簿 | 資料列 | " + " | ".join(f"`{L}`" for L in "STUVWXYZ") + " |")
    print("|---|---|---:|" + "---|" * 8)
    for b in books[:8]:
        try:
            w = openpyxl.load_workbook(b, read_only=True, data_only=True)
            s = next((w[n] for n in w.sheetnames if n.startswith("Test Case Specification")), None)
            if s is None:
                continue
            rs = [r for r in s.iter_rows(min_row=10, values_only=True) if r[3] not in (None, "")]
            cells = []
            for i in range(18, 26):            # S..Z = 0-indexed 18..25
                vs = Counter(_txt(r[i]) for r in rs if len(r) > i)
                cells.append("／".join(f"`{k or '空'}`×{v}" for k, v in vs.most_common(3)))
            print(f"| {b.parts[-3]} | `{b.name[:18]}…` | {len(rs)} | " + " | ".join(cells) + " |")
        except Exception as e:                  # noqa: BLE001
            print(f"| {b.parts[-3]} | `{b.name[:18]}…` | — | 讀取失敗：{e} |")


def t35d():
    by, d, owner = _load()
    tc = [i for i in sorted(d, key=NUM) if owner[i] == "Telematics Client"]
    print("\n\n## T35d —— `Telematics Client` 5 列之 `Verification Method`\n")
    print("| 037 列 | 標題 | `Verification Method` |")
    print("|---|---|---|")
    for i in tc:
        print(f"| `{i}` | {_txt(by[i][C_TITLE])[:40]} | `{_txt(by[i][C_VM])}` |")
    print("\n### `366`／`367` 之 `Verification Criteria` 全文（上繳包 20 §2.1 未展開）\n")
    for k in ("366", "367"):
        i = f"SWE1-FOTA-{k}"
        print(f"\n#### `{i}` — {_txt(by[i][C_TITLE])}\n")
        for ln in _txt(by[i][C_VC]).split("\n"):
            if ln.strip():
                print(f"> {ln.strip()}")
                print(">")
    return tc


if __name__ == "__main__":
    want = set(sys.argv[1:]) or {"35a", "35b", "35c", "35d"}
    if "35a" in want:
        t35a()
    if "35b" in want:
        t35b()
    if "35c" in want:
        t35c()
    if "35d" in want:
        t35d()
