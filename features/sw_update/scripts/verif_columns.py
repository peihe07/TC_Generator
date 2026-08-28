#!/usr/bin/env python3
"""T34a–d —— Verification 二欄之傾印、素材欄位全覽（下放包 21 §四）。

Usage: python3 scripts/verif_columns.py 34a 34b 34c 34d
"""

import re
import sys
import warnings
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from corpus_v2 import _rows_desc                                          # noqa: E402
from anchor_table import A03, CFTS, C_ID, C_TITLE, C_CAT, IN_SCOPE        # noqa: E402
from framework_survey import a03_rows, group_by_heading, SYS1             # noqa: E402
from layer2_close import SETS, H                                          # noqa: E402
from observability import classify                                        # noqa: E402
from write_back_036 import MASTER, SHEET_NAME, FEAT                       # noqa: E402

warnings.filterwarnings("ignore")
C_VC, C_VM = 16, 17               # `Verification Criteria` / `Verification Method`
NUM = lambda s: int(str(s).rsplit("-", 1)[1])
# IN §5.1 之禁用步驟動詞（`lint036.RE_A` 逐字）
BANNED = ("Observe", "Verify", "See if", "Watch", "Monitor", "Inspect")


def _load():
    rows = [r for r in a03_rows() if r[C_CAT] in IN_SCOPE]
    by = {str(r[C_ID]).strip(): r for r in rows}
    _, d = _rows_desc()
    gmap = {g["id"]: g for g in group_by_heading(a03_rows())[1:]}
    owner = {}
    for name, items in SETS:
        for it in items:
            if isinstance(it, str):
                for r in gmap[H(it)]["rows"]:
                    owner[str(r[C_ID]).strip()] = name
            else:
                for i in d:
                    if it[1] <= NUM(i) <= it[2]:
                        owner[i] = name
    return by, d, owner


def _txt(v):
    return re.sub(r"[ \t]+", " ", str(v or "")).strip()


def t34a():
    by, d, owner = _load()
    internal = [i for i in sorted(d, key=NUM) if classify(d[i])[0]]
    tc_first = [i for i in internal if owner[i] == "Telematics Client"]
    rest = [i for i in internal if owner[i] != "Telematics Client"]
    order = tc_first + rest

    empty = [i for i in internal if not _txt(by[i][C_VC])]
    print("## T34a —— 126 個內部列之 `Verification Criteria` 傾印\n")
    print(f"- 內部列 **{len(internal)}**（上繳包 19 §T33b）")
    print(f"- 其中 `Verification Criteria` **為空者 {len(empty)} 列**"
          + (f"：{'、'.join('`'+x.rsplit('-',1)[1]+'`' for x in empty)}"
             if empty else "：**無**"))
    print(f"- `Telematics Client` 之 {len(tc_first)} 列置於最前（下放包 21 §二 #3 待判）\n")
    print("> **執行層不裁定其觀測面。** 本節為分析層逐列裁定之材料。\n")
    print("> ⚠ 本欄之語形含 `Monitor`／`Observe` —— 二者為 **IN §5.1 之禁用步驟動詞**"
          "（`lint036.RE_A`）。取用其所述之觀測面時須改寫動詞，見上繳包 20 §自評。\n")

    # ── VC 本身是否給出外部面（下放包 21 §五.6 之量測）
    from observability import RE_EXTERNAL
    ext = [i for i in internal
           if any(r.search(_txt(by[i][C_VC])) for r in RE_EXTERNAL.values())]
    mixed = [i for i in d if i not in set(internal)]
    ext2 = [i for i in mixed
            if any(r.search(_txt(by[i][C_VC])) for r in RE_EXTERNAL.values())]
    verbs = Counter()
    for i in d:
        for ln in _txt(by[i][C_VC]).split("\n"):
            m = re.match(r"([A-Za-z]+)", ln.strip())
            if m:
                verbs[m.group(1)] += 1
    tot = sum(verbs.values())
    ban = sum(v for k, v in verbs.items() if k in BANNED)

    print("### 0. 本欄能否供給觀測面 —— 先量再傾印\n")
    print("| 量 | 值 |")
    print("|---|---:|")
    print(f"| VC 之總行數（310 列） | {tot} |")
    print(f"| 行首為 IN §5.1 禁用動詞者 | **{ban}（{ban/tot*100:.0f}%）** |")
    print(f"| **126 內部列中，其 VC 含外部面語形者** | **{len(ext)}／126（{len(ext)/126*100:.0f}%）** |")
    print(f"| **126 內部列中，其 VC 亦無任何外部面者** | **{126-len(ext)}／126（{(126-len(ext))/126*100:.0f}%）** |")
    print(f"| 對照：185 非內部列中，其 VC 含外部面者 | {len(ext2)}／185（{len(ext2)/185*100:.0f}%） |")
    print("\n行首動詞（前 10）：")
    print("\n| 動詞 | 行數 | IN §5.1 |")
    print("|---|---:|:--:|")
    for k, v in verbs.most_common(10):
        print(f"| `{k}` | {v} | {'**禁用**' if k in BANNED else '—'} |")
    print(f"\n> **本表即下放包 21 §五.6 之答案**：本欄之觀測面與需求本文**同源** ——"
          f" 需求提及外部面者其 VC 亦提及（93%），需求未提者其 VC 多半亦未提"
          f"（{(126-len(ext))/126*100:.0f}%）。詳見上繳包 20 §自評。\n")

    print("### 甲 —— `Telematics Client`（5 列，全組皆內部列）\n")
    _dump(tc_first, by, owner)
    print(f"\n### 乙 —— 其餘 {len(rest)} 列（依 037 列序）\n")
    _dump(rest, by, owner)
    return internal, empty


def _dump(ids, by, owner):
    for n, i in enumerate(ids, 1):
        r = by[i]
        vc, vm = _txt(r[C_VC]), _txt(r[C_VM])
        print(f"\n---\n\n#### {n}. `{i}` — {_txt(r[C_TITLE])}\n")
        print(f"- Test Set：**`{owner[i]}`**｜`Verification Method`：`{vm or '(空)'}`")
        if vc:
            print(f"\n**`Verification Criteria` 全文**：\n")
            for ln in vc.split("\n"):
                if ln.strip():
                    print(f"> {ln.strip()}")
                    print(">")
        else:
            print("\n**`Verification Criteria`**：**(空)**")
        print()


def t34b():
    by, d, owner = _load()
    hits = [i for i in sorted(d, key=NUM) if "HMI Validation" in _txt(by[i][C_VM])]
    internal = {i for i in d if classify(d[i])[0]}
    print("\n\n## T34b —— `HMI Validation Testing` 之 32 列（正向樣本）\n")
    print(f"- 命中 **{len(hits)}** 列；其中內部列 **{len(set(hits) & internal)}** 列"
          f" —— **交集為 {'0，完美分離' if not (set(hits) & internal) else '非 0'}**")
    print("- 用途：作為「有 HMI 可觀測面」之正向樣本，供分析層校準 R-SU25(c) 之取用方式\n")
    print("| # | 037 列 | Test Set | `Verification Method` 之串接 |")
    print("|---:|---|---|---|")
    for n, i in enumerate(hits, 1):
        print(f"| {n} | `{i}` | `{owner[i]}` | {_txt(by[i][C_VM])} |")
    print(f"\n### 逐列之 `Verification Criteria`\n")
    _dump(hits, by, owner)
    return hits


# ── T34c／T34d —— 欄位全覽（R-SU26）─────────────────────────────────
# 用途標記：已用（何處）／不用（理由）／未定
USE_037 = {
    0: ("已用", "`req_id`／全案之列 id（R-SU3）"),
    1: ("不用", "R-SU5 v2：三形態並存，該欄不取為 spec_reference"),
    2: ("已用", "Layer 2 切分之材料（T28a／T28b）；TC 不 verbatim 抄"),
    3: ("已用", "路徑 A 之查詢側（R-SU12）、TC `test_item` 上半之 verbatim 來源"),
    4: ("未定", "`Release Version` —— 未查其值型態與用途"),
    5: ("已用", "in-scope 判定（FR／NFR／Heading，R-SU3）"),
    6: ("已用", "HMI／Service 之分面（T28c、原則 3）"),
    7: ("未定", "`Feasibility` —— 未讀"),
    8: ("未定", "`Description/Action for Feasibility` —— 未讀"),
    9: ("未定", "`Impact` —— 未讀"),
    10: ("未定", "`Description/Action for Impact` —— 未讀"),
    11: ("未定", "`Risk Factor` —— 未讀"),
    12: ("未定", "`Description/Action for Risk` —— 未讀"),
    13: ("未定", "`Reusable` —— 未讀"),
    14: ("未定", "`Description/Action for Reusable` —— 未讀"),
    15: ("已用", "R-SU22：僅作參考訊號，不作 P 值之唯一依據"),
    16: ("已用", "**R-SU27(a)：R-SU25(c) 外部可觀測後果之候選來源**（本輪起）"),
    17: ("已用", "**R-SU27(b)：測試層級之參考訊號**（本輪起）"),
}


def t34c():
    wb = openpyxl.load_workbook(A03, read_only=True, data_only=True)
    ws = wb["AnalysisReport_FULL"]
    allr = [r for r in ws.iter_rows(min_row=8, values_only=True) if r[0] not in (None, "")]
    hdr = [_txt(c) for c in next(ws.iter_rows(min_row=7, max_row=7, values_only=True))]
    inscope = {str(r[C_ID]).strip() for r in a03_rows() if r[C_CAT] in IN_SCOPE}
    print("\n\n## T34c —— 037 欄位全覽（R-SU26(a)(b)(c)）\n")
    print(f"來源：`AnalysisReport_FULL`，表頭列 7、資料列 8 起，"
          f"全 **{len(allr)}** 資料列；驗證母體 **{len(inscope)}** 列。\n")
    print("| 欄 | 標頭原文 | 非空（383） | 非空（311 母體） | 值型態摘要 | 用途 |")
    print("|---:|---|---:|---:|---|---|")
    undef = []
    for c in range(18):
        n_all = sum(1 for r in allr if r[c] not in (None, ""))
        n_in = sum(1 for r in allr if str(r[0]).strip() in inscope and r[c] not in (None, ""))
        vals = [_txt(r[c]) for r in allr if r[c] not in (None, "")]
        uniq = len(set(vals))
        if uniq <= 8 and vals:
            shape = "枚舉 " + "／".join(f"`{v[:22]}`" for v in sorted(set(vals))[:5])
        else:
            ln = sorted(len(v) for v in vals) if vals else [0]
            shape = (f"自由文字，unique {uniq}，長度中位 {ln[len(ln)//2]}"
                     if vals else "**全空**")
        use, why = USE_037[c]
        if use == "未定":
            undef.append((c, hdr[c] if c < len(hdr) else "", n_all))
        print(f"| {c} | `{hdr[c] if c < len(hdr) else ''}` | {n_all} | {n_in} | "
              f"{shape[:70]} | **{use}** —— {why} |")
    print(f"\n**用途統計**：已用 **{sum(1 for v in USE_037.values() if v[0]=='已用')}**／"
          f"不用 {sum(1 for v in USE_037.values() if v[0]=='不用')}／"
          f"**未定 {len(undef)}**（共 18）")
    print(f"\n### ⚠ `未定` 清單（R-SU26(b)：不得留存跨輪，下一輪須裁）\n")
    print("| 欄 | 標頭原文 | 非空（383） |")
    print("|---:|---|---:|")
    for c, h, n in undef:
        print(f"| {c} | `{h}` | {n} |")
    return undef


def t34d():
    print("\n\n## T34d —— 其餘素材之欄位全覽（R-SU26(c)）\n")
    # SYS1 export
    wb = openpyxl.load_workbook(SYS1, read_only=True, data_only=True)
    ws = wb["Basic Report"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [_txt(c) for c in rows[0]]
    data = [r for r in rows[1:] if r[0] not in (None, "")]
    print(f"### SYS1 export（`Basic Report`，{len(data)} 資料列，{len(hdr)} 欄）\n")
    print("| 欄 | 標頭原文 | 非空 | 值型態摘要 | 用途 |")
    print("|---:|---|---:|---|---|")
    USE_SYS1 = {"Outline Number": ("已用", "R-SU11：SYS1 之接點為 HMI 87 列；T18b 分群"),
                "Description": ("已用", "T18b／T18d 之比對側")}
    for i, h in enumerate(hdr):
        n = sum(1 for r in data if r[i] not in (None, ""))
        vals = [_txt(r[i]) for r in data if r[i] not in (None, "")]
        uniq = len(set(vals))
        shape = (("枚舉 " + "／".join(f"`{v[:18]}`" for v in sorted(set(vals))[:4]))
                 if uniq <= 8 and vals else f"自由文字，unique {uniq}")
        u, why = USE_SYS1.get(h, ("未定", "未讀"))
        print(f"| {i} | `{h}` | {n} | {shape[:56]} | **{u}** —— {why} |")

    # 036 母本之 TC 分頁
    wbm = openpyxl.load_workbook(FEAT / "inputs" / MASTER, read_only=True, data_only=True)
    wsm = wbm[SHEET_NAME]
    hdr9 = list(next(wsm.iter_rows(min_row=9, max_row=9, values_only=True)))
    import yaml
    cfg = yaml.safe_load((FEAT / "feature.yaml").read_text())
    used = {v: k for k, v in cfg["workbook"]["columns"].items()}
    letter = lambda n: openpyxl.utils.get_column_letter(n + 1)
    print(f"\n### 036 母本（`{SHEET_NAME}`，表頭列 9，{len(hdr9)} 欄）\n")
    print("| 欄 | 標頭原文 | 用途 |")
    print("|---|---|---|")
    for i, h in enumerate(hdr9):
        L = letter(i)
        t = _txt(h).replace("\n", " ／ ")
        if not t:
            continue
        if L in used:
            u = f"**已用** —— `{used[L]}`（`feature.yaml` §workbook.columns）"
        elif L == "F":
            u = "**已用** —— TC ID（R-SU24；`feature.yaml` 未列，實測 `F9`）"
        elif L == "B":
            u = "**不用** —— 共用公式之宿主（`t=\"shared\"` 1401 處），賦值即毀"
        else:
            u = "**未定** —— 未讀"
        print(f"| `{L}` | {t[:52]} | {u} |")

    print("\n### 文字型素材之結構元素（CFTS_57／SYSAD／VF747／HMI PDF）\n")
    print("| 素材 | 已用之結構元素 | 未用之結構元素 |")
    print("|---|---|---|")
    print("| **CFTS_57**（docx） | heading style 1–4 之 `{7位}` 章節（87）；"
          "`[Artifact Type:Subsystem Functional Requirement]` 宣告之需求物件（487）；"
          "其後至下一宣告之全文（語料 v2） | **表格**（`<w:tbl>`）、**圖**（drawing）、"
          "註腳、`[Artifact Type:…]` 之**其他型別**（Description 137 已用其歸屬，"
          "其餘型別未列）、修訂標記 |")
    print("| **SYSAD**（docx） | 迄今**僅作 T33c 之全文語形掃描** | "
          "章節結構、介面節、架構圖、**全部表格** |")
    print("| **VF747**（docx） | 已綁 `reference.vf747`；**內容未讀** | 全部 |")
    print("| **HMI 規格 PDF** | R-SU6 v2：真 PDF、68 頁全文字層；"
          "T5' 之 popup id 抽取（52 個） | 頁面版面、圖、**全文之語意內容** |")


if __name__ == "__main__":
    want = set(sys.argv[1:]) or {"34a", "34b", "34c", "34d"}
    if "34a" in want:
        t34a()
    if "34b" in want:
        t34b()
    if "34c" in want:
        t34c()
    if "34d" in want:
        t34d()
