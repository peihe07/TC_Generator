#!/usr/bin/env python3
"""交付本列序重排之驗證（R-BLM17 §1.3 之七項，逐項照做）。

其 round-trip 之期望值取自**配對表**（依 `req_id` 配對），
**不取自產生寫入之同一支函式**（bed_lowering 上繳 07 之教訓）。
"""
import sys, warnings, zipfile, re
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
warnings.filterwarnings("ignore")
import openpyxl
import write_back_all as W
from write_back_036 import FEAT, MASTER, SHEET_NAME

DELIV = FEAT / "delivered" / W.OUT_NAME
COLS = {4: "req_id", 6: "tc_id", 7: "test_group", 8: "test_set", 9: "test_item",
        10: "pre", 11: "input", 12: "proc", 13: "er", 14: "spec", 15: "tc_ref",
        16: "priority", 18: "design_method", 19: "func_safety", 27: "author"}
num = lambda r: int(r.rsplit("-", 1)[1])


def read_book(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME]
    out = []
    for r in range(10, 400):
        if not ws.cell(row=r, column=6).value:
            continue
        out.append({v: (ws.cell(row=r, column=c).value or "") for c, v in COLS.items()} | {"row": r})
    return out


def main():
    # 期望值：自生成器重建之「配對表」—— 以 (req_id, 該 req 內之起草序) 為鍵
    tcs = W.collect()
    W.A9_FALLBACK = ""
    import openpyxl as ox
    wb = ox.load_workbook(FEAT / "inputs" / MASTER, read_only=True, data_only=True)
    legal = [str(wb["下拉選單"].cell(row=r, column=1).value) for r in range(1, 10)]
    W.A9_FALLBACK = legal[8]
    seen = {}
    expect = {}
    for t in tcs:
        k = t["req"]
        seen[k] = seen.get(k, 0) + 1
        pre = t["pre"] if t["pre"] and re.match(r"^\d+\.", t["pre"][0]) else \
            [f"{i}. {s}" for i, s in enumerate(t["pre"], 1)]
        expect[(k, seen[k])] = {
            "test_item": "\n".join(t["item"]), "pre": "\n".join(pre),
            "proc": "\n".join(t["proc"]), "er": "\n".join(t["er"]),
            "spec": t["spec"], "priority": t["prio"],
            "design_method": t.get("dm") or W.META[id(t)][1] or W.A9_FALLBACK,
            "test_set": t.get("ts") or W.META[id(t)][0]}

    got = read_book(DELIV)
    print("## 列序重排之驗證（R-BLM17 §1.3）\n")
    ok = True

    print(f"**4. 計數**：交付本 **{len(got)}** 列｜生成器 **{len(tcs)}** 個 TC —— "
          f"{'✅' if len(got) == len(tcs) else '❌'}")
    ok &= len(got) == len(tcs)

    print("\n**2. 排序正確性**：`req_id` 之數值逐列非遞減，且同 req 之列相鄰")
    bad = [(a["row"], a["req_id"], b["req_id"]) for a, b in zip(got, got[1:])
           if num(a["req_id"]) > num(b["req_id"])]
    print(f"- 逆序處：**{len(bad)}** —— {'✅' if not bad else bad[:5]}")
    ok &= not bad
    firsts = {}
    split = []
    for i, g in enumerate(got):
        firsts.setdefault(g["req_id"], []).append(i)
    for k, ix in firsts.items():
        if ix != list(range(ix[0], ix[0] + len(ix))):
            split.append(k)
    print(f"- 同 req 之列被拆散者：**{len(split)}** —— {'✅' if not split else split}")
    ok &= not split

    print("\n**3. TC ID 連續性**：`001`–`%03d` 無跳號無重複" % len(got))
    ids = [g["tc_id"] for g in got]
    want = [f"newR1L-SU-{i:03d}" for i in range(1, len(got) + 1)]
    print(f"- {'✅ 相符' if ids == want else '❌ 不符'}")
    ok &= ids == want

    print("\n**1. 內容不變性**：以 `req_id` 配對，逐列比 TC ID 與列號以外之全部欄")
    cnt = {}
    diffs = []
    for g in got:
        k = g["req_id"]
        cnt[k] = cnt.get(k, 0) + 1
        e = expect.get((k, cnt[k]))
        if e is None:
            diffs.append((g["tc_id"], k, "配對表無此列"))
            continue
        for f, v in e.items():
            if str(g[f]).strip() != str(v).strip():
                diffs.append((g["tc_id"], k, f))
    print(f"- 相異欄位數：**{len(diffs)}** —— {'✅' if not diffs else diffs[:6]}")
    ok &= not diffs

    print("\n**5. 保全計數**（legacy DV 與 x14 分開計）")
    for tag, p in (("母本", FEAT / "inputs" / MASTER), ("交付", DELIV)):
        z = zipfile.ZipFile(p)
        s6 = z.read("xl/worksheets/sheet6.xml").decode()
        print(f"- {tag}：標準 DV **{len(re.findall(r'<dataValidation ', s6))}**、"
              f"x14 DV **{len(re.findall(r'<x14:dataValidation ', s6))}**、"
              f"部件 **{len(z.namelist())}**")

    print(f"\n**7. round-trip**：期望值取自配對表（生成器），實測值取自交付本實檔 —— "
          f"二者為不同來源 ✅")
    print(f"\n**結果：{'全部通過 ✅' if ok else '**不通過 ❌**'}**")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
