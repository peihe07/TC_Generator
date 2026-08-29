#!/usr/bin/env python3
"""T39a —— batch 1（下放包 26 §五）之工作簿產出，供 lint 用。

pilot（`newR1L-SU-001`～`005`）之後之首個正式批次。
沿革：
- **v1**（下放包 26 §五）：4 列（TC-6～TC-9），`181` 待其 VC。
- **v2**（本檔，T41b／T41c）：TC-8 依下放包 28 §2.4 改寫
  （去階段歸屬、明載觀測窗之起訖）；**增 TC-10（`181`，下放包 27 §3.2）**。

TC 內容**逐字取自下放包 26 §五**，執行層不改寫（T32b）。
特記二處：

- **TC-6（`180`）之 `test_item` 上半保留原文之 `shalll`**（037 之拼寫殘留，
  登 `DESCRIPTION_DEFECTS.md` D-4）。R-4 僅允許句首大寫之正規化，
  拼寫不在其列 —— **不得改正**。
- **TC-8（`184`）改判為第三型並掛三個 `PENDING`**（**v3**，下放包 30 §2.2）——
  下放包 28 §2.2 之否證係以一段**不存在之引文**（TC-1「自更新開始執行起錄」）為據，
  已撤銷；上繳包 25 §6.1(丙) 之原結論成立：其增額驗證點**為空**
  （`no confirmation screen` 已由 TC-6／TC-7 覆蓋、`no prompt／progress notification`
  已由 TC-1 覆蓋、`across the three phases` 不可觀測）。所求為**區辨手段**。
- 〔已撤銷〕v2 曾記：不掛 `PENDING` —— 其驗證點可觀測且非空（R-SU33）：
  全稱之否定式由「窗內無違例」驗證即足，**ER 不作階段歸屬之宣稱**（R-SU33(a)），
  且**窗之起訖已明載**（R-SU33(b)）。與 `newR1L-SU-001`（`175`）之區分在**窗之起點**。
- **TC-10（`181`）掛二個 `PENDING`** —— **第三型，且其列不屬 105 列**
  （R-SU32 v2(e) 之首例：語形判準未攔下它，而它同樣不可完整驗證）。
  其限定詞 `immediately after download` 需 (i) 下載完成時點可觀測、
  (ii) 時間閾值 —— 二者皆無，自訂秒數即造值。**其餘部分可觀測，故不整列掛起。**
- **TC-9（`179`）掛三個 `PENDING`**，其成因為 **R-SU32(iii)「不可區辨」**，
  **非「無後果」** —— 其後果與 `175`（`newR1L-SU-001`）完全相同。
  DR-SU2 對本列所求為**區辨手段**，不是觀測手段。

欄集同 pilot v3／v4：`S`（functional_safety）填 `NA`、`T`–`Z` 七個車型旗標留空、
`AH`（remarks）未寫。輸出落 `sandbox/batch01/`（R-G25），`inputs/` 之母本一字不動。

預期 lint：**U=8**（TC-8 之 3 + TC-9 之 3 + TC-10 之 2），其餘各項 0。
"""
import re
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from write_back_036 import _set_row, MASTER, SHEET_NAME, HEADER_ROW, FEAT  # noqa: E402

TAG = "batch01"
OUT = FEAT / "sandbox" / TAG / MASTER
TEST_GROUP, TEST_SET = "SW Update", "Silent Update"
AUTHOR = "PeiPYHsu"
FN = "功能測試 (Functional based ; no specific technique)"
START_N = 6                        # pilot 已用 001–005，本批自 006 起

PRE_STD = [
    "1. The head unit is connected to a Wi-Fi network with internet access",
    "2. An update package with update type Silent Update is staged on the OTA Server for this head unit",
]

TCS = [
 # TC-6 —— `shalll` 為 037 原文之拼寫殘留（D-4），verbatim 不改
 dict(req="SWE1-FOTA-180", spec="CFTS057-4907482", dm=FN, prio="P2",
  item=["When the update type is identified as Silent Update, the WiFi Update Service shalll not trigger the SW Update HMI to display a download confirmation screen.",
        "(No download confirmation screen shown for a silent update)"],
  pre=PRE_STD,
  proc=["1. Read the software version shown on the head unit and record it as Version_initial",
        "2. Trigger an update availability check to the OTA Server",
        "3. Record the head unit screen content as continuous video capture from the availability check until the software version changes",
        "4. Check that no download confirmation screen appears in the recorded screen content"],
  er=["1. Version_initial is recorded",
      "2. The update availability check completes and an update is reported as available",
      "3. The head unit screen content from the availability check until the software version changes is recorded as continuous video capture",
      "4. The recorded screen content contains no download confirmation screen"]),
 # TC-7 —— 與 TC-6 之 sibling 區分：download vs deployment confirmation
 dict(req="SWE1-FOTA-182", spec="CFTS057-4907484", dm=FN, prio="P2",
  item=["The WiFi Update Service shall not trigger the SW Update HMI to display a deployment confirmation screen when the update type is identified as Silent Update.",
        "(No deployment confirmation screen shown for a silent update)"],
  pre=PRE_STD,
  proc=["1. Read the software version shown on the head unit and record it as Version_initial",
        "2. Trigger an update availability check to the OTA Server",
        "3. Record the head unit screen content as continuous video capture from the availability check until the software version changes",
        "4. Check that no deployment confirmation screen appears in the recorded screen content"],
  er=["1. Version_initial is recorded",
      "2. The update availability check completes and an update is reported as available",
      "3. The head unit screen content from the availability check until the software version changes is recorded as continuous video capture",
      "4. The recorded screen content contains no deployment confirmation screen"]),
 # TC-8 —— 獨有驗證點為「三階段皆適用」（ER 第 5 行）
 dict(req="SWE1-FOTA-184", spec="CFTS057-4907486", dm=FN, prio="P1",
  item=["The WiFi Update Service shall apply Silent Update execution rules to all supported update session flows, including update check, deployment package download and installation processing.",
        "(No user-facing interaction from the availability check through installation)"],
  pre=PRE_STD + [
      "3. PENDING: DR-SU2 means of identifying the boundaries between the update check, download and installation phases"],
  proc=["1. Read the software version shown on the head unit and record it as Version_initial",
        "2. Trigger an update availability check to the OTA Server",
        "3. PENDING: DR-SU2 step to record the head unit screen content with the check, download and installation phases identifiable",
        "4. Read the software version shown on the head unit and record it as Version_after",
        "5. Check that Version_after differs from Version_initial and that the recorded screen content contains no SW Update prompt, progress notification or confirmation screen"],
  er=["1. Version_initial is recorded",
      "2. The update availability check completes and an update is reported as available",
      "3. PENDING: DR-SU2 observable evidence delimiting the check, download and installation phases",
      "4. Version_after is recorded",
      "5. Version_after differs from Version_initial; the recorded screen content, taken continuously from the availability check until the software version changes, contains no SW Update prompt, no progress notification and no confirmation screen"]),
 # TC-9 —— PENDING，成因為 R-SU32(iii) 不可區辨（非無後果）
 dict(req="SWE1-FOTA-179", spec="CFTS057-4907481", dm=FN, prio="P1",
  item=["If the DD metadata indicates a Silent Update, the WiFi Update Service shall automatically request SWMC to initiate deployment package download.",
        "(Download request is issued automatically on silent classification)"],
  pre=PRE_STD + [
      "3. PENDING: DR-SU2 means of distinguishing the automatic download request from the overall silent background execution"],
  proc=["1. Trigger an update availability check to the OTA Server",
        "2. PENDING: DR-SU2 step to observe that the deployment package download request has been issued",
        "3. Check that the download request is issued without any user interaction"],
  er=["1. The update availability check completes and an update is reported as available",
      "2. PENDING: DR-SU2 observable evidence that the download request has been issued",
      "3. No user interaction occurs before the download request is issued"]),
 # TC-10 —— `181`，第三型（R-SU32 v2(e)：**不屬 105 列**而同樣不可完整驗證）
 dict(req="SWE1-FOTA-181", spec="CFTS057-4907483", dm=FN, prio="P1",
  item=["Upon receiving deployment package download completion status, the WiFi Update Service shall immediately start installation prechecks and deployment.",
        "(Installation follows download completion with no further interaction)"],
  pre=PRE_STD,
  proc=["1. Read the software version shown on the head unit and record it as Version_initial",
        "2. Trigger an update availability check to the OTA Server",
        "3. Record the head unit screen content as continuous video capture, and any user input, from the availability check until the software version changes",
        "4. PENDING: DR-SU2 step to observe the point at which deployment package download completes",
        "5. Read the software version shown on the head unit and record it as Version_after",
        "6. Check that Version_after differs from Version_initial and that no user input was required between download completion and installation"],
  er=["1. Version_initial is recorded",
      "2. The update availability check completes and an update is reported as available",
      "3. The head unit screen content and user input from the availability check until the software version changes are recorded, the screen content as continuous video capture",
      "4. PENDING: DR-SU2 observable evidence of the download completion point",
      "5. Version_after is recorded",
      "6. Version_after differs from Version_initial; no user input occurred between download completion and installation"]),
]


def project_name():
    """036 母本之專案名稱，實測 D2（標籤 C2）。**不推定**（R-SU24）。"""
    import openpyxl, warnings
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(FEAT / "inputs" / MASTER, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    return str(ws["D2"].value).strip(), str(ws["C2"].value).strip()


def main():
    proj, label = project_name()
    src = FEAT / "inputs" / MASTER
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(src) as z:
        wbxml = z.read("xl/workbook.xml").decode("utf-8")
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        rid = re.search(r'<sheet[^>]*name="%s"[^>]*r:id="([^"]+)"'
                        % re.escape(SHEET_NAME), wbxml).group(1)
        tgt = re.search(r'Id="%s"[^>]*Target="([^"]+)"' % re.escape(rid), rels).group(1)
        sheet_path = "xl/" + tgt.lstrip("/")
        sx = z.read(sheet_path).decode("utf-8")

    print("## T39a —— batch 1 之產出（受檢物，非交付本）\n")
    print(f"- 036 母本專案名稱（實測 `D2`，標籤 `C2` = `{label}`）：**`{proj}`**")
    print(f"- 工作表：`{SHEET_NAME}`｜表頭列 {HEADER_ROW}｜輸出 `sandbox/{TAG}/`（母本未動）\n")

    rows = []
    for i, t in enumerate(TCS):
        n = START_N + i
        tcid = f"{proj}-SU-{n:03d}"
        vals = {"D": t["req"], "F": tcid, "G": TEST_GROUP, "H": TEST_SET,
                "I": "\n".join(t["item"]), "J": "\n".join(t["pre"]),
                "K": "NA", "L": "\n".join(t["proc"]), "M": "\n".join(t["er"]),
                "N": t["spec"], "O": "NEW", "P": t["prio"], "R": t["dm"],
                "S": "NA", "AA": AUTHOR}
        # `T`–`Z`（車型適用旗標）留空；`AH`（remarks）未寫 —— 同 pilot v3
        sx = _set_row(sx, HEADER_ROW + i + 1, vals)
        pend = sum(s.count("PENDING:") for s in t["pre"] + t["proc"] + t["er"])
        rows.append((HEADER_ROW + i + 1, tcid, t["req"], t["spec"], t["prio"], pend))

    with zipfile.ZipFile(src) as zin, \
            zipfile.ZipFile(OUT, "w", zipfile.ZIP_DEFLATED) as zo:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_path:
                data = sx.encode("utf-8")
            zo.writestr(item, data)

    print("| 列 | TC ID | 037 列 | spec_reference | P | PENDING |")
    print("|---|---|---|---|---|---:|")
    for r, tid, req, sp, pr, pd in rows:
        print(f"| {r} | `{tid}` | `{req}` | `{sp}` | {pr} | {pd} |")
    pend = [(tid, req, pd) for _, tid, req, _, _, pd in rows if pd]
    print(f"\n- **PENDING 合計 {sum(p[2] for p in pend)}**，分佈於 {len(pend)} 個 TC —— "
          + "；".join(f"`{tid}`（`{req}`）{pd} 行" for tid, req, pd in pend))
    print("  三者之成因皆為 **R-SU32(iii) 不可區辨**，非「無後果」——"
          "`184` 求三階段界線之辨識手段（下放包 30 §2.2）、"
          "`179` 求下載請求已發出之跡象、`181` 求下載完成時點；"
          "`181` 另為 **R-SU32 v2(e)** 之首例（其列不屬 105 列）。")
    print("- `S` 填 `NA`；`T`–`Z` 留空；`AH` 未寫（欄集同 pilot v3／v4）。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
