#!/usr/bin/env python3
"""T37a／T37b —— `AB`–`AG` 六欄之實測佐證、`Silent Update` 9 列之 VC／VM（下放包 24 §五）。

Usage: python3 scripts/batch1_material.py 37a 37b
"""

import sys
import warnings
from collections import Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from anchor_table import C_ID, C_TITLE, C_CAT, C_SUB, IN_SCOPE          # noqa: E402
from corpus_v2 import _rows_desc                                        # noqa: E402
from observability import classify, RE_EXTERNAL                         # noqa: E402
from verif_columns import _load, _txt, C_VC, C_VM, NUM                  # noqa: E402
from write_back_036 import FEAT                                         # noqa: E402

warnings.filterwarnings("ignore")
ROOT = FEAT.parent.parent
# 036 之 `AB`–`AG`（0-indexed 27–32）
ABAG = {27: "AB Test Version", 28: "AC Test Vehicle (Bench)", 29: "AD Test Period",
        30: "AE Tester", 31: "AF Test Result", 32: "AG Defect ID"}
PILOT9 = ["175", "176", "177", "179", "180", "181", "182", "183", "184"]


def t37a():
    print("## T37a —— 036 `AB`–`AG` 六欄之實測佐證（R-SU28 v3 三之拘束）\n")
    print("> **若任一本有填值，即停並回報 —— R-SU28 v3 三之裁定失效。**\n")
    books = sorted(ROOT.glob("features/*/delivered/*.xlsx")) + \
        sorted(ROOT.glob("features/*/output/*.xlsx")) + \
        sorted(ROOT.glob("features/*/sandbox/*/*.xlsx"))
    print(f"掃描範圍：`features/*/delivered`／`output`／`sandbox/*` 之 xlsx，"
          f"共 **{len(books)}** 本。\n")
    print("| feature | 簿 | 資料列 | " + " | ".join(ABAG.values()) + " |")
    print("|---|---|---:|" + "---|" * len(ABAG))
    bad = []
    scanned = 0
    for b in books:
        try:
            w = openpyxl.load_workbook(b, read_only=True, data_only=True)
            s = next((w[n] for n in w.sheetnames
                      if n.startswith("Test Case Specification")), None)
            if s is None:
                continue
            rs = [r for r in s.iter_rows(min_row=10, values_only=True)
                  if len(r) > 3 and r[3] not in (None, "")]
            if not rs:
                continue
            scanned += 1
            cells = []
            for c in ABAG:
                vals = [_txt(r[c]) for r in rs if len(r) > c]
                nonempty = [v for v in vals if v]
                cells.append(f"**{len(nonempty)} 非空**" if nonempty else f"空×{len(vals)}")
                if nonempty:
                    bad.append((b.parts[b.parts.index("features") + 1], b.name, ABAG[c],
                                Counter(nonempty).most_common(3)))
            feat = b.parts[-4] if b.parts[-3] not in ("delivered", "output") else b.parts[-3]
            feat = b.parts[b.parts.index("features") + 1]
            print(f"| {feat} | `{b.parent.name}/{b.name[:14]}…` | {len(rs)} | "
                  + " | ".join(cells) + " |")
        except Exception as e:                                    # noqa: BLE001
            print(f"| {b.parts[b.parts.index('features') + 1]} | `{b.name[:16]}…` | — | 讀取失敗：{e} |")
    print(f"\n**實掃 {scanned} 本含 TC 分頁且有資料列者。**\n")
    if bad:
        print("### ❌ 有填值者 —— **R-SU28 v3 三之裁定失效**\n")
        print("| feature | 簿 | 欄 | 值（前 3） |")
        print("|---|---|---|---|")
        for f, n, col, vs in bad:
            print(f"| {f} | `{n[:20]}…` | {col} | "
                  + "／".join(f"`{k}`×{v}" for k, v in vs) + " |")
        sys.exit("T37a：`AB`–`AG` 有填值，R-SU28 v3 三失效，停")
    print("### ✅ 六欄於全部已掃之簿**皆為空**\n")
    print("**R-SU28 v3 三之裁定成立**（其依據為欄位語意，本項為其實測佐證）。\n")
    print("> ⚠ **「皆為空」不蘊含「應為空」** —— 亦可能全部漏填。"
          "本項只證明**本 feature 之留空與既有實務一致**，"
          "不證明該六欄不需填（同上繳包 21 §4.1 對 `T`–`Z` 之記明）。")
    return scanned


def t37b():
    by, d, owner = _load()
    internal = {i for i in d if classify(d[i])[0]}
    has_ext_vc = lambda i: any(r.search(_txt(by[i][C_VC])) for r in RE_EXTERNAL.values())
    g105 = {i for i in internal if not has_ext_vc(i)}

    print("\n\n## T37b —— `Silent Update` 9 列之 `Verification Criteria`／`Method`\n")
    print("- 用途：分析層起草 batch 1 時依 **R-SU27(a)** 取其觀測面候選")
    print("- 其 Description 全文與路徑 A 前 5 候選**已備於 "
          "`docs/upstream/17_pilot_material.md` §2**，本節不重複傾印\n")
    print("| # | 037 列 | 標題 | 126 內部列？ | **105 列？** | `Verification Method` |")
    print("|---:|---|---|:--:|:--:|---|")
    for n, k in enumerate(PILOT9, 1):
        i = f"SWE1-FOTA-{k}"
        print(f"| {n} | `{i}` | {_txt(by[i][C_TITLE])[:34]} | "
              f"{'**✅**' if i in internal else '—'} | "
              f"{'**⚠ 是**' if i in g105 else '—'} | `{_txt(by[i][C_VM])[:46]}` |")
    ni = [k for k in PILOT9 if f"SWE1-FOTA-{k}" in internal]
    n105 = [k for k in PILOT9 if f"SWE1-FOTA-{k}" in g105]
    print(f"\n- 屬 126 內部列者：**{len(ni)}** —— {'、'.join('`'+x+'`' for x in ni) or '無'}")
    print(f"- **屬 105 列者：{len(n105)}** —— "
          f"{'、'.join('`'+x+'`' for x in n105) or '**無**'}")
    print(f"\n### 逐列全文\n")
    for n, k in enumerate(PILOT9, 1):
        i = f"SWE1-FOTA-{k}"
        vc = _txt(by[i][C_VC])
        tag = ("**105 列**（內部列且 VC 亦無外部面）" if i in g105
               else "**126 內部列**（但 VC 有外部面）" if i in internal else "非內部列")
        print(f"\n---\n\n#### {n}. `{i}` — {_txt(by[i][C_TITLE])}\n")
        print(f"- 分類：{tag}｜`Verification Method`：`{_txt(by[i][C_VM])}`")
        print(f"\n**`Verification Criteria` 全文**：\n")
        if vc:
            for ln in vc.split("\n"):
                if ln.strip():
                    print(f"> {ln.strip()}")
                    print(">")
        else:
            print("**(空)**")
        print()
    return ni, n105




# ── T38b —— batch 1 材料重列（下放包 25 §四，一次性）─────────────────
BATCH1 = ["179", "180", "181", "182", "184"]


def t38b():
    """五列之 Title、Description 全文、前 5 候選（含候選全文）。

    下放包 24 §四原令「已備於 `17_pilot_material.md` §2，不重複傾印」，
    **下放包 25 §四撤回該指示，改令重列一次**（理由：分析層之閱讀預算 ——
    `17_pilot_material.md` 為 632 行）。**一次性，不立為通例。**
    """
    from corpus_v2 import corpus_v2
    from anchor_table import TfIdf, C_SRC, C_SUB
    from framework_survey import group_by_heading

    objs = corpus_v2()[0]
    tf = TfIdf([o["text"] for o in objs])
    rows_a, d = _rows_desc()
    cand = {_txt(r[C_ID]): [(s, objs[j]) for s, j in tf.query(d[_txt(r[C_ID])], top=20)]
            for r in rows_a}
    by, dd, owner = _load()
    internal = {i for i in dd if classify(dd[i])[0]}
    has_ext_vc = lambda i: any(r.search(_txt(by[i][C_VC])) for r in RE_EXTERNAL.values())
    g105 = {i for i in internal if not has_ext_vc(i)}
    gmap = {g["id"]: g for g in group_by_heading([r for r in _rows_desc()[0]]) [1:]} \
        if False else None

    print("\n\n## T38b —— batch 1 材料重列（`Silent Update` 其餘 5 列）\n")
    print("- Test Set：**`Silent Update`**｜Layer 3 provisional：`4.7.3.2`")
    print("- 其 `Verification Criteria` **已於上繳包 23 §2 備妥，本節不再列**")
    print("- 機制 3 之門檻（R-SU23(b) 改 `≤`）：首選分 **≤ 0.267**\n")
    print("> **執行層不撰寫 TC、不裁定錨。**\n")
    print("| # | 037 列 | 標題 | Sub Cat | **105 列？** | 首選分 | 機制 3 |")
    print("|---:|---|---|---|:--:|---:|:--:|")
    for n, k in enumerate(BATCH1, 1):
        i = f"SWE1-FOTA-{k}"
        s = cand[i][0][0]
        print(f"| {n} | `{i}` | {_txt(by[i][C_TITLE])[:34]} | {by[i][C_SUB] or '(blank)'} | "
              f"{'**⚠ 是**' if i in g105 else '—'} | {s:.3f} | "
              f"{'**⚠ 攔下**' if s <= 0.26716366259482566 else '—'} |")
    print("\n---\n\n### 逐列材料\n")
    for n, k in enumerate(BATCH1, 1):
        i = f"SWE1-FOTA-{k}"
        r = by[i]
        tag = ("**105 列** —— 其 TC 將撞上 R-SU29(c)（R-SU31(c)：本 feature 首個進入撰寫之 105 列）"
               if i in g105 else "**126 內部列**（但 VC 有外部面）" if i in internal else "非內部列")
        print(f"\n---\n\n#### {n}. `{i}` — {_txt(r[C_TITLE])}\n")
        print(f"- 分類：{tag}")
        print(f"- Sub Cat：{r[C_SUB] or '(blank)'}｜Priority：{r[15] or '(blank)'}"
              f"｜Source：`{r[C_SRC]}`")
        print(f"\n**Requirement Description 全文**：\n")
        print("> " + (d[i] or "(空)") + "\n")
        print("**路徑 A（語料 v2）前 5 候選**：\n")
        for j, (sc, o) in enumerate(cand[i][:5], 1):
            print(f"{j}. `{o['oid']}` — 章 **{o['chap']}** {o['chap_title']} — 分 **{sc:.3f}**")
            print(f"   > {o['text'][:420]}{'…' if len(o['text'])>420 else ''}\n")


if __name__ == "__main__":
    want = set(sys.argv[1:]) or {"37a", "37b"}
    if "37a" in want:
        t37a()
    if "37b" in want:
        t37b()
    if "38b" in want:
        t38b()
