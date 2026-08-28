#!/usr/bin/env python3
"""T13a–c —— 下放包 08 §七 之三項量測。

T13a  `LID Proxi & Configuration` 全分頁掃重複 Logical Identifier：
      列出所有重複之 LID 名、列號、各列非空欄數、各架構欄值。
      **只給量測與分布，不下結論**（r420／r421 由分析層裁）。
T13b  037 全 28 列以 MPH／mph／mile／km/h／kph 五組字樣重掃門檻表述，
      補上繳包 05 §2.3 之已知邊界（該輪限 `MPH` 字樣）。
T13c  `Gear_Box_Type` 於二 DBC 之存在性實測 ——
      關掉上繳包 05 §3.5(丙) 之未量測斷言。

引用格式依 R-DD10：Excel 書欄名（不書 c{n}）、LID 列須標架構欄、
計數須書母體判準與排除項、列號一律 1-based。
四庫綁 `features/vehicle_setting/inputs/`（R-DD5），全程 read_only。
"""
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
VS = ROOT.parent / "vehicle_setting" / "inputs"
LID = VS / "Logical Identifiers and CAN Mapping v1_76.xlsx"
DBC = [VS / "PDT27_E2A_R4_BHCAN.dbc", VS / "PDT27_E2A_R5_FDCAN8.dbc"]
SWE1 = ROOT / "inputs" / "DD_SWE1_0807_EN.xlsx"
SHEET = "Proxi & Configuration"


def cn(j):
    """0-based 欄索引 → Excel 欄名（R-DD10(a)）。"""
    s, j = "", j + 1
    while j:
        j, r = divmod(j - 1, 26)
        s = chr(65 + r) + s
    return s


def load_sheet():
    wb = openpyxl.load_workbook(LID, read_only=True, data_only=True)
    rows = [list(r) for r in wb[SHEET].iter_rows(values_only=True)]
    wb.close()
    bands = {j: str(v).strip() for j, v in enumerate(rows[1]) if v not in (None, "")}
    names = [str(v) if v is not None else "" for v in rows[2]]
    return rows, bands, names


def band_start(bands, label):
    return next(j for j, v in bands.items() if v == label)


def t13a():
    rows, bands, names = load_sheet()
    print("=" * 72)
    print("T13a —— LID `Proxi & Configuration` 重複 Logical Identifier 之分布")
    print(f"母體判準：r4 起至 r{len(rows)}（r1 表題／r2 架構帶／r3 欄名，排除）；")
    print("          A 欄 `Logical Identifier` 非空者計入；空白列排除。")
    print("架構帶（自 r2 讀取）：" + "／".join(f"{cn(j)}={v}" for j, v in sorted(bands.items())))

    pn, cu, at = (band_start(bands, x) for x in
                  ("Powernet", "CUSW", "Atlantis & Atlantis High"))
    sig = {"Powernet": pn, "CUSW": cu, "Atlantis": at}

    groups = defaultdict(list)
    total = 0
    for i, row in enumerate(rows[3:], 4):
        lid = row[0]
        if lid in (None, ""):
            continue
        total += 1
        groups[str(lid)].append(i)

    dups = {k: v for k, v in groups.items() if len(v) > 1}
    print(f"\n非空 Logical Identifier 列數 = {total}；unique = {len(groups)}；"
          f"**重複之 LID 名 = {len(dups)}**，佔 {sum(len(v) for v in dups.values())} 列")

    # 分布：每組之列數
    dist = defaultdict(int)
    for v in dups.values():
        dist[len(v)] += 1
    print("重複組之列數分布：" + "／".join(f"{k} 列 × {dist[k]} 組" for k in sorted(dist)))

    # 「稀疏＋完整」之形態量測：非空欄數 與 各架構欄值
    print("\n逐組逐列（非空欄數 = 該列非 None 且非空字串之欄數）：")
    print(f"{'LID 名':<34}{'列':>6} {'非空欄':>6}  "
          f"{'Powernet 欄':<30}{'CUSW 欄':<26}{'Atlantis 欄':<26}")
    print("-" * 130)
    shapes = defaultdict(int)
    for k in sorted(dups, key=lambda x: dups[x][0]):
        for i in dups[k]:
            row = rows[i - 1]
            ne = sum(1 for v in row if v not in (None, ""))
            vals = {b: (row[j] if j < len(row) else None) for b, j in sig.items()}
            print(f"{k[:33]:<34}{i:>6} {ne:>6}  "
                  f"{str(vals['Powernet'])[:29]:<30}"
                  f"{str(vals['CUSW'])[:25]:<26}"
                  f"{str(vals['Atlantis'])[:25]:<26}")
        # 形態鍵：組內各列之非空欄數序列
        nes = tuple(sum(1 for v in rows[i - 1] if v not in (None, "")) for i in dups[k])
        shapes[("遞增" if list(nes) == sorted(nes) and len(set(nes)) > 1
                else "遞減" if list(nes) == sorted(nes, reverse=True) and len(set(nes)) > 1
                else "相等" if len(set(nes)) == 1 else "無序")] += 1
        print()
    print("組內非空欄數之序（量測，不判讀）：" +
          "／".join(f"{k} {v} 組" for k, v in sorted(shapes.items())))

    # 各架構欄之 'Not Applicable' 分布（r420 之值）
    na = defaultdict(int)
    for i, row in enumerate(rows[3:], 4):
        for b, j in sig.items():
            if j < len(row) and str(row[j]).strip() == "Not Applicable":
                na[b] += 1
    print("\n全分頁 `Not Applicable` 之分布（母體同上）：" +
          "／".join(f"{b} 欄 {na[b]} 列" for b in sig))


def t13b():
    print("\n" + "=" * 72)
    print("T13b —— 037 全 28 列之門檻表述重掃（五組字樣）")
    pats = {"MPH": r"\bMPH\b", "mph": r"\bmph\b", "mile": r"mile",
            "km/h": r"km\s*/\s*h", "kph": r"\bkph\b"}
    wb = openpyxl.load_workbook(SWE1, read_only=True, data_only=True)
    rows = list(wb["Analysis Report"].iter_rows(values_only=True))
    wb.close()
    print("母體判準：`Analysis Report` 全列中含 `SWE1-RA-Driver_Distraction-{n}` 者；")
    print("          比對為全列各格串接後之正則掃描（大小寫敏感者已分列）。\n")
    print(f"{'leaf':<8}{'037 列':>7}  " + "".join(f"{k:<8}" for k in pats) + " 命中之字樣")
    print("-" * 92)
    n_any = 0
    for i, row in enumerate(rows, 1):
        txt = " ".join(str(c) for c in row if c is not None)
        m = re.search(r"SWE1-RA-Driver_Distraction-(\d+)", txt)
        if not m:
            continue
        hits = {k: bool(re.search(p, txt)) for k, p in pats.items()}
        lits = sorted(set(re.findall(r"[\d.]+\s*(?:MPH|mph|kph|km\s*/\s*h)", txt)))
        if any(hits.values()):
            n_any += 1
        print(f"-{m.group(1):<7}{i:>7}  " +
              "".join(f"{'Y' if hits[k] else '.':<8}" for k in pats) +
              " " + (", ".join(lits) if lits else "—"))
    print(f"\n任一字樣命中之 leaf 數 = {n_any}")


def t13c():
    print("\n" + "=" * 72)
    print("T13c —— `Gear_Box_Type` 於二綁定 DBC 之存在性實測")
    print("母體：二 DBC 全文；比對 `SG_` 定義名、`BO_` 訊息名、`VAL_` 列舉行，")
    print("      另加全文裸字串掃描（大小寫不敏感）—— 三者分開計。\n")
    for p in DBC:
        txt = p.read_text("utf-8", errors="replace")
        nmsg = len(re.findall(r"^BO_ ", txt, re.M))
        nsg = len(re.findall(r"^\s*SG_ ", txt, re.M))
        print(f"[{p.name}]  BO_ {nmsg} 訊息／SG_ {nsg} 訊號")
        for label, pat in (
            ("SG_ 定義名", r"^\s*SG_\s+\w*Gear_Box_Type\w*\b"),
            ("BO_ 訊息名", r"^BO_ \d+ \w*Gear_Box_Type\w*\b"),
            ("VAL_ 列舉行", r"^VAL_.*Gear_Box_Type"),
            ("全文裸字串（大小寫不敏感）", r"gear_box_type"),
        ):
            hits = re.findall(pat, txt, re.M | (re.I if "不敏感" in label else 0))
            print(f"    {label:<26}: {len(hits)} 命中"
                  + ("" if not hits else "  -> " + str(hits[:3])))
        # 對照：確認掃描法有效 —— 已知在庫之名
        for known in ("VehicleSpeedVSOSig", "GearEngagedForDisplay_PT"):
            n = len(re.findall(rf"^\s*SG_\s+{known}\b", txt, re.M))
            print(f"    [對照] SG_ {known:<24}: {n} 命中")
        print()





def t13a2():
    """T13a 之二 —— 17 組配對之逐欄比對：同欄皆非空且值不同者，逐字列出。

    這才是 §四 所問之實質：`Not Applicable` vs `Gear_Box_Type` 之互斥，
    是該表之個案還是系統性形態。**只給分布，不下結論。**
    """
    rows, bands, names = load_sheet()
    groups = defaultdict(list)
    for i, row in enumerate(rows[3:], 4):
        if row[0] not in (None, ""):
            groups[str(row[0])].append(i)
    dups = {k: v for k, v in groups.items() if len(v) > 1}

    print("\n" + "=" * 72)
    print("T13a-2 —— 17 組配對之逐欄比對")
    ncol = max(len(rows[i - 1]) for v in dups.values() for i in v)
    tally = defaultdict(int)
    conflict_rows = []
    for k in sorted(dups, key=lambda x: dups[x][0]):
        a, b = (rows[i - 1] for i in dups[k])
        both_diff, both_same, one_only = [], 0, 0
        for j in range(ncol):
            va = a[j] if j < len(a) else None
            vb = b[j] if j < len(b) else None
            ea, eb = va in (None, ""), vb in (None, "")
            if ea and eb:
                continue
            if ea != eb:
                one_only += 1
            elif str(va) == str(vb):
                both_same += 1
            else:
                both_diff.append((j, va, vb))
        kind = ("完全一致" if not both_diff and not one_only else
                "純補全（一空一有，無衝突）" if not both_diff else
                "**同欄異值**")
        tally[kind] += 1
        if both_diff:
            conflict_rows.append((k, dups[k], both_diff))
        print(f"  {k:<26} r{dups[k][0]}/r{dups[k][1]}  "
              f"同欄同值 {both_same:>2}／一空一有 {one_only:>2}／同欄異值 {len(both_diff):>2}  {kind}")

    print("\n形態分布（母體 = 17 組重複之 LID）：")
    for k, v in sorted(tally.items(), key=lambda x: -x[1]):
        print(f"  {k:<28} {v:>2} 組")

    print("\n同欄異值者之逐字（欄名依 R-DD10(a) 書 Excel 欄名）：")
    for k, (ra, rb), diffs in conflict_rows:
        print(f"\n  [{k}] r{ra} vs r{rb}")
        for j, va, vb in diffs:
            b = next(v for s, v in sorted(bands.items(), reverse=True) if j >= s)
            print(f"    {cn(j):>3} [{b} 帶 · {names[j]}]")
            print(f"        r{ra} = {va!r}")
            print(f"        r{rb} = {vb!r}")




def t13a3():
    """T13a 之三 —— 衝突欄之集中度，與 `Not Applicable` 之出現位置。

    量測 16 組「同欄異值」之衝突落在哪些欄、集中程度如何，
    以及 r420／r421 之形態在該分布中的位置。**只給分布，不下結論。**
    """
    rows, bands, names = load_sheet()
    groups = defaultdict(list)
    for i, row in enumerate(rows[3:], 4):
        if row[0] not in (None, ""):
            groups[str(row[0])].append(i)
    dups = {k: v for k, v in groups.items() if len(v) > 1}

    ncol = max(len(rows[i - 1]) for v in dups.values() for i in v)
    per_col = defaultdict(list)      # 欄 -> 有衝突之組
    sig_only = defaultdict(list)     # 衝突欄集合 -> 組
    na_pairs = []
    for k in sorted(dups, key=lambda x: dups[x][0]):
        a, b = (rows[i - 1] for i in dups[k])
        cols = []
        for j in range(ncol):
            va = a[j] if j < len(a) else None
            vb = b[j] if j < len(b) else None
            if va in (None, "") or vb in (None, ""):
                continue
            if str(va) != str(vb):
                cols.append(j)
                per_col[j].append(k)
                if "Not Applicable" in (str(va).strip(), str(vb).strip()):
                    if k not in [x[0] for x in na_pairs]:
                        na_pairs.append((k, dups[k]))
        if cols:
            sig_only[tuple(cn(j) for j in cols)].append(k)

    print("\n" + "=" * 72)
    print("T13a-3 —— 衝突欄之集中度（母體 = 16 組「同欄異值」）")
    print("\n逐欄之衝突組數：")
    for j in sorted(per_col, key=lambda x: -len(per_col[x])):
        b = next(v for s, v in sorted(bands.items(), reverse=True) if j >= s)
        print(f"  {cn(j):>3} [{b} 帶 · {names[j]}]  {len(per_col[j]):>2} 組")

    print("\n依「衝突欄之組合」分群：")
    for combo, ks in sorted(sig_only.items(), key=lambda x: -len(x[1])):
        print(f"  {'＋'.join(combo):<16} {len(ks):>2} 組   {', '.join(ks)[:88]}")

    print("\n衝突值含 `Not Applicable` 者：")
    if not na_pairs:
        print("  無")
    for k, v in na_pairs:
        print(f"  {k}  r{v[0]}/r{v[1]}")

    print("\nG 欄（Powernet CAN）之衝突值配對逐字：")
    gpairs = defaultdict(int)
    for k in per_col.get(6, []):
        a, b = (rows[i - 1] for i in dups[k])
        gpairs[(str(a[6]), str(b[6]))] += 1
    for (x, y), n in sorted(gpairs.items(), key=lambda t: -t[1]):
        print(f"  上列 {x!r} → 下列 {y!r}   {n} 組")


if __name__ == "__main__":
    t13a()
    t13a2()
    t13a3()
    t13b()
    t13c()
