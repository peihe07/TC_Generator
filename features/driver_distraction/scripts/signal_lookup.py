#!/usr/bin/env python3
"""T6 —— R-DD5 之逐訊號對照初查（下放包 02 §四）。

五個 `$…$` 參數對四庫逐一查：DBC 之 `SG_`／`VAL_`、LID 之對照表、
PROXI 之參數名。**查得／查無皆逐項輸出** —— 查無者不得代以語意相近之
他訊號（R-13／IN §8.7.5(d)(g)），其 DR 由分析層擬稿。

四庫綁 `features/vehicle_setting/inputs/` 之原件（R-DD5），不複製入本 feature。
"""
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
VS = ROOT.parent / "vehicle_setting" / "inputs"
LID = VS / "Logical Identifiers and CAN Mapping v1_76.xlsx"
DBC = [VS / "PDT27_E2A_R4_BHCAN.dbc", VS / "PDT27_E2A_R5_FDCAN8.dbc"]
PROXI = VS / "PROXI_HDCC27_R3_20250424.xlsx"

SIGNALS = ["Speedometer", "VC_Trans_Equipped", "PresentGear",
           "PARK_BRK_EGD", "Country_Code"]


def dbc_hits(name):
    """`SG_ <name>` 之定義列、與含該名之 `VAL_` 列舉。逐字，不做模糊比對。"""
    out = []
    for p in DBC:
        txt = p.read_text("utf-8", errors="replace")
        for m in re.finditer(rf"^\s*SG_\s+{re.escape(name)}\b.*$", txt, re.M):
            out.append((p.name, "SG_", m.group(0).strip()[:150]))
        for m in re.finditer(rf"^VAL_\s+\d+\s+{re.escape(name)}\b.*$", txt, re.M):
            out.append((p.name, "VAL_", m.group(0).strip()[:220]))
        # 其所屬之 MESSAGE 全名
        for m in re.finditer(rf"^BO_ (\d+) (\w+)[^\n]*\n(?:\s*SG_[^\n]*\n)*?"
                             rf"\s*SG_\s+{re.escape(name)}\b", txt, re.M):
            out.append((p.name, "BO_", f"{m.group(2)} (id {m.group(1)})"))
    return out


def xlsx_hits(path, name, maxhits=4):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for sn in wb.sheetnames:
        for i, row in enumerate(wb[sn].iter_rows(values_only=True), 1):
            for j, c in enumerate(row):
                if c is not None and name.lower() in str(c).lower():
                    ctx = " | ".join(str(x)[:40] for x in row
                                     if x not in (None, ""))[:200]
                    out.append((sn, i, ctx))
                    break
            if len(out) >= maxhits:
                wb.close()
                return out
    wb.close()
    return out


def main():
    print("## T6 —— 五訊號對四庫之逐項查對\n")
    print(f"- LID：`{LID.name}`")
    print(f"- DBC：`{DBC[0].name}`／`{DBC[1].name}`")
    print(f"- PROXI：`{PROXI.name}`")
    print("- **四庫皆綁 `features/vehicle_setting/inputs/` 之原件**（R-DD5），"
          "未複製入本 feature\n")
    summary = []
    for s in SIGNALS:
        print(f"### `${s}$`\n")
        d = dbc_hits(s)
        print(f"**DBC**：{len(d)} 處" + ("" if d else " —— **查無**"))
        for src, kind, txt in d[:6]:
            print(f"  - `{src}` [{kind}] {txt}")
        l = xlsx_hits(LID, s)
        print(f"\n**LID**：{len(l)} 處" + ("" if l else " —— **查無**"))
        for sn, i, ctx in l:
            print(f"  - `{sn}` r{i}: {ctx}")
        p = xlsx_hits(PROXI, s)
        print(f"\n**PROXI**：{len(p)} 處" + ("" if p else " —— **查無**"))
        for sn, i, ctx in p:
            print(f"  - `{sn}` r{i}: {ctx}")
        found = [k for k, v in (("DBC", d), ("LID", l), ("PROXI", p)) if v]
        summary.append((s, found))
        print(f"\n**小結**：{'／'.join(found) if found else '**四庫皆查無**'}\n")
    print("---\n\n## 彙總\n")
    print("| 訊號 | 查得於 | 判 |\n|---|---|---|")
    for s, found in summary:
        print(f"| `${s}$` | {'／'.join(found) or '—'} | "
              f"{'查得' if found else '**查無 —— 須登 DR**'} |")
    n = sum(1 for _, f in summary if not f)
    print(f"\n**查無 {n} / {len(SIGNALS)}** —— 查無者依 IN §8.7.5(d)(g) "
          f"保留來源名稱並逐項登 DR，**不得代以語意相近之他訊號**（R-13）。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
