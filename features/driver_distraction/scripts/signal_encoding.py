#!/usr/bin/env python3
"""T9a–c —— 訊號之 DBC／PROXI 編碼實測（下放包 04 §三）。

**只查證、只列原值**：
  T9a  逐訊號輸出 BO_ 歸屬、位長、scaling、`VAL_` 列舉逐字；查無者列表
  T9b  PROXI 之 `Country_Code` 值域，輸出 Hong Kong 之列舉值
  T9c  5/3 MPH 之 raw 換算 —— **只列 DBC factor/offset 原值**，
       換算式與結果由分析層覆核，**執行層不逕填 TC 值**（§8.4.1）

`STATUS_BH_BCM1.ParkBrakeSts`／`BCM_FD_9.ParkBrakeSts` 為 A-DD2 之
**候選對應，僅查證不採用**（下放包 04 T9a 明文；R-DD5／R-13）。
"""
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
VS = ROOT.parent / "vehicle_setting" / "inputs"
DBC = [VS / "PDT27_E2A_R4_BHCAN.dbc", VS / "PDT27_E2A_R5_FDCAN8.dbc"]
PROXI = VS / "PROXI_HDCC27_R3_20250424.xlsx"

# (顯示名, BO_ 名, SG_ 名, 是否僅查證不採用)
TARGETS = [
    ("$Speedometer$",        "GW_C1",           "VEH_SPEED",          False),
    ("$VC_Trans_Equipped$",  "VehCfg7",         "VC_Trans_Equipped",  False),
    ("$PresentGear$",        "GW_C1",           "Gr",                 False),
    ("$PARK_BRK_EGD$ 候選",  "STATUS_BH_BCM1",  "ParkBrakeSts",       True),
    ("$PARK_BRK_EGD$ 候選",  "BCM_FD_9",        "ParkBrakeSts",       True),
]

SG = re.compile(
    r"^\s*SG_\s+(?P<name>\w+)\s*:\s*(?P<start>\d+)\|(?P<len>\d+)@(?P<order>[01])"
    r"(?P<sign>[+-])\s*\(\s*(?P<factor>[^,]+),\s*(?P<offset>[^)]+)\)\s*"
    r"\[(?P<min>[^|]*)\|(?P<max>[^\]]*)\]\s*\"(?P<unit>[^\"]*)\"", re.M)


def dbc_index(path):
    """BO_ → [SG_ 定義]；並收 VAL_ 列舉。"""
    txt = path.read_text("utf-8", errors="replace")
    msgs, cur = {}, None
    for line in txt.splitlines():
        m = re.match(r"^BO_\s+(\d+)\s+(\w+)\s*:\s*(\d+)", line)
        if m:
            cur = {"id": m.group(1), "name": m.group(2), "dlc": m.group(3), "sgs": []}
            msgs[m.group(2)] = cur
            continue
        if cur is not None and line.strip().startswith("SG_"):
            g = SG.match(line)
            cur["sgs"].append(g.groupdict() if g else {"name": line.strip()[:60],
                                                       "raw": line.strip()})
        elif line.strip() and not line.startswith(" "):
            cur = None
    vals = {}
    for m in re.finditer(r"^VAL_\s+(\d+)\s+(\w+)\s+(.*?);\s*$", txt, re.M):
        vals.setdefault(m.group(2), []).append((m.group(1), m.group(3).strip()))
    return msgs, vals


def main():
    idx = {p.name: dbc_index(p) for p in DBC}
    print("## T9a —— DBC 逐訊號實測\n")
    print("四庫之 DBC 綁 `features/vehicle_setting/inputs/`（R-DD5），"
          "**未複製入本 feature**。\n")
    missing = []
    for label, bo, sg, probe_only in TARGETS:
        tag = "　**（僅查證不採用，A-DD2）**" if probe_only else ""
        print(f"### {label} → `{bo}.{sg}`{tag}\n")
        hit = False
        for fname, (msgs, vals) in idx.items():
            m = msgs.get(bo)
            if not m:
                continue
            for s in m["sgs"]:
                if s.get("name") != sg:
                    continue
                hit = True
                print(f"**`{fname}`**")
                print(f"- `BO_` **{m['name']}**（id {m['id']}、DLC {m['dlc']}）")
                if "start" in s:
                    print(f"- `SG_ {s['name']}` : "
                          f"start **{s['start']}**、長度 **{s['len']}** bit、"
                          f"byte order `@{s['order']}`（{'Intel' if s['order']=='1' else 'Motorola'}）、"
                          f"符號 `{s['sign']}`")
                    print(f"- **factor `{s['factor'].strip()}`、offset `{s['offset'].strip()}`**、"
                          f"範圍 `[{s['min']}|{s['max']}]`、單位 `\"{s['unit']}\"`")
                else:
                    print(f"- 定義行未匹配樣式，逐字：`{s.get('raw','')}`")
                if sg in vals:
                    print(f"- **`VAL_` 列舉（逐字）**：")
                    for mid, body in vals[sg]:
                        print(f"  - msg {mid}: `{body[:300]}`")
                else:
                    print(f"- `VAL_` 列舉：**無**")
                print()
        if not hit:
            missing.append(f"{bo}.{sg}")
            print("**二個 DBC 皆查無。**\n")
    print(f"**查無清單**：{missing or '無'}\n")

    print("---\n\n## T9b —— PROXI `Country_Code` 值域\n")
    wb = openpyxl.load_workbook(PROXI, read_only=True, data_only=True)
    found = []
    for sn in wb.sheetnames:
        for i, row in enumerate(wb[sn].iter_rows(values_only=True), 1):
            blob = " | ".join(str(c) for c in row if c not in (None, ""))
            if re.search(r"country[_ ]?code", blob, re.I):
                found.append((sn, i, blob))
    wb.close()
    print(f"命中 **{len(found)}** 列\n")
    for sn, i, blob in found[:8]:
        print(f"- `{sn}` r{i}: `{blob[:420]}`")
    hk = [f for f in found if re.search(r"hong ?kong|\bHK\b", f[2], re.I)]
    print(f"\n**含 `Hong Kong`／`HK` 者：{len(hk)} 列**"
          + ("" if hk else " —— **查無 → DR 候選（分析層擬 DR-DD3）**"))
    for sn, i, blob in hk[:5]:
        print(f"- `{sn}` r{i}: `{blob[:420]}`")

    print("\n---\n\n## T9c —— 5／3 MPH 之 raw 換算：**只列原值**\n")
    for fname, (msgs, vals) in idx.items():
        m = msgs.get("GW_C1")
        if not m:
            continue
        for s in m["sgs"]:
            if s.get("name") == "VEH_SPEED" and "factor" in s:
                print(f"`{fname}` → `GW_C1.VEH_SPEED`：")
                print(f"- **factor = `{s['factor'].strip()}`**")
                print(f"- **offset = `{s['offset'].strip()}`**")
                print(f"- 範圍 `[{s['min']}|{s['max']}]`、單位 `\"{s['unit']}\"`、"
                      f"長度 {s['len']} bit")
    print("\n> **執行層不逕填 TC 值**（下放包 04 T9c、§8.4.1）——")
    print("> 換算式與結果由分析層覆核後入 profile。")
    print("> ⚠ **單位須先確認**：spec 之門檻為 **MPH**；若 DBC 之單位為 km/h，")
    print("> 換算涉及單位轉換，**其係數不在 DBC 內**，須另有依據。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
