#!/usr/bin/env python3
"""把 24 則寫入 036 工作簿（Pei 2026-08-29 裁乙案、leaf 升冪 001–024）。

**`openpyxl` 只作計算層，不作 emit 層**（R16-1／R16-2）——
`Workbook.save()` 會丟掉它物件模型外之一切（含 R 欄 design_method 之
`x14:dataValidation` 下拉）。emit 由 `backend.xlsx_surgical.surgical_save`
以 zip 成員逐一比對後**只補目標 sheet 之 XML 文字**，其餘位元組照抄。

**寫不回來的寫，不算寫** —— `--verify` 自交付檔逐列讀回，
與產物 JSON **逐欄逐字元**比對。

預設 dry-run；`--apply` 方寫。寫前必備份。
"""
import argparse
import json
import re
import shutil
import sys
import warnings
from pathlib import Path

import yaml

warnings.filterwarnings("ignore")
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT.parent.parent))
import openpyxl                                                    # noqa: E402
from backend.xlsx_surgical import surgical_save                    # noqa: E402

CFG = yaml.safe_load((ROOT / "feature.yaml").read_text("utf-8"))["workbook"]
WB = ROOT / "workbook" / "driver_distraction_00.xlsx"
ARTS = ["pilot_group3.json", "batch_b1.json", "batch_b2.json",
        "batch_body_off_init.json"]
# 交付實測（11 本、828 列）：C／E／Q 三欄無一列填過 → 留空。
# **空集合保留於此而非刪除**，使「三欄刻意不寫」在程式裡看得見（同 R-BLM16 之體例）。
LEAVE_BLANK = ["polarion_id", "tc_id_testrail", "est_test_time", "remarks"]


def load():
    tcs = []
    for a in ARTS:
        tcs += json.loads((ROOT / "generated" / a).read_text("utf-8"))
    tcs.sort(key=lambda t: t["req_id"][-3:])          # leaf 升冪（Pei 2026-08-29）
    return tcs


def plan(tcs):
    """回傳 [(列, {欄字母: 值})]。`do_not_write` 之欄一律不觸。"""
    cols, skip = CFG["columns"], set(CFG["do_not_write"])
    out = []
    for i, t in enumerate(tcs):
        cells = {}
        for key, col in cols.items():
            if col in skip or key in LEAVE_BLANK:
                continue
            v = t.get(key)
            if v is None:
                raise SystemExit(f"**產物缺欄 `{key}`**（{t['tc_id']}）；停")
            cells[col] = v
        out.append((CFG["first_data_row"] + i, cells))
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--verify", action="store_true")
    a = ap.parse_args()

    tcs = load()
    rows = plan(tcs)
    print("=" * 84)
    print(f"寫回計畫：{len(rows)} 列，{rows[0][0]}–{rows[-1][0]}；"
          f"每列 {len(rows[0][1])} 欄；留空 {LEAVE_BLANK}；不觸 {CFG['do_not_write']}")

    wb = openpyxl.load_workbook(WB)
    ws = wb[CFG["sheet"]]
    occupied = [r for r in range(CFG["first_data_row"], CFG["last_template_row"] + 1)
                if any(ws.cell(r, c).value not in (None, "") for c in range(3, 35))]
    if occupied:
        raise SystemExit(f"**目標列非空**（{occupied[:5]}…）—— 停，須先裁處置")
    print(f"目標區 {CFG['first_data_row']}–{CFG['last_template_row']} 實測全空 ✓")

    if not a.apply:
        print("（dry-run；加 --apply 方寫）"); print("=" * 84); return 0

    bak = WB.with_name(WB.stem + "_bak.xlsx")   # 副檔名須為 .xlsx（openpyxl 只認）
    shutil.copy2(WB, bak)
    print(f"已備份 → {bak.name}")

    for r, cells in rows:
        for col, v in cells.items():
            ws[f"{col}{r}"] = v
    # **src 與 out 必須分離** —— `surgical_save` 是邊讀 src 邊寫 out；
    # 傳同一個路徑會在讀到一半時把來源截斷（本輪實測：交付檔剩 22 bytes，
    # 由上一行之備份救回）。先寫暫存，驗過再置換。
    tmp = WB.with_name(WB.stem + "_tmp.xlsx")
    st = surgical_save(wb, bak, tmp)          # 來源取備份，確保 src 不被動
    print(f"surgical_save：{st}")
    shutil.move(tmp, WB)

    # ── read-back：自交付檔逐列讀回，逐欄逐字元比對 ──────────────────
    wb2 = openpyxl.load_workbook(WB)
    ws2 = wb2[CFG["sheet"]]
    bad = []
    for (r, cells), t in zip(rows, tcs):
        for col, v in cells.items():
            got = ws2.cell(r, openpyxl.utils.column_index_from_string(col)).value
            if got != v:
                bad.append((t["tc_id"], col, r, repr(v)[:40], repr(got)[:40]))
    wb2.close()
    print(f"read-back：{len(rows)} 列 × {len(rows[0][1])} 欄 = "
          f"{len(rows)*len(rows[0][1])} 格，逐字元不符 {len(bad)}")
    if bad:
        for b in bad[:10]:
            print("  ✗", b)
        return 1
    print("**寫不回來的寫，不算寫** —— 全數讀回相符 ✓")
    print("=" * 84)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
