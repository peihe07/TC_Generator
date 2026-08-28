#!/usr/bin/env python3
"""T12a–c —— `$VC_Trans_Equipped$` 之 PROXI／LID 窮盡量測（下放包 07 §四）。

T12a  PROXI 全檔查 `Gear_Box_Type`：所在 component／byte／bit、值域與列舉逐字
T12b  PROXI 全檔查 `VC_Trans_Equipped` 之直接命中（T6 記為 0 處，本輪覆核）
T12c  LID `Proxi & Configuration` r420／r421 全欄逐字，並標明架構帶歸屬

拘束：**只量測，不寫 profile**。查得或查無皆逐字輸出。
架構帶自分頁 r2 之合併標題列讀取（非硬編），欄名取 r3；
引用格式依 R-DD6(c)：`LID {分頁名} r{n} [{架構}欄]`。

四庫綁 `features/vehicle_setting/inputs/` 之原件（R-DD5），全程 read_only。
"""
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
VS = ROOT.parent / "vehicle_setting" / "inputs"
LID = VS / "Logical Identifiers and CAN Mapping v1_76.xlsx"
PROXI = VS / "PROXI_HDCC27_R3_20250424.xlsx"

SHEET = "Proxi & Configuration"


def col(j):
    """0-based 欄索引 → Excel 欄名。"""
    s, j = "", j + 1
    while j:
        j, r = divmod(j - 1, 26)
        s = chr(65 + r) + s
    return s


def bands(row2):
    """自 r2 之合併標題列讀架構帶：{起始欄索引: 帶名}，不硬編。"""
    return {j: str(v).strip() for j, v in enumerate(row2) if v not in (None, "")}


def band_of(j, bnd):
    """欄索引落在哪一個架構帶。"""
    starts = sorted(bnd)
    cur = None
    for s in starts:
        if j >= s:
            cur = bnd[s]
        else:
            break
    return cur


def scan(path, needles, label):
    """全分頁逐格 substring 查（大小寫不敏感），逐列輸出命中。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"\n{'=' * 72}\n{label} —— {path.name}\nsheets: {wb.sheetnames}")
    found = {n: [] for n in needles}
    for sn in wb.sheetnames:
        for i, row in enumerate(wb[sn].iter_rows(values_only=True), 1):
            for j, c in enumerate(row):
                if c is None:
                    continue
                s = str(c).lower()
                for n in needles:
                    if n.lower() in s:
                        found[n].append((sn, i, j, c, row))
                        break
    for n in needles:
        print(f"\n-- TARGET {n!r}: {len(found[n])} 命中")
        for sn, i, j, c, row in found[n]:
            print(f"   [{sn}] r{i} c{j}/{col(j)} = {c!r}")
    wb.close()
    return found


def dump_row(rows, i, bnd, names):
    """整列逐欄輸出，附架構帶與欄名。"""
    print(f"\n-- r{i}")
    for j, v in enumerate(rows[i - 1]):
        if v in (None, ""):
            continue
        b = band_of(j, bnd) or "?"
        nm = names[j] if j < len(names) and names[j] else "?"
        print(f"   c{j}/{col(j)} [{b} 帶 · {nm}] = {v!r}")


def main():
    # T12a / T12b —— PROXI
    f = scan(PROXI, ["Gear_Box_Type", "VC_Trans_Equipped"], "T12a/T12b PROXI 全檔")
    wb = openpyxl.load_workbook(PROXI, read_only=True, data_only=True)
    rows = list(wb["Format"].iter_rows(values_only=True))
    hdr = [str(v) if v is not None else "" for v in rows[1]]  # Format 之欄名在 r2
    print(f"\n{'=' * 72}\nT12a 逐字 —— PROXI `Format` r443（Gear_Box_Type）")
    for j, v in enumerate(rows[442]):
        if v not in (None, ""):
            print(f"   c{j}/{col(j)} [{hdr[j] if j < len(hdr) else '?'}] = {v!r}")
    print("\n對照 —— PROXI `Format` r468（Country_Code，命名慣例之基準）")
    for j in (0, 1, 2, 3, 4, 5, 7):
        print(f"   c{j}/{col(j)} [{hdr[j]}] = {rows[467][j]!r}")
    wb.close()

    # T12c —— LID
    scan(LID, ["Gear_Box_Type"], "T12c 前置：LID 全分頁掃 Gear_Box_Type")
    wb = openpyxl.load_workbook(LID, read_only=True, data_only=True)
    rows = list(wb[SHEET].iter_rows(values_only=True))
    bnd = bands(rows[1])
    names = [str(v) if v is not None else "" for v in rows[2]]
    print(f"\n{'=' * 72}\nT12c —— LID `{SHEET}` 架構帶（自 r2 讀取）")
    for j in sorted(bnd):
        print(f"   c{j}/{col(j)} 起 = {bnd[j]!r}")
    print(f"\nT12c —— r420／r421 全欄逐字（另附 r43 Country_Code 為對照）")
    for i in (420, 421, 43):
        dump_row(rows, i, bnd, names)
    wb.close()


if __name__ == "__main__":
    main()
