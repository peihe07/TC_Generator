#!/usr/bin/env python3
"""T20c（下放包 14 §3.2）—— Lockout Table 之「非 L/O 側」行為是否為本 leaf 所有。

母體二：
  (a) 037 `Analysis Report` 之 `-013`~`-016`，**全 20 欄**
  (b) CFTS022 SYSRA `Basic Report` 之 `-120`／`-121`，**全欄**

搜非 `L/O` 側之表述（`not marked`／`remain available`／`no lockout` 之類）。

**載 → 屬本 leaf 所有，於同則加負向斷言（不另立 TC）。**
**未載 → 登 `COVERAGE_GAPS.md`，不造、不擴入（IN §8.4.2）。**

唯讀。
"""
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
A03 = ROOT / "inputs" / "DD_SWE1_0807_EN.xlsx"
CFTS = next(ROOT.joinpath("inputs").glob("CFTS022_Driver_Distraction*.xlsx"))

# 非 L/O 側之表述候選（大小寫不敏感；**刻意放寬**，寧可多命中再人讀）
PATS = [
    r"not\s+marked", r"unmarked", r"remain\w*\s+available", r"still\s+available",
    r"no\s+lock\s*out", r"no\s+lockout", r"not\s+locked", r"remain\w*\s+accessible",
    r"accessible", r"unrestricted", r"NOT_RESTRICTED", r"not\s+restricted",
    r"available", r"allow", r"except", r"other\s+features", r"non-?L/?O",
    r"features\s+not", r"shall\s+not\s+lock",
]


def sq(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def scan(label, rows, idx_desc):
    print("\n" + "=" * 78)
    print(label)
    print("=" * 78)
    total = 0
    for tag, row in rows:
        hits = []
        for j, v in enumerate(row):
            if v is None:
                continue
            s = str(v)
            for p in PATS:
                for m in re.finditer(p, s, re.I):
                    a, b = max(0, m.start() - 70), min(len(s), m.end() + 70)
                    hits.append((j, m.group(0), sq(s[a:b])))
        total += len(hits)
        print(f"\n-- {tag}：{len(hits)} 命中")
        seen = set()
        for j, tok, ctx in hits:
            k = (j, ctx)
            if k in seen:
                continue
            seen.add(k)
            print(f"   c{j} [{tok}] …{ctx[:150]}…")
    print(f"\n{label} 命中合計 = {total}")
    return total


def main():
    wb = openpyxl.load_workbook(A03, read_only=True, data_only=True)
    a = [list(r) for r in wb["Analysis Report"].iter_rows(values_only=True)]
    wb.close()
    rows_a = []
    for i, r in enumerate(a, 1):
        m = re.match(r"SWE1-RA-Driver_Distraction-(\d+)$", str(r[0] or ""))
        if m and m.group(1) in ("013", "014", "015", "016"):
            rows_a.append((f"037 r{i} `-{m.group(1)}`", r))
    print(f"母體 (a)：037 `Analysis Report` {len(rows_a)} 列 × 全 {len(a[8])} 欄")
    na = scan("(a) 037 `-013`~`-016` 全 20 欄", rows_a, 3)

    wb = openpyxl.load_workbook(CFTS, read_only=True, data_only=True)
    print("\nCFTS022 sheets:", wb.sheetnames)
    sh = "Basic Report" if "Basic Report" in wb.sheetnames else wb.sheetnames[0]
    c = [list(r) for r in wb[sh].iter_rows(values_only=True)]
    wb.close()
    rows_c = []
    for i, r in enumerate(c, 1):
        s = " ".join(str(x) for x in r[:3] if x is not None)
        m = re.search(r"SYS-RA-Driver_Distraction-(1(?:19|20|21))\b", s)
        if m:
            rows_c.append((f"CFTS022 `{sh}` r{i} `-{m.group(1)}`", r))
    print(f"\n母體 (b)：CFTS022 `{sh}` {len(rows_c)} 列 × 全 {max(len(r[1]) for r in rows_c) if rows_c else 0} 欄")
    for tag, r in rows_c:
        print(f"   {tag}: {sq(r[3] if len(r) > 3 else '')[:170]}")
    nc = scan("(b) CFTS022 `-119`／`-120`／`-121` 全欄", rows_c, 3)

    print("\n" + "=" * 78)
    print(f"T20c 合計：(a) {na} ／ (b) {nc}")
    print("**命中為機器之寬鬆掃描，是否構成『非 L/O 側之行為』須人讀判定。**")


if __name__ == "__main__":
    main()
