#!/usr/bin/env python3
"""T20d（下放包 14 §四）—— `-001`／`-002` 之激勵：窮盡內部來源。

三項**唯讀**查證：
  1. CFTS022 全表搜 `Body OFF`／`Body Off`／`sleep`／`wake`
  2. 037 全 28 leaf：`-001`／`-002` 以外之列是否於他處給該激勵之名
  3. `features/power/` 與 `features/power_moding/` 之 RULINGS.md／profile／
     feature.yaml：是否已有 Body OFF 電源時序之已裁綁定

**(3) 只回報「該線如何施加 Body OFF」，不作同一性判斷；
不得代改他線任何檔。** 本檔全程唯讀。
"""
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
REPO = ROOT.parent.parent
CFTS = next(ROOT.joinpath("inputs").glob("CFTS022_Driver_Distraction*.xlsx"))
A03 = ROOT / "inputs" / "DD_SWE1_0807_EN.xlsx"

KW = [r"body\s*off", r"\bsleep\b", r"\bwake", r"cold\s*start", r"power\s*sequence",
      r"power\s*mode", r"\bIGN\b", r"ignition"]
# 具名識別碼之形態：`$X$`／`CamelCase.Signal`／全大寫底線
ID_PATS = [r"\$[^$\s]+\$", r"\b[A-Z][A-Za-z0-9_]*\.[A-Za-z][A-Za-z0-9_]*\b",
           r"\b[A-Z][A-Z0-9_]{4,}\b"]


def sq(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def kw_hits(text):
    return sorted({m.group(0) for p in KW for m in re.finditer(p, text, re.I)})


def probe_xlsx(path, label, sheets=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = sheets or wb.sheetnames
    print("\n" + "=" * 78)
    print(label)
    print("=" * 78)
    print("分頁:", wb.sheetnames)
    tot = 0
    out = []
    for sn in names:
        for i, row in enumerate(wb[sn].iter_rows(values_only=True), 1):
            for j, v in enumerate(row):
                if v is None:
                    continue
                s = str(v)
                hs = kw_hits(s)
                if hs:
                    tot += 1
                    out.append((sn, i, j, hs, s))
    print(f"命中格數 = {tot}")
    for sn, i, j, hs, s in out[:25]:
        print(f"\n  [{sn}] r{i} c{j}  關鍵詞 {hs}")
        for m in re.finditer("|".join(KW), s, re.I):
            a, b = max(0, m.start() - 90), min(len(s), m.end() + 90)
            print(f"      …{sq(s[a:b])}…")
            break
        ids = sorted({x for p in ID_PATS for x in re.findall(p, s)})
        ids = [x for x in ids if x not in ("HU", "HMI", "SWE1", "SYS", "NOTE")]
        print(f"      該格之具名識別碼候選: {ids if ids else '（無）'}")
    if len(out) > 25:
        print(f"\n  …另 {len(out)-25} 格未列")
    wb.close()
    return out


def main():
    # ── (1) CFTS022 全表 ──────────────────────────────────────────
    o1 = probe_xlsx(CFTS, "(1) CFTS022 全表 —— 搜 Body OFF／sleep／wake 等")

    # ── (2) 037 全 28 leaf ────────────────────────────────────────
    wb = openpyxl.load_workbook(A03, read_only=True, data_only=True)
    rows = [list(r) for r in wb["Analysis Report"].iter_rows(values_only=True)]
    wb.close()
    print("\n" + "=" * 78)
    print("(2) 037 全 28 leaf —— `-001`／`-002` 以外之列是否給該激勵之名")
    print("=" * 78)
    n = 0
    for i, r in enumerate(rows, 1):
        m = re.match(r"SWE1-RA-Driver_Distraction-(\d+)$", str(r[0] or ""))
        if not m:
            continue
        lf = m.group(1)
        s = " ".join(str(c) for c in r if c is not None)
        hs = kw_hits(s)
        if hs:
            n += 1
            tag = "（本身即 -001／-002）" if lf in ("001", "002") else "**他列**"
            print(f"\n  -{lf} (r{i}) {tag}  關鍵詞 {hs}")
            ids = sorted({x for p in ID_PATS for x in re.findall(p, s)})
            ids = [x for x in ids if x not in ("HU", "HMI", "SWE1", "SYS", "NOTE", "RESTRICTED", "NOT_RESTRICTED")]
            print(f"      具名識別碼候選: {ids if ids else '（無）'}")
    print(f"\n  命中 leaf 數 = {n}")

    # ── (3) 他線之已裁綁定（唯讀，不判同一性）────────────────────
    print("\n" + "=" * 78)
    print("(3) `features/power/`／`features/power_moding/` —— **只回報，不判同一性**")
    print("=" * 78)
    for slug in ("power", "power_moding"):
        d = REPO / "features" / slug
        print(f"\n-- {slug}: {'存在' if d.exists() else '**不存在**'}")
        if not d.exists():
            continue
        for fn in ("RULINGS.md", "feature.yaml", "framework.md"):
            f = d / fn
            if not f.exists():
                print(f"   {fn}: 不存在")
                continue
            txt = f.read_text("utf-8", errors="replace")
            hs = []
            for p in KW:
                for m in re.finditer(p, txt, re.I):
                    a, b = max(0, m.start() - 110), min(len(txt), m.end() + 110)
                    hs.append((m.group(0), sq(txt[a:b])))
            print(f"   {fn}: {len(hs)} 命中")
            seen = set()
            for tok, ctx in hs:
                if ctx in seen:
                    continue
                seen.add(ctx)
                print(f"      [{tok}] …{ctx[:170]}…")
                if len(seen) >= 8:
                    print(f"      …另 {len(hs)-8} 命中未列")
                    break
        prof = REPO / "docs/runtime/profiles"
        for f in sorted(prof.glob("*.md")):
            if slug.split("_")[0] in f.name.lower() or slug in f.name.lower():
                txt = f.read_text("utf-8", errors="replace")
                hs = kw_hits(txt)
                print(f"   profile `{f.name}`: 關鍵詞 {hs if hs else '（無命中）'}")


if __name__ == "__main__":
    main()
