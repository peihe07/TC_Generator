#!/usr/bin/env python3
"""ICS 寫回 —— 依 `Requirement or Design ID` 升序排列 31 條並重編列位號。

**本檔補兩個洞。**

一、**b19b 未留寫回腳本。** v1 工作簿（`sandbox/v1/ics_management_v1.xlsx`）
    產出時未落任何程式於 repo，其產出不可再現 —— 一個交付件若只能由
    當時那次對話重建，它的 sha 保護的是一份沒有人能再做一次的東西。

二、**列序錯。** v1 之列序取 `generated/` 之批次序（b01→b07，批內依 json
    陣列序），而 D 欄（`req_id`）因此有 6 個逆序斷點。comfort 96 §1 一
    早已立「列序依 `Requirement or Design ID` 遞增，**不依批次順序、
    不依 Test Set 順序**」，並以 `row-order-by-reqid` 一道 gate 守之
    （`features/comfort/scripts/verify_row_order_gates.py`）。
    ICS 因無寫回腳本，該 gate 從未對其跑過。
    power 之交付件實測 389 列 0 斷點，其不變量與 comfort 之條文一致。

**內容不重新投影。** 31 條之 13 欄逐格自 v1 讀出後原樣搬位，只有 B（序號）
與 F（tc_id）依新列位重編。故本檔之產出與 v1 之差別**只有列序與該二欄**，
且此事由 `g_content_preserved` 以多重集比對證明，不是宣稱。

排序鍵為 `leaf_sort_key`，非字串序 —— 與 comfort 同型，使 `SWE-ICS-9` 與
`SWE-ICS-10` 之序由數值決定。本語料之 id 皆三位補零，兩者結果相同，
但依賴補零是依賴一個沒有人保證過的事。

Usage:
    python3 features/ics_management/scripts/write_back.py            # dry run
    python3 features/ics_management/scripts/write_back.py --write    # 產出 v2
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

FEATURE = Path(__file__).resolve().parents[1]
ROOT = FEATURE.parents[1]
sys.path.insert(0, str(ROOT))

from backend.xlsx_surgical import surgical_save  # noqa: E402

MASTER = ROOT / "forms" / (
    "FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT STLA Test Case "
    "Specification & Result_SWQT_20260817_ext.xlsx"
)
V1 = FEATURE / "sandbox" / "v1" / "ics_management_v1.xlsx"
SHEET = "Test Case Specification 測試用例規範"
FIRST_ROW = 10

# v1 實測有值之 13 欄（`scripts/write_back.py` 落檔前以 Counter 普查）。
COLS = ["B", "D", "F", "G", "H", "I", "J", "K", "L", "M", "N", "P", "R"]
COL_NO, COL_REQ, COL_TC = "B", "D", "F"

TC_ID_FMT = "NR1L-ICS-{n:03d}"          # R-ICS56(a)
LEAF_KEY = re.compile(r"SWE-ICS-(\d+)$")


def leaf_sort_key(req_id: str) -> tuple:
    """037 之 leaf 序。數值比較，不依賴補零。"""
    m = LEAF_KEY.match(req_id.strip())
    if not m:
        raise ValueError(f"req_id 不合 SWE-ICS-<n> 之形: {req_id!r}")
    return (int(m.group(1)),)


def leaf_universe() -> list:
    """`generated/` 所載之 leaf 全集，依 id 遞增。

    ICS 之 037 實測恰為 SWE-ICS-001..010 十個，且十個皆有 TC ——
    故無 comfort 96 §1 二 之留空列。**該前提由 `g_all_leaves` 現場複驗**，
    不以本註解為據。
    """
    ids = {t["req_id"] for t in load_tcs_json()}
    return sorted(ids, key=leaf_sort_key)


def load_tcs_json() -> list:
    """`generated/b01..b07` 之 31 條，依批次序（即 v1 之列序）。"""
    tcs = []
    for p in sorted((FEATURE / "generated").glob("b*/b*_tcs.json")):
        doc = json.loads(p.read_text(encoding="utf-8"))
        tcs.extend(doc["tcs"] if isinstance(doc, dict) else doc)
    return tcs


def read_v1_rows() -> list:
    """自 v1 讀出 31 列之 13 欄。回傳 [(原列序, {欄: 值})]。"""
    wb = openpyxl.load_workbook(V1, data_only=True)
    ws = wb[SHEET]
    rows = []
    r = FIRST_ROW
    while ws[f"{COL_REQ}{r}"].value:
        rows.append((r - FIRST_ROW, {c: ws[f"{c}{r}"].value for c in COLS}))
        r += 1
    wb.close()
    return rows


def row_plan(rows: list) -> list:
    """列序：D 欄升序；同一 leaf 內維持其原相對序（穩定排序）。

    同一 leaf 拆出多條 TC 時該幾列 D 欄相同 —— 那是規則所要之形態，
    故排序鍵不含任何用以打破同值之欄，只以原位置穩定之。
    """
    plan = sorted(rows, key=lambda it: (leaf_sort_key(it[1][COL_REQ]), it[0]))
    for i, (_, cells) in enumerate(plan, start=1):
        cells[COL_NO] = str(i)
        cells[COL_TC] = TC_ID_FMT.format(n=i)
    return plan


# ----------------------------------------------------------------- 各道 gate
def g_row_order(col_d: list) -> str | None:
    """首個**倒退**之列；無則 None。判準為 `<` 而非 `<=`（同 comfort）。"""
    for i in range(1, len(col_d)):
        if leaf_sort_key(col_d[i]) < leaf_sort_key(col_d[i - 1]):
            return f"row{FIRST_ROW + i}: {col_d[i]} 在 {col_d[i - 1]} 之後"
    return None


def g_all_leaves(col_d: list, universe: list) -> list:
    return sorted(set(universe) - set(col_d), key=leaf_sort_key)


def g_tc_sequence(col_f: list) -> str | None:
    want = [TC_ID_FMT.format(n=i) for i in range(1, len(col_f) + 1)]
    return None if col_f == want else f"首個不符: 列位 {col_f.index(next(a for a, b in zip(col_f, want) if a != b)) + 1}"


def g_content_preserved(before: list, after: list) -> list:
    """除 B／F 外之 11 欄，其列之多重集須前後相同 —— 重排不是編輯。"""
    def key(cells):
        return tuple(cells[c] for c in COLS if c not in (COL_NO, COL_TC))
    lost = sorted(set(map(key, [c for _, c in before])) - set(map(key, [c for _, c in after])))
    return [k[0] for k in lost]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true", help="產出 sandbox/v2/")
    ap.add_argument("--out", help="輸出路徑（預設 sandbox/v2/ics_management_v2.xlsx）")
    args = ap.parse_args()

    rows = read_v1_rows()
    plan = row_plan([(i, dict(c)) for i, c in rows])
    col_d = [c[COL_REQ] for _, c in plan]
    col_f = [c[COL_TC] for _, c in plan]

    fails = []
    print(f"# ICS 寫回 —— 列序依 {COL_REQ} 欄（Requirement or Design ID）升序\n")
    print(f"自 v1 讀出 {len(rows)} 列 × {len(COLS)} 欄；重排後 {len(plan)} 列\n")

    print("| 列位 | tc_id | req_id | Test Set |")
    print("|---|---|---|---|")
    for i, (_, c) in enumerate(plan, start=1):
        print(f"| {i} | {c[COL_TC]} | {c[COL_REQ]} | {c['H']} |")

    print("\n## Gates\n")
    bad = g_row_order(col_d)
    print(f"- row-order-by-reqid: {'PASS' if not bad else 'FAIL — ' + bad}")
    fails += [] if not bad else [bad]

    missing = g_all_leaves(col_d, leaf_universe())
    print(f"- all-leaves-present: {'PASS' if not missing else 'FAIL — 缺 ' + ', '.join(missing)}")
    fails += [] if not missing else missing

    seq = g_tc_sequence(col_f)
    print(f"- tc-id-sequence: {'PASS' if not seq else 'FAIL — ' + seq}")
    fails += [] if not seq else [seq]

    lost = g_content_preserved(rows, plan)
    print(f"- content-preserved（11 欄多重集）: {'PASS' if not lost else 'FAIL — 遺失 ' + ', '.join(lost)}")
    fails += lost

    if fails:
        print("\n**有 gate 未過，不產出。**")
        return 1

    if not args.write:
        print("\nDRY RUN —— 未寫出任何檔。加 --write 產出 v2。")
        return 0

    out = Path(args.out) if args.out else FEATURE / "sandbox" / "v2" / "ics_management_v2.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(MASTER)
    ws = wb[SHEET]
    for i, (_, cells) in enumerate(plan):
        for c in COLS:
            ws[f"{c}{FIRST_ROW + i}"] = cells[c]
    report = surgical_save(wb, MASTER, out)      # StructureError 即中止
    print(f"\n產出: {out}")
    print(f"surgical_save: {report}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
