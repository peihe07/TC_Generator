#!/usr/bin/env python3
"""覆蓋缺口實測（下放包 05 作業 D）。唯讀，不寫任何檔、不改任何 TC JSON。

本腳本產生 `docs/reports/05_coverage_gaps.md` 之全部數字。三個獨立量測：

  A. `1.8.1.3 Button Press Events` 逐物件三軸實值與 R-ICS2 v2(b) 判定
     —— 直接呼叫 `cfts020_probe.py` 之 `parse()`（其內建 v2 判定），
     不重寫判準、不抄 `03_cfts020_recon_v2.md`。
  B. 全文檢索 `Short Press`／`Long Press`／`<Tpress>` 於 **判適用** 之物件中
     是否另有母條（用以證「ICS 側無母條」為實測而非推論）。
  C. RD 覆蓋矩陣：SWRA `SWE1 Requirements` 之 ID 全集
     × `generated/b01`～`b04` 之 `req_id` 計數，
     並列 `SYS2 Traceability` 之 SWE1 ID 全集以顯示 011／012 之缺列。
  D. TC 中「未定義按壓時長之裸按壓步驟」計數（A-ICS33 之受影響面）。

掃描條件（逐項揭露）：
  - 物件母數、屬性抓取、軸值比對一律沿用 `cfts020_probe.py` 之定義
    （區分大小寫之精確字串集合交集；軸不存在記 `None`）。
  - 關鍵詞比對：`Short Press` / `Long Press` / `<Tpress>` 為
    **區分大小寫之子字串**比對（不作詞界正規化，不去連字號）；
    另以 `casefold()` 之不分大小寫掃描並列，二者數字皆報。
  - RD ID 取 `SWE1 Requirements` 之 A 欄（表頭列為第 7 列，資料自第 8 列起），
    `strip()` 後非空且符合 `^SWE-ICS-\\d{3}$` 者入集合。
  - `req_id` 取各 `b0N_tcs.json` 之 `tcs[].req_id`，逐字比對（不正規化）。
  - 裸按壓步驟：test_procedure 之行內含 `Press the ICS` 且該行不含
    `and hold` / ` for ` —— 二條件皆為區分大小寫之子字串比對。

用法：
  python3 features/ics_management/scripts/gap_probe_05.py
  python3 features/ics_management/scripts/gap_probe_05.py --json
"""
from __future__ import annotations

import argparse
import importlib.util
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SWRA = ROOT / "inputs/ICS_Management_FM-WI-FSM-037-A03_STLA_Report_SWRA.xlsx"
RD_RE = re.compile(r"^SWE-ICS-\d{3}$")
SWE1_RE = re.compile(r"^SWE1-ICS-\d{3}$")
KEYWORDS = ("Short Press", "Long Press", "<Tpress>")


def load_probe():
    """以檔案路徑載入既有之 `cfts020_probe.py`（不修改該檔）。"""
    spec = importlib.util.spec_from_file_location(
        "cfts020_probe", Path(__file__).with_name("cfts020_probe.py")
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def probe_objects(probe) -> list[dict]:
    """取全部 2180 物件並附 v2 判定（`parse()` 已內含 v2 判定）。"""
    return probe.parse()


def section_1813(objs: list[dict]) -> list[dict]:
    return [o for o in objs if o["section_no"] == "1.8.1.3"
            or o["section_no"].startswith("1.8.1.3.")]


def ecu_class(o: dict) -> str:
    """A-ICS33 之成因分類（實測，非沿用 R-ICS23(a) 之字面）。"""
    ecu = o["ecu"]
    if ecu is None:
        return "ECU 軸缺"
    if "FPDM" in ecu:
        return "ECU 含 FPDM"
    return "ECU=" + "/".join(ecu)


def keyword_scan(objs: list[dict]) -> dict:
    res = {}
    for kw in KEYWORDS:
        cs = [o for o in objs if kw in o["text"]]
        ci = [o for o in objs if kw.casefold() in o["text"].casefold()]
        res[kw] = {
            "區分大小寫_命中物件數": len(cs),
            "區分大小寫_其中判適用": [o["id"] for o in cs if o["verdict"] == "適用"],
            "不分大小寫_命中物件數": len(ci),
            "不分大小寫_其中判適用": [o["id"] for o in ci if o["verdict"] == "適用"],
            "區分大小寫_命中之章節": sorted({o["section_no"] for o in cs}),
        }
    return res


def swra_ids() -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(SWRA, read_only=True, data_only=True)
    ws = wb["SWE1 Requirements"]
    reqs = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        rid = (row[0] or "").strip() if isinstance(row[0], str) else None
        if rid and RD_RE.match(rid):
            reqs.append({"rd": rid, "src": row[1], "title": row[2],
                         "vc": row[16], "vm": row[17]})
    ws2 = wb["SYS2 Traceability"]
    trace = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        sid = (row[0] or "").strip() if isinstance(row[0], str) else None
        if sid and SWE1_RE.match(sid):
            trace.append({"swe1": sid, "sysra": row[2], "title": row[4]})
    return {"requirements": reqs, "traceability": trace}


def tc_rows() -> list[dict]:
    rows = []
    for f in sorted((ROOT / "generated").glob("b0*/b0*_tcs.json")):
        d = json.loads(f.read_text())
        for t in d["tcs"]:
            rows.append({"batch": f.parent.name, "req_id": t["req_id"],
                         "tc_title": t["tc_title"], "test_set": t["test_set"],
                         "anchor": t["specification_reference"],
                         "design_method": t["design_method"],
                         "has_pending": t["has_pending"],
                         "procedure": t["test_procedure"]})
    return rows


def bare_presses(rows: list[dict]) -> list[dict]:
    out = []
    for t in rows:
        for line in t["procedure"].split("\n"):
            if "Press the ICS" in line and "and hold" not in line and " for " not in line:
                out.append({"req_id": t["req_id"], "tc_title": t["tc_title"],
                            "step": line.strip()})
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()

    probe = load_probe()
    objs = probe_objects(probe)
    sec = section_1813(objs)
    swra = swra_ids()
    rows = tc_rows()

    counts: dict[str, int] = {}
    for t in rows:
        counts[t["req_id"]] = counts.get(t["req_id"], 0) + 1

    cls: dict[str, int] = {}
    for o in sec:
        if o["verdict"] == "不適用":
            k = ecu_class(o)
            cls[k] = cls.get(k, 0) + 1

    result = {
        "物件母數": len(objs),
        "1.8.1.3 物件數": len(sec),
        "1.8.1.3 判定分佈": {
            "適用": sum(1 for o in sec if o["verdict"] == "適用"),
            "不適用": sum(1 for o in sec if o["verdict"] == "不適用"),
        },
        "1.8.1.3 不適用之成因分類": cls,
        "1.8.1.3 逐物件": [
            {"id": o["id"], "section_no": o["section_no"],
             "artifact_type": o["artifact_type"], "ecu": o["ecu"],
             "radio": o["radio"], "ee": o["ee"], "verdict": o["verdict"],
             "reasons": o["reasons"]} for o in sec],
        "關鍵詞全文掃描": keyword_scan(objs),
        "SWRA SWE1 Requirements ID 全集": [r["rd"] for r in swra["requirements"]],
        "SYS2 Traceability SWE1 ID 全集": [r["swe1"] for r in swra["traceability"]],
        "TC 總數": len(rows),
        "每 RD 之 TC 數": counts,
        "無 TC 之 RD": sorted(
            {r["rd"] for r in swra["requirements"]} - set(counts)),
        "Verification Criteria": {r["rd"]: r["vc"] for r in swra["requirements"]},
        "Verification Method": {r["rd"]: r["vm"] for r in swra["requirements"]},
        "裸按壓步驟": bare_presses(rows),
    }

    if args.json:
        json.dump(result, sys.stdout, ensure_ascii=False, indent=1)
        print()
        return

    for k, v in result.items():
        print(f"== {k}")
        print(json.dumps(v, ensure_ascii=False, indent=1))
        print()


if __name__ == "__main__":
    main()
