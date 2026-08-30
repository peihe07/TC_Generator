#!/usr/bin/env python3
"""下放包 12 作業 C 之量測腳本（唯讀；不改任何既有檔）。

C-1  CFTS022 全文族層之 `{CFTS020}` 外引點清單、所指章節、
     落在 CFTS020 §1.8／§1.18 者之數、七個現用錨之間接影響檢定。
C-2  第三種變體表達（CAN node 在否 ＋ `$Head_Unit_Screen_Size$`）
     於 CFTS022／CFTS020 之全部出現，及代入本 DUT 之判定。

抽取條件（逐項揭露，與報告 §0 一致）：
  - `.docx` → 讀 `word/document.xml`；`</w:p>`→換行、`</w:tc>`→tab、
    `re.sub(r"<[^>]+>","")` 去標籤、`html.unescape`
  - **空白正規化**：`U+00A0`(NBSP)、`U+2007`、`U+202F`、`U+2009`、
    `U+2002`、`U+2003`、`U+2005`、`U+2006`、`U+2008`、`U+205F`、`U+3000`
    一律換為 `U+0020`；`U+200B`(ZWSP)／`U+FEFF` 刪除。
    **另備一份「折疊連續空白」之比對面**（`re.sub(r"[ \t]+"," ")`），
    因二檔皆大量使用 NBSP 且實測存在**雙 NBSP**，不折疊會漏命中。
    原始行（未折疊）保留供逐字引用。
  - 章節行：`^(\d+(?:\.\d+)*) (.+?) \{(\d{7})\}$` 且該行不含 `PAGEREF`
  - 物件屬性頭：`^(\d{7}): \[`（對 `line.strip()`）；本文取次一行
  - 屬性：`\[([^:\]]+):([^\]]*)\]`，同 key 取首見
  - `.xlsx`：`openpyxl`，`read_only=True`、`data_only=True`
  - `.dbc`：**一律 `latin-1` 開檔**

用法：
  python3 features/ics_management/scripts/crossref_probe_12.py c1
  python3 features/ics_management/scripts/crossref_probe_12.py c2
  python3 features/ics_management/scripts/crossref_probe_12.py lid
"""
from __future__ import annotations

import html
import re
import sys
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
REPO = ROOT.parents[1]

DOC022 = ROOT / ("inputs/R1LR_Atl-H_26PI2.5 Jun Release-Privacy_CFTS_022 "
                 "Functional Specification_20260608-1205.docx")
DOC020 = ROOT / ("inputs/R1LR_Atl-H_26PI1.5 Mar Release-Cabin_CFTS_020 ICS and "
                 "DCSD _20260310-1533.docx")
LID = REPO / "forms/Logical Identifiers and CAN Mapping v1_78.xlsx"
DBCS = [REPO / "forms/PDT27_E2A_R1_BHCAN2.dbc",
        REPO / "forms/PDT27_E2A_R1_FDCAN8.dbc",
        REPO / "features/vehicle_setting/inputs/PDT27_E2A_R4_BHCAN.dbc",
        REPO / "features/vehicle_setting/inputs/PDT27_E2A_R5_FDCAN8.dbc"]

SPACE_CHARS = "          　"
DROP_CHARS = "​﻿"
SEC_RE = re.compile(r"^(\d+(?:\.\d+)*) (.+?) \{(\d{7})\}\s*$")
OBJ_RE = re.compile(r"^(\d{7}): \[")
ATTR_RE = re.compile(r"\[([^:\]]+):([^\]]*)\]")

ANCHORS = ["4914956", "4914957", "4914958", "4914974",
           "4914975", "4914976", "4914993"]


def norm(s: str) -> str:
    """空白正規化（不折疊）。"""
    for c in SPACE_CHARS:
        s = s.replace(c, " ")
    for c in DROP_CHARS:
        s = s.replace(c, "")
    return s


def fold(s: str) -> str:
    """折疊連續空白／tab（供比對用）。"""
    return re.sub(r"[ \t]+", " ", s)


def doc_lines(path: Path) -> list[str]:
    xml = zipfile.ZipFile(path).read("word/document.xml").decode("utf8")
    xml = re.sub(r"</w:p>", "\n", xml)
    xml = re.sub(r"</w:tc>", "\t", xml)
    txt = html.unescape(re.sub(r"<[^>]+>", "", xml))
    return [norm(x) for x in txt.split("\n")]


def parse(path: Path) -> tuple[list[str], list[dict], list[dict]]:
    """回傳 (lines, objects, sections)。"""
    lines = doc_lines(path)
    objs, secs = [], []
    cur = None
    for i, raw in enumerate(lines):
        s = raw.strip()
        m = SEC_RE.match(s)
        if m and "PAGEREF" not in s:
            cur = {"num": m.group(1), "title": m.group(2),
                   "oid": m.group(3), "line": i + 1}
            secs.append(cur)
            continue
        mo = OBJ_RE.match(s)
        if mo:
            attrs = {}
            for k, v in ATTR_RE.findall(s):
                attrs.setdefault(k.strip(), v.strip())
            body = lines[i + 1].strip() if i + 1 < len(lines) else ""
            objs.append({"oid": mo.group(1), "line": i + 1, "attrs": attrs,
                         "body": body, "sec": cur["num"] if cur else None,
                         "sec_title": cur["title"] if cur else None,
                         "sec_oid": cur["oid"] if cur else None})
    return lines, objs, secs


# ---------------------------------------------------------------- C-1

# 外引比對式（施於**折疊後**之文字，`re.I`；逐式獨立計數後取聯集）
XREF_PATTERNS = {
    "brace_plain":  r"\{ ?CFTS ?0?20 ?\}",                 # {CFTS020}
    "brace_id":     r"\{ ?CFTS ?0?20 ?- ?\d{5,8} ?\}",     # {CFTS020-nnnnnnn}
    "bare":         r"CFTS ?0?20",                          # 裸 CFTS020（含上二式）
    "title_quoted": r"['‘’\"“”] ?ICS and DCSD ?"
                    r"['‘’\"“”]",       # 'ICS and DCSD'
    "title_bare":   r"ICS and DCSD",                        # 章名裸寫
}
XREF_ANY = re.compile("|".join(f"(?:{p})" for p in XREF_PATTERNS.values()), re.I)


def c1() -> None:
    lines, objs, secs = parse(DOC022)
    print(f"[CFTS022] lines={len(lines)} objects={len(objs)} sections={len(secs)}")

    # 逐式命中行數（全文行面，含目次）
    print("\n-- 逐式命中（行面，折疊後，re.I）--")
    for name, pat in XREF_PATTERNS.items():
        n = sum(1 for l in lines if re.search(pat, fold(l), re.I))
        print(f"  {name:14s} {pat:55s} {n}")
    n_any = sum(1 for l in lines if XREF_ANY.search(fold(l)))
    print(f"  {'UNION':14s} {'':55s} {n_any}")

    # 物件面外引點
    print("\n-- 外引點（物件面：屬性頭行 + 本文行）--")
    pts = []
    for o in objs:
        hay = fold(o["body"])
        hits = XREF_ANY.findall(hay)
        if not hits:
            continue
        # 判定所指之 CFTS020 章節：本文中若有 {CFTS020-nnn} 或 §號
        target = []
        for m in re.finditer(r"\{ ?CFTS ?0?20 ?- ?(\d{5,8}) ?\}", hay, re.I):
            target.append("OID:" + m.group(1))
        for m in re.finditer(r"CFTS ?0?20[^.]{0,40}?§ ?(\d+(?:\.\d+)*)", hay, re.I):
            target.append("SEC:" + m.group(1))
        for m in re.finditer(r"[Ss]ection ?(\d+(?:\.\d+)+)", hay):
            target.append("SEC:" + m.group(1))
        pts.append({"o": o, "target": target})
        print(f"\n  OID {o['oid']}  L{o['line']}  §{o['sec']} {o['sec_title']}"
              f"  [{o['attrs'].get('Artifact Type','-')}]")
        print(f"    ECU={o['attrs'].get('ECU')} | Radio={o['attrs'].get('Radio')}"
              f" | EE={o['attrs'].get('EE Architecture')}")
        print(f"    target={target or '未明示'}")
        print(f"    BODY: {o['body']}")

    # 章節標題行本身之外引
    print("\n-- 外引點（章節標題行）--")
    for s in secs:
        if XREF_ANY.search(fold(s["title"])):
            print(f"  §{s['num']} {s['title']} {{{s['oid']}}} L{s['line']}")

    # 非物件、非章節之散行（表格列等）
    print("\n-- 外引命中行中，未歸入任何物件本文者 --")
    body_lines = {o["line"] + 1 for o in objs}
    obj_lines = {o["line"] for o in objs}
    sec_lines = {s["line"] for s in secs}
    for i, l in enumerate(lines, 1):
        if not XREF_ANY.search(fold(l)):
            continue
        if i in body_lines or i in obj_lines or i in sec_lines:
            continue
        tag = "TOC" if "PAGEREF" in l else "OTHER"
        print(f"  L{i} [{tag}] {l.strip()[:200]}")

    print(f"\n外引物件數 = {len(pts)}")

    # 七錨之同節其他物件
    print("\n-- 七錨所在章節內之全部物件與其外引狀態 --")
    anchor_secs = sorted({o["sec"] for o in objs if o["oid"] in ANCHORS},
                         key=lambda x: [int(y) for y in x.split(".")])
    print(f"  錨所在節：{anchor_secs}")
    for sn in anchor_secs:
        # 同節＝節號相同，或為其子節（前綴）
        members = [o for o in objs if o["sec"] and
                   (o["sec"] == sn or o["sec"].startswith(sn + "."))]
        print(f"\n  === §{sn}（含子節）物件 {len(members)} 個 ===")
        for o in members:
            x = "XREF" if XREF_ANY.search(fold(o["body"])) else "    "
            mark = "*" if o["oid"] in ANCHORS else " "
            print(f"   {mark}{x} {o['oid']} L{o['line']} §{o['sec']}"
                  f" [{o['attrs'].get('Artifact Type','-')}]"
                  f" Radio={(o['attrs'].get('Radio') or '')[:40]}"
                  f" EE={(o['attrs'].get('EE Architecture') or '')[:40]}")
            print(f"        {o['body'][:220]}")

    # 錨本文自身之外引複驗
    print("\n-- 七錨本文自身之外引複驗 --")
    for a in ANCHORS:
        o = next(x for x in objs if x["oid"] == a)
        print(f"  {a}: {'有外引' if XREF_ANY.search(fold(o['body'])) else '無外引'}"
              f"  §{o['sec']}")

    # 全文所有 {XXXnnn} 型引用之種類統計（看還有沒有別的外引目標）
    print("\n-- 全文 {..} 引用 token 統計（非 7 位數 ObjectID 者）--")
    from collections import Counter
    c = Counter()
    for l in lines:
        if "PAGEREF" in l:
            continue
        for m in re.finditer(r"\{([^{}]{1,40})\}", fold(l)):
            t = m.group(1).strip()
            if re.fullmatch(r"\d{7}", t):
                continue
            c[t] += 1
    for t, n in c.most_common(60):
        print(f"  {n:4d}  {t}")


def c1_sections020() -> None:
    """列 CFTS020 之頂層節與 §1.8／§1.18 之 ObjectID 範圍，供 target 對映。"""
    lines, objs, secs = parse(DOC020)
    print(f"[CFTS020] lines={len(lines)} objects={len(objs)} sections={len(secs)}")
    print("\n-- 頂層 §1.x 節 --")
    for s in secs:
        if re.fullmatch(r"1\.\d+", s["num"]):
            print(f"  §{s['num']:6s} {{{s['oid']}}} L{s['line']:5d} {s['title']}")
    for tgt in ("1.8", "1.18"):
        mem = [o for o in objs if o["sec"] and
               (o["sec"] == tgt or o["sec"].startswith(tgt + "."))]
        ids = sorted(int(o["oid"]) for o in mem)
        print(f"\n  §{tgt} 子樹物件 {len(mem)} 個，OID {ids[0]}~{ids[-1]}"
              if mem else f"\n  §{tgt} 無物件")


# ---------------------------------------------------------------- C-2

C2_TERMS = {
    "CAN node":              r"CAN ?node",
    "ICS_R":                 r"ICS_R\b",
    "Head_Unit_Screen_Size": r"Head_Unit_Screen_Size",
    "node 119":              r"node ?119",
    "node 94":               r"node ?94",
    "Screen_Size":           r"Screen_Size",
    "are present":           r"(?:is|are) present",
    "ICS_R (無詞界)":        r"ICS_R",
}


def c2() -> None:
    for tag, path in (("CFTS022", DOC022), ("CFTS020", DOC020)):
        lines, objs, secs = parse(path)
        print(f"\n================ {tag} "
              f"(lines={len(lines)} objects={len(objs)}) ================")
        print("-- 詞面命中行數（折疊後，re.I）--")
        for name, pat in C2_TERMS.items():
            n = sum(1 for l in lines if re.search(pat, fold(l), re.I))
            print(f"  {name:24s} {n}")

        print("\n-- 逐命中行（任一詞）--")
        big = re.compile("|".join(f"(?:{p})" for p in
                                  [C2_TERMS["CAN node"], C2_TERMS["ICS_R (無詞界)"],
                                   C2_TERMS["Head_Unit_Screen_Size"]]), re.I)
        by_line = {o["line"] + 1: o for o in objs}
        for i, l in enumerate(lines, 1):
            if not big.search(fold(l)):
                continue
            o = by_line.get(i)
            if o:
                print(f"\n  L{i}  OID {o['oid']}  §{o['sec']} {o['sec_title']}"
                      f"  [{o['attrs'].get('Artifact Type','-')}]")
                print(f"    ECU={o['attrs'].get('ECU')}")
                print(f"    Radio={o['attrs'].get('Radio')}")
                print(f"    EE={o['attrs'].get('EE Architecture')}")
                print(f"    Market={o['attrs'].get('Market')}")
                print(f"    BODY: {l.strip()}")
            else:
                print(f"\n  L{i}  (非物件本文行) {l.strip()[:300]}")

        # 所有 CAN node 編號之出現
        print("\n-- 全文出現之 CAN node 編號 --")
        from collections import Counter
        c = Counter()
        for l in lines:
            for m in re.finditer(r"CAN ?node ?(\d+)", fold(l), re.I):
                c[int(m.group(1))] += 1
        print("  ", dict(sorted(c.items())) or "無")

        # $..$ 訊號 token 中含 Screen 者
        print("\n-- $..$ token 含 Screen 者 --")
        c2t = Counter()
        for l in lines:
            for m in re.finditer(r"\$([^$]{1,60})\$", fold(l)):
                if "screen" in m.group(1).lower():
                    c2t[m.group(1)] += 1
        for t, n in c2t.most_common():
            print(f"   {n:4d}  ${t}$")


def lid() -> None:
    """LID 表：查 CAN node 編號 / ICS_R / Head_Unit_Screen_Size。"""
    from openpyxl import load_workbook
    wb = load_workbook(LID, read_only=True, data_only=True)
    print("sheets:", wb.sheetnames)
    pats = [("ICS_R", re.compile(r"ICS_R")),
            ("Head_Unit_Screen_Size", re.compile(r"Head_Unit_Screen_Size", re.I)),
            ("119", re.compile(r"\b119\b")),
            ("node", re.compile(r"\bnode\b", re.I)),
            ("DCSD", re.compile(r"DCSD"))]
    for ws in wb.worksheets:
        rows = [["" if c is None else str(c) for c in r]
                for r in ws.iter_rows(values_only=True)]
        for name, p in pats:
            hits = [(i + 1, r) for i, r in enumerate(rows)
                    if any(p.search(c) for c in r)]
            if not hits:
                continue
            print(f"\n=== [{ws.title}] {name}: {len(hits)} 列 ===")
            for i, r in hits[:40]:
                cells = [c for c in r if c.strip()]
                print(f"  R{i}: {' | '.join(cells)[:400]}")
            if len(hits) > 40:
                print(f"  … 另 {len(hits)-40} 列")
    wb.close()


def dbc() -> None:
    for p in DBCS:
        if not p.exists():
            print(f"\n### {p} — 不存在")
            continue
        txt = p.read_text(encoding="latin-1")
        bu = re.search(r"^BU_:(.*)$", txt, re.M)
        nodes = bu.group(1).split() if bu else []
        print(f"\n### {p.name}  BU_ {len(nodes)}: {' '.join(nodes)}")
        for kw in ("ICS_R", "DCSD", "Screen_Size", "ICS"):
            n = len(re.findall(kw, txt))
            print(f"    '{kw}' 出現 {n} 次")


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "c1"
    {"c1": c1, "c1s": c1_sections020, "c2": c2,
     "lid": lid, "dbc": dbc}[cmd]()
