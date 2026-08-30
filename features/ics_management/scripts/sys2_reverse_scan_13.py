#!/usr/bin/env python3
"""下放包 13 作業 B：以 SYS2 為起點之反向掃描（只量不裁）。

沿用 `sys2_87_probe_12.py` 之讀法（同一檔、同一 sheet、同一表頭列、同一欄名），
以確保與前十二包可比。本腳本不作任何「應否納入驗證範圍」之判斷。

模式：
  --sec1     §1 來源欄空白之列逐一（含四欄逐字）
  --sec2     §2 來源欄非空但不指向 CFTS020／CFTS022 之列逐一 + 總盤點
  --census   全表來源桶盤點
"""
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
REPO = ROOT.parents[1]
SYS2 = ROOT / ("inputs/SYS2_CFTS_020_DISP_TCH_ICS_20260616_All_HW_System_"
               "Accepted & Released.xlsx")
SOURCES_DIR = REPO / "spec-index" / "sources"

SHEET = "Basic Report"
HEADER_ROW = 1
COL_ID = "ID"
COL_DESC = "Description"
COL_SRC = "SYS2 來源需求項目ID  Source Requirement items"
COL_CAT = "SYS2 分類 Category"
COL_SUB = "SYS2 子分類 Sub Category Function Name"
COL_SWHW = ("SYS2 SW/HW/System (如果是HW+SW，就選System) "
            "( software, hardware, or system (both software and hardware).)")
COL_MD = "SYS2 MD Feedback"
COL_DOC = "SYS2 文件識別碼 Document ID"
WANTED = [COL_ID, COL_DESC, COL_SRC, COL_CAT, COL_SUB, COL_SWHW, COL_MD, COL_DOC]

ID_RE = re.compile(r"\d{7}")


def norm(s) -> str:
    """正規化：NBSP→space、其他 Unicode 空白→space、collapse 連續空白、strip。"""
    if s is None:
        return ""
    t = str(s).replace(" ", " ")
    t = "".join(" " if unicodedata.category(ch) == "Zs" else ch for ch in t)
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def rows() -> list[dict]:
    wb = openpyxl.load_workbook(SYS2, data_only=True, read_only=True)
    ws = wb[SHEET]
    raw = list(ws.iter_rows(values_only=True))
    header = [str(c) if c is not None else "" for c in raw[HEADER_ROW - 1]]
    for w in WANTED:
        if w not in header:
            raise SystemExit(f"表頭查無欄名：{w!r}")
    idx = {w: header.index(w) for w in WANTED}
    out = []
    for n, r in enumerate(raw[HEADER_ROW:], start=HEADER_ROW + 1):
        if all(c is None for c in r):
            continue
        rec = {w: (r[idx[w]] if idx[w] < len(r) else None) for w in WANTED}
        rec["_row"] = n
        rec["_src_raw"] = rec[COL_SRC]
        rec["_src_norm"] = norm(rec[COL_SRC])
        rec["_src_ids"] = ID_RE.findall(rec["_src_norm"])
        out.append(rec)
    return out


def src_bucket(r: dict) -> str:
    """來源桶分類（互斥，逐列一桶）。"""
    s = r["_src_norm"]
    if s == "":
        return "空白"
    u = s.upper().replace("_", "").replace(" ", "")
    has20 = "CFTS020" in u
    has22 = "CFTS022" in u
    if has20 and has22:
        return "CFTS020+CFTS022"
    if has20:
        return "CFTS020"
    if has22:
        return "CFTS022"
    if r["_src_ids"]:
        return "純 7 位 ID（無文件名）"
    return "其他（非空、非 CFTS020/022、無 7 位 ID）"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--sec1", action="store_true")
    ap.add_argument("--sec2", action="store_true")
    ap.add_argument("--census", action="store_true")
    ap.add_argument("--cols", action="store_true")
    a = ap.parse_args()

    rs = rows()
    print(f"# SYS2 檔：{SYS2.name}")
    print(f"# sheet={SHEET} 表頭列={HEADER_ROW} 資料列數={len(rs)}"
          f" 列號範圍 {rs[0]['_row']}~{rs[-1]['_row']}")

    if a.cols:
        wb = openpyxl.load_workbook(SYS2, data_only=True, read_only=True)
        print("sheets:", wb.sheetnames)
        ws = wb[SHEET]
        hdr = next(ws.iter_rows(values_only=True))
        for i, h in enumerate(hdr):
            print(f"  [{i}] {h!r}")
        return 0

    blank = [r for r in rs if r["_src_norm"] == ""]
    blank_fr = [r for r in blank
                if norm(r[COL_CAT]).lower() == "functional requirement"]
    print(f"# 來源欄空白之列 {len(blank)}；其中 Category=Functional Requirement "
          f"{len(blank_fr)}")

    if a.census:
        c = Counter(src_bucket(r) for r in rs)
        print("\n== 來源桶盤點（互斥，總和應等於資料列數）==")
        tot = 0
        for k, v in sorted(c.items(), key=lambda kv: -kv[1]):
            print(f"  {k}: {v}")
            tot += v
        print(f"  合計: {tot}")
        print("\n== 各桶 x Category ==")
        for k in sorted(c):
            sub = [r for r in rs if src_bucket(r) == k]
            print(f"  {k}: {dict(Counter(norm(r[COL_CAT]) for r in sub))}")
        print("\n== 空白桶之 Category 分佈 ==")
        print(" ", dict(Counter(norm(r[COL_CAT]) for r in blank)))
        print("== 空白桶之 Document ID 分佈 ==")
        print(" ", dict(Counter(norm(r[COL_DOC]) for r in blank)))
        print("== 空白桶之 SW/HW/System 分佈 ==")
        print(" ", dict(Counter(norm(r[COL_SWHW]) for r in blank)))
        print("== 空白桶之子分類分佈 ==")
        print(" ", dict(Counter(norm(r[COL_SUB]) for r in blank)))
        return 0

    if a.sec1:
        for i, r in enumerate(blank_fr, 1):
            print(f"\n----- #{i} (xlsx 列 {r['_row']}) -----")
            print(f"ID: {r[COL_ID]!r}")
            print(f"Document ID: {r[COL_DOC]!r}")
            print(f"Category: {r[COL_CAT]!r}")
            print(f"SubCategory: {r[COL_SUB]!r}")
            print(f"SW/HW/System: {r[COL_SWHW]!r}")
            print(f"Source(raw): {r['_src_raw']!r}")
            print(f"Description: {r[COL_DESC]!r}")
            print(f"MD Feedback: {r[COL_MD]!r}")
        print(f"\n# §1 列數（自列舉長度取得）= {len(blank_fr)}")
        return 0

    if a.sec2:
        others = [r for r in rs if src_bucket(r) in
                  ("其他（非空、非 CFTS020/022、無 7 位 ID）", "純 7 位 ID（無文件名）")]
        for i, r in enumerate(others, 1):
            print(f"\n----- #{i} (xlsx 列 {r['_row']}) 桶={src_bucket(r)} -----")
            print(f"ID: {r[COL_ID]!r}")
            print(f"Document ID: {r[COL_DOC]!r}")
            print(f"Category: {r[COL_CAT]!r}")
            print(f"SubCategory: {r[COL_SUB]!r}")
            print(f"Source(raw): {r['_src_raw']!r}")
            print(f"Description: {str(r[COL_DESC])[:200]!r}")
            print(f"MD Feedback: {str(r[COL_MD])[:300]!r}")
        print(f"\n# §2 列數 = {len(others)}")
        return 0

    return 0


if __name__ == "__main__":
    sys.exit(main())
