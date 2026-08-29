#!/usr/bin/env python3
"""LID -> DBC 綁定探針（下放包 03 作業 E）。

對指定之 LID 清單，於 `Logical Identifiers and CAN Mapping` 之 `CAN Mapping`
分頁抽出 `Atlantis High` 群組欄，再於二個綁定 DBC 中逐候選驗證
`BO_` 訊息存在性、發送節點、`SG_` 是否屬於該訊息，並抽 `VAL_` 逐字。

取捨依 R-ICS13：發送節點 = ICS 者為主路徑，其餘記備援。
升級條件：二 DBC 皆查無 -> E4；多名皆在但無一發送節點為 ICS -> E1。

所有表頭列號/欄號皆於執行時自驗，不沿用前包常數。
"""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[3]
XLSX = REPO / "forms" / "Logical Identifiers and CAN Mapping v1_78.xlsx"
SHEET = "CAN Mapping"
DBCS = {
    "R4_BHCAN": REPO / "features" / "vehicle_setting" / "inputs" / "PDT27_E2A_R4_BHCAN.dbc",
    "R5_FDCAN8": REPO / "features" / "vehicle_setting" / "inputs" / "PDT27_E2A_R5_FDCAN8.dbc",
}
OUT = REPO / "features" / "ics_management" / "generated" / "b03" / "lid_dbc_map.json"

TARGET_LIDS = [
    "ICS_KNOB1_DIR",
    "ICS_KNOB1_VAL",
    "ICS_KNOB2_DIR",
    "ICS_KNOB2_VAL",
    "ICSPowerButton",
    "ICSScreenOffButton",
    "Enter_Button",
    "Back_Button",
]

GROUP_LABEL = "Atlantis High"
SUBCOLS = ["Signal Name", "CAN", "Format", "SNA", "VFs"]


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def find_header(ws) -> dict:
    """自驗群組列、欄名列、Atlantis High 五欄之欄號、LID 欄號與資料起始列。"""
    group_row = None
    group_col = None
    for r in range(1, 11):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() == GROUP_LABEL:
                group_row, group_col = r, c
                break
        if group_row:
            break
    if group_row is None:
        raise RuntimeError("找不到 Atlantis High 群組列")

    name_row = group_row + 1
    cols = {}
    for off, want in enumerate(SUBCOLS):
        c = group_col + off
        got = ws.cell(name_row, c).value
        got = got.strip() if isinstance(got, str) else got
        if got != want:
            raise RuntimeError(f"欄名列 {name_row} 欄 {c} 期望 {want!r} 實得 {got!r}")
        cols[want] = c

    lid_col = None
    comment_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(name_row, c).value
        if isinstance(v, str):
            s = v.strip()
            if s == "Logical Identifier" and lid_col is None:
                lid_col = c
            if s == "Usage Comment" and comment_col is None:
                comment_col = c
    if lid_col is None:
        raise RuntimeError("找不到 Logical Identifier 欄")

    return {
        "group_row": group_row,
        "group_col_atlantis_high": group_col,
        "header_name_row": name_row,
        "data_start_row": name_row + 1,
        "lid_col": lid_col,
        "usage_comment_col": comment_col,
        "atlantis_high_cols": cols,
    }


def cell_text(ws, r, c):
    if c is None:
        return None
    v = ws.cell(r, c).value
    if v is None:
        return None
    return str(v)


def parse_dbc(path: Path) -> dict:
    """回傳 {msg_name: {"id": int, "dlc": int, "node": str, "signals": {sig: raw_line}}}。"""
    msgs: dict[str, dict] = {}
    cur = None
    bo_re = re.compile(r"^BO_\s+(\d+)\s+([A-Za-z0-9_]+)\s*:\s*(\d+)\s+(\S+)")
    sg_re = re.compile(r"^\s+SG_\s+([A-Za-z0-9_]+)")
    # 註：本專案之 DBC 為 ISO-8859-1（latin-1）且 BO_ 區塊間「無空行分隔」，
    # 故訊息邊界僅由下一個 BO_ 判定，不得以空行重置。
    with path.open("r", encoding="latin-1") as fh:
        for line in fh:
            m = bo_re.match(line)
            if m:
                cur = m.group(2)
                msgs[cur] = {
                    "id": int(m.group(1)),
                    "dlc": int(m.group(3)),
                    "node": m.group(4),
                    "signals": {},
                }
                continue
            if cur is not None:
                m = sg_re.match(line)
                if m:
                    msgs[cur]["signals"][m.group(1)] = line.rstrip("\r\n")
    return msgs


def collect_vals(path: Path) -> dict[tuple[int, str], str]:
    """VAL_ 逐字，鍵為 (msg_id, signal)。"""
    out: dict[tuple[int, str], str] = {}
    val_re = re.compile(r"^VAL_\s+(\d+)\s+([A-Za-z0-9_]+)\s")
    with path.open("r", encoding="latin-1") as fh:
        for line in fh:
            m = val_re.match(line)
            if m:
                out[(int(m.group(1)), m.group(2))] = line.strip()
    return out


def split_candidates(raw: str | None) -> list[str]:
    if not raw:
        return []
    parts = [p.strip() for p in re.split(r"[\r\n]+", raw)]
    return [p for p in parts if p]


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    hdr = find_header(ws)

    dbc_msgs = {k: parse_dbc(p) for k, p in DBCS.items()}
    dbc_vals = {k: collect_vals(p) for k, p in DBCS.items()}

    # 建 LID -> 列號索引
    lid_rows: dict[str, list[int]] = {}
    for r in range(hdr["data_start_row"], ws.max_row + 1):
        v = cell_text(ws, r, hdr["lid_col"])
        if v:
            lid_rows.setdefault(v.strip(), []).append(r)

    results = []
    for lid in TARGET_LIDS:
        rows = lid_rows.get(lid, [])
        entry: dict = {"lid": lid, "rows_found": rows}
        if len(rows) != 1:
            entry["status"] = "E4" if not rows else "E1"
            entry["note"] = f"LID 於 CAN Mapping 命中 {len(rows)} 列，非唯一"
            results.append(entry)
            continue
        row = rows[0]
        ah = hdr["atlantis_high_cols"]
        raw_signal = cell_text(ws, row, ah["Signal Name"])
        entry.update(
            {
                "row": row,
                "atlantis_high": {
                    "signal_name_raw": raw_signal,
                    "can": cell_text(ws, row, ah["CAN"]),
                    "format": cell_text(ws, row, ah["Format"]),
                    "sna": cell_text(ws, row, ah["SNA"]),
                    "vfs": cell_text(ws, row, ah["VFs"]),
                },
                "usage_comment": cell_text(ws, row, hdr["usage_comment_col"]),
            }
        )
        candidates = split_candidates(raw_signal)
        entry["candidates"] = candidates

        checks = []
        for cand in candidates:
            if "." in cand:
                msg_name, sig_name = cand.split(".", 1)
            else:
                msg_name, sig_name = cand, None
            per_dbc = {}
            for key, msgs in dbc_msgs.items():
                info = msgs.get(msg_name)
                d: dict = {
                    "message_present": info is not None,
                    "message_id": info["id"] if info else None,
                    "message_node": info["node"] if info else None,
                    "signal_in_message": bool(info and sig_name in info["signals"]),
                    "val_verbatim": None,
                }
                if info and sig_name in info["signals"]:
                    d["signal_raw"] = info["signals"][sig_name].strip()
                    d["val_verbatim"] = dbc_vals[key].get((info["id"], sig_name))
                per_dbc[key] = d
            checks.append(
                {"candidate": cand, "message": msg_name, "signal": sig_name, "dbc": per_dbc}
            )
        entry["candidate_checks"] = checks

        # R-ICS13 取捨
        hits = []
        for ch in checks:
            for key, d in ch["dbc"].items():
                if d["message_present"] and d["signal_in_message"]:
                    hits.append({"candidate": ch["candidate"], "dbc": key, **d})
        entry["hits"] = hits
        ics_hits = [h for h in hits if h["message_node"] == "ICS"]
        if not hits:
            entry["status"] = "E4"
            entry["primary"] = None
            entry["fallbacks"] = []
            entry["val_verbatim"] = "無 VAL_"
        elif not ics_hits:
            entry["status"] = "E1"
            entry["primary"] = None
            entry["fallbacks"] = hits
            entry["val_verbatim"] = None
            entry["note"] = "候選皆在 DBC 但無一發送節點為 ICS（R-ICS13 未涵蓋，停下回報）"
        else:
            primary = ics_hits[0]
            entry["status"] = "RESOLVED"
            entry["primary"] = primary
            entry["fallbacks"] = [h for h in hits if h is not primary]
            entry["val_verbatim"] = primary["val_verbatim"] or "無 VAL_"
        results.append(entry)

    doc = {
        "generated_by": "features/ics_management/scripts/lid_dbc_probe.py",
        "ruling_basis": "R-ICS13",
        "scan_conditions": {
            "xlsx": str(XLSX.relative_to(REPO)),
            "sheet": SHEET,
            "header": hdr,
            "dbc": {
                key: {
                    "path": str(p.relative_to(REPO)),
                    "sha256": sha256(p),
                    "message_count": len(dbc_msgs[key]),
                }
                for key, p in DBCS.items()
            },
        },
        "lids": results,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT}")
    for e in results:
        p = e.get("primary")
        print(
            e["lid"],
            e.get("row"),
            e["status"],
            (p["candidate"] + " @" + p["dbc"] + " node=" + p["message_node"]) if p else "-",
            "VAL_" if e.get("val_verbatim") not in (None, "無 VAL_") else "no-VAL_",
        )


if __name__ == "__main__":
    main()
