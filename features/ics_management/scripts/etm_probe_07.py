"""ETM = DUT 先決驗證探針（下放包 07 作業 C）。

三路交叉：SYSAD(docx) / SWRA(xlsx) / LID(xlsx)，另加 DBC 佐證。
唯讀，不寫任何來源檔。
"""
import html
import re
import sys
import zipfile
from pathlib import Path

ROOT = Path("/Users/peihe/Work_Projects/TC_Generator")
ICS = ROOT / "features/ics_management"


def docx_text(path: Path) -> list[str]:
    """docx 抽取：</w:p>→換行、</w:tc>→tab、去標籤、html.unescape。"""
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8")
    xml = xml.replace("</w:p>", "\n").replace("</w:tc>", "\t")
    txt = re.sub(r"<[^>]+>", "", xml)
    return [html.unescape(line) for line in txt.split("\n")]


def scan_lines(lines, keywords, label, ci=False):
    """逐行掃關鍵詞，印出 段落序號 + 逐字內容。"""
    hits = 0
    for i, line in enumerate(lines):
        hay = line.upper() if ci else line
        for kw in keywords:
            needle = kw.upper() if ci else kw
            if needle in hay:
                s = line.strip()
                if s:
                    print(f"[{label}] p{i}\t{kw}\t{s[:600]}")
                    hits += 1
                break
    print(f"[{label}] total hits={hits}, total paragraphs={len(lines)}")


def cmd_sysad():
    p = ICS / "inputs/SYS3_CFTS020_ICS_FM-WI-FSM-011-A01_System Architectural Design_SYSAD_v1.0.docx"
    lines = docx_text(p)
    scan_lines(lines, ["ETM", "LTM", "SGW", "TBM", "BU_", "node", "Node", "DBC"], "SYSAD")


def cmd_sysad_dump():
    p = ICS / "inputs/SYS3_CFTS020_ICS_FM-WI-FSM-011-A01_System Architectural Design_SYSAD_v1.0.docx"
    for i, line in enumerate(docx_text(p)):
        s = line.strip()
        if s:
            print(f"p{i}\t{s}")


def cmd_cfts020():
    p = ICS / "inputs/R1LR_Atl-H_26PI1.5 Mar Release-Cabin_CFTS_020 ICS and DCSD _20260310-1533.docx"
    scan_lines(docx_text(p), ["ETM", "LTM", "SGW", "TBM"], "CFTS020")


def cmd_cfts022():
    p = ICS / "inputs/R1LR_Atl-H_26PI2.5 Jun Release-Privacy_CFTS_022 Functional Specification_20260608-1205.docx"
    scan_lines(docx_text(p), ["ETM", "LTM", "SGW", "TBM"], "CFTS022")


def cmd_swra():
    import openpyxl
    p = ICS / "inputs/ICS_Management_FM-WI-FSM-037-A03_STLA_Report_SWRA.xlsx"
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    kws = ["ETM", "LTM", "SGW", "TBM"]
    for ws in wb.worksheets:
        print(f"=== sheet: {ws.title} dims={ws.calculate_dimension()}")
        for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for c, v in enumerate(row, start=1):
                if v is None:
                    continue
                s = str(v)
                up = s.upper()
                for kw in kws:
                    # 詞邊界，避免 SYSTEM 之類誤中
                    if re.search(rf"(?<![A-Z0-9]){kw}(?![A-Z0-9])", up):
                        print(f"[SWRA] {ws.title} r{r} c{c}\t{kw}\t{s[:500]}")
                        break


def cmd_swra_head():
    import openpyxl
    p = ICS / "inputs/ICS_Management_FM-WI-FSM-037-A03_STLA_Report_SWRA.xlsx"
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    for ws in wb.worksheets:
        print(f"=== sheet: {ws.title} dims={ws.calculate_dimension()}")
        for r, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
            print(f"r{r}: {[str(v)[:60] if v is not None else None for v in row]}")


def cmd_lid_head():
    import openpyxl
    p = ROOT / "forms/Logical Identifiers and CAN Mapping v1_78.xlsx"
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    print("sheets:", wb.sheetnames)
    ws = wb["CAN Mapping"]
    print("dims:", ws.calculate_dimension())
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        for c, v in enumerate(row, start=1):
            if v is not None:
                print(f"r{r} c{c}\t{str(v)[:120]}")


def cmd_lid_tgw():
    import openpyxl
    p = ROOT / "forms/Logical Identifiers and CAN Mapping v1_78.xlsx"
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    ws = wb["CAN Mapping"]
    rows = list(ws.iter_rows(values_only=True))
    grp = rows[1]
    names = rows[2]
    for r, row in enumerate(rows, start=1):
        joined = " ".join(str(v) for v in row if v is not None)
        if "TGW_DISP_STAT" in joined.upper():
            print(f"--- row {r}")
            for c, v in enumerate(row, start=1):
                if v is not None:
                    g = grp[c - 1] if c - 1 < len(grp) else None
                    n = names[c - 1] if c - 1 < len(names) else None
                    print(f"  c{c}\tgrp={g}\tcol={n}\tval={str(v)[:200]}")


def cmd_dbc():
    for name in ["PDT27_E2A_R4_BHCAN.dbc", "PDT27_E2A_R5_FDCAN8.dbc"]:
        p = ROOT / "features/vehicle_setting/inputs" / name
        txt = p.read_text(encoding="latin-1")
        print(f"=== {name}")
        for line in txt.splitlines():
            if line.startswith("BU_:"):
                print("BU_ line:", line.strip())
        # 訊息邊界：下一個 BO_
        lines = txt.splitlines()
        idx = [i for i, l in enumerate(lines) if l.startswith("BO_ ")]
        for target in ["TELEMATIC_DISPLAY2", "TELEMATIC_FD_4"]:
            for k, i in enumerate(idx):
                if re.match(rf"BO_ \d+ {target}\s*:", lines[i]):
                    end = idx[k + 1] if k + 1 < len(idx) else len(lines)
                    print(f"--- {target} BO_ line: {lines[i].strip()}")
                    for l in lines[i:end]:
                        if "TGW_DISP_STAT" in l or "Telematic_Power" in l:
                            print("   SG_:", l.strip())
        for l in lines:
            if l.startswith("VAL_") and "TGW_DISP_STAT" in l:
                print("VAL_:", l.strip())
        # ETM 節點所發的訊息數
        cnt = {}
        for i in idx:
            m = re.match(r"BO_ \d+ \S+\s*:\s*\d+\s+(\S+)", lines[i])
            if m:
                cnt[m.group(1)] = cnt.get(m.group(1), 0) + 1
        print("sender counts:", cnt)


if __name__ == "__main__":
    globals()[f"cmd_{sys.argv[1]}"]()
