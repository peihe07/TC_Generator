#!/usr/bin/env python3
"""b01 之 test_item 上半逐字比對（IN §4.3.1 之 verbatim 要求）。

比對對象自**原檔**重抽，不用任何暫存副本（charter §探測與工具）：
  - CFTS022 = features/privacy/inputs/R1LR_Atl-H_25PI3.5_Privacy_CFTS_022 …docx
  - SWRA    = features/ics_management/inputs/ICS_Management_…_SWRA.xlsx
  - CFTS020 = features/ics_management/inputs/R1LR_Atl-H_26PI1.5 … CFTS_020 …docx

正規化條件（逐項揭露，比對前雙方同套）：
  1. 彎引號 “ ” ‘ ’ → 直引號 " '  —— 排版正規化，IN §11 令 UI 標籤用直雙引號
  2. NBSP(U+00A0) → 空格；非斷字連字號 U+2011 → '-'
  3. 連續空白摺為單一空格、頭尾去空白
  4. 去除句末之單一句號 —— IN §11 令交付欄無尾句號，故上半抄入時必去
  5. 句首字母之大小寫（R-4：自原句中段起抄時句首轉大寫屬排版正規化）

上述四項皆為**形而非義**之變更；除此之外一字不動。
"""
from __future__ import annotations

import html
import json
import re
import sys
import warnings
import zipfile
from pathlib import Path

warnings.filterwarnings("ignore")
from openpyxl import load_workbook  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
REPO = ROOT.parents[1]
CFTS022 = REPO / ("features/privacy/inputs/R1LR_Atl-H_25PI3.5_Privacy_CFTS_022 "
                  "Functional Specification_20250910_1708.docx")
SWRA = ROOT / "inputs/ICS_Management_FM-WI-FSM-037-A03_STLA_Report_SWRA.xlsx"
CFTS020 = ROOT / ("inputs/R1LR_Atl-H_26PI1.5 Mar Release-Cabin_CFTS_020 ICS and "
                  "DCSD _20260310-1533.docx")
DEFAULT_BATCHES = [ROOT / "generated/b01/b01_tcs.json",
                   ROOT / "generated/b02/b02_tcs.json",
                   ROOT / "generated/b03/b03_tcs.json",
                   ROOT / "generated/b04/b04_tcs.json"]


def docx_text(path: Path) -> str:
    xml = zipfile.ZipFile(path).read("word/document.xml").decode("utf8")
    xml = re.sub(r"</w:p>", "\n", xml)
    xml = re.sub(r"</w:tc>", "\t", xml)
    return html.unescape(re.sub(r"<[^>]+>", "", xml))


def norm(s: str) -> str:
    s = (s.replace("“", '"').replace("”", '"')
          .replace("‘", "'").replace("’", "'")
          .replace("\xa0", " ").replace("‑", "-"))
    return re.sub(r"\s+", " ", s).strip().rstrip(".")


def main() -> int:
    src = norm(docx_text(CFTS022))
    ws = load_workbook(SWRA, data_only=True)["SWE1 Requirements"]
    swra = {norm(r[3]) for r in ws.iter_rows(values_only=True)
            if r[0] and str(r[0]).startswith("SWE-ICS") and r[3]}

    src020 = norm(docx_text(CFTS020))
    paths = [Path(a) for a in sys.argv[1:]] or DEFAULT_BATCHES
    paths = [p for p in paths if p.exists()]
    tcs = []
    for bp in paths:
        tcs += json.loads(bp.read_text())["tcs"]
    print("受檢批次：" + "、".join(p.parent.name for p in paths))
    bad = []
    for t in tcs:
        upper = norm(t["test_item"].split("\n")[0])
        # R-4：自原句中段起抄時，句首字母轉大寫屬排版正規化，允許。
        # 故比對取「原樣」與「句首轉小寫」二式，任一命中即算逐字。
        cands = {upper, (upper[:1].lower() + upper[1:]) if upper else upper}
        origin = None
        for u in cands:
            if u in src:
                origin = "CFTS022"; break
            if u in src020:
                origin = "CFTS020"; break
            if u in swra:
                origin = "SWRA"; break
        if origin and upper not in {u for u in cands if u == upper} | set():
            pass
        print(f'{t["tc_title"]:46} {t["req_id"]:13} '
              f'{origin or "**MISS**"}  {t["specification_reference"].splitlines()[0]}')
        if origin is None:
            bad.append(t["tc_title"])
    print()
    print(f"總判：{'**FAIL**' if bad else 'PASS'} —— {len(tcs)} 條，"
          f"逐字命中 {len(tcs) - len(bad)}，未命中 {len(bad)}")
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())
