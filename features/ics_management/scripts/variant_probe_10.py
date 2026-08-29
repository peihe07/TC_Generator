#!/usr/bin/env python3
"""下放包 10 作業 A —— 本 DUT 之變體歸屬量測（Associated / Disassociated）。

唯讀量測腳本。不寫入任何 repo 檔（輸出僅到 stdout）。

掃描條件（逐項揭露）：
  - CFTS020 物件面：以 `importlib` 唯讀載入 `cfts020_probe.parse()`（不修改該檔）。
    其抽取法見該檔 docstring；判定為 R-ICS2 v2(b)。
  - CFTS020 行層級（章節標題／目次）：重用 `cfts020_probe.doc_lines()`。
  - 其他 .docx：`word/document.xml` → `</w:p>` 換行、`</w:tc>` tab、去標籤、`html.unescape`。
  - .xlsx：openpyxl，`read_only=True`、`data_only=True`。
  - .dbc：**一律 `latin-1` 開檔**；訊息邊界由下一個 `BO_` 判定。
  - 詞界：`\\b` 之 ASCII 詞界；`DCSD`／`TLM`／`LTM` 等縮寫**區分大小寫**；
    `Associated`／`Disassociated`／`Silver Box` 亦區分大小寫（另附不分大小寫之對照數）。
  - 「出現」計數單位於物件面 = 物件數；於行面 = 行數（同行多次記 1，另附總次數）。

用法：
  python3 features/ics_management/scripts/variant_probe_10.py --m1
  python3 features/ics_management/scripts/variant_probe_10.py --m2
  python3 features/ics_management/scripts/variant_probe_10.py --m3
  python3 features/ics_management/scripts/variant_probe_10.py --m4
  python3 features/ics_management/scripts/variant_probe_10.py --m5
  python3 features/ics_management/scripts/variant_probe_10.py --m6
  python3 features/ics_management/scripts/variant_probe_10.py --all
"""
from __future__ import annotations

import argparse
import html
import importlib.util
import re
import sys
import zipfile
from pathlib import Path

FEAT = Path(__file__).resolve().parents[1]
ROOT = FEAT.parents[1]
SCRIPTS = FEAT / "scripts"

DBC_FILES = [
    ROOT / "features/vehicle_setting/inputs/PDT27_E2A_R4_BHCAN.dbc",
    ROOT / "features/vehicle_setting/inputs/PDT27_E2A_R5_FDCAN8.dbc",
]


def load_probe():
    """以 importlib 唯讀載入既有 cfts020_probe（禁改該檔）。"""
    spec = importlib.util.spec_from_file_location(
        "cfts020_probe_ro", SCRIPTS / "cfts020_probe.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def docx_lines(path: Path) -> list[str]:
    """與 cfts020_probe.doc_lines() 同一抽取法，套用於任意 .docx。"""
    xml = zipfile.ZipFile(path).read("word/document.xml").decode("utf8")
    xml = re.sub(r"</w:p>", "\n", xml)
    xml = re.sub(r"</w:tc>", "\t", xml)
    return html.unescape(re.sub(r"<[^>]+>", "", xml)).split("\n")


def axes(o: dict) -> str:
    return (f"ECU={o['ecu']} Radio={o['radio']} EE={o['ee']}")


# ---------------------------------------------------------------- 量測項 1
def m1(probe) -> None:
    print("=" * 78)
    print("§1 變體定義之完整逐字（CFTS020-4819134 及鄰接物件）")
    print("=" * 78)
    objs = probe.parse()
    print(f"CFTS020 物件母數：{len(objs)}")
    ids = [o["id"] for o in objs]
    for target in ("4819134",):
        i = ids.index(target)
        for j in range(max(0, i - 3), min(len(objs), i + 4)):
            o = objs[j]
            mark = " <<< 目標" if o["id"] == target else ""
            print(f"\n--- [{j - i:+d}] {o['id']}  §{o['section_no']} {o['section']}{mark}")
            print(f"    {axes(o)}  Artifact={o['artifact_type']} State={o['state']}")
            print(f"    判定(v2)={o['verdict']} reasons={o['reasons']}")
            print(f"    逐字：{o['text']}")

    print("\n" + "-" * 78)
    print("§1b `_ADspl` / `_DDspl` 於 CFTS020 之全部出現（物件面，區分大小寫）")
    print("-" * 78)
    for suf in ("_ADspl", "_DDspl"):
        hits = [o for o in objs if suf in o["text"]]
        print(f"\n### `{suf}`：命中物件 {len(hits)} 個")
        for o in hits:
            n = o["text"].count(suf)
            print(f"  - {o['id']} §{o['section_no']}  n={n}  {axes(o)}  判定={o['verdict']}")
            print(f"    逐字：{o['text']}")

    # 行層級（含目次、表格、標題）
    lines = probe.doc_lines()
    print("\n" + "-" * 78)
    print("§1c `_ADspl` / `_DDspl` 於 CFTS020 之行層級命中（含目次／表格／標題）")
    print("-" * 78)
    for suf in ("_ADspl", "_DDspl"):
        rows = [(i, l.strip()) for i, l in enumerate(lines) if suf in l]
        tot = sum(l.count(suf) for _, l in rows)
        print(f"\n### `{suf}`：命中行 {len(rows)}，總次數 {tot}")
        for i, l in rows:
            print(f"  L{i}: {l[:400]}")

    # LTM_ADspl / LTM_DDspl 前綴組合
    print("\n" + "-" * 78)
    print("§1d 前綴組合統計（行層級，區分大小寫）")
    print("-" * 78)
    joined = "\n".join(lines)
    for tok in ("LTM_ADspl", "LTM_DDspl", "ETM_ADspl", "ETM_DDspl",
                "RRM_ADspl", "RRM_DDspl"):
        print(f"  {tok}: {joined.count(tok)}")

    # R1L-R 與後綴同物件共現
    print("\n" + "-" * 78)
    print("§1e `R1L-R` / `R1L` 與 `_ADspl` / `_DDspl` 之同物件共現")
    print("-" * 78)
    for suf in ("_ADspl", "_DDspl"):
        for rv in ("R1L-R", "R1L"):
            n = sum(1 for o in objs
                    if suf in o["text"] and o["radio"] and rv in o["radio"])
            m = sum(1 for o in objs if suf in o["text"] and rv in o["text"])
            print(f"  {suf} × Radio軸含 {rv}：{n} 物件；{suf} × 本文含字串 {rv}：{m} 物件")


# ---------------------------------------------------------------- 量測項 2
def m2(probe) -> None:
    print("=" * 78)
    print("§2 Configuration parameters 四節（§1.7 / §1.10 / §1.13 / §1.17）")
    print("=" * 78)
    objs = probe.parse()

    lines = probe.doc_lines()
    print("\n--- 全部含 `Configuration parameters` 之章節標題行（區分大小寫，行層級）")
    for i, l in enumerate(lines):
        s = l.strip()
        if "PAGEREF" in s:
            continue
        m = probe.SEC_RE.match(s)
        if m and "onfiguration parameter" in m.group(2):
            print(f"  §{m.group(1)}  {m.group(2)} {{{m.group(3)}}}")

    for sec in ("1.7", "1.10", "1.13", "1.17"):
        sub = [o for o in objs
               if o["section_no"] == sec or o["section_no"].startswith(sec + ".")]
        ap = [o for o in sub if o["verdict"] == "適用"]
        title = sub[0]["section"] if sub else "(無物件)"
        # 章節標題逐字（自行層級取）
        head = [l.strip() for l in lines
                if "PAGEREF" not in l and probe.SEC_RE.match(l.strip())
                and probe.SEC_RE.match(l.strip()).group(1) == sec]
        print(f"\n{'=' * 70}")
        print(f"§{sec} 標題逐字：{head[0] if head else '(查無)'}")
        print(f"  物件數 {len(sub)}；R-ICS2 v2(b) 判適用 {len(ap)}")
        for o in sub:
            print(f"  - {o['id']} §{o['section_no']} 判定={o['verdict']} {axes(o)}")
            print(f"      reasons={o['reasons']}")
            print(f"      逐字：{o['text'][:600]}")


# ---------------------------------------------------------------- 量測項 3
def m3(probe) -> None:
    print("=" * 78)
    print("§3 DCSD 之存在與否（結構證據）")
    print("=" * 78)

    # --- 3a LID 表
    lid = ROOT / "forms/Logical Identifiers and CAN Mapping v1_78.xlsx"
    print(f"\n--- 3a LID 表：{lid}")
    print(f"    存在={lid.exists()}")
    if lid.exists():
        import openpyxl
        wb = openpyxl.load_workbook(lid, read_only=True, data_only=True)
        print(f"    分頁：{wb.sheetnames}")
        if "CAN Mapping" in wb.sheetnames:
            ws = wb["CAN Mapping"]
            rows = list(ws.iter_rows(values_only=True))
            print(f"    CAN Mapping 列數（含表頭）：{len(rows)}")
            # 表頭自驗：找第一列非空欄數最多者（前 5 列內）
            hdr_i, hdr = 0, rows[0]
            for i in range(min(5, len(rows))):
                if sum(1 for c in rows[i] if c not in (None, "")) > \
                   sum(1 for c in hdr if c not in (None, "")):
                    hdr_i, hdr = i, rows[i]
            print(f"    表頭自驗：第 {hdr_i} 列，欄名 = {[c for c in hdr]}")
            data = rows[hdr_i + 1:]
            print(f"    資料列數：{len(data)}")
            hits = [(hdr_i + 1 + n, r) for n, r in enumerate(data)
                    if any("DCSD" in str(c) for c in r if c is not None)]
            print(f"    含 `DCSD`（區分大小寫）之資料列：{len(hits)}")
            for n, r in hits:
                print(f"      L{n}: {[c for c in r if c not in (None, '')]}")
            # LTM 對照
            for tok in ("LTM", "TLM", "ICS"):
                cnt = sum(1 for r in data
                          if any(re.search(rf"\b{tok}\b", str(c)) for c in r if c is not None))
                print(f"    含詞界 `{tok}` 之資料列：{cnt}")
        wb.close()

    # --- 3b DBC
    print("\n--- 3b 二綁定 DBC（latin-1 開檔；訊息邊界由下一個 `BO_` 判定）")
    bo_re = re.compile(r"^BO_ (\d+) (\S+)\s*:\s*(\d+)\s+(\S+)")
    sg_re = re.compile(r"^\s*SG_ (\S+)")
    for dbc in DBC_FILES:
        print(f"\n### {dbc}  存在={dbc.exists()}")
        if not dbc.exists():
            continue
        text = dbc.read_text(encoding="latin-1")
        lines = text.split("\n")
        # 節點宣告
        nodes = []
        for l in lines:
            if l.startswith("BU_:"):
                nodes = l[4:].split()
                break
        print(f"  BU_ 節點數 {len(nodes)}；含 DCSD? {'DCSD' in nodes}；"
              f"含 LTM? {'LTM' in nodes}；含 TLM? {'TLM' in nodes}；含 ICS? {'ICS' in nodes}")
        print(f"  節點含 'DCSD' 子字串者：{[n for n in nodes if 'DCSD' in n]}")
        print(f"  節點含 'LTM'/'TLM' 子字串者："
              f"{[n for n in nodes if 'LTM' in n or 'TLM' in n]}")
        # 訊息切塊
        msgs, cur = [], None
        for i, l in enumerate(lines):
            m = bo_re.match(l)
            if m:
                cur = {"line": i, "id": m.group(1), "name": m.group(2),
                       "sender": m.group(4), "sigs": []}
                msgs.append(cur)
                continue
            if cur is not None:
                sm = sg_re.match(l)
                if sm:
                    cur["sigs"].append((sm.group(1), l))
        print(f"  BO_ 訊息數：{len(msgs)}")
        # DCSD 為發方
        snd = [m for m in msgs if "DCSD" in m["sender"]]
        print(f"  DCSD 為發方（BO_ 之 transmitter 欄）之訊息：{len(snd)}")
        for m in snd:
            print(f"    - {m['name']} (id={m['id']}, sender={m['sender']}, "
                  f"訊號 {len(m['sigs'])})")
        # 訊息名含 DCSD
        nm = [m for m in msgs if "DCSD" in m["name"]]
        print(f"  訊息名含 `DCSD` 之訊息：{len(nm)}  {[m['name'] for m in nm]}")
        # 訊號名含 DCSD
        sg_hits = [(m["name"], s) for m in msgs for s, _ in m["sigs"] if "DCSD" in s]
        print(f"  訊號名含 `DCSD`：{len(sg_hits)}")
        for mn, s in sg_hits:
            print(f"    - {mn}.{s}")
        # BO_TX_BU_ / 接收方（SG_ 行末之接收節點清單）
        rec = []
        for m in msgs:
            for s, raw in m["sigs"]:
                tail = raw.rsplit(" ", 1)[-1].strip()
                if "DCSD" in tail:
                    rec.append((m["name"], s, tail))
        print(f"  SG_ 行末接收節點含 `DCSD`：{len(rec)}")
        for mn, s, t in rec[:40]:
            print(f"    - {mn}.{s} -> {t}")
        # LTM 為發方 / 收方
        for tok in ("LTM", "TLM", "ICS"):
            s_ = [m["name"] for m in msgs if m["sender"] == tok]
            r_ = set()
            for m in msgs:
                for s, raw in m["sigs"]:
                    tail = raw.rsplit(" ", 1)[-1].strip()
                    if tok in tail.split(","):
                        r_.add(m["name"])
            print(f"  節點 `{tok}`：發送 {len(s_)} 則、接收 {len(r_)} 則")
            if s_:
                print(f"      發送：{s_[:20]}")
            if r_:
                print(f"      接收：{sorted(r_)[:20]}")
        # DCSD 與 LTM 是否共現於同一訊息（收發關係）
        both = []
        for m in msgs:
            parties = {m["sender"]}
            for s, raw in m["sigs"]:
                parties |= set(x.strip() for x in raw.rsplit(" ", 1)[-1].split(","))
            if any("DCSD" in p for p in parties) and any(p in ("LTM", "TLM") for p in parties):
                both.append(m["name"])
        print(f"  DCSD 與 LTM/TLM 同現於同一訊息之收發方：{len(both)}  {both}")


# ---------------------------------------------------------------- 量測項 4
def m4() -> None:
    print("=" * 78)
    print("§4 DUT 自身文件之 DCSD / Silver Box / Associated / Disassociated")
    print("=" * 78)
    targets: list[Path] = []
    for d in (FEAT / "inputs", ROOT / "spec-index/sources"):
        print(f"\n--- 目錄 {d} 存在={d.exists()}")
        if d.exists():
            for p in sorted(d.rglob("*")):
                if p.is_file() and p.suffix.lower() in (".docx", ".xlsx"):
                    print(f"    {p.relative_to(ROOT)}")
                    targets.append(p)

    pats = {
        "DCSD": re.compile(r"DCSD"),
        "Silver Box": re.compile(r"Silver Box"),
        "Silver Box(i)": re.compile(r"silver\s*box", re.I),
        "Associated": re.compile(r"\bAssociated\b"),
        "Disassociated": re.compile(r"\bDisassociated\b"),
        "assoc(i)": re.compile(r"\b(dis)?associated\b", re.I),
        "_ADspl": re.compile(r"_ADspl"),
        "_DDspl": re.compile(r"_DDspl"),
    }
    for p in targets:
        print(f"\n{'=' * 70}\n### {p.relative_to(ROOT)}")
        try:
            if p.suffix.lower() == ".docx":
                lines = docx_lines(p)
            else:
                import openpyxl
                wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                lines = []
                for ws in wb.worksheets:
                    for r in ws.iter_rows(values_only=True):
                        lines.append("\t".join("" if c is None else str(c) for c in r))
                wb.close()
        except Exception as e:  # noqa: BLE001
            print(f"  讀取失敗：{e!r}")
            continue
        print(f"  行數 {len(lines)}")
        for name, rx in pats.items():
            rows = [(i, l.strip()) for i, l in enumerate(lines) if rx.search(l)]
            tot = sum(len(rx.findall(l)) for l in lines)
            print(f"  `{name}`：命中行 {len(rows)}，總次數 {tot}")
            for i, l in rows[:25]:
                print(f"      L{i}: {l[:400]}")


# ---------------------------------------------------------------- 量測項 5
def m5(probe) -> None:
    print("=" * 78)
    print("§5 CFTS020 章節樹（頂層＋二層標題逐字）")
    print("=" * 78)
    lines = probe.doc_lines()
    seen = set()
    for l in lines:
        s = l.strip()
        if "PAGEREF" in s:
            continue
        m = probe.SEC_RE.match(s)
        if not m:
            continue
        no = m.group(1)
        if no.count(".") <= 1 and no not in seen:
            seen.add(no)
            print(f"  §{no}\t{m.group(2)} {{{m.group(3)}}}")

    objs = probe.parse()
    print("\n--- 各頂層節之物件數與 v2 適用數")
    from collections import Counter
    tot, ap = Counter(), Counter()
    for o in objs:
        top = ".".join(o["section_no"].split(".")[:2])
        tot[top] += 1
        if o["verdict"] == "適用":
            ap[top] += 1
    for k in sorted(tot, key=lambda x: [int(y) for y in x.split(".") if y.isdigit()] or [0]):
        print(f"  §{k}\t物件 {tot[k]}\t適用 {ap[k]}")

    print("\n--- §1.18 / §1.8 之 DCSD 提及數（物件面）")
    for sec in ("1.8", "1.18", "1.15", "1.19", "1.11"):
        sub = [o for o in objs
               if o["section_no"] == sec or o["section_no"].startswith(sec + ".")]
        d = [o for o in sub if "DCSD" in o["text"]]
        print(f"  §{sec}：物件 {len(sub)}，含 `DCSD` {len(d)}，"
              f"適用 {sum(1 for o in sub if o['verdict'] == '適用')}")


# ---------------------------------------------------------------- 量測項 6
# 分支配對檢定：以「章節標題所宣告之變體分支」為分組，比對各組之 v2 適用數。
# **排除 §1.8 與 §1.18**（交辦之禁循環令），故本檢定完全脫離該二節。
BRANCH = {
    "Associated": ["1.5", "1.6", "1.7", "1.14", "1.16", "1.17"],
    "Disassociated/SilverBox": ["1.3", "1.9", "1.10", "1.11", "1.12", "1.13", "1.15", "1.19"],
    "中性（Common/Intro）": ["1.1", "1.2", "1.4"],
    "【禁區·循環·不計入】": ["1.8", "1.18"],
}


def m6(probe) -> None:
    from collections import Counter
    print("=" * 78)
    print("§6 分支配對檢定（排除 §1.8／§1.18）＋ SYS2 交叉表")
    print("=" * 78)
    objs = probe.parse()

    def sub(sec):
        return [o for o in objs
                if o["section_no"] == sec or o["section_no"].startswith(sec + ".")]

    for grp, secs in BRANCH.items():
        t = a = 0
        print(f"\n### {grp}")
        for s in secs:
            g = sub(s)
            ga = sum(1 for o in g if o["verdict"] == "適用")
            t += len(g)
            a += ga
            print(f"  §{s}\t物件 {len(g)}\t適用 {ga}")
        print(f"  小計：物件 {t}，適用 {a}")

    # SYS2 交叉表：來源節 × Category
    sys2 = FEAT / ("inputs/SYS2_CFTS_020_DISP_TCH_ICS_20260616_All_HW_"
                   "System_Accepted & Released.xlsx")
    print(f"\n{'=' * 70}\nSYS2 交叉表（來源 ObjectID 之 CFTS020 節 × SYS2 分類 Category）")
    print(f"檔：{sys2.name}  存在={sys2.exists()}")
    if sys2.exists():
        import openpyxl
        wb = openpyxl.load_workbook(sys2, read_only=True, data_only=True)
        rows = list(wb["Basic Report"].iter_rows(values_only=True))
        wb.close()
        hdr, data = rows[0], rows[1:]
        print(f"表頭自驗：欄0={hdr[0]!r} 欄3={hdr[3]!r} 欄7={hdr[7]!r} 欄10={hdr[10]!r}")
        print(f"資料列 {len(data)}")
        secmap = {o["id"]: o["section_no"] for o in objs}
        cross, unmapped = {}, 0
        for r in data:
            sid = str(r[7] or "").strip()
            s = secmap.get(sid)
            if s is None:
                unmapped += 1
                continue
            top = ".".join(s.split(".")[:2])
            cat = str(r[10])
            cross.setdefault(top, Counter())[cat] += 1
        print(f"來源 ObjectID 無法對映 CFTS020 物件之列：{unmapped}")
        for k in sorted(cross, key=lambda x: -sum(cross[x].values())):
            print(f"  §{k}\t總 {sum(cross[k].values())}\t{cross[k].most_common()}")
        # DCSD 面
        d = [r for r in data if r[3] and "DCSD" in str(r[3])]
        print(f"\n  Description 含 `DCSD` 之列：{len(d)}；"
              f"Category 分佈 {Counter(str(r[10]) for r in d).most_common()}")
        print("  MD/HARMAN 回饋含 `not LTM` 之列（逐字）：")
        for r in data:
            fb = " ".join(str(r[i]) for i in (16, 18) if r[i])
            if "not LTM" in fb:
                print(f"    {r[0]} src={r[7]} cat={r[10]} | {fb[:180]}")

    # LID 焦點列
    lid = ROOT / "forms/Logical Identifiers and CAN Mapping v1_78.xlsx"
    print(f"\n{'=' * 70}\nLID 焦點 LID（RQ_DISP_INTS / TGW_DISP_STAT / DCSD_DISP_STAT）")
    if lid.exists():
        import openpyxl
        wb = openpyxl.load_workbook(lid, read_only=True, data_only=True)
        rows = list(wb["CAN Mapping"].iter_rows(values_only=True))
        wb.close()
        for r in rows[3:]:
            if str(r[0]) in ("RQ_DISP_INTS", "TGW_DISP_STAT", "DCSD_DISP_STAT"):
                print(f"  {r[0]}\n     Function 逐字：{r[1]}"
                      f"\n     Arch Basis={r[3]!r} Signal={r[5]!r} CAN={r[6]!r} VFs={r[9]!r}")
        dc = [r for r in rows[3:] if str(r[0]).startswith("DCSD")]
        print(f"  LID 名以 `DCSD` 起首之列：{len(dc)}；"
              f"Arch Basis 分佈 {Counter(str(r[3]) for r in dc).most_common()}")


def main() -> int:
    ap = argparse.ArgumentParser()
    for f in ("m1", "m2", "m3", "m4", "m5", "m6"):
        ap.add_argument(f"--{f}", action="store_true")
    ap.add_argument("--all", action="store_true")
    a = ap.parse_args()
    probe = load_probe()
    run = {k: getattr(a, k) or a.all for k in ("m1", "m2", "m3", "m4", "m5", "m6")}
    if not any(run.values()):
        ap.print_help()
        return 1
    if run["m1"]:
        m1(probe)
    if run["m2"]:
        m2(probe)
    if run["m3"]:
        m3(probe)
    if run["m4"]:
        m4()
    if run["m5"]:
        m5(probe)
    if run["m6"]:
        m6(probe)
    return 0


if __name__ == "__main__":
    sys.exit(main())
