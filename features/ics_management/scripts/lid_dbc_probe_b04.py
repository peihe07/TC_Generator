#!/usr/bin/env python3
"""LID -> DBC 綁定探針（下放包 04 作業 B，DR-ICS15 四訊號）。

對 `TGW_DISP_STAT`、`RQ_DISP_INTS`、`DCSD_DISP_STAT`、`Telematic_Power`
四個 Logical Identifier，於 `Logical Identifiers and CAN Mapping` 之
`CAN Mapping` 分頁抽出 `Atlantis High` 群組欄，再於二個綁定 DBC 中
逐候選驗證 `BO_` 訊息存在性、發送節點、`SG_` 是否在該訊息界內，
並抽 `VAL_` 逐字。

方法約束（A-ICS25 實測所得，不得放寬）：
- DBC 一律以 latin-1 開檔。
- `BO_` 區塊間無空行分隔，訊息邊界僅由下一個 `BO_` 判定。
- LID 表頭列號/欄號執行時自驗，不沿用前包常數。

取捨依 R-ICS13：一格多名且皆在庫時，取發送節點 = ICS 者為主路徑。
本包四訊號依 CFTS020 語意未必為 ICS 發送，故本腳本另記
`ics13_applicable`：僅在「候選 >= 2 且皆在庫」時 R-ICS13 之情境成立；
單名或無 ICS 候選時如實回報節點，不硬套。

升級條件：
- E5：LID 查無，或 LID 有而二 DBC 皆無 -> 狀態「維持佔位」。
- E1：多名並列且皆在庫而無一發送節點為 ICS -> 不自選。

輸出為**累計**對照表：本輪四筆 + 併入 b03 已解八筆（沿用 b03 實測）。
"""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[3]
XLSX = REPO / "forms" / "Logical Identifiers and CAN Mapping v1_78.xlsx"
SHEET = "CAN Mapping"
DBCS = {
    "R4_BHCAN": REPO / "features" / "vehicle_setting" / "inputs" / "PDT27_E2A_R4_BHCAN.dbc",
    "R5_FDCAN8": REPO / "features" / "vehicle_setting" / "inputs" / "PDT27_E2A_R5_FDCAN8.dbc",
}
B03_MAP = REPO / "features" / "ics_management" / "generated" / "b03" / "lid_dbc_map.json"
OUT = REPO / "features" / "ics_management" / "generated" / "b04" / "lid_dbc_map.json"

TARGET_LIDS = [
    "TGW_DISP_STAT",
    "RQ_DISP_INTS",
    "DCSD_DISP_STAT",
    "Telematic_Power",
]

GROUP_LABEL = "Atlantis High"
SUBCOLS = ["Signal Name", "CAN", "Format", "SNA", "VFs"]
DBC_ENCODING = "latin-1"


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def find_header(ws) -> dict:
    """自驗群組列、欄名列、Atlantis High 五欄之欄號、LID 欄號與資料起始列。"""
    group_row = None
    group_col = None
    for r in range(1, 11):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() == GROUP_LABEL:
                group_row, group_col = r, c
                break
        if group_row:
            break
    if group_row is None:
        raise RuntimeError("找不到 Atlantis High 群組列")

    name_row = group_row + 1
    cols = {}
    for off, want in enumerate(SUBCOLS):
        c = group_col + off
        got = ws.cell(name_row, c).value
        got = got.strip() if isinstance(got, str) else got
        if got != want:
            raise RuntimeError(f"欄名列 {name_row} 欄 {c} 期望 {want!r} 實得 {got!r}")
        cols[want] = c

    lid_col = None
    comment_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(name_row, c).value
        if isinstance(v, str):
            s = v.strip()
            if s == "Logical Identifier" and lid_col is None:
                lid_col = c
            if s == "Usage Comment" and comment_col is None:
                comment_col = c
    if lid_col is None:
        raise RuntimeError("找不到 Logical Identifier 欄")

    return {
        "group_row": group_row,
        "group_col_atlantis_high": group_col,
        "header_name_row": name_row,
        "data_start_row": name_row + 1,
        "lid_col": lid_col,
        "usage_comment_col": comment_col,
        "atlantis_high_cols": cols,
    }


def cell_text(ws, r, c):
    if c is None:
        return None
    v = ws.cell(r, c).value
    if v is None:
        return None
    return str(v)


def parse_dbc(path: Path) -> dict:
    """回傳 {msg_name: {"id", "dlc", "node", "signals": {sig: raw_line}}}。

    訊息邊界僅由下一個 `BO_` 判定（BO_ 區塊間無空行），故不以空行重置 cur。
    """
    msgs: dict[str, dict] = {}
    cur = None
    bo_re = re.compile(r"^BO_\s+(\d+)\s+([A-Za-z0-9_]+)\s*:\s*(\d+)\s+(\S+)")
    sg_re = re.compile(r"^\s+SG_\s+([A-Za-z0-9_]+)")
    with path.open("r", encoding=DBC_ENCODING) as fh:
        for line in fh:
            m = bo_re.match(line)
            if m:
                cur = m.group(2)
                msgs[cur] = {
                    "id": int(m.group(1)),
                    "dlc": int(m.group(3)),
                    "node": m.group(4),
                    "signals": {},
                }
                continue
            if cur is not None:
                m = sg_re.match(line)
                if m:
                    msgs[cur]["signals"][m.group(1)] = line.rstrip("\r\n")
    return msgs


def collect_vals(path: Path) -> dict[tuple[int, str], str]:
    """VAL_ 逐字，鍵為 (msg_id, signal)。"""
    out: dict[tuple[int, str], str] = {}
    val_re = re.compile(r"^VAL_\s+(\d+)\s+([A-Za-z0-9_]+)\s")
    with path.open("r", encoding=DBC_ENCODING) as fh:
        for line in fh:
            m = val_re.match(line)
            if m:
                out[(int(m.group(1)), m.group(2))] = line.strip()
    return out


def signal_index(msgs: dict) -> dict[str, list[str]]:
    """裸訊號名 -> 含此訊號之訊息名清單（供「訊息名不符但訊號在庫」之旁證）。"""
    idx: dict[str, list[str]] = {}
    for mname, info in msgs.items():
        for sig in info["signals"]:
            idx.setdefault(sig, []).append(mname)
    return idx


def split_candidates(raw: str | None) -> list[str]:
    if not raw:
        return []
    parts = [p.strip() for p in re.split(r"[\r\n]+", raw)]
    return [p for p in parts if p]


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    hdr = find_header(ws)

    dbc_msgs = {k: parse_dbc(p) for k, p in DBCS.items()}
    dbc_vals = {k: collect_vals(p) for k, p in DBCS.items()}
    dbc_sigidx = {k: signal_index(m) for k, m in dbc_msgs.items()}

    lid_rows: dict[str, list[int]] = {}
    for r in range(hdr["data_start_row"], ws.max_row + 1):
        v = cell_text(ws, r, hdr["lid_col"])
        if v:
            lid_rows.setdefault(v.strip(), []).append(r)

    results = []
    for lid in TARGET_LIDS:
        rows = lid_rows.get(lid, [])
        entry: dict = {"lid": lid, "source": "b04 實測", "rows_found": rows}
        if not rows:
            entry.update(
                {
                    "status": "E5",
                    "disposition": "維持佔位",
                    "note": (
                        f"LID {lid!r} 於 CAN Mapping 之 Logical Identifier 欄"
                        f"（c{hdr['lid_col']}，資料列 {hdr['data_start_row']}~{ws.max_row}）"
                        "逐列 strip 後精確比對，0 命中"
                    ),
                }
            )
            results.append(entry)
            continue
        if len(rows) > 1:
            entry.update(
                {
                    "status": "E1",
                    "disposition": "維持佔位",
                    "note": f"LID 於 CAN Mapping 命中 {len(rows)} 列，非唯一，不自選",
                }
            )
            results.append(entry)
            continue

        row = rows[0]
        ah = hdr["atlantis_high_cols"]
        raw_signal = cell_text(ws, row, ah["Signal Name"])
        entry.update(
            {
                "row": row,
                "atlantis_high": {
                    "signal_name_raw": raw_signal,
                    "can": cell_text(ws, row, ah["CAN"]),
                    "format": cell_text(ws, row, ah["Format"]),
                    "sna": cell_text(ws, row, ah["SNA"]),
                    "vfs": cell_text(ws, row, ah["VFs"]),
                },
                "usage_comment": cell_text(ws, row, hdr["usage_comment_col"]),
            }
        )
        candidates = split_candidates(raw_signal)
        entry["candidates"] = candidates

        if not candidates:
            entry.update(
                {
                    "status": "E5",
                    "disposition": "維持佔位",
                    "note": (
                        f"LID 在列 {row}，但 Atlantis High 之 Signal Name"
                        f"（c{ah['Signal Name']}）為空，無候選可綁 DBC"
                    ),
                }
            )
            results.append(entry)
            continue

        checks = []
        for cand in candidates:
            if "." in cand:
                msg_name, sig_name = cand.split(".", 1)
            else:
                msg_name, sig_name = cand, None
            msg_name = msg_name.strip()
            sig_name = sig_name.strip() if sig_name else None
            per_dbc = {}
            for key, msgs in dbc_msgs.items():
                info = msgs.get(msg_name)
                in_msg = bool(info and sig_name and sig_name in info["signals"])
                d: dict = {
                    "message_present": info is not None,
                    "message_id": info["id"] if info else None,
                    "message_node": info["node"] if info else None,
                    "signal_in_message": in_msg,
                    "val_verbatim": None,
                }
                if in_msg:
                    d["signal_raw"] = info["signals"][sig_name].strip()
                    d["val_verbatim"] = dbc_vals[key].get((info["id"], sig_name))
                # 旁證：該裸訊號名於本 DBC 出現在哪些其他訊息中
                if sig_name:
                    d["signal_seen_in_messages"] = sorted(dbc_sigidx[key].get(sig_name, []))
                per_dbc[key] = d
            checks.append(
                {"candidate": cand, "message": msg_name, "signal": sig_name, "dbc": per_dbc}
            )
        entry["candidate_checks"] = checks

        hits = []
        for ch in checks:
            for key, d in ch["dbc"].items():
                if d["message_present"] and d["signal_in_message"]:
                    hits.append(
                        {
                            "candidate": ch["candidate"],
                            "dbc": key,
                            "message": ch["message"],
                            "signal": ch["signal"],
                            "message_id": d["message_id"],
                            "message_node": d["message_node"],
                            "val_verbatim": d["val_verbatim"],
                            "signal_raw": d.get("signal_raw"),
                        }
                    )
        entry["hits"] = hits
        ics_hits = [h for h in hits if h["message_node"] == "ICS"]
        multi = len(candidates) >= 2
        all_present = bool(hits) and len(hits) >= len(candidates)
        entry["ics13_context"] = {
            "candidate_count": len(candidates),
            "hit_count": len(hits),
            "ics_node_hit_count": len(ics_hits),
            "applicable": bool(multi and all_present),
            "reason": (
                "多名並列且皆在庫，R-ICS13 之取捨情境成立"
                if (multi and all_present)
                else (
                    "單名候選，非 R-ICS13 之『一格多名』情境，不套用取 ICS"
                    if not multi
                    else "多名但非皆在庫，非 R-ICS13 之『皆查有』情境，不套用取 ICS"
                )
            ),
        }

        if not hits:
            entry.update(
                {
                    "status": "E5",
                    "disposition": "維持佔位",
                    "primary": None,
                    "fallbacks": [],
                    "val_verbatim": None,
                    "note": "LID 有列且有候選，但二 DBC 皆查無該 BO_/SG_ 綁定",
                }
            )
        elif multi and all_present and not ics_hits:
            entry.update(
                {
                    "status": "E1",
                    "disposition": "維持佔位",
                    "primary": None,
                    "fallbacks": hits,
                    "val_verbatim": None,
                    "note": "多名並列且皆在庫但無一發送節點為 ICS，不自選（R-ICS13 未涵蓋）",
                }
            )
        else:
            # R-ICS13 適用者取 ICS 節點；不適用者取唯一/首個在庫命中，如實記節點
            primary = ics_hits[0] if ics_hits else hits[0]
            entry.update(
                {
                    "status": "RESOLVED",
                    "disposition": "可綁定",
                    "primary": primary,
                    "fallbacks": [h for h in hits if h is not primary],
                    "val_verbatim": primary["val_verbatim"] or "無 VAL_",
                    "primary_node_is_ics": primary["message_node"] == "ICS",
                }
            )
        results.append(entry)

    # 併入 b03 八筆（沿用 b03 實測，不重新量測）
    b03 = json.loads(B03_MAP.read_text(encoding="utf-8"))
    carried = []
    for e in b03["lids"]:
        e = dict(e)
        e["source"] = "沿用 b03 實測（未重新量測）"
        e["source_file"] = str(B03_MAP.relative_to(REPO))
        carried.append(e)

    doc = {
        "generated_by": "features/ics_management/scripts/lid_dbc_probe_b04.py",
        "purpose": "DR-ICS15 四訊號 LID->CAN 解析；累計對照表（b04 四筆 + b03 八筆）",
        "ruling_basis": "R-ICS13（適用性逐筆判定，見各筆 ics13_context）",
        "scan_conditions": {
            "xlsx": str(XLSX.relative_to(REPO)),
            "sheet": SHEET,
            "header_self_verified": hdr,
            "sheet_max_row": ws.max_row,
            "sheet_max_column": ws.max_column,
            "dbc_encoding": DBC_ENCODING,
            "dbc": {
                key: {
                    "path": str(p.relative_to(REPO)),
                    "sha256": sha256(p),
                    "message_count": len(dbc_msgs[key]),
                }
                for key, p in DBCS.items()
            },
            "method_notes": [
                "DBC 以 latin-1 開檔；以 UTF-8 讀而得之『查無』不算數（A-ICS25）",
                "BO_ 區塊間無空行，訊息邊界僅由下一個 BO_ 判定；SG_ 必須在該訊息界內",
                "LID 表頭列號/欄號本輪自驗，未沿用前包欄號",
            ],
        },
        "lids_b04": results,
        "lids_b03_carried": carried,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT}")
    for e in results:
        p = e.get("primary")
        print(
            e["lid"],
            "row=" + str(e.get("row")),
            e["status"],
            (p["candidate"] + " @" + p["dbc"] + " id=" + str(p["message_id"]) + " node=" + p["message_node"]) if p else "-",
            "VAL_" if e.get("val_verbatim") not in (None, "無 VAL_") else "no-VAL_",
        )


if __name__ == "__main__":
    main()
