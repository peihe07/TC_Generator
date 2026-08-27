#!/usr/bin/env python3
"""丙″ 之可行性驗證（下放包 29 §2.4，T152）。

丙″ ＝ 丙′ **減**圖片還原、**加** `xmlns:xr` 補宣告。

下放包 29 §2.2 之裁定：圖片還原一項**多餘** —— 障礙 1 之 `KeyError`
只發生在還原圖片之變體，而變體 B 根本沒碰圖片；其失敗另有原因（障礙 2）。
§2.3 之裁定：障礙 2 採**甲**（補宣告），不採乙（剝除 `xr:uid`）——
修復應朝「還原母本狀態」，不朝「讓片段適應被破壞的輸出」。

**只驗可行性，不產出交付本。** 3 筆假資料，全程 `/tmp`。
**母本不得被開啟寫入。** 第 6 步任一項不符即停，不自行調整方案。
"""
import csv
import hashlib
import json
import re
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / ("inputs/FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT STLA "
                 "Test Case Specification & Result_SWQT_20260817_ext.xlsx")
TMP = Path("/tmp/vc_writeback_c2")
SHEET = "Test Case Specification 測試用例規範"
COLS = {"C": "polarion_id", "D": "leaf_id", "G": "test_group", "H": "test_set",
        "I": "test_item", "J": "pre_conditions", "K": "input_test_data",
        "L": "test_procedure", "M": "expected_result",
        "N": "specification_reference", "P": "priority", "R": "design_method",
        "S": "functional_safety"}


def sheet_xml(zf, name):
    wbxml = zf.read("xl/workbook.xml").decode("utf-8")
    rid = re.search(rf'<sheet[^>]*name="{re.escape(name)}"[^>]*r:id="([^"]+)"',
                    wbxml).group(1)
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    for tag in re.findall(r"<Relationship [^>]*/>", rels):
        if re.search(rf'Id="{rid}"', tag):
            t = re.search(r'Target="([^"]+)"', tag).group(1).lstrip("/")
            return t if t.startswith("xl/") else "xl/" + t
    raise KeyError(rid)


def facts(path):
    z = zipfile.ZipFile(path)
    sx = sheet_xml(z, SHEET)
    d = z.read(sx).decode("utf-8", "ignore")
    n = z.namelist()
    return {"x14_dv": len(re.findall(r"<x14:dataValidation[ >]", d)),
            "extLst": len(re.findall(r"<extLst>", d)),
            "std_dv": len(re.findall(r"<dataValidation[ >]", d)),
            "printer": len([x for x in n if "printerSettings" in x
                            and x.endswith(".bin")]),
            "media": sorted(x for x in n if x.startswith("xl/media/")),
            "sheet_xml": sx, "size": path.stat().st_size}


def main():
    if TMP.exists():
        shutil.rmtree(TMP)
    TMP.mkdir(parents=True)
    src = TMP / "master_copy.xlsx"
    shutil.copy(MASTER, src)
    before = hashlib.sha256(MASTER.read_bytes()).hexdigest()

    tcs = json.loads((ROOT / "generated/pilot_glovebox.json")
                     .read_text("utf-8"))["tcs"][:3]
    recon = {r["req_id"]: r for r in csv.DictReader(
        (ROOT / "data/recon_leaf_to_section.tsv").open(encoding="utf-8"),
        delimiter="\t")}
    for t in tcs:
        t["polarion_id"] = recon[t["leaf_id"]]["polarion_id"]

    print("步驟 1 —— openpyxl 開副本、寫入 14 欄、save")
    wb = openpyxl.load_workbook(src)
    ws = wb[SHEET]
    for i, t in enumerate(tcs):
        for col, key in COLS.items():
            ws[f"{col}{10 + i}"] = t[key]
    out = TMP / "step1_openpyxl.xlsx"
    wb.save(out)
    wb.close()
    fm, f1 = facts(src), facts(out)
    print(f"  母本  x14={fm['x14_dv']} extLst={fm['extLst']} std={fm['std_dv']} "
          f"printer={fm['printer']} media={len(fm['media'])} size={fm['size']}")
    print(f"  step1 x14={f1['x14_dv']} extLst={f1['extLst']} std={f1['std_dv']} "
          f"printer={f1['printer']} media={len(f1['media'])} size={f1['size']}")

    print("步驟 2 —— 解壓輸出檔與母本")
    zm, zo = zipfile.ZipFile(src), zipfile.ZipFile(out)
    print(f"  母本 {len(zm.namelist())} 項；step1 {len(zo.namelist())} 項")

    print("步驟 3 —— 注入 extLst ＋ printerSettings（**不動 media 與 drawing**）")
    dm = zm.read(fm["sheet_xml"]).decode("utf-8", "ignore")
    ext = re.search(r"<extLst>.*?</extLst>", dm, re.S)
    doc = zo.read(f1["sheet_xml"]).decode("utf-8", "ignore")
    doc = re.sub(r"</worksheet>\s*$", ext.group(0) + "</worksheet>", doc)
    printer = [n for n in zm.namelist()
               if "printerSettings" in n and n not in zo.namelist()]
    print(f"  extLst 注入：{'是' if ext else '**否**'}；"
          f"自母本補回 printerSettings 相關 {len(printer)} 項")

    print("步驟 4 —— 於輸出之 <worksheet> 根元素補宣告 xmlns:xr（取母本同一字串）")
    decl = re.search(r'xmlns:xr="[^"]+"', re.search(r"<worksheet\b[^>]*>", dm)
                     .group(0)).group(0)
    root = re.search(r"<worksheet\b[^>]*?>", doc).group(0)
    if "xmlns:xr=" in root:
        print("  輸出根元素已有 xmlns:xr，不重複補")
    else:
        doc = doc.replace(root, root[:-1] + " " + decl + ">", 1)
        print(f"  已補：{decl}")

    print("步驟 5 —— 重打包")
    final = TMP / "step5_repacked.xlsx"
    with zipfile.ZipFile(final, "w", zipfile.ZIP_DEFLATED) as zf:
        for n in zo.namelist():
            zf.writestr(n, doc if n == f1["sheet_xml"] else zo.read(n))
        for n in printer:
            zf.writestr(n, zm.read(n))
    print(f"  {final.name}  {final.stat().st_size} bytes")

    print("\n步驟 6 —— 六項驗收")
    ff = facts(final)
    chk = []
    chk.append(("x14:dataValidation 1 條", ff["x14_dv"] == fm["x14_dv"] == 1,
                f"母本 {fm['x14_dv']} → 修復後 {ff['x14_dv']}"))
    chk.append(("extLst 1 個", ff["extLst"] == fm["extLst"] == 1,
                f"母本 {fm['extLst']} → 修復後 {ff['extLst']}"))
    chk.append(("標準 dataValidation 3 條", ff["std_dv"] == fm["std_dv"] == 3,
                f"母本 {fm['std_dv']} → 修復後 {ff['std_dv']}"))
    chk.append(("printerSettings 5 個", ff["printer"] == fm["printer"] == 5,
                f"母本 {fm['printer']} → 修復後 {ff['printer']}"))
    opened, err = True, ""
    bad, b10 = [], None
    try:
        wb2 = openpyxl.load_workbook(final)
        ws2 = wb2[SHEET]
        for i, t in enumerate(tcs):
            for col, key in COLS.items():
                if ws2[f"{col}{10 + i}"].value != t[key]:
                    bad.append(f"{col}{10 + i}")
        b10 = ws2["B10"].value
        wb2.close()
    except Exception as e:                                  # noqa: BLE001
        opened, err = False, f"{type(e).__name__}: {e}"
    chk.append((f"資料 14 欄 × 3 筆逐格與 JSON 相符（B10 公式 {b10!r}）",
                opened and not bad and isinstance(b10, str)
                and b10.startswith("="),
                f"不符 {len(bad)} 格 {bad[:6] or '無'}" if opened else err))
    chk.append(("可被 openpyxl 開啟", opened, err or "無例外"))

    print(f"{'#':>2}  {'項':<46} 判")
    print("-" * 84)
    failed = 0
    for i, (name, ok, detail) in enumerate(chk, 1):
        failed += 0 if ok else 1
        print(f"{i:>2}  {name:<46} {'PASS' if ok else '**FAIL**'}")
        print(f"    {detail}")
    print("-" * 84)

    # 第二讀者 —— LibreOffice headless。**非 Excel**，其通過不等於 Excel 不報修復。
    so = shutil.which("soffice")
    if so:
        r = subprocess.run([so, "--headless", "--convert-to", "csv",
                            "--outdir", str(TMP / "lo"), str(final)],
                           capture_output=True, text=True, timeout=180)
        made = list((TMP / "lo").glob("*.csv")) if (TMP / "lo").exists() else []
        print(f"第二讀者 LibreOffice headless：返回碼 {r.returncode}；"
              f"產出 {[p.name for p in made]}")
        print("  ⚠ **此非 Excel** —— 其通過只證明套件結構可被另一個實作讀出，"
              "**不證明 Excel 開啟不報修復**。")
    else:
        print("第二讀者：環境無 soffice，**未驗**")

    print("\n步驟 7 —— 圖片之變化與實害評估（§2.2）")
    print(f"  母本 media {fm['media']}")
    print(f"  修復後 media {ff['media']}")
    print(f"  檔案大小 母本 {fm['size']} → 修復後 {ff['size']} "
          f"（{ff['size'] - fm['size']:+d}）")

    after = hashlib.sha256(MASTER.read_bytes()).hexdigest()
    print(f"\n母本 SHA256 前 {before[:16]} → 後 {after[:16]}  "
          f"{'**未變**' if before == after else '**⚠ 母本被改動**'}")
    print(f"{len(chk)} checked / {failed} failed")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
