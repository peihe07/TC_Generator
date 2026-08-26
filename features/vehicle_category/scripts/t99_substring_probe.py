#!/usr/bin/env python3
"""T99 —— 第 7b 項之候選判準：整段子串比對（下放包 18 §二）。

**只跑不判**（§2.3 第 1 步）：對既有 34 筆（pilot 12 ＋ 第 1 批 22）
計算其 `test_item` 上半是否為三來源之逐字子串，逐筆回報，
**不改變其收斂狀態、不改判準、不改 TC**。

判準（候選）：
    上半（去首尾空白、首字母大小寫正規化）
      ∈ substring( 037 Title ∪ 037 Description ∪ SYS1 對應節之全文 )

依據：R-S4 上半為「需求／規格原句 verbatim」；R-3 得摘句；
R-4 得將句首字母轉大寫。**故其應為某來源字串之逐字子串。**
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
A03 = ROOT / "inputs/FM-WI-FSM-037-A03-N1L-SWE1-VehicleCategory-HMI-V0.1 STLA 報告.xlsx"
S1 = ROOT / ("inputs/SYS1_HMI_Vehicle_Category_HMI_Logic_and_Flow_"
             "R1_SR24_Post_2A_(December_27_2023).xlsx")
BATCHES = ["generated/pilot_glovebox.json",
           "generated/batch1_category_structure.json"]


def load_037():
    wb = openpyxl.load_workbook(A03, read_only=True, data_only=True)
    out = {}
    for r in list(wb["Analysis Report"].iter_rows(values_only=True))[7:]:
        if r[0] not in (None, ""):
            out[str(r[0]).strip()] = (str(r[3]).strip(), str(r[4]).strip())
    return out


def load_sys1():
    """回傳 {章節號: Description 全文}。"""
    wb = openpyxl.load_workbook(S1, read_only=True, data_only=True)
    rows = list(wb["Basic Report"].iter_rows(values_only=True))
    h = [str(c).strip() if c else "" for c in rows[0]]
    oi, di = h.index("Outline Number"), h.index("Description")
    out = {}
    for r in rows[1:]:
        o = str(r[oi]).strip() if r[oi] else ""
        if o:
            out[o] = ((str(r[di]) if r[di] else "")
                      .replace("_x000D_\n", "\n").replace("_x000D_", " "))
    return out


def variants(s: str):
    """R-4 之首字母正規化：原樣、首字大寫、首字小寫，三者皆試。"""
    s = s.strip()
    if not s:
        return {s}
    return {s, s[0].upper() + s[1:], s[0].lower() + s[1:]}


def main():
    src, sys1 = load_037(), load_sys1()
    rows, bad = [], []
    for b in BATCHES:
        p = ROOT / b
        if not p.exists():
            continue
        J = json.loads(p.read_text("utf-8"))
        for t in J["tcs"]:
            lid = t["leaf_id"]
            top = t["test_item"].split("\n\n")[0].strip()
            title, desc = src.get(lid, ("", ""))
            sec = t["specification_reference"].rsplit("_", 1)[-1]
            s1 = sys1.get(sec, "")
            hit = None
            for v in variants(top):
                if v and v in title:
                    hit = "037 Title"
                    break
                if v and v in desc:
                    hit = "037 Description"
                    break
                if v and v in s1:
                    hit = f"SYS1 §{sec}"
                    break
            rows.append((Path(b).stem, lid, hit, top))
            if hit is None:
                bad.append((Path(b).stem, lid, top))

    print("T99 —— 第 7b 項候選判準（整段子串）之只跑不判")
    print(f"母體: {len(rows)} 筆（pilot 12 ＋ 第 1 批 22）\n")
    print(f"{'批':<28}{'leaf':<22}{'命中來源':<18}判")
    print("-" * 78)
    for b, lid, hit, _ in rows:
        print(f"{b:<28}{lid.replace('SWE1-HMI-VC-', ''):<22}"
              f"{(hit or '—'):<18}{'PASS' if hit else '**FAIL**'}")
    print("-" * 78)
    print(f"{len(rows)} 筆 / 不通過 {len(bad)} 筆")
    if bad:
        print("\n=== 不通過者之上半全文（供裁定）")
        for b, lid, top in bad:
            title, desc = src.get(lid, ("", ""))
            print(f"\n--- {lid}  [{b}]")
            print(f"    上半 : {top}")
            print(f"    Title: {title}")
            print(f"    Desc : {desc}")
    return len(bad)


if __name__ == "__main__":
    main()
    sys.exit(0)
