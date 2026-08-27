#!/usr/bin/env python3
"""126 筆之寫回 —— 工作版（丙″，下放包 34 §四 ＋ Pei 2026-08-27 授權）。

**工作版之定義**（下放包 34 §四）：14 處 PENDING 依 IN §8.4.3 內嵌佔位、
表 B 仍為草稿。**出貨版仍卡 PENDING 結案**，此門檻不因本檔改變。

方法為**丙″**（上繳包 29，六項全過）：
  1. openpyxl 開母本副本、寫入 14 欄、save
  2. 自母本注入 `<extLst>`（含 x14 dataValidation）與 printerSettings
  3. 於輸出之 `<worksheet>` 根元素補宣告 `xmlns:xr`
  4. 重打包
  5. 驗六項 ＋ 126 列逐格

**母本不得被寫入** —— 全程 `read_bytes()` 與唯讀載入，SHA256 前後實測。

TC ID 依 **R-VC28**：`newR1L-VC-001` … `newR1L-VC-126`，
**依工作簿列序**（＝ spec section 序）連續，拆分筆各佔一獨立序號。
"""
import csv
import hashlib
import json
import re
import shutil
import sys
import zipfile
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / ("inputs/FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT STLA "
                 "Test Case Specification & Result_SWQT_20260817_ext.xlsx")
OUTDIR = ROOT / "output"
SHEET = "Test Case Specification 測試用例規範"
FIRST_ROW = 10

# 上繳包 27 §3.2 之映射。**B 欄為公式，不得覆寫**；
# E／O／Q／T–Z／AA／AB–AH 無來源，留空。
COLS = {"C": "polarion_id", "D": "leaf_id", "F": "tc_id",
        "G": "test_group", "H": "test_set", "I": "test_item",
        "J": "pre_conditions", "K": "input_test_data", "L": "test_procedure",
        "M": "expected_result", "N": "specification_reference",
        "P": "priority", "R": "design_method", "S": "functional_safety"}

BATCH_ORDER = ["pilot_glovebox", "batch1_category_structure",
               "batch2_settings_list", "batch3_controls",
               "batch4_settings_behavior", "batch5_ignition_availability",
               "batch6_brake_service", "batch7_cabrio_widget"]


def outline_key(o):
    return tuple(int(x) for x in o.split("."))


def collect():
    """126 筆，依 spec section 序 → leaf_id 序 → 批內原序。"""
    recon = {r["req_id"]: r for r in csv.DictReader(
        (ROOT / "data/recon_leaf_to_section.tsv").open(encoding="utf-8"),
        delimiter="\t")}
    rows = []
    on_disk = {p.stem for p in (ROOT / "generated").glob("*.json")}
    missing = on_disk - set(BATCH_ORDER)
    if missing:
        raise SystemExit(f"**BATCH_ORDER 未涵蓋之批檔 {sorted(missing)}** —— 停")
    for name in BATCH_ORDER:
        d = json.loads((ROOT / "generated" / f"{name}.json").read_text("utf-8"))
        for i, t in enumerate(d["tcs"]):
            rows.append({**t, "polarion_id": recon[t["leaf_id"]]["polarion_id"],
                         "_o": recon[t["leaf_id"]]["outline"], "_i": i})
    rows.sort(key=lambda r: (outline_key(r["_o"]), r["leaf_id"], r["_i"]))
    for n, r in enumerate(rows, 1):                    # R-VC28
        r["tc_id"] = f"newR1L-VC-{n:03d}"
    return rows


def sheet_xml(zf, name):
    wb = zf.read("xl/workbook.xml").decode("utf-8")
    rid = re.search(rf'<sheet[^>]*name="{re.escape(name)}"[^>]*r:id="([^"]+)"',
                    wb).group(1)
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    for tag in re.findall(r"<Relationship [^>]*/>", rels):
        if re.search(rf'Id="{rid}"', tag):
            t = re.search(r'Target="([^"]+)"', tag).group(1).lstrip("/")
            return t if t.startswith("xl/") else "xl/" + t
    raise KeyError(rid)


def facts(p):
    z = zipfile.ZipFile(p)
    d = z.read(sheet_xml(z, SHEET)).decode("utf-8", "ignore")
    n = z.namelist()
    return {"x14": len(re.findall(r"<x14:dataValidation[ >]", d)),
            "ext": len(re.findall(r"<extLst>", d)),
            "std": len(re.findall(r"<dataValidation[ >]", d)),
            "prn": len([x for x in n if "printerSettings" in x
                        and x.endswith(".bin")]),
            "sx": sheet_xml(z, SHEET)}


def main():
    rows = collect()
    OUTDIR.mkdir(exist_ok=True)
    before = hashlib.sha256(MASTER.read_bytes()).hexdigest()
    tmp = OUTDIR / "_step1.xlsx"
    shutil.copy(MASTER, tmp)

    print(f"步驟 1 —— 寫入 {len(rows)} 列 × 14 欄")
    wb = openpyxl.load_workbook(tmp)
    ws = wb[SHEET]
    for i, r in enumerate(rows):
        for col, key in COLS.items():
            ws[f"{col}{FIRST_ROW + i}"] = r[key]
    step1 = OUTDIR / "_step1_written.xlsx"
    wb.save(step1)
    wb.close()

    print("步驟 2／3 —— 注入 extLst ＋ printerSettings，補宣告 xmlns:xr")
    zm, zo = zipfile.ZipFile(tmp), zipfile.ZipFile(step1)
    fm, f1 = facts(tmp), facts(step1)
    dm = zm.read(fm["sx"]).decode("utf-8", "ignore")
    ext = re.search(r"<extLst>.*?</extLst>", dm, re.S)
    doc = zo.read(f1["sx"]).decode("utf-8", "ignore")
    doc = re.sub(r"</worksheet>\s*$", ext.group(0) + "</worksheet>", doc)
    decl = re.search(r'xmlns:xr="[^"]+"',
                     re.search(r"<worksheet\b[^>]*>", dm).group(0)).group(0)
    root = re.search(r"<worksheet\b[^>]*?>", doc).group(0)
    if "xmlns:xr=" not in root:
        doc = doc.replace(root, root[:-1] + " " + decl + ">", 1)
    printer = [n for n in zm.namelist()
               if "printerSettings" in n and n not in zo.namelist()]

    print("步驟 4 —— 重打包")
    stamp = date.today().strftime("%Y%m%d")
    final = OUTDIR / (f"FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT STLA "
                      f"Test Case Specification & Result_SWQT_"
                      f"VehicleCategory_{stamp}_working.xlsx")
    with zipfile.ZipFile(final, "w", zipfile.ZIP_DEFLATED) as zf:
        for n in zo.namelist():
            zf.writestr(n, doc if n == f1["sx"] else zo.read(n))
        for n in printer:
            zf.writestr(n, zm.read(n))
    zm.close()
    zo.close()
    tmp.unlink()
    step1.unlink()

    print("\n步驟 5 —— 驗收")
    ff = facts(final)
    chk = [("x14:dataValidation 1 條", ff["x14"] == fm["x14"] == 1,
            f"母本 {fm['x14']} → 輸出 {ff['x14']}"),
           ("extLst 1 個", ff["ext"] == fm["ext"] == 1,
            f"母本 {fm['ext']} → 輸出 {ff['ext']}"),
           ("標準 dataValidation 3 條", ff["std"] == fm["std"] == 3,
            f"母本 {fm['std']} → 輸出 {ff['std']}"),
           ("printerSettings 5 個", ff["prn"] == fm["prn"] == 5,
            f"母本 {fm['prn']} → 輸出 {ff['prn']}")]
    wb2 = openpyxl.load_workbook(final)
    ws2 = wb2[SHEET]
    bad = []
    for i, r in enumerate(rows):
        for col, key in COLS.items():
            if ws2[f"{col}{FIRST_ROW + i}"].value != r[key]:
                bad.append(f"{col}{FIRST_ROW + i}")
    b10 = ws2[f"B{FIRST_ROW}"].value
    after_last = ws2[f"D{FIRST_ROW + len(rows)}"].value
    wb2.close()
    chk.append((f"{len(rows)} 列 × 14 欄逐格與 JSON 相符",
                not bad, f"不符 {len(bad)} 格 {bad[:5] or '無'}"))
    chk.append(("B 欄公式未覆寫", isinstance(b10, str) and b10.startswith("="),
                f"B{FIRST_ROW} = {b10!r}"))
    chk.append((f"第 {FIRST_ROW + len(rows)} 列（末列之後）仍為空",
                after_last in (None, ""), f"D 欄值 {after_last!r}"))
    ids = [r["tc_id"] for r in rows]
    chk.append(("TC ID 連續且無重複（R-VC28）",
                ids == [f"newR1L-VC-{i:03d}" for i in range(1, len(rows) + 1)],
                f"{ids[0]} … {ids[-1]}；相異 {len(set(ids))}"))

    print(f"{'#':>2}  {'項':<42} 判")
    print("-" * 80)
    failed = 0
    for i, (name, ok, det) in enumerate(chk, 1):
        failed += 0 if ok else 1
        print(f"{i:>2}  {name:<42} {'PASS' if ok else '**FAIL**'}")
        print(f"    {det}")
    print("-" * 80)
    after = hashlib.sha256(MASTER.read_bytes()).hexdigest()
    print(f"母本 SHA256 前 {before[:16]} → 後 {after[:16]}  "
          f"{'**未變**' if before == after else '**⚠ 被改動**'}")
    print(f"{len(chk)} checked / {failed} failed")
    print(f"\n輸出：{final.relative_to(ROOT)}  ({final.stat().st_size} bytes)")
    print("**工作版** —— 14 處 PENDING 內嵌，表 B 仍為草稿；出貨版仍卡 PENDING 結案。")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
