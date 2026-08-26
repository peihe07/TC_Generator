#!/usr/bin/env python3
"""CONT 表之二層防護（下放包 20a）。

`CONT` 表（`data/cont_table.tsv`）記「037 之片段出自 SYS1 之完整句」者，
供收斂條件第 16 項比對。上繳包 16 §9 已揭露其二個風險：

  1. **該進表的沒進表** —— 新批次之續行型／指涉型 leaf 未登記，
     第 16 項對其 N/A，片段上半無人攔。
  2. **表之內容錯** —— 登記之 SYS1 節指錯，第 16 項拿錯的來源比對。

本檔以二層解之：第一層候選偵測（風險 1）、第二層內容驗證（風險 2）。

**第三層（下放包 23 §2.2）**：`resolved-by-structure` 之聲稱驗證 ——
登記 `resolution=PC`／`Step` 者，其 `resolution_key` 須真的出現於
該 leaf 之 TC 之對應欄位。**聲稱「結構會解」而結構裡沒有 → FAIL。**
第三個風險至此有承載者：**判準之擴充自帶其檢查**。

**self-test 前置**（PLAYBOOK §7.1.1，下放包 20 §1.1）——
main 先跑八個斷言，任一不過即非零碼退出且**不輸出正式結果**。
「先」由紀律變為程式結構。
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
A03 = ROOT / "inputs/FM-WI-FSM-037-A03-N1L-SWE1-VehicleCategory-HMI-V0.1 STLA 報告.xlsx"
S1 = ROOT / ("inputs/SYS1_HMI_Vehicle_Category_HMI_Logic_and_Flow_"
             "R1_SR24_Post_2A_(December_27_2023).xlsx")
CONT = ROOT / "data/cont_table.tsv"
EXCL = ROOT / "data/cont_exclusions.tsv"
# 列冊 —— 候選已被看見且歸屬於未來批次，其判定隨該批勘查為之。
# 與排除清單之別：排除者為「已判定非此類」，列冊者為「已看見、待判定」。
# 二者皆使「已考慮過」與「沒看到」可區分（20a §2.1）。
DEFER = ROOT / "data/cont_deferred.tsv"

# 條號前綴樣式表 —— **第一層 (a) 與第二層共用同一張**（20a §2.2 註：
# 二處不得分歧，承機讀行之同一原則）。
PREFIX = re.compile(r"^\s*(?:[A-Z]{1,4}\d[\d.]*\s*\)|\d[\d.]*\s*\))\s*")
PRONOUN = re.compile(r"^(It|They|This|These|That|Those)\b")
# `short` 已自 CONT 候選移除（下放包 21 §三）——
# 其對 CONT 而言是偽陽性（完整短句非片段），對 profile §8 才是真陽性候選。
# 本檔仍掃它，但**輸出至 `data/short_source_leaves.tsv`**，不進候選判定。
# 60 字元之耦合隨之解除：CONT 不再用該數字，該清單即 §8 之清單。
SHORT = 60
SHORT_OUT = ROOT / "data/short_source_leaves.tsv"
GEN = ROOT / "generated"

# 第三層之欄位對照 —— `resolution` 之值決定**看哪個欄位**。
# 這張表是第三層之全部語意：`PC` 看 pre_conditions、`Step` 看 test_procedure。
# 分開查是本層之要義 —— 若改為「整份 TC 的 JSON 裡有沒有這個詞」，
# 則聲稱 `Step` 而其實只寫在 PC 者會被放過，
# **而那正是層次 2 所要區分的二種承載方式**（self-test 8 即測此）。
RES_FIELD = {"PC": "pre_conditions", "Step": "test_procedure"}


def strip_prefix(s):
    return PREFIX.sub("", s.strip()).strip()


def norm(s):
    """**沿用第 7b 項之同一正規化**（下放包 21 §二，丁案）：
    去條號前綴 ＋ 小寫 ＋ 去標點 ＋ 壓縮空白。

    全案自此只有一套正規化 —— 第二層原本自帶「首字母正規化」，
    與第 7b 之定義是二套；丁之後合一（承機讀行與前綴樣式表之同一原則：
    單一來源，二處必分歧）。

    `013-02` 之改寫（`, and` → `.`）**恰好全落在標點層** ——
    去標點後其片段回復為 SYS1 之純前段，子串成立。
    """
    s = strip_prefix(s).lower()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def load_037():
    wb = openpyxl.load_workbook(A03, read_only=True, data_only=True)
    out = {}
    for r in list(wb["Analysis Report"].iter_rows(values_only=True))[7:]:
        if r[0] not in (None, ""):
            out[str(r[0]).strip()] = (str(r[2]).split("\n")[0].strip(),
                                      str(r[4]).strip())
    return out


def load_sys1():
    wb = openpyxl.load_workbook(S1, read_only=True, data_only=True)
    rows = list(wb["Basic Report"].iter_rows(values_only=True))
    h = [str(c).strip() if c else "" for c in rows[0]]
    oi, di = h.index("Outline Number"), h.index("Description")
    out = {}
    for r in rows[1:]:
        o = str(r[oi]).strip() if r[oi] else ""
        if o:
            out[o] = ((str(r[di]) if r[di] else "")
                      .replace("_x000D_\n", "\n").replace("_x000D_", " "))
    return out


def read_tsv(p):
    if not p.exists():
        return []
    lines = p.read_text("utf-8").splitlines()
    if len(lines) < 2:
        return []
    head = lines[0].split("\t")
    return [dict(zip(head, ln.split("\t"))) for ln in lines[1:] if ln.strip()]


def classify(desc):
    """回傳命中之候選特徵（可多個），空 list 表非候選。"""
    s = strip_prefix(desc)
    hits = []
    if s and s[0].islower():
        hits.append("continuation")
    if PRONOUN.match(s):
        hits.append("reference")
    return hits                      # `short` 已移出（下放包 21 §三）


def layer1(src, cont_leaves, excl_leaves, defer_leaves=frozenset()):
    """候選偵測 —— 候選 ∉ CONT ∧ ∉ 排除清單 ∧ ∉ 列冊 → 未處置。"""
    unhandled, all_cand = [], []
    for lid, (_, desc) in sorted(src.items()):
        hits = classify(desc)
        if not hits:
            continue
        all_cand.append((lid, hits, strip_prefix(desc)[:70]))
        if (lid not in cont_leaves and lid not in excl_leaves
                and lid not in defer_leaves):
            unhandled.append((lid, hits, strip_prefix(desc)[:70]))
    return all_cand, unhandled


SENT_SPLIT = re.compile(r"(?<=\.)\s+(?=[A-Z])")


def sentence(sys1_text, idx):
    """取指定句；支援單句 `3`、**範圍 `1-2`**、`*`／空值取整段。

    範圍為下放包 22 §二所明文（「reference 型：登記為**範圍**或 `*`」）——
    指涉型之先行詞常在前句，取單句不足以解其指涉，取整段又可能逾
    R-3 之 50 token。
    """
    if idx in ("*", "", None):
        return sys1_text
    parts = [s.strip() for s in SENT_SPLIT.split(sys1_text.strip())]
    m = re.fullmatch(r"(\d+)\s*-\s*(\d+)", str(idx).strip())
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if 1 <= a <= b <= len(parts):
            return " ".join(parts[a - 1:b])
        return ""
    try:
        n = int(idx)
    except ValueError:
        return sys1_text
    return parts[n - 1] if 1 <= n <= len(parts) else ""


def layer2(src, sys1, cont_rows):
    """內容驗證 —— 037 片段（正規化）須為登記**句**之子串。

    **句級（下放包 22 §二）**：句序由程式硬推改為表中 `sentence_index`
    登記，驗證隨之自節級細化至句級 —— 指錯句則子串關係不成立
    （同節他句之文字不同）。**sentence_index 之正確性至此有機器承載者。**
    """
    bad = []
    for row in cont_rows:
        lid, sec = row["leaf"], row["sys1_section"]
        idx = row.get("sentence_index", "*")
        raw = sys1.get(sec, "")
        if not raw:
            bad.append((lid, sec, idx, "節不存在"))
            continue
        tgt = norm(sentence(raw, idx))
        frag = norm(src.get(lid, ("", ""))[1])
        if not frag or not tgt or frag not in tgt:
            bad.append((lid, f"{sec} s{idx}", frag[:44],
                        "指定句不存在" if not tgt else "片段非該句之子串"))
    return bad


def load_generated():
    """generated/*.json → leaf_id → [tc, ...]（拆分者多筆）。"""
    out = {}
    for f in sorted(GEN.glob("*.json")):
        for t in json.loads(f.read_text("utf-8")).get("tcs", []):
            out.setdefault(t["leaf_id"], []).append(t)
    return out


def layer3(cont_rows, by_leaf):
    """`resolved-by-structure` 之聲稱驗證（profile §9.2／§9.4）。

    回傳 `(bad, applied, pending)`：
      bad     —— 聲稱與 TC 結構不符者
      applied —— 已生成且已驗者
      pending —— 已登記但該 leaf 尚未生成（**非 FAIL，但須顯示**）

    `pending` 獨立成一態之理由同 20a §2.1 之三態：
    「尚未生成」與「驗過了」不得看起來一樣。
    """
    bad, applied, pending = [], [], []
    for row in cont_rows:
        res = (row.get("resolution") or "").strip()
        if not res:
            continue
        lid = row["leaf"]
        key = (row.get("resolution_key") or "").strip()
        head = res.split("-")[0]
        if not key or head not in RES_FIELD:
            bad.append((lid, res, key, "登記不完整（resolution 未知或 key 空）"))
            continue
        tcs = by_leaf.get(lid)
        if not tcs:
            pending.append((lid, res, key))
            continue
        field = RES_FIELD[head]
        m = re.fullmatch(r"\w+-(\d+)", res)
        for t in tcs:
            text = t.get(field, "")
            if m:                        # `Step-n` —— 限該步驟
                n = int(m.group(1))
                lines = [x for x in text.split("\n") if x.strip()]
                text = lines[n - 1] if 1 <= n <= len(lines) else ""
            if key not in text:          # **逐字**，不做大小寫／連字號寬鬆（profile §9.3）
                bad.append((lid, res, key, f"{field} 未含該 key"))
            else:
                applied.append((lid, res, key))
    return bad, applied, pending


# 第三層之 self-test 夾具 —— **TC 之欄位形狀**，非裸字串。
# 用真實欄位名之理由：若第三層查錯欄位（如查 `preconditions`），
# 裸字串夾具測不出來，欄位形狀之夾具測得出來。
FIX_TC = {
    "leaf_id": "FIXTURE-01",
    "pre_conditions": "1. The language-change pop-up is displayed",
    "test_procedure": "1. Press the X button on the dialog\n2. Record the screen",
}


def self_test(src, sys1):
    """五個斷言（20a §2.4 ＋ 下放包 22 §二之錯句斷言）。任一不過即非零碼退出。"""
    ok = True
    # 第一層 (b) 已知標的：019-02 為指涉型，未登記時應被列為候選
    h = classify(src["SWE1-HMI-VC-019-02"][1])
    a1 = "reference" in h
    print(f"  self-test 1  第一層(b) 已知標的 019-02 應為候選  "
          f"{'PASS' if a1 else '**FAIL**'}  命中={h}")
    ok &= a1
    # 第一層 (a) 反向：017 為完整句，不應成候選
    h = classify(src["SWE1-HMI-VC-017"][1])
    a2 = not h
    print(f"  self-test 2  第一層(a) 反向 017 不應為候選        "
          f"{'PASS' if a2 else '**FAIL**'}  命中={h or '無'}")
    ok &= a2
    # 第二層 (b) 已知標的：第 1 批四筆之既有登記應全過
    rows = read_tsv(CONT)
    base = [r for r in rows if r["leaf"].startswith(
        ("SWE1-HMI-VC-012", "SWE1-HMI-VC-013"))]
    bad = layer2(src, sys1, base)
    a3 = base and not bad
    print(f"  self-test 3  第二層(b) 第 1 批四筆登記應全過      "
          f"{'PASS' if a3 else '**FAIL**'}  母體={len(base)} 不符={bad}")
    ok &= a3
    # 第二層 (a) 反向之一：錯節
    p1 = [{"leaf": "SWE1-HMI-VC-012-03", "sys1_section": "2.5",
           "sentence_index": "3"}]
    a4 = bool(layer2(src, sys1, p1))
    print(f"  self-test 4  第二層(a) 反向 012-03→§2.5 應 FAIL   "
          f"{'PASS' if a4 else '**FAIL**'}")
    ok &= a4
    # 第二層 (a) 反向之二：**錯句**（下放包 22 §二之新斷言）——
    # 節對而句序錯，句級細化後應被抓到；節級時抓不到。
    p2 = [{"leaf": "SWE1-HMI-VC-013-02", "sys1_section": "2.6.3",
           "sentence_index": "1"}]
    a5 = bool(layer2(src, sys1, p2))
    print(f"  self-test 5  第二層(a) 反向 013-02→§2.6.3 s1 應 FAIL "
          f"{'PASS' if a5 else '**FAIL**'}")
    ok &= a5
    # ── 第三層（下放包 23 §2.2）——夾具三斷言 ────────────────────
    fx = {"FIXTURE-01": [FIX_TC]}
    r_pos = [{"leaf": "FIXTURE-01", "resolution": "PC",
              "resolution_key": "pop-up"}]
    a6 = not layer3(r_pos, fx)[0]
    print(f"  self-test 6  第三層(b) 已知標的 PC 含 key 應過        "
          f"{'PASS' if a6 else '**FAIL**'}")
    ok &= a6
    r_neg = [{"leaf": "FIXTURE-01", "resolution": "PC",
              "resolution_key": "thermostat"}]
    a7 = bool(layer3(r_neg, fx)[0])
    print(f"  self-test 7  第三層(a) 反向 key 不在 PC 應 FAIL       "
          f"{'PASS' if a7 else '**FAIL**'}")
    ok &= a7
    # 反向之二：**欄位須分開查** —— key 只在 PC，卻聲稱 Step，應 FAIL。
    # 此斷言所測者為「第三層有沒有查對欄位」，非「有沒有找到字」。
    r_fld = [{"leaf": "FIXTURE-01", "resolution": "Step",
              "resolution_key": "pop-up"}]
    a8 = bool(layer3(r_fld, fx)[0])
    print(f"  self-test 8  第三層(a) 反向 key 只在 PC 而聲稱 Step 應 FAIL "
          f"{'PASS' if a8 else '**FAIL**'}")
    ok &= a8
    return ok


def main():
    src, sys1 = load_037(), load_sys1()
    print("cont_guard —— self-test 前置（PLAYBOOK §7.1.1）")
    if not self_test(src, sys1):
        print("\n**self-test 未全過 —— 不輸出正式結果，非零碼退出。**")
        return 2
    print("  → 八個斷言全過，開始跑正式母體\n")

    cont_rows = read_tsv(CONT)
    excl_rows = read_tsv(EXCL)
    cont_leaves = {r["leaf"] for r in cont_rows}
    excl_leaves = {r["leaf"] for r in excl_rows}
    defer_rows = read_tsv(DEFER)
    defer_leaves = {r["leaf"] for r in defer_rows}

    # `short` 清單（profile §8 之適用對象；不參與 CONT 候選判定）
    shorts = sorted((lid, len(strip_prefix(d)), strip_prefix(d))
                    for lid, (_, d) in src.items()
                    if len(strip_prefix(d)) < SHORT)
    SHORT_OUT.write_text(
        "leaf\tlen\tdescription\n"
        + "".join(f"{a}\t{b}\t{c}\n" for a, b, c in shorts), encoding="utf-8")

    all_cand, unhandled = layer1(src, cont_leaves, excl_leaves, defer_leaves)
    bad2 = layer2(src, sys1, cont_rows)

    print(f"第一層 —— 候選偵測（母體 {len(src)} 列 / 全 117 leaf 之 037）")
    print(f"  候選 {len(all_cand)} 筆；已處置 {len(all_cand) - len(unhandled)}；"
          f"**未處置 {len(unhandled)}**")
    for lid, hits, d in all_cand:
        state = ("CONT" if lid in cont_leaves
                 else "排除" if lid in excl_leaves
                 else f"列冊(第{next(r['batch'] for r in defer_rows if r['leaf'] == lid)}批)"
                 if lid in defer_leaves else "**未處置**")
        print(f"    {lid.replace('SWE1-HMI-VC-', ''):<10}{','.join(hits):<26}"
              f"{state:<12}{d}")
    print(f"\nprofile §8 短來源清單（< {SHORT} 字元）：{len(shorts)} 筆 → "
          f"{SHORT_OUT.relative_to(ROOT)}")
    print(f"\n第二層 —— 內容驗證（CONT 表 {len(cont_rows)} 條）")
    print(f"  不符 {len(bad2)} 條 {bad2 or '無'}")

    by_leaf = load_generated()
    bad3, ap3, pd3 = layer3(cont_rows, by_leaf)
    print(f"\n第三層 —— `resolved-by-structure` 之聲稱驗證"
          f"（profile §9.2；已生成 {len(by_leaf)} leaf）")
    print(f"  已驗 {len(ap3)} 條 {ap3 or '無'}；"
          f"待生成 {len(pd3)} 條 {pd3 or '無'}；不符 {len(bad3)} 條 {bad3 or '無'}")

    failed = bool(unhandled) or bool(bad2) or bool(bad3)
    print(f"\n{'**FAIL**' if failed else 'PASS'} —— "
          f"未處置候選 {len(unhandled)}；內容不符 {len(bad2)}；"
          f"結構聲稱不符 {len(bad3)}")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
