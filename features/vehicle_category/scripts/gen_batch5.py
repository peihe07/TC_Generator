#!/usr/bin/env python3
"""第 5 批 `Ignition Availability` 之生成（下放包 25 T134）。

16 leaf → 20 TC（`058-03`／`062-02`／`063-02`／`064-01` 各拆 2，R-VC26）。
a 段 16、b 段 0、**PENDING 7 筆**（皆為 DR-VC10）。

上半之取材一律**自來源檔逐字取出**，不在本檔內手抄 ——
手抄之逐字與來源之逐字是二件事，而第 7b 項只驗前者對得上後者，
**驗不出「我抄的時候看的是不是這一份」**。故本檔以 leaf_id 取值。
"""
import csv
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
A03 = ROOT / ("inputs/FM-WI-FSM-037-A03-N1L-SWE1-VehicleCategory-HMI-V0.1"
              " STLA 報告.xlsx")
S1 = ROOT / ("inputs/SYS1_HMI_Vehicle_Category_HMI_Logic_and_Flow_"
             "R1_SR24_Post_2A_(December_27_2023).xlsx")
OUT = ROOT / "generated/batch5_ignition_availability.json"

src = {}
for r in list(openpyxl.load_workbook(A03, read_only=True, data_only=True)
              ["Analysis Report"].iter_rows(values_only=True))[7:]:
    if r[0] not in (None, ""):
        src[str(r[0]).strip()] = (str(r[3]).strip(), str(r[4]).strip())

sys1 = {}
_rows = list(openpyxl.load_workbook(S1, read_only=True, data_only=True)
             ["Basic Report"].iter_rows(values_only=True))
_h = [str(c).strip() if c else "" for c in _rows[0]]
_oi, _di = _h.index("Outline Number"), _h.index("Description")
for _r in _rows[1:]:
    _o = str(_r[_oi]).strip() if _r[_oi] else ""
    if _o:
        sys1[_o] = ((str(_r[_di]) if _r[_di] else "")
                    .replace("_x000D_\n", "\n").replace("_x000D_", " "))

recon = {r["req_id"]: r for r in csv.DictReader(
    (ROOT / "data/recon_leaf_to_section.tsv").open(encoding="utf-8"),
    delimiter="\t")}
prio = {r["req_id"]: r["final_p"] for r in csv.DictReader(
    (ROOT / "data/priority_final.tsv").open(encoding="utf-8"), delimiter="\t")}


def sent(sec, idx):
    """SYS1 之句 —— 切分沿用全案同一式（`(?<=\\.)\\s+(?=[A-Z])`）。"""
    parts = [s.strip() for s in
             re.split(r"(?<=\.)\s+(?=[A-Z])", sys1[sec].strip())]
    if isinstance(idx, tuple):
        return " ".join(parts[idx[0] - 1:idx[1]])
    return parts[idx - 1]


TAB = 'Open the Vehicle Category screen and select the "Settings" tab'
TAB_ER = 'The Vehicle Category screen is displayed with the "Settings" tab active'

FUNC = "功能測試 (Functional based ; no specific technique)"
STATE = "狀態轉換 (State Transition Testing)"
TABLE = "決策表 (Decision Table Testing)"
EP = "等價劃分 (Equivalence Partitioning, EP)"
NEG = "負向測試 (Negative / Invalid)"

# `upper` 之取法：("D", leaf) = 037 Description；("T", leaf) = 037 Title；
# ("S", 節, 句) = SYS1。**無字面上半** —— 皆自來源取。
PU_STR = "PENDING: DR-VC10 PU0091 popup string"
PU_PATH = "PENDING: DR-VC10 Software Updates entry path in Key Off"
OBS = ("An observation period long enough to exceed any UI timeout, set to "
       "5 minutes as a test-design parameter. The source states no timeout value")
PATH_NOTE = ("**引上繳包 22 §2 之路徑解**：§13.1 擋的是 Settings 頁籤這一條進入"
             "路徑，§13.2–§13.4 所載之他路徑於 Key Off／ACC 明文可用 —— "
             "`057` 與本筆**並存不悖**，非例外，**不發 DR**。")
PU0237 = ("**追溯佐證（不入 ER）**：`Pop Up List HMI R1 (26PI)` `Main` 第 239 列 "
          "`PU0237` 之 `String/Popup Message` 與本筆之彈窗文字逐字相同，"
          "其 `Timeout (sec)` = `N/A`、`Exit Conditions` = `<X>`／`<OK>`。"
          "**PU 編號未載於 SYS1／037**（不同於 `PU0091` 之明載），"
          "故只入本欄作追溯，不入 ER（沿 DR-VC1 對 `VC-021` 之分寸）。")
PU0319 = ("**追溯佐證（不入 ER）**：`Pop Up List` 第 321 列 `PU0319` 之文字與本筆"
          "逐字相同，其 `Timeout (sec)` = `N/A`、**`Exit Conditions` = `N/A`**，"
          "其 `Description` 作 `Ignition Status: In RUN and then turned to Key Off "
          "or ACC … Pop-up is shown and is unable to be closed`。"
          "**`058` 與 `064` 確為二個不同彈窗**（文字相同、行為不同）。"
          "PU 編號未載於 SYS1／037，只入本欄。")
R26 = ("**拆分依 R-VC26（互相消耗）**：按下其一，彈窗即消失 —— "
       "另一條路徑之情境不復存在，須完整重建才能走。"
       "單一 TC 之 Procedure 結構上只能走一條，不拆即必有覆蓋洞。"
       "**二筆之 Procedure 各自完整重建情境**（下放包 25 §三）。")
A18 = ("**⚠ PENDING（IN §8.4.3；A-VC18／DR-VC10(一)）**：彈窗文字二源相左 —— "
       "SYS1 §13.4.1／§13.4.2 與 037 作 `“**Feature** not available while vehicle "
       "is in motion”`；`Pop Up List` 第 93 列 `PU0091` 之 `String/Popup Message` "
       "作 `**Function** not available while vehicle is in motion**.**`（含句末句點），"
       "`HMI Settings List` 第 150 列亦作 `Function`。"
       "**二份獨立來源對規格一份**，且該欄位就是彈窗的字。**不自行擇一**（§8.4.1）。")

SPEC = [
 dict(leaf="SWE1-HMI-VC-057", upper=("D", None), dm=EP,
      title="Settings tab unavailable in three ignition states",
      low="Tab availability -- the tab itself is blocked in the three ignition states",
      pc=["The vehicle under test can be placed in Key Off, Timed Mode and ACC"],
      data="NA",
      pr=["Place the vehicle in Key Off and record whether the Settings tab can be opened",
          "Place the vehicle in Timed Mode and record whether the Settings tab can be opened",
          "Place the vehicle in ACC and record whether the Settings tab can be opened"],
      er=["The Settings tab is unavailable in Key Off",
          "The Settings tab is unavailable in Timed Mode",
          "The Settings tab is unavailable in ACC"],
      axis="頁籤本身之可用性（對 058-* 之彈窗、059/060/061 之他路徑）",
      why="**驗證目標**：Settings 頁籤於 Key Off／Timed Mode／ACC 三個狀態皆不可用。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**為什麼三狀態同一筆（R-VC26）**：三者**不互相消耗** —— "
          "點火狀態可循環切換，走完其一不使另一之情境消失，"
          "故各以一個步驟／一條 ER 涵蓋即可，不拆。"
          + PATH_NOTE),
 dict(leaf="SWE1-HMI-VC-058-01", upper=("D", None), dm=FUNC,
      title="Pop-up on a blocked Settings tab attempt",
      low="Attempt in a blocked state -- the pop-up, its text and its two options",
      pc=["The vehicle under test can be placed in Key Off, Timed Mode and ACC"],
      data="NA",
      pr=["Place the vehicle in Key Off and attempt to access the Settings tab",
          "Record the pop-up that is displayed, its text and the options it carries",
          "Repeat step 1 in Timed Mode and in ACC and record the pop-up each time"],
      er=["The attempt to access the Settings tab is registered",
          'A pop-up is displayed, its text reads "Turn vehicle to Run or Key On to '
          'access menu." and it carries an "OK" option and an "X" option',
          "The same pop-up is displayed in Timed Mode and in ACC"],
      axis="嘗試進入之後果：彈窗出現（對 -02 之持續、-03 之關閉落點）",
      why="**驗證目標**：於三個受阻狀態嘗試進入 Settings 頁籤時，"
          "顯示所載文字之彈窗，帶 OK 與 X 二個選項。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**⚠ 二欄記法不對稱（A-VC10 第三面）**：Title 用直單引號、"
          "Description 用彎單 `‘…’` 與彎雙 `“…”` —— **取 Description 一欄，不混用**。"
          "**為什麼彈窗與其二選項同一筆**：同一觸發（嘗試進入）之數個後果，"
          "IN §398 明文不拆，列為 ER 之內容。" + PU0237),
 dict(leaf="SWE1-HMI-VC-058-02", upper=("S", "13.1.1", (1, 2)), dm=FUNC,
      title="Blocked-tab pop-up does not time out",
      low="Persistence -- the blocked-tab pop-up does not dismiss itself over time",
      pc=["The Settings key-off access pop-up is displayed after an attempt made "
          "in Key Off"],
      data=OBS,
      pr=["Record the screen at the start of the observation period",
          "Leave the pop-up untouched for the observation period named in the test data",
          "Record the screen at the end of the observation period"],
      er=["The pop-up is displayed",
          "No control on the pop-up is operated during the observation period",
          "The pop-up is still displayed"],
      axis="嘗試進入之後果：彈窗不逾時（對 -01 之出現、-03 之關閉）",
      why="**驗證目標**：該彈窗不會自行逾時消失。"
          "**⚠ 取材為 CONT 之指涉型（R-VC25 優先序第 2）**：037 之 `This pop-up` "
          "其先行詞在 SYS1 §13.1.1 s1。**s1+s2 共 43 token，未逾 R-3 之 50** → "
          "**profile §9.2 層次 1 之預設處置成立，不採第三處置類**（層次不得跳層）。"
          "與 `064-02` 之對照值得記：二筆句型幾乎相同，"
          "而 `064-02` 因其 s1 較長（42 vs 37 token）落入層次 2。"
          "**觀察期之 5 分鐘為測試設計參數，非來源所載**（§8.4.1）——"
          "來源未給逾時值。" + PU0237),
 dict(leaf="SWE1-HMI-VC-058-03", upper=("S", "13.1.1", 3), dm=FUNC,
      title="Closing the blocked-tab pop-up with X",
      low="Closing with X -- return to the screen the attempt was made from",
      pc=["The vehicle is in Key Off and the Settings key-off access pop-up is "
          "displayed after an attempt made from a known screen"],
      data="NA",
      pr=["Record the screen the attempt to enter the Settings tab was made from",
          'Press the "X" option on the pop-up',
          "Record the screen that is displayed"],
      er=["The screen the attempt was made from is recorded as the baseline",
          "The X press is accepted",
          "The screen recorded in step 1 is displayed"],
      axis="關閉之控制項：X（對 OK 支之同一落點）",
      why="**驗證目標**：以 X 關閉該彈窗，返回嘗試進入時所在之畫面。"
          "**⚠ 取材為第三處置類 `resolved-by-structure`（profile §9.2 層次 2）**："
          "037 之 `Closing **the pop-up**…` 為**定冠詞回指**，其先行詞在 s1；"
          "整段 s1-s3 為 **68 token，逾 R-3 之 50**，故不取整段。"
          "單句 s3 ＋ 指涉由 TC 結構承載，CONT 登記 "
          "`resolution=PC`／`resolution_key=pop-up`。"
          "**⚠ 本筆為第一層之偽陰性**（定冠詞回指，非代名詞起首）——"
          "由勘查 (d) 之 SYS1 對照發現，非由候選偵測發現（profile §9.4.1）。"
          + R26 + PU0237),
 dict(leaf="SWE1-HMI-VC-058-03", upper=("S", "13.1.1", 3), dm=FUNC,
      title="Closing the blocked-tab pop-up with OK",
      low="Closing with OK -- return to the screen the attempt was made from",
      pc=["The vehicle is in Key Off and the Settings key-off access pop-up is "
          "displayed after an attempt made from a known screen"],
      data="NA",
      pr=["Record the screen the attempt to enter the Settings tab was made from",
          'Press the "OK" option on the pop-up',
          "Record the screen that is displayed"],
      er=["The screen the attempt was made from is recorded as the baseline",
          "The OK press is accepted",
          "The screen recorded in step 1 is displayed"],
      axis="關閉之控制項：OK（對 X 支之同一落點）",
      why="**驗證目標**：以 OK 關閉該彈窗，返回嘗試進入時所在之畫面。"
          "**取材同 X 支**（第三處置類，單句 s3）。" + R26
          + "**二支之落點相同而控制項不同** —— 括號下半以控制項區分。"),
 dict(leaf="SWE1-HMI-VC-059-01", upper=("D", None), dm=FUNC,
      title="Phone settings reached through the Phone screens",
      low="Access path -- Phone settings reached through the Phone screens",
      pc=["The vehicle under test is equipped with the Phone screens"],
      data="NA",
      pr=["Open the Phone screens",
          "Open the Phone settings from the Phone screens and record the screen "
          "that is displayed"],
      er=["The Phone screens are displayed",
          "The Phone settings are displayed"],
      axis="Phone settings 之路徑（對 -02 之點火狀態）",
      why="**驗證目標**：使用者可經 Phone screens 進入 Phone settings。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**為什麼與 -02 分立**：本筆驗**路徑存在**，`-02` 驗**該路徑於受阻狀態"
          "仍可用** —— 二者之失效不同（路徑不通 vs 路徑於 Key Off 被擋）。"),
 dict(leaf="SWE1-HMI-VC-059-02", upper=("D", None), dm=EP,
      title="Phone settings available in Key Off and ACC",
      low="Ignition states -- Phone settings stay reachable in Key Off and in ACC",
      pc=["The vehicle under test is equipped with the Phone screens"],
      data="NA",
      pr=["Place the vehicle in Key Off and open the Phone settings through the "
          "Phone screens",
          "Record the screen that is displayed",
          "Repeat step 1 in ACC and record the screen that is displayed"],
      er=["The Phone settings are opened while the vehicle is in Key Off",
          "The Phone settings are displayed",
          "The Phone settings are displayed in ACC"],
      axis="Phone settings 之點火狀態（對 -01 之路徑）",
      why="**驗證目標**：Phone settings 於 Key Off 與 ACC 仍可用。"
          "**取材（R-VC25）**：上半取自 037 `Description`。" + PATH_NOTE
          + "**二狀態不互相消耗**（R-VC26）—— 不拆，各以步驟／ER 涵蓋。"),
 dict(leaf="SWE1-HMI-VC-060-01", upper=("D", None), dm=FUNC,
      title="Audio settings reached through the Media",
      low="Access path -- Audio settings reached through the Media",
      pc=["The vehicle under test is equipped with the Media screens"],
      data="NA",
      pr=["Open the Media screens",
          "Open the Audio settings from the Media and record the screen that is "
          "displayed"],
      er=["The Media screens are displayed",
          "The Audio settings are displayed"],
      axis="Audio settings 之路徑（對 -02 之點火狀態）",
      why="**驗證目標**：使用者可經 Media 進入 Audio settings。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**來源用語逐字為 `through the Media`**（非 `Media screens`）——"
          "上半保留其原字；Procedure 之 `the Media screens` 為作者散文，"
          "其所指同一（§13.3 之標的）。"),
 dict(leaf="SWE1-HMI-VC-060-02", upper=("D", None), dm=EP,
      title="Audio settings available in Key Off and ACC",
      low="Ignition states -- Audio settings stay reachable in Key Off and in ACC",
      pc=["The vehicle under test is equipped with the Media screens"],
      data="NA",
      pr=["Place the vehicle in Key Off and open the Audio settings through the Media",
          "Record the screen that is displayed",
          "Repeat step 1 in ACC and record the screen that is displayed"],
      er=["The Audio settings are opened while the vehicle is in Key Off",
          "The Audio settings are displayed",
          "The Audio settings are displayed in ACC"],
      axis="Audio settings 之點火狀態（對 -01 之路徑）",
      why="**驗證目標**：Audio settings 於 Key Off 與 ACC 仍可用。"
          "**取材（R-VC25）**：上半取自 037 `Description`。" + PATH_NOTE),
 dict(leaf="SWE1-HMI-VC-061", upper=("D", None), dm=EP,
      title="Software Updates available in Key Off and ACC",
      low="Ignition states -- Software Updates stay reachable in Key Off and in ACC",
      pc=["The vehicle under test is equipped with Software Updates in its "
          "Settings list"],
      data="NA",
      pr=[f"Place the vehicle in Key Off and open Software Updates through "
          f"{PU_PATH}",
          "Record the screen that is displayed",
          "Repeat step 1 in ACC and record the screen that is displayed"],
      er=["Software Updates are opened while the vehicle is in Key Off",
          "The Software Updates screen is displayed",
          "The Software Updates screen is displayed in ACC"],
      axis="Software Updates 之點火狀態（對 059/060 之已載路徑）",
      why="**驗證目標**：Software Updates 於 Key Off 與 ACC 仍可用。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**⚠ PENDING（IN §8.4.3；A-VC19／DR-VC10(二)）**：章 13 為三個"
          "「他路徑仍可用」之需求給出路徑，**獨缺本筆** —— `059-*` 有 §13.2 之 "
          "`through the Phone screens`、`060-*` 有 §13.3 之 `through the Media`，"
          "而 §13.4 **只斷言可用，未載經何路徑**。"
          "**執行層之實測**：SYS1 全表搜 `Software Update|FOTA|Wi-Fi` "
          "**僅命中 §13.4／§13.4.1／§13.4.2**，三節皆無路徑；"
          "`HMI Settings List` `Settings` 分頁之 `Software Updates` 為**第 27 類**"
          "（第 650 列），即在被 §13.1 擋住的頁籤後方，"
          "其第 651 列作 `See Software Updates Logic and Flow for logic` ——"
          "**委派至我方未持有之文件**。"
          "**為何不以通稱表述帶過（下放包 25 §2.1）**：`034-02` 所缺者為"
          "**測試資料**，通稱後 Procedure 仍可執行；本筆所缺者為**進入路徑**，"
          "「經一條於 Key Off 仍可用之路徑進入」**不是可執行的步驟**。"),
 dict(leaf="SWE1-HMI-VC-062-01", upper=("D", None), dm=NEG,
      title="Wi-Fi download setting blocked while in motion",
      low="In-motion block -- pressing the Wi-Fi download setting is refused with the pop-up",
      pc=["The vehicle under test is equipped with the Software Downloads Over "
          "Wi-Fi setting"],
      data="NA",
      pr=["Set the vehicle in motion",
          'Press the "Software Downloads Over Wi-Fi" setting',
          "Record whether the setting is entered, and record the pop-up that is "
          "displayed and its text"],
      er=["The vehicle is in motion",
          "The press on the setting is registered",
          f"The setting is not entered and a pop-up is displayed whose text is {PU_STR}"],
      axis="攔阻之觸發：按下設定（對 063-01 之流程中起步）",
      why="**驗證目標**：行進中按下 Wi-Fi 下載設定時，操作被攔阻並顯示彈窗。"
          "**取材（R-VC25）**：上半取自 037 `Description`。" + A18
          + "**不需 §5.6 之 baseline（下放包 25 §三）**：本筆攔的是**動作**"
          "（設定未被進入）而非**值** —— 與 `035-03` 之值比對不同型，"
          "「未進入」由該次操作之結果直接可判，不需操作前之基準值。"
          "**⚠ 記法不對稱（A-VC10 第三面）**：Title 直單、Description 彎單＋彎雙"
          " —— 取 Description 一欄。"
          "**`Software Downloads Over Wi-Fi` 之大小寫**：`HMI Settings List` "
          "第 651 列作 `over`（小寫 o），SYS1／037 作 `Over` —— "
          "**依 R-VC7 以 SYS1／037 為準**，記明以免誤判為抄錯。"),
 dict(leaf="SWE1-HMI-VC-062-02", upper=("S", "13.4.1", (1, 2)), dm=FUNC,
      title="OK on the in-motion pop-up returns to the Settings list",
      low="In-motion pop-up dismissed with OK -- return to the Settings list",
      pc=["The in-motion pop-up launched from the Software Downloads Over Wi-Fi "
          "setting is displayed"],
      data="NA",
      pr=[f"Record the pop-up that is displayed and compare its text against {PU_STR}",
          'Press the "OK" option on the pop-up',
          "Record the screen that is displayed"],
      er=["The pop-up is displayed",
          "The OK press is accepted",
          "The Settings list is displayed"],
      axis="離開之控制項：OK（對 X 支之同一落點）",
      why="**驗證目標**：按 OK 離開行進中攔阻彈窗，返回 Settings 清單。"
          "**⚠ 取材為 CONT 之指涉型（R-VC25 優先序第 2）**：037 之 `If **they** "
          "press…` 其代名詞**非句首**（句首為 `If`），且其所按之標的為 s1 之彈窗。"
          "**s1+s2 共 46 token，未逾 R-3 之 50** → **profile §9.2 層次 1**，"
          "取 s1-s2，不採第三處置類。"
          "**⚠ 本筆為第一層之偽陰性**（非句首代名詞）——由勘查 (d) 發現。"
          "**R-VC24 判別**：Title 含 `Software Downloads Over Wi-Fi`（屬 `062-01`），"
          "其謂語為 `return them to the Settings list`（本 leaf 之行為），"
          "該詞用以定位是哪一個 in-motion 彈窗 —— **情境脈絡，非行為主張，非越界**。"
          + R26 + "**PENDING 置於 Procedure 而非 ER**：本筆之驗證標的為"
          "**返回落點**，非彈窗文字（後者屬 `062-01`）；"
          "文字於此只用於**辨識按的是哪個彈窗**，故置於步驟。"
          "另立 ER 斷言其文字會與 `062-01` 之驗證點重複（IN §527）。"),
 dict(leaf="SWE1-HMI-VC-062-02", upper=("S", "13.4.1", (1, 2)), dm=FUNC,
      title="X on the in-motion pop-up returns to the Settings list",
      low="In-motion pop-up dismissed with X -- return to the Settings list",
      pc=["The in-motion pop-up launched from the Software Downloads Over Wi-Fi "
          "setting is displayed"],
      data="NA",
      pr=[f"Record the pop-up that is displayed and compare its text against {PU_STR}",
          'Press the "X" option on the pop-up',
          "Record the screen that is displayed"],
      er=["The pop-up is displayed",
          "The X press is accepted",
          "The Settings list is displayed"],
      axis="離開之控制項：X（對 OK 支之同一落點）",
      why="**驗證目標**：按 X 離開行進中攔阻彈窗，返回 Settings 清單。"
          "**取材同 OK 支**（層次 1，s1-s2）。" + R26),
 dict(leaf="SWE1-HMI-VC-063-01", upper=("D", None), dm=NEG,
      title="Motion during a FOTA via Wi-Fi flow raises the block",
      low="Motion starts mid-flow -- the FOTA via Wi-Fi flow is interrupted by the pop-up",
      pc=["The vehicle under test is stationary and the user is part way through "
          "the FOTA via Wi-Fi flow"],
      data="NA",
      pr=["Record the step of the FOTA via Wi-Fi flow that is on screen",
          "Set the vehicle in motion",
          "Record the pop-up that is displayed and its text"],
      er=["A step of the FOTA via Wi-Fi flow is displayed",
          "The vehicle is in motion",
          f"A pop-up is displayed whose text is {PU_STR}"],
      axis="攔阻之觸發：流程中起步（對 062-01 之按下設定）",
      why="**驗證目標**：FOTA via Wi-Fi 流程中車輛起步時，顯示攔阻彈窗。"
          "**取材（R-VC25）**：上半取自 037 `Description`。" + A18
          + "**與 `062-01` 之區分**：`062-01` 之觸發為**使用者按下設定**"
          "（先靜後動之進入嘗試），本筆之觸發為**車輛開始移動**"
          "（先進入後起步）—— 二個不同觸發，IN §402 之既有判準即足，"
          "不需援引 R-VC26。"
          "**不需 baseline**（同 `062-01`）—— 攔的是動作。"
          "**範圍（§8.4.2）**：`any of the logic for FOTA via Wi-Fi` 之流程內容"
          "屬 Software Updates 側，本筆只驗**起步時之攔阻**，不驗流程本身。"),
 dict(leaf="SWE1-HMI-VC-063-02", upper=("S", "13.4.2", 2), dm=FUNC,
      title="OK on the FOTA in-motion popup returns to the pre-flow screen",
      low="FOTA popup dismissed with OK -- return to the screen in context before the flow",
      pc=["The in-motion popup raised during a FOTA via Wi-Fi flow is displayed, "
          "and the screen that was in context before the flow was entered is known"],
      data="NA",
      pr=[f"Record the popup that is displayed and compare its text against {PU_STR}",
          'Press the "OK" option on the popup',
          "Record the screen that is displayed"],
      er=["The popup is displayed",
          "The OK press is accepted",
          "The screen that was in context before the FOTA via Wi-Fi flow was "
          "entered is displayed"],
      axis="離開之控制項：OK（對 X 支之同一落點）",
      why="**驗證目標**：按 OK 離開該彈窗，返回進入 FOTA via Wi-Fi 流程前之畫面。"
          "**⚠ 取材為第三處置類 `resolved-by-structure`（profile §9.2 層次 2）**："
          "037 之 `If **they** press…` 為非句首代名詞，其標的為 s1 之彈窗；"
          "**s1+s2 共 68 token，逾 R-3 之 50**，故不取整段，"
          "單句 s2 ＋ CONT 登記 `resolution=PC`。"
          "**⚠ `resolution_key` 為 `popup` 而非 `pop-up`** —— "
          "SYS1 §13.4.2 原文即作 `popup`（§13.1.1／§13.5 作 `pop-up`），"
          "依 profile §9.3 **逐字不寬鬆**：不去連字號、不同義展開。"
          "故本筆之 Pre-Condition 與 Procedure 一律書 `popup`。"
          "**與 `062-02` 之落點不同**：`062-02` 返回 Settings 清單，"
          "本筆返回**進入流程前之畫面** —— 二者非同一落點，故非重複。"
          + R26),
 dict(leaf="SWE1-HMI-VC-063-02", upper=("S", "13.4.2", 2), dm=FUNC,
      title="X on the FOTA in-motion popup returns to the pre-flow screen",
      low="FOTA popup dismissed with X -- return to the screen in context before the flow",
      pc=["The in-motion popup raised during a FOTA via Wi-Fi flow is displayed, "
          "and the screen that was in context before the flow was entered is known"],
      data="NA",
      pr=[f"Record the popup that is displayed and compare its text against {PU_STR}",
          'Press the "X" option on the popup',
          "Record the screen that is displayed"],
      er=["The popup is displayed",
          "The X press is accepted",
          "The screen that was in context before the FOTA via Wi-Fi flow was "
          "entered is displayed"],
      axis="離開之控制項：X（對 OK 支之同一落點）",
      why="**驗證目標**：按 X 離開該彈窗，返回進入流程前之畫面。"
          "**取材同 OK 支**（第三處置類，單句 s2，`resolution_key=popup`）。" + R26),
 dict(leaf="SWE1-HMI-VC-064-01", upper=("D", None), dm=STATE,
      title="Transition to Key Off with the Settings tab open",
      low="Settings tab open then transition -- the pop-up appears over the tab",
      pc=["The vehicle is in Run and the Settings tab is open"],
      data="NA",
      pr=["Turn the vehicle to Key Off",
          "Record the pop-up that is displayed and its text"],
      er=["The vehicle is in Key Off",
          'A pop-up is displayed and its text reads "Turn vehicle to Run or Key '
          'On to access menu."'],
      axis="開啟之範圍：Settings 頁籤（對 category 支）",
      why="**驗證目標**：Settings 頁籤開啟中車輛轉入 Key Off 時，顯示彈窗。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**⚠ 拆 2 依 R-VC26（下放包 25 §2.3）**：`tab` 與 `category` 為二個"
          "**範圍**，且 `064-02` 已載該彈窗**不可被使用者關閉** ——"
          "走完其一須**整輪點火循環**才能重建另一之情境，即互相消耗。"
          "**⚠ 記法不對稱（A-VC10 第三面）**：Title 作 `Key Off/Timed Mode/ACC`"
          "（斜線），Description 作 `Key Off, Timed Mode or ACC` —— "
          "取 Description 一欄，其形態隨之。" + PU0319),
 dict(leaf="SWE1-HMI-VC-064-01", upper=("D", None), dm=STATE,
      title="Transition to ACC with a Settings category open",
      low="Settings category open then transition -- the pop-up appears over the category",
      pc=["The vehicle is in Run and a Settings category that is not available "
          "in Key Off is open"],
      data="NA",
      pr=["Turn the vehicle to ACC",
          "Record the pop-up that is displayed and its text"],
      er=["The vehicle is in ACC",
          'A pop-up is displayed and its text reads "Turn vehicle to Run or Key '
          'On to access menu."'],
      axis="開啟之範圍：Settings 類別（對 tab 支）",
      why="**驗證目標**：不可用之 Settings 類別開啟中車輛轉入 ACC 時，顯示彈窗。"
          "**取材同 tab 支**。"
          "**§13.5 之 `tab **or a Settings category**` 是本拆分之依據** ——"
          "該措辭明文承認 category 可獨立於 tab 被開啟（上繳包 22 §2.2 之旁證）。"
          "**二支之點火目標狀態分取 Key Off 與 ACC**：來源之 "
          "`turned to Key Off or ACC` 為二個狀態，二者**不互相消耗**"
          "（可循環），本可同筆涵蓋；分置二支使二個範圍各配一個狀態，"
          "**不增加 TC 數而涵蓋二者**。"),
 dict(leaf="SWE1-HMI-VC-064-02", upper=("S", "13.5", 2), dm=FUNC,
      title="Transition pop-up neither times out nor closes",
      low="Persistence -- the transition pop-up neither times out nor yields to the user",
      pc=["The Key Off transition pop-up is displayed and the vehicle remains in "
          "Key Off"],
      data=OBS,
      pr=["Leave the pop-up untouched for the observation period named in the "
          "test data and record the screen",
          "Press the pop-up and each area around it where a close control would "
          "normally be, and record the screen after each press"],
      er=["The pop-up is still displayed at the end of the observation period",
          "The pop-up is still displayed after each press"],
      axis="轉換彈窗之持續：不逾時且不可關（對 -03 之自動解除）",
      why="**驗證目標**：該彈窗不逾時，且使用者關不掉。"
          "**⚠ 取材為第三處置類 `resolved-by-structure`（profile §9.2 層次 2）**："
          "037 之 `This pop-up` 先行詞在 s1；**s1+s2 共 54 token，逾 R-3 之 50** →"
          "單句 s2 ＋ `resolution=PC`／`resolution_key=pop-up`。"
          "**與 `058-02` 之對照**：二筆句型幾乎相同，"
          "而 `058-02` 之 s1 較短（37 vs 42 token）使其整段未逾限、落在層次 1。"
          "**層次不得跳層** —— 差別只在來源句之長度。"
          "**為什麼二個斷言不拆（R-VC26）**：「不逾時」與「不可關」"
          "**不互相消耗** —— 同一個彈窗可連續觀察，走完其一另一之情境仍在。"
          "沿 `045`（不逾時＋選取後不關閉）之既有處置，以二條 ER 涵蓋。"
          "**觀察期之 5 分鐘為測試設計參數，非來源所載**。" + PU0319),
 dict(leaf="SWE1-HMI-VC-064-03", upper=("S", "13.5", 3), dm=STATE,
      title="Returning to Run clears the transition pop-up",
      low="Return to Run -- the pop-up clears itself and the prior Settings screen comes back",
      pc=["The Key Off transition pop-up is displayed, and the Settings screen "
          "the user was on before the transition is known"],
      data="NA",
      pr=["Record the Settings screen the user was on before the pop-up was triggered",
          "Turn the vehicle to Run",
          "Record the screen that is displayed"],
      er=["The Settings screen before the transition is recorded as the baseline",
          "The vehicle is in Run",
          "The pop-up is no longer displayed and the Settings screen recorded in "
          "step 1 is displayed"],
      axis="轉換彈窗之解除：回到 Run（對 -02 之持續）",
      why="**驗證目標**：車輛回到 Run 時彈窗自動關閉，並回到轉換前之 Settings 畫面。"
          "**⚠ 取材為第三處置類**：037 之 `while **pop-up** is on screen` 為"
          "**無冠詞名詞回指**；整段 s1-s3 為 **83 token，逾 R-3 之 50** →"
          "單句 s3 ＋ `resolution=PC`／`resolution_key=pop-up`。"
          "**⚠ 本筆為第一層之偽陰性**（無冠詞名詞回指）——由勘查 (d) 發現。"
          "**ER 之 baseline（§5.6）**：「回到轉換前之畫面」須先記錄那是哪一個。"
          "**⚠ 未涵蓋者，請上游注意**：來源作 `turned to Run **or Key On**` ——"
          "本筆只走 Run。二者**互相消耗**（轉到 Run 後彈窗已消失，"
          "須整輪重建才能走 Key On），依 R-VC26 應拆 2；"
          "**惟本包授權為 20 筆，其拆分清單未含本筆**，故不自行增筆。"
          "若上游認 Run 與 Key On 為同一等價類（「離開受阻狀態」），"
          "則本筆之涵蓋完整；若認為二個狀態，則此處為覆蓋洞。**請裁。**"
          + PU0319),
]

def upper_of(sp, leaf):
    kind = sp["upper"][0]
    if kind == "T":
        return src[leaf][0]
    if kind == "D":
        return src[leaf][1]
    return sent(sp["upper"][1], sp["upper"][2])


def numbered(xs):
    return "\n".join(f"{i}. {x}" for i, x in enumerate(xs, 1))


tcs = []
for sp in SPEC:
    leaf = sp["leaf"]
    up = upper_of(sp, leaf)
    tcs.append({
        "leaf_id": leaf,
        "test_group": "Vehicle Category",
        "test_set": "Ignition Availability",
        "tc_title": sp["title"],
        "test_item": f"{up}\n\n({sp['low']})",
        "pre_conditions": numbered(sp["pc"]),
        "input_test_data": sp["data"],
        "test_procedure": numbered(sp["pr"]),
        "expected_result": numbered(sp["er"]),
        "specification_reference": recon[leaf]["spec_reference"],
        "design_method": sp["dm"],
        "priority": prio[leaf],
        "split_flag": False,
        "split_reason": "",
        "functional_safety": "NA",
        "reasoning": sp["why"],
        "distinguishing_axis": sp["axis"],
    })

doc = {
    "batch": "batch5_ignition_availability",
    "feature": "vehicle_category",
    "test_group": "Vehicle Category",
    "test_set": "Ignition Availability",
    "handoff": "docs/handoff/25_batch5_tc.md（勘查見 24）",
    "ruling": "R-VC21／R-VC22／R-VC23／R-VC24／R-VC25／**R-VC26** ＋ profile §9.2 第三處置類",
    "segment": "a",
    "segment_note": "**本批無 b 段** —— 16 leaf 全數生成，`held_leaves` 為空。",
    "split_delta": 4,
    "tc_id_status": "provisional",
    "leaf_scope": [x["leaf"] for i, x in enumerate(SPEC)
                   if x["leaf"] not in [y["leaf"] for y in SPEC[:i]]],
    "held_leaves": [],
    # 宣告 —— 第 8b 項比對其與實際出現者是否逐筆相符（同第 15 項之作法）
    "pending_scope": [
        {"leaf": t["leaf_id"], "dr": m.group(1), "marker": m.group(2).strip()}
        for t in tcs
        for f in ("pre_conditions", "input_test_data", "test_procedure",
                  "expected_result", "test_item", "tc_title")
        for m in [__import__("re").search(
            r"PENDING:\s*(DR-VC\d+)\s+([^\n\"]*)", t[f])] if m
    ],
    "write_back": "凍結 —— 本輪只產出 JSON，不寫回工作簿",
    "reasoning": "**驗證目標**：點火狀態對 Settings 可用性之影響 —— "
                 "頁籤於 Key Off／Timed Mode／ACC 之受阻與其彈窗、"
                 "Phone／Audio／Software Updates 三條他路徑之續可用、"
                 "行進中之二個攔阻、開啟中途轉入受阻狀態之彈窗與其解除。"
                 "**為什麼這樣切**：16 leaf 中 12 筆一 leaf 一 TC；"
                 "`058-03`／`062-02`／`063-02`／`064-01` 各拆 2 —— "
                 "**依 R-VC26 之互相消耗**（走了其一，另一之情境須重建），"
                 "`split_delta: 4`。"
                 "**PENDING 7 筆 TC**（皆為 DR-VC10）：`062-*`／`063-*` 六筆之"
                 "彈窗字（A-VC18）＋ `061` 之進入路徑（A-VC19）。"
                 "**本批非全潔批。**"
                 "**狀態並存**：`057`（頁籤受阻）與 `059-02`／`060-02`／`061`"
                 "（他路徑可用）並存不悖 —— 一管入口、一管他入口，"
                 "沿上繳包 22 §2 之路徑解，**不視為例外、不發 DR**。"
                 "**`split_flag` 恆 `False`、`split_reason` 恆空**（profile §11）。"
                 "**未涵蓋**：FOTA via Wi-Fi 之流程內容（§8.4.2 委派至 "
                 "Software Updates 側）、彈窗之 PU 編號（只入 reasoning 作追溯）。",
    "tcs": tcs,
}
OUT.write_text(json.dumps(doc, ensure_ascii=False, indent=1) + "\n", "utf-8")
print(f"{OUT.relative_to(ROOT)} — {len(tcs)} TC / "
      f"{len(doc['leaf_scope'])} leaf / split_delta {doc['split_delta']}")
