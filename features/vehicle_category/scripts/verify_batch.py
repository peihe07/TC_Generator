#!/usr/bin/env python3
"""收斂條件之驗證 —— pilot 與全量批次共用（`verify_batch.py`）。

**更名（T75，下放包 13）**：原名 `verify_pilot.py`。自第 13、14 項加入後
其已非 pilot 專用 —— 第 13 項驗「該批之 Test Set 與 framework §2 相符」、
第 14 項驗「setup 片語取自 profile §5」，二者皆為**每批**之收斂條件
（下放包 13 §4.4）。名實相符優於相容性，故更名。

沿用之路徑：`BATCH` 常數指向待驗之批次 JSON，預設為 pilot。

第 3–8 項可機械化者一律機械化；第 2、9、10 項含人工判斷成分，
本腳本只驗其**可機械化之部分**，其餘於上繳包標明主觀範圍。

母體標註（R-VC15）：本輪 TC 數 12，其 leaf 取自 **117 leaf 母體**之
Test Set `Glove Box`（12 leaf）。
"""
import csv
import json
import re
import sys
import os
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
BATCH = Path(os.environ.get(
    "BATCH", ROOT / "generated" / "pilot_glovebox.json"))
J = json.loads(BATCH.read_text("utf-8"))
# R-VC22(b)：收斂條件之母體為**本批 a 段之筆數**，非 Test Set 之 leaf 總數。
# R-VC22(b)：母體為本批 a 段之筆數。
# ⚠ 第 2 批起 `tcs ≠ leaf_scope` —— 拆分使 TC 數多於 leaf 數（IN §8.2.2）。
# `split_delta` 須於 JSON 明列且可驗（下放包 18 §4.4）。
SPLIT_DELTA = J.get("split_delta", 0)
EXPECT_N = len(J["leaf_scope"]) + SPLIT_DELTA
TCS = J["tcs"]

REQ_KEYS = ["tc_title", "pre_conditions", "input_test_data", "test_procedure",
            "expected_result", "specification_reference", "design_method",
            "priority", "split_flag", "split_reason"]
FIELDS = ["pre_conditions", "input_test_data", "test_procedure",
          "expected_result"]

res = []


def chk(n, name, ok, detail):
    res.append((n, name, ok, detail))


# 1 —— 10 個必要 key（IN §10.1）
missing = {t["leaf_id"]: [k for k in REQ_KEYS if k not in t] for t in TCS}
bad = {k: v for k, v in missing.items() if v}
chk(1, f"{EXPECT_N} 筆 JSON 完整，10 個必要 key 齊備（IN §10.1）",
    len(TCS) == EXPECT_N and not bad,
    f"TC 數 {len(TCS)}；缺 key {bad or '無'}")

# 2 —— IN §9 十七項：見上繳包逐項；此處只驗可機械化之子項
chk(2, "IN §9 十七項自檢（機械化子項見第 3–8、11、12 項；全項見上繳包）",
    True, "本腳本不代替逐項判讀，見上繳包 10 §4.2")


# 11 —— §4.4 之三類禁項（下放包 12 §四.1）。
# ⚠ 前一輪之第 3 項人工判讀**只驗了三類中的一類**（feature under test as
# premise），漏掉 system defaults 與 step-controlled state，
# 導致 12/12 帶錯往下走。判準太鬆之漏檢，與前三件「太嚴而誤報」方向相反。
PC_DEFAULT = re.compile(
    r"\b(head unit|HU)\b.*\b(powered on|booted|on)\b|\bignition is on\b",
    re.I)
PC_PREMISE = re.compile(
    r"\b(is accessible|is available|can be reached)\b", re.I)
# 第三類最難機械化：「該狀態是否步驟可達成」需語意判斷。
# 可機械化之近似 —— 該 pre_condition 之文字若與任一 procedure 步驟之標的
# 重疊（共用一個引號標的），即列為候選並人工判讀。
QUOTED = re.compile(r'"([^"]+)"')
c11 = {"default": [], "premise": [], "step_overlap": []}
for t in TCS:
    for ln in t["pre_conditions"].split("\n"):
        s = re.sub(r"^\d+\.\s*", "", ln).strip()
        if not s:
            continue
        if PC_DEFAULT.search(s):
            c11["default"].append(f"{t['leaf_id']}: {s[:56]}")
        if PC_PREMISE.search(s):
            c11["premise"].append(f"{t['leaf_id']}: {s[:56]}")
        pcq = set(QUOTED.findall(s))
        prq = set(QUOTED.findall(t["test_procedure"]))
        if pcq & prq:
            c11["step_overlap"].append(
                f"{t['leaf_id']}: 共用標的 {sorted(pcq & prq)}")
chk(11, "pre_conditions 無 §4.4 三類禁項（system defaults／premise／step-controlled）",
    not any(c11.values()),
    "；".join(f"{k} {len(v)}" for k, v in c11.items())
    + (f" -- {sum(c11.values(), [])[:3]}" if any(c11.values()) else ""))

# 12 —— 對他筆之值的隱性依賴（下放包 12 §五）
RE_CROSSREF = re.compile(
    r"\b(comparable|corresponding|similar|equivalent)\b[^.]{0,40}"
    r"\b(ceiling|threshold|limit|count|timeout|duration|interval)\b", re.I)
c12 = []
for t in TCS:
    for f in ("test_procedure", "expected_result", "pre_conditions",
              "input_test_data", "test_item"):
        for m in RE_CROSSREF.finditer(t[f]):
            c12.append(f"{t['leaf_id']}/{f}: {m.group(0)[:50]}")
chk(12, "無對他筆之值的隱性依賴（comparable/corresponding/... ＋ 門檻類名詞）",
    not c12, f"命中 {len(c12)} 處 {c12 or '無'}")

# 3 —— test_item 括號下半 12 筆兩兩不同
low = {}
nobracket = []
for t in TCS:
    m = re.search(r"\(([^()]*)\)\s*$", t["test_item"].strip())
    if not m:
        nobracket.append(t["leaf_id"])
    else:
        low.setdefault(m.group(1).strip(), []).append(t["leaf_id"])
dup = {k: v for k, v in low.items() if len(v) > 1}
chk(3, f"test_item 括號下半 {EXPECT_N} 筆兩兩不同（機械）",
    not nobracket and not dup and len(low) == EXPECT_N,
    f"缺括號 {nobracket or '無'}；重複 {dup or '無'}；相異 {len(low)}")

# 3b —— 下半一律英文（R-S4 / R-PMH153）
cjk = [t["leaf_id"] for t in TCS
       if re.search(r"[一-鿿]",
                    re.search(r"\(([^()]*)\)\s*$",
                              t["test_item"].strip()).group(1))]
chk("3b", "test_item 括號下半無中文（R-S4）", not cjk, f"含中文 {cjk or '無'}")

# 4 —— specification_reference 與 recon_leaf_to_section.tsv 逐字相符
ref = {r[0]: r[3] for r in csv.reader(
    (ROOT / "data" / "recon_leaf_to_section.tsv").open(encoding="utf-8"),
    delimiter="\t")}
bad4 = [(t["leaf_id"], t["specification_reference"], ref.get(t["leaf_id"]))
        for t in TCS if t["specification_reference"] != ref.get(t["leaf_id"])]
chk(4, f"specification_reference {EXPECT_N} 筆與 recon_leaf_to_section.tsv 逐字相符",
    not bad4, f"不符 {len(bad4)} 筆 {bad4 or ''}")

# 5 —— priority 與 priority_final.tsv 逐字相符
pf = {r["req_id"]: r["final_p"] for r in csv.DictReader(
    (ROOT / "data" / "priority_final.tsv").open(encoding="utf-8"),
    delimiter="\t")}
bad5 = [(t["leaf_id"], t["priority"], pf.get(t["leaf_id"]))
        for t in TCS if t["priority"] != pf.get(t["leaf_id"])]
chk(5, f"priority {EXPECT_N} 筆與 priority_final.tsv 逐字相符",
    not bad5, f"不符 {len(bad5)} 筆 {bad5 or ''}")

# 6 —— Test Set 皆為 Glove Box，無變體
ts = {t["test_set"] for t in TCS}
tg = {t["test_group"] for t in TCS}
chk(6, f"Test Set {EXPECT_N} 筆一致，Test Group 皆為 `Vehicle Category`",
    len(ts) == 1 and tg == {"Vehicle Category"},
    f"test_set={sorted(ts)}；test_group={sorted(tg)}")

# 7 —— 尾句號、引號規則
trail, sq, br, ws = [], [], [], []
for t in TCS:
    for f in FIELDS:
        for ln in t[f].split("\n"):
            s = ln.rstrip()
            if s and re.match(r"^\s*(\d+\.|[a-c]\.)", s) and s.endswith((".", "。")):
                trail.append(f"{t['leaf_id']}/{f}: {s[-40:]}")
            if ln != ln.strip():
                ws.append(f"{t['leaf_id']}/{f}")
    for f in FIELDS + ["tc_title"]:
        if re.search(r"\[[^\]]+\]|<[^>]+>", t[f]):
            br.append(f"{t['leaf_id']}/{f}")
        if re.search(r"(?<![A-Za-z])'[^']+'(?![A-Za-z])", t[f]):
            sq.append(f"{t['leaf_id']}/{f}")
chk(7, "尾句號／方括號／單引號／行首尾空白（IN §11，作者欄位）",
    not trail and not sq and not br and not ws,
    f"尾句號 {len(trail)}；單引號 {len(sq)}；方括號角括號 {len(br)}；"
    f"空白 {len(ws)}" + (f" -- {trail[:2]}{sq[:2]}{br[:2]}" if (trail or sq or br) else ""))

# 7b —— **整段子串比對**（下放包 18 §二；T99 之 34/34 通過後改採）。
#
# 前一版為逐 token 比對，其保護建立在一張樣式表上 —— R-VC23 通則化後
# **掃不到即等於未受保護**（`「…」`／`｢…｣`／任何未來記法皆在其外）。
# 而 R-VC23(c) 是通則化後之唯一實質保護。
#
# 新判準：上半（去首尾空白、首字母大小寫正規化）須為
#   037 `Title` ∪ 037 `Description` ∪ SYS1 對應節全文
# 之**逐字子串**。依據：R-S4 上半為規格原句 verbatim、R-3 得摘句、
# R-4 得轉首字母大小寫。
#
# 三項優於 token 比對：不需樣式表（任何記法自動受驗）、
# token 對得上而其間文字被竄改者亦抓得到、零閾值無語意判斷。
SRC, SYS1 = {}, {}
try:
    import openpyxl
    _wb = openpyxl.load_workbook(
        ROOT / "inputs/FM-WI-FSM-037-A03-N1L-SWE1-VehicleCategory-HMI-V0.1 STLA 報告.xlsx",
        read_only=True, data_only=True)
    for r in list(_wb["Analysis Report"].iter_rows(values_only=True))[7:]:
        if r[0] not in (None, ""):
            SRC[str(r[0]).strip()] = (str(r[3]).strip(), str(r[4]).strip())
    _s1 = ROOT / ("inputs/SYS1_HMI_Vehicle_Category_HMI_Logic_and_Flow_"
                  "R1_SR24_Post_2A_(December_27_2023).xlsx")
    if _s1.exists():
        _rr = list(openpyxl.load_workbook(_s1, read_only=True, data_only=True)
                   ["Basic Report"].iter_rows(values_only=True))
        _hh = [str(c).strip() if c else "" for c in _rr[0]]
        _oi, _di = _hh.index("Outline Number"), _hh.index("Description")
        for _x in _rr[1:]:
            _o = str(_x[_oi]).strip() if _x[_oi] else ""
            if _o:
                SYS1[_o] = ((str(_x[_di]) if _x[_di] else "")
                            .replace("_x000D_\n", "\n").replace("_x000D_", " "))
except Exception as e:                      # noqa: BLE001
    SRC = {}
    print(f"（警告：來源不可讀，7b 未實測：{e}）", file=sys.stderr)


def _vars(s):
    s = s.strip()
    return {s, s[:1].upper() + s[1:], s[:1].lower() + s[1:]} if s else {s}


origins, unsourced = {}, []
for t in TCS:
    top = t["test_item"].split("\n\n")[0].strip()
    ti, de = SRC.get(t["leaf_id"], ("", ""))
    sec = t["specification_reference"].rsplit("_", 1)[-1]
    s1 = SYS1.get(sec, "")
    hit = None
    for v in _vars(top):
        if v and v in ti:
            hit = "Title"
        elif v and v in de:
            hit = "Description"
        elif v and v in s1:
            hit = "SYS1"
        if hit:
            break
    origins[t["leaf_id"]] = hit
    if hit is None:
        unsourced.append(t["leaf_id"])
from collections import Counter as _C
chk("7b", "test_item 上半為來源之逐字子串（R-VC23(c)；整段，不倚樣式表）",
    bool(SRC) and not unsourced,
    f"取材來源分布 {dict(_C(v for v in origins.values() if v))}；"
    f"未對上來源 {len(unsourced)} 筆 {unsourced or '無'}"
    + ("" if SRC else "；**來源不可讀，本項未實測**"))

# 8 —— VC-033-01 帶且僅帶一處 PENDING
pat = "PENDING: DR-VC8 Glove Box lockout threshold"
cnt = {t["leaf_id"]: json.dumps(t, ensure_ascii=False).count("PENDING:")
       for t in TCS}
others = {k: v for k, v in cnt.items() if k != "SWE1-HMI-VC-033-01" and v}
one = cnt.get("SWE1-HMI-VC-033-01") == 1
# ⚠ 原為 `TCS[10]` —— **硬編位置索引**，只在該批恰有 ≥11 筆且
# `VC-033-01` 恰在第 11 位時才對。1 筆之探針即 IndexError。
# 改為依 leaf_id 查，位置無關。
_t331 = next((x for x in TCS if x["leaf_id"] == "SWE1-HMI-VC-033-01"), None)
exact = bool(_t331) and pat in _t331["test_procedure"]
chk(8, "PENDING 之分布與其字串（pilot 專屬；他批以第 8b 項驗）",
    (one and exact and not others) if "SWE1-HMI-VC-033-01" in cnt else True,
    f"033-01 之 PENDING 數 {cnt.get('SWE1-HMI-VC-033-01')}；"
    f"字串相符 {exact}；他筆帶 PENDING {others or '無'}")

# 8b —— PENDING 之字串與其 DR 編號（下放包 25 §3.1）。
#
# ⚠ **本項至此才實作。** 第 8 項之名稱自下放包 10 起即寫著
# 「pilot 專屬；他批以第 8b 項驗」—— **而第 8b 項從來不存在**。
# 第 2／3／4 批之 PENDING（`014`／`021`／`033-01`）因此**只受第 8 項之
# pilot 專屬分支保護，即等於未受檢查**。與「17 項當成 19 項交」同型：
# 名稱寫了，承載者沒有。
#
# 判準三件（零閾值，無語意判斷）：
#   (a) 每一處 `PENDING:` 皆須匹配 `PENDING: DR-VC<n> <text>`，`<text>` 非空
#   (b) 其 `DR-VC<n>` 須實際存在於 `DATA_REQUESTS.md` 之 `## DR-VC<n>` 標題
#   (c) JSON 頂層之 `pending_scope` 須與實際出現者**逐筆相符**
#       —— 宣告與實際之比對，同第 15 項對 `split_delta` 之作法
RE_PEND = re.compile(r"PENDING:\s*(DR-VC\d+)\s+([^\n\"]*)")
DRDOC = (ROOT / "DATA_REQUESTS.md").read_text("utf-8")
known_dr = set(re.findall(r"^##\s+(DR-VC\d+)", DRDOC, re.M))
found8b, bad8b = [], []
for t in TCS:
    for f in FIELDS + ["test_item", "tc_title"]:
        for m in RE_PEND.finditer(t[f]):
            dr, txt = m.group(1), m.group(2).strip()
            found8b.append((t["leaf_id"], f, dr, txt))
            if not txt:
                bad8b.append(f"{t['leaf_id']}/{f}: {dr} 無說明文字")
            if dr not in known_dr:
                bad8b.append(f"{t['leaf_id']}/{f}: {dr} 不存在於 DATA_REQUESTS.md")
    # 裸 `PENDING` 而不合樣式者（含 `PENDING: DR-XX` 之他 feature 前綴）
    blob = json.dumps({k: t[k] for k in FIELDS + ["test_item", "tc_title"]},
                      ensure_ascii=False)
    if blob.count("PENDING") != len(RE_PEND.findall(blob)):
        bad8b.append(f"{t['leaf_id']}: 有 PENDING 不合 `PENDING: DR-VC<n> <text>` 樣式")
decl = J.get("pending_scope")
if decl is not None:
    got = sorted((a, c, d) for a, _, c, d in found8b)
    want = sorted((x["leaf"], x["dr"], x["marker"]) for x in decl)
    if got != want:
        bad8b.append(f"pending_scope 宣告 {len(want)} 筆與實際 {len(got)} 筆不符")
chk("8b", "PENDING 之樣式、DR 存在性、與 pending_scope 之宣告相符",
    not bad8b and (decl is not None or not found8b),
    f"實際 {len(found8b)} 處；宣告 {'無' if decl is None else len(decl)} 筆；"
    f"DR 分布 {dict(Counter(d for _, _, d, _ in found8b)) or '無'}；"
    f"不符 {len(bad8b)} 處 {bad8b or '無'}"
    + ("" if decl is not None or not found8b else "；**有 PENDING 而 JSON 未宣告 pending_scope**"))

# 9 —— 流程區分（機械部分：括號下半須含 activation / deactivation）
f9 = []
for lid in ("SWE1-HMI-VC-028-02", "SWE1-HMI-VC-033-01"):
    t = next((x for x in TCS if x["leaf_id"] == lid), None)
    if t is None:
        continue
    b = re.search(r"\(([^()]*)\)\s*$", t["test_item"].strip()).group(1).lower()
    if "activation" not in b and "deactivation" not in b:
        f9.append(lid)
chk(9, "`028-02`／`033-01` 之括號下半明載其流程（pilot 專屬）",
    not f9, f"未載者 {f9 or '無'}")

# 10 —— VC-021 委派載於各 TC 之 reasoning
if any(x["leaf_id"].startswith("SWE1-HMI-VC-02") and "Glove Box" in x["test_set"]
       for x in TCS):
    f10 = [t["leaf_id"] for t in TCS if "VC-021" not in t.get("reasoning", "")]
    chk(10, f"`VC-021` 之委派載於全部 {EXPECT_N} 筆之 reasoning（§8.2.1）",
        not f10, f"未載者 {f10 or '無'}")
else:
    chk(10, "`VC-021` 之委派（pilot 專屬；本批不適用）", True, "N/A")

# 另：test_procedure ≥ 2 步、Procedure ↔ ER 1:1、ER 無 modal
p2 = [t["leaf_id"] for t in TCS
      if len(t["test_procedure"].split("\n")) < 2]
mism = [(t["leaf_id"], len(t["test_procedure"].split("\n")),
         len(t["expected_result"].split("\n")))
        for t in TCS
        if len(t["test_procedure"].split("\n")) != len(t["expected_result"].split("\n"))]
# modal 之檢查須**排除引號內之 UI 文字** —— 彈窗原文
# "PIN must be 4 digits / OK" 之 must 是規格所載之畫面文字，
# 不是作者之 modal。初版未排除，誤報 VC-033-02（見上繳包 10 §4.3）。
def _unquoted(s):
    return re.sub(r'"[^"]*"', " ", s)


modal = [t["leaf_id"] for t in TCS
         if re.search(r"\b(shall|will|should|must)\b",
                      _unquoted(t["expected_result"]), re.I)]
verb = [t["leaf_id"] for t in TCS
        if re.search(r"^\d+\.\s*(observe|verify|check that)\b",
                     t["test_procedure"], re.I | re.M)]
chk("A", "Procedure ≥2 步 ∧ Procedure↔ER 1:1 ∧ ER 無 modal ∧ 步驟無 observe/verify 起首",
    not p2 and not mism and not modal and not verb,
    f"步數不足 {p2 or '無'}；1:1 不符 {mism or '無'}；"
    f"ER 含 modal {modal or '無'}；禁用起首動詞 {verb or '無'}")

# 13 —— 該批之 Test Set 與 framework.md §2 逐字相符（下放包 13 §4.4）
#      名稱自 framework 解析，**不硬編** —— 硬編會使本項只驗到腳本自己。
FW = (ROOT / "framework.md").read_text("utf-8")
fw_names = set(re.findall(r"^\|\s*\d+\s*\|\s*`([^`]+)`\s*\|", FW, re.M))
batch_ts = {t["test_set"] for t in TCS}
chk(13, "該批 Test Set 全筆一致且與 framework.md §2 逐字相符",
    len(batch_ts) == 1 and batch_ts <= fw_names,
    f"批內 test_set={sorted(batch_ts)}；framework §2 之 8 組="
    f"{len(fw_names)} 個；相符={batch_ts <= fw_names}")

# 14 —— §5.3 所防者為**變體擴散，不是位置**（下放包 16 §2.1）。
#
# 前一版判「Procedure 首步是否為常數」，隱含「首步必為 setup」之假設。
# 該假設對 `VC-001-02` 不成立 —— 其首步 `Open the Vehicle Category screen`
# **本身即受測動作**（「進入」正是該需求之觸發）。
#
# 新判準不問位置、不問角色，只問「**用到常數就不許走樣**」：
#   正規化 = 小寫 + 去標點（含引號）+ 壓縮連續空白 + 去首尾空白
#   任一步驟之正規化形式若等於某常數之正規化形式，
#   而其原字串與該常數**不逐字相同** → FAIL
#
# 零閾值。**不用編輯距離作硬判準** —— `Open the Vehicle Category screen`
# 與 `…and select the "Controls" tab` 之距離大，但其關係是**前綴**不是變體；
# 閾值鬆則誤報、緊則漏報，因為它量錯了東西（同 R-G39 對停止條件 87 之判詞）。
PROF = ROOT.parent.parent / "docs/runtime/profiles/FW036_R1L_VehicleCategory_Profile.md"
consts: set[str] = set()
if PROF.exists():
    ptxt = PROF.read_text("utf-8")
    m = re.search(r"### 5\.1 常數\s*\n\s*```\n(.*?)```", ptxt, re.S)
    if m:
        blk = m.group(1)
        # 值域自機讀行解析（profile §5.1 之 `values(<p>) = A | B`）
        domains = {p_: [v.strip() for v in vals.split("|")]
                   for p_, vals in re.findall(
                       r"values\(<(\w+)>\)\s*=\s*(.+)", blk)}
        for chunk in re.split(r"\n(?=\S)", blk.strip()):
            lines = [x for x in chunk.split("\n")]
            if not lines or ":" not in lines[0]:
                continue
            body = [x.strip() for x in lines[1:]
                    if x.strip() and not x.strip().startswith(("<", "values("))]
            if not body:
                continue
            tmpl = body[0]
            holes = re.findall(r"<(\w+)>", tmpl)
            if not holes:
                consts.add(tmpl)
            else:
                h = holes[0]
                for v in domains.get(h, []):
                    consts.add(tmpl.replace(f"<{h}>", v))


def _norm(s: str) -> str:
    s = s.lower()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


norm_const = {_norm(c): c for c in consts}
variants, soft = [], []
for t in TCS:
    for ln in t["test_procedure"].split("\n"):
        step = re.sub(r"^\d+\.\s*", "", ln).strip()
        if not step:
            continue
        c = norm_const.get(_norm(step))
        if c is not None and step != c:
            variants.append(f"{t['leaf_id']}: {step[:60]!r} ≠ {c[:60]!r}")
        elif c is None:
            # 軟檢查：近似但正規化後不等者，列為候選供人工判讀，**不自動 FAIL**
            for cn, co in norm_const.items():
                a, b = set(_norm(step).split()), set(cn.split())
                if a and b and len(a & b) / len(a | b) >= 0.7:
                    soft.append(f"{t['leaf_id']}: {step[:52]!r} ~ {co[:52]!r}")
                    break
chk(14, "常數之變體擴散（正規化後相等而原字不同 → FAIL；§5.3）",
    bool(consts) and not variants,
    f"profile 常數（展開後）{len(consts)} 條；變體 {len(variants)} 處 "
    f"{variants or '無'}"
    + (f"；軟檢查候選 {len(soft)} 處（人工判讀，不 FAIL）{soft}" if soft else "")
    + ("" if consts else "；**profile §5.1 未解析到常數**"))

# 15 —— 母體為本批 a 段之筆數（R-VC22(b)）。
#      已由 EXPECT_N 貫穿第 1／3／4／5／6 項，此處另立一項使其顯式可見。
# 15 —— 母體 ＝ leaf_scope ＋ split_delta，且 split_delta 須與實際拆分相符。
#      「實際拆分」＝ 同一 leaf_id 出現多筆之超額數，自 tcs 直接數得。
_lc = Counter(x["leaf_id"] for x in TCS)
actual_delta = sum(v - 1 for v in _lc.values() if v > 1)
split_detail = {k: v for k, v in _lc.items() if v > 1}
chk(15, f"母體 = leaf_scope + split_delta = {len(J['leaf_scope'])} + "
        f"{SPLIT_DELTA} = {EXPECT_N}（R-VC22(b)／IN §8.2.2）",
    len(TCS) == EXPECT_N and SPLIT_DELTA == actual_delta,
    f"tcs={len(TCS)}；leaf_scope={len(J['leaf_scope'])}；"
    f"宣告 split_delta={SPLIT_DELTA}；實際拆分增量={actual_delta}"
    + (f"（{split_detail}）" if split_detail else "")
    + f"；held={len(J.get('held_leaves', []))}（b 段不計入母體）")

# 16 —— 續行型 leaf 之 test_item 上半須與 SYS1 之完整句逐字相符
#      （下放包 15 §5.3 第 16 項）。SYS1 為權威複本（R-VC7）。
SENT = {}
_sys1 = ROOT / ("inputs/SYS1_HMI_Vehicle_Category_HMI_Logic_and_Flow_"
                "R1_SR24_Post_2A_(December_27_2023).xlsx")
# CONT 表**自 `data/cont_table.tsv` 讀取**，不再硬編（下放包 21 T116-5）。
# 句序仍需指定 —— 續行型取其所屬句、指涉型取整段（sec_idx = None）。
CONT = {}
_ct = ROOT / "data/cont_table.tsv"
if _ct.exists():
    _ln = _ct.read_text("utf-8").splitlines()
    _hd = _ln[0].split("\t")
    for _l in _ln[1:]:
        if not _l.strip():
            continue
        _r = dict(zip(_hd, _l.split("\t")))
        # 句序**自表中 `sentence_index` 讀取**（下放包 22 §二）——
        # 原為由節號硬推（`2.6.2`→第 3 句、`2.6.3`→第 2 句），該推導無檢查。
        # `*` 或空值 = 取整段（指涉型）。
        _si = (_r.get("sentence_index") or "*").strip()
        CONT[_r["leaf"]] = (_r["sys1_section"], _si)   # 原樣傳遞，含範圍
targets = [k for k in CONT if any(x["leaf_id"] == k for x in TCS)]
if not targets:
    chk(16, "續行型 leaf 之上半取 SYS1 完整句（本批無適用對象）", True, "N/A")
elif not _sys1.exists():
    chk(16, "續行型 leaf 之上半取 SYS1 完整句", False, "**SYS1 不可讀**")
else:
    import openpyxl as _ox
    _r = list(_ox.load_workbook(_sys1, read_only=True, data_only=True)
              ["Basic Report"].iter_rows(values_only=True))
    _h = [str(c).strip() if c else "" for c in _r[0]]
    _oi, _di = _h.index("Outline Number"), _h.index("SYSRE_HMI_Source ID")
    _di = _h.index("Description")
    for _row in _r[1:]:
        _o = str(_row[_oi]).strip() if _row[_oi] else ""
        if _o in {v[0] for v in CONT.values()}:
            _txt = (str(_row[_di]) if _row[_di] else "").replace(
                "_x000D_\n", "\n").replace("_x000D_", " ").strip()
            SENT[_o] = [s.strip() for s in
                        re.split(r"(?<=\.)\s+(?=[A-Z])", _txt)]
    bad16 = []
    for lid in targets:
        sec, idx = CONT[lid]
        if sec not in SENT:
            want = None
        elif idx in ("*", ""):
            want = " ".join(SENT[sec]).strip()      # 整段
        elif re.fullmatch(r"\d+\s*-\s*\d+", idx):
            _a, _b = (int(x) for x in idx.split("-"))
            want = (" ".join(SENT[sec][_a - 1:_b])
                    if 1 <= _a <= _b <= len(SENT[sec]) else None)
        else:
            _n = int(idx) - 1
            want = SENT[sec][_n] if 0 <= _n < len(SENT[sec]) else None
        got = next(x for x in TCS if x["leaf_id"] == lid)[
            "test_item"].split("\n\n")[0].strip()
        if want is None or got != want:
            bad16.append(f"{lid}: 上半 {got[:48]!r} ≠ SYS1 {str(want)[:48]!r}")
    chk(16, "續行型 leaf 之 test_item 上半與 SYS1 完整句逐字相符（R-VC7）",
        not bad16,
        f"適用 {len(targets)} 筆；不符 {len(bad16)} 筆 {bad16 or '無'}")

# 17 —— CONT 表之二層防護常駐（下放包 21 T116-5）。
#      其 self-test 前置由 cont_guard 自帶；本項只收其結果。
import subprocess as _sp
_cg = ROOT / "scripts" / "cont_guard.py"
if _cg.exists():
    _p = _sp.run([sys.executable, str(_cg)], capture_output=True, text=True)
    _tail = [x for x in _p.stdout.splitlines()
             if x.startswith(("PASS —", "**FAIL** —"))]
    chk(17, "CONT 表二層防護（候選無未處置 ∧ 內容驗證全過；含 self-test）",
        _p.returncode == 0,
        (_tail[-1] if _tail else f"cont_guard 離開碼 {_p.returncode}")
        + f"；離開碼 {_p.returncode}")
else:
    chk(17, "CONT 表二層防護", False, "**cont_guard.py 不存在**")

# 18 —— 項號指涉之存在性（下放包 26 §一，T136）。
#
# 由來：第 8 項之名稱自下放包 10 起寫著「他批以**第 8b 項**驗」，
# **而第 8b 項不存在**，直到下放包 25 才實作。五處 PENDING 走完三個批次無人看過。
#
# ⚠ **其形態異於「17 項當成 19 項交」**：那次是清單上的項沒做，一數即現形；
# 本次是**該項做了且 PASS，而它的名稱指涉一個不存在的項** ——
# **數量對、狀態綠，只有讀名稱的人會發現**。故需一項機械檢查。
#
# 判準：檢查之名稱／說明中凡出現 `第<id>項`（含範圍 `3–8` 與頓號列舉），
# 該 `<id>` 須存在於**本次執行之檢查清單**。不存在 → FAIL。零閾值。
RE_REF = re.compile(r"第\s*([0-9][0-9a-zA-Z]*(?:\s*[–—-]\s*[0-9][0-9a-zA-Z]*)?"
                    r"(?:\s*[、,]\s*[0-9][0-9a-zA-Z]*)*)\s*項")


def _ids(txt):
    """自一段文字取出其所引之項號（展開範圍與頓號列舉）。"""
    out = set()
    for grp in RE_REF.findall(txt):
        for part in re.split(r"[、,]", grp):
            part = part.strip()
            m = re.fullmatch(r"([0-9]+)\s*[–—-]\s*([0-9]+)", part)
            if m:                       # 範圍 —— 純數字者才展開
                out.update(str(x) for x in range(int(m.group(1)),
                                                int(m.group(2)) + 1))
            elif part:
                out.add(part)
    return out


def refcheck(items):
    """items = [(id, name, detail), …] → 回傳無主之項號。"""
    have = {str(i) for i, _, _ in items}
    dangling = []
    for i, name, detail in items:
        for ref in sorted(_ids(f"{name} {detail}")):
            if ref not in have:
                dangling.append(f"第{i}項 之名稱／說明引用不存在之「第{ref}項」")
    return dangling


# self-test 前置（PLAYBOOK §7.1.1）—— 不過則本項直接 FAIL，不看正式母體
_items = [(n, str(nm), str(dt)) for n, nm, ok_, dt in res]
_st_b = not refcheck(_items)                       # (a) 反向：現行全部名稱應全過
_probe = list(_items)
_probe[0] = (_probe[0][0], _probe[0][1] + "（見第 99 項）", _probe[0][2])
_st_a = bool(refcheck(_probe))                     # (b) 已知標的：插入第 99 項應 FAIL
print("第 18 項 —— self-test 前置（PLAYBOOK §7.1.1）")
print(f"  self-test 1  (b) 已知標的 插入「見第 99 項」應 FAIL  "
      f"{'PASS' if _st_a else '**FAIL**'}")
print(f"  self-test 2  (a) 反向 現行全部名稱應全過          "
      f"{'PASS' if _st_b else '**FAIL**'}")
_dangling = refcheck(_items)
chk(18, "檢查之名稱／說明所引之項號皆存在（下放包 26 §一）",
    _st_a and _st_b and not _dangling,
    f"self-test {'全過' if _st_a and _st_b else '**未過**'}；"
    f"掃描 {len(_items)} 項；引用之項號 "
    f"{sorted(set().union(*[_ids(f'{n} {d}') for _, n, d in _items]) or set())}；"
    f"無主 {len(_dangling)} 處 {_dangling or '無'}")

print(f"verify_batch — {BATCH.name}（收斂條件；下放包 10 §四 ＋ 13 §4.4）")
print(f"{'#':>3}  {'條件':<62} 判")
print("-" * 96)
failed = 0
for n, name, ok, detail in res:
    if not ok:
        failed += 1
    print(f"{str(n):>3}  {name:<62} {'PASS' if ok else '**FAIL**'}")
    print(f"     {detail}")
print("-" * 96)
print(f"{len(res)} checked / {failed} failed")
sys.exit(1 if failed else 0)
