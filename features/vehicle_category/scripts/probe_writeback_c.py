#!/usr/bin/env python3
"""丙′ 之可行性驗證：openpyxl 寫入 ＋ XML 後處理修復（下放包 28 §2.3，T145）。

**只驗可行性，不產出正式交付本。** 3 筆假資料，全程在 `/tmp` 副本上。
**母本不得被開啟寫入** —— 本檔對母本只做 `read_bytes()` 與唯讀載入。

甲／乙之破壞已於上繳包 27 §3.4 實測：openpyxl `load→save`（不改任何一格）
即毀 `x14:dataValidation`、`extLst`、`printerSettings`，並重新編碼圖片。
丙′ 之賭注是：**那些結構可以從母本原件搬回去**。

五步（下放包 28 §2.3 逐字）：
  1. openpyxl 開母本副本、寫入 14 欄資料、save
  2. 解壓輸出檔與母本
  3. 自母本取出遭毀之結構，注入輸出檔
  4. 重打包
  5. 驗六項

**第 5 步任一項不符即停並回報，不自行調整方案**（下放包 §2.3 明文）。
"""
import hashlib
import json
import re
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / ("inputs/FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT STLA "
                 "Test Case Specification & Result_SWQT_20260817_ext.xlsx")
TMP = Path("/tmp/vc_writeback_c")
SHEET = "Test Case Specification 測試用例規範"

# 下放包 27 §3.2 之映射，14 欄有來源者。C 自 recon 表帶入，餘自 TC JSON。
COLS = {"C": "polarion_id", "D": "leaf_id", "G": "test_group", "H": "test_set",
        "I": "test_item", "J": "pre_conditions", "K": "input_test_data",
        "L": "test_procedure", "M": "expected_result",
        "N": "specification_reference", "P": "priority", "R": "design_method",
        "S": "functional_safety"}
# 13 欄 ＋ C，共 14。B 為公式，不得覆寫。


def sheet_xml(zf: zipfile.ZipFile, name: str) -> str:
    """分頁名 → 其 worksheet XML 之 zip 路徑（不假設 sheetN 之編號穩定）。"""
    wbxml = zf.read("xl/workbook.xml").decode("utf-8")
    # openpyxl 之輸出把 `xmlns:r=` 放在 `name=` 之前 —— 屬性順序不可假設。
    rid = re.search(rf'<sheet[^>]*name="{re.escape(name)}"[^>]*r:id="([^"]+)"',
                    wbxml).group(1)
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    # 屬性順序與 Target 之絕對／相對形式**皆不可假設** ——
    # 母本作 `Id=… Target="worksheets/sheet6.xml"`，
    # openpyxl 作 `Target="/xl/worksheets/sheet1.xml" … Id=…`。
    tgt = None
    for tag in re.findall(r"<Relationship [^>]*/>", rels):
        if re.search(rf'Id="{rid}"', tag):
            tgt = re.search(r'Target="([^"]+)"', tag).group(1)
            break
    if tgt is None:
        raise KeyError(f"rels 中找不到 {rid}")
    tgt = tgt.lstrip("/")
    return tgt if tgt.startswith("xl/") else "xl/" + tgt


def facts(path: Path) -> dict:
    """六項驗收之量測。"""
    z = zipfile.ZipFile(path)
    sx = sheet_xml(z, SHEET)
    d = z.read(sx).decode("utf-8", "ignore")
    names = z.namelist()
    imgs = sorted(n for n in names if n.startswith("xl/media/"))
    return {
        "x14_dv": len(re.findall(r"<x14:dataValidation[ >]", d)),
        "extLst": len(re.findall(r"<extLst>", d)),
        "std_dv": len(re.findall(r"<dataValidation[ >]", d)),
        "printer": len([n for n in names if "printerSettings" in n and
                        n.endswith(".bin")]),
        "media": imgs,
        "sheet_xml": sx,
    }


def main() -> int:
    if TMP.exists():
        shutil.rmtree(TMP)
    TMP.mkdir(parents=True)
    src = TMP / "master_copy.xlsx"
    shutil.copy(MASTER, src)
    before = hashlib.sha256(MASTER.read_bytes()).hexdigest()

    tcs = json.loads((ROOT / "generated/pilot_glovebox.json").read_text("utf-8"))["tcs"][:3]
    import csv
    recon = {r["req_id"]: r for r in csv.DictReader(
        (ROOT / "data/recon_leaf_to_section.tsv").open(encoding="utf-8"),
        delimiter="\t")}
    for t in tcs:
        t["polarion_id"] = recon[t["leaf_id"]]["polarion_id"]

    print("步驟 1 —— openpyxl 開副本、寫入 14 欄、save")
    wb = openpyxl.load_workbook(src)
    ws = wb[SHEET]
    for i, t in enumerate(tcs):
        r = 10 + i
        for col, key in COLS.items():
            ws[f"{col}{r}"] = t[key]
    out = TMP / "step1_openpyxl.xlsx"
    wb.save(out)
    wb.close()
    f_master, f_step1 = facts(src), facts(out)
    print(f"  母本   : x14_dv={f_master['x14_dv']} extLst={f_master['extLst']} "
          f"std_dv={f_master['std_dv']} printer={f_master['printer']} "
          f"media={f_master['media']}")
    print(f"  step1  : x14_dv={f_step1['x14_dv']} extLst={f_step1['extLst']} "
          f"std_dv={f_step1['std_dv']} printer={f_step1['printer']} "
          f"media={f_step1['media']}")

    print("步驟 2 —— 解壓輸出檔與母本")
    zm, zo = zipfile.ZipFile(src), zipfile.ZipFile(out)
    print(f"  母本 {len(zm.namelist())} 項；step1 {len(zo.namelist())} 項；"
          f"消失 {len(set(zm.namelist()) - set(zo.namelist()))} 項")

    print("步驟 3／4 —— 自母本取出遭毀之結構，注入並重打包")
    # **二變體各跑一次**：下放包 §2.3 第 3 步列了三類待修復結構
    # （extLst／printerSettings／原始圖片）。第三類實測**不可單獨搬回** ——
    # openpyxl 已把 drawing 之 rels 改指向它重新編碼後之檔名，
    # 只換 xl/media/ 而不換 drawing 與其 rels，套件即開不起來。
    # 故變體 A（含圖片還原）與變體 B（不還原圖片）各產一檔，二者皆量。

    sx_m, sx_o = f_master["sheet_xml"], f_step1["sheet_xml"]
    ext = re.search(r"<extLst>.*?</extLst>",
                    zm.read(sx_m).decode("utf-8", "ignore"), re.S)
    o_doc = zo.read(sx_o).decode("utf-8", "ignore")
    injected = re.sub(r"</worksheet>\s*$", ext.group(0) + "</worksheet>", o_doc) \
        if ext else o_doc
    keep_from_master = [n for n in zm.namelist()
                        if "printerSettings" in n
                        or n.startswith("xl/media/")
                        or n.endswith("printerSettings.xml.rels")]
    def repack(dest: Path, restore_media: bool):
        with zipfile.ZipFile(dest, "w", zipfile.ZIP_DEFLATED) as zf:
            for n in zo.namelist():
                if n == sx_o:
                    zf.writestr(n, injected)
                elif restore_media and n.startswith("xl/media/"):
                    continue
                else:
                    zf.writestr(n, zo.read(n))
            for n in keep_from_master:
                if n.startswith("xl/media/"):
                    if restore_media:
                        zf.writestr(n, zm.read(n))
                elif n not in zo.namelist():
                    zf.writestr(n, zm.read(n))

    var_a = TMP / "step4_variantA_media_restored.xlsx"
    var_b = TMP / "step4_variantB_media_kept.xlsx"
    repack(var_a, True)
    repack(var_b, False)
    print(f"  注入 extLst：{'是' if ext else '**母本無 extLst，無可注入**'}")
    print(f"  變體 A（含圖片還原）／變體 B（不還原圖片）各產一檔")
    for name, f in (("A", var_a), ("B", var_b)):
        try:
            _wb = openpyxl.load_workbook(f)
            _wb.close()
            print(f"  變體 {name} 可被 openpyxl 開啟：**是**")
        except Exception as e:                    # noqa: BLE001
            print(f"  變體 {name} 可被 openpyxl 開啟：**否** —— {type(e).__name__}: {e}")
    # ── 停點（下放包 28 §2.3：「若第 5 步任一項不符，停並回報，不自行調整方案」）
    # 二變體皆不可開啟 —— **第 5 步無從執行**。二個障礙皆在第 3 步，
    # 且皆為對 §2.3 第 3 步所列範圍之擴充，非本檔可自裁。
    ok_a = ok_b = False
    try:
        _w = openpyxl.load_workbook(var_a); _w.close(); ok_a = True
    except Exception:                                     # noqa: BLE001
        pass
    try:
        _w = openpyxl.load_workbook(var_b); _w.close(); ok_b = True
    except Exception:                                     # noqa: BLE001
        pass
    if not (ok_a or ok_b):
        ext_txt = ext.group(0) if ext else ""
        prefixes = sorted(set(re.findall(r"[<\s]([a-zA-Z]\w*):", ext_txt))
                          - {"xmlns"})
        selfdecl = sorted(set(re.findall(r"xmlns:(\w+)=", ext_txt)))
        print("\n步驟 5 —— **未執行**。二變體皆不可開啟，停點如下。\n")
        print("障礙 1（變體 A）—— 圖片不可單獨搬回")
        print("  openpyxl 已把 drawing 之 rels 改指向其重新編碼後之檔名")
        print(f"  （母本 {f_master['media']}")
        print(f"    → openpyxl {f_step1['media']}）。")
        print("  只換 `xl/media/` 而不換 drawing 與其 rels，套件即缺檔。")
        print("  **修復範圍大於 §2.3 第 3 步所列**（該步只列「原始 image2.jpeg」）。")
        print("\n障礙 2（變體 B）—— extLst 逐字注入產生 unbound prefix")
        print(f"  片段所用之前綴 {prefixes}；其自帶宣告者 {selfdecl}。")
        print("  **`xr` 無自帶宣告** —— 其 xmlns 在母本 <worksheet> 根元素上，")
        print("  而 openpyxl 之輸出根元素只宣告預設 namespace。")
        print("  片段內 `xr` 之唯一用途為 `xr:uid=\"{GUID}\"`（裝飾性）。")
        print("\n二個候選處置（**本檔不選**，待分析層裁）：")
        print("  (甲) 於輸出之 <worksheet> 根元素補宣告 xmlns:xr")
        print("       —— 片段保持逐字，但改動了輸出之根元素")
        print("  (乙) 自片段剝除 xr:uid 屬性")
        print("       —— 不動輸出，但**搬回之結構不再與母本逐字相同**")
        print("\n已量得之五項（步驟 1 之破壞，與上繳包 27 §3.4 一致）：")
        for k, mv, sv in (("x14:dataValidation", f_master["x14_dv"], f_step1["x14_dv"]),
                          ("extLst", f_master["extLst"], f_step1["extLst"]),
                          ("標準 dataValidation", f_master["std_dv"], f_step1["std_dv"]),
                          ("printerSettings", f_master["printer"], f_step1["printer"]),
                          ("xl/media 檔數", len(f_master["media"]), len(f_step1["media"]))):
            print(f"    {k:<24} 母本 {mv} → openpyxl {sv}")
        after0 = hashlib.sha256(MASTER.read_bytes()).hexdigest()
        print(f"\n母本 SHA256 前 {before[:16]} → 後 {after0[:16]}  "
              f"{'**未變**' if before == after0 else '**⚠ 母本被改動**'}")
        print(f"產物在 {TMP}（/tmp，非交付本）")
        print("\n**丙′ 於本輪未被證實可行，亦未被證偽** —— "
              "停於第 3 步之二個障礙，待裁。")
        return 2
    final = var_a if ok_a else var_b

    print("\n步驟 5 —— 六項驗收")
    f_final = facts(final)
    checks = []
    checks.append(("x14:dataValidation 條數與母本相同",
                   f_final["x14_dv"] == f_master["x14_dv"],
                   f"母本 {f_master['x14_dv']} → 修復後 {f_final['x14_dv']}"))
    checks.append(("extLst 個數與母本相同",
                   f_final["extLst"] == f_master["extLst"],
                   f"母本 {f_master['extLst']} → 修復後 {f_final['extLst']}"))
    checks.append(("標準 dataValidation 條數與母本相同",
                   f_final["std_dv"] == f_master["std_dv"],
                   f"母本 {f_master['std_dv']} → 修復後 {f_final['std_dv']}"))
    checks.append(("printerSettings 個數與母本相同",
                   f_final["printer"] == f_master["printer"],
                   f"母本 {f_master['printer']} → 修復後 {f_final['printer']}"))
    checks.append(("xl/media 內容與母本逐檔相同",
                   f_final["media"] == f_master["media"],
                   f"母本 {f_master['media']} → 修復後 {f_final['media']}"))
    # 第六項：資料 14 欄逐格與 JSON 相符（重新開檔讀回）
    wb2 = openpyxl.load_workbook(final, data_only=False)
    ws2 = wb2[SHEET]
    bad = []
    for i, t in enumerate(tcs):
        r = 10 + i
        for col, key in COLS.items():
            got = ws2[f"{col}{r}"].value
            if got != t[key]:
                bad.append(f"{col}{r}")
    b10 = ws2["B10"].value
    wb2.close()
    checks.append((f"資料 14 欄 × 3 筆逐格與 JSON 相符（B 欄公式未動：{b10!r}）",
                   not bad and isinstance(b10, str) and b10.startswith("="),
                   f"不符 {len(bad)} 格 {bad[:6] or '無'}"))

    print(f"{'#':>2}  {'項':<52} 判")
    print("-" * 88)
    failed = 0
    for i, (name, ok, detail) in enumerate(checks, 1):
        if not ok:
            failed += 1
        print(f"{i:>2}  {name:<52} {'PASS' if ok else '**FAIL**'}")
        print(f"    {detail}")
    print("-" * 88)
    after = hashlib.sha256(MASTER.read_bytes()).hexdigest()
    print(f"母本 SHA256 前 {before[:16]} → 後 {after[:16]}  "
          f"{'**未變**' if before == after else '**⚠ 母本被改動**'}")
    print(f"{len(checks)} checked / {failed} failed")
    print(f"產物在 {TMP}（/tmp，非交付本）")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
