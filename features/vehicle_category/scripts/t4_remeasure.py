#!/usr/bin/env python3
"""T4 —— 對 repo 內複本重測下放包 01 §三／§四之全部數字。

量測條件（R-G8 揭露）：
  - 讀檔一律 openpyxl `read_only=True, data_only=True`（取快取值，不重算公式）
  - 字串比對前一律 `strip()`，且視 U+00A0（NBSP）為空白（A-VC1）
  - 集合比對為**逐字**（大小寫敏感、不正規化底線／括號／空白）
  - 章節號取自 SYS1 `Basic Report` 之 Outline Number 欄，逐字為鍵

輸出：逐項「項目 / 下放包值 / 實測值 / = 或 ≠」。
本腳本只讀不寫，不修改任何素材。
"""
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
A03 = ROOT / "inputs/FM-WI-FSM-037-A03-N1L-SWE1-VehicleCategory-HMI-V0.1 STLA 報告.xlsx"
SYS1 = ROOT / "inputs/SYS1_HMI_Vehicle_Category_HMI_Logic_and_Flow_R1_SR24_Post_2A_(December_27_2023).xlsx"

NBSP = "\xa0"


def norm(v):
    """None / NBSP-only / 空白 一律歸為空字串；其餘 strip 後回傳原字。"""
    if v is None:
        return ""
    return str(v).replace(NBSP, " ").strip()


rows_out = []


def check(item, expected, measured, note=""):
    ok = expected == measured
    rows_out.append((item, expected, measured, "=" if ok else "≠", note))
    return ok


# ---------------------------------------------------------------- 037
wb = openpyxl.load_workbook(A03, read_only=True, data_only=True)
ws = wb["Analysis Report"]
raw = list(ws.iter_rows(values_only=True))

HDR = 7                                   # 表頭列（實測）
hdr = [norm(c) for c in raw[HDR - 1]]
col = {h: i for i, h in enumerate(hdr) if h}

# 資料列 = 表頭以下、A 欄（SWE-Requirement ID）非空者
data = []
first_row = last_row = None
for n, r in enumerate(raw[HDR:], start=HDR + 1):
    if norm(r[0]):
        data.append(r)
        first_row = first_row or n
        last_row = n

check("Analysis Report 資料列數", 145, len(data))
check("資料列起始列號", 8, first_row)
check("資料列結束列號", 152, last_row)

ids = [norm(r[0]) for r in data]

PARENT = re.compile(r"^SWE1-HMI-VC-(\d{3})$")
CHILD = re.compile(r"^SWE1-HMI-VC-(\d{3})-(\d{2})$")

parents = [i for i in ids if PARENT.match(i)]
children = [i for i in ids if CHILD.match(i)]
outside = [i for i in ids if not PARENT.match(i) and not CHILD.match(i)]

check("父需求 SWE1-HMI-VC-NNN 筆數", 66, len(parents))
check("子需求 SWE1-HMI-VC-NNN-MM 筆數", 79, len(children))
check("形態外之 id 筆數", 0, len(outside), repr(outside[:5]) if outside else "")

pnums = sorted(int(PARENT.match(i).group(1)) for i in parents)
check("父需求無重號（相異數）", 66, len(set(pnums)))
check("父需求連號 001–066（min,max）", (1, 66), (pnums[0], pnums[-1]))
check("父需求跳號數", 0, sum(1 for a, b in zip(pnums, pnums[1:]) if b - a != 1))

parent_of_child = {CHILD.match(i).group(1) for i in children}
has_child = [i for i in parents if PARENT.match(i).group(1) in parent_of_child]
no_child = [i for i in parents if PARENT.match(i).group(1) not in parent_of_child]
check("有子之父（不入 leaf）", 28, len(has_child))
check("無子之父（本身即 leaf）", 38, len(no_child))

leaves = set(no_child) | set(children)
check("leaf 全集", 117, len(leaves))

# --- 欄位分布
def dist(colname):
    return Counter(norm(r[col[colname]]) for r in data)

c = dist("Categorization")
check("Categorization = Functional Requirement", 145, c.get("Functional Requirement", 0),
      f"全分布={dict(c)}")
rv = dist("Release Version")
check("Release Version = 1.00.00", 145, rv.get("1.00.00", 0), f"全分布={dict(rv)}")

frop_key = next(h for h in col if h.startswith("FROP"))
fr = dist(frop_key)
check("FROP = Vehicle Settings", 128, fr.get("Vehicle Settings", 0))
check("FROP = Power Management", 16, fr.get("Power Management", 0))
check("FROP = Audio Management", 1, fr.get("Audio Management", 0))

sc = dist("Sub Categorization")
check("Sub Categorization = HMI", 103, sc.get("HMI", 0))
check("Sub Categorization = Service", 42, sc.get("Service", 0))

srid = {norm(r[col["Source Requirement ID"]]) for r in data} - {""}
check("Source Requirement ID 相異值", 61, len(srid))

hmi_col = col["HMI Source ID"]
hmi_vals = {norm(r[hmi_col]) for r in data} - {""}
check("HMI Source ID 相異值", 66, len(hmi_vals))

# --- Verification Method 於 117 leaf 之單一起首
vm_col = col[next(h for h in col if h.startswith("Verification Method"))]
PREFIX = "Manual functional test on the target head unit"
leaf_rows = [r for r in data if norm(r[0]) in leaves]
hit = sum(1 for r in leaf_rows if norm(r[vm_col]).startswith(PREFIX))
check("Verification Method 起首一致（117 leaf）", 117, hit,
      f"母體={len(leaf_rows)} 起首={PREFIX!r}")

# --- 第 10–18 欄之形態。
# 基準已由 R-VC6（下放包 02 §二）取代下放包 01 §3.3 —— 後者為分析層
# 未經全表掃描之全稱斷言，作廢；A-VC1 一併撤銷。
# R-VC6 所裁之形態：117 個 leaf 皆有實質內容；28 個「有子之父」為 `\xa0`；
# 欄 16／17 之該 28 列中另有 3 列為 None（VC-034 / VC-052 / VC-063）。
NONE_ROWS = {"SWE1-HMI-VC-034", "SWE1-HMI-VC-052", "SWE1-HMI-VC-063"}
parents_all = set(has_child) | set(no_child)
non_leaf = {i for i in ids if i not in leaves}          # 28 個有子之父
rvc6_cols, rvc6_detail = [], []
for idx in range(9, 18):                       # 0-based 9..17 = 第 10..18 欄
    filled = {norm(r[0]) for r in data
              if idx < len(r) and r[idx] is not None and str(r[idx]) != NBSP}
    nbsp = {norm(r[0]) for r in data if idx < len(r) and str(r[idx]) == NBSP}
    none_ = {norm(r[0]) for r in data if idx < len(r) and r[idx] is None}
    ok = (filled == leaves
          and nbsp | none_ == non_leaf
          and none_ == (NONE_ROWS if hdr[idx].startswith(
              ("Reusable", "Description/Action for Reusable")) else set()))
    if ok:
        rvc6_cols.append(hdr[idx])
    else:
        rvc6_detail.append(
            f"{hdr[idx]}: filled={len(filled)} nbsp={len(nbsp)} none={sorted(none_)}")
check("第 10–18 欄符合 R-VC6 所裁之形態（欄數）", 9, len(rvc6_cols),
      "; ".join(rvc6_detail) if rvc6_detail else "九欄全符")

# ---------------------------------------------------------------- SYS1
wb1 = openpyxl.load_workbook(SYS1, read_only=True, data_only=True)
sheets = wb1.sheetnames
ws1 = wb1["Basic Report"]
raw1 = list(ws1.iter_rows(values_only=True))
hdr1 = [norm(x) for x in raw1[0]]
col1 = {h: i for i, h in enumerate(hdr1) if h}

out_idx = col1["Outline Number"]
sysre_idx = col1[next(h for h in col1 if "SYSRE_HMI_Source ID" in h)]

body = raw1[1:]
outlines = [norm(r[out_idx]) for r in body]
valid_out = [o for o in outlines if o]
check("SYS1 Basic Report 資料列（列 2–110）", 109, len(body))
check("有效 Outline Number", 108, len(valid_out))

sysre_vals = {norm(r[sysre_idx]) for r in body} - {""}
hits = hmi_vals & sysre_vals
check("037 HMI Source ID → SYS1 SYSRE_HMI_Source ID 命中", 66, len(hits))
check("命中但不在 SYS1 者", 0, len(hmi_vals - sysre_vals))

# --- SYS-HMI-RA 於三分頁全儲存格
total = 0
per_sheet = {}
for sn in sheets:
    w = wb1[sn]
    n = 0
    for r in w.iter_rows(values_only=True):
        for cell in r:
            if cell is not None and "SYS-HMI-RA" in str(cell):
                n += 1
    per_sheet[sn] = n
    total += n
check("字串 SYS-HMI-RA 於 SYS1 全簿出現次數", 0, total, f"分頁={sheets} 逐頁={per_sheet}")

# --- 未引用章節
cited = {v.rsplit("_", 1)[-1] for v in hmi_vals}
sec_by_outline = {}
for r in body:
    o = norm(r[out_idx])
    if o:
        sec_by_outline[o] = norm(r[sysre_idx])
cited_outlines = {o for o, s in sec_by_outline.items() if s in hmi_vals}
uncited = sorted(set(sec_by_outline) - cited_outlines,
                 key=lambda s: [int(x) for x in s.split(".") if x.isdigit()])
check("引用章節數", 66, len(cited_outlines))
check("未引用章節數", 42, len(uncited), f"清單={uncited}")

# ---------------------------------------------------------------- 輸出
w1 = max(len(r[0]) for r in rows_out) + 2
print(f"{'項目':<{w1}} {'下放包':>18} {'實測':>18}  判")
print("-" * (w1 + 46))
bad = 0
for item, exp, mea, mark, note in rows_out:
    if mark == "≠":
        bad += 1
    print(f"{item:<{w1}} {str(exp):>18} {str(mea):>18}  {mark}")
    if note:
        print(f"{'':<{w1}}   note: {note}")
print("-" * (w1 + 46))
print(f"{len(rows_out)} 項；≠ {bad} 項")
sys.exit(1 if bad else 0)
