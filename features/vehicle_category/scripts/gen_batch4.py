#!/usr/bin/env python3
"""第 4 批 `Settings Behavior` 之生成（下放包 23 T124）。

15 leaf → 16 TC（`038-05` 拆 2）。a 段 15、b 段 0、**PENDING 0**。

上半之取材一律**自來源檔逐字取出**，不在本檔內手抄 ——
手抄之逐字與來源之逐字是二件事，而第 7b 項只驗前者對得上後者，
**驗不出「我抄的時候看的是不是這一份」**。故本檔以 leaf_id 取值。
"""
import csv
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
A03 = ROOT / ("inputs/FM-WI-FSM-037-A03-N1L-SWE1-VehicleCategory-HMI-V0.1"
              " STLA 報告.xlsx")
S1 = ROOT / ("inputs/SYS1_HMI_Vehicle_Category_HMI_Logic_and_Flow_"
             "R1_SR24_Post_2A_(December_27_2023).xlsx")
OUT = ROOT / "generated/batch4_settings_behavior.json"

src = {}
for r in list(openpyxl.load_workbook(A03, read_only=True, data_only=True)
              ["Analysis Report"].iter_rows(values_only=True))[7:]:
    if r[0] not in (None, ""):
        src[str(r[0]).strip()] = (str(r[3]).strip(), str(r[4]).strip())

sys1 = {}
_rows = list(openpyxl.load_workbook(S1, read_only=True, data_only=True)
             ["Basic Report"].iter_rows(values_only=True))
_h = [str(c).strip() if c else "" for c in _rows[0]]
_oi, _di = _h.index("Outline Number"), _h.index("Description")
for _r in _rows[1:]:
    _o = str(_r[_oi]).strip() if _r[_oi] else ""
    if _o:
        sys1[_o] = ((str(_r[_di]) if _r[_di] else "")
                    .replace("_x000D_\n", "\n").replace("_x000D_", " "))

recon = {r["req_id"]: r for r in csv.DictReader(
    (ROOT / "data/recon_leaf_to_section.tsv").open(encoding="utf-8"),
    delimiter="\t")}
prio = {r["req_id"]: r["final_p"] for r in csv.DictReader(
    (ROOT / "data/priority_final.tsv").open(encoding="utf-8"), delimiter="\t")}


def sent(sec, idx):
    """SYS1 之句 —— 切分沿用全案同一式（`(?<=\\.)\\s+(?=[A-Z])`）。"""
    parts = [s.strip() for s in
             re.split(r"(?<=\.)\s+(?=[A-Z])", sys1[sec].strip())]
    if isinstance(idx, tuple):
        return " ".join(parts[idx[0] - 1:idx[1]])
    return parts[idx - 1]


TAB = 'Open the Vehicle Category screen and select the "Settings" tab'
TAB_ER = 'The Vehicle Category screen is displayed with the "Settings" tab active'

FUNC = "功能測試 (Functional based ; no specific technique)"
STATE = "狀態轉換 (State Transition Testing)"
TABLE = "決策表 (Decision Table Testing)"

# `upper` 之取法：("D", leaf) = 037 Description；("T", leaf) = 037 Title；
# ("S", 節, 句) = SYS1。**無字面上半** —— 皆自來源取。
SPEC = [
 dict(leaf="SWE1-HMI-VC-034-01", upper=("D", None), dm=FUNC,
      title="Settings absent from the vehicle are not listed",
      low="Applicability -- a setting the vehicle does not contain is absent from its list",
      pc=["The vehicle under test is a configuration that does not contain the "
          "setting named in the test data"],
      data="A setting that the vehicle under test does not contain, named from "
           "the HMI Settings List",
      pr=[TAB,
          "Record every item in the Settings list and compare it against the "
          "setting named in the test data"],
      er=[TAB_ER,
          "The setting named in the test data is not present anywhere in the "
          "Settings list"],
      axis="不適用者之處置：隱藏（對 -02 之灰化）",
      why="**驗證目標**：不屬於該車之設定不出現於其 Settings 清單。"
          "**取材（R-VC25）**：上半取自 037 `Description`（優先序第 1）。"
          "**為什麼這樣切**：§11.1 之二規則由 037 拆為 -01／-02 二 leaf，"
          "本筆只驗「隱藏」，灰化屬 -02（§8.2.1）。"
          "**範圍**：不驗該設定於其他車型是否出現 —— 那是他車之組態，"
          "非本需求所斷言（§8.4.2）。"),
 dict(leaf="SWE1-HMI-VC-034-02", upper=("D", None), dm=FUNC,
      title="Key Off greys out unavailable settings",
      low="Key Off -- an unavailable setting greys out instead of disappearing",
      pc=["The vehicle under test contains a setting that the system does not "
          "offer while in Key Off"],
      data="A setting that is available to the vehicle but not in key-off",
      pr=["Place the vehicle in Key Off",
          "Open the Phone settings through the Phone screens",
          "Record how the setting named in the test data is rendered in that list"],
      er=["The system is in Key Off",
          "The Phone settings are reachable through the Phone screens while the "
          "system is in Key Off",
          "The setting named in the test data is present in the list and is "
          "rendered grey"],
      axis="不適用者之處置：灰化（對 -01 之隱藏）",
      why="**驗證目標**：車輛有、但 key-off 下不可用之設定，於 key-off 呈灰而非消失。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**⚠ 進入路徑之拘束（上繳包 22 §2.3）**：SYS1 §13.1 於 Key Off／"
          "Timed Mode／ACC 擋住 **Settings 頁籤**，故本筆**不得以該頁籤進入**"
          "，否則於待測狀態下不可執行。改經 §13.2 明文於 Key Off 可用之 "
          "Phone screens 進入 Phone settings。§11.1 與 §13.1 **非衝突** ——"
          "一管入口、一管入口之後的呈現，§13.5 之 `tab or a Settings category` "
          "為旁證，故不發 DR。"
          "**⚠ 測試資料未具名（§8.4.1）**：`HMI Settings List R1 SR25 Post "
          "R1L-R (Feb 13 2026)` 之 `Settings` 分頁已實測搜尋 "
          "`key-off`／`key off`／`ignition`／`ACC`／`Timed Mode`（命中 21 列）"
          "與 `grey`／`gray`（命中 15 列）—— **其所載之灰化成因為頭燈關閉、"
          "Display Mode 為 Auto、sync 選項、Steering only、車輛行進中，"
          "無一為 key-off**；該清單無「key-off 可用性」之欄位。"
          "故不自行指定某一設定，以規格語言之通稱表述之。"),
 dict(leaf="SWE1-HMI-VC-035-01", upper=("D", None), dm=FUNC,
      title="Restore defaults resets the settings",
      low="Restore defaults confirmed -- the settings actually return to their default values",
      pc=["At least one setting holds a value other than its default"],
      data="NA",
      pr=[TAB,
          "Record the current value of the setting that holds a non-default value",
          'Open the restore defaults prompt and press "Yes"',
          "Record the value of that setting again"],
      er=[TAB_ER,
          "The setting reads its non-default value",
          "The restore defaults action is accepted",
          "The setting reads its default value"],
      axis="回復預設之三段：值之生效（對 -02 之確認彈窗、-03 之取消）",
      why="**驗證目標**：於回復預設之提示選 Yes，設定確實回到預設值。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**為什麼含前後二次記錄**：「回到預設」為值之變化，"
          "無變化前之值即無從判其已變（§5.6 之 baseline）。"),
 dict(leaf="SWE1-HMI-VC-035-02", upper=("D", None), dm=FUNC,
      title="Confirmation pop-up after a reset to default",
      low="Confirmation pop-up -- its appearance and its wording after the reset",
      pc=["At least one setting holds a value other than its default"],
      data="NA",
      pr=[TAB,
          'Open the restore defaults prompt and press "Yes"',
          "Record the pop-up that is displayed and its text"],
      er=[TAB_ER,
          "The restore defaults action is accepted",
          'A pop-up is displayed and its text reads "Settings reset to default"'],
      axis="回復預設之三段：確認彈窗（對 -01 之值生效、-03 之取消）",
      why="**驗證目標**：回復完成後顯示 `Settings reset to default` 之彈窗。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**⚠ 二欄記法不對稱（A-VC10 第三面）**：本 leaf 之 Description 用彎單引號 "
          "`‘…’`、Title 用直單引號 `'…'`。**取 Description 一欄，不混用**；"
          "上半之彎單引號依 R-VC23 逐字保留，"
          "ER 之彈窗文字為作者散文，依 R-VC23(b) 用直雙引號。"),
 dict(leaf="SWE1-HMI-VC-035-03", upper=("T", None), dm=FUNC,
      title="Cancel leaves the settings unchanged",
      low="Restore-defaults prompt cancelled -- baseline recorded before the cancel and re-read after",
      pc=["The settings named in the test data hold known values other than "
          "their defaults"],
      data="Three settings that currently hold non-default values, read before "
           "and after the cancel. The count of three is a test-design parameter "
           "and is not stated by the source",
      pr=[TAB,
          "Record the current value of each setting named in the test data",
          'Open the restore defaults prompt and press "Cancel"',
          "Record the value of each setting named in the test data again"],
      er=[TAB_ER,
          "The recorded values form the baseline for step 4",
          "The previous screen is displayed",
          "Each recorded value is identical to the baseline recorded in step 2"],
      axis="回復預設之三段：取消（對 -01 之值生效、-02 之確認彈窗）",
      why="**驗證目標**：於回復預設之提示選 Cancel，返回前一畫面且設定未變。"
          "**⚠ 取材為 R-VC25 之例外路徑（Title），三件逐筆記**："
          "(a) **理由** —— 本筆與 `036-02` 之 Description **逐字相同**"
          "（`Selecting cancel will take the user back to the previous screen.`），"
          "而其 P0 之依據 `without changing any settings` **只在 Title**；"
          "Description 未載該條件，取之則本 TC 之驗證標的落空（A-VC10 第一面）。"
          "(b) **R-VC24 判別結果** —— Title 之謂語為 `returns the user…`，"
          "為本 leaf 之行為；`restore-defaults prompt` 用以定位是哪一個提示，"
          "屬**情境脈絡**。(c) **非行為主張** —— 由 (b) 滿足。"
          "**ER 之 baseline（§5.6）**：「未變」無變前之值即不可判，"
          "故第 2 步記錄、第 4 步回讀。"
          "**測試資料之三筆為測試設計參數**，非來源所載（§8.4.1）。"),
 dict(leaf="SWE1-HMI-VC-036-01", upper=("D", None), dm=FUNC,
      title="Clear personal data confirmed",
      low="Clear personal data confirmed -- the data is cleared and the pop-up carries its X button",
      pc=["Personal data is stored in the vehicle under test"],
      data="NA",
      pr=[TAB,
          "Record the personal data that is stored",
          'Open the clear personal data prompt and press "Yes"',
          "Record the pop-up that is displayed, its text and the controls it carries",
          "Record the personal data that is stored again"],
      er=[TAB_ER,
          "The stored personal data is present",
          "The clear personal data action is accepted",
          'A pop-up is displayed, its text reads "Personal data cleared" and it '
          'carries an "X" button in the top right corner',
          "The personal data recorded in step 2 is no longer stored"],
      axis="清除個人資料之二支：執行（對 -02 之取消）",
      why="**驗證目標**：選 Yes 清除個人資料，並顯示帶右上 X 之確認彈窗。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**⚠ 二欄記法不對稱（A-VC10 第三面）**：Description 用彎單引號、"
          "Title 用直單引號 —— 取 Description 一欄。"
          "**⚠ R-VC14(b) 之分歧揭露**：037 給 Medium，本地初判 P0（資料遺失風險），"
          "**改判 P1** —— 本筆之失效為「該清而未清」，資料仍在，"
          "**非 data-loss**；其風險為隱私外洩，依 R-VC11(c) 記於本欄而"
          "**不入 priority**。取消支（-02）之失效才是靜默清除，故其為 P0。"
          "**為什麼一 TC 不拆**：清除與彈窗為同一觸發（按 Yes）之二個後果，"
          "IN §398 明文不拆，列為二條 ER。"),
 dict(leaf="SWE1-HMI-VC-036-02", upper=("T", None), dm=FUNC,
      title="Cancel leaves the personal data intact",
      low="Clear-personal-data prompt cancelled -- stored data re-read against the baseline",
      pc=["Personal data is stored in the vehicle under test"],
      data="The stored personal data that is read before and after the cancel",
      pr=[TAB,
          "Record the personal data that is stored",
          'Open the clear personal data prompt and press "Cancel"',
          "Record the personal data that is stored again"],
      er=[TAB_ER,
          "The recorded personal data forms the baseline for step 4",
          "The previous screen is displayed",
          "The stored personal data is identical to the baseline recorded in step 2"],
      axis="清除個人資料之二支：取消（對 -01 之執行）",
      why="**驗證目標**：於清除個人資料之提示選 Cancel，返回前一畫面且資料未被清除。"
          "**⚠ 取材為 R-VC25 之例外路徑（Title），三件逐筆記**："
          "(a) **理由** —— 本筆與 `035-03` 之 Description 逐字相同，"
          "而其 P0 之依據 `without clearing any data` **只在 Title**（A-VC10 第一面）。"
          "(b) **R-VC24 判別結果** —— Title 之謂語為 `returns the user…`，"
          "為本 leaf 之行為；`clear-personal-data prompt` 為**情境脈絡**。"
          "(c) **非行為主張** —— 由 (b) 滿足。"
          "**與 `035-03` 之區分**：二筆之 Title 僅以提示之別區分"
          "（restore-defaults／clear-personal-data），括號下半即以此區分。"
          "**ER 之 baseline（§5.6）**：「未被清除」須有清除前之內容可比。"),
 dict(leaf="SWE1-HMI-VC-037-01", upper=("D", None), dm=FUNC,
      title="Only one suspension mode is on",
      low="The rule itself -- at most one suspension mode reads on at any moment",
      pc=["The vehicle under test is equipped with Suspension settings that "
          "offer more than one suspension mode"],
      data="NA",
      pr=[TAB,
          "Open the Suspension settings and record the state of every suspension mode",
          "Press each suspension mode in turn and record the state of every "
          "suspension mode after each press"],
      er=[TAB_ER,
          "Exactly one suspension mode reads on",
          "Exactly one suspension mode reads on after each press"],
      axis="互斥之靜態面：任一時刻之狀態（對 -02 之切換動作）",
      why="**驗證目標**：懸吊模式於任一時刻至多一者為開。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**一靜一動之區分（下放包 20 §3.4／上繳包 22 §5.2）**："
          "本筆驗**規則**（同時僅一），`-02` 驗**行為**（開一關餘）；"
          "括號下半以此區分。本筆之第 3 步逐一按過每個模式，"
          "其驗的是每次之後不變式仍成立，非某一次之轉換結果。"),
 dict(leaf="SWE1-HMI-VC-037-02", upper=("D", None), dm=STATE,
      title="Activating a mode turns the others off",
      low="The transition -- turning one on drives the others off",
      pc=["The vehicle under test is equipped with Suspension settings that "
          "offer more than one suspension mode"],
      data="NA",
      pr=[TAB,
          "Open the Suspension settings and record the state of every suspension mode",
          "Press a suspension mode that currently reads off",
          "Record the state of every suspension mode again"],
      er=[TAB_ER,
          "The recorded states form the baseline for step 4",
          "The pressed suspension mode is accepted",
          "The pressed suspension mode reads on and every suspension mode that "
          "read on in the baseline now reads off"],
      axis="互斥之動態面：開一關餘之轉換（對 -01 之不變式）",
      why="**驗證目標**：開啟一個懸吊模式時，其餘自動關閉。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**ER 之 baseline（§5.6，下放包 23 §3.2）**：「其餘被關掉」"
          "須先知道原本哪些是開的 —— 無 baseline 則「餘者為 off」"
          "與「餘者本來就 off」不可分。"
          "**一靜一動之區分**：本筆為行為，`-01` 為規則。"),
 dict(leaf="SWE1-HMI-VC-038-01", upper=("D", None), dm=FUNC,
      title="Progress pop-up on a language change",
      low="Language change triggered -- the progress pop-up appears with its stated text",
      pc=["The vehicle under test offers more than one language in the language "
          "settings"],
      data="NA",
      pr=[TAB,
          "Open the language settings and select a language other than the current one",
          "Record the pop-up that is displayed and its text"],
      er=[TAB_ER,
          "The language selection is accepted",
          'A pop-up is displayed and its text reads "Language updated, voice '
          'command change in process…"'],
      axis="語言變更五段：彈窗之出現（對 -02 之語言、-03 之持續、-04 之返回、-05 之清單呈現）",
      why="**驗證目標**：選擇變更語言時出現進度彈窗，其文字為所載者。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**⚠ 二欄記法不對稱（A-VC10 第三面）**：Description 用彎單引號、"
          "Title 用直單引號 —— 取 Description 一欄。"
          "ER 之彈窗文字含來源之刪節號 `…`，逐字保留（R-VC23(c)）；"
          "其引號為作者散文之直雙引號（R-VC23(b)）。"),
 dict(leaf="SWE1-HMI-VC-038-02", upper=("S", "11.5", (1, 2)), dm=FUNC,
      title="Pop-up appears in the new language",
      low="Pop-up language -- rendered in the newly selected language, not English",
      pc=["The current language of the vehicle under test is English"],
      data="A target language other than English that the vehicle offers",
      pr=[TAB,
          "Open the language settings and select the target language named in "
          "the test data",
          "Record the language in which the pop-up text is rendered"],
      er=[TAB_ER,
          "The language selection is accepted",
          "The pop-up text is rendered in the target language named in the test "
          "data and not in English"],
      axis="語言變更五段：彈窗之呈現語言（對 -01 之出現）",
      why="**驗證目標**：進度彈窗以新選之語言呈現，非固定英文。"
          "**⚠ 取材為 CONT 之指涉型（R-VC25 優先序第 2）**：037 `Description` "
          "之 `It` 其先行詞（pop-up）在 s1，取單句則指涉無解。"
          "**登記 SYS1 §11.5 範圍 `1-2`，33 token，未逾 R-3 之 50**"
          "（profile §9.2 層次 1 之預設處置對本筆成立，故不採第三處置類）。"
          "上半即 s1＋s2 之逐字，收斂第 16 項逐字比對。"),
 dict(leaf="SWE1-HMI-VC-038-03", upper=("S", "11.5", 3), dm=STATE,
      title="Pop-up stays until completion or X",
      low="Persistence -- the pop-up leaves only on completion or on X",
      pc=["The language-change pop-up is displayed and the system has not "
          "completed changing the voice commands"],
      data="NA",
      pr=["Record the screen while the system is still changing the voice commands",
          "Press the X button on the pop-up",
          "Record the screen again"],
      er=["The pop-up is still displayed",
          "The X press is accepted",
          "The pop-up is no longer displayed"],
      axis="語言變更五段：彈窗之持續與離開條件（對 -01 之出現、-04 之返回）",
      why="**驗證目標**：彈窗持續顯示，直到系統完成或使用者按 X。"
          "**⚠ 取材為第三處置類 `resolved-by-structure`（profile §9.2 層次 2，"
          "下放包 23 §2.2）**：`This pop-up` 之先行詞在 SYS1 §11.5 s1，"
          "而連續 `1-3` 為 **54 token，逾 R-3 之 50**（profile §10 之工作定義）；"
          "非連續 `1,3` 為 42 但破壞 verbatim 之連續性，"
          "使第 7b 項與第二層之子串判準對本筆失效。"
          "**採單句 s3 ＋ 指涉由 TC 結構承載** —— 其先行詞「語言變更彈窗已顯示」"
          "**本即本 TC 驗證持續性之前提**，不解指涉也必須建立它。"
          "CONT 登記 `resolution=PC`／`resolution_key=pop-up`，"
          "**第三檢查點驗其 pre_conditions 確含該詞**。"
          "**⚠ 不拆之判讀（人工，記明）**：本筆含二個離開條件"
          "（系統完成、使用者按 X）。§8.3 壓測下二者為同一規則"
          "「彈窗不自行消失」之二個邊界，而非二個獨立觸發，故依授權不拆；"
          "惟此為人工判讀，若上游認為應拆，本筆為第一候選。"),
 dict(leaf="SWE1-HMI-VC-038-04", upper=("S", "11.5", 4), dm=STATE,
      title="Return to the language settings screen",
      low="Return target -- the screen the user lands on once the pop-up is gone",
      pc=["A language change has been selected from the language settings screen"],
      data="NA",
      pr=["Press the X button on the language-change pop-up",
          "Record the screen that is displayed"],
      er=["The X press is accepted",
          "The language settings screen is displayed"],
      axis="語言變更五段：彈窗關閉後之落點（對 -03 之持續）",
      why="**驗證目標**：彈窗結束後使用者返回語言設定畫面。"
          "**⚠ 第一層偽陰性之第二實例（上繳包 22 §6）**：037 之 "
          "`The user is **then** taken back…` 以 `then` 承接前句，"
          "但首字大寫、非代名詞起首 —— 二特徵皆不命中，"
          "候選偵測看不到，由勘查之 SYS1 對照發現。"
          "**取材為第三處置類**：`then` 所承接之「按 X 或系統完成」"
          "**即本 TC 之必然步驟**，故取單句 s4，"
          "CONT 登記 `resolution=Step`／`resolution_key=pop-up`，"
          "**第三檢查點驗其 test_procedure 確含該詞**。"),
 dict(leaf="SWE1-HMI-VC-038-05", upper=("D", None), dm=TABLE,
      title="Language screen normal once updating completes",
      low="Voice commands complete -- the language settings screen renders as normal",
      pc=["A language change has been made and the system has completed "
          "changing the voice commands"],
      data="NA",
      pr=["Open the language settings screen",
          "Record how every language option is rendered"],
      er=["The language settings screen is displayed",
          "Every language option is rendered as normal and none of them is grey"],
      axis="更新完成與否：完成支（對未完成支之勾選＋灰化）",
      why="**驗證目標**：語音命令更新完成後，語言設定畫面呈現如常。"
          "**取材（R-VC25）**：上半取自 037 `Description`（含 s5＋s6 二句）。"
          "**⚠ 拆 2 之理由（IN §8.2.2／§5.2，上繳包 22 §5.1）**："
          "s5 自身即二個 if 分支（完成 → 如常；未完成 → 現用語言勾選、"
          "其餘灰化），二者為**二個獨立失效**，單一 TC 之判準不明；"
          "且 IN §5.2 禁一 TC 內寫條件分支。二筆同 req_id，"
          "括號下半以分支區分（IN §8.2.2：sub-id 數 ≠ TC 數）。"),
 dict(leaf="SWE1-HMI-VC-038-05", upper=("D", None), dm=TABLE,
      title="Other languages grey while updating runs",
      low="Voice commands not complete -- current language checked, the rest greyed until completion",
      pc=["A language change has been made and the system has not completed "
          "changing the voice commands"],
      data="NA",
      pr=["Open the language settings screen",
          "Record how every language option is rendered",
          "Wait until the system has completed changing the voice commands and "
          "record how every language option is rendered again"],
      er=["The language settings screen is displayed",
          "The current language is rendered checked and every other language "
          "option is rendered grey",
          "The other language options are no longer grey"],
      axis="更新完成與否：未完成支（對完成支之如常呈現）",
      why="**驗證目標**：更新未完成時，現用語言勾選、其餘灰化，且灰化持續至完成。"
          "**取材（R-VC25）**：上半同完成支，取自 037 `Description`。"
          "**為什麼第 3 步在本支而不在完成支**：s6 之「持續至完成」"
          "是對**未完成**狀態之時間性斷言，其終點才是完成；"
          "把它放到完成支則無灰化可觀察。"
          "**與完成支之區分**：括號下半載其分支條件。"),
 dict(leaf="SWE1-HMI-VC-039", upper=("D", None), dm=FUNC,
      title="Chinese language change pop-up",
      low="Chinese selected -- the head unit pop-up and its X/Close button",
      pc=["The vehicle under test offers Chinese in the language settings"],
      data="NA",
      pr=[TAB,
          "Open the language settings and select Chinese",
          "Record the pop-up that is displayed, its text and the controls it carries"],
      er=[TAB_ER,
          "The Chinese language selection is accepted",
          'A pop-up is displayed, its text reads "Language updates in '
          'progress...Driver screen only will display language in Chinese." '
          "and it carries an X/Close button"],
      axis="語言之特定值：中文（對 -038 各筆之語言無關流程）",
      why="**驗證目標**：語言改為中文時顯示所載之彈窗，帶 X/Close 按鍵。"
          "**取材（R-VC25）**：上半取自 037 `Description`。"
          "**記法（R-VC23）**：上半之彎雙引號 `“…”` 與 `X/Close` 之斜線逐字保留；"
          "ER 之彈窗文字為作者散文，依 R-VC23(b) 用直雙引號，"
          "其內容含來源之三點刪節 `...` 逐字。"
          "**⚠ 範圍（§8.4.2）**：來源之 `Driver screen only will display "
          "language in Chinese` 為彈窗**文字之內容**，"
          "其所述之叢集（Driver screen）中文顯示屬**叢集側之行為**，"
          "非本 HMI 需求所斷言 —— 本筆之 ER 限於 HU 彈窗之出現與其文字，"
          "叢集顯示委派記於本欄，不入 ER。"),
]


def upper_of(sp, leaf):
    kind = sp["upper"][0]
    if kind == "T":
        return src[leaf][0]
    if kind == "D":
        return src[leaf][1]
    return sent(sp["upper"][1], sp["upper"][2])


def numbered(xs):
    return "\n".join(f"{i}. {x}" for i, x in enumerate(xs, 1))


tcs = []
for sp in SPEC:
    leaf = sp["leaf"]
    up = upper_of(sp, leaf)
    tcs.append({
        "leaf_id": leaf,
        "test_group": "Vehicle Category",
        "test_set": "Settings Behavior",
        "tc_title": sp["title"],
        "test_item": f"{up}\n\n({sp['low']})",
        "pre_conditions": numbered(sp["pc"]),
        "input_test_data": sp["data"],
        "test_procedure": numbered(sp["pr"]),
        "expected_result": numbered(sp["er"]),
        "specification_reference": recon[leaf]["spec_reference"],
        "design_method": sp["dm"],
        "priority": prio[leaf],
        "split_flag": False,
        "split_reason": "",
        "functional_safety": "NA",
        "reasoning": sp["why"],
        "distinguishing_axis": sp["axis"],
    })

doc = {
    "batch": "batch4_settings_behavior",
    "feature": "vehicle_category",
    "test_group": "Vehicle Category",
    "test_set": "Settings Behavior",
    "handoff": "docs/handoff/23_batch4_tc.md（勘查見 22）",
    "ruling": "R-VC21／R-VC22／R-VC23／R-VC24／R-VC25 ＋ profile §9.2 第三處置類",
    "segment": "a",
    "segment_note": "**本批無 b 段** —— 15 leaf 全數生成，"
                    "為首個 a 段即全批之批次。`held_leaves` 為空。",
    "split_delta": 1,
    "tc_id_status": "provisional",
    "leaf_scope": [x["leaf"] for i, x in enumerate(SPEC)
                   if x["leaf"] not in [y["leaf"] for y in SPEC[:i]]],
    "held_leaves": [],
    "write_back": "凍結 —— 本輪只產出 JSON，不寫回工作簿",
    "reasoning": "**驗證目標**：Settings 之行為面 —— 清單可見性（隱藏／灰化）、"
                 "回復預設與清除個人資料之確認與取消、懸吊模式互斥、"
                 "語言變更之彈窗流程與中文之特定彈窗。"
                 "**為什麼這樣切**：15 leaf 中 14 筆一 leaf 一 TC；"
                 "`038-05` 因 s5 自身即二個 if 分支而拆 2（`split_delta: 1`）。"
                 "**PENDING 0** —— 本批為首個全潔批：15 筆皆僅需 037／SYS1 文字層。"
                 "**取材分布之預期**：`Title` 2 筆（`035-03`／`036-02`，"
                 "R-VC25 例外路徑之首次動用）、`SYS1` 3 筆（CONT 之 "
                 "`038-02`／`038-03`／`038-04`）、其餘 `Description`。"
                 "**未涵蓋**：Settings 清單之版面與操作（屬 `Settings List`）、"
                 "點火可用性之狀態規則（屬 `Ignition Availability`）、"
                 "`039` 所涉之叢集中文顯示（§8.4.2 委派）。",
    "tcs": tcs,
}
OUT.write_text(json.dumps(doc, ensure_ascii=False, indent=1) + "\n", "utf-8")
print(f"{OUT.relative_to(ROOT)} — {len(tcs)} TC / "
      f"{len(doc['leaf_scope'])} leaf / split_delta {doc['split_delta']}")
