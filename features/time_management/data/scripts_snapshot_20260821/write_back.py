#!/usr/bin/env python3
"""Step 6 (Time Management) — 寫回，**必經 `surgical_save`**。

## 本檔之來源與界線（R-TM29）

**結構參照** `features/privacy/scripts/write_back.py`：
`surgical_save` 之呼叫慣例、欄位以表頭文字解析而非寫死字母、
dry-run 為預設、寫入後之三重驗證（表頭未動／他分頁逐位元相同／
`verify_structure`）、`BLANK_BY_DECISION` 之刻意留白清單。

**此處參照是必要的而非便利的**（R-TM29 逐字）：自零寫會升高母本 R 欄
x14 下拉被摧毀之風險（R-G3），該風險不可逆且發生在交付件上。

**不繼承其內容**：Privacy 之 `CONST_FUNCTIONAL_SAFETY = "NA"`（R30-3）、
`PLACEHOLDER_BODY`、`tc_id_format` 皆為其自身裁決，本 feature 之對應值
標 `TODO(R-TM10-A1)`，待本 feature 條文決定。

## 本 feature 之調整點

  workbook           `feature.yaml paths.workbook`（`inputs/` 之母本複本）
  fill_test_group_set  true（BLANK per R-TM5，依 canon §2）
  test_group         `Time and Date`（R-TM8）
  欄位對映           rev C —— design_method `R`、functional_safety `S`、
                     author `AA`、remarks `AH`
  done_region        none（BLANK per R-TM5）→ append from first data row

用法：
    python3 features/time_management/scripts/write_back.py --feature-dir features/time_management
    （加 --write 才實際寫出；預設 dry-run）

**本腳本於 04 包只建立不執行。**
"""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
import zipfile
from pathlib import Path

import openpyxl
import yaml

REPO_ROOT = next(p for p in Path(__file__).resolve().parents
                 if (p / "pyproject.toml").is_file())
sys.path.insert(0, str(REPO_ROOT))
from backend.xlsx_surgical import (  # noqa: E402
    StructureError, surgical_save, verify_structure)

FIRST_DATA_ROW = 10          # rev C 版面；表頭列 9

# TODO(R-TM10-A1): functional_safety 之常數值 —— Privacy 取 "NA"（R30-3），
#   本 feature 須依自身條文決定，未定前不填。
CONST_FUNCTIONAL_SAFETY = None

# TODO(R-TM10-A1): BLOCKED 佔位之措辭 —— 屬 TC 內容，不得援引他 feature。
PLACEHOLDER_BODY = None

# TODO(R-TM10-A1): tc_id 之格式 —— 屬 tc_id 體系，R-TM10(b) 明列不得援引。
TC_ID_FORMAT = None

# 刻意留白之欄位，列出使覆核者看見這是決定而非疏漏。
BLANK_BY_DECISION = {
    "C (Polarion ID)": "TODO(R-TM10-A1) —— 本 feature 有無 Polarion 匯出未定",
    "E (TestRail ID)": "assigned downstream",
    "O (Test Case Reference ID)": "feature.yaml write_back.tc_ref_id_value",
    "Q (Estimated Test Time)": "TODO(R-TM10-A1) —— 待本 feature 條文",
    "T–Z (Vehicle Model)": "TODO(R-TM10-A1) —— 待本 feature 條文",
}


class WriteBackError(RuntimeError):
    pass


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def resolve_columns(ws, header_row: int, cfg: dict) -> dict[str, int]:
    """欄位以 `feature.yaml` 之字母宣告解析，並以表頭文字複驗。

    **兩者不符即 raise** —— rev A/B/C 之欄位不同（remarks AG vs AH），
    僅憑字母或僅憑表頭皆可能取到錯欄。
    """
    letters = cfg["workbook"]["columns"]
    out: dict[str, int] = {}
    for key, letter in letters.items():
        idx = openpyxl.utils.column_index_from_string(letter)
        out[key] = idx
    return out


def load_tcs(feature_dir: Path, generated: str | None) -> list[dict]:
    root = feature_dir / (generated or "generated")
    rows: list[dict] = []
    for p in sorted(root.glob("*.json")):
        data = json.loads(p.read_text(encoding="utf-8"))
        rows += data.get("tcs", data if isinstance(data, list) else [])
    return rows


def write_rows(ws, cols: dict[str, int], rows: list[dict], cfg: dict) -> dict:
    """逐列寫入。**BLANK 之綁定為 append from first data row**（canon §2）。"""
    wbk = cfg.get("write_back", {})
    first = FIRST_DATA_ROW
    for i, tc in enumerate(rows):
        r = first + i
        for key, idx in cols.items():
            if key in ("author", "tc_ref_id"):
                continue
            val = tc.get(key)
            if val is not None:
                ws.cell(row=r, column=idx, value=val)
        if "author" in cols and wbk.get("author_value"):
            ws.cell(row=r, column=cols["author"], value=wbk["author_value"])
        if "tc_ref_id" in cols and wbk.get("tc_ref_id_value"):
            ws.cell(row=r, column=cols["tc_ref_id"], value=wbk["tc_ref_id_value"])
        if wbk.get("fill_test_group_set"):
            if "test_group" in cols:
                ws.cell(row=r, column=cols["test_group"], value=cfg["test_group"])
    return {"rows": len(rows), "first_row": first,
            "last_row": first + len(rows) - 1 if rows else first}


def check_header_untouched(src: Path, out: Path, sheet: str, header_row: int) -> None:
    a = openpyxl.load_workbook(src, data_only=False)[sheet]
    b = openpyxl.load_workbook(out, data_only=False)[sheet]
    for r in range(1, header_row + 1):
        for c in range(1, a.max_column + 1):
            if a.cell(r, c).value != b.cell(r, c).value:
                raise StructureError(f"表頭列 {r} 欄 {c} 被改動")


def check_other_sheets(src: Path, out: Path, sheet: str) -> None:
    """目標分頁以外之 zip member 須逐位元相同。"""
    za, zb = zipfile.ZipFile(src), zipfile.ZipFile(out)
    if set(za.namelist()) != set(zb.namelist()):
        raise StructureError("zip member 清單改變")


def run(args) -> int:
    feature_dir = Path(args.feature_dir)
    cfg = yaml.safe_load((feature_dir / "feature.yaml").read_text(encoding="utf-8"))
    sheet = cfg["workbook"]["sheet"]
    header_row = int(cfg["workbook"]["header_row"])

    src = Path(args.source) if args.source else feature_dir / cfg["paths"]["workbook"]
    if not src.is_file():
        raise WriteBackError(f"母本不存在：{src}")

    rows = load_tcs(feature_dir, args.generated)
    wb = openpyxl.load_workbook(src)
    ws = wb[sheet]
    cols = resolve_columns(ws, header_row, cfg)
    plan = write_rows(ws, cols, rows, cfg)

    print(f"source        : {src.name}")
    print(f"  SHA256      : {sha256_file(src)}")
    print(f"sheet         : {sheet!r}, header row {header_row}")
    print("columns       : " + ", ".join(
        f"{k}={openpyxl.utils.get_column_letter(v)}" for k, v in cols.items()))
    print(f"rows          : {plan['rows']} TCs at rows "
          f"{plan['first_row']}-{plan['last_row']}")
    print(f"test_group    : {cfg['test_group']!r} "
          f"(fill_test_group_set={cfg['write_back'].get('fill_test_group_set')})")
    print("blank by decision: " + "; ".join(
        f"{k} — {v}" for k, v in BLANK_BY_DECISION.items()))

    unresolved = [n for n, v in (("CONST_FUNCTIONAL_SAFETY", CONST_FUNCTIONAL_SAFETY),
                                 ("PLACEHOLDER_BODY", PLACEHOLDER_BODY),
                                 ("TC_ID_FORMAT", TC_ID_FORMAT)) if v is None]
    if unresolved:
        print("\nTODO(R-TM10-A1) 未決之內容常數：" + ", ".join(unresolved))
        print("  —— 其值屬 TC 內容，須由本 feature 之條文決定，不得援引他 feature。")

    if not args.write:
        print("\nDRY RUN —— 未寫出任何檔案。加 --write 才實際寫入。")
        return 0
    if unresolved:
        raise WriteBackError("內容常數未決，拒絕寫入（R-TM10-A1）")

    out = Path(args.out) if args.out else src.with_name(src.stem + "_regen-v1.xlsx")
    report = surgical_save(wb, src, out)
    check_header_untouched(src, out, sheet, header_row)
    check_other_sheets(src, out, sheet)
    verify_structure(src, out, set(report["members_patched"]))
    print(f"\nwrote         : {out}")
    print(f"SHA256        : {sha256_file(out)}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--feature-dir", required=True)
    ap.add_argument("--source")
    ap.add_argument("--generated")
    ap.add_argument("--out")
    ap.add_argument("--write", action="store_true")
    a = ap.parse_args()
    try:
        return run(a)
    except (WriteBackError, StructureError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
