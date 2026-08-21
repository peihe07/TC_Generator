#!/usr/bin/env python3
"""Step 3 (Time Management) — 機械漂移檢查，掃 `generated/*.json`。

## 本檔之來源與界線（R-TM29）

**結構參照** `features/privacy/scripts/lint_tcs.py`（同為 BLANK workbook、
rev C 母本、spec_mode D，形態最近）：權威讀取而非寫死、閘門逐條回傳
`(gate, message)`、`--self-test` 對每閘造紅綠兩向、exit 0/1/2 之語義。

**不繼承其內容**（R-TM10-A1，射程由 R-TM29 界定）：步驟措辭常數、
ER 樣板字串、Test Set 值、priority 預設一律留 `TODO(R-TM10-A1)`，
待本 feature 依條文決定。Privacy 之 `step-actions`／`negative-scope`
兩閘編碼的是 R33-5／R33-1(d)，**那是 Privacy 之裁決，本 feature 不援引**。

## 權威一律讀取，不寫死

  design_method 詞彙  → 母本 `下拉選單` 分頁（`feature.yaml`
                        `lint.design_method_source: dropdown_sheet`）
  欄位對映            → `feature.yaml` `workbook.columns`（rev C：
                        design_method `R`、functional_safety `S`、
                        author `AA`、remarks `AH`）
  test_group          → `feature.yaml` `test_group`（`Time and Date`，R-TM8）
  spec_reference 形態 → `feature.yaml` `spec_reference_template`
                        ＋ CFTS015 docx 內實際存在之物件 id（**查得，不推算**）

exit 0 = 無發現，1 = 至少一項發現，2 = 呼叫錯誤。

用法：
    python3 features/time_management/scripts/lint_tcs.py --feature-dir features/time_management
    python3 features/time_management/scripts/lint_tcs.py --feature-dir features/time_management --self-test

**本腳本於 04 包只建立不執行**（下放包 §8「不執行 T5 所建之任何腳本」）。
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
import html
from pathlib import Path

try:
    import yaml
except ImportError:  # pragma: no cover
    yaml = None


class LintError(RuntimeError):
    """呼叫或權威讀取之失敗 —— 與「發現」區分，走 exit 2。"""


# ── 權威讀取 ────────────────────────────────────────────────
def load_authorities(feature_dir: Path) -> dict:
    """自 `feature.yaml`、母本與 spec docx 讀出全部權威值。

    **任何一項讀不到即 raise** —— 不以預設值頂替（R-TM7 同族：
    寫死之預設會使閘門在權威缺席時仍然全綠）。
    """
    if yaml is None:
        raise LintError("PyYAML 不可用")
    cfg_path = feature_dir / "feature.yaml"
    if not cfg_path.exists():
        raise LintError(f"缺 {cfg_path}")
    cfg = yaml.safe_load(cfg_path.read_text(encoding="utf-8"))

    wb_rel = cfg["paths"]["workbook"]
    wb_path = feature_dir / wb_rel
    if not wb_path.exists():
        raise LintError(f"母本不存在：{wb_path}")

    auth = {
        "cfg": cfg,
        "feature_dir": feature_dir,
        "workbook": wb_path,
        "sheet": cfg["workbook"]["sheet"],
        "header_row": int(cfg["workbook"]["header_row"]),
        "columns": dict(cfg["workbook"]["columns"]),
        "test_group": cfg["test_group"],
        "spec_template": cfg.get("spec_reference_template", ""),
        "design_methods": read_design_methods(wb_path, cfg),
        "spec_objects": read_spec_objects(feature_dir, cfg),
    }
    return auth


def read_design_methods(workbook: Path, cfg: dict) -> set[str]:
    """自母本之 `下拉選單` 分頁讀 design_method 詞彙（逐字，不正規化）。"""
    if cfg.get("lint", {}).get("design_method_source") != "dropdown_sheet":
        raise LintError("feature.yaml 未宣告 design_method_source=dropdown_sheet")
    import openpyxl
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    if "下拉選單" not in wb.sheetnames:
        wb.close()
        raise LintError("母本無 `下拉選單` 分頁")
    ws = wb["下拉選單"]
    vals: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        for c in row:
            if c is None:
                continue
            s = str(c).strip()
            # 設計方法詞條之形態為「中文 (English)」；其餘欄位不收
            if s and "(" in s and ")" in s:
                vals.add(s)
    wb.close()
    if not vals:
        raise LintError("`下拉選單` 分頁未讀到任何設計方法詞條")
    return vals


def read_spec_objects(feature_dir: Path, cfg: dict) -> set[str]:
    """自 CFTS015 docx 讀出實際存在之物件 id。

    **查得，不推算** —— Privacy 之 R30-1 記載了以偏移量推算 id 產生兩個
    錯誤 id 之實例；本 feature 不重蹈，每次執行都重讀 docx。
    """
    rel = cfg["paths"].get("spec_pdf")
    if not rel:
        raise LintError("feature.yaml 未宣告 spec_pdf（CFTS015 docx）")
    path = feature_dir / rel
    if not path.exists():
        raise LintError(f"spec docx 不存在：{path}")
    xml = zipfile.ZipFile(path).read("word/document.xml").decode("utf-8")
    ids: set[str] = set()
    for p in re.findall(r"<w:p[ >].*?</w:p>", xml, flags=re.S):
        t = "".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", p, flags=re.S))
        t = html.unescape(re.sub(r"<[^>]+>", "", t)).strip()
        m = re.match(r"^(\d{6,8})\s*:", t)
        if m:
            ids.add(m.group(1))
    if not ids:
        raise LintError("spec docx 未讀到任何物件 id")
    return ids


# ── 逐條閘門 ────────────────────────────────────────────────
def _steps(text: str) -> list[str]:
    return [s for s in str(text or "").split("\n") if s.strip()]


def lint_required_fields(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """必填欄位齊全 —— 欄名取自 `feature.yaml`，不寫死。"""
    out = []
    for key in auth["columns"]:
        if key in ("tc_ref_id", "author"):      # 寫回時填，生成時不要求
            continue
        if key not in tc:
            out.append(("required-fields", f"{where}: 缺欄位 `{key}`"))
    return out


def lint_test_group(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """`test_group` 須逐字等於 `feature.yaml` 之值（R-TM8）。"""
    want = auth["test_group"]
    got = str(tc.get("test_group", ""))
    if got != want:
        return [("test-group", f"{where}: test_group `{got}` ≠ `{want}`")]
    return []


def lint_design_method(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """`design_method` 須為母本 `下拉選單` 之詞條之一（逐字）。"""
    got = str(tc.get("design_method", "")).strip()
    if got and got not in auth["design_methods"]:
        return [("design-method",
                 f"{where}: design_method `{got[:40]}` 不在母本下拉選單詞彙內")]
    return []


def lint_spec_reference(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """`specification_reference` 之物件 id 須實際存在於 CFTS015 docx。"""
    out = []
    raw = str(tc.get("specification_reference", ""))
    for oid in re.findall(r"\b\d{6,8}\b", raw):
        if oid not in auth["spec_objects"]:
            out.append(("spec-reference",
                        f"{where}: 物件 id `{oid}` 不存在於 CFTS015 docx"))
    return out


def lint_step_er_count(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """步驟數與 ER 行數相等。

    **此為結構性不變量，非措辭判準** —— 故不受 R-TM10-A1 拘束。
    """
    proc = _steps(tc.get("test_procedure"))
    er = _steps(tc.get("expected_result"))
    if proc and er and len(proc) != len(er):
        return [("step-er-count",
                 f"{where}: 步驟 {len(proc)} 條 vs ER {len(er)} 條")]
    return []


# TODO(R-TM10-A1): 步驟措辭之閘門（禁用動詞、步長上限、最終步須帶查核目標）
#   —— 其詞彙與門檻屬 TC 內容，須由本 feature 之條文決定，不得援引他 feature。
#   Privacy 之 `step-actions`（R33-5：一步一動作，判動詞數不判連接詞）
#   為其自身裁決之編碼，本 feature 未有對應條文前不實作。

# TODO(R-TM10-A1): Test Set 值域之閘門 —— 待本 feature 之 framework 定案。

# TODO(R-TM10-A1): priority 分佈之閘門 —— 待本 feature 之條文決定。

# TODO(R-TM10-A1): Input Test Data 填法之閘門 —— 待本 feature 之條文決定。


GATES = (
    lint_required_fields,
    lint_test_group,
    lint_design_method,
    lint_spec_reference,
    lint_step_er_count,
)


def lint_tc(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    for gate in GATES:
        out += gate(tc, auth, where)
    return out


def lint_file(path: Path, auth: dict) -> list[tuple[str, str]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    tcs = data.get("tcs", data if isinstance(data, list) else [])
    out: list[tuple[str, str]] = []
    for i, tc in enumerate(tcs, 1):
        out += lint_tc(tc, auth, f"{path.name}#{i}")
    return out


# ── 範圍向自驗（R-G9）───────────────────────────────────────
def base_tc(auth: dict) -> dict:
    """一條應全綠之最小 TC —— 其值全部取自權威，不寫死措辭。"""
    tc = {k: "" for k in auth["columns"]}
    tc["test_group"] = auth["test_group"]
    tc["design_method"] = sorted(auth["design_methods"])[0]
    tc["specification_reference"] = sorted(auth["spec_objects"])[0]
    tc["test_procedure"] = "1. a\n2. b"
    tc["expected_result"] = "1. a\n2. b"
    return tc


def self_test(auth: dict) -> int:
    """每閘一紅一綠。**綠向證明其不對合規者誤報，紅向證明其抓得到。**"""
    bad = 0
    green = base_tc(auth)
    v = lint_tc(green, auth, "GREEN")
    if v:
        bad += 1
        print(f"**FAIL** 綠向：合規之 TC 轉紅 → {v}")
    else:
        print("PASS 綠向：合規之 TC 未轉紅")

    reds = [
        ("test-group", {**green, "test_group": "Wrong Group"}),
        ("design-method", {**green, "design_method": "不存在之方法 (Nope)"}),
        ("spec-reference", {**green, "specification_reference": "9999999"}),
        ("step-er-count", {**green, "expected_result": "1. a"}),
        ("required-fields", {k: v for k, v in green.items() if k != "test_item"}),
    ]
    for gate, tc in reds:
        v = lint_tc(tc, auth, f"RED[{gate}]")
        hit = any(g == gate for g, _ in v)
        bad += not hit
        print(f"{'PASS' if hit else '**FAIL**'} 紅向 {gate}: "
              f"{[m for g, m in v if g == gate] or '未叫 —— 閘失效'}")
    total = len(reds) + 1
    print(f"\n自驗：{total - bad} / {total}")
    return 1 if bad else 0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--feature-dir", required=True)
    ap.add_argument("--self-test", action="store_true")
    a = ap.parse_args()
    fd = Path(a.feature_dir)
    try:
        auth = load_authorities(fd)
    except LintError as e:
        print(f"LintError: {e}", file=sys.stderr)
        return 2
    if a.self_test:
        return self_test(auth)
    findings: list[tuple[str, str]] = []
    gen = sorted((fd / "generated").glob("*.json"))
    if not gen:
        print("generated/ 無 json —— 尚未生成 TC")
        return 0
    for p in gen:
        findings += lint_file(p, auth)
    for g, m in findings:
        print(f"[{g}] {m}")
    print(f"\n檔 {len(gen)}；發現 {len(findings)} 項")
    return 1 if findings else 0


if __name__ == "__main__":
    raise SystemExit(main())
