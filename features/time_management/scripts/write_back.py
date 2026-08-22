#!/usr/bin/env python3
"""Step 6 (Time Management) — 寫回，**必經 `surgical_save`**。

modified by TC_Generator analysis round 05 under G-TM1/G-TM2/G-TM3

## 本檔之來源與界線（R-TM29）

**結構參照** `features/privacy/scripts/write_back.py`：
`surgical_save` 之呼叫慣例、欄位以表頭文字解析而非寫死字母、
dry-run 為預設、寫入後之三重驗證（表頭未動／他分頁逐位元相同／
`verify_structure`）、`BLANK_BY_DECISION` 之刻意留白清單。

**此處參照是必要的而非便利的**（R-TM29 逐字）：自零寫會升高母本 R 欄
x14 下拉被摧毀之風險（R-G3），該風險不可逆且發生在交付件上。

**不繼承其內容**：Privacy 之 `CONST_FUNCTIONAL_SAFETY = "NA"`（R30-3）、
`PLACEHOLDER_BODY`、`tc_id_format` 皆為其自身裁決。本 feature 之對應值
現況（2026-08-22）：

  CONST_FUNCTIONAL_SAFETY  已決 `"NA"`（**R-TM57**）—— 依據為本 feature
                           自身之交付件實測與 037 分類實測，與 Privacy
                           之值巧合相同但**依據不同**，非援引
  tc_id_format             已決（**R-TM32**），單一來源為 `feature.yaml`，
                           模組層不另存值（**R-TM59**）
  PLACEHOLDER_BODY         仍為 `TODO(R-TM10-A1)`，但**無使用點**，
                           故不列入 unresolved（R-TM59）

## 本 feature 之調整點

  workbook           `feature.yaml paths.workbook`（`inputs/` 之母本複本）
  fill_test_group_set  true（BLANK per R-TM5，依 canon §2）
  test_group         `Time and Date`（R-TM8）
  欄位對映           rev C —— design_method `R`、functional_safety `S`、
                     author `AA`、remarks `AH`
  done_region        none（BLANK per R-TM5）→ append from first data row

用法：
    python3 features/time_management/scripts/write_back.py --feature-dir features/time_management
    （加 --write 才實際寫出；預設 dry-run）

**本腳本於 04 包只建立不執行。**
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

import openpyxl
import yaml

REPO_ROOT = next(p for p in Path(__file__).resolve().parents
                 if (p / "pyproject.toml").is_file())
sys.path.insert(0, str(REPO_ROOT))
from backend.xlsx_surgical import (  # noqa: E402
    StructureError, surgical_save, verify_structure)

FIRST_DATA_ROW = 10          # rev C 版面；表頭列 9

# R-TM57（08 §1）—— 值已裁定，TODO 撤除。
#   "NA" 為 canon §8.4.3 之「確認不適用」，非缺件佔位：Time and Date 與
#   User Profiles 同屬 HMI 功能，無功能安全需求分派。兩條獨立證據：
#   (a) 交付件 UserProfiles_20260820 之 S 欄 189/189 皆 "NA"（08 §1，
#       執行層已獨立複驗，見 upstream/08_style.md §1.1）
#   (b) 037 之 Categorization 22/22 皆 Functional，SYS2 與 037 皆無
#       ASIL / FTTI 欄（04Z-A2 §2）
#   值與 Privacy 之 R30-3 巧合相同，但**依據不同**：此處為本 feature
#   自身之實測與條文，非援引他 feature（R-TM10(b)）。
CONST_FUNCTIONAL_SAFETY = "NA"

# TODO(R-TM10-A1): BLOCKED 佔位之措辭 —— 屬 TC 內容，不得援引他 feature。
#   **不列入 unresolved**（R-TM59）：本常數無任何使用點，其未決不影響任何
#   寫入；留在 unresolved 內等於以一個不生效之未決項阻擋整條寫回路徑。
#   BLOCKED 佔位之寫入路徑實作時須移回 —— 屆時它才真的會影響輸出。
PLACEHOLDER_BODY = None

# R-TM59 —— tc_id 格式之**單一來源為 feature.yaml**，模組層不另存一份值。
# 本識別字保留為**來源指標**（R-TM13 之精神：留下「此處曾有一個值」之痕跡），
# 其內容刻意不是格式字串本身，使誤用者立即失敗而非靜默產出壞 tc_id。
# 原 `TC_ID_FORMAT = None` 為死常數：從不被讀用（真正生效者在 yaml，
# 值早由 R-TM32 裁定），卻列入 unresolved 而擋住 --write（08 上繳 §5）。
TC_ID_FORMAT = "<see feature.yaml: write_back.tc_id_format>"


def resolve_tc_id_format(cfg: dict) -> str:
    """tc_id 格式之**唯一**取值入口（R-TM59）。

    凡需要該格式者一律呼叫本函式，不得自行 `cfg["write_back"][...]` ——
    自行讀取即製造第二個來源，而兩個來源不會互相比對，漂移時無人發現。
    """
    fmt = (cfg.get("write_back") or {}).get("tc_id_format")
    if not fmt:
        raise WriteBackError(
            "feature.yaml 缺 `write_back.tc_id_format` —— tc_id 格式為 "
            "R-TM32 之裁決值，其單一來源為 feature.yaml（R-TM59）。"
            "不得於程式碼內補寫預設值。")
    if "{n" not in fmt:
        raise WriteBackError(
            f"`write_back.tc_id_format` = {fmt!r} 不含序號欄位 `{{n...}}`，"
            "無法賦號（R-TM32：`NR1L-TimeAndDate-{n:03d}`）")
    return fmt


def assert_tc_id_single_source(cfg: dict, used: str) -> None:
    """R-TM56 之可獨立呼叫守衛 —— 實際用於寫入者須與唯一入口同值。

    `used` 為 write_rows 實際賦號所用之格式字串。此檢查之作用不是驗
    yaml 內容（那是 resolve 的事），而是驗**沒有第二條取值路徑** ——
    若日後有人在 write_rows 內改回自行讀 cfg，本檢查即在兩者分岔時失敗。
    """
    want = resolve_tc_id_format(cfg)
    if used != want:
        raise WriteBackError(
            f"tc_id 格式雙來源：write_rows 用 {used!r}，"
            f"resolve_tc_id_format 得 {want!r}（R-TM59）")

# 刻意留白之欄位，列出使覆核者看見這是決定而非疏漏。
BLANK_BY_DECISION = {
    "C (Polarion ID)": "TODO(R-TM10-A1) —— 本 feature 有無 Polarion 匯出未定",
    "E (TestRail ID)": "assigned downstream",
    "O (Test Case Reference ID)": "feature.yaml write_back.tc_ref_id_value",
    "Q (Estimated Test Time)": "TODO(R-TM10-A1) —— 待本 feature 條文",
    "T–Z (Vehicle Model)": "R-TM77 —— 交付件 UserProfiles_20260820 之該七欄 "
                           "189/189 全空；車型欄在既有交付實務中不作為範圍"
                           "標示之用，範圍改由 Pre-Condition 承載（R-TM76）",
}


class WriteBackError(RuntimeError):
    pass


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# 表頭文字之判準：欄位鍵 -> (須含之字串, 不得含之字串或 None)。
# 取自母本 rev C 之實測表頭（FORMS.md），比對前先 norm（小寫、空白折疊）。
HEADER_NEEDLE = {
    "req_id": ("requirement or design id", "polarion"),
    "tc_id": ("test case id", "testrail"),
    "test_group": ("test group", None),
    "test_set": ("test set", None),
    "test_item": ("test item", None),
    "pre_conditions": ("pre-condition", None),
    "input_test_data": ("input test data", None),
    "test_procedure": ("test procedure", None),
    "expected_result": ("expected result", None),
    "spec_reference": ("specification reference", None),
    "tc_ref_id": ("test case reference id", None),
    "priority": ("test case priority", None),
    "design_method": ("test case design", None),
    "functional_safety": ("functional safety", None),
    "author": ("test case author", None),
    "remarks": ("remarks", None),
}


def _norm(v) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip().lower()


def resolve_columns(ws, header_row: int, cfg: dict) -> dict[str, int]:
    """欄位以 `feature.yaml` 之字母宣告解析，並以表頭文字複驗。

    **兩者不符即 raise**（A-TM21(a) / G-TM2 項 1）—— rev A/B/C 之欄位不同
    （design_method Q→R、author Z→AA、remarks AG→AH），僅憑字母或僅憑
    表頭皆可能取到錯欄。

    此為本 feature 現存唯一「錯了會被執行而非被攔」之盲區之防線：
    `verify_structure` 保護檔案結構，不驗欄位對映是否取對 —— 寫進錯欄時
    錯欄仍在目標分頁內、屬 patched，三層結構檢查全綠。
    """
    letters = cfg["workbook"]["columns"]
    header = [ws.cell(header_row, c).value
              for c in range(1, ws.max_column + 1)]
    norm = [_norm(v) for v in header]
    out: dict[str, int] = {}
    drift: list[str] = []
    for key, letter in letters.items():
        idx = openpyxl.utils.column_index_from_string(letter)
        out[key] = idx
        needle = HEADER_NEEDLE.get(key)
        if needle is None:
            continue                      # 未列於判準表者不複驗，但不靜默取代
        need, forbid = needle
        actual = norm[idx - 1] if idx <= len(norm) else ""
        if need not in actual or (forbid and forbid in actual):
            hits = [i + 1 for i, h in enumerate(norm)
                    if need in h and not (forbid and forbid in h)]
            where = (openpyxl.utils.get_column_letter(hits[0])
                     if len(hits) == 1 else f"{len(hits)} 個候選")
            drift.append(
                f"{key}: feature.yaml 宣告 {letter}，該欄表頭為 "
                f"{header[idx - 1]!r}；表頭文字指向 {where}")
    if drift:
        raise WriteBackError(
            "欄位字母與表頭文字不符（A-TM21(a)）：\n  " + "\n  ".join(drift)
            + "\n以表頭文字為準修正 feature.yaml，勿改本檔。")
    return out


def load_tcs(feature_dir: Path, generated: str | None) -> list[dict]:
    """讀取待寫回之 TC。

    **`.pre-arch.json` 排除**（`20` T3 之 R-TM13 軌跡備份）—— 其為改動前
    之完整副本，讀入會使每條 TC 寫入兩次。本項由 dry-run 抓到：
    57 條顯示為 114 條。

    **排除清單以副檔名樣式為準而非白名單**：白名單須隨批次增減維護，
    而漏維護之後果是「少寫一批」——那比多寫一批更難察覺。
    """
    root = feature_dir / (generated or "generated")
    rows: list[dict] = []
    skipped: list[str] = []
    for p in sorted(root.glob("*.json")):
        if p.name.endswith(".pre-arch.json"):
            skipped.append(p.name)
            continue
        data = json.loads(p.read_text(encoding="utf-8"))
        rows += data.get("tcs", data if isinstance(data, list) else [])
    if skipped:
        print(f"skipped       : {len(skipped)} 個軌跡備份 —— {', '.join(skipped)}")
    return rows


def existing_data_rows(ws, cols: dict[str, int], header_row: int) -> int:
    """既有資料列數 —— tc_id 跨批連續之**唯一起點來源**（A3）。

    R-TM32 要求序號跨批連續不重設，故需一個起點。候選有二：
    讀工作簿之既有列數，或讀 generated/ 之累計 TC 數。**本檔取前者，
    且只取前者** —— 兩者在正常情形一致，但某批若曾撤回或重生成即分歧，
    而工作簿是交付件、是唯一真相。G-TM2 項 3 訂正要求「單一且可查」。

    判準取 req_id（D 欄）非空之列數：B 欄為公式（恆存在），
    其餘欄可能因 BLOCKED 列而留空。
    """
    col = cols["req_id"]
    n = 0
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(r, col).value not in (None, ""):
            n += 1
    return n


def write_rows(ws, cols: dict[str, int], rows: list[dict], cfg: dict,
               start_seq: int) -> dict:
    """逐列寫入。**BLANK 之綁定為 append from first data row**（canon §2）。

    tc_id 由本函式**依列位置賦號**，不自 `tc` 讀取 —— canon §10.3 末句：
    `the generator handles assignment, the LLM does not emit tc_id`
    （ASPICE_SWE6_AI_Instruction.md:521-525）。故 tc.get(key) 迴圈排除
    tc_id，序號在迴圈外由 start_seq 遞增。
    """
    wbk = cfg.get("write_back", {})
    fmt = resolve_tc_id_format(cfg)       # R-TM59：唯一取值入口
    first = FIRST_DATA_ROW + start_seq
    expected: list[dict] = []             # A5 正向驗證之預期值
    for i, tc in enumerate(rows):
        r = first + i
        seq = start_seq + i + 1
        for key, idx in cols.items():
            # tc_id 由賦號決定、author/tc_ref_id/functional_safety 由條文
            # 決定，四者皆不從 tc 讀 —— 欄位值之決定權不下放給生成端。
            if key in ("author", "tc_ref_id", "tc_id", "functional_safety"):
                continue
            val = tc.get(key)
            if val is not None:
                ws.cell(row=r, column=idx, value=val)
        tc_id = fmt.format(n=seq)
        ws.cell(row=r, column=cols["tc_id"], value=tc_id)
        if "author" in cols and wbk.get("author_value"):
            ws.cell(row=r, column=cols["author"], value=wbk["author_value"])
        if "tc_ref_id" in cols and wbk.get("tc_ref_id_value"):
            ws.cell(row=r, column=cols["tc_ref_id"], value=wbk["tc_ref_id_value"])
        # A4 —— 原為死碼（宣告了卻不寫入）。接上使其實際寫入 S 欄。
        # 值已由 R-TM57 裁定為 "NA"；unresolved 檢查仍在，用以攔下
        # 日後任一常數被改回 None 之情形（08 T3 紅向已證其未被削弱）。
        if "functional_safety" in cols:
            ws.cell(row=r, column=cols["functional_safety"],
                    value=CONST_FUNCTIONAL_SAFETY)
        if wbk.get("fill_test_group_set"):
            if "test_group" in cols:
                ws.cell(row=r, column=cols["test_group"], value=cfg["test_group"])
        expected.append({"row": r, "tc_id": tc_id,
                         "test_item": tc.get("test_item")})
    return {"rows": len(rows), "first_row": first,
            "last_row": first + len(rows) - 1 if rows else first,
            "start_seq": start_seq, "expected": expected,
            "tc_id_format_used": fmt}      # R-TM59 一致性檢查之輸入


# R-TM65 —— 啟動檢查之豁免欄。**列出並附理由，不以「大概不用驗」帶過。**
#   四者由條文決定而非自 TC 讀（write_rows 之迴圈明文排除），
#   remarks 為條件性欄位（僅缺口 TC 需要），故未出現不代表失效。
KEY_CHECK_EXEMPT = {
    "tc_id": "由 write_rows 依列位置賦號（canon §10.3）",
    "author": "feature.yaml write_back.author_value",
    "tc_ref_id": "feature.yaml write_back.tc_ref_id_value",
    "functional_safety": "常數 CONST_FUNCTIONAL_SAFETY（R-TM57）",
    "remarks": "條件性 —— 僅有缺口宣告之 TC 需要（R-TM64）",
    "test_group": "fill_test_group_set 時由 cfg 填（R-TM5 / R-TM8）",
}


def check_keys_present(cols: dict, rows: list[dict]) -> None:
    """R-TM65 —— `cols` 之每個 key 至少在一條 TC 出現，否則 raise。

    緣起：lint 讀 `specification_reference` 而 feature.yaml 宣告
    `spec_reference`，兩者不一致時 **write_rows 之 `tc.get(key)` 取不到值，
    該欄靜默空白寫入工作簿**（`13` 上繳 §5.1）。lint 不知道 write_back 用
    什麼鍵，write_back 不跑 lint —— 兩支工具之間無人把關。

    判準刻意是「**至少一條**」而非「每條都有」：後者會誤攔條件性欄位。
    豁免清單見 `KEY_CHECK_EXEMPT`。
    """
    if not rows:
        return
    seen = {k for r in rows for k in r}
    missing = [k for k in cols if k not in seen and k not in KEY_CHECK_EXEMPT]
    if missing:
        raise WriteBackError(
            "下列欄位在 feature.yaml 之 columns 有宣告，但**沒有任何一條 TC "
            f"帶此鍵**，寫入後將全欄空白：{missing}\n"
            "最可能之成因為 TC 產生端與 feature.yaml 之鍵名不一致"
            "（R-TM65）。豁免欄見 KEY_CHECK_EXEMPT。")


def check_header_untouched(src: Path, out: Path, sheet: str, header_row: int) -> None:
    a = openpyxl.load_workbook(src, data_only=False)[sheet]
    b = openpyxl.load_workbook(out, data_only=False)[sheet]
    for r in range(1, header_row + 1):
        for c in range(1, a.max_column + 1):
            if a.cell(r, c).value != b.cell(r, c).value:
                raise StructureError(f"表頭列 {r} 欄 {c} 被改動")


# check_other_sheets() 已移除（G-TM2 項 2 訂正 / A-TM21(b)）。
# 其功能被 backend/xlsx_surgical.py:268-275 之 verify_structure 第三層
# 完全涵蓋且後者更嚴格：逐 member 位元組比對 `a.read(m) != b.read(m)`，
# 且限定僅 patched 之 member 得有差異。原函式只比對 member 名稱集合，
# 內容被改寫而名稱不變則全綠 —— 保留一個更弱且被完全涵蓋之檢查，
# 只會製造「有兩道獨立防護」之假象。


def check_written_back(out: Path, sheet: str, cols: dict[str, int],
                       expected: list[dict]) -> None:
    """G-TM3 —— 寫回後之**正向驗證**：重開輸出檔，讀指定 cell 比對預期。

    現有三層（verify_structure）全為反向驗證「不該變的沒變」，
    **無一層驗證「該變的變對了地方」**（A-TM22）。反向驗證再嚴格也無法
    發現「寫對了內容但寫錯了地方」。

    **主要防護對象為 column 層（A-TM21(a)），非 member 層** —— 後者之
    sheet 名→member 對映走 workbook rels 且失敗會 raise，有自身保證；
    前者之字母宣告若整體位移一格，寫入落在鄰欄而結構檢查全綠。

    取樣以 `tc_id` 為首選（依序號賦值必逐列互異，可排除「兩欄值恰同」
    之偽陰性），`test_item` 次之。不取 design_method / priority 一類
    值域小之欄 —— 位移一格仍可能取到合法值（R-TM42）。
    """
    if not expected:
        return
    picks = [expected[0], expected[len(expected) // 2], expected[-1]]
    ws = openpyxl.load_workbook(out)[sheet]
    bad: list[str] = []
    for e in picks:
        for field in ("tc_id", "test_item"):
            want = e[field]
            if want is None:
                continue
            got = ws.cell(e["row"], cols[field]).value
            if got != want:
                bad.append(
                    f"列 {e['row']} 欄 "
                    f"{openpyxl.utils.get_column_letter(cols[field])}"
                    f"（{field}）：預期 {want!r}，實得 {got!r}")
    if bad:
        raise WriteBackError(
            "正向驗證失敗（G-TM3）—— 寫入之值不在預期位置：\n  "
            + "\n  ".join(bad)
            + "\n最可能之成因為 feature.yaml 之欄位字母與實際欄位不符"
              "（A-TM21(a) / column 層位移）。")


def run(args) -> int:
    feature_dir = Path(args.feature_dir)
    cfg = yaml.safe_load((feature_dir / "feature.yaml").read_text(encoding="utf-8"))
    sheet = cfg["workbook"]["sheet"]
    header_row = int(cfg["workbook"]["header_row"])

    src = Path(args.source) if args.source else feature_dir / cfg["paths"]["workbook"]
    if not src.is_file():
        raise WriteBackError(f"母本不存在：{src}")

    rows = load_tcs(feature_dir, args.generated)
    wb = openpyxl.load_workbook(src)
    ws = wb[sheet]
    cols = resolve_columns(ws, header_row, cfg)
    start_seq = existing_data_rows(ws, cols, header_row)
    check_keys_present(cols, rows)          # R-TM65 啟動檢查
    plan = write_rows(ws, cols, rows, cfg, start_seq)

    print(f"source        : {src.name}")
    print(f"  SHA256      : {sha256_file(src)}")
    print(f"sheet         : {sheet!r}, header row {header_row}")
    print("columns       : " + ", ".join(
        f"{k}={openpyxl.utils.get_column_letter(v)}" for k, v in cols.items()))
    print(f"rows          : {plan['rows']} TCs at rows "
          f"{plan['first_row']}-{plan['last_row']}")
    # R-TM59：本處原為 tc_id_format 之**第三個**自讀點（模組常數、
    # write_rows、此處各一）。全部改走 resolve_tc_id_format。
    assert_tc_id_single_source(cfg, plan["tc_id_format_used"])
    _fmt = resolve_tc_id_format(cfg)
    print(f"tc_id         : 起點序號 {plan['start_seq']}（既有資料列數，"
          f"A3 之唯一起點來源）；本批 "
          f"{_fmt.format(n=plan['start_seq'] + 1)}"
          f" … {_fmt.format(n=plan['start_seq'] + plan['rows'])}")
    print(f"test_group    : {cfg['test_group']!r} "
          f"(fill_test_group_set={cfg['write_back'].get('fill_test_group_set')})")
    print("blank by decision: " + "; ".join(
        f"{k} — {v}" for k, v in BLANK_BY_DECISION.items()))

    # R-TM59 —— 清單只列**其未決會實際影響寫入**者。
    # PLACEHOLDER_BODY 無使用點故移出；TC_ID_FORMAT 之值改由 yaml 供給，
    # 其缺失由 resolve_tc_id_format 直接 raise，不走本清單。
    # 判準由 `v is None` 改為 `v is None or v == ""`：空字串同樣是未決
    # （08 上繳 §4.1 紅向 2 之已知射程缺口）。
    unresolved = [n for n, v in (("CONST_FUNCTIONAL_SAFETY",
                                  CONST_FUNCTIONAL_SAFETY),)
                  if v is None or v == ""]
    if unresolved:
        print("\n未決之內容常數（其值屬 TC 內容，須由本 feature 之條文決定）："
              + ", ".join(unresolved))
    else:
        print("\n內容常數        : 全部已決 —— unresolved 為空（R-TM57 / R-TM59）")

    if not args.write:
        print("\nDRY RUN —— 未寫出任何檔案。加 --write 才實際寫入。")
        return 0
    if unresolved:
        raise WriteBackError(
            "內容常數未決，拒絕寫入：" + ", ".join(unresolved))

    out = Path(args.out) if args.out else src.with_name(src.stem + "_regen-v1.xlsx")
    report = surgical_save(wb, src, out)
    check_header_untouched(src, out, sheet, header_row)
    # 他分頁之內容比對由 verify_structure 第三層為之（見上方註解）
    verify_structure(src, out, set(report["members_patched"]))
    check_written_back(out, sheet, cols, plan["expected"])   # G-TM3
    print(f"\nwrote         : {out}")
    print(f"SHA256        : {sha256_file(out)}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--feature-dir", required=True)
    ap.add_argument("--source")
    ap.add_argument("--generated")
    ap.add_argument("--out")
    ap.add_argument("--write", action="store_true")
    a = ap.parse_args()
    try:
        return run(a)
    except (WriteBackError, StructureError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
