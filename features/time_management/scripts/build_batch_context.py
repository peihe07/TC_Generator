#!/usr/bin/env python3
"""Step 2 (Time and Date) — B1 之 context 產生器。

modified by TC_Generator analysis round 06 under G-TM1 更正 / R-TM33

## 本檔之來源與界線

**重新設計，非還原**（`06` §4）。原執行層版於 2026-08-21 09:15 被覆蓋
且已失落（A-TM20 記載更正），現存版無 A-TM13 缺口與界線之編碼 ——
G-TM1 更正即為此。

**裁決值一律取自 `tm_rulings.py`**，與 `lint_tcs.py` **同一來源**
（`06` §4.2）。兩份會漂移，且漂移時 lint 全綠：context 說 A、lint 驗 B，
生成照 A 寫則被 B 攔，呈現為「模型出錯」而非「規則不一致」。

## 六類編碼（`06` §4.1）

  C-1  五條 §8.2.1 界線（owns / not_ours / 物件 id）  R-TM17 + R-TM25
  C-2  A-TM13 兩片之缺口指示（PENDING: DR-5）        R-TM41 訂正 + §8.4.3
  C-3  spec_reference 候選（v2 格式）                R-TM40 + canon §10.7(a)
  C-4  test_item 兩段式（上半 ≤50 token + 下半括號）  canon §4.3.1 + R-TM24
  C-5  訊號三件組（segment 依 R-TM51 判定）          canon §8.7.5 + R-TM49/51
  C-6  Test Set 值                                  R-TM17 / Part VII

**不進入 context**：`tc_id`（canon §10.3 明訂 generator 賦號，LLM 不得
emit）、`functional_safety`（A-TM24 未決且由條文定）、`priority` 分佈。

## C-5 之實作限制

R-TM51 判準 (a) 同物件為純位置判斷，可自動化；判準 (b) 同述語為**句法
判斷**，程式無法可靠自動化。故本檔**列出同物件內之網段候選並附其原句**，
由生成端依 (b) 判定；**不自行斷定 (b) 成立**。同物件無候選者直接標
`PENDING: DR-6`。

用法：
    python3 features/time_management/scripts/build_batch_context.py \
        --feature-dir features/time_management --batch B1
    加 --out data/context_B1.json 寫檔；預設印出

本檔只建 context，不呼叫模型、不生成 TC。
"""
from __future__ import annotations

import argparse
import html
import json
import re
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from tm_rulings import (                                       # noqa: E402
    BATCHES, BOUNDARY_NOTES, BOUNDARY_SIGNALS, SEGMENT_PLACEHOLDER,
    SPEC_GAP, SPEC_REF_PREFIX, TEST_GROUP, TEST_ITEM_TOKEN_MAX,
    TEST_SET_OF, TEST_SETS, spec_gap_placeholder)

REPO_ROOT = next(p for p in Path(__file__).resolve().parents
                 if (p / "pyproject.toml").is_file())

LEAF_RE = re.compile(r"^(SWE-RA-TIME&DATE-(\d{3}))\s*\|\s*(.*)$")
OBJ_RE = re.compile(r"^(\d{6,8})\s*:")
HEAD_RE = re.compile(r"^(\d+(?:\.\d+)*)\s+.*\{(\d+)\}$")
SEGMENT_RE = re.compile(r"\b([A-Z]{1,3}H?-CAN|CAN-[A-Z]+)\b")
SIGNAL_RE = re.compile(r"\$[A-Za-z][A-Za-z0-9_]*\$")


class ContextError(RuntimeError):
    pass


# ── 來源讀取（皆為既有產物，不重算）──────────────────────────

def load_leaf_text(feature_dir: Path) -> dict[str, dict]:
    """C-4 之上半來源 —— R-TM24：037 之直接輸出為唯一許可來源。"""
    path = feature_dir / "data" / "leaf_descriptions.txt"
    if not path.is_file():
        raise ContextError(
            f"{path} 不存在 —— 該檔為 test_item 上半之唯一許可來源"
            "（R-TM24）。其對策為來源隔離，非人工記得不抄下放包之簡寫")
    out: dict[str, dict] = {}
    lines = path.read_text(encoding="utf-8").splitlines()
    for i, line in enumerate(lines):
        m = LEAF_RE.match(line)
        if m:
            out[m.group(2)] = {
                "leaf": m.group(1), "title": m.group(3).strip(),
                "description": lines[i + 1].strip() if i + 1 < len(lines) else "",
            }
    if len(out) != 22:
        raise ContextError(f"{path.name} 讀到 {len(out)} 筆 leaf，期望 22")
    return out


def load_sys2_items(feature_dir: Path, cfg: dict) -> dict[str, list[str]]:
    """SYS-RA 序號 -> 來源物件 id（C-3 之來源）。"""
    import openpyxl
    path = feature_dir / cfg["paths"]["sys1_export"]
    if not path.is_file():
        raise ContextError(f"SYS2 匯出不存在：{path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, list[str]] = {}
    for r in wb["Basic Report"].iter_rows(min_row=2, values_only=True):
        m = re.search(r"-(\d{3})\s*$", str(r[1] or ""))
        if m:
            out[m.group(1)] = [x.strip() for x in
                               re.split(r"[,\n]+", str(r[4] or "")) if x.strip()]
    wb.close()
    if not out:
        raise ContextError("SYS2 第 5 欄未讀到任何來源物件 id")
    return out


def load_leaf_refs(feature_dir: Path, cfg: dict) -> dict[str, list[str]]:
    """leaf -> 其全部引用之 SYS-RA 序號（範圍展開）。"""
    import openpyxl
    path = feature_dir / cfg["paths"]["a03_report"]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, list[str]] = {}
    for r in wb["Analysis Report"].iter_rows(min_row=9, values_only=True):
        if not r[0]:
            continue
        t = str(r[1]).replace("–", "-").replace("—", "-")
        nums: set[int] = set()
        for a, b in re.findall(r"(\d{3})\s*-\s*(\d{3})", t):
            nums |= set(range(int(a), int(b) + 1))
        nums |= {int(n) for n in
                 re.findall(r"\d{3}", re.sub(r"\d{3}\s*-\s*\d{3}", "", t))}
        out[str(r[0]).strip()[-3:]] = [str(n).zfill(3) for n in sorted(nums)]
    wb.close()
    return out


def load_spec_objects(feature_dir: Path, cfg: dict) -> dict[str, dict]:
    """CFTS015 物件 id -> {章節, 內文} —— C-3 之存在性判定與 C-5 之來源。"""
    path = feature_dir / cfg["paths"]["spec_pdf"]
    if not path.is_file():
        raise ContextError(f"CFTS015 docx 不存在：{path}")
    xml = zipfile.ZipFile(path).read("word/document.xml").decode("utf-8")
    paras = []
    for para in re.findall(r"<w:p[ >].*?</w:p>", xml, flags=re.S):
        txt = "".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", para, flags=re.S))
        paras.append(html.unescape(re.sub(r"<[^>]+>", "", txt)).strip())
    out: dict[str, dict] = {}
    cur = None
    open_id = None          # 目前累積中之物件
    for line in paras:
        if "\t" in line:
            continue
        m = HEAD_RE.match(line)
        if m:
            cur, open_id = m.group(1), None
            out.setdefault(m.group(2), {"section": cur, "text": line,
                                        "body": ""})
            continue
        m2 = OBJ_RE.match(line)
        if m2 and cur:
            open_id = m2.group(1)
            # 物件行本身為屬性列（[Artifact Type:…][ECU:…]…），
            # **需求文字在其後續段落** —— 訊號名 `$X$` 出現於 body 而非
            # 屬性列。只存屬性列會使 C-5 抓到零個訊號（06 T4 實測）。
            out.setdefault(open_id, {"section": cur, "text": line, "body": ""})
            continue
        if open_id and line:
            out[open_id]["body"] += (" " if out[open_id]["body"] else "") + line
    if not out:
        raise ContextError("CFTS015 docx 未讀到任何物件 id")
    if not any(v["body"] for v in out.values()):
        raise ContextError(
            "全部物件之 body 皆空 —— 物件行之後續段落未被收入。"
            "訊號名出現於 body 而非屬性列，body 全空則 C-5 必然抓到零個訊號")
    return out


# ── 六類編碼 ─────────────────────────────────────────────────

def assert_rendered(r: str) -> None:
    """C-3 執行期守衛（06Z T3）—— `rendered` 之格式在**產出當下**即驗。

    抽為獨立函式使其可被 self-test 直接以壞值呼叫 —— 守衛若只存在於
    產生函式內部，紅向無從注入壞值（本包首次構造即因此失敗：
    在產生函式回傳**之後**改值，守衛已跑過）。

    canon §10.7 排列段：文件前綴僅敘明一次、禁用 `;`、同文件內 ID 升冪。
    """
    if r.count(SPEC_REF_PREFIX) != 1 or ";" in r:
        raise ContextError(
            f"C-3 守衛：rendered {r!r} 之前綴非恰一次或含 `;` —— "
            "canon §10.7 排列段：文件前綴僅敘明一次、禁用 `;`")
    got = re.findall(r"\d{7}", r)
    if got != sorted(got):
        raise ContextError(
            f"C-3 守衛：rendered {r!r} 之物件 id 未升冪 —— canon §10.7 排列段")


def build_spec_reference(num, refs, sys2, objs) -> dict:
    """C-3 ＋ C-2 —— 候選清單（v2 格式）與缺口佔位。"""
    ids = sorted({i for n in refs.get(num, []) for i in sys2.get(n, [])})
    ok = [i for i in ids if i in objs]
    missing = [i for i in ids if i not in objs]
    entry: dict = {"format": f"{SPEC_REF_PREFIX}-{{7 位物件 id}}",
                   "candidates_upper_bound": ok}
    if ok:
        entry["rendered"] = (f"{SPEC_REF_PREFIX}-{ok[0]}"
                             + "".join(f", {i}" for i in ok[1:]))
    if missing:                                   # C-2
        g = SPEC_GAP.get(num, {})
        entry["gap"] = {
            "objects": missing,
            "sys_ra": g.get("sys_ra"),
            "note": g.get("note"),
            "placeholder": spec_gap_placeholder(missing),
            "instruction": (
                f"物件 {' / '.join(missing)} 於 CFTS015 SR26 全檔零命中"
                "（A-TM13）。該部分**無章節可引**：不得以鄰近物件填充"
                "（§8.4.1），亦**不得留空**（canon §8.4.3）——"
                f"於 Remarks 寫 `{spec_gap_placeholder(missing)}`。"),
        }
    if entry.get("rendered") is not None:
        assert_rendered(entry["rendered"])          # C-3 執行期守衛
    entry["scope_note"] = (
        "本清單為該 leaf 之**聯集上限**，非每條 TC 之預設值 ——"
        " canon §10.7：只列該 TC 直接驗證或作為 setup 依賴之物件，"
        "僅供背景者放 reasoning。排列：前綴僅敘明一次、升冪、禁用 `;`。")
    return entry


def build_signals(num, refs, sys2, objs) -> dict:
    """C-5 —— 訊號三件組。segment 依 R-TM51 (a) 自動篩，(b) 交生成端。"""
    ids = [i for n in refs.get(num, []) for i in sys2.get(n, []) if i in objs]
    signals: dict[str, dict] = {}
    for oid in sorted(set(ids)):
        text = objs[oid]["body"] or objs[oid]["text"]
        segs = sorted(set(SEGMENT_RE.findall(text)))
        for sig in sorted(set(SIGNAL_RE.findall(text))):
            e = signals.setdefault(sig, {"objects": [], "segment_candidates": []})
            e["objects"].append(oid)
            for s in segs:                        # R-TM51 (a) 同物件
                e["segment_candidates"].append(
                    {"segment": s, "object": oid,
                     "sentence": text[:200]})
    for sig, e in signals.items():
        if not e["segment_candidates"]:
            e["segment"] = SEGMENT_PLACEHOLDER    # (a) 不成立即佔位
        else:
            e["segment"] = None                   # 待生成端依 (b) 判定
    return {
        "form": "<Signal> in <MESSAGE> on <segment>",
        "rule": ("canon §8.7.5。Signal 與 MESSAGE 取自 CFTS015 內文；"
                 "**segment 須依 R-TM51 兩項判定**：(a) 網段敘述與訊號名"
                 "出現於同一物件（本檔已篩，見 segment_candidates）；"
                 "(b) 該敘述之句法上網段為該訊號或其 MESSAGE 之修飾語，"
                 "非另一句之主題 —— **此項須由你判定，本檔不代為斷定**。"
                 f"兩項有一不成立即填 `{SEGMENT_PLACEHOLDER}`。"
                 "填寫時於 reasoning 註明來源物件 id，使該判定可被覆核。"),
        "signals": signals,
    }


def build(feature_dir: Path, batch: str) -> dict:
    import yaml
    cfg = yaml.safe_load(
        (feature_dir / "feature.yaml").read_text(encoding="utf-8"))
    if batch not in BATCHES:
        raise ContextError(f"未知批次 {batch!r}；已知：{sorted(BATCHES)}")
    text = load_leaf_text(feature_dir)
    leaf_src = (feature_dir / "data" / "leaf_descriptions.txt").read_text(
        encoding="utf-8")                      # C-4 守衛之比對基準
    sys2 = load_sys2_items(feature_dir, cfg)
    refs = load_leaf_refs(feature_dir, cfg)
    objs = load_spec_objects(feature_dir, cfg)

    items = []
    for num in BATCHES[batch]:
        if num not in text:
            raise ContextError(f"leaf {num} 不在 leaf_descriptions.txt 內")
        leaf = text[num]["leaf"]
        e: dict = {
            "leaf": leaf,
            "title": text[num]["title"],
            "test_group": TEST_GROUP,                       # R-TM8
            "test_set": TEST_SET_OF[num],                   # C-6
            "test_item": {                                  # C-4
                "upper_verbatim": text[num]["description"],
                "upper_token_count": len(text[num]["description"].split()),
                "upper_token_max": TEST_ITEM_TOKEN_MAX,
                "rule": ("canon §4.3.1 兩段式。上半為需求原句 **verbatim**"
                         f"（來源限 data/leaf_descriptions.txt，R-TM24），"
                         f"token 上限 {TEST_ITEM_TOKEN_MAX}，超限須摘句"
                         "（取與括號下半直接相關之句），全文以 "
                         "specification_reference 指回，不得整段傾倒。"
                         "下半為你所擬之測試目的，獨立成行，格式 `(...)`。"
                         "**缺括號下半 = FAIL，不得出貨。** 同一 leaf 衍生"
                         "之多列，其括號內容不得逐字相同。"),
            },
            "specification_reference": build_spec_reference(num, refs, sys2, objs),
            "signals": build_signals(num, refs, sys2, objs),
        }
        if leaf in BOUNDARY_SIGNALS:                        # C-1
            b = BOUNDARY_SIGNALS[leaf]
            e["boundary"] = {
                "owns": b["owns"], "not_ours": b["not_ours"],
                "objects": b["objects"], "why": b["why"],
                "rule": ("Part VII §8.2.1 拘束條款（R-TM23 + R-TM25，"
                         "R-TM17 簽核）。**不得於本片之任何欄位提及 "
                         "not_ours 之訊號** —— 那屬鄰片，提及即重複覆蓋。"),
            }
        if leaf in BOUNDARY_NOTES:                          # C-1（敘述型）
            e.setdefault("boundary", {})["note"] = BOUNDARY_NOTES[leaf]
        # C-4 執行期守衛（06Z T3）—— upper_verbatim 須逐字出現於來源檔。
        # R-TM24 之來源隔離若只靠「產生器讀對檔案」，仍可能於後續處理
        # 被改寫；此處在產出當下比對來源，使該隔離成為必然而非約定。
        v = e["test_item"]["upper_verbatim"]
        if not v or v not in leaf_src:
            raise ContextError(
                f"C-4 守衛：{leaf} 之 upper_verbatim 未逐字出現於 "
                "data/leaf_descriptions.txt —— test_item 上半之唯一許可來源"
                "為該檔（R-TM24），不得取自任何下放包或上繳包之敘述")
        items.append(e)

    return {
        "batch": batch,
        "feature": "Time Management",
        "test_group": TEST_GROUP,
        "n_leaves": len(items),
        "leaves": items,
        "constraints": [
            "無 done region（BLANK workbook，R-TM5）—— canon §1.1 之第三層"
            "（以證據仲裁）在本 feature 不存在，pilot review 是唯一人工閘門。",
            "R-TM10-A1 SUSPENDED：不得援引任何他 feature 之 TC 樣式。"
            "步驟措辭與 ER 句式依 canon §4–§12 與本 feature profile 自訂。",
            "tc_id 不得產出 —— canon §10.3：generator 賦號，LLM 不得 emit。",
            "functional_safety 不得填 —— 其值由條文定且未決（A-TM24）。",
            "缺件一律 `PENDING: DR-{n}`，不得留空、不得填 NA；"
            "NA 僅限「確認不適用」（canon §8.4.3）。",
        ],
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--feature-dir", default="features/time_management")
    ap.add_argument("--batch", required=True, help="B1 | B2 | B3 | B4")
    ap.add_argument("--out")
    ap.add_argument("--self-test", action="store_true")
    args = ap.parse_args()
    fd = (REPO_ROOT / args.feature_dir
          if not Path(args.feature_dir).is_absolute() else Path(args.feature_dir))
    if args.self_test:
        return self_test(fd, args.batch)
    ctx = build(fd, args.batch)
    blob = json.dumps(ctx, ensure_ascii=False, indent=2)
    if args.out:
        path = fd / args.out if not Path(args.out).is_absolute() else Path(args.out)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(blob + "\n", encoding="utf-8")
        print(f"wrote {path} — {ctx['n_leaves']} leaves, batch {ctx['batch']}")
    else:
        print(blob)
    return 0


def self_test(feature_dir: Path, batch: str) -> int:
    """六類各驗**內容非退化**（R-TM52），並各附一退化紅向。

    R-TM52 之判準：「若該項之產生邏輯完全失效而只回傳空結構，本綠向會不會
    通過？」會 → 判準無效。故綠向不得以「鍵存在」為滿足條件。

    退化紅向之構造：monkeypatch 該類之產生函式使其回傳空結構，綠向須轉紅。
    **不觸及來源檔案。**
    """
    import copy
    bad = 0
    ctx = build(feature_dir, batch)
    leaf_src = (feature_dir / "data" / "leaf_descriptions.txt").read_text(
        encoding="utf-8")

    def chk_c1(c) -> bool:
        """三條自動判準各至少一訊號名，owns / not_ours 皆非空。"""
        n = 0
        for l in c["leaves"]:
            b = l.get("boundary", {})
            if "owns" in b:
                if not b["owns"] or not b["not_ours"]:
                    return False
                n += 1
        return n > 0

    def chk_c2(c) -> bool:
        """gap 之 placeholder 逐字含 PENDING: DR-5 且 objects 非空。

        **本判準要求 gaps 非空** —— 若回傳「無 gap 即通過」，則在不含
        A-TM13 兩片之批次上恆真，退化紅向無從轉紅（R-TM42：受測物對該項
        無鑑別力）。故 C-2 一律在含該兩片之批次上驗，見 `gap_ctx`。
        """
        gaps = [l["specification_reference"]["gap"] for l in c["leaves"]
                if "gap" in l["specification_reference"]]
        return bool(gaps) and all(
            g["objects"] and "PENDING: DR-5" in g["placeholder"] for g in gaps)

    def chk_c3(c) -> bool:
        """rendered 非空、含前綴、**前綴只出現一次**、無 `;`、升冪。"""
        for l in c["leaves"]:
            r = l["specification_reference"].get("rendered", "")
            if not r or SPEC_REF_PREFIX not in r or ";" in r:
                return False
            if r.count(SPEC_REF_PREFIX) != 1:        # canon §10.7 排列段
                return False
            ids = re.findall(r"\d{7}", r)
            if not ids or ids != sorted(ids):
                return False
        return True

    def chk_c4(c) -> bool:
        """upper_verbatim 須**逐字出現於 leaf_descriptions.txt**，非僅非空。"""
        for l in c["leaves"]:
            v = l["test_item"]["upper_verbatim"]
            if not v or v not in leaf_src:
                return False
        return True

    def chk_c5(c) -> bool:
        """至少一片有實際訊號，且其 segment 非全為 None（值來自實際來源）。"""
        tot = sum(len(l["signals"]["signals"]) for l in c["leaves"])
        return tot > 0

    def chk_c6(c) -> bool:
        """test_set 皆為七值之一；全 22 片之覆蓋恰為七組。"""
        from tm_rulings import BATCHES as _B
        if not all(l["test_set"] in TEST_SETS for l in c["leaves"]):
            return False
        cover = {TEST_SET_OF[n] for nums in _B.values() for n in nums}
        return cover == TEST_SETS and len(cover) == 7

    checks = [("C-1 界線", chk_c1), ("C-2 缺口", chk_c2),
              ("C-3 spec_ref", chk_c3), ("C-4 test_item", chk_c4),
              ("C-5 訊號", chk_c5), ("C-6 Test Set", chk_c6)]
    # C-2 之受測物須含 A-TM13 兩片（002 / 005）。當前批次若不含，
    # 改以含之者建 context —— R-TM42：受測物須能區分正確與錯誤實作。
    gap_batch = next(b for b, ns in BATCHES.items()
                     if any(n in SPEC_GAP for n in ns))
    gap_ctx = ctx if batch == gap_batch else build(feature_dir, gap_batch)
    if batch != gap_batch:
        print(f"（C-2 之受測物改用 {gap_batch} —— {batch} 不含 A-TM13 兩片，"
              f"對該項無鑑別力，R-TM42）")

    for name, fn in checks:
        ok = fn(gap_ctx if name.startswith("C-2") else ctx)
        bad += not ok
        print(f"{'PASS' if ok else '**FAIL**'} 綠向 {name}")

    # ── 退化紅向：把每一類之內容抽空，綠向須轉紅 ──────────────
    print()
    degens = [
        ("C-1 界線", lambda c: [l.pop("boundary", None) for l in c["leaves"]]),
        ("C-2 缺口", lambda c: [l["specification_reference"]["gap"].update(
            {"objects": [], "placeholder": ""})
            for l in c["leaves"] if "gap" in l["specification_reference"]]),
        ("C-3 spec_ref", lambda c: [l["specification_reference"].update(
            {"rendered": ""}) for l in c["leaves"]]),
        ("C-4 test_item", lambda c: [l["test_item"].update(
            {"upper_verbatim": "SOMETHING NOT IN THE SOURCE FILE"})
            for l in c["leaves"]]),
        ("C-5 訊號", lambda c: [l["signals"].update({"signals": {}})
                                for l in c["leaves"]]),
    ]
    for (name, fn), (_, degen) in zip(checks, degens):
        d = copy.deepcopy(gap_ctx if name.startswith("C-2") else ctx)
        degen(d)
        ok = fn(d)
        bad += ok                       # 退化後仍通過 = 綠向無鑑別力
        print(f"{'PASS' if not ok else '**FAIL**'} 退化紅向 {name}: "
              f"{'抽空後轉紅' if not ok else '抽空後仍綠 —— 綠向無鑑別力'}")

    # C-6 之退化：test_set 換成非七組之值
    d = copy.deepcopy(ctx)
    for l in d["leaves"]:
        l["test_set"] = "Time"
    ok = chk_c6(d)
    bad += ok
    print(f"{'PASS' if not ok else '**FAIL**'} 退化紅向 C-6 Test Set: "
          f"{'非七組之值已轉紅' if not ok else '仍綠 —— 無鑑別力'}")

    # ── 來源缺件紅向（原有）────────────────────────────────
    print()
    import shutil, tempfile
    tmp = Path(tempfile.mkdtemp()) / "fd"
    shutil.copytree(feature_dir, tmp, symlinks=True,
                    ignore=shutil.ignore_patterns("__pycache__"))
    (tmp / "data" / "leaf_descriptions.txt").unlink()
    try:
        build(tmp, batch)
        print("**FAIL** 紅向 抽掉 leaf_descriptions.txt：未報錯 —— 靜默略過")
        bad += 1
    except ContextError as exc:
        print(f"PASS 紅向 抽掉 leaf_descriptions.txt："
              f"{str(exc).splitlines()[0][:64]}")
    shutil.rmtree(tmp, ignore_errors=True)

    total = len(checks) + len(degens) + 2
    print(f"\n自驗：{total - bad} / {total}")
    return 1 if bad else 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except ContextError as exc:
        print(f"ABORTED: {exc}", file=sys.stderr)
        sys.exit(2)
