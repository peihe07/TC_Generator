#!/usr/bin/env python3
"""Step 3 (Time Management) — 機械漂移檢查，掃 `generated/*.json`。

modified by TC_Generator analysis round 05R under G-TM1/G-TM2

## 本檔之來源與界線（R-TM29）

**結構參照** `features/privacy/scripts/lint_tcs.py`（同為 BLANK workbook、
rev C 母本、spec_mode D，形態最近）：權威讀取而非寫死、閘門逐條回傳
`(gate, message)`、`--self-test` 對每閘造紅綠兩向、exit 0/1/2 之語義。

**不繼承其內容**（R-TM10-A1，射程由 R-TM29 界定）。各項現況
（2026-08-22 更新 —— 原文稱「一律留 TODO」已過時）：

  Test Set 值      **已實作**（`lint_test_set`）—— Part VII 七組，R-TM17 簽核
  priority 值域    **已實作**（`lint_priority_domain`）—— 自母本 P 欄 DV 讀取
  priority 分佈    仍未決，但標 `TODO(內容裁決)` 而非 R-TM10-A1（見檔末註）
  步驟措辭常數     仍為 `TODO(R-TM10-A1)`；已確定**自訂**（08 §3.2 實測：
                   既有交付件之步驟全為他 feature 專屬 UI，無可援引者），
                   擬定中（09 §3.1 常數表 v2 [PROPOSED]）
  ER 樣板字串      **刻意不做**（08 上繳 §6.3）—— 樣板化會使
                   `lint_step_er_count` 之 1:1 以湊行數方式通過而內容仍錯

Privacy 之 `step-actions`／`negative-scope`
兩閘編碼的是 R33-5／R33-1(d)，**那是 Privacy 之裁決，本 feature 不援引**。

## 權威一律讀取，不寫死

  design_method 詞彙  → 母本 `下拉選單` 分頁（`feature.yaml`
                        `lint.design_method_source: dropdown_sheet`）
  欄位對映            → `feature.yaml` `workbook.columns`（rev C：
                        design_method `R`、functional_safety `S`、
                        author `AA`、remarks `AH`）
  test_group          → `feature.yaml` `test_group`（`Time and Date`，R-TM8）
  spec_reference 形態 → `feature.yaml` `spec_reference_template`
                        ＋ CFTS015 docx 內實際存在之物件 id（**查得，不推算**）

exit 0 = 無發現，1 = 至少一項發現，2 = 呼叫錯誤。

用法：
    python3 features/time_management/scripts/lint_tcs.py --feature-dir features/time_management
    python3 features/time_management/scripts/lint_tcs.py --feature-dir features/time_management --self-test

**本腳本於 04 包只建立不執行**（下放包 §8「不執行 T5 所建之任何腳本」）。
"""
from __future__ import annotations

import argparse
import json
import tempfile
import re
import sys
import zipfile
import html
from pathlib import Path

try:
    import yaml
except ImportError:  # pragma: no cover
    yaml = None


PENDING_RE = re.compile(r"PENDING:\s*DR-\d+")      # canon §8.4.3 佔位形式
TEST_ITEM_TOKEN_MAX = 50                            # canon §4.3.1 上半上限
TEST_ITEM_TAIL_RE = re.compile(r"\([^)]+\)\s*$")    # canon §4.3.1 下半括號
DESIGN_METHOD_COUNT = 9          # B6 —— 母本 下拉選單!$A$1:$A$9
LEAF_COUNT = 22                  # B2 —— 037 之 leaf 全集
# B3 / B4 / C1 之裁決值取自 tm_rulings —— **context 層與本檔之單一來源**
# （06 §4.2）。各寫一份會漂移，且漂移時 lint 全綠：context 說 A、lint 驗 B，
# 生成照 A 寫則被 B 攔，呈現為「模型出錯」而非「規則不一致」。
sys.path.insert(0, str(Path(__file__).resolve().parent))
from tm_rulings import (                      # noqa: E402
    SPEC_GAP_LEAVES, SPEC_GAP, SPEC_GAP_DR, TEST_SETS, BOUNDARY_SIGNALS,
    load_ee_arch, atl_hi_placeholder, load_lid_table)

SPEC_REF_RE = re.compile(r"^CFTS015-(\d{7})$")      # B7(i) —— canon §10.7(a)

# R-TM64 —— 零真值之片，spec_reference 之唯一合法佔位（不含物件 id）
ATL_HI_BARE_PLACEHOLDER = "PENDING: DR-11 Atl-H 對應需求"

# A-TM26 之強制記錄字樣 —— 與 tm_rulings 之 tsv `ArchColumn` 欄同字面
ARCH_COL_MARK = "Atlantis High (col 26-30)"

# R-TM68 —— 佔位行之 DR 號擷取（供數值升冪比較）
PENDING_DR_RE = re.compile(r"PENDING:\s*DR-(\d+)")

class LintError(RuntimeError):
    """呼叫或權威讀取之失敗 —— 與「發現」區分，走 exit 2。"""


# ── 權威讀取 ────────────────────────────────────────────────
def load_authorities(feature_dir: Path) -> dict:
    """自 `feature.yaml`、母本與 spec docx 讀出全部權威值。

    **任何一項讀不到即 raise** —— 不以預設值頂替（R-TM7 同族：
    寫死之預設會使閘門在權威缺席時仍然全綠）。
    """
    if yaml is None:
        raise LintError("PyYAML 不可用")
    cfg_path = feature_dir / "feature.yaml"
    if not cfg_path.exists():
        raise LintError(f"缺 {cfg_path}")
    cfg = yaml.safe_load(cfg_path.read_text(encoding="utf-8"))

    wb_rel = cfg["paths"]["workbook"]
    wb_path = feature_dir / wb_rel
    if not wb_path.exists():
        raise LintError(f"母本不存在：{wb_path}")

    auth = {
        "cfg": cfg,
        "feature_dir": feature_dir,
        "workbook": wb_path,
        "sheet": cfg["workbook"]["sheet"],
        "header_row": int(cfg["workbook"]["header_row"]),
        "columns": dict(cfg["workbook"]["columns"]),
        "test_group": cfg["test_group"],
        "spec_template": cfg.get("spec_reference_template", ""),
        "design_methods": read_design_methods(wb_path, cfg),
        "spec_objects": read_spec_objects(feature_dir, cfg),
        "leaves": read_leaves(feature_dir),          # B2
        "priority": read_priority_domain(wb_path, cfg),   # C2
        "sys2_items": read_sys2_items(feature_dir, cfg),  # B7(ii)
    }
    auth["ee_arch"] = load_ee_arch(feature_dir)   # R-TM63 之單一來源
    auth["lid_table"] = load_lid_table(feature_dir)   # A-TM26
    return auth


def read_leaves(feature_dir: Path) -> set[str]:
    """B2 —— leaf 全集，唯一來源為 data/leaf_descriptions.txt（R-TM24）。"""
    path = feature_dir / "data" / "leaf_descriptions.txt"
    if not path.exists():
        raise LintError(
            f"缺 {path} —— 該檔為 leaf id 與 test_item 上半之唯一許可來源"
            "（R-TM24）。其對策為來源隔離，非人工記得")
    leaves = set(re.findall(r"SWE-RA-TIME&DATE-\d{3}",
                            path.read_text(encoding="utf-8")))
    if len(leaves) != LEAF_COUNT:
        raise LintError(
            f"{path.name} 讀到 {len(leaves)} 筆 leaf，期望 {LEAF_COUNT}")
    return leaves


def read_priority_domain(workbook: Path, cfg: dict) -> set[str]:
    """C2 —— priority 值域自母本 P 欄 DV 讀出，不寫死字面。"""
    import zipfile
    z = zipfile.ZipFile(workbook)
    xml = z.read("xl/worksheets/sheet6.xml").decode("utf-8")
    z.close()
    for m in re.finditer(r"<dataValidation\b[^>]*>(.*?)</dataValidation>",
                         xml, re.S):
        f1 = re.search(r"<formula1>\"?([^<\"]*)\"?</formula1>", m.group(1))
        if not f1:
            continue
        vals = {x.strip() for x in f1.group(1).split(",") if x.strip()}
        if vals and all(re.fullmatch(r"P[0-3]", v) for v in vals):
            return vals
    raise LintError(
        "母本工作表未讀到 P0..P3 之 dataValidation —— 不以寫死字面頂替，"
        "母本自身之清單才是值域權威")


def read_sys2_items(feature_dir: Path, cfg: dict) -> set[str]:
    """B7(ii) —— SYS2 第 5 欄之物件 id 全集。"""
    import openpyxl
    rel = cfg["paths"].get("sys1_export")
    if not rel:
        raise LintError("feature.yaml 未宣告 sys1_export（SYS2 匯出）")
    path = feature_dir / rel
    if not path.exists():
        raise LintError(f"SYS2 匯出不存在：{path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ids: set[str] = set()
    for r in wb["Basic Report"].iter_rows(min_row=2, values_only=True):
        for tok in re.split(r"[,\n]+", str(r[4] or "")):
            tok = tok.strip()
            if re.fullmatch(r"\d{7}", tok):
                ids.add(tok)
    wb.close()
    if not ids:
        raise LintError("SYS2 第 5 欄未讀到任何 7 位物件 id")
    return ids


def read_design_methods(workbook: Path, cfg: dict) -> set[str]:
    """自母本之 `下拉選單` 分頁讀 design_method 詞彙（逐字，不正規化）。"""
    if cfg.get("lint", {}).get("design_method_source") != "dropdown_sheet":
        raise LintError("feature.yaml 未宣告 design_method_source=dropdown_sheet")
    import openpyxl
    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    if "下拉選單" not in wb.sheetnames:
        wb.close()
        raise LintError("母本無 `下拉選單` 分頁")
    ws = wb["下拉選單"]
    vals: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        for c in row:
            if c is None:
                continue
            s = str(c).strip()
            # 設計方法詞條之形態為「中文 (English)」；其餘欄位不收
            if s and "(" in s and ")" in s:
                vals.add(s)
    wb.close()
    if not vals:
        raise LintError("`下拉選單` 分頁未讀到任何設計方法詞條")
    return vals


def read_spec_objects(feature_dir: Path, cfg: dict) -> set[str]:
    """自 CFTS015 docx 讀出實際存在之物件 id。

    **查得，不推算** —— Privacy 之 R30-1 記載了以偏移量推算 id 產生兩個
    錯誤 id 之實例；本 feature 不重蹈，每次執行都重讀 docx。
    """
    rel = cfg["paths"].get("spec_pdf")
    if not rel:
        raise LintError("feature.yaml 未宣告 spec_pdf（CFTS015 docx）")
    path = feature_dir / rel
    if not path.exists():
        raise LintError(f"spec docx 不存在：{path}")
    xml = zipfile.ZipFile(path).read("word/document.xml").decode("utf-8")
    ids: set[str] = set()
    for p in re.findall(r"<w:p[ >].*?</w:p>", xml, flags=re.S):
        t = "".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", p, flags=re.S))
        t = html.unescape(re.sub(r"<[^>]+>", "", t)).strip()
        m = re.match(r"^(\d{6,8})\s*:", t)
        if m:
            ids.add(m.group(1))
    if not ids:
        raise LintError("spec docx 未讀到任何物件 id")
    return ids


# ── 逐條閘門 ────────────────────────────────────────────────
def _steps(text: str) -> list[str]:
    return [s for s in str(text or "").split("\n") if s.strip()]


def lint_required_fields(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """必填欄位齊全**且非空** —— 欄名取自 `feature.yaml`，不寫死。

    B5（G-TM2 項 4 / A-TM21(f)）—— 原實作只檢查鍵存在。其 `base_tc()`
    為全空字串，故一條所有欄位皆為空之 TC 會全綠通過。
    """
    out = []
    for key in auth["columns"]:
        # 寫回端填之欄位，生成時不要求：tc_id 由序號賦值（canon §10.3）、
        # author / tc_ref_id / functional_safety 由條文決定（A4 之推論）。
        if key in ("tc_ref_id", "author", "tc_id", "functional_safety"):
            continue
        # remarks 之必要性是**條件式**的：僅 A-TM13 兩片與 BLOCKED 列需填，
        # 其餘列本應為空（BLANK_BY_DECISION）。該條件由 lint_spec_gap（B3）
        # 管，不由本閘管 —— 列為無條件必填會使 20 片正常 leaf 全部誤報。
        if key == "remarks":
            continue
        if key not in tc:
            out.append(("required-fields", f"{where}: 缺欄位 `{key}`"))
        elif not str(tc.get(key) or "").strip():
            out.append(("required-fields", f"{where}: 欄位 `{key}` 為空"))
    return out


def lint_leaf_source(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """B2（G-TM1 項 2）—— leaf 文字來源隔離（R-TM24）。

    `req_id` 須為 `data/leaf_descriptions.txt` 之 22 筆之一。該檔為 037
    `Requirement Description` 欄之直接輸出，是 test_item 上半之唯一許可
    來源 —— R-TM24 之對策為**來源隔離**，非人工記得不要抄下放包之簡寫。
    """
    out = []
    leaf = str(tc.get("req_id", ""))
    if leaf not in auth["leaves"]:
        out.append(("leaf-source",
                    f"{where}: req_id `{leaf}` 不在 data/leaf_descriptions.txt "
                    f"之 {LEAF_COUNT} 筆 leaf 全集內"))
    # canon §4.3.1（R-TM48 使其生效）—— test_item 兩段式。
    # 本閘只管**結構**（下半括號之存在、上半長度），不管措辭 —— 措辭屬
    # TC 內容，仍受 R-TM10-A1 拘束。
    item = str(tc.get("test_item") or "")
    if item.strip():
        if not TEST_ITEM_TAIL_RE.search(item.strip()):
            out.append(("test-item-shape",
                        f"{where}: test_item 缺下半之 `(...)` 測試目的。"
                        "canon §4.3.1：缺括號下半 = FAIL，不得出貨"))
        head = TEST_ITEM_TAIL_RE.sub("", item.strip()).strip()
        n = len(head.split())
        if n > TEST_ITEM_TOKEN_MAX:
            out.append(("test-item-shape",
                        f"{where}: test_item 上半 {n} token，超過 canon "
                        f"§4.3.1 之上限 {TEST_ITEM_TOKEN_MAX}。須摘句"
                        "（以與括號下半直接相關之句為限），全文以 "
                        "specification_reference 指回，不得整段傾倒"))
    return out


def lint_test_set(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """C1 —— Test Set 值域閘門。

    原標 `TODO(R-TM10-A1)`「待本 feature 之 framework 定案」，**該標記
    已過時**：framework Part VII 之七組已由 R-TM17 簽核（2026-08-20）。
    """
    got = str(tc.get("test_set", ""))
    if got not in TEST_SETS:
        return [("test-set",
                 f"{where}: test_set `{got}` 不是 framework Part VII 之七組"
                 "（R-TM17 已簽）")]
    return []


def lint_priority_domain(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """C2 —— priority **值域**閘門（非分佈）。

    值域自母本 P 欄之 DV 讀取，非 TC 內容裁決，故可立即實作。
    **分佈**（各 P 級之比例）才是內容裁決，見檔末 TODO(內容裁決)。
    """
    got = str(tc.get("priority", ""))
    if got and got not in auth["priority"]:
        return [("priority-domain",
                 f"{where}: priority `{got}` 不在母本 P 欄 DV 之值域 "
                 f"{sorted(auth['priority'])} 內")]
    return []


def lint_spec_gap(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """B3（G-TM1 項 3）—— A-TM13 兩片之缺口須於 Remarks 宣告。

    `-005` 與 `-002` 各引用一筆 SYS-RA，其來源物件（6151328 / 6151331）
    於 CFTS015 SR26 全檔零命中。該兩筆無章節可寫，**不得以鄰近章節填充**
    （§8.4.1）。Remarks 為空即為把缺口藏起來。
    """
    leaf = str(tc.get("req_id", ""))
    if leaf not in SPEC_GAP_LEAVES:
        return []
    rem = str(tc.get("remarks") or "")
    if not rem.strip():
        return [("spec-gap",
                 f"{where}: {leaf} 為 A-TM13 之受影響 leaf，其 Remarks 為空。"
                 "canon §8.4.3 明訂缺件不得留空，須寫 "
                 "`PENDING: DR-5 CFTS015 缺件物件 …`（R-TM41 處置訂正）")]
    # R-TM69(1) —— 驗**該 leaf 之特定 DR 號**，非「含任一 PENDING: DR-」。
    # 舊判準在 R-TM64 之前等值於「缺口已宣告」（當時 Remarks 只承載 DR-5）；
    # R-TM64 使 Remarks 成為全部佔位之單一落點後，一個 DR-11 佔位即可
    # 使本閘放行 —— 靜默失效兩輪（`15` §5.2）。
    want = f"PENDING: DR-{SPEC_GAP_DR}"
    if want not in rem:
        others = sorted({f"DR-{n}" for n in PENDING_DR_RE.findall(rem)})
        return [("spec-gap",
                 f"{where}: {leaf} 之 Remarks 未含 `{want}` 佔位"
                 + (f"（現有 {others}，但那是別的缺件）" if others else "")
                 + "。該 leaf 有 SYS-RA 之來源物件不在 CFTS015 基線內，"
                 "缺口須以其自身之 DR 號宣告（R-TM69）")]
    return []


def lint_boundary(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """B4（G-TM1 項 4）—— §8.2.1 鄰接對之訊號歸屬（R-TM23 / R-TM25 / R-TM55）。

    TC 全文若命中鄰片所擁有之訊號，即為跨界重複覆蓋。

    **本閘只讀 `not_ours`，不讀 `owns`**（R-TM55，07 T3）：判準是「命中了
    別人的」，不是「有沒有提自己的」。故 018 / 017 之 `owns` 為空集不影響
    其判定 —— 該二片之能力是行為（值之初始化／日期通道）而非某一訊號，
    本就無「自己的訊號」可列。以 `owns` 非空為前提會把該二片排除在射程外，
    正是 R-TM53 所記三處「可測而未測」之一部分。

    未列於 `BOUNDARY_SIGNALS` 者一律不判（如 022 —— R-TM55 明文駁回列入，
    因 B-2 之區辨軸為條件與值而非訊號名）。
    """
    leaf = str(tc.get("req_id", ""))
    rule = BOUNDARY_SIGNALS.get(leaf)
    if rule is None:
        return []
    # R-TM70 —— 掃描範圍**排除 test_item 之上半 verbatim 與 reasoning**。
    #   上半為逐字照錄之上游文字（canon §4.3.1），作者無權改寫；若界線及於
    #   上半，唯一的遵守方式就是改寫需求原文 —— 那正是 §4.3.1 所禁。
    #   reasoning 排除之理由見 `15` §5.3：掃 reasoning 會把「本條不涵蓋
    #   SNA」這句**聲明本身**當成違規。
    ti = str(tc.get("test_item") or "")
    lower = ti.split("\n", 1)[1] if "\n" in ti else ""
    body = " ".join([lower] + [str(tc.get(k) or "") for k in
                    ("pre_conditions", "input_test_data",
                     "test_procedure", "expected_result")])
    out = []
    for sig in rule["not_ours"]:
        if sig in body:
            out.append(("boundary",
                        f"{where}: {leaf} 之內文命中 `{sig}`，該訊號屬鄰片。"
                        f"{rule['why']}（Part VII §8.2.1，R-TM23 / R-TM25）"))
    return out


def lint_arch(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """R-TM63 —— Atl-Mid 專屬物件不得以 `CFTS015-{id}` 形式出現。

    **與 context 層共用同一 tsv**（R-TM63 第 5 項）—— 判準不在本檔複寫，
    由 `tm_rulings.load_ee_arch()` 供給。兩層若各自判定，漂移時 lint 全綠
    而內容錯（R-TM59 / R-TM61 之同族）。

    本閘只管**該不該寫真值**，不管佔位字串之措辭 —— 後者由 B3 之
    PENDING 規則與 R-TM40 之格式閘各自負責，三者射程不重疊。
    """
    arch = auth.get("ee_arch")
    if not arch:
        return []
    leaf = str(tc.get("req_id", ""))[-3:]
    per = arch.get(leaf)
    if not per:
        return []
    body = str(tc.get("spec_reference") or "")
    out = []
    for oid, info in sorted(per.items()):
        if info["is_atl_hi"]:
            continue
        if re.search(rf"(?<!\d){re.escape(oid)}(?!\d)", body) and \
                atl_hi_placeholder(oid) not in body:
            out.append(("arch",
                        f"{where}: 物件 {oid} 標為 `{info['ee']}`（非 Atl-H），"
                        f"不得以 `CFTS015-{oid}` 形式寫入 —— R-TM63 第 2 項："
                        f"該條目須為 `{atl_hi_placeholder(oid)}`"))
    return out


def lint_arch_column(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """A-TM26 之強制記錄 —— 用了 LID 訊號者，reasoning 須註明架構欄與來源列。

    A-TM26 逐字：「凡自 LID 表取值者，須於同一處記錄取自哪一組架構欄，
    **無此記錄之取值一律視為未驗**」。該判準原先只靠人工遵守 ——
    `13` 生成時 30 處訊號中僅 2 處記錄，靠自檢才發現（`13` §5.3）。

    本閘只驗**記錄之存在**，不驗其正確性：記了 `Atlantis High (col 26-30)`
    卻其實取自 Powernet 欄，本閘攔不住。**該層由 context 之單一來源
    （`load_lid_table`）保證** —— 值不由生成端自行查表，故兩者射程互補。
    """
    lid = auth.get("lid_table")
    if not lid:
        return []
    body = " ".join(str(tc.get(k) or "") for k in
                    ("pre_conditions", "input_test_data",
                     "test_procedure", "expected_result"))
    used = sorted({n for n in lid if f"${n}$" in body})
    if not used:
        return []
    rsn = str(tc.get("reasoning") or "")
    out = []
    if ARCH_COL_MARK not in rsn:
        out.append(("arch-column",
                    f"{where}: 內文用了 LID 訊號 {used}，但 reasoning 未註明 "
                    f"`{ARCH_COL_MARK}` —— A-TM26：無此記錄之取值視為未驗"))
        return out
    for n in used:
        row = lid[n]["source_row"]
        if row not in rsn:
            out.append(("arch-column",
                        f"{where}: 用了 `${n}$` 但 reasoning 未載其來源列 "
                        f"{row} —— A-TM26 要求可回溯至該筆對映"))
    return out


def lint_remarks_order(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """R-TM68 —— Remarks 之佔位行須依 DR 號**數值**升冪，一佔位一行。

    **數值比較，非字串比較**：`DR-11` 之字串序在 `DR-5` 之前（`1` < `5`），
    以字串排序會把一個已正確升冪之 Remarks 判為錯 —— 而該誤判之方向是
    「對的被判錯」，比放行更難察覺其成因。
    """
    raw = str(tc.get("remarks") or "")
    if not raw.strip():
        return []
    out = []
    nums, lines = [], []
    for ln in raw.splitlines():
        s = ln.strip()
        if not s:
            continue
        found = PENDING_DR_RE.findall(s)
        if not found:
            continue
        if len(found) > 1:
            out.append(("remarks-order",
                        f"{where}: Remarks 單行含 {len(found)} 個佔位 "
                        f"—— R-TM68 第 1 項：一佔位一行，不合併"))
        nums.append(int(found[0]))
        lines.append(s[:40])
    if nums != sorted(nums):
        out.append(("remarks-order",
                    f"{where}: Remarks 之佔位非依 DR 號升冪 —— 現為 "
                    f"{nums}，應為 {sorted(nums)}（R-TM68 第 2 項）"))
    return out


def lint_data_placement(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """canon §4.5 / R-TM66 —— 同一資料不得在 input_test_data 與 Procedure 重複。

    **判準為「逐字相同之行」，不涉任何特定字串** —— R-TM66 所記之同型風險
    （日後新增之 DR 號字串不同即漏掃）於此消除：本閘不認得 DR 號，
    只比對兩欄之行。

    只驗**重複**，不驗歸屬是否正確：一筆真正的獨立資料集寫在
    input_test_data 而 Procedure 未提，本閘不報 —— 該判斷需讀懂資料性質，
    非本層所能。射程刻意窄，以免誤攔合法之資料欄。
    """
    itd = str(tc.get("input_test_data") or "").strip()
    if not itd or itd.upper() == "NA":
        return []
    proc = {l.strip() for l in str(tc.get("test_procedure") or "").splitlines()}
    proc |= {re.sub(r"^\d+\.\s*", "", l).strip() for l in proc}
    out = []
    for ln in [l.strip() for l in itd.splitlines() if l.strip()]:
        if ln in proc:
            out.append(("data-placement",
                        f"{where}: input_test_data 之 `{ln[:44]}` 與 "
                        "test_procedure 逐字重複 —— canon §4.5：互動操作屬 "
                        "Procedure，本欄應為 `NA`（R-TM66）"))
    return out


def lint_placeholder_completeness(tc: dict, auth: dict,
                                  where: str) -> list[tuple[str, str]]:
    """R-TM69(2) —— Remarks 之佔位集合須與「應有」相等，**缺與多皆報**。

    應有集合由兩處推得，皆為既有之單一來源：
      DR-11  該 leaf 之 Atl-Mid 物件（`ee_architecture_by_leaf.tsv`）
      DR-5   該 leaf 若在 `SPEC_GAP` 內

    **報「多」而非只報「缺」**：多出之佔位表示宣告了一個本片沒有的缺口，
    其後果是 DR 答覆回來時對不上任何欄位 —— 與缺漏同樣使佔位失去指向。

    **本閘只比對 DR-11 之物件層級與 DR-5 之有無**，不驗其他 DR 號
    （DR-8/9/10/12/20 為步驟措辭之佔位，寫在 procedure 而非 Remarks，
    其齊全性無資料可據）。射程寫窄，理由同 `15` §3.2。
    """
    arch = auth.get("ee_arch")
    if not arch:
        return []
    leaf3 = str(tc.get("req_id", ""))[-3:]
    per = arch.get(leaf3)
    if per is None:
        return []
    ref = str(tc.get("spec_reference") or "")
    rem = str(tc.get("remarks") or "")
    # 應有：本 TC 之 spec_reference 所涉 leaf 中，屬 Atl-Mid 者。
    # **以該 leaf 之全集為準**（canon §10.7 之「只列該 TC 驗證之物件」
    # 拘束真值欄；缺口宣告則須及於該 leaf 之全部 Atl-Mid 物件，
    # 否則佔位會隨 TC 之取捨而漏）。
    want_mid = {o for o, v in per.items() if not v["is_atl_hi"]
                and v["ee"] != "(不在 docx —— A-TM13)"}
    have_mid = set(re.findall(r"CFTS015-(\d{7}) 標為", rem))
    out = []
    for o in sorted(want_mid - have_mid):
        out.append(("placeholder-completeness",
                    f"{where}: 缺 DR-11 佔位 —— 物件 {o} 標為 "
                    f"`{per[o]['ee']}`，應於 Remarks 宣告（R-TM69）"))
    for o in sorted(have_mid - want_mid):
        out.append(("placeholder-completeness",
                    f"{where}: 多出 DR-11 佔位 —— 物件 {o} 非本片之 "
                    "Atl-Mid 引用，該佔位無對應欄位（R-TM69）"))
    want_gap = str(tc.get("req_id", "")) in SPEC_GAP_LEAVES
    have_gap = f"PENDING: DR-{SPEC_GAP_DR}" in rem
    if want_gap and not have_gap:
        out.append(("placeholder-completeness",
                    f"{where}: 缺 DR-{SPEC_GAP_DR} 佔位（A-TM13 受影響 leaf）"))
    if have_gap and not want_gap:
        out.append(("placeholder-completeness",
                    f"{where}: 多出 DR-{SPEC_GAP_DR} 佔位 —— 本片不在 "
                    "SPEC_GAP 內"))
    return out


def lint_no_tc_id(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """B8（G-TM2 項 3 訂正）—— TC JSON 不得攜帶 tc_id。

    canon §10.3 末句：`the generator handles assignment, the LLM does not
    emit tc_id`。本閘為生成端違反該條之偵測點；賦號由 write_back 依列
    位置為之。
    """
    if "tc_id" in tc:
        return [("no-tc-id",
                 f"{where}: TC JSON 含 `tc_id` 鍵。canon §10.3 明訂 "
                 "generator 賦號、LLM 不得產出；賦號由 write_back 依列位置為之")]
    return []


def lint_test_group(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """`test_group` 須逐字等於 `feature.yaml` 之值（R-TM8）。"""
    want = auth["test_group"]
    got = str(tc.get("test_group", ""))
    if got != want:
        return [("test-group", f"{where}: test_group `{got}` ≠ `{want}`")]
    return []


def lint_design_method(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """`design_method` 須為母本 `下拉選單` 之詞條之一（逐字）。"""
    got = str(tc.get("design_method", "")).strip()
    if got and got not in auth["design_methods"]:
        return [("design-method",
                 f"{where}: design_method `{got[:40]}` 不在母本下拉選單詞彙內")]
    return []


def lint_spec_reference(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """B7 —— spec_reference 三重閘門（R-TM40 / R-TM41 / canon §10.7(a)）。

    (i)   形式須為 `CFTS015-{7 位}`（canon §10.7(a)：CFTS 母文件 →
          `CFTS{nnn}-{ObjectID}`，ObjectID 為 Polarion 7 位號碼）
    (ii)  該 7 位須存在於 SYS2 第 5 欄之全集 —— 即真有此物件被本 feature
          之需求引用，非任意 7 位數
    (iii) 該 7 位須存在於 CFTS015 docx —— **此項擋掉 6151328 / 6151331**
          （R-TM41）。格式湊得出來不等於來源有此內容（§8.4.1）

    第 (iii) 項為現存版之 `lint_spec_reference` 所有，G-TM2 項 6 明列為
    「不得回退」；本次由「額外保護」升為「格式正確性之必要條件」。

    **排列**依 canon §10.7 之排列段：同一文件內多個 ObjectID 以 `, ` 續列
    且**文件前綴僅敘明一次**、**禁用 `;`**、升冪。
    """
    # R-TM65 —— 鍵名以 feature.yaml 之 workbook.columns 宣告為準。
    raw = str(tc.get("spec_reference", "") or "")
    out: list[tuple[str, str]] = []
    if not raw.strip():
        return [("spec-reference", f"{where}: spec_reference 為空")]
    # R-TM64 例外 —— 零真值之片（全部引用物件皆非 Atl-Hi，如 020 / 021）：
    # 欄值**恰為**單一佔位即通過。逐項明細在 Remarks，N 欄只標「待補」。
    # 用 `==` 而非 `startswith` —— 後者會放行「佔位 + 任何尾巴」，
    # 使 R-TM64 所廢止之 ` / ` 並列從這個豁免口溜回來。
    if raw.strip() == ATL_HI_BARE_PLACEHOLDER:
        return []
    if ";" in raw:
        out.append(("spec-reference",
                    f"{where}: 含 `;` —— canon §10.7 排列段明文禁用，"
                    "同一文件內多個 ObjectID 以 `, ` 續列"))
    # 一來源文件一行；同一行內前綴僅首個 token 帶
    for line in [l.strip() for l in raw.splitlines() if l.strip()]:
        toks = [x.strip() for x in line.split(",") if x.strip()]
        if not toks:
            continue
        m = SPEC_REF_RE.match(toks[0])
        if not m:
            out.append(("spec-reference",
                        f"{where}: 首 token `{toks[0]}` 不符 `CFTS015-<7 位>`"
                        "（canon §10.7(a)）"))
            continue
        ids = [m.group(1)]
        for tok in toks[1:]:
            if SPEC_REF_RE.match(tok):
                out.append(("spec-reference",
                            f"{where}: `{tok}` 重複帶前綴 —— canon §10.7 排列段"
                            "「文件前綴僅敘明一次」"))
                ids.append(tok.split("-", 1)[1])
            elif re.fullmatch(r"\d{7}", tok):
                ids.append(tok)
            else:
                out.append(("spec-reference",
                            f"{where}: 續列 token `{tok}` 非 7 位物件 id"))
        if ids != sorted(ids):
            out.append(("spec-reference",
                        f"{where}: 物件 id 未升冪 —— canon §10.7 排列段"))
        for oid in ids:
            if oid not in auth["sys2_items"]:
                out.append(("spec-reference",
                            f"{where}: 物件 `{oid}` 不在 SYS2 "
                            "`Source Requirement items` 之全集內"))
            if oid not in auth["spec_objects"]:
                out.append(("spec-reference",
                            f"{where}: 物件 `{oid}` 不存在於 CFTS015 docx。"
                            "格式湊得出來不等於來源有此內容（R-TM41 / §8.4.1）"))
    return out


# R-TM58 —— 合法 037 檔名形態（去副檔名）。**只判形態不判歸屬** ——
# 「是 037 檔名」不等於「是**本**工作簿之 037」（A-H26 即他 feature 之 037）。
D5_037_RE = re.compile(r"FM-WI-FSM-037-[A-Z]\d{2}-")


def lint_d5_scope(auth: dict) -> list[tuple[str, str]]:
    """B1（G-TM1 項 1）—— D5 Scope 守衛，**具名失敗，不與 header drift 混列**。

    **R-TM58（08 §2）改判**：D5 為空即通過，不再要求 `PENDING: DR-` 佔位。
    canon §8.4.3 之射程為逐列 TC 資料欄，非工作簿層之表頭格；且唯一可測
    之正確交付實例（UserProfiles_20260820）之 D5 實測為空。

    本閘之存在意義隨之改變：由「使 D5 空著這件事現形」改為
    **偵測 D5 被誤填**——若 D5 非空且非合法 037 檔名形態，報 A-H26 同型缺陷。
    A-TM02a 不因此結案（037 身分仍未定），但其阻塞 D5 之性質解除。
    """
    import openpyxl
    wb = openpyxl.load_workbook(auth["workbook"], read_only=True, data_only=True)
    ws = wb[auth["sheet"]]
    v = ws["D5"].value
    wb.close()
    s = str(v or "")
    if not s.strip():
        return []                       # R-TM58：空為交付先例所支持之狀態
    if PENDING_RE.search(s):
        return [("spec-scope-pending",
                 f"D5 為佔位 {s!r} —— R-TM58 撤回 PENDING 佔位之要求，D5 應"
                 "維持空白。此非缺陷，但該佔位須移除以與交付先例一致")]
    if s.upper() == "NA":
        return [("d5-scope",
                 "D5 為 `NA` —— canon §8.4.3 明訂 NA 僅限「確認不適用」。"
                 "D5 之空白非「不適用」之宣告（R-TM58），須留空而非填 NA")]
    if D5_037_RE.search(s):
        return [("spec-scope-pending",
                 f"D5 為 037 檔名形態 {s!r} —— 須先確認其確為**本**工作簿之"
                 "依據 037（A-TM02a 身分未定），並更新 A-TM02a 與 A-TM11")]
    return [("d5-scope",
             f"D5 已填 {s!r} —— 既非空白（R-TM58 之應然狀態）亦非 037 檔名"
             "形態。A-H26 同型缺陷：不得以 feature 名或 spec 標題組值填入"
             "（R-TM9-A2），亦不得填入他 feature 之 037")]


def lint_step_er_count(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    """步驟數與 ER 行數相等。

    **此為結構性不變量，非措辭判準** —— 故不受 R-TM10-A1 拘束。
    """
    proc = _steps(tc.get("test_procedure"))
    er = _steps(tc.get("expected_result"))
    if proc and er and len(proc) != len(er):
        return [("step-er-count",
                 f"{where}: 步驟 {len(proc)} 條 vs ER {len(er)} 條")]
    return []


# TODO(R-TM10-A1): 步驟措辭之閘門（禁用動詞、步長上限、最終步須帶查核目標）
#   —— 其詞彙與門檻屬 TC 內容，須由本 feature 之條文決定，不得援引他 feature。
#   Privacy 之 `step-actions`（R33-5：一步一動作，判動詞數不判連接詞）
#   為其自身裁決之編碼，本 feature 未有對應條文前不實作。

# C1 —— Test Set 值域閘門已實作（lint_test_set）。原 TODO(R-TM10-A1)
#   「待本 feature 之 framework 定案」已過時：Part VII 七組經 R-TM17 簽核。

# C2 —— priority **值域**閘門已實作（lint_priority_domain），自母本 P 欄
#   DV 讀取，非 TC 內容裁決。
# TODO(內容裁決): priority **分佈**（各 P 級之比例）—— 屬 TC 內容裁決，
#   待本 feature 之條文決定。**與 TODO(R-TM10-A1) 區分**：後者管跨 feature
#   樣式參照，本項管本 feature 自身之內容決定，不會隨 R-TM10-A1 解除。

# TODO(R-TM10-A1): Input Test Data 填法之閘門 —— 待本 feature 之條文決定。


GATES = (
    lint_required_fields,
    lint_leaf_source,        # B2
    lint_test_set,           # C1
    lint_priority_domain,    # C2
    lint_spec_gap,           # B3
    lint_boundary,           # B4
    lint_arch,               # R-TM63
    lint_arch_column,        # A-TM26
    lint_remarks_order,      # R-TM68
    lint_data_placement,     # R-TM66 / canon §4.5
    lint_placeholder_completeness,   # R-TM69(2)
    lint_no_tc_id,           # B8
    lint_test_group,
    lint_design_method,
    lint_spec_reference,
    lint_step_er_count,
)


def lint_tc(tc: dict, auth: dict, where: str) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    for gate in GATES:
        out += gate(tc, auth, where)
    return out


def lint_file(path: Path, auth: dict) -> list[tuple[str, str]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    tcs = data.get("tcs", data if isinstance(data, list) else [])
    out: list[tuple[str, str]] = []
    for i, tc in enumerate(tcs, 1):
        out += lint_tc(tc, auth, f"{path.name}#{i}")
    return out


# ── 範圍向自驗（R-G9）───────────────────────────────────────
def base_tc(auth: dict) -> dict:
    """一條應全綠之最小 TC —— 其值全部取自權威，不寫死措辭。

    B5 之後，全空字串不再算合規，故各欄須有實質值。
    """
    leaf = sorted(auth["leaves"])[0]          # SWE-RA-TIME&DATE-001
    oid = sorted(auth["spec_objects"] & auth["sys2_items"])[0]
    tc = {k: "x" for k in auth["columns"]
          if k not in ("tc_id", "author", "tc_ref_id", "functional_safety")}
    tc["req_id"] = leaf
    tc["test_group"] = auth["test_group"]
    tc["test_set"] = "Manual Setting"
    tc["design_method"] = sorted(auth["design_methods"])[0]
    tc["spec_reference"] = f"CFTS015-{oid}"
    tc["priority"] = sorted(auth["priority"])[0]
    tc["test_procedure"] = "1. a\n2. b"
    tc["expected_result"] = "1. a\n2. b"
    # R-TM69(2) —— 合規之最小 TC 須帶該 leaf 應有之全部 DR-11 佔位。
    # **動態取自 auth["ee_arch"]，不寫死物件 id** —— 寫死會使本 vector 在
    # tsv 更新後仍然全綠，而那正是它該偵測的漂移（R-TM52）。
    mid = sorted(o for o, v in (auth.get("ee_arch", {}).get(leaf[-3:]) or {}).items()
                 if not v["is_atl_hi"] and "不在 docx" not in v["ee"])
    tc["remarks"] = "\n".join(atl_hi_placeholder(o) for o in mid)
    # canon §4.3.1 —— 上半 verbatim + 下半 `(...)` 測試目的
    tc["test_item"] = "The software shall set time (manual entry)"
    return tc


def self_test(auth: dict) -> int:
    """每閘一紅一綠。**綠向證明其不對合規者誤報，紅向證明其抓得到。**

    紅向一律以**刻意構造之壞輸入**觸發，不以「理論上會 raise」代替。
    構造時避免與他閘混淆（R-TM45 之同層版本）：例如 B5 之紅向只留一欄
    為空，不用全空 TC —— 後者會同時觸發多閘，看不出是哪一閘抓到的。
    """
    bad = 0
    green = base_tc(auth)
    v = lint_tc(green, auth, "GREEN")
    if v:
        bad += 1
        print(f"**FAIL** 綠向：合規之 TC 轉紅 → {v}")
    else:
        print("PASS 綠向：合規之 TC 未轉紅")

    gap_leaf = sorted(SPEC_GAP_LEAVES)[0]
    spec_gap_ph = "PENDING: DR-5 CFTS015 缺件物件 " + SPEC_GAP[gap_leaf[-3:]]["object"]
    oid_ok = green["spec_reference"].split("-")[1]
    reds = [
        ("required-fields  (B5 空值，只留一欄空)",
         {**green, "pre_conditions": "   "}),
        ("leaf-source      (B2 不在 22 筆內)",
         {**green, "req_id": "SWE-RA-TIME&DATE-099"}),
        ("test-set         (C1 非七組之一)",
         {**green, "test_set": "Time"}),
        ("priority-domain  (C2 值域外)",
         {**green, "priority": "P9"}),
        ("spec-gap         (B3 A-TM13 leaf 而 Remarks 空)",
         {**green, "req_id": gap_leaf, "remarks": ""}),
        ("boundary         (B4 011 命中 008 之訊號)",
         {**green, "req_id": "SWE-RA-TIME&DATE-011",
          "expected_result": "HU transmits $DateTmHour$ within 1000 ms"}),
        ("boundary         (B4 R-TM55：018 命中 011 之 $DateTmFormat$，owns 為空)",
         {**green, "req_id": "SWE-RA-TIME&DATE-018",
          "expected_result": "IPC restores $DateTmFormat$ after reset"}),
        ("boundary         (B4 R-TM55：017 命中 014 之 $GPSDateTm，owns 為空)",
         {**green, "req_id": "SWE-RA-TIME&DATE-017",
          "expected_result": "HU sends $GPSDateTmHour$ on the date channel"}),
        ("no-tc-id         (B8 JSON 攜帶 tc_id)",
         {**green, "tc_id": "NR1L-TimeAndDate-001"}),
        ("test-group       (R-TM8：feature 名不是 Test Group)",
         {**green, "test_group": "Time Management"}),
        ("design-method    (不在母本下拉選單內)",
         {**green, "design_method": "不存在之方法 (Nope)"}),
        ("spec-reference i (B7 形式不符 CFTS015-<7位>)",
         {**green, "spec_reference": "4813905"}),
        ("spec-reference ii(B7 7 位不在 SYS2 全集)",
         {**green, "spec_reference": "CFTS015-9999999"}),
        ("spec-reference iii(B7 R-TM41：不在 CFTS015 docx)",
         {**green, "spec_reference": "CFTS015-6151328"}),
        ("spec-reference   (B7 canon §10.7 禁用 ;)",
         {**green, "spec_reference": f"CFTS015-{oid_ok}; CFTS015-{oid_ok}"}),
        ("spec-reference   (B7 前綴重複敘明)",
         {**green, "spec_reference": f"CFTS015-{oid_ok}, CFTS015-{oid_ok}"}),
        ("step-er-count    (步驟數 vs ER 行數)",
         {**green, "expected_result": "1. a"}),
        ("spec-gap         (L2：Remarks 有值但非 PENDING 佔位)",
         {**green, "req_id": gap_leaf, "remarks": "缺口見 A-TM13"}),
        ("test-item-shape  (L3：缺下半括號)",
         {**green, "test_item": "The software shall set time"}),
        ("test-item-shape  (L3：上半 51 token 未摘句)",
         {**green, "test_item": " ".join(["word"] * 51) + " (purpose)"}),
    ]
    for name, tc in reds:
        gate = name.split()[0]
        v = lint_tc(tc, auth, f"RED[{gate}]")
        hit = any(g == gate for g, _ in v)
        bad += not hit
        msg = [m for g, m in v if g == gate]
        print(f"{'PASS' if hit else '**FAIL**'} 紅向 {name}: "
              f"{msg[0].split(': ', 1)[-1][:78] if msg else '未叫 —— 閘失效'}")

    # 綠向 2：A-TM13 leaf 且 Remarks 為合規佔位 → 不應報 spec-gap
    ok_gap = {**green, "req_id": gap_leaf,
              "remarks": "PENDING: DR-5 CFTS015 缺件物件 6151328"}
    v = [g for g, _ in lint_tc(ok_gap, auth, "GREEN2") if g == "spec-gap"]
    bad += bool(v)
    print(f"{'PASS' if not v else '**FAIL**'} 綠向 2 (A-TM13 leaf 帶 "
          f"PENDING 佔位): {'未誤報 spec-gap' if not v else v}")

    # ── R-TM63 之 arch 閘（13 T4）──────────────────────────
    # 綠向：003 之 Atl-Hi 物件 4813923 寫真值 → 不報
    # 紅向：003 之 Atl-Mid 物件 4814088 寫真值 → 報 arch
    # 綠向 2：同一物件寫佔位 → 不報（佔位與真值並存之證明）
    a_cases = [
        ("綠向 (003 之 Atl-Hi 物件 4813923 寫真值)",
         {**green, "req_id": "SWE-RA-TIME&DATE-003",
          "spec_reference": "CFTS015-4813923"}, False),
        ("紅向 (003 之 Atl-Mid 物件 4814088 寫真值 → 須報)",
         {**green, "req_id": "SWE-RA-TIME&DATE-003",
          "spec_reference": "CFTS015-4813923, 4814088"}, True),
        # R-TM64 後：佔位移 Remarks，spec_reference 只留真值。
        ("綠向 (Atl-Mid 物件之佔位移至 Remarks → 不報)",
         {**green, "req_id": "SWE-RA-TIME&DATE-003",
          "spec_reference": "CFTS015-4813923",
          "remarks": atl_hi_placeholder("4814088")}, False),
        # 020 之四個物件全為 Atl-Mid（R-TM63 第 4 項），任一寫真值即報。
        # 首版此處用了不存在之 id 4814035 而未報 —— 構造錯誤，非閘失效
        # （R-TM56 所記之同一形態：兩者現象相同）。id 改自 tsv 實取。
        ("紅向 (020 全片 Atl-Mid，寫任一真值即報)",
         {**green, "req_id": "SWE-RA-TIME&DATE-020",
          "spec_reference": "CFTS015-4814064"}, True),
    ]
    for name, tc, want in a_cases:
        got = [m for g, m in lint_tc(tc, auth, "ARCH") if g == "arch"]
        ok = bool(got) is want
        bad += not ok
        print(f"{'PASS' if ok else '**FAIL**'} arch {name}: "
              f"{(got[0][:64] if got else '未報')}")

    # ── R-TM64 之零真值例外 + A-TM26 之 ArchColumn 閘（14 T3）──
    lid_tc = {**green, "req_id": "SWE-RA-TIME&DATE-008",
              "expected_result": "1. $DateTmHour$ carries the hour value\n2. b",
              "test_procedure": "1. a\n2. b"}
    mark = "訊號取自 LID 表 Atlantis High (col 26-30) 欄，來源列 409"
    x_cases = [
        # R-TM64 —— 零真值之片，欄值恰為單一佔位
        ("綠向 R-TM64 (spec_reference 恰為單一 DR-11 佔位 → 不報)",
         {**green, "spec_reference": ATL_HI_BARE_PLACEHOLDER},
         "spec-reference", False),
        ("紅向 R-TM64 (佔位後綴任何尾巴 → 仍報)",
         {**green, "spec_reference": ATL_HI_BARE_PLACEHOLDER + " / CFTS015-4813919"},
         "spec-reference", True),
        ("紅向 R-TM64 (佔位帶物件 id —— 明細應在 Remarks)",
         {**green, "spec_reference": atl_hi_placeholder("4814088")},
         "spec-reference", True),
        # A-TM26 —— 用了 LID 訊號而 reasoning 無記錄
        ("紅向 A-TM26 (用 $DateTmHour$ 而 reasoning 無 ArchColumn)",
         lid_tc, "arch-column", True),
        ("紅向 A-TM26 (有 ArchColumn 但缺來源列號)",
         {**lid_tc, "reasoning": "訊號取自 LID 表 Atlantis High (col 26-30) 欄"},
         "arch-column", True),
        ("綠向 A-TM26 (ArchColumn 與來源列俱全 → 不報)",
         {**lid_tc, "reasoning": mark}, "arch-column", False),
        ("綠向 A-TM26 (未用任何 LID 訊號 → 不判)",
         {**green, "reasoning": ""}, "arch-column", False),
    ]
    for name, tc, gate, want in x_cases:
        got = [m for g, m in lint_tc(tc, auth, "X") if g == gate]
        ok = bool(got) is want
        bad += not ok
        print(f"{'PASS' if ok else '**FAIL**'} {name}: "
              f"{(got[0][:60] if got else '未報')}")

    # ── R-TM68 / R-TM66 兩閘（15 T3）──────────────────────
    # **紅向皆附構造複驗（R-TM67）**：先證明壞值確實違反該閘所檢之條件，
    # 再看守衛反應 —— 否則「守衛失效」與「壞值不壞」在現象上相同。
    import re as _re
    def _drs(s):
        return [int(x) for x in _re.findall(r"PENDING:\s*DR-(\d+)", s)]
    r_ok = "PENDING: DR-5 CFTS015 缺件物件 6151328\nPENDING: DR-11 Atl-H 對應需求"
    r_bad = "PENDING: DR-11 Atl-H 對應需求\nPENDING: DR-5 CFTS015 缺件物件 6151328"
    r_merged = "PENDING: DR-5 缺件 / PENDING: DR-11 Atl-H 對應需求"
    itd_dup = {"input_test_data": "Wake the CAN bus",
               "test_procedure": "1. Wake the CAN bus\n2. b",
               "expected_result": "1. a\n2. b"}
    itd_ok = {**itd_dup, "input_test_data": "Hours = 10, Minutes = 30"}
    # ── 構造複驗（R-TM67）──
    print(f"   [構造複驗] r_ok 之 DR 序 {_drs(r_ok)} 已升冪 = {_drs(r_ok)==sorted(_drs(r_ok))}"
          f"；r_bad 之 DR 序 {_drs(r_bad)} 已升冪 = {_drs(r_bad)==sorted(_drs(r_bad))}")
    print(f"   [構造複驗] r_bad 之字串序若以字典序排 = "
          f"{sorted(r_bad.splitlines()) == r_bad.splitlines()}"
          f"（若為 True，則字串比較會放行此壞值 —— 本閘須用數值）")
    _p = {l.strip() for l in itd_dup['test_procedure'].splitlines()}
    _p |= {_re.sub(r'^\d+\.\s*','',l).strip() for l in _p}
    print(f"   [構造複驗] itd_dup 之 input_test_data 確實出現於 procedure = "
          f"{itd_dup['input_test_data'] in _p}"
          f"；itd_ok 則為 {itd_ok['input_test_data'] in _p}")
    y_cases = [
        ("綠向 R-TM68 (DR-5 在 DR-11 之前 → 不報)",
         {**green, "remarks": r_ok}, "remarks-order", False),
        ("紅向 R-TM68 (DR-11 在 DR-5 之前 → 報)",
         {**green, "remarks": r_bad}, "remarks-order", True),
        ("紅向 R-TM68 (兩佔位擠一行 → 報)",
         {**green, "remarks": r_merged}, "remarks-order", True),
        ("綠向 R-TM68 (Remarks 為空 → 不判)",
         {**green, "remarks": ""}, "remarks-order", False),
        ("紅向 R-TM66 (input_test_data 與 procedure 逐字重複 → 報)",
         {**green, **itd_dup}, "data-placement", True),
        ("綠向 R-TM66 (獨立資料集，procedure 未提 → 不報)",
         {**green, **itd_ok}, "data-placement", False),
        ("綠向 R-TM66 (input_test_data 為 NA → 不判)",
         {**green, **itd_dup, "input_test_data": "NA"}, "data-placement", False),
    ]
    for name, tc, gate, want in y_cases:
        got = [m for g, m in lint_tc(tc, auth, "Y") if g == gate]
        ok = bool(got) is want
        bad += not ok
        print(f"{'PASS' if ok else '**FAIL**'} {name}: {(got[0][:56] if got else '未報')}")

    # ── R-TM69 兩項 + R-TM70（16 T2）──────────────────────
    gap_leaf3 = gap_leaf[-3:]
    gap_mid = sorted(o for o, v in (auth["ee_arch"].get(gap_leaf3) or {}).items()
                     if not v["is_atl_hi"] and "不在 docx" not in v["ee"])
    gap_rem = "\n".join([spec_gap_ph] + [atl_hi_placeholder(o) for o in gap_mid])
    b011 = "SWE-RA-TIME&DATE-011"
    # ── 構造複驗（R-TM67）──
    print(f"   [構造複驗] gap_leaf={gap_leaf} 在 SPEC_GAP_LEAVES = "
          f"{gap_leaf in SPEC_GAP_LEAVES}；其 Atl-Mid 物件 {gap_mid}")
    print(f"   [構造複驗] 只帶 DR-11 之 Remarks 確實不含 `PENDING: DR-{SPEC_GAP_DR}` = "
          f"{'PENDING: DR-%d' % SPEC_GAP_DR not in atl_hi_placeholder(gap_mid[0] if gap_mid else '9999999')}")
    _v = "$DateTmHour$"
    print(f"   [構造複驗] {_v} 屬 011 之 not_ours = "
          f"{_v in BOUNDARY_SIGNALS[b011]['not_ours']}"
          f"；置於上半 verbatim 時 R-TM70 應豁免、置於下半時應報")
    z_cases = [
        ("紅向 R-TM69(1) (A-TM13 leaf 只帶 DR-11 → 報 spec-gap)",
         {**green, "req_id": gap_leaf,
          "remarks": atl_hi_placeholder(gap_mid[0]) if gap_mid else "PENDING: DR-11 x"},
         "spec-gap", True),
        ("綠向 R-TM69(1) (帶該 leaf 之 DR-5 → 不報)",
         {**green, "req_id": gap_leaf, "remarks": gap_rem}, "spec-gap", False),
        ("紅向 R-TM69(2) (缺一個應有之 DR-11 → 報)",
         {**green, "req_id": gap_leaf, "remarks": spec_gap_ph},
         "placeholder-completeness", bool(gap_mid)),
        ("紅向 R-TM69(2) (多出一個不屬本片之 DR-11 → 報)",
         {**green, "req_id": gap_leaf,
          "remarks": gap_rem + "\n" + atl_hi_placeholder("4814064")},
         "placeholder-completeness", True),
        ("綠向 R-TM69(2) (應有與實有相等 → 不報)",
         {**green, "req_id": gap_leaf, "remarks": gap_rem},
         "placeholder-completeness", False),
        ("綠向 R-TM70 (鄰片訊號只出現於 verbatim 上半 → 豁免)",
         {**green, "req_id": b011,
          "test_item": f"Requirement text mentioning {_v} verbatim\n(purpose)"},
         "boundary", False),
        ("紅向 R-TM70 (同一訊號出現於下半括號 → 仍報)",
         {**green, "req_id": b011,
          "test_item": f"Requirement text\n(confirm {_v} is transmitted)"},
         "boundary", True),
    ]
    for name, tc, gate, want in z_cases:
        got = [m for g, m in lint_tc(tc, auth, "Z") if g == gate]
        ok = bool(got) is want
        bad += not ok
        print(f"{'PASS' if ok else '**FAIL**'} {name}: {(got[0][:56] if got else '未報')}")

    # ── R-TM55 之綠向與負控（07 T3）─────────────────────────
    # 空 `owns` 之片，其 not_ours 以外之訊號不得誤報；且 022 須維持不判。
    b_cases = [
        ("綠向 3 (018 提 $DateTmHour$ —— 非其 not_ours)",
         {**green, "req_id": "SWE-RA-TIME&DATE-018",
          "expected_result": "IPC restores $DateTmHour$ to 00 after reset"}, False),
        ("綠向 4 (017 提 $DateTmMinute$ —— 非其 not_ours)",
         {**green, "req_id": "SWE-RA-TIME&DATE-017",
          "expected_result": "IPC shows $DateTmMinute$ on the date channel"}, False),
        ("負控   (022 提 $GPSDateTm —— R-TM55 駁回列入，須仍不叫)",
         {**green, "req_id": "SWE-RA-TIME&DATE-022",
          "expected_result": "HU sets SNA into $GPSDateTmHour$"}, False),
    ]
    for name, tc, want in b_cases:
        got = [m for g, m in lint_tc(tc, auth, "B4") if g == "boundary"]
        ok = bool(got) is want
        bad += not ok
        print(f"{'PASS' if ok else '**FAIL**'} {name}: "
              f"{'未誤報 boundary' if not got else got[0][:70]}")

    # ── B1 D5 守衛（工作簿層，非逐 TC）—— R-TM58 改判後之紅綠 ─────
    # 綠向：母本現況 D5 為空 → 不報（R-TM58：空為應然狀態）
    d5 = lint_d5_scope(auth)
    ok = not d5
    bad += not ok
    print(f"{'PASS' if ok else '**FAIL**'} B1 綠向 (D5 為空 → 不報): "
          f"{'未誤報' if ok else d5}")

    # 紅向：D5 填他 feature 之 037（A-H26 形態）→ 須報。
    # 以**新建之臨時工作簿**構造，不對任何既有工作簿存回。
    d5_reds = [
        ("A-H26 形態：他 feature 之 037 檔名",
         "FM-WI-FSM-037-A03-N1L-SWE1-AppDrawer-HMI-V0.1 STLA 報告",
         "spec-scope-pending"),
        ("以 feature 名組值（R-TM9-A2 所禁）",
         "Time and Date", "d5-scope"),
        ("填 NA（canon §8.4.3：NA 僅限確認不適用）",
         "NA", "d5-scope"),
        ("殘留之 PENDING 佔位（R-TM58 撤回）",
         "PENDING: DR-2 037 正式報告檔名", "spec-scope-pending"),
    ]
    with tempfile.TemporaryDirectory() as td:
        for name, val, want in d5_reds:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = auth["sheet"]
            ws["C5"] = "範圍 Scope："
            ws["D5"] = val
            fp = Path(td) / "d5.xlsx"
            wb.save(fp)          # 新建之臨時檔，非存回母本（R-TM49 不適用）
            got = lint_d5_scope({**auth, "workbook": fp})
            hit = any(g == want for g, _ in got)
            bad += not hit
            print(f"{'PASS' if hit else '**FAIL**'} B1 紅向 ({name}) → "
                  f"{want}: {got[0][1][:56] if got else '未叫 —— 閘失效'}")

    total = (len(reds) + 3 + len(b_cases) + 1 + len(d5_reds)
             + len(a_cases) + len(x_cases) + len(y_cases)
             + len(z_cases))
    print(f"\n自驗：{total - bad} / {total}")
    return 1 if bad else 0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--feature-dir", required=True)
    ap.add_argument("--self-test", action="store_true")
    a = ap.parse_args()
    fd = Path(a.feature_dir)
    try:
        auth = load_authorities(fd)
    except LintError as e:
        print(f"LintError: {e}", file=sys.stderr)
        return 2
    if a.self_test:
        return self_test(auth)
    findings: list[tuple[str, str]] = []
    # B1 為**工作簿層**檢查，與是否已生成 TC 無關 —— 故置於 early return
    # 之前。原置於 TC 迴圈之後，`generated/` 為空時提前 return 即被跳過，
    # 而 B1 之設計意圖正是「使 D5 之狀態每次 lint 都現形」（L1）。
    findings += lint_d5_scope(auth)
    gen = sorted((fd / "generated").glob("*.json"))
    if not gen:
        print("generated/ 無 json —— 尚未生成 TC（工作簿層閘門仍已執行）")
        for g, m in findings:
            print(f"[{g}] {m}")
        return 1 if any(g != "spec-scope-pending" for g, _ in findings) else 0
    for p in gen:
        findings += lint_file(p, auth)
    for g, m in findings:
        print(f"[{g}] {m}")
    print(f"\n檔 {len(gen)}；發現 {len(findings)} 項")
    return 1 if findings else 0


if __name__ == "__main__":
    raise SystemExit(main())
