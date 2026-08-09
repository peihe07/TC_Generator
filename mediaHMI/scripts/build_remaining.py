#!/usr/bin/env python3
"""Build remaining_leaves.json + sibling_map.json.

Diff 037 A03 Analysis Report (leaf Functional Requirements) against the
completed region of FW036 (rows 10-332). Everything not covered there is
the remaining work for TC generation.

Usage:
    python build_remaining.py --a03 <037.xlsx> --fw036 <036.xlsx> --out data/
"""
import argparse
import json
import re
from collections import OrderedDict
from pathlib import Path

import openpyxl

DONE_REGION_LAST_ROW = 332  # rows 10..332 = human-authored, compliant TCs
TC_SHEET = "Test Case Specification 測試用例規範"
A03_SHEET = "Analysis Report"


def load_a03(path: str) -> "OrderedDict[str, dict]":
    """Return all leaf FRs from the 037 analysis report, in document order."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[A03_SHEET]
    leaves = OrderedDict()
    for r in ws.iter_rows(min_row=8, values_only=True):
        rid, cat = r[0], str(r[6] or "")
        if not rid or not str(rid).startswith("SWE1"):
            continue
        if cat != "Functional Requirement":
            continue
        src = str(r[2] or "")
        m = re.search(r"_(\d{1,2}(?:\.\d+)*)$", src)
        leaves[str(rid)] = {
            "req_id": str(rid),
            "parent": str(rid).rsplit("-", 1)[0],
            "title": str(r[3] or "").strip(),
            "description": str(r[4] or "").strip(),
            "hmi_source_id": src,
            "section": m.group(1) if m else "",
            "source_req_id": str(r[1] or ""),
            "frop": str(r[7] or "") if len(r) > 7 else "",
        }
    return leaves


def load_fw036_done(path: str) -> set:
    """Req IDs already covered in the compliant (done) region of FW036."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[TC_SHEET]
    done = set()
    rownum = 9
    for r in ws.iter_rows(min_row=10, values_only=True):
        rownum += 1
        if rownum > DONE_REGION_LAST_ROW:
            break
        if r[3]:
            done.add(str(r[3]))
    return done


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--a03", required=True)
    ap.add_argument("--fw036", required=True)
    ap.add_argument("--out", default="data")
    args = ap.parse_args()

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    leaves = load_a03(args.a03)
    done = load_fw036_done(args.fw036)

    remaining = OrderedDict(
        (rid, info) for rid, info in leaves.items() if rid not in done
    )

    # sibling_map: parent -> ordered list of ALL its leaf sub-ids (done ones
    # included, so the generator can see full sibling context per docs.md §4.6)
    sibling_map = OrderedDict()
    for rid, info in leaves.items():
        sibling_map.setdefault(info["parent"], []).append(
            {"req_id": rid, "title": info["title"], "remaining": rid in remaining}
        )
    # keep only parents that still have remaining work
    sibling_map = OrderedDict(
        (p, subs) for p, subs in sibling_map.items()
        if any(s["remaining"] for s in subs)
    )

    (out / "remaining_leaves.json").write_text(
        json.dumps(list(remaining.values()), ensure_ascii=False, indent=2)
    )
    (out / "sibling_map.json").write_text(
        json.dumps(sibling_map, ensure_ascii=False, indent=2)
    )

    print(f"A03 leaves: {len(leaves)}, done: {len(done & set(leaves))}, "
          f"remaining: {len(remaining)} across {len(sibling_map)} parents")
    print(f"-> {out/'remaining_leaves.json'}\n-> {out/'sibling_map.json'}")


if __name__ == "__main__":
    main()
