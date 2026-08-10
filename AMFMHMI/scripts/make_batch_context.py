#!/usr/bin/env python3
"""Step 2 (AMFM) — assemble the generation context for one batch.

Gathers into one JSON everything the generator needs and nothing that would
let it invent a value:

  - the batch's leaves: 037 title (which carries the requirement text
    verbatim, with its STLA id tail) and description
  - the CFTS section each leaf brackets into, with that section's full text
    read out of the docx — spec_mode D has no export to read, so the document
    is the only source
  - sibling leaves sharing a section, so the generator can see the split axis
    it is working inside (§8.3) instead of re-deriving it per leaf
  - style-only exemplars from the frozen legacy region
  - a `column_conventions` block stating the R7 rulings as data

That last block exists because the workbook in front of the generator
disagrees with the rules it must follow: the legacy rows write Test Group
`Radio` and a band-based Test Set, and R7-Q1/Q2 replaced both for new rows.
An exemplar is persuasive; a convention has to be stated louder than the
example contradicting it.

Leaves allocated to an external CFTS document (A-AM06/A-AM07) carry no
section text — the file is not in `inputs/`. They are emitted with the reason
and the anomaly id, never silently short of context.

Usage:
    python AMFMHMI/scripts/make_batch_context.py --feature-dir AMFMHMI --list
    python AMFMHMI/scripts/make_batch_context.py --feature-dir AMFMHMI \\
        --batch "Tuner Availability"
    python AMFMHMI/scripts/make_batch_context.py --feature-dir AMFMHMI --pilot
"""
from __future__ import annotations

import argparse
import difflib
import json
import re
import sys
from pathlib import Path

import docx
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent / "scripts"))
from feature_config import load_feature_config, resolve_path  # noqa: E402

BATCH_ROW = re.compile(r"^\|\s*([^|]+?)\s*\|\s*([^|]*?)\s*\|\s*(\d+)\s*\|"
                       r"\s*([^|]+?)\s*\|\s*([^|]*?)\s*\|")
ANOMALY = re.compile(r"A-AM\d\d")
# A batch cites a spec worksheet as `[[table:NAME]]` in its context note, so
# membership and its evidence stay in the one batch table.
TABLE_CITE = re.compile(r"\[\[table:([A-Za-z0-9_.-]+)\]\]")
HEADING_NUM = re.compile(r"^(\d+(?:\.\d+)*)\s")
REQ_PREFIX = "SWE-RA-RAD-"
# Above this the 037 title is a verbatim quote of the CFTS clause (the id tail
# and whitespace account for the gap); below it the two texts genuinely differ.
WORDING_VERBATIM = 0.95


def _fold(s: str) -> str:
    """Compare wording, not typography or the trailing STLA id tag."""
    s = re.sub(r"[({（]\s*\d{7}\s*[)}）]", " ", str(s or ""))
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9$ ]", " ", s.lower())).strip()


class ContextError(RuntimeError):
    pass


def parse_leaf_selector(spec: str) -> list[str]:
    out = []
    for part in str(spec).split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = (int(x) for x in part.split("-", 1))
            out.extend(f"{REQ_PREFIX}{n:03d}" for n in range(a, b + 1))
        else:
            out.append(f"{REQ_PREFIX}{int(part):03d}")
    return out


def load_batches(md: Path) -> dict:
    if not md.exists():
        raise ContextError(f"missing {md}")
    batches = {}
    for line in md.read_text(encoding="utf-8").splitlines():
        m = BATCH_ROW.match(line)
        if not m or m.group(3) == "n":
            continue
        name, secs, count, ids, note = m.groups()
        req_ids = parse_leaf_selector(ids)
        if len(req_ids) != int(count):
            raise ContextError(
                f"{name}: table says {count} leaves but lists {len(req_ids)}")
        batches[name] = {
            "test_set": name, "req_ids": req_ids,
            "declared_sections": [s.strip() for s in re.split(r"[,;]", secs)
                                  if s.strip()],
            "context_note": note.strip(),
            "anomalies": sorted(set(ANOMALY.findall(note))),
            "spec_tables": sorted(set(TABLE_CITE.findall(note))),
        }
    if not batches:
        raise ContextError(f"no batch rows parsed from {md}")
    return batches


def section_texts(docx_path: Path) -> dict[str, dict]:
    """{section number: {title, level, text, children, subtree_chars}}.

    `text` is the section's OWN body — the paragraphs between its heading and
    the next heading of any level. That is what a leaf bracketing into this
    section is governed by: anything under a subsection belongs to a leaf that
    brackets into the subsection instead.

    The distinction is not cosmetic. §1.3 "HU Analog Tuner" runs to §1.4, so a
    same-or-higher-level span hands 82k characters — the entire analog tuner
    chapter — to the two leaves whose requirement is the section's own
    two-paragraph AM-presence gate. `children` and `subtree_chars` are kept so
    a batch that genuinely needs the subtree can ask for it knowingly.
    """
    doc = docx.Document(docx_path)
    paras = doc.paragraphs
    heads = []
    for i, p in enumerate(paras):
        if not p.style.name.startswith("Heading"):
            continue
        m = HEADING_NUM.match(p.text.strip())
        if not m:
            continue
        try:
            level = int(p.style.name.split()[-1])
        except ValueError:
            level = m.group(1).count(".") + 1
        heads.append((i, m.group(1), p.text.strip(), level))

    out = {}
    for k, (i, num, title, level) in enumerate(heads):
        own_end = heads[k + 1][0] if k + 1 < len(heads) else len(paras)
        subtree_end = len(paras)
        children = []
        for j, nnum, _, nlevel in heads[k + 1:]:
            if nlevel <= level:
                subtree_end = j
                break
            if nlevel == level + 1:
                children.append(nnum)
        body = [p.text.strip() for p in paras[i + 1:own_end] if p.text.strip()]
        subtree = [p.text.strip() for p in paras[i + 1:subtree_end] if p.text.strip()]
        out[num] = {"title": title, "level": level, "text": "\n".join(body),
                    "children": children,
                    "subtree_chars": len("\n".join(subtree))}
    return out


def load_spec_tables(cfg, root: Path, wanted: list[str]) -> dict:
    """Inject the spec worksheets a batch cites, as rows rather than prose.

    Mode D's spec is not only the CFTS docx: several clauses delegate their
    detail to a companion workbook ("See 'CIP_Radio_Tables*', 'SEEK
    Cancel_Stop Transitions' worksheet"). Those tables ARE the requirement for
    a State Transition TC — a seek batch generated without the cancel/stop
    matrix has no source for which event cancels and which stops, and would
    have to guess.

    Declared per batch in feature.yaml `spec_tables`; a batch that names a
    sheet which is not in the file fails loud rather than generating without
    it. Merged header rows are forward-filled, because the event columns are
    grouped under a banner row and a lone column label loses its group.
    """
    out = {}
    for name in wanted:
        spec = (cfg.get("spec_tables") or {}).get(name)
        if spec is None:
            raise ContextError(
                f"batch cites spec table {name!r} but feature.yaml "
                "spec_tables has no such entry")
        path = resolve_path(cfg, spec["source"])
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = spec["sheet"]
        if sheet not in wb.sheetnames:
            wb.close()
            raise ContextError(
                f"{path.name} has no worksheet {sheet!r}; present: "
                f"{wb.sheetnames}")
        rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        wb.close()
        # Some worksheets stack several tables (the market configuration sheet
        # opens with a per-market feature grid, then the tuner configuration
        # table below it). `first_row` points at the wanted table's own header
        # instead of forcing a second sheet or a hand-copied extract.
        first = spec.get("first_row", 0)
        if first >= len(rows):
            raise ContextError(
                f"{path.name}/{sheet}: first_row {first} is past the last row "
                f"({len(rows)})")
        rows = rows[first:]
        hdr_rows = spec.get("header_rows", 1)
        banner, header = [], []
        for i in range(hdr_rows):
            cells = [str(v).strip().replace("\n", " ") if v else ""
                     for v in rows[i]]
            if i < hdr_rows - 1:
                run = ""
                filled = []
                for c in cells:
                    run = c or run
                    filled.append(run)
                banner = filled
            else:
                header = cells
        records = []
        for r in rows[hdr_rows:]:
            label = str(r[0]).strip() if r[0] else ""
            if not label:
                continue
            cells = {}
            for j, v in enumerate(r):
                if j == 0 or v in (None, ""):
                    continue
                key = header[j] if j < len(header) else f"col{j}"
                if banner and j < len(banner) and banner[j]:
                    key = f"{banner[j]} / {key}"
                cells[key] = str(v).replace("\n", " ").strip()
            records.append({"state": label, "events": cells})
        out[name] = {"source": path.name, "sheet": sheet,
                     "row_label": header[0] if header else "",
                     "rows": records}
    return out


def leaf_citations(citations: dict, req_id: str) -> list[dict]:
    """Cross-document citations made by this leaf's own clause.

    A leaf whose clause says `HU shall play the rejection tone {See CFTS019-718}`
    borrows an outcome it does not define. R11 rules this CITE-FORM, not
    absorption: the outcome is asserted where the citing clause puts it, but
    anchored to the token, and the cited document's own rule surface stays out
    of scope. The instruction travels with the leaf because the two failure
    modes are opposite — silently dropping the outcome under-covers the citing
    clause, and silently adopting it claims verification of another delivery's
    requirement.
    """
    out = []
    for token, e in sorted(citations.items()):
        if req_id not in e.get("req_ids", []):
            continue
        item = {"token": token, "doc": e["doc"], "status": e["status"],
                "cited_in": e["citing_clauses"][0]["context"],
                "handling": "cite-form (R11) — reference, do not absorb",
                "instruction": (
                    f"Cite {token} verbatim as a second citation in "
                    "specification_reference, after this leaf's own clause. "
                    "The ER may assert the borrowed outcome, anchored to the "
                    f"citation — e.g. 'the key press rejection tone is played, "
                    f"as defined by {token}'. Do NOT test that document's own "
                    f"rule surface (which conditions qualify, the specification "
                    f"of the behaviour itself): that is {e['doc']}'s delivery. "
                    "Verify only that the outcome occurs in the citing "
                    "clause's scenario."),
                }
        if e.get("resolved_text"):
            item |= {"resolved_clause": e["resolved_clause"],
                     "resolved_section": e.get("resolved_section"),
                     "referenced_text": e["resolved_text"],
                     "ruling": e.get("ruling")}
        else:
            item["candidates"] = e.get("candidates", [])
        out.append(item)
    return out


def build(cfg, batch: dict, stla_map: dict, sections: dict,
          exemplars: dict, spec_docs: dict, spec_tables: dict | None = None,
          citations: dict | None = None) -> dict:
    primary = spec_docs.get("primary")
    leaves, needed, blocked = [], set(), []
    for rid in batch["req_ids"]:
        e = stla_map.get(rid)
        if e is None:
            raise ContextError(f"{rid} is in the batch table but not the "
                               "STLA map — rerun build_stla_map.py")
        entry = {"req_id": rid, "stla_id": e["stla_id"], "doc": e["doc"],
                 "section": e["section"], "section_title": e["section_title"],
                 "requirement_text": e["title"],
                 "analysis_note": e["description"],
                 "source_components": e["source_components"],
                 "spec_paragraph": e.get("spec_paragraph"),
                 "spec_paragraph_metadata": e.get("spec_paragraph_metadata"),
                 "spec_reference": cfg["spec_reference_template"].format(
                     doc=e["doc"], stla_id=e["stla_id"])}
        # The 037 title is supposed to quote the CFTS clause verbatim. Where it
        # does not, §8.6 says the source spec wins — so the divergence has to
        # reach the generator rather than be averaged away by having both texts
        # side by side with no comment.
        if e.get("spec_paragraph"):
            ratio = difflib.SequenceMatcher(
                None, _fold(e["title"]), _fold(e["spec_paragraph"])).ratio()
            entry["wording_agreement"] = round(ratio, 3)
            if ratio < WORDING_VERBATIM:
                entry["wording_note"] = (
                    "037 title and CFTS clause differ; §8.6 — the source spec "
                    "is authority for wording, the 037 title for scope")
        refs = leaf_citations(citations or {}, rid)
        if refs:
            entry["cross_references"] = refs
        if e["section"]:
            needed.add(e["section"])
        else:
            reason = ("spec file not in inputs/ — external CFTS document"
                      if e["doc"] else "no document allocated")
            entry["section_text_absent"] = reason
            blocked.append(rid)
        leaves.append(entry)

    # Siblings: every other leaf bracketing into a section this batch uses.
    # They are the split axis made visible, not extra targets.
    in_batch = set(batch["req_ids"])
    siblings = [{"req_id": rid, "section": e["section"],
                 "requirement_text": e["title"]}
                for rid, e in stla_map.items()
                if e["section"] in needed and rid not in in_batch]

    spec = {}
    missing_sections = []
    for num in sorted(needed):
        if num in sections:
            spec[num] = sections[num]
        else:
            missing_sections.append(num)
    if missing_sections:
        raise ContextError(
            f"sections {missing_sections} are cited by the bracket map but "
            f"absent from the {primary} document — the map and the spec file "
            "have diverged")

    return {
        "feature": cfg["feature"],
        "batch": batch["test_set"],
        "spec_mode": cfg["spec_mode"],
        "leaves": leaves,
        "spec_sections": spec,
        "siblings": siblings,
        "spec_tables": spec_tables or {},
        "exemplars": exemplars.get("exemplars", []),
        "exemplar_basis": exemplars.get("basis"),
        "column_conventions": {
            "test_group": {"value": cfg["test_group"],
                           "note": "R7-Q1 — new rows write this; the legacy "
                                   "region's 'Radio' is frozen, not a model"},
            "test_set": {"value": batch["test_set"],
                         "note": "R7-Q2 — capability scheme from framework "
                                 "Part III; the legacy FM/AM/USB band scheme "
                                 "is NOT adopted"},
            "author": {"value": cfg["write_back"]["author_value"],
                       "note": "R3 — rows generated here; rows quoted wholly "
                               "from another author keep that author"},
            "tc_ref_id": {"value": cfg["write_back"]["tc_ref_id_value"]},
            "spec_reference": {
                "template": cfg["spec_reference_template"],
                "note": "R7-Q3 — {doc} comes from the STLA map, never guessed"},
        },
        "context_note": batch["context_note"],
        "anomalies": batch["anomalies"],
        "blocked_section_text": blocked,
    }


def run(args) -> int:
    root = Path(args.feature_dir)
    cfg = load_feature_config(args.feature_dir)
    batches = load_batches(root / args.batches_md)

    if args.list:
        for name, b in batches.items():
            print(f"{len(b['req_ids']):3}  {name:22} "
                  f"{', '.join(b['declared_sections'])}")
        return 0

    wanted = list(args.batch or [])
    if args.pilot:
        wanted = [n for n in batches if n in ("Tuner Availability", "Tune")]
    if not wanted:
        raise ContextError("give --batch NAME (repeatable), --pilot, or --list")

    stla_path = root / "data" / "stla_to_cfts.json"
    if not stla_path.exists():
        raise ContextError("data/stla_to_cfts.json missing — run "
                           "build_stla_map.py first")
    stla_map = json.loads(stla_path.read_text(encoding="utf-8"))
    ex_path = root / "data" / "exemplars.json"
    exemplars = (json.loads(ex_path.read_text(encoding="utf-8"))
                 if ex_path.exists() else {})
    if not exemplars:
        print("WARNING: no data/exemplars.json — generating without style "
              "anchors", file=sys.stderr)
    sections = section_texts(resolve_path(cfg, "cfts_doc"))
    spec_docs = cfg.get("spec_docs") or {}
    cite_path = root / "data" / "cross_doc_citations.json"
    citations = (json.loads(cite_path.read_text(encoding="utf-8"))
                 if cite_path.exists() else {})
    if not citations:
        print("WARNING: no data/cross_doc_citations.json — leaves citing "
              "another CFTS will reach the generator unflagged; run "
              "build_stla_map.py", file=sys.stderr)

    out_dir = root / "batches"
    out_dir.mkdir(exist_ok=True)
    for name in wanted:
        if name not in batches:
            raise ContextError(f"unknown batch {name!r}; --list to see them")
        tables = load_spec_tables(cfg, root, batches[name]["spec_tables"])
        ctx = build(cfg, batches[name], stla_map, sections, exemplars,
                    spec_docs, tables, citations)
        slug = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
        path = out_dir / f"{slug}.json"
        path.write_text(json.dumps(ctx, ensure_ascii=False, indent=1),
                        encoding="utf-8")
        chars = sum(len(s["text"]) for s in ctx["spec_sections"].values())
        tbl = ctx.get("spec_tables") or {}
        diverged = [l["req_id"] for l in ctx["leaves"] if "wording_note" in l]
        cross = [f"{l['req_id']}:{r['token']}" for l in ctx["leaves"]
                 for r in l.get("cross_references", [])]
        if cross:
            print(f"  cross-document citations (R11 cite-form, multi-cite + "
                  f"anchored ER): {cross}", file=sys.stderr)
        if diverged:
            print(f"  037 title diverges from the CFTS clause: {diverged} "
                  "(§8.6 — spec wins on wording)", file=sys.stderr)
        print(f"{name:22} {len(ctx['leaves'])} leaves, "
              f"{len(ctx['spec_sections'])} sections ({chars} chars), "
              f"{len(ctx['siblings'])} siblings, "
              f"{len(ctx['exemplars'])} exemplars"
              + (", tables " + "+".join(
                  f"{k}({len(v['rows'])} rows)" for k, v in tbl.items())
                 if tbl else "")
              + (f", {len(ctx['blocked_section_text'])} without section text"
                 if ctx["blocked_section_text"] else "")
              + f" -> {path}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--feature-dir", default=".")
    ap.add_argument("--batches-md", default="docs/batches-amfm.md")
    ap.add_argument("--batch", action="append", help="batch name (repeatable)")
    ap.add_argument("--pilot", action="store_true",
                    help="the pilot pair proposed in batches-amfm.md")
    ap.add_argument("--list", action="store_true")
    args = ap.parse_args()
    try:
        return run(args)
    except ContextError as exc:
        print(f"ABORTED: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
