#!/usr/bin/env python3
"""Step 3 (AMFM) — hard gate over generated/*.json.

Authorities are read, never hard-coded: `feature.yaml` supplies the Test
Group, the spec_reference template and the author; the design-method
vocabulary comes from the workbook's own `下拉選單` sheet; the legal
spec_reference values come from `data/stla_to_cfts.json`.

Two gates are AMFM rulings rather than generic rules, and both exist because
the failure they catch is invisible in a diff:

- **R10-2 (`absorption-cite`)** — absorbing an unallocated CFTS clause into
  the leaf that it elaborates is allowed, on condition that the absorbed
  clause is cited. An `[A-AM10]` marker without a second citation is a TC
  verifying behaviour it does not trace to; a multi-cite without the marker
  is a citation nobody can explain. The gate makes the two sides equivalent.
- **R10-4 (`remarks-internal`)** — Remarks ships to the customer. Internal
  ruling ids (`R9`, `A-AM08`) are how we talk to ourselves; an external
  reader sees an unexplained code in a controlled document.

Exit 0 = clean, 1 = at least one finding, 2 = bad invocation.

Usage:
    python AMFMHMI/scripts/lint_tcs.py --feature-dir AMFMHMI
    python AMFMHMI/scripts/lint_tcs.py --feature-dir AMFMHMI --json-report lint_report.json
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent / "scripts"))
from feature_config import load_feature_config, resolve_path  # noqa: E402

PRIORITIES = ("P0", "P1", "P2", "P3")
MODALS = re.compile(r"\b(shall|will|should|would|must)\b", re.I)
# Internal identifiers: ruling ids and anomaly markers. Deliberately narrow —
# "R1L" (the program name) and years must not trip it.
INTERNAL_ID = re.compile(r"\bA-AM\d{2}\b|\bR\d{1,2}(?:-Q?\d)?\b(?!L)")
ABSORPTION_MARKER = "[A-AM10]"
CITE_SEP = "; "
# Source-quoted signal values keep the spec's square brackets (Profile §3.4);
# everything else must use double quotes for UI labels (§11).
SIGNAL_TOKEN = re.compile(r"\$[A-Za-z0-9_]+\$\s*=\s*\[[^\]]*\]")
MULTILINE_FIELDS = ("pre_conditions", "input_test_data", "test_procedure",
                    "expected_result")


def numbered(text: str) -> list[str]:
    return [l for l in str(text or "").split("\n") if l.strip()]


def absorbed_ids(doc: dict) -> set[str]:
    """The clause ids an `[A-AM10]` assumption names as absorbed.

    R10-2a makes absorption conditional on citing what was absorbed, so the
    marker has to say WHICH clause — otherwise the condition is unverifiable
    and the gate degenerates into counting citations.
    """
    out: set[str] = set()
    for a in doc.get("assumptions") or []:
        text = str(a)
        if ABSORPTION_MARKER not in text:
            continue
        out |= set(re.findall(r"CFTS\d{3}-(\d{7})", text))
    return out


def load_design_methods(workbook: Path) -> list[str]:
    wb = openpyxl.load_workbook(workbook, read_only=True)
    if "下拉選單" not in wb.sheetnames:
        raise SystemExit("workbook has no 下拉選單 sheet")
    out = [str(r[0].value).strip() for r in wb["下拉選單"].iter_rows()
           if r and r[0].value]
    wb.close()
    return out


def cited_tokens(citations: dict, req_id: str) -> set[str]:
    """Cross-document tokens this leaf's own clause writes (R11 cite-form).

    Membership is per leaf, not global: `CFTS019-718` is a legitimate citation
    for 014, whose clause writes it, and a fabricated one anywhere else.
    """
    return {t for t, e in (citations or {}).items()
            if req_id in e.get("req_ids", [])}


def lint_tc(tag: str, tc: dict, doc: dict, cfg: dict, methods: list[str],
            stla: dict, citations: dict | None = None) -> list[dict]:
    f = []

    def bad(gate, msg):
        f.append({"tc": tag, "gate": gate, "message": msg})

    proc, er = numbered(tc.get("test_procedure")), numbered(tc.get("expected_result"))
    if len(proc) < 2:
        bad("step-count", f"{len(proc)} numbered steps, minimum 2 (§10.5)")
    if len(proc) != len(er):
        bad("er-alignment", f"{len(proc)} steps vs {len(er)} ER lines (§6)")

    for field in MULTILINE_FIELDS:
        for line in str(tc.get(field) or "").split("\n"):
            if line.strip().endswith((".", "。")):
                bad("trailing-period", f"{field}: {line.strip()[-40:]!r} (§11)")
        stripped = SIGNAL_TOKEN.sub("", str(tc.get(field) or ""))
        if "[" in stripped or "]" in stripped:
            bad("square-bracket",
                f"{field}: bracket outside a source-quoted $SIGNAL$ = [value] (§11)")

    if MODALS.search(str(tc.get("expected_result") or "")):
        bad("er-modal", f"modal verb in expected_result (§6)")
    title = str(tc.get("tc_title") or "")
    if not title:
        bad("keys", "tc_title is empty (§4.3)")
    elif MODALS.search(title):
        bad("title-modal", f"modal verb in tc_title (§4.3)")
    elif len(title.split()) > 14:
        bad("title-length", f"tc_title is {len(title.split())} words, max 14 (§4.3)")

    if tc.get("priority") not in PRIORITIES:
        bad("priority", f"{tc.get('priority')!r} is not one of {PRIORITIES} (§10.2)")
    if tc.get("design_method") not in methods:
        bad("design-method",
            f"{tc.get('design_method')!r} is not one of the 下拉選單 strings (§12)")
    if tc.get("test_group") != cfg["test_group"]:
        bad("test-group",
            f"{tc.get('test_group')!r} != {cfg['test_group']!r} (R7-Q1)")
    if not str(tc.get("test_set") or "").strip():
        bad("test-set", "empty; AMFM fills Test Set (R7-Q2)")

    # ---- spec_reference resolves through the STLA map, every citation
    refs = [r.strip() for r in str(tc.get("specification_reference") or "").split(";")
            if r.strip()]
    if not refs:
        bad("spec-reference", "empty (§10.7)")
    entry = stla.get(tc.get("req_id"))
    if entry is None:
        bad("unknown-req-id", f"{tc.get('req_id')!r} is not a 037 leaf")
    else:
        primary = cfg["spec_reference_template"].format(
            doc=entry["doc"], stla_id=entry["stla_id"])
        if refs and refs[0] != primary:
            bad("spec-reference",
                f"first citation {refs[0]!r} is not the leaf's own clause "
                f"{primary!r} (§10.7 most-specific first)")
    # R11: a cross-document citation is quoted in the CITING document's own
    # short-id scheme (`CFTS019-718`), so it cannot be held to the 7-digit
    # anchor shape. It is allowed only where the spec actually writes it —
    # per leaf, from the citation sweep — which is what stops the short form
    # from becoming a hole in the format check.
    allowed_tokens = cited_tokens(citations, tc.get("req_id"))
    for r in refs:
        if re.fullmatch(r"CFTS\d{3}-\d{7}", r) or r in allowed_tokens:
            continue
        if re.fullmatch(r"CFTS\d{3}-\d{1,6}", r):
            bad("cross-reference",
                f"{r!r} is a cross-document citation, but this leaf's clause "
                f"does not write it (R11; leaf cites {sorted(allowed_tokens)})")
        else:
            bad("spec-reference", f"{r!r} does not match CFTSnnn-nnnnnnn")

    # R11 cite-form: the referenced outcome may be asserted, but only anchored
    # to the citation ("the key press rejection tone is played, as defined by
    # CFTS019-718"). An unanchored assertion reads as this leaf's own verified
    # behaviour, which is exactly the absorption R11 declined.
    er_text = str(tc.get("expected_result") or "")
    for token in sorted(set(refs) & allowed_tokens):
        if token not in er_text:
            bad("cross-reference-anchor",
                f"{token} is cited but no ER line anchors to it (R11 — write "
                f"the borrowed outcome 'as defined by {token}', or drop the "
                "citation)")

    # ---- R10-2a: every absorbed clause the marker names must be cited
    # somewhere under this parent, and a multi-cite must be explained by a
    # marker. Checking the ids the assumption itself names is what makes this
    # exact: a blanket "marked parent => every TC multi-cites" over-fires on a
    # parent where only one TC absorbs, and a blanket count check misses the
    # second absorbed clause entirely.
    # Cite-form cross-references are excluded from the count: R11 rules them
    # NOT absorption — nothing is taken into this leaf's scope, so demanding an
    # absorption marker for them would misfile the very distinction R11 draws.
    absorbing_refs = [r for r in refs if r not in allowed_tokens]
    if len(absorbing_refs) > 1 and not absorbed_ids(doc):
        bad("absorption-cite",
            f"{len(refs)} clauses cited but the parent has no "
            f"{ABSORPTION_MARKER} assumption naming an absorbed clause (R10-2a)")

    # ---- R10-4: Remarks is external-facing
    remarks = str(tc.get("remarks") or "")
    hit = INTERNAL_ID.search(remarks)
    if hit:
        bad("remarks-internal",
            f"Remarks contains the internal identifier {hit.group(0)!r}; "
            "Remarks is written for external readers (R10-4)")
    return f


def run(args) -> int:
    cfg = load_feature_config(args.feature_dir)
    root = Path(args.feature_dir)
    methods = load_design_methods(resolve_path(cfg, "workbook"))
    stla_path = root / "data" / "stla_to_cfts.json"
    if not stla_path.exists():
        raise SystemExit("data/stla_to_cfts.json missing — run build_stla_map.py")
    stla = json.loads(stla_path.read_text(encoding="utf-8"))
    cite_path = root / "data" / "cross_doc_citations.json"
    citations = (json.loads(cite_path.read_text(encoding="utf-8"))
                 if cite_path.exists() else {})

    files = sorted((root / args.generated).glob("*.json"))
    if not files:
        raise SystemExit(f"no generated JSON under {root / args.generated}")

    findings, n_tc, titles = [], 0, {}
    for path in files:
        doc = json.loads(path.read_text(encoding="utf-8"))
        parent = doc.get("parent") or path.stem
        if not str(doc.get("reasoning") or "").strip():
            findings.append({"tc": parent, "gate": "reasoning",
                             "message": "top-level reasoning is empty (§10.4)"})
        tcs = doc.get("tcs") or []
        axis = (doc.get("distinguishing_axis") or {}).get("axis")
        if len(tcs) > 1 and axis in (None, "", "none"):
            findings.append({"tc": parent, "gate": "distinguishing-axis",
                             "message": "multiple TCs but axis is none (§4.6)"})
        cited_here: set[str] = set()
        for i, tc in enumerate(tcs, 1):
            tag = f"{parent}-{i:02d}"
            n_tc += 1
            cited_here |= set(re.findall(
                r"CFTS\d{3}-(\d{7})",
                str(tc.get("specification_reference") or "")))
            if tc.get("req_id") != parent:
                findings.append({"tc": tag, "gate": "keys",
                                 "message": f"req_id {tc.get('req_id')!r} != parent"})
            title = str(tc.get("tc_title") or "")
            if title in titles:
                findings.append({"tc": tag, "gate": "sibling-distinction",
                                 "message": f"tc_title duplicates {titles[title]}"})
            titles[title] = tag
            findings.extend(lint_tc(tag, tc, doc, cfg, methods, stla,
                                    citations))
        for missing in sorted(absorbed_ids(doc) - cited_here):
            findings.append({
                "tc": parent, "gate": "absorption-cite",
                "message": f"assumption absorbs CFTS clause {missing} but no "
                           "TC under this leaf cites it (R10-2a)"})

    print(f"linted {n_tc} TCs from {len(files)} leaf file(s) against "
          f"{len(methods)} design methods, {len(stla)} leaves")
    if findings:
        print(f"\nFAIL — {len(findings)} finding(s):")
        for f in findings:
            print(f"  [{f['gate']}] {f['tc']}: {f['message']}")
    else:
        print("PASS — no findings")
    if args.json_report:
        Path(args.json_report).write_text(json.dumps(
            {"tcs": n_tc, "files": len(files), "findings": findings},
            ensure_ascii=False, indent=1), encoding="utf-8")
        print(f"-> {args.json_report}")
    return 1 if findings else 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--feature-dir", default=".")
    ap.add_argument("--generated", default="generated")
    ap.add_argument("--json-report")
    return run(ap.parse_args())


if __name__ == "__main__":
    sys.exit(main())
