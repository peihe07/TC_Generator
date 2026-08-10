"""Tests for AMFMHMI/scripts/write_back.py (Step 4).

AMFM appends after a frozen prefix rather than rewriting interleaved segments,
so the failure modes differ from Home's and each has a test here:

- the 158 legacy rows must not move, change, or be re-authored. The hash is
  content-based and selected by author value, because after the write the
  appended rows also carry an author.
- column B is a PER-ROW formula (`=IF(ISBLANK($D168),"",ROW()-9)`). Home's
  column B is a constant string; copying that shape here would point every row
  at row 9's requirement id and blank the whole numbering column.
- Test Group / Test Set ARE written on new rows (R7-Q1/Q2). Home's profile
  bans them, so the shared-looking code path is the one most likely to be
  "fixed" back to Home's behaviour by someone reading both.
"""
import importlib.util
import sys
from pathlib import Path

import openpyxl
import pytest

AMFM = Path(__file__).resolve().parent.parent / "AMFMHMI"

_spec = importlib.util.spec_from_file_location(
    "amfm_write_back", AMFM / "scripts" / "write_back.py")
if _spec is None or _spec.loader is None:
    pytest.skip("AMFM write_back not present", allow_module_level=True)
wb_mod = importlib.util.module_from_spec(_spec)
sys.modules["amfm_write_back"] = wb_mod
_spec.loader.exec_module(wb_mod)


CFG = {
    "col": {"req_id": 3, "test_group": 6, "test_set": 7, "test_item": 8,
            "pre_conditions": 9, "input_test_data": 10, "test_procedure": 11,
            "expected_result": 12, "spec_reference": 13, "tc_ref_id": 14,
            "priority": 15, "design_method": 16, "functional_safety": 17,
            "author": 25, "remarks": 32},
    "workbook": {"header_row": 9},
    "write_back": {"author_value": "PeiPYHsu", "tc_ref_id_value": "NEW",
                   "fill_test_group_set": True, "scope_label": "範圍 Scope"},
}


def row(n, req_id, author, cells=None):
    return {"row": n, "req_id": req_id, "author": author,
            "cells": cells or [req_id, author]}


# ------------------------------------------------------------ legacy-region hash

def test_hash_selects_by_author_value_not_by_non_empty():
    """After the append the new rows also carry an author."""
    before = [row(10, "A", "Wilson"), row(11, "B", "Wilson")]
    after = before + [row(12, "SWE-RA-RAD-001", "PeiPYHsu")]
    assert (wb_mod.ordered_content_hash(before, "Wilson")
            == wb_mod.ordered_content_hash(after, "Wilson"))


def test_hash_changes_when_a_legacy_row_changes():
    before = [row(10, "A", "Wilson", ["A", "x"])]
    after = [row(10, "A", "Wilson", ["A", "y"])]
    assert (wb_mod.ordered_content_hash(before, "Wilson")
            != wb_mod.ordered_content_hash(after, "Wilson"))


def test_hash_is_order_sensitive():
    a = [row(10, "A", "Wilson"), row(11, "B", "Wilson")]
    b = [row(10, "B", "Wilson"), row(11, "A", "Wilson")]
    assert (wb_mod.ordered_content_hash(a, "Wilson")
            != wb_mod.ordered_content_hash(b, "Wilson"))


def test_segments_are_legacy_then_regen():
    rows = [row(10, "A", "Wilson"), row(11, "B", "Wilson"),
            row(12, "SWE-RA-RAD-001", "PeiPYHsu")]
    segs = wb_mod.segments_of(rows, "Wilson")
    assert [s["kind"] for s in segs] == ["LEGACY", "REGEN"]
    assert (segs[0]["start"], segs[0]["end"]) == (10, 11)
    assert (segs[1]["start"], segs[1]["end"]) == (12, 12)


# ------------------------------------------------------------------ cell values

def base_tc(**over):
    tc = {"req_id": "SWE-RA-RAD-001", "test_group": "AMFM",
          "test_set": "Tuner Availability", "test_item": "item",
          "pre_conditions": "1. pre", "input_test_data": "NA",
          "test_procedure": "1. a\n2. b", "expected_result": "1. a\n2. b",
          "specification_reference": "CFTS024-4872377", "priority": "P1",
          "design_method": "功能測試 (Functional based ; no specific technique)",
          "remarks": ""}
    tc.update(over)
    return tc


def test_test_group_and_set_are_written_unlike_home():
    """R7-Q1/Q2 — the AMFM workbook convention is filled columns."""
    v = wb_mod.cell_values(base_tc(), CFG)
    assert v[CFG["col"]["test_group"]] == "AMFM"
    assert v[CFG["col"]["test_set"]] == "Tuner Availability"


def test_an_empty_test_set_aborts_when_the_columns_are_filled():
    with pytest.raises(wb_mod.WriteBackError, match="Test Group / Test Set"):
        wb_mod.cell_values(base_tc(test_set=""), CFG)


def test_constant_and_configured_columns_are_written():
    v = wb_mod.cell_values(base_tc(), CFG)
    assert v[CFG["col"]["functional_safety"]] == "NA"
    assert v[CFG["col"]["tc_ref_id"]] == "NEW"
    assert v[CFG["col"]["author"]] == "PeiPYHsu"


def test_empty_strings_become_none_not_empty_cells():
    v = wb_mod.cell_values(base_tc(remarks=""), CFG)
    assert v[CFG["col"]["remarks"]] is None


def test_placeholder_must_leave_priority_blank():
    tc = base_tc(placeholder=True, priority="P1",
                 test_procedure=wb_mod.PLACEHOLDER_BODY,
                 expected_result=wb_mod.PLACEHOLDER_BODY)
    with pytest.raises(wb_mod.WriteBackError, match="priority"):
        wb_mod.cell_values(tc, CFG)


def test_placeholder_body_must_be_the_fixed_string():
    tc = base_tc(placeholder=True, priority="", design_method="",
                 test_procedure="something else",
                 expected_result=wb_mod.PLACEHOLDER_BODY)
    with pytest.raises(wb_mod.WriteBackError, match="test_procedure"):
        wb_mod.cell_values(tc, CFG)


# --------------------------------------------------------------- row numbering

def sheet_with(rows: int, first: int = 10):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(rows):
        ws.cell(first + i, CFG["col"]["req_id"] + 1).value = f"REQ-{i}"
    return ws


def test_row_number_formula_references_its_own_row():
    """A constant `=ROW()-9` would leave every row testing $D9 for blankness."""
    ws = sheet_with(3)
    wb_mod.reemit_row_numbers(ws, CFG)
    assert ws.cell(10, 2).value == '=IF(ISBLANK($D10),"",ROW()-9)'
    assert ws.cell(12, 2).value == '=IF(ISBLANK($D12),"",ROW()-9)'


def test_rows_without_a_requirement_id_get_no_number():
    ws = sheet_with(2)
    ws.cell(12, CFG["col"]["req_id"] + 1).value = None
    n = wb_mod.reemit_row_numbers(ws, CFG)
    assert n == 2
    assert ws.cell(12, 2).value is None


# ------------------------------------------------------------------ appending

def test_template_rows_are_reused_before_new_ones_are_inserted():
    """The template carries formula-only rows below the data; consuming them
    first keeps the inserted count honest in the dry-run arithmetic."""
    ws = sheet_with(2)                      # data at rows 10-11
    for r in range(12, 15):                 # three spare template rows
        ws.cell(r, 2).value = "=formula"
    plan = wb_mod.append_rows(ws, CFG, 12, [base_tc(req_id=f"R{i}")
                                            for i in range(5)])
    assert (plan["first_row"], plan["rows"]) == (12, 5)
    assert (plan["template_rows_used"], plan["inserted"]) == (3, 2)


def test_appended_rows_land_in_document_order():
    ws = sheet_with(1)
    tcs = [base_tc(req_id="R0"), base_tc(req_id="R1")]
    wb_mod.append_rows(ws, CFG, 11, tcs)
    col = CFG["col"]["req_id"] + 1
    assert [ws.cell(11, col).value, ws.cell(12, col).value] == ["R0", "R1"]


# --------------------------------------------------------------- Scope field

def _header_book(label="範圍 Scope", label_at="C5", value="wrong-report"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws[label_at] = label
    ws.cell(int(label_at[1:]), openpyxl.utils.column_index_from_string(
        label_at[0]) + 1).value = value
    return ws


def test_scope_cell_is_located_by_its_label_not_a_coordinate():
    ws = _header_book(label_at="C5")
    assert wb_mod.find_scope_cell(ws, CFG) == (5, 4)


def test_scope_cell_follows_the_label_when_the_layout_moves():
    ws = _header_book(label_at="B7")
    assert wb_mod.find_scope_cell(ws, CFG) == (7, 3)


def test_missing_scope_label_aborts():
    ws = openpyxl.Workbook().active
    with pytest.raises(wb_mod.WriteBackError, match="範圍 Scope"):
        wb_mod.find_scope_cell(ws, CFG)


def test_scope_is_rewritten_and_reports_the_change():
    ws = _header_book()
    got = wb_mod.fix_scope(ws, CFG, "the-037-report")
    assert got["changed"] is True
    assert got["before"] == "wrong-report"
    assert ws["D5"].value == "the-037-report"


def test_fixing_an_already_correct_scope_is_a_no_op():
    ws = _header_book(value="the-037-report")
    assert wb_mod.fix_scope(ws, CFG, "the-037-report")["changed"] is False


# ------------------------------------------------------------------ integration

def test_the_real_dry_run_holds_every_invariant(tmp_path, capsys):
    """Runs the actual dry run against the customer workbook when present."""
    if not (AMFM / "data" / "stla_to_cfts.json").exists():
        pytest.skip("AMFM data/ not built")
    if not list((AMFM / "generated").glob("*.json")):
        pytest.skip("AMFM generated/ empty")
    import argparse
    args = argparse.Namespace(
        write=False, init_baseline=False, feature_dir=str(AMFM), data="data",
        generated="generated", workbook=None, out=None, date=None)
    try:
        rc = wb_mod.run(args)
    except SystemExit:
        pytest.skip("customer workbook not available")
    out = capsys.readouterr().out
    assert rc == 0
    assert "unchanged, order unchanged" in out
    assert "exact match" in out
    assert "DRY RUN" in out


# ------------------------------------------------------------------- TC IDs

TC_ID_CFG = dict(CFG)
TC_ID_CFG["col"] = dict(CFG["col"], tc_id=5)
TC_ID_CFG["write_back"] = dict(CFG["write_back"],
                               tc_id_format="newR1L-AMFM-{n:03d}")


def test_tc_ids_are_a_new_series_not_a_continuation(tmp_path):
    """R13 — the frozen newR1L-Radio-158 series is NOT continued, so removing
    the legacy region later cannot leave a hole in this one."""
    ids = wb_mod.assign_tc_ids(TC_ID_CFG, [base_tc()] * 3,
                               {"newR1L-Radio-001", "newR1L-Radio-158"})
    assert ids == ["newR1L-AMFM-001", "newR1L-AMFM-002", "newR1L-AMFM-003"]


def test_a_tc_id_collision_aborts():
    with pytest.raises(wb_mod.WriteBackError, match="collide"):
        wb_mod.assign_tc_ids(TC_ID_CFG, [base_tc()] * 2, {"newR1L-AMFM-002"})


def test_no_ruled_scheme_leaves_the_column_alone():
    """A feature without tc_id_format must not invent ids."""
    assert wb_mod.assign_tc_ids(CFG, [base_tc()] * 2, set()) == [None, None]
    assert wb_mod.cell_values(base_tc(), CFG, None).get(5) is None


def test_the_tc_id_is_written_into_its_column():
    v = wb_mod.cell_values(base_tc(), TC_ID_CFG, "newR1L-AMFM-007")
    assert v[TC_ID_CFG["col"]["tc_id"]] == "newR1L-AMFM-007"
