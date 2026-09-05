"""來源抽取（R-G66，27 包 §D-7）。

**不以 openpyxl 寫任何 xlsx**（R-G3／27 包 §A）——
故本檔之工作簿一律取自 repo 內既有原檔，只讀不寫。
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

import extract_source as es  # noqa: E402

REPO = Path(__file__).resolve().parents[1]
SAMPLE = REPO / "features/vehicle_setting/inputs/PROXI_HDCC27_R3_20250424.xlsx"


@pytest.fixture
def staged(tmp_path: Path) -> Path:
    if not SAMPLE.exists():
        pytest.skip(f"樣本原檔不在：{SAMPLE.name}")
    raw = tmp_path / es.RAW_ROOT / "PROXI_HDCC27_R3"
    raw.mkdir(parents=True)
    shutil.copy2(SAMPLE, raw / SAMPLE.name)
    return tmp_path


# --- 抽取物之形 ---------------------------------------------------------------

def test_每個_sheet_一份_tsv_且首列帶來源_sha(staged: Path) -> None:
    doc_id, written = es.extract_one(staged, staged / es.RAW_ROOT
                                     / "PROXI_HDCC27_R3" / SAMPLE.name)
    out = staged / es.EXTRACTED_ROOT / doc_id
    assert len(list(out.glob("*.tsv"))) == len(written) > 1

    sha = es.sha256_of(staged / es.RAW_ROOT / "PROXI_HDCC27_R3" / SAMPLE.name)
    for tsv in out.glob("*.tsv"):
        head = tsv.read_text(encoding="utf-8").splitlines()[0]
        assert head.startswith("# source_sha256\t" + sha)


def test_行數與非空儲存格數逐_sheet_相符(staged: Path) -> None:
    """§F-6 之量：抽取物回讀之兩個量須等於 read_only 實測。"""
    doc_id, written = es.extract_one(staged, staged / es.RAW_ROOT
                                     / "PROXI_HDCC27_R3" / SAMPLE.name)
    out = staged / es.EXTRACTED_ROOT / doc_id
    for name, n_rows, n_filled in written:
        body = (out / f"{es.safe_name(name)}.tsv").read_text(
            encoding="utf-8").splitlines()[1:]
        filled = sum(1 for line in body for cell in line.split("\t") if cell.strip())
        assert (len(body), filled) == (n_rows, n_filled), f"sheet {name}"


# --- 序列化：抽取失真之實際發生處 ---------------------------------------------

def test_cell_text_跳脫_tab_與換行() -> None:
    """未跳脫者，一個含 tab 之儲存格會把該列之後續欄位全部推位。"""
    assert es.cell_text("a\tb\nc") == "a\\tb\\nc"
    assert es.cell_text(None) == ""
    assert es.cell_text("a\\b") == "a\\\\b"


def test_sheet_名不得逃出目標目錄() -> None:
    assert "/" not in es.safe_name("../../etc/passwd")
    assert es.safe_name("Revision Notes") == "Revision_Notes"
    assert es.safe_name("...") == "sheet"


def test_measure_不把空白字串算成非空() -> None:
    assert es.measure([("a", "", None, "  "), ("b", "c", None, None)]) == (2, 3)


def test_不支援之型別具名拒絕(tmp_path: Path) -> None:
    raw = tmp_path / es.RAW_ROOT / "SOME_DBC"
    raw.mkdir(parents=True)
    (raw / "bus.dbc").write_text("x", encoding="utf-8")
    with pytest.raises(ValueError, match="dbc"):
        es.extract_one(tmp_path, raw / "bus.dbc")


# --- A-POP1：抽取檔名撞名之靜默覆蓋 -------------------------------------------

def test_前導底線不被剝掉() -> None:
    """`_polarion` 與 `Polarion` 同簿；剝前導底線會讓兩者撞名。"""
    assert es.safe_name("_polarion") == "_polarion"
    assert es.safe_name("Polarion") == "Polarion"


def test_撞名之_sheet_停下而非靜默覆蓋(monkeypatch, tmp_path: Path) -> None:
    """大小寫不敏感之檔案系統上，後寫者會覆蓋前者；§F-6 自驗測不到。"""
    monkeypatch.setattr(es, "safe_name", lambda name: "same")
    fake = {"Polarion": [("a",)], "_polarion": [("b",)]}

    import openpyxl
    monkeypatch.setattr(openpyxl, "load_workbook",
                        lambda *a, **k: _Stub(fake))
    with pytest.raises(es.ExtractionMismatch, match="覆蓋"):
        es.extract_xlsx(Path("fake.xlsx"), tmp_path, "deadbeef")


class _Stub:
    """openpyxl workbook 之最小替身 —— 只供撞名測試用。"""

    def __init__(self, sheets: dict[str, list[tuple]]) -> None:
        self._sheets = sheets
        self.sheetnames = list(sheets)

    def __getitem__(self, name: str) -> "_Sheet":
        return _Sheet(self._sheets[name])

    def close(self) -> None:
        pass


class _Sheet:
    def __init__(self, rows: list[tuple]) -> None:
        self._rows = rows

    def iter_rows(self, values_only: bool = True):
        return iter(self._rows)
