"""Tests for features/home/scripts/write_back.py (Step 4).

The invariant that matters is content-based: the 144 Arif rows must survive a
rewrite that inserts and deletes rows around them. The integration test runs
the real dry run and asserts the hash is unchanged; it skips when the customer
inputs are absent (they are gitignored).

The unit tests cover the parts that were easy to get wrong: which regen segment
a leaf lands in, and the guards on placeholder rows.
"""
import importlib.util
import sys
from pathlib import Path

import pytest

HOME = Path(__file__).resolve().parent.parent / "features" / "home"

# features/media/scripts/write_back.py exists too — load this one under a unique
# module name so the two projects' tests cannot hand each other the wrong one.
sys.path.append(str(HOME / "scripts"))
_spec = importlib.util.spec_from_file_location(
    "home_write_back", HOME / "scripts" / "write_back.py")
if _spec is None or _spec.loader is None:
    pytest.skip("features/home write_back not present", allow_module_level=True)
write_back = importlib.util.module_from_spec(_spec)
sys.modules["home_write_back"] = write_back
_spec.loader.exec_module(write_back)


CFG = {
    "col": {"req_id": 3, "test_group": 6, "test_set": 7, "test_item": 8,
            "pre_conditions": 9, "input_test_data": 10, "test_procedure": 11,
            "expected_result": 12, "spec_reference": 13, "tc_ref_id": 14,
            "priority": 15, "design_method": 16, "functional_safety": 17,
            "author": 25, "remarks": 32},
    "write_back": {"author_value": "PeiPYHsu"},
}


def row(n, req_id, author, cells=None):
    return {"row": n, "req_id": req_id, "author": author,
            "cells": cells or [req_id, author]}


# ------------------------------------------------------------ done-region hash

def test_hash_selects_by_author_value_not_by_non_empty():
    """After write-back the regen rows also have an author; the selector must
    still pick out exactly the done region."""
    before = [row(10, "A", "Arif"), row(11, "B", ""), row(12, "C", "Arif")]
    after = [row(10, "A", "Arif"), row(11, "B", "PeiPYHsu"),
             row(12, "C", "Arif")]
    h1 = write_back.ordered_content_hash(before, "Arif")
    h2 = write_back.ordered_content_hash(after, "Arif")
    assert h1 == h2


def test_hash_changes_when_a_done_row_changes():
    before = [row(10, "A", "Arif", ["A", "Arif", "text"])]
    after = [row(10, "A", "Arif", ["A", "Arif", "edited"])]
    assert (write_back.ordered_content_hash(before, "Arif")
            != write_back.ordered_content_hash(after, "Arif"))


def test_hash_is_order_sensitive():
    a = [row(10, "A", "Arif"), row(11, "B", "Arif")]
    b = [row(10, "B", "Arif"), row(11, "A", "Arif")]
    assert (write_back.ordered_content_hash(a, "Arif")
            != write_back.ordered_content_hash(b, "Arif"))


def test_segments_classify_by_author_value():
    rows = [row(10, "A", "Arif"), row(11, "B", "PeiPYHsu"),
            row(12, "C", "Arif")]
    kinds = [s["kind"] for s in write_back.segments_of(rows, "Arif")]
    assert kinds == ["ARIF", "REGEN", "ARIF"]


# --------------------------------------------------------- segment assignment

LEAVES = ["L0", "L1", "L2", "L3", "L4", "L5", "L6"]


def seg(start, end, ids):
    return {"kind": "REGEN", "start": start, "end": end, "ids": ids}


def test_leaves_are_assigned_to_their_draft_segment():
    regen = [seg(87, 88, ["L0", "L1"]), seg(125, 126, ["L2", "L3"]),
             seg(162, 163, ["L5", "L6"])]
    a = write_back.assign_segments(LEAVES, regen)
    # Each segment's bound is the LAST of its own draft ids, so L1 stays in
    # segment 0 and L4 (no draft row) falls to the segment that brackets it.
    assert [a[x] for x in LEAVES] == [0, 0, 1, 1, 2, 2, 2]


def test_leaf_with_no_draft_row_lands_in_the_bracketing_segment():
    """055-03 and 066 have no draft row of their own; they must not be lost."""
    regen = [seg(87, 88, ["L0"]), seg(125, 126, ["L2"]), seg(162, 163, ["L6"])]
    a = write_back.assign_segments(LEAVES, regen)
    assert a["L1"] == 1, "L1 sits between the first and second segment bounds"
    assert a["L3"] == 2 and a["L4"] == 2 and a["L5"] == 2


def test_non_monotonic_segment_bounds_abort():
    regen = [seg(87, 88, ["L5"]), seg(125, 126, ["L0"])]
    with pytest.raises(write_back.WriteBackError, match="monotonic"):
        write_back.assign_segments(LEAVES, regen)


def test_segment_with_no_recognisable_draft_id_aborts():
    regen = [seg(87, 88, ["gone"])]
    with pytest.raises(write_back.WriteBackError, match="no draft req_id"):
        write_back.assign_segments(LEAVES, regen)


# ------------------------------------------------------------------ cell rules

def base_tc(**over):
    tc = {"req_id": "SWE1-HMI-HOME-020", "test_group": "", "test_set": "",
          "test_item": "item", "pre_conditions": "pre", "input_test_data": "",
          "test_procedure": "1. a\n2. b", "expected_result": "1. c\n2. d",
          "specification_reference": "ref", "priority": "P1",
          "design_method": "功能測試 (Functional based ; no specific technique)",
          "remarks": ""}
    tc.update(over)
    return tc


def test_constant_columns_are_written():
    v = write_back.cell_values(base_tc(), CFG)
    assert v[CFG["col"]["tc_ref_id"]] == "NEW"
    assert v[CFG["col"]["functional_safety"]] == "NA"
    assert v[CFG["col"]["author"]] == "PeiPYHsu"


def test_test_group_and_set_are_written_blank():
    v = write_back.cell_values(base_tc(test_group="X", test_set="Y"), CFG)
    assert v[CFG["col"]["test_group"]] is None
    assert v[CFG["col"]["test_set"]] is None


def test_empty_strings_become_none_not_empty_cells():
    v = write_back.cell_values(base_tc(), CFG)
    assert v[CFG["col"]["input_test_data"]] is None
    assert v[CFG["col"]["remarks"]] is None


def test_placeholder_must_leave_priority_blank():
    tc = base_tc(placeholder=True, priority="P1", design_method="",
                 test_procedure=write_back.PLACEHOLDER_BODY,
                 expected_result=write_back.PLACEHOLDER_BODY)
    with pytest.raises(write_back.WriteBackError, match="priority"):
        write_back.cell_values(tc, CFG)


def test_placeholder_body_must_be_the_fixed_string():
    tc = base_tc(placeholder=True, priority="", design_method="",
                 test_procedure="1. do a thing\n2. do another",
                 expected_result=write_back.PLACEHOLDER_BODY)
    with pytest.raises(write_back.WriteBackError, match="test_procedure"):
        write_back.cell_values(tc, CFG)


def test_valid_placeholder_passes():
    tc = base_tc(placeholder=True, priority="", design_method="",
                 test_procedure=write_back.PLACEHOLDER_BODY,
                 expected_result=write_back.PLACEHOLDER_BODY,
                 remarks="Covered by 066-01/066-02 (A-H01)")
    v = write_back.cell_values(tc, CFG)
    assert v[CFG["col"]["priority"]] is None
    assert v[CFG["col"]["design_method"]] is None
    assert v[CFG["col"]["remarks"]].endswith("(A-H01)")


# --------------------------------------------------------------- Scope field

SCOPE_CFG = {"workbook": {"header_row": 9},
             "write_back": {"scope_label": "範圍 Scope"}}


def _header_book(label="範圍 Scope：", label_at="C5", value="wrong-deliverable",
                 merge=None):
    import openpyxl as _pyxl
    wb = _pyxl.Workbook()
    ws = wb.active
    ws["A1"] = "STLA Test Case Specification & Result"
    ws["C2"] = "專案名稱 Project  Name："
    ws["D2"] = "newR1L"
    ws[label_at] = label
    r, c = _pyxl.utils.cell.coordinate_to_tuple(label_at)
    ws.cell(r, c + 1).value = value
    if merge:
        ws.merge_cells(merge)
    return ws


def test_scope_cell_is_located_by_its_label_not_a_coordinate():
    ws = _header_book()
    assert write_back.find_scope_cell(ws, SCOPE_CFG) == (5, 4)


def test_scope_cell_follows_the_label_when_the_layout_moves():
    """A workbook whose header block sits one column over must still resolve —
    AM/FM is the same template with a different offset."""
    ws = _header_book(label_at="D5")
    assert write_back.find_scope_cell(ws, SCOPE_CFG) == (5, 5)


def test_missing_scope_label_aborts():
    ws = _header_book(label="目的 Purpose：")
    with pytest.raises(write_back.WriteBackError, match="Scope"):
        write_back.find_scope_cell(ws, SCOPE_CFG)


def test_scope_search_stops_at_the_header_row():
    """Column N's own header text must never be mistaken for the Scope field."""
    ws = _header_book(label="目的 Purpose：")
    ws.cell(9, 14).value = "Specification Reference \n規格參考 Scope"
    with pytest.raises(write_back.WriteBackError, match="Scope"):
        write_back.find_scope_cell(ws, SCOPE_CFG)


def test_scope_is_rewritten_to_the_expected_deliverable():
    ws = _header_book(value="FM-WI-FSM-037-A03-N1L-SWE1-AppDrawer-Projection"
                            "-SWE1HMI-V0.1 STLA 報告")
    result = write_back.fix_scope(ws, SCOPE_CFG,
                                  "FM-WI-FSM-037-A03-N1L-SWE1-Home-HMI-V0.1 "
                                  "STLA 報告")
    assert result["changed"] is True
    assert "AppDrawer" in result["before"]
    assert ws.cell(5, 4).value == ("FM-WI-FSM-037-A03-N1L-SWE1-Home-HMI-V0.1 "
                                   "STLA 報告")


def test_fixing_an_already_correct_scope_is_a_no_op():
    ws = _header_book(value="FM-WI-FSM-037-A03-N1L-SWE1-Home-HMI-V0.1")
    result = write_back.fix_scope(ws, SCOPE_CFG,
                                  "FM-WI-FSM-037-A03-N1L-SWE1-Home-HMI-V0.1")
    assert result["changed"] is False


def test_scope_write_targets_the_anchor_of_a_merged_range():
    """D5:H5 is merged in the real workbook; writing to a non-anchor member
    raises inside openpyxl."""
    ws = _header_book(value=None, merge="D5:H5")
    write_back.fix_scope(ws, SCOPE_CFG, "correct-name")
    assert ws["D5"].value == "correct-name"


# ---------------------------------------------------------- normalisation

def test_normalisation_keeps_the_workbook_openable(tmp_path):
    """The dcterms substitution once used `\\1` immediately before a stamp
    starting with "2", which Python read as group 12 and silently destroyed
    docProps/core.xml. The file still hashed fine — only reopening caught it."""
    import openpyxl as _pyxl
    src = tmp_path / "wb.xlsx"
    wb = _pyxl.Workbook()
    wb.active["A1"] = "hello"
    wb.save(src)

    write_back.normalize_for_reproducibility(src)

    reopened = _pyxl.load_workbook(src)
    assert reopened.active["A1"].value == "hello"

    import zipfile
    core = zipfile.ZipFile(src).read("docProps/core.xml")
    assert b"2000-01-01T00:00:00Z" in core
    assert b"</dcterms:created>" in core, "closing tag must survive"


def test_normalisation_is_idempotent_and_stable(tmp_path):
    import openpyxl as _pyxl
    digests = []
    for name in ("a.xlsx", "b.xlsx"):
        p = tmp_path / name
        wb = _pyxl.Workbook()
        wb.active["A1"] = "same"
        wb.save(p)
        write_back.normalize_for_reproducibility(p)
        digests.append(write_back.sha256_file(p))
    assert digests[0] == digests[1], "two identical books must hash identically"


# ------------------------------------------------------------------ integration

@pytest.fixture(scope="module")
def dry_run():
    """Run the real dry run against the real workbook, or skip."""
    if not (HOME / "data" / "row_segments.json").exists():
        pytest.skip("features/home data/ not built")
    if not list((HOME / "generated").glob("*.json")):
        pytest.skip("features/home generated/ empty")
    import argparse
    import contextlib
    import io
    import os
    args = argparse.Namespace(write=False, feature_dir=str(HOME),
                              data=str(HOME / "data"),
                              generated=str(HOME / "generated"),
                              workbook=None, out=None, date=None)
    buf = io.StringIO()
    cwd = os.getcwd()
    try:
        os.chdir(HOME)
        with contextlib.redirect_stdout(buf):
            rc = write_back.run(args)
    except SystemExit as exc:  # resolve_path could not find the inputs
        pytest.skip(f"features/home inputs/ not present: {exc}")
    finally:
        os.chdir(cwd)
    return rc, buf.getvalue()


def test_dry_run_holds_every_invariant(dry_run):
    rc, out = dry_run
    assert rc == 0
    assert "144 Arif rows unchanged" in out
    assert "DRY RUN" in out


def test_dry_run_corrects_the_scope_field(dry_run):
    """A-H26 — the shipped workbook must claim its own 037 report."""
    _, out = dry_run
    line = next(l for l in out.splitlines() if l.startswith("Scope "))
    assert "AppDrawer-Projection" in line, "the residue must be reported"
    assert "SWE1-Home-HMI-V0.1" in out


def test_dry_run_writes_every_remaining_leaf(dry_run):
    _, out = dry_run
    assert "62 leaves" in out


def test_dry_run_preserves_the_segment_order(dry_run):
    _, out = dry_run
    line = next(l for l in out.splitlines() if l.startswith("segments"))
    kinds = [tok for tok in line.split() if tok in ("ARIF", "REGEN")]
    assert kinds == ["ARIF", "REGEN", "ARIF", "REGEN", "ARIF", "REGEN"]


def test_dry_run_is_deterministic(dry_run):
    """A second run must produce byte-identical output — the workbook is only
    ever read, so a differing second run would mean hidden state."""
    rc1, out1 = dry_run
    import argparse
    import contextlib
    import io
    import os
    args = argparse.Namespace(write=False, feature_dir=str(HOME),
                              data=str(HOME / "data"),
                              generated=str(HOME / "generated"),
                              workbook=None, out=None, date=None)
    buf = io.StringIO()
    cwd = os.getcwd()
    try:
        os.chdir(HOME)
        with contextlib.redirect_stdout(buf):
            rc2 = write_back.run(args)
    finally:
        os.chdir(cwd)
    assert (rc1, out1) == (rc2, buf.getvalue())
