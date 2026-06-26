"""Tests for Excel parser module."""
import os
import pytest
from openpyxl import Workbook

from parser import parse_tc_xlsx, parse_test_group_from_filename


@pytest.fixture
def sample_xlsx(tmp_path):
    """Create a sample TC xlsx file for testing."""
    filepath = tmp_path / "SomeProject_SWQT_DeviceManager_20260408.xlsx"
    wb = Workbook()

    # Product Document sheet
    ws_pd = wb.active
    ws_pd.title = "Product Document"
    ws_pd.cell(row=3, column=2, value="newR1L")

    # Test Case Specification&Result sheet
    ws_tc = wb.create_sheet("Test Case Specification&Result")
    # Header row 9
    headers = {
        2: "Test Level", 3: "Verification Criteria", 4: "Requirement or Design ID",
        5: "Functional Safety", 6: "Test Case ID", 7: "Test Group",
        8: "Test Set", 9: "Test Item", 10: "Pre-Conditions",
        11: "Input Test Data", 12: "Test Procedure", 13: "Expected Result",
        14: "Specification Reference", 15: "Remark", 16: "Test Case Priority",
        17: "Test Case Design Method",
    }
    for col, name in headers.items():
        ws_tc.cell(row=9, column=col, value=name)

    # Data rows (10, 11, 12)
    rows_data = [
        {
            "D": "SWE1-HMI-DM-001-01",
            "I": "PDM01.1) The Device Manager can be added to the status bar.",
            "L": "1. Old procedure",
            "M": "1. Old result",
        },
        {
            "D": "SWE1-HMI-DM-001-02",
            "I": "PDM01.2) The Device Manager icon shows connected device count.",
            "L": "",
            "M": "",
        },
        {
            "D": "SWE1-HMI-DM-002-01",
            "I": "PDM02) User can open Device Manager from the menu bar.",
            "L": "1. Some procedure",
            "M": "1. Some result",
        },
    ]
    col_map = {"D": 4, "I": 9, "L": 12, "M": 13}
    for i, row_data in enumerate(rows_data):
        row_num = 10 + i
        for col_letter, value in row_data.items():
            ws_tc.cell(row=row_num, column=col_map[col_letter], value=value)

    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def empty_xlsx(tmp_path):
    """Create an xlsx with no data rows."""
    filepath = tmp_path / "Empty_SWQT_TestGroup_20260408.xlsx"
    wb = Workbook()
    ws_pd = wb.active
    ws_pd.title = "Product Document"
    ws_pd.cell(row=3, column=2, value="projX")

    ws_tc = wb.create_sheet("Test Case Specification&Result")
    ws_tc.cell(row=9, column=4, value="Requirement or Design ID")
    ws_tc.cell(row=9, column=9, value="Test Item")

    wb.save(filepath)
    return str(filepath)


class TestParseTestGroupFromFilename:
    def test_standard_filename(self):
        result = parse_test_group_from_filename(
            "SomeProject_SWQT_DeviceManager_20260408.xlsx"
        )
        assert result == "DeviceManager"

    def test_filename_with_path(self):
        result = parse_test_group_from_filename(
            "/path/to/SomeProject_SWQT_MediaPlayer_20260410.xlsx"
        )
        assert result == "MediaPlayer"

    def test_filename_no_match(self):
        result = parse_test_group_from_filename("random_file.xlsx")
        assert result is None

    def test_filename_without_date_suffix(self):
        """Smoke / debug / adhoc 變體：`_SWQT_<group>_<anything>` 都能抽出 group。"""
        assert parse_test_group_from_filename(
            "FM-WI_SWQT_DeviceManager_smoke.xlsx"
        ) == "DeviceManager"
        assert parse_test_group_from_filename(
            "Proj_SWQT_BT_debug.xlsx"
        ) == "BT"

    def test_filename_group_only_no_suffix(self):
        """檔名結束於 group 本身（例：`..._SWQT_DeviceManager.xlsx`）也能抽出。"""
        assert parse_test_group_from_filename(
            "Proj_SWQT_DeviceManager.xlsx"
        ) == "DeviceManager"


class TestParseTcXlsx:
    def test_project_name(self, sample_xlsx):
        result = parse_tc_xlsx(sample_xlsx)
        assert result["project"] == "newR1L"

    def test_test_group(self, sample_xlsx):
        result = parse_tc_xlsx(sample_xlsx)
        assert result["test_group"] == "DeviceManager"

    def test_row_count(self, sample_xlsx):
        result = parse_tc_xlsx(sample_xlsx)
        assert result["row_count"] == 3

    def test_tc_sheet_with_bilingual_specification_name(self, tmp_path):
        filepath = tmp_path / "SomeProject_SWQT_Projection_20260502_EN.xlsx"
        wb = Workbook()
        ws_pd = wb.active
        ws_pd.title = "Product Document 記錄封面頁"
        ws_pd.cell(row=3, column=2, value="new R1L")

        ws_tc = wb.create_sheet("Test Case Specification 測試用例規範")
        ws_tc.cell(row=9, column=4, value="Requirement or Design ID")
        ws_tc.cell(row=9, column=9, value="Test Item")
        ws_tc.cell(row=10, column=4, value="SWE1-PROJ-001")
        ws_tc.cell(row=10, column=9, value="Projection phrase should be parsed.")
        wb.save(filepath)

        result = parse_tc_xlsx(str(filepath))

        assert result["row_count"] == 1
        assert result["rows"][0]["test_item"] == "Projection phrase should be parsed."

    def test_rows_structure(self, sample_xlsx):
        result = parse_tc_xlsx(sample_xlsx)
        rows = result["rows"]
        assert len(rows) == 3

        row0 = rows[0]
        assert row0["row_num"] == 10
        assert row0["req_id"] == "SWE1-HMI-DM-001-01"
        assert "PDM01.1" in row0["test_item"]

    def test_column_fill_status(self, sample_xlsx):
        result = parse_tc_xlsx(sample_xlsx)
        fill = result["column_fill_status"]
        # Col D (req_id) should be 3/3 filled
        assert fill["D"] == 3
        # Col L (procedure) has 2 filled, 1 empty
        assert fill["L"] == 2

    def test_empty_file(self, empty_xlsx):
        result = parse_tc_xlsx(empty_xlsx)
        assert result["row_count"] == 0
        assert result["rows"] == []

    def test_file_not_found(self):
        with pytest.raises(FileNotFoundError):
            parse_tc_xlsx("/nonexistent/file.xlsx")

    def test_skips_blank_rows_and_continues(self, tmp_path):
        filepath = tmp_path / "SomeProject_SWQT_DeviceManager_20260408.xlsx"
        wb = Workbook()
        ws_pd = wb.active
        ws_pd.title = "Product Document"
        ws_pd.cell(row=3, column=2, value="newR1L")

        ws_tc = wb.create_sheet("Test Case Specification&Result")
        ws_tc.cell(row=9, column=4, value="Requirement or Design ID")
        ws_tc.cell(row=9, column=9, value="Test Item")
        ws_tc.cell(row=10, column=4, value="SWE1-HMI-DM-001-01")
        ws_tc.cell(row=10, column=9, value="First row")
        ws_tc.cell(row=12, column=4, value="SWE1-HMI-DM-002-01")
        ws_tc.cell(row=12, column=9, value="Second row after blank line")
        wb.save(filepath)

        result = parse_tc_xlsx(str(filepath))
        assert result["row_count"] == 2
        assert [row["row_num"] for row in result["rows"]] == [10, 12]


class TestHeaderBasedColumnResolution:
    """Regression: real workbooks insert columns (e.g. 'Estimated Test Time'),
    shifting 'Design Methods' from Q to R. Resolution must follow the header
    text, not a fixed letter, and must disambiguate twin columns."""

    def _build(self, tmp_path, headers, data):
        filepath = tmp_path / "Proj_SWQT_Player_20260626.xlsx"
        wb = Workbook()
        ws_pd = wb.active
        ws_pd.title = "Product Document"
        ws_pd.cell(row=3, column=2, value="newR1L")
        ws_tc = wb.create_sheet("Test Case Specification&Result")
        for col, name in headers.items():
            ws_tc.cell(row=9, column=col, value=name)
        for col, value in data.items():
            ws_tc.cell(row=10, column=col, value=value)
        wb.save(filepath)
        return str(filepath)

    def test_shifted_design_method_read_from_correct_column(self, tmp_path):
        # Layout matching the real Player file: Q = Estimated Test Time (empty),
        # R = the actual Design Methods column.
        headers = {
            4: "Requirement or Design ID 需求/設計 ID",
            6: "Test Case ID 測試用例ID",
            9: "Test Item 測試項目",
            16: "Test Case Priority 測試用例優先級別",
            17: "Estimated Test Time (mins) 預估測試時間",
            18: "Test Case Design  Methods 測試用例設計方法",
        }
        data = {
            4: "SWE1-PLA-001",
            6: "NR1L-Player-001",
            9: "Verify popup PU0003 appears.",
            16: "P1",
            17: None,  # Estimated Test Time empty
            18: "負向測試 (Negative / Invalid)",
        }
        result = parse_tc_xlsx(self._build(tmp_path, headers, data))
        row = result["rows"][0]
        assert row["design_method"] == "負向測試 (Negative / Invalid)"
        assert row["priority"] == "P1"

    def test_twin_columns_disambiguated(self, tmp_path):
        # Polarion req-id (C) and TestRail tc-id (E) must NOT be picked; the
        # working req-id (D) and tc-id (F) must win.
        headers = {
            3: "Requirement or Design ID (Polarion) 設計/需求 ID (Polarion)",
            4: "Requirement or Design ID 需求/設計 ID",
            5: "Test Case ID (TestRail) 測試用例 ID (TestRail)",
            6: "Test Case ID 測試用例ID",
            9: "Test Item 測試項目",
        }
        data = {
            3: "POLARION-999",
            4: "SWE1-PLA-007",
            5: "TR-12345",
            6: "NR1L-Player-007",
            9: "Verify repeat mode.",
        }
        result = parse_tc_xlsx(self._build(tmp_path, headers, data))
        row = result["rows"][0]
        assert row["req_id"] == "SWE1-PLA-007"
        assert row["tc_id"] == "NR1L-Player-007"
