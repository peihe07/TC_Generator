"""Agent dispatcher 單元測試。

全部透過注入的 `llm_fn` 測試，完全不打真的 OpenAI API。
"""

from __future__ import annotations

from typing import Callable
from unittest.mock import patch

import pytest

from agent_dispatcher import (
    AgentEvent,
    DispatchContext,
    LLMStep,
    _requires_confirmation,
    build_system_prompt,
    dispatch_tool,
    run_agent_turn,
)
from trace_store import SqliteTraceStore


@pytest.fixture
def trace(tmp_path):
    s = SqliteTraceStore(tmp_path / "trace.db")
    yield s
    s.close()


@pytest.fixture
def ctx(trace):
    return DispatchContext(
        job_store={},
        trace_store=trace,
        session_id="test-session",
        rules_text="RULES_STUB",
        auto_approve=True,
    )


def _mock_llm(script: list[LLMStep]) -> Callable:
    """回傳一個 `llm_fn`：依序輸出 script 裡的 step。"""
    calls = iter(script)

    def fn(messages, tool_definitions):
        return next(calls)

    return fn


# ---------------------------------------------------------------------------
# System prompt
# ---------------------------------------------------------------------------

def test_build_system_prompt_includes_rules_when_provided():
    p = build_system_prompt("SOME_RULES_MARKDOWN")
    assert "SOME_RULES_MARKDOWN" in p
    assert "TC Generator operations co-pilot" in p


def test_build_system_prompt_without_rules():
    p = build_system_prompt(None)
    assert "TC Generator operations co-pilot" in p
    assert "authoritative" not in p


# ---------------------------------------------------------------------------
# dispatch_tool
# ---------------------------------------------------------------------------

def test_dispatch_tool_success_writes_trace(ctx, trace):
    out = dispatch_tool(
        "group_tests",
        {"rows": [{"id": "r1", "reqId": "R-001", "testItem": "PDM01 foo"}]},
        ctx,
    )

    assert out["ok"] is True
    assert "groups" in out["result"]
    assert out["duration_ms"] >= 0

    events = trace.replay("test-session")
    assert len(events) == 1
    assert events[0].type == "tool_call"
    assert events[0].tool == "group_tests"
    assert events[0].result_hash.startswith("sha256:")


def test_dispatch_tool_wraps_tool_error_as_ok_false(ctx, trace):
    out = dispatch_tool(
        "match_spec",
        {"rows": [], "reference_workbook_path": "/tmp/does_not_exist.xlsx"},
        ctx,
    )

    assert out["ok"] is False
    assert out["error"]["code"] == "not_found"

    events = trace.replay("test-session")
    assert events[0].type == "tool_error"
    assert events[0].tool == "match_spec"


def test_dispatch_tool_injects_job_store(ctx, tmp_path):
    # 用真實 parse_workbook 驗證注入鏈路：建最小 xlsx 避免 parse 失敗
    from openpyxl import Workbook

    wb = Workbook()
    pd = wb.active
    pd.title = "Product Document"
    pd.cell(row=3, column=2, value="newR1L")
    tc = wb.create_sheet("Test Case Specification&Result")
    tc.cell(row=9, column=4, value="Requirement or Design ID")
    tc.cell(row=9, column=9, value="Test Item")
    tc.cell(row=10, column=4, value="SWE1-DM-001")
    tc.cell(row=10, column=9, value="PDM01 text")
    path = tmp_path / "SomeProject_SWQT_DeviceManager_20260408.xlsx"
    wb.save(path)

    out = dispatch_tool(
        "parse_workbook",
        {"raw_path": str(path), "raw_filename": path.name},
        ctx,
    )

    assert out["ok"] is True
    # job_store 被寫入
    job_id = out["result"]["jobId"]
    assert job_id in ctx.job_store


# ---------------------------------------------------------------------------
# confirmation gate
# ---------------------------------------------------------------------------

def test_read_only_tool_never_requires_confirm():
    need, _ = _requires_confirmation("group_tests", {"rows": []})
    assert need is False


def test_destructive_tool_always_requires_confirm():
    need, _ = _requires_confirmation("write_excel", {})
    assert need is True


def test_generate_tc_confirm_depends_on_budget():
    # 少量 rows → 成本低於 $0.50 → 不用 confirm
    small, cost_small = _requires_confirmation(
        "generate_tc", {"rows": [{}], "model": "gpt-4.1-mini"}
    )
    assert cost_small > 0
    # 大量 rows → 估算成本超過 $0.50 → 要 confirm
    big, cost_big = _requires_confirmation(
        "generate_tc",
        {"rows": [{}] * 500, "model": "gpt-4.1"},
    )
    assert cost_big > cost_small
    assert big is True


# ---------------------------------------------------------------------------
# run_agent_turn — full loop
# ---------------------------------------------------------------------------

def test_loop_direct_text_reply_no_tool_calls(ctx):
    llm = _mock_llm([LLMStep(text="Hi there", tool_calls=[])])

    events, history = run_agent_turn(
        user_message="say hi",
        history=[],
        ctx=ctx,
        llm_fn=llm,
    )

    types = [e.type for e in events]
    assert types == ["message", "done"]
    assert events[0].data["text"] == "Hi there"
    assert history[-1] == {"role": "assistant", "content": "Hi there"}


def test_loop_tool_call_then_final_reply(ctx):
    llm = _mock_llm(
        [
            LLMStep(
                text=None,
                tool_calls=[
                    {
                        "id": "call-1",
                        "name": "group_tests",
                        "args": {"rows": [{"id": "r1", "reqId": "R-001", "testItem": "PDM01"}]},
                    }
                ],
            ),
            LLMStep(text="done grouping", tool_calls=[]),
        ]
    )

    events, _ = run_agent_turn(user_message="group them", history=[], ctx=ctx, llm_fn=llm)

    types = [e.type for e in events]
    assert types == ["tool_call", "tool_result", "message", "done"]
    assert events[0].data["tool"] == "group_tests"
    assert "groups" in events[1].data["result"]


def test_loop_emits_tool_error_on_tool_failure(ctx):
    llm = _mock_llm(
        [
            LLMStep(
                tool_calls=[
                    {
                        "id": "call-x",
                        "name": "match_spec",
                        "args": {"rows": [], "reference_workbook_path": "/nope.xlsx"},
                    }
                ]
            ),
            LLMStep(text="recovered", tool_calls=[]),
        ]
    )

    events, _ = run_agent_turn(user_message="match", history=[], ctx=ctx, llm_fn=llm)

    types = [e.type for e in events]
    assert types == ["tool_call", "tool_error", "message", "done"]
    assert events[1].data["error"]["code"] == "not_found"


def test_loop_halts_on_require_confirm_when_not_approved(ctx):
    ctx.auto_approve = False  # 強制走確認流程

    llm = _mock_llm(
        [
            LLMStep(
                tool_calls=[
                    {
                        "id": "call-risky",
                        "name": "write_excel",
                        "args": {"source_path": "s", "output_path": "o", "rows": [{"a": 1}]},
                    }
                ]
            ),
        ]
    )

    events, history = run_agent_turn(
        user_message="export it", history=[], ctx=ctx, llm_fn=llm
    )

    types = [e.type for e in events]
    assert types == ["require_confirm"]
    assert events[0].data["tool"] == "write_excel"
    # tool role message 要被加回 history 讓下輪 LLM 看到
    assert history[-1]["role"] == "tool"


def test_loop_proceeds_when_approved_call_ids_match(ctx):
    ctx.auto_approve = False
    ctx.approved_call_ids = {"call-ok"}

    llm = _mock_llm(
        [
            LLMStep(
                tool_calls=[
                    {
                        "id": "call-ok",
                        "name": "validate_tc",  # READ_ONLY，其實不會觸發 confirm
                        "args": {"tc": {"tc_id": "x", "test_item": "", "pre_conditions": "",
                                        "test_procedure": "", "expected_result": "",
                                        "design_method": "", "priority": ""}},
                    }
                ]
            ),
            LLMStep(text="ok", tool_calls=[]),
        ]
    )

    events, _ = run_agent_turn(user_message="check", history=[], ctx=ctx, llm_fn=llm)
    types = [e.type for e in events]
    assert "require_confirm" not in types


def test_loop_max_steps_emits_error(ctx):
    # 讓 LLM 永遠回 tool_call，看是否會 bail out
    def runaway(messages, tool_defs):
        return LLMStep(
            tool_calls=[
                {
                    "id": f"call-{len(messages)}",
                    "name": "group_tests",
                    "args": {"rows": []},
                }
            ]
        )

    events, _ = run_agent_turn(
        user_message="loop",
        history=[],
        ctx=ctx,
        llm_fn=runaway,
        max_steps=3,
    )
    assert events[-1].type == "error"
    assert events[-1].data["code"] == "max_steps_exceeded"


def test_loop_writes_llm_response_trace_when_usage_present(ctx, trace):
    llm = _mock_llm(
        [
            LLMStep(
                text="quick answer",
                tool_calls=[],
                usage={"input_tokens": 50, "output_tokens": 10},
            )
        ]
    )
    run_agent_turn(user_message="hi", history=[], ctx=ctx, llm_fn=llm)

    evs = trace.replay("test-session")
    llm_events = [e for e in evs if e.type == "llm_response"]
    assert len(llm_events) == 1
    assert llm_events[0].usage["output_tokens"] == 10


def test_loop_history_carries_over(ctx):
    first_llm = _mock_llm([LLMStep(text="hi", tool_calls=[])])
    _, history_after_first = run_agent_turn(
        user_message="你好", history=[], ctx=ctx, llm_fn=first_llm
    )

    captured_messages: list[list[dict]] = []

    def capture(messages, tool_defs):
        captured_messages.append(messages)
        return LLMStep(text="second reply", tool_calls=[])

    run_agent_turn(
        user_message="再來",
        history=history_after_first,
        ctx=ctx,
        llm_fn=capture,
    )

    # 第二輪看到的 messages 應包含上一輪的 user + assistant + 本輪的 user
    roles = [m["role"] for m in captured_messages[0]]
    assert roles[0] == "system"
    assert roles.count("user") == 2
    assert "assistant" in roles
