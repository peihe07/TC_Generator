"""`surgical_insert_rows` — structural mid-sheet row insertion.

`surgical_save` is append-only by design: it aborts rather than shift rows
(`patch_sheet_xml`). Row insertion is a separate structural pass, and these
tests pin the four things that make it safe to run against a controlled form:

1. rows below the anchor shift down, carrying their values
2. inserted rows are blank but keep the anchor's row height and cell styles —
   an unstyled row is a visible defect in a delivered workbook
3. every row-addressed reference outside `<sheetData>` moves with its rows
   (dropdowns, autofilter, colour scale) instead of silently detaching
4. the delivery invariant still holds: member set, data-validation counts

A reverse test pins the abort path: an anchor that is not an existing row
cannot be honoured, because there is no format to clone.
"""
from __future__ import annotations

import shutil
import zipfile
from pathlib import Path

import openpyxl
import pytest

from backend.xlsx_surgical import (
    StructureError,
    build_shift,
    insert_rows_xml,
    sheet_members,
    shift_ref_list,
    surgical_insert_rows,
    surgical_save,
)

REPO_ROOT = Path(__file__).resolve().parents[1]
BASE = REPO_ROOT / "features/power/sandbox/b19/pm_19.xlsx"
SHEET = "Test Case Specification&Result"

pytestmark = pytest.mark.skipif(
    not BASE.is_file(), reason="power b19 base workbook not present")


@pytest.fixture
def source(tmp_path: Path) -> Path:
    dst = tmp_path / "source.xlsx"
    shutil.copy(BASE, dst)
    return dst


def sheet_xml(path: Path) -> str:
    with zipfile.ZipFile(path) as z:
        return z.read(sheet_members(path)[SHEET]).decode("utf-8")


# ------------------------------------------------------------ offset algebra

def test_shift_counts_only_anchors_strictly_before():
    """Rows inserted after row k land below k, so k itself does not move."""
    shift = build_shift({11: 3, 12: 2, 23: 2})
    assert shift(10) == 10
    assert shift(11) == 11           # the anchor stays put
    assert shift(12) == 15           # 3 rows inserted after 11
    assert shift(13) == 18           # + 2 inserted after 12
    assert shift(24) == 31           # + 2 inserted after 23


def test_shift_ref_list_handles_multi_range_sqref():
    shift = build_shift({11: 3})
    assert shift_ref_list("P10:P221 Q10:Q11", shift) == "P10:P224 Q10:Q11"
    assert shift_ref_list("N9", shift) == "N9"


# ----------------------------------------------------------- value migration

def test_rows_below_the_anchor_shift_down_with_their_values(source, tmp_path):
    before = openpyxl.load_workbook(source)[SHEET]
    row12, row293 = before["F12"].value, before["F293"].value

    out = tmp_path / "expanded.xlsx"
    surgical_insert_rows(source, out, {11: 3}, SHEET)
    after = openpyxl.load_workbook(out)[SHEET]

    assert after["F11"].value == before["F11"].value    # anchor unmoved
    assert after["F15"].value == row12                  # shifted by 3
    assert after["F296"].value == row293
    assert after.max_row == before.max_row + 3


def test_inserted_rows_are_blank_but_keep_the_anchor_format(source, tmp_path):
    out = tmp_path / "expanded.xlsx"
    surgical_insert_rows(source, out, {11: 3}, SHEET)
    ws = openpyxl.load_workbook(out)[SHEET]

    for row in (12, 13, 14):
        assert all(ws.cell(row, c).value is None for c in range(1, 35))
        # Height and cell style are the anchor's, not openpyxl's defaults.
        assert ws.row_dimensions[row].height == ws.row_dimensions[11].height
        assert ws.cell(row, 12).border.left.style == \
            ws.cell(11, 12).border.left.style
        assert ws.cell(row, 12).alignment.wrap_text == \
            ws.cell(11, 12).alignment.wrap_text


# ------------------------------------------------------- reference migration

def test_row_addressed_references_move_with_their_rows(source, tmp_path):
    out = tmp_path / "expanded.xlsx"
    surgical_insert_rows(source, out, {11: 3}, SHEET)
    xml = sheet_xml(out)

    assert 'ref="A9:AH224"' in xml                  # autoFilter
    assert 'sqref="H10:H148"' in xml                # conditionalFormatting
    assert 'sqref="P10:P224 Q10:Q11"' in xml        # classic dataValidation
    assert "<xm:sqref>R10:R224</xm:sqref>" in xml   # x14 dataValidation
    assert 'ref="A1:AH296"' in xml                  # dimension restated
    # Merges live above the insertion zone and must be left exactly as found.
    assert 'ref="T8:Z8"' in xml and 'ref="A1:AE1"' in xml


def test_delivery_invariant_holds_after_insertion(source, tmp_path):
    out = tmp_path / "expanded.xlsx"
    report = surgical_insert_rows(source, out, {11: 3, 179: 1, 293: 2}, SHEET)

    assert report["inserted"] == 6
    assert report["differing"] == [sheet_members(source)[SHEET]]
    # x14 dropdown survives — the member-set clause alone cannot see this.
    assert any(x14 for _, x14 in report["dv_counts"].values())


def test_insertion_composes_with_a_following_surgical_save(source, tmp_path):
    """The two-pass shape this package runs: expand, then write values."""
    expanded = tmp_path / "expanded.xlsx"
    surgical_insert_rows(source, expanded, {11: 3}, SHEET)

    wb = openpyxl.load_workbook(expanded)
    wb[SHEET]["L12"] = "1. inserted step"
    out = tmp_path / "final.xlsx"
    report = surgical_save(wb, expanded, out)

    assert report["differing"] == [sheet_members(expanded)[SHEET]]
    assert openpyxl.load_workbook(out)[SHEET]["L12"].value == "1. inserted step"


# ----------------------------------------------------------- abort behaviour

def test_anchor_that_is_not_an_existing_row_aborts(source, tmp_path):
    with pytest.raises(StructureError) as exc:
        surgical_insert_rows(source, tmp_path / "x.xlsx", {6: 1}, SHEET)
    assert "insertion anchors absent" in str(exc.value)


def test_out_of_order_rows_abort():
    xml = ('<worksheet><dimension ref="A1:A2"/><sheetData>'
           '<row r="2"><c r="A2"/></row><row r="1"><c r="A1"/></row>'
           '</sheetData></worksheet>')
    with pytest.raises(StructureError) as exc:
        insert_rows_xml(xml, {1: 1})
    assert "ascending document order" in str(exc.value)
