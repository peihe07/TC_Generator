"""Parser regression tests focused on review-mode requirements.

Review pipeline depends on two parser invariants the generate path may not exercise:

1. Multi-Req-ID cells (newline-separated `req_id`) must be preserved verbatim
   so Tier 1 §6.7 detection can split on `\n`.
2. The `Specification Reference` column (col N) must surface as `spec_reference`
   so Tier 2 §7.4 can check whether a fabricated value is documented upstream.
"""
import pytest
from openpyxl import Workbook

from parser import parse_tc_xlsx


def _build_workbook(tmp_path, rows):
    """Build a minimal TC workbook with given rows. `rows` is a list of dicts
    keyed by column letter. Columns not provided default to empty."""
    filepath = tmp_path / "SomeProject_SWQT_Projection_20260502.xlsx"
    wb = Workbook()

    ws_pd = wb.active
    ws_pd.title = "Product Document"
    ws_pd.cell(row=3, column=2, value="ProjectionUnit")

    ws_tc = wb.create_sheet("Test Case Specification&Result")
    headers = {
        2: "Test Level", 3: "Verification Criteria", 4: "Requirement or Design ID",
        5: "Functional Safety", 6: "Test Case ID", 7: "Test Group",
        8: "Test Set", 9: "Test Item", 10: "Pre-Conditions",
        11: "Input Test Data", 12: "Test Procedure", 13: "Expected Result",
        14: "Specification Reference", 15: "Remark", 16: "Test Case Priority",
        17: "Test Case Design Method",
    }
    for col, name in headers.items():
        ws_tc.cell(row=9, column=col, value=name)

    col_letter_to_idx = {chr(ord("A") + i): i + 1 for i in range(17)}
    for offset, row in enumerate(rows):
        r = 10 + offset
        for letter, value in row.items():
            ws_tc.cell(row=r, column=col_letter_to_idx[letter], value=value)

    wb.save(filepath)
    return str(filepath)


def test_single_req_id_parsed_unchanged(tmp_path):
    fp = _build_workbook(tmp_path, [{
        "D": "SWE1-PROJ-071-001",
        "F": "TC-PROJ-071-001-001",
        "I": "the HU shall determine if the device supports wireless projection",
        "N": "Spec §3.2.1",
    }])
    parsed = parse_tc_xlsx(fp)
    assert parsed["row_count"] == 1
    row = parsed["rows"][0]
    assert row["req_id"] == "SWE1-PROJ-071-001"
    assert row["spec_reference"] == "Spec §3.2.1"


def test_multi_req_id_preserves_newline(tmp_path):
    """§6.7 trigger — newline-separated Req IDs must be preserved verbatim."""
    multi = "SWE1-PROJ-212\nSWE1-PROJ-213"
    fp = _build_workbook(tmp_path, [{
        "D": multi,
        "F": "TC-PROJ-212-001",
        "I": "[SWE1-PROJ-212]\nAudio sensitivity MUST be 2500 RMS\n[SWE1-PROJ-213]\nAudio distortion MUST be < 1%",
    }])
    parsed = parse_tc_xlsx(fp)
    row = parsed["rows"][0]
    assert "\n" in row["req_id"], "newline lost — §6.7 detection will break"
    assert row["req_id"].split("\n") == ["SWE1-PROJ-212", "SWE1-PROJ-213"]


def test_empty_spec_reference_is_empty_string(tmp_path):
    """Empty col N must parse as '' (not None) for consistent §7.4 / §8.3.6 checks."""
    fp = _build_workbook(tmp_path, [{
        "D": "SWE1-PROJ-229",
        "F": "TC-PROJ-229-001",
        "I": "車機應在五秒內完成連線",
        "L": "1. 等待5秒\n2. 確認連線",
        "N": None,
    }])
    parsed = parse_tc_xlsx(fp)
    row = parsed["rows"][0]
    assert row["spec_reference"] == ""
    assert isinstance(row["spec_reference"], str)
