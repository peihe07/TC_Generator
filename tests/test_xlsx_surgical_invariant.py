"""R18-4 — reverse test for the delivery-structure invariant.

An invariant that has never been seen to fail is not a verified invariant.
`verify_structure` guards every deliverable this repo emits (R18-3 rule 2),
so its ABORT path is exercised here directly: a check that cannot fail must
not be reported as passing.

Two damage modes, matching the two clauses of the invariant:

1. **member-set damage** — an output produced through `openpyxl`'s save path,
   which is exactly how AMFM v1 lost 21 zip members. Must ABORT naming the
   lost and added members.
2. **data-validation damage** — an output whose zip member set is IDENTICAL
   to the source and whose only defect is a stripped `x14` dropdown. This is
   the harder case and the reason clause 2 exists at all: member-set equality
   alone would wave it through.

A third case pins the positive control, so an always-raising implementation
cannot make the first two pass.
"""
from __future__ import annotations

import re
import shutil
import zipfile
from pathlib import Path

import openpyxl
import pytest

from backend.xlsx_surgical import (
    StructureError,
    sheet_members,
    surgical_save,
    verify_structure,
)

REPO_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = (REPO_ROOT / "features/privacy/inputs/FM-WI-FSM-036-A01 STLA "
            "測試用例規範與結果_SWQT STLA Test Case Specification & "
            "Result_SWQT_20260121.xlsx")
SHEET = "Test Case Specification 測試用例規範"

pytestmark = pytest.mark.skipif(
    not TEMPLATE.is_file(), reason="FW036 template not present in inputs/")


@pytest.fixture
def source(tmp_path: Path) -> Path:
    dst = tmp_path / "source.xlsx"
    shutil.copy(TEMPLATE, dst)
    return dst


def test_openpyxl_save_output_is_rejected(source: Path, tmp_path: Path):
    """Damage mode 1 — the exact failure that produced AMFM v1."""
    out = tmp_path / "via_openpyxl.xlsx"
    wb = openpyxl.load_workbook(source)
    wb[SHEET]["AF1"] = "damaged"
    wb.save(out)

    with pytest.raises(StructureError) as exc:
        verify_structure(source, out, {sheet_members(source)[SHEET]})

    message = str(exc.value)
    assert "zip member set changed" in message
    # The report must name what was lost, not merely that something was.
    assert "xl/sharedStrings.xml" in message
    assert "xl/printerSettings/printerSettings1.bin" in message
    print("\n[reverse test 1 — member-set damage] ABORTED:\n" + message)


def test_stripped_dropdown_is_rejected_despite_intact_member_set(
        source: Path, tmp_path: Path):
    """Damage mode 2 — member set intact, only the x14 dropdown removed.

    Clause 1 of the invariant cannot see this: every member is present and
    named identically. Only the per-sheet data-validation count catches it.
    """
    member = sheet_members(source)[SHEET]
    out = tmp_path / "stripped_dv.xlsx"
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(
            out, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == member:
                xml = data.decode("utf-8")
                xml = re.sub(r"<extLst>.*?</extLst>", "", xml, flags=re.DOTALL)
                data = xml.encode("utf-8")
            zout.writestr(info, data)

    # Precondition: the damage is invisible to the member-set clause.
    with zipfile.ZipFile(source) as a, zipfile.ZipFile(out) as b:
        assert set(a.namelist()) == set(b.namelist())

    with pytest.raises(StructureError) as exc:
        verify_structure(source, out, {member})

    message = str(exc.value)
    assert "data-validation counts changed" in message
    assert member in message
    print("\n[reverse test 2 — data-validation damage] ABORTED:\n" + message)


def test_surgical_output_passes(source: Path, tmp_path: Path):
    """Positive control — an always-raising check would fail here."""
    out = tmp_path / "via_surgical.xlsx"
    wb = openpyxl.load_workbook(source)
    wb[SHEET]["AF1"] = "written"
    report = surgical_save(wb, source, out)

    assert report["differing"] == [sheet_members(source)[SHEET]]
    assert openpyxl.load_workbook(out)[SHEET]["AF1"].value == "written"
