"""TC 分頁名之兩個真實變體（R-G48(b)，GC-03 §二-5-3）。

背景：`Test Case Specification&Result` 一度被當成「不存在之表名」而擬自夾具移除。
全 repo 145 本 FW036 工作簿實測 —— 帶中文副標者 121 本、帶 `&Result` 者 24 本
（`docs/reports/tc_sheetname_census_20260905.tsv`），**兩者皆為真實變體**。
既有夾具只覆蓋 `&Result` 一支，本檔補另一支，並守住「模板宣告之名確在母本內」。
"""
import re
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from parser import TC_SHEET_NAME_CANDIDATES, parse_tc_xlsx, resolve_tc_sheet
from writer import find_tc_sheet_name

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "forms" / (
    "FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT STLA "
    "Test Case Specification & Result_SWQT_20260817_ext.xlsx"
)
TEMPLATE = ROOT / "docs" / "fw036" / "templates" / "feature.yaml"

VARIANTS = [
    pytest.param("Test Case Specification 測試用例規範", id="zh-subtitle"),
    pytest.param("Test Case Specification&Result", id="ampersand-result"),
]

HEADERS = {
    4: "Requirement or Design ID", 6: "Test Case ID", 7: "Test Group",
    8: "Test Set", 9: "Test Item", 10: "Pre-Conditions",
    11: "Input Test Data", 12: "Test Procedure", 13: "Expected Result",
    14: "Specification Reference", 16: "Test Case Priority",
}


def _book(tmp_path, sheet_name):
    """造一本只在分頁名上有差異的工作簿。"""
    path = tmp_path / "P_SWQT_DeviceManager_20260408.xlsx"
    wb = Workbook()
    ws_pd = wb.active
    ws_pd.title = "Product Document"
    ws_pd.cell(row=3, column=2, value="newR1L")
    ws = wb.create_sheet(sheet_name)
    for col, name in HEADERS.items():
        ws.cell(row=9, column=col, value=name)
    ws.cell(row=10, column=4, value="SWE1-HMI-DM-001-01")
    ws.cell(row=10, column=9, value="PDM01.1) The Device Manager can be added.")
    wb.save(path)
    return path


@pytest.mark.parametrize("sheet_name", VARIANTS)
def test_resolver_finds_either_variant(tmp_path, sheet_name):
    """讀路徑：兩變體皆須解得其**自己**的名，不得回退到另一個。"""
    wb = load_workbook(_book(tmp_path, sheet_name))
    assert resolve_tc_sheet(wb) == sheet_name


@pytest.mark.parametrize("sheet_name", VARIANTS)
def test_writer_resolves_the_same_sheet_as_the_reader(tmp_path, sheet_name):
    """讀與寫必須走同一支 —— 否則會讀 A 寫 B 而看似成功。"""
    wb = load_workbook(_book(tmp_path, sheet_name))
    assert find_tc_sheet_name(wb) == resolve_tc_sheet(wb) == sheet_name


@pytest.mark.parametrize("sheet_name", VARIANTS)
def test_parser_reads_rows_from_either_variant(tmp_path, sheet_name):
    result = parse_tc_xlsx(str(_book(tmp_path, sheet_name)))
    assert [r["req_id"] for r in result["rows"]] == ["SWE1-HMI-DM-001-01"]


def test_feature_yaml_preference_wins_over_the_candidate_order(tmp_path):
    """feature.yaml 之 `workbook.sheet` 優先於候選集之順序（R-G48(b)）。"""
    path = tmp_path / "two.xlsx"
    wb = Workbook()
    wb.active.title = "Product Document"
    for name in ("Test Case Specification 測試用例規範", "Test Case Specification&Result"):
        wb.create_sheet(name)
    wb.save(path)
    loaded = load_workbook(path)
    assert resolve_tc_sheet(loaded, "Test Case Specification&Result") == \
        "Test Case Specification&Result"
    assert resolve_tc_sheet(loaded) == TC_SHEET_NAME_CANDIDATES[0]


def test_missing_tc_sheet_names_the_sheets_it_did_find(tmp_path):
    """解不出時要報錯並列出實際 sheetnames，不得靜默回退（否則寫到別的分頁）。"""
    path = tmp_path / "none.xlsx"
    wb = Workbook()
    wb.active.title = "Cover 封面"
    wb.save(path)
    with pytest.raises(KeyError, match="Cover 封面"):
        resolve_tc_sheet(load_workbook(path))


def test_template_sheet_name_exists_in_the_master_workbook():
    """守衛（R-G48(b)）：模板宣告之 `workbook.sheet` 必在 R-G1 母本之 sheetnames 內。

    模板值曾寫一個母本沒有的名（R-G48），本測試把「宣告」與「實物」綁在一起。
    """
    assert MASTER.exists(), f"R-G1 母本不在：{MASTER}"
    declared = re.search(r'^\s*sheet:\s*"(.+?)"', TEMPLATE.read_text(encoding="utf-8"), re.MULTILINE)
    assert declared, "模板未宣告 workbook.sheet"
    wb = load_workbook(MASTER, read_only=True)
    try:
        assert declared.group(1) in wb.sheetnames, (
            f"模板宣告 {declared.group(1)!r}，母本實有 {wb.sheetnames}"
        )
    finally:
        wb.close()
