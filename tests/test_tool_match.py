"""`match_spec_tool` 單元測試。"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from backend.tools import ToolError, match_spec_tool, get_tool, SafetyLevel


def _build_reference_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Basic Report"
    ws.cell(row=1, column=1, value="NRL ID")
    ws.cell(row=1, column=3, value="Outline")
    ws.cell(row=1, column=4, value="Description")
    ws.cell(row=1, column=5, value="Source ID")
    ws.cell(row=2, column=1, value="NRL-001")
    ws.cell(row=2, column=3, value="2.1")
    ws.cell(row=2, column=4, value="PDM01 behaviour description")
    ws.cell(row=2, column=5, value="SPEC_REF_PDM01")
    wb.save(path)
    return path


def test_match_without_reference_returns_unmatched():
    rows = [{"id": "r1", "req_id": "SWE1-DM-001", "test_item": "PDM01 foo"}]
    out = match_spec_tool(rows=rows, reference_workbook_path=None)

    assert out["summary"]["hasReferenceWorkbook"] is False
    assert out["summary"]["total"] == 1
    assert out["summary"]["unmatched"] == 1
    assert out["matches"][0]["matchType"] == "unmatched"
    assert out["matches"][0]["specReference"] is None


def test_match_with_reference_finds_exact(tmp_path):
    ref = _build_reference_workbook(tmp_path / "sys1.xlsx")
    rows = [{"id": "r1", "req_id": "SWE1-DM-001", "test_item": "PDM01 behaviour"}]

    out = match_spec_tool(rows=rows, reference_workbook_path=str(ref))

    assert out["summary"]["hasReferenceWorkbook"] is True
    assert out["summary"]["exact"] == 1
    assert out["matches"][0]["matchType"] == "exact"
    assert out["matches"][0]["specReference"] is not None


def test_match_missing_reference_path_raises(tmp_path):
    missing = tmp_path / "nope.xlsx"
    with pytest.raises(ToolError) as info:
        match_spec_tool(rows=[], reference_workbook_path=str(missing))
    assert info.value.code == "not_found"


def test_match_response_shape_keys():
    rows = [{"id": "r1", "req_id": "SWE1-DM-001", "test_item": "text"}]
    out = match_spec_tool(rows=rows)

    assert set(out.keys()) == {"summary", "matches"}
    assert set(out["matches"][0].keys()) == {
        "id",
        "reqId",
        "testItem",
        "specReference",
        "matchType",
        "matchScore",
    }


def test_match_tool_registered():
    spec = get_tool("match_spec")
    assert spec.safety is SafetyLevel.READ_ONLY
