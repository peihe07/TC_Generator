"""R-G42 DELIVERY-SPEC 閘之測試（scripts/lint_delivery_spec.py）。

以 openpyxl 造一本最小 036 形制工作簿（表頭列 9、資料列 10 起），
正向：全合規 → 零紅；反向：逐項破壞一格 → 恰命中該項（PLAYBOOK §7.1 雙向）。
"""

from __future__ import annotations

import hashlib
import importlib.util
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SPEC = importlib.util.spec_from_file_location("lds", ROOT / "scripts" / "lint_delivery_spec.py")
lds = importlib.util.module_from_spec(SPEC)
sys.modules["lds"] = lds
SPEC.loader.exec_module(lds)

FILENAME = ("FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT STLA Test Case Specification"
            " & Result_SWQT_BedLoweringMode_20260901.xlsx")
HEADERS = {
    3: "Requirement or Design ID\n需求", 5: "Test Case ID\n測試用例ID", 6: "Test Group\n測試組",
    7: "Test Set\n測試集", 8: "Test Item\n測試項目", 9: "Pre-Conditions\n先前條件",
    10: "Input Test Data\n輸入條件", 11: "Test procedure\n測試程序", 12: "Expected Result\n預期結果",
    13: "Specification Reference \n規格", 15: "Test Case Priority\n優先", 16: "Estimated Test Time (mins)",
    26: "Test Case Author\n作者",
}
ROWS = [  # (req, tc, group, prio, author, est, proc)
    ("SWE1-HMI-BLM-001-01", "NR1L-BLM-001", "Bed Lowering Mode", "P1", "PeiPYHsu", None, "1. Press"),
    ("SWE1-HMI-BLM-001-02", "NR1L-BLM-002", "Bed Lowering Mode", "P2", "PeiPYHsu", None, "1. Press"),
    ("SWE1-HMI-BLM-002-01", None, None, None, None, None, None),          # 無 TC 之需求空列
    ("SWE1-HMI-BLM-010-01", "NR1L-BLM-003", "Bed Lowering Mode", "P1", "PeiPYHsu", None, "1. Press"),
]


def build(tmp: Path, rows=ROWS, name=FILENAME, *, yaml=True, note=True, manifest=True,
          manifest_note="", dr_line="未結 DR：DR-BLM1\n") -> Path:
    fdir = tmp / "features" / "bed_lowering"
    ddir = fdir / "delivered"
    ddir.mkdir(parents=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Case Specification 測試用例規範"
    ws.cell(row=9, column=4, value=HEADERS[3])
    for col, text in HEADERS.items():
        ws.cell(row=9, column=col + 1, value=text)
    for i, (req, tc, group, prio, author, est, proc) in enumerate(rows, 10):
        ws.cell(row=i, column=4, value=req)
        if tc:
            ws.cell(row=i, column=6, value=tc); ws.cell(row=i, column=7, value=group)
            ws.cell(row=i, column=8, value="Lowering"); ws.cell(row=i, column=9, value="x\n(y)")
            ws.cell(row=i, column=12, value=proc); ws.cell(row=i, column=16, value=prio)
            ws.cell(row=i, column=17, value=est); ws.cell(row=i, column=27, value=author)
    book = ddir / name
    wb.save(book)
    if yaml:
        (fdir / "feature.yaml").write_text(
            'feature: "Bed Lowering Mode"\ndelivery:\n  tc_id_abbr: BLM\n  test_group: "Bed Lowering Mode"\n',
            encoding="utf-8")
    if manifest:
        sha = hashlib.sha256(book.read_bytes()).hexdigest()
        (ddir / "MANIFEST.tsv").write_text(
            "filename\tsha256\tsource_path\tdelivered_round\tnote\n"
            f"{name}\t{sha}\tsandbox/x.xlsx\t01\t{manifest_note}\n", encoding="utf-8")
    if note:
        (ddir / "DELIVERY_NOTE.md").write_text("# DELIVERY_NOTE\n" + dr_line, encoding="utf-8")
    (tmp / "docs" / "fw036").mkdir(parents=True)
    return book


def reds(tmp: Path, book: Path) -> list[str]:
    return lds.check_book(tmp, book).red


def test_compliant_book_has_no_red(tmp_path):
    book = build(tmp_path)
    assert reds(tmp_path, book) == []


def test_req_key_is_numeric_not_string():
    assert lds.req_key("SWE1-HMI-BLM-010-01") > lds.req_key("SWE1-HMI-BLM-002-01")
    assert lds.req_key("SWE-PM-2") < lds.req_key("SWE-PM-10")


def test_row_order_break_is_red(tmp_path):
    rows = [ROWS[3], ROWS[0], ROWS[1], ROWS[2]]
    book = build(tmp_path, rows)
    assert any(r.startswith("一(a)") for r in reds(tmp_path, book))


def test_empty_req_row_with_other_cells_is_red(tmp_path):
    rows = list(ROWS)
    rows[2] = ("SWE1-HMI-BLM-002-01", None, "Bed Lowering Mode", None, None, None, None)
    book = build(tmp_path, rows)
    # group cell is written only when tc present in build(); write it directly
    from openpyxl import load_workbook
    wb = load_workbook(book); wb.active.cell(row=12, column=7, value="Bed Lowering Mode"); wb.save(book)
    build_manifest_refresh(tmp_path, book)
    assert any(r.startswith("一(b) row 12") for r in reds(tmp_path, book))


def build_manifest_refresh(tmp: Path, book: Path):
    sha = hashlib.sha256(book.read_bytes()).hexdigest()
    m = book.parent / "MANIFEST.tsv"
    lines = m.read_text(encoding="utf-8").splitlines()
    cells = lines[1].split("\t"); cells[1] = sha
    m.write_text(lines[0] + "\n" + "\t".join(cells) + "\n", encoding="utf-8")


@pytest.mark.parametrize("bad_tc,expect", [
    ("newR1L-BLM-001", "二 TC ID 形制"),
    ("TC-DM-001", "二 TC ID 形制"),
    ("NR1L-VC-001", "二 ABBR 實為"),
])
def test_tc_id_shape_and_abbr(tmp_path, bad_tc, expect):
    rows = list(ROWS); rows[0] = (ROWS[0][0], bad_tc) + ROWS[0][2:]
    book = build(tmp_path, rows)
    assert any(r.startswith(expect) for r in reds(tmp_path, book))


def test_test_group_abbreviation_is_red(tmp_path):
    rows = [(r[0], r[1], "BLM" if r[1] else None) + r[3:] for r in ROWS]
    book = build(tmp_path, rows, name=FILENAME.replace("BedLoweringMode", "BLM"))
    assert any(r.startswith("三 Test Group 實為") for r in reds(tmp_path, book))


def test_author_priority_esttime(tmp_path):
    rows = list(ROWS)
    rows[0] = ROWS[0][:3] + ("High", "Arif", 5, ROWS[0][6])
    book = build(tmp_path, rows)
    v = lds.check_book(tmp_path, book)
    assert any(r.startswith("四 Author") for r in v.red)
    assert any(r.startswith("四 Priority") for r in v.red)
    assert any(w.startswith("四 Est. Time") for w in v.warn)


def test_filename_suffix_and_sandbox_name(tmp_path):
    book = build(tmp_path, name="bed_lowering_12.xlsx")
    assert any(r.startswith("五 檔名不合") for r in reds(tmp_path, book))


def test_filename_feature_name_must_match_group(tmp_path):
    book = build(tmp_path, name=FILENAME.replace("BedLoweringMode", "BedLowering"))
    assert any("≠ Test Group" in r for r in reds(tmp_path, book))


def test_manifest_missing_row_is_red(tmp_path):
    book = build(tmp_path, manifest=False)
    assert any(r.startswith("五 MANIFEST.tsv 無此檔") for r in reds(tmp_path, book))


def test_delivery_note_and_dr_list_required(tmp_path):
    book = build(tmp_path, note=False)
    out = reds(tmp_path, book)
    assert any(r.startswith("六 delivered/ 無 DELIVERY_NOTE") for r in out)
    assert any(r.startswith("六 未結 DR 清單缺") for r in out)


def test_pending_without_exception_is_red_with_exception_is_info(tmp_path):
    rows = list(ROWS); rows[0] = ROWS[0][:6] + ("1. PENDING: DR-BLM1 timeout",)
    book = build(tmp_path, rows)
    assert any(r.startswith("七 PENDING") for r in reds(tmp_path, book))
    build_manifest_refresh(tmp_path, book)
    m = book.parent / "MANIFEST.tsv"
    m.write_text(m.read_text(encoding="utf-8").rstrip("\n") + "S6 例外 R-BLM99\n", encoding="utf-8")
    v = lds.check_book(tmp_path, book)
    assert not any(r.startswith("七") for r in v.red)
    assert any(i.startswith("七 PENDING") for i in v.info)


def test_undeclared_yaml_is_red(tmp_path):
    book = build(tmp_path, yaml=False)
    out = reds(tmp_path, book)
    assert any("未宣告 delivery.tc_id_abbr" in r for r in out)
    assert any("未宣告 delivery.test_group" in r for r in out)


def test_baseline_grandfathers_but_gate_fails_on_fresh(tmp_path, monkeypatch, capsys):
    book = build(tmp_path, yaml=False)
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(sys, "argv", ["x", "--gate"])
    assert lds.main() == 1
    monkeypatch.setattr(sys, "argv", ["x", "--emit-baseline"])
    assert lds.main() == 0
    monkeypatch.setattr(sys, "argv", ["x", "--gate"])
    assert lds.main() == 0
    assert "基線（警示計數）" in capsys.readouterr().out
