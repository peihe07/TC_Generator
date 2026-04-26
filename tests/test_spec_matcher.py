"""Tests for spec matcher module (Layer 1: PDM code exact match)."""
import pytest
from openpyxl import Workbook

from spec_matcher import (
    SpecIndex,
    extract_pdm_codes,
    build_spec_index,
    load_spec_index,
    match_spec_references,
    save_spec_index,
)


class TestExtractPdmCodes:
    def test_single_pdm(self):
        assert extract_pdm_codes("PDM01.1) The Device Manager can be added") == ["PDM01.1"]

    def test_multiple_pdm(self):
        codes = extract_pdm_codes("PDM01 and PDM02.3 are related features")
        assert codes == ["PDM01", "PDM02.3"]

    def test_pdms_code(self):
        assert extract_pdm_codes("PDMS01 setup screen") == ["PDMS01"]

    def test_mpdm_code(self):
        assert extract_pdm_codes("MPDM1 media player") == ["MPDM1"]

    def test_td_code(self):
        assert extract_pdm_codes("TD1.7 touch display") == ["TD1.7"]

    def test_dnds_code(self):
        assert extract_pdm_codes("DNDS2 do not disturb") == ["DNDS2"]

    def test_pdee_code(self):
        assert extract_pdm_codes("PDEE01 power delivery") == ["PDEE01"]

    def test_apac_code(self):
        assert extract_pdm_codes("APAC0.1 asia pacific") == ["APAC0.1"]

    def test_psr_code(self):
        assert extract_pdm_codes("PSR1 passenger") == ["PSR1"]

    def test_no_codes(self):
        assert extract_pdm_codes("No codes in this text") == []

    def test_mixed_codes(self):
        codes = extract_pdm_codes("PDM01 and TD2.5 and PDMS03")
        assert set(codes) == {"PDM01", "TD2.5", "PDMS03"}


@pytest.fixture
def sys1_xlsx(tmp_path):
    """Create a sample SYS1 spec xlsx."""
    filepath = tmp_path / "SYS1_spec.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Basic Report"
    # Headers
    ws.cell(row=1, column=1, value="NRL ID")
    ws.cell(row=1, column=3, value="Outline Number")
    ws.cell(row=1, column=4, value="Description")
    ws.cell(row=1, column=5, value="Source ID")

    # Data
    data = [
        ("NRL-144757", "2.2", "PDM01.) The user can access the Device Manager",
         "Device_Manager_HMI Logic_and_Flow_R1_SR24_Post_2A_(March_13_2023)_2.2"),
        ("NRL-144758", "2.3", "PDM01.1) Device Manager can be added to status bar",
         "Device_Manager_HMI Logic_and_Flow_R1_SR24_Post_2A_(March_13_2023)_2.3"),
        ("NRL-144760", "3.1", "PDM02) Open Device Manager from menu",
         "Device_Manager_HMI Logic_and_Flow_R1_SR24_Post_2A_(March_13_2023)_3.1"),
        ("NRL-144770", "5.1", "TD1.7) Touch display calibration",
         "Device_Manager_HMI Logic_and_Flow_R1_SR24_Post_2A_(March_13_2023)_5.1"),
    ]
    for i, (nrl, outline, desc, source) in enumerate(data):
        row = i + 2
        ws.cell(row=row, column=1, value=nrl)
        ws.cell(row=row, column=3, value=outline)
        ws.cell(row=row, column=4, value=desc)
        ws.cell(row=row, column=5, value=source)

    wb.save(filepath)
    return str(filepath)


class TestBuildSpecIndex:
    def test_basic_index(self, sys1_xlsx):
        index = build_spec_index(sys1_xlsx)
        assert "PDM01" in index
        assert index["PDM01"]["nrl_id"] == "NRL-144757"
        assert index["PDM01"]["outline"] == "2.2"

    def test_sub_code_index(self, sys1_xlsx):
        index = build_spec_index(sys1_xlsx)
        assert "PDM01.1" in index
        assert index["PDM01.1"]["nrl_id"] == "NRL-144758"

    def test_td_code_index(self, sys1_xlsx):
        index = build_spec_index(sys1_xlsx)
        assert "TD1.7" in index

    def test_all_entries(self, sys1_xlsx):
        index = build_spec_index(sys1_xlsx)
        assert len(index) == 4

    def test_save_and_load_index_uses_json_cache(self, sys1_xlsx, tmp_path):
        index = build_spec_index(sys1_xlsx)
        cache_path = tmp_path / "SYS1_spec.json"

        save_spec_index(index, cache_path, name="SYS1_spec", source_file=sys1_xlsx)

        assert cache_path.read_text(encoding="utf-8").startswith("{")
        loaded = load_spec_index(["SYS1_spec"], index_dir=tmp_path)
        assert isinstance(loaded, SpecIndex)
        assert loaded["PDM01"]["nrl_id"] == "NRL-144757"
        assert len(loaded.entries) == len(index.entries)

    def test_load_index_rejects_legacy_binary_pickle(self, tmp_path):
        (tmp_path / "legacy.json").write_bytes(b"\x80\x05bad-pickle-payload")

        with pytest.raises(ValueError, match="expected JSON"):
            load_spec_index(["legacy"], index_dir=tmp_path)


class TestMatchSpecReferences:
    def test_single_match(self, sys1_xlsx):
        index = build_spec_index(sys1_xlsx)
        rows = [
            {"req_id": "R1", "test_item": "PDM01.1) Device Manager added to status bar"},
        ]
        results = match_spec_references(rows, index)
        assert results[0]["spec_reference"] == (
            "Device_Manager_HMI Logic_and_Flow_R1_SR24_Post_2A_(March_13_2023)_2.3"
        )
        assert results[0]["match_type"] == "exact"

    def test_multiple_matches(self, sys1_xlsx):
        index = build_spec_index(sys1_xlsx)
        rows = [
            {"req_id": "R1", "test_item": "PDM01 and PDM02 combined feature"},
        ]
        results = match_spec_references(rows, index)
        refs = results[0]["spec_reference"].split("; ")
        assert len(refs) == 2

    def test_no_match(self, sys1_xlsx):
        index = build_spec_index(sys1_xlsx)
        rows = [
            {"req_id": "R1", "test_item": "No PDM code here, just a description"},
        ]
        results = match_spec_references(rows, index)
        assert results[0]["spec_reference"] is None
        assert results[0]["match_type"] == "unmatched"


class TestFuzzyMatch:
    """Layer 1.5: token Jaccard fallback when Layer 1 PDM regex misses."""

    def test_fuzzy_hit_when_no_pdm_code(self, sys1_xlsx):
        index = build_spec_index(sys1_xlsx)
        # 沒 PDM code，但描述重疊關鍵字 (device, manager, access)
        rows = [{"req_id": "R1", "test_item": "The operator can access Device Manager screen"}]
        results = match_spec_references(rows, index)
        assert results[0]["match_type"] == "fuzzy"
        assert results[0]["spec_reference"] is not None
        # fuzzy score 介於 0 與 1
        assert 0 < results[0]["match_score"] <= 1

    def test_exact_preferred_over_fuzzy(self, sys1_xlsx):
        index = build_spec_index(sys1_xlsx)
        # 同時有 PDM code + 描述性文字 → 應優先 exact
        rows = [{"req_id": "R1", "test_item": "PDM01 Device Manager access"}]
        results = match_spec_references(rows, index)
        assert results[0]["match_type"] == "exact"
        assert "match_score" not in results[0]

    def test_fuzzy_below_threshold_stays_unmatched(self, sys1_xlsx):
        index = build_spec_index(sys1_xlsx)
        # 完全不同領域的文字
        rows = [{"req_id": "R1", "test_item": "Kitchen timer beeps when food cooks"}]
        results = match_spec_references(rows, index, fuzzy_threshold=0.5)
        assert results[0]["match_type"] == "unmatched"

    def test_threshold_controls_aggressiveness(self, sys1_xlsx):
        index = build_spec_index(sys1_xlsx)
        rows = [{"req_id": "R1", "test_item": "manager handles device status"}]
        # 低門檻容易命中，高門檻較嚴
        permissive = match_spec_references(rows, index, fuzzy_threshold=0.01)
        strict = match_spec_references(rows, index, fuzzy_threshold=0.99)
        assert permissive[0]["match_type"] == "fuzzy"
        assert strict[0]["match_type"] == "unmatched"
