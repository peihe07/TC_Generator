"""`write_excel_tool` 單元測試。"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from backend.tools import ToolError, get_tool, SafetyLevel, write_excel_tool


def _build_source_workbook(path: Path) -> Path:
    wb = Workbook()
    ws_pd = wb.active
    ws_pd.title = "Product Document"
    ws_pd.cell(row=3, column=2, value="newR1L")

    ws_tc = wb.create_sheet("Test Case Specification&Result")
    # header row 9
    ws_tc.cell(row=9, column=4, value="Requirement or Design ID")
    ws_tc.cell(row=9, column=9, value="Test Item")
    ws_tc.cell(row=10, column=4, value="SWE1-HMI-DM-001-01")
    ws_tc.cell(row=10, column=9, value="PDM01 original")
    wb.save(path)
    return path


def _export_row(row_num: int = 10) -> dict:
    return {
        "row_num": row_num,
        "tc_id": "newR1L-DMS-001",
        "test_group": "DeviceManager",
        "test_set": "Access",
        "tc_title": "(user taps X -> Y shows)",
        "pre_conditions": "1. System normal",
        "input_test_data": "NA",
        "test_procedure": "1. Do A.\n2. Verify B.",
        "expected_result": "1. A happens.\n2. B is visible.",
        "priority": "P1",
        "design_method": "功能測試 (Functional based ; no specific technique)",
        "spec_reference": "SPEC_REF_PDM01",
    }


def test_write_excel_writes_rows(tmp_path):
    source = _build_source_workbook(tmp_path / "src.xlsx")
    out = tmp_path / "out.xlsx"

    result = write_excel_tool(
        source_path=str(source),
        output_path=str(out),
        rows=[_export_row()],
    )

    assert result["rowsWritten"] == 1
    assert result["frameworkSheetWritten"] is False
    assert out.is_file()

    wb = load_workbook(out)
    ws = wb["Test Case Specification&Result"]
    # TC ID 落在 F 欄（第 6 欄）
    assert ws.cell(row=10, column=6).value == "newR1L-DMS-001"


def test_write_excel_with_framework_sheet(tmp_path):
    source = _build_source_workbook(tmp_path / "src.xlsx")
    out = tmp_path / "out.xlsx"

    result = write_excel_tool(
        source_path=str(source),
        output_path=str(out),
        rows=[_export_row()],
        framework_rows=[{"test_group": "DeviceManager", "test_set": "Access", "req_count": 1}],
    )

    assert result["frameworkSheetWritten"] is True
    wb = load_workbook(out)
    assert "Test Case Framework" in wb.sheetnames


def test_write_excel_missing_source_raises(tmp_path):
    with pytest.raises(ToolError) as info:
        write_excel_tool(
            source_path=str(tmp_path / "nope.xlsx"),
            output_path=str(tmp_path / "out.xlsx"),
            rows=[_export_row()],
        )
    assert info.value.code == "not_found"


def test_write_excel_empty_rows_raises(tmp_path):
    source = _build_source_workbook(tmp_path / "src.xlsx")
    with pytest.raises(ToolError) as info:
        write_excel_tool(
            source_path=str(source),
            output_path=str(tmp_path / "out.xlsx"),
            rows=[],
        )
    assert info.value.code == "bad_request"


def test_write_excel_tool_registered():
    spec = get_tool("write_excel")
    assert spec.safety is SafetyLevel.DESTRUCTIVE


def test_write_excel_strips_illegal_control_characters(tmp_path):
    source = _build_source_workbook(tmp_path / "src.xlsx")
    out = tmp_path / "out.xlsx"
    row = _export_row()
    row["input_test_data"] = "1. password: Ab1!abc\x07"
    row["tc_title"] = "(bell\x07 removed)"

    write_excel_tool(
        source_path=str(source),
        output_path=str(out),
        rows=[row],
    )

    wb = load_workbook(out)
    ws = wb["Test Case Specification&Result"]
    assert ws.cell(row=10, column=11).value == "1. password: Ab1!abc"
    assert "\x07" not in (ws.cell(row=10, column=9).value or "")
