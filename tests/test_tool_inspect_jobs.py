"""`inspect_workbook` / `list_jobs` / `estimate_cost` / `get_job_validation` 單元測試。"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from backend.tools import (
    SafetyLevel,
    ToolError,
    aggregate_metrics_tool,
    diff_jobs_tool,
    estimate_cost_tool,
    get_job_detail_tool,
    get_job_validation_tool,
    get_tool,
    inspect_workbook_tool,
    list_jobs_tool,
)


def _build_tc_workbook(path: Path) -> Path:
    wb = Workbook()
    pd = wb.active
    pd.title = "Product Document"
    pd.cell(row=3, column=2, value="newR1L")

    tc = wb.create_sheet("Test Case Specification&Result")
    tc.cell(row=9, column=4, value="Requirement or Design ID")
    tc.cell(row=9, column=9, value="Test Item")
    for i in range(3):
        tc.cell(row=10 + i, column=4, value=f"SWE1-DM-00{i + 1}-01")
        tc.cell(row=10 + i, column=9, value=f"PDM0{i + 1} text")
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# inspect_workbook
# ---------------------------------------------------------------------------

def test_inspect_workbook_reports_sheets_and_rows(tmp_path):
    path = _build_tc_workbook(tmp_path / "Proj_SWQT_DeviceManager_20260408.xlsx")
    out = inspect_workbook_tool(path=str(path))

    assert out["fileName"].startswith("Proj_SWQT_DeviceManager")
    assert out["fileSizeBytes"] > 0
    assert "Test Case Specification&Result" in out["sheets"]
    assert out["testGroup"] == "DeviceManager"
    assert out["rowsEstimate"] == 3
    assert out["hasTcSheet"] is True


def test_inspect_workbook_rejects_missing_file(tmp_path):
    with pytest.raises(ToolError) as info:
        inspect_workbook_tool(path=str(tmp_path / "nope.xlsx"))
    assert info.value.code == "not_found"


def test_inspect_workbook_rejects_bad_extension(tmp_path):
    bad = tmp_path / "thing.txt"
    bad.write_text("nope")
    with pytest.raises(ToolError) as info:
        inspect_workbook_tool(path=str(bad))
    assert info.value.code == "bad_request"


def test_inspect_workbook_handles_no_tc_sheet(tmp_path):
    wb = Workbook()
    wb.active.title = "Some Random Sheet"
    path = tmp_path / "nope_SWQT_X_20260101.xlsx"
    wb.save(path)

    out = inspect_workbook_tool(path=str(path))
    assert out["hasTcSheet"] is False
    assert out["rowsEstimate"] == 0


def test_inspect_tool_registered():
    spec = get_tool("inspect_workbook")
    assert spec.safety is SafetyLevel.READ_ONLY


# ---------------------------------------------------------------------------
# list_jobs
# ---------------------------------------------------------------------------

def test_list_jobs_empty_store():
    out = list_jobs_tool(job_store={})
    assert out["jobs"] == []
    assert out["total"] == 0


def test_list_jobs_returns_shape_and_limit():
    store = {
        "j1": {
            "jobId": "j1",
            "rawFileName": "a.xlsx",
            "status": "parsed",
            "parsedData": {"row_count": 10, "test_group": "DM", "project": "P"},
        },
        "j2": {
            "jobId": "j2",
            "rawFileName": "b.xlsx",
            "status": "completed",
            "parsedData": {"row_count": 5, "test_group": "CM", "project": "P"},
        },
    }
    out = list_jobs_tool(limit=1, job_store=store)
    assert out["total"] == 1
    assert len(out["jobs"]) == 1
    job = out["jobs"][0]
    assert {"jobId", "fileName", "status", "rowCount", "testGroup", "project"} <= set(job)


def test_list_jobs_status_filter():
    store = {
        "j1": {"jobId": "j1", "status": "parsed", "parsedData": {"row_count": 1}},
        "j2": {"jobId": "j2", "status": "completed", "parsedData": {"row_count": 2}},
    }
    out = list_jobs_tool(status_filter="completed", job_store=store)
    assert len(out["jobs"]) == 1
    assert out["jobs"][0]["jobId"] == "j2"


def test_list_jobs_invalid_limit_raises():
    with pytest.raises(ToolError):
        list_jobs_tool(limit=0, job_store={})


def test_list_jobs_registered():
    spec = get_tool("list_jobs")
    assert spec.safety is SafetyLevel.READ_ONLY


# ---------------------------------------------------------------------------
# estimate_cost
# ---------------------------------------------------------------------------

def test_estimate_cost_positive():
    store = {
        "jX": {
            "jobId": "jX",
            "parsedData": {"row_count": 20},
        }
    }
    out = estimate_cost_tool(job_id="jX", model="gpt-5.4-mini", job_store=store)
    assert out["rowCount"] == 20
    assert out["estCostUsd"] > 0
    assert out["jobId"] == "jX"


def test_estimate_cost_missing_job():
    with pytest.raises(ToolError) as info:
        estimate_cost_tool(job_id="ghost", job_store={})
    assert info.value.code == "not_found"


def test_estimate_cost_invalid_batch_size():
    with pytest.raises(ToolError):
        estimate_cost_tool(job_id="x", batch_size=0, job_store={"x": {"parsedData": {"row_count": 1}}})


def test_estimate_cost_registered():
    spec = get_tool("estimate_cost")
    assert spec.safety is SafetyLevel.READ_ONLY


# ---------------------------------------------------------------------------
# get_job_validation
# ---------------------------------------------------------------------------

def _row_with_generated(tc_id: str, *, priority: str = "Medium") -> dict:
    return {
        "id": f"row-{tc_id}",
        "tcId": tc_id,
        "originalRequirement": "Original",
        "generated": {
            "testItemRewrite": "(A → B)",
            "preConditions": "NA",
            "testProcedure": "1. Do A to start.\n2. Verify B.",
            "expectedResult": "1. A happens.\n2. B is visible.",
            "designMethod": "功能測試 (Functional based ; no specific technique)",
            "priority": priority,
        },
    }


def test_get_job_validation_aggregates_counts():
    rows = [
        _row_with_generated("tc-1"),
        _row_with_generated("tc-2", priority="Bogus"),  # 一定有 warning
    ]
    out = get_job_validation_tool(rows=rows)

    assert out["total"] == 2
    assert out["pass"] + out["warnings"] == 2
    assert out["warnings"] >= 1
    assert len(out["perRow"]) == 2


def test_get_job_validation_skips_rows_without_generated():
    rows = [
        {"id": "r1", "tcId": "t1"},              # 無 generated
        _row_with_generated("tc-2"),
    ]
    out = get_job_validation_tool(rows=rows)

    assert out["total"] == 2
    skipped = [r for r in out["perRow"] if r.get("skipped")]
    assert len(skipped) == 1


def test_get_job_validation_registered():
    spec = get_tool("get_job_validation")
    assert spec.safety is SafetyLevel.READ_ONLY


# ---------------------------------------------------------------------------
# get_job_detail
# ---------------------------------------------------------------------------

def test_get_job_detail_returns_core_fields():
    store = {
        "j1": {
            "jobId": "j1",
            "rawFileName": "Proj_SWQT_DM_20260408.xlsx",
            "rawBytes": b"\x00\x01\x02",  # 不應出現在輸出
            "status": "parsed",
            "createdAt": "2026-04-18T10:00:00Z",
            "updatedAt": "2026-04-18T10:05:00Z",
            "parsedData": {
                "row_count": 12,
                "project": "Proj",
                "test_group": "DM",
                "rows": [{"id": "r1"}, {"id": "r2"}],
            },
        }
    }
    out = get_job_detail_tool(job_id="j1", job_store=store)

    assert out["jobId"] == "j1"
    assert out["fileName"] == "Proj_SWQT_DM_20260408.xlsx"
    assert out["status"] == "parsed"
    assert out["rowCount"] == 12
    assert out["project"] == "Proj"
    assert out["testGroup"] == "DM"
    assert out["hasRawFile"] is True
    assert out["createdAt"] == "2026-04-18T10:00:00Z"
    assert out["updatedAt"] == "2026-04-18T10:05:00Z"
    # 不能外洩 rawBytes
    assert "rawBytes" not in out


def test_get_job_detail_summarises_match_and_group():
    store = {
        "j2": {
            "jobId": "j2",
            "status": "matched",
            "parsedData": {"row_count": 3},
            "matchResults": {
                "results": [
                    {"id": "r1", "matched": True},
                    {"id": "r2", "matched": False},
                    {"id": "r3", "matched": True},
                ]
            },
            "groupResults": {
                "groups": [
                    {"test_set": "Alpha", "count": 2},
                    {"test_set": "Beta", "count": 1},
                ]
            },
        }
    }
    out = get_job_detail_tool(job_id="j2", job_store=store)

    assert out["matchSummary"] == {"total": 3, "matched": 2, "unmatched": 1}
    assert out["groupSummary"] == {"groupCount": 2}


def test_get_job_detail_generated_summary_and_cost():
    store = {
        "j3": {
            "jobId": "j3",
            "status": "completed",
            "parsedData": {"row_count": 4},
            "generatedRows": [
                {"id": "r1", "generated": {"testItemRewrite": "x"}},
                {"id": "r2", "generated": None},
                {"id": "r3", "generated": {"testItemRewrite": "y"}},
                {"id": "r4"},
            ],
            "costUsd": 0.1234,
        }
    }
    out = get_job_detail_tool(job_id="j3", job_store=store)

    assert out["generatedRowCount"] == 2
    assert out["costUsd"] == 0.1234


def test_get_job_detail_missing_job_raises():
    with pytest.raises(ToolError) as info:
        get_job_detail_tool(job_id="ghost", job_store={})
    assert info.value.code == "not_found"


def test_get_job_detail_empty_job_id_rejected():
    with pytest.raises(ToolError) as info:
        get_job_detail_tool(job_id="", job_store={})
    assert info.value.code == "bad_request"


def test_get_job_detail_tolerates_minimal_record():
    store = {"jMin": {"jobId": "jMin"}}
    out = get_job_detail_tool(job_id="jMin", job_store=store)

    assert out["jobId"] == "jMin"
    assert out["rowCount"] == 0
    assert out["hasRawFile"] is False
    assert out["matchSummary"] is None
    assert out["groupSummary"] is None
    assert out["generatedRowCount"] == 0


def test_get_job_detail_registered():
    spec = get_tool("get_job_detail")
    assert spec.safety is SafetyLevel.READ_ONLY


# ---------------------------------------------------------------------------
# diff_jobs
# ---------------------------------------------------------------------------

def _job_record(*, job_id: str, row_count: int, **extra: Any) -> dict:
    base = {
        "jobId": job_id,
        "rawFileName": f"{job_id}.xlsx",
        "status": "completed",
        "parsedData": {"row_count": row_count, "project": "P", "test_group": "DM"},
    }
    base.update(extra)
    return base


def test_diff_jobs_returns_both_details_and_deltas():
    store = {
        "jA": _job_record(
            job_id="jA",
            row_count=10,
            costUsd=0.10,
            matchResults={
                "results": [{"matched": True}, {"matched": True}, {"matched": False}]
            },
        ),
        "jB": _job_record(
            job_id="jB",
            row_count=14,
            costUsd=0.16,
            matchResults={
                "results": [{"matched": True}] * 4 + [{"matched": False}]
            },
        ),
    }

    out = diff_jobs_tool(job_a_id="jA", job_b_id="jB", job_store=store)

    assert out["jobA"]["jobId"] == "jA"
    assert out["jobB"]["jobId"] == "jB"
    diff = out["diff"]
    assert diff["rowCountDelta"] == 4       # 14 - 10
    assert diff["costUsdDelta"] == pytest.approx(0.06)
    assert diff["statusChanged"] is False
    assert diff["matchedDelta"] == 2        # 4 - 2
    assert diff["unmatchedDelta"] == 0      # 1 - 1


def test_diff_jobs_detects_status_change():
    store = {
        "jA": _job_record(job_id="jA", row_count=5, status="parsed"),
        "jB": _job_record(job_id="jB", row_count=5, status="completed"),
    }
    out = diff_jobs_tool(job_a_id="jA", job_b_id="jB", job_store=store)
    assert out["diff"]["statusChanged"] is True


def test_diff_jobs_same_job_id_rejected():
    store = {"jA": _job_record(job_id="jA", row_count=1)}
    with pytest.raises(ToolError) as info:
        diff_jobs_tool(job_a_id="jA", job_b_id="jA", job_store=store)
    assert info.value.code == "bad_request"


def test_diff_jobs_missing_job_raises_not_found():
    store = {"jA": _job_record(job_id="jA", row_count=1)}
    with pytest.raises(ToolError) as info:
        diff_jobs_tool(job_a_id="jA", job_b_id="ghost", job_store=store)
    assert info.value.code == "not_found"


def test_diff_jobs_handles_null_cost_and_match():
    store = {
        "jA": _job_record(job_id="jA", row_count=3),  # 無 costUsd / matchResults
        "jB": _job_record(job_id="jB", row_count=5, costUsd=0.02),
    }
    out = diff_jobs_tool(job_a_id="jA", job_b_id="jB", job_store=store)

    assert out["diff"]["rowCountDelta"] == 2
    assert out["diff"]["costUsdDelta"] is None   # jA 無成本資料，delta 不可計算
    assert out["diff"]["matchedDelta"] is None   # 任一 job 無 matchResults → None


def test_diff_jobs_registered():
    spec = get_tool("diff_jobs")
    assert spec.safety is SafetyLevel.READ_ONLY


# ---------------------------------------------------------------------------
# aggregate_metrics
# ---------------------------------------------------------------------------

def test_aggregate_metrics_across_multiple_jobs():
    store = {
        "j1": _job_record(job_id="j1", row_count=10, costUsd=0.10),
        "j2": _job_record(job_id="j2", row_count=20, costUsd=0.30),
        "j3": _job_record(job_id="j3", row_count=30, costUsd=0.60),
    }
    out = aggregate_metrics_tool(job_store=store)

    assert out["jobCount"] == 3
    assert out["totalRowCount"] == 60
    assert out["totalCostUsd"] == pytest.approx(1.0)
    assert out["avgRowCount"] == pytest.approx(20.0)
    assert out["avgCostUsd"] == pytest.approx(1.0 / 3, abs=1e-4)  # 4 位小數四捨五入


def test_aggregate_metrics_respects_explicit_job_ids():
    store = {
        "j1": _job_record(job_id="j1", row_count=10, costUsd=0.10),
        "j2": _job_record(job_id="j2", row_count=20, costUsd=0.30),
        "j3": _job_record(job_id="j3", row_count=30, costUsd=0.60),
    }
    out = aggregate_metrics_tool(job_ids=["j1", "j3"], job_store=store)

    assert out["jobCount"] == 2
    assert out["totalRowCount"] == 40
    assert out["totalCostUsd"] == pytest.approx(0.70)


def test_aggregate_metrics_match_rate_only_counts_jobs_with_match():
    store = {
        "j1": _job_record(
            job_id="j1",
            row_count=4,
            matchResults={"results": [{"matched": True}] * 3 + [{"matched": False}]},
        ),
        "j2": _job_record(
            job_id="j2",
            row_count=2,
            matchResults={"results": [{"matched": True}, {"matched": False}]},
        ),
        "j3": _job_record(job_id="j3", row_count=5),  # 無 match results → 不納入
    }
    out = aggregate_metrics_tool(job_store=store)

    # matched: 3+1 = 4; total matched-phase rows: 4+2 = 6 → rate = 2/3
    assert out["matchRate"] == pytest.approx(4 / 6)
    assert out["jobsWithMatch"] == 2


def test_aggregate_metrics_handles_empty_store():
    out = aggregate_metrics_tool(job_store={})

    assert out["jobCount"] == 0
    assert out["totalRowCount"] == 0
    assert out["totalCostUsd"] is None
    assert out["avgCostUsd"] is None
    assert out["avgRowCount"] is None
    assert out["matchRate"] is None


def test_aggregate_metrics_missing_cost_yields_null_totals():
    store = {
        "j1": _job_record(job_id="j1", row_count=3),  # 無 costUsd
        "j2": _job_record(job_id="j2", row_count=5, costUsd=0.10),
    }
    out = aggregate_metrics_tool(job_store=store)

    # 至少一筆有 cost → totalCostUsd 就是有 cost 那些的加總；但 avgCostUsd 以
    # 「有 cost 的 job 數」為分母，避免把無資料的 job 當成 $0 稀釋平均
    assert out["totalCostUsd"] == pytest.approx(0.10)
    assert out["avgCostUsd"] == pytest.approx(0.10)
    assert out["jobsWithCost"] == 1


def test_aggregate_metrics_unknown_job_id_raises():
    store = {"j1": _job_record(job_id="j1", row_count=3)}
    with pytest.raises(ToolError) as info:
        aggregate_metrics_tool(job_ids=["ghost"], job_store=store)
    assert info.value.code == "not_found"


def test_aggregate_metrics_registered():
    spec = get_tool("aggregate_metrics")
    assert spec.safety is SafetyLevel.READ_ONLY
