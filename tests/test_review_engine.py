"""Tests for the ASPICE SWE.6 Review engine.

Covers:
  - 4 deep-dive cases referenced by CLAUDE_CODE_BRIEF.md (synthesized
    fixtures rather than pulling rows out of a real workbook):
      A: fabricated value Critical (§7.4) — group has spec句
      B: multiple fabricated values across siblings
      C: multi-Req-ID + tool launch (§6.7 + §7.5 + §6.2-style boundary)
      D: tier1_skipped + §8.3.6 fallback (§6.6 + §8.3.6, NOT §7.4)
  - severity ceiling enforcement (Tier 3 cannot emit Critical)
  - §7.4 ⊕ §8.3.6 mutual exclusion
  - §6.4 → §8.1.4 suppression
  - dry_run does not invoke OpenAI
"""
from __future__ import annotations

import json
from types import SimpleNamespace

from openpyxl import Workbook

import pytest

import review_engine
from review_engine import (
    ReqGroup,
    ReviewEngineError,
    TCRecord,
    _build_groups,
    _detect_7_4_or_8_3_6,
    _detect_7_5,
    _detect_8_3_5,
    _enforce_severity_ceiling,
    _normalize_row,
    requirement_anchor,
    _run_llm_pipeline,
    _run_regex_pipeline,
    extract_spec_sentence,
    review_workbook,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _build_workbook(tmp_path, rows):
    fp = tmp_path / "Review_SWQT_Projection_20260502.xlsx"
    wb = Workbook()
    ws_pd = wb.active
    ws_pd.title = "Product Document"
    ws_pd.cell(row=3, column=2, value="Projection")

    ws_tc = wb.create_sheet("Test Case Specification&Result")
    headers = {
        2: "Test Level", 3: "Verification Criteria", 4: "Requirement or Design ID",
        5: "Functional Safety", 6: "Test Case ID", 7: "Test Group",
        8: "Test Set", 9: "Test Item", 10: "Pre-Conditions",
        11: "Input Test Data", 12: "Test Procedure", 13: "Expected Result",
        14: "Specification Reference", 15: "Remark", 16: "Test Case Priority",
        17: "Test Case Design Method",
    }
    for col, name in headers.items():
        ws_tc.cell(row=9, column=col, value=name)

    col_letter_to_idx = {chr(ord("A") + i): i + 1 for i in range(17)}
    for offset, row in enumerate(rows):
        r = 10 + offset
        for letter, value in row.items():
            ws_tc.cell(row=r, column=col_letter_to_idx[letter], value=value)
    wb.save(fp)
    return str(fp)


def _make_tc(**kw) -> TCRecord:
    """Cheap TCRecord factory with sensible defaults."""
    defaults = {
        "row_num": kw.pop("row_num", 10),
        "tc_id": kw.pop("tc_id", "TC-X-001"),
        "req_id_raw": kw.pop("req_id_raw", "REQ-A"),
        "req_ids": kw.pop("req_ids", ["REQ-A"]),
        "test_item": kw.pop("test_item", ""),
        "pre_conditions": kw.pop("pre_conditions", ""),
        "input_test_data": kw.pop("input_test_data", ""),
        "test_procedure": kw.pop("test_procedure", ""),
        "expected_result": kw.pop("expected_result", ""),
        "spec_reference": kw.pop("spec_reference", ""),
        "priority": kw.pop("priority", "P1"),
        "design_method": kw.pop("design_method", "Functional"),
    }
    return TCRecord(**defaults)


# ---------------------------------------------------------------------------
# Severity ceiling
# ---------------------------------------------------------------------------


def test_severity_ceiling_tier3_cannot_emit_critical():
    with pytest.raises(ReviewEngineError):
        _enforce_severity_ceiling(3, "§8.X", "Critical")


def test_severity_ceiling_passes_at_ceiling():
    assert _enforce_severity_ceiling(3, "§8.5.2", "Major") == "Major"
    assert _enforce_severity_ceiling(2, "§7.4", "Critical") == "Critical"
    assert _enforce_severity_ceiling(1, "§6.5", "Critical") == "Critical"


def test_unknown_severity_raises():
    with pytest.raises(ReviewEngineError):
        _enforce_severity_ceiling(2, "§7.5", "Severe")


# ---------------------------------------------------------------------------
# Spec句 extraction
# ---------------------------------------------------------------------------


def test_extract_english_spec_sentence():
    s = extract_spec_sentence("the HU shall connect to the device")
    assert s is not None and "shall" in s


def test_extract_returns_none_when_no_modal():
    assert extract_spec_sentence("車機應該完成連線（中文 only）") is None


def test_multi_req_id_segment_match():
    text = (
        "[SWE1-PROJ-212]\nAudio sensitivity MUST be 2500 RMS\n"
        "[SWE1-PROJ-213]\nAudio distortion MUST be < 1%"
    )
    s_a = extract_spec_sentence(text, "SWE1-PROJ-212")
    s_b = extract_spec_sentence(text, "SWE1-PROJ-213")
    assert s_a and "2500 RMS" in s_a
    assert s_b and "< 1%" in s_b


# ---------------------------------------------------------------------------
# Group A — fabricated value Critical (§7.4)
# ---------------------------------------------------------------------------


def test_group_a_fabricated_value_fires_7_4_when_spec_present():
    """Spec句 exists; spec_reference empty; "等待5秒" not in spec — §7.4 Critical fires.
    §8.3.6 must NOT fire (mutual exclusion)."""
    tc = _make_tc(
        test_item="the HU shall complete pairing",
        test_procedure="1. Bond the device.\n2. 等待5秒.\n3. Confirm pairing.",
        spec_reference="",
    )
    group = ReqGroup(req_id="REQ-A", tcs=[tc], spec_sentence="the HU shall complete pairing", tier1_skipped=False)
    findings = _detect_7_4_or_8_3_6(tc, group)
    refs = [f["rule_ref"] for f in findings]
    assert "§7.4" in refs
    assert "§8.3.6" not in refs
    f74 = next(f for f in findings if f["rule_ref"] == "§7.4")
    assert f74["severity"] == "Critical"
    assert f74["tier"] == 2
    assert f74["evidence_req_spec"] == "the HU shall complete pairing"


def test_value_in_spec_reference_does_not_fire_7_4():
    tc = _make_tc(
        test_item="the HU shall complete pairing",
        test_procedure="1. 等待 5 秒.",
        spec_reference="Spec §3.2: timeout 5 sec",
    )
    group = ReqGroup(req_id="REQ-A", tcs=[tc], spec_sentence="the HU shall complete pairing")
    assert _detect_7_4_or_8_3_6(tc, group) == []


def test_value_in_spec_sentence_does_not_fire_7_4():
    tc = _make_tc(
        test_item="the HU shall complete pairing in 5 seconds",
        test_procedure="1. 等待 5 秒.",
    )
    group = ReqGroup(req_id="REQ-A", tcs=[tc], spec_sentence="the HU shall complete pairing in 5 seconds")
    assert _detect_7_4_or_8_3_6(tc, group) == []


# ---------------------------------------------------------------------------
# Group B — multiple fabricated values (§7.4 fires per value)
# ---------------------------------------------------------------------------


def test_group_b_multiple_fabricated_values():
    tc = _make_tc(
        test_item="the system shall handle long sessions",
        test_procedure="1. 持續 30 分鐘.\n2. 重複 20 次.\n3. 維持 3 小時.",
        spec_reference="",
    )
    group = ReqGroup(req_id="REQ-B", tcs=[tc],
                     spec_sentence="the system shall handle long sessions")
    findings = _detect_7_4_or_8_3_6(tc, group)
    assert len(findings) == 3
    assert all(f["rule_ref"] == "§7.4" and f["severity"] == "Critical" for f in findings)


# ---------------------------------------------------------------------------
# Group C — multi-Req-ID + Tool launch (§6.7 + §7.5)
# ---------------------------------------------------------------------------


def test_group_c_multi_req_and_tool_launch(tmp_path):
    fp = _build_workbook(tmp_path, [{
        "D": "SWE1-PROJ-212\nSWE1-PROJ-213",
        "F": "TC-PROJ-212-001",
        "I": (
            "[SWE1-PROJ-212]\nAudio sensitivity MUST be 2500 RMS.\n"
            "[SWE1-PROJ-213]\nAudio distortion MUST be < 1%."
        ),
        "L": (
            "1. Setup phone.\n"
            "2. Connect HFP.\n"
            "3. Open PCTS-MT1 panel.\n"
            "4. Configure test profile.\n"
            "5. Run PCTS-MT1 to start measurement."
        ),
        "M": "1. ok\n2. ok\n3. ok\n4. ok\n5. measurement started.",
        "N": "",
        "P": "P1",
        "Q": "Functional",
    }])

    report = review_workbook(fp, dry_run=True)

    # §6.7 fires once with comma-joined req_id
    f67 = [f for f in report["per_req_findings"] if f["rule_ref"] == "§6.7"]
    assert len(f67) == 1
    assert f67[0]["req_id"] == "SWE1-PROJ-212, SWE1-PROJ-213"
    assert f67[0]["severity"] == "Major"

    # §7.5 fires on the multi-Req TC (last step launches PCTS-MT1, no follow-up)
    tc_entry = report["per_tc_findings"][0]
    refs = [f["rule_ref"] for f in tc_entry["findings"]]
    assert "§7.5" in refs


# ---------------------------------------------------------------------------
# Group D — tier1_skipped + §8.3.6 fallback (NOT §7.4)
# ---------------------------------------------------------------------------


def test_group_d_tier1_skipped_triggers_8_3_6_not_7_4(tmp_path):
    """When the Req group has no English spec句, §6.6 fires (Major),
    Tier 2 §7.4 is skipped, and §8.3.6 fallback (Major) takes the
    fabricated-value finding."""
    fp = _build_workbook(tmp_path, [{
        "D": "SWE1-PROJ-229",
        "F": "TC-PROJ-229-001",
        "I": "車機應在配對流程中完成裝置綁定。",
        "L": "1. 觸發配對\n2. 重複 5 次\n3. 確認結果",
        "M": "1. 開始\n2. 完成\n3. 通過",
        "N": "",
        "P": "P1",
        "Q": "Functional",
    }])
    report = review_workbook(fp, dry_run=True)

    refs_req = [f["rule_ref"] for f in report["per_req_findings"]]
    assert "§6.6" in refs_req

    tc_entry = report["per_tc_findings"][0]
    refs = [f["rule_ref"] for f in tc_entry["findings"]]
    assert "§8.3.6" in refs
    assert "§7.4" not in refs
    assert "§7.5" not in refs  # Tier 2 entirely skipped


# ---------------------------------------------------------------------------
# Mutual exclusion sanity (synthesized)
# ---------------------------------------------------------------------------


def test_mutual_exclusion_74_836_never_both_on_same_value():
    tc = _make_tc(
        test_procedure="1. 等待 5 秒",
        spec_reference="",
    )
    # Tier1 active group → §7.4 only
    g_active = ReqGroup(req_id="REQ", tcs=[tc],
                        spec_sentence="the HU shall connect", tier1_skipped=False)
    f1 = _detect_7_4_or_8_3_6(tc, g_active)
    refs1 = {f["rule_ref"] for f in f1}
    assert refs1 == {"§7.4"}

    # Tier1 skipped group → §8.3.6 only
    g_skipped = ReqGroup(req_id="REQ", tcs=[tc], spec_sentence=None, tier1_skipped=True)
    f2 = _detect_7_4_or_8_3_6(tc, g_skipped)
    refs2 = {f["rule_ref"] for f in f2}
    assert refs2 == {"§8.3.6"}


# ---------------------------------------------------------------------------
# §6.4 ↔ §8.1.4 suppression
# ---------------------------------------------------------------------------


def test_64_fires_when_siblings_have_identical_test_item(tmp_path):
    fp = _build_workbook(tmp_path, [
        {"D": "REQ-DUP", "F": "TC-A", "I": "the device shall connect via BT",
         "L": "1. Step.", "M": "1. Done.", "P": "P1", "Q": "Functional"},
        {"D": "REQ-DUP", "F": "TC-B", "I": "the device shall connect via BT",
         "L": "1. Step.", "M": "1. Done.", "P": "P1", "Q": "Functional"},
    ])
    report = review_workbook(fp, dry_run=True)
    refs = [f["rule_ref"] for f in report["per_req_findings"]]
    assert "§6.4" in refs


# ---------------------------------------------------------------------------
# Tier 3 detectors (sanity)
# ---------------------------------------------------------------------------


def test_8_5_2_design_method_missing(tmp_path):
    fp = _build_workbook(tmp_path, [{
        "D": "REQ-A",
        "F": "TC-A-1",
        "I": "the HU shall display the icon",
        "L": "1. Open menu.\n2. Confirm icon visible.",
        "M": "1. Menu opens.\n2. Icon shown.",
        "P": "P1",
        # Q (design_method) intentionally blank
    }])
    report = review_workbook(fp, dry_run=True)
    refs = [f["rule_ref"] for f in report["per_tc_findings"][0]["findings"]]
    assert "§8.5.2" in refs


def test_8_5_1_priority_outside_p0_p3(tmp_path):
    fp = _build_workbook(tmp_path, [{
        "D": "REQ-A", "F": "TC-A-1",
        "I": "the HU shall display the icon",
        "L": "1. Open menu.\n2. Confirm icon visible.",
        "M": "1. Menu opens.\n2. Icon shown.",
        "P": "High",
        "Q": "Functional",
    }])
    report = review_workbook(fp, dry_run=True)
    refs = [f["rule_ref"] for f in report["per_tc_findings"][0]["findings"]]
    assert "§8.5.1" in refs


class TestRequirementAnchor:
    def test_shall_sentence(self):
        req = {"id": "R", "title": "T", "desc": "The HU shall display PU0003."}
        assert "shall display" in requirement_anchor(req)

    def test_will_sentence_anchors(self):
        # requirements legitimately use "will" — must anchor, not be rejected.
        req = {"id": "R", "title": "Skip Back Completion",
               "desc": "If the HU finds the file it will exit Skip Back and play."}
        assert requirement_anchor(req) is not None
        assert "will exit" in requirement_anchor(req)

    def test_declarative_requirement_anchors(self):
        req = {"id": "R", "title": "AUX",
               "desc": "The HU AUX source does not have any command capability."}
        assert requirement_anchor(req) is not None

    def test_cross_reference_pointer_does_not_anchor(self):
        req = {"id": "R", "title": "Video Player in Dealer Mode",
               "desc": "Refer to CFTS012-696 and CFTS022-2237 for more requirements."}
        assert requirement_anchor(req) is None


class TestTier1AnchorsOnRequirement:
    def test_group_anchors_on_swe1_req_when_tc_has_no_spec_sentence(self):
        # TC carries no shall/must/should — Tier 1 must still anchor via SWE1 req.
        tc = _make_tc(req_id_raw="SWE1-PLA-022", req_ids=["SWE1-PLA-022"],
                      test_item="Skip forward and observe the next file plays")
        reqs = [{"id": "SWE1-PLA-022", "title": "Skip Forward Completion",
                 "desc": "If the HU finds the Next File it will exit Skip Forward "
                         "and play the new file from the beginning."}]
        groups = _build_groups([tc], reqs)
        grp = groups["SWE1-PLA-022"]
        assert grp.tier1_skipped is False
        assert grp.spec_sentence and "will exit" in grp.spec_sentence

    def test_pointer_requirement_still_tier1_skipped(self):
        tc = _make_tc(req_id_raw="SWE1-PLA-052", req_ids=["SWE1-PLA-052"],
                      test_item="Refer to CFTS012; select a Showroom Demo Video")
        reqs = [{"id": "SWE1-PLA-052", "title": "Video Player in Dealer Mode",
                 "desc": "Refer to CFTS012-696 for more requirements."}]
        groups = _build_groups([tc], reqs)
        assert groups["SWE1-PLA-052"].tier1_skipped is True

    def test_no_swe1_reqs_keeps_legacy_behavior(self):
        # Without --swe1-reqs, a TC with no spec句 is still tier1_skipped.
        tc = _make_tc(req_id_raw="SWE1-PLA-022", req_ids=["SWE1-PLA-022"],
                      test_item="Skip forward and observe")
        groups = _build_groups([tc])
        assert groups["SWE1-PLA-022"].tier1_skipped is True


class TestInteractiveBridge:
    def test_export_then_assemble_merges_llm_with_regex(self, tmp_path):
        from review_engine import export_review_bundle, assemble_review
        fp = _build_workbook(tmp_path, [{
            "D": "REQ-A", "F": "TC-A-1",
            "I": "the HU shall display the icon",
            "L": "1. Open menu.\n2. Confirm icon visible.",
            "M": "1. Menu.\n2. Icon shown.",
            "P": "P1", "Q": "Functional",
        }])
        bundle = export_review_bundle(fp)

        # exported deterministically, no model called
        assert bundle["schema"] == "review-bundle/v1"
        assert bundle["total_tcs"] == 1
        assert bundle["batches"] and bundle["batches"][0]["answer"] is None
        assert "user_prompt" in bundle["batches"][0]
        # regex findings already present in the bundle
        assert "per_tc" in bundle["regex_findings"]

        # Claude fills the batch answer in-session (simulated here)
        bundle["batches"][0]["answer"] = {
            "per_req_findings": [],
            "per_tc_findings": [
                {"row": 10, "tc_id": "TC-A-1",
                 "findings": [{"rule_ref": "§7.6", "tier": 2, "severity": "Major",
                               "field": "expected_result", "issue": "reality gap demo"}]},
            ],
        }
        report = assemble_review(bundle, output_dir=str(tmp_path / "out"))

        assert report["batch_meta"]["llm_stats"]["mode"] == "interactive"
        row10 = next(e for e in report["per_tc_findings"] if e["row"] == 10)
        refs = [f["rule_ref"] for f in row10["findings"]]
        assert "§7.6" in refs  # LLM finding merged
        assert (tmp_path / "out" / "findings.json").is_file()

    def test_assemble_rejects_bad_schema(self, tmp_path):
        from review_engine import assemble_review, ReviewEngineError
        with pytest.raises(ReviewEngineError):
            assemble_review({"schema": "nope", "batches": []})


def test_8_3_1_forbidden_verb(tmp_path):
    fp = _build_workbook(tmp_path, [{
        "D": "REQ-A", "F": "TC-A-1",
        "I": "the HU shall display the icon",
        "L": "1. Open menu.\n2. Observe whether icon appears.",
        "M": "1. Menu.\n2. Icon shown.",
        "P": "P1", "Q": "Functional",
    }])
    report = review_workbook(fp, dry_run=True)
    refs = [f["rule_ref"] for f in report["per_tc_findings"][0]["findings"]]
    assert "§8.3.1" in refs


def test_8_3_5_accepts_real_check_verbs():
    """Regression: a final step that verifies an outcome must NOT be flagged.
    Real test cases write 'Check the …' / 'Verify that …', not only the narrow
    'Check that' the old regex required (which also omitted 'Verify')."""
    for last_step in (
        "5. Check the system response after selecting the USB source.",
        "8. Verify that all controls except Source are disabled.",
        "3. Validate the repeat mode status is highlighted.",
        "4. 確認重複模式顯示為 Repeat All。",
    ):
        tc = _make_tc(test_procedure="1. Open the player.\n" + last_step)
        assert _detect_8_3_5(tc) == [], f"false positive on: {last_step}"


def test_8_3_5_flags_bare_action_final_step():
    """A final step that is a bare action with no verification is still flagged."""
    tc = _make_tc(test_procedure="1. Open the player.\n2. Select the USB source.")
    findings = _detect_8_3_5(tc)
    assert [f["rule_ref"] for f in findings] == ["§8.3.5"]


def test_8_1_1_exempts_requirement_sentence_test_item(tmp_path):
    """House convention: Test Item carries the full shall/must/should
    requirement句; §8.1.1 length limit must not fire on it."""
    fp = _build_workbook(tmp_path, [{
        "D": "REQ-A", "F": "TC-A-1",
        "I": ("When a No Supported Files Found Error occurs, the HU shall "
              "display Pop Up ID PU0003 with the corresponding message text."),
        "L": "1. Trigger the error.\n2. Check the popup ID.",
        "M": "1. Error triggered.\n2. PU0003 shown.",
        "P": "P1", "Q": "Functional",
    }])
    report = review_workbook(fp, dry_run=True)
    refs = [f["rule_ref"] for f in report["per_tc_findings"][0]["findings"]]
    assert "§8.1.1" not in refs


def test_8_4_1_successfully_with_concrete_verb_not_vague(tmp_path):
    """Regression: 'connected successfully' is observable (mirrors the ZH
    成功(?!連線) carve-out) and must NOT be flagged as vague."""
    fp = _build_workbook(tmp_path, [{
        "D": "REQ-A", "F": "TC-A-1",
        "I": "the HU shall connect the BTSA device",
        "L": "1. Pair the device.\n2. Check the connection state.",
        "M": "1. The BTSA device is connected successfully.",
        "P": "P1", "Q": "Functional",
    }])
    report = review_workbook(fp, dry_run=True)
    refs = [f["rule_ref"] for f in report["per_tc_findings"][0]["findings"]]
    assert "§8.4.1" not in refs


def test_8_4_1_bare_vague_words_still_flagged(tmp_path):
    """Genuinely vague outcomes (no concrete observable) are still flagged."""
    fp = _build_workbook(tmp_path, [{
        "D": "REQ-A", "F": "TC-A-1",
        "I": "the HU shall play the track",
        "L": "1. Start playback.\n2. Check the result.",
        "M": "1. The system works correctly and behaves as expected.",
        "P": "P1", "Q": "Functional",
    }])
    report = review_workbook(fp, dry_run=True)
    refs = [f["rule_ref"] for f in report["per_tc_findings"][0]["findings"]]
    assert "§8.4.1" in refs


def test_8_1_1_still_flags_long_title_without_spec_sentence(tmp_path):
    """A long Test Item that is NOT a normative requirement句 is still flagged."""
    fp = _build_workbook(tmp_path, [{
        "D": "REQ-A", "F": "TC-A-1",
        "I": ("Player USB source browsing folder category navigation and item "
              "selection across every supported media type end to end"),
        "L": "1. Open menu.\n2. Check the list.",
        "M": "1. Menu.\n2. List shown.",
        "P": "P1", "Q": "Functional",
    }])
    report = review_workbook(fp, dry_run=True)
    refs = [f["rule_ref"] for f in report["per_tc_findings"][0]["findings"]]
    assert "§8.1.1" in refs


# ---------------------------------------------------------------------------
# Dry run never calls the LLM
# ---------------------------------------------------------------------------


def test_dry_run_does_not_invoke_openai(tmp_path, monkeypatch):
    """If dry_run=True, _run_llm_pipeline must not be called and no OpenAI
    import should happen via review_engine."""
    called = {"flag": False}

    def _boom(*a, **kw):
        called["flag"] = True
        raise AssertionError("LLM pipeline must not run in dry-run")

    monkeypatch.setattr(review_engine, "_run_llm_pipeline", _boom)
    fp = _build_workbook(tmp_path, [{
        "D": "REQ-A", "F": "TC-A-1",
        "I": "the HU shall display the icon",
        "L": "1. Open menu.\n2. Confirm icon shown.",
        "M": "1. Menu.\n2. Icon shown.",
        "P": "P1", "Q": "Functional",
    }])
    review_workbook(fp, dry_run=True)
    assert called["flag"] is False


def test_llm_pipeline_keeps_empty_tc_id_rows_distinct(monkeypatch):
    tcs = [
        _make_tc(row_num=10, tc_id=""),
        _make_tc(row_num=11, tc_id=""),
    ]

    def _fake_chat(**_kwargs):
        body = {
            "per_req_findings": [],
            "per_tc_findings": [
                {
                    "tc_id": "",
                    "row": 10,
                    "findings": [{
                        "tier": 3,
                        "field": "expected_result",
                        "rule_ref": "§8.4.2",
                        "severity": "Major",
                        "issue": "第 10 列 finding",
                    }],
                },
                {
                    "tc_id": "",
                    "row": 11,
                    "findings": [{
                        "tier": 3,
                        "field": "expected_result",
                        "rule_ref": "§8.4.2",
                        "severity": "Major",
                        "issue": "第 11 列 finding",
                    }],
                },
            ],
        }
        from providers import LLMResponse, LLMUsage
        return LLMResponse(text=json.dumps(body), usage=LLMUsage(), model="fake")

    monkeypatch.setattr("generator._chat", _fake_chat)
    _, per_tc, _stats = _run_llm_pipeline(tcs, {"REQ-A": ReqGroup(req_id="REQ-A", tcs=tcs)}, model="fake")

    assert [entry["row"] for entry in per_tc] == [10, 11]
    assert [entry["findings"][0]["issue"] for entry in per_tc] == ["第 10 列 finding", "第 11 列 finding"]


# ---------------------------------------------------------------------------
# Output schema invariants
# ---------------------------------------------------------------------------


def test_output_writes_json_and_markdown(tmp_path):
    fp = _build_workbook(tmp_path, [{
        "D": "REQ-A", "F": "TC-A-1",
        "I": "the HU shall display the icon",
        "L": "1. Open menu.\n2. Confirm icon shown.",
        "M": "1. Menu.\n2. Icon shown.",
        "P": "P1", "Q": "Functional",
    }])
    out = tmp_path / "out"
    review_workbook(fp, output_dir=str(out), dry_run=True)
    assert (out / "findings.json").is_file()
    assert (out / "findings_report.md").is_file()


def test_tier2_findings_have_evidence_req_spec(tmp_path):
    fp = _build_workbook(tmp_path, [{
        "D": "REQ-A", "F": "TC-A-1",
        "I": "the HU shall complete pairing",
        "L": "1. Trigger.\n2. 等待 5 秒.\n3. Confirm.",
        "M": "1. ok\n2. ok\n3. ok",
        "P": "P1", "Q": "Functional",
    }])
    report = review_workbook(fp, dry_run=True)
    for entry in report["per_tc_findings"]:
        for f in entry["findings"]:
            if f["tier"] == 2:
                assert "evidence_req_spec" in f and f["evidence_req_spec"]
            else:
                assert "evidence_req_spec" not in f


def test_llm_pipeline_injects_domain_block_and_reality_gap(monkeypatch):
    """Stage 6: domain pack is injected into the review prompt, and §7.6
    reality-gap findings flow through with their flag intact."""
    from review_prompt_builder import LLM_RULE_HINTS
    assert "§7.6" in LLM_RULE_HINTS  # reality-gap rule registered

    tcs = [_make_tc(row_num=10, tc_id="T10")]
    captured = {}

    def _fake_chat(system, user, model, json_mode=True, max_tokens=None):
        captured["user"] = user
        from providers import LLMResponse, LLMUsage
        body = {
            "per_req_findings": [],
            "per_tc_findings": [{
                "tc_id": "T10", "row": 10,
                "findings": [{
                    "tier": 2, "field": "test_procedure", "rule_ref": "§7.6",
                    "severity": "Major",
                    "issue": "假設了 spec 未定義的 No Repeat 態",
                    "evidence": "step 3", "reality_gap": True,
                }],
            }],
        }
        return LLMResponse(text=json.dumps(body), usage=LLMUsage(), model="fake")

    monkeypatch.setattr("generator._chat", _fake_chat)
    _, per_tc, _stats = _run_llm_pipeline(
        tcs, {"REQ-A": ReqGroup(req_id="REQ-A", tcs=tcs)},
        model="fake", domain_block="# Domain Pack — Player\nRepeat only All/One Track",
    )

    # Domain block reached the prompt.
    assert "Domain Pack" in captured["user"]
    assert "Repeat only All/One Track" in captured["user"]
    # Reality-gap finding flowed through with its flag.
    gaps = [f for tc in per_tc for f in tc["findings"] if f.get("reality_gap")]
    assert gaps and gaps[0]["rule_ref"] == "§7.6"


def test_review_workbook_dry_run_feeds_scorecard(tmp_path):
    """Cross-module: real findings.json shape from review_workbook parses cleanly
    into the Stage 7 scorecard."""
    from scorecard import compute_scorecard
    fp = _build_workbook(tmp_path, [{
        "D": "SWE1-PROJ-001", "F": "TC-PROJ-001-001", "I": "open the page",
        "L": "1. open", "M": "1. page shown", "P": "P1",
    }])
    report = review_workbook(fp, dry_run=True)
    sc = compute_scorecard(report)
    assert sc.total_tcs == report["batch_meta"]["total_tcs"]
    assert sc.total_requirements == report["batch_meta"]["total_req_groups"]
    # first_pass_rate is computable from the real findings shape.
    assert sc.kpis["first_pass_rate"].denominator == sc.total_tcs


def test_llm_pipeline_injects_content_req(monkeypatch):
    """Stage 6 + content-traceability: the content-matched requirement is fed
    into the review payload so the reviewer anchors on it, not the stale id."""
    tcs = [_make_tc(row_num=10, tc_id="T10", test_item="Repeat All loops to first")]
    captured = {}

    def _fake_chat(system, user, model, json_mode=True, max_tokens=None):
        captured["user"] = user
        from providers import LLMResponse, LLMUsage
        return LLMResponse(text='{"per_req_findings":[],"per_tc_findings":[]}',
                           usage=LLMUsage(), model="fake")

    monkeypatch.setattr("generator._chat", _fake_chat)
    content_map = {10: {"req_id": "SWE1-PLA-006-02", "title": "Repeat All Behavior",
                        "desc": "plays each item sequentially", "score": 0.4}}
    _run_llm_pipeline(tcs, {"REQ-A": ReqGroup(req_id="REQ-A", tcs=tcs)},
                      model="fake", content_map=content_map)
    assert "content_req" in captured["user"]
    assert "SWE1-PLA-006-02" in captured["user"]


def test_llm_pipeline_reports_failed_batches(monkeypatch):
    """Empty/truncated LLM responses (e.g. reasoning model ran out of budget)
    must be COUNTED, not silently swallowed into a regex-only result."""
    tcs = [_make_tc(row_num=10, tc_id="T10"), _make_tc(row_num=11, tc_id="T11")]

    def _empty_chat(system, user, model, json_mode=True, max_tokens=None):
        from providers import LLMResponse, LLMUsage
        return LLMResponse(text="", usage=LLMUsage(), model="fake")  # truncated -> ""

    monkeypatch.setattr("generator._chat", _empty_chat)
    _, _, stats = _run_llm_pipeline(
        tcs, {"REQ-A": ReqGroup(req_id="REQ-A", tcs=tcs)}, model="fake", batch_size=1)
    assert stats["llm_batches"] == 2
    assert stats["llm_failed"] == 2  # both empty -> both counted as failed
