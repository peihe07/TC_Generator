"""Tests for features/sxm/scripts/write_back.py — the BLANK-workbook write.

SXM has no done region and no legacy region, so the invariants Home and AMFM
rely on (an ordered content hash over frozen rows) have nothing to hold. Two
different guards replace them, and both are tested here because both fail
silently in production:

- **the source must be blank.** A workbook that acquired rows between the
  scaffold copy and the write would be appended to, shipping someone else's
  content under our ChangeHistory entry.
- **template residue must not survive.** The blank form ships sample values,
  including `NR1L-AntiTheft-001` in the TC ID column — a *different feature's*
  identifier, in a column this feature does not write and therefore does not
  overwrite. Residue is checked only in unwritten columns, because Functional
  Safety is legitimately `NA` both on the sample row and on every row we write.
"""
import importlib.util
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent

_spec = importlib.util.spec_from_file_location(
    "sxm_write_back", ROOT / "features" / "sxm" / "scripts" / "write_back.py")
if _spec is None or _spec.loader is None:
    pytest.skip("SXM write_back not present", allow_module_level=True)
wb_mod = importlib.util.module_from_spec(_spec)
sys.modules["sxm_write_back"] = wb_mod
_spec.loader.exec_module(wb_mod)


CFG = {
    "workbook": {"header_row": 9, "sheet": "S"},
    "col": {"req_id": 3, "tc_id": 5, "test_group": 6, "test_set": 7,
            "test_item": 8, "pre_conditions": 9, "input_test_data": 10,
            "test_procedure": 11, "expected_result": 12, "spec_reference": 13,
            "tc_ref_id": 14, "priority": 15, "estimated_test_time": 16,
            "design_method": 17, "functional_safety": 18, "author": 26,
            "remarks": 33},
    "write_back": {"author_value": "PeiPYHsu", "tc_ref_id_value": "NEW",
                   "fill_test_group_set": True,
                   "tc_id_format": "NR1L-SXM-{n:03d}",
                   "scope_label": "範圍 Scope", "scope_source": "a03_report"},
}


def sheet(rows=None):
    """A minimal sheet with the header at row 9 and optional data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(5, 3).value = "範圍 Scope："
    ws.cell(9, 4).value = "Requirement or Design ID"
    for r, cells in (rows or {}).items():
        for key, value in cells.items():
            ws.cell(r, CFG["col"][key] + 1).value = value
    return ws


def tc(req_id="SWE-RA-SXM-001", **over):
    base = {"req_id": req_id, "test_group": "SXM", "test_set": "Seek",
            "test_item": "item", "pre_conditions": "1. x",
            "input_test_data": "NA", "test_procedure": "1. a\n2. b",
            "expected_result": "1. c\n2. d",
            "specification_reference": "CFTS024-4872752", "priority": "P1",
            "design_method": "功能測試", "remarks": ""}
    base.update(over)
    return base


# ---------------------------------------------------------- blank-source gate

TEMPLATE = {10: {"req_id": "xxx", "tc_id": "NR1L-AntiTheft-001",
                 "test_group": "AntiTheft", "functional_safety": "NA"},
            11: {"req_id": "xxx"}}


def test_the_forms_own_sample_rows_are_not_authored_content():
    ws = sheet(TEMPLATE)
    assert wb_mod.assert_blank_source(wb_mod.data_rows(ws, CFG))


def test_an_authored_row_aborts_the_write():
    ws = sheet(TEMPLATE | {12: {"req_id": "SWE-RA-SXM-001"}})
    with pytest.raises(wb_mod.WriteBackError, match="authored"):
        wb_mod.assert_blank_source(wb_mod.data_rows(ws, CFG))


def test_a_half_written_row_with_no_req_id_still_aborts():
    """Keying the check on req_id alone would wave this through."""
    ws = sheet(TEMPLATE | {12: {"test_procedure": "1. someone started this"}})
    with pytest.raises(wb_mod.WriteBackError, match="authored"):
        wb_mod.assert_blank_source(wb_mod.data_rows(ws, CFG))


def test_residue_is_scoped_per_column():
    """`NA` is sample content in Functional Safety and nowhere else; treating
    it as a global residue string would tolerate it in req_id."""
    assert wb_mod.is_residue("functional_safety", "NA")
    assert not wb_mod.is_residue("req_id", "NA")
    assert wb_mod.is_residue("req_id", "xxx")
    assert not wb_mod.is_residue("test_item", "xxx")


def test_a_real_row_whose_functional_safety_is_na_is_still_authored():
    ws = sheet(TEMPLATE | {12: {"functional_safety": "NA",
                                "test_item": "a real test item"}})
    with pytest.raises(wb_mod.WriteBackError, match="authored"):
        wb_mod.assert_blank_source(wb_mod.data_rows(ws, CFG))


# ------------------------------------------------------------ residue sweep

WRITTEN = {"req_id", "tc_id", "test_group", "test_set", "test_item",
           "pre_conditions", "input_test_data", "test_procedure",
           "expected_result", "spec_reference", "tc_ref_id", "priority",
           "design_method", "functional_safety", "author", "remarks"}


def test_residue_in_a_column_the_write_does_not_touch_is_caught():
    """The sweep only earns its keep on columns nothing overwrites, so it is
    exercised with tc_id held out of the written set."""
    ws = sheet({10: {"req_id": "SWE-RA-SXM-001",
                     "tc_id": "NR1L-AntiTheft-001"}})
    found = wb_mod.residue_left(ws, CFG, 10, WRITTEN - {"tc_id"})
    assert any("tc_id" in f for f in found)


def test_the_ruled_tc_id_series_overwrites_the_sample_row_identifier():
    """`NR1L-AntiTheft-001` used to survive on row 10 because column F was
    left alone; the ruled series is what now removes it."""
    assert "tc_id" in WRITTEN
    v = wb_mod.cell_values(tc(), CFG, "NR1L-SXM-001")
    assert v[CFG["col"]["tc_id"]] == "NR1L-SXM-001"


def test_our_own_na_in_a_written_column_is_not_reported_as_residue():
    ws = sheet({10: {"req_id": "SWE-RA-SXM-001", "functional_safety": "NA"}})
    assert wb_mod.residue_left(ws, CFG, 10, WRITTEN) == []


def test_a_clean_sheet_reports_no_residue():
    ws = sheet({10: {"req_id": "SWE-RA-SXM-001"}})
    assert wb_mod.residue_left(ws, CFG, 10, WRITTEN) == []


# --------------------------------------------------------------- cell values

def test_written_row_carries_the_ruled_constants():
    v = wb_mod.cell_values(tc(), CFG)
    assert v[CFG["col"]["tc_ref_id"]] == "NEW"
    assert v[CFG["col"]["functional_safety"]] == "NA"
    assert v[CFG["col"]["author"]] == "PeiPYHsu"
    assert v[CFG["col"]["test_group"]] == "SXM"
    assert v[CFG["col"]["test_set"]] == "Seek"


def test_empty_strings_become_none_not_empty_cells():
    assert wb_mod.cell_values(tc(), CFG)[CFG["col"]["remarks"]] is None


def test_a_missing_test_set_aborts_when_fill_is_on():
    with pytest.raises(wb_mod.WriteBackError, match="Test Set"):
        wb_mod.cell_values(tc(test_set=""), CFG)


def test_the_tc_id_column_is_left_alone_without_a_ruled_series():
    """A feature that declares no series gets no invented value."""
    assert CFG["col"]["tc_id"] not in wb_mod.cell_values(tc(), CFG)


# ------------------------------------------------------------- TC ID series

def test_the_series_starts_at_one_and_is_monotonic():
    ids = wb_mod.assign_tc_ids(CFG, [tc()] * 3, set())
    assert ids == ["NR1L-SXM-001", "NR1L-SXM-002", "NR1L-SXM-003"]


def test_the_series_is_unique_across_the_whole_write():
    ids = wb_mod.assign_tc_ids(CFG, [tc()] * 215, set())
    assert len(set(ids)) == 215
    assert ids[-1] == "NR1L-SXM-215"


def test_an_id_already_in_the_sheet_aborts_rather_than_duplicating():
    with pytest.raises(wb_mod.WriteBackError, match="collision"):
        wb_mod.assign_tc_ids(CFG, [tc()] * 3, {"NR1L-SXM-002"})


def test_the_sample_rows_own_id_is_not_a_collision():
    """`NR1L-AntiTheft-001` is a different series — it is cleared, not
    collided with."""
    assert wb_mod.assign_tc_ids(CFG, [tc()], {"NR1L-AntiTheft-001"})


def test_a_feature_without_a_declared_format_gets_no_ids():
    cfg = CFG | {"write_back": {k: v for k, v in CFG["write_back"].items()
                                if k != "tc_id_format"}}
    assert wb_mod.assign_tc_ids(cfg, [tc()] * 3, set()) == [None, None, None]


# ------------------------------------------------------------- Scope filling

def test_scope_is_filled_from_empty():
    ws = sheet()
    result = wb_mod.fix_scope(ws, CFG, "SWE1_SXM_FM-WI-FSM-037-A03")
    assert result["changed"] is True
    assert result["before"] == ""
    assert ws.cell(5, 4).value == "SWE1_SXM_FM-WI-FSM-037-A03"


def test_a_missing_scope_label_aborts():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(5, 3).value = "目的 Purpose："
    with pytest.raises(wb_mod.WriteBackError, match="Scope"):
        wb_mod.fix_scope(ws, CFG, "x")


# --------------------------------------------------------------- integration

def test_the_real_write_holds_every_invariant(tmp_path):
    home = ROOT / "features" / "sxm"
    if not (home / "data" / "stla_to_cfts.json").exists():
        pytest.skip("SXM data/ not built")
    if not list((home / "generated").glob("*.json")):
        pytest.skip("SXM generated/ empty")
    import argparse
    import contextlib
    import io
    args = argparse.Namespace(feature_dir=str(home), data="data",
                              generated="generated", workbook=None,
                              out=None, date=None, write=False)
    buf = io.StringIO()
    try:
        with contextlib.redirect_stdout(buf):
            rc = wb_mod.run(args)
    except SystemExit as exc:
        pytest.skip(f"SXM inputs/ not present: {exc}")
    out = buf.getvalue()
    assert rc == 0, out
    assert "0 authored" in out
    assert "exact match" in out
    assert "DRY RUN" in out
