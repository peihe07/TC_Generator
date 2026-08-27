"""lint036 檢查 A–N 之單元測試（每項一正一反例）。"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

import lint036  # noqa: E402

FIELD_KEYS = ("test_set", "test_item", "pre", "input", "proc", "er", "spec", "author")


def make_fields(**overrides) -> dict[str, str]:
    """建立一列預設全合規的欄位值，再套用覆寫。"""
    base = {
        "test_set": "General Anatomy",
        "test_item": "The system shall display the Media screen\n(Media screen shown)",
        "pre": "1. The Home screen is displayed",
        "input": "NA",
        "proc": "1. Press \"Media\" on the Main Menu Bar",
        "er": "1. The Media screen is displayed",
        "spec": "Media_HMI_Logic_R1.docx",
        "author": "PeiPYHsu",
    }
    base.update(overrides)
    return base


def run(**overrides) -> list[str]:
    """跑單列檢查，回傳觸發的檢查代號清單。"""
    violations = lint036.check_row(make_fields(**overrides), 10, "TC-001",
                                   lint036.DEFAULT_LENGTH_LIMIT)
    return [v.check for v in violations]


def test_clean_row_has_no_violations():
    """預設列應零違規（作為所有反例之基準）。"""
    assert run() == []


# --- header 定位 -------------------------------------------------------------


def build_sheet(header_row: int):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Case Specification 測試用例規範"
    headers = ["No.#\n序號", "Requirement or Design ID\n需求ID", "Test Case ID\n測試用例ID",
               "Test Group\n測試組", "Test Set\n測試集", "Test Item\n測試項目",
               "Pre-Conditions\n先前條件", "Input Test Data\n輸入條件",
               "Test procedure\n測試程序", "Expected Result\n預期結果",
               "Specification Reference \n規格參考", "Test Case Author\n測試案例作者"]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=value)
    return ws


def test_find_header_row_locates_anchor():
    """header 定位：找到含 Specification Reference 之列。"""
    assert lint036.find_header_row(build_sheet(9)) == 9


def test_find_header_row_raises_when_absent():
    """header 定位：前 15 列無錨點時報錯。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="無關內容")
    with pytest.raises(ValueError):
        lint036.find_header_row(ws)


def test_build_column_map_covers_all_field_keys():
    """欄位對照：八個欄位鍵全數命中，且 tc_id/req_id 可定位。"""
    ws = build_sheet(9)
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    columns = lint036.build_column_map(list(rows[8]))
    assert set(lint036.FIELD_HEADERS) <= set(columns)
    assert columns["tc_id"] == 2
    assert columns["req_id"] == 1
    assert columns["test_set"] == 4


# --- A–N 逐項 ----------------------------------------------------------------


def test_a_flags_forbidden_verb():
    assert "A" in run(proc="1. Observe the Media screen.")


def test_a_allows_imperative_step():
    assert "A" not in run(proc="1. Press the Media icon.")


def test_b_flags_modal_in_expected_result():
    assert "B" in run(er="1. The screen shall be displayed.")


def test_b_exempts_modal_inside_quotes():
    assert "B" not in run(er='1. The label "system shall reboot" is displayed.')


def test_c_flags_hedge_word_in_paren_half():
    """R-6b：括號下半為作者生成內容，hedge 仍受規制。"""
    assert "C" in run(test_item="The system displays the screen\n"
                                "(read the screen -> The screen loads properly)")


def test_c_allows_specific_wording():
    assert "C" not in run(test_item="The system displays the screen\n"
                                    "(read the screen -> The Media screen is shown)")


def test_c_exempts_hedge_in_verbatim_upper_half():
    """R-6b：上半為需求原句 verbatim，其用語屬來源文件，不受 C 規制。"""
    assert "C" not in run(test_item="TLM is able to work properly again\n"
                                    "(read the state -> The TLM restores the values)")


def test_d_flags_powered_on_precondition():
    assert "D" in run(pre="1. The HU is powered on.")


def test_d_flags_action_verb_in_precondition():
    assert "D" in run(pre="1. Insert a USB drive.")


def test_d_allows_state_precondition():
    assert "D" not in run(pre="1. The Home screen is displayed.")


def test_e_flags_step_count_mismatch():
    assert "E" in run(proc="1. Press A.\n2. Press B.", er="1. Screen shown.")


def test_e_allows_matching_step_counts():
    assert "E" not in run(proc="1. Press A.\n2. Press B.",
                          er="1. Screen A shown.\n2. Screen B shown.")


def test_f_flags_bracket_placeholder():
    assert "F" in run(proc="1. Press [Media] on the bar.")


def test_f_allows_quoted_label():
    assert "F" not in run(proc='1. Press "Media" on the bar.')


def test_g_flags_empty_test_set():
    assert "G" in run(test_set="   ")


def test_g_allows_filled_test_set():
    assert "G" not in run(test_set="General Anatomy")


def test_h_flags_vague_expected_result():
    assert "H" in run(er="1. The screen works normally.")


def test_h_allows_concrete_expected_result():
    assert "H" not in run(er="1. The Media screen is displayed.")


def test_i_flags_missing_paren_half():
    assert "I" in run(test_item="The system shall display the Media screen.")


def test_i_allows_paren_line():
    assert "I" not in run(test_item="The system displays it.\n(Media screen shown)")


def test_i_allows_trailing_paren():
    assert "I" not in run(test_item="The system displays it (Media screen shown)")


def test_i_sibling_flags_verbatim_duplicate_paren_lines():
    rows = [
        (10, "TC-001", "REQ-1", "Item A.\n(same paren)"),
        (11, "TC-002", "REQ-1", "Item B.\n(same paren)"),
    ]
    assert len(lint036.check_sibling_parens(rows)) == 2


def test_i_sibling_allows_distinct_paren_lines():
    rows = [
        (10, "TC-001", "REQ-1", "Item A.\n(paren one)"),
        (11, "TC-002", "REQ-1", "Item B.\n(paren two)"),
    ]
    assert lint036.check_sibling_parens(rows) == []


def test_j_flags_lowercase_line_start():
    assert "J" in run(er="1. the Media screen is displayed")


def test_j_exempts_numeric_first_token():
    """00b 修訂 2：行號後第一個 token 為數字者整行豁免，不再往後尋找。"""
    assert "J" not in run(er="4. 5 sources are displayed in the Source Menu")


def test_j_exempts_symbol_first_token():
    """第一個 token 以 $ 開頭者整行豁免。"""
    assert "J" not in run(proc="1. $Telematic_Power$ is read")


def test_j_exempts_whitelisted_token():
    assert "J" not in run(proc="1. adb shell dumpsys media.")


def test_j_exempts_camel_case_token():
    assert "J" not in run(proc="1. mediaPlayer is invoked.")


def test_j_exempts_dotted_call():
    assert "J" not in run(proc="1. media.play() is invoked.")


def test_k_flags_cjk_characters():
    assert "K" in run(er="1. 媒體畫面顯示.")


def test_k_allows_ascii_only():
    assert "K" not in run(er="1. The Media screen is displayed")


def test_k_covers_test_set_column():
    """00b 修訂 3：test_set 屬六欄之一。"""
    assert "K" in run(test_set="一般外觀")


def test_k_excludes_spec_column():
    """00b 修訂 3：spec 不屬六欄。"""
    assert "K" not in run(spec="媒體規格_R1.docx")


def test_k_excludes_author_column():
    """00b 修訂 3：author 不屬六欄。"""
    assert "K" not in run(author="許沛")


def test_l_flags_overlong_upper_half():
    long_item = " ".join(["word"] * 60) + ".\n(paren)"
    assert "L" in run(test_item=long_item)


def test_l_ignores_paren_lines_in_length():
    item = " ".join(["word"] * 40) + ".\n(" + " ".join(["x"] * 40) + ")"
    assert "L" not in run(test_item=item)


def test_m_flags_empty_required_column():
    assert "M" in run(spec="")


def test_m_allows_na_placeholder():
    assert "M" not in run(spec="NA")


def test_m_allows_pending_marker():
    assert "M" not in run(spec="PENDING: awaiting spec id")


def test_n_flags_trailing_period():
    """00b 修訂 1：命中 `[.。]$` 即違規（canon §11 禁尾句號）。"""
    assert "N" in run(er="1. The Media screen is displayed.")


def test_n_flags_trailing_cjk_period():
    assert "N" in run(er="1. The Media screen is displayed。")


def test_n_allows_line_without_period():
    assert "N" not in run(er="1. The Media screen is displayed")


def test_n_allows_na_placeholder():
    """`NA` 不以句號結尾，正確方向下零命中（推翻 00 包所稱 M/N 衝突）。"""
    assert "N" not in run(input="NA")


def test_n_exempts_shell_command_line():
    assert "N" not in run(proc="1. Run the command\n$ adb shell dumpsys media.")


def test_n_exempts_indented_continuation():
    assert "N" not in run(proc="1. Press the icon\n    continued fragment.")


# --- 00c 裁定 1：N 子步驟納入，且不得外溢至 E ---


def test_n_flags_lettered_substep_with_period():
    """a./b./c. 縮排子步驟為實質測試步驟（canon §6.1），尾句號計入 N。"""
    proc = "1. Power cycle the HU\n   a. Hold H/K[POWER] for 2 secs."
    assert "N" in run(proc=proc)


def test_n_flags_lettered_substep_with_paren_marker():
    """`a)` 形式之子步驟同樣納入 N 行定義。"""
    assert "N" in run(proc="1. Power cycle the HU\n   a) Hold POWER for 2 secs.")


def test_n_still_exempts_true_continuation():
    """真續行（無 a./b. 標記）維持豁免。"""
    proc = "1. Press the icon\n   duplicate contact entry is created."
    assert "N" not in run(proc=proc)


def test_e_ignores_lettered_substeps():
    r"""迴歸：E 沿用全域 `^\s*\d+[.)]`，不因 N 行定義擴充而漂移。"""
    proc = ("1. Power cycle the HU\n   a. Hold POWER for 2 secs\n"
            "   b. Press MUTE for 1 sec\n2. Observe nothing")
    er = "1. The HU restarts\n2. The screen is shown"
    violations = lint036.check_row(make_fields(proc=proc, er=er), 10, "TC-001",
                                   lint036.DEFAULT_LENGTH_LIMIT)
    assert "E" not in [v.check for v in violations]


def test_numbered_lines_helper_excludes_lettered_substeps():
    """全域 numbered_lines() 不得納入 a./b. —— N 之擴充限定於 n_exempt()。"""
    text = "1. Step one\n   a. Sub step\n2. Step two"
    assert len(lint036.numbered_lines(text)) == 2
    assert lint036.N_STEP_LINE.match("   a. Sub step")
    assert not lint036.NUMBERED_LINE.match("   a. Sub step")


# --- 03 §三 R-6：P 訊號記法之施用範圍 ---


def test_p_flags_withdrawn_triplet():
    """R-1 v1 之三件組已撤銷，殘留即違規。"""
    assert "P" in run(proc="1. Drive Radio_btn0 in CLIMATIC_PANEL on BH-CAN to Pressed")


def test_p_allows_v2_send_form():
    """R-1 v2(a)：`Send CAN: <MSG>.<Sig> = <raw> (<label>)`。"""
    assert "P" not in run(proc="1. Send CAN: BCM_FD_14.Command_02Sts = 1 (PSD)")


def test_p_flags_assignment_without_label():
    """R-7：值須帶 DBC `VAL_` 括號標籤。"""
    assert "P" in run(proc="1. Send CAN: BCM_FD_14.Command_02Sts = 1")


def test_p_flags_procedure_assignment_without_send_prefix():
    assert "P" in run(proc="1. Set BCM_FD_14.Command_02Sts = 1 (PSD)")


def test_p_allows_multiple_assignments_on_one_line():
    """SWC 一行可載多個賦值，逐出現判定不得誤報。"""
    assert "P" not in run(
        proc="1. Send CAN: BCM_FD_14.Command_01Sts = 1 (Pressed) and "
             "BCM_FD_14.Command_03Sts = 0 (Not_Pressed)")


def test_p_er_does_not_constrain_trailing_wording():
    """ER 之收尾語於基準本不固定（is sent／is set／during …），不得設限。"""
    assert "P" not in run(er="2. BCM_FD_14.Command_09Sts = 0 (NOT_PSD) is set after release")
    assert "P" not in run(er="2. BCM_FD_14.Command_09Sts = 1 (PSD) during press window")


def test_p_ignores_internal_signal_notation():
    """內部訊號 `X.Info`／`X.Req` 之 message 段含小寫，不得誤判為 CAN。"""
    assert "P" not in run(er="1. TLM_Status.Info reads \"Standby\"",
                          pre="1. Phone_Call.Info reads \"Not_Active\"")


def test_p_ignores_proxi_parameter():
    """PROXI 層 `$X$` 無兩段記法，且不得被套三件組（A-PM03）。"""
    assert "P" not in run(proc="1. Read $Radio_Theme$ on the bench")


def test_p_exempts_test_item_verbatim_upper_half():
    """R-6：test_item 上半為需求原句 verbatim，其記法保留來源原文。"""
    item = ("When STATUS_LIN.PN14_LS_Actv = 1 the TLM shall react\n"
            "(read the volume -> The maximum volume is reduced)")
    assert "P" not in run(test_item=item)


def test_p_flags_legacy_notation_in_test_item_paren():
    """R-6：括號下半屬作者生成內容，仍受 R-1 v2 規制。"""
    item = ("The TLM shall react to the load shed signal\n"
            "(drive PN14_LS_Actv in STATUS_LIN on BH-CAN -> The volume is reduced)")
    assert "P" in run(test_item=item)


# --- 計數口徑 ---


def test_row_counts_distinguish_line_and_row_basis():
    """同一列多行違規：行計 > 列計。"""
    result = lint036.SheetResult(sheet="s", header_row=9, data_rows=1)
    result.violations = lint036.check_row(
        make_fields(er="1. the screen is shown\n2. the icon is shown"),
        10, "TC-001", lint036.DEFAULT_LENGTH_LIMIT)
    assert lint036.count_by_check([result])["J"] == 2
    assert lint036.rows_by_check([result])["J"] == 1


def test_every_check_has_status_and_granularity():
    """報告表頭所需之三張對照表須涵蓋全部檢查（含 profile 專屬者）。"""
    full = set(lint036.check_order("power"))
    assert set(lint036.CHECK_STATUS) == full
    assert set(lint036.CHECK_GRANULARITY) == full
    assert set(lint036.CHECK_TITLES) == full
    # profile 覆寫表不得引入 CHECK_ORDER 以外之代號
    assert set(lint036.CHECK_TITLE_PROFILE) <= full
    assert set(lint036.CHECK_STATUS_PROFILE) <= full


# --- 報告命名 ----------------------------------------------------------------


def test_report_stem_strips_date_and_annotation():
    path = Path("FM-WI-FSM-036-A01 x_SWQT_CFTS012_DealerMode_20260417(done).xlsx")
    assert lint036.report_stem(path) == "CFTS012_DealerMode"


# --- 日期起首之 tag 併入身分（25 包 §D-6）-----------------------------------
#
# 缺陷之原文以**字面**釘入（G-N）：036 母本副本之檔名於 `SWQT_` 之後
# 直接是日期，去日期後 tag 為 `20260817_ext`，三個 feature 之報告互相覆寫。

_MASTER = ("FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT STLA "
           "Test Case Specification & Result_SWQT_20260817_ext.xlsx")


def test_date_only_tag_collides_before_identity_is_added(tmp_path):
    """缺陷之形狀：三份不同 feature 之母本副本，其 `SWQT_` 段只有日期。"""
    bare = [Path(_MASTER)]
    # 未帶目錄脈絡時，tag 只能是日期段本身 —— 這正是覆寫之成因
    assert lint036.report_stem(bare[0]).endswith("20260817_ext")


def test_identity_dir_disambiguates_the_three_features(tmp_path):
    """修正後不得再命中（G-N 之回歸向）：三份之 tag 須相異。"""
    tags = set()
    for feature in ("user_profiles", "time_management", "power_moding"):
        p = tmp_path / "features" / feature / "inputs" / _MASTER
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_bytes(b"")
        tags.add(lint036.report_stem(p))
    assert tags == {
        "user_profiles_20260817_ext",
        "time_management_20260817_ext",
        "power_moding_20260817_ext",
    }, tags


def test_identity_falls_back_to_nearest_non_generic_dir(tmp_path):
    """無 `features/` 者取最近之非通用容器祖先（`docs/test/<name>/SWE6/`）。"""
    name = ("FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT STLA "
            "Test Case Specification & Result_SWQT_20260121.xlsx")
    got = {}
    for area in ("Dealer Mode", "Player"):
        p = tmp_path / "docs" / "test" / area / "SWE6" / name
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_bytes(b"")
        got[area] = lint036.report_stem(p)
    assert got == {"Dealer Mode": "Dealer_Mode_20260121",
                   "Player": "Player_20260121"}, got


@pytest.mark.parametrize("stem, expected", [
    # 範圍向（R-G9）：非日期起首者 tag 一律不變
    ("FM-WI-FSM-036-A01 x_SWQT_AMFM_20260821", "AMFM"),
    ("FM-WI-FSM-036-A01 x_SWQT_Home_20260809", "Home"),
    ("FM-WI-FSM-036-A01 x_SWQT_CFTS012_DealerMode_20260417(done)", "CFTS012_DealerMode"),
    ("FM-WI-FSM-036-A01 x_SWQT_PowerManagement_20260821(Revise)", "PowerManagement"),
    ("FM-WI-FSM-036-A01 x_SWQT_UserProfiles_20260820_itemgap", "UserProfiles"),
])
def test_named_tags_are_untouched_by_the_fix(tmp_path, stem, expected):
    p = tmp_path / "features" / "amfm" / "inputs" / f"{stem}.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_bytes(b"")
    assert lint036.report_stem(p) == expected, "feature 名不得混入已具身分之 tag"


# --- 報告檔名 v2（26 包 §C 裁定 3）------------------------------------------
#
# 25 包之回歸向為「既有報告**檔名**不變」。裁定 3 改其方向：
# 新產報告採 `{tag}_{來源檔sha8}_{YYYYMMDD}`，故新檔名本就會變；
# **回歸判準改為「既有報告檔案不被重命名」** —— 工具不得碰 report_dir
# 內既存之檔。二者並存：tag 之範圍向（上）＋ 既有檔案之不動（下）。

EXISTING_REPORTS = [                      # 字面釘入（G-N）：現存之既有報告檔名
    "AMFM_20260821.md", "AMFM_20260821.json",
    "Home_20260821.md", "CFTS012_DealerMode_20260821.md",
    "PowerManagement_20260821.json", "Projection_20260821.md",
    "pm_25__power_20260824.md",
]


def test_source_sha8_is_content_addressed(tmp_path):
    """sha8 取自位元組 —— 檔名可以改，位元組不會。"""
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    a.write_bytes(b"same-bytes")
    b.write_bytes(b"same-bytes")
    assert lint036.source_sha8(a) == lint036.source_sha8(b)
    b.write_bytes(b"other-bytes")
    assert lint036.source_sha8(a) != lint036.source_sha8(b)


def test_source_sha8_declares_its_own_absence(tmp_path):
    """讀不到者回 `nosha`，不回退為空字串（空字串會讓檔名退回舊式而看似正常）。"""
    assert lint036.source_sha8(tmp_path / "missing.xlsx") == "nosha"


def test_same_tag_different_content_no_longer_collides(tmp_path):
    """字面案例：同一 feature 之兩個來源日期，其 tag 相同而內容不同。

    25 包實測 `PowerManagement` 有 12 檔共用一個 tag —— 同日 lint 即互相覆寫。
    """
    names = set()
    for day, content in (("20260816", b"v1"), ("20260820", b"v2"), ("20260821", b"v3")):
        p = (tmp_path / "features" / "power" / "inputs" /
             f"FM-WI-FSM-036-A01 x_SWQT_PowerManagement_{day}.xlsx")
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_bytes(content)
        assert lint036.report_stem(p) == "PowerManagement"      # tag 仍相同
        names.add(f"{lint036.report_stem(p)}_{lint036.source_sha8(p)}")
    assert len(names) == 3, f"三份相異內容須得三個相異檔名，實得 {names}"


def test_identical_copies_share_one_report_name(tmp_path):
    """範圍向（R-G9）：位元組相同之多處副本**應**共用檔名，不得被當成碰撞。"""
    names = set()
    for feature in ("comfort", "time_management"):
        p = tmp_path / "features" / feature / "inputs" / "SR24 Table v1.6.xlsx"
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_bytes(b"identical")
        names.add(f"{lint036.report_stem(p)}_{lint036.source_sha8(p)}")
    assert len(names) == 1


@pytest.mark.parametrize("existing", EXISTING_REPORTS)
def test_existing_report_files_are_never_renamed(existing):
    """裁定 3 之回歸判準：既有報告**檔案**不得被重命名。

    工具只寫新檔名，從不移動或改名 `report_dir` 內既存之檔 ——
    以原始碼為證：`lint036` 全檔無任何 rename／move／unlink 呼叫。
    """
    root = Path(__file__).resolve().parent.parent
    src = (root / "scripts" / "lint036.py").read_text(encoding="utf-8")
    for forbidden in (".rename(", ".replace(Path", "shutil.move", ".unlink(", "os.remove"):
        assert forbidden not in src, f"lint036 不得改動既存檔案：發現 {forbidden}"
    report = root / "docs" / "fw036" / "lint_reports" / existing
    if report.exists():
        assert report.name == existing        # 檔名逐字不變


# --- profile 專屬檢查（21 包：`--profile <feature>`）-------------------------


def run_profile(**overrides) -> list[str]:
    """以 profile 模式跑單列檢查，回傳觸發的檢查代號清單。"""
    violations = lint036.check_row(make_fields(**overrides), 10, "TC-001",
                                   lint036.DEFAULT_LENGTH_LIMIT,
                                   profile="power")
    return [v.check for v in violations]


def test_profile_off_by_default() -> None:
    """未指定 profile 時不跑 Q／R／T —— 既有八本之基線不動。"""
    dirty = make_fields(pre="1. NBSP\xa0here\nunnumbered line")
    checks = [v.check for v in lint036.check_row(
        dirty, 10, "TC-001", lint036.DEFAULT_LENGTH_LIMIT)]
    assert "Q" not in checks and "R" not in checks and "T" not in checks


def test_check_order_extends_only_with_profile() -> None:
    assert lint036.check_order(None) == lint036.CHECK_ORDER
    assert lint036.check_order("power") == \
        lint036.CHECK_ORDER + ["Q", "R", "T", "U", "V"]


# Q —— 不可見字元（R-10(a)），全欄位含 verbatim 上半

def test_q_nbsp_in_verbatim_half() -> None:
    assert "Q" in run_profile(
        test_item="The system\xa0shall display it\n(shown)")


def test_q_ideographic_space_and_trailing_ws() -> None:
    assert "Q" in run_profile(proc="1. Press　the button")
    assert "Q" in run_profile(er="1. The screen is shown   ")


def test_q_clean_row_passes() -> None:
    assert "Q" not in run_profile()


# V —— 行首空白（IN §11，27 包 §D-4）

def test_v_leading_space_on_body_and_numbered_lines() -> None:
    assert "V" in run_profile(test_item="  The system shall display it\n\n(shown)")
    assert "V" in run_profile(proc="1. Press the button\n 2. Release it")
    assert "V" in run_profile(er="\t1. The screen is shown")


def test_v_whitespace_only_line() -> None:
    assert "V" in run_profile(pre="1. The TLM is in Idle state\n   \n2. Ignition On")


def test_v_in_6_1_and_5_4_indents_are_exempt() -> None:
    """IN §11 之唯二例外不得判紅（G-9 範圍向）。"""
    assert "V" not in run_profile(proc="1. Press the button\n   a. Hold for 3 s")
    assert "V" not in run_profile(proc="1. Press the button\n      - the LED turns on")
    assert "V" not in run_profile(proc="1. Run the tool\n   $ adb shell dumpsys")


def test_v_near_miss_indents_still_red() -> None:
    """例外是**定格**，不是「有縮排就放行」—— 格數或記號不符者照紅。"""
    assert "V" in run_profile(proc="1. Press the button\n  a. Hold for 3 s")
    assert "V" in run_profile(proc="1. Press the button\n    - the LED turns on")
    assert "V" in run_profile(proc="1. Press the button\n   b) Hold for 3 s")


def test_v_clean_row_passes() -> None:
    assert "V" not in run_profile()


def test_v_does_not_double_count_trailing_ws_with_q() -> None:
    """行尾空白屬 Q，V 不重複計 —— 否則量化矩陣之命中數雙倍膨脹。"""
    checks = run_profile(er="1. The screen is shown   ")
    assert "Q" in checks and "V" not in checks


# R —— Pre-Condition 版面（R-9(a)）

def test_r_unnumbered_line() -> None:
    assert "R" in run_profile(
        pre="1. The TLM is in Idle state\nThe ignition is On")


def test_r_multiple_conditions_on_one_line() -> None:
    assert "R" in run_profile(
        pre="1. An SDCARD is inserted and a BT device is connected")


def test_r_tool_line_is_not_a_multi_condition() -> None:
    """工具行本身含 `and`，不得誤判（R-12(a) 之固定措辭）。"""
    assert "R" not in run_profile(
        pre="1. The TLM is in Idle state\n2. LIN and CAN tool is available on HU")


# T —— PENDING 說明之語言（R-14）

def test_t_non_ascii_pending_description() -> None:
    assert "T" in run_profile(
        proc="1. Read the screen and check that PENDING: DR-PW22 擇一判準")


def test_t_english_pending_description_passes() -> None:
    assert "T" not in run_profile(
        proc="1. Read the screen and check that PENDING: DR-PW22 "
             "(which of the two is shown)")


# P —— R-1 v3 判準（profile 專屬，取代 v2）

def test_p_v3_rejects_send_can_prefix() -> None:
    assert "P" in run_profile(
        proc="1. Send CAN: STATUS_BH_BCM2.RemStActvSts = 1 (Remote Start Active)")


def test_p_v3_rejects_bare_can_token_assignment() -> None:
    assert "P" in run_profile(
        proc="1. Set STATUS_BH_BCM1.OperationalModeSts = 2 (Ignition_Off)")


def test_p_v3_rejects_assignment_without_val_label() -> None:
    assert "P" in run_profile(
        proc="1. Send the signal $STATUS_BH_BCM1.OperationalModeSts$ = 2")


def test_p_v3_rejects_proxi_with_dollar() -> None:
    assert "P" in run_profile(pre='1. PROXI $Rear_View_Camera$ = "Present"')


def test_p_v3_accepts_canonical_form() -> None:
    assert "P" not in run_profile(
        pre="1. PROXI Rear_View_Camera = Present",
        proc="1. Send the signal $STATUS_BH_BCM1.OperationalModeSts$ = 2 "
             "(Ignition_Off)\n"
             "2. Read the signal $STATUS_TELEMATIC.PowerSts_Telematic$ and "
             "check that it is 1 (Standby)",
        er="1. The signal value $STATUS_TELEMATIC.PowerSts_Telematic$ = "
           "1 (Standby) is received")


def test_p_v3_allows_pending_placeholder_as_value() -> None:
    """R-14 佔位不視為缺 VAL_ 標籤 —— 其缺件由 M／T 承接。"""
    assert "P" not in run_profile(
        proc="1. Send the signal $STATUS_BH_BCM1.OperationalModeSts$ = "
             "PENDING: DR-PW20")


def test_p_v3_still_rejects_v1_triplet() -> None:
    assert "P" in run_profile(
        proc="1. Drive Radio_btn0 in CLIMATIC_PANEL on BH-CAN to Pressed")


# U —— PENDING 佔位之可見性（A-PM16：ER 側原不受任何檢查覆蓋）

def test_u_counts_pending_on_the_er_side() -> None:
    checks = run_profile(er="1. PENDING: DR-PW22 (which of the two is shown)")
    assert "U" in checks and "T" not in checks


def test_u_counts_pending_in_procedure_too() -> None:
    assert "U" in run_profile(
        proc="1. Send the signal $STATUS_BH_BCM1.OperationalModeSts$ = "
             "PENDING: DR-PW20")


def test_u_silent_when_no_placeholder() -> None:
    assert "U" not in run_profile()
