"""Tests for Excel writer module (RULES.md §10.3)."""
import os
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment

from writer import (
    write_generated_results,
    write_framework_sheet,
    build_output_path,
)


@pytest.fixture
def input_xlsx(tmp_path):
    """Create a minimal input xlsx to write results into."""
    filepath = tmp_path / "SomeProject_SWQT_DeviceManager_20260408.xlsx"
    wb = Workbook()

    # Product Document sheet
    ws_pd = wb.active
    ws_pd.title = "Product Document"
    ws_pd.cell(row=3, column=2, value="newR1L")

    # TC sheet with header + 2 data rows
    ws_tc = wb.create_sheet("Test Case Specification&Result")
    headers = {
        4: "Requirement or Design ID", 6: "Test Case ID", 7: "Test Group",
        8: "Test Set", 9: "Test Item", 10: "Pre-Conditions",
        11: "Input Test Data", 12: "Test Procedure", 13: "Expected Result",
        14: "Specification Reference", 16: "Test Case Priority",
        17: "Test Case Design Method",
    }
    for col, name in headers.items():
        ws_tc.cell(row=9, column=col, value=name)

    # Row 10: has req_id and original test item
    ws_tc.cell(row=10, column=4, value="SWE1-HMI-DM-001-01")
    ws_tc.cell(row=10, column=9, value="PDM01.1) Original text here.")

    # Row 11: second row
    ws_tc.cell(row=11, column=4, value="SWE1-HMI-DM-002-01")
    ws_tc.cell(row=11, column=9, value="PDM02) Another original text.")

    # Untouched column B with data (should be preserved)
    ws_tc.cell(row=10, column=2, value="SWE6")
    ws_tc.cell(row=11, column=2, value="SWE6")

    # Test Case Framework sheet (empty)
    wb.create_sheet("Test Case Framework")

    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def generated_rows():
    return [
        {
            "row_num": 10,
            "tc_id": "newR1L-DMR-001",
            "test_group": "DeviceManager",
            "test_set": "Access & Entry",
            "tc_title": "(User adds DM → DM icon displayed)",
            "pre_conditions": "1. BT enabled on HU",
            "input_test_data": "Status bar menu",
            "test_procedure": "1. Open settings to access.\n2. Add DM and verify icon appears.",
            "expected_result": "1. Settings shown.\n2. DM icon displayed.",
            "spec_reference": "Device_Manager_HMI Logic_2.3",
            "priority": "P1",
            "design_method": "功能測試 (Functional based ; no specific technique)",
        },
        {
            "row_num": 11,
            "tc_id": "newR1L-DMR-002",
            "test_group": "DeviceManager",
            "test_set": "Access & Entry",
            "tc_title": "(User opens DM from menu → DM screen shown)",
            "pre_conditions": "NA",
            "input_test_data": "Menu bar icon",
            "test_procedure": "1. Tap menu bar to open.\n2. Select DM and verify it opens.",
            "expected_result": "1. Menu bar expanded.\n2. DM main screen displayed.",
            "spec_reference": "Device_Manager_HMI Logic_3.1",
            "priority": "P1",
            "design_method": "功能測試 (Functional based ; no specific technique)",
        },
    ]


@pytest.fixture
def framework_data():
    return [
        {"test_group": "DeviceManager", "test_set": "Access & Entry", "req_count": 5},
        {"test_group": "DeviceManager", "test_set": "Device List", "req_count": 3},
    ]


class TestBuildOutputPath:
    def test_default(self):
        path = build_output_path("/data/input.xlsx")
        assert path.endswith("input_generated.xlsx")

    def test_output_dir(self):
        path = build_output_path("/data/input.xlsx", output_dir="/out")
        assert path.startswith("/out/")
        assert path.endswith("input_generated.xlsx")

    def test_strips_macos_duplicate_suffix_tc(self):
        # macOS Finder duplicates as "foo 拷貝.xlsx" — must not leak into output.
        path = build_output_path("/data/Project_SWQT_DeviceManager_20260408 拷貝.xlsx")
        assert path.endswith("Project_SWQT_DeviceManager_20260408_generated.xlsx")

    def test_strips_macos_duplicate_suffix_no_space(self):
        # Chinese duplicate marker without separating space.
        path = build_output_path("/data/foo拷貝.xlsx")
        assert path.endswith("foo_generated.xlsx")

    def test_strips_windows_copy_suffix(self):
        path = build_output_path("/data/report - Copy.xlsx")
        assert path.endswith("report_generated.xlsx")

    def test_strips_windows_copy_with_index(self):
        path = build_output_path("/data/report - Copy (2).xlsx")
        assert path.endswith("report_generated.xlsx")

    def test_strips_macos_english_copy(self):
        path = build_output_path("/data/foo copy 3.xlsx")
        assert path.endswith("foo_generated.xlsx")

    def test_keeps_clean_basename_unchanged(self):
        path = build_output_path("/data/Project_SWQT_DeviceManager_20260408.xlsx")
        assert path.endswith("Project_SWQT_DeviceManager_20260408_generated.xlsx")

    def test_does_not_double_append_generated_suffix(self):
        path = build_output_path("/data/foo_generated.xlsx")
        assert path.endswith("foo_generated.xlsx")
        assert not path.endswith("foo_generated_generated.xlsx")

    def test_collapses_repeated_generated_suffixes(self):
        path = build_output_path("/data/foo_generated_generated.xlsx")
        assert path.endswith("foo_generated.xlsx")


class TestWriteGeneratedResults:
    def test_writes_tc_id(self, input_xlsx, generated_rows, tmp_path):
        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, generated_rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        assert ws.cell(row=10, column=6).value == "newR1L-DMR-001"
        assert ws.cell(row=11, column=6).value == "newR1L-DMR-002"

    def test_writes_test_group(self, input_xlsx, generated_rows, tmp_path):
        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, generated_rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        assert ws.cell(row=10, column=7).value == "DeviceManager"

    def test_appends_tc_title(self, input_xlsx, generated_rows, tmp_path):
        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, generated_rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        cell_value = ws.cell(row=10, column=9).value
        # Original text preserved, rewrite appended
        assert "PDM01.1) Original text here." in cell_value
        assert "(User adds DM → DM icon displayed)" in cell_value

    def test_each_row_gets_its_own_rewrite_even_when_req_id_shared(self, input_xlsx, tmp_path):
        """Per §1.2：同 req 下每筆 TC 必須有自己帶 scenario tag 的 rewrite，
        不再跨列 dedup。"""
        rows = [
            {
                "row_num": 10,
                "req_id": "SWE1-HMI-DM-001-01",
                "tc_id": "newR1L-DMR-001",
                "tc_title": "(Trigger → Outcome A)",
            },
            {
                "row_num": 11,
                "req_id": "SWE1-HMI-DM-001-01",
                "tc_id": "newR1L-DMR-002",
                "tc_title": "(Trigger → Outcome B)",
            },
        ]
        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        assert "(Trigger → Outcome A)" in ws.cell(row=10, column=9).value
        # Row 11 也要有自己的 rewrite
        row11_value = ws.cell(row=11, column=9).value or ""
        assert "(Trigger → Outcome B)" in row11_value
        # 不應該混到 A
        assert "Outcome A" not in row11_value
        assert "PDM02) Another original text." in row11_value

    def test_rewrite_applied_once_per_distinct_req_id(self, input_xlsx, tmp_path):
        """Different req_ids each get their own rewrite even if text matches."""
        rows = [
            {
                "row_num": 10,
                "req_id": "SWE1-HMI-DM-001-01",
                "tc_id": "newR1L-DMR-001",
                "tc_title": "(First → A)",
            },
            {
                "row_num": 11,
                "req_id": "SWE1-HMI-DM-002-01",  # different req_id
                "tc_id": "newR1L-DMR-002",
                "tc_title": "(Second → B)",
            },
        ]
        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        assert "(First → A)" in ws.cell(row=10, column=9).value
        assert "(Second → B)" in ws.cell(row=11, column=9).value

    def test_writes_procedure(self, input_xlsx, generated_rows, tmp_path):
        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, generated_rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        assert "Open settings" in ws.cell(row=10, column=12).value

    def test_preserves_template_wrap_text_setting(self, input_xlsx, generated_rows, tmp_path):
        wb = load_workbook(input_xlsx)
        ws = wb["Test Case Specification&Result"]
        ws.cell(row=10, column=12).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=11, column=12).alignment = Alignment(wrap_text=False, vertical="center")
        wb.save(input_xlsx)

        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, generated_rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        assert ws.cell(row=10, column=12).alignment.wrap_text is True
        assert ws.cell(row=10, column=12).alignment.vertical == "center"
        assert ws.cell(row=11, column=12).alignment.wrap_text is not True
        assert ws.cell(row=11, column=12).alignment.vertical == "center"

    def test_preserves_untouched_columns(self, input_xlsx, generated_rows, tmp_path):
        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, generated_rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        # Col B should be untouched
        assert ws.cell(row=10, column=2).value == "SWE6"
        # Col D (req_id) should be untouched
        assert ws.cell(row=10, column=4).value == "SWE1-HMI-DM-001-01"

    def test_writes_all_generated_columns(self, input_xlsx, generated_rows, tmp_path):
        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, generated_rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        row = 10
        assert ws.cell(row=row, column=10).value is not None  # Pre-Conditions (J)
        assert ws.cell(row=row, column=11).value is not None  # Input Test Data (K)
        assert ws.cell(row=row, column=13).value is not None  # Expected Result (M)
        assert ws.cell(row=row, column=14).value is not None  # Spec Reference (N)
        assert ws.cell(row=row, column=16).value == "P1"  # Priority (P)
        assert "功能測試" in ws.cell(row=row, column=17).value  # Design Method (Q)

    def test_replaces_existing_rewrite_instead_of_appending_again(self, input_xlsx, generated_rows, tmp_path):
        first_output = str(tmp_path / "first.xlsx")
        second_output = str(tmp_path / "second.xlsx")

        write_generated_results(input_xlsx, generated_rows, first_output)
        write_generated_results(first_output, generated_rows, second_output)

        wb = load_workbook(second_output)
        ws = wb["Test Case Specification&Result"]
        cell_value = ws.cell(row=10, column=9).value
        assert cell_value.count("(User adds DM → DM icon displayed)") == 1

    def test_clears_stale_optional_fields(self, input_xlsx, tmp_path):
        output = str(tmp_path / "output.xlsx")
        rows = [{
            "row_num": 10,
            "tc_id": "newR1L-DMR-001",
            "test_group": "DeviceManager",
            "test_set": "Access & Entry",
            "tc_title": "(User adds DM → DM icon displayed)",
            "pre_conditions": "NA",
            "input_test_data": "NA",
            "test_procedure": "1. Open settings.\n2. Verify the result.",
            "expected_result": "1. Settings shown.\n2. Result verified.",
            "spec_reference": "Spec_1",
            "priority": "P1",
            "design_method": "功能測試 (Functional based ; no specific technique)",
        }]
        write_generated_results(input_xlsx, rows, output)

        rows[0]["test_set"] = None
        rows[0]["spec_reference"] = None
        write_generated_results(output, rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        assert ws.cell(row=10, column=8).value is None
        assert ws.cell(row=10, column=14).value is None

    def test_writes_reasoning_to_column_r_for_primary_tc(self, input_xlsx, generated_rows, tmp_path):
        # B 方案：AI 解讀寫到 column R（18）。Primary TC 帶 reasoning，
        # writer 必須把它寫進去且不影響其他欄位。
        output = str(tmp_path / "output.xlsx")
        rows = [{**generated_rows[0], "reasoning": "§9 列舉 3 種格式，各一筆 TC。"}]
        write_generated_results(input_xlsx, rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        assert ws.cell(row=10, column=18).value == "§9 列舉 3 種格式，各一筆 TC。"

    def test_skips_reasoning_when_empty_string(self, input_xlsx, generated_rows, tmp_path):
        # Sub TC 的 reasoning 會是 ""；writer 對非 clearable 的空值會保留 template
        # 既有內容（這裡 template column R 本來就空）→ cell 維持空值。
        output = str(tmp_path / "output.xlsx")
        rows = [{**generated_rows[0], "reasoning": ""}]
        write_generated_results(input_xlsx, rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        # column R 仍應是 None / 空（不會被空字串污染）
        assert not ws.cell(row=10, column=18).value

    def test_reasoning_cell_has_wrap_text_alignment(self, input_xlsx, generated_rows, tmp_path):
        # 2–5 句的 reasoning 一定要 wrap_text 才不會溢出 column R。
        output = str(tmp_path / "output.xlsx")
        rows = [{**generated_rows[0], "reasoning": "原子需求，無分支、無列舉、無狀態切換。"}]
        write_generated_results(input_xlsx, rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        cell = ws.cell(row=10, column=18)
        assert cell.alignment.wrap_text is True
        assert cell.alignment.vertical == "top"

    def test_reasoning_column_has_explicit_width(self, input_xlsx, generated_rows, tmp_path):
        # Column R 是新加的欄，沒 template 預設寬度，writer 必須自己設一個合理值。
        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, generated_rows, output)
        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        assert ws.column_dimensions["R"].width == 60

    def test_writes_reasoning_header_at_template_header_row(self, input_xlsx, generated_rows, tmp_path):
        # _ensure_reasoning_header 應該掃 J 欄找到 Pre-Conditions header（row 9），
        # 在同一列的 R 欄寫上 "AI Reasoning"。
        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, generated_rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        assert ws.cell(row=9, column=18).value == "AI Reasoning"

    def test_does_not_overwrite_existing_reasoning_header(self, input_xlsx, generated_rows, tmp_path):
        # 使用者自己在 template 改過 column R 的 header → 不要覆蓋。
        from openpyxl import load_workbook as lw, Workbook  # noqa
        wb = lw(input_xlsx)
        ws = wb["Test Case Specification&Result"]
        ws.cell(row=9, column=18, value="自訂解讀標題")
        wb.save(input_xlsx)

        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, generated_rows, output)
        wb_out = load_workbook(output)
        ws_out = wb_out["Test Case Specification&Result"]
        assert ws_out.cell(row=9, column=18).value == "自訂解讀標題"

    def test_respects_selected_fields_when_writing(self, input_xlsx, generated_rows, tmp_path):
        output = str(tmp_path / "output.xlsx")
        write_generated_results(
            input_xlsx,
            generated_rows,
            output,
            selected_fields={"expected_result", "priority"},
        )

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        assert ws.cell(row=10, column=6).value is None  # TC ID untouched
        assert ws.cell(row=10, column=9).value == "PDM01.1) Original text here."  # no rewrite appended
        assert ws.cell(row=10, column=10).value is None  # Pre-Conditions untouched
        assert ws.cell(row=10, column=13).value == "1. Settings shown.\n2. DM icon displayed."
        assert ws.cell(row=10, column=16).value == "P1"


    def test_multi_tc_per_row_inserts_rows(self, input_xlsx, tmp_path):
        """AI 把 1 個 req 拆成 3 筆 TC 時，writer 應該在原列下方插 2 列，
        複製 C/D/I 欄，然後各列寫入自己的 TC 內容。"""
        rows = [
            {
                "row_num": 10,
                "req_id": "SWE1-HMI-DM-001-01",
                "tc_id": "newR1L-DMR-001",
                "tc_title": "(Trigger → Outcome A)",
                "pre_conditions": "pre-A",
                "test_procedure": "1. step A",
                "expected_result": "1. result A",
                "priority": "P0",
                "design_method": "Functional",
            },
            {
                "row_num": 10,
                "req_id": "SWE1-HMI-DM-001-01",
                "tc_id": "newR1L-DMR-002",
                "tc_title": "(Trigger → Outcome B)",
                "pre_conditions": "pre-B",
                "test_procedure": "1. step B",
                "expected_result": "1. result B",
                "priority": "P1",
                "design_method": "Functional",
            },
            {
                "row_num": 10,
                "req_id": "SWE1-HMI-DM-001-01",
                "tc_id": "newR1L-DMR-003",
                "tc_title": "(Trigger → Outcome C)",
                "pre_conditions": "pre-C",
                "test_procedure": "1. step C",
                "expected_result": "1. result C",
                "priority": "P2",
                "design_method": "Functional",
            },
            # 第二個 req 放在原本的 row 11，確認 insert 後的 offset 正確。
            {
                "row_num": 11,
                "req_id": "SWE1-HMI-DM-002-01",
                "tc_id": "newR1L-DMR-004",
                "tc_title": "(Other trigger → outcome)",
                "pre_conditions": "pre-D",
                "test_procedure": "1. step D",
                "expected_result": "1. result D",
                "priority": "P0",
                "design_method": "Functional",
            },
        ]
        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        # Row 10–12 都屬於第一個 req，C/D/I 欄位值應相同。
        for r in (10, 11, 12):
            assert ws.cell(row=r, column=4).value == "SWE1-HMI-DM-001-01"
            assert "PDM01.1) Original text here." in (ws.cell(row=r, column=9).value or "")
        # 三個 TC ID 各自寫入各自的列。
        assert ws.cell(row=10, column=6).value == "newR1L-DMR-001"
        assert ws.cell(row=11, column=6).value == "newR1L-DMR-002"
        assert ws.cell(row=12, column=6).value == "newR1L-DMR-003"
        # procedure / expected 各自分開。
        assert ws.cell(row=10, column=12).value == "1. step A"
        assert ws.cell(row=11, column=12).value == "1. step B"
        assert ws.cell(row=12, column=12).value == "1. step C"
        assert ws.cell(row=10, column=13).value == "1. result A"
        assert ws.cell(row=12, column=13).value == "1. result C"
        # 每筆 sub-TC 都有自己專屬的 rewrite（§1.2），且不會混到別筆的 tag。
        row10_i = ws.cell(row=10, column=9).value
        row11_i = ws.cell(row=11, column=9).value or ""
        row12_i = ws.cell(row=12, column=9).value or ""
        assert "(Trigger → Outcome A)" in row10_i
        assert "(Trigger → Outcome B)" in row11_i
        assert "(Trigger → Outcome C)" in row12_i
        # row 11 不應混到 A 或 C
        assert "Outcome A" not in row11_i
        assert "Outcome C" not in row11_i
        # 第二個 req 被 offset 後應該落在 row 13（10 + 2 extras + 1）。
        assert ws.cell(row=13, column=4).value == "SWE1-HMI-DM-002-01"
        assert ws.cell(row=13, column=6).value == "newR1L-DMR-004"

    def test_does_not_overwrite_template_with_empty_values(self, input_xlsx, tmp_path):
        """Template 既有內容（例如 spec 帶進來的 test_procedure）不應該被
        空字串 / None 覆蓋。"""
        # 先模擬 template 原本就有 procedure/expected 內容。
        wb = load_workbook(input_xlsx)
        ws = wb["Test Case Specification&Result"]
        ws.cell(row=10, column=12, value="template procedure from spec")
        ws.cell(row=10, column=13, value="template expected from spec")
        wb.save(input_xlsx)

        rows = [{
            "row_num": 10,
            "req_id": "SWE1-HMI-DM-001-01",
            "tc_id": "newR1L-DMR-001",
            "tc_title": "(X → Y)",
            "pre_conditions": "",         # 空字串：不應蓋掉 template
            "test_procedure": "",
            "expected_result": "",
            "priority": "P1",
            "design_method": "Functional",
        }]
        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        # Template 原本的內容保留。
        assert ws.cell(row=10, column=12).value == "template procedure from spec"
        assert ws.cell(row=10, column=13).value == "template expected from spec"
        # 非空欄位仍有寫入。
        assert ws.cell(row=10, column=16).value == "P1"

    def test_preserves_existing_cell_format_and_data_validation(self, input_xlsx, generated_rows, tmp_path):
        wb = load_workbook(input_xlsx)
        ws = wb["Test Case Specification&Result"]
        ref = wb.create_sheet("reference")
        ref.sheet_state = "hidden"
        for idx, value in enumerate(
            [
                "功能測試 (Functional based ; no specific technique)",
                "狀態轉換 (State Transition Testing)",
                "決策表 (Decision Table Testing)",
                "等價劃分 (Equivalence Partitioning, EP)",
                "邊界值分析 (Boundary Value Analysis, BVA)",
                "組合測試 (Combinatorial Testing ; Pairwise / t-wise)",
                "情境 / 用例 (Scenario / Use Case Testing)",
                "負向測試 (Negative / Invalid)",
                "基礎故障注入 (Fault Injection Lite)",
            ],
            start=4,
        ):
            ref.cell(row=idx, column=3, value=value)
        target = ws.cell(row=10, column=16)
        target.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        priority_dv = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=True)
        priority_dv.add(target)
        ws.add_data_validation(priority_dv)
        design_method_dv = DataValidation(
            type="list",
            formula1="'reference'!$C$4:$C$12",
            allow_blank=True,
        )
        design_method_dv.add(ws.cell(row=10, column=17))
        ws.add_data_validation(design_method_dv)
        wb.save(input_xlsx)

        output = str(tmp_path / "output.xlsx")
        write_generated_results(input_xlsx, generated_rows, output)

        wb = load_workbook(output)
        ws = wb["Test Case Specification&Result"]
        written = ws.cell(row=10, column=16)
        assert written.value == "P1"
        assert written.alignment.horizontal == "center"
        assert written.alignment.vertical == "center"
        assert written.alignment.wrap_text is not True
        assert any("P10" in str(dv.sqref) for dv in ws.data_validations.dataValidation)
        assert any(
            "Q10" in str(dv.sqref) and dv.formula1 == "'reference'!$C$4:$C$12"
            for dv in ws.data_validations.dataValidation
        )


class TestWriteFrameworkSheet:
    def test_writes_data(self, input_xlsx, framework_data, tmp_path):
        output = str(tmp_path / "output.xlsx")
        write_framework_sheet(input_xlsx, framework_data, output)

        wb = load_workbook(output)
        ws = wb["Test Case Framework"]
        assert ws.cell(row=1, column=1).value == "Test Group"
        assert ws.cell(row=1, column=3).value == "TC Count"
        assert ws.cell(row=1, column=4).value == "Req Count"
        assert ws.cell(row=2, column=1).value == "DeviceManager"
        assert ws.cell(row=2, column=2).value == "Access & Entry"
        # req_count only → falls back to both TC and Req counts
        assert ws.cell(row=2, column=3).value == 5
        assert ws.cell(row=2, column=4).value == 5
        assert ws.cell(row=3, column=2).value == "Device List"

    def test_distinguishes_tc_count_from_req_count(self, input_xlsx, tmp_path):
        """AI 把一個 req 拆成多筆 TC 時，TC Count > Req Count。"""
        data = [
            {"test_group": "DM", "test_set": "Formats", "tc_count": 6, "req_count": 1},
            {"test_group": "DM", "test_set": "Boundary", "tc_count": 3, "req_count": 3},
        ]
        output = str(tmp_path / "output.xlsx")
        write_framework_sheet(input_xlsx, data, output)

        wb = load_workbook(output)
        ws = wb["Test Case Framework"]
        assert ws.cell(row=2, column=3).value == 6  # TC Count
        assert ws.cell(row=2, column=4).value == 1  # Req Count
        assert ws.cell(row=3, column=3).value == 3
        assert ws.cell(row=3, column=4).value == 3
