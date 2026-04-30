"""Tests for FastAPI integration endpoints."""
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

import json
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from api_server import app


def _build_workbook_bytes() -> bytes:
    wb = Workbook()
    ws_pd = wb.active
    ws_pd.title = "Product Document"
    ws_pd.cell(row=3, column=2, value="newR1L")

    ws_tc = wb.create_sheet("Test Case Specification&Result")
    ws_tc.cell(row=9, column=4, value="Requirement or Design ID")
    ws_tc.cell(row=9, column=8, value="Test Set")
    ws_tc.cell(row=9, column=9, value="Test Item")
    ws_tc.cell(row=10, column=4, value="SWE1-HMI-DM-001-01")
    ws_tc.cell(row=10, column=8, value="Workbook Legacy Set 1")
    ws_tc.cell(row=10, column=9, value="PDM01 original text")
    ws_tc.cell(row=11, column=4, value="SWE1-HMI-DM-002-01")
    ws_tc.cell(row=11, column=8, value="Workbook Legacy Set 2")
    ws_tc.cell(row=11, column=9, value="PDM02 original text")

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _build_duplicate_req_workbook_bytes() -> bytes:
    wb = Workbook()
    ws_pd = wb.active
    ws_pd.title = "Product Document"
    ws_pd.cell(row=3, column=2, value="newR1L")

    ws_tc = wb.create_sheet("Test Case Specification&Result")
    ws_tc.cell(row=9, column=4, value="Requirement or Design ID")
    ws_tc.cell(row=9, column=6, value="Test Case ID")
    ws_tc.cell(row=9, column=9, value="Test Item")
    ws_tc.cell(row=10, column=4, value="SWE1-HMI-DM-001-01")
    ws_tc.cell(row=10, column=6, value="newR1L-DM-001")
    ws_tc.cell(row=10, column=9, value="PDM01 original text")
    ws_tc.cell(row=11, column=4, value="SWE1-HMI-DM-001-01")
    ws_tc.cell(row=11, column=6, value="newR1L-DM-002")
    ws_tc.cell(row=11, column=9, value="PDM01 split text")

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _build_reference_workbook_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Basic Report"
    ws.cell(row=1, column=1, value="NRL ID")
    ws.cell(row=1, column=3, value="Outline")
    ws.cell(row=1, column=4, value="Description")
    ws.cell(row=1, column=5, value="Source ID")
    ws.cell(row=2, column=1, value="NRL-001")
    ws.cell(row=2, column=3, value="Device Manager")
    ws.cell(row=2, column=4, value="PDM01 behavior description")
    ws.cell(row=2, column=5, value="SPEC_REF_PDM01")
    ws.cell(row=3, column=1, value="NRL-002")
    ws.cell(row=3, column=3, value="Device Manager")
    ws.cell(row=3, column=4, value="PDM02 behavior description")
    ws.cell(row=3, column=5, value="SPEC_REF_PDM02")

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


client = TestClient(app)


def test_healthcheck():
    response = client.get("/api/health")
    assert response.status_code == 200
    assert response.json()["status"] == "ok"


def test_admin_reset_wipes_stores_from_localhost():
    # Seed a job so we can verify it disappears.
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    from api_server import JOB_REGISTRY
    assert JOB_REGISTRY.get(job_id) is not None

    response = client.delete("/api/admin/reset")
    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "ok"
    assert payload["jobsRemoved"] >= 1

    # Job is gone.
    assert JOB_REGISTRY.get(job_id) is None


def test_parse_workbook_rejects_path_separator_in_filename():
    response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "../evil.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 400
    assert "filename" in response.json()["detail"]


def test_attach_raw_rejects_path_separator_in_filename():
    response = client.post(
        "/api/jobs/job-sec/attach-raw",
        files={
            "raw_file": (
                r"..\\evil.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 400
    assert "filename" in response.json()["detail"]


def test_metrics_aggregate_returns_shape_with_empty_store():
    # 測試環境 JOB_REGISTRY 可能有也可能沒有資料；只驗回傳 shape 正確
    response = client.get("/api/metrics/aggregate")
    assert response.status_code == 200
    payload = response.json()
    for field in (
        "jobCount", "totalRowCount", "totalCostUsd", "avgCostUsd",
        "avgRowCount", "matchRate", "jobsWithCost", "jobsWithMatch",
    ):
        assert field in payload


def test_metrics_aggregate_rejects_unknown_job_id():
    response = client.get("/api/metrics/aggregate?job_ids=__does_not_exist__")
    assert response.status_code == 404


def test_job_usage_returns_per_model_breakdown():
    """Per-job usage endpoint — drives CostMeter popup's per-model section."""
    # Seed a job with a known per-model cost breakdown.
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    from api_server import JOB_REGISTRY, _persist_job_usage
    job = JOB_REGISTRY.get(job_id)
    # Simulate two AI tasks on the same job at different model tiers.
    _persist_job_usage(
        job_id, job,
        cost=0.10, input_tokens=1000, output_tokens=200,
        cache_creation_tokens=0, cache_read_tokens=0,
        cost_delta=0.10, model="gpt-5",
    )
    job = JOB_REGISTRY.get(job_id)
    _persist_job_usage(
        job_id, job,
        cost=0.12, input_tokens=1200, output_tokens=250,
        cache_creation_tokens=0, cache_read_tokens=0,
        cost_delta=0.02, model="gpt-5-mini",
    )

    response = client.get(f"/api/jobs/{job_id}/usage")
    assert response.status_code == 200
    payload = response.json()
    assert payload["jobId"] == job_id
    assert payload["cost"] == 0.12
    # Per-model bucket preserves separate attribution.
    assert payload["costByModel"]["gpt-5"] == 0.10
    assert payload["costByModel"]["gpt-5-mini"] == 0.02


def test_job_usage_404_for_unknown_job():
    response = client.get("/api/jobs/does-not-exist/usage")
    assert response.status_code == 404


def test_parse_workbook():
    files = {
        "raw_file": (
            "SomeProject_SWQT_DeviceManager_20260408.xlsx",
            _build_workbook_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    response = client.post("/api/parse", files=files)
    assert response.status_code == 200
    payload = response.json()
    assert payload["project"] == "newR1L"
    assert payload["testGroup"] == "DeviceManager"
    assert payload["rowCount"] == 2
    assert payload["previewHeaders"] == ["req_id", "test_item", "test_set", "priority"]
    assert payload["rows"][0]["reqId"] == "SWE1-HMI-DM-001-01"
    assert payload["rows"][0]["tcId"] == ""
    assert payload["rows"][0]["testSet"] == ""
    assert payload["previewRows"][0]["test_set"] == ""


@patch("tools.group.classify_test_sets")
def test_group_uses_imported_test_set_as_ai_hint_not_existing_assignment(mock_classify):
    from generator import ClassificationResult

    mock_classify.return_value = ClassificationResult(
        assignments={"row-10": "Generated Set"},
    )
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    response = client.post(
        "/api/group",
        json={
            "jobId": job_id,
            "forceRegroup": False,
            "rows": [
                {
                    "id": "row-10",
                    "rowNum": 10,
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "PDM01 original text",
                    "testSet": "",
                }
            ],
        },
    )

    assert response.status_code == 200
    sent_rows = mock_classify.call_args.args[0]
    assert sent_rows[0]["current_test_set"] == "Workbook Legacy Set 1"
    assignment = response.json()["assignments"][0]
    assert assignment["testSet"] == "Generated Set"
    assert assignment["source"] == "derived"


def test_parse_workbook_accepts_reference_workbook():
    files = {
        "raw_file": (
            "SomeProject_SWQT_DeviceManager_20260408.xlsx",
            _build_workbook_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        "reference_file": (
            "ReferenceWorkbook.xlsx",
            _build_reference_workbook_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }
    response = client.post("/api/parse", files=files)
    assert response.status_code == 200
    payload = response.json()
    assert payload["files"]["referenceWorkbookName"] == "ReferenceWorkbook.xlsx"


def test_parse_rejects_invalid_extension():
    files = {
        "raw_file": ("bad.txt", b"not-an-excel", "text/plain"),
    }
    response = client.post("/api/parse", files=files)
    assert response.status_code == 400


def test_create_generate_job():
    response = client.post(
        "/api/generate",
        json={
            "jobId": "parse-20260416-123456",
            "rows": [
                {
                    "id": "row-10",
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "PDM01 original text",
                    "originalRequirement": "PDM01 original text",
                    "testSet": "",
                    "priority": "",
                }
            ],
            "config": {
                "model": "gpt-5.4-mini",
                "batchSize": 5,
                "budget": 2,
                "strictValidation": False,
            },
        },
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["jobId"] == "parse-20260416-123456"
    assert payload["status"] == "queued"
    assert payload["totalRows"] == 1
    assert "/api/generate/stream?jobId=parse-20260416-123456" in payload["streamUrl"]


@patch("tools.generate.generate_batch_multi")
def test_stream_generate_job(mock_generate_batch):
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    payload = parse_response.json()

    # Multi-TC 模式：每個 input row 回傳一個 list of TCs（此測試每 req 只 1 筆）
    mock_generate_batch.return_value = SimpleNamespace(
        tc_data=[
            [
                {
                    "tc_title": "(Condition → Outcome A)",
                    "pre_conditions": "NA",
                    "input_test_data": "NA",
                    "test_procedure": "1. Perform setup.\n2. Verify the result.",
                    "expected_result": "1. Setup completes.\n2. Result is verified.",
                    "design_method": "功能測試 (Functional based ; no specific technique)",
                    "priority": "P0",
                    "split_flag": False,
                    "split_reason": "",
                },
            ],
            [
                {
                    "tc_title": "(Condition → Outcome B)",
                    "pre_conditions": "NA",
                    "input_test_data": "NA",
                    "test_procedure": "1. Perform setup.\n2. Verify the result.",
                    "expected_result": "1. Setup completes.\n2. Result is verified.",
                    "design_method": "功能測試 (Functional based ; no specific technique)",
                    "priority": "P1",
                    "split_flag": False,
                    "split_reason": "",
                },
            ],
        ],
        input_tokens=10,
        output_tokens=20,
        cost=0.001,
    )

    client.post(
        "/api/generate",
        json={
            "jobId": payload["jobId"],
            "rows": payload["rows"],
            "config": {
                "model": "gpt-5.4-mini",
                "batchSize": 2,
                "budget": 2,
                "strictValidation": False,
            },
        },
    )

    response = client.get("/api/generate/stream", params={"jobId": payload["jobId"]})
    assert response.status_code == 200
    assert mock_generate_batch.called
    assert '"type": "job.started"' in response.text
    assert '"type": "row.completed"' in response.text
    assert '"type": "job.completed"' in response.text
    assert '"tcId": "newR1L-DM-001"' in response.text


@patch("tools.generate.generate_batch_multi")
def test_stream_generate_assigns_tc_ids_in_final_display_order(mock_generate_batch):
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    payload = parse_response.json()

    mock_generate_batch.return_value = SimpleNamespace(
        tc_data=[
            [
                {
                    "tc_title": "(Condition → Outcome A1)",
                    "pre_conditions": "NA",
                    "input_test_data": "NA",
                    "test_procedure": "1. Perform setup.\n2. Verify the result.",
                    "expected_result": "1. Setup completes.\n2. Result is verified.",
                    "design_method": "功能測試 (Functional based ; no specific technique)",
                    "priority": "P0",
                    "split_flag": True,
                    "split_reason": "Split into two cases",
                },
                {
                    "tc_title": "(Condition → Outcome A2)",
                    "pre_conditions": "NA",
                    "input_test_data": "NA",
                    "test_procedure": "1. Perform alternate setup.\n2. Verify alternate result.",
                    "expected_result": "1. Alternate setup completes.\n2. Alternate result is verified.",
                    "design_method": "功能測試 (Functional based ; no specific technique)",
                    "priority": "P0",
                    "split_flag": True,
                    "split_reason": "Split into two cases",
                },
            ],
            [
                {
                    "tc_title": "(Condition → Outcome B1)",
                    "pre_conditions": "NA",
                    "input_test_data": "NA",
                    "test_procedure": "1. Perform setup.\n2. Verify the result.",
                    "expected_result": "1. Setup completes.\n2. Result is verified.",
                    "design_method": "功能測試 (Functional based ; no specific technique)",
                    "priority": "P1",
                    "split_flag": False,
                    "split_reason": "",
                },
            ],
        ],
        input_tokens=10,
        output_tokens=20,
        cost=0.001,
        split_meta=[
            {"req_id": "SWE1-HMI-DM-001-01", "reasoning": "split", "keywords": []},
            {"req_id": "SWE1-HMI-DM-002-01", "reasoning": "single", "keywords": []},
        ],
    )

    client.post(
        "/api/generate",
        json={
            "jobId": payload["jobId"],
            "rows": payload["rows"],
            "config": {
                "model": "gpt-4.1",
                "batchSize": 2,
                "budget": 2,
                "strictValidation": False,
            },
        },
    )

    response = client.get("/api/generate/stream", params={"jobId": payload["jobId"]})
    assert response.status_code == 200

    events = _parse_sse(response.content)
    tc_ids = [
        event["row"]["tcId"]
        for event in events
        if event.get("type") in {"row.completed", "row.added"}
    ]
    assert tc_ids == [
        "newR1L-DM-001",
        "newR1L-DM-002",
        "newR1L-DM-003",
    ]


def _build_mixed_workbook_bytes() -> bytes:
    """Row 10 is a reference example (Pre-Cond/Procedure/Expected filled).
    Row 11 is a new row that needs generation.
    """
    wb = Workbook()
    ws_pd = wb.active
    ws_pd.title = "Product Document"
    ws_pd.cell(row=3, column=2, value="newR1L")

    ws_tc = wb.create_sheet("Test Case Specification&Result")
    ws_tc.cell(row=9, column=4, value="Requirement or Design ID")
    ws_tc.cell(row=9, column=9, value="Test Item")

    # Row 10: reference example — fully populated, must not be regenerated.
    ws_tc.cell(row=10, column=4, value="SWE1-REF-001")
    ws_tc.cell(row=10, column=9, value="Reference test item text")
    ws_tc.cell(row=10, column=10, value="1. Reference pre-condition")
    ws_tc.cell(row=10, column=12, value="1. Reference procedure step")
    ws_tc.cell(row=10, column=13, value="1. Reference expected outcome")

    # Row 11: needs generation.
    ws_tc.cell(row=11, column=4, value="SWE1-NEW-002")
    ws_tc.cell(row=11, column=9, value="New test item text")

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


@patch("tools.generate.generate_tcs_for_row")
def test_stream_generate_sends_reviewer_prefills_as_hints(mock_generate_single):
    """每列都走 AI，template 既有的 pre-conditions / procedure / expected 會被
    當作 reviewer hints 塞進 prompt，不再當「已完成內容」跳過 AI。"""
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_mixed_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    parsed = parse_response.json()

    mock_generate_single.return_value = SimpleNamespace(
        tc_data=[{
            "tc_title": "(Condition → Outcome)",
            "pre_conditions": "NA",
            "input_test_data": "NA",
            "test_procedure": "1. Setup.\n2. Verify.",
            "expected_result": "1. Setup ok.\n2. Verified.",
            "design_method": "功能測試 (Functional based ; no specific technique)",
            "priority": "P1",
            "split_flag": False,
            "split_reason": "",
        }],
        input_tokens=10,
        output_tokens=20,
        cache_creation_tokens=0,
        cache_read_tokens=0,
        cost=0.001,
        model="gpt-5.4-mini",
        split_meta=[{"req_id": "R", "reasoning": "r", "keywords": []}],
    )

    client.post(
        "/api/generate",
        json={
            "jobId": parsed["jobId"],
            "rows": parsed["rows"],
            "config": {
                "model": "gpt-4.1",
                "batchSize": 1,
                "budget": 2,
                "strictValidation": False,
            },
        },
    )

    response = client.get("/api/generate/stream", params={"jobId": parsed["jobId"]})
    assert response.status_code == 200
    # 兩列都應該走 AI（一個 batch_size=1 時每列各 1 次）
    assert mock_generate_single.call_count == 2
    # 被送給 AI 的 row dict 仍帶原始 pre-fills，供 prompt_builder 轉成 reviewer hints。
    sent_rows = [call.args[0] for call in mock_generate_single.call_args_list]
    prefilled_row = next(
        r for r in sent_rows if r["req_id"] == "SWE1-REF-001"
    )
    assert "Reference pre-condition" in prefilled_row.get("pre_conditions", "")
    assert "Reference procedure step" in prefilled_row.get("test_procedure", "")
    # 不再有 preserved / preserved true 的訊息
    assert '"preserved": true' not in response.text


@patch("tools.generate.generate_batch_multi")
def test_stream_generate_regenerate_all_skips_preservation(mock_generate_batch):
    """regenerateAll=True forces AI regeneration for every row."""
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_mixed_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    parsed = parse_response.json()

    # Multi-TC：外層 per-req list，內層 per-TC list（這裡每 req 1 筆 TC）
    mock_generate_batch.return_value = SimpleNamespace(
        tc_data=[
            [{
                "tc_title": "(Condition → Outcome)",
                "pre_conditions": "NA",
                "input_test_data": "NA",
                "test_procedure": "1. Setup.\n2. Verify.",
                "expected_result": "1. Setup ok.\n2. Verified.",
                "design_method": "功能測試 (Functional based ; no specific technique)",
                "priority": "P1",
                "split_flag": False,
                "split_reason": "",
            }]
        ] * 2,
        input_tokens=10,
        output_tokens=20,
        cost=0.001,
    )

    client.post(
        "/api/generate",
        json={
            "jobId": parsed["jobId"],
            "rows": parsed["rows"],
            "config": {
                "model": "gpt-4.1",
                "batchSize": 2,
                "budget": 2,
                "strictValidation": False,
                "regenerateAll": True,
            },
        },
    )

    response = client.get("/api/generate/stream", params={"jobId": parsed["jobId"]})
    assert response.status_code == 200
    # Both rows in one batch — AI called once with 2 rows.
    assert mock_generate_batch.call_count == 1
    sent_rows = mock_generate_batch.call_args.args[0]
    assert len(sent_rows) == 2


@patch("tools.generate.generate_tcs_for_row")
def test_stream_generate_job_marks_strict_validation_failures(mock_generate_single_tc):
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    payload = parse_response.json()

    mock_generate_single_tc.return_value = SimpleNamespace(
        tc_data=[{
            "tc_title": "(Condition → Outcome)",
            "pre_conditions": "1. Open settings menu",
            "input_test_data": "NA",
            "test_procedure": "1. Perform setup.\n2. Execute action without verification.",
            "expected_result": "1. Setup completes.\n2. Works as expected.",
            "design_method": "invalid method",
            "priority": "Critical",
            "split_flag": False,
            "split_reason": "",
        }],
        input_tokens=10,
        output_tokens=20,
        cost=0.001,
    )

    client.post(
        "/api/generate",
        json={
            "jobId": payload["jobId"],
            "rows": [payload["rows"][0]],
            "config": {
                "model": "gpt-4.1",
                "batchSize": 1,
                "budget": 2,
                "strictValidation": True,
            },
        },
    )

    response = client.get("/api/generate/stream", params={"jobId": payload["jobId"]})
    assert response.status_code == 200
    assert '"type": "row.failed"' in response.text
    assert '"status": "error"' in response.text
    assert '"generated": null' in response.text


@patch("tools.generate.generate_tcs_for_row")
def test_stream_generate_job_keeps_warning_rows_completed_in_non_strict_mode(mock_generate_single_tc):
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    payload = parse_response.json()

    mock_generate_single_tc.return_value = SimpleNamespace(
        tc_data=[{
            "tc_title": "(Condition → Outcome)",
            "pre_conditions": "1. Open settings menu",
            "input_test_data": "NA",
            "test_procedure": "1. Perform setup.\n2. Execute action without verification.",
            "expected_result": "1. Setup completes.\n2. Works as expected.",
            "design_method": "invalid method",
            "priority": "Critical",
            "split_flag": False,
            "split_reason": "",
        }],
        input_tokens=10,
        output_tokens=20,
        cost=0.001,
    )

    client.post(
        "/api/generate",
        json={
            "jobId": payload["jobId"],
            "rows": [payload["rows"][0]],
            "config": {
                "model": "gpt-4.1",
                "batchSize": 1,
                "budget": 2,
                "strictValidation": False,
            },
        },
    )

    response = client.get("/api/generate/stream", params={"jobId": payload["jobId"]})
    assert response.status_code == 200
    assert '"type": "row.completed"' in response.text
    assert '"status": "ready"' in response.text
    assert '"generated": {' in response.text


def test_export_job_and_download():
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    export_response = client.post(
        "/api/export",
        json={
            "jobId": job_id,
            "scope": "accepted",
            "outputMode": "new-file",
            "includeFrameworkSheet": True,
            "selectedColumns": ["TC Title", "Expected Result"],
            "rows": [
                {
                    "id": "row-10",
                    "rowNum": 10,
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "PDM01 original text",
                    "reviewStatus": "accepted",
                    "testSet": "Smoke",
                    "generated": {
                        "tcTitle": "(PDM01 original text → Observable outcome confirmed)",
                        "preConditions": "1. Vehicle profile loaded\n2. Required subsystem available",
                        "testProcedure": (
                            "1. Open the source screen and prepare the feature.\n"
                            "2. Trigger the target behavior and verify the visible outcome."
                        ),
                        "expectedResult": (
                            "1. The setup screen is ready for the operator.\n"
                            "2. The requested behavior is shown with the correct visible outcome."
                        ),
                        "designMethod": "功能測試 (Functional based ; no specific technique)",
                        "priority": "P0",
                    },
                }
            ],
        },
    )
    assert export_response.status_code == 200
    payload = export_response.json()
    assert payload["status"] == "ready"
    assert payload["exportedRows"] == 1
    assert payload["fileName"].endswith("_generated.xlsx")

    download_response = client.get(f"/api/export/download/{job_id}")
    assert download_response.status_code == 200
    assert (
        download_response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_export_respects_selected_columns():
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    export_response = client.post(
        "/api/export",
        json={
            "jobId": job_id,
            "scope": "accepted",
            "outputMode": "new-file",
            "includeFrameworkSheet": False,
            "selectedColumns": ["Expected Result", "Priority"],
            "rows": [
                {
                    "id": "row-10",
                    "rowNum": 10,
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "PDM01 original text",
                    "reviewStatus": "accepted",
                    "testSet": "Smoke",
                    "generated": {
                        "tcTitle": "(PDM01 original text → Observable outcome confirmed)",
                        "preConditions": "1. Vehicle profile loaded\n2. Required subsystem available",
                        "inputTestData": "NA",
                        "testProcedure": (
                            "1. Open the source screen and prepare the feature.\n"
                            "2. Trigger the target behavior and verify the visible outcome."
                        ),
                        "expectedResult": (
                            "1. The setup screen is ready for the operator.\n"
                            "2. The requested behavior is shown with the correct visible outcome."
                        ),
                        "designMethod": "功能測試 (Functional based ; no specific technique)",
                        "priority": "P0",
                    },
                }
            ],
        },
    )
    assert export_response.status_code == 200

    download_response = client.get(f"/api/export/download/{job_id}")
    workbook = load_workbook(BytesIO(download_response.content))
    ws = workbook["Test Case Specification&Result"]
    assert ws.cell(row=10, column=6).value is None
    assert ws.cell(row=10, column=9).value == "PDM01 original text"
    assert ws.cell(row=10, column=13).value == (
        "1. The setup screen is ready for the operator.\n"
        "2. The requested behavior is shown with the correct visible outcome."
    )
    assert ws.cell(row=10, column=16).value == "P0"


def test_export_includes_metadata_only_rows():
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    export_response = client.post(
        "/api/export",
        json={
            "jobId": job_id,
            "scope": "accepted",
            "outputMode": "new-file",
            "includeFrameworkSheet": False,
            "selectedColumns": ["Priority", "Design Method"],
            "rows": [
                {
                    "id": "row-10",
                    "rowNum": 10,
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "PDM01 original text",
                    "reviewStatus": "accepted",
                    "testSet": "Smoke",
                    "generated": {
                        "priority": "P0",
                        "designMethod": "功能測試 (Functional based ; no specific technique)",
                    },
                }
            ],
        },
    )
    assert export_response.status_code == 200
    assert export_response.json()["exportedRows"] == 1

    download_response = client.get(f"/api/export/download/{job_id}")
    workbook = load_workbook(BytesIO(download_response.content))
    ws = workbook["Test Case Specification&Result"]
    assert ws.cell(row=10, column=16).value == "P0"
    assert ws.cell(row=10, column=17).value == "功能測試 (Functional based ; no specific technique)"


def test_export_prefers_tc_id_when_row_num_is_missing_for_duplicate_req_ids():
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "DupReq_SWQT_DeviceManager_20260408.xlsx",
                _build_duplicate_req_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    export_response = client.post(
        "/api/export",
        json={
            "jobId": job_id,
            "scope": "accepted",
            "outputMode": "new-file",
            "includeFrameworkSheet": False,
            "selectedColumns": ["Expected Result", "AI 需求解讀"],
            "rows": [
                {
                    "id": "row-11",
                    "rowNum": None,
                    "tcId": "newR1L-DM-002",
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "PDM01 split text",
                    "reviewStatus": "accepted",
                    "testSet": "Smoke",
                    "generated": {
                        "expectedResult": "2nd row only",
                    },
                    "splitDecision": {
                        "reasoning": "AI 判定兩列完全重複，保留第二列覆寫原始 row。",
                    },
                }
            ],
        },
    )
    assert export_response.status_code == 200

    download_response = client.get(f"/api/export/download/{job_id}")
    workbook = load_workbook(BytesIO(download_response.content))
    ws = workbook["Test Case Specification&Result"]
    assert ws.cell(row=10, column=4).value == "SWE1-HMI-DM-001-01"
    assert ws.cell(row=10, column=9).value == "PDM01 split text"
    assert ws.cell(row=10, column=13).value == "2nd row only"
    assert ws.cell(row=9, column=18).value == "AI 需求解讀"
    assert ws.cell(row=10, column=18).value == "AI 判定兩列完全重複，保留第二列覆寫原始 row。"
    assert ws.cell(row=11, column=4).value is None


@patch("api_server.classify_test_sets")
def test_export_defaults_blank_input_data_and_derives_test_set(mock_classify):
    from generator import ClassificationResult
    # Per-row classification: assignments keyed by row id (uuid).
    mock_classify.return_value = ClassificationResult(
        assignments={"row-10": "BT Switch"},
    )
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    export_response = client.post(
        "/api/export",
        json={
            "jobId": job_id,
            "scope": "accepted",
            "outputMode": "new-file",
            "includeFrameworkSheet": False,
            "selectedColumns": ["Test Set", "Input Test Data"],
            "rows": [
                {
                    "id": "row-10",
                    "rowNum": 10,
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "When HMI sends Bluetooth switch request, vehicle provides checkbox to enable/disable Bluetooth",
                    "reviewStatus": "accepted",
                    "testSet": "",
                    "generated": {
                        "inputTestData": "",
                    },
                }
            ],
        },
    )
    assert export_response.status_code == 200

    download_response = client.get(f"/api/export/download/{job_id}")
    workbook = load_workbook(BytesIO(download_response.content))
    ws = workbook["Test Case Specification&Result"]
    assert ws.cell(row=10, column=8).value == "BT Switch"
    assert ws.cell(row=10, column=11).value == "NA"


def test_export_rejects_overwrite_mode():
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    export_response = client.post(
        "/api/export",
        json={
            "jobId": job_id,
            "scope": "accepted",
            "outputMode": "overwrite",
            "includeFrameworkSheet": False,
            "selectedColumns": ["Expected Result"],
            "rows": [
                {
                    "id": "row-10",
                    "rowNum": 10,
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "PDM01 original text",
                    "reviewStatus": "accepted",
                    "testSet": "Smoke",
                    "generated": {
                        "expectedResult": "should not write",
                    },
                }
            ],
        },
    )
    assert export_response.status_code == 400
    assert export_response.json()["detail"] == "overwrite export mode is not supported"


def test_export_returns_detail_for_unexpected_exception():
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    with patch("api_server.write_excel_tool", side_effect=ValueError("boom")):
        export_response = client.post(
            "/api/export",
            json={
                "jobId": job_id,
                "scope": "accepted",
                "outputMode": "new-file",
                "includeFrameworkSheet": False,
                "selectedColumns": ["Expected Result"],
                "rows": [
                    {
                        "id": "row-10",
                        "rowNum": 10,
                        "reqId": "SWE1-HMI-DM-001-01",
                        "testItem": "PDM01 original text",
                        "reviewStatus": "accepted",
                        "testSet": "Smoke",
                        "generated": {
                            "expectedResult": "1. Result",
                        },
                    }
                ],
            },
        )

    assert export_response.status_code == 500
    assert export_response.json()["detail"] == "export failed: ValueError: boom"


def test_group_preview_returns_assignments():
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    payload = parse_response.json()

    response = client.post(
        "/api/group",
        json={
            "jobId": payload["jobId"],
            "rows": payload["rows"],
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["groups"]
    assert body["assignments"]
    assert body["assignments"][0]["testSet"]


def test_match_preview_uses_reference_workbook():
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "reference_file": (
                "ReferenceWorkbook.xlsx",
                _build_reference_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )
    payload = parse_response.json()

    response = client.post(
        "/api/match",
        json={
            "jobId": payload["jobId"],
            "rows": payload["rows"],
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["summary"]["hasReferenceWorkbook"] is True
    assert body["summary"]["exact"] == 2
    assert body["matches"][0]["specReference"] == "SPEC_REF_PDM01"


# ── Quick Generate Stream ─────────────────────────────────────────────────────

VALID_TC_JSON = {
    "tc_title": "Button pressed → LED turns on",
    "pre_conditions": "1. System is powered.",
    "input_test_data": "NA",
    "test_procedure": "1. Press button.\n2. Observe LED.",
    "expected_result": "1. LED turns on.",
    "design_method": "Functional",
    "priority": "P1",
    "split_flag": False,
    "split_reason": "",
}

VALID_DECOMPOSE_RESPONSE = {
    "reasoning": "Two distinct scenarios: normal and boundary.",
    "scenarios": [
        {"id": 1, "name": "Normal", "description": "Primary path.", "test_item": "Button pressed → LED on"},
        {"id": 2, "name": "Boundary", "description": "Edge case.", "test_item": "Button held → LED blink"},
    ],
}


@patch("api_server.generate_tc_tool")
def test_regenerate_stream_accepts_sub_tc_row_from_payload(mock_tool):
    """Sub-TC rows (id like `row__tc2`) only live in frontend state — regenerate
    must fall back to payload.rows, otherwise UI gets stuck on 'generating'."""
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    mock_tool.return_value = {
        "tcData": [[{
            "tc_title": "(Condition → Outcome)",
            "pre_conditions": "NA",
            "input_test_data": "NA",
            "test_procedure": "1. Setup.\n2. Verify.",
            "expected_result": "1. Setup ok.\n2. Verified.",
            "design_method": "功能測試 (Functional based ; no specific technique)",
            "priority": "P1",
            "split_flag": False,
            "split_reason": "",
        }]],
        "cost": 0.0,
        "inputTokens": 0,
        "outputTokens": 0,
        "cacheCreationTokens": 0,
        "cacheReadTokens": 0,
    }

    sub_tc_id = "row-10__tc2"
    response = client.post(
        f"/api/jobs/{job_id}/regenerate/stream",
        json={
            "rowIds": [sub_tc_id],
            "rows": [
                {
                    "id": sub_tc_id,
                    "rowNum": None,
                    "tcId": "newR1L-DM-002",
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "PDM01 split text",
                    "testSet": "Smoke",
                    "preConditions": "",
                    "inputTestData": "",
                    "steps": "",
                    "expectedResults": "",
                    "status": "pending",
                }
            ],
        },
    )
    assert response.status_code == 200
    events = _parse_sse(response.content)
    regenerated = [e for e in events if e["type"] == "row.regenerated"]
    assert len(regenerated) == 1
    assert regenerated[0]["row"]["id"] == sub_tc_id


@patch("api_server.generate_tc_tool")
def test_regenerate_stream_passes_reason_and_emits_insert_plan(mock_tool):
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    def _tc(label: str) -> dict:
        return {
            "tc_title": label,
            "pre_conditions": "NA",
            "input_test_data": "NA",
            "test_procedure": "1. Setup.\n2. Verify.",
            "expected_result": "1. Setup ok.\n2. Verified.",
            "design_method": "功能測試 (Functional based ; no specific technique)",
            "priority": "P1",
            "split_flag": False,
            "split_reason": "",
        }

    mock_tool.return_value = {
        "tcData": [[_tc("primary"), _tc("negative")]],
        "splitMeta": [{"req_id": "SWE1-HMI-DM-001-01", "reasoning": "補負向情境。", "keywords": []}],
        "cost": 0.0,
        "inputTokens": 0,
        "outputTokens": 0,
        "cacheCreationTokens": 0,
        "cacheReadTokens": 0,
    }

    response = client.post(
        f"/api/jobs/{job_id}/regenerate/stream",
        json={
            "rowIds": ["row-10"],
            "regenerateReason": "Missing negative validation case",
            "rows": [
                {
                    "id": "row-10",
                    "rowNum": 10,
                    "tcId": "newR1L-DM-001",
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "Add DM to status bar",
                    "testSet": "Smoke",
                    "preConditions": "",
                    "inputTestData": "",
                    "steps": "",
                    "expectedResults": "",
                    "status": "pending",
                }
            ],
        },
    )

    assert response.status_code == 200
    _, kwargs = mock_tool.call_args
    assert kwargs["allow_split"] is True
    assert kwargs["rows"][0]["regenerate_reason"] == "Missing negative validation case"

    events = _parse_sse(response.content)
    split_events = [e for e in events if e["type"] == "req.split"]
    assert split_events[0]["insertPlan"] == {
        "needsInsert": True,
        "insertAfterId": "row-10",
        "newCount": 1,
        "renumberRequired": True,
    }
    assert len([e for e in events if e["type"] == "row.added"]) == 1


@patch("api_server.generate_tc_tool")
def test_rerun_stream_emits_split_and_added_for_multi_tc(mock_tool):
    """Rerun 走完整 pipeline：一筆 req → 多筆 TC 時要發 req.split + row.regenerated
    (primary) + row.added (extras)，而且帶 parentId 讓前端 insert-after。"""
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]

    def _tc(label: str) -> dict:
        return {
            "tc_title": label,
            "tc_title": "Trigger action → Observable outcome",
            "pre_conditions": "NA",
            "input_test_data": "NA",
            "test_procedure": "1. Setup.\n2. Verify.",
            "expected_result": "1. Setup ok.\n2. Verified.",
            "design_method": "功能測試 (Functional based ; no specific technique)",
            "priority": "P1",
            "split_flag": False,
            "split_reason": "",
        }

    # 一筆 input → AI 拆成 2 筆 TC
    mock_tool.return_value = {
        "tcData": [[_tc("primary"), _tc("secondary")]],
        "splitMeta": [{"req_id": "SWE1-HMI-DM-001-01", "reasoning": "boundary split", "keywords": []}],
        "cost": 0.0,
        "inputTokens": 0,
        "outputTokens": 0,
        "cacheCreationTokens": 0,
        "cacheReadTokens": 0,
    }

    row_id = "row-10"
    response = client.post(
        f"/api/jobs/{job_id}/rerun/stream",
        json={
            "rowIds": [row_id],
            "rows": [
                {
                    "id": row_id,
                    "rowNum": 10,
                    "tcId": "newR1L-DM-001",
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "Add DM to status bar",
                    "testSet": "Smoke",
                    "preConditions": "",
                    "inputTestData": "",
                    "steps": "",
                    "expectedResults": "",
                    "status": "pending",
                }
            ],
        },
    )
    assert response.status_code == 200
    events = _parse_sse(response.content)
    types = [e["type"] for e in events]

    # 必須有 rerun.started / req.split / 至少一筆 row.regenerated / 一筆 row.added
    assert "rerun.started" in types
    assert "req.split" in types
    regenerated = [e for e in events if e["type"] == "row.regenerated"]
    added = [e for e in events if e["type"] == "row.added"]
    assert len(regenerated) == 1
    assert regenerated[0]["row"]["id"] == row_id
    assert len(added) == 1
    assert added[0]["row"]["parentId"] == row_id
    split = [e for e in events if e["type"] == "req.split"][0]
    assert split["insertPlan"]["newCount"] == 1
    assert split["insertPlan"]["insertAfterId"] == row_id
    assert "rerun.completed" in types
    # 呼叫時必須帶 allow_split=True
    _, kwargs = mock_tool.call_args
    assert kwargs.get("allow_split") is True


@patch("api_server.generate_tc_tool")
def test_rerun_stream_uses_payload_workbook_context_fields(mock_tool):
    """Re-run should send the current row context from the UI to AI analysis."""
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]
    mock_tool.return_value = {
        "tcData": [[VALID_TC_JSON]],
        "splitMeta": [{"req_id": "SWE1-HMI-DM-001-01", "reasoning": "", "keywords": []}],
        "cost": 0.0,
        "inputTokens": 0,
        "outputTokens": 0,
        "cacheCreationTokens": 0,
        "cacheReadTokens": 0,
    }

    client.post(
        f"/api/jobs/{job_id}/rerun/stream",
        json={
            "rowIds": ["row-10"],
            "rows": [
                {
                    "id": "row-10",
                    "rowNum": 10,
                    "tcId": "newR1L-DM-001",
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "Payload Test Item",
                    "testSet": "Payload Test Set",
                    "preConditions": "Payload pre-condition",
                    "inputTestData": "Payload input data",
                    "steps": "Payload procedure",
                    "expectedResults": "Payload expected result",
                    "designMethod": "Payload design method",
                    "priority": "P0",
                    "status": "pending",
                }
            ],
        },
    )

    sent_row = mock_tool.call_args.kwargs["rows"][0]
    assert sent_row["test_set"] == "Payload Test Set"
    assert sent_row["test_item"] == "Payload Test Item"
    assert sent_row["pre_conditions"] == "Payload pre-condition"
    assert sent_row["test_procedure"] == "Payload procedure"
    assert sent_row["expected_result"] == "Payload expected result"


@patch("api_server.generate_tc_tool")
def test_rerun_stream_payload_overrides_existing_job_rows(mock_tool):
    """When a job already has rows, re-run still uses the current UI row context."""
    from api_server import JOB_REGISTRY

    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]
    JOB_REGISTRY[job_id]["rows"] = [
        {
            "id": "row-10",
            "rowNum": 10,
            "tcId": "newR1L-DM-001",
            "reqId": "SWE1-HMI-DM-001-01",
            "testItem": "Stale backend Test Item",
            "testSet": "Stale backend Test Set",
            "specReference": None,
            "priority": "P2",
        }
    ]
    mock_tool.return_value = {
        "tcData": [[VALID_TC_JSON]],
        "splitMeta": [{"req_id": "SWE1-HMI-DM-001-01", "reasoning": "", "keywords": []}],
        "cost": 0.0,
        "inputTokens": 0,
        "outputTokens": 0,
        "cacheCreationTokens": 0,
        "cacheReadTokens": 0,
    }

    client.post(
        f"/api/jobs/{job_id}/rerun/stream",
        json={
            "rowIds": ["row-10"],
            "rows": [
                {
                    "id": "row-10",
                    "rowNum": 10,
                    "tcId": "newR1L-DM-001",
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "Fresh payload Test Item",
                    "testSet": "Fresh payload Test Set",
                    "preConditions": "Fresh payload pre-condition",
                    "steps": "Fresh payload procedure",
                    "expectedResults": "Fresh payload expected result",
                    "priority": "P0",
                    "status": "pending",
                }
            ],
        },
    )

    sent_row = mock_tool.call_args.kwargs["rows"][0]
    assert sent_row["test_set"] == "Fresh payload Test Set"
    assert sent_row["test_item"] == "Fresh payload Test Item"
    assert sent_row["pre_conditions"] == "Fresh payload pre-condition"
    assert sent_row["test_procedure"] == "Fresh payload procedure"
    assert sent_row["expected_result"] == "Fresh payload expected result"


@patch("api_server.generate_tc_tool")
def test_rerun_stream_stops_on_quota_exceeded(mock_tool):
    from tools.errors import ToolError

    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    job_id = parse_response.json()["jobId"]
    mock_tool.side_effect = ToolError(
        "generation failed: API call failed: insufficient_quota",
        code="quota_exceeded",
    )

    response = client.post(
        f"/api/jobs/{job_id}/rerun/stream",
        json={
            "rowIds": ["row-10", "row-11"],
            "rows": [
                {
                    "id": "row-10",
                    "rowNum": 10,
                    "tcId": "newR1L-DM-001",
                    "reqId": "SWE1-HMI-DM-001-01",
                    "testItem": "A",
                    "status": "pending",
                },
                {
                    "id": "row-11",
                    "rowNum": 11,
                    "tcId": "newR1L-DM-002",
                    "reqId": "SWE1-HMI-DM-001-02",
                    "testItem": "B",
                    "status": "pending",
                },
            ],
            "config": {
                "model": "gpt-5",
                "batchSize": 1,
                "budget": 0,
                "strictValidation": False,
            },
        },
    )

    assert response.status_code == 200
    events = _parse_sse(response.content)
    types = [event["type"] for event in events]
    assert "rerun.failed" in types
    assert "row.regen_failed" not in types
    assert mock_tool.call_count == 1


@patch("api_server.generate_tc_tool")
def test_generate_stream_passes_matched_reference_context_to_ai(mock_tool):
    parse_response = client.post(
        "/api/parse",
        files={
            "raw_file": (
                "SomeProject_SWQT_DeviceManager_20260408.xlsx",
                _build_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "reference_file": (
                "ReferenceWorkbook.xlsx",
                _build_reference_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )
    parsed = parse_response.json()
    job_id = parsed["jobId"]
    mock_tool.return_value = {
        "tcData": [[VALID_TC_JSON] for _ in parsed["rows"]],
        "splitMeta": [{"req_id": row["reqId"], "reasoning": "", "keywords": []} for row in parsed["rows"]],
        "cost": 0.0,
        "inputTokens": 0,
        "outputTokens": 0,
        "cacheCreationTokens": 0,
        "cacheReadTokens": 0,
    }

    generate_response = client.post(
        "/api/generate",
        json={
            "jobId": job_id,
            "rows": parsed["rows"],
            "config": {
                "model": "gpt-5.4-mini",
                "batchSize": 2,
                "budget": 0,
                "strictValidation": False,
            },
        },
    )
    stream_url = generate_response.json()["streamUrl"]
    client.get(stream_url)

    sent_rows = mock_tool.call_args.kwargs["rows"]
    assert sent_rows[0]["spec_reference"] == "SPEC_REF_PDM01"
    assert "PDM01 behavior description" in sent_rows[0]["matched_spec_context"]


def _parse_sse(content: bytes) -> list[dict]:
    """Parse SSE response body into a list of event dicts."""
    events = []
    for line in content.decode().split("\n"):
        line = line.strip()
        if line.startswith("data: "):
            try:
                events.append(json.loads(line[len("data: "):]))
            except json.JSONDecodeError:
                pass
    return events


@patch("api_server.generate_tcs_for_row")
def test_quick_generate_unified_single_tc(mock_gen):
    """AI 判斷只需要 1 筆 TC 時，quick-generate 仍會走 auto-split 流程，
    event 順序：job.started → decompose.analysis → tc.generating → tc.completed → job.completed。
    """
    from generator import GenerationResult
    mock_gen.return_value = GenerationResult(
        tc_data=[VALID_TC_JSON],
        input_tokens=300,
        output_tokens=150,
        cost=0.003,
        model="gpt-4.1",
        split_meta=[{"req_id": "QUICK", "reasoning": "原子行為，不需拆分。", "keywords": []}],
    )

    response = client.post(
        "/api/quick-generate/stream",
        json={"testItem": "Button pressed → LED on", "context": None, "model": "gpt-5.4-mini"},
    )
    assert response.status_code == 200
    events = _parse_sse(response.content)
    types = [e["type"] for e in events]
    assert types[0] == "job.started"
    assert "decompose.analysis" in types
    assert "tc.completed" in types
    assert types[-1] == "job.completed"

    analysis = next(e for e in events if e["type"] == "decompose.analysis")
    assert len(analysis["scenarios"]) == 1
    assert "原子行為" in analysis["reasoning"]

    tc_done = [e for e in events if e["type"] == "tc.completed"]
    assert len(tc_done) == 1
    assert tc_done[0]["tc"]["priority"] == "P1"


@patch("api_server.generate_tcs_for_row")
def test_quick_generate_unified_multi_tc(mock_gen):
    """AI 拆成 3 筆 TC 時應該發 3 個 tc.completed，analysis.scenarios 對應 3 筆。"""
    from generator import GenerationResult
    tcs = [
        {**VALID_TC_JSON, "tc_title": f"Trigger {i} → outcome {i}"} for i in (1, 2, 3)
    ]
    mock_gen.return_value = GenerationResult(
        tc_data=tcs,
        input_tokens=600,
        output_tokens=900,
        cost=0.012,
        model="gpt-5.4-mini",
        split_meta=[{
            "req_id": "QUICK",
            "reasoning": "§1.4 列舉 3 種格式，各一筆 TC。",
            "keywords": [{"keyword": "format", "meaning": "支援格式", "covered_by": [1, 2, 3]}],
        }],
    )

    response = client.post(
        "/api/quick-generate/stream",
        json={"testItem": "Supports .mp4, .avi, .mpg", "context": None, "model": "gpt-5.4-mini"},
    )
    assert response.status_code == 200
    events = _parse_sse(response.content)
    tc_done = [e for e in events if e["type"] == "tc.completed"]
    assert len(tc_done) == 3

    analysis = next(e for e in events if e["type"] == "decompose.analysis")
    assert len(analysis["scenarios"]) == 3
    assert analysis["scenarios"][0]["name"] == "Trigger 1 → outcome 1"
    assert analysis["keywords"][0]["scenarios"] == [1, 2, 3]

    tc_done = [e for e in events if e["type"] == "tc.completed"]
    assert tc_done[0]["scenarioName"] == "Trigger 1 → outcome 1"


@patch("api_server.generate_tcs_for_row")
def test_quick_generate_with_context_included_in_prompt(mock_gen):
    """Quick Generate should send context through the same row fields used for split analysis."""
    from generator import GenerationResult
    mock_gen.return_value = GenerationResult(
        tc_data=[VALID_TC_JSON],
        input_tokens=100, output_tokens=50, cost=0.001, model="gpt-5.4-mini",
        split_meta=[{"req_id": "QUICK", "reasoning": "", "keywords": []}],
    )

    client.post(
        "/api/quick-generate/stream",
        json={
            "testItem": "Button pressed → LED on",
            "context": "System must be powered",
            "model": "gpt-5.4-mini",
        },
    )
    call_kwargs = mock_gen.call_args[1]
    assert call_kwargs["context"]["test_group"] == "QuickGenerate"
    assert call_kwargs["row"]["test_set"] == "Quick Generate"
    assert call_kwargs["row"]["test_item"] == "Button pressed → LED on"
    assert call_kwargs["row"]["pre_conditions"] == "System must be powered"


@patch("api_server.generate_tcs_for_row")
def test_quick_generate_accepts_structured_workbook_context_fields(mock_gen):
    from generator import GenerationResult
    mock_gen.return_value = GenerationResult(
        tc_data=[VALID_TC_JSON],
        input_tokens=100, output_tokens=50, cost=0.001, model="gpt-5.4-mini",
        split_meta=[{"req_id": "QUICK", "reasoning": "", "keywords": []}],
    )

    client.post(
        "/api/quick-generate/stream",
        json={
            "testItem": "Structured Test Item",
            "context": "fallback context",
            "testGroup": "DeviceManager",
            "testSet": "Status Bar",
            "preConditions": "Structured pre-condition",
            "testProcedure": "Structured procedure",
            "expectedResult": "Structured expected result",
            "model": "gpt-5.4-mini",
        },
    )

    call_kwargs = mock_gen.call_args[1]
    assert call_kwargs["context"]["test_group"] == "DeviceManager"
    assert call_kwargs["row"]["test_set"] == "Status Bar"
    assert call_kwargs["row"]["test_item"] == "Structured Test Item"
    assert call_kwargs["row"]["pre_conditions"] == "Structured pre-condition"
    assert call_kwargs["row"]["test_procedure"] == "Structured procedure"
    assert call_kwargs["row"]["expected_result"] == "Structured expected result"


@patch("api_server.generate_tcs_for_row")
def test_quick_generate_api_error(mock_gen):
    from generator import GenerationError
    mock_gen.side_effect = GenerationError("API timeout")

    response = client.post(
        "/api/quick-generate/stream",
        json={"testItem": "some item", "context": None, "model": "gpt-5.4-mini"},
    )
    assert response.status_code == 200
    events = _parse_sse(response.content)
    types = [e["type"] for e in events]
    assert "job.failed" in types
    failed = next(e for e in events if e["type"] == "job.failed")
    assert "API timeout" in failed["message"]


class TestResolveDuplicateOf:
    """duplicate_of normalizer: AI strings → row_num matched against siblings."""

    siblings = [
        {"id": "uuid-a", "row_num": 11, "test_item": "scenario A"},
        {"id": "uuid-b", "row_num": 14, "test_item": "scenario B"},
    ]

    def test_plain_digit_string_maps_to_row_num(self):
        from api_server import _resolve_duplicate_of

        assert _resolve_duplicate_of("11", self.siblings) == "11"
        assert _resolve_duplicate_of("14", self.siblings) == "14"

    def test_row_prefixed_strings_are_normalized(self):
        from api_server import _resolve_duplicate_of

        assert _resolve_duplicate_of("row #11", self.siblings) == "11"
        assert _resolve_duplicate_of("row 11", self.siblings) == "11"

    def test_legacy_uuid_resolves_to_row_num(self):
        from api_server import _resolve_duplicate_of

        # AI 沒照新指示、回了 sibling 的 uuid → 仍要對應到該 sibling 的 row_num。
        assert _resolve_duplicate_of("uuid-b", self.siblings) == "14"

    def test_unknown_value_returns_empty(self):
        from api_server import _resolve_duplicate_of

        # AI hallucinate（回 row_num 99 但沒有那個 sibling）→ 直接回空，前端
        # 不顯示「(已刪除或未找到)」這種誤導訊息。
        assert _resolve_duplicate_of("99", self.siblings) == ""
        assert _resolve_duplicate_of("garbage", self.siblings) == ""

    def test_empty_or_no_siblings_returns_empty(self):
        from api_server import _resolve_duplicate_of

        assert _resolve_duplicate_of("11", None) == ""
        assert _resolve_duplicate_of("11", []) == ""
        assert _resolve_duplicate_of("", self.siblings) == ""
        assert _resolve_duplicate_of(None, self.siblings) == ""


class TestResequenceExportTcIds:
    """Export-time renumbering: gaps from deletes are closed before write."""

    def _row(self, tc_id, row_num):
        return {"tc_id": tc_id, "row_num": row_num, "req_id": "R-1"}

    def test_no_change_when_already_contiguous(self):
        from api_server import _resequence_export_tc_ids

        rows = [self._row("newR1L-DM-001", 10), self._row("newR1L-DM-002", 11)]
        changed = _resequence_export_tc_ids(rows)
        assert changed == 0
        assert [r["tc_id"] for r in rows] == ["newR1L-DM-001", "newR1L-DM-002"]

    def test_gaps_are_closed_within_bucket(self):
        from api_server import _resequence_export_tc_ids

        # User deleted -002 and -004; remaining should renumber to 001/002/003.
        rows = [
            self._row("newR1L-DM-001", 10),
            self._row("newR1L-DM-003", 12),
            self._row("newR1L-DM-005", 14),
        ]
        changed = _resequence_export_tc_ids(rows)
        assert changed == 2  # 003→002, 005→003
        assert [r["tc_id"] for r in rows] == [
            "newR1L-DM-001",
            "newR1L-DM-002",
            "newR1L-DM-003",
        ]

    def test_separate_buckets_renumber_independently(self):
        from api_server import _resequence_export_tc_ids

        rows = [
            self._row("newR1L-DM-003", 10),
            self._row("newR1L-MP-007", 11),
            self._row("newR1L-DM-009", 12),
        ]
        _resequence_export_tc_ids(rows)
        # DM bucket (rows 10, 12) → 001, 002
        # MP bucket (row 11) → 001
        by_id = {r["tc_id"] for r in rows}
        assert "newR1L-DM-001" in by_id
        assert "newR1L-DM-002" in by_id
        assert "newR1L-MP-001" in by_id

    def test_rows_without_tc_id_are_skipped(self):
        from api_server import _resequence_export_tc_ids

        rows = [
            {"tc_id": "", "row_num": 10},
            self._row("newR1L-DM-005", 11),
        ]
        _resequence_export_tc_ids(rows)
        assert rows[0]["tc_id"] == ""
        assert rows[1]["tc_id"] == "newR1L-DM-001"

    def test_row_order_drives_assignment(self):
        from api_server import _resequence_export_tc_ids

        # rows out of row_num order → renumber follows row_num ascending.
        rows = [
            self._row("newR1L-DM-005", 14),
            self._row("newR1L-DM-002", 11),
            self._row("newR1L-DM-008", 17),
        ]
        _resequence_export_tc_ids(rows)
        # After: row_num 11 → 001, row_num 14 → 002, row_num 17 → 003.
        by_row = {r["row_num"]: r["tc_id"] for r in rows}
        assert by_row[11] == "newR1L-DM-001"
        assert by_row[14] == "newR1L-DM-002"
        assert by_row[17] == "newR1L-DM-003"


class TestPrepareGenerationRowsSiblings:
    """Sibling annotation: rows sharing reqId carry each other's test_items."""

    def _job(self, rows_input):
        return {
            "rows": rows_input,
            "parsedData": {
                "rows": [
                    {"row_num": r.get("rowNum"), "req_id": r["reqId"], "test_item": r.get("testItem", "")}
                    for r in rows_input
                ],
                "project": "newR1L",
                "test_group": "DeviceManager",
            },
            "config": {"model": "gpt-test"},
        }

    def test_unique_req_ids_get_no_siblings(self):
        from api_server import _prepare_generation_rows

        job = self._job([
            {"id": "a", "rowNum": 10, "reqId": "R-1", "testItem": "alpha"},
            {"id": "b", "rowNum": 11, "reqId": "R-2", "testItem": "beta"},
        ])
        prepared = _prepare_generation_rows(job)
        for row in prepared:
            assert "siblings" not in row or not row["siblings"]

    def test_duplicate_req_id_annotates_other_rows_as_siblings(self):
        from api_server import _prepare_generation_rows

        job = self._job([
            {"id": "a", "rowNum": 10, "reqId": "R-DUP", "testItem": "scenario A"},
            {"id": "b", "rowNum": 11, "reqId": "R-DUP", "testItem": "scenario B"},
            {"id": "c", "rowNum": 12, "reqId": "R-DUP", "testItem": "scenario C"},
            {"id": "d", "rowNum": 13, "reqId": "R-OTHER", "testItem": "alone"},
        ])
        prepared = _prepare_generation_rows(job)
        by_id = {r["id"]: r for r in prepared}

        # Each duplicate row sees the other two — and only those.
        siblings_a = {s["id"] for s in by_id["a"].get("siblings", [])}
        siblings_b = {s["id"] for s in by_id["b"].get("siblings", [])}
        siblings_c = {s["id"] for s in by_id["c"].get("siblings", [])}
        assert siblings_a == {"b", "c"}
        assert siblings_b == {"a", "c"}
        assert siblings_c == {"a", "b"}

        # Self never appears as own sibling.
        assert all(s["id"] != "a" for s in by_id["a"]["siblings"])

        # The standalone row gets no siblings.
        assert not by_id["d"].get("siblings")

        # Sibling entries carry test_item content for AI context.
        a_sibs = {s["id"]: s["test_item"] for s in by_id["a"]["siblings"]}
        assert a_sibs["b"] == "scenario B"
        assert a_sibs["c"] == "scenario C"


# ─────────────────────────────────────────────────────────────────────
# Telemetry collector
# ─────────────────────────────────────────────────────────────────────


def test_events_appends_to_jsonl(tmp_path, monkeypatch):
    log = tmp_path / "events.jsonl"
    monkeypatch.setenv("TC_EVENTS_LOG", str(log))

    response = client.post(
        "/api/events",
        json={
            "events": [
                {"name": "home_new_run_click", "props": {"source": "top-nav"}},
                {
                    "name": "builder_step_next",
                    "props": {"from": "data", "to": "configure"},
                    "ts": 1714000000000,
                },
            ]
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["ok"] is True
    assert body["count"] == 2

    lines = log.read_text("utf-8").splitlines()
    assert len(lines) == 2
    first = json.loads(lines[0])
    assert first["name"] == "home_new_run_click"
    assert first["props"]["source"] == "top-nav"
    assert "received_at" in first
    second = json.loads(lines[1])
    assert second["ts"] == 1714000000000


def test_events_accepts_experiment_exposure_and_assignment_map(tmp_path, monkeypatch):
    log = tmp_path / "events.jsonl"
    monkeypatch.setenv("TC_EVENTS_LOG", str(log))

    response = client.post(
        "/api/events",
        json={
            "events": [
                {
                    "name": "experiment_exposure",
                    "props": {
                        "experiment": "home_layout_emphasis",
                        "variant": "action_first",
                    },
                    "experiments": {
                        "home_layout_emphasis": "action_first",
                    },
                }
            ]
        },
    )

    assert response.status_code == 200
    record = json.loads(log.read_text("utf-8").splitlines()[0])
    assert record["name"] == "experiment_exposure"
    assert record["experiments"] == {
        "home_layout_emphasis": "action_first",
    }


def test_events_aggregate_empty_log(tmp_path, monkeypatch):
    monkeypatch.setenv("TC_EVENTS_LOG", str(tmp_path / "missing.jsonl"))

    response = client.get("/api/events/aggregate")

    assert response.status_code == 200
    payload = response.json()
    assert payload["totalEvents"] == 0
    assert payload["variants"] == {}
    assert payload["unknownVariant"]["eventCount"] == 0


def test_events_aggregate_by_experiment_variant(tmp_path, monkeypatch):
    log = tmp_path / "events.jsonl"
    monkeypatch.setenv("TC_EVENTS_LOG", str(log))
    records = [
        {
            "name": "experiment_exposure",
            "props": {
                "experiment": "home_layout_emphasis",
                "variant": "action_first",
            },
            "experiments": {"home_layout_emphasis": "action_first"},
            "ts": 1,
        },
        {
            "name": "home_new_run_click",
            "props": {"source": "quick-actions"},
            "experiments": {"home_layout_emphasis": "action_first"},
            "ts": 2,
        },
        {
            "name": "run_execute_start",
            "props": {"jobId": "j1", "rowCount": 4},
            "experiments": {"home_layout_emphasis": "action_first"},
            "ts": 3,
        },
        {
            "name": "run_execute_success",
            "props": {"jobId": "j1", "rowCount": 4},
            "experiments": {"home_layout_emphasis": "action_first"},
            "ts": 4,
        },
        {
            "name": "run_execute_fail",
            "props": {"jobId": "j2", "reason": "boom"},
            "experiments": {"home_layout_emphasis": "kpi_first"},
            "ts": 5,
        },
        {"bad": "json"},
    ]
    log.write_text(
        "\n".join(json.dumps(r) for r in records) + "\nnot-json\n",
        encoding="utf-8",
    )

    response = client.get(
        "/api/events/aggregate?experiment=home_layout_emphasis"
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["experiment"] == "home_layout_emphasis"
    assert payload["totalEvents"] == 6
    assert payload["malformedLines"] == 1
    action = payload["variants"]["action_first"]
    assert action["eventCount"] == 4
    assert action["exposures"] == 1
    assert action["newRunClicks"] == 1
    assert action["runStarts"] == 1
    assert action["runSuccesses"] == 1
    assert action["runFailures"] == 0
    assert action["completionRate"] == 1.0
    assert action["failureRate"] == 0.0
    kpi = payload["variants"]["kpi_first"]
    assert kpi["runFailures"] == 1
    assert kpi["completionRate"] == 0.0
    assert kpi["failureRate"] == 1.0


def test_events_aggregate_unknown_variant_bucket(tmp_path, monkeypatch):
    log = tmp_path / "events.jsonl"
    monkeypatch.setenv("TC_EVENTS_LOG", str(log))
    log.write_text(
        json.dumps({"name": "home_new_run_click", "props": {}, "ts": 1}) + "\n",
        encoding="utf-8",
    )

    response = client.get(
        "/api/events/aggregate?experiment=home_layout_emphasis"
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["unknownVariant"]["eventCount"] == 1
    assert payload["unknownVariant"]["newRunClicks"] == 1


def test_events_rejects_unknown_name(tmp_path, monkeypatch):
    monkeypatch.setenv("TC_EVENTS_LOG", str(tmp_path / "events.jsonl"))
    response = client.post(
        "/api/events",
        json={"events": [{"name": "totally_made_up", "props": {}}]},
    )
    assert response.status_code == 400
    assert "unknown event" in response.json()["detail"]


def test_events_rejects_empty_batch(tmp_path, monkeypatch):
    monkeypatch.setenv("TC_EVENTS_LOG", str(tmp_path / "events.jsonl"))
    response = client.post("/api/events", json={"events": []})
    assert response.status_code == 400


def test_events_rejects_too_large_batch(tmp_path, monkeypatch):
    monkeypatch.setenv("TC_EVENTS_LOG", str(tmp_path / "events.jsonl"))
    payload = {
        "events": [
            {"name": "home_new_run_click", "props": {}}
            for _ in range(101)
        ]
    }
    response = client.post("/api/events", json=payload)
    assert response.status_code == 400
    assert "batch too large" in response.json()["detail"]


def test_events_rejects_oversized_props(tmp_path, monkeypatch):
    monkeypatch.setenv("TC_EVENTS_LOG", str(tmp_path / "events.jsonl"))
    huge = "x" * (5 * 1024)  # > 4 KB
    response = client.post(
        "/api/events",
        json={
            "events": [
                {"name": "builder_validation_fail", "props": {"detail": huge}}
            ]
        },
    )
    assert response.status_code == 413


# ─────────────────────────────────────────────────────────────────────
# Job timeline
# ─────────────────────────────────────────────────────────────────────


def test_timeline_404_for_unknown_job():
    response = client.get("/api/jobs/does-not-exist/timeline")
    assert response.status_code == 404


def test_timeline_records_queued_after_generate():
    """POST /api/generate 應寫入 'queued' timeline event。"""
    import api_server

    job_id = "timeline-job-1"
    api_server.JOB_REGISTRY[job_id] = {
        "jobId": job_id,
        "parsedData": {"project": "P", "test_group": "G"},
    }

    payload = {
        "jobId": job_id,
        "rows": [
            {
                "id": "r1",
                "reqId": "REQ-1",
                "testGroup": "G",
                "testSet": "Set",
                "testItem": "Item",
                "preConditions": "",
                "inputTestData": "",
                "steps": "",
                "expectedResults": "",
                "status": "pending",
            }
        ],
        "config": {
            "model": "gpt-5",
            "batchSize": 5,
            "budget": 10,
            "strictValidation": False,
        },
    }
    res = client.post("/api/generate", json=payload)
    assert res.status_code == 200, res.text

    response = client.get(f"/api/jobs/{job_id}/timeline")
    assert response.status_code == 200
    events = response.json()["events"]
    assert events[0]["kind"] == "queued"
    assert events[0]["rowCount"] == 1
    assert isinstance(events[0]["ts"], int)


def test_timeline_helper_running_completed_messages():
    import api_server

    job_id = "timeline-job-helper"
    api_server.JOB_REGISTRY[job_id] = {"jobId": job_id}
    api_server._append_job_event(job_id, "running", rowCount=3)
    api_server._append_job_event(
        job_id, "completed", rowCount=3, processed=3, cost=0.05
    )

    events = client.get(f"/api/jobs/{job_id}/timeline").json()["events"]
    assert [e["kind"] for e in events] == ["running", "completed"]
    assert events[1]["processed"] == 3
    assert events[1]["cost"] == 0.05


def test_timeline_caps_at_50_events():
    import api_server

    job_id = "timeline-job-bound"
    api_server.JOB_REGISTRY[job_id] = {"jobId": job_id}
    for i in range(60):
        api_server._append_job_event(job_id, "running", rowCount=i)
    events = client.get(f"/api/jobs/{job_id}/timeline").json()["events"]
    assert len(events) == 50
    # 上限保護下保留最後 50 筆
    assert events[0]["rowCount"] == 10
    assert events[-1]["rowCount"] == 59


# ─────────────────────────────────────────────────────────────────────
# Output compare
# ─────────────────────────────────────────────────────────────────────


def _build_export_workbook_bytes(rows: list[dict]) -> bytes:
    """Tiny TC workbook matching writer's column layout."""
    from writer import TC_SHEET_NAME, WRITE_COLUMNS

    wb = Workbook()
    ws = wb.active
    ws.title = TC_SHEET_NAME
    for ridx, row in enumerate(rows, start=10):
        if "reqId" in row:
            ws.cell(row=ridx, column=4, value=row["reqId"])
        for field, col in WRITE_COLUMNS.items():
            if field in row:
                ws.cell(row=ridx, column=col, value=row[field])
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_compare_diff_tcs_changed_added_removed(tmp_path):
    import api_server

    a_path = tmp_path / "a.xlsx"
    b_path = tmp_path / "b.xlsx"
    a_path.write_bytes(
        _build_export_workbook_bytes(
            [
                {
                    "reqId": "REQ-1",
                    "tc_id": "T-001",
                    "pre_conditions": "old pre",
                    "test_procedure": "step",
                    "expected_result": "ok",
                },
                {
                    "reqId": "REQ-2",
                    "tc_id": "T-002",
                    "pre_conditions": "stable",
                },
            ]
        )
    )
    b_path.write_bytes(
        _build_export_workbook_bytes(
            [
                {
                    "reqId": "REQ-1",
                    "tc_id": "T-001",
                    "pre_conditions": "new pre",  # changed
                    "test_procedure": "step",
                    "expected_result": "ok",
                },
                {
                    "reqId": "REQ-3",  # added
                    "tc_id": "T-003",
                    "pre_conditions": "fresh",
                },
            ]
        )
    )

    api_server.JOB_REGISTRY["job-a"] = {"jobId": "job-a", "exportPath": str(a_path)}
    api_server.JOB_REGISTRY["job-b"] = {"jobId": "job-b", "exportPath": str(b_path)}

    response = client.post(
        "/api/outputs/compare", json={"a": "job-a", "b": "job-b"}
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["summary"]["total"] == 3
    assert body["summary"]["changed"] == 1
    assert body["summary"]["added"] == 1
    assert body["summary"]["removed"] == 1

    by_id = {r["tcId"]: r for r in body["rows"]}
    assert by_id["T-001"]["status"] == "changed"
    fields = {c["field"] for c in by_id["T-001"]["changes"]}
    assert "pre_conditions" in fields
    assert by_id["T-002"]["status"] == "removed"
    assert by_id["T-003"]["status"] == "added"


def test_compare_rejects_same_job():
    import api_server

    api_server.JOB_REGISTRY["job-self"] = {
        "jobId": "job-self",
        "exportPath": "/nope.xlsx",
    }
    response = client.post(
        "/api/outputs/compare", json={"a": "job-self", "b": "job-self"}
    )
    assert response.status_code == 400


def test_compare_404_when_job_missing():
    response = client.post(
        "/api/outputs/compare",
        json={"a": "ghost-1", "b": "ghost-2"},
    )
    assert response.status_code == 404


def test_compare_409_when_export_missing():
    import api_server

    api_server.JOB_REGISTRY["job-no-export-a"] = {"jobId": "job-no-export-a"}
    api_server.JOB_REGISTRY["job-no-export-b"] = {"jobId": "job-no-export-b"}
    response = client.post(
        "/api/outputs/compare",
        json={"a": "job-no-export-a", "b": "job-no-export-b"},
    )
    assert response.status_code == 409
    assert "Export first" in response.json()["detail"]


# ─────────────────────────────────────────────────────────────────────
# Bulk output download (zip)
# ─────────────────────────────────────────────────────────────────────


def test_bulk_download_returns_zip(tmp_path):
    import api_server
    import io
    import zipfile

    a_path = tmp_path / "alpha.xlsx"
    b_path = tmp_path / "beta.xlsx"
    a_path.write_bytes(_build_export_workbook_bytes([{"reqId": "REQ-A", "tc_id": "T-A"}]))
    b_path.write_bytes(_build_export_workbook_bytes([{"reqId": "REQ-B", "tc_id": "T-B"}]))

    api_server.JOB_REGISTRY["bulk-a"] = {"jobId": "bulk-a", "exportPath": str(a_path)}
    api_server.JOB_REGISTRY["bulk-b"] = {"jobId": "bulk-b", "exportPath": str(b_path)}

    res = client.post(
        "/api/outputs/bulk-download",
        json={"jobIds": ["bulk-a", "bulk-b"]},
    )
    assert res.status_code == 200
    assert res.headers["content-type"].startswith("application/zip")
    cd = res.headers.get("content-disposition", "")
    assert ".zip" in cd
    with zipfile.ZipFile(io.BytesIO(res.content)) as zf:
        names = zf.namelist()
        assert "alpha.xlsx" in names
        assert "beta.xlsx" in names


def test_bulk_download_dedupes_and_handles_collisions(tmp_path):
    """Repeating the same jobId is silently dropped; same filename for two jobs gets a -2 suffix."""
    import api_server
    import io
    import zipfile

    same_name_a = tmp_path / "dup.xlsx"
    same_name_a.write_bytes(_build_export_workbook_bytes([{"reqId": "A", "tc_id": "T-A"}]))
    sub = tmp_path / "sub"
    sub.mkdir()
    same_name_b = sub / "dup.xlsx"
    same_name_b.write_bytes(_build_export_workbook_bytes([{"reqId": "B", "tc_id": "T-B"}]))

    api_server.JOB_REGISTRY["bulk-d1"] = {"jobId": "bulk-d1", "exportPath": str(same_name_a)}
    api_server.JOB_REGISTRY["bulk-d2"] = {"jobId": "bulk-d2", "exportPath": str(same_name_b)}

    res = client.post(
        "/api/outputs/bulk-download",
        json={"jobIds": ["bulk-d1", "bulk-d1", "bulk-d2"]},
    )
    assert res.status_code == 200
    with zipfile.ZipFile(io.BytesIO(res.content)) as zf:
        names = zf.namelist()
        assert "dup.xlsx" in names
        assert "dup-2.xlsx" in names
        assert len(names) == 2


def test_bulk_download_empty_jobids_400():
    res = client.post("/api/outputs/bulk-download", json={"jobIds": []})
    assert res.status_code == 400


def test_bulk_download_too_many_jobids_400():
    res = client.post(
        "/api/outputs/bulk-download",
        json={"jobIds": [f"x{i}" for i in range(60)]},
    )
    assert res.status_code == 400
    assert "too many" in res.json()["detail"]


def test_bulk_download_404_unknown_job():
    res = client.post(
        "/api/outputs/bulk-download",
        json={"jobIds": ["bulk-ghost"]},
    )
    assert res.status_code == 404


def test_bulk_download_409_no_export(tmp_path):
    import api_server

    api_server.JOB_REGISTRY["bulk-noexp"] = {"jobId": "bulk-noexp"}
    res = client.post(
        "/api/outputs/bulk-download",
        json={"jobIds": ["bulk-noexp"]},
    )
    assert res.status_code == 409


# ─────────────────────────────────────────────────────────────────────
# Job config snapshot
# ─────────────────────────────────────────────────────────────────────


def test_config_404_for_unknown_job():
    response = client.get("/api/jobs/nope/config")
    assert response.status_code == 404


def test_config_returns_snapshot():
    import api_server

    job_id = "config-snapshot-job"
    api_server.JOB_REGISTRY[job_id] = {
        "jobId": job_id,
        "status": "completed",
        "totalRows": 4,
        "parsedData": {"project": "Demo", "test_group": "Core"},
        "config": {
            "model": "gpt-5",
            "batchSize": 5,
            "budget": 12.5,
            "strictValidation": True,
        },
    }
    response = client.get(f"/api/jobs/{job_id}/config")
    assert response.status_code == 200
    body = response.json()
    assert body["jobId"] == job_id
    assert body["projectName"] == "Demo"
    assert body["testGroup"] == "Core"
    assert body["totalRows"] == 4
    assert body["status"] == "completed"
    assert body["config"]["model"] == "gpt-5"
    assert body["config"]["strictValidation"] is True


def test_config_returns_none_when_no_config_stored():
    import api_server

    job_id = "config-no-snapshot"
    api_server.JOB_REGISTRY[job_id] = {"jobId": job_id}
    body = client.get(f"/api/jobs/{job_id}/config").json()
    assert body["config"] is None
    assert body["projectName"] is None


# ─────────────────────────────────────────────────────────────────────
# Validation logs
# ─────────────────────────────────────────────────────────────────────


def test_validation_logs_append_and_get():
    import api_server

    job_id = "vl-job-1"
    api_server.JOB_REGISTRY[job_id] = {"jobId": job_id}

    res = client.post(
        f"/api/jobs/{job_id}/validation-logs",
        json={
            "entries": [
                {
                    "rowId": "r1",
                    "reqId": "REQ-1",
                    "severity": "error",
                    "field": "steps",
                    "message": "missing steps",
                },
                {
                    "rowId": "r2",
                    "reqId": "REQ-2",
                    "severity": "warning",
                    "message": "soft warning",
                },
            ]
        },
    )
    assert res.status_code == 200
    assert res.json() == {"ok": True, "count": 2}

    body = client.get(f"/api/jobs/{job_id}/validation-logs").json()
    rows = {e["rowId"]: e for e in body["entries"]}
    assert rows["r1"]["severity"] == "error"
    assert rows["r1"]["field"] == "steps"
    assert rows["r2"]["severity"] == "warning"
    assert all(isinstance(e["ts"], int) for e in body["entries"])


def test_validation_logs_replace_by_row_id():
    import api_server

    job_id = "vl-job-replace"
    api_server.JOB_REGISTRY[job_id] = {"jobId": job_id}

    client.post(
        f"/api/jobs/{job_id}/validation-logs",
        json={"entries": [{"rowId": "r1", "severity": "error", "message": "v1"}]},
    )
    client.post(
        f"/api/jobs/{job_id}/validation-logs",
        json={"entries": [{"rowId": "r1", "severity": "warning", "message": "v2"}]},
    )
    entries = client.get(f"/api/jobs/{job_id}/validation-logs").json()["entries"]
    assert len(entries) == 1
    assert entries[0]["message"] == "v2"


def test_record_stream_validation_failure_writes_log_entry():
    import api_server

    job_id = "stream-vlog-job"
    api_server.JOB_REGISTRY[job_id] = {"jobId": job_id}
    api_server._record_stream_validation_failure(
        job_id,
        {"id": "row-7", "req_id": "REQ-7", "row_num": 7},
        "Generation failed: quota",
        field="generation",
    )
    body = client.get(f"/api/jobs/{job_id}/validation-logs").json()
    assert len(body["entries"]) == 1
    entry = body["entries"][0]
    assert entry["rowId"] == "row-7"
    assert entry["reqId"] == "REQ-7"
    assert entry["severity"] == "error"
    assert entry["field"] == "generation"
    assert "quota" in entry["message"]
    assert isinstance(entry["ts"], int)


def test_record_stream_validation_failure_replaces_per_row():
    import api_server

    job_id = "stream-vlog-replace"
    api_server.JOB_REGISTRY[job_id] = {"jobId": job_id}
    api_server._record_stream_validation_failure(
        job_id, {"id": "row-1", "req_id": "REQ-1"}, "first"
    )
    api_server._record_stream_validation_failure(
        job_id, {"id": "row-1", "req_id": "REQ-1"}, "second"
    )
    entries = client.get(f"/api/jobs/{job_id}/validation-logs").json()["entries"]
    assert len(entries) == 1
    assert entries[0]["message"] == "second"


def test_validation_logs_404_unknown_job():
    assert client.get("/api/jobs/no-such/validation-logs").status_code == 404
    res = client.post(
        "/api/jobs/no-such/validation-logs",
        json={"entries": [{"rowId": "r1", "severity": "error", "message": "x"}]},
    )
    assert res.status_code == 404


# ─────────────────────────────────────────────────────────────────────
# Template usage analytics
# ─────────────────────────────────────────────────────────────────────


def test_spec_library_usage_no_matches():
    response = client.get("/api/spec-library/template-with-no-runs/usage")
    assert response.status_code == 200
    assert response.json() == {
        "name": "template-with-no-runs",
        "usageCount": 0,
        "lastUsedAt": None,
        "recentRunIds": [],
    }


def test_spec_library_usage_counts_matching_jobs():
    import api_server

    api_server.JOB_REGISTRY["tpl-usage-a"] = {
        "jobId": "tpl-usage-a",
        "templateId": "tpl-X",
        "status": "completed",
        "timeline": [{"kind": "queued", "ts": 100}],
    }
    api_server.JOB_REGISTRY["tpl-usage-b"] = {
        "jobId": "tpl-usage-b",
        "templateId": "tpl-X",
        "status": "completed",
        "timeline": [{"kind": "completed", "ts": 200}],
    }
    api_server.JOB_REGISTRY["tpl-usage-other"] = {
        "jobId": "tpl-usage-other",
        "templateId": "tpl-Y",
    }

    body = client.get("/api/spec-library/tpl-X/usage").json()
    assert body["usageCount"] == 2
    assert body["lastUsedAt"] == 200
    assert body["recentRunIds"][0] == "tpl-usage-b"
    assert "tpl-usage-other" not in body["recentRunIds"]


# ─────────────────────────────────────────────────────────────────────
# Dataset re-hydrate
# ─────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────
# Spec library changelog
# ─────────────────────────────────────────────────────────────────────


def _seed_manifest(tmp_path, specs: list[dict]) -> str:
    path = tmp_path / "manifest.json"
    path.write_text(json.dumps({"specs": specs}, ensure_ascii=False, indent=2))
    return str(path)


def test_spec_library_returns_version_and_changelog(tmp_path, monkeypatch):
    manifest = _seed_manifest(
        tmp_path,
        [
            {
                "name": "tpl-A",
                "source_file": "a.xlsx",
                "entries_count": 5,
                "version": "1.2.0",
                "changelog": [
                    {"version": "1.0.0", "message": "initial", "ts": 1},
                ],
            }
        ],
    )
    monkeypatch.setenv("TC_SPEC_INDEX_MANIFEST", manifest)
    body = client.get("/api/spec-library").json()
    assert len(body["specs"]) == 1
    spec = body["specs"][0]
    assert spec["version"] == "1.2.0"
    assert spec["changelog"][0]["message"] == "initial"


def test_spec_library_changelog_append(tmp_path, monkeypatch):
    manifest = _seed_manifest(
        tmp_path,
        [{"name": "tpl-CL", "version": "1.0.0", "changelog": []}],
    )
    monkeypatch.setenv("TC_SPEC_INDEX_MANIFEST", manifest)

    res = client.post(
        "/api/spec-library/tpl-CL/changelog",
        json={"version": "1.1.0", "message": "added boundary cases"},
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["ok"] is True
    assert body["spec"]["version"] == "1.1.0"
    assert body["spec"]["changelog"][-1]["message"] == "added boundary cases"

    persisted = json.loads(open(manifest).read())
    assert persisted["specs"][0]["version"] == "1.1.0"
    assert (
        persisted["specs"][0]["changelog"][-1]["message"]
        == "added boundary cases"
    )


def test_spec_library_changelog_404_unknown(tmp_path, monkeypatch):
    manifest = _seed_manifest(tmp_path, [])
    monkeypatch.setenv("TC_SPEC_INDEX_MANIFEST", manifest)
    res = client.post(
        "/api/spec-library/ghost/changelog",
        json={"message": "anything"},
    )
    assert res.status_code == 404


def test_spec_library_changelog_rejects_empty_message(tmp_path, monkeypatch):
    manifest = _seed_manifest(tmp_path, [{"name": "tpl-X", "changelog": []}])
    monkeypatch.setenv("TC_SPEC_INDEX_MANIFEST", manifest)
    res = client.post(
        "/api/spec-library/tpl-X/changelog",
        json={"message": "   "},
    )
    assert res.status_code == 400


def test_output_preview_404_for_unknown_job():
    assert client.get("/api/jobs/no-such/output-preview").status_code == 404


def test_output_preview_409_when_no_export():
    import api_server

    api_server.JOB_REGISTRY["preview-no-export"] = {"jobId": "preview-no-export"}
    res = client.get("/api/jobs/preview-no-export/output-preview")
    assert res.status_code == 409


def test_output_preview_returns_rows(tmp_path):
    import api_server

    path = tmp_path / "preview.xlsx"
    path.write_bytes(
        _build_export_workbook_bytes(
            [
                {"reqId": "REQ-1", "tc_id": "T-001", "pre_conditions": "p1"},
                {"reqId": "REQ-2", "tc_id": "T-002", "pre_conditions": "p2"},
                {"reqId": "REQ-3", "tc_id": "T-003", "pre_conditions": "p3"},
            ]
        )
    )
    api_server.JOB_REGISTRY["preview-job"] = {
        "jobId": "preview-job",
        "exportPath": str(path),
    }
    body = client.get("/api/jobs/preview-job/output-preview").json()
    assert body["jobId"] == "preview-job"
    assert body["totalRows"] == 3
    assert body["limit"] == 200
    assert len(body["rows"]) == 3
    assert body["rows"][0]["tc_id"] == "T-001"


def test_output_preview_respects_limit(tmp_path):
    import api_server

    path = tmp_path / "preview.xlsx"
    path.write_bytes(
        _build_export_workbook_bytes(
            [
                {"reqId": f"REQ-{i}", "tc_id": f"T-{i:03d}"}
                for i in range(5)
            ]
        )
    )
    api_server.JOB_REGISTRY["preview-limit-job"] = {
        "jobId": "preview-limit-job",
        "exportPath": str(path),
    }
    body = client.get(
        "/api/jobs/preview-limit-job/output-preview?limit=2"
    ).json()
    assert body["totalRows"] == 5
    assert body["limit"] == 2
    assert len(body["rows"]) == 2


def test_dataset_404_for_unknown_job():
    assert client.get("/api/jobs/no-such/dataset").status_code == 404


def test_dataset_returns_camel_case_rows_for_builder():
    import api_server

    job_id = "dataset-rehydrate-job"
    api_server.JOB_REGISTRY[job_id] = {
        "jobId": job_id,
        "parsedData": {
            "project": "Demo",
            "test_group": "Core",
            "rows": [
                {
                    "row_num": 10,
                    "req_id": "REQ-1",
                    "test_set": "Workbook Set",
                    "test_item": "First requirement",
                },
                {
                    "row_num": 11,
                    "req_id": "REQ-2",
                    "test_set": "Workbook Set",
                    "test_item": "Second requirement",
                },
            ],
        },
        "rows": [
            {
                "id": "row-10",
                "rowNum": 10,
                "reqId": "REQ-1",
                "testItem": "First requirement",
                "testSet": "User Override",
            }
        ],
    }
    body = client.get(f"/api/jobs/{job_id}/dataset").json()
    assert body["jobId"] == job_id
    assert body["projectName"] == "Demo"
    assert body["testGroup"] == "Core"
    assert body["rowCount"] == 2

    by_req = {r["reqId"]: r for r in body["rows"]}
    # 用戶覆寫的 testSet 應優先於 parser 原值
    assert by_req["REQ-1"]["testSet"] == "User Override"
    assert by_req["REQ-2"]["testSet"] == "Workbook Set"
    # generated content 還沒填，狀態應為 pending
    assert by_req["REQ-1"]["status"] == "pending"
    assert by_req["REQ-1"]["preConditions"] == ""
    assert by_req["REQ-1"]["testGroup"] == "Core"
    assert by_req["REQ-2"]["id"] == "row-11"


def test_dataset_empty_rows_when_no_parsed_data():
    import api_server

    job_id = "dataset-no-parsed"
    api_server.JOB_REGISTRY[job_id] = {"jobId": job_id}
    body = client.get(f"/api/jobs/{job_id}/dataset").json()
    assert body["rowCount"] == 0
    assert body["rows"] == []
    assert body["projectName"] is None


def test_generate_persists_template_id_for_attribution():
    import api_server

    job_id = "tpl-attrib-job"
    api_server.JOB_REGISTRY[job_id] = {"jobId": job_id}
    payload = {
        "jobId": job_id,
        "templateId": "tpl-Z",
        "rows": [
            {
                "id": "r1",
                "reqId": "REQ-1",
                "testGroup": "G",
                "testSet": "S",
                "testItem": "Item",
                "preConditions": "",
                "inputTestData": "",
                "steps": "",
                "expectedResults": "",
                "status": "pending",
            }
        ],
        "config": {
            "model": "gpt-5",
            "batchSize": 5,
            "budget": 5,
            "strictValidation": False,
        },
    }
    res = client.post("/api/generate", json=payload)
    assert res.status_code == 200
    assert api_server.JOB_REGISTRY[job_id]["templateId"] == "tpl-Z"
    usage = client.get("/api/spec-library/tpl-Z/usage").json()
    assert usage["usageCount"] >= 1
    assert job_id in usage["recentRunIds"]
