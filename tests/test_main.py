"""Tests for CLI entry point."""
import pytest
from openpyxl import Workbook
from unittest.mock import patch
from openpyxl import load_workbook

from main import parse_args, _filter_rows, _estimate_cost, _reject_generate_flags, main, run


@pytest.fixture
def sample_xlsx(tmp_path):
    """Create minimal TC xlsx for CLI testing."""
    filepath = tmp_path / "Test_SWQT_DeviceManager_20260408.xlsx"
    wb = Workbook()
    ws_pd = wb.active
    ws_pd.title = "Product Document"
    ws_pd.cell(row=3, column=2, value="newR1L")

    ws_tc = wb.create_sheet("Test Case Specification&Result")
    for col, name in {4: "Req ID", 6: "TC ID", 9: "Test Item"}.items():
        ws_tc.cell(row=9, column=col, value=name)

    ws_tc.cell(row=10, column=4, value="SWE1-HMI-DM-001-01")
    ws_tc.cell(row=10, column=9, value="PDM01 test feature A")
    ws_tc.cell(row=11, column=4, value="SWE1-HMI-DM-002-01")
    ws_tc.cell(row=11, column=9, value="PDM02 test feature B")

    wb.create_sheet("Test Case Framework")
    wb.save(filepath)
    return str(filepath)


class TestParseArgs:
    def test_required_input(self):
        args = parse_args(["--input", "test.xlsx"])
        assert args.input == "test.xlsx"
        assert args.mode == "full"
        assert args.dry_run is False

    def test_all_options(self):
        args = parse_args([
            "--input", "in.xlsx",
            "--sys1", "sys1.xlsx",
            "--spec", "spec.pdf",
            "--model", "gpt-4.1-mini",
            "--batch-size", "10",
            "--mode", "incremental",
            "--dry-run",
            "--budget", "2.5",
        ])
        assert args.sys1 == "sys1.xlsx"
        assert args.batch_size == 10
        assert args.mode == "incremental"
        assert args.dry_run is True
        assert args.budget == 2.5

    def test_strict_validation_flag(self):
        args = parse_args(["--input", "test.xlsx", "--strict-validation"])
        assert args.strict_validation is True


class TestFilterRows:
    def test_full_mode(self):
        rows = [{"row_num": 10}, {"row_num": 11}]
        assert len(_filter_rows(rows, "full", None)) == 2

    def test_incremental_skips_existing(self):
        rows = [
            {"row_num": 10, "tc_id": "p-A-001"},
            {"row_num": 11, "tc_id": ""},
            {"row_num": 12},
        ]
        result = _filter_rows(rows, "incremental", None)
        assert len(result) == 2
        assert result[0]["row_num"] == 11

    def test_regenerate_by_row_num(self):
        rows = [
            {"row_num": 10, "req_id": "R001"},
            {"row_num": 11, "req_id": "R002"},
            {"row_num": 12, "req_id": "R003"},
        ]
        result = _filter_rows(rows, "regenerate", "10,12")
        assert len(result) == 2
        assert result[0]["row_num"] == 10
        assert result[1]["row_num"] == 12

    def test_regenerate_by_req_id(self):
        rows = [
            {"row_num": 10, "req_id": "R001"},
            {"row_num": 11, "req_id": "R002"},
        ]
        result = _filter_rows(rows, "regenerate", "R002")
        assert len(result) == 1
        assert result[0]["req_id"] == "R002"


class TestEstimateCost:
    def test_nonzero(self):
        cost = _estimate_cost(10, "gpt-4.1", 5)
        assert cost > 0

    def test_mini_cheaper(self):
        standard = _estimate_cost(10, "gpt-5", 5)
        mini = _estimate_cost(10, "gpt-5-mini", 5)
        assert mini < standard


class TestDryRun:
    def test_dry_run_exits_zero(self, sample_xlsx, tmp_path):
        args = parse_args([
            "--input", sample_xlsx,
            "--output-dir", str(tmp_path / "out"),
            "--dry-run",
        ])
        exit_code = run(args)
        assert exit_code == 0


class TestRunGeneration:
    @patch("main.generate_single_tc")
    def test_invalid_existing_tc_ids_do_not_crash(self, mock_generate_single_tc, sample_xlsx, tmp_path):
        wb = Workbook()
        # placeholder to silence linter-style expectations in some environments
        del wb

        from openpyxl import load_workbook

        loaded = load_workbook(sample_xlsx)
        ws = loaded["Test Case Specification&Result"]
        ws.cell(row=10, column=6, value="bad-id")
        loaded.save(sample_xlsx)
        loaded.close()

        mock_generate_single_tc.return_value.tc_data = {
            "tc_title": "Condition → Outcome",
            "pre_conditions": "NA",
            "input_test_data": "NA",
            "test_procedure": "1. Perform setup.\n2. Check that the result is displayed.",
            "expected_result": "1. Setup completes.\n2. Result is verified.",
            "design_method": "功能測試 (Functional based ; no specific technique)",
            "priority": "P1",
            "split_flag": False,
            "split_reason": "",
        }
        mock_generate_single_tc.return_value.input_tokens = 10
        mock_generate_single_tc.return_value.output_tokens = 20
        mock_generate_single_tc.return_value.cost = 0.001

        args = parse_args([
            "--input", sample_xlsx,
            "--output-dir", str(tmp_path / "out"),
            "--mode", "incremental",
        ])
        exit_code = run(args)
        assert exit_code == 0

    @patch("main.generate_batch")
    def test_budget_limit_blocks_batch_before_api_call(self, mock_generate_batch, sample_xlsx, tmp_path):
        args = parse_args([
            "--input", sample_xlsx,
            "--output-dir", str(tmp_path / "out"),
            "--budget", "0.0001",
            "--batch-size", "2",
        ])

        exit_code = run(args)

        assert exit_code == 1
        mock_generate_batch.assert_not_called()

    def test_unsupported_spec_format_fails_fast(self, sample_xlsx, tmp_path):
        spec_path = tmp_path / "unsupported.txt"
        spec_path.write_text("not supported")

        args = parse_args([
            "--input", sample_xlsx,
            "--spec", str(spec_path),
            "--dry-run",
        ])

        with pytest.raises(ValueError, match="Unsupported spec format"):
            run(args)

    @patch("main.generate_single_tc")
    def test_strict_validation_fails_invalid_rows(self, mock_generate_single_tc, sample_xlsx, tmp_path):
        mock_generate_single_tc.return_value.tc_data = {
            "tc_title": "Condition → Outcome",
            "pre_conditions": "1. Open settings menu",
            "input_test_data": "NA",
            "test_procedure": "1. Perform setup.\n2. Execute action without verification.",
            "expected_result": "1. Setup completes.\n2. Works as expected.",
            "design_method": "invalid method",
            "priority": "Critical",
            "split_flag": False,
            "split_reason": "",
        }
        mock_generate_single_tc.return_value.input_tokens = 10
        mock_generate_single_tc.return_value.output_tokens = 20
        mock_generate_single_tc.return_value.cost = 0.001

        args = parse_args([
            "--input", sample_xlsx,
            "--output-dir", str(tmp_path / "out"),
            "--mode", "incremental",
            "--batch-size", "1",
            "--strict-validation",
        ])

        exit_code = run(args)

        assert exit_code == 1
        output_path = tmp_path / "out" / "Test_SWQT_DeviceManager_20260408_generated.xlsx"
        assert not output_path.exists()

    @patch("main.generate_single_tc")
    def test_non_strict_validation_keeps_warning_behavior(self, mock_generate_single_tc, sample_xlsx, tmp_path):
        mock_generate_single_tc.return_value.tc_data = {
            "tc_title": "Condition → Outcome",
            "pre_conditions": "1. Open settings menu",
            "input_test_data": "NA",
            "test_procedure": "1. Perform setup.\n2. Execute action without verification.",
            "expected_result": "1. Setup completes.\n2. Works as expected.",
            "design_method": "invalid method",
            "priority": "Critical",
            "split_flag": False,
            "split_reason": "",
        }
        mock_generate_single_tc.return_value.input_tokens = 10
        mock_generate_single_tc.return_value.output_tokens = 20
        mock_generate_single_tc.return_value.cost = 0.001

        args = parse_args([
            "--input", sample_xlsx,
            "--output-dir", str(tmp_path / "out"),
            "--mode", "incremental",
            "--batch-size", "1",
        ])

        exit_code = run(args)

        assert exit_code == 0
        output_path = tmp_path / "out" / "Test_SWQT_DeviceManager_20260408_generated.xlsx"
        assert output_path.exists()


# ---------------------------------------------------------------------------
# --review CLI mode
# ---------------------------------------------------------------------------


def _build_review_workbook(tmp_path):
    fp = tmp_path / "Review_SWQT_Projection_20260502.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Document"
    ws.cell(row=3, column=2, value="Projection")
    ws_tc = wb.create_sheet("Test Case Specification&Result")
    headers = {4: "Req ID", 6: "TC ID", 9: "Test Item", 10: "Pre-Cond",
               12: "Procedure", 13: "ER", 14: "Spec Ref", 16: "Priority", 17: "Design Method"}
    for col, name in headers.items():
        ws_tc.cell(row=9, column=col, value=name)
    ws_tc.cell(row=10, column=4, value="REQ-A")
    ws_tc.cell(row=10, column=6, value="TC-A-1")
    ws_tc.cell(row=10, column=9, value="the HU shall display the icon")
    ws_tc.cell(row=10, column=12, value="1. Open menu.\n2. Confirm icon visible.")
    ws_tc.cell(row=10, column=13, value="1. Menu opens.\n2. Icon shown.")
    ws_tc.cell(row=10, column=16, value="P1")
    ws_tc.cell(row=10, column=17, value="Functional")
    wb.save(fp)
    return str(fp)


class TestReviewMode:
    def test_review_flag_parses(self):
        args = parse_args(["--input", "x.xlsx", "--review"])
        assert args.review is True

    def test_review_rejects_generate_only_flags(self):
        args = parse_args([
            "--input", "x.xlsx", "--review",
            "--mode", "incremental", "--batch-size", "10",
        ])
        err = _reject_generate_flags(args)
        assert err is not None
        assert "--mode" in err and "--batch-size" in err

    def test_review_accepts_shared_flags(self):
        args = parse_args([
            "--input", "x.xlsx", "--review",
            "--output-dir", "out", "--model", "gpt-5", "--dry-run", "--budget", "2.0",
        ])
        assert _reject_generate_flags(args) is None

    def test_review_dry_run_writes_outputs(self, tmp_path):
        fp = _build_review_workbook(tmp_path)
        out = tmp_path / "review_out"
        exit_code = main([
            "--input", fp, "--review",
            "--output-dir", str(out), "--dry-run",
        ])
        assert exit_code == 0
        assert (out / "findings.json").is_file()
        assert (out / "findings_report.md").is_file()

    def test_review_main_returns_2_when_misused(self, tmp_path, capsys):
        fp = _build_review_workbook(tmp_path)
        exit_code = main([
            "--input", fp, "--review",
            "--mode", "incremental",
        ])
        assert exit_code == 2
        captured = capsys.readouterr()
        assert "generate-only flags" in captured.err


# ---------------------------------------------------------------------------
# Stage 2.5 / 7 / 6 — new CLI subcommands (zero API)
# ---------------------------------------------------------------------------

import json as _json


def _write_findings(path):
    path.write_text(_json.dumps({
        "batch_meta": {"source_file": "x.xlsx", "total_tcs": 4, "total_req_groups": 2},
        "per_req_findings": [
            {"req_id": "R1", "tier": 1, "severity": "Critical", "rule_ref": "§6.3"},
        ],
        "per_tc_findings": [
            {"tc_id": "T1", "row": 10, "findings": [
                {"tier": 2, "rule_ref": "§7.4", "severity": "Critical"}]},
        ],
        "batch_summary": {},
    }, ensure_ascii=False), encoding="utf-8")


class TestScorecardCli:
    def test_scorecard_writes_outputs(self, tmp_path):
        findings = tmp_path / "findings.json"
        _write_findings(findings)
        out = tmp_path / "sc_out"
        assert main(["--scorecard", "--findings", str(findings),
                     "--output-dir", str(out)]) == 0
        assert (out / "scorecard.json").is_file()
        assert (out / "scorecard.md").is_file()
        data = _json.loads((out / "scorecard.json").read_text(encoding="utf-8"))
        assert data["total_tcs"] == 4
        assert data["kpis"]["tier1_critical_req_rate"]["numerator"] == 1

    def test_scorecard_missing_findings_returns_2(self, tmp_path):
        assert main(["--scorecard", "--findings",
                     str(tmp_path / "nope.json"), "--output-dir", str(tmp_path)]) == 2


class TestBudgetCli:
    def test_preflight_uncalibrated_waits(self, tmp_path, monkeypatch, capsys):
        monkeypatch.chdir(tmp_path)  # no config/budget.json -> defaults (uncalibrated)
        assert main(["--preflight", "--remaining-pct", "0.65",
                     "--n-light", "10", "--n-deep", "2"]) == 0
        assert "WAIT" in capsys.readouterr().out

    def test_preflight_requires_remaining_pct(self):
        assert main(["--preflight", "--n-light", "10"]) == 2

    def test_calibrate_roundtrips_to_budget_json(self, tmp_path, monkeypatch):
        monkeypatch.chdir(tmp_path)
        assert main(["--calibrate", "--start-pct", "1.0", "--end-pct", "0.88",
                     "--n-probe", "10", "--regime", "deep"]) == 0
        cfg = _json.loads((tmp_path / "config" / "budget.json").read_text(encoding="utf-8"))
        assert abs(cfg["per_req_pct_deep"] - 0.012) < 1e-9


class TestReviewDomainPackCli:
    def test_review_dry_run_accepts_domain_pack(self, tmp_path):
        fp = _build_review_workbook(tmp_path)
        dp = tmp_path / "dp.json"
        dp.write_text(_json.dumps({"project": "X", "glossary": [
            {"term": "A", "definition": "B"}]}), encoding="utf-8")
        out = tmp_path / "rev_out"
        # dry-run skips the LLM, but the --domain-pack arg must parse and not crash.
        assert main(["--input", fp, "--review", "--dry-run",
                     "--output-dir", str(out), "--domain-pack", str(dp)]) == 0
        assert (out / "scorecard.json").is_file()


class TestTraceCli:
    def test_trace_writes_outputs(self, tmp_path):
        fp = _build_review_workbook(tmp_path)
        reqs = tmp_path / "swe1.json"
        reqs.write_text(_json.dumps([
            {"id": "SWE1-X-001", "title": "Pairing", "desc": "the HU shall complete pairing"},
        ]), encoding="utf-8")
        out = tmp_path / "trace_out"
        assert main(["--trace", "--input", fp, "--swe1-reqs", str(reqs),
                     "--output-dir", str(out)]) == 0
        assert (out / "traceability.json").is_file()
        assert (out / "traceability.md").is_file()

    def test_trace_requires_swe1_reqs(self, tmp_path):
        fp = _build_review_workbook(tmp_path)
        assert main(["--trace", "--input", fp, "--output-dir", str(tmp_path)]) == 2
