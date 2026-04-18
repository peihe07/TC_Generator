"""`parse_workbook_tool` 單元測試。

不經 FastAPI，直接驗證 pure function 行為。Route 層的整合測試由
`tests/test_api_server.py::test_parse_workbook*` 繼續守。
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from backend.tools import ToolError, get_tool, parse_workbook_tool, SafetyLevel


def _build_workbook_bytes() -> bytes:
    wb = Workbook()
    pd_sheet = wb.active
    pd_sheet.title = "Product Document"
    pd_sheet.cell(row=3, column=2, value="newR1L")

    tc_sheet = wb.create_sheet("Test Case Specification&Result")
    tc_sheet.cell(row=9, column=4, value="Requirement or Design ID")
    tc_sheet.cell(row=9, column=9, value="Test Item")
    tc_sheet.cell(row=10, column=4, value="SWE1-HMI-DM-001-01")
    tc_sheet.cell(row=10, column=9, value="PDM01 original text")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_workbook(path: Path) -> Path:
    path.write_bytes(_build_workbook_bytes())
    return path


def test_parse_tool_returns_expected_shape(tmp_path):
    raw = _write_workbook(tmp_path / "SomeProject_SWQT_DeviceManager_20260408.xlsx")
    store: dict = {}

    result = parse_workbook_tool(
        raw_path=str(raw),
        raw_filename=raw.name,
        job_store=store,
    )

    assert result["project"] == "newR1L"
    assert result["testGroup"] == "DeviceManager"
    assert result["rowCount"] == 1
    assert result["rows"][0]["reqId"] == "SWE1-HMI-DM-001-01"
    assert result["files"]["rawFileName"] == raw.name

    # job_store 被正確寫入
    job_id = result["jobId"]
    assert job_id in store
    assert store[job_id]["rawFileName"] == raw.name
    assert store[job_id]["rawBytes"] == raw.read_bytes()
    assert store[job_id]["status"] == "parsed"


def test_parse_tool_rejects_invalid_extension(tmp_path):
    raw = tmp_path / "bad_file.txt"
    raw.write_text("not an xlsx")

    with pytest.raises(ToolError) as info:
        parse_workbook_tool(
            raw_path=str(raw),
            raw_filename=raw.name,
            job_store={},
        )
    assert info.value.code == "bad_request"
    assert info.value.http_status == 400


def test_parse_tool_missing_file(tmp_path):
    missing = tmp_path / "nope.xlsx"

    with pytest.raises(ToolError) as info:
        parse_workbook_tool(
            raw_path=str(missing),
            raw_filename=missing.name,
            job_store={},
        )
    assert info.value.code == "not_found"
    assert info.value.http_status == 404


def test_parse_tool_stores_reference_and_spec(tmp_path):
    raw = _write_workbook(tmp_path / "SomeProject_SWQT_DeviceManager_20260408.xlsx")
    ref = _write_workbook(tmp_path / "reference.xlsx")
    spec = tmp_path / "supplement.docx"
    spec.write_bytes(b"dummy-docx-content")

    store: dict = {}
    result = parse_workbook_tool(
        raw_path=str(raw),
        raw_filename=raw.name,
        reference_path=str(ref),
        reference_filename=ref.name,
        spec_path=str(spec),
        spec_filename=spec.name,
        job_store=store,
    )

    record = store[result["jobId"]]
    assert record["referenceWorkbookName"] == "reference.xlsx"
    assert record["referenceWorkbookBytes"] == ref.read_bytes()
    assert record["specFileName"] == "supplement.docx"
    assert record["specFormat"] == "docx"
    assert record["specBytes"] == b"dummy-docx-content"
    assert result["files"]["specFormat"] == "docx"


def test_parse_tool_registered():
    spec = get_tool("parse_workbook")
    assert spec.func is parse_workbook_tool
    assert spec.safety is SafetyLevel.WRITE_SAFE
