"""Agent golden scenario 整合測試（ROADMAP §4.3.3）。

每個 scenario 用 scripted LLM 驗證 dispatcher 把 tool 串起來的行為：
1. 「parse 這份檔」→ parse_workbook
2. 「跑完整個 pipeline」→ parse_workbook → group_tests → generate_tc → validate_tc
3. 「上週那個 job 補 REQ-045」→ list_jobs → generate_tc (auto_approve)
4. 「跨 job 查 match rate」→ list_jobs → get_job_validation ×N
5. 「estimate 這個檔多少錢」→ inspect_workbook → LLM 報價

LLM 全部 mock。`generate_tc` 底層的 OpenAI 呼叫也 mock。
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Callable
from unittest.mock import patch

import pytest
from openpyxl import Workbook

from agent_dispatcher import DispatchContext, LLMStep, run_agent_turn
from trace_store import SqliteTraceStore


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _mock_llm(script: list[LLMStep]) -> Callable:
    it = iter(script)

    def fn(messages, tool_defs):
        return next(it)

    return fn


def _build_tc_workbook(path: Path, *, rows: int = 2) -> Path:
    wb = Workbook()
    pd = wb.active
    pd.title = "Product Document"
    pd.cell(row=3, column=2, value="newR1L")

    tc = wb.create_sheet("Test Case Specification&Result")
    tc.cell(row=9, column=4, value="Requirement or Design ID")
    tc.cell(row=9, column=9, value="Test Item")
    for i in range(rows):
        tc.cell(row=10 + i, column=4, value=f"SWE1-HMI-DM-{i + 1:03d}-01")
        tc.cell(row=10 + i, column=9, value=f"PDM0{i + 1} original text")
    wb.save(path)
    return path


def _fake_generation_result(row_count: int = 1) -> SimpleNamespace:
    return SimpleNamespace(
        tc_data=[
            {
                "test_item_rewrite": "(Condition → Outcome)",
                "pre_conditions": "NA",
                "input_test_data": "NA",
                "test_procedure": "1. Setup.\n2. Verify result.",
                "expected_result": "1. Setup complete.\n2. Result verified.",
                "design_method": "功能測試 (Functional based ; no specific technique)",
                "priority": "Medium",
                "split_flag": False,
                "split_reason": "",
            }
            for _ in range(row_count)
        ] if row_count > 1 else {
            "test_item_rewrite": "(Condition → Outcome)",
            "pre_conditions": "NA",
            "input_test_data": "NA",
            "test_procedure": "1. Setup.\n2. Verify result.",
            "expected_result": "1. Setup complete.\n2. Result verified.",
            "design_method": "功能測試 (Functional based ; no specific technique)",
            "priority": "Medium",
            "split_flag": False,
            "split_reason": "",
        },
        input_tokens=100 * row_count,
        output_tokens=50 * row_count,
        cache_creation_tokens=0,
        cache_read_tokens=0,
        cost=0.001 * row_count,
        model="gpt-4.1-mini",
    )


@pytest.fixture
def ctx(tmp_path):
    trace = SqliteTraceStore(tmp_path / "trace.db")
    yield DispatchContext(
        job_store={},
        trace_store=trace,
        session_id="golden",
        rules_text="",
        auto_approve=True,
    )
    trace.close()


# ---------------------------------------------------------------------------
# Scenario 1: parse 這份檔
# ---------------------------------------------------------------------------

def test_scenario_1_parse_workbook(ctx, tmp_path):
    workbook = _build_tc_workbook(tmp_path / "Proj_SWQT_DeviceManager_20260408.xlsx")

    script = [
        LLMStep(
            tool_calls=[
                {
                    "id": "call-1",
                    "name": "parse_workbook",
                    "args": {"raw_path": str(workbook), "raw_filename": workbook.name},
                }
            ]
        ),
        LLMStep(text="已完成 parse，共 2 rows，Test Group = DeviceManager。", tool_calls=[]),
    ]

    events, _ = run_agent_turn(
        user_message=f"請 parse 這份檔：{workbook}",
        history=[],
        ctx=ctx,
        llm_fn=_mock_llm(script),
    )

    types = [e.type for e in events]
    # parse_workbook 是 WRITE_SAFE 且 result 帶 jobId → 跟一個 state_update
    assert types == ["tool_call", "tool_result", "state_update", "message", "done"]
    assert events[0].data["tool"] == "parse_workbook"
    result = events[1].data["result"]
    assert result["rowCount"] == 2
    assert result["testGroup"] == "DeviceManager"
    # job_store 被寫入
    assert result["jobId"] in ctx.job_store


# ---------------------------------------------------------------------------
# Scenario 2: 跑完整個 pipeline
# ---------------------------------------------------------------------------

def test_scenario_2_full_pipeline(ctx, tmp_path):
    workbook = _build_tc_workbook(tmp_path / "Proj_SWQT_DeviceManager_20260408.xlsx", rows=1)

    # 固定 script：parse → group → generate → validate → 報告
    gen_row = {
        "id": "row-10",
        "row_num": 10,
        "tc_id": "",
        "req_id": "SWE1-HMI-DM-001-01",
        "test_item": "PDM01 original text",
        "original_requirement": "PDM01 original text",
        "test_set": "Access",
        "priority": "Medium",
    }

    script = [
        LLMStep(tool_calls=[{
            "id": "c-parse", "name": "parse_workbook",
            "args": {"raw_path": str(workbook), "raw_filename": workbook.name},
        }]),
        LLMStep(tool_calls=[{
            "id": "c-group", "name": "group_tests",
            "args": {"rows": [{"id": "row-10", "reqId": "SWE1-HMI-DM-001-01", "testItem": "PDM01 text"}]},
        }]),
        LLMStep(tool_calls=[{
            "id": "c-gen", "name": "generate_tc",
            "args": {
                "rows": [gen_row],
                "context": {"project": "newR1L", "test_group": "DeviceManager"},
                "rules_text": "RULES_STUB",
                "model": "gpt-4.1-mini",
            },
        }]),
        LLMStep(tool_calls=[{
            "id": "c-val", "name": "validate_tc",
            "args": {
                "tc": {
                    "tc_id": "x",
                    "test_item": "orig\n\n(Condition → Outcome)",
                    "pre_conditions": "NA",
                    "test_procedure": "1. Setup.\n2. Verify result.",
                    "expected_result": "1. Setup complete.\n2. Result verified.",
                    "design_method": "功能測試 (Functional based ; no specific technique)",
                    "priority": "Medium",
                }
            },
        }]),
        LLMStep(text="Pipeline 完成：1 row 走完 parse→group→generate→validate。", tool_calls=[]),
    ]

    with patch(
        "tools.generate.generate_single_tc",
        return_value=_fake_generation_result(row_count=1),
    ):
        events, _ = run_agent_turn(
            user_message="跑完整個 pipeline",
            history=[],
            ctx=ctx,
            llm_fn=_mock_llm(script),
            max_steps=8,
        )

    tool_sequence = [e.data["tool"] for e in events if e.type == "tool_call"]
    assert tool_sequence == ["parse_workbook", "group_tests", "generate_tc", "validate_tc"]
    assert events[-1].type == "done"
    assert not any(e.type == "tool_error" for e in events)


# ---------------------------------------------------------------------------
# Scenario 3: 上週那個 job 補 REQ-045
# ---------------------------------------------------------------------------

def test_scenario_3_list_then_regenerate(ctx):
    # 預先塞一筆 job（模擬過往 parse 結果）
    ctx.job_store["j-old"] = {
        "jobId": "j-old",
        "rawFileName": "old.xlsx",
        "status": "completed",
        "parsedData": {
            "project": "newR1L",
            "test_group": "DeviceManager",
            "row_count": 44,
            "rows": [
                {"row_num": 10, "req_id": "REQ-045", "test_item": "missing thing", "tc_id": "newR1L-DMS-045"},
            ],
        },
    }

    script = [
        LLMStep(
            tool_calls=[
                {"id": "c-list", "name": "list_jobs", "args": {"limit": 10}}
            ]
        ),
        LLMStep(
            tool_calls=[
                {
                    "id": "c-regen",
                    "name": "generate_tc",
                    "args": {
                        "rows": [
                            {
                                "id": "row-10",
                                "row_num": 10,
                                "tc_id": "newR1L-DMS-045",
                                "req_id": "REQ-045",
                                "test_item": "missing thing",
                                "original_requirement": "missing thing",
                                "test_set": "Access",
                                "priority": "Medium",
                            }
                        ],
                        "context": {"project": "newR1L", "test_group": "DeviceManager"},
                        "rules_text": "RULES",
                        "model": "gpt-4.1-mini",
                    },
                }
            ]
        ),
        LLMStep(text="已為 REQ-045 重生一筆 TC。", tool_calls=[]),
    ]

    with patch(
        "tools.generate.generate_single_tc",
        return_value=_fake_generation_result(row_count=1),
    ):
        events, _ = run_agent_turn(
            user_message="上週那個 job 幫我補 REQ-045",
            history=[],
            ctx=ctx,
            llm_fn=_mock_llm(script),
            max_steps=6,
        )

    tools = [e.data["tool"] for e in events if e.type == "tool_call"]
    assert tools == ["list_jobs", "generate_tc"]
    # list_jobs 的 result 有 j-old
    list_result = events[1].data["result"]
    assert list_result["jobs"][0]["jobId"] == "j-old"


# ---------------------------------------------------------------------------
# Scenario 4: 跨 job 查 match rate（這裡改為跑 validation）
# ---------------------------------------------------------------------------

def test_scenario_4_cross_job_validation(ctx):
    # 兩個 job，各自帶 generated rows
    def _gen_row(i: int, priority: str = "Medium") -> dict:
        return {
            "id": f"row-{i}",
            "tcId": f"tc-{i}",
            "originalRequirement": "Original",
            "generated": {
                "testItemRewrite": "(A → B)",
                "preConditions": "NA",
                "testProcedure": "1. Do A.\n2. Verify B.",
                "expectedResult": "1. Done.\n2. OK.",
                "designMethod": "功能測試 (Functional based ; no specific technique)",
                "priority": priority,
            },
        }

    ctx.job_store["j1"] = {
        "jobId": "j1", "status": "completed",
        "parsedData": {"row_count": 2, "project": "P", "test_group": "G1"},
    }
    ctx.job_store["j2"] = {
        "jobId": "j2", "status": "completed",
        "parsedData": {"row_count": 1, "project": "P", "test_group": "G2"},
    }

    # scripted：list_jobs → validate j1 rows → validate j2 rows → summary message
    script = [
        LLMStep(tool_calls=[{"id": "c-list", "name": "list_jobs", "args": {}}]),
        LLMStep(
            tool_calls=[
                {
                    "id": "c-v1",
                    "name": "get_job_validation",
                    "args": {"rows": [_gen_row(1), _gen_row(2, priority="Bogus")]},
                }
            ]
        ),
        LLMStep(
            tool_calls=[
                {
                    "id": "c-v2",
                    "name": "get_job_validation",
                    "args": {"rows": [_gen_row(3)]},
                }
            ]
        ),
        LLMStep(text="j1: 1 pass / 1 warning；j2: 1 pass。", tool_calls=[]),
    ]

    events, _ = run_agent_turn(
        user_message="幫我查兩個 job 的 match rate",
        history=[],
        ctx=ctx,
        llm_fn=_mock_llm(script),
        max_steps=6,
    )

    tools = [e.data["tool"] for e in events if e.type == "tool_call"]
    assert tools == ["list_jobs", "get_job_validation", "get_job_validation"]

    # 結果聚合正確
    r1 = events[3].data["result"]
    r2 = events[5].data["result"]
    assert r1["total"] == 2 and r1["warnings"] >= 1
    assert r2["total"] == 1 and r2["pass"] + r2["warnings"] == 1


# ---------------------------------------------------------------------------
# Scenario 5: estimate 這個檔多少錢
# ---------------------------------------------------------------------------

def test_scenario_5_inspect_then_quote(ctx, tmp_path):
    workbook = _build_tc_workbook(
        tmp_path / "Proj_SWQT_CM_20260101.xlsx", rows=5
    )

    script = [
        LLMStep(
            tool_calls=[
                {
                    "id": "c-insp",
                    "name": "inspect_workbook",
                    "args": {"path": str(workbook)},
                }
            ]
        ),
        LLMStep(
            text="這份檔 5 rows，用 gpt-4.1-mini 粗估約 $0.006 — 精確數字需 parse 後以 estimate_cost 查詢。",
            tool_calls=[],
        ),
    ]

    events, _ = run_agent_turn(
        user_message="estimate 這個檔多少錢",
        history=[],
        ctx=ctx,
        llm_fn=_mock_llm(script),
    )

    tools = [e.data["tool"] for e in events if e.type == "tool_call"]
    assert tools == ["inspect_workbook"]

    insp = events[1].data["result"]
    assert insp["rowsEstimate"] == 5
    assert insp["testGroup"] == "CM"
    # job_store 沒被寫（inspect 不建 job）
    assert ctx.job_store == {}
