#!/usr/bin/env python3
"""Phase 1 Recon — automated feature survey per FEATURE_ONBOARDING.md.

Reads feature.yaml, surveys the FW036 workbook and the 037 report, and emits:

- RECON.md      — human-readable survey (evidence for every [AUTO] value)
- DECISIONS.md  — pre-filled decision sheet ([AUTO] filled, [PROPOSED]
                  suggested from per-state strategy bindings, [PEI] left open)
- data/recon.json — machine-readable results for downstream scripts

Tier 0/1 only: this script DETECTS and PROPOSES; it never rules. Ambiguity is
surfaced with row-level evidence, not resolved.

Regression-validated against features/home (2026-08-09): reproduces the manually
verified survey exactly — PARTIAL_INTERLEAVED; 140 leaves; 62 regen targets;
done segments 10-86/91-124/129-161; uncovered {055-03, 066}; parent/child
dupe {066}; orphan done req {035} (A-H06); 9 design-method strings;
27 done-region compliance notes (14 single-step procedures + 13 blank
priorities).

NOTE on the spec text-layer probe: it measures the file it is given. A PDF
copy that has been re-rendered elsewhere may probe differently from the
original — always trust the probe run against the repo inputs/ copy.

Usage:
    python scripts/recon.py --feature features/home --root .
"""

import argparse
import hashlib
import json
import re
import shutil
import subprocess
import sys
from collections import Counter
from pathlib import Path

import openpyxl
import yaml

# ---------------------------------------------------------------- helpers

PLACEHOLDER_PROCEDURES = {"test", "tbd", "todo", "na", "n/a", "-"}


def norm(v) -> str:
    """Header text, comparable: lowered, newlines and runs of space collapsed."""
    return re.sub(r"\s+", " ", str(v or "")).strip().lower()


# --- workbook column resolution by header TEXT, not by letter ------------
#
# feature.yaml letters are a prior, not the authority. The FM-WI-FSM-036 form
# has at least two revisions in circulation: revision C (2026-01-21) inserted
# "Estimated Test Time (mins)" at Q, shifting design_method, functional_safety,
# the vehicle block, author and remarks one column right. An instance created
# after that date may still be built on the older layout — the AM/FM
# CFTS024_Radio workbook (2026-01-29) is, and its ChangeHistory stops at B. So
# the layout has to be read off the instance in front of us.
#
# Each entry is (required tokens, forbidden tokens); a field resolves only on
# exactly one matching column, because near-duplicate headers are the trap:
# "Requirement or Design ID" (D) vs "...ID (Polarion)" (C), "Test Case ID" (F)
# vs "Test Case ID (TestRail)" (E) vs "Test Case Reference ID" (O).
HEADER_SPEC = {
    "req_id": (("requirement or design id",), ("polarion",)),
    "test_group": (("test group",), ()),
    "test_set": (("test set",), ()),
    "test_item": (("test item",), ()),
    "pre_conditions": (("pre-condition",), ()),
    "input_test_data": (("input test data",), ()),
    "test_procedure": (("test procedure",), ()),
    "expected_result": (("expected result",), ()),
    "spec_reference": (("specification reference",), ()),
    "tc_ref_id": (("test case reference id",), ()),
    "priority": (("test case priority",), ()),
    "design_method": (("test case design", "methods"), ()),
    "functional_safety": (("functional safety",), ()),
    "author": (("test case author",), ()),
    "remarks": (("remarks",), ()),
    # Revision marker, not a pipeline field: present => rev C layout.
    "estimated_test_time": (("estimated test time",), ()),
}
OPTIONAL_COLUMNS = {"estimated_test_time"}


def resolve_columns(header: tuple) -> tuple[dict, list[str]]:
    """header row values -> {field: 0-based index}, plus unresolved notes."""
    resolved, notes = {}, []
    for key, (need, forbid) in HEADER_SPEC.items():
        hits = [j for j, v in enumerate(header)
                if all(t in norm(v) for t in need)
                and not any(t in norm(v) for t in forbid)]
        if len(hits) == 1:
            resolved[key] = hits[0]
        elif not hits and key in OPTIONAL_COLUMNS:
            continue
        else:
            notes.append(f"{key}: {len(hits)} header matches"
                         + (f" at {[idx_to_letter(h) for h in hits]}" if hits else ""))
    return resolved, notes


def idx_to_letter(idx: int) -> str:
    """0-based index -> Excel column letter."""
    s, n = "", idx + 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def col_letter_to_idx(letter: str) -> int:
    """Excel column letter -> 0-based index."""
    n = 0
    for ch in letter.strip().upper():
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def numbered_steps(text: str) -> int:
    if not text:
        return 0
    return len(re.findall(r"^\s*\d+[\.\)]", str(text), flags=re.M))


def outline_key(s: str) -> tuple:
    """Sort '10.9.1' after '10.9' and before '10.10' — string sort does not."""
    return tuple(int(p) if p.isdigit() else -1 for p in str(s).split("."))


# ---------------------------------------------------------------- outline map

def build_outline_map(sys1_path: Path | None) -> tuple[dict, str]:
    """outline number -> {id, title} from a SYS1 Polarion export.

    This is the lookup every spec_reference is linted against under spec_mode
    A: a cited section that is not in here does not exist in the ruled
    baseline, whatever a newer revision of the document may contain.

    Returns ({}, reason) when the declared file cannot carry an outline —
    `sys1_export` does not always mean "SYS1 spec export" (Privacy declares a
    SYSRA safety-analysis export under that key), and a feature that never
    cites document sections must not be blocked by a lookup it does not use.
    Whether the empty map MATTERS is decided by run_assertions, which reports
    it as a failure for any feature whose leaves do cite sections.
    """
    if not sys1_path:
        return {}, "paths.sys1_export is null"
    if not sys1_path.exists():
        return {}, f"{sys1_path.name} does not exist"
    wb = openpyxl.load_workbook(sys1_path, read_only=True)
    if "Basic Report" not in wb.sheetnames:
        wb.close()
        return {}, f"{sys1_path.name}: no 'Basic Report' sheet — not a SYS1 export"
    rows = list(wb["Basic Report"].iter_rows(values_only=True))
    header = rows[0]
    idx = {norm(v): j for j, v in enumerate(header) if v}
    out_j = next((j for h, j in idx.items() if "outline number" in h), None)
    id_j = next((j for h, j in idx.items() if h == "id"), 0)
    desc_j = next((j for h, j in idx.items() if "description" in h), None)
    if out_j is None:
        wb.close()
        return {}, (f"{sys1_path.name}: no 'Outline Number' column — this "
                    "export does not carry a document outline")
    omap, dupes = {}, []
    for r in rows[1:]:
        if out_j >= len(r) or not r[out_j]:
            continue
        key = str(r[out_j]).strip()
        title = re.sub(r"\s+", " ", str(r[desc_j] or ""))[:80] if desc_j is not None else ""
        if key in omap:
            dupes.append(key)
            continue
        omap[key] = {"id": str(r[id_j] or "").strip(), "title": title}
    wb.close()
    if dupes:
        # Not a soft return: a duplicated outline number means the lookup is
        # not a function, and every spec_reference built from it would be
        # ambiguous. That is a defect in the export, not a feature that
        # happens to lack an outline.
        sys.exit(f"{sys1_path.name}: duplicate outline numbers {sorted(set(dupes))}"
                 " — the map would not be a function; refusing to build it")
    return omap, ""


# ------------------------------------------------- signed-DECISIONS guard

# A placeholder is the template's own text: empty, or nothing but underscores
# (the form ships "____", Projection's carries "____________").
PLACEHOLDER_RE = re.compile(r"^[\s_]*$")


def read_signoff(decisions_path: Path) -> dict:
    """What the EXISTING DECISIONS.md says about its own sign-off state.

    Two accepted shapes, per R-C10: a filled Sign-off block, or SXM-style
    `- Amendment (date, nth pass): … directive「…」` entries that record the
    ruling and its directive verbatim. Either is repo evidence that a human
    signed; neither present means the sign-off state is not knowable from the
    repo, whatever may have happened in a chat.
    """
    if not decisions_path.exists():
        return {"exists": False, "signed": False, "reviewed_by": "",
                "date": "", "amendments": 0, "has_proposed": False}
    text = decisions_path.read_text(encoding="utf-8")
    m = re.search(r"Reviewed by:\s*(.*?)(?:\s\s+|\s*\|)Date:\s*(.*)", text)
    who, when = (m.group(1).strip(), m.group(2).strip()) if m else ("", "")
    who_filled = bool(who) and not PLACEHOLDER_RE.match(who)
    amendments = len(re.findall(r"^\s*-\s*Amendment\b", text, flags=re.M))
    return {
        "exists": True,
        "signed": who_filled or amendments > 0,
        "reviewed_by": "" if not who_filled else who,
        "date": "" if PLACEHOLDER_RE.match(when) else when,
        "amendments": amendments,
        "has_proposed": "[PROPOSED" in text,
    }


def write_decisions(feature_dir: Path, body: str, signoff: dict) -> tuple[Path, bool]:
    """R-C9 — never overwrite a signed decision sheet.

    recon.py rewrites DECISIONS.md whole. That is fine for an unsigned sheet
    and destructive for a signed one, and it fires for anyone who re-runs
    recon on an existing feature for any reason — which no amount of
    remembering-not-to can prevent. So the check is here, in the writer,
    rather than in a policy someone has to recall.

    Returns (path_written, diverted).
    """
    target = feature_dir / "DECISIONS.md"
    if signoff["signed"]:
        alt = feature_dir / "DECISIONS.new.md"
        alt.write_text(body, encoding="utf-8")
        return alt, True
    target.write_text(body, encoding="utf-8")
    return target, False


# ---------------------------------------------------------------- assertions

class Assertions:
    """Ruled constants checked mechanically, reported as PASS/FAIL + the
    MEASURED value.

    A recon that only prints counts leaves the comparison to whoever reads it,
    and the ruling that says 403 is then enforced by attention rather than by
    the script (Comfort R-C3 requires the opposite explicitly). Expected values
    come from feature.yaml `recon_assertions`; a feature that declares none
    runs with an empty list and no behaviour change.
    """

    def __init__(self) -> None:
        self.results: list[dict] = []

    def check(self, name: str, expected, actual, note: str = "") -> bool:
        ok = expected == actual
        self.results.append({"name": name, "expected": expected,
                             "actual": actual, "pass": ok, "note": note})
        return ok

    @property
    def failed(self) -> list[dict]:
        return [r for r in self.results if not r["pass"]]

    def lines(self) -> list[str]:
        out = []
        for r in self.results:
            mark = "PASS" if r["pass"] else "**FAIL**"
            out.append(f"- {mark} — {r['name']}: expected `{r['expected']}`, "
                       f"measured `{r['actual']}`"
                       + (f" — {r['note']}" if r["note"] else ""))
        return out or ["- (no assertions declared in feature.yaml)"]


def run_assertions(cfg: dict, a03res: dict, omap: dict,
                   omap_reason: str = "") -> tuple[Assertions, list]:
    """Declared-constant checks + the outline lookup. Returns (results, misses)."""
    a = Assertions()
    want = cfg.get("recon_assertions") or {}
    cat = a03res["categorization_distribution"]

    if "functional_requirement_count" in want:
        a.check("leaf count == Functional Requirement rows",
                want["functional_requirement_count"], len(a03res["leaves"]),
                f"categorization distribution: {cat}; the banned id-suffix "
                f"criterion would have selected {a03res['naive_leaf_shape_count']} "
                f"({len(a03res['parent_shape_functional'])} parent-shaped "
                "requirements dropped)")

    if "distinct_spec_sections" in want:
        a.check("distinct spec sections after citation parse",
                want["distinct_spec_sections"], len(a03res["distinct_sections"]),
                f"{a03res['multiline_citations']} citation cells carry extra "
                "lines below the section (Polarion item ids), not parsed")

    if "spec_reference_stem" in want:
        # Sorted, so a second stem fails the check on its presence rather than
        # on the order the report happened to list them in.
        stems = sorted(a03res["citation_stems"])
        a.check("citation stem is the ruled baseline, and only that",
                [want["spec_reference_stem"]], stems,
                "a second stem means the report cites more than one baseline")

    # Outline lookup: every cited section must exist in the ruled export.
    # NOTE the asymmetry, which is the point of A-CF08: a cited section that
    # is absent from the baseline is an error (assert below), while a baseline
    # section nobody cites is a coverage question — reported, never asserted.
    misses = []
    if omap:
        misses = [s for s in a03res["distinct_sections"] if s not in omap]
        a.check("cited sections found in the ruled SYS1 outline",
                0, len(misses),
                f"{len(a03res['distinct_sections'])} cited / {len(omap)} "
                "outline entries in the export"
                + (f"; missing: {misses[:10]}" if misses else ""))
    elif a03res["distinct_sections"]:
        a.check("SYS1 export available for outline lookup", True, False,
                f"{omap_reason or 'no outline map'} — the "
                f"{len(a03res['distinct_sections'])} cited sections cannot be "
                "verified against any baseline")

    if a03res["unparsed_citations"]:
        a.check("every leaf's citation parses to a section", 0,
                len(a03res["unparsed_citations"]),
                f"samples: {a03res['unparsed_citations'][:5]}")
    return a, misses


# ---------------------------------------------------------------- survey

def survey_workbook(cfg: dict, wb_path: Path) -> dict:
    wbc = cfg["workbook"]
    header_row = wbc["header_row"]
    cols = {k: col_letter_to_idx(v) for k, v in wbc["columns"].items()}

    wb = openpyxl.load_workbook(wb_path, read_only=True)
    if wbc["sheet"] not in wb.sheetnames:
        sys.exit(f"sheet {wbc['sheet']!r} not found in {wb_path.name}")
    ws = wb[wbc["sheet"]]
    rows = list(ws.iter_rows(values_only=True))

    # --- columns resolved FROM the header text; feature.yaml is only a prior
    header = rows[header_row - 1]
    resolved, unresolved = resolve_columns(header)
    if unresolved:
        sys.exit("cannot resolve workbook columns from the header row "
                 f"{header_row}: {unresolved}")
    layout_rev = "C (has Estimated Test Time)" if "estimated_test_time" in \
        resolved else "A/B (no Estimated Test Time)"
    # Disagreements are reported, never silently honoured: a stale letter in
    # feature.yaml is exactly the defect this resolution exists to catch.
    col_conflicts = []
    for key, idx in resolved.items():
        want = cols.get(key)
        if want is not None and want != idx:
            col_conflicts.append(
                f"{key}: feature.yaml says {idx_to_letter(want)}, "
                f"header says {idx_to_letter(idx)}")
    cols = {k: v for k, v in resolved.items() if k != "estimated_test_time"}
    col_check = {k: True for k in cols}
    col_ok = len(cols)

    # --- who authored the existing rows? The configured value is a prior too;
    # a fresh feature.yaml carries the previous feature's name.
    dr = cfg["done_region"]
    author_val = dr.get("author_value")
    seen_authors = {}
    for r in rows[header_row:]:
        a = str(r[cols["author"]] or "").strip() if cols["author"] < len(r) else ""
        if a:
            seen_authors[a] = seen_authors.get(a, 0) + 1
    author_note = ""
    if seen_authors and author_val not in seen_authors:
        detected = max(seen_authors, key=seen_authors.get)
        author_note = (f"feature.yaml done_region.author_value={author_val!r} "
                       f"matches 0 rows; surveying with the dominant author "
                       f"{detected!r} ({seen_authors[detected]} rows) — "
                       "Tier 2 must confirm before Phase 4")
        author_val = detected
    row_class = []  # (excel_row, cls, req) cls in {EMPTY, DONE, DRAFT}
    for i, r in enumerate(rows[header_row:], start=header_row + 1):
        def cell(key):
            idx = cols.get(key)
            return r[idx] if idx is not None and idx < len(r) else None
        filled = bool(cell("test_item")) or bool(cell("req_id"))
        if not filled:
            row_class.append((i, "EMPTY", None))
            continue
        author = str(cell("author") or "").strip()
        proc = str(cell("test_procedure") or "").strip()
        # Done-region membership is about authorship + non-placeholder
        # content, NOT current-lint compliance (done region is frozen, not
        # re-linted; single-step done rows are compliance NOTES, below).
        is_placeholder = (proc.lower() in PLACEHOLDER_PROCEDURES
                          or (numbered_steps(proc) == 0 and len(proc) < 30))
        qualifies = (
            bool(author)
            and (author == author_val if dr.get("detection") == "author" else True)
            and bool(proc) and not is_placeholder
        )
        req = str(cell("req_id") or "").strip()
        row_class.append((i, "DONE" if qualifies else "DRAFT", req))

    # ambiguity: done-author rows carrying PLACEHOLDER content — genuine
    # segmentation doubt, surfaced not ruled
    ambiguous = []
    # compliance notes: done rows deviating from current rules (recorded,
    # never fixed — the done region is frozen; A-H05 pattern)
    compliance = []
    for i, cls, _ in row_class:
        r = rows[i - 1]
        author = str(r[cols["author"]] or "").strip()
        proc = str(r[cols["test_procedure"]] or "").strip()
        prio = str(r[cols["priority"]] or "").strip() if "priority" in cols else ""
        if cls == "DRAFT" and author_val and author == author_val:
            ambiguous.append((i, "done-author row with placeholder procedure"))
        if cls == "DONE":
            if numbered_steps(proc) < 2:
                compliance.append((i, "procedure < 2 numbered steps (§10.5)"))
            if not prio:
                compliance.append((i, "blank priority"))

    # --- segmentation
    segs = []
    for i, cls, _ in row_class:
        if cls == "EMPTY":
            continue
        if segs and segs[-1]["cls"] == cls and segs[-1]["end"] == i - 1:
            segs[-1]["end"] = i
            segs[-1]["n"] += 1
        else:
            segs.append({"cls": cls, "start": i, "end": i, "n": 1})

    done_segs = [s for s in segs if s["cls"] == "DONE"]
    draft_segs = [s for s in segs if s["cls"] == "DRAFT"]
    if not done_segs and not draft_segs:
        state = "BLANK"
    elif not draft_segs:
        state = "FULL"
    elif not done_segs:
        state = "BLANK"  # drafts only: no done region; treat drafts per ruling
    else:
        interleaved = any(d["start"] < done_segs[-1]["end"] for d in draft_segs)
        state = "PARTIAL_INTERLEAVED" if interleaved else "PARTIAL_CLEAN"

    done_reqs = sorted({req for i, cls, req in row_class if cls == "DONE" and req})
    draft_reqs = sorted({req for i, cls, req in row_class if cls == "DRAFT" and req})

    # --- design method vocabulary
    vocab = []
    if "下拉選單" in wb.sheetnames:
        for r in wb["下拉選單"].iter_rows(values_only=True):
            for c in r:
                if c:
                    vocab.append(str(c).strip())
    wb.close()

    return {
        "state": state, "col_check": col_check, "col_ok": col_ok,
        "col_total": len(cols), "segments": segs, "done_segments": done_segs,
        "draft_segments": draft_segs, "done_rows": sum(s["n"] for s in done_segs),
        "draft_rows": sum(s["n"] for s in draft_segs),
        "done_reqs": done_reqs, "draft_reqs": draft_reqs,
        "ambiguous_rows": ambiguous, "compliance_notes": compliance,
        "design_method_vocab": vocab,
        "columns": {k: idx_to_letter(v) for k, v in resolved.items()},
        "col_conflicts": col_conflicts, "layout_rev": layout_rev,
        "authors": seen_authors, "author_used": author_val,
        "author_note": author_note,
    }


def survey_a03(a03_path: Path) -> dict:
    """Survey the requirement report — columns located by header text.

    The Analysis Report template is not stable across features: Home's
    Categorization sits at column 7 and its values read "Functional
    Requirement"/"Heading"; AM/FM's sits at column 31 (behind 24 review-
    criteria columns) and reads "Functional" with no Heading rows at all. The
    old positional read (`row[6] == "Functional Requirement"`) classifies every
    AM/FM row as a heading and reports zero leaves — silently, since an empty
    leaf list is a legal state for a workbook that is already complete.

    Safety attributes (ASIL / FTTI) are reported when the template carries
    them: whether the safety-analysis layer belongs in the trace chain is a
    property of the ruled requirement source, not of what files are present.
    """
    wb = openpyxl.load_workbook(a03_path, read_only=True)
    ws = wb["Analysis Report"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = next((i for i, r in enumerate(rows)
                if any("requirement description" in norm(v) for v in r)), None)
    if hdr is None:
        sys.exit(f"{a03_path.name}: no header row (no 'Requirement Description'"
                 " cell) in the Analysis Report sheet")
    header = rows[hdr]

    def find(*need, forbid=()):
        hits = [j for j, v in enumerate(header)
                if all(t in norm(v) for t in need)
                and not any(t in norm(v) for t in forbid)]
        return hits[0] if len(hits) == 1 else None

    cat_i = find("categorization", forbid=("sub",))
    asil_i, ftti_i = find("asil"), find("ftti")
    # Document-citation column. "HMI Source" is checked first because a report
    # can carry both it and a generic "Source Requirement ID" that holds
    # upstream req ids rather than document sections.
    src_i = find("hmi source")
    if src_i is None:
        src_i = find("source", forbid=("description", "requirement id"))

    leaves, headings, parent_child = [], [], []
    safety = {}
    cat_dist: Counter = Counter()
    # R-C4 — the citation cell may carry the document section on line 1 and
    # Polarion item ids on the lines below. Only line 1 takes part in section
    # parsing; the rest is audit evidence, kept but never parsed.
    sections: dict[str, str] = {}
    stems: Counter = Counter()
    multiline, unparsed = 0, []
    for r in rows[hdr + 1:]:
        if not r[0]:
            continue
        rid = str(r[0]).strip()
        cat = str(r[cat_i] or "").strip() if cat_i is not None else ""
        cat_dist[cat or "(blank)"] += 1
        # "Functional" (AM/FM) and "Functional Requirement" (Home) are the same
        # classification written two ways; anything else is a heading/other.
        is_leaf = cat.lower().startswith("functional")
        (leaves if is_leaf else headings).append(rid)
        if is_leaf and src_i is not None:
            raw = str(r[src_i] or "")
            if "\n" in raw.strip():
                multiline += 1
            first = raw.split("\n")[0].strip()
            m = re.match(r"^(?P<stem>.+)_(?P<sec>\d+(?:\.\d+)*)$", first)
            if m:
                stems[m.group("stem")] += 1
                sections[rid] = m.group("sec")
            elif first:
                unparsed.append((rid, first[:60]))
        if asil_i is not None:
            val = str(r[asil_i] or "").strip()
            if val and len(val) < 12:      # skip the template's help text row
                safety[val] = safety.get(val, 0) + 1
    leafset = set(leaves)
    for rid in leaves:
        if re.search(r"-\d\d$", rid) and rid.rsplit("-", 1)[0] in leafset:
            parent_child.append(rid.rsplit("-", 1)[0])
    # R-C3 evidence: the ID-suffix heuristic that the ruling BANS, measured
    # side by side with the ruled criterion so the gap is a number in the
    # report rather than a claim in a document. `naive_leaf_shape` counts
    # leaves whose id carries a -NN child suffix; the difference is the set a
    # suffix-based selector would silently drop.
    naive_leaf_shape = [r for r in leaves if re.search(r"-\d\d$", r)]
    parent_shape_functional = [r for r in leaves if not re.search(r"-\d\d$", r)]
    wb.close()
    return {"leaves": leaves, "headings": headings,
            "parent_child_dupes": sorted(set(parent_child)),
            "categorization_col": idx_to_letter(cat_i) if cat_i is not None else None,
            "categorization_distribution": dict(cat_dist.most_common()),
            "source_col": idx_to_letter(src_i) if src_i is not None else None,
            "sections": sections,
            "distinct_sections": sorted(set(sections.values()), key=outline_key),
            "citation_stems": dict(stems),
            "multiline_citations": multiline,
            "unparsed_citations": unparsed,
            "naive_leaf_shape_count": len(naive_leaf_shape),
            "parent_shape_functional": sorted(parent_shape_functional),
            "has_safety_columns": asil_i is not None or ftti_i is not None,
            "asil_distribution": safety}


def survey_spec_text_layer(pdf_path: Path | None) -> str:
    """Does this PDF carry a text layer, or is it a scan needing OCR?

    Two extractors, because reporting "unknown" when the answer is sitting on
    disk is its own defect: pymupdf is preferred, and `pdftotext` (poppler,
    already present on this machine) is the fallback. Only when NEITHER is
    available is the answer genuinely unknown, and the message then names
    both so the reader knows what to install. (A-CF06 / handoff 09 §4.)
    """
    if not pdf_path or not pdf_path.exists():
        return "no-pdf"
    chars, how = None, ""
    try:
        import pymupdf as fitz
        doc = fitz.open(pdf_path)
        chars = sum(len(p.get_text()) for p in doc)
        doc.close()
        how = "pymupdf"
    except ImportError:
        exe = shutil.which("pdftotext")
        if exe:
            proc = subprocess.run([exe, "-q", str(pdf_path), "-"],
                                  capture_output=True)
            if proc.returncode == 0:
                chars, how = len(proc.stdout.decode("utf-8", "replace")), "pdftotext"
    if chars is None:
        return "unknown (neither pymupdf nor pdftotext available)"
    return (f"text-layer: {chars} chars (via {how})" if chars > 500
            else f"scanned (OCR path) — {chars} chars via {how}")


# -------------------------------------------------- uncited baseline sections

UNCITED_TSV = "sr24_uncited_sections.tsv"


def uncited_section_report(feature_dir: Path, a03res: dict, omap: dict) -> str:
    """Baseline sections the requirement report never cites (A-CF08).

    Computed here so it can never go stale against the export; the four-value
    classification is produced separately (the judgement rules live in
    `features/<f>/scripts/classify_uncited_sections.py`) and folded in when
    its TSV is present. If the TSV is missing the count is still reported —
    an unclassified gap is a visible state, an omitted section is not.

    This is a coverage OBSERVATION, not a disposition: nothing here generates
    a TC, enters a denominator, or files an RD item.
    """
    if not omap:
        return "- (no outline map — cannot tell cited from uncited)"
    uncited = sorted(set(omap) - set(a03res["distinct_sections"]), key=outline_key)
    lines = [f"- baseline outline entries: {len(omap)}; cited by the leaves: "
             f"{len(a03res['distinct_sections'])}; **uncited: {len(uncited)}**",
             "- these sit INSIDE the ruled baseline — a different question from"
             " content in an out-of-scope revision, and not answered by any"
             " ruling about that revision"]
    tsv = feature_dir / "data" / UNCITED_TSV
    if not tsv.exists():
        lines.append(f"- classification: **not produced** — `data/{UNCITED_TSV}`"
                     " absent; run the feature's classify_uncited_sections.py")
        return "\n".join(lines)
    rows = [r.split("\t") for r in
            tsv.read_text(encoding="utf-8").strip().split("\n")[1:]]
    hdr = tsv.read_text(encoding="utf-8").split("\n")[0].split("\t")
    ci, oi = hdr.index("classification"), hdr.index("outline")
    counts = Counter(r[ci] for r in rows)
    classified = {r[oi] for r in rows}
    lines.append(f"- classification (`data/{UNCITED_TSV}`, "
                 f"{len(rows)} rows): "
                 + "、".join(f"{k} {counts[k]}" for k in
                             ("container", "assumption", "figure", "substantive")))
    subs = [r[oi] for r in rows if r[ci] == "substantive"]
    lines.append(f"- **substantive: {len(subs)}** — {subs}")
    lines.append("  classify only; disposition of these is a ruling, not a"
                 " detection (§8.2, §8.4.2)")
    stale = sorted(set(uncited) ^ classified, key=outline_key)
    if stale:
        lines.append(f"- **STALE**: classification covers a different section"
                     f" set than the current export — symmetric difference"
                     f" {stale}; re-run the classifier")
    return "\n".join(lines)


# ---------------------------------------------------------------- emit

def emit(feature_dir: Path, cfg: dict, wbres: dict, a03res: dict,
         textlayer: str, hashes: dict, asserts: "Assertions", omap: dict,
         misses: list, omap_reason: str = "",
         signoff: dict | None = None) -> dict:
    state = wbres["state"]
    leaves = a03res["leaves"]
    targets = sorted(set(leaves) - set(wbres["done_reqs"]))
    uncovered = sorted(set(leaves) - set(wbres["done_reqs"])
                       - set(wbres["draft_reqs"]))
    all_037 = set(leaves) | set(a03res["headings"])
    # req_ids written in the workbook but absent from 037 (A-H06 pattern):
    # a traceability audit sees these as pointing at nothing
    orphan_done = sorted(set(wbres["done_reqs"]) - all_037)
    orphan_draft = sorted(set(wbres["draft_reqs"]) - all_037)

    def cap(items, n=8):
        """RECON.md is read by a human; recon.json carries the full list."""
        items = list(items)
        if len(items) <= n:
            return str(items) if items else "(none)"
        return (f"{items[:n]} … +{len(items) - n} more "
                "(full list in data/recon.json)")

    # A workbook can be FULL and still cover none of the ruled requirement
    # source — AM/FM 2026-08: 158 authored rows, all tracing a requirement
    # family that a ruling superseded. The state machine keys on draft rows, so
    # it cannot see this; naming it here keeps "FULL" from reading as "done".
    covered_n = len(set(leaves) & set(wbres["done_reqs"]))
    foreign = (wbres["done_rows"] > 0 and covered_n == 0 and bool(leaves))

    seg_lines = "\n".join(
        f"  - rows {s['start']}-{s['end']}: {s['cls']} ({s['n']} rows)"
        for s in wbres["segments"])
    amb_lines = "\n".join(f"  - row {i}: {why}"
                          for i, why in wbres["ambiguous_rows"]) or "  (none)"
    comp_lines = "\n".join(f"  - row {i}: {why}"
                           for i, why in wbres["compliance_notes"]) or "  (none)"
    pc = a03res["parent_child_dupes"]

    def chapter_hist(values) -> str:
        h = Counter(str(s).split(".")[0] for s in values)
        return "、".join(f"{ch}({n})" for ch, n in
                        sorted(h.items(), key=lambda kv: outline_key(kv[0]))) or "(none)"

    # Two different quantities, both reported because they are easy to
    # conflate: how many distinct sections each chapter contributes, and how
    # many leaves cite into it. A batch plan grouped "by spec chapter" is
    # sized by the second.
    sec_hist_line = chapter_hist(a03res["distinct_sections"])
    leaf_hist_line = chapter_hist(a03res["sections"].values())
    uncited_block = uncited_section_report(feature_dir, a03res, omap)

    recon = f"""# RECON — {cfg['feature']} (generated by recon.py)

## Inputs
{chr(10).join(f"- {k}: `{v['name']}` sha256={v['sha256'][:16]}…" for k, v in hashes.items())}
- spec text layer: {textlayer}

## Assertions — ruled constants, checked mechanically
{chr(10).join(asserts.lines())}

**{len(asserts.failed)} failed / {len(asserts.results)} checked.**
{"An assertion failure blocks DECISIONS.md; RECON.md is still written because it is the evidence." if asserts.failed else "DECISIONS.md written."}

## Spec outline map
- ruled export: `{(hashes.get('sys1_export') or {}).get('name', '(none)')}`
- outline entries in the export: {len(omap)}{f' — no map built: {omap_reason}' if omap_reason else ''}
- distinct sections cited by the leaves: {len(a03res['distinct_sections'])}
- cited sections NOT in the ruled export: **{len(misses)}** {misses[:10] if misses else ''}
- citation column: {a03res['source_col'] or 'NOT FOUND'}; stems: {a03res['citation_stems']}
- citation cells with extra lines below the section (audit evidence, not
  parsed): {a03res['multiline_citations']}
- distinct sections by chapter: {sec_hist_line}
- leaves by chapter: {leaf_hist_line}
- map written to `data/spec_id_to_outline.tsv` (tracked — a diff on it is the
  signal that the spec export moved underneath us)

## Uncited baseline sections
{uncited_block}

## Workbook
- workbook_state: **{state}**
- form layout revision: {wbres['layout_rev']}
- column mapping: {wbres['col_ok']} fields resolved from header text
  (authority; feature.yaml letters are only a prior)
{chr(10).join(f"  - {k} = {v}" for k, v in wbres['columns'].items())}
- feature.yaml column conflicts:
{chr(10).join(f"  - {c}" for c in wbres['col_conflicts']) or "  (none)"}
- authors present: {wbres['authors'] or '(none)'}
{("- **author detection**: " + wbres['author_note']) if wbres['author_note'] else ""}
- design-method vocabulary: {len(wbres['design_method_vocab'])} strings
- segments:
{seg_lines}
- done rows: {wbres['done_rows']} / draft rows: {wbres['draft_rows']}
- ambiguous rows (Tier 2 if any):
{amb_lines}
- done-region compliance notes (recorded, not fixed):
{comp_lines}

## Requirement report
- Categorization column: {a03res['categorization_col'] or 'NOT FOUND'}
- Categorization distribution: {a03res['categorization_distribution']}
- leaf criterion in force: **Categorization == Functional** → {len(a03res['leaves'])} leaves
{f'''- id-suffix criterion (`-NN` child shape, NOT in force): {a03res['naive_leaf_shape_count']} leaves
  — it drops {len(a03res['parent_shape_functional'])} parent-shaped rows that
  are themselves Functional Requirements
  {cap(a03res['parent_shape_functional'], 6)}'''
 if a03res['naive_leaf_shape_count'] else
 "- id-suffix criterion: not applicable — no leaf id in this family carries a "
 "`-NN` child suffix, so the shape carries no information here"}
- safety attributes (ASIL/FTTI) in the ruled source: {
    'PRESENT — ' + str(a03res['asil_distribution'])
    if a03res['has_safety_columns'] else '**ABSENT**'}
  - absent means the safety-analysis layer (SYS2/SYSRA) has no attachment
    point on these leaves and does NOT enter the trace chain. It says nothing
    about whether the requirements are safety-relevant — only that the ruled
    source does not claim they are.

## Coverage (vs 037)
- 037 leaves: {len(leaves)}; headings: {len(a03res['headings'])}
- covered by done region: {covered_n}
- regen targets: **{len(targets)}**
- covered nowhere (done nor draft): {len(uncovered)} {cap(uncovered)}
- parent/child both-leaf duplications: {cap(pc)}
- workbook req_ids ABSENT from 037 (traceability orphans, RD-1 candidates):
  - done region: {len(orphan_done)} {cap(orphan_done)}
  - draft region: {len(orphan_draft)} {cap(orphan_draft)}
{f'''
## Requirement-family mismatch — Tier 2

The workbook is **{state}** ({wbres['done_rows']} authored rows, no drafts) yet
covers **0 of {len(leaves)}** leaves in the ruled requirement source. Every
authored row traces a req_id the ruled source does not contain
({len(orphan_done)} distinct). Against the ruled source this workbook is
effectively BLANK, and "{state}" must not be read as "done".

This is not a state the canon's state machine can express — it keys on draft
rows, and there are none. Disposition of the existing rows (freeze as a legacy
region / replace / re-map) is a ruling, not a detection.
''' if foreign else ''}"""
    (feature_dir / "RECON.md").write_text(recon, encoding="utf-8")

    # ---- outline map: req_id -> section -> the exact spec_reference string
    (feature_dir / "data").mkdir(exist_ok=True)
    tsv = ["req_id\toutline\tpolarion_id\tspec_reference\ttitle"]
    tpl = cfg.get("spec_reference_template", "{outline}")
    for rid in sorted(a03res["sections"]):
        sec = a03res["sections"][rid]
        hit = omap.get(sec, {})
        tsv.append("\t".join([rid, sec, hit.get("id", ""),
                              tpl.replace("{outline}", sec), hit.get("title", "")]))
    (feature_dir / "data" / "spec_id_to_outline.tsv").write_text(
        "\n".join(tsv) + "\n", encoding="utf-8")

    (feature_dir / "data" / "recon.json").write_text(json.dumps({
        "workbook_state": state, "segments": wbres["segments"],
        "done_reqs": wbres["done_reqs"], "draft_reqs": wbres["draft_reqs"],
        "leaves": leaves, "regen_targets": targets, "uncovered": uncovered,
        "parent_child_dupes": pc,
        "orphan_done_reqs": orphan_done, "orphan_draft_reqs": orphan_draft,
        "design_method_vocab": wbres["design_method_vocab"],
        "ambiguous_rows": wbres["ambiguous_rows"],
        "compliance_notes": wbres["compliance_notes"],
        "columns": wbres["columns"], "layout_rev": wbres["layout_rev"],
        "col_conflicts": wbres["col_conflicts"],
        "authors": wbres["authors"], "author_used": wbres["author_used"],
        "has_safety_columns": a03res["has_safety_columns"],
        "asil_distribution": a03res["asil_distribution"],
        "assertions": asserts.results,
        "categorization_distribution": a03res["categorization_distribution"],
        "sections": a03res["sections"],
        "distinct_sections": a03res["distinct_sections"],
        "outline_misses": misses,
        "parent_shape_functional": a03res["parent_shape_functional"],
        "multiline_citations": a03res["multiline_citations"],
    }, ensure_ascii=False, indent=1), encoding="utf-8")

    # A failed assertion means a ruled constant does not hold against the files
    # in front of us. Pre-filling a decision sheet from that survey would hand
    # Tier 2 a sheet to sign whose [AUTO] values are known-wrong, so the sheet
    # is not written — RECON.md and recon.json carry the evidence and the run
    # exits non-zero.
    if asserts.failed:
        return {"decisions_written": False, "diverted": False}

    # ---- DECISIONS prefill from per-state strategy bindings (canon §2)
    blank = state == "BLANK"
    d = []
    d.append(f"# DECISIONS — {cfg['feature']} (FW036)\n")
    d.append("Pre-filled by recon.py. Markers per FEATURE_ONBOARDING §4; an\n"
             "unsigned sheet blocks Phase 4+. `[PROPOSED]` untouched at\n"
             "sign-off = binding as proposed.\n")
    d.append("## 1. Intake")
    d.append(f"- spec_mode: [AUTO] {cfg.get('spec_mode')}")
    d.append(f"- spec text layer: [AUTO] {textlayer}")
    d.append(f"- source files: [AUTO] {len(hashes)} present (SHA256 in RECON.md)")
    d.append(f"- ruled-constant assertions: [AUTO] {len(asserts.results)} checked,"
             f" {len(asserts.results) - len(asserts.failed)} PASS,"
             f" {len(asserts.failed)} FAIL (measured values in RECON.md)")
    if omap:
        d.append(f"- spec outline map: [AUTO] {len(a03res['distinct_sections'])}"
                 f" cited sections, all found in a {len(omap)}-entry ruled"
                 " export; map at data/spec_id_to_outline.tsv")
    d.append("\n## 2. Workbook survey")
    d.append(f"- workbook_state: [AUTO] {state}")
    d.append(f"- form layout revision: [AUTO] {wbres['layout_rev']}")
    d.append(f"- column mapping: [AUTO] {wbres['col_ok']} fields resolved from"
             " header text")
    if wbres["col_conflicts"]:
        d.append("- feature.yaml column letters: [PEI] "
                 f"{len(wbres['col_conflicts'])} disagree with the header —"
                 " update feature.yaml before Phase 4 (see RECON.md)")
    if wbres["author_note"]:
        d.append(f"- done_region.author_value: [PROPOSED: {wbres['author_used']}"
                 f" — feature.yaml value matches 0 rows]")
    if wbres["done_segments"]:
        d.append("- done segments: [AUTO] " + ", ".join(
            f"{s['start']}-{s['end']}" for s in wbres["done_segments"]))
    else:
        d.append("- done segments: [AUTO] none")
    if wbres["ambiguous_rows"]:
        d.append("- ambiguous rows: [PEI] see RECON.md — per-row disposition required")
    else:
        d.append("- ambiguous rows: [AUTO] none")
    if wbres["draft_rows"]:
        d.append("- draft disposition: [PROPOSED: discard & regenerate — lint"
                 " consistency cheaper than row salvage]")
    d.append(f"- design-method vocabulary: [AUTO] {len(wbres['design_method_vocab'])}"
             " exact strings from 下拉選單")
    if wbres["compliance_notes"]:
        d.append(f"- done-region compliance notes: [AUTO] "
                 f"{len(wbres['compliance_notes'])} recorded (frozen, not fixed)"
                 " — see RECON.md; register in ANOMALIES if new")
    d.append("\n## 3. Coverage")
    d.append(f"- 037 leaves: [AUTO] {len(leaves)}")
    if foreign:
        d.append(f"- requirement-family mismatch: [PEI] the {wbres['done_rows']}"
                 f" authored rows cover 0 of {len(leaves)} ruled leaves and"
                 f" trace {len(orphan_done)} req_ids absent from the ruled"
                 " source. Rule their disposition (freeze as legacy /"
                 " replace / re-map) BEFORE the write-back strategy — it"
                 " defines what 'done region' and 'completeness' mean here")
    if a03res["has_safety_columns"]:
        d.append(f"- safety attributes: [AUTO] present — "
                 f"{a03res['asil_distribution']}; safety layer joins the chain")
    else:
        d.append("- safety attributes: [PROPOSED: ruled source carries no"
                 " ASIL/FTTI column, so the SYS2/SYSRA safety layer does NOT"
                 " enter the trace chain]")
    d.append(f"- regen targets: [AUTO] {len(targets)} (list in recon.json)")
    if uncovered and blank:
        # Under BLANK every leaf is uncovered by construction — there is no
        # done region for it to be covered by. Demanding an anomaly per leaf
        # (the non-BLANK wording) would ask for 403 entries recording that a
        # blank workbook is blank.
        d.append(f"- covered nowhere: [AUTO] {len(uncovered)} = all leaves —"
                 " expected under BLANK, not an anomaly; this is the Phase 4"
                 " work list, not a gap")
    elif uncovered:
        d.append(f"- covered nowhere: [AUTO] {len(uncovered)} {cap(uncovered)}"
                 " — ANOMALIES entries required")
    if pc:
        d.append(f"- parent/child dupes: [PROPOSED: proportion test per case — {pc}]")
    if orphan_done or orphan_draft:
        d.append(f"- workbook req_ids absent from 037: [AUTO] done="
                 f"{len(orphan_done)} {cap(orphan_done, 4)} draft="
                 f"{len(orphan_draft)} {cap(orphan_draft, 4)} — ANOMALIES +"
                 " RD-1 required; scope the write-back traceability invariant"
                 " to regen rows only"
                 + (". NOTE: under BLANK these are template sample rows before"
                    " they are anything else — check the rows themselves"
                    " before filing an RD-1" if blank else ""))
    d.append("\n## 4. Style bindings")
    if blank:
        d.append("- style authority: [PROPOSED: fallback chain — no done region]")
        d.append("- test item shape: [PROPOSED: standard §4.3 tc_title]")
        d.append("- test group/set columns: [PROPOSED: FILL per framework Part N]")
        d.append("- exemplar source: [PROPOSED: nearest sibling feature done"
                 " region, cross-feature: style only]")
    else:
        if foreign:
            d.append("- style authority: [PROPOSED: the existing rows — they"
                     " are the workbook's only precedent — but they are NOT"
                     " the traceability authority; style may be borrowed from"
                     " rows whose req_ids the ruled source does not contain]")
        else:
            d.append("- style authority: [AUTO] done region")
        d.append("- test item shape: [PROPOSED: follow done-region first-row shape"
                 " — verify against profile]")
        d.append("- test group/set columns: [PROPOSED: match done-region"
                 " (blank if blank)]")
        d.append("- exemplar source: [AUTO] own done region")
    d.append(f"- author on new rows: [PROPOSED: {cfg['write_back']['author_value']}]")
    d.append(f"- spec_reference: [PROPOSED: {cfg['spec_reference_template']}]")
    # A tc_id scheme present in feature.yaml is one a ruling already froze, so
    # it is marked [RULED], not [PROPOSED]: re-offering it as a proposal
    # invites a change at sign-off, which is what freezing it forbids.
    if cfg["write_back"].get("tc_id_format"):
        d.append(f"- tc_id scheme: [RULED] {cfg['write_back']['tc_id_format']}"
                 " — frozen per this feature's RULINGS.md, not open at sign-off")
    d.append("\n## 5. Split & scope")
    d.append("- split_mode: [PROPOSED: standard]")
    d.append("\n## 6. Framework & profile")
    d.append("- Test Set table (Part N): [PEI — draft with Claude, Tier 2]")
    d.append("- profile [OVERRIDE] clauses: [PEI — draft with Claude, Tier 2]")
    d.append("\n## 7. Execution")
    d.append(f"- batch plan: [PROPOSED: group {len(targets)} targets by spec"
             " chapter, pilot = smallest coherent batch]")
    d.append("\n---\n\n## Sign-off\n\n- Reviewed by: ____  Date: ____\n"
             "- Overridden items: ____\n- Ruling notes:\n")
    written, diverted = write_decisions(
        feature_dir, "\n".join(d), signoff or {"signed": False})
    return {"decisions_written": True, "diverted": diverted,
            "decisions_path": written}


# ---------------------------------------------------------------- main

def resolve_glob(feature_dir: Path, pattern: str) -> Path | None:
    if not pattern:
        return None
    hits = sorted(feature_dir.glob(pattern))
    if not hits:
        sys.exit(f"input not found: {pattern} under {feature_dir}")
    if len(hits) > 1:
        sys.exit(f"ambiguous input {pattern}: {[h.name for h in hits]}")
    return hits[0]


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--feature", required=True,
                    help="feature directory (contains feature.yaml)")
    ap.add_argument("--root", default=".")
    args = ap.parse_args()

    feature_dir = Path(args.root).resolve() / args.feature
    cfg = yaml.safe_load((feature_dir / "feature.yaml").read_text("utf-8"))

    # Every declared input is hashed, not just the five the pipeline consumes:
    # a feature may declare its own (AM/FM's cfts_doc, sysra_export) and an
    # un-hashed input is one nobody can prove they surveyed.
    paths, hashes = {}, {}
    for key in dict.fromkeys(list(cfg["paths"]) + ["workbook", "a03_report",
                                                   "sys1_export", "spec_pdf",
                                                   "popup_list"]):
        pat = cfg["paths"].get(key)
        p = resolve_glob(feature_dir, pat) if pat else None
        paths[key] = p
        if p:
            hashes[key] = {"name": p.name, "sha256": sha256_file(p)}

    wbres = survey_workbook(cfg, paths["workbook"])
    a03res = survey_a03(paths["a03_report"])
    textlayer = survey_spec_text_layer(paths["spec_pdf"])
    omap, omap_reason = build_outline_map(paths.get("sys1_export"))
    asserts, misses = run_assertions(cfg, a03res, omap, omap_reason)
    signoff = read_signoff(feature_dir / "DECISIONS.md")
    outcome = emit(feature_dir, cfg, wbres, a03res, textlayer, hashes, asserts,
                   omap, misses, omap_reason, signoff)

    print("\n".join(["assertions:"] + [
        re.sub(r"\*\*|`", "", ln) for ln in asserts.lines()]), file=sys.stderr)
    print(f"recon complete: state={wbres['state']}, "
          f"leaves={len(a03res['leaves'])}, "
          f"sections={len(a03res['distinct_sections'])}, "
          f"targets={len(set(a03res['leaves']) - set(wbres['done_reqs']))}",
          file=sys.stderr)
    if wbres["ambiguous_rows"]:
        print(f"WARNING: {len(wbres['ambiguous_rows'])} ambiguous rows — "
              "Tier 2 review required (see RECON.md)", file=sys.stderr)

    # R-C10 — non-blocking. A sheet full of [PROPOSED] with an untouched
    # Sign-off block is indistinguishable, in the repo, from one nobody ever
    # looked at. Say so rather than let the silence read as approval.
    if (signoff["exists"] and not signoff["signed"]
            and signoff["has_proposed"] and not outcome["diverted"]):
        print("WARNING (R-C10): DECISIONS.md carries [PROPOSED] items and its "
              "Sign-off block is an unfilled placeholder — this feature's "
              "sign-off state is not knowable from the repo. Not blocking.",
              file=sys.stderr)

    # R-C9 — blocking. Reported after the survey so the run's evidence is on
    # screen alongside the refusal.
    if outcome["diverted"]:
        sys.exit(
            "REFUSED (R-C9): "
            f"{feature_dir / 'DECISIONS.md'} is signed "
            f"(Reviewed by: {signoff['reviewed_by'] or '—'}"
            f"{', ' + signoff['date'] if signoff['date'] else ''}"
            f"{'; ' + str(signoff['amendments']) + ' Amendment entries' if signoff['amendments'] else ''})"
            " and was NOT overwritten.\n"
            f"The freshly generated sheet was written to "
            f"{outcome['decisions_path']} instead — diff the two and merge by "
            "hand if the new survey should supersede the signed one.")

    if asserts.failed:
        sys.exit(f"FAILED: {len(asserts.failed)} ruled-constant assertion(s) do "
                 "not hold against these files. DECISIONS.md was NOT written; "
                 "see RECON.md for the measured values.")


if __name__ == "__main__":
    main()
