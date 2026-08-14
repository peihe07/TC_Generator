#!/usr/bin/env python3
"""Phase 0 Intake — drop-folder classification and missing-spec detection.

The user dumps whatever files they have into `_intake/<Feature>/` without
classifying them. This script:

1. Sniffs each file by CONTENT, not filename: workbook (Test Case
   Specification sheet, plus its Scope field and traced req-id family),
   swra_report (Analysis Report sheet, plus req family / row count / source
   column semantics / ASIL presence), polarion_export (Basic Report sheet,
   subtyped SYS1 spec export vs SYS2 safety export), popup_list (String/
   Popup Message header or PU-number ids), spec_pdf (text-layer probe),
   cfts_doc (.doc/.docx)
2. Derives the need list from the requirement report's source column WHEN it
   holds document citations (name_{section} shape). When the template's
   source column holds component or Polarion ids instead (the AM/FM SWRA
   family), it says so explicitly rather than crashing or emitting an empty
   list.
3. With multiple requirement reports present, the WORKBOOK's Scope field
   arbitrates which one is the TC source; the others are flagged for a
   Tier 2 role confirmation.
4. Proposes spec_mode (A–E per FEATURE_ONBOARDING §3)
5. With --scaffold, creates features/<feature>/ via new_feature.py
   conventions, moves classified files into inputs/, and pre-fills
   feature.yaml paths

Regression-validated (2026-08-09) against BOTH template families:
- Home corpus: all files classified under scrambled names; Last Mode absent
  → MISSING with citing-leaf count (15); present under a different release
  label → matched with an A-H03-pattern label warning
- AM/FM corpus: two Analysis-Report files disambiguated by workbook Scope
  (SWRA-A02 = TC source; the SWE1 037-A03 flagged for role confirmation);
  need list correctly reported as not derivable; spec_mode D proposed

Tier 0 output: INTAKE.md + intake.json. Obtaining missing files stays
Tier 3 (Pei).

Usage:
    python scripts/intake.py AMFM                # classify _intake/AMFM/
    python scripts/intake.py AMFM --scaffold     # + create features/amfm/, move files
"""

import argparse
import json
import re
import shutil
import sys
from collections import Counter
from pathlib import Path

import openpyxl

# ------------------------------------------------------------ sniffers

SHEET_SIGNATURES = [
    ("swra_report", lambda names: "Analysis Report" in names),
    ("workbook", lambda names: any("Test Case Specification" in n for n in names)),
    ("polarion_export", lambda names: "Basic Report" in names),
]


def sniff_xlsx(path: Path) -> tuple[str, str]:
    """Return (kind, note) for an xlsx by sheet structure."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception as e:  # not a real xlsx / corrupt
        return "unreadable", str(e)[:60]
    names = wb.sheetnames
    kind, note = "spec_xlsx", ""
    for k, test in SHEET_SIGNATURES:
        if test(names):
            kind = k
            if kind == "swra_report":
                note = _swra_profile(wb)
            elif kind == "polarion_export":
                note = _polarion_subtype(wb)
            elif kind == "workbook":
                note = _workbook_profile(wb, names)
            break
    else:
        if _is_popup_list(wb, names):
            kind = "popup_list"
        else:
            # supplemental spec xlsx (e.g. Last Mode Table): note its title
            title = _xlsx_title(wb, names)
            note = title or f"sheets: {', '.join(names[:3])}"
    wb.close()
    return kind, note


def _find_req_header(rows, id_words=("Requirement ID", "ID")):
    """Locate the requirement-table header row within the first 15 rows.
    Returns (index, row) or (None, None) — caller decides how loud to fail."""
    for i, r in enumerate(rows[:15]):
        cells = [str(c) for c in r if c]
        if not cells:
            continue
        has_id = any("Requirement ID" in c or c.strip() == "ID" for c in cells)
        has_desc = any("Description" in c for c in cells)
        if has_id and has_desc:
            return i, r
    return None, None


def _swra_profile(wb) -> str:
    """Summarize an Analysis Report: req-id family, row count, source shape."""
    ws = wb["Analysis Report"]
    rows = list(ws.iter_rows(values_only=True, max_row=1500))
    hdr_i, hdr = _find_req_header(rows)
    if hdr_i is None:
        return "Analysis Report present, header not recognized"
    cols = {str(c).strip(): j for j, c in enumerate(hdr) if c}
    id_col = next((j for h, j in cols.items() if "Requirement ID" in h or h == "ID"), 0)
    # Same preference order as cited_documents(): a report may carry both an
    # "HMI Source ID" holding document citations and a "Source Requirement ID"
    # holding upstream req ids. Describing the wrong one made this summary
    # contradict the need list computed two functions down.
    src_col = next((j for h, j in cols.items() if "HMI Source" in h), None)
    if src_col is None:
        src_col = next((j for h, j in cols.items()
                        if "Source" in h and "Description" not in h), None)
    data = [r for r in rows[hdr_i + 1:] if r[id_col]
            and not str(r[id_col]).startswith("<")]
    ids = [str(r[id_col]).strip() for r in data]
    fam = re.sub(r"-?\d+(-\d+)?$", "", ids[0]).rstrip("-") if ids else "?"
    shape = "no source column"
    if src_col is not None and data:
        sample = str(data[0][src_col] or "").split("\n")[0]
        if re.search(r"_[\d.]+\s*$", sample) and len(sample) > 20:
            shape = "sources: document citations (need-list derivable)"
        elif re.match(r"[A-Z]+R?\d*L?-\d+", sample.split("\n")[0].strip()):
            shape = "sources: Polarion work items (trace via export)"
        elif sample.strip():
            shape = "sources: component/architecture ids (trace via SYS3)"
    asil = "with ASIL/FTTI" if any("ASIL" in h for h in cols) else ""
    return f"req family {fam}, {len(ids)} rows, {shape}{', ' + asil if asil else ''}"


def _polarion_subtype(wb) -> str:
    ws = wb["Basic Report"]
    hdr = next(ws.iter_rows(values_only=True, max_row=1))
    cells = " ".join(str(c) for c in hdr if c)
    if "Sys-RA" in cells or "SYS2" in cells:
        return "SYS2 safety-analysis export (Sys-RA ids)"
    if "Outline Number" in cells:
        return "SYS1 spec export (outline numbers)"
    return "Polarion export, subtype unknown"


def _workbook_profile(wb, names) -> str:
    name = next(n for n in names if "Test Case Specification" in n)
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True, max_row=60))
    scope = ""
    for r in rows[:8]:
        cells = [str(c) for c in r if c]
        for j, c in enumerate(cells):
            if "Scope" in c and j + 1 < len(cells):
                scope = cells[j + 1][:60]
    req_fam = None
    for r in rows[9:]:
        if r and len(r) > 3 and r[3]:
            req_fam = re.sub(r"-?\d+(-\d+)?$", "", str(r[3]).strip()).rstrip("-")
            break
    return (f"Scope: {scope or '(blank)'}"
            + (f"; rows trace {req_fam}" if req_fam else "; no data rows"))


def _is_popup_list(wb, names) -> bool:
    """Detect by content: a sheet whose header row carries 'String/Popup
    Message' and whose id column holds PU-prefixed numbers."""
    for n in names[:4]:
        try:
            rows = wb[n].iter_rows(values_only=True)
        except Exception:
            continue
        for i, r in enumerate(rows):
            if i > 5:
                break
            cells = [str(c) for c in r if c]
            if any("String/Popup Message" in c for c in cells):
                return True
            if any(re.match(r"PU\d{3,4}$", c.strip()) for c in cells):
                return True
    return False


def _xlsx_title(wb, names) -> str:
    """Best-effort document title from a Title/cover sheet."""
    for n in names:
        if n.strip().lower() in ("title", "cover", "封面", "cover 封面"):
            for row in wb[n].iter_rows(values_only=True):
                cells = [str(c) for c in row if c]
                if len(cells) >= 2 and cells[0].strip().lower() in ("title",):
                    return cells[1][:80]
                if cells and len(cells[0]) > 8 and "title" not in cells[0].lower():
                    return cells[0][:80]
    return ""


def sniff_pdf(path: Path) -> tuple[str, str]:
    try:
        import pymupdf
    except ImportError:
        return "spec_pdf", "text layer unknown (pymupdf not installed)"
    doc = pymupdf.open(path)
    per_page = [len(p.get_text()) for p in doc]
    doc.close()
    text_pages = sum(1 for n in per_page if n > 100)
    if text_pages == 0:
        return "spec_pdf", f"scanned ({len(per_page)} pages, OCR path)"
    return "spec_pdf", f"text-layer on {text_pages}/{len(per_page)} pages"


def classify(folder: Path) -> list[dict]:
    out = []
    for p in sorted(folder.iterdir()):
        if p.name.startswith(".") or p.is_dir():
            continue
        if p.name in ("INTAKE.md", "intake.json"):
            continue
        ext = p.suffix.lower()
        if ext in (".xlsx", ".xlsm"):
            kind, note = sniff_xlsx(p)
        elif ext == ".pdf":
            kind, note = sniff_pdf(p)
        elif ext in (".doc", ".docx"):
            kind, note = "cfts_doc", "CFTS/Word candidate (spec_mode D)"
        else:
            kind, note = "unclassified", f"unhandled extension {ext}"
        out.append({"file": p.name, "kind": kind, "note": note})
    return out


# ------------------------------------------------------------ cited-spec diff

def doc_stem(cited: str) -> str:
    """Strip trailing _{n}/_{x.y} suffix and release-label noise → title stem."""
    s = re.sub(r"_[\d.]+\s*$", "", cited).strip()
    # cut at the first release-ish token to get the stable title prefix
    m = re.search(r"\b(R\d|SR\d|\()", s)
    if m and m.start() > 10:
        s = s[: m.start()]
    return re.sub(r"\s+", " ", s).strip().lower()


def cited_documents(a03_path: Path):
    """title-stem -> citing leaf count, from a source column that holds
    DOCUMENT citations. Returns a Counter; when this report template's
    source column holds component/Polarion ids instead, the Counter is
    empty and .reason explains why — the need list is then not derivable
    from this file, and intake says so rather than crashing (the original
    StopIteration bug) or emitting an empty list silently."""
    wb = openpyxl.load_workbook(a03_path, read_only=True)
    ws = wb["Analysis Report"]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i, hdr = _find_req_header(rows)
    if hdr_i is None:
        wb.close()
        r = Counter(); r.samples = {}; r.reason = (
            "requirement-table header not found in the first 15 rows — "
            "unknown 037 template, need list not derivable")
        return r
    cols = {str(c).strip(): j for j, c in enumerate(hdr) if c}
    id_col = next((j for h, j in cols.items()
                   if "Requirement ID" in h or h == "ID"), 0)
    # candidate source columns, most-specific first
    cand = [j for h, j in cols.items() if "HMI Source" in h]
    cand += [j for h, j in cols.items()
             if "Source" in h and "Description" not in h and j not in cand]
    cat_col = next((j for h, j in cols.items() if "Categorization" in h
                    and "Sub" not in h), None)
    data = [r for r in rows[hdr_i + 1:] if r[id_col]
            and not str(r[id_col]).startswith("<")]
    if cat_col is not None:
        data = [r for r in data
                if str(r[cat_col] or "").strip().startswith("Functional")]
    cnt: Counter = Counter()
    samples: dict = {}
    for src_col in cand:
        for r in data:
            # FIRST LINE ONLY. An HMI Source ID cell may carry the document
            # citation on line 1 and Polarion item ids on the lines below
            # (Comfort R-C4; 57 of 403 leaves). Matching the whole cell makes
            # the `_{n}` suffix test fail on exactly those rows, so the need
            # list silently undercounts — 346 instead of 403 — while still
            # naming the right document, which is the shape of undercount
            # nobody notices.
            s = str(r[src_col] or "").strip().split("\n")[0].strip()
            # document citation shape: long prefix + _{n} section suffix
            if s and re.search(r"_[\d.]+\s*$", s) and len(s) > 20:
                stem = doc_stem(s)
                cnt[stem] += 1
                samples.setdefault(stem, s)
        if cnt:
            break
    wb.close()
    cnt.samples = samples  # type: ignore[attr-defined]
    cnt.reason = ("" if cnt else
                  "source column holds component/Polarion ids, not document "
                  "names — need list not derivable from this template; trace "
                  "runs through the architecture/export files instead")
    return cnt


LABEL_TOKEN = re.compile(r"\b(R\d+[A-Z\-]*[A-Z]|R\d+L?-?[A-Z]?|SR\d+\s*\w*)\b", re.I)


def _labels(s: str) -> set:
    return {t.upper().replace(" ", "") for t in LABEL_TOKEN.findall(s)}


def match_cited_to_present(cited: Counter, files: list[dict]) -> list[dict]:
    """For each cited doc stem, find a present file whose name/title contains it."""
    results = []
    ranked = cited.most_common()
    for rank, (stem, n) in enumerate(ranked):
        hit, how = None, ""
        for f in files:
            hay = (f["file"] + " " + f.get("note", "")).lower()
            hay = re.sub(r"[_\-]", " ", hay)
            if stem and stem in re.sub(r"\s+", " ", hay):
                hit, how = f["file"], "name/title match"
                cited_labels = _labels(getattr(cited, "samples", {}).get(stem, ""))
                file_labels = _labels(f["file"])
                if cited_labels and file_labels and not (cited_labels & file_labels):
                    how = (f"name/title match — release label differs "
                           f"({'/'.join(sorted(cited_labels))} vs "
                           f"{'/'.join(sorted(file_labels))}), confirm same "
                           f"document (A-H03 pattern)")
                break
        if not hit and rank == 0:
            # dominant spec: a SYS1 export IS its text carrier even when the
            # filename does not repeat the title (content-based fallback)
            carrier = next((f for f in files if f["kind"] == "polarion_export"
                            and "SYS1" in f["note"]), None)
            if carrier:
                hit, how = carrier["file"], "satisfied by SYS1 export (dominant spec)"
        results.append({"cited": stem, "leaves": n, "present": hit, "how": how})
    return results


# ------------------------------------------------------------ spec_mode

def propose_spec_mode(files: list[dict]) -> tuple[str, str]:
    kinds = {f["kind"] for f in files}
    notes = " ".join(f["note"] for f in files)
    sys1 = any(f["kind"] == "polarion_export" and "SYS1" in f["note"]
               for f in files)
    if sys1:
        pdf = "with figure PDF" if "spec_pdf" in kinds else "no PDF"
        return "A", f"SYS1 spec export present ({pdf})"
    if "cfts_doc" in kinds:
        return "D", "CFTS/Word is the spec source; spec_reference by lookup"
    if "spec_pdf" in kinds:
        if "scanned" in notes:
            return "C", "scanned PDF only (OCR pipeline)"
        return "B", "text-layer PDF only"
    return "E", "no spec source found — expect heavy BLOCKED + RD-1"


# ------------------------------------------------------------ emit / scaffold

def emit(feature: str, folder: Path, files: list[dict], cited,
         mode: tuple[str, str], missing_a03: bool) -> None:
    lines = [f"# INTAKE — {feature} (generated by intake.py)", ""]
    lines.append("## Classified files")
    for f in files:
        note = f" — {f['note']}" if f["note"] else ""
        lines.append(f"- `{f['file']}` → **{f['kind']}**{note}")
    lines.append("")
    swras = [f for f in files if f["kind"] == "swra_report"]
    wbk = next((f for f in files if f["kind"] == "workbook"), None)
    if len(swras) > 1:
        lines.append("## Multiple requirement reports present — workbook Scope arbitrates")
        scope = (wbk or {}).get("note", "")
        for s in swras:
            in_scope = wbk and any(
                tok and tok.lower() in scope.lower()
                for tok in re.findall(r"[A-Za-z0-9][A-Za-z0-9\-]{6,}", s["file"]))
            tag = ("**TC source (named in workbook Scope)**" if in_scope
                   else "present, NOT named in workbook Scope — confirm its role (Tier 2)")
            lines.append(f"- `{s['file']}` — {s['note']} → {tag}")
        lines.append("")
    lines.append("## Documents cited by the requirement report (need list)")
    if missing_a03:
        lines.append("- **NO requirement report found** — cannot derive the "
                     "need list; this is the first file to obtain")
    elif getattr(cited, "reason", ""):
        lines.append(f"- not derivable: {cited.reason}")
    else:
        for c in cited:
            if c["present"]:
                mark = f"present: `{c['present']}` ({c['how']})"
            else:
                mark = f"**MISSING** — cited by {c['leaves']} leaves"
            lines.append(f"- `{c['cited']}` ({c['leaves']} leaves) → {mark}")
    lines.append("")
    lines.append(f"## Proposed spec_mode: **{mode[0]}** — {mode[1]}")
    missing = ([c for c in cited if isinstance(c, dict) and not c["present"]]
               if isinstance(cited, list) else [])
    lines.append("")
    if missing:
        total = sum(c["leaves"] for c in missing)
        lines.append(f"## Action: obtain {len(missing)} missing document(s) "
                     f"(else {total} leaves start BLOCKED), then re-run intake")
    else:
        lines.append("## Action: complete — run `--scaffold` then Phase 1 recon")
    (folder / "INTAKE.md").write_text("\n".join(lines), encoding="utf-8")
    cited_json = cited if isinstance(cited, list) else {
        "not_derivable": getattr(cited, "reason", "")}
    (folder / "intake.json").write_text(json.dumps(
        {"feature": feature, "files": files, "cited": cited_json,
         "spec_mode": mode[0]}, ensure_ascii=False, indent=1), encoding="utf-8")
    print("\n".join(lines))


# intake kind -> feature.yaml paths key. swra_report fills a03_report only
# when it is the (single or Scope-designated) TC source — handled in scaffold.
KIND_TO_YAML = {
    "workbook": "workbook", "swra_report": "a03_report",
    "polarion_export": "sys1_export", "spec_pdf": "spec_pdf",
    "popup_list": "popup_list",
}


def scaffold(feature: str, folder: Path, files: list[dict], mode: str,
             root: Path, a03_pick: dict | None = None) -> None:
    """Create features/<feature>/, move classified files in, pre-fill feature.yaml.

    `a03_pick` is the Scope-arbitrated requirement report. It matters when
    several are present: without it every swra_report writes a03_report in
    turn and the last one in sorted order silently wins, which is both
    non-deterministic in intent and free to disagree with the arbitration the
    report just printed. When the choice was contested the written line is
    annotated so the Tier 2 reviewer sees a decision was made, not a default.
    """
    feat_dir = root / "features" / feature.lower()
    if not feat_dir.exists():
        import subprocess
        subprocess.run(
            [sys.executable, str(root / "scripts" / "new_feature.py"),
             feature, "--root", str(root)], check=True)
    inputs = feat_dir / "inputs"
    inputs.mkdir(exist_ok=True)
    yaml_path = feat_dir / "feature.yaml"
    text = yaml_path.read_text(encoding="utf-8")
    swras = [f for f in files if f["kind"] == "swra_report"]
    for f in files:
        if f["kind"] in ("unclassified", "unreadable"):
            continue
        shutil.move(str(folder / f["file"]), str(inputs / f["file"]))
        key = KIND_TO_YAML.get(f["kind"])
        if key == "a03_report" and a03_pick and f["file"] != a03_pick["file"]:
            continue          # not the Scope-designated TC source
        if key:
            text = re.sub(
                rf'({key}:\s*)"[^"]*"', rf'\g<1>"inputs/{f["file"]}"', text)
    if len(swras) > 1:
        others = "\n".join(f'  #   inputs/{s["file"]}'
                           for s in swras if not a03_pick
                           or s["file"] != a03_pick["file"])
        text = re.sub(
            r'(  a03_report:[^\n]*\n)',
            rf'\g<1>  # CONFIRM (Tier 2): {len(swras)} requirement reports '
            'present. The line above is\n'
            "  # the one the workbook's Scope field names; a ruling may "
            'override it. Others:\n' + others.replace("\\", "\\\\") + "\n",
            text, count=1)
    text = re.sub(r'(spec_mode:\s*)"[A-E]"', rf'\g<1>"{mode}"', text)
    yaml_path.write_text(text, encoding="utf-8")
    print(f"\nscaffolded {feat_dir}; classified files moved to inputs/; "
          f"feature.yaml paths + spec_mode pre-filled")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("feature")
    ap.add_argument("--root", default=".")
    ap.add_argument("--scaffold", action="store_true")
    args = ap.parse_args()
    root = Path(args.root).resolve()
    folder = root / "_intake" / args.feature
    if not folder.is_dir():
        sys.exit(f"drop folder not found: {folder}")

    files = classify(folder)
    swras = [f for f in files if f["kind"] == "swra_report"]
    cited, matched, pick = Counter(), [], None
    if swras:
        # prefer the report the workbook Scope names; else first
        wbk = next((f for f in files if f["kind"] == "workbook"), None)
        pick = swras[0]
        if wbk and len(swras) > 1:
            for s in swras:
                toks = re.findall(r"[A-Za-z0-9][A-Za-z0-9\-]{6,}", s["file"])
                if any(t.lower() in wbk["note"].lower() for t in toks):
                    pick = s
                    break
        cited = cited_documents(folder / pick["file"])
        matched = match_cited_to_present(cited, files)
    mode = propose_spec_mode(files)
    emit(args.feature, folder, files,
         matched if not getattr(cited, "reason", "") else cited,
         mode, missing_a03=not swras)
    if args.scaffold:
        missing = [c for c in matched if not c["present"]]
        if missing:
            print("\nNOTE: scaffolding with missing documents — the affected "
                  "leaves will start BLOCKED", file=sys.stderr)
        scaffold(args.feature, folder, files, mode[0], root, a03_pick=pick)


if __name__ == "__main__":
    main()
