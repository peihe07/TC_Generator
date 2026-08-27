#!/usr/bin/env python3
"""行首空白之量化矩陣（27 包 §D-4；IN §11 之檢測面）。

兩層語料各掃一次，**共用 lint036 之 V 檢查本體**（同一判準，含其例外表）：

* (a) `features/<f>/generated/*.json` —— LLM 產出之七欄位
* (b) 已寫回工作簿 —— openpyxl `read_only`（R-G3：不得開啟寫入）

**本工具只量化，不修語料**（27 包 §D-4）。

G-D：掃不到的簿／檔記 `未掃` 並載明原因，**不以 0 代**——
一個永遠空的清單與一個壞掉的清單，其輸出相同。
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import lint036  # noqa: E402

# json 鍵 -> lint036 之欄位鍵（Q_FIELDS 之七欄）
JSON_FIELD_MAP = {
    "test_item": "test_item",
    "test_set": "test_set",
    "pre_conditions": "pre",
    "input_test_data": "input",
    "test_procedure": "proc",
    "expected_result": "er",
    "specification_reference": "spec",
}
FIELDS = tuple(lint036.Q_FIELDS)
OUT_DEFAULT = "docs/reports/whitespace_matrix.tsv"
COLUMNS = ("layer", "unit", "status", "rows") + FIELDS + ("total", "note")


def scan_text(text: str) -> str | None:
    """單一 cell 之判定；回傳 `blank`／`leading`／None。**判準取自 lint036。**"""
    for line in text.split("\n"):
        if lint036.RE_V_BLANK_WS.match(line):
            return "blank"
        if lint036.RE_V_LEADING_WS.match(line) and not lint036.RE_V_EXEMPT.match(line):
            return "leading"
    return None


def records_of(payload) -> list[dict]:
    """遞迴取出**全部**帶七欄位之 dict。

    **不得以「第一個 dict 之 list」為準**——`SWE1-HMI-PROF-111-china.json`
    之 TC 列在 `tcs` 鍵下，而其前之 `outline` 亦為 dict 之 list，
    取首者遂掃到 outline 而回報 0 命中。**那個 0 是掃錯地方掃出來的 0**，
    與「掃過且乾淨」不可分辨 —— G-D 所指之同一形態。
    """
    found: list[dict] = []

    def walk(node) -> None:
        if isinstance(node, dict):
            if any(k in node for k in JSON_FIELD_MAP):
                found.append(node)
            for value in node.values():
                walk(value)
        elif isinstance(node, list):
            for value in node:
                walk(value)

    walk(payload)
    return found


def scan_json(path: Path) -> tuple[str, int, Counter, str]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return "未掃", 0, Counter(), f"讀取失敗：{type(exc).__name__}"
    records = records_of(payload)
    if not records:
        return "未掃", 0, Counter(), "檔內無 TC 列（非語料檔）"
    hits: Counter = Counter()
    for record in records:
        for json_key, field in JSON_FIELD_MAP.items():
            value = record.get(json_key)
            if isinstance(value, str) and scan_text(value):
                hits[field] += 1
    return "已掃", len(records), hits, ""


def scan_workbook(path: Path) -> tuple[str, int, Counter, str]:
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:                      # 損壞／非 xlsx／受保護
        return "未掃", 0, Counter(), f"開啟失敗：{type(exc).__name__}"
    hits: Counter = Counter()
    rows = 0
    try:
        sheets = [s for s in wb.sheetnames if s.startswith(lint036.TC_SHEET_PREFIX)]
        if not sheets:
            return "未掃", 0, Counter(), f"無 {lint036.TC_SHEET_PREFIX!r} 開頭之 sheet"
        for name in sheets:
            ws = wb[name]
            try:
                header_row = lint036.find_header_row(ws)
            except ValueError as exc:
                return "未掃", 0, Counter(), str(exc)[:60]
            raw_rows = list(ws.iter_rows(min_row=1, values_only=True))
            columns = lint036.build_column_map(list(raw_rows[header_row - 1]))
            missing = [k for k in FIELDS if k not in columns]
            if missing:
                return "未掃", 0, Counter(), f"sheet {name} 缺欄位 {missing}"
            for raw in raw_rows[header_row:]:
                cells = {k: lint036.cell_text(raw[i]) if i < len(raw) else ""
                         for k, i in columns.items() if k in FIELDS}
                if not any(cells[k].strip() for k in ("test_item", "proc", "er")):
                    continue
                rows += 1
                for field, text in cells.items():
                    if scan_text(text):
                        hits[field] += 1
    finally:
        wb.close()
    return "已掃", rows, hits, ""


def row_of(layer: str, unit: str, status: str, rows: int,
           hits: Counter, note: str) -> str:
    counts = [str(hits.get(f, 0)) if status == "已掃" else "未掃" for f in FIELDS]
    total = str(sum(hits.values())) if status == "已掃" else "未掃"
    return "\t".join([layer, unit, status, str(rows), *counts, total, note])


def main() -> int:
    ap = argparse.ArgumentParser(description="行首空白量化矩陣（IN §11）")
    ap.add_argument("--root", default=".")
    ap.add_argument("--out", default=OUT_DEFAULT)
    ap.add_argument("--skip-workbooks", action="store_true",
                    help="只跑 (a) 層；(b) 層逐簿於報表標未掃")
    args = ap.parse_args()

    root = Path(args.root).resolve()
    lines: list[str] = ["\t".join(COLUMNS)]
    tally = {"已掃": 0, "未掃": 0}
    grand: Counter = Counter()

    for path in sorted(root.glob("features/*/generated/*.json")):
        status, rows, hits, note = scan_json(path)
        tally[status] += 1
        grand.update(hits)
        lines.append(row_of("json", str(path.relative_to(root)), status, rows, hits, note))

    books = [p for p in sorted(root.glob("features/*/**/*.xlsx"))
             if "inputs" not in p.parts]
    for path in books:
        if args.skip_workbooks:
            lines.append(row_of("xlsx", str(path.relative_to(root)), "未掃", 0,
                                Counter(), "--skip-workbooks"))
            tally["未掃"] += 1
            continue
        status, rows, hits, note = scan_workbook(path)
        tally[status] += 1
        grand.update(hits)
        lines.append(row_of("xlsx", str(path.relative_to(root)), status, rows, hits, note))

    out = root / args.out
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"寫入 {args.out}：{len(lines) - 1} 單位"
          f"（已掃 {tally['已掃']}／未掃 {tally['未掃']}）")
    print("  逐欄命中：" + "／".join(f"{f} {grand.get(f, 0)}" for f in FIELDS))
    print(f"  合計 {sum(grand.values())}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
