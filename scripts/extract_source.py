#!/usr/bin/env python3
"""來源文件之文字形抽取（R-G27，27 包 §D-7）。

`sources/raw/<doc_id>/` 之原檔 → `sources/extracted/<doc_id>/`：

* `.xlsx` → 逐 sheet 一份 tsv（openpyxl **read_only**；R-G3 禁寫）
* `.pdf`  → 逐頁一份 md

每份抽取物之首列（tsv）／首行（md）帶**來源 sha256**，
其後每次比對即可判知抽取物是否對得上現行原檔。

**自驗（§F-6）**：每個 sheet 抽完後重讀原檔，比對
`行數` 與 `非空儲存格數`；不符即抽取失真，**停**（exit 2），不寫該檔。
"""

from __future__ import annotations

import argparse
import hashlib
import sys
from pathlib import Path

RAW_ROOT = "sources/raw"
EXTRACTED_ROOT = "sources/extracted"
MANIFEST = "sources/MANIFEST.tsv"
MANIFEST_COLUMNS = ("doc_id", "filename", "sha256", "version", "features", "note")


class ExtractionMismatch(RuntimeError):
    """抽取後之量測與原檔不符（§F-6）。"""


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def safe_name(text: str) -> str:
    """sheet 名 → 檔名 —— 路徑分隔與空白一律換掉，不得逃出目標目錄。

    A-POP1：原本以 `strip("._")` 剝掉**前導**底線，`_polarion` 遂與同簿之
    `Polarion` 撞名；macOS 之 APFS 大小寫不敏感，後寫者靜默覆蓋前者，
    而 §F-6 自驗回讀的是自己剛寫的檔，測不到這種遺失。改為只剝尾端 ——
    前導 `_` 不會逃出目標目錄，`..` 則因 `.` 不在保留集以外而由下方
    的 `startswith(".")` 檔掉。
    """
    out = "".join(ch if ch.isalnum() or ch in "-_." else "_" for ch in text.strip())
    out = out.rstrip("._")
    if out.startswith("."):
        out = "sheet_" + out.lstrip(".")
    return out or "sheet"


def cell_text(value) -> str:
    """tsv 之單一儲存格 —— tab／換行以字面轉義，否則欄位會錯位。"""
    if value is None:
        return ""
    return (str(value).replace("\\", "\\\\")
            .replace("\t", "\\t").replace("\n", "\\n").replace("\r", ""))


def measure(rows: list[tuple]) -> tuple[int, int]:
    """`(行數, 非空儲存格數)` —— 自驗之兩個量。"""
    filled = sum(1 for row in rows for v in row
                 if v is not None and str(v).strip() != "")
    return len(rows), filled


def extract_xlsx(path: Path, out_dir: Path, source_sha: str) -> list[tuple[str, int, int]]:
    import openpyxl

    def read_all() -> dict[str, list[tuple]]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            return {name: list(wb[name].iter_rows(values_only=True))
                    for name in wb.sheetnames}
        finally:
            wb.close()

    sheets = read_all()
    written: list[tuple[str, int, int]] = []
    taken: dict[str, str] = {}
    for name, rows in sheets.items():
        # A-POP1：撞名一律停，不靜默覆蓋。比對以 casefold 為之 ——
        # 檔案系統（APFS／NTFS）不敏感，只比原字串測不到。
        key = safe_name(name).casefold()
        if key in taken:
            raise ExtractionMismatch(
                f"{path.name}：sheet {name!r} 與 {taken[key]!r} 之抽取檔名"
                f"皆為 {safe_name(name)!r}（大小寫不敏感檔案系統下會互相覆蓋）")
        taken[key] = name
        n_rows, n_filled = measure(rows)
        body = "\n".join(
            [f"# source_sha256\t{source_sha}\t sheet\t{name}"]
            + ["\t".join(cell_text(v) for v in row) for row in rows]
        ) + "\n"
        target = out_dir / f"{safe_name(name)}.tsv"
        target.write_text(body, encoding="utf-8")

        # 自驗（§F-6）：**回讀所寫之 tsv**，其兩個量須與 read_only 實測相符。
        # **不重讀原檔** —— 那只證明 openpyxl 兩次讀出同一份東西，
        # 對序列化（跳脫、欄位對齊）恆真，測不到抽取失真之實際發生處。
        back = target.read_text(encoding="utf-8").splitlines()[1:]
        r2 = len(back)
        f2 = sum(1 for line in back for cell in line.split("\t") if cell.strip())
        if (r2, f2) != (n_rows, n_filled):
            target.unlink()
            raise ExtractionMismatch(
                f"{path.name} sheet {name!r}：行數／非空儲存格 "
                f"原檔 ({n_rows}, {n_filled})，抽取物 ({r2}, {f2})")
        written.append((name, n_rows, n_filled))
    return written


def extract_pdf(path: Path, out_dir: Path, source_sha: str) -> list[tuple[str, int, int]]:
    import pdfplumber

    written: list[tuple[str, int, int]] = []
    with pdfplumber.open(path) as pdf:
        parts = [f"<!-- source_sha256: {source_sha} -->", f"# {path.name}", ""]
        for number, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ""
            parts += [f"## p.{number}", "", text, ""]
            written.append((f"p.{number}", len(text.splitlines()), len(text)))
    (out_dir / f"{path.stem}.md").write_text("\n".join(parts) + "\n", encoding="utf-8")
    return written


def extract_one(root: Path, path: Path) -> tuple[str, list[tuple[str, int, int]]]:
    doc_id = path.parent.name
    out_dir = root / EXTRACTED_ROOT / doc_id
    out_dir.mkdir(parents=True, exist_ok=True)
    source_sha = sha256_of(path)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return doc_id, extract_xlsx(path, out_dir, source_sha)
    if suffix == ".pdf":
        return doc_id, extract_pdf(path, out_dir, source_sha)
    raise ValueError(f"{path.name}：不支援之型別 {suffix}（.dbc 之抽取待裁）")


def manifest_rows(root: Path) -> list[list[str]]:
    rows = []
    for path in sorted((root / RAW_ROOT).rglob("*")):
        if not path.is_file():
            continue
        rows.append([path.parent.name, path.name, sha256_of(path),
                     "未載明", "未載明", ""])
    return rows


def main() -> int:
    ap = argparse.ArgumentParser(description="來源文件抽取（R-G27）")
    ap.add_argument("--root", default=".")
    ap.add_argument("--doc-id", default=None, help="只跑該 doc_id；預設全跑")
    ap.add_argument("--refresh-manifest", action="store_true",
                    help="以 raw/ 現況重寫 MANIFEST 之 doc_id／檔名／sha 三欄")
    args = ap.parse_args()

    root = Path(args.root).resolve()
    raw = root / RAW_ROOT
    if not raw.is_dir():
        print(f"FAIL: {RAW_ROOT} 不存在", file=sys.stderr)
        return 1

    targets = sorted(p for p in raw.rglob("*")
                     if p.is_file() and (args.doc_id is None
                                         or p.parent.name == args.doc_id))
    if not targets:
        print(f"{RAW_ROOT} 內無原檔 —— 機制已備，本輪不造例（27 包 §D-7）")
    for path in targets:
        try:
            doc_id, written = extract_one(root, path)
        except ExtractionMismatch as exc:
            print(f"FAIL（§F-6 抽取失真）: {exc}", file=sys.stderr)
            return 2
        except ValueError as exc:
            print(f"跳過 {path.name}：{exc}")
            continue
        print(f"{doc_id}／{path.name} → {len(written)} 份")
        for name, n_rows, n_filled in written:
            print(f"    {name:<28} 行 {n_rows:>6}   非空儲存格 {n_filled:>7}")

    if args.refresh_manifest:
        rows = manifest_rows(root)
        body = "\n".join(["\t".join(MANIFEST_COLUMNS)]
                         + ["\t".join(r) for r in rows]) + "\n"
        (root / MANIFEST).write_text(body, encoding="utf-8")
        print(f"寫入 {MANIFEST}：{len(rows)} 列")
    return 0


if __name__ == "__main__":
    sys.exit(main())
