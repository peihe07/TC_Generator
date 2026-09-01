#!/usr/bin/env python3
"""交付規格表之讀者（R-G42，DELIVERY-SPEC 閘）。

只掃 `features/*/delivered/*.xlsx`；不掃 output/、sandbox/、generated/。
R-G42 生效前已在 delivered/ 之檔（基線 `docs/fw036/DELIVERY_SPEC_BASELINE.tsv`）
只計警示不判紅（Pei 2026-08-30：不回歸）；基線外之檔判紅。基線只減不增。

逐檔判準（R-G42 一～七）：
  1a  D 欄 req_id 數值升冪（數字 token 逐段比較）
  1b  037 有列而無 TC 之需求須有僅填 D 欄之空列 —— 母體取
      feature.yaml `delivery.leaf_ids`（tsv，一行一 id）；未宣告則報「未比對」
  2   F 欄全部 `NR1L-{ABBR}-{nnn}`，ABBR 單一且 = feature.yaml `delivery.tc_id_abbr`
  3   G 欄唯一值 = feature.yaml `delivery.test_group`（缺則取頂層 `feature`）
  4   Author 全 PeiPYHsu；Priority 全 ∈ P0–P3；Est. Time 全空（警）
  5   檔名為客戶檔名形制且 {FeatureName} = G 欄去空白；MANIFEST.tsv 有列且 sha 相符
  6   同目錄有 DELIVERY_NOTE.md；未結 DR 清單存在（同目錄檔名含 DR／PENDING，
      或 DELIVERY_NOTE.md 內有 `DR-` 段）
  7   I～N 欄 PENDING 計數 = 0，或 MANIFEST note 欄含 `R-` 號

唯讀；`--emit-baseline` 時只寫基線 tsv。
"""

from __future__ import annotations

import argparse
import hashlib
import re
import sys
import warnings
from dataclasses import dataclass, field
from pathlib import Path

warnings.filterwarnings("ignore")

BASELINE_DEFAULT = "docs/fw036/DELIVERY_SPEC_BASELINE.tsv"
MANIFEST_NAME = "MANIFEST.tsv"
SHEET_KEY = "Test Case Specification"
AUTHOR = "PeiPYHsu"
PRIORITIES = {"P0", "P1", "P2", "P3"}
TC_ID_RE = re.compile(r"^NR1L-([A-Za-z]+)-(\d{3})$")
FILENAME_RE = re.compile(
    r"^FM-WI-FSM-036-A01 STLA 測試用例規範與結果_SWQT STLA Test Case Specification"
    r" & Result_SWQT_(?P<name>[A-Za-z0-9]+)_(?P<date>\d{8})\.xlsx$"
)
HEADER_KEYS = {
    "req": "Requirement or Design ID",
    "tc": "Test Case ID",
    "group": "Test Group",
    "item": "Test Item",
    "pre": "Pre-Condition",
    "input": "Input Test Data",
    "proc": "Test procedure",
    "er": "Expected Result",
    "spec": "Specification Reference",
    "prio": "Test Case Priority",
    "est": "Estimated Test Time",
    "author": "Test Case Author",
}


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def req_key(s: str) -> tuple:
    # Numeric tokens compare as int, alphabetic tokens as str: "-10" sorts after "-2".
    return tuple(int(t) if t.isdigit() else t
                 for t in re.findall(r"\d+|[A-Za-z]+", s))


@dataclass
class Sheet:
    cols: dict[str, int]
    rows: list[tuple]           # data rows (row 10 onward) with any non-empty cell
    header_row: int


def read_sheet(path: Path) -> Sheet:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    names = [n for n in wb.sheetnames if SHEET_KEY in n]
    ws = wb[names[0]] if names else wb[wb.sheetnames[0]]
    cols: dict[str, int] = {}
    header_row = 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        hits = {k: j for j, v in enumerate(row) if isinstance(v, str)
                for k, key in HEADER_KEYS.items() if key in v and k not in cols}
        # Later columns overwrite earlier hits, so the merged C-column label yields to column D.
        for k, j in hits.items():
            cols[k] = j
        if "tc" in cols and "req" in cols:
            header_row = i
            break
    if not header_row:
        raise ValueError(f"{path.name}: 表頭未命中（{list(cols)}）")
    rows = [r for r in ws.iter_rows(min_row=header_row + 1, values_only=True) if any(v not in (None, "") for v in r)]
    return Sheet(cols, rows, header_row)


def cell(row: tuple, idx: int):
    return row[idx] if idx < len(row) else None


def load_yaml(feature_dir: Path) -> dict:
    p = feature_dir / "feature.yaml"
    if not p.exists():
        return {}
    import yaml
    with p.open(encoding="utf-8") as fh:
        return yaml.safe_load(fh) or {}


def load_manifest(folder: Path) -> dict[str, dict]:
    p = folder / MANIFEST_NAME
    out: dict[str, dict] = {}
    if not p.exists():
        return out
    lines = p.read_text(encoding="utf-8").splitlines()
    if not lines:
        return out
    head = lines[0].split("\t")
    for line in lines[1:]:
        if not line.strip():
            continue
        cells = line.split("\t")
        rec = {head[i]: cells[i] for i in range(min(len(head), len(cells)))}
        out[cells[0]] = rec
    return out


@dataclass
class Verdict:
    path: str
    red: list[str] = field(default_factory=list)
    warn: list[str] = field(default_factory=list)
    info: list[str] = field(default_factory=list)


def check_book(root: Path, book: Path) -> Verdict:
    rel = str(book.relative_to(root))
    v = Verdict(rel)
    feature_dir = book.parent.parent
    cfg = load_yaml(feature_dir)
    delivery = cfg.get("delivery") or {}
    manifest = load_manifest(book.parent)
    sheet = read_sheet(book)
    c = sheet.cols
    req_col, tc_col = c["req"], c["tc"]

    # Header scan keeps the rightmost hit, so `req` resolves to column D (C is a merged label).
    req_rows = [(i, str(cell(r, req_col) or "").strip(), r)
                for i, r in enumerate(sheet.rows, sheet.header_row + 1)]
    tc_rows = [(i, rq, r) for i, rq, r in req_rows if cell(r, tc_col) not in (None, "")]

    # 1a order
    prev = None
    breaks = 0
    for i, rq, _ in req_rows:
        if not rq:
            continue
        k = req_key(rq)
        if prev and k < prev[0]:
            breaks += 1
            if breaks == 1:
                v.red.append(f"一(a) 列序：row {i} `{rq}` < 前列 `{prev[1]}`")
        prev = (k, rq)
    if breaks > 1:
        v.red.append(f"一(a) 列序：共 {breaks} 處逆序")

    # 1b coverage rows
    empty_req_rows = [i for i, rq, r in req_rows if rq and cell(r, tc_col) in (None, "")]
    leaf_path = delivery.get("leaf_ids")
    if leaf_path:
        leaf_file = feature_dir / leaf_path
        if leaf_file.exists():
            leaves = {l.strip() for l in leaf_file.read_text(encoding="utf-8").splitlines() if l.strip()}
            present = {rq for _, rq, _ in req_rows if rq}
            missing = sorted(leaves - present, key=req_key)
            if missing:
                v.red.append(f"一(b) 037 有列而簿上無列：{len(missing)}（{', '.join(missing[:5])}…）")
        else:
            v.red.append(f"一(b) feature.yaml delivery.leaf_ids 指向不存在之檔：{leaf_path}")
    else:
        v.info.append(f"一(b) 未比對母體（feature.yaml 無 delivery.leaf_ids）；簿上無 TC 之需求列 {len(empty_req_rows)}")
    for i in empty_req_rows:
        r = sheet.rows[i - sheet.header_row - 1]
        others = [cell(r, j) for j in range(len(r)) if j not in (0, 1, req_col - 1, req_col)]
        if any(x not in (None, "") for x in others):
            v.red.append(f"一(b) row {i} 無 TC 之需求列其餘欄非空")

    # 2 TC id
    abbrs: set[str] = set()
    bad_ids = []
    seqs = []
    for i, _, r in tc_rows:
        tc = str(cell(r, tc_col)).strip()
        m = TC_ID_RE.match(tc)
        if not m:
            bad_ids.append((i, tc))
        else:
            abbrs.add(m.group(1))
            seqs.append(int(m.group(2)))
    if bad_ids:
        v.red.append(f"二 TC ID 形制：{len(bad_ids)} 列不合 NR1L-{{ABBR}}-{{nnn}}（首例 row {bad_ids[0][0]} `{bad_ids[0][1]}`）")
    if len(abbrs) > 1:
        v.red.append(f"二 ABBR 不唯一：{sorted(abbrs)}")
    declared_abbr = delivery.get("tc_id_abbr")
    if declared_abbr is None:
        v.red.append("二 feature.yaml 未宣告 delivery.tc_id_abbr")
    elif abbrs and abbrs != {declared_abbr}:
        v.red.append(f"二 ABBR 實為 {sorted(abbrs)}，宣告為 {declared_abbr}")
    if seqs and seqs != sorted(seqs):
        v.info.append("二 tc_id 非依列序遞增（PARTIAL 本保留舊 ID 時屬允許，須附對照表）")

    # 3 Test Group
    groups = {str(cell(r, c["group"])).strip() for _, _, r in tc_rows if cell(r, c["group"]) not in (None, "")}
    expected_group = delivery.get("test_group") or cfg.get("feature")
    if len(groups) != 1:
        v.red.append(f"三 Test Group 非單一值：{sorted(groups)}")
    if expected_group is None:
        v.red.append("三 feature.yaml 未宣告 delivery.test_group（亦無頂層 feature）")
    elif groups and groups != {expected_group}:
        v.red.append(f"三 Test Group 實為 {sorted(groups)}，應為 `{expected_group}`（037 全名）")
    group = next(iter(groups)) if len(groups) == 1 else None

    # 4 fixed columns
    bad_author = sum(1 for _, _, r in tc_rows if str(cell(r, c["author"]) or "").strip() != AUTHOR)
    if bad_author:
        v.red.append(f"四 Author 非 {AUTHOR}：{bad_author} 列")
    bad_prio = sum(1 for _, _, r in tc_rows if str(cell(r, c["prio"]) or "").strip() not in PRIORITIES)
    if bad_prio:
        v.red.append(f"四 Priority 非 P0–P3：{bad_prio} 列")
    est = sum(1 for _, _, r in tc_rows if cell(r, c["est"]) not in (None, ""))
    if est:
        v.warn.append(f"四 Est. Time 非空：{est} 列（[DEFAULT] 留空）")

    # 5 filename + manifest
    m = FILENAME_RE.match(book.name)
    if not m:
        v.red.append("五 檔名不合客戶檔名形制（含尾綴或 sandbox 名）")
    elif group and m.group("name") != group.replace(" ", ""):
        v.red.append(f"五 檔名 {{FeatureName}} `{m.group('name')}` ≠ Test Group 去空白 `{group.replace(' ', '')}`")
    rec = manifest.get(book.name)
    if rec is None:
        v.red.append("五 MANIFEST.tsv 無此檔之列")
    elif rec.get("sha256") != sha256_of(book):
        v.red.append("五 MANIFEST.tsv sha256 與檔不符")

    # 6 contents
    note = book.parent / "DELIVERY_NOTE.md"
    if not note.exists():
        v.red.append("六 delivered/ 無 DELIVERY_NOTE.md")
    dr_files = [p for p in book.parent.iterdir()
                if p.is_file() and re.search(r"DR|PENDING", p.name, re.I)]
    note_has_dr = note.exists() and re.search(r"\bDR-", note.read_text(encoding="utf-8", errors="ignore")) is not None
    if not dr_files and not note_has_dr:
        v.red.append("六 未結 DR 清單缺（同目錄無 DR／PENDING 檔，DELIVERY_NOTE 亦無 DR- 段）")

    # 7 PENDING
    content_cols = [c[k] for k in ("item", "pre", "input", "proc", "er", "spec") if k in c]
    pend = sum(1 for _, _, r in tc_rows for j in content_cols
               if isinstance(cell(r, j), str) and "PENDING" in cell(r, j))
    if pend:
        note_txt = (rec or {}).get("note", "")
        if re.search(r"\bR-[A-Za-z]*\d+", note_txt):
            v.info.append(f"七 PENDING {pend} 格，MANIFEST note 載例外 R- 號")
        else:
            v.red.append(f"七 PENDING {pend} 格且 MANIFEST note 無 R- 例外號")
    return v


def load_baseline(path: Path) -> set[str]:
    if not path.exists():
        return set()
    return {r.split("\t")[0] for r in path.read_text(encoding="utf-8").splitlines()[1:] if r.strip()}


def delivered_books(root: Path) -> list[Path]:
    return sorted(p for p in root.glob("features/*/delivered/*.xlsx") if not p.name.startswith("~$"))


def main() -> int:
    ap = argparse.ArgumentParser(description="交付規格表檢查（R-G42，DELIVERY-SPEC）")
    ap.add_argument("--root", default=".")
    ap.add_argument("--baseline", default=BASELINE_DEFAULT)
    ap.add_argument("--emit-baseline", action="store_true",
                    help="以現況 delivered/ 內全部 xlsx 產出基線（生效日一次性；其後只減不增）")
    ap.add_argument("--gate", action="store_true", help="基線外之檔有紅時 exit 1")
    args = ap.parse_args()
    root = Path(args.root).resolve()
    books = delivered_books(root)

    if args.emit_baseline:
        out = root / args.baseline
        out.parent.mkdir(parents=True, exist_ok=True)
        lines = ["path\treason"] + [f"{b.relative_to(root)}\tR-G42 生效前已在 delivered/，不回歸" for b in books]
        out.write_text("\n".join(lines) + "\n", encoding="utf-8")
        print(f"寫入基線 {args.baseline}：{len(books)} 列")
        return 0

    baseline = load_baseline(root / args.baseline)
    fail = 0
    for book in books:
        try:
            v = check_book(root, book)
        except Exception as exc:  # unreadable workbook is itself a red
            v = Verdict(str(book.relative_to(root)), red=[f"讀取失敗：{exc}"])
        grandfathered = v.path in baseline
        tag = "基線（警示計數）" if grandfathered else "判紅"
        print(f"\n{v.path}  [{tag}]")
        for s in v.red:
            print(f"  {'警' if grandfathered else '紅'}  {s}")
        for s in v.warn:
            print(f"  警  {s}")
        for s in v.info:
            print(f"  註  {s}")
        if not grandfathered:
            fail += len(v.red)
    print(f"\n{'FAIL' if fail else 'PASS'}: 基線外判紅 {fail}（掃 {len(books)} 檔，基線 {len(baseline)} 列）")
    return 1 if (args.gate and fail) else 0


if __name__ == "__main__":
    sys.exit(main())
