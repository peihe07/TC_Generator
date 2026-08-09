#!/usr/bin/env python3
"""Compare two requirement reports that claim the same scope.

When a ruling supersedes one requirement report with another (AM/FM R1:
037-A03 replaces SWRA-A02), three questions decide what happens to the
existing test cases, and all three are answerable from the two files:

1. Do the families correspond at all? If the new report is a restructuring of
   the old one, the existing rows can be re-traced instead of discarded.
2. Which leaves of the new report have no ancestor? Those are new work.
3. Which leaves of the OLD report survive nowhere? Those are the ones a
   ruling silently drops, and the ones an assessor asks about.

The alignment axis is discovered, not assumed. AM/FM's 037-A03 carries the
SWRA-A02's *Description* text in its *Title* field — comparing title-to-title
finds almost nothing and would have supported the opposite conclusion. So
every field pairing is scored and the best-aligning one is reported with its
evidence.

Similarity is difflib on case- and punctuation-folded text. It is a screening
instrument: it establishes that a correspondence exists and bounds its size.
Every pair it proposes is a candidate for human confirmation, never a mapping.

Usage:
    python scripts/compare_req_families.py --new <report.xlsx> \\
        --old <report.xlsx> --out <dir> [--label-new A03 --label-old A02]

Exit 0 always — this reports, it does not gate.
"""
from __future__ import annotations

import argparse
import difflib
import json
import re
import sys
from pathlib import Path

import openpyxl

SHEET = "Analysis Report"
# Confidence bands. 0.85 is where near-verbatim restructuring (a sentence
# reflowed, a trailing clause dropped) stops and paraphrase begins; 0.60 is
# where "plausibly the same requirement" stops. Both are reporting
# thresholds, not rules — the bands are printed so a reader can move them.
STRONG, PLAUSIBLE = 0.85, 0.60
# Template help-text rows sit under the header in these workbooks and are not
# requirements; they are recognisable by the angle-bracket prose.
TEMPLATE_MARKERS = ("< mention", "<mention", "< provide", "<the ", "< the ")


def norm(s) -> str:
    """Fold case and punctuation so wording, not typography, is compared."""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", " ", str(s or "").lower())).strip()


def load_report(path: Path) -> tuple[list[str], list[tuple]]:
    """-> (normalised header texts, requirement rows)."""
    wb = openpyxl.load_workbook(path, read_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"{path.name}: no {SHEET!r} sheet")
    rows = list(wb[SHEET].iter_rows(values_only=True))
    wb.close()
    hdr = next((i for i, r in enumerate(rows)
                if any("requirement description" in norm(v) for v in r)), None)
    if hdr is None:
        sys.exit(f"{path.name}: no header row (no 'Requirement Description')")
    header = [norm(v) for v in rows[hdr]]
    data = [r for r in rows[hdr + 1:]
            if r[0] and not any(m in str(r[0]).lower() for m in TEMPLATE_MARKERS)]
    return header, data


def find_col(header: list[str], *need: str, forbid: tuple = ()) -> int | None:
    hits = [i for i, v in enumerate(header)
            if all(t in v for t in need) and not any(t in v for t in forbid)]
    return hits[0] if len(hits) == 1 else None


def field_texts(header: list[str], rows: list[tuple]) -> dict[str, list[str]]:
    """The two free-text fields worth aligning on, per row."""
    ti = find_col(header, "title")
    di = find_col(header, "requirement", "description")
    out = {}
    if ti is not None:
        out["title"] = [norm(r[ti]) for r in rows]
    if di is not None:
        out["description"] = [norm(r[di]) for r in rows]
    return out


def best_matches(src: list[str], dst: list[str], dst_ids: list[str]):
    """For each src text, its best dst match as (id, ratio). Empty src -> 0."""
    out = []
    for s in src:
        if not s:
            out.append((None, 0.0))
            continue
        best_i, best_r = None, 0.0
        for j, d in enumerate(dst):
            if not d:
                continue
            r = difflib.SequenceMatcher(None, s, d).ratio()
            if r > best_r:
                best_i, best_r = j, r
        out.append((dst_ids[best_i] if best_i is not None else None, best_r))
    return out


def score_axis(pairs) -> float:
    """An axis is judged by how many STRONG matches it finds, not by mean
    similarity — a uniformly mediocre axis is noise, a bimodal one is signal."""
    return sum(1 for _, r in pairs if r >= STRONG)


def run(args) -> int:
    new_path, old_path = Path(args.new), Path(args.old)
    hn, rn = load_report(new_path)
    ho, ro = load_report(old_path)
    ids_new = [str(r[0]).strip() for r in rn]
    ids_old = [str(r[0]).strip() for r in ro]
    fn, fo = field_texts(hn, rn), field_texts(ho, ro)
    if not fn or not fo:
        sys.exit("could not locate Title/Description columns in both reports")

    # --- discover the alignment axis
    axes = {}
    for a in fn:
        for b in fo:
            axes[(a, b)] = best_matches(fn[a], fo[b], ids_old)
    axis = max(axes, key=lambda k: score_axis(axes[k]))
    pairs = axes[axis]

    strong = [(i, m, r) for i, (m, r) in zip(ids_new, pairs) if r >= STRONG]
    plausible = [(i, m, r) for i, (m, r) in zip(ids_new, pairs)
                 if PLAUSIBLE <= r < STRONG]
    unmatched = [i for i, (_, r) in zip(ids_new, pairs) if r < PLAUSIBLE]
    consumed = {m for _, m, _ in strong}
    # An old row with only a PLAUSIBLE descendant is still represented, just
    # not verbatim. Counting it as dropped overstates what the ruling loses,
    # which is the number an assessor will read most closely.
    represented = consumed | {m for _, m, _ in plausible}
    orphaned_old = [i for i in ids_old if i not in represented]
    weakly_represented = [i for i in ids_old
                          if i in represented and i not in consumed]

    ln, lo = args.label_new, args.label_old
    lines = [
        f"# Requirement-family overlap — {ln} vs {lo}",
        "",
        f"- new report: `{new_path.name}` — {len(ids_new)} requirement rows",
        f"- old report: `{old_path.name}` — {len(ids_old)} requirement rows",
        f"- alignment axis (discovered): **{ln}.{axis[0]} ↔ {lo}.{axis[1]}**",
        "",
        "Axis scores (count of near-verbatim matches; the axis is discovered,",
        "not assumed — see the module docstring for why that matters here):",
        "",
        "| new field | old field | strong matches |",
        "|---|---|---|",
    ]
    for (a, b), p in sorted(axes.items(), key=lambda kv: -score_axis(kv[1])):
        mark = " ←" if (a, b) == axis else ""
        lines.append(f"| {a} | {b} | {score_axis(p)}{mark} |")
    lines += [
        "",
        "## Result",
        "",
        f"| | count | of |",
        "|---|---|---|",
        f"| {ln} leaves with a near-verbatim ancestor (≥{STRONG}) | "
        f"**{len(strong)}** | {len(ids_new)} |",
        f"| {ln} leaves with a plausible ancestor ({PLAUSIBLE}–{STRONG}) | "
        f"{len(plausible)} | {len(ids_new)} |",
        f"| {ln} leaves with no ancestor (<{PLAUSIBLE}) — new work | "
        f"**{len(unmatched)}** | {len(ids_new)} |",
        f"| {lo} rows that are a leaf's near-verbatim ancestor | "
        f"{len(consumed)} | {len(ids_old)} |",
        f"| {lo} rows represented only by a paraphrase | "
        f"{len(weakly_represented)} | {len(ids_old)} |",
        f"| {lo} rows represented nowhere — **dropped by the ruling** | "
        f"**{len(orphaned_old)}** | {len(ids_old)} |",
        "",
        f"Mapping shape: {len(strong)} strong matches consume {len(consumed)} "
        f"distinct {lo} rows — {'1:1' if len(strong) == len(consumed) else 'NOT 1:1'}.",
        "",
        f"## Candidate ancestry ({ln} → {lo}, ≥{PLAUSIBLE})",
        "",
        "Screening output. Every row is a candidate for human confirmation,",
        "never a mapping.",
        "",
        f"| {ln} | {lo} | similarity |",
        "|---|---|---|",
    ]
    for i, m, r in sorted(strong + plausible, key=lambda x: -x[2]):
        lines.append(f"| {i} | {m} | {r:.3f} |")
    lines += [
        "",
        f"## {lo} rows represented nowhere in {ln}",
        "",
        "These are what the ruling drops. If any is still a live requirement",
        "it needs a home, or the coverage claim is lost silently.",
        "",
        "".join(f"`{i}` " for i in orphaned_old) or "(none)",
        "",
    ]

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    (out / "family_overlap.md").write_text("\n".join(lines), encoding="utf-8")
    (out / "family_overlap.json").write_text(json.dumps({
        "new": new_path.name, "old": old_path.name,
        "axis": {"new_field": axis[0], "old_field": axis[1]},
        "thresholds": {"strong": STRONG, "plausible": PLAUSIBLE},
        "counts": {"new_rows": len(ids_new), "old_rows": len(ids_old),
                   "strong": len(strong), "plausible": len(plausible),
                   "unmatched_new": len(unmatched),
                   "weakly_represented_old": len(weakly_represented),
                   "orphaned_old": len(orphaned_old)},
        "weakly_represented_old": weakly_represented,
        "strong": [{"new": i, "old": m, "ratio": round(r, 4)} for i, m, r in strong],
        "plausible": [{"new": i, "old": m, "ratio": round(r, 4)}
                      for i, m, r in plausible],
        "unmatched_new": unmatched, "orphaned_old": orphaned_old,
    }, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"axis {axis[0]}<->{axis[1]}: {len(strong)} strong, "
          f"{len(plausible)} plausible, {len(unmatched)} new-only, "
          f"{len(orphaned_old)} old-only -> {out}", file=sys.stderr)
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--new", required=True, help="the ruled requirement source")
    ap.add_argument("--old", required=True, help="the superseded report")
    ap.add_argument("--out", required=True)
    ap.add_argument("--label-new", default="new")
    ap.add_argument("--label-old", default="old")
    return run(ap.parse_args())


if __name__ == "__main__":
    sys.exit(main())
