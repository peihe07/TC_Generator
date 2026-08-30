#!/usr/bin/env python3
"""Power 之列序四道 gate（74 包，Pei「Requirement or Design ID 要照順序排」）。

判準取自 **comfort 96 §6** 之三道（`features/comfort/scripts/verify_row_order_gates.py`）
＋ **ICS b35e651** 之第四道，逐道對 power 之交付工作簿跑：

    row-order-by-reqid   D 欄自上而下須為 037 之 leaf 序（`<` 為違規，`==` 為同 leaf 多條，合法）
    tc-id-sequence       同一 leaf 內 `tc_id` 須遞增
    all-leaves-present   037 之 leaf 每一個皆須出現至少一次（受 DR 阻斷者例外，逐一列名）
    blank-row-shape      留空列除 B（序號，A-PW102）與 D（req_id）外各欄皆須為空

⚠ **一道只會通過的檢查證明不了任何事**（comfort R-C41 之同一理由）——
本檔對前二道各注入一個壞序，斷言其轉紅；再以實檔斷言其轉綠。

用法：
    python scripts/verify_row_order_74.py [xlsx]
"""

from __future__ import annotations

import re
import sys
from itertools import groupby
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]
DEFAULT = ROOT / "features/power/delivered/pm_73.xlsx"
BLOCKED = {"SWE-PM-008": "DR-PW11", "SWE-PM-010": "DR-PW11", "SWE-PM-112": "DR-PW9"}


def leaf_key(s: str) -> int:
    return int(re.match(r"SWE-PM-(\d+)", s).group(1))


def g_row_order(col_d: list[str], first_row: int) -> str | None:
    for i in range(1, len(col_d)):
        if leaf_key(col_d[i]) < leaf_key(col_d[i - 1]):
            return f"row{first_row + i}: {col_d[i]} 在 {col_d[i - 1]} 之後"
    return None


def g_tc_sequence(pairs: list[tuple[str, str]]) -> str | None:
    for leaf, grp in groupby(pairs, key=lambda x: x[0]):
        ids = [t for _, t in grp if t]
        if ids != sorted(ids):
            return f"{leaf}: {[i.rsplit('-', 1)[-1] for i in ids]}"
    return None


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT
    cfg = yaml.safe_load((ROOT / "features/power/feature.yaml").read_text())
    col = cfg["workbook"]["columns"]
    ws = openpyxl.load_workbook(path, data_only=True)[cfg["workbook"]["sheet"]]
    rows = [r for r in range(10, ws.max_row + 1)
            if ws[f"{col['req_id']}{r}"].value]
    d = [str(ws[f"{col['req_id']}{r}"].value) for r in rows]
    pairs = [(d[i], str(ws[f"{col['tc_id']}{rows[i]}"].value or ""))
             for i in range(len(rows))]

    sys.path.insert(0, str(ROOT / "features/power/scripts"))
    import verify_gates as vg
    leaves = {x[0] for x in vg.leaves()}
    missing = sorted(leaves - set(d))

    blank = [r for r in rows if not ws[f"{col['tc_id']}{r}"].value]
    shape_bad = [r for r in blank
                 if [ws.cell(r, c).column_letter for c in range(2, 35)
                     if ws.cell(r, c).value not in (None, "")] != ["B", "D"]]

    ok1, ok2 = g_row_order(d, rows[0]), g_tc_sequence(pairs)
    print(f"檔：{path.relative_to(ROOT)}　資料列 {len(rows)}")
    print(f"1 row-order-by-reqid : {'PASS' if not ok1 else 'FAIL ' + ok1}")
    print(f"2 tc-id-sequence     : {'PASS' if not ok2 else 'FAIL ' + ok2}")
    unexpected = [m for m in missing if m not in BLOCKED]
    miss_txt = "、".join(f"{m}（{BLOCKED.get(m, '未預期')}）" for m in missing) or "無"
    verdict = "PASS（缺者皆為 DR 阻斷）" if not unexpected else "FAIL"
    print(f"3 all-leaves-present : {len(set(d))} / {len(leaves)}；缺 {miss_txt}　{verdict}")
    print(f"4 blank-row-shape    : {'PASS' if not shape_bad else 'FAIL ' + str(shape_bad)}"
          f"　留空列 {blank}")

    # ── 反向驗證：注入壞序，斷言轉紅 ──
    bad_d = d[:]
    # 須挑**相鄰而 leaf 號不同**之一對 —— 同 leaf 之相鄰列互換不構成逆序
    # （判準為 `<` 而非 `<=`，同 comfort 之定義），互換之則證明不了 gate 有效
    for i in range(1, len(bad_d)):
        if leaf_key(bad_d[i]) > leaf_key(bad_d[i - 1]):
            bad_d[i - 1], bad_d[i] = bad_d[i], bad_d[i - 1]
            break
    bad_p = pairs[:]
    for i, (leaf, tid) in enumerate(bad_p):
        if i and bad_p[i - 1][0] == leaf and tid:
            bad_p[i - 1], bad_p[i] = bad_p[i], bad_p[i - 1]
            break
    print("\n反向驗證（注入壞資料，須轉紅）：")
    print(f"  row-order-by-reqid : {'紅 ✓' if g_row_order(bad_d, rows[0]) else '**未轉紅 ✗**'}")
    print(f"  tc-id-sequence     : {'紅 ✓' if g_tc_sequence(bad_p) else '**未轉紅 ✗**'}")
    sys.exit(0 if not (ok1 or ok2 or unexpected or shape_bad) else 1)


if __name__ == "__main__":
    main()
