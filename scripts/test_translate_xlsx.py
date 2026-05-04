"""Tests for translate_xlsx — uses a stub translate_fn instead of OpenAI."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest

from translate_xlsx import (
    DEFAULT_COLUMNS,
    DEFAULT_START_ROW,
    XlsxTranslator,
    build_output_path,
    contains_cjk,
    parse_args,
    parse_columns,
)


# --------------------------------------------------------------------------- #
# Pure helpers                                                                #
# --------------------------------------------------------------------------- #


class TestContainsCjk:
    @pytest.mark.parametrize(
        "value, expected",
        [
            ("手機和車機藍芽處於可配對狀態", True),
            ("配对蓝牙成功后", True),  # Simplified Chinese
            ("Mixed 中文 and English", True),
            ("Pure English text", False),
            ("123 + 456 = 579", False),
            ("", False),
            (None, False),
            (12345, False),
            ("CarPlay", False),
        ],
    )
    def test_detects_cjk(self, value, expected):
        assert contains_cjk(value) is expected


class TestParseColumns:
    def test_basic(self):
        assert parse_columns(["A", "B", "Z"]) == [1, 2, 26]

    def test_handles_lowercase_and_whitespace(self):
        assert parse_columns([" j ", "k", "l", "m"]) == [10, 11, 12, 13]

    def test_skips_blank(self):
        assert parse_columns(["J", "", "  ", "K"]) == [10, 11]


class TestBuildOutputPath:
    def test_explicit_override_wins(self, tmp_path):
        src = tmp_path / "foo.xlsx"
        src.touch()
        override = tmp_path / "custom" / "out.xlsx"
        assert build_output_path(src, override) == override

    def test_appends_en_suffix_in_output_dir(self, tmp_path):
        # Layout: tmp_path/output (sibling of source)
        (tmp_path / "output").mkdir()
        src = tmp_path / "report.xlsx"
        src.touch()
        result = build_output_path(src, None)
        assert result == tmp_path / "output" / "report_EN.xlsx"


# --------------------------------------------------------------------------- #
# XlsxTranslator                                                              #
# --------------------------------------------------------------------------- #


class TestXlsxTranslator:
    def _build_workbook(self, sheet_name: str = "Test Case Specification 測試用例規範"):
        wb = openpyxl.Workbook()
        # Rename the default sheet to match production sheet name
        ws = wb.active
        ws.title = sheet_name
        # Header rows (rows 1-9) are intentionally empty / English-only
        ws["J9"] = "Pre-Conditions\nPre-Conditions"
        # Data rows starting at row 10
        ws["J10"] = "1.手機和車機藍芽處於可配對狀態\n2.手機支援CarPlay功能"
        ws["K10"] = ""  # empty
        ws["L10"] = "1.開啟車機藍芽，搜尋並選擇手機進行配對"
        ws["M10"] = "1.車機彈出CarPlay連線確認對話方塊"
        ws["J11"] = "Pure English — should be skipped"
        ws["L11"] = "1.手機和車機藍芽處於可配對狀態"  # duplicate of J10's first line
        return wb, ws

    def test_translates_only_cjk_cells_in_target_columns(self):
        wb, ws = self._build_workbook()
        translate_fn = MagicMock(side_effect=lambda s: f"[EN]{s[:20]}")
        translator = XlsxTranslator(
            translate_fn=translate_fn,
            columns=(10, 11, 12, 13),  # J K L M
            start_row=10,
        )
        translator.translate_workbook(wb)

        # J10, L10, M10, L11 contain CJK → translated
        # K10 empty, J11 English → untouched
        assert ws["J10"].value.startswith("[EN]")
        assert ws["K10"].value == ""
        assert ws["L10"].value.startswith("[EN]")
        assert ws["M10"].value.startswith("[EN]")
        assert ws["J11"].value == "Pure English — should be skipped"
        assert ws["L11"].value.startswith("[EN]")
        assert translator.stats.cells_translated == 4

    def test_caches_identical_strings(self):
        """The same Chinese string should only hit the API once."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Case Specification 測試用例規範"
        repeated = "1.手機和車機藍芽處於可配對狀態"
        ws["J10"] = repeated
        ws["L10"] = repeated
        ws["M10"] = repeated

        translate_fn = MagicMock(return_value="translated")
        translator = XlsxTranslator(
            translate_fn=translate_fn,
            columns=(10, 11, 12, 13),
            start_row=10,
        )
        translator.translate_workbook(wb)

        assert translate_fn.call_count == 1
        assert translator.stats.api_calls == 1
        assert translator.stats.cache_hits == 2
        assert translator.stats.cells_translated == 3

    def test_continues_on_failure(self):
        """If one cell errors, the translator logs and moves on."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Case Specification 測試用例規範"
        ws["J10"] = "失敗的儲存格"
        ws["L10"] = "成功的儲存格"

        def flaky(text):
            if "失敗" in text:
                raise RuntimeError("API down")
            return f"OK: {text}"

        translator = XlsxTranslator(
            translate_fn=flaky,
            columns=(10, 12),
            start_row=10,
        )
        translator.translate_workbook(wb)

        assert ws["J10"].value == "失敗的儲存格"  # unchanged on failure
        assert ws["L10"].value == "OK: 成功的儲存格"
        assert translator.stats.failures == 1
        assert translator.stats.cells_translated == 1

    def test_missing_sheet_raises(self):
        wb = openpyxl.Workbook()
        translator = XlsxTranslator(
            translate_fn=lambda x: x,
            columns=(10,),
            sheet_name="non-existent",
        )
        with pytest.raises(KeyError, match="not found"):
            translator.translate_workbook(wb)

    def test_skips_rows_before_start_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Case Specification 測試用例規範"
        ws["J5"] = "標頭中文不應翻譯"
        ws["J10"] = "資料中文應翻譯"

        translate_fn = MagicMock(return_value="translated")
        translator = XlsxTranslator(
            translate_fn=translate_fn,
            columns=(10,),
            start_row=10,
        )
        translator.translate_workbook(wb)

        assert ws["J5"].value == "標頭中文不應翻譯"
        assert ws["J10"].value == "translated"
        assert translate_fn.call_count == 1


# --------------------------------------------------------------------------- #
# CLI parsing                                                                 #
# --------------------------------------------------------------------------- #


class TestParseArgs:
    def test_defaults(self):
        args = parse_args(["foo.xlsx"])
        assert args.input == Path("foo.xlsx")
        assert args.columns == list(DEFAULT_COLUMNS)
        assert args.start_row == DEFAULT_START_ROW
        assert args.dry_run is False

    def test_overrides(self):
        args = parse_args(
            [
                "foo.xlsx",
                "--output",
                "out.xlsx",
                "--sheet",
                "MySheet",
                "--columns",
                "A",
                "B",
                "--start-row",
                "5",
                "--dry-run",
            ]
        )
        assert args.output == Path("out.xlsx")
        assert args.sheet == "MySheet"
        assert args.columns == ["A", "B"]
        assert args.start_row == 5
        assert args.dry_run is True
