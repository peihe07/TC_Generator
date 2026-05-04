"""Translate Chinese cells in JKLM columns of an STLA test case xlsx to English.

Usage:
    export OPENAI_API_KEY=sk-...
    python scripts/translate_xlsx.py <input.xlsx> [--output output/foo_EN.xlsx] \
        [--sheet "Test Case Specification 測試用例規範"] [--start-row 10] \
        [--columns J K L M] [--model gpt-4o-mini] [--dry-run]

Behavior:
    - Loads the workbook with openpyxl (preserves styles/merges).
    - Iterates the target sheet from --start-row, replacing any cell in the
      target columns whose value contains CJK characters with an English
      translation produced by OpenAI Chat Completions.
    - Identical strings are translated only once (in-process LRU cache).
    - The original file is left untouched; result is saved to output/<name>_EN.xlsx
      unless --output is provided.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Iterable

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Auto-load .env from the project root if python-dotenv is installed.
# This lets the script pick up OPENAI_API_KEY without an explicit `export`.
try:
    from dotenv import load_dotenv

    _project_root = Path(__file__).resolve().parent.parent
    load_dotenv(_project_root / ".env")
except ImportError:
    pass

# Match any CJK Unified Ideograph (covers Traditional & Simplified Chinese)
CJK_PATTERN = re.compile(r"[一-鿿㐀-䶿]")

DEFAULT_SHEET = "Test Case Specification 測試用例規範"
DEFAULT_COLUMNS = ("J", "K", "L", "M")
DEFAULT_START_ROW = 10
DEFAULT_MODEL = "gpt-4o-mini"

SYSTEM_PROMPT = (
    "You are a professional technical translator for automotive infotainment "
    "test case documents. Translate the user's text from Chinese (Traditional "
    "or Simplified) to clear, concise English suitable for a software test "
    "specification. Rules: (1) Preserve the original line breaks and numbered "
    "list structure exactly. (2) Keep English/product names, acronyms, IDs, "
    "and codes as-is (e.g. CarPlay, Android Auto, HU, HMI, Bluetooth, AA, "
    "Polarion, TestRail). (3) Do not add explanations, prefaces, or quotes. "
    "(4) Output ONLY the translated text."
)

logger = logging.getLogger("translate_xlsx")


# --------------------------------------------------------------------------- #
# Pure helpers (easy to unit-test)                                            #
# --------------------------------------------------------------------------- #


def contains_cjk(text: object) -> bool:
    """Return True iff text is a string containing any CJK character."""
    return isinstance(text, str) and bool(CJK_PATTERN.search(text))


def parse_columns(values: Iterable[str]) -> list[int]:
    """Convert column letters (e.g. 'J', 'K') to 1-based indices."""
    return [column_index_from_string(v.strip().upper()) for v in values if v.strip()]


def build_output_path(input_path: Path, override: Path | None = None) -> Path:
    """Default to <repo>/output/<stem>_EN.xlsx unless override is provided."""
    if override is not None:
        return override
    project_root = input_path.resolve().parents[0]
    # Walk up looking for an existing 'output' folder; fall back to CWD/output
    for parent in [input_path.resolve()] + list(input_path.resolve().parents):
        candidate = parent / "output"
        if candidate.is_dir():
            project_root = candidate
            break
    else:
        project_root = Path.cwd() / "output"
    project_root.mkdir(parents=True, exist_ok=True)
    return project_root / f"{input_path.stem}_EN.xlsx"


# --------------------------------------------------------------------------- #
# Translator                                                                  #
# --------------------------------------------------------------------------- #


@dataclass
class TranslationStats:
    cells_scanned: int = 0
    cells_translated: int = 0
    cache_hits: int = 0
    api_calls: int = 0
    failures: int = 0


@dataclass
class XlsxTranslator:
    """Translate Chinese cells in an xlsx file to English.

    `translate_fn` accepts a string and returns its English translation.
    Inject a fake function in tests to avoid real API calls.
    """

    translate_fn: Callable[[str], str]
    columns: tuple[int, ...]
    start_row: int = DEFAULT_START_ROW
    sheet_name: str = DEFAULT_SHEET
    cache: dict[str, str] = field(default_factory=dict)
    stats: TranslationStats = field(default_factory=TranslationStats)

    def translate_text(self, text: str) -> str:
        """Translate with in-process cache."""
        if text in self.cache:
            self.stats.cache_hits += 1
            return self.cache[text]
        try:
            result = self.translate_fn(text)
        except Exception:
            self.stats.failures += 1
            raise
        self.stats.api_calls += 1
        self.cache[text] = result
        return result

    def translate_sheet(self, ws: Worksheet) -> None:
        """Translate all CJK cells in the configured columns."""
        last_row = ws.max_row
        for row in range(self.start_row, last_row + 1):
            for col in self.columns:
                cell = ws.cell(row=row, column=col)
                value = cell.value
                self.stats.cells_scanned += 1
                if not contains_cjk(value):
                    continue
                # Skip merged-cell continuations (they have value=None usually)
                try:
                    translated = self.translate_text(value)
                except Exception as exc:
                    logger.error(
                        "Translation failed at %s%d: %s", get_column_letter(col), row, exc
                    )
                    continue
                cell.value = translated
                self.stats.cells_translated += 1
                logger.info(
                    "Translated %s%d (%d chars)",
                    get_column_letter(col),
                    row,
                    len(value),
                )

    def translate_workbook(self, wb: Workbook) -> None:
        if self.sheet_name not in wb.sheetnames:
            raise KeyError(
                f"Sheet '{self.sheet_name}' not found. Available: {wb.sheetnames}"
            )
        self.translate_sheet(wb[self.sheet_name])


# --------------------------------------------------------------------------- #
# OpenAI translator factory                                                   #
# --------------------------------------------------------------------------- #


def make_openai_translate_fn(
    model: str = DEFAULT_MODEL,
    api_key: str | None = None,
    max_retries: int = 3,
    retry_backoff: float = 2.0,
) -> Callable[[str], str]:
    """Return a callable that translates a single string via OpenAI."""
    from openai import OpenAI  # imported lazily so tests don't require the package

    client = OpenAI(api_key=api_key or os.environ.get("OPENAI_API_KEY"))

    def _translate(text: str) -> str:
        last_exc: Exception | None = None
        for attempt in range(1, max_retries + 1):
            try:
                resp = client.chat.completions.create(
                    model=model,
                    temperature=0,
                    messages=[
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": text},
                    ],
                )
                content = resp.choices[0].message.content or ""
                return content.strip()
            except Exception as exc:  # noqa: BLE001 — retry-all-then-raise pattern
                last_exc = exc
                wait = retry_backoff ** attempt
                logger.warning(
                    "OpenAI call failed (attempt %d/%d): %s — retrying in %.1fs",
                    attempt,
                    max_retries,
                    exc,
                    wait,
                )
                time.sleep(wait)
        assert last_exc is not None
        raise last_exc

    return _translate


# --------------------------------------------------------------------------- #
# CLI entrypoint                                                              #
# --------------------------------------------------------------------------- #


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("input", type=Path, help="Path to the source xlsx file")
    parser.add_argument(
        "--output", type=Path, default=None, help="Output xlsx path (default: output/<name>_EN.xlsx)"
    )
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Worksheet name to translate")
    parser.add_argument(
        "--columns",
        nargs="+",
        default=list(DEFAULT_COLUMNS),
        help="Column letters to translate (default: J K L M)",
    )
    parser.add_argument(
        "--start-row", type=int, default=DEFAULT_START_ROW, help="First row of data (default: 10)"
    )
    parser.add_argument("--model", default=DEFAULT_MODEL, help="OpenAI model name")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Scan and report cells that WOULD be translated; no API calls, no save.",
    )
    parser.add_argument("--verbose", "-v", action="store_true")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    if not args.input.is_file():
        logger.error("Input file not found: %s", args.input)
        return 2

    columns = tuple(parse_columns(args.columns))
    output_path = build_output_path(args.input, args.output)

    if args.dry_run:
        translate_fn: Callable[[str], str] = lambda t: t  # noqa: E731 — simple sentinel
    else:
        if not os.environ.get("OPENAI_API_KEY"):
            logger.error("OPENAI_API_KEY environment variable is required (or use --dry-run)")
            return 2
        translate_fn = make_openai_translate_fn(model=args.model)

    logger.info("Loading workbook: %s", args.input)
    wb = openpyxl.load_workbook(args.input)

    translator = XlsxTranslator(
        translate_fn=translate_fn,
        columns=columns,
        start_row=args.start_row,
        sheet_name=args.sheet,
    )
    translator.translate_workbook(wb)

    logger.info(
        "Done. scanned=%d translated=%d cache_hits=%d api_calls=%d failures=%d",
        translator.stats.cells_scanned,
        translator.stats.cells_translated,
        translator.stats.cache_hits,
        translator.stats.api_calls,
        translator.stats.failures,
    )

    if args.dry_run:
        logger.info("Dry run — workbook NOT saved.")
        return 0

    wb.save(output_path)
    logger.info("Saved: %s", output_path)
    print(json.dumps({"output": str(output_path), **translator.stats.__dict__}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
