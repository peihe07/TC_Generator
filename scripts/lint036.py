#!/usr/bin/env python3
"""FM-WI-FSM-036 工作簿靜態檢查（報告模式）。

檢查 A–N 之定義見 docs/fw036/handoff/00_lint_spec.md。
本工具唯讀開啟 xlsx，絕不寫回任何 xlsx。

`--profile <feature>` （21 包）：指定時 P 改採 **R-1 v3** 判準
（`$MESSAGE.Signal$` ＋ DBC `VAL_` 標籤），並另跑
Q（不可見字元，R-10(a)）／R（Pre-Condition 版面，R-9(a)）／
T（PENDING 說明語言，R-14）。**未指定時行為與 21 包之前完全一致** ——
既有八本之報告基線因而不動（迴歸基準見上繳 22）。

gate 政策（S3）：`--gate` 旗標保留但**尚不啟用**。啟用時機為尾批
（全數回修完成後）；現階段啟用將使所有既有交付本 exit 1，阻斷正常
作業。裁決條文見 docs/fw036/RULINGS_LEDGER.md。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from dataclasses import dataclass, field as dc_field
from datetime import date
from pathlib import Path

import openpyxl

# --- 欄位鍵與標頭關鍵字（startswith 匹配） ---------------------------------

FIELD_HEADERS: dict[str, str] = {
    "test_set": "Test Set",
    "test_item": "Test Item",
    "pre": "Pre-Conditions",
    "input": "Input Test Data",
    "proc": "Test procedure",
    "er": "Expected Result",
    "spec": "Specification Reference",
    "author": "Test Case Author",
}

# 輔助欄（非檢查對象，供報告定位與 I-sibling 分組用）
REQ_ID_HEADER = "Requirement or Design ID"
TC_ID_FIRSTLINE = "Test Case ID"

HEADER_ANCHOR = "Specification Reference"
HEADER_SCAN_ROWS = 15
TC_SHEET_PREFIX = "Test Case Specification"

# --- 各檢查所涵蓋之欄位 ------------------------------------------------------

# K「六欄」（00b 修訂 3 明確化）：不含 spec、author、remarks
K_FIELDS = ("test_item", "test_set", "pre", "input", "proc", "er")
M_FIELDS = ("pre", "proc", "er", "spec")
N_FIELDS = ("pre", "input", "proc", "er")
# P（R-1／R-6）：僅施於作者生成之內容 —— 四欄 ＋ test_item 之括號下半。
# test_item 上半為需求原句 verbatim，其訊號記法保留來源原文，不套 R-1。
P_FIELDS = ("pre", "input", "proc", "er")
J_NUMBERED_FIELDS = ("pre", "proc", "er")
# Q（R-10(a)）：不可見字元不構成內容，全欄位適用（含 verbatim 上半與 spec）
Q_FIELDS = ("test_item", "test_set", "pre", "input", "proc", "er", "spec")
# T（R-14）：`PENDING:` 佔位之說明須為英文
T_FIELDS = ("pre", "input", "proc", "er")

# --- 正則 --------------------------------------------------------------------

NUMBERED_LINE = re.compile(r"^\s*\d+[.)]")          # 全域編號行定義（E 等沿用，勿動）
NUMBER_PREFIX = re.compile(r"^\s*\d+[.)]\s*")
# N 檢查自身之行定義：另納 a./b./c. 縮排子步驟（canon §6.1 子層為實質測試步驟）
# 限定僅供 n_exempt() 使用，不得外溢至 NUMBERED_LINE 之使用點
N_STEP_LINE = re.compile(r"^\s*(\d+|[a-z])[.)]")

RE_A = re.compile(
    r"(^\s*\d+[.)]\s*(Observe|Verify|See if|Watch|Monitor|Inspect)\b)"
    r"|\b(observe whether|check whether|confirm whether|see if)\b",
    re.I | re.M,
)
RE_B = re.compile(r"\b(shall|should|will)\b")
RE_C = re.compile(r"\b(properly|successfully|within reasonable time)\b", re.I)
RE_D_POWERED = re.compile(r"\b(HU|system|unit) is powered on\b", re.I)
RE_D_VERB = re.compile(
    r"^(Insert|Connect|Press|Open|Enable|Disable|Launch|Select|Tap|Trigger|Perform|Set)\b",
    re.I,
)
RE_F = re.compile(r"\[[A-Za-z][^\]]{0,30}\]")
# `--profile` 專屬例外（下放包 43 §二 #1）：緊接於 `$<name>$ =` 之後之方括號
# **不是未填佔位，是車輛屬性之值**（037 逐字之訊號記法 `$FOTA_Status$ = [值]`）。
# 未指定 `--profile` 時本例外不生效 —— 既有八本之基線因而完全不動。
RE_F_SIGNAL_VALUE = re.compile(r"\$[A-Za-z][A-Za-z0-9_.]*\$\s*=\s*(\[[^\]]{0,60}\])")
RE_H = re.compile(r"\b(as expected|works? normally|normal(ly)? operation)\b", re.I)
# `--profile` 專屬（下放包 47 §二 #6）：**關係模糊詞** ——
# 其與上式之**程度／狀態模糊詞**不同族，現行詞表**整族未收**。
# 程度模糊詞（`properly`）一望即知其模糊；**關係模糊詞看起來很具體** ——
# 它指名了二個被比較的量，只是沒說「相符」到什麼程度算相符。
RE_H_RELATION = re.compile(
    r"\b(corresponds? to|matches?|is consistent with|in line with|aligns? with)\b", re.I)
RE_PAREN_LINE = re.compile(r"^\(.+\)$")
RE_PAREN_TAIL = re.compile(r"\([^)]{3,}\)\s*$")
RE_CJK = re.compile(r"[一-鿿]")
RE_TOKEN = re.compile(r"[A-Za-z0-9$_.'\"-]+")
RE_TRAILING_PERIOD = re.compile(r"[.。]$")
# CAN 訊號 token `MESSAGE.Signal`（message 段全大寫）。
# 內部訊號 `TLM_Status.Info`／`Phone_Call.Info` 之 message 段含小寫，不命中。
RE_CAN_TOKEN = re.compile(r"\b[A-Z][A-Z0-9_]{2,}\.[A-Za-z][A-Za-z0-9_]*\b")
# R-1 v1 之三件組，已撤銷；殘留即違規。
RE_P_TRIPLET = re.compile(
    r"\b[A-Za-z0-9_]+\s+in\s+[A-Z][A-Z0-9_]{2,}\s+on\s+[A-Za-z0-9-]+\b")
# 賦值之偵測：CAN token 後接 `=`／`from`／`to`（不論是否合式）。
RE_P_ASSIGNMENT = re.compile(
    r"\b[A-Z][A-Z0-9_]{2,}\.[A-Za-z][A-Za-z0-9_]*\s*(?:=|from\b|to\b)")
# R-1 v2(a)(b) 共通之賦值形：`<MSG>.<Sig> = <raw> (<label>)`。
# 括號標籤即 R-7 之 DBC `VAL_` 語意標籤。收尾語不設限（見 check_signal_line）。
RE_P_VALUE_FORM = re.compile(
    r"[A-Z][A-Z0-9_]{2,}\.[A-Za-z][A-Za-z0-9_]*\s*=\s*[^\s(]+\s*\([^)]+\)")
# PROXI 行（R-1 v2(c)）
RE_P_PROXI = re.compile(r"\bPROXI\s+(\$[^$]+\$|[A-Za-z][A-Za-z0-9_]*)\s*=")
# --- profile 專屬（未指定 --profile 時全部不啟用）-----------------------------
# P v3（R-1 v3）：訊號一律 `$<MSG>.<Sig>$`；賦值須帶 `(<VAL_ label>)`
RE_P3_DOLLAR_ASSIGN = re.compile(
    r"\$[A-Z][A-Z0-9_]{2,}\.[A-Za-z][A-Za-z0-9_]*\$\s*=\s*(?P<val>[^\s(]+)\s*(?P<lab>\([^)]+\))?")
# `$` 未包覆之 CAN token 賦值（v3(a) 要求包覆）
RE_P3_BARE_ASSIGN = re.compile(
    r"(?<!\$)\b[A-Z][A-Z0-9_]{2,}\.[A-Za-z][A-Za-z0-9_]*\s*=")
RE_P3_SEND_CAN = re.compile(r"\bSend CAN:")
RE_P3_PROXI_DOLLAR = re.compile(r"\bPROXI\s+\$")
# 下放包 43 §二 #1：無點之車輛屬性記法 `$<Name>$ = [值]`（037 逐字）。
# **P v3 之二式皆要求訊號名含一個點，故本形態原本不被任何式命中** ——
# `P=0` 是沉默不是核可（上繳包 37 §2.2）。本式使其**被檢查**：
# `$` 包覆之名、`=`、方括號包覆之值，三者齊備即通過；缺一即報。
RE_P3_PROP_OK = re.compile(r"\$[A-Za-z][A-Za-z0-9_]*\$\s*=\s*\[[^\]]{1,60}\]")
# 候選：`$` 包覆之無點名後接 `=`，其值形態不拘 —— 用以偵測「有名有等號而值不合式」
RE_P3_PROP_ANY = re.compile(r"\$([A-Za-z][A-Za-z0-9_]*)\$\s*=\s*(?P<val>\S+)")
# Q（R-10(a)）
RE_Q_TRAILING_WS = re.compile(r"[ \t]+$")
# V 行首空白（IN §11）。**行尾空白不在本檢查**——其已由 Q 覆蓋，
# 兩處同時計數會使量化矩陣之命中數雙倍膨脹（G-D：數字須可解釋）。
RE_V_LEADING_WS = re.compile(r"^[ \t]+")
RE_V_BLANK_WS = re.compile(r"^\s+$")
# IN §11 之唯二例外：§6.1 子層記法（`a./b./c.` 縮排 3 格、`-` 子彈 6 格）
# 與 §5.4 之 `$` 命令行（縮排 3 格）。
RE_V_EXEMPT = re.compile(r"^(?: {3}(?:[a-z]\.\s|\$ )| {6}- )")
# R（R-9(a)）：多條件並列之謂詞計數
RE_R_PREDICATE = re.compile(r"\b(is|are|was|were|reads?|holds?|has|have)\b")
R_TOOL_PHRASE = "tool is available on HU"
# T（R-14）
RE_T_PENDING = re.compile(r"PENDING:\s*(?P<dr>[A-Za-z0-9-]+)?(?P<desc>.*)$")

RE_CAMEL = re.compile(r"^[a-z][a-zA-Z0-9_]*[A-Z]")
RE_DOTCALL = re.compile(r"^[a-z][a-z0-9_]*\.[a-zA-Z(]")

J_WHITELIST = {
    "adb", "tmpfs", "iPod", "iOS", "iPhone", "dd", "cat", "mount",
    "btsnoop", "hciconfig", "hcitool", "logcat", "sdptool",
}

DEFAULT_LENGTH_LIMIT = 50

CHECK_TITLES = {
    "A": "禁用動詞 (proc)",
    "B": "ER 情態詞 (er)",
    "C": "hedge (test_item 括號下半)",
    "D": "PC 違規 (pre)",
    "E": "proc/er 編號行數不對齊",
    "F": "方括號佔位 (proc)",
    "G": "Test Set 空值",
    "H": "ER 模糊語 (er)",
    "I": "test_item 括號下半缺失",
    "I-sibling": "同 Requirement ID 括號行逐字重複",
    "J": "行首大寫",
    "K": "CJK 字元",
    "L": f"test_item 上半過長 (>{DEFAULT_LENGTH_LIMIT} tokens)",
    "M": "空欄三態",
    "N": "行尾多餘句號",
    "P": "訊號寫法不合 R-1 v2",
    "Q": "不可見字元（NBSP／全形空格／行尾空白）",
    "R": "Pre-Condition 版面（未編號行／多條件並列）",
    "T": "PENDING 說明非英文",
    "U": "PENDING 佔位（四欄全掃，含 ER 側）",
    "V": "行首空白（IN §11）",
    "I-cross": "跨 req_id：觀測窗相同且違例類有交集（R-SU34 v3）",
    "W": "ER 含比較關係而 test_item 上半無數值（下放包 47 §二 #6）",
    "X": "導航路徑無固定入口（§5.8／R-G71）",
}
# `--profile` 啟用時 P 改以 R-1 v3 判準，標題隨之替換
CHECK_TITLE_PROFILE = {"P": "訊號寫法不合 R-1 v3"}

# 校準狀態（00c 最終版）：M、J 經全語料分佈補校，改標已校準
CHECK_STATUS = {
    "A": "已校準", "B": "已校準", "C": "已校準（R-6b 範圍：Media 錨值 1→0）",
    "D": "已校準",
    "E": "已校準", "F": "已校準", "G": "已校準（詞彙表外值待接入）",
    "H": "已校準", "I": "已校準", "I-sibling": "未校準（M15）",
    "J": "已校準（行計口徑）", "K": "已校準（分級待 R-5）",
    "L": "已校準（閾值待 R-3）", "M": "已校準", "N": "已校準",
    "P": "已校準（SWC 0708：195 —— proc 11／er 184，見上繳 09）",
    "Q": "未校準（R-10(a)，21 包新增）",
    "R": "未校準（R-9(a)，21 包新增）",
    "T": "未校準（R-14，21 包新增）",
    "U": "計數用（A-PM16：ER 側原不受任何檢查覆蓋）",
    "V": "未校準（IN §11，27 包新增）",
    "I-cross": "警示器非判準（R-SU34 v3(c)）—— 命中一律送人裁，不自動判 FAIL",
    "W": "**待人裁非 FAIL** —— 輸出分二段（下放包 48 §二）：(a) 已裁段只報列數、(b) 新命中段逐列陳述",
    "X": "未校準（§5.8／R-G71，GC-07 新增）—— **WARN 只報不改**",
}
CHECK_STATUS_PROFILE = {"P": "未校準（R-1 v3，21 包改寫；profile 專屬）"}
CHECK_ORDER = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "I-sibling",
               "J", "K", "L", "M", "N", "P"]
# profile 專屬檢查：僅於 `--profile <feature>` 指定時啟用。
# 未指定時 CHECK_ORDER 不變 —— 既有八本之報告基線因而完全不動。
PROFILE_CHECKS = ["Q", "R", "T", "U", "V", "I-cross", "W", "X"]


def check_order(profile: str | None) -> list[str]:
    """本次執行所啟用之檢查序列。"""
    return CHECK_ORDER + PROFILE_CHECKS if profile else list(CHECK_ORDER)


def check_title(key: str, profile: str | None) -> str:
    if profile and key in CHECK_TITLE_PROFILE:
        return CHECK_TITLE_PROFILE[key]
    return CHECK_TITLES[key]


def check_status(key: str, profile: str | None) -> str:
    if profile and key in CHECK_STATUS_PROFILE:
        return CHECK_STATUS_PROFILE[key]
    return CHECK_STATUS[key]

# 各檢查之記錄粒度（報告表頭「行計」欄之語意）
CHECK_GRANULARITY = {
    "A": "每次命中", "B": "每次命中", "C": "每次命中",
    "D": "每次命中／每編號行", "E": "每列", "F": "每次命中",
    "G": "每列", "H": "每次命中", "I": "每列", "I-sibling": "每列",
    "J": "每行", "K": "每列每欄", "L": "每列", "M": "每列每欄", "N": "每行",
    "P": "每次命中",
    "Q": "每行每欄", "R": "每行", "T": "每次命中", "U": "每次命中",
    "V": "每行每欄",
    "I-cross": "每列每配對（一組命中記二列）",
    "W": "每次命中",
    "X": "每行",
}


# --- 資料結構 ----------------------------------------------------------------


@dataclass
class Violation:
    """單筆違規記錄。"""

    check: str
    row: int
    tc_id: str
    field: str
    detail: str
    snippet: str = ""

    def as_dict(self) -> dict:
        return {
            "check": self.check,
            "row": self.row,
            "tc_id": self.tc_id,
            "field": self.field,
            "detail": self.detail,
            "snippet": self.snippet,
        }


@dataclass
class SheetResult:
    """單一 TC sheet 的檢查結果。"""

    sheet: str
    header_row: int
    data_rows: int
    violations: list[Violation] = dc_field(default_factory=list)
    # I-cross 之原料（`--merge` 用）：(列, TC id, req id, Test Set, proc, er)
    cross_rows: list[tuple] = dc_field(default_factory=list)


# --- 工具函式 ----------------------------------------------------------------


def cell_text(value) -> str:
    """儲存格轉純文字；None 視為空字串。"""
    if value is None:
        return ""
    return str(value)


def split_lines(text: str) -> list[str]:
    """欄內以 \\n 切分為行。"""
    return text.split("\n")


def numbered_lines(text: str) -> list[str]:
    """取出編號行。"""
    return [ln for ln in split_lines(text) if NUMBERED_LINE.match(ln)]


def find_header_row(ws) -> int:
    """掃前 15 列，找含 Specification Reference 之列作為 header。"""
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS,
                                           values_only=True), start=1):
        for value in row:
            if value is not None and HEADER_ANCHOR in str(value):
                return idx
    raise ValueError(f"sheet {ws.title!r} 前 {HEADER_SCAN_ROWS} 列找不到 "
                     f"{HEADER_ANCHOR!r} 標頭")


def build_column_map(header_values: list) -> dict[str, int]:
    """依 startswith 建立 欄位鍵 -> 0-based 欄索引 對照。"""
    columns: dict[str, int] = {}
    for idx, value in enumerate(header_values):
        if value is None:
            continue
        text = str(value).strip()
        first_line = text.split("\n", 1)[0].strip()
        for key, keyword in FIELD_HEADERS.items():
            if key not in columns and text.startswith(keyword):
                columns[key] = idx
        if "req_id" not in columns and text.startswith(REQ_ID_HEADER):
            columns["req_id"] = idx
        if "tc_id" not in columns and first_line == TC_ID_FIRSTLINE:
            columns["tc_id"] = idx
    return columns


def quoted_spans(text: str) -> list[tuple[int, int]]:
    """回傳成對引號（" " 與 “ ”）涵蓋之區間，供 B 豁免用。"""
    spans: list[tuple[int, int]] = []
    positions = [i for i, ch in enumerate(text) if ch == '"']
    for i in range(0, len(positions) - 1, 2):
        spans.append((positions[i], positions[i + 1]))
    start = None
    for i, ch in enumerate(text):
        if ch == "“":
            start = i
        elif ch == "”" and start is not None:
            spans.append((start, i))
            start = None
    return spans


def inside_spans(span: tuple[int, int], spans: list[tuple[int, int]]) -> bool:
    """判斷 match 區間是否完整落於任一引號區間內。"""
    return any(lo < span[0] and span[1] <= hi for lo, hi in spans)


def snippet_of(text: str, start: int = 0, width: int = 80) -> str:
    """取違規上下文片段，換行改為 ⏎ 以利單行呈現。"""
    lo = max(0, start - 20)
    return text[lo:lo + width].replace("\n", " ⏎ ").strip()


def upper_half(test_item: str) -> str:
    """test_item 上半：去除整行為 (…) 之括號行。"""
    kept = [ln for ln in split_lines(test_item)
            if not RE_PAREN_LINE.match(ln.strip())]
    return "\n".join(kept)


def paren_lines(test_item: str) -> list[str]:
    """test_item 中整行為 (…) 之括號行。"""
    return [ln.strip() for ln in split_lines(test_item)
            if RE_PAREN_LINE.match(ln.strip())]


def first_token(text: str) -> str | None:
    """取第一個 token（00b 修訂 2：不再跳過非字母 token）。"""
    tokens = text.split()
    return tokens[0] if tokens else None


def j_violating_token(line_body: str) -> str | None:
    """J 判定：回傳違規 token，合規或豁免則回傳 None。

    第一個 token 若非以字母開頭（數字、$、引號、符號），該行豁免。
    """
    token = first_token(line_body)
    if token is None:
        return None
    if not token[0].isalpha():
        return None
    if not token[0].islower():
        return None
    if j_exempt(token):
        return None
    return token


def j_exempt(token: str) -> bool:
    """J 檢查之豁免判斷。"""
    bare = token.strip(".,;:)!?")
    if token in J_WHITELIST or bare in J_WHITELIST:
        return True
    if RE_CAMEL.match(token) or RE_CAMEL.match(bare):
        return True
    if RE_DOTCALL.match(token) or RE_DOTCALL.match(bare):
        return True
    if token[:1] in ('$', '"', "'", "“"):
        return True
    return False


def n_exempt(line: str) -> bool:
    """N 檢查之豁免：空行、$ 指令行、縮排續行。"""
    if not line.strip():
        return True
    if line.strip().startswith("$"):
        return True
    if line[:1] in (" ", "\t") and not N_STEP_LINE.match(line):
        return True
    return False


# --- 逐列檢查 ----------------------------------------------------------------


def check_row(fields: dict[str, str], row_no: int, tc_id: str,
              length_limit: int, profile: str | None = None) -> list[Violation]:
    """對單列跑 A–N（除 I-sibling 外）之檢查。

    `profile` 為 None 時行為與 21 包之前完全一致；指定時另跑 Q／R／T，
    且 P 改以 R-1 v3 判準（見 `check_signal_line_v3`）。
    """
    out: list[Violation] = []
    proc = fields["proc"]
    er = fields["er"]
    pre = fields["pre"]
    item = fields["test_item"]

    def add(check: str, field_key: str, detail: str, snippet: str = "") -> None:
        out.append(Violation(check, row_no, tc_id, field_key, detail, snippet))

    # A 禁用動詞
    for m in RE_A.finditer(proc):
        add("A", "proc", f"禁用動詞 {m.group(0).strip()!r}",
            snippet_of(proc, m.start()))

    # B ER 情態詞（引號內豁免）
    spans = quoted_spans(er)
    for m in RE_B.finditer(er):
        if inside_spans((m.start(), m.end()), spans):
            continue
        add("B", "er", f"情態詞 {m.group(0)!r}", snippet_of(er, m.start()))

    # C hedge（範圍依 R-6b：僅括號下半。上半為需求原句 verbatim，
    # 其用語屬來源文件而非作者所擇，不受「作者用語品質」類檢查規制）
    for line in paren_lines(item):
        for m in RE_C.finditer(line):
            add("C", "test_item(括號下半)", f"hedge {m.group(0)!r}", line[:80])

    # D PC 違規
    for m in RE_D_POWERED.finditer(pre):
        add("D", "pre", f"通電前提 {m.group(0)!r}", snippet_of(pre, m.start()))
    for line in split_lines(pre):
        if not NUMBERED_LINE.match(line):
            continue
        body = NUMBER_PREFIX.sub("", line)
        m = RE_D_VERB.match(body)
        if m:
            add("D", "pre", f"編號行行首動詞 {m.group(0)!r}", line.strip()[:80])

    # E 對齊
    n_proc, n_er = len(numbered_lines(proc)), len(numbered_lines(er))
    if n_proc > 0 and n_er > 0 and n_proc != n_er:
        add("E", "proc/er", f"proc {n_proc} 步 vs er {n_er} 步", "")

    # W：ER 有比較而上半無數值（profile 專屬）
    if profile:
        _up = split_lines(fields["test_item"])
        if _up and not RE_W_NUMERAL.search(_up[0]):
            for m in RE_W_COMPARE.finditer(er):
                add("W", "er", f"比較關係 {m.group(0)!r}，而 test_item 上半無數值",
                    snippet_of(er, m.start()))

    # F 方括號 —— profile 啟用時，`$<name>$ = [值]` 之值不判（下放包 43 §二 #1）
    exempt = {m.span(1) for m in RE_F_SIGNAL_VALUE.finditer(proc)} if profile else set()
    for m in RE_F.finditer(proc):
        if m.span() in exempt:
            continue
        add("F", "proc", f"方括號佔位 {m.group(0)!r}", snippet_of(proc, m.start()))

    # G Test Set 空值
    if not fields["test_set"].strip():
        add("G", "test_set", "Test Set 為空", "")

    # H ER 模糊
    for m in RE_H.finditer(er):
        add("H", "er", f"模糊語 {m.group(0)!r}", snippet_of(er, m.start()))
    if profile:
        for m in RE_H_RELATION.finditer(er):
            add("H", "er", f"關係模糊語 {m.group(0)!r}", snippet_of(er, m.start()))

    # I 括號下半（缺括號）
    if item.strip():
        has_paren_line = bool(paren_lines(item))
        has_paren_tail = bool(RE_PAREN_TAIL.search(item.strip()))
        if not has_paren_line and not has_paren_tail:
            add("I", "test_item", "缺括號下半", snippet_of(item))

    # J 行首大寫
    first_line = split_lines(item)[0] if item else ""
    token = j_violating_token(NUMBER_PREFIX.sub("", first_line))
    if token:
        add("J", "test_item", f"首字小寫 {token!r}", first_line.strip()[:80])
    for key in J_NUMBERED_FIELDS:
        for line in split_lines(fields[key]):
            if not NUMBERED_LINE.match(line):
                continue
            token = j_violating_token(NUMBER_PREFIX.sub("", line))
            if token:
                add("J", key, f"首字小寫 {token!r}", line.strip()[:80])

    # K CJK
    for key in K_FIELDS:
        m = RE_CJK.search(fields[key])
        if m:
            add("K", key, "含 CJK 字元", snippet_of(fields[key], m.start()))

    # L 長度
    head = upper_half(item)
    n_tokens = len(RE_TOKEN.findall(head))
    if n_tokens > length_limit:
        add("L", "test_item", f"上半 {n_tokens} tokens > {length_limit}",
            snippet_of(head))

    # M 空欄三態
    for key in M_FIELDS:
        value = fields[key].strip()
        if value:
            continue
        add("M", key, "空欄（非 NA、非 PENDING:）", "")

    # N 尾句號：命中 [.。]$ 即違規（canon §11 禁尾句號；00b 修訂 1 反轉）
    for key in N_FIELDS:
        for line in split_lines(fields[key]):
            if n_exempt(line):
                continue
            if RE_TRAILING_PERIOD.search(line.rstrip()):
                add("N", key, "行尾多餘句號", line.strip()[:80])

    # P 訊號寫法（範圍依 R-6：作者生成內容，不含 test_item 上半）
    signal_check = check_signal_line_v3 if profile else check_signal_line
    for key in P_FIELDS:
        for line in split_lines(fields[key]):
            out.extend(signal_check(line, key, row_no, tc_id))
    for line in paren_lines(item):                    # test_item 括號下半
        out.extend(signal_check(line, "test_item(括號下半)", row_no, tc_id))

    if not profile:
        return out

    # --- 以下僅於 --profile 指定時啟用 --------------------------------------

    # Q 不可見字元（R-10(a)，全欄位含 verbatim 上半）
    for key in Q_FIELDS:
        for line in fields.get(key, "").split("\n"):
            hits = []
            if "\xa0" in line:
                hits.append("NBSP")
            if "\u3000" in line:
                hits.append("全形空格")
            if RE_Q_TRAILING_WS.search(line):
                hits.append("行尾空白")
            if hits:
                add("Q", key, "／".join(hits), line.strip()[:80])

    # V 行首空白（IN §11，27 包 §D-4）
    for key in Q_FIELDS:
        for line in fields.get(key, "").split("\n"):
            if RE_V_BLANK_WS.match(line):
                add("V", key, "整行僅空白", "")
            elif RE_V_LEADING_WS.match(line) and not RE_V_EXEMPT.match(line):
                add("V", key, "行首空白", line[:80])

    # R Pre-Condition 版面（R-9(a)）
    for line in split_lines(pre):
        if not NUMBERED_LINE.match(line):
            add("R", "pre", "未編號行", line.strip()[:80])
            continue
        body = NUMBER_PREFIX.sub("", line).replace(R_TOOL_PHRASE, "")
        if (" and " in body or ", " in body) and \
                len(RE_R_PREDICATE.findall(body)) >= 2:
            add("R", "pre", "多條件並列於同一行", line.strip()[:80])

    # T PENDING 說明之語言（R-14）
    for key in T_FIELDS:
        for line in split_lines(fields[key]):
            m = RE_T_PENDING.search(line)
            if not m:
                continue
            desc = m.group("desc") or ""
            bad = [c for c in desc if ord(c) > 127]
            if bad:
                add("T", key, f"PENDING 說明含非 ASCII 字元 {bad[:3]!r}",
                    line.strip()[:80])
            # U 佔位之可見性（A-PM16）：ER 側原不受任何檢查覆蓋，
            # 致 `PENDING` 行「未被覆蓋」與「通過」無從分辨。逐一列出。
            add("U", key, f"PENDING 佔位（{m.group('dr') or '未標 DR'}）",
                line.strip()[:80])

    # X 導航路徑之固定入口（§5.8／R-G71）—— WARN 只報不改。
    # 入口以**整列** proc＋pre 為範圍（§5.8(a) 之「同 TC 內」）：入口常寫在
    # Pre-Condition 或前一步驟，逐行判會把正確的多步路徑全報成違規。
    nav_scope = pre + "\n" + proc
    if not RE_X_ENTRY.search(nav_scope):
        for line in split_lines(proc):
            if RE_X_PENDING.search(line):
                continue
            m = RE_X_TARGET.search(line)
            if m:
                add("X", "proc", f"導航標的 {m.group(0)!r} 而同 TC 無固定入口",
                    line.strip()[:80])

    return out


def check_signal_line(line: str, field: str, row_no: int, tc_id: str
                      ) -> list[Violation]:
    """R-1 v2 之單行判定（逐賦值出現，非逐行）。

    三項：(1) 撤銷之三件組殘留；(2) CAN 賦值未寫成
    `<MSG>.<Sig> = <raw> (<label>)`；(3) Procedure 之賦值行缺
    `Send CAN:` 前綴。

    逐「出現」而非逐「行」判定，是因 SWC 語料一行可載多個賦值
    （`… = 1 (Pressed) and BCM_FD_14.Command_09Sts = 0 (Not_Pressed)`），
    且 ER 之收尾語不固定（`is sent`／`is set`／`during …`／`then …`），
    對收尾語設限即與基準本相牴觸。
    """
    out: list[Violation] = []

    def add(detail: str) -> None:
        out.append(Violation("P", row_no, tc_id, field, detail, line.strip()[:80]))

    for m in RE_P_TRIPLET.finditer(line):
        add(f"三件組已撤銷（R-1 v1）{m.group(0)!r}")

    assignments = list(RE_P_ASSIGNMENT.finditer(line))
    if not assignments:
        return out

    for m in assignments:
        if not RE_P_VALUE_FORM.match(line, m.start()):
            add(f"賦值未寫成 `<MSG>.<Sig> = <raw> (<label>)`：{m.group(0)!r}")

    if field == "proc" and "Send CAN:" not in line:
        add("Procedure 之 CAN 賦值行缺 `Send CAN:` 前綴（R-1 v2(a)）")
    return out


def check_signal_line_v3(line: str, field: str, row_no: int, tc_id: str
                         ) -> list[Violation]:
    """R-1 v3 之單行判定（`--profile` 專屬；取代 v2 之 `check_signal_line`）。

    四項：(1) v1 三件組殘留；(2) v2 之 `Send CAN:` 前綴殘留；
    (3) 訊號賦值未以 `$` 包覆全名；(4) `$MSG.Sig$` 之賦值缺 `(<VAL_ label>)`。
    另 (5) `PROXI $X$` —— v3(c) 明定 PROXI 不加 `$`。
    """
    out: list[Violation] = []

    def add(detail: str) -> None:
        out.append(Violation("P", row_no, tc_id, field, detail, line.strip()[:80]))

    for m in RE_P_TRIPLET.finditer(line):
        add(f"三件組已撤銷（R-1 v1）{m.group(0)!r}")
    for m in RE_P3_SEND_CAN.finditer(line):
        add("`Send CAN:` 為 R-1 v2 舊式，v3 改 `Send the signal $MSG.Sig$ = …`")
    for m in RE_P3_BARE_ASSIGN.finditer(line):
        add(f"訊號賦值未以 `$` 包覆全名（R-1 v3(a)）：{m.group(0).strip()!r}")
    for m in RE_P3_DOLLAR_ASSIGN.finditer(line):
        if m.group("lab"):
            continue
        if m.group("val").startswith("PENDING"):     # R-14 佔位，另由 T 檢查
            continue
        add(f"賦值缺 DBC `VAL_` 標籤 `(<label>)`（R-1 v3(a)／R-7）："
            f"{m.group(0).strip()!r}")
    for m in RE_P3_PROXI_DOLLAR.finditer(line):
        add("PROXI 不加 `$`（R-1 v3(c)）")
    # (6) 無點之車輛屬性 `$<Name>$ = [值]`（下放包 43 §二 #1）——
    #     使其**被檢查而非被忽略**：值須以方括號包覆。
    ok = {m.start() for m in RE_P3_PROP_OK.finditer(line)}
    for m in RE_P3_PROP_ANY.finditer(line):
        if "." in m.group(1):          # 含點者由 RE_P3_DOLLAR_ASSIGN 管
            continue
        if m.start() in ok:
            continue
        if m.group("val").startswith("PENDING"):
            continue
        add(f"車輛屬性之值須以 `[…]` 包覆（037 逐字記法）：{m.group(0).strip()!r}")
    return out


def check_sibling_parens(rows: list[tuple[int, str, str, str]]) -> list[Violation]:
    """I-sibling：同 Requirement ID 下多列括號行內容逐字相同。"""
    out: list[Violation] = []
    groups: dict[tuple[str, str], list[tuple[int, str]]] = {}
    for row_no, tc_id, req_id, item in rows:
        content = "\n".join(paren_lines(item))
        if not req_id.strip() or not content:
            continue
        groups.setdefault((req_id.strip(), content), []).append((row_no, tc_id))
    for (req_id, content), members in groups.items():
        if len(members) < 2:
            continue
        for row_no, tc_id in members:
            out.append(Violation(
                "I-sibling", row_no, tc_id, "test_item",
                f"與 {req_id} 下另 {len(members) - 1} 列括號行逐字相同",
                content.replace("\n", " ⏎ ")[:80],
            ))
    return out


# --- W：ER 含比較關係而 `test_item` 上半無數值（下放包 47 §二 #6）------------
#
# 其形態為「指名二個被比較的量，而規格未給任何數值」—— **精度遂由讀者決定**。
# **待人裁非 FAIL**：有些列之精度由畫面粒度給出（`041` 改寫後即是），合法。
RE_W_COMPARE = re.compile(
    r"\b(corresponds? to|equals?|differs? from|matches?|greater than|less than|"
    r"same as|identical to)\b", re.I)
RE_W_NUMERAL = re.compile(r"\d")

# --- X 導航路徑（§5.8／R-G71）------------------------------------------------
# 觸發詞：步驟指向某個畫面／層級而未必寫出入口。
RE_X_TARGET = re.compile(r"\b(menu|page|screen|settings|tab)\b", re.I)
# 固定入口之閉合清單（§5.8(a)）。`H/K "<button>"` 之標籤自由，故只認前綴。
RE_X_ENTRY = re.compile(
    r'Menu Bar|App Drawer|Home Screen|Status Bar|H/K\s*"|Dealer Mode|Eng Mode')
# 已依 §5.8(d) 登記者不重複報 —— PENDING 行本身由 U 承擔。
RE_X_PENDING = re.compile(r"PENDING:\s*DR-")


# --- I-cross（R-SU34 v3）------------------------------------------------------
#
# `I-sibling` 之分組鍵含 `req_id`，故跨 `Requirement ID` 之偽通過**結構上永不觸發**。
# 本檢查補該缺口，其指標為 **觀測窗 × 違例類**（非行文相似度 —— v1 之比率指標
# 經回測與欲測性質負相關而作廢，見 `features/sw_update/scripts/i_cross.py`）。
#
# ⚠ **本檢查有一處前提被寫死在檢查裡**（R-SU34 v3(b) 之明令、PLAYBOOK (33)）：
#   `IX_NORMALISE` 把「未指定之起點」正規化為可用性查詢、
#   把 `until the update finishes` 正規化為版本號改變。
#   **其來源為下放包 30 §2.1 之裁定，不是 TC 之文字。**
#   **若該裁定改動，本表須同步改** —— 否則本檢查會沉默地沿用一個已失效之前提。

IX_START = [(r"from the availability check", "availability-check"),
            (r"from the start of the session", "session-start")]

# ── 訖點之抽取（下放包 43 §二 #4：改語形抽取，不寫死片語）────────────────
#
# **首版為一張寫死之片語表**（`until the software version changes`／
# `until the update finishes`），二者皆出自 `Silent Update` 那一批。
# `ROV Installation` 之 `until the installation ends` 二者皆不匹配，
# 遂被靜默算成半窗（上繳包 37 §2.3）。**每進一個新 Test Set 該表即落後一次。**
#
# **正規化規則（須隨結果揭露）**：
#   1. 取 `until` 之後至行尾／逗號／分號為止之整段子句；
#   2. 去冠詞（`the`／`a`／`an`）、轉小寫、空白收斂為單一連字號；
#   3. 查 `IX_END_ALIAS` —— **其只收「由裁定導出之等價」**，不收語形近似。
RE_IX_UNTIL = re.compile(r"\buntil\s+([^,;]+?)(?=[,;]|$)", re.M)
IX_END_ALIAS = {
    # 下放包 30 §2.1 之裁定：更新完成之唯一外部表徵為版本號改變，
    # 故「更新結束」與「版本改變」為同一訖點。**此為裁定，非語形。**
    "update-finishes": "software-version-changes",
    "update-finish": "software-version-changes",
}


def _ix_end_label(txt: str) -> str | None:
    m = RE_IX_UNTIL.search(txt)
    if not m:
        return None
    s = re.sub(r"\b(the|a|an)\b", " ", m.group(1).lower())
    lab = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return IX_END_ALIAS.get(lab, lab) or None
# ⚠ **正規化之射程限於「起點」**（下放包 30 §2.1 所裁者為起點）。
# **首版誤把同一規則同時套在訖點上**，致無 `until …` 片語之 TC 之窗
# 退化為 `availability-check → availability-check` —— **它長得像一個窗，
# 故沒有任何檢查會攔它**（下放包 34 §1.2）。訖點缺失時**不正規化**。
IX_NORMALISE_START = {None: "availability-check"}      # 未指定之起點 → 唯一可觀測者
# 訖點之等價由 `IX_END_ALIAS` 承擔（見上），此處不再另設對照
# 取最細之類；交集以上下位關係判（R-SU34 v3(b)）
IX_VIOLATION = [
    (r"download confirmation screen", "confirmation-screen/download"),
    (r"deployment confirmation screen", "confirmation-screen/deployment"),
    (r"\bconfirmation screen", "confirmation-screen"),
    (r"SW Update prompt", "prompt"),
    (r"progress notification", "progress-notification"),
    (r"opt-out control", "opt-out"),
    (r"defer control", "defer"),
]
IX_NEG = re.compile(r"contains no |no SW Update prompt|no progress notification"
                    r"|no download confirmation|no deployment confirmation"
                    r"|no confirmation screen|no opt-out|no defer")


def _ix_window(proc: str, er: str) -> tuple[str | None, str | None]:
    """回傳 (起, 訖)。訖點無片語可抽時為 `None` —— **半窗，不參與比對**。"""
    txt = proc + " " + er
    s = next((v for p, v in IX_START if re.search(p, txt)), None)
    e = _ix_end_label(txt)
    return IX_NORMALISE_START.get(s, s), e


def _ix_violations(er: str) -> set[str]:
    """僅取**否定式**之 ER 行；同行命中概括式與子類時只留子類。"""
    out: set[str] = set()
    for ln in er.split("\n"):
        if not IX_NEG.search(ln):
            continue
        hit = {v for p, v in IX_VIOLATION if re.search(p, ln)}
        out |= {v for v in hit
                if not any(o != v and o.startswith(v + "/") for o in hit)}
    return out


def _ix_subsumes(a: str, b: str) -> bool:
    return a == b or a.startswith(b + "/") or b.startswith(a + "/")


def check_cross(rows: list[tuple[int, str, str, str, str, str]]) -> list[Violation]:
    """I-cross：**同一 Test Set 內**，觀測窗相同且違例類有交集之跨 req_id 配對。

    **警示器非判準**（R-SU34 v3(c)）—— 窗同而違例類不同者合法。
    人裁所問為「本 TC 是否有屬於其需求單元之驗證點」，
    **不是**「其驗證點是否被他 TC 涵蓋」——**覆蓋是允許的**（R-SU34 v3(e)）。
    """
    out: list[Violation] = []
    info: list = []
    half: list = []          # 窗未完整宣告者（待補），不參與比對
    for row_no, tc_id, req_id, test_set, proc, er in rows:
        w = _ix_window(proc, er)
        if w[0] is None or w[1] is None:
            # **半窗**：R-SU36(b)／R-SU33 v1(b) 令 ER 須明載窗之起訖，
            # 抽不出訖點者其窗未完整宣告 —— **不參與比對，列為待補**。
            half.append((row_no, tc_id, w))
            continue
        info.append((row_no, tc_id, req_id, test_set, w, _ix_violations(er)))
    for i, (ra, ta, qa, sa, wa, va) in enumerate(info):
        for rb, tb, qb, sb, wb, vb in info[i + 1:]:
            if qa.strip() == qb.strip() or sa.strip() != sb.strip():
                continue          # 同 req_id 由 I-sibling 管；跨 Test Set 不比
            if wa != wb:
                continue
            inter = {min(x, y, key=len) for x in va for y in vb
                     if _ix_subsumes(x, y)}
            if not inter:
                continue
            for rn, tid, other in ((ra, ta, tb), (rb, tb, ta)):
                out.append(Violation(
                    "I-cross", rn, tid, "expected_result",
                    f"與 {other} 之觀測窗相同（{wa[0]} → {wa[1]}）且違例類有交集",
                    "／".join(sorted(inter))[:80],
                ))
    for rn, tid, w in half:
        out.append(Violation(
            "I-cross", rn, tid, "expected_result",
            "**窗未完整宣告** —— 訖點無片語可抽，本列不參與 I-cross 比對"
            "（R-SU33(b)：ER 須明載窗之起訖）",
            f"起 {w[0] or '—'} → 訖 **未載**",
        ))
    return out


# --- 工作簿層 ----------------------------------------------------------------


def lint_workbook(path: Path, length_limit: int = DEFAULT_LENGTH_LIMIT,
                  profile: str | None = None) -> list[SheetResult]:
    """唯讀開啟工作簿，對每個 TC sheet 跑檢查。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        results: list[SheetResult] = []
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith(TC_SHEET_PREFIX):
                continue
            results.append(lint_sheet(wb[sheet_name], length_limit, profile))
        if not results:
            raise ValueError(f"{path.name}：找不到以 {TC_SHEET_PREFIX!r} 開頭之 sheet")
        return results
    finally:
        wb.close()


def lint_sheet(ws, length_limit: int, profile: str | None = None) -> SheetResult:
    """單一 sheet 的檢查流程。"""
    header_row = find_header_row(ws)
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    columns = build_column_map(list(rows[header_row - 1]))
    missing = [k for k in FIELD_HEADERS if k not in columns]
    if missing:
        raise ValueError(f"sheet {ws.title!r} 缺欄位：{missing}")

    result = SheetResult(sheet=ws.title, header_row=header_row, data_rows=0)
    sibling_input: list[tuple[int, str, str, str]] = []
    cross_input: list[tuple[int, str, str, str, str, str]] = []

    for offset, raw in enumerate(rows[header_row:], start=header_row + 1):
        fields = {key: cell_text(raw[idx]) if idx < len(raw) else ""
                  for key, idx in columns.items() if key in FIELD_HEADERS}
        if not any(fields[k].strip() for k in ("test_item", "proc", "er")):
            continue
        tc_id = cell_text(raw[columns["tc_id"]]) if "tc_id" in columns else ""
        req_id = cell_text(raw[columns["req_id"]]) if "req_id" in columns else ""
        result.data_rows += 1
        result.violations.extend(
            check_row(fields, offset, tc_id, length_limit, profile))
        sibling_input.append((offset, tc_id, req_id, fields["test_item"]))
        test_set = (cell_text(raw[columns["test_set"]])
                    if "test_set" in columns else "")
        cross_input.append((offset, tc_id, req_id, test_set,
                            fields["proc"], fields["er"]))

    result.violations.extend(check_sibling_parens(sibling_input))
    result.cross_rows = cross_input
    if profile:                       # I-cross 為 profile 專屬（PROFILE_CHECKS）
        result.violations.extend(check_cross(cross_input))
    order = check_order(profile)
    result.violations.sort(key=lambda v: (order.index(v.check), v.row))
    return result


def count_by_check(results: list[SheetResult],
                   profile: str | None = None) -> dict[str, int]:
    """彙總各檢查之行計（違規記錄數）。"""
    counts = {key: 0 for key in check_order(profile)}
    for result in results:
        for violation in result.violations:
            counts[violation.check] += 1
    return counts


def rows_by_check(results: list[SheetResult],
                  profile: str | None = None) -> dict[str, int]:
    """彙總各檢查之列計（涉及之相異資料列數）。"""
    seen: dict[str, set[tuple[str, int]]] = {key: set()
                                             for key in check_order(profile)}
    for result in results:
        for violation in result.violations:
            seen[violation.check].add((result.sheet, violation.row))
    return {key: len(value) for key, value in seen.items()}


# --- 報告輸出 ----------------------------------------------------------------


def render_report(path: Path, results: list[SheetResult], length_limit: int,
                  profile: str | None = None) -> str:
    """產生 markdown 報告。"""
    order = check_order(profile)
    counts = count_by_check(results, profile)
    row_counts = rows_by_check(results, profile)
    total_rows = sum(r.data_rows for r in results)
    lines = [
        f"# lint036 報告：{path.name}",
        "",
        f"- 來源：`{path}`（唯讀）",
        f"- 資料列數：{total_rows}",
        f"- sheet：" + ", ".join(f"`{r.sheet}`（header 第 {r.header_row} 列）"
                                for r in results),
        f"- L 閾值：{length_limit} tokens",
        "",
        "## 違規統計",
        "",
        "計數口徑：**行計為主**（違規記錄數，粒度見「粒度」欄），"
        "**附列計**（涉及之相異資料列數）。兩者不可互相加總。",
        "",
        "| 檢查 | 項目 | 行計 | 列計 | 粒度 | 校準 |",
        "| --- | --- | ---: | ---: | --- | --- |",
    ]
    for key in order:
        lines.append(
            f"| {key} | {check_title(key, profile)} | {counts[key]} "
            f"| {row_counts[key]} "
            f"| {CHECK_GRANULARITY[key]} | {check_status(key, profile)} |"
        )
    if profile:
        lines.insert(6, f"- profile：`{profile}`（P 採 R-1 v3；另跑 Q／R／T）")
    lines += ["",
              f"**總計：行計 {sum(counts.values())}**"
              f"（列計不加總——同一列可觸發多項檢查）",
              "", "## 明細", ""]

    for key in order:
        items = [v for r in results for v in r.violations if v.check == key]
        if not items:
            continue
        affected = len({(v.row) for v in items})
        lines += [f"### {key} — {check_title(key, profile)}"
                  f"（行計 {len(items)}／列計 {affected}）", "",
                  "| 列 | TC ID | 欄位 | 說明 | 片段 |",
                  "| ---: | --- | --- | --- | --- |"]
        for v in items:
            detail = v.detail.replace("|", "\\|")
            snippet = v.snippet.replace("|", "\\|")
            lines.append(f"| {v.row} | {v.tc_id} | {v.field} | {detail} | {snippet} |")
        lines.append("")
    return "\n".join(lines) + "\n"


def source_sha8(path: Path) -> str:
    """來源工作簿之 sha256 前 8 碼（26 包 §C 裁定 3）。

    報告檔名自本裁定後之新產報告起採 `{tag}_{來源檔sha8}_{YYYYMMDD}`。
    **`tag` 本身不足以識別報告**：同一 feature 之兩個來源日期於同一天被
    lint，其 `{tag}_{今日}` 相同，後者靜默覆寫前者（25 上繳 §九-2 實測
    18 組）。sha8 帶回來源之身分，且它比檔名可靠 —— **檔名可以改，
    位元組不會**。

    讀不到者回 `nosha`，**不回退為空字串** —— 空字串會使檔名退回舊式而
    看起來正常，`nosha` 則在檔名上自陳其缺。
    """
    try:
        return hashlib.sha256(path.read_bytes()).hexdigest()[:8]
    except OSError:
        return "nosha"


def report_stem(path: Path) -> str:
    """由檔名推出報告 tag（取 SWQT_ 之後、日期之前的段）。

    **日期起首之 tag 併入 feature 名**（25 包 §D-6）：036 母本之副本檔名為
    `…_SWQT_20260817_ext.xlsx`，`SWQT_` 之後直接是日期，去日期後 tag 為
    `20260817_ext` —— **不帶任何 workbook 身分**，故 `user_profiles`／
    `time_management`／`power_moding` 三份之報告互相覆寫。
    此形態下以 `features/<name>/` 之 name 前置，使 tag 帶回身分。

    **非此形態者 tag 一律不變** —— `AMFM`／`Home`／`CFTS012_DealerMode`
    等既有八本之報告檔名須維持（G-N 之回歸向）。
    """
    stem = path.stem
    m = re.search(r"SWQT_(.+)$", stem)
    tag = m.group(1) if m else stem
    tag = re.sub(r"_\d{8}.*$", "", tag)          # 去除尾端日期與 (done)/(Refine) 註記
    tag = re.sub(r"[^\w.-]+", "_", tag).strip("_")
    if re.match(r"^\d{8}(?:[_.-]|$)", tag):
        tag = f"{_identity_dir(path)}_{tag}"
    return tag


# 通用容器目錄 —— 其名不帶 workbook 身分，取身分時跳過
_GENERIC_DIRS = {"inputs", "SWE6", "docs", "test", "data", "generated", "output", "_intake"}


def _identity_dir(path: Path) -> str:
    """日期起首之 tag 取身分用之目錄名。

    `features/<name>/…` 取 `<name>`；否則取**最近之非通用容器**祖先目錄
    （`docs/test/Dealer Mode/SWE6/x.xlsx` → `Dealer_Mode`）。
    皆無者取 `unknown` —— 不回退為空字串，否則又得到一個不帶身分之 tag。
    """
    parts = path.resolve().parts
    if "features" in parts:
        i = parts.index("features")
        if i + 1 < len(parts):
            return parts[i + 1]
    for name in reversed(parts[:-1]):
        if name not in _GENERIC_DIRS and not name.startswith("."):
            return re.sub(r"[^\w.-]+", "_", name).strip("_") or "unknown"
    return "unknown"


# --- CLI ---------------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="lint036.py",
        description="FM-WI-FSM-036 工作簿靜態檢查（報告模式，唯讀）",
    )
    parser.add_argument("files", metavar="FILES", nargs="+",
                        help="一個或多個 .xlsx 路徑")
    parser.add_argument("--report-dir", default="docs/fw036/lint_reports",
                        help="報告輸出目錄（預設 docs/fw036/lint_reports/）")
    parser.add_argument("--gate", action="store_true",
                        help="任一違規 exit 1（本包不啟用）")
    parser.add_argument("--json", action="store_true",
                        help="另輸出機讀 json")
    parser.add_argument("--length-limit", type=int, default=DEFAULT_LENGTH_LIMIT,
                        help=f"L 檢查 token 閾值（預設 {DEFAULT_LENGTH_LIMIT}）")
    parser.add_argument("--merge", action="store_true",
                        help="把所有 FILES 之列視為同一本簿再跑一次 I-cross —— "
                             "使比對範圍等同**交付簿**。開發期之 sandbox 分簿"
                             "會使跨簿配對逐簿比不到（PLAYBOOK (36)）。"
                             "逐簿之報告不受影響。")
    parser.add_argument("--profile", default=None, metavar="FEATURE",
                        help="feature 專屬判準：P 改採 R-1 v3，另跑 Q／R／T。"
                             "未指定時行為與既有八本之報告基線完全一致")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    report_dir = Path(args.report_dir)
    report_dir.mkdir(parents=True, exist_ok=True)

    total_violations = 0
    for raw_path in args.files:
        path = Path(raw_path)
        if not path.is_file():
            print(f"錯誤：找不到檔案 {path}", file=sys.stderr)
            return 2
        results = lint_workbook(path, args.length_limit, args.profile)
        counts = count_by_check(results, args.profile)
        total_violations += sum(counts.values())

        tag = report_stem(path)
        if args.profile:
            tag = f"{tag}__{args.profile}"
        report_path = report_dir / f"{tag}_{source_sha8(path)}_{date.today():%Y%m%d}.md"
        report_path.write_text(
            render_report(path, results, args.length_limit, args.profile),
            encoding="utf-8")
        print(f"{path.name}\n  -> {report_path}")
        print("  行計 " + "  ".join(f"{k}={counts[k]}"
                                   for k in check_order(args.profile)))

        if args.json:
            json_path = report_path.with_suffix(".json")
            payload = {
                "source": str(path),
                "profile": args.profile,
                "counts": counts,
                "row_counts": rows_by_check(results, args.profile),
                "granularity": CHECK_GRANULARITY,
                "status": {k: check_status(k, args.profile)
                           for k in check_order(args.profile)},
                "sheets": [
                    {
                        "sheet": r.sheet,
                        "header_row": r.header_row,
                        "data_rows": r.data_rows,
                        "violations": [v.as_dict() for v in r.violations],
                    }
                    for r in results
                ],
            }
            json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2),
                                 encoding="utf-8")
            print(f"  -> {json_path}")

    # --- `--merge`：跨簿之 I-cross（R-SU34 v1(d)／PLAYBOOK (36)）-------------
    if args.merge:
        if not args.profile:
            print("錯誤：--merge 須與 --profile 併用（I-cross 為 profile 專屬）",
                  file=sys.stderr)
            return 2
        pooled: list[tuple] = []
        for raw_path in args.files:
            for r in lint_workbook(Path(raw_path), args.length_limit, args.profile):
                pooled.extend(r.cross_rows)
        merged_v = check_cross(pooled)
        pairs = sorted({tuple(sorted((v.tc_id, v.detail.split("與 ")[1].split(" 之")[0])))
                        for v in merged_v if v.detail.startswith("與 ")})
        half = sorted({v.tc_id for v in merged_v if "窗未完整宣告" in v.detail})
        print(f"\n=== --merge：{len(args.files)} 簿併為一，共 {len(pooled)} 列 ===")
        print(f"  I-cross(merged) = {len(merged_v)}"
              f"（配對 {len(pairs)} 組；窗未完整宣告 {len(half)} 列）")
        for a, b in pairs:
            print(f"    · {a} ↔ {b}")
        if half:
            print("    窗未完整宣告（不參與比對，待補）：" + "、".join(half))

    if args.gate and total_violations:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
