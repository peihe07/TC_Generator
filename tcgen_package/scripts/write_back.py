#!/usr/bin/env python3
"""Step 4 — write generated TCs back into the FW036 workbook.

Produces `output/FW036_regen.xlsx` from the source workbook plus
`generated/*.json`:

  - rows 10-332 (the compliant human-authored region) are left untouched
  - rows 333+ are replaced by the generated TCs in 037 document order
  - blocked parents still get a row, so the deliverable shows the hole
  - the `Test Case Framework` sheet gains the two Test Sets ruled in by Pei
  - `ChangeHistory` gains a revision row

This is not a one-shot script. Every RD-1 ruling that lands (the A-026 label
pass, the A-029 flag pass, assumption rework) means regenerating this workbook,
so the run must be reproducible and self-checking rather than something a
person reconstructs from memory.

Three invariants are asserted while writing, not verified afterwards:

  1. **Traceability** — every req_id written must exist in 037. Row-level lint
     cannot catch an invented `-02` sub-id: such a row is well-formed in every
     other respect. Only reconciliation against the leaf list catches it.
  2. **Completeness** — the set of leaves written must equal the set of
     remaining leaves exactly: none missing, none extra.
  3. **Done region untouched** — rows 10-332 are out of scope for this
     regeneration (cf. ANOMALIES A-005) and are hashed before and after.

Usage:
    python scripts/write_back.py --src <036.xlsx> --data data --generated generated
    python scripts/write_back.py ... --out output/FW036_regen.xlsx --rev D
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sys
import warnings
from datetime import date
from pathlib import Path

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402

REPO = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(REPO / "backend"))
from validator import VALID_DESIGN_METHODS, VALID_PRIORITIES  # noqa: E402

SHEET = "Test Case Specification 測試用例規範"
FRAMEWORK_SHEET = "Test Case Framework"
HISTORY_SHEET = "ChangeHistory 修訂履歷"

FIRST_DATA_ROW = 10
DONE_REGION_LAST_ROW = 332
FIRST_NEW_ROW = DONE_REGION_LAST_ROW + 1

# 1-based column numbers, per RUNBOOK Step 4 plus AH (Remarks), which the
# runbook's original mapping omitted — the blocked declaration lives there.
COL = {
    "no": 2, "req_id": 4, "tc_id": 6, "test_group": 7, "test_set": 8,
    "test_item": 9, "pre_conditions": 10, "input_test_data": 11,
    "test_procedure": 12, "expected_result": 13, "spec_ref": 14,
    "tc_ref": 15, "priority": 16, "design_method": 18, "functional_safety": 19,
    "author": 27, "remarks": 34,
}
VEHICLE_COLS = range(20, 27)  # T..Z
NEW_TEST_SETS = {15: "Preset Management", 16: "Media Widget"}
AUTHOR = "PeiPYHsu"
BLOCKED_PLACEHOLDER = "BLOCKED - see Remarks"


class WriteBackError(RuntimeError):
    """An invariant failed; no workbook is produced."""


# ---------------------------------------------------------------------------
# inputs
# ---------------------------------------------------------------------------

def load_leaves(data_dir: Path) -> list[dict]:
    return json.loads((data_dir / "remaining_leaves.json").read_text(encoding="utf-8"))


def parent_order(leaves: list[dict]) -> list[str]:
    """037 document order — the order parents first appear in the leaf list."""
    seen: list[str] = []
    for leaf in leaves:
        if leaf["parent"] not in seen:
            seen.append(leaf["parent"])
    return seen


def collect_rows(generated_dir: Path, leaves: list[dict]) -> list[dict]:
    """Flatten generated parents into workbook rows, in 037 document order."""
    by_id = {leaf["req_id"]: leaf for leaf in leaves}
    rows: list[dict] = []

    for parent in parent_order(leaves):
        path = generated_dir / f"{parent}.json"
        if not path.exists():
            raise WriteBackError(f"no generated file for parent {parent}")
        payload = json.loads(path.read_text(encoding="utf-8"))
        blocked = payload.get("blocked")
        write_back = payload.get("write_back") or {}

        if blocked and not payload.get("tcs"):
            spec = blocked.get("write_back", {})
            if not spec.get("emit_row"):
                raise WriteBackError(
                    f"{parent} is blocked without emit_row; the leaf would vanish "
                    "from the deliverable and break traceability")
            for req_id in blocked.get("req_ids", []):
                leaf = by_id.get(req_id)
                if leaf is None:
                    raise WriteBackError(f"{parent}: blocked req_id {req_id} is not a leaf in 037")
                rows.append({
                    "req_id": req_id,
                    "test_set": spec.get("test_set") or _blocked_test_set(payload),
                    "test_item": leaf["description"],
                    "pre_conditions": "NA",
                    "input_test_data": "NA",
                    "test_procedure": spec.get("test_procedure", BLOCKED_PLACEHOLDER),
                    "expected_result": spec.get("expected_result", BLOCKED_PLACEHOLDER),
                    "spec_ref": leaf["hmi_source_id"],
                    "priority": spec.get("priority", ""),
                    "design_method": spec.get("design_method", ""),
                    "remarks": spec.get("remarks", ""),
                    "blocked": True,
                })
            continue

        for tc in payload.get("tcs", []):
            rows.append({
                "req_id": tc["req_id"],
                "test_set": tc["test_set"],
                "test_item": tc["test_item"],
                "pre_conditions": tc["pre_conditions"],
                "input_test_data": tc["input_test_data"],
                "test_procedure": tc["test_procedure"],
                "expected_result": tc["expected_result"],
                "spec_ref": tc["specification_reference"],
                "priority": tc["priority"],
                "design_method": tc["design_method"],
                "remarks": write_back.get("remarks", ""),
                "blocked": False,
            })
    return rows


def _blocked_test_set(payload: dict) -> str:
    """A blocked parent has no TC to take the Test Set from; use its declaration."""
    ts = (payload.get("blocked", {}).get("write_back", {}) or {}).get("test_set")
    if ts:
        return ts
    ts = payload.get("test_set")
    if ts:
        return ts
    raise WriteBackError(f"{payload.get('parent')}: blocked row needs a test_set")


# ---------------------------------------------------------------------------
# invariants
# ---------------------------------------------------------------------------

def assert_traceable_and_complete(rows: list[dict], leaves: list[dict]) -> None:
    """Invariants 1 and 2, checked before anything is written."""
    leaf_ids = {leaf["req_id"] for leaf in leaves}
    written = {r["req_id"] for r in rows}

    invented = sorted(written - leaf_ids)
    if invented:
        raise WriteBackError(
            "req_ids not present in 037 — several TCs may share one sub-id (§8.2.2), "
            f"but a new sub-id invents a requirement: {invented}")

    missing = sorted(leaf_ids - written)
    if missing:
        raise WriteBackError(f"leaves with no row: {missing}")


def assert_row_shape(rows: list[dict]) -> None:
    """Dropdown-constrained fields must hold legal values (blank only when blocked)."""
    bad = []
    for r in rows:
        if r["blocked"]:
            if r["priority"] or r["design_method"]:
                bad.append((r["req_id"], "blocked row must leave priority/design_method blank"))
            continue
        if r["priority"] not in VALID_PRIORITIES:
            bad.append((r["req_id"], f"priority {r['priority']!r}"))
        if r["design_method"] not in VALID_DESIGN_METHODS:
            bad.append((r["req_id"], f"design_method {r['design_method']!r}"))
    if bad:
        raise WriteBackError(f"illegal dropdown values: {bad}")


def hash_region(ws, first: int, last: int) -> str:
    """Stable digest of a row range, for the done-region invariant."""
    h = hashlib.sha256()
    for row in range(first, last + 1):
        for col in range(1, 35):
            h.update(repr(ws.cell(row, col).value).encode("utf-8"))
        h.update(b"\n")
    return h.hexdigest()


# ---------------------------------------------------------------------------
# writing
# ---------------------------------------------------------------------------

def write_rows(ws, rows: list[dict]) -> int:
    ws.delete_rows(FIRST_NEW_ROW, max(ws.max_row - DONE_REGION_LAST_ROW, 0))
    row_no = FIRST_NEW_ROW
    for r in rows:
        ws.cell(row_no, COL["no"]).value = "=ROW()-9"
        ws.cell(row_no, COL["tc_id"]).value = (
            f'=IF(ISBLANK($AA{row_no}),"","NR1L-MediaHMI-"&TEXT(ROW()-9,"000"))')
        ws.cell(row_no, COL["req_id"]).value = r["req_id"]
        ws.cell(row_no, COL["test_group"]).value = "MediaHMI"
        ws.cell(row_no, COL["test_set"]).value = r["test_set"]
        ws.cell(row_no, COL["test_item"]).value = r["test_item"]
        ws.cell(row_no, COL["pre_conditions"]).value = r["pre_conditions"]
        ws.cell(row_no, COL["input_test_data"]).value = r["input_test_data"]
        ws.cell(row_no, COL["test_procedure"]).value = r["test_procedure"]
        ws.cell(row_no, COL["expected_result"]).value = r["expected_result"]
        ws.cell(row_no, COL["spec_ref"]).value = r["spec_ref"]
        ws.cell(row_no, COL["tc_ref"]).value = "NEW"
        if r["priority"]:
            ws.cell(row_no, COL["priority"]).value = r["priority"]
        if r["design_method"]:
            ws.cell(row_no, COL["design_method"]).value = r["design_method"]
        ws.cell(row_no, COL["functional_safety"]).value = "NA"
        for col in VEHICLE_COLS:
            ws.cell(row_no, col).value = 1
        ws.cell(row_no, COL["author"]).value = AUTHOR
        if r["remarks"]:
            ws.cell(row_no, COL["remarks"]).value = r["remarks"]
        row_no += 1
    return row_no - 1


def sync_framework_sheet(wb) -> list[str]:
    """Add the two Test Sets ruled in for ch18/ch23 (see docs/framework.md)."""
    ws = wb[FRAMEWORK_SHEET]
    added = []
    for row, label in NEW_TEST_SETS.items():
        if ws.cell(row, 1).value != label:
            ws.cell(row, 1).value = label
            added.append(label)
    return added


def append_history(wb, revision: str, when: str, last_row: int, n_rows: int) -> None:
    """Controlled document: the revision history must record the change."""
    ws = wb[HISTORY_SHEET]
    row = 5
    while ws.cell(row, 1).value not in (None, ""):
        row += 1
    ws.cell(row, 1).value = revision
    ws.cell(row, 2).value = (
        f"Rows {FIRST_NEW_ROW}-{last_row} regenerated ({n_rows} rows) from "
        "FMWIFSM037A03 remaining leaves; rows 10-332 unchanged.\n"
        f"Added Test Sets to Test Case Framework: {', '.join(NEW_TEST_SETS.values())}.\n"
        "Blocked leaves carry a row with the reason in Remarks (see ANOMALIES A-009, A-011)."
    )
    ws.cell(row, 3).value = AUTHOR
    ws.cell(row, 4).value = when


def next_revision(wb) -> str:
    ws = wb[HISTORY_SHEET]
    row, last = 5, "A"
    while ws.cell(row, 1).value not in (None, ""):
        last = str(ws.cell(row, 1).value).strip()
        row += 1
    return chr(ord(last[-1]) + 1) if len(last) == 1 and last.isalpha() else f"{last}+1"


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


# An xlsx is a zip, and both its entries and docProps/core.xml carry a
# wall-clock timestamp. Left alone, two runs over identical content produce
# different bytes — which would defeat recording the file's SHA256 against the
# commit that produced it. Normalising makes the digest mean "this content".
_EPOCH = (1980, 1, 1, 0, 0, 0)
# The back-reference matters: without it the alternation can pair a `created`
# opening tag with a `modified` closing tag and swallow everything between,
# producing XML that Excel refuses to open.
_DCTERMS_RE = re.compile(rb"(<dcterms:(created|modified)[^>]*>)[^<]*(</dcterms:\2>)")
_FIXED_STAMP = b"2000-01-01T00:00:00Z"
# \g<n> rather than \n: the replacement is built by concatenation, so a bare
# \1 followed by the stamp reads as \120 — an octal escape, not group 1.


def normalize_for_reproducibility(path: Path) -> None:
    """Rewrite the xlsx so identical content yields identical bytes."""
    import zipfile

    with zipfile.ZipFile(path) as src:
        entries = [(i.filename, src.read(i.filename)) for i in src.infolist()]

    entries = [
        (name, _DCTERMS_RE.sub(rb"\g<1>" + _FIXED_STAMP + rb"\g<3>", data)
               if name == "docProps/core.xml" else data)
        for name, data in entries
    ]

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as out:
        for name, data in entries:
            info = zipfile.ZipInfo(name, date_time=_EPOCH)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            out.writestr(info, data)


# ---------------------------------------------------------------------------

def run(src: Path, data: Path, generated: Path, out: Path,
        revision: str | None = None, when: str | None = None) -> dict:
    leaves = load_leaves(data)
    rows = collect_rows(generated, leaves)
    assert_traceable_and_complete(rows, leaves)
    assert_row_shape(rows)

    src_wb = openpyxl.load_workbook(src)
    before = hash_region(src_wb[SHEET], FIRST_DATA_ROW, DONE_REGION_LAST_ROW)

    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(src, out)
    wb = openpyxl.load_workbook(out)
    ws = wb[SHEET]

    last_row = write_rows(ws, rows)
    added = sync_framework_sheet(wb)
    rev = revision or next_revision(wb)
    append_history(wb, rev, when or date.today().isoformat(), last_row, len(rows))
    wb.save(out)
    normalize_for_reproducibility(out)

    after = hash_region(openpyxl.load_workbook(out)[SHEET], FIRST_DATA_ROW, DONE_REGION_LAST_ROW)
    if before != after:
        raise WriteBackError("done region rows 10-332 changed; they are out of scope (A-005)")

    return {
        "out": str(out), "rows": len(rows), "last_row": last_row,
        "blocked": sum(1 for r in rows if r["blocked"]),
        "leaves": len(leaves), "revision": rev,
        "framework_added": added, "sha256": sha256_file(out),
        "done_region_sha256": before,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--src", required=True, help="source FMWIFSM036A01 workbook")
    ap.add_argument("--data", default="data")
    ap.add_argument("--generated", default="generated")
    ap.add_argument("--out", default="output/FW036_regen.xlsx")
    ap.add_argument("--revision", help="ChangeHistory revision letter (default: next)")
    ap.add_argument("--date", dest="when", help="ChangeHistory date (default: today)")
    args = ap.parse_args()

    try:
        result = run(Path(args.src), Path(args.data), Path(args.generated),
                     Path(args.out), args.revision, args.when)
    except WriteBackError as e:
        print(f"write-back aborted: {e}", file=sys.stderr)
        return 1

    print(f"{result['out']}")
    print(f"  leaves       : {result['leaves']} (all covered)")
    print(f"  rows written : {result['rows']} -> {FIRST_NEW_ROW}..{result['last_row']}"
          f" ({result['blocked']} blocked)")
    print(f"  framework    : {result['framework_added'] or 'already in sync'}")
    print(f"  revision     : {result['revision']}")
    print(f"  sha256       : {result['sha256']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
